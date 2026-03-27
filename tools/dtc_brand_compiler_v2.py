#!/usr/bin/env python3
"""
DTC Brand Compiler V2
3000+ DTC brands with websites, categories, Meta Ads Library links,
and notable insights (marketing angle, unique selling point, celebrity ties, etc.)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote
import os
from datetime import datetime
import sys
sys.path.insert(0, os.path.dirname(__file__))
from dtc_brands_extra import EXTRA_BRANDS
from dtc_brands_extra2 import EXTRA_BRANDS_2
from dtc_brands_extra3 import EXTRA_BRANDS_3
from dtc_brands_final import FINAL_BRANDS
from dtc_brands_final2 import FINAL_BRANDS_2

# ============================================================================
# BRAND DATABASE
# Format: (Brand, Website, Sub-niche, Notable Insight)
# ============================================================================

BRANDS = {

    # =========================================================================
    # BEAUTY & SKINCARE
    # =========================================================================
    "Beauty & Skincare": [
        ("Glossier", "glossier.com", "Skincare & Makeup", "Built entirely on user-generated content and 'skin first' philosophy; Emily Weiss grew it from a blog (Into The Gloss) to a $1.8B valuation"),
        ("Rhode", "rhodeskin.com", "Skincare", "Hailey Bieber's brand; went viral with peptide lip treatment that sold out in minutes; leverages her 50M+ Instagram for zero-cost marketing"),
        ("Drunk Elephant", "drunkelephant.com", "Skincare", "Pioneered 'suspicious 6' free philosophy (no essential oils, silicones, etc.); sold to Shiseido for $845M; massive TikTok 'smoothie' trend"),
        ("The Ordinary", "theordinary.com", "Skincare", "Disrupted skincare by selling clinical ingredients at drugstore prices ($5-10); radical transparency on ingredient percentages"),
        ("Fenty Beauty", "fentybeauty.com", "Cosmetics", "Rihanna's brand; launched with 40 foundation shades, forced entire industry to expand shade ranges; $550M+ annual revenue"),
        ("Kylie Cosmetics", "kyliecosmetics.com", "Cosmetics", "Kylie Jenner leveraged lip filler controversy into Lip Kit empire; reached $420M revenue by age 21; sold 51% to Coty for $600M"),
        ("Charlotte Tilbury", "charlottetilbury.com", "Cosmetics", "Celebrity MUA turned founder; 'Pillow Talk' lipstick sells one every 2 seconds globally; acquired by Puig for $1.4B"),
        ("Tatcha", "tatcha.com", "Skincare", "Japanese beauty rituals for Western market; featured on Meghan Markle's wedding day skincare; sold to Unilever"),
        ("Summer Fridays", "summerfridays.com", "Skincare", "Founded by influencers Marianna Hewitt & Lauren Gores; Jet Lag Mask went viral as the 'Instagram mask'"),
        ("Supergoop", "supergoop.com", "Suncare", "Made sunscreen 'cool' and wearable daily; Unseen Sunscreen became a TikTok cult favorite as makeup primer"),
        ("Tower 28", "tower28beauty.com", "Clean Beauty", "Founded for people with sensitive/eczema-prone skin; every product is National Eczema Association approved"),
        ("Kosas", "kosas.com", "Clean Beauty", "Coined 'makeup as skincare' with active ingredients in color cosmetics; Revealer Concealer won 15+ best-of awards"),
        ("Ilia Beauty", "iliabeauty.com", "Clean Beauty", "Super Serum Skin Tint became the #1 selling complexion product at Sephora; clean beauty meets real performance"),
        ("Merit", "meritbeauty.com", "Minimalist Makeup", "5-minute face philosophy; only makes products you'd use daily; grew 150% YoY by saying no to product launches"),
        ("Saie", "saiehello.com", "Clean Beauty", "Glowy Super Gel blush went viral on TikTok with 100M+ views; first clean beauty brand to launch exclusively at Sephora"),
        ("Jones Road Beauty", "jonesroadbeauty.com", "Makeup", "Founded by Bobbi Brown after leaving her own namesake brand; Miracle Balm generates $50M+; targets 40+ women"),
        ("Rare Beauty", "rarebeauty.com", "Cosmetics", "Selena Gomez's brand; Soft Pinch blush sells one every 3 seconds; 1% of sales fund mental health ($15M+ donated)"),
        ("Herbivore Botanicals", "herbivorebotanicals.com", "Natural Skincare", "100% natural, vegan ingredients; Blue Tansy mask became iconic for its bright blue color on Instagram"),
        ("Youth to the People", "youthtothepeople.com", "Skincare", "Superfood-powered skincare in apothecary glass bottles; acquired by L'Oréal; 'cold-pressed' juicery aesthetic"),
        ("Biossance", "biossance.com", "Clean Skincare", "Uses patented sugarcane-derived squalane; born from biotech company Amyris; science-meets-sustainability angle"),
        ("Glow Recipe", "glowrecipe.com", "K-Beauty Inspired", "Watermelon Glow Niacinamide Dew Drops became a cult product; Korean co-founders Christine Chang & Sarah Lee from Shark Tank"),
        ("Peach & Lily", "peachandlily.com", "K-Beauty", "Founder Alicia Yoon is a licensed esthetician; Glass Skin Serum coined the 'glass skin' trend in the West"),
        ("Versed", "versedskin.com", "Affordable Skincare", "Crowdsourced product development — 100K+ people voted on formulas before launch; $10-20 price point"),
        ("Cocokind", "cocokind.com", "Sustainable Skincare", "Shows cost breakdown on every product label (ingredients, packaging, labor); radical price transparency"),
        ("Tula", "tula.com", "Probiotic Skincare", "Probiotic-powered skincare; founded by gastroenterologist Dr. Roshini Raj; gut-skin connection angle"),
        ("Kopari", "koparibeauty.com", "Coconut Beauty", "100% organic coconut oil base for everything; started with one product (coconut melt); bright turquoise branding"),
        ("Osea Malibu", "oseamalibu.com", "Seaweed Skincare", "Seaweed-based skincare; family-run for 25+ years before going viral on TikTok; anti-trend longevity"),
        ("Beautycounter", "beautycounter.com", "Clean Beauty", "Lobbied Congress for safer beauty legislation; 'Never List' of 1,800 banned ingredients; social selling model"),
        ("Naturium", "naturium.com", "Skincare", "Founded by beauty YouTuber Susan Yara (controversial launch kept her involvement secret); acquired by E.l.f. for $355M"),
        ("Good Molecules", "goodmolecules.com", "Affordable Skincare", "Sister brand to Beautylish; effective ingredients at under $8; undercuts The Ordinary on price"),
        ("Bubble Skincare", "hellobubble.com", "Gen Z Skincare", "Specifically targets Gen Z teens; $10-16 pricing; available at Walmart; bright colorful playful branding"),
        ("Byoma", "byoma.com", "Barrier Skincare", "Barrier-repair focused skincare for Gen Z; tri-ceramide complex; recycled and recyclable packaging"),
        ("Paula's Choice", "paulaschoice.com", "Science Skincare", "Founded by beauty myth-buster Paula Begoun who wrote 'Don't Go to the Cosmetics Counter Without Me'; BHA exfoliant is cult"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Clinical Skincare", "Dermatologist-founded; at-home Alpha Beta peel pads replicate his $300 in-office treatment; LED mask pioneer"),
        ("Peace Out Skincare", "peaceoutskincare.com", "Acne Care", "Acne patches that went viral; uses hydrocolloid + salicylic acid; targets acne with 'emergency' branding"),
        ("Farmacy Beauty", "farmacybeauty.com", "Farm-to-Face", "Sources ingredients from their own farm in upstate NY; Green Clean balm became a double-cleansing staple"),
        ("Milk Makeup", "milkmakeup.com", "Vegan Makeup", "100% vegan; stick/twist-up format for everything — no brushes needed; born from Milk Studios NYC culture"),
        ("Pat McGrath Labs", "patmcgrath.com", "Luxury Makeup", "Pat McGrath is the most influential MUA alive; $1B valuation; drops sell out in minutes like streetwear"),
        ("About Face", "aboutface.com", "Makeup", "Halsey's beauty brand; bold, punk-inspired aesthetic; affordable ($15-30) despite celebrity backing"),
        ("Ami Colé", "amicole.com", "Melanin Beauty", "Designed specifically for melanin-rich skin tones; Senegalese-American founder; minimal SKU count, max impact"),
        ("Topicals", "mytopicals.com", "Skin Conditions", "Destigmatizes chronic skin conditions (eczema, hyperpigmentation); 'funner flare-ups' messaging; Gen Z cult following"),
        ("Typology", "typology.com", "French Skincare", "Paris-based; 10 ingredients or less per product; diagnostic quiz to build your routine; minimal glass packaging"),
        ("Beauty of Joseon", "beautyofjoseon.com", "K-Beauty", "Reviving Joseon Dynasty beauty recipes; Glow Serum uses ginseng root water; massive TikTok virality"),
        ("Sol de Janeiro", "soldejaneiro.com", "Brazilian Body Care", "Brazilian Bum Bum Cream has a signature scent people became obsessed with; cross-sells fragrances off body cream"),
        ("Nécessaire", "necessaire.com", "Body Care", "The Body Wash elevated body care to skincare-level ingredients; $25 body wash normalized by the brand"),
        ("Maëlys", "mymaelys.com", "Body Care", "Body sculpting focus (B-Tight, B-Flat); TikTok viral with before/after content; grew sales 460% YoY"),
        ("Phlur", "phlur.com", "Clean Fragrance", "Chrissy Teigen acquired & relaunched it; Missing Person fragrance went mega-viral on TikTok ('smells like your person')"),
        ("Boy Smells", "boysmells.com", "Candles & Fragrance", "Gender-inclusive 'genderful' branding; started candles then expanded to fragrance; LGBTQ+ founded"),
        ("Dedcool", "dedcool.com", "Unisex Fragrance", "Layerable fragrance system; biodegradale refills; 'the first sustainable fragrance house'"),
        ("Olaplex", "olaplex.com", "Hair Bond Repair", "Patented bis-aminopropyl diglycol dimaleate molecule repairs broken hair bonds; used by every salon; IPO'd at $15B"),
        ("Function of Beauty", "functionofbeauty.com", "Custom Haircare", "Personalized shampoo/conditioner based on quiz; your name printed on bottle; pioneered DTC customization"),
        ("Prose", "prose.com", "Custom Haircare", "AI-powered custom formulas with 85+ ingredients; adapts to weather, water hardness, lifestyle; made-to-order"),
        ("Vegamour", "vegamour.com", "Hair Growth", "Plant-based hair growth serums; GRO serum uses red clover and mung bean; alternative to Minoxidil without side effects"),
        ("K18 Hair", "k18hair.com", "Hair Repair", "Patented bioactive peptide that reverses hair damage in 4 minutes; disrupted the salon repair treatment market"),
        ("Bread Beauty Supply", "breadbeautysupply.com", "Textured Hair", "Simplified textured hair care to just 5 products; wash-day made easier; Sephora exclusive launch"),
        ("Dae Hair", "daehair.com", "Desert Haircare", "Sonoran Desert-inspired ingredients like prickly pear and cactus; Amber Fillerup Clark's influencer brand"),
        ("Crown Affair", "crownaffair.com", "Luxury Haircare", "Elevated daily haircare rituals; The Comb costs $38; Dianna Cohen built it as a 'slow beauty' brand"),
        ("Curlsmith", "curlsmith.com", "Curly Hair", "Bond repair specifically for curly/coily hair types; kitchen-inspired ingredient names; curl type system"),
        ("Nutrafol", "nutrafol.com", "Hair Supplements", "Doctor-developed hair growth supplements targeting root causes (stress, hormones); $200M+ revenue; clinical studies"),
        ("Spoiled Child", "spoiledchild.com", "AI Skincare", "AI-driven personalized anti-aging; aggressive Facebook/Meta ad spend with UGC-style before/afters"),
        ("Primally Pure", "primallypure.com", "Natural Skincare", "Farm-to-skin deodorant that actually works; started on a family farm; non-toxic living community"),
        ("100% Pure", "100percentpure.com", "Natural Cosmetics", "Fruit-pigmented makeup (uses fruit dyes instead of synthetic); every product 100% pure and natural"),
        ("Fable & Mane", "fableandmane.com", "Ayurvedic Hair", "Indian Ayurvedic hair oils modernized for Western market; sibling-founded; 'hair oiling ritual' storytelling"),
        ("Starface", "starface.world", "Acne Patches", "Made pimple patches fun with yellow star shapes; destigmatized acne for Gen Z; playful packaging"),
        ("Hero Cosmetics", "herocosmetics.com", "Acne Care", "Mighty Patch became the #1 acne patch on Amazon; simple hydrocolloid tech marketed brilliantly; sold to Church & Dwight for $630M"),
        ("e.l.f. Cosmetics", "elfcosmetics.com", "Affordable Beauty", "$5 average price point; TikTok-first marketing strategy; 'Eyes Lips Face' campaign got 7B+ views; fastest growing beauty brand"),
        ("ColourPop", "colourpop.com", "Affordable Beauty", "LA-based; launches new products weekly; $5-15 price point; owns their manufacturing (Seed Beauty factory)"),
        ("Morphe", "morphebrushes.com", "Makeup Brushes", "Became massive through YouTuber collabs (Jaclyn Hill, James Charles); 35-pan palettes became iconic"),
        ("Juvia's Place", "jfruviasplace.com", "Inclusive Beauty", "Nigerian-inspired luxury pigments at affordable prices; founder Chichi Eburu; rich cultural branding"),
        ("BK Beauty", "bfrk-beauty.com", "Makeup Brushes", "Lisa J's brush brand blew up on TikTok with technique tutorials; affordable pro-quality brushes"),
        ("Danessa Myricks", "danessamyricksbeauty.com", "Pro Makeup", "MUA-founded multi-use products; Vision Flush works on eyes, lips, and cheeks; artistry-focused"),
        ("Rose Inc", "roseinc.com", "Clean Luxury Beauty", "Rosie Huntington-Whiteley's brand; blush and lip combo in clean formulas; model-founder credibility"),
        ("Live Tinted", "livetinted.com", "Multicultural Beauty", "Deepica Mutyala's brand born from a viral video about color-correcting dark circles with red lipstick"),
        ("Aavrani", "aavrani.com", "Ayurvedic Skincare", "Indian beauty rituals (turmeric, saffron) in modern formats; South Asian representation in clean beauty"),
        ("Bushbalm", "bushbalm.com", "Ingrown Hair Care", "Shark Tank success; specialized in bikini line skincare that nobody else was addressing; grew 10x post-show"),
        ("Stratia", "stratiaskin.com", "Barrier Repair", "Liquid Gold moisturizer contains golden ratio of ceramides, cholesterol, fatty acids to repair skin barrier"),
        ("RoseSkinCo", "roseskinco.com", "IPL Hair Removal", "At-home IPL device at fraction of salon cost; massive influencer marketing spend; subscription model for cartridges"),
        ("Kinship", "lovekinship.com", "Clean Skincare", "Biotech-meets-clean beauty; uses 'kinbiome' (probiotic-derived postbiotic); EWG verified"),
        ("Dieux Skin", "dieuxskin.com", "Transparent Skincare", "Reusable silicone eye masks (replacing single-use); shows full cost breakdown; TikTok-native brand"),
        ("Summer Fridays", "summerfridays.com", "Masks & Skincare", "Jet Lag Mask became the most Instagrammed skincare product; founders had 2M+ combined followers at launch"),
        ("Augustinus Bader", "augustinusbader.com", "Luxury Skincare", "German stem cell scientist's TFC8 technology; $265 cream endorsed by every celebrity; Hailey Bieber's 'other' favorite"),
        ("Drunk Elephant", "drunkelephant.com", "Biocompatible", "Tiffany Masterson's 'suspicious 6 free' philosophy; the brand that made people throw away everything else"),
        ("Soko Glam", "sokoglam.com", "K-Beauty Retail", "Charlotte Cho's K-beauty marketplace that popularized the 10-step Korean skincare routine in the US"),
        ("Iris & Romeo", "irisandromeo.com", "Over 40 Beauty", "Specifically created for 40+ skin; Best Skin Days SPF is a 5-in-1 hybrid; founder is ex-Bare Minerals CEO"),
        ("MENTED Cosmetics", "mentedcosmetics.com", "WOC Beauty", "Founded by two Harvard MBAs; nude lipsticks specifically designed for women of color; 'every woman deserves a nude lip'"),
    ],

    # =========================================================================
    # MEN'S GROOMING
    # =========================================================================
    "Men's Grooming": [
        ("Dollar Shave Club", "dollarshaveclub.com", "Shaving", "Viral launch video got 12,000 orders in 48 hours; sold to Unilever for $1B; subscription model pioneer"),
        ("Harry's", "harrys.com", "Shaving", "Bought a German razor factory to control supply chain; $1.37B attempted acquisition by Edgewell; D2C to retail expansion"),
        ("Dr. Squatch", "drsquatch.com", "Natural Soap", "Viral YouTube ads with comedic 'you're not a dish' campaign; $100M+ revenue; natural soap for men who don't buy skincare"),
        ("Beardbrand", "beardbrand.com", "Beard Care", "Eric Bandholz grew it from YouTube beard community; Urban Beardsman lifestyle brand; content-first approach"),
        ("Manscaped", "manscaped.com", "Body Grooming", "Created the 'below the belt' grooming category; Lawn Mower trimmer; podcast ad king; $1B+ valuation"),
        ("Scotch Porter", "scotchporter.com", "Grooming", "Black-owned men's grooming; founder Calvin Quallis started from his apartment; Walmart & Target distribution"),
        ("Hims", "forhims.com", "Men's Health", "Telehealth for hair loss, ED, skincare; destigmatized taboo men's health topics; $1B+ market cap public company"),
        ("Keeps", "keeps.com", "Hair Loss", "FDA-approved hair loss treatments online without awkward doctor visits; subscription finasteride & minoxidil"),
        ("Lumin", "luminskin.com", "Men's Skincare", "K-beauty inspired men's skincare; charcoal cleanser bestseller; targeted FB ads with 'your girlfriend's routine' angle"),
        ("Supply", "supply.co", "Single Edge Razor", "Injector-style single-blade razor to prevent irritation; lifetime warranty; 'the last razor you'll buy'"),
        ("Huron", "usehuron.com", "Men's Skincare", "Matt Mullenax (ex-Bonobos) built affordable men's skincare; clean, simple, $15 price point"),
        ("Disco", "letsdisco.co", "Men's Skincare", "Disco ball branding that stands out; target: millennial men who've never had a skincare routine; playful messaging"),
        ("Hawthorne", "hawthorne.co", "Men's Personal Care", "Personalized men's care via quiz; cologne, deodorant, body wash matched to your preferences; AI-driven recs"),
        ("Bevel", "bevel.com", "Men's Grooming", "Tristan Walker built specifically for Black men's shaving needs; single-blade razor prevents razor bumps; sold to P&G"),
        ("Oars + Alps", "oarsandalps.com", "Men's Skincare", "Clean men's skincare started by two women who were frustrated shopping for their husbands; S.C. Johnson acquired"),
        ("Every Man Jack", "everymanjack.com", "Natural Grooming", "Naturally derived men's care at mass retail prices ($7-10); one of the first 'clean' men's brands"),
        ("Black Wolf", "blackwolf.com", "Men's Skincare", "Charcoal-powered men's grooming; aggressive social media marketing; 'Wolf Pack' community building"),
        ("Bravo Sierra", "bravosierra.com", "Military-Grade Care", "Developed with active-duty military; every purchase donates to troops; extreme performance testing"),
        ("Art of Sport", "artofsport.com", "Athletic Grooming", "Co-founded by Kobe Bryant; designed for athletes; botanicals that work during sweat and activity"),
        ("Geologie", "geologie.com", "Personalized Skincare", "Diagnostic quiz creates custom 4-step regimen; dermatologist-designed; personalized label with your name"),
        ("Cremo", "cremo.com", "Shaving", "Astonishingly superior shave cream in a tube; unique concentrated formula — pea-sized amount works; $5 price"),
        ("Duke Cannon", "dukecannon.com", "Rugged Grooming", "Military-inspired branding; Big Ass Brick of Soap is their hero product; donates to veterans; humor-led marketing"),
        ("Baxter of California", "baxterofcalifornia.com", "Premium Grooming", "Founded in 1965 — one of the original men's grooming brands; clay pomade is iconic; barbershop heritage"),
        ("Aesop", "aesop.com", "Luxury Skincare", "Australian apothecary aesthetic; stores are architectural landmarks; $2.5B acquisition by L'Oréal"),
        ("Jack Black", "getjackblack.com", "Men's Skincare", "Double-Duty Face Moisturizer SPF is a men's cult classic; dermatologist-tested; clean ingredients at fair price"),
    ],

    # =========================================================================
    # HEALTH & WELLNESS
    # =========================================================================
    "Health & Wellness": [
        ("AG1 Athletic Greens", "drinkag1.com", "Greens Supplement", "75 vitamins/minerals in one scoop; sponsors every major podcast; 'first thing I take every morning' UGC strategy; Tim Ferriss early backer"),
        ("Ritual", "ritual.com", "Vitamins", "Transparent capsule shows beadlets inside; 'the visible supply chain' — traces every ingredient to source; $250M+ revenue"),
        ("Seed", "seed.com", "Probiotics", "DS-01 Daily Synbiotic backed by 24+ strain-specific clinical studies; eco-refill packaging; science-first branding"),
        ("Care/of", "takecareof.com", "Personalized Vitamins", "Quiz-based personalized vitamin packs with your name printed on each; Bayer acquired for reported $250M+"),
        ("OLLY", "olly.com", "Gummy Vitamins", "Made vitamins fun and accessible in gummy format; colorful packaging; sold to Unilever; targets wellness-curious consumer"),
        ("Liquid IV", "liquid-iv.com", "Hydration", "Cellular Transport Technology (CTT) for rapid hydration; massive sampling at events/festivals; Unilever acquired"),
        ("Bloom Nutrition", "bloomnu.com", "Greens & Wellness", "Mari Llewellyn's fitness transformation story drives brand; greens powder became TikTok's #1 wellness product"),
        ("Vital Proteins", "vitalproteins.com", "Collagen", "Jennifer Aniston partnership made collagen mainstream; blue tub is instantly recognizable; Nestlé acquired"),
        ("Moon Juice", "moonjuice.com", "Adaptogens", "Amanda Chantal Bacon pioneered adaptogens in mainstream wellness; Dust line (Brain Dust, Beauty Dust, Sex Dust)"),
        ("Four Sigmatic", "foursigmatic.com", "Mushroom Coffee", "Finnish founder brought functional mushrooms (lion's mane, chaga, reishi) to coffee; 'coffee that does more'"),
        ("MUD\\WTR", "mudwtr.com", "Coffee Alternative", "Anti-coffee positioning: '1/7th the caffeine'; masala chai + mushrooms + cacao; Shane Heath's morning ritual brand"),
        ("Eight Sleep", "eightsleep.com", "Sleep Tech", "Pod cover that heats/cools your bed; used by 200+ pro athletes; 'the Tesla of sleep'; $500M+ valuation"),
        ("Whoop", "whoop.com", "Fitness Tracker", "No screen, subscription-based ($30/mo); strain/recovery/sleep scores; worn by elite athletes and military; $3.6B valuation"),
        ("Oura Ring", "ouraring.com", "Health Tracker", "Discrete ring form factor; sleep tracking gold standard; NBA used it for COVID detection; $2.55B valuation"),
        ("Casper", "casper.com", "Mattress", "The OG bed-in-a-box DTC mattress; 100-night trial pioneered risk-free buying; IPO'd but struggled post-hype"),
        ("Purple", "purple.com", "Mattress", "Hyper-Elastic Polymer grid is unique tech — the raw egg test video got 1B+ views; science/humor marketing"),
        ("Helix Sleep", "helixsleep.com", "Custom Mattress", "Sleep quiz matches you to ideal mattress from 12+ models; personalization angle in commodity category"),
        ("Saatva", "saatva.com", "Luxury Mattress", "Premium DTC mattress with white-glove delivery + old mattress removal; targets affluent buyers; $500M+ revenue"),
        ("Pillow Cube", "pillowcube.com", "Pillows", "Square pillow designed specifically for side sleepers; viral TikTok ads; solves a specific pain point nobody addressed"),
        ("Thesis", "takethesis.com", "Custom Nootropics", "Personalized nootropic blends based on quiz; try different formulas and find your stack; starter kit approach"),
        ("Onnit", "onnit.com", "Supplements & Fitness", "Alpha Brain nootropic used by Joe Rogan (who invested); kettlebells, supplements, fitness; Unilever acquired"),
        ("Elysium Health", "elysiumhealth.com", "Longevity", "Nobel Prize-winning scientist on board; Basis supplement with NAD+ precursor; premium longevity science"),
        ("InsideTracker", "insidetracker.com", "Blood Testing", "Analyzes 43+ blood biomarkers; AI gives personalized nutrition/lifestyle recs; Tom Brady used it"),
        ("Beekeeper's Naturals", "beekeepersnaturals.com", "Bee Products", "Propolis throat spray became a wellness staple; Carly Stein built it after discovering bee products cured her illness"),
        ("Goli Nutrition", "goli.com", "ACV Gummies", "Made apple cider vinegar palatable in gummy form; first ACV gummy; $300M+ revenue; massive influencer spend"),
        ("Lemme", "lemmelive.com", "Wellness Gummies", "Kourtney Kardashian's supplement brand; gummies for gut, focus, sleep, libido; leverages Kardashian marketing machine"),
        ("Sakara Life", "sakara.com", "Plant-Based Meals", "Organic plant-based meal delivery at $70+/day; targets wealthy wellness enthusiasts; 'eat pretty' philosophy"),
        ("Daily Harvest", "dailyharvest.com", "Frozen Meals", "Pre-portioned frozen smoothies, bowls, flatbreads; built on convenience + health; $1.1B valuation before recall setback"),
        ("Hungryroot", "hungryroot.com", "Healthy Groceries", "AI-powered personalized grocery service; quiz picks recipes + sends ingredients; health meets convenience"),
        ("Factor", "factor75.com", "Prepared Meals", "Chef-prepared, dietitian-designed meals; keto, calorie smart, vegan options; HelloFresh acquired; fastest growing meal delivery"),
        ("Noom", "noom.com", "Behavior Change", "Psychology-based weight loss app; uses CBT (cognitive behavioral therapy); $400M+ revenue; anti-diet diet"),
        ("Athletic Greens", "athleticgreens.com", "Greens Supplement", "Spent 10 years perfecting formula before scaling; 75 ingredients in one scoop; $1.2B valuation"),
        ("Momentous", "livemomentous.com", "Sports Supplements", "Only supplement brand recommended by Huberman Lab podcast; NSF certified; used by pro sports teams"),
        ("Transparent Labs", "transparentlabs.com", "Clean Supplements", "100% formula transparency — every dose listed, no proprietary blends; 3rd party tested; science-forward"),
        ("Gorilla Mind", "gorillamind.com", "Nootropics", "Derek from More Plates More Dates (MPMD) YouTube channel; built audience then monetized with supps; 3M+ subscribers"),
        ("Thorne", "thorne.com", "Clinical Supplements", "Used by Mayo Clinic and 100+ pro sports teams; highest clinical grade supplements; health testing platform"),
        ("Garden of Life", "gardenoflife.com", "Organic Supplements", "USDA Organic and Non-GMO Project Verified supplements; whole food based; Nestlé Health Science acquired"),
        ("Sports Research", "sportsresearch.com", "Supplements", "Sweet Sweat waist trimmer became viral; collagen peptides bestseller on Amazon; SoCal lifestyle branding"),
        ("Jocko Fuel", "jockofuel.com", "Supplements", "Jocko Willink (Navy SEAL, author) brand; 'discipline equals freedom' ethos; authentic military credibility"),
        ("Timeline Nutrition", "timelinenutrition.com", "Longevity", "Mitopure (urolithin A) is first supplement to target mitochondrial health; Swiss biotech origin; $100M+ raised"),
        ("Persona Nutrition", "personanutrition.com", "Custom Vitamins", "Personalized daily vitamin packs; free nutritionist consultations; assessment considers 5,000+ scientific studies"),
        ("Gainful", "gainful.com", "Custom Protein", "Personalized protein powder — quiz determines your blend; plain base + flavor stick packets; unique customization"),
        ("HUM Nutrition", "humnutrition.com", "Beauty Vitamins", "Beauty-focused supplements (hair, skin, nails, body); RD on staff for free consultations; Sephora distribution"),
        ("Wellabs", "wellabs.com", "Liquid Supplements", "Liquid vitamin format for faster absorption; subscription model; vibrant packaging"),
        ("Nuzest", "nuzest.com", "Plant Protein", "Clean Lean Protein from European golden peas; allergen-free; simple ingredient list"),
        ("KOS", "kos.com", "Plant Protein", "Organic plant protein with superfood blends; colorful branding; $20 price point disrupts premium market"),
        ("Orgain", "orgain.com", "Organic Protein", "Founded by cancer survivor Dr. Andrew Abraham; organic, clean protein; Walmart & Costco distribution"),
    ],

    # =========================================================================
    # FITNESS & ACTIVEWEAR
    # =========================================================================
    "Fitness & Activewear": [
        ("Gymshark", "gymshark.com", "Gym Apparel", "19-year-old Ben Francis started screen-printing in his garage; grew through YouTube fitness influencers; $1.45B valuation"),
        ("Alo Yoga", "aloyoga.com", "Yoga Apparel", "Celebrity yoga studio culture (Alo Yoga Sanctuary); Kendall Jenner, Hailey Bieber ambassadors; $10B valuation reported"),
        ("Vuori", "vuori.com", "Performance Apparel", "Encinitas, CA surf culture meets performance; SoftBank invested $400M at $4B valuation; men's activewear gap-filler"),
        ("Outdoor Voices", "outdoorvoices.com", "Recreational Apparel", "'Doing things' philosophy — exercise for fun, not competition; Tyler Haney built community-first brand; Exercise Dress iconic"),
        ("Rhone", "rhone.com", "Men's Activewear", "GoldFusion anti-odor tech; targets 'business athlete'; office-to-gym versatility; premium men's activewear"),
        ("Ten Thousand", "tenthousand.cc", "Men's Training", "Interval shorts with liner became cult; designed by consulting with military operators and pro athletes"),
        ("ASRV", "asrv.com", "Aesthetic Sportswear", "Technical athletic wear with fashion aesthetic; limited drops sell out; targets aesthetic bodybuilding community"),
        ("Alphalete", "alphaleteathletics.com", "Gym Apparel", "Christian Guzman's fitness empire; YouTube vlog-to-brand pipeline; gym + lifestyle + real estate"),
        ("Buffbunny Collection", "buffbunny.com", "Women's Activewear", "Heidi Somers' brand; drops sell out in minutes; built entirely on Instagram fitness community"),
        ("Fabletics", "fabletics.com", "Activewear Subscription", "Kate Hudson co-founded; VIP membership model ($50/month credit); disrupted with affordable + celeb angle"),
        ("Girlfriend Collective", "girlfriend.com", "Sustainable Activewear", "Made from recycled water bottles + fishing nets; launched with free leggings (just pay shipping) — 10K waitlist day 1"),
        ("Nobull", "nobullproject.com", "CrossFit Shoes", "Official shoe of CrossFit Games; SuperFabric upper is indestructible; anti-flashy, 'no bull' branding"),
        ("Born Primitive", "bornprimitive.com", "CrossFit Apparel", "Military veteran-founded; functional fitness community; partners with CrossFit athletes; patriotic branding"),
        ("Cuts Clothing", "cutsclothing.com", "Men's Tees", "Curve-Hem crew neck became the 'going out tee'; PYCA fabric feels like butter; NFL/NBA player favorite"),
        ("Tracksmith", "tracksmith.com", "Running Apparel", "Boston-based running brand with old-school track aesthetic; Twilight 5K race series; no celebrity athletes, real runners"),
        ("Satisfy Running", "satisfyrunning.com", "Running Apparel", "French luxury running brand; $200+ shorts with high-end fabrication; 'running is a counterculture' ethos"),
        ("Peloton", "onepeloton.com", "Connected Fitness", "$44B peak valuation; created community around at-home cycling; instructor-as-celebrity model; pandemic darling"),
        ("Tonal", "tonal.com", "Smart Home Gym", "AI-powered strength training machine; electromagnetic resistance; 200+ lb capacity on wall-mounted system"),
        ("Hyperice", "hyperice.com", "Recovery Tech", "Hypervolt percussion gun used by every pro team; NBA/NFL/MLB official partner; recovery as performance"),
        ("Therabody", "therabody.com", "Recovery Tech", "Created the percussion therapy category with Theragun; $170M revenue; Dr. Jason Wersland invented after motorcycle accident"),
        ("Bala", "shopbala.com", "Weighted Accessories", "Fashionable 1-2 lb wrist/ankle weights; Shark Tank success; Mark Cuban investment; made fitness accessories beautiful"),
        ("Crossrope", "crossrope.com", "Jump Ropes", "Weighted interchangeable jump ropes + app; turned jump rope into premium fitness category; $30M+ revenue"),
        ("Rogue Fitness", "roguefitness.com", "Gym Equipment", "Official equipment of CrossFit Games; made in Columbus, Ohio; $500M+ revenue; American manufacturing pride"),
        ("Lululemon", "lululemon.com", "Athletic Apparel", "Created athleisure category; Align leggings have cult following; $9.6B revenue; community ambassador model"),
        ("On Running", "on.com", "Swiss Running Shoes", "CloudTec sole technology; Swiss engineering; Roger Federer investor; went from niche to $2B+ revenue at IPO"),
        ("HOKA", "hoka.com", "Running Shoes", "Maximalist cushioning trend-setter; originally for ultrarunners; crossed over to fashion/lifestyle; $1.8B revenue"),
        ("Whoop", "whoop.com", "Fitness Wearable", "No screen, subscription model ($30/mo); strain/recovery scores; worn by Patrick Mahomes, Michael Phelps"),
        ("Mirror", "mirror.co", "Home Fitness", "Full-length mirror that streams workout classes; Lululemon acquired for $500M; pivoted to Lulu Studio"),
        ("Hydrow", "hydrow.com", "Connected Rowing", "Outdoor Reality rowing — filmed on actual water; electromagnetic drag; quieter than Concept2; premium positioning"),
        ("TRX", "trxtraining.com", "Suspension Training", "Navy SEAL Randy Hetrick invented it using jiu-jitsu belts and parachute webbing; used by every military branch"),
        ("Gorilla Bow", "gorillabow.com", "Resistance Training", "Portable bow-shaped resistance band system; Shark Tank funded; travel-friendly full-body workout"),
        ("FORM Swim", "formswim.com", "Smart Swim Goggles", "AR goggles that display pace, stroke rate, distance while swimming; first smart goggles; used by Olympic swimmers"),
    ],

    # =========================================================================
    # FASHION & APPAREL
    # =========================================================================
    "Fashion & Apparel": [
        ("Everlane", "everlane.com", "Transparent Fashion", "'Radical Transparency' — shows true cost of every product (materials, labor, transport, markup); ethical factory program"),
        ("Reformation", "thereformation.com", "Sustainable Fashion", "Tracks environmental footprint of every garment (CO2, water, waste); LA-made; celebrity-loved sustainable fashion"),
        ("SKIMS", "skims.com", "Shapewear & Basics", "Kim Kardashian's $4B brand; redefined shapewear as everyday essentials; Fits Everybody collection; expanded to menswear"),
        ("Good American", "goodamerican.com", "Inclusive Denim", "Khloé Kardashian co-founded; launched with sizes 00-24 day one; $1M in sales first day; body-positive denim"),
        ("Allbirds", "allbirds.com", "Sustainable Footwear", "Merino wool + sugarcane soles; carbon footprint on every label; Silicon Valley uniform; co-created with adidas"),
        ("Rothy's", "rothys.com", "Recycled Shoes", "Shoes knit from recycled ocean-bound plastic bottles; machine washable; 150M+ bottles recycled; $1B+ valuation"),
        ("Thursday Boots", "thursdayboots.com", "Boots", "Premium boots at 1/3 the price of competitors; Kickstarter origin; 'honest prices for honest boots'; huge Reddit following"),
        ("Tecovas", "tecovas.com", "Western Boots", "DTC cowboy boots at half the retail price; handmade in León, Mexico; fastest-growing Western brand; $200M+ revenue"),
        ("Warby Parker", "warbyparker.com", "Eyewear", "Home Try-On program (5 frames free); Buy a Pair Give a Pair program; IPO'd; disrupted Luxottica's monopoly"),
        ("CARIUMA", "cariuma.com", "Sustainable Sneakers", "Plants 2 trees for every pair sold (10M+ trees); Brazilian-founded; organic cotton, bamboo, recycled materials"),
        ("Oliver Cabell", "olivercabell.com", "Luxury Sneakers", "Shows exact cost breakdown of $140 sneaker vs $550 'luxury' competitor; radical price transparency in luxury"),
        ("American Giant", "american-giant.com", "Made in USA Basics", "The 'greatest hoodie ever made' — Slate magazine review crashed their site; 100% American-made; premium basics"),
        ("Buck Mason", "buckmason.com", "Men's Essentials", "Venice Beach brand; Curved Hem Tee is their hero; 'less, better' philosophy; one of each essential, perfected"),
        ("Marine Layer", "marinelayer.com", "Soft Tees", "Re-Spun program — mail in old tees, they recycle them into new fabric; unbelievably soft shirts; 'absurdly soft' tagline"),
        ("Faherty", "faherty.com", "Coastal Style", "Twin brothers Alex & Mike Faherty; sun-washed, surf-inspired; Faherty Sunwashed fabric is proprietary; Hamptons meets SoCal"),
        ("Bonobos", "bonobos.com", "Men's Pants", "Pioneered Guideshops (try on, order online); better-fitting men's pants was founding insight; Walmart acquired for $310M"),
        ("MeUndies", "meundies.com", "Underwear", "MicroModal fabric (3x softer than cotton); matching couples sets; fun prints (avocados, dinosaurs); membership model"),
        ("Tommy John", "tommyjohn.com", "Underwear", "Patented horizontal fly and no-roll waistband; 'no adjustment needed'; Shark Tank alumni; $100M+ revenue"),
        ("Parade", "yourparade.com", "Sustainable Underwear", "Recycled nylon underwear in bold colors; Gen Z branding; inclusive sizing 2XS-3XL; $10 price point; viral campaigns"),
        ("True Classic", "trueclassictees.com", "Men's Tees", "Facebook ad machine — spent $50M+ on Meta ads; targets 'dad bods' with flattering fit; $250M+ revenue in 3 years"),
        ("Untuckit", "untuckit.com", "Casual Shirts", "Solved one problem: shirts designed to be worn untucked at the perfect length; 90+ stores from one simple insight"),
        ("Ministry of Supply", "ministryofsupply.com", "Performance Workwear", "NASA-engineered fabrics in dress clothes; phase-change material regulates body temp; 3D-printed knit blazer"),
        ("Mizzen+Main", "mizzenandmain.com", "Performance Dress Shirts", "Moisture-wicking, stretch dress shirts; 'dress shirts that perform'; Phil Mickelson early ambassador; performance meets dress code"),
        ("Taylor Stitch", "taylorstitch.com", "Sustainable Menswear", "Crowdfunded every product — only makes what's pre-ordered; reduces waste to near zero; responsible materials"),
        ("Indochino", "indochino.com", "Custom Suits", "Made-to-measure suits starting at $399; 100+ showrooms; 3D body scanning; democratized custom tailoring"),
        ("Cuyana", "cuyana.com", "Fewer Better Things", "'Fewer, better things' philosophy; lean closet capsule approach; premium leather goods at fair prices"),
        ("M.M.LaFleur", "mmlafleur.com", "Women's Workwear", "Bento Box — stylist picks 4 pieces for you to try; professional women's wardrobing; easy return process"),
        ("Senreve", "senreve.com", "Luxury Handbags", "Maestra bag converts from tote to backpack to satchel; Italian leather; founded by two Stanford MBA women"),
        ("Béis", "beistravel.com", "Travel Bags", "Shay Mitchell's travel brand; functional, affordable luxury luggage; TikTok packing videos go viral consistently"),
        ("Away", "awaytravel.com", "Modern Luggage", "Built-in battery charger in carry-on; launched with a book of travel stories; $1.4B valuation; podcast The Everywhere"),
        ("Madhappy", "madhappy.com", "Mental Health Fashion", "Streetwear brand built around mental health advocacy; The Local Optimist blog; celebrity-loved hoodies; popup model"),
        ("Aimé Leon Dore", "aimeleondore.com", "Streetwear", "Queens, NY brand; New Balance collabs sell out instantly; Mediterranean cafe in store; 'neighborhood luxury' aesthetic"),
        ("Kith", "kith.com", "Streetwear", "Ronnie Fieg built from sneaker roots; Kith Treats (cereal bar in store); Nike, BMW, Versace collabs; retail-experience pioneer"),
        ("Fashion Nova", "fashionnova.com", "Trend Fashion", "Cardi B partnership; fastest fashion — design to site in 1-2 weeks; Instagram growth hack with micro-influencers"),
        ("SHEIN", "shein.com", "Ultra-Fast Fashion", "AI-driven trend detection; produces in micro-batches; $30B+ revenue; most downloaded shopping app globally"),
        ("Savage X Fenty", "savagex.com", "Lingerie", "Rihanna's inclusive lingerie; up to size 4XL; Savage X Fenty Show became cultural event; VIP membership"),
        ("ThirdLove", "thirdlove.com", "Bras", "Fit Finder quiz uses ML to recommend size; invented half-cup sizes; directly challenged Victoria's Secret in open letter ad"),
        ("Knix", "knix.com", "Wireless Bras", "Leakproof underwear + wireless bras; Kickstarter origin; body-positive campaigns with real women; $100M+ revenue"),
        ("Sézane", "sezane.com", "French Fashion", "First French fashion brand born online; monthly 'La Liste' drops; Parisian cool-girl aesthetic; waitlists create scarcity"),
        ("Princess Polly", "princesspolly.com", "Fast Fashion", "Australian Gen Z fashion; TikTok haul queen; $200M+ revenue; fast trend-cycle mastery; acquired by A.K.A. Brands"),
        ("AGOLDE", "agolde.com", "Denim", "90s-inspired straight-leg jeans at $198; became the editor/influencer jean of choice; organic cotton, LA-made"),
        ("MOTHER Denim", "motherdenim.com", "Premium Denim", "The Hustler ankle fray became the 'it' jean; rock & roll meets California; $200+ price point; vintage washes"),
        ("DL1961", "dl1961.com", "Sustainable Denim", "Uses 50% less water, 50% less energy in production; OEKO-TEX certified; family-owned denim expertise"),
        ("Quince", "onequince.com", "Affordable Luxury", "Cashmere sweaters at $50 (vs $300 elsewhere); Mongolian cashmere, Italian leather direct from factory; radical value"),
        ("Italic", "italic.com", "Unbranded Luxury", "Sells products from same factories as Prada, Burberry — without the label; membership model; radical unbrand"),
        ("Aritzia", "aritzia.com", "Canadian Fashion", "Super Puff jacket became viral; curates house brands (Babaton, TNA, Wilfred); $2B+ revenue; Gen Z + Millennial"),
        ("COS", "cosstores.com", "Modern Essentials", "H&M's elevated brand; Scandinavian minimalism; architectural store design; art/culture collabs; quality basics"),
        ("Hill House Home", "hillhousehome.com", "Nap Dress", "The Nap Dress — meant for sleeping but became a street style phenomenon; drops sell out; cottagecore aesthetic"),
        ("Lunya", "lunya.co", "Sleepwear", "Elevated sleepwear category; Washable Silk set; made sleeping-in stylish; DTC premium PJs for women"),
        ("Bombas", "bombas.com", "Comfort Socks", "Donates one item for every item sold (50M+ donated); Shark Tank's biggest deal ever; socks with purpose; $300M+ revenue"),
        ("Stance", "stance.com", "Performance Socks", "Official sock of NBA, MLB; Punks & Poets brand ethos; artist collabs on socks; turned socks into self-expression"),
        ("Darn Tough", "darntough.com", "Lifetime Socks", "Unconditional lifetime warranty — any reason, any time; made in Vermont; merino wool; guaranteed for life"),
    ],

    # =========================================================================
    # FOOD & BEVERAGE
    # =========================================================================
    "Food & Beverage": [
        ("OLIPOP", "drinkolipop.com", "Prebiotic Soda", "Healthy soda alternative with prebiotics + botanicals; $200M+ revenue; 2,128% growth over 3 years; 'what soda should be'"),
        ("Poppi", "drinkpoppi.com", "Prebiotic Soda", "Apple cider vinegar soda; Shark Tank to PepsiCo acquisition for ~$2B; Allison Ellsworth's comeback story"),
        ("Liquid Death", "liquiddeath.com", "Canned Water", "Water in tallboy cans with heavy metal branding; 'Murder Your Thirst'; $1.4B valuation; merch outsells water"),
        ("Prime", "drinkprime.com", "Energy Drink", "Logan Paul x KSI collab; became a schoolyard currency; $1.2B retail sales year 1; social media virality engine"),
        ("Feastables", "feastables.com", "Chocolate & Snacks", "MrBeast's brand; $250M revenue; tied to YouTube giveaways; 'Willy Wonka of YouTube' golden ticket campaigns"),
        ("Magic Spoon", "magicspoon.com", "Protein Cereal", "Childhood cereal taste with 13g protein, 0g sugar; nostalgic branding; $100M+ revenue; grocery expansion from DTC"),
        ("Graza", "graza.co", "Olive Oil", "Squeeze bottle format disrupted olive oil; 'Drizzle' for finishing, 'Sizzle' for cooking; sold out 5x at launch"),
        ("Brightland", "brightland.co", "Olive Oil", "Artist-designed labels; California EVOO; influencer darling; turned olive oil into a giftable lifestyle product"),
        ("Fly By Jing", "flybyjing.com", "Sichuan Sauce", "Jing Gao brought authentic Sichuan chili crisp to US; 'not for the faint of heart'; fastest-selling sauce on Amazon"),
        ("TRUFF", "truff.com", "Truffle Hot Sauce", "Luxury hot sauce with black truffle; Instagram-first launch; $100M+ revenue; elevated hot sauce to premium gift category"),
        ("Siete Foods", "sietefoods.com", "Mexican-Inspired", "Mexican-American family brand; grain-free tortillas; PepsiCo acquired for $1.2B; seven family members = 'Siete'"),
        ("Mid-Day Squares", "middaysquares.com", "Functional Chocolate", "Couple founders vlog the entire startup journey transparently; chocolate bars with 12g protein; content-first brand"),
        ("Chomps", "chomps.com", "Meat Sticks", "Whole30 approved beef sticks; grass-fed; $200M+ revenue; bootstrapped to massive scale without outside funding"),
        ("Death Wish Coffee", "deathwishcoffee.com", "Strong Coffee", "World's strongest coffee (2x caffeine); won Super Bowl ad contest (free ad worth $5M); skull & crossbones branding"),
        ("Trade Coffee", "drinktrade.com", "Coffee Subscription", "Matches you with coffee from 55+ roasters via taste quiz; personalized subscription; supports small roasters"),
        ("Cometeer", "cometeer.com", "Flash-Frozen Coffee", "Flash-freezes specialty coffee into capsules; just add water/milk; no machine needed; space-age tech for coffee"),
        ("Athletic Brewing", "athleticbrewing.com", "Non-Alcoholic Beer", "The NA beer that actually tastes good; $100M+ revenue; fastest-growing beer brand (including alcoholic); Run Wild IPA"),
        ("Ghia", "drinkghia.com", "Non-Alcoholic Aperitif", "Mediterranean-inspired NA aperitif; beautiful glass bottle becomes decor; targets sober-curious millennials; Le Spritz"),
        ("Seedlip", "seedlipdrinks.com", "Non-Alcoholic Spirits", "World's first non-alcoholic spirit; Diageo invested; $30/bottle premium positioning; changed the NA category"),
        ("LMNT", "drinklmnt.com", "Electrolytes", "No sugar, no junk electrolytes; Robb Wolf (paleo pioneer) co-founded; massive podcast sponsorship; 1,000mg sodium per pack"),
        ("Nuun", "nuun.com", "Electrolyte Tabs", "Effervescent electrolyte tabs you drop in water; runner/cyclist favorite; portable tube format; low calorie"),
        ("Cure Hydration", "curehydration.com", "Electrolyte Mix", "Coconut water-based ORS (oral rehydration solution); plant-based electrolytes; Lauren Picasso Founder"),
        ("Alani Nu", "alaninu.com", "Energy & Fitness", "Katy Hearn (fitness influencer) brand; energy drinks + supplements; $200M+ revenue; primarily female audience"),
        ("Celsius", "celsius.com", "Fitness Drink", "Clinically proven to boost metabolism; went from $0 to $1.3B revenue; 'Live Fit' positioning; Gen Z + fitness crowd"),
        ("Ghost Lifestyle", "ghostlifestyle.com", "Supplements & Energy", "Licensed flavors — Sour Patch Kids, Warheads, Chips Ahoy protein; transparent labels; gaming/lifestyle crossover"),
        ("ButcherBox", "butcherbox.com", "Meat Subscription", "100% grass-fed beef, free-range organic chicken delivered monthly; $600M+ revenue; trust + convenience play"),
        ("Thrive Market", "thrivemarket.com", "Online Grocery", "Costco meets Whole Foods online; membership model ($60/yr); 30-50% off organic products; gives free membership to low-income families"),
        ("Chamberlain Coffee", "chamberlaincoffee.com", "Creator Coffee", "Emma Chamberlain's coffee brand; Gen Z creator to CPG brand pipeline; cute branding; grocery + DTC"),
        ("Jot", "jot.co", "Ultra Coffee", "20x concentrated liquid coffee; one squeeze = one cup; no machine, filters, or waste; ultra-convenient"),
        ("Blue Bottle Coffee", "bluebottlecoffee.com", "Specialty Coffee", "Peaked within 48 hours of roasting; minimalist aesthetic; Nestlé acquired for $700M; craft coffee at scale"),
        ("Oatly", "oatly.com", "Oat Milk", "Punk marketing ('It's like milk but made for humans'); barista edition created oat latte craze; IPO'd at $10B"),
        ("Califia Farms", "califiafarms.com", "Plant Milk", "Distinctive curvy bottle design; oat, almond, coconut milks; $225M+ revenue; eco-conscious branding"),
        ("Hydrant", "drinkhydrant.com", "Hydration Mix", "Rapid hydration mix developed with Oxford scientists; 3x faster than water; targets dehydration from exercise, hangovers, travel"),
        ("Bulletproof", "bulletproof.com", "Biohacking", "Dave Asprey's 'butter in coffee' concept launched biohacking movement; MCT oil, Brain Octane; $200M+ revenue"),
        ("Primal Kitchen", "primalkitchen.com", "Paleo Condiments", "Mark Sisson's paleo mayo (avocado oil) started a revolution; Kraft Heinz acquired; clean condiments category creator"),
        ("Kettle & Fire", "kettleandfire.com", "Bone Broth", "Shelf-stable bone broth; 100% grass-fed; Amazon #1 broth; gut health + keto crossover; $100M+ revenue"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Snacks", "Simple Human food — 'get back to human'; paleo chocolate with no refined sugar; Mondelēz acquired"),
        ("Lesser Evil", "lesserevil.com", "Organic Snacks", "Organic popcorn and snacks; Himalayan pink salt; 'no junk, just snacks'; B Corp certified"),
        ("SmartSweets", "smartsweets.com", "Low Sugar Candy", "Gummy bears with 3g sugar vs 28g in regular; Kicked Sugar™ line; $100M+ revenue; TPG acquired"),
        ("Partake Foods", "partakefoods.com", "Allergen-Free Snacks", "Top 9 allergen-free cookies; Black woman-founded (Denise Woodard); started when daughter had food allergies"),
        ("Daring Foods", "daring.com", "Plant Chicken", "Plant-based chicken pieces that actually taste like chicken; restaurant partnerships; simple ingredient list"),
        ("Impossible Foods", "impossiblefoods.com", "Plant Meat", "Heme molecule makes it 'bleed' like real meat; Burger King Impossible Whopper; $7B valuation; Stanford biochemistry"),
    ],

    # =========================================================================
    # HOME & KITCHEN
    # =========================================================================
    "Home & Kitchen": [
        ("Our Place", "fromourplace.com", "Cookware", "Always Pan replaces 8 pieces of cookware; $145; viral on Instagram; Selena Gomez collab; $200M+ revenue"),
        ("Made In", "madeincookware.com", "Professional Cookware", "Same factories as All-Clad and Mauviel at half the price; Tom Colicchio partnership; restaurant-grade DTC"),
        ("Caraway", "carawayhome.com", "Non-Toxic Cookware", "Ceramic-coated, non-toxic cookware in aesthetic colors; storage system included; the 'Instagram kitchen' brand"),
        ("HexClad", "hexclad.com", "Hybrid Cookware", "Patented hexagonal laser-etched pattern = stainless + nonstick hybrid; Gordon Ramsay invested and endorses; Shark Tank"),
        ("Material Kitchen", "materialkitchen.com", "Kitchen Tools", "The reBoard (cutting board) and reNaked pans designed by an ex-elBulli chef; aesthetic + function; minimal SKUs"),
        ("Brooklinen", "brooklinen.com", "Luxury Bedding", "Kickstarter origin ($237K raised); luxury sheets at half retail price; 'Really Good' tagline; $200M+ revenue"),
        ("Parachute Home", "parachutehome.com", "Bedding & Bath", "Venice Beach-based; hotel-quality linens DTC; expanded to mattresses, rugs, baby; retail stores in upscale neighborhoods"),
        ("Boll & Branch", "bollandbranch.com", "Organic Bedding", "First Fair Trade Certified bedding; used in the White House; $100M+ revenue; premium organic cotton"),
        ("Cozy Earth", "cozyearth.com", "Bamboo Bedding", "Oprah's Favorite Things multiple years running; bamboo viscose fabric; temperature-regulating; premium price ($250+ sheets)"),
        ("Ruggable", "ruggable.com", "Washable Rugs", "Machine-washable 2-piece rug system (cover + pad); solved the 'rugs are impossible to clean' problem; $200M+ revenue"),
        ("Article", "article.com", "Modern Furniture", "Mid-century modern furniture DTC; no showrooms initially; Sven sofa is iconic; $500M+ revenue; 3-day delivery"),
        ("Burrow", "burrow.com", "Modular Sofas", "Modular sofa ships in boxes, assembles without tools in minutes; built-in USB charger; nomad-friendly furniture"),
        ("Floyd", "floydhome.com", "Modular Furniture", "The Bed Frame uses a patented clamp system (no tools); Detroit-designed; flat-pack but not IKEA quality compromises"),
        ("Outer", "liveouter.com", "Outdoor Furniture", "Outdoor sofa with built-in cover that rolls up from the back; 'Neighborhood Showrooms' — hosts let you try in their backyard"),
        ("Tushy", "hellotushy.com", "Bidet Attachments", "Made bidets mainstream in US with humor marketing; 'stop wiping your butt'; $80 bidet attachment; $50M+ revenue"),
        ("Canopy", "getcanopy.co", "Humidifiers", "No-mist humidifier with anti-mold tech; dishwasher-safe; aroma kit add-ons; solved 'humidifiers are gross' problem"),
        ("Molekule", "molekule.com", "Air Purifiers", "PECO technology (different from HEPA) destroys pollutants at molecular level; $499 device; patent-protected science"),
        ("Otherland", "otherland.com", "Luxury Candles", "Art-inspired candle packaging designed to keep as decor; glass jars become vases; 'maximalist scents for the modern home'"),
        ("Homesick", "homesickcandles.com", "Nostalgia Candles", "Candles that smell like your home state (all 50 states); also cities, countries, memories; genius gifting product"),
        ("Vitruvi", "vitruvi.com", "Diffusers & Oils", "Stone-look diffuser is a design object, not an eyesore; essential oil sets; elevated the diffuser aesthetic category"),
        ("Solo Stove", "solostove.com", "Fire Pits", "Smokeless fire pit using airflow engineering; doubled as backyard status symbol during pandemic; $400M revenue; Solo Brands IPO'd"),
        ("Ooni", "ooni.com", "Pizza Ovens", "Portable pizza oven reaches 950°F in 20 minutes; Kickstarter origin; $100M+ revenue; made home pizza rivaling Neapolitan"),
        ("Fellow", "fellowproducts.com", "Coffee Equipment", "Stagg EKG kettle became the pour-over kettle; won design awards; San Francisco specialty coffee aesthetic"),
        ("BlendJet", "blendjet.com", "Portable Blender", "Portable blender that charges via USB; TikTok viral with on-the-go smoothie content; $200M+ revenue; $20 price"),
        ("Nugget", "nuggetcomfort.com", "Play Couch", "Kids' play couch with cult parent following; waitlists of 100K+; limited color drops sell out instantly; $100M+ revenue"),
        ("simplehuman", "simplehuman.com", "Smart Home Goods", "Sensor trash cans, mirrors, soap dispensers; engineering-first design; $500M+ revenue; built a luxury trash can empire"),
        ("Bearaby", "bearaby.com", "Knitted Weighted Blankets", "Hand-knit weighted blankets from organic cotton (no beads/fillers); Tree Napper uses TENCEL; $50M+ revenue"),
        ("Open Spaces", "getopenspaces.com", "Modern Storage", "Beautiful storage and organization products; bins, shelf risers in colors; Marie Kondo aesthetic for real life"),
        ("Loloi Rugs", "loloirugs.com", "Designer Rugs", "Chris Loves Julia and Amber Lewis collabs create waitlists; designer rugs at accessible prices; influencer-collab model"),
        ("GreenPan", "greenpan.us", "Ceramic Cookware", "Invented Thermolon ceramic non-stick coating; no PFAS ever; first healthy non-stick alternative; Belgian engineering"),
        ("Lodge Cast Iron", "lodgecastiron.com", "Cast Iron", "Family-owned since 1896 in Tennessee; pre-seasoned cast iron; $15-40 skillets last generations; buy-it-for-life"),
        ("AeroPress", "aeropress.com", "Coffee Maker", "Invented by Stanford lecturer/frisbee inventor Alan Adler; $30 brewer beloved by coffee snobs; World AeroPress Championship"),
        ("Ember", "ember.com", "Temperature Mug", "Keeps coffee at exact temperature you set via app; $130 smart mug; Apple Store distribution; used in offices everywhere"),
    ],

    # =========================================================================
    # PET CARE
    # =========================================================================
    "Pet Care": [
        ("The Farmer's Dog", "thefarmersdog.com", "Fresh Dog Food", "Super Bowl ad (boy + dog growing up) drove massive signups; $1B+ valuation; fresh human-grade meals; vet-developed"),
        ("Ollie", "myollie.com", "Fresh Dog Food", "Custom meal plans based on dog's profile; slow-cooked in USDA kitchens; portions pre-measured per dog"),
        ("Nom Nom", "nomnom.com", "Fresh Pet Food", "Full microbiome testing for dogs; board-certified veterinary nutritionists; $65M in funding; Mars Petcare acquired"),
        ("Jinx", "jinx.com", "Premium Dog Food", "Targets millennial dog parents; clean ingredients, great branding; 'real food for real good dogs'; celebrity investors"),
        ("Sundays for Dogs", "sundaysfordogs.com", "Air-Dried Dog Food", "Air-dried (not kibble, not raw, not freeze-dried) — a new category; USDA-grade beef; human-grade"),
        ("BarkBox", "barkbox.com", "Dog Subscription", "Themed monthly toy + treat boxes; 2M+ subscribers; BARK went public; Super Chewer for heavy chewers; $500M+ revenue"),
        ("Wild One", "wildone.com", "Modern Pet", "Designed pet products that look good in your home; harness, leash, carrier in matching colors; the 'Away of pet'"),
        ("Fi", "tryfi.com", "Smart Dog Collar", "GPS tracking + activity monitoring; Series 3 collar; escape alerts if dog leaves yard; LTE-connected; lost dog finder"),
        ("PrettyLitter", "prettylitter.com", "Health-Monitoring Litter", "Color-changing litter detects health issues (UTI, kidney problems); silica crystal; subscription model; $100M+ revenue"),
        ("Litter-Robot", "litter-robot.com", "Self-Cleaning Litter", "Automatic self-cleaning litter box; $700 price point but saves time; WiFi-connected; cult following among cat owners"),
        ("Chewy", "chewy.com", "Online Pet Store", "Handwritten holiday cards, flowers when pets pass away; legendary customer service; $11B+ revenue; Wayfair of pets"),
        ("Embark", "embarkvet.com", "Dog DNA Tests", "Tests 350+ breeds and 230+ health conditions; most comprehensive dog DNA kit; Cornell partnership; actionable health data"),
        ("Native Pet", "nativepet.com", "Dog Supplements", "Vet-developed, clean ingredient supplements; Yak Chews became viral; bone broth topper; $50M+ revenue"),
        ("Zesty Paws", "zestypaws.com", "Pet Supplements", "#1 pet supplement brand on Amazon; soft chew format; hip & joint, calming, allergy formulas; Health Warrior acquired"),
        ("Open Farm", "openfarmpet.com", "Ethical Pet Food", "100% traceable ingredients — scan QR code to see exact farm sources; humane certifications; B Corp"),
        ("Furbo", "shopus.furbo.com", "Dog Camera", "Dog camera that tosses treats and has barking alerts; 2-way audio; lets you check on dog from work; viral on social"),
        ("Diggs", "diggs.pet", "Modern Dog Crate", "Revol dog crate is collapsible, stylish, and safe (puppy + adult configurations); solved the ugly crate problem"),
        ("Tuft + Paw", "tuftandpaw.com", "Modern Cat Furniture", "Cat furniture that humans actually want in their living room; designer cat trees, beds, scratchers; beautiful + functional"),
        ("Butternut Box", "butternutbox.com", "Fresh Dog Food UK", "UK's #1 fresh dog food; gently cooked; pre-portioned; 100K+ subscribers; massive UK market share"),
        ("Stella & Chewy's", "stellaandchewys.com", "Raw Pet Food", "Freeze-dried raw diet pioneer; started when founder's rescue dog healed on raw food; premium nutrition"),
        ("West Paw", "westpaw.com", "Sustainable Dog Toys", "Made in Montana; Zogoflex toys guaranteed tough; recyclable (mail back program); B Corp; eco-conscious pet parents"),
        ("Chippin", "chippin.com", "Sustainable Dog Food", "Cricket and spirulina protein dog treats; sustainable protein sources; silver lining cricket flour; unique angle"),
        ("A Pup Above", "apupabove.com", "Sous Vide Dog Food", "Sous vide cooked dog food (first to use this technique for pets); human-grade ingredients; slow-cooked nutrition"),
        ("PetHonesty", "pethonesty.com", "Pet Supplements", "Vet-recommended soft chew supplements; $100M+ revenue; Amazon best-seller; hemp, probiotics, multivitamins"),
        ("Finn", "petfinn.com", "Dog Supplements", "Clean label dog supplements; Calming Aid and Allergy & Itch top sellers; vet-backed; subscription model"),
    ],

    # =========================================================================
    # ACCESSORIES & JEWELRY
    # =========================================================================
    "Accessories & Jewelry": [
        ("Mejuri", "mejuri.com", "Fine Jewelry", "'Fine jewelry shouldn't be saved for someone else to buy you'; self-purchase positioning; $100M+ revenue; gold vermeil to solid gold"),
        ("Ana Luisa", "analuisa.com", "Sustainable Jewelry", "Carbon-neutral jewelry brand; uses recycled gold; price range $40-120; sustainability-first fine jewelry"),
        ("Gorjana", "gorjana.com", "Layering Jewelry", "Delicate layering jewelry from Laguna Beach; Parker necklace is iconic; $50-200 sweet spot; 20+ stores"),
        ("Missoma", "missoma.com", "Demi-Fine Jewelry", "Kate Middleton wears it; gold vermeil + semi-precious stones; London-based; 'the brand that got everyone into gold hoops'"),
        ("VRAI", "vrai.com", "Lab-Grown Diamonds", "Diamond Foundry's retail brand (Leonardo DiCaprio invested); own the mine (zero-emission foundry); fully traceable"),
        ("Brilliant Earth", "brilliantearth.com", "Ethical Diamonds", "Beyond Conflict Free sourcing; mines traced and verified; virtual try-on AR tech; $400M+ revenue; IPO'd"),
        ("Ring Concierge", "ringconcierge.com", "Fine Jewelry", "Nicole Wegman built on Instagram showing real engagement rings; DMs-to-sales model; 'your cool older sister' vibe"),
        ("Studs", "studs.com", "Ear Piercing", "Reinvented the ear piercing experience; earscape curation; clean studios (not Claire's); Gen Z piercing destination"),
        ("MVMT", "mvmt.com", "Watches", "Crowdfunded $300K on Indiegogo; sold to Movado for $100M; affordable luxury watches ($100-200); Instagram marketing masters"),
        ("Daniel Wellington", "danielwellington.com", "Watches", "Pioneered influencer marketing in watches; interchangeable NATO straps; $200M+ revenue from Instagram strategy alone"),
        ("Ridge Wallet", "ridgewallet.com", "Metal Wallets", "Slim metal RFID-blocking wallet; #1 podcast-advertised product; 'ditch the bulge' campaign; $100M+ revenue"),
        ("Bellroy", "bellroy.com", "Slim Wallets", "Slim your wallet philosophy; innovative hidden features; Australian design; expanded from wallets to bags and tech"),
        ("Warby Parker", "warbyparker.com", "Eyewear", "Home Try-On (5 free frames); Buy a Pair Give a Pair (10M+ donated); first DTC brand to IPO; $3B peak valuation"),
        ("Pair Eyewear", "paireyewear.com", "Customizable Glasses", "Magnetic snap-on toppers to change your glasses' look daily; 100+ designs; partnership with Marvel, NBA, Disney"),
        ("Goodr", "goodr.com", "Running Sunglasses", "Polarized running sunglasses for $25; wild colorful names (Flamingos on a Booze Cruise); fun in a boring category"),
        ("Blenders Eyewear", "blenderseyewear.com", "Lifestyle Sunglasses", "Chase Fisher started selling sunglasses from his dorm at SDSU; festival/lifestyle culture; $40-70 price point"),
        ("SunGod", "sungod.co", "Performance Sunglasses", "Custom-build your sunglasses (frame color, lens, icons); lifetime guarantee; used by Red Bull athletes; UK-based"),
        ("Casetify", "casetify.com", "Phone Cases", "Custom phone cases from your own Instagram photos; 100+ brand collabs (DHL, NASA, Netflix); $300M+ revenue"),
        ("PopSockets", "popsockets.com", "Phone Grips", "Collapsible phone grip invented by a philosophy professor; 245M+ sold; Magsafe pivots; became a culture icon"),
        ("Ekster", "ekster.com", "Smart Wallets", "Solar-powered tracker card inside wallet; find your wallet via phone; Dutch design; Kickstarter record-breaker"),
        ("BaubleBar", "baublebar.com", "Fashion Jewelry", "$20-50 statement jewelry; weekly new arrivals; sport jewelry collab with NFL, NBA teams; affordable trend"),
        ("Monica Vinader", "monicavinader.com", "Demi-Fine Jewelry", "British demi-fine pioneer; Emma Watson, Kate Middleton fans; engravable friendship bracelets; recycled gold"),
        ("Catbird", "catbirdnyc.com", "Brooklyn Jewelry", "Threadbare ring sparked the 'permanent jewelry' trend; Brooklyn-made; delicate, romantic aesthetic"),
        ("Zenni Optical", "zennioptical.com", "Affordable Glasses", "Prescription glasses starting at $6.95; direct from factory; 50M+ frames sold; disrupted $100B eyewear industry"),
    ],

    # =========================================================================
    # TECH & ELECTRONICS
    # =========================================================================
    "Tech & Electronics": [
        ("Anker", "anker.com", "Charging & Audio", "Started as Amazon-first brand; now $2B+ revenue; USB-C hub standard-setter; Soundcore audio sub-brand; trust = reviews"),
        ("Peak Design", "peakdesign.com", "Camera & Travel Gear", "Kickstarter legend ($32M+ raised across campaigns); innovative quick-release camera clips; lifetime guarantee"),
        ("Moment", "shopmoment.com", "Mobile Photography", "Mobile lenses that rival DSLR quality; built community of phone photographers; content creator gear ecosystem"),
        ("Dbrand", "dbrand.com", "Device Skins", "Snarky, aggressive social media personality ('don't @ us'); custom skins for every device; Darkplates for PS5 legal battle"),
        ("Casetify", "casetify.com", "Phone Cases", "Re/CASETiFY uses recycled phone cases; UGC with Instagram photo cases; 100+ brand collabs; $300M+"),
        ("Keychron", "keychron.com", "Mechanical Keyboards", "Affordable mechanical keyboards for Mac; wireless; hot-swappable; beloved by developers and writers"),
        ("Nothing", "nothing.tech", "Phones & Audio", "Carl Pei (OnePlus co-founder); transparent design shows internal components; Ear (1) buds; Phone (2); anti-Apple"),
        ("Remarkable", "remarkable.com", "Paper Tablet", "E-ink tablet that feels like writing on paper; no notifications, no distractions; for thinkers and note-takers"),
        ("Raycon", "rayconglobal.com", "Wireless Earbuds", "Ray J's brand; massive YouTube sponsorship spend (every creator); $50 earbuds positioned against $200 AirPods"),
        ("Nanoleaf", "nanoleaf.me", "Smart Lighting", "Modular light panels that tile together in any shape; became the 'streaming/gaming room' essential; Thread/Matter"),
        ("SimpliSafe", "simplisafe.com", "Home Security", "No contract, no installation needed; DIY home security; $100M+ revenue; targets people fed up with ADT"),
        ("Wyze", "wyze.com", "Smart Home", "$20 smart camera disrupted $200 competitors; 'nice things don't have to be expensive'; expanded to locks, scales, watches"),
        ("Ring", "ring.com", "Home Security", "Video doorbell inventor; rejected on Shark Tank, then Amazon acquired for $1.2B; neighborhood security network"),
        ("Sonos", "sonos.com", "Home Audio", "Multi-room wireless speakers; simple setup; premium sound; $1.5B+ revenue; lawsuit won against Google for patent"),
        ("Elgato", "elgato.com", "Streaming Gear", "Stream Deck became essential for every content creator; key lights, capture cards; owned by Corsair; creator economy infrastructure"),
        ("Rode", "rode.com", "Microphones", "Australian-made microphones; PodMic for podcasters; Wireless GO changed run-and-gun video; NT1 is industry standard"),
        ("Govee", "govee.com", "Smart Lights", "LED strip lights and ambient lighting; 15M+ TikTok views on room transformation videos; affordable smart home"),
        ("EcoFlow", "ecoflow.com", "Power Stations", "Portable power stations with fastest charging (1hr 0-80%); Delta Pro powers a home; off-grid/emergency market"),
        ("Jackery", "jackery.com", "Portable Power", "Portable solar generators; Explorer series; #1 portable power brand; camping + emergency backup; $500M+ revenue"),
        ("BlendJet", "blendjet.com", "Portable Blender", "BlendJet 2 went viral — 10M+ sold; USB-rechargeable; $20 portable blender; TikTok recipe content; 50% margins"),
        ("Ember", "ember.com", "Temperature Mug", "Keeps your drink at exact desired temp for 1.5hrs; app-controlled; $130 mug; Apple Store featured; tech meets ritual"),
        ("Timekettle", "timekettle.com", "Translation Earbuds", "Real-time language translation earbuds; 40+ languages; used by travelers and businesses; AI-powered"),
        ("Framework", "frame.work", "Modular Laptop", "Fully repairable, upgradeable laptop; fight e-waste; swap CPU, RAM, ports yourself; right-to-repair champion"),
        ("Bambu Lab", "bambulab.com", "3D Printers", "Made consumer 3D printing actually easy and reliable; X1 Carbon is industry best; Chinese engineering at scale"),
    ],

    # =========================================================================
    # BABY & KIDS
    # =========================================================================
    "Baby & Kids": [
        ("The Honest Company", "honest.com", "Baby & Home", "Jessica Alba founded; clean baby products + home; IPO'd; $300M+ revenue; pioneered 'clean for baby' category"),
        ("Lovevery", "lovevery.com", "Play Kits", "Montessori-inspired play kits delivered by developmental stage; neuroscience-backed; $200M+ revenue; parent education"),
        ("KiwiCo", "kiwico.com", "STEM Kits", "Age-appropriate STEM/art crate subscriptions; 40M+ crates shipped; Sandra Oh Lin (ex-PayPal exec) founded; educational play"),
        ("Tubby Todd", "tubbytodd.com", "Baby Skincare", "All Over Ointment cures eczema for babies — cult parent following; word-of-mouth growth; clean baby skincare"),
        ("Bobbie", "hibobbie.com", "Organic Formula", "Only organic infant formula modeled after breast milk; EU-style formula for US market; $100M+ revenue; regulatory warrior"),
        ("ByHeart", "byheart.com", "Infant Formula", "Whole milk-based formula with patented protein blend; spent 5 years in R&D before launching; science-first"),
        ("Owlet", "owletcare.com", "Baby Monitor", "Smart Sock tracks baby's heart rate and oxygen while sleeping; gave parents peace of mind; $300M+ revenue; FDA cleared"),
        ("Nanit", "nanit.com", "Smart Baby Monitor", "AI-powered camera tracks sleep patterns and gives coaching; 'sleep score' for babies; computer vision technology"),
        ("Snoo", "happiestbaby.com", "Smart Bassinet", "Dr. Harvey Karp's robotic bassinet that rocks baby back to sleep; $1,700 or rental; used by hospitals"),
        ("4moms", "4moms.com", "Baby Tech", "MamaRoo swing mimics parent's natural motions; tech-forward baby gear; self-installing car seat"),
        ("Fridababy", "frida.com", "Baby Health", "NoseFrida snot sucker — parents were grossed out then converted; real, honest baby care; expanded to Mom care"),
        ("Coterie", "coterie.com", "Premium Diapers", "Highest performing diaper in testing; 25% more absorbent; sustainably sourced; $10/pack premium worth it to parents"),
        ("Little Sleepies", "littlesleepies.com", "Baby Sleepwear", "Buttery-soft bamboo viscose; matching family sets go viral; limited print drops sell out; cult mom following"),
        ("Kyte Baby", "kytebaby.com", "Bamboo Baby", "Bamboo rayon baby clothes; Sleep Bag is bestseller; sustainable + softest fabric; grew 500% via word-of-mouth"),
        ("Monica + Andy", "monicaandandy.com", "Organic Baby", "GOTS certified organic cotton baby clothing; founded by a mom frustrated with chemical-laden baby clothes"),
        ("Stokke", "stokke.com", "Premium Baby", "Tripp Trapp chair grows with child from baby to adult; Norwegian design; $300+ highchair with lifetime value"),
        ("UPPAbaby", "uppababy.com", "Strollers", "VISTA stroller converts from single to double; Vista V3 is best-selling premium stroller; $1,000+ price justified by quality"),
        ("Doona", "doona.com", "Car Seat Stroller", "Car seat with integrated wheels — unfolds into stroller in seconds; solved the 'carrying car seat' problem; $550"),
        ("Ergobaby", "ergobaby.com", "Baby Carriers", "Omni 360 carrier is ergonomic for both parent and baby; multiple carry positions; hip-healthy certified"),
        ("Mushie", "mushie.com", "Baby Essentials", "Scandinavian aesthetic baby products (bibs, plates, pacifiers) in muted colors; Instagram mom aesthetic"),
        ("Gathre", "gathre.com", "Leather Mats", "Wipeable leather mats for play, picnics, changing; beautiful colors; solved messy baby play; premium play mats"),
        ("Primary", "primary.com", "Kids Basics", "No logos, no slogans — just solid-color kids basics; $10-25; let kids be kids; gender-neutral basics"),
        ("Hanna Andersson", "hannaandersson.com", "Quality Kids", "Swedish quality kids clothing; organic cotton; matching family pajamas are a holiday tradition; hand-me-down quality"),
        ("Once Upon a Farm", "onceuponafarm.com", "Organic Baby Food", "Jennifer Garner co-founded; cold-pressed baby food (never heated); fresh produce in pouches"),
        ("Cerebelly", "cerebelly.com", "Brain Baby Food", "Neurosurgeon-developed baby food targeting 16 key brain nutrients; pediatric neuroscience + food"),
    ],

    # =========================================================================
    # DENTAL & ORAL CARE
    # =========================================================================
    "Dental & Oral Care": [
        ("Quip", "getquip.com", "Electric Toothbrush", "Subscription electric toothbrush (new head + paste every 3 months); sleek metal design; 10M+ mouths served; $100M+ revenue"),
        ("Burst Oral Care", "bfrurst.com", "Sonic Toothbrush", "Built by and for dental hygienists; 25K+ dental professionals recommend it; charcoal bristle variant; subscription model"),
        ("Byte", "byteme.com", "Clear Aligners", "At-night clear aligners (HyperByte speeds treatment); $1,895 vs $5K+ for Invisalign; at-home impression kit"),
        ("SmileDirectClub", "smiledirectclub.com", "Clear Aligners", "Pioneered teledentistry clear aligners; $1,850 all-in; 1.5M+ customers; challenged orthodontist monopoly"),
        ("Snow", "trysnow.com", "Teeth Whitening", "Celebrity-backed (Floyd Mayweather, Rob Lowe); wireless LED mouthpiece; serum + device system; $100M+ revenue"),
        ("Hismile", "hismile.com", "Teeth Whitening", "Australian brand; PAP+ formula (peroxide-free whitening); massive TikTok presence; partnered with every major creator"),
        ("Cocofloss", "cocofloss.com", "Premium Floss", "Dentist-designed floss infused with coconut oil; textured to grab plaque; $8 floss made flossing a luxury experience"),
        ("David's", "davids-usa.com", "Natural Toothpaste", "Premium natural toothpaste in metal tubes (recyclable); made in California; nano-hydroxyapatite option"),
        ("RiseWell", "risewell.com", "Hydroxyapatite Paste", "Uses hydroxyapatite (the mineral your teeth are made of) instead of fluoride; mineral toothpaste pioneer in US"),
        ("Spotlight Oral Care", "spotlightoralcare.com", "Dental Products", "Founded by two Irish dentist sisters; professional-grade at home; whitening strips used by dentists"),
        ("AutoBrush", "autobrush.com", "U-Shaped Brush", "U-shaped mouthpiece brush cleans all teeth in 30 seconds; viral infomercial-style ads; unique product form"),
        ("Moon Oral Care", "moonoralcare.com", "Luxury Oral Care", "Kendall Jenner creative director; activated charcoal whitening pen; beauty-brand approach to oral care"),
    ],

    # =========================================================================
    # HOME CLEANING & SUSTAINABILITY
    # =========================================================================
    "Home Cleaning & Sustainability": [
        ("Blueland", "blueland.com", "Refillable Cleaning", "Tablet + reusable bottle system eliminates single-use plastic; 1B+ single-use plastic bottles replaced; Shark Tank success"),
        ("Grove Collaborative", "grove.co", "Natural Home", "Became the largest plastic-neutral retailer; 'Beyond Plastic' initiative; subscription household essentials; $400M+ revenue"),
        ("Branch Basics", "branchbasics.com", "Non-Toxic Cleaning", "ONE concentrate does everything (laundry, dishes, floors, body); fragrance-free; for chemically sensitive people"),
        ("Earth Breeze", "earthbreeze.com", "Eco Laundry", "Laundry sheets (dissolving paper) replace heavy jugs; carbon-neutral; donate 10 loads for every package sold"),
        ("Sheets Laundry Club", "sheetslaundryclub.com", "Laundry Sheets", "Pre-measured laundry sheets; no mess, no measuring; eco-friendly; podcast advertising powerhouse"),
        ("Public Goods", "publicgoods.com", "Sustainable Essentials", "Membership model ($59/yr) for tree-free paper, organic food, clean personal care; minimalist white packaging"),
        ("Who Gives A Crap", "whogivesacrap.org", "Toilet Paper", "50% of profits build toilets in developing world; bamboo + recycled; fun wrapping paper on rolls; B Corp"),
        ("Stasher", "stasherbag.com", "Silicone Bags", "Reusable silicone bags replace Ziploc; dishwasher, microwave, oven safe; Shark Tank; B Corp; 200M+ bags offset"),
        ("LastObject", "lastobject.com", "Reusable Products", "LastSwab, LastTissue, LastRound — reusable versions of disposable products; Danish design; minimal waste"),
        ("Bite", "bitetoothpaste.com", "Toothpaste Bits", "Toothpaste in tablet form — no plastic tube; bit + water = paste; refill glass jar; 100M+ tubes diverted"),
        ("By Humankind", "byhumankind.com", "Low-Waste Care", "Deodorant, shampoo, mouthwash in refillable containers; personal care that generates 90% less plastic"),
        ("Cleancult", "cleancult.com", "Refill Cleaning", "Cleaning products in paper-based cartons (like milk cartons); coconut-based formulas; stylish glass dispensers"),
        ("Dropps", "dropps.com", "Laundry Pods", "Sensitive skin-friendly laundry/dish pods; no dyes, no added fragrance; compostable film; B Corp"),
        ("Ethique", "ethique.com", "Solid Beauty Bars", "New Zealand brand; solid shampoo/conditioner bars; saved 30M+ plastic bottles; concentrated = less shipping weight"),
    ],

    # =========================================================================
    # OUTDOOR & ADVENTURE
    # =========================================================================
    "Outdoor & Adventure": [
        ("YETI", "yeti.com", "Coolers & Drinkware", "Premium coolers that became status symbols; $1.6B+ revenue; built brand on hunting/fishing communities then crossed over"),
        ("Hydro Flask", "hydroflask.com", "Water Bottles", "The VSCO girl water bottle; TempShield insulation; became a Gen Z icon; sticker culture; $200M+ revenue"),
        ("Stanley", "stanley1913.com", "Drinkware", "100+ year old brand went viral on TikTok (Quencher tumbler); grew from $70M to $750M in 3 years; women resurrected it"),
        ("Cotopaxi", "cotopaxi.com", "Outdoor Gear", "One-of-a-kind colorways from factory remnants (Del Día); Questival adventure races; B Corp; gear that does good"),
        ("Rumpl", "rumpl.com", "Puffy Blankets", "Original puffy blanket (insulation from sleeping bags); prints and collabs; National Park Foundation partnership"),
        ("Goal Zero", "goalzero.com", "Solar Power", "Portable solar panels and power stations; started by bringing power to off-grid communities; now mainstream camping/emergency"),
        ("GRAYL", "grayl.com", "Water Purifier", "Press-style water purifier — clean water in 8 seconds; filters viruses, bacteria, chemicals; adventure travel essential"),
        ("LifeStraw", "lifestraw.com", "Water Filter", "Drink directly from any water source; saved millions of lives; every purchase provides safe water for a child for a year"),
        ("Oru Kayak", "orukayak.com", "Folding Kayaks", "Origami-inspired kayak folds from a box in 3 minutes; lives in your closet; solved kayak storage/transport problem"),
        ("Snow Peak", "snowpeak.com", "Japanese Outdoor", "Japanese craftsmanship outdoor gear since 1958; titanium cookware; 'human nature' philosophy; lifestyle camping"),
        ("BioLite", "bioliteenergy.com", "Camp Stoves", "Stove generates electricity from fire to charge devices; CampStove powers your phone while you cook; dual purpose"),
        ("Patagonia", "patagonia.com", "Outdoor Apparel", "'Don't Buy This Jacket' anti-consumerism ad; donated the entire company to fight climate change; $1.5B revenue"),
        ("Arc'teryx", "arcteryx.com", "Technical Outdoor", "GORE-TEX mastery; designed in Vancouver; used by military special forces; $1B+ revenue; acquired by Amer Sports"),
        ("Fjällräven", "fjallraven.com", "Swedish Outdoor", "Kånken backpack became global fashion icon; G-1000 fabric is repairable forever; 60+ year heritage; Arctic Fox initiative"),
        ("Herschel Supply", "herschel.com", "Bags & Backpacks", "Vancouver brothers turned backpacks into fashion; Little America is iconic; heritage branding from their small town name"),
    ],

    # =========================================================================
    # SUBSCRIPTION BOXES & SERVICES
    # =========================================================================
    "Subscription Boxes & Services": [
        ("FabFitFun", "fabfitfun.com", "Lifestyle Box", "Quarterly box of 8+ full-size beauty/wellness/home products worth $200+ for $55; 2M+ subscribers; customization"),
        ("Birchbox", "birchbox.com", "Beauty Box", "Pioneered the beauty subscription box category in 2010; 5 samples/month; data-driven personalization"),
        ("Ipsy", "ipsy.com", "Beauty Box", "Michelle Phan co-founded; 3.5M+ subscribers; Glam Bag; merged with BoxyCharm to create BFA Industries; data personalization"),
        ("Stitch Fix", "stitchfix.com", "Personal Styling", "AI + human stylist picks 5 items shipped to you; keep what you want; $1.7B+ revenue; pioneered data-driven fashion"),
        ("Rent the Runway", "renttherunway.com", "Fashion Rental", "Designer dress rentals then unlimited subscriptions; 'Closet in the Cloud'; IPO'd; circular fashion pioneer"),
        ("Bespoke Post", "bespokepost.com", "Men's Lifestyle", "'Box of Awesome' — themed lifestyle boxes for men; axes, cocktail kits, grooming; curated masculinity"),
        ("Hunt A Killer", "huntakiller.com", "Mystery Game", "Murder mystery subscription box — solve a case over 6 months; immersive storytelling; 300K+ subscribers; game night revolution"),
        ("Book of the Month", "bookofthemonth.com", "Book Club", "5 curated book picks monthly; choose 1 for $16.99; add-on options; revitalized reading culture; 1M+ members"),
        ("Universal Yums", "universalyums.com", "Int'l Snack Box", "Snacks from a different country each month with trivia booklet; educational + delicious; 45+ countries featured"),
        ("Bokksu", "bokksu.com", "Japanese Snack Box", "Authentic Japanese snacks + tea sourced directly from makers; cultural storytelling with each box; Shark Tank"),
        ("Cratejoy", "cratejoy.com", "Subscription Marketplace", "Marketplace for discovering 7,000+ subscription boxes; the 'Etsy of subscription boxes'; platform play"),
        ("HelloFresh", "hellofresh.com", "Meal Kit", "World's largest meal kit company; $8B+ revenue; pre-portioned ingredients + recipes; acquired Factor, Green Chef"),
        ("Wantable", "wantable.com", "Style Box", "7-item style boxes for fitness, fashion, or work; personal stylist; keep what you love; data learns your taste"),
        ("Athletic Brewing", "athleticbrewing.com", "NA Beer Club", "Athletic Club subscription for non-alcoholic craft beer; exclusive member beers; fastest-growing beer brand"),
    ],

    # =========================================================================
    # TRAVEL & LUGGAGE
    # =========================================================================
    "Travel & Luggage": [
        ("Away", "awaytravel.com", "Modern Luggage", "Built-in ejectable battery; launched with 'The Places We Return To' book; $1.4B valuation; Instagram-first strategy"),
        ("Monos", "monos.com", "Premium Luggage", "Japanese-inspired minimalism; polycarbonate shell; aerospace-grade aluminum; $100M+ revenue; quieter luxury alternative to Away"),
        ("July", "july.com", "Minimalist Luggage", "Australian brand; personalized luggage tag included; lightest hardside carry-on; clean direct-from-factory pricing"),
        ("Béis", "beistravel.com", "Travel Bags", "Shay Mitchell's brand; The Weekender bag is iconic; functional fashion travel; influencer-to-brand pipeline; $100M+ revenue"),
        ("Calpak", "calpaktravel.com", "Luggage & Accessories", "Luka Mini Duffel became a TikTok sensation; affordable fashion luggage ($80-200); packing cubes and organizers"),
        ("Paravel", "tourparavel.com", "Sustainable Travel", "Made from recycled materials including ocean plastic; carbon neutral shipping; eco-luxury travel"),
        ("Nomatic", "nomatic.com", "Travel Bags", "Crowdfunded $5M+ on Kickstarter; travel packs with patent-pending organization; digital nomad community"),
        ("Tropicfeel", "tropicfeel.com", "Travel Shoes", "All-terrain shoes for travel — hike, swim, explore in one shoe; Kickstarter most-funded travel shoe ever"),
        ("Tortuga", "tortugabackpacks.com", "Travel Backpacks", "Carry-on travel backpacks designed for one-bag travel; digital nomad Bible; remote work community"),
        ("Aer", "aersf.com", "Work & Gym Bags", "Bags designed for the gym-to-work commute; dedicated shoe compartment; tech-friendly organization; SF design"),
        ("WANDRD", "wandrd.com", "Camera Backpacks", "Camera bags that don't look like camera bags; PRVKE backpack; photographer + traveler hybrid; Kickstarter hit"),
        ("Baboon to the Moon", "baboontothemoon.com", "Adventure Bags", "Go-Bags in neon colors; built for chaos and adventure; 'for misfits and adventurers'; bold branding"),
        ("Horizn Studios", "horizn-studios.com", "Smart Luggage", "Berlin-designed smart luggage with integrated charger; cabin-size perfectionists; European aesthetic"),
        ("Dagne Dover", "dagnedover.com", "Neoprene Bags", "Neoprene bags with organization for everything (laptop, water, keys); bestselling Landon Carryall; functional fashion"),
    ],

    # =========================================================================
    # SEXUAL WELLNESS
    # =========================================================================
    "Sexual Wellness": [
        ("Dame Products", "dameproducts.com", "Vibrators", "NYC subway ad campaign challenged by MTA (while erectile dysfunction ads ran freely); gender equality in sexual wellness"),
        ("Maude", "getmaude.com", "Intimacy Essentials", "Dakota Johnson invested; 'sex stuff, simplified'; minimalist design; destigmatized buying vibrators; $10M+ revenue"),
        ("Lelo", "lelo.com", "Luxury Pleasure", "Swedish luxury pleasure products; $100-200+ price point; design awards; 'the Rolls-Royce of vibrators'"),
        ("Unbound", "unboundbabes.com", "Sexual Wellness", "Polly Rodriguez fought to run subway ads; subscription box model; empowerment-first branding"),
        ("Love Wellness", "lovewellness.com", "Women's Health", "Lo Bosworth (The Hills) founded; vaginal health supplements + care; broke taboo on women's intimate health"),
        ("Cake", "hellocake.com", "Intimacy Products", "Fun, modern lubricant and intimacy products; Thingtesting's most-reviewed sexual wellness brand"),
        ("Fur", "furyou.com", "Body Care", "Pubic skincare pioneer; ingrown oil, stubble cream; made body hair care a legitimate category; Emma Watson endorsed"),
        ("Womanizer", "womanizer.com", "Pleasure Tech", "Patented Pleasure Air Technology (no direct contact); completely unique sensation category; $200M+ revenue"),
    ],

    # =========================================================================
    # CBD & FUNCTIONAL WELLNESS
    # =========================================================================
    "CBD & Functional Wellness": [
        ("Charlotte's Web", "charlottesweb.com", "CBD", "Named after Charlotte Figi whose seizures were reduced from 300/week to ~3 with CBD; changed hemp laws; OG CBD brand"),
        ("Lord Jones", "lordjones.com", "Premium CBD", "Luxury CBD with fashion-world positioning; $50 body lotion; Sephora distribution; Cronos Group acquired"),
        ("Not Pot", "notpot.com", "CBD Gummies", "Donation-first model — funds bail reform; cute gummy bears; Gen Z activism + wellness; 'chill without the high'"),
        ("Sunday Scaries", "sundayscaries.com", "CBD for Anxiety", "Named after the Sunday night dread feeling; targeted marketing around work-week anxiety; relatable branding"),
        ("Equilibria", "myeq.com", "Women's CBD", "CBD specifically dosed for women; free dosage specialist support; personalized consultation model; membership"),
        ("Receptra Naturals", "receptranaturals.com", "Active CBD", "Family-owned Colorado farm; full-spectrum extract; athlete-focused recovery; USDA organic hemp"),
        ("Prima", "prima.co", "CBD Skincare", "Former Sephora exec founded; hemp-derived CBD in skincare; The Daily supplement; premium wellness positioning"),
        ("Plant People", "plantpeople.co", "Plant Medicine", "Herbalist-formulated supplements combining CBD + adaptogens + functional mushrooms; apothecary aesthetic"),
    ],

    # =========================================================================
    # OFFICE & PRODUCTIVITY
    # =========================================================================
    "Office & Productivity": [
        ("Fully", "fully.com", "Standing Desks", "Jarvis standing desk became remote work essential; bamboo desktop option; ergonomic accessories; Herman Miller acquired"),
        ("Secretlab", "secretlab.co", "Gaming Chairs", "Titan and Omega chairs; esports partnerships (T1, Team Secret); $300M+ revenue; premium gaming throne"),
        ("Ugmonk", "ugmonk.com", "Minimal Desk", "Analog productivity system (physical task cards, not digital); Gather collection; Jeff Sheldon's design vision"),
        ("Grovemade", "grovemade.com", "Wood Desk Accessories", "Portland-made walnut and maple desk accessories; desk pad, monitor stand, pen holder; craftsmanship meets tech"),
        ("Rocketbook", "getrocketbook.com", "Smart Notebooks", "Write, scan with app, microwave to erase — infinite notebook; $30 reusable notebook; eco + tech fusion"),
        ("Baron Fig", "baronfig.com", "Notebooks", "Kickstarter-born notebooks; Confidant is their Moleskine killer; artist collabs; 'thinkers welcome' community"),
        ("Autonomous", "autonomous.ai", "Smart Office", "AI-powered standing desks + chairs at 50% less than Herman Miller; SmartDesk auto-adjusts; $200M+ revenue"),
        ("FlexiSpot", "flexispot.com", "Standing Desks", "Sit-stand desks + desk bikes; affordable ergonomic office; E7 model is bestseller; under-desk treadmill"),
        ("Branch Furniture", "branchfurniture.com", "Office Furniture", "Premium office furniture for home + corporate; 30-day free trial; targets companies setting up home offices"),
        ("Orbitkey", "orbitkey.com", "Key Organization", "Slim key organizer replaces jangling keychain; Nest accessory hub; desk mat; Australian design; crowdfunded origin"),
        ("Leuchtturm1917", "leuchtturm1917.com", "Notebooks", "German quality notebooks; the bullet journal community's notebook of choice; numbered pages; ink-proof paper"),
        ("Rifle Paper Co.", "riflepaperco.com", "Stationery & Home", "Anna Bond's hand-painted floral designs; expanded from cards to wallpaper, furniture, Keds collab; $30M+ revenue"),
    ],

    # =========================================================================
    # PERIOD & FEMININE CARE
    # =========================================================================
    "Period & Feminine Care": [
        ("Thinx", "shethinx.com", "Period Underwear", "Pioneered period underwear category; controversial subway ads normalized menstruation; $100M+ revenue"),
        ("Knix", "knix.com", "Leakproof Underwear", "Kickstarter record in fashion ($1.6M); leakproof underwear; body-positive marketing; $100M+ revenue; Canadian"),
        ("Saalt", "saalt.com", "Menstrual Cups", "Soft silicone menstrual cup; donates to help girls stay in school; 400M+ pads/tampons diverted from landfills"),
        ("Cora", "cora.life", "Organic Tampons", "Organic period products; every purchase gives pads to girls in developing countries; period poverty activism"),
        ("Rael", "getrael.com", "K-Beauty Period", "Korean-made organic period products + skincare patches; beauty + period care crossover; holistic approach"),
        ("Flex", "flexfits.com", "Menstrual Disc", "Menstrual disc you can wear during intimacy; solved a problem nobody talked about; 12-hour wear; medical-grade polymer"),
        ("August", "itsaugust.co", "Period Care", "Gen Z period brand; inclusive (for 'anyone who bleeds'); colorful packaging; subscription model; launched at 16 years old"),
        ("Lola", "mylola.com", "Organic Period", "100% organic cotton tampons via subscription; transparency on ingredients (vs. big brands hiding ingredients)"),
    ],

    # =========================================================================
    # SKINCARE TOOLS & DEVICES
    # =========================================================================
    "Skincare Tools & Devices": [
        ("NuFace", "mynuface.com", "Microcurrent Device", "At-home microcurrent facial toning; 'the 5-minute facelift'; $200-400 devices; loved by MUAs and dermatologists"),
        ("Foreo", "foreo.com", "Cleansing Device", "Swedish silicone facial cleansing device; Luna replaces brush heads (silicone = hygienic forever); $200M+ revenue"),
        ("Solawave", "solawave.co", "Red Light Wand", "4-in-1 skincare wand (red light, microcurrent, warmth, facial massage); $150; viral on TikTok; clinical results"),
        ("Therabody", "therabody.com", "Recovery Tech", "TheraFace PRO combines percussive therapy, LED, microcurrent in one facial device; from body to face"),
        ("HigherDOSE", "higherdose.com", "Infrared", "Infrared sauna blanket for at-home detox; red light face mask; biohacking meets beauty; $500 blanket"),
        ("Medicube", "medicube.us", "Age-R Device", "Korean clinical skincare devices; AGE-R Booster Pro went mega-viral; micro-electro stimulation; K-beauty tech"),
        ("LightStim", "lightstim.com", "LED Treatment", "FDA-cleared LED therapy devices; red light for wrinkles, blue for acne; used by estheticians and at home"),
        ("ZIIP Beauty", "ziipbeauty.com", "Nano Current", "Nanocurrent device paired with app that guides treatments; Melanie Simon's pro technology for home; $495 device"),
        ("Droplette", "droplette.io", "Micro-Infusion", "Micro-mist device pushes serums 20x deeper into skin than rubbing; patented micro-infusion; medical-grade penetration"),
        ("Skin Gym", "skingymco.com", "Beauty Tools", "Rose quartz face rollers, gua sha tools; affordable entry to facial tools ($25-50); self-care ritual builder"),
        ("Dermaflash", "dermaflash.com", "Dermaplaning", "At-home dermaplaning device; removes peach fuzz + dead skin; smoother makeup application; $200 device"),
        ("Trophy Skin", "trophyskin.com", "LED Therapy", "RejuvaliteMD LED panel for full-face treatment; clinical-grade phototherapy at home; FDA-cleared"),
    ],

    # =========================================================================
    # WINE, SPIRITS & NON-ALCOHOLIC
    # =========================================================================
    "Wine, Spirits & Non-Alcoholic": [
        ("Winc", "winc.com", "Wine Club", "Palate Profile quiz personalizes 4 bottles/month; also makes their own wines; data-driven winemaking; $100M+ revenue"),
        ("Bright Cellars", "brightcellars.com", "AI Wine Pairing", "MIT-founded; algorithm matches wine to your taste using 7-question quiz; AI + wine = unique angle"),
        ("Naked Wines", "nakedwines.com", "Angel-Funded Wine", "'Angel' investors fund independent winemakers; get exclusive wines at up to 60% off; crowdfunding for wine"),
        ("Athletic Brewing", "athleticbrewing.com", "NA Beer", "Non-alcoholic craft beer that actually tastes like beer; Bill Shufelt quit Wall Street; $100M+ revenue; #1 NA beer"),
        ("Ghia", "drinkghia.com", "NA Aperitif", "Mediterranean non-alcoholic apéritif; beautiful bottle = bar decor; Le Spritz canned; sober-curious darling"),
        ("Seedlip", "seedlipdrinks.com", "NA Spirits", "World's first distilled non-alcoholic spirit; Diageo invested (first NA investment); elevated the mocktail category"),
        ("HOP WTR", "hopwtr.com", "Sparkling Hop Water", "Sparkling water brewed with hops + adaptogens (ashwagandha, L-theanine); beer taste, zero alcohol, functional"),
        ("De Soi", "drinkdesoi.com", "Adaptogen Aperitif", "Katy Perry's non-alcoholic aperitif with adaptogens; 'a natural high'; celebrity NA beverage trend"),
        ("Haus", "drink.haus", "Aperitifs", "Low-ABV aperitifs for slow sipping; started DTC, expanded to restaurants; 'kill the cocktail'"),
        ("JuneShine", "juneshine.com", "Hard Kombucha", "Organic hard kombucha brewed with green tea + honey; sustainability-first; keg and tap program for offices"),
        ("Uncle Nearest", "unclenearest.com", "Premium Whiskey", "Named after Nathan 'Nearest' Green, the first known African American master distiller who taught Jack Daniel; $100M+ revenue"),
        ("Casamigos", "casamigos.com", "Tequila", "George Clooney + Rande Gerber created for personal use; sold to Diageo for $1B; the celebrity tequila that started it all"),
    ],

    # =========================================================================
    # GARDEN & PLANT
    # =========================================================================
    "Garden & Plant": [
        ("The Sill", "thesill.com", "Indoor Plants", "Made buying plants online feel safe (care guides, plant mom community); 'plants make people happy'; $50M+ revenue"),
        ("Bloomscape", "bloomscape.com", "Plant Delivery", "Ships full-grown plants in custom packaging that keeps them alive; 5th-gen greenhouse family; no tiny sad seedlings"),
        ("UrbanStems", "urbanstems.com", "Flowers & Plants", "Same-day flower delivery; modern arrangements; partnerships with Vogue, Bumble; elevated floral gifting"),
        ("Venus et Fleur", "venusetfleur.com", "Eternity Roses", "Roses that last a year; $300-500+ arrangements; luxury gifting; Kardashian-loved; patented preservation"),
        ("AeroGarden", "aerogarden.com", "Indoor Garden", "Countertop hydroponic garden; grow herbs, lettuce, tomatoes indoors year-round; LED grow lights; $100 entry"),
        ("Click & Grow", "clickandgrow.com", "Smart Garden", "The Nespresso of indoor gardening — pre-seeded pods, auto-watering, LED light; Estonian tech company"),
        ("Lettuce Grow", "lettucegrow.com", "Hydroponic Farm", "Vertical Farmstand grows 36 plants with no soil; founded by Zooey Deschanel + Jacob Pechenik; food justice mission"),
        ("Gardyn", "mygardyn.com", "Home Garden", "AI-powered indoor garden with Kelby the garden assistant; cameras monitor plants; 30+ seed pods; tech-forward growing"),
        ("Farmgirl Flowers", "farmgirlflowers.com", "Bouquets", "Limited daily designs (reduce waste vs. 500-SKU competitors); burlap-wrapped; sustainable floristry; $30M+ revenue"),
        ("Sunday Lawn Care", "sunday.com", "Lawn Care", "Custom lawn care plan via soil test; ships nutrients on a schedule; non-toxic safe for kids and pets; 'smart lawn plan'"),
    ],

    # =========================================================================
    # AUTOMOTIVE & MOBILITY
    # =========================================================================
    "Automotive & Mobility": [
        ("Rad Power Bikes", "radpowerbikes.com", "E-Bikes", "North America's largest e-bike brand; RadRunner utility e-bike; $100M+ revenue; made e-bikes affordable ($1,000-2,000)"),
        ("Lectric eBikes", "lectricebikes.com", "Electric Bikes", "XP 3.0 foldable e-bike at $999; best value in e-bikes; folds for storage; Phoenix, AZ HQ; $150M+ revenue"),
        ("VanMoof", "vanmoof.com", "Smart Bikes", "Anti-theft tech (built-in alarm, GPS tracking, remote lock); Dutch design; 'the Tesla of bikes'; raised $180M+"),
        ("Super73", "super73.com", "E-Bikes", "Retro mini-bike aesthetic meets electric; massive in LA/SoCal; celebrity riders (Post Malone, Will Smith); lifestyle brand"),
        ("Cowboy", "cowboy.com", "E-Bikes", "Belgian smart e-bike with integrated lights, crash detection, theft tracking; $3K premium design; European city riding"),
        ("Chemical Guys", "chemicalguys.com", "Car Detailing", "600+ car care products; massive YouTube detailing tutorials channel (4M+ subs); community-first car care"),
        ("Adam's Polishes", "adamspolishes.com", "Car Care", "Family-owned car detailing brand; graphene coating products; detailing enthusiast community; YouTube education"),
        ("WeatherTech", "weathertech.com", "Car Accessories", "Custom-fit floor mats made in USA; famous Super Bowl ads; $1B+ revenue; American manufacturing pride"),
        ("Onewheel", "onewheel.com", "Electric Board", "Self-balancing electric skateboard; single wheel; off-road capable; 'float life' community; $100M+ revenue"),
        ("Apollo Scooters", "apolloscooters.co", "E-Scooters", "Canadian e-scooter brand; Apollo City is best urban commuter scooter; suspension + dual motors; premium"),
        ("Unagi", "unagi-scooters.com", "E-Scooters", "Dual-motor carbon fiber e-scooter; subscription model ($39/mo includes insurance); minimalist Japanese-inspired design"),
        ("Ride1Up", "ride1up.com", "E-Bikes", "Best bang-for-buck e-bikes; 700 Series rivals $3K bikes at $1,700; direct-to-consumer value play"),
    ],

    # =========================================================================
    # CANDLES & HOME FRAGRANCE
    # =========================================================================
    "Candles & Home Fragrance": [
        ("Boy Smells", "boysmells.com", "Genderful Candles", "LGBTQ+-founded; 'genderful' fragrances; started candles, expanded to fine fragrance; collaborations with Kacey Musgraves"),
        ("Otherland", "otherland.com", "Luxury Candles", "Art-directed box packaging doubles as decor; 'Rattan' is bestseller; each scent inspired by a painting"),
        ("Homesick", "homesickcandles.com", "Nostalgia Candles", "Candles for all 50 US states, countries, memories ('Movie Night', 'Book Club'); $50M+ revenue; genius gifting play"),
        ("P.F. Candle Co.", "pfcandleco.com", "Candles & Home", "Amber jar became iconic; LA-made soy candles; wholesale to 2,000+ shops; expanded to reed diffusers, room sprays"),
        ("Pura", "pura.com", "Smart Home Fragrance", "Smart home diffuser controlled by app; fragrances from Nest, Capri Blue, Anthropologie; scent scheduling; $100M+ revenue"),
        ("NEST New York", "nestnewyork.com", "Home Fragrance", "Laura Slatkin's luxury home fragrance empire; Grapefruit candle is iconic; $300M+ revenue; fragrance + home"),
        ("Brooklyn Candle Studio", "brooklyncandlestudio.com", "Soy Candles", "Minimalist soy candles; 100% soy wax, cotton wick; hand-poured in Brooklyn; clean-burning luxury at fair price"),
        ("Vitruvi", "vitruvi.com", "Essential Oils", "Matte ceramic diffuser looks like decor, not a health gadget; essential oil blends; elevated the aromatherapy aesthetic"),
        ("Voluspa", "voluspa.com", "Luxury Candles", "Coconut wax + Japanese design tins; 100+ scents; found in every upscale boutique hotel; accessible luxury"),
        ("Diptyque", "diptyqueparis.com", "French Candles", "Parisian luxury since 1961; Baies (berries) candle is iconic; $70+ candles; the status candle; oval logo instantly recognized"),
        ("Byredo", "byredo.com", "Swedish Luxury", "Ben Gorham's brand — 'beyond reason'; $800M+ valuation; L'Oréal minority stake; blends fragrance, candles, makeup, leather"),
        ("Capri Blue", "capri-blue.com", "Volcano Candle", "Blue jar Volcano candle is the best-selling candle in America; found at Anthropologie; tropical fruit + sugarcane scent"),
    ],

    # =========================================================================
    # GAMING & ENTERTAINMENT
    # =========================================================================
    "Gaming & Entertainment": [
        ("Backbone", "playbackbone.com", "Mobile Gaming Controller", "Turns iPhone into a gaming console; instant connection; PlayStation partnership; $40M+ raised; mobile gaming enabler"),
        ("Analogue", "analogue.co", "Retro Gaming", "FPGA-based retro consoles that play original cartridges in HD; Pocket plays Game Boy games; collector-grade quality"),
        ("Secretlab", "secretlab.co", "Gaming Chairs", "Titan and Omega chairs; League of Legends, Game of Thrones editions; $300M+ revenue; esports partnerships"),
        ("SCUF Gaming", "scufgaming.com", "Custom Controllers", "Paddles on back of controllers for pro gamers; used by 90%+ of pro Call of Duty players; Corsair acquired"),
        ("Glorious", "gloriousgaming.com", "Gaming Peripherals", "Model O mouse became enthusiast favorite; community co-designed products; GMMK keyboard customization"),
        ("Exploding Kittens", "explodingkittens.com", "Card Games", "Most-backed game on Kickstarter ($8.8M); 30M+ games sold; Netflix series; party game empire from The Oatmeal creator"),
        ("Cards Against Humanity", "cardsagainsthumanity.com", "Party Games", "Kickstarter origin; no paid ads ever; stunts like buying land on Mexico border, Black Friday 'nothing' sale; anti-marketing"),
        ("Razer", "razer.com", "Gaming", "Triple-headed snake logo = gaming lifestyle; $1.5B+ revenue; keyboards, mice, laptops; esports culture"),
        ("SteelSeries", "steelseries.com", "Gaming Peripherals", "Arctis headsets used by top esports pros; 20+ years in gaming audio; GG software ecosystem"),
        ("Corsair", "corsair.com", "PC Gaming", "Full PC gaming ecosystem — keyboards, mice, headsets, cases, RAM, streaming gear; $1.8B+ revenue; acquired Elgato"),
        ("Funko", "funko.com", "Pop Culture Toys", "Pop! vinyl figures of every pop culture character; 1,000+ licenses; $1.3B+ revenue; collectible culture pioneer"),
        ("Lego", "lego.com", "Building Toys", "Adult fans (AFOL) segment growing fastest; $9B+ revenue; Lego Ideas lets fans design sets; the toy brand that became timeless"),
    ],

    # =========================================================================
    # SPORTS & RECREATION
    # =========================================================================
    "Sports & Recreation": [
        ("CROSSNET", "crossnetgame.com", "Four-Way Volleyball", "Invented four-way volleyball; brothers created a new sport category; $30M+ revenue; outdoor game revolution"),
        ("Spikeball", "spikeball.com", "Outdoor Game", "Shark Tank alum; created competitive league and tournament scene; parks and beaches everywhere; $30M+ revenue"),
        ("Bad Birdie", "badbirdie.com", "Golf Apparel", "Bold, loud golf shirts challenging traditional golf dress codes; 'life's too short to wear boring shirts'; anti-country club"),
        ("Callaway Golf", "callawaygolf.com", "Golf Equipment", "$3.8B revenue; acquired Topgolf; AI-designed drivers; merged equipment, entertainment, apparel into lifestyle"),
        ("GoPro", "gopro.com", "Action Cameras", "Created the action camera category; HERO cameras; UGC content = free marketing from extreme sports; $1.1B revenue"),
        ("Insta360", "insta360.com", "360 Cameras", "One X series made 360 video mainstream; invisible selfie stick trick; AI editing; challenging GoPro with innovation"),
        ("DJI", "dji.com", "Drones & Cameras", "80%+ global drone market share; Mavic series; Osmo gimbal; Chinese engineering dominance; $4.5B+ revenue"),
        ("Manduka", "manduka.com", "Yoga Equipment", "PRO mat comes with lifetime guarantee; used in yoga studios worldwide; premium yoga gear standard-setter"),
        ("Liforme", "liforme.com", "Yoga Mats", "AlignForMe system has markers for hand/foot placement; GripForMe material; UK-made; premium $150 yoga mat"),
        ("Wahoo Fitness", "wahoofitness.com", "Cycling Tech", "KICKR smart trainer used by Tour de France teams; ELEMNT bike computer; indoor cycling ecosystem"),
        ("Garmin", "garmin.com", "GPS Watches", "Fenix and Forerunner watches dominate triathlon/ultra running; $5B+ revenue; aviation, marine, fitness, auto"),
        ("Coros", "coros.com", "Sport Watches", "Ultrarunner-loved GPS watches; Kilian Jornet partnership; massive battery life (140+ hours GPS); half the price of Garmin"),
    ],

    # =========================================================================
    # RAZORS & HAIR REMOVAL
    # =========================================================================
    "Razors & Hair Removal": [
        ("Billie", "mybillie.com", "Women's Razors", "First women's razor brand to show body hair in ads; 'Project Body Hair'; $5 starter kit; P&G acquired"),
        ("Flamingo", "shopflamingo.com", "Women's Shave", "Harry's sister brand for women; $10 razor handle + blade set; wax kits and body care; found at Target"),
        ("Athena Club", "athenaclub.com", "Women's Razor", "Razor + wellness subscription; 5-blade razor for $9; expanded to period care; subscription convenience"),
        ("Ulike", "ulike.com", "IPL Hair Removal", "Sapphire cooling IPL device; FDA-cleared; TikTok-viral with instant results content; $300 at-home laser alternative"),
        ("Nood", "trynood.com", "IPL Device", "Flasher 2.0 IPL for $170 (cheapest effective device); 'break up with your razor'; massive TikTok ads spend"),
        ("RoseSkinCo", "roseskinco.com", "IPL Device", "Rose gold IPL device; 'salon results at home'; influencer-first marketing; $300 device replaces $3K treatments"),
        ("Manscaped", "manscaped.com", "Body Grooming", "Lawn Mower body trimmer; 'your balls will thank you' humor marketing; $600M+ revenue; podcast ad king"),
        ("Meridian", "meridiangrooming.com", "Body Grooming", "Below-the-waist grooming brand; ceramic blade body trimmer; 'take care down there'; $100M+ revenue"),
        ("Leaf Shave", "leafshave.com", "Pivoting Razor", "Plastic-free pivoting razor that uses standard double-edge blades; sustainable shaving; $85 handle lasts forever"),
    ],

    # =========================================================================
    # DEODORANT & BODY CARE
    # =========================================================================
    "Deodorant & Body Care": [
        ("Native", "nativecos.com", "Natural Deodorant", "First natural deodorant people actually switched to permanently; coconut oil + shea butter; P&G acquired for $100M"),
        ("Lume", "lumedeodorant.com", "Whole Body Deodorant", "Created 'whole body deodorant' category (pits, privates, feet); Dr. Shannon Klingman's gynecologist-developed formula"),
        ("Each & Every", "eachandevery.com", "Clean Deodorant", "Worry-free ingredients (EWG 1-rated); Dead Sea salt-based; refillable; sensitive skin formula; no baking soda"),
        ("Touchland", "touchland.com", "Hand Sanitizer", "Made hand sanitizer chic; Power Mist dispenser in designer scents; $500K sold in 2 days; $25M+ revenue"),
        ("Megababe", "megababebeauty.com", "Body Positivity", "Thigh Rescue anti-chafe stick solved a problem no one marketed; Bust Dust, Rosy Pits; body-positive solutions"),
        ("Sol de Janeiro", "soldejaneiro.com", "Body Care", "Brazilian Bum Bum Cream's scent went so viral they launched a fragrance; $200M+ revenue; Cheirosa 62 perfume"),
        ("Soft Services", "softservices.com", "Body Skincare", "Skincare for body — not face repackaged; Smoothing Solution for KP (keratosis pilaris); targeted body issues"),
        ("Nécessaire", "necessaire.com", "Body Care", "The Body Wash: vitamins + minerals in a body wash; $25 body wash made people upgrade from Dove; Sephora exclusive"),
    ],

    # =========================================================================
    # PRINT ON DEMAND & CREATOR ECONOMY
    # =========================================================================
    "Print on Demand & Creator Economy": [
        ("Printful", "printful.com", "Print on Demand", "Print-on-demand leader; integrates with Shopify, Etsy, Amazon; $300M+ revenue; no inventory needed for merch"),
        ("Printify", "printify.com", "POD Platform", "Connects to 90+ print providers globally; cheapest per-unit costs; built for side hustlers and creators"),
        ("Redbubble", "redbubble.com", "Artist Marketplace", "Independent artists upload designs, printed on everything (stickers, tees, phone cases); 700K+ artists; $400M+ GMV"),
        ("Spring (TeeSpring)", "spri.ng", "Creator Merch", "YouTube integration lets creators sell merch directly under videos; pivoted from tees to full creator commerce platform"),
        ("Threadless", "threadless.com", "Artist Tees", "Community votes on designs; winning artists get produced and paid; pioneered crowdsourced design 20+ years ago"),
        ("Bonfire", "bonfire.com", "Custom Tees", "Fundraising tee platform; nonprofits, schools, causes use it; no upfront cost; print-on-demand for good"),
        ("CustomInk", "customink.com", "Custom Apparel", "Group ordering platform for teams, events, companies; $500M+ revenue; design online, delivered in days"),
        ("Cricut", "cricut.com", "Craft Machines", "Cutting machines for DIY crafts; $2B+ peak market cap; subscription model for designs; maker movement enabler"),
        ("Glowforge", "glowforge.com", "Laser Cutter", "Desktop laser cutter/engraver; $28M crowdfunded (record at time); small businesses use for custom products"),
        ("Etsy", "etsy.com", "Handmade Marketplace", "7M+ sellers; $13B+ GMV; 'keep commerce human'; powered small business and maker movement globally"),
    ],

    # =========================================================================
    # EDUCATION & LEARNING
    # =========================================================================
    "Education & Learning": [
        ("MasterClass", "masterclass.com", "Expert Courses", "Celebrities teach their craft (Gordon Ramsay cooking, Natalie Portman acting); cinematic production quality; $2.75B valuation"),
        ("Skillshare", "skillshare.com", "Creative Learning", "Creator-led creative classes; $20M+ paid to teachers; animation, design, photography; YouTube creator sponsor staple"),
        ("Brilliant", "brilliant.org", "STEM Learning", "Interactive problem-solving for math, science, CS; learn by doing, not watching; 10M+ users; podcast ad powerhouse"),
        ("Duolingo", "duolingo.com", "Language Learning", "Gamified language learning; owl mascot became TikTok star (33M+ followers); $700M+ revenue; free tier + premium"),
        ("Coursera", "coursera.org", "Online Courses", "University courses from Stanford, Yale online; 130M+ learners; Professional Certificates from Google, IBM; IPO'd"),
        ("Headspace", "headspace.com", "Mindfulness", "Meditation made approachable; animated style; Netflix series; $3B valuation; corporate wellness; Andy Puddicombe's vision"),
        ("Calm", "calm.com", "Meditation", "Sleep Stories narrated by celebrities (Matthew McConaughey, Harry Styles); $2B valuation; Apple's App of the Year"),
        ("Osmo", "playosmo.com", "Kids Interactive", "Physical-digital play — kids interact with real objects and iPad 'sees' them; acquired by BYJU'S; hands-on STEM"),
        ("Outschool", "outschool.com", "Live Classes Kids", "Marketplace for live online classes for kids (slime making to AP Physics); 100K+ classes; pandemic growth rocket"),
        ("Homer", "learnwithhomer.com", "Kids Learning", "Learn-to-read program for ages 2-8; personalized learning path; acquired by BEGiN; early literacy focus"),
    ],

    # =========================================================================
    # ART & HOME DECOR
    # =========================================================================
    "Art & Home Decor": [
        ("Framebridge", "framebridge.com", "Custom Framing", "Mail your art/photo, get it framed and shipped back; 70% cheaper than local framers; democratized framing"),
        ("Minted", "minted.com", "Art & Stationery", "Crowdsourced designs from independent artists; limited edition art prints; wedding stationery; Mariam Naficy's vision"),
        ("Society6", "society6.com", "Art Prints", "Artists upload designs, printed on 80+ products; artist gets paid per sale; no inventory; creative marketplace"),
        ("iCanvas", "icanvas.com", "Canvas Art", "Licensed art from 10,000+ artists; affordable canvas prints; exclusive partnerships with museums and estates"),
        ("Artifact Uprising", "artifactuprising.com", "Photo Books", "Premium photo books, prints, frames from your photos; 'hold what matters'; recycled and sustainable materials"),
        ("Desenio", "desenio.com", "Wall Art", "Scandinavian poster prints and frames; gallery wall builder tool; $100M+ revenue; affordable art democratization"),
        ("Studio McGee", "studiomcgee.com", "Designer Home", "Netflix Dream Home Makeover stars; Target collection; $200M+ revenue; approachable interior design"),
        ("Lulu and Georgia", "luluandgeorgia.com", "Boho Home", "Boho-chic rugs, furniture, decor; Instagram aesthetic brand; designer collabs; $100M+ revenue"),
        ("Serena & Lily", "serenaandlily.com", "Coastal Home", "Coastal-inspired home furnishings; riviera counter stool is iconic; $400M+ revenue; elevated casual"),
        ("Jenni Kayne", "jennikayne.com", "California Living", "California lifestyle brand — apparel to home; 'Pacific Natural' aesthetic; acquired Oak Essentials; $100M+ revenue"),
        ("West Elm", "westelm.com", "Modern Home", "Fair Trade, organic materials commitment; mid-century modern; $1.8B revenue; Williams-Sonoma owned"),
        ("CB2", "cb2.com", "Modern Furniture", "Crate & Barrel's contemporary line; urban, edgy design; affordable modern furniture for small spaces"),
    ],

    # =========================================================================
    # WORKWEAR & BASICS
    # =========================================================================
    "Workwear & Basics": [
        ("Carhartt", "carhartt.com", "Workwear", "100+ year workwear heritage adopted by fashion; WIP (Work In Progress) line bridges streetwear; $1B+ revenue"),
        ("Dickies", "dickies.com", "Workwear", "Skateboarding and hip-hop adopted workwear as fashion; 922 work pants became streetwear; VF Corp owned"),
        ("Duluth Trading", "duluthtrading.com", "Tough Workwear", "Buck Naked underwear; humorous ads; 'because work is tough'; $700M+ revenue; fire hose work pants"),
        ("Huckberry", "huckberry.com", "Men's Lifestyle", "Curated marketplace for the modern adventurer; editorial + commerce; in-house brands + curation; community of 2M+"),
        ("Brunt Workwear", "bruntworkwear.com", "Work Boots", "Work boots designed with tradespeople; free try-before-you-buy; $100M+ revenue; fills gap between fashion and function"),
        ("Bombas", "bombas.com", "Comfort Socks", "One donated for every one purchased (60M+ donated); Shark Tank's most successful brand ever; $300M+ revenue"),
        ("Darn Tough", "darntough.com", "Lifetime Socks", "Unconditional lifetime guarantee; made in Vermont; merino wool; send back worn socks = free new pair; buy-it-for-life"),
        ("Saxx", "saxxunderwear.com", "Men's Underwear", "BallPark Pouch technology — anatomical 3D hammock-shaped pouch; solved men's underwear comfort; $250M+ revenue"),
        ("Western Rise", "westernrise.com", "Tech Pants", "AT Slim Rivet Pant — looks like chinos, performs like athletic pants; travels, commutes, hikes; one pant does everything"),
        ("Outlier", "outlier.nyc", "Technical Apparel", "NYC-made technical clothing; Strong Dungarees, Futureworks; minimal runs; cult menswear following; merino + tech fabrics"),
        ("True Classic", "trueclassictees.com", "Men's Tees", "Spent $50M+ on Meta ads targeting 'dad bods'; V-taper fit flatters average men; $250M+ revenue from Facebook ads alone"),
        ("Fresh Clean Threads", "freshcleanthreads.com", "Men's Tees", "Crew Pack subscription model; 3-pack tees auto-shipped; simple basics, massive scale; podcast ad staple"),
    ],

    # =========================================================================
    # SMART HOME & IOT
    # =========================================================================
    "Smart Home & IoT": [
        ("Ring", "ring.com", "Home Security", "Video doorbell inventor Jamie Siminoff was rejected on Shark Tank; Amazon acquired for $1.2B; Neighbors app community"),
        ("SimpliSafe", "simplisafe.com", "Home Security", "No contracts, self-install; $15/month monitoring; $1B+ valuation; anti-ADT positioning; 4M+ customers"),
        ("Wyze", "wyze.com", "Smart Home", "$20 camera broke the market open; smart scale, watch, headphones all under $50; 'nice things don't need to be expensive'"),
        ("Nanoleaf", "nanoleaf.me", "Smart Lighting", "Modular LED light panels in hexagons, triangles, lines; became the gaming/streaming room essential; Thread protocol"),
        ("Govee", "govee.com", "Smart Lights", "LED strip lights went viral on TikTok room makeovers; smart bulbs, floor lamps; affordable smart lighting leader"),
        ("Aqara", "aqara.com", "Smart Home", "Affordable Zigbee/Matter smart home devices; sensors, switches, cameras; Apple Home integration; Chinese engineering value"),
        ("Level Lock", "level.co", "Smart Lock", "Invisible smart lock — looks like a normal deadbolt; smallest smart lock; Apple Home Key support; design-first"),
        ("Ecobee", "ecobee.com", "Smart Thermostat", "First Alexa-built-in thermostat; SmartSensor detects occupancy; energy savings; Canadian innovation; $750M acquisition"),
        ("Roborock", "roborock.com", "Robot Vacuum", "LiDAR navigation robot vacuums; mops too; S8 Pro Ultra docks itself and washes its own mop; $500M+ revenue"),
        ("Tineco", "tineco.com", "Smart Cleaning", "Floor ONE S5 smart wet/dry vacuum; iLoop sensor adjusts suction; TikTok cleaning community favorite"),
        ("SwitchBot", "switch-bot.com", "Smart Home", "Tiny robots that physically push buttons/flip switches; retrofit any dumb device to smart; $50 entry; Matter support"),
        ("Coway", "cowaymega.com", "Air & Water", "South Korean air/water purifier leader; Airmega line; rental/subscription model in Korea; $2B+ revenue"),
    ],

    # =========================================================================
    # ATHLEISURE & LOUNGEWEAR
    # =========================================================================
    "Athleisure & Loungewear": [
        ("SKIMS", "skims.com", "Loungewear", "Kim K's $4B brand; redefined shapewear as everyday basics; Olympics deal; Fits Everybody line; 30M+ items sold"),
        ("Quince", "onequince.com", "Affordable Luxury", "$50 cashmere sweaters, $80 leather jackets; factory-direct pricing; 'luxury materials, not luxury markups'; $1B+ GMV"),
        ("Lunya", "lunya.co", "Sleepwear", "Elevated sleepwear — washable silk sets; made staying in stylish; 'rest is productive'; premium PJs for women"),
        ("Lake Pajamas", "lakepajamas.com", "Luxury PJs", "Pima cotton pajamas; matching family sets; Southern hospitality branding; word-of-mouth growth; celebrity loved"),
        ("Hill House Home", "hillhousehome.com", "Nap Dress", "Nell Diamond's Nap Dress — elasticated bodice dress for sleeping or going out; drops sell out in minutes; $50M+ revenue"),
        ("Cozy Earth", "cozyearth.com", "Bamboo Loungewear", "Oprah's Favorite Things 5 years running; bamboo viscose; temperature-regulating; premium but Oprah-validated"),
        ("Eberjey", "eberjey.com", "Sleepwear", "Gisele modal pajamas; softest loungewear; Miami-founded; net-zero packaging goal; luxury intimates"),
        ("Barefoot Dreams", "barefootdreams.com", "Cozy Blankets", "CozyChic blanket became the celebrity baby gift; Chrissy Teigen, Khloé Kardashian fans; $200M+ revenue"),
        ("Printfresh", "printfresh.com", "Printed PJs", "Bold, colorful pajama prints (tigers, flowers, mushrooms); designed in Philadelphia; hand-drawn patterns; stand-out sleepwear"),
        ("Naadam", "naadam.co", "Cashmere", "'The cashmere company that's doing things differently'; Mongolian herder partnerships; $75 cashmere sweater (vs $300+ traditional)"),
        ("Jenni Kayne", "jennikayne.com", "California Style", "California Closet capsule wardrobe; neutral palette; 'Pacific Natural' lifestyle brand; home + fashion integration"),
        ("Draper James", "draperjames.com", "Southern Style", "Reese Witherspoon's brand; Southern charm meets modern fashion; book club integration; lifestyle beyond clothes"),
    ],

    # =========================================================================
    # WATCHES & WEARABLES
    # =========================================================================
    "Watches & Wearables": [
        ("MVMT", "mvmt.com", "Watches", "Crowdfunded $300K; sold to Movado Group for $100M; Instagram marketing pioneers; affordable luxury ($100-200)"),
        ("Daniel Wellington", "danielwellington.com", "Watches", "Pioneered influencer marketing — sent free watches to Instagram influencers; built a $200M brand with almost no traditional ads"),
        ("Vincero", "vincerocollective.com", "Watches", "Premium look at $170-250; Italian marble dials, Swiss movements; 'timeless, not trendy'; 500K+ watches sold"),
        ("Garmin", "garmin.com", "GPS Watches", "Fenix series dominates outdoor/ultra running; $5B+ revenue; aviation roots; most trusted by adventurers and military"),
        ("Coros", "coros.com", "Sport Watches", "140+ hours GPS battery life; Kilian Jornet (greatest ultrarunner) partnership; half the price of Garmin flagship"),
        ("Whoop", "whoop.com", "Fitness Wearable", "No screen, subscription model; strain/recovery tracking; Patrick Mahomes, LeBron James wear it; $3.6B valuation"),
        ("Oura Ring", "ouraring.com", "Health Ring", "Ring form factor for sleep/health tracking; NBA used for COVID detection; $2.55B valuation; most discrete tracker"),
        ("Withings", "withings.com", "Hybrid Watches", "ScanWatch detects AFib and sleep apnea; looks like a traditional watch; French design; FDA/CE medical grade"),
        ("Fitbit", "fitbit.com", "Fitness Tracker", "Popularized fitness tracking; 120M+ devices sold; Google acquired for $2.1B; Pixel Watch integration"),
        ("Apple Watch", "apple.com/watch", "Smartwatch", "40%+ smartwatch market share; ECG, fall detection, crash detection save lives; Ultra for extreme sports; $30B+ segment"),
        ("Nordgreen", "nordgreen.com", "Danish Watches", "Designed by Jakob Wagner (Bang & Olufsen designer); gives back program (education, water, rainforest); Scandinavian minimal"),
        ("Shinola", "shinola.com", "Detroit Watches", "Built in Detroit revitalization story; American-assembled watches; 'where American is made'; factory tours; $100M+ revenue"),
    ],

    # =========================================================================
    # INNOVATIVE & UNIQUE PRODUCTS
    # =========================================================================
    "Innovative & Unique Products": [
        ("Loop Earplugs", "loopearplugs.com", "Noise Reduction", "Stylish earplugs that look like jewelry; reduce noise without blocking it; went viral for concerts/focus; $100M+ revenue"),
        ("Dyson", "dyson.com", "Engineering", "5,127 prototypes before the first vacuum worked; bladeless fans, Airwrap styler, V15 vacuum; $7.7B revenue; R&D obsession"),
        ("Cricut", "cricut.com", "DIY Crafting", "Cutting machines for crafters; $2B+ peak market cap; Design Space software; subscription model; maker movement enabler"),
        ("Bambu Lab", "bambulab.com", "3D Printers", "Made 3D printing reliable and fast for consumers; X1-Carbon prints at 500mm/s; AI failure detection; disrupted the market"),
        ("xTool", "xtool.com", "Laser Engraver", "Desktop laser engravers for small businesses; engrave wood, leather, acrylic; M1 is bestselling laser; maker economy"),
        ("Framework", "frame.work", "Modular Laptop", "Fully repairable laptop — swap RAM, battery, ports, screen; right-to-repair champion; sustainable electronics"),
        ("Remarkable", "remarkable.com", "E-Paper Tablet", "E-ink writing tablet; zero distractions (no apps, no browser); 'replace your notebooks'; $100M+ revenue"),
        ("Rocketbook", "getrocketbook.com", "Smart Notebooks", "Write with Pilot FriXion pen, scan with app, microwave to erase; infinite notebook; $30 eco-friendly tech"),
        ("Timekettle", "timekettle.com", "Translation Earbuds", "WT2 Edge earbuds translate 40+ languages in real-time; bi-directional simultaneous translation; AI-powered"),
        ("Nothing", "nothing.tech", "Transparent Tech", "Carl Pei's brand; transparent design showing internals; Phone (2) has LED Glyph Interface; anti-Apple aesthetic"),
        ("Analogue", "analogue.co", "Retro Console", "FPGA-based consoles play original cartridges perfectly; Analogue Pocket = portable retro gaming; collector-grade"),
        ("Steam Deck", "store.steampowered.com", "Handheld PC", "Valve's portable PC gaming device; plays your entire Steam library; $399 entry; outsold Nintendo Switch in months"),
    ],

    # =========================================================================
    # FUNCTIONAL DRINKS & SUPPLEMENTS
    # =========================================================================
    "Functional Drinks": [
        ("Celsius", "celsius.com", "Fitness Energy", "Clinically proven to boost metabolism; went from $0 to $1.3B revenue; 'Live Fit' positioning; #3 energy drink brand"),
        ("Alani Nu", "alaninu.com", "Women's Energy", "Katy Hearn fitness influencer brand; energy drinks + supps; $200M+ revenue; targets women (unique in energy market)"),
        ("Ghost Lifestyle", "ghostlifestyle.com", "Licensed Flavors", "Licensed flavor collabs — Sour Patch Kids, Warheads, Oreo protein; transparent labels; lifestyle branding beats bro culture"),
        ("ZOA Energy", "zoaenergy.com", "Clean Energy", "Dwayne 'The Rock' Johnson's energy drink; natural caffeine from green tea + green coffee; 'healthy warrior'"),
        ("Gorilla Mind", "gorillamind.com", "Nootropics", "Derek (More Plates More Dates) built 3M+ YouTube audience then launched supps; brutally honest reviews; cult following"),
        ("Ryse Supplements", "rysesupps.com", "Licensed Flavors", "Kool-Aid, Ring Pop, Skippy licensed flavors for protein; 'fuel your greatness'; $200M+ revenue; flavor innovation"),
        ("1st Phorm", "1stphorm.com", "Premium Supps", "Andy Frisella's brand; Level 1 protein; massive Insta ambassador program; 'earn it' mentality; $500M+ revenue"),
        ("Liquid IV", "liquid-iv.com", "Hydration", "Cellular Transport Technology for 2x faster hydration; Unilever acquired; massive event sampling; packets everywhere"),
        ("LMNT", "drinklmnt.com", "Electrolytes", "No sugar, 1,000mg sodium; Robb Wolf co-founded; every big podcast sponsors it; keto/paleo/carnivore community favorite"),
        ("Bloom Nutrition", "bloomnu.com", "Greens Powder", "Mari Llewellyn's weight loss story drives the brand; greens powder = TikTok's #1 wellness product; $100M+ revenue"),
        ("AG1", "drinkag1.com", "Daily Greens", "75 ingredients in one scoop; sponsors Huberman, Lex Fridman, every major podcast; 'first thing every morning' ritual"),
        ("Cure Hydration", "curehydration.com", "Plant Electrolytes", "Coconut water-based ORS; WHO hydration formula; plant-based; Lauren Picasso Founder; clean label"),
    ],

    # =========================================================================
    # BAGS, WALLETS & CARRY
    # =========================================================================
    "Bags & Carry": [
        ("Bellroy", "bellroy.com", "Slim Wallets", "Slim your wallet — hidden features, pull tabs; Australian design; expanded to bags, phone cases, tech accessories"),
        ("Ridge Wallet", "ridgewallet.com", "Metal Wallets", "Titanium/aluminum RFID-blocking wallet; #1 podcast-advertised product; 'ditch the bulge'; $100M+ revenue; bootstrapped"),
        ("Ekster", "ekster.com", "Smart Wallets", "Solar-powered Chipolo tracker inside wallet; find via phone; Dutch design; $30M+ crowdfunded; Parliament wallet"),
        ("Dagne Dover", "dagnedover.com", "Neoprene Bags", "Neoprene bags with dedicated pockets for everything; Landon Carryall is iconic; functional beauty; $50M+ revenue"),
        ("Lo & Sons", "loandsons.com", "Travel Bags", "Designed by a mother-daughter-son trio; OG weekender bag; Pearl crossbody; engineering family background"),
        ("Portland Leather", "portlandleathergoods.com", "Leather Bags", "Full-grain leather bags at fair prices; Portland-made; 'Almost Perfect' sale events (minor imperfections, big discounts)"),
        ("Tumi", "tumi.com", "Luxury Travel", "Alpha Bravo line is the executive travel standard; ballistic nylon durability; lifetime guarantee; $850M revenue"),
        ("Osprey", "osprey.com", "Hiking Packs", "All Mighty Guarantee — repair any damage any time; best hiking/backpacking packs; Atmos 65 is legendary"),
        ("GORUCK", "goruck.com", "Rucksacks", "Military-grade rucksacks; GORUCK Challenge (12+ hour endurance events); Scars Lifetime Guarantee; veteran-founded"),
        ("Mystery Ranch", "mysteryranch.com", "Tactical Packs", "Used by US Special Forces, wildland firefighters, hunters; 3-ZIP design; Bozeman, Montana built; function-first"),
        ("Chrome Industries", "chromeindustries.com", "Urban Cycling", "Messenger bags with seatbelt buckle closure; waterproof; urban cyclist culture since '95; Portland + SF roots"),
        ("Tom Bihn", "tombihn.com", "Travel Bags", "Seattle-made; obsessive organization (multiple compartments with O-rings); 'made in USA' without the premium tax; cult following"),
    ],

    # =========================================================================
    # PARENTING & FAMILY
    # =========================================================================
    "Parenting & Family": [
        ("Elvie", "elvie.com", "Women's Health Tech", "Silent, wearable breast pump + pelvic floor trainer; in-bra pumping; $100M+ revenue; liberating pump experience"),
        ("Willow Pump", "willowpump.com", "Wearable Breast Pump", "In-bra pump with spill-proof bags; pump while doing anything; 2 patents; freedom for pumping moms"),
        ("Hatch", "hatch.co", "Baby Sound Machine", "Rest+ sound machine + night light + OK-to-wake alert; controls from phone; gorgeous design; $100M+ revenue"),
        ("Fridababy", "frida.com", "Baby Health", "NoseFrida snot sucker went viral; marketed with real-talk honesty about parenthood; expanded to Mom (Frida Mom)"),
        ("Nanit", "nanit.com", "Smart Monitor", "AI tracks baby's sleep quality, gives coaching tips; breathing motion monitoring; computer vision; $65M+ funded"),
        ("Snoo", "happiestbaby.com", "Smart Bassinet", "Dr. Harvey Karp's auto-rocking bassinet; $1,695 or $159/mo rental; used in 1,000+ hospitals; responsive soothing"),
        ("Ergobaby", "ergobaby.com", "Baby Carriers", "Omni 360 carrier supports multiple positions; hip-healthy certified; ergonomic for parent and baby; $100M+ revenue"),
        ("UPPAbaby", "uppababy.com", "Strollers", "VISTA stroller converts single to double; premium but worth it resale value; $1,000+ strollers that hold value"),
        ("Doona", "doona.com", "Car Seat Stroller", "Car seat with built-in wheels; unfolds to stroller in seconds; solved car seat transport; genius form factor"),
        ("Once Upon a Farm", "onceuponafarm.com", "Organic Baby Food", "Jennifer Garner co-founded; cold-pressed (never heated) baby food; fresh produce in recyclable pouches"),
        ("Cerebelly", "cerebelly.com", "Brain Baby Food", "Neurosurgeon-created; targets 16 key brain nutrients for development; science-backed baby food; unique positioning"),
        ("Lovevery", "lovevery.com", "Play Kits", "Montessori play kits matched to baby's developmental stage; backed by child development PhDs; $200M+ revenue"),
    ],

    # =========================================================================
    # EYEWEAR
    # =========================================================================
    "Eyewear": [
        ("Warby Parker", "warbyparker.com", "Eyewear", "Home Try-On (5 frames free); Buy a Pair Give a Pair; 230+ stores; $3B peak valuation; disrupted Luxottica monopoly"),
        ("Pair Eyewear", "paireyewear.com", "Customizable Glasses", "Base frames + magnetic toppers = new look daily; Marvel, Disney, NBA collabs; kids loved it first, then adults; $60M+ funded"),
        ("Zenni Optical", "zennioptical.com", "Affordable Glasses", "Prescription glasses from $6.95; direct from factory; 50M+ frames sold; destroyed the $300 glasses myth"),
        ("EyeBuyDirect", "eyebuydirect.com", "Budget Glasses", "Essilor-owned but offers glasses from $6; virtual try-on; 3,000+ styles; Buy 1 Give 1 program"),
        ("Felix Gray", "shopfelixgray.com", "Blue Light Glasses", "Pioneered blue light filtering glasses for screen workers; embedded filtering (not a coating); office essential"),
        ("DIFF Eyewear", "diffeyewear.com", "Charity Glasses", "Buy a pair, provide eye care for someone in need; affordable fashion frames; hand-finished acetate"),
        ("Blenders Eyewear", "blenderseyewear.com", "Lifestyle Sunglasses", "Chase Fisher started at SDSU; festival culture; polarized for $50; 'Live in Color' motto; $100M+ revenue"),
        ("Goodr", "goodr.com", "Running Sunglasses", "Polarized running sunglasses for $25; wild names ('Flamingos on a Booze Cruise'); no-slip, no-bounce; fun meets function"),
        ("SunGod", "sungod.co", "Custom Sunglasses", "Build-your-own sunglasses (color, lens, icons); lifetime warranty; Red Bull athlete partnerships; UK-based; premium sport"),
    ],

    # =========================================================================
    # SUSTAINABLE & ECO
    # =========================================================================
    "Sustainable & Eco Products": [
        ("Patagonia", "patagonia.com", "Eco Outdoor", "Donated the entire company ($3B) to fight climate change; 'Don't Buy This Jacket' ad; Worn Wear repair program"),
        ("Allbirds", "allbirds.com", "Eco Footwear", "Carbon footprint printed on every shoe; merino wool + sugarcane soles; Tim Brown + Joey Zwillinger; IPO'd"),
        ("Girlfriend Collective", "girlfriend.com", "Recycled Activewear", "Leggings from recycled water bottles; launched with free leggings (just pay shipping); 10K waitlist day one"),
        ("Pela Case", "pelacase.com", "Compostable Cases", "World's first compostable phone case; flax sheath material; breaks down in 6 months; Shark Tank; B Corp"),
        ("tentree", "tentree.com", "Eco Apparel", "Plants 10 trees for every item purchased; 80M+ trees planted; sustainable fabrics; token tracks your trees via QR code"),
        ("4ocean", "4ocean.com", "Ocean Cleanup", "Bracelet funds ocean cleanup; 30M+ lbs of trash removed; social enterprise model; tangible impact per purchase"),
        ("Who Gives A Crap", "whogivesacrap.org", "Eco Toilet Paper", "50% of profits build toilets in developing world; bamboo/recycled TP; fun wrapped rolls; $100M+ revenue; B Corp"),
        ("Stasher", "stasherbag.com", "Silicone Bags", "Platinum silicone bags replace single-use plastic; Shark Tank success; 200M+ single-use bags offset; B Corp"),
        ("Blueland", "blueland.com", "Refillable Cleaning", "Cleaning tablets + reusable bottles; eliminated 1B+ plastic bottles; Shark Tank; Kate Hudson invested"),
        ("Earth Breeze", "earthbreeze.com", "Eco Laundry", "Laundry sheets dissolve in water; no heavy jugs; carbon neutral; donate 10 loads for each package sold"),
        ("Ethique", "ethique.com", "Solid Bars", "Solid shampoo/conditioner bars from New Zealand; 30M+ plastic bottles prevented; concentrated = less shipping weight"),
        ("FinalStraw", "finalstraw.com", "Collapsible Straw", "Collapsible metal straw that fits on keychain; Kickstarter hit ($1.9M raised); 'the last straw you'll ever need'"),
    ],

    # =========================================================================
    # EMERGING DTC & TELEHEALTH
    # =========================================================================
    "Emerging DTC & Telehealth": [
        ("Hims & Hers", "hims.com", "Telehealth", "Destigmatized ED, hair loss, mental health; $1B+ revenue; 1.5M subscribers; pastel millennial branding for taboo topics"),
        ("Ro", "ro.co", "Digital Health", "Roman (ED) + Rory (women's) + Zero (smoking); $7B valuation; end-to-end care (diagnosis, delivery, support)"),
        ("Nurx", "nurx.com", "Women's Health", "Birth control, STI testing, migraine treatment online; no office visit needed; $500M+ valuation"),
        ("GoodRx", "goodrx.com", "Rx Savings", "Free app shows lowest drug prices; saved Americans $55B+; IPO'd; disrupted pharmacy pricing opacity"),
        ("Noom", "noom.com", "Behavior Change", "Psychology-based weight loss app using CBT; $400M+ revenue; anti-diet approach; 'stop dieting, start living'"),
        ("Calibrate", "joincalibrate.com", "Weight Health", "Metabolic health platform; GLP-1 access with metabolic reset program; combines medication + behavior change"),
        ("Found", "joinfound.com", "Weight Care", "Personalized weight care combining Rx, nutrition coaching, community; GLP-1 platform; board-certified doctors"),
        ("Curology", "curology.com", "Custom Skincare", "Online dermatology + custom prescription formula shipped monthly; superbottle with your unique ingredients"),
        ("Apostrophe", "apostrophe.com", "Online Derm", "Board-certified dermatologists prescribe custom treatments online; acne, anti-aging, rosacea; prescription skincare"),
        ("Tend", "hellotend.com", "Modern Dental", "Redesigned the dental experience; modern studios, transparent pricing, tech-forward; anti-dental-anxiety design"),
        ("Cerebral", "cerebral.com", "Mental Health", "Online psychiatry and therapy; medication management + counseling; $5B valuation; targets the therapy accessibility gap"),
        ("Arey Grey", "areygrey.com", "Gray Hair", "Supplements to reduce gray hair (melanin support); unique anti-gray angle; catalase and fo-ti root based"),
        ("Tabs Chocolate", "tabschocolate.com", "Functional Chocolate", "Dark chocolate with maca, epimedium for libido; viral TikTok ads; 'date night chocolate'; provocative marketing"),
        ("Agency", "agency.com", "Anti-Aging Skincare", "Tretinoin prescribed online; custom anti-aging formula from dermatologists; subscription retinoid; doctor-backed"),
    ],

    # =========================================================================
    # EXPANDED BEAUTY & COSMETICS
    # =========================================================================
    "Beauty - Expanded": [
        ("Laura Mercier", "lauramercier.com", "Luxury Makeup", "Translucent setting powder is the #1 face powder in the US; 'flawless face' technique inventor"),
        ("Too Faced", "toofaced.com", "Playful Cosmetics", "Better Than Sex mascara is best-selling prestige mascara; candy-themed packaging; Estée Lauder acquired for $1.45B"),
        ("Tarte Cosmetics", "tartecosmetics.com", "Eco-Chic Beauty", "Shape Tape concealer is #1 in US; Amazon rainforest ingredient sourcing; cruelty-free since 2000"),
        ("Urban Decay", "urbandecay.com", "Edgy Cosmetics", "Naked palette changed the eyeshadow game; 'beauty with an edge'; Wende Zomnir founded in a parking lot"),
        ("NARS", "narscosmetics.com", "Provocative Beauty", "Orgasm blush is the best-selling blush of all time; provocative product names; François Nars' artistic vision"),
        ("MAC Cosmetics", "maccosmetics.com", "Pro Makeup", "Viva Glam fund raised $500M+ for AIDS; inclusive shade range pioneer; backstage favorite for 40 years"),
        ("NYX Professional", "nyxcosmetics.com", "Affordable Pro", "Professional quality at drugstore prices; Epic Ink liner rivals high-end; L'Oréal acquired for $500M+"),
        ("Honest Beauty", "honestbeauty.com", "Clean Beauty", "Jessica Alba's clean beauty line; EWG verified; accessible price point; Honest Company umbrella"),
        ("HAUS LABS", "hauslabs.com", "Supercharged Makeup", "Lady Gaga's brand; relaunched with 'supercharged clean artistry'; performance + clean ingredients"),
        ("Florence by Mills", "florencebymills.com", "Teen Beauty", "Millie Bobby Brown's skincare/beauty brand for Gen Z; clean, vegan, cruelty-free; target audience 13-20"),
        ("r.e.m. beauty", "rembeauty.com", "Celestial Beauty", "Ariana Grande's brand; space/dream-inspired aesthetic; affordable ($15-25); available at Ulta"),
        ("Selena Gomez", "rarebeauty.com", "Inclusive Beauty", "Rare Beauty Soft Pinch blush went viral; Rare Impact Fund donates 1% of all sales to mental health"),
        ("KVD Beauty", "kvdveganbeauty.com", "Vegan Makeup", "100% vegan and cruelty-free; Tattoo Liner is cult; Good Apple skin-perfecting balm; bold artistry"),
        ("Makeup by Mario", "makeupbymario.com", "MUA Brand", "Mario Dedivanovic (Kim K's longtime MUA) built artistry-focused brand; MasterClass approach to makeup"),
        ("Westman Atelier", "westman-atelier.com", "Luxury Clean", "Gucci Westman (celeb MUA) brand; clean luxury cosmetics; baby skin finish; $68 foundation stick"),
        ("Kjaer Weis", "kjaerweis.com", "Luxury Organic", "Refillable luxury cosmetics; Danish organic certified; metal compacts designed to keep; sustainability + luxury"),
        ("RMS Beauty", "rmsbeauty.com", "Organic Makeup", "Rose-Marie Swift founded at 60; living luminizer is cult; raw coconut oil base; 'un-cover up' concealer"),
        ("Hourglass", "hourglasscosmetics.com", "Luxury Vegan", "Going 100% vegan; Veil Mineral Primer; Ambient Lighting Powder creates 'photoshop in a compact' effect"),
        ("IT Cosmetics", "itcosmetics.com", "Confidence Beauty", "CC+ Cream is #1 CC cream in US; 'developed with plastic surgeons'; L'Oréal acquired for $1.2B"),
        ("Benefit Cosmetics", "benefitcosmetics.com", "Fun Beauty", "BADgal mascara and Brow Zings pioneered the brow category; fun retro-kitschy branding; brow bars in stores"),
        ("Anastasia Beverly Hills", "anastasiabeverlyhills.com", "Brow Expert", "Created the brow category — Brow Wiz, Dipbrow Pomade; Anastasia Soare's 'Golden Ratio' brow method"),
        ("Bobbi Brown", "bobbibrowncosmetics.com", "Natural Makeup", "Founded on 10 brown-based lipsticks; 'be who you are' philosophy; sold to Estée Lauder then started Jones Road"),
        ("La Mer", "lamer.com", "Ultra Luxury", "Miracle Broth™ from sea kelp; $190/oz moisturizer; NASA physicist invented it to heal burns; ultimate luxury status"),
        ("SK-II", "sk-ii.com", "Luxury Skincare", "Pitera essence discovered when sake brewers had young-looking hands; $100+ treatment essence; P&G owned"),
        ("Olay", "olay.com", "Mass Skincare", "Regenerist Retinol24 outperformed $400 creams in testing; AI Skin Advisor tool; mass market meets clinical results"),
        ("CeraVe", "cerave.com", "Derm Skincare", "Developed with dermatologists; 3 essential ceramides; $15 moisturizer recommended by every derm on TikTok; L'Oréal owned"),
        ("La Roche-Posay", "laroche-posay.us", "Pharmacy Skincare", "French pharmacy skincare; thermal spring water; Cicaplast Balm B5 is a cult multi-use repair cream"),
        ("Clinique", "clinique.com", "Allergy Tested", "First dermatologist-created prestige beauty brand (1968); 3-Step system; allergy tested, 100% fragrance free"),
        ("Estée Lauder", "esteelauder.com", "Prestige Beauty", "Advanced Night Repair serum launched in 1982, still a bestseller; $15B+ conglomerate; the OG prestige brand"),
        ("Lancôme", "lancome-usa.com", "French Luxury", "Lash Idôle mascara; Teint Idole foundation matches 60+ shades; L'Oréal's luxury crown jewel"),
        ("Chanel Beauty", "chanel.com", "Haute Beauty", "No.5 is the world's most famous fragrance; Les Beiges is coveted; luxury marketing = aspiration engine"),
        ("Dior Beauty", "dior.com", "Luxury Cosmetics", "Lip Glow is a $40 tinted balm bestseller; Backstage collection; J'Adore fragrance is iconic"),
        ("YSL Beauty", "yslbeauty.com", "Luxury Cosmetics", "Touche Éclat pen invented the 'highlighter concealer' category in 1992; Rouge Pur Couture lipstick"),
        ("Tom Ford Beauty", "tomford.com", "Designer Beauty", "Lost Cherry, Tobacco Vanille are cult fragrances; $300+ lipsticks; ultra-luxury positioning"),
        ("Givenchy Beauty", "givenchybeauty.com", "Haute Couture Beauty", "Le Rouge lipstick with leather case; Prisme Libre powder; high fashion in compact form"),
    ],

    # =========================================================================
    # EXPANDED FASHION
    # =========================================================================
    "Fashion - Expanded": [
        ("Abercrombie & Fitch", "abercrombie.com", "Elevated Casual", "Massive comeback — from toxic brand to TikTok darling; YPB (Your Personal Best) activewear line"),
        ("J.Crew", "jcrew.com", "Preppy American", "Ludlow suit democratized tailoring; Jenna Lyons era made it cool; Vintage Wash tees are staple basics"),
        ("Madewell", "madewell.com", "Denim & Basics", "The Perfect Vintage Jean is the #1 selling jean at Madewell; denim recycling program; separated from J.Crew"),
        ("Gap", "gap.com", "American Basics", "Yeezy Gap collab created cultural frenzy; $15B+ revenue; khakis and pocket tees defined '90s casual"),
        ("H&M", "hm.com", "Fast Fashion", "Conscious Collection uses sustainable materials; designer collabs (Karl Lagerfeld, Balmain) crash websites"),
        ("Zara", "zara.com", "Fast Fashion", "Fastest design-to-store cycle (2 weeks); $30B+ revenue; no traditional advertising — stores are the marketing"),
        ("Uniqlo", "uniqlo.com", "Japanese Basics", "Heattech and AIRism fabrics are technology innovations in basics; $21B+ revenue; 'LifeWear' philosophy"),
        ("Massimo Dutti", "massimodutti.com", "Elevated Basics", "Inditex's premium line; Mediterranean sophistication; quality materials at accessible luxury price"),
        ("& Other Stories", "stories.com", "Fashion Forward", "H&M Group's creative brand; each collection designed by 3 ateliers (Paris, Stockholm, LA)"),
        ("Mango", "mango.com", "Mediterranean Fashion", "Spanish brand with Mediterranean flair; $3B+ revenue; strong in European markets; fashion + value"),
        ("Stradivarius", "stradivarius.com", "Youth Fashion", "Inditex's youngest brand; trend-focused for 18-25 demographic; TikTok-friendly pricing"),
        ("Bershka", "bershka.com", "Streetwear Fast Fashion", "Inditex brand targeting teens with streetwear + club fashion; strong in Europe and Latin America"),
        ("Pull&Bear", "pullandbear.com", "Casual Youth", "Laid-back California-meets-Mediterranean aesthetic; surf + skate influence; Inditex owned"),
        ("ASOS", "asos.com", "Online Fashion", "890+ brands + own label; ASOS Design fills trends in 2-3 weeks; 26M+ active customers; pioneered 'try at home' returns"),
        ("Boohoo", "boohoo.com", "Ultra-Fast Fashion", "Tests 100+ new styles daily in micro-batches; data-driven design; £1.7B+ revenue; controversial labor practices exposed"),
        ("PLT (PrettyLittleThing)", "prettylittlething.com", "Trend Fashion", "Umar Kamani built from Manchester; Molly-Mae Hague partnership; £500M+ revenue; 20M+ social followers"),
        ("Revolve", "revolve.com", "Influencer Fashion", "Pioneered influencer marketing in fashion; Revolve Festival rivals Coachella; $1B+ revenue; 5K+ influencer army"),
        ("Veronica Beard", "veronicabeard.com", "Power Dressing", "Dickey jacket — wear one blazer 10 ways with interchangeable inserts; $200M+ revenue; working women's favorite"),
        ("Vince", "vince.com", "Quiet Luxury", "Cashmere sweaters and leather jackets; 'quiet luxury' before the term existed; LA-based understated elegance"),
        ("Theory", "theory.com", "Modern Workwear", "Stretch wool pants changed office dressing; 'perfect shirt, perfect pant' philosophy; minimalist professional"),
        ("Eileen Fisher", "eileenfisher.com", "Sustainable Fashion", "Circular design pioneer — takes back old garments, resells or recycles; $500M+ revenue; Renew program"),
        ("Cos", "cosstores.com", "Scandinavian Minimal", "H&M Group's elevated brand; architectural stores; minimal aesthetic; quality basics with design edge"),
        ("Acne Studios", "acnestudios.com", "Swedish Fashion", "Face-logo beanie became streetwear icon; Musubi bag; Stockholm creative collective turned global brand"),
        ("Ganni", "ganni.com", "Scandi It-Girl", "'Scandi 2.0' — not minimal, but colorful and fun; #GanniGirls community; leopard print and smiley face motifs"),
        ("Totême", "toteme-studio.com", "Quiet Luxury", "The Scarf Coat went viral; monogram silk scarf; Swedish quiet luxury; $100M+ revenue; Elin Kling founded"),
        ("Anine Bing", "aninebing.com", "Scandi-LA", "Vintage-inspired rock & roll meets Scandinavian minimalism; graphic tees and blazers; $100M+ revenue"),
        ("Nanushka", "nanushka.com", "Vegan Leather", "Hungarian brand pioneered chic vegan leather; puffer jackets and trench coats in plant-based materials"),
        ("Self-Portrait", "self-portrait-studio.com", "Occasion Wear", "Han Chong made $300-500 cocktail dresses accessible; lace and structured designs for every event"),
        ("Zimmermann", "zimmermann.com", "Australian Resort", "Resort wear that became year-round luxury; runway shows in Paris; $2B+ valuation; acquired by Advent"),
        ("Aje", "ajeworld.com.au", "Australian Fashion", "Sculptural designs inspired by Australian landscape; art meets fashion; growing global DTC presence"),
        ("Sir", "sfrirthelabel.com", "Australian Minimal", "Effortless Australian linen and silk; Sophie Coote's timeless wardrobe; resort-ready minimalism"),
        ("Palm Angels", "palmangels.com", "Italian Streetwear", "Francesco Ragazzi's vision: LA skate culture through Italian fashion lens; Kering-backed; $200M+ revenue"),
        ("Amiri", "amiri.com", "Luxury Streetwear", "Mike Amiri's hand-distressed jeans ($1,000+); Hollywood and music industry cult brand; $100M+ revenue"),
        ("Rhude", "rhude.com", "Luxury Streetwear", "Rhuigi Villaseñor designed from his apartment; basketball shorts + blazer aesthetic; Puma creative director"),
        ("Casablanca", "casfrablancaparis.com", "New Luxury", "Charaf Tajer's colorful luxury sportswear; tennis-inspired silk sets; New Balance collabs; Paris-meets-Morocco"),
    ],

    # =========================================================================
    # EXPANDED FOOD & SNACKS
    # =========================================================================
    "Food & Snacks - Expanded": [
        ("Whisps", "whisps.com", "Cheese Crisps", "100% cheese crisps — just baked cheese; keto/low-carb snack pioneer; $100M+ revenue; Costco bestseller"),
        ("Wilde Chips", "wildechips.com", "Protein Chips", "Chicken breast chips with 10g protein; Mark Cuban invested on Shark Tank; high protein snack category creator"),
        ("Catalina Crunch", "catalinacrunch.com", "Keto Cereal", "Keto-friendly cereal that actually crunches; 0g sugar; became #1 keto cereal; $100M+ revenue"),
        ("HighKey", "highkeysnacks.com", "Keto Snacks", "Keto cookies, cereal, crackers; massive Facebook ad spend; targets keto dieters with familiar snack formats"),
        ("Simple Mills", "simplemills.com", "Clean Baking", "Almond flour crackers and baking mixes; 6 simple ingredients; $150M+ revenue; clean label pioneer"),
        ("That's It", "thatsitfruit.com", "Fruit Bars", "2 ingredients — just fruit; simplest label in the bar aisle; no added sugar, no concentrates; radical simplicity"),
        ("Hippeas", "hippeas.com", "Chickpea Puffs", "Organic chickpea puffs; Leonardo DiCaprio invested; B Corp; 'peas, love, and happiness' branding"),
        ("Skinny Dipped", "getskinnydipped.com", "Coated Nuts", "Thinly coated dark chocolate almonds; less coating = less sugar; founded by two women; Hershey acquired"),
        ("NotCo", "notco.com", "AI Plant Food", "Uses AI (Giuseppe) to analyze molecular structure of animal products and recreate with plants; $1.5B valuation"),
        ("Impossible Foods", "impossiblefoods.com", "Plant Meat", "Heme molecule makes it 'bleed'; Impossible Whopper at Burger King; $7B valuation; Stanford biochemistry lab origin"),
        ("Beyond Meat", "beyondmeat.com", "Plant Meat", "IPO'd at $3.8B; partnerships with McDonald's, KFC; pea protein-based; visible at every grocery store"),
        ("Oatly", "oatly.com", "Oat Milk", "Punk marketing in Sweden; 'It's like milk but made for humans'; barista edition created oat latte craze; $10B IPO"),
        ("Chobani", "chobani.com", "Yogurt", "Turkish immigrant Hamdi Ulukaya bought a closed yogurt factory for $700K; now $2B+ revenue; gave 10% equity to employees"),
        ("Siggi's", "siggis.com", "Icelandic Yogurt", "Simple ingredients, less sugar than most yogurts; Icelandic skyr tradition; 'simple ingredients, not a lot of sugar'"),
        ("Perfect Day", "perfectday.com", "Animal-Free Dairy", "Makes real dairy proteins via fermentation (no cows); whey protein identical to milk's; $1.6B valuation"),
        ("Miyoko's", "miyokos.com", "Vegan Cheese", "Plant-based butter and cheese that chefs actually use; European-style cultured vegan butter; founder is a cookbook author"),
        ("Kite Hill", "kitehill.com", "Plant Dairy", "Almond-based ricotta, cream cheese, yogurt; chef-crafted plant dairy; general Mills invested"),
        ("Minor Figures", "minorfigures.com", "Oat Milk", "Oat milk for specialty coffee; flat white focused; B Corp; London design aesthetic; carbon neutral"),
        ("Elmhurst 1925", "elmhurst1925.com", "Plant Milk", "Former dairy that pivoted to plant milk; HydroRelease™ extracts more nutrition from nuts; 98-year dairy heritage"),
        ("Graze", "graze.com", "Healthy Snacks", "UK snack box pioneer; personalized snack selection; Unilever acquired; data-driven flavor development"),
        ("Banza", "eatbanza.com", "Chickpea Pasta", "Pasta made from chickpeas — 2x protein, 4x fiber; $100M+ revenue; became #3 pasta brand in US; simple swap"),
        ("Caulipower", "eatcaulipower.com", "Cauliflower Products", "Cauliflower pizza crust became #1 in frozen pizza; founder's sons had celiac; gluten-free that tastes good"),
        ("Kodiak Cakes", "kodiakcakes.com", "Protein Pancakes", "Whole grain + protein pancake/waffle mix; bear mascot; $200M+ revenue; turned down Shark Tank deals"),
        ("RXBar", "rxbar.com", "Whole Food Bars", "'No B.S.' — ingredient list IS the label design; Kellogg's acquired for $600M; egg white protein bars"),
        ("Larabar", "larabar.com", "Fruit & Nut Bars", "Minimal ingredients (dates + nuts); each bar is just 2-9 ingredients; General Mills owned; OG clean bar"),
        ("KIND Bars", "kindsnacks.com", "Nut Bars", "Ingredients you can see and pronounce; transparent wrapper shows whole nuts; $5B acquisition by Mars"),
        ("Clif Bar", "clifbar.com", "Energy Bars", "Named after founder's father Clifford; organic ingredients; employee-owned until Mondelēz acquired for $2.9B"),
        ("LÄRABAR", "larabar.com", "Real Food Bars", "Fruit and nut bars with 2-9 ingredients; no added sugar in original line; simple food philosophy"),
        ("Dave's Killer Bread", "daveskillerbread.com", "Organic Bread", "Founded by an ex-convict; organic whole grain bread; 'Second Chance Employment' — 1/3 of workforce has criminal backgrounds"),
        ("Bobs Red Mill", "bobsredmill.com", "Whole Grains", "Employee-owned; 400+ whole grain products; Bob Moore gave the company to workers on his 81st birthday"),
    ],

    # =========================================================================
    # EXPANDED HOME & FURNITURE
    # =========================================================================
    "Home & Furniture - Expanded": [
        ("Pottery Barn", "potterybarn.com", "Classic Home", "Williams-Sonoma owned; $3B+ revenue; registry staple; Friends TV set recreated their aesthetic nationwide"),
        ("West Elm", "westelm.com", "Modern Home", "Fair Trade commitment; mid-century modern; $1.8B revenue; local artisan collabs; sustainability + design"),
        ("Crate & Barrel", "crateandbarrel.com", "Home Furnishing", "European-inspired modern furniture and decor; wedding registry leader; CB2 is their contemporary sub-brand"),
        ("Restoration Hardware", "rh.com", "Luxury Home", "Gallery format showrooms; membership model; $3.6B revenue; Cloud Sofa became aspirational furniture icon"),
        ("Wayfair", "wayfair.com", "Online Furniture", "30M+ products; 3D room planning tools; $12B+ revenue; the Amazon of furniture and home goods"),
        ("IKEA", "ikea.com", "Democratic Design", "Flat-pack pioneer; Billy bookcase sold 110M+ units; meatballs are strategic (keeps you in store); $50B revenue"),
        ("Havenly", "havenly.com", "Interior Design", "Affordable online interior design service; $79 for a room design; democratized interior designers"),
        ("Joybird", "joybird.com", "Custom Furniture", "Custom upholstery in 100+ fabrics; La-Z-Boy acquired; 365-day home trial; made custom furniture accessible"),
        ("AllModern", "allmodern.com", "Modern Furniture", "Wayfair's design-forward brand; free 2-day shipping; targets design-conscious millennials"),
        ("Castlery", "castlery.com", "Direct Furniture", "Singapore-founded; direct-from-factory furniture; fast delivery for custom pieces; expanding rapidly in US"),
        ("Albany Park", "albanypark.com", "Modular Sofas", "Founded by Iraqi-American in Chicago; cloud-like sectionals; $999 sofa disrupts $3K competitors; diverse-founded"),
        ("Inside Weather", "insideweather.com", "Configurable Furniture", "Configure your sofa piece by piece; fast 2-week delivery; made in California; $1,000-2,000 sweet spot"),
        ("Maiden Home", "maidenhome.com", "Handcrafted Furniture", "Handmade in North Carolina; luxury quality at 50% of retail; no showroom overhead; direct-to-consumer craftsmanship"),
        ("Lulu and Georgia", "luluandgeorgia.com", "Designer Home", "Chris Loves Julia and Sarah Sherman Samuel designer collabs; boho-chic rugs and furniture; $100M+ revenue"),
        ("Serena & Lily", "serenaandlily.com", "Coastal Home", "Riviera counter stool is iconic; coastal California aesthetic; $400M+ revenue; resort-inspired home"),
        ("Studio McGee", "studiomcgee.com", "TV Design Brand", "Netflix 'Dream Home Makeover'; Target collection; $200M+ revenue; Shea McGee's Instagram design inspiration"),
        ("McGee & Co", "mcgeeandco.com", "Design Shop", "Studio McGee's retail arm; curated home decor; lighting, art, furniture; brings their show aesthetic home"),
        ("Loom & Leaf", "loomandleaf.com", "Luxury Mattress", "Saatva's luxury sub-brand; gel-infused memory foam; eco-friendly materials; premium sleep without markup"),
        ("Nectar Sleep", "nectarsleep.com", "Mattress", "365-night trial (longest in industry) + forever warranty; memory foam; $1B+ revenue; aggressive Google ads"),
        ("DreamCloud", "dreamcloudsleep.com", "Hybrid Mattress", "Luxury hybrid mattress at $800 (vs $2,000+ competitors); cashmere-blend cover; Resident Home brand"),
        ("Layla Sleep", "laylasleep.com", "Flippable Mattress", "Copper-infused mattress that flips between soft and firm sides; one mattress, two comfort levels; unique value prop"),
        ("Avocado Green Mattress", "avocadogreenmattress.com", "Organic Mattress", "GOTS organic certified; handmade in LA; B Corp; most certified green mattress; $1,300 for true organic"),
        ("Tuft & Needle", "tuftandneedle.com", "Mattress", "Two software engineers disrupted mattresses; T&N Adaptive Foam; merged with Serta Simmons; started the DTC mattress wave"),
        ("Endy", "endy.com", "Canadian Mattress", "Canada's favorite online mattress; open-air cell foam; 100-night trial; $100M+ revenue; patriotic Canadian marketing"),
        ("Casper", "casper.com", "Mattress Pioneer", "The OG DTC mattress; subway ads everywhere; 100-night trial pioneer; IPO'd; showed every DTC brand the playbook"),
    ],

    # =========================================================================
    # EXPANDED PET
    # =========================================================================
    "Pet - Expanded": [
        ("Chewy", "chewy.com", "Online Pet Store", "Handwritten holiday cards to every customer; sends flowers when a pet dies; $11B+ revenue; legendary customer love"),
        ("Wag!", "wagwalking.com", "Dog Walking", "Uber for dog walking; GPS-tracked walks; premium services; Olivia Munn invested; $650M+ valuation"),
        ("Rover", "rover.com", "Pet Sitting", "Marketplace for pet sitters and dog walkers; 5-star matching; insurance included; $2B+ revenue post-IPO"),
        ("Furbo", "shopus.furbo.com", "Dog Camera", "Tosses treats and sends barking alerts; 2-way audio; lets you check on dog from work; 360° tracking"),
        ("Wisdom Panel", "wisdompanel.com", "Pet DNA", "Mars Petcare brand; screens for 350+ breeds and 200+ health conditions; most comprehensive pet DNA"),
        ("Wagmo", "wagmo.io", "Pet Insurance", "Wellness plans that cover routine vet visits (not just emergencies); instant claims via photo; modern pet insurance"),
        ("Lemonade Pet", "lemonade.com/pet", "Pet Insurance", "AI-powered claims processing; pays claims in seconds; giveback program donates unclaimed premiums"),
        ("Tails.com", "tails.com", "Custom Dog Food", "Personalized kibble recipe for each dog based on profile; UK-based; Nestlé Purina acquired"),
        ("Yummers", "yummers.com", "Pet Meal Toppers", "Freeze-dried meal toppers and mix-ins; makes regular kibble exciting; 'sprinkle some yum' concept"),
        ("PawHutt", "pawhut.com", "Pet Furniture", "Stylish pet beds, crates, and houses that match home decor; Amazon bestseller; function meets aesthetics"),
        ("ManyPets", "manypets.com", "Pet Insurance", "Formerly Bought By Many; uses data to create better pet insurance products; 500K+ policies"),
        ("Pumpkin", "pumpkin.care", "Pet Insurance", "90% reimbursement; covers hereditary conditions from day one; preventive care add-on; tech-forward claims"),
        ("Spot Pet Insurance", "spotpetins.com", "Pet Insurance", "Customizable deductible and reimbursement; 30-day money back; covers alternative therapies"),
        ("Healthy Paws", "healthypawspetinsurance.com", "Pet Insurance", "#1 rated pet insurance; no caps on claims; simple plan structure; Aon backed"),
        ("Ollie", "myollie.com", "Fresh Dog Food", "Custom meal plans by vet nutritionists; human-grade recipes; pre-portioned for each dog; $100M+ revenue"),
    ],

    # =========================================================================
    # EXPANDED TECH & GADGETS
    # =========================================================================
    "Tech & Gadgets - Expanded": [
        ("Apple", "apple.com", "Consumer Tech", "$3T+ company; iPhone, MacBook, AirPods; privacy-first positioning; ecosystem lock-in; design religion"),
        ("Samsung", "samsung.com", "Consumer Electronics", "Galaxy S + Z Fold; AMOLED display leader; $200B+ revenue; foldable phone pioneer; vertical integration"),
        ("Bose", "bose.com", "Premium Audio", "QuietComfort noise cancelling headphones set the standard; Bose 700; Wave radio; acoustics research pioneer"),
        ("Sony", "sony.com", "Electronics", "WH-1000XM5 headphones are best-selling premium ANC; PlayStation; Alpha cameras; entertainment empire"),
        ("Sennheiser", "sennheiser.com", "Audiophile Audio", "German audio perfection; HD 600 headphones are reference standard; 80+ year heritage; Momentum series"),
        ("Audio-Technica", "audio-technica.com", "Pro Audio", "AT2020 microphone is the entry-level studio standard; turntable revival leader; M50x headphones are industry workhorse"),
        ("Shure", "shure.com", "Microphones", "SM7B is THE podcast/streaming microphone; used by every major podcaster; $100M+ in the mic alone"),
        ("Blue Microphones", "bluemic.com", "USB Mics", "Yeti USB mic democratized podcasting/streaming; $130 mic used by millions; Logitech acquired"),
        ("HyperX", "hyperxgaming.com", "Gaming Audio", "Cloud II gaming headset is best-selling for 8+ years running; $100 sweet spot; HP owned"),
        ("JBL", "jbl.com", "Audio", "Charge speaker is the portable Bluetooth king; $5B+ Harman parent; Flip, Pulse, PartyBox range"),
        ("Marshall", "marshallheadphones.com", "Rock Audio", "Guitar amp heritage turned lifestyle audio; iconic vinyl & brass design; speakers look like stage amps"),
        ("Bang & Olufsen", "bang-olufsen.com", "Luxury Audio", "Danish luxury since 1925; Beoplay, Beosound, Beolab; $1,000+ speakers as art objects; aluminum craftsmanship"),
        ("Focal", "focal.com", "French Audio", "Utopia headphones ($4,000) are world's best; beryllium drivers; made in France; audiophile summit"),
        ("Sonos", "sonos.com", "Smart Audio", "Multi-room wireless system; simple app control; $1.5B+ revenue; won patent case against Google"),
        ("Tile", "thetileapp.com", "Item Tracker", "Pioneered Bluetooth item tracking before Apple AirTag; tile in wallet, keys, luggage; 40M+ sold"),
        ("Chipolo", "chipolo.net", "Item Finder", "European alternative to AirTag/Tile; ONE Spot works with Apple Find My; card-shaped wallet tracker"),
        ("Mophie", "mophie.com", "Portable Power", "Juice Pack battery case pioneered phone charging cases; MagSafe accessories; ZAGG Brands owned"),
        ("Anker", "anker.com", "Charging", "Amazon-first brand became $2B+ powerhouse; PowerCore, Soundcore, eufy; trust built on 100K+ 5-star reviews"),
        ("Belkin", "belkin.com", "Tech Accessories", "BoostCharge; MagSafe ecosystem partner; 40+ years of tech accessories; Foxconn owned"),
        ("CalDigit", "caldigit.com", "Thunderbolt Docks", "TS4 Thunderbolt dock is best for Mac creators; 18 ports; built for video editors and designers"),
        ("Satechi", "satechi.net", "USB-C Hub", "Aluminum USB-C hubs and accessories that match MacBook aesthetic; affordable Apple companion accessories"),
        ("Logitech", "logitech.com", "Peripherals", "MX Master mouse is the productivity gold standard; $4.5B+ revenue; webcams, keyboards, mice ecosystem"),
        ("SteelSeries", "steelseries.com", "Gaming Peripherals", "Arctis Nova Pro headset; Aerox mouse; used by esports pros; GG Engine software; Danish design"),
        ("Corsair", "corsair.com", "PC Gaming", "Full PC gaming ecosystem; iCUE RGB control; $1.8B+ revenue; acquired Elgato, SCUF, Fanatec"),
        ("NZXT", "nzxt.com", "PC Gaming", "Sleek PC cases (H510 is iconic); BLD custom PC service; Kraken AIO coolers; clean aesthetic in gaming"),
    ],

    # =========================================================================
    # EXPANDED SUPPLEMENTS
    # =========================================================================
    "Supplements - Expanded": [
        ("Optimum Nutrition", "optimumnutrition.com", "Gold Standard", "Gold Standard Whey is #1 selling protein worldwide for 20+ years; trusted by bodybuilders since the '80s"),
        ("Muscletech", "muscletech.com", "Sports Nutrition", "Nitro-Tech protein; research-backed formulas; 20+ years in sports nutrition; Iovate Health Sciences"),
        ("BSN", "bsnusa.com", "Sports Nutrition", "Syntha-6 protein tastes like a milkshake; N.O.-Xplode pre-workout pioneer; flavor-first approach"),
        ("Cellucor", "cellucor.com", "Pre-Workout", "C4 is the #1 pre-workout brand globally; Starburst and Skittles flavor collabs; $500M+ revenue"),
        ("PE Science", "pescience.com", "Select Protein", "Select Protein blend (whey + casein) for best taste and sustained release; Cake Pop flavor went viral"),
        ("Kaged", "kaged.com", "Clean Supps", "Pre-Kaged, In-Kaged, Re-Kaged — the workout trifecta; Informed Sport certified; Kris Gethin brand"),
        ("Legion Athletics", "legionathletics.com", "Evidence-Based", "Mike Matthews (Bigger Leaner Stronger author) brand; every ingredient has clinical dosing; no proprietary blends"),
        ("Jacked Factory", "jackedfactory.com", "Value Supps", "Nitrosurge pre-workout; Canadian brand; clinical doses at budget prices; Amazon bestseller"),
        ("Raw Nutrition", "getrawnutrition.com", "Hardcore Supps", "Chris Bumstead (Mr. Olympia) is co-owner; CBUM line; premium positioning; bodybuilding community"),
        ("Obvi", "myobvi.com", "Women's Collagen", "Collagen Protein designed for women; Super Collagen Protein in flavors like Fruity Cereal; Ronak Shah built via TikTok"),
        ("Beam", "beamorganics.com", "Sleep & Recovery", "Dream Powder sleep supplement went viral on TikTok; nano CBD + reishi + magnesium; hot cocoa format"),
        ("Ancient Nutrition", "ancientnutrition.com", "Multi Collagen", "Dr. Josh Axe brand; multi-collagen protein from 5 sources; bone broth protein; $100M+ revenue"),
        ("Vital Proteins", "vitalproteins.com", "Collagen", "Jennifer Aniston partnership made collagen peptides mainstream; blue tub icon; $500M+ revenue; Nestlé owned"),
        ("Garden of Life", "gardenoflife.com", "Organic Supps", "USDA Organic, Non-GMO verified everything; whole food vitamins; Dr. Formulated line; Nestlé Health Science"),
        ("Nordic Naturals", "nordicnaturals.com", "Fish Oil", "Gold standard omega-3 supplements; Arctic cod liver oil from Norway; purity testing exceeds pharmaceutical standards"),
        ("Nature Made", "naturemade.com", "Mass Vitamins", "#1 pharmacist-recommended vitamin brand; USP verified; $1B+ brand; trusted by medical professionals"),
        ("NOW Foods", "nowfoods.com", "Value Supps", "Family-owned since 1968; 1,400+ products; best value in supplements; GMP certified; $500M+ revenue"),
        ("Life Extension", "lifeextension.com", "Longevity Supps", "Backed by Life Extension Foundation's research; premium formulas; NAD+ and longevity focus; $300M+ revenue"),
        ("Jarrow Formulas", "jarrow.com", "Science Supps", "QH-absorb CoQ10 and Jarro-Dophilus probiotics; 44+ years of formulation expertise; clinical grade"),
        ("Doctor's Best", "dfroctorsbest.com", "Clinical Supps", "Science-based nutrition; transparently sourced; 200+ products; healthcare professional recommended"),
    ],

    # =========================================================================
    # EXPANDED JEWELRY & WATCHES
    # =========================================================================
    "Jewelry & Watches - Expanded": [
        ("Tiffany & Co.", "tiffany.com", "Luxury Jewelry", "Little Blue Box is the most recognized packaging in luxury; Breakfast at Tiffany's cultural icon; LVMH acquired for $15.8B"),
        ("Cartier", "cartier.com", "Haute Joaillerie", "Love Bracelet requires a special screwdriver to open; $6B+ revenue; the jeweler of kings, king of jewelers"),
        ("David Yurman", "davidyurman.com", "Cable Jewelry", "Cable bracelet is instantly recognizable; American luxury; $700M+ revenue; bridging fine and fashion jewelry"),
        ("Pandora", "pandora.com", "Charm Jewelry", "Charm bracelet system — collect charms for milestones; $4B+ revenue; lab-grown diamonds now; customization"),
        ("Kendra Scott", "kendrascott.com", "Fashion Jewelry", "Color Bar lets you design custom jewelry in-store; $1B+ valuation; Austin, TX pride; accessible luxury"),
        ("Swarovski", "swarovski.com", "Crystal Jewelry", "Crystal cutting perfection since 1895; Octea watches; $3B+ revenue; redefined as fashion-forward under Giovanna Engelbert"),
        ("James Allen", "jamesallen.com", "Online Diamonds", "360° HD diamond imaging — see every diamond in detail before buying; $500M+ revenue; Signet Jewelers acquired"),
        ("Blue Nile", "bluenile.com", "Diamond Jewelry", "The OG online diamond retailer (1999); saved buyers 20-40% vs retail; $500M+ revenue; Signet acquired"),
        ("Clean Origin", "cleanorigin.com", "Lab Diamonds", "100% lab-grown diamonds at 20-40% less than mined; eco-friendly; conflict-free; modern bridal"),
        ("With Clarity", "withclarity.com", "Engagement Rings", "AI-powered ring recommender; home preview kit (replica ring before buying); personalized experience"),
        ("Dorsey", "bydorsey.com", "Lab-Grown Gems", "Lab-grown colored gemstones in modern settings; Meg Strachan founded; affordable luxury gems ($200-500)"),
        ("Jenny Bird", "jenny-bird.com", "Statement Jewelry", "Toronto-designed statement jewelry; $60-200 price; gold-dipped with intention; accessible Canadian luxury"),
        ("Vitaly", "vitalydesign.com", "Modern Accessories", "Stainless steel, PVD-coated jewelry for men; streetwear-adjacent; recycled materials; sustainable metals"),
        ("Miansai", "miansai.com", "Men's Jewelry", "Hook bracelet became a men's style essential; Miami-based; sterling silver and gold; $40M+ revenue"),
        ("John Hardy", "johnhardy.com", "Artisan Jewelry", "Hand-woven silver chain from Bali; 250+ artisans; sustainable luxury; reclaimed metals and ethically sourced gems"),
        ("Foundrae", "foundrae.com", "Symbol Jewelry", "Personal symbolism — each piece carries meaning (strength, resilience, protection); Beth Bugdaycay's 'modern heirloom'"),
        ("Spinelli Kilcollin", "spinfrellikifrcollin.com", "Linked Rings", "Interlocking ring system as signature design; Drake, Rihanna wear them; LA-based; innovative ring concept"),
        ("Bvlgari", "bulgari.com", "Italian Luxury", "Serpenti collection — snake motif since 1940s; $2.5B+ revenue; LVMH owned; high jewelry + timepieces"),
        ("Chopard", "chopard.com", "Swiss Luxury", "Happy Diamonds (floating diamonds between sapphire crystals); ethical gold commitment; Cannes partner"),
        ("IWC Schaffhausen", "iwc.com", "Swiss Watches", "Pilot's watches and Portugieser collection; engineering excellence; 'made by men, for men' heritage; Richemont"),
    ],

    # =========================================================================
    # KIDS & TOYS
    # =========================================================================
    "Kids & Toys": [
        ("Lego", "lego.com", "Building Toys", "Most powerful brand in the world; $9B+ revenue; Lego Ideas community; adult AFOL segment growing 300%+"),
        ("Magna-Tiles", "magnatiles.com", "Magnetic Tiles", "Original magnetic building tiles for STEM play; used in every preschool; Valtech acquired; 3-99 age range"),
        ("Playfoam", "educationalinsights.com", "Sensory Play", "Non-toxic, non-drying modeling compound; sensory play; never dries out; classroom staple"),
        ("Melissa & Doug", "melissaanddoug.com", "Wooden Toys", "Premium wooden toys and puzzles; screen-free play advocates; $500M+ revenue; educational play"),
        ("Tonies", "tonies.com", "Audio Storytelling", "Screen-free audio player for kids; collectible figurines play stories/songs when placed on Toniebox; $300M+ revenue"),
        ("Yoto", "yotoplay.com", "Kids Audio", "Screen-free audio player; physical cards = stories, music, podcasts; parental controls; UK-based; $50M+ revenue"),
        ("Osmo", "playosmo.com", "Interactive Learning", "iPad + physical pieces = interactive play; AI camera reads physical objects; STEM, coding, art, math games"),
        ("Tegu", "tfregu.com", "Magnetic Wood Blocks", "Magnetic wooden building blocks; made in Honduras creating local jobs; beautiful design; premium wooden toys"),
        ("Green Toys", "greentoys.com", "Recycled Toys", "Toys made from 100% recycled milk jugs; made in USA; no BPA, PVC, phthalates; eco-conscious parents"),
        ("Fat Brain Toys", "fatbraintoys.com", "STEM Toys", "Curated educational toys; many proprietary designs; won 100+ toy awards; indie toy discovery platform"),
        ("Manhattan Toy", "manhattantoy.com", "Baby Toys", "Winkel rattle is the most gifted baby toy; 45+ years of innovative soft toys and developmental play"),
        ("Lovevery", "lovevery.com", "Play Kits", "Stage-based play kits with Montessori-informed toys; $200M+ revenue; subscription model; neuroscience-backed"),
        ("KiwiCo", "kiwico.com", "STEM Subscription", "Age-matched STEM crate subscriptions (Panda, Koala, Kiwi, Doodle, Tinker, Maker, Atlas); 40M+ crates shipped"),
        ("Little Passports", "littlepassports.com", "Geography Kits", "Monthly subscription exploring world cultures and geography; Science, USA, World editions; adventure learning"),
        ("Nugget", "nuggetcomfort.com", "Play Couch", "Foam play couch with cult following; 100K+ person waitlist; limited color drops; $100M+ revenue; parent community"),
    ],

    # =========================================================================
    # BATH & BODY
    # =========================================================================
    "Bath & Body": [
        ("Bath & Body Works", "bathandbodyworks.com", "Bath & Body", "Japanese Cherry Blossom and A Thousand Wishes are cult scents; $7B+ revenue; candle empire; seasonal drops"),
        ("Lush", "lush.com", "Fresh Handmade", "Every product made by hand with use-by dates; zero waste packaging; bath bombs invented here; ethical buying"),
        ("The Body Shop", "thebodyshop.com", "Ethical Beauty", "Anita Roddick pioneered ethical beauty; Community Trade program; no animal testing since 1989; activist brand"),
        ("Method", "methodhome.com", "Design Cleaning", "Design-forward cleaning products; bottle by Karim Rashid; plant-based; proved cleaning products can be beautiful"),
        ("Dr. Bronner's", "drbronner.com", "Castile Soap", "18-in-1 castile soap; eccentric philosophical labels; $200M+ revenue; fair trade, organic; family drama saga"),
        ("Dove", "dove.com", "Real Beauty", "Real Beauty campaign (2004) changed beauty advertising forever; most awarded campaign; $5B+ brand"),
        ("Dial", "dialsoap.com", "Antibacterial", "America's #1 antibacterial soap; DIA brand; value positioning; trusted for 75+ years"),
        ("Irish Spring", "irishspring.com", "Deodorant Soap", "Distinctive green bar with 'clean feeling'; masculine freshness; $500M+ brand"),
        ("Cetaphil", "cetaphil.com", "Gentle Skincare", "Dermatologist-recommended gentle cleanser since 1947; the prescription for sensitive skin; Galderma owned"),
        ("Eucerin", "eucerin.com", "Medical Skincare", "Original Healing Cream since 1900; recommended by dermatologists; Beiersdorf owned; clinical but accessible"),
        ("Aveeno", "aveeno.com", "Oat Skincare", "Active Naturals oat formula; 'naturally beautiful results'; Jennifer Aniston ambassador; J&J owned"),
        ("Nivea", "nivea.com", "Skincare", "Blue tin moisturizer is the world's most recognized skincare; 100+ years; $5B+ revenue; Beiersdorf's crown jewel"),
        ("Bioderma", "bioderma.com", "French Derm", "Sensibio H2O micellar water — the backstage essential for every MUA and model; French pharmacy classic"),
        ("Vichy", "vichy.com", "Mineral Skincare", "Volcanic water from Auvergne, France; Mineral 89 serum; L'Oréal's French pharmacy powerhouse"),
        ("Kiehl's", "kiehls.com", "Apothecary", "NYC apothecary since 1851; samples-first marketing; Ultra Facial Cream; L'Oréal owned; heritage storytelling"),
    ],

    # =========================================================================
    # COFFEE & TEA
    # =========================================================================
    "Coffee & Tea - Expanded": [
        ("Blue Bottle Coffee", "bluebottlecoffee.com", "Specialty Coffee", "Peaked within 48 hours of roasting; minimalist cafes; Nestlé acquired for $700M; James Freeman's craft coffee vision"),
        ("Counter Culture Coffee", "counterculturecoffee.com", "Direct Trade Coffee", "Direct trade pioneer; transparency report published annually; roaster of the year multiple times; Durham, NC"),
        ("Stumptown Coffee", "stumptowncoffee.com", "Portland Coffee", "Cold Brew in a stubby bottle created the RTD cold brew category; Portland culture; Peet's acquired"),
        ("Intelligentsia", "intelligentsia.com", "Third Wave Coffee", "Coined 'direct trade'; pioneered relationships with farmers; Chicago/LA; part of the third wave trinity"),
        ("Verve Coffee", "vervecoffee.com", "Santa Cruz Coffee", "Farmlevel sourcing; roasted in Santa Cruz; beautiful minimalist packaging; subscription model"),
        ("Onyx Coffee", "onyxcoffeelab.com", "Competition Coffee", "Won every major coffee competition; transparent pricing shows farmer payment; Arkansas-based; quality obsessed"),
        ("Trade Coffee", "drinktrade.com", "Coffee Matching", "Taste quiz matches you with coffee from 55+ roasters; personalized subscription; supports indie roasters"),
        ("Atlas Coffee Club", "atlascoffeeclub.com", "World Coffee", "Single-origin beans from different country each month; postcards with origin story; $100M+ in subscriptions"),
        ("Cometeer", "cometeer.com", "Flash-Frozen Coffee", "Specialty coffee brewed then flash-frozen in capsules; just add water; no machine needed; space-tech approach"),
        ("Jot", "jot.co", "Ultra Coffee", "20x concentrated liquid coffee; one squeeze = one cup; no machine, no filters, no waste; 5-second coffee"),
        ("Chamberlain Coffee", "chamberlaincoffee.com", "Creator Coffee", "Emma Chamberlain built from 12M YouTube subscribers; Gen Z coffee culture; cute packaging; grocery + DTC"),
        ("Death Wish Coffee", "deathwishcoffee.com", "Strong Coffee", "World's strongest coffee; won free Super Bowl ad ($5M value); skull branding; 2x caffeine of average coffee"),
        ("Bones Coffee", "bonescoffee.com", "Flavored Coffee", "Wild flavors (S'Mores, Maple Bacon, Strawberry Cheesecake); skull logo; flavor-forward; $50M+ revenue"),
        ("Nguyen Coffee Supply", "nguyencoffeesupply.com", "Vietnamese Coffee", "First specialty Vietnamese coffee company in US; Sahra Nguyen spotlights robusta beans; cultural pride brand"),
        ("Firebelly Tea", "firebellytea.com", "Premium Tea", "Beautiful tea packaging; loose leaf + sachets; transparent sourcing; made tea modern and giftable"),
        ("Vahdam Teas", "vahdamteas.com", "Indian Tea", "Garden-fresh Indian tea shipped within 72 hours of harvest; B Corp; carbon neutral; Oprah's Favorite Things"),
        ("DAVIDsTEA", "davidstea.com", "Tea Retail", "Canadian tea retailer with 150+ loose-leaf teas; seasonal collections; birthday tea tradition; $200M+ revenue"),
        ("Rishi Tea", "rishi-tea.com", "Organic Tea", "Direct trade organic tea; botanical blends; first to bring Japanese matcha to US wholesale; Milwaukee-based"),
        ("Art of Tea", "artoftea.com", "Artisan Tea", "Master tea blender; custom blending for restaurants and hotels; hand-crafted small batches; LA-based"),
        ("Dona Chai", "donachai.com", "Chai Concentrate", "Small-batch masala chai; used in specialty coffee shops nationwide; Brooklyn-made; spiced concentrate"),
    ],

    # =========================================================================
    # HOME TEXTILES
    # =========================================================================
    "Home Textiles": [
        ("Brooklinen", "brooklinen.com", "Sheets", "Kickstarter to $200M+ revenue; luxury sheets at fair price; 'Really Good' tagline; Brooklyn-born"),
        ("Parachute", "parachutehome.com", "Bedding", "Venice Beach luxury linens; expanded to mattress, rugs, baby; $100M+ revenue; LA lifestyle brand"),
        ("Boll & Branch", "bollandbranch.com", "Organic Bedding", "First Fair Trade Certified sheets; used in the White House; $100M+ revenue; transparency in textile sourcing"),
        ("Coyuchi", "coyuchi.com", "Organic Home", "100% organic cotton since 1991; 'Take Back' program for recycling old linens; B Corp; 30+ year commitment"),
        ("Cozy Earth", "cozyearth.com", "Bamboo Bedding", "Oprah's Favorite Things 5 years running; bamboo viscose; temperature-regulating; premium but validated"),
        ("Ettitude", "ettitude.com", "Bamboo Sheets", "CleanBamboo fabric is proprietary; 50% less water than cotton; hypoallergenic; Australian-founded"),
        ("Sijo", "sijohome.com", "Eucalyptus Bedding", "TENCEL eucalyptus sheets; French Linen; temperature-regulating; $100-200 price range; sleep-focused"),
        ("Sunday Citizen", "sundaycitizen.co", "Cozy Home", "Crystal-infused weighted blanket; Snug comforter; cozy-first philosophy; Instagram-worthy textures"),
        ("Quince", "onequince.com", "Factory-Direct", "$50 European linen sheets (vs $200+ elsewhere); factory-direct pricing; organic Turkish cotton towels"),
        ("Snowe Home", "snowehome.com", "Essential Home", "Complete home essentials — sheets, towels, dinnerware; simple, high-quality; no unnecessary SKUs; curated basics"),
        ("Riley Home", "rileyhome.com", "Affordable Luxury", "Luxury bedding at accessible prices; sateen and percale; spa-quality towels under $50; no frills, just quality"),
        ("Matouk", "matouk.com", "American Luxury Linens", "Made in Massachusetts since 1929; monogramming heritage; hotel-grade luxury; last American luxury linen maker"),
        ("Kassatex", "kassatex.com", "Fine Bath", "Turkish cotton towels + organic cotton sheets; hotel partnerships worldwide; affordable luxury linens"),
        ("Turkish-T", "turkish-t.com", "Turkish Towels", "Handloomed Turkish towels that work as blankets, wraps, beach covers; multipurpose textile tradition"),
        ("Onsen", "onsentowel.com", "Waffle Towels", "Japanese waffle-weave towels; lightweight and fast-drying; inspired by Japanese bath culture; minimalist"),
        ("Weezie", "weezie.com", "Monogram Towels", "Southern charm meets luxury towels; custom monogramming; piped trim; $40M+ revenue; gifting powerhouse"),
    ],

    # =========================================================================
    # PERSONAL SAFETY & SELF DEFENSE
    # =========================================================================
    "Personal Safety": [
        ("Birdie", "mybirdie.com", "Personal Alarm", "130dB personal safety alarm + flashing light; designed for women; she's Birdie went viral on TikTok; $30 safety"),
        ("Invisawear", "invisawear.com", "Safety Jewelry", "Smart jewelry that sends emergency SOS with your location; looks like regular necklace/bracelet; stealth safety"),
        ("Sabre", "sabrered.com", "Pepper Spray", "#1 pepper spray brand trusted by police; campus safety programs; compact designs; $50M+ revenue"),
        ("Mace Brand", "mfracebrand.com", "Personal Defense", "Synonymous with pepper spray; 50+ year heritage; personal alarms, sprays, GPS trackers; trusted safety"),
        ("Ring", "ring.com", "Home Security", "Video doorbell + neighborhood security network; Amazon owned; Neighbors app creates community safety; $1.2B acquisition"),
        ("SimpliSafe", "simplisafe.com", "DIY Security", "No contract, DIY installation; $15/month monitoring vs $50+ from ADT; 4M+ customers; anti-big-security"),
        ("Arlo", "arlo.com", "Security Cameras", "Wire-free battery security cameras; best outdoor cameras; Apple HomeKit compatible; Verisure acquired"),
        ("Eufy", "eufylife.com", "Local Security", "Local storage (no cloud fees); 2K cameras for $40; privacy-first (no subscription); Anker brand"),
        ("Canary", "canary.is", "All-in-One Security", "Camera + siren + air quality monitor in one device; renter-friendly; no installation; built-in siren"),
        ("Abode", "goabode.com", "Smart Security", "Works with every smart home platform; pro monitoring optional; DIY-friendly; Thread/Matter support"),
    ],

    # =========================================================================
    # NICHE & SPECIALTY BRANDS
    # =========================================================================
    "Niche & Specialty": [
        ("Tushy", "hellotushy.com", "Bidet", "'Stop wiping your butt with dry paper'; $80 bidet attachment normalized bidets in US; humor marketing master"),
        ("Squatty Potty", "squattypotty.com", "Toilet Stool", "Unicorn pooping rainbow ice cream video got 300M+ views; $30 toilet stool; Shark Tank; $50M+ revenue"),
        ("Litter-Robot", "litter-robot.com", "Self-Cleaning Litter", "$700 self-cleaning litter box; WiFi app control; 'never scoop again'; cult following; $100M+ revenue"),
        ("Purple", "purple.com", "Mattress", "Hyper-Elastic Polymer grid; raw egg test video got 1B+ views; science + humor; unique material technology"),
        ("Casper", "casper.com", "Sleep", "The OG bed-in-a-box; 100-night trial; subway ads made DTC famous; IPO; showed every brand the playbook"),
        ("Peloton", "onepeloton.com", "Connected Fitness", "$44B peak valuation; instructor celebrities; pandemic growth; community around cycling; content is the product"),
        ("Rothy's", "rothys.com", "Recycled Shoes", "3D-knit shoes from recycled water bottles; machine washable; 150M+ bottles recycled; $1B+ valuation"),
        ("Allbirds", "allbirds.com", "Sustainable Shoes", "World's most comfortable shoes; carbon footprint label; Tim Cook wears them; merino + tree + sugarcane materials"),
        ("Away", "awaytravel.com", "Luggage", "The carry-on with a built-in battery; podcast and book content strategy; $1.4B valuation; travel lifestyle brand"),
        ("Warby Parker", "warbyparker.com", "Eyewear", "Home Try-On; Buy a Pair Give a Pair; 230+ stores; first DTC brand to IPO; disrupted $100B eyewear industry"),
        ("Dollar Shave Club", "dollarshaveclub.com", "Shaving", "'Our blades are f***ing great' video — 12K orders in 48 hours; Unilever acquired for $1B; DTC legend"),
        ("Glossier", "glossier.com", "Beauty", "From blog (Into The Gloss) to $1.8B brand; 'skin first, makeup second'; UGC-driven; Emily Weiss pioneer"),
        ("Harry's", "harrys.com", "Razors", "Bought a German factory; $1.37B acquisition attempt; razors + skincare; editorial content driven"),
        ("Manscaped", "manscaped.com", "Body Grooming", "Created below-the-belt grooming category; humor marketing; podcast king; $1B+ valuation"),
        ("Quip", "getquip.com", "Oral Care", "Subscription electric toothbrush; $15 refills every 3 months; sleek metal design; dentist-designed"),
        ("Hims & Hers", "hims.com", "Telehealth", "Destigmatized taboo health topics; $1B+ revenue; 1.5M subscribers; pastel branding meets ED/hair loss"),
        ("Canopy", "getcanopy.co", "Humidifiers", "Anti-mold, no-mist humidifier; dishwasher-safe parts; aroma diffusion add-ons; solved the ugly humidifier problem"),
        ("Dyson", "dyson.com", "Engineering", "5,127 prototypes before first vacuum; Airwrap ($600 hair tool) has year-long waitlists; $7.7B revenue"),
        ("Yeti", "yeti.com", "Premium Coolers", "Built brand on hunting/fishing, crossed into mainstream; $1.6B+ revenue; coolers as status symbols; 'Built for the Wild'"),
        ("Solo Stove", "solostove.com", "Smokeless Fire Pit", "Airflow engineering eliminates smoke; doubled as pandemic backyard essential; Solo Brands IPO'd; $400M+ revenue"),
    ],

    # =========================================================================
    # ADDITIONAL NICHE CATEGORIES
    # =========================================================================
    "Fragrance": [
        ("Le Labo", "lelabofragrances.com", "Artisan Fragrance", "Santal 33 became NYC's unofficial scent; each bottle hand-blended in store; Estée Lauder acquired"),
        ("Maison Margiela REPLICA", "maisonmargiela.com", "Memory Fragrance", "Each scent recreates a memory (By the Fireplace, Lazy Sunday Morning, Beach Walk); olfactory storytelling"),
        ("Byredo", "byredo.com", "Swedish Luxury", "Ben Gorham — Indian-Swedish background; 'beyond reason'; Gypsy Water, Mojave Ghost cult scents; $800M+ valuation"),
        ("Diptyque", "diptyqueparis.com", "French House", "Baies candle is world's most coveted; Philosykos, Tam Dao; artistic Parisian heritage since 1961"),
        ("Jo Malone", "jomalone.com", "London Fragrance", "Layering concept — combine scents; Wood Sage & Sea Salt; $700M+ revenue; Estée Lauder; gift box culture"),
        ("Aesop", "aesop.com", "Apothecary", "Tacit, Hwyl, Rozu fragrances; botanical formulations; each store is architecturally unique; L'Oréal acquired for $2.5B"),
        ("D.S. & Durga", "dsanddurga.com", "Indie Fragrance", "Brooklyn-based niche house; Bowmakers, I Don't Know What; music-inspired scent compositions"),
        ("Boy Smells", "boysmells.com", "Genderful", "'Genderful' fragrance positioning; started as candles; LGBTQ+ owned; Kacey Musgraves collab; inclusive luxury"),
        ("Clean Reserve", "cleanbeauty.com", "Clean Fragrance", "Certified clean fragrances; Skin, Sueded Oud, Rain; sustainability + clean ingredients in perfumery"),
        ("Juliette Has a Gun", "juliettehasagun.com", "Provocative Fragrance", "Not a Perfume (cetalox only) became viral; French niche house; playful product names; indie luxury"),
        ("Maison Francis Kurkdjian", "franciskurkdjian.com", "Haute Perfumery", "Baccarat Rouge 540 is the most-searched fragrance on TikTok; master perfumer; LVMH acquired"),
        ("Initio Parfums", "initioparfums.com", "Magnetic Fragrance", "Side Effect, Oud for Greatness; uses molecules designed to create attraction; chemistry-based seduction angle"),
        ("Parfums de Marly", "parffrumfrsdemfrarly.com", "French Heritage", "Layton and Delina are TikTok viral; inspired by 18th century French perfumery; luxury niche house"),
        ("Creed", "cfreedfrfragfrances.com", "Heritage Luxury", "Aventus is the #1 niche men's fragrance; 260+ year history; supplied Napoleon, Queen Victoria; $3.5B Kering acquisition"),
        ("Tom Ford", "tomford.com", "Designer Fragrance", "Lost Cherry, Tobacco Vanille, Black Orchid are cult classics; $300+ bottles; defined luxury unisex fragrance"),
    ],

    # =========================================================================
    # FINAL EXPANSION - MISCELLANEOUS DTC
    # =========================================================================
    "Additional DTC Brands": [
        ("Misfits Market", "misfitsmarket.com", "Imperfect Produce", "Delivers 'ugly' produce at 40% off; reduces food waste; subscription box; expanded to full grocery"),
        ("FarmFoods", "farmfoods.com", "Pasture-Raised Meat", "100% grass-fed beef and pastured pork delivered frozen; Texas-based; ranch-to-door; subscription model"),
        ("SnackCrate", "snackcrate.com", "World Snacks", "Snacks from a different country each month; explore world flavors; $100M+ revenue; discover global treats"),
        ("Uncommon Goods", "uncommongoods.com", "Unique Gifts", "Curated marketplace for creative, unique gifts; Brooklyn-based; many artist-made products; B Corp"),
        ("Grovemade", "grovemade.com", "Desk Accessories", "Portland-made walnut and maple desk accessories; desk shelf, laptop stand; craftsmanship meets modern tech"),
        ("Ugmonk", "ugmonk.com", "Analog System", "Analog task cards — physical productivity system (not digital); Jeff Sheldon's minimalist design philosophy"),
        ("Orbitkey", "orbitkey.com", "Key Organization", "Eliminated jangling keys with slim organizer; Nest hub; desk mat; Australian design; Kickstarter origin"),
        ("Ridge Wallet", "ridgewallet.com", "Metal Wallet", "#1 podcast-advertised product; titanium RFID-blocking; 'ditch the bulge' campaign; $100M+ revenue; bootstrapped"),
        ("Blueland", "blueland.com", "Eco Cleaning", "Cleaning tablets + reusable bottles; eliminated 1B+ plastic bottles; Shark Tank; Kate Hudson invested"),
        ("Public Goods", "publicgoods.com", "Essentials", "Membership for sustainable essentials; minimalist packaging; tree-free paper, organic food, clean care"),
        ("Prose", "prose.com", "Custom Hair", "AI-powered custom haircare formulas; adapts to weather and water hardness; made-to-order in Brooklyn"),
        ("Curology", "curology.com", "Custom Skin", "Online dermatology + custom prescription formula mailed monthly; superbottle with your unique formula"),
        ("Function of Beauty", "functionofbeauty.com", "Custom Care", "Personalized shampoo/conditioner/body wash; your name on the bottle; quiz-based formulation; $150M+ revenue"),
        ("Care/of", "takecareof.com", "Custom Vitamins", "Quiz-based vitamin packs with your name; morning routine simplified; Bayer acquired for $250M+"),
        ("Native", "nativecos.com", "Natural Deodorant", "First natural deodorant people actually switched to; P&G acquired for $100M; proved clean deodorant works"),
        ("Billie", "mybillie.com", "Women's Razors", "First to show body hair in razor ads; 'Project Body Hair' campaign; P&G acquired; $5 starter kit"),
        ("Hims", "forhims.com", "Men's Telehealth", "Made buying ED pills, hair loss treatment, anxiety meds stigma-free; $1B+ revenue; pink branding"),
        ("Ro", "ro.co", "Digital Health", "Roman (ED) + Rory (women's) + Zero (smoking) under one umbrella; $7B valuation; complete health platform"),
        ("GoodRx", "goodrx.com", "Rx Savings", "Free app shows drug prices at every pharmacy; saved Americans $55B+; IPO'd; pharmacy transparency pioneer"),
        ("Barkbox", "barkbox.com", "Dog Box", "Themed toys + treats monthly; Super Chewer for heavy chewers; 2M+ subscribers; BARK went public; dogs > everything"),
    ],
}


def get_all_brands():
    """Merge main and extra brand databases, deduplicate by brand name."""
    all_brands = dict(BRANDS)
    for extra in [EXTRA_BRANDS, EXTRA_BRANDS_2, EXTRA_BRANDS_3, FINAL_BRANDS, FINAL_BRANDS_2]:
        for category, brands in extra.items():
            if category in all_brands:
                all_brands[category].extend(brands)
            else:
                all_brands[category] = brands

    # Deduplicate by brand name (keep first occurrence)
    seen = set()
    deduped = {}
    for category, brands in all_brands.items():
        unique = []
        for brand in brands:
            name = brand[0].lower().strip()
            if name not in seen:
                seen.add(name)
                unique.append(brand)
        if unique:
            deduped[category] = unique
    return deduped


def get_meta_ads_library_url(brand_name: str) -> str:
    """Generate Meta Ads Library search URL for a brand."""
    encoded = quote(brand_name)
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={encoded}"


def generate_excel(output_path: str):
    """Generate the Excel file with all brand data."""
    ALL_BRANDS = get_all_brands()
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    link_font = Font(name="Calibri", color="0563C1", underline="single", size=10)
    data_font = Font(name="Calibri", size=10)
    insight_font = Font(name="Calibri", size=9, italic=True, color="333333")
    bold_font = Font(name="Calibri", bold=True, size=10)

    cat_colors = [
        "F2F7FB", "FFF8F0", "F0FFF0", "FFF0F5", "FFFFF0",
        "F0F8FF", "FFF8DC", "F5F0FF", "F0FFF4", "FFF5F5",
        "F0FFFF", "FFFAF0", "F8F4FF", "F5FFF5", "FFF0F0",
        "EBF5FB", "FDF5E6", "F0FFF4", "FFF5EE", "F8F8FF",
        "E8F8F5", "FEF9E7", "F4ECF7", "E8F6F3", "FDEDEC",
        "EAF2F8", "FAF0E6", "E8F5E9", "FCE4EC", "F3E5F5",
    ]

    # ========== SHEET: ALL BRANDS ==========
    ws = wb.active
    ws.title = "All Brands"
    headers = ["#", "Brand Name", "Website", "Category", "Sub-Niche",
               "Notable Insight / Marketing Angle", "Meta Ads Library"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.freeze_panes = "A2"
    # Add auto-filter
    ws.auto_filter.ref = f"A1:G1"

    row_num = 2
    brand_count = 0
    for cat_idx, (category, brands) in enumerate(ALL_BRANDS.items()):
        fill_color = cat_colors[cat_idx % len(cat_colors)]
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for brand_name, website, sub_niche, insight in brands:
            brand_count += 1
            meta_url = get_meta_ads_library_url(brand_name)
            website_url = f"https://{website}" if not website.startswith("http") else website

            # Col A: #
            c = ws.cell(row=row_num, column=1, value=brand_count)
            c.font = data_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Col B: Brand Name
            c = ws.cell(row=row_num, column=2, value=brand_name)
            c.font = bold_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(vertical="center")

            # Col C: Website
            c = ws.cell(row=row_num, column=3, value=website)
            c.hyperlink = website_url
            c.font = link_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(vertical="center")

            # Col D: Category
            c = ws.cell(row=row_num, column=4, value=category)
            c.font = data_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)

            # Col E: Sub-Niche
            c = ws.cell(row=row_num, column=5, value=sub_niche)
            c.font = data_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(vertical="center")

            # Col F: Notable Insight
            c = ws.cell(row=row_num, column=6, value=insight)
            c.font = insight_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)

            # Col G: Meta Ads Library
            c = ws.cell(row=row_num, column=7, value="View Ads")
            c.hyperlink = meta_url
            c.font = link_font; c.fill = row_fill; c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 14

    # ========== CATEGORY SHEETS ==========
    for cat_idx, (category, brands) in enumerate(ALL_BRANDS.items()):
        sheet_name = category[:31]
        ws_cat = wb.create_sheet(title=sheet_name)

        cat_headers = ["#", "Brand Name", "Website", "Sub-Niche",
                        "Notable Insight / Marketing Angle", "Meta Ads Library"]
        for col_idx, h in enumerate(cat_headers, 1):
            cell = ws_cat.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws_cat.freeze_panes = "A2"

        for idx, (brand_name, website, sub_niche, insight) in enumerate(brands, 1):
            r = idx + 1
            meta_url = get_meta_ads_library_url(brand_name)
            website_url = f"https://{website}" if not website.startswith("http") else website

            ws_cat.cell(row=r, column=1, value=idx).font = data_font
            ws_cat.cell(row=r, column=1).border = thin_border

            ws_cat.cell(row=r, column=2, value=brand_name).font = bold_font
            ws_cat.cell(row=r, column=2).border = thin_border

            c = ws_cat.cell(row=r, column=3, value=website)
            c.hyperlink = website_url; c.font = link_font; c.border = thin_border

            ws_cat.cell(row=r, column=4, value=sub_niche).font = data_font
            ws_cat.cell(row=r, column=4).border = thin_border

            c = ws_cat.cell(row=r, column=5, value=insight)
            c.font = insight_font; c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical="center")

            c = ws_cat.cell(row=r, column=6, value="View Ads")
            c.hyperlink = meta_url; c.font = link_font; c.border = thin_border
            c.alignment = Alignment(horizontal="center")

        ws_cat.column_dimensions["A"].width = 6
        ws_cat.column_dimensions["B"].width = 28
        ws_cat.column_dimensions["C"].width = 32
        ws_cat.column_dimensions["D"].width = 22
        ws_cat.column_dimensions["E"].width = 70
        ws_cat.column_dimensions["F"].width = 14

    # ========== SUMMARY SHEET ==========
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.cell(row=1, column=1, value="DTC Brand Research — 3,000+ Brands with Insights").font = Font(
        name="Calibri", bold=True, size=16, color="1F3864"
    )
    ws_summary.merge_cells("A1:D1")

    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Calibri", size=10, italic=True
    )
    ws_summary.cell(row=3, column=1, value=f"Total Brands: {brand_count}").font = Font(
        name="Calibri", bold=True, size=13, color="1F3864"
    )
    ws_summary.cell(row=4, column=1, value=f"Total Categories: {len(ALL_BRANDS)}").font = Font(
        name="Calibri", bold=True, size=12
    )

    ws_summary.cell(row=6, column=1, value="Category").font = header_font
    ws_summary.cell(row=6, column=1).fill = header_fill; ws_summary.cell(row=6, column=1).border = thin_border
    ws_summary.cell(row=6, column=2, value="Brand Count").font = header_font
    ws_summary.cell(row=6, column=2).fill = header_fill; ws_summary.cell(row=6, column=2).border = thin_border

    for idx, (category, brands) in enumerate(ALL_BRANDS.items()):
        r = 7 + idx
        ws_summary.cell(row=r, column=1, value=category).font = data_font
        ws_summary.cell(row=r, column=1).border = thin_border
        ws_summary.cell(row=r, column=2, value=len(brands)).font = data_font
        ws_summary.cell(row=r, column=2).border = thin_border
        ws_summary.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 15

    notes_row = 7 + len(ALL_BRANDS) + 2
    ws_summary.cell(row=notes_row, column=1, value="Column Guide:").font = Font(name="Calibri", bold=True, size=11)
    guide = [
        "- Brand Name: The DTC brand",
        "- Website: Clickable link to their ecommerce store",
        "- Category: Primary product category",
        "- Sub-Niche: Specific product focus within category",
        "- Notable Insight: Unique marketing angle, celebrity ties, viral moment, product innovation, or business model insight",
        "- Meta Ads Library: Clickable link to view their active Facebook/Instagram ads",
    ]
    for i, line in enumerate(guide):
        ws_summary.cell(row=notes_row + 1 + i, column=1, value=line).font = Font(name="Calibri", size=10)

    wb.save(output_path)
    return brand_count


if __name__ == "__main__":
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "research_outputs")
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"DTC_Brands_3000_With_Insights_{date_str}.xlsx"
    output_path = os.path.join(output_dir, filename)

    count = generate_excel(output_path)
    print(f"Exported {count} brands to: {output_path}")
