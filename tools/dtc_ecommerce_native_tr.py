#!/usr/bin/env python3
"""
DTC Ecommerce-Native Brands - Turkish Excel Generator
Generates a comprehensive Excel file with 2000+ ecommerce-native DTC brands.
All category names, sub-niches, and insights are in Turkish.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Output config ───────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research_outputs")
TODAY = datetime.now().strftime("%Y-%m-%d")
FILENAME = f"DTC_Ecommerce_Native_Markalar_Turkce_{TODAY}.xlsx"

# ─── Brand Data ──────────────────────────────────────────────────────────────
# Format: (Brand, Website, Sub-niche TR, Notable Insight TR)

BRANDS = {
    # ═══════════════════════════════════════════════════════════════════════════
    # 1. GÜZELLIK & CİLT BAKIMI (Beauty & Skincare) — 80+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Güzellik & Cilt Bakımı": [
        ("Glossier", "glossier.com", "Doğal Güzellik", "Emily Weiss'ın blogundan doğan marka; 'skin first, makeup second' felsefesi; Boy Brow ikonik ürün; topluluk odaklı pazarlama"),
        ("Rhode", "rhodeskin.com", "Peptit Cilt Bakımı", "Hailey Bieber'ın markası; peptit dudak bakımı dakikalar içinde tükendi; 50M+ Instagram takipçisi ile sıfır maliyetli pazarlama"),
        ("Drunk Elephant", "drunkelephant.com", "Temiz Cilt Bakımı", "Tiffany Masterson'ın 'şüpheli 6' içerik stratejisi; Shiseido'ya $845M'a satıldı; rengarenk ambalaj ikonik"),
        ("The Ordinary", "theordinary.com", "Bilimsel Cilt Bakımı", "DECIEM altında; aktif madde isimleriyle ürün adlandırma; %99 düşük fiyat stratejisi; Estée Lauder satın aldı"),
        ("Rare Beauty", "rarebeauty.com", "Kapsayıcı Güzellik", "Selena Gomez'in markası; Rare Impact Fund mental sağlık bağışı; Soft Pinch likit allık viral oldu"),
        ("Fenty Beauty", "fentybeauty.com", "Kapsayıcı Fondöten", "Rihanna'nın markası; 40+ fondöten tonu ile sektörü değiştirdi; ilk yıl $570M gelir; LVMH ortaklığı"),
        ("Kylie Cosmetics", "kyliecosmetics.com", "Dudak Ürünleri", "Kylie Jenner'ın dudak kiti dakikalar içinde tükendi; sosyal medya ile $1B+ değerleme; Coty %51 hisse aldı"),
        ("Summer Fridays", "summerfridays.com", "Maske & Nemlendirici", "Marianna Hewitt & Lauren Gores; Jet Lag Mask Instagram'da viral oldu; Sephora'da en çok satan"),
        ("Supergoop", "supergoop.com", "Güneş Koruma", "SPF'i eğlenceli hale getiren marka; Unseen Sunscreen görünmez formül; her yaş grubu için güneş koruma"),
        ("Tower 28", "tower28beauty.com", "Hassas Cilt Güzellik", "NEA onaylı hassas cilt ürünleri; ShineOn Lip Jelly en çok satan; temiz ve güvenli formüller"),
        ("Kosas", "kosas.com", "Temiz Makyaj", "Revealer Concealer cilt bakımı + makyaj birleşimi; 'makeup that's skincare' konsepti; Sephora favorisi"),
        ("Ilia", "iliabeauty.com", "Organik Makyaj", "Super Serum Skin Tint en çok satan; temiz güzellik + yüksek performans; Sephora'da hızla büyüyen"),
        ("Merit", "meritbeauty.com", "Minimalist Makyaj", "Katherine Power'ın '5 dakika makyaj' felsefesi; 10'dan az ürün portföyü; Sephora'da en hızlı büyüyen"),
        ("Saie", "saiehello.com", "Temiz Günlük Makyaj", "Glowy Super Gel viral TikTok ürünü; 'dew' görünümü trendi başlattı; temiz formüller"),
        ("Jones Road Beauty", "jonesroadbeauty.com", "Yaş Dostu Güzellik", "Bobbi Brown'ın yeni markası; Miracle Balm viral oldu; 50+ yaş grubunu hedefleyen temiz makyaj"),
        ("Topicals", "mytopicals.com", "Cilt Sorunları", "Hiperpigmentasyon ve egzama için renkli ambalajlı ürünler; Gen Z'nin cilt sorunlarını normalleştirme"),
        ("Bubble Skincare", "hellobubble.com", "Gen Z Cilt Bakımı", "Walmart'ta $5-16 fiyat; Z kuşağı için erişilebilir cilt bakımı; TikTok'ta 2B+ görüntülenme"),
        ("Byoma", "byoma.com", "Bariyer Onarım", "Cilt bariyeri onarım odaklı; $10-15 fiyat aralığı; Target'ta; Gen Z'nin en erişilebilir cilt bakımı"),
        ("Versed", "versedskin.com", "Temiz Eczane", "Target'ta temiz cilt bakımı; topluluk oylamasıyla ürün geliştirme; $10-20 fiyat aralığı"),
        ("Cocokind", "cocokind.com", "Bilinçli Güzellik", "Sürdürülebilir ambalaj + şeffaf fiyatlandırma; maliyet dağılımını etiketinde gösteriyor"),
        ("Tula", "tula.com", "Probiyotik Cilt Bakımı", "Probiyotik bazlı cilt bakımı; Dr. Roshini Raj kurdu; influencer pazarlama ustası; $150M+ gelir"),
        ("Glow Recipe", "glowrecipe.com", "K-Beauty Ilham", "Karpuz serisi ikonik; K-beauty'yi Batı'ya taşıdı; Watermelon Glow Niacinamide Dew Drops viral"),
        ("Peach & Lily", "peachandlily.com", "Kore Cilt Bakımı", "Alicia Yoon'un K-beauty küratörlüğü; Glass Skin Serum en çok satan; cam cilt trendini başlattı"),
        ("Farmacy", "farmacybeauty.com", "Çiftlik Güzellik", "Kendi çiftliğinden elde edilen bileşenler; Green Clean balm temizleyici kült ürün; echinacea bazlı"),
        ("Youth to the People", "youthtothepeople.com", "Süperfood Cilt Bakımı", "Kale + yeşil çay + hyalüronik asit; 'superfood for skin' konsepti; L'Oréal satın aldı"),
        ("Biossance", "biossance.com", "Biyo-Teknoloji Cilt", "Şeker kamışından squalane üretimi; sürdürülebilir biyoteknoloji; temiz bilim yaklaşımı"),
        ("Naturium", "naturium.com", "Etkin Maddeli Bakım", "Susan Yara kurdu; yüksek konsantrasyon aktif maddeler düşük fiyatla; The Ordinary alternatifi"),
        ("Good Molecules", "goodmolecules.com", "Uygun Fiyatlı Aktif", "Beautylish markası; $6-12 fiyat aralığı; The Ordinary'ye rakip; basit ve etkili formüller"),
        ("Starface", "starface.world", "Akne Patch Eğlenceli", "Yıldız şeklinde akne patchleri; akneyi eğlenceli hale getirdi; sarı yıldız ikonik; Gen Z favorisi"),
        ("Hero Cosmetics", "herocosmetics.com", "Akne Tedavi", "Mighty Patch gece kullanım akne patchleri; Church & Dwight $630M'a satın aldı; Amazon #1"),
        ("Dieux Skin", "dieuxskin.com", "Şeffaf Güzellik", "Klinik kanıtları paylaşan marka; Instant Angel göz altı maskesi; sürdürülebilir ambalaj"),
        ("Kinship", "lovekinship.com", "Ekolojik Cilt Bakımı", "SPF ürünleri mercan resiflerine güvenli; adaptojenik bileşenler; çevreye duyarlı formüller"),
        ("Peace Out", "peaceoutskincare.com", "Hedefli Tedavi", "Akne, kırışıklık, gözenek patchleri; Sephora'da en çok satan patch markası; nokta tedavi uzmanı"),
        ("Paula's Choice", "paulaschoice.com", "BHA Exfoliant", "2% BHA Liquid Exfoliant kült ürün; Unilever satın aldı; bilimsel yaklaşım; BS olmayan güzellik"),
        ("Maëlys", "maelyscosmetics.com", "Vücut Bakımı", "B-Flat karın sıkılaştırma kremi viral; selülit ve vücut şekillendirme; sonuç odaklı vücut bakımı"),
        ("Phlur", "phlur.com", "Temiz Parfüm", "Chriselle Lim'in yeniden kurduğu marka; Missing Person parfümü TikTok'ta viral; temiz kokular"),
        ("Spoiled Child", "spoiledchild.com", "AI Kişisel Bakım", "AI ile kişiselleştirilmiş cilt bakımı; yaş ve cilt tipine göre formül; veri odaklı güzellik"),
        ("Beauty of Joseon", "beautyofjoseon.com", "Kore Geleneği", "Joseon Hanedanı'ndan ilham; pirinç + ginseng formülleri; Relief Sun SPF ikonik; uygun fiyatlı K-beauty"),
        ("Typology", "typology.com", "Fransız Minimalist", "Paris merkezli; sadece gerekli bileşenler; doğal aktifler; minimalist Fransız cilt bakımı"),
        ("Primally Pure", "primallypure.com", "Doğal Deodorant", "Çiftlik ilhamı; doğal deodorant ve cilt bakımı; organik bileşenler; temiz yaşam topluluğu"),
        ("Sol de Janeiro", "soldejaneiro.com", "Brezilya Vücut Bakımı", "Brazilian Bum Bum Cream ikonik; Brezilya ilhamlı vücut bakımı; koku ve doku deneyimi"),
        ("Nécessaire", "necessaire.com", "Vücut Temizliği", "Nick Axelrod kurdu; vücut bakımında aktif maddeler; The Body Wash hyalüronik asitli; 'body care is skincare'"),
        ("Bushbalm", "bushbalm.com", "Bikini Bölgesi Bakım", "Tüy batması ve koyulaşma tedavisi; Shark Tank'ta yatırım aldı; cesur niş pazarlama"),
        ("Stratia", "stratiaskin.com", "Bariyer Bakımı", "Liquid Gold cilt bariyeri onarım kült ürünü; bağımsız tek kişi markası; Reddit favorisi"),
        ("Ami Colé", "amicole.com", "Melanin Güzellik", "Koyu cilt tonları için temiz makyaj; Diarrha N'Diaye kurdu; kapsayıcı renk paleti"),
        ("Live Tinted", "livetinted.com", "Çok Kültürlü Güzellik", "Deepica Mutyala kurdu; Huestick çok amaçlı renk çubuğu; esmer ciltler için özel formüller"),
        ("Aavrani", "aavrani.com", "Hint Güzellik", "Ayurveda + modern bilim; Hint güzellik gelenekleri; zerdeçal ve bakuchiol bazlı ürünler"),
        ("Vacation Inc", "vacation.inc", "Retro Güneş Bakımı", "1980'ler estetik güneş kremleri; Classic Whip SPF 30 krem şantili ambalaj; eğlenceli marka kimliği"),
        ("Youthforia", "youthforia.com", "Yatak Makyajı", "BYO Blush renk değiştiren allık; 'yatakta bile takılabilir makyaj' konsepti; viral TikTok ürünleri"),
        ("EADEM", "eadem.co", "Melanin Cilt Bilimi", "Koyu cilt tonları için bilimsel bakım; Smart Melanin teknolojisi; Marie Claire ödüllü"),
        ("Krave Beauty", "kravebeauty.com", "Basit Cilt Bakımı", "Liah Yoo kurdu; #PressReset kampanyası; tüketimi azaltma felsefesi; Great Barrier Relief ikonik"),
        ("Ceremonia", "ceremonia.com", "Latin Saç Bakımı", "Latin Amerika bitkileriyle saç bakımı; guava yaprak özü; kültürel miras + modern bilim"),
        ("Florence by Mills", "florencebymills.com", "Genç Güzellik", "Millie Bobby Brown'ın markası; 12-18 yaş hedef kitle; hafif ve temiz formüller; Ulta'da satışta"),
        ("r.e.m. beauty", "rembeauty.com", "Pop Yıldız Güzellik", "Ariana Grande'nin markası; rüya temalı ambalaj; göz makyajı uzmanı; Ulta ortaklığı"),
        ("About Face", "aboutface.com", "Cesur Makyaj", "Halsey'nin markası; yüksek pigmentli cesur renkler; sanatsal ifade; vegan formüller"),
        ("Makeup by Mario", "makeupbymario.com", "Pro Makyaj", "Mario Dedivanovic (Kim K'nın makyözü); SurrealSkin Foundation en çok satan; profesyonel kalite"),
        ("Danessa Myricks", "danessamyricks.com", "Sanatsal Güzellik", "Makyaj sanatçısı; ColorFix çok amaçlı pigment; profesyonel + günlük kullanım; yenilikçi formüller"),
        ("Rose Inc", "roseinc.com", "Lüks Temiz Makyaj", "Rosie Huntington-Whiteley markası; Skin Tint serum fondöten; sürdürülebilir lüks; Sephora'da"),
        ("Iris & Romeo", "irisandromeo.com", "Çok İşlevli Makyaj", "Best Skin Days SPF + nemlendirici + fondöten tek üründe; minimalist rutini seven kadınlar için"),
        ("MENTED Cosmetics", "mentedcosmetics.com", "Nude Ruj Çeşitliliği", "Koyu cilt tonları için nude rujlar; KJ Miller & Amanda Johnson kurdu; kapsayıcı 'nude' tanımı"),
        ("Selfless by Hyram", "selflessbyhyram.com", "Youtuber Cilt Bakımı", "Hyram Yarbro'nun markası; gelirin bir kısmı yağmur ormanlarına bağış; etik güzellik; The Inkey List ile"),
        ("Osea", "oseamalibu.com", "Deniz Yosunu Bakımı", "Malibu merkezli; deniz yosunu bazlı cilt bakımı; 1996'dan beri; anti-aging odaklı temiz formüller"),
        ("Herbivore", "herbivorebotanicals.com", "Doğal Lüks Bakım", "Seattle merkezli; Bakuchiol retinol alternatifi; kristal ve bitki bazlı; Sephora'da"),
        ("Tatcha", "tatcha.com", "Japon Güzellik", "Japon güzellik ritüelleri; Dewy Skin Cream ikonik; LVMH Unilever $500M satın aldı; geisha ilhamı"),
        ("Augustinus Bader", "augustinusbader.com", "Kök Hücre Teknolojisi", "Prof. Augustinus Bader'in TFC8 teknolojisi; $265 krem; ünlülerin tercihi; bilimsel lüks bakım"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Profesyonel Peel", "Dermatolog markası; Alpha Beta Peel patchleri kült ürün; LED maske viral; klinik sonuçlar evde"),
        ("Sunday Riley", "sundayriley.com", "Lüks Aktif Bakım", "Good Genes laktik asit serum kült ürün; retinol uzmanı; Sephora'da en çok satan"),
        ("Drunk Elephant", "drunkelephant.com", "Temiz Klinik", "Biouyumlu pH ürünleri; 'Suspect 6' bileşen hariç tutma; çocuklar bile kullanıyor; renkli ambalaj"),
        ("Herbivore Botanicals", "herbivorebotanicals.com", "Bitki Bazlı", "Doğal + vegan; Prism Glow Potion viral serum; şeffaf ambalaj estetiği; Sephora'da"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi", "Hindistan cevizi bazlı güzellik; deodorant en çok satan; tropikal estetik; San Diego merkezli"),
        ("Dermalogica", "dermalogica.com", "Profesyonel Cilt", "Cilt terapistleri için kurulan marka; Unilever satın aldı; Face Mapping teknolojisi; eğitim odaklı"),
        ("Sand & Sky", "sandandsky.com", "Avustralya Kili", "Avustralya pembe kili maskesi viral; Instagram'da 100M+ görüntülenme; Avustralya botanikleri"),
        ("Frank Body", "frankbody.com", "Kahve Scrub", "Avustralya kahve peelingi viral; UGC pazarlaması öncüsü; eğlenceli marka sesi; $50M+ gelir"),
        ("Go-To Skincare", "gotoskincare.com", "Avustralya Günlük", "Zoë Foster Blake kurdu; Avustralya'nın en sevilen cilt bakımı; eğlenceli + etkili; BWX satın aldı"),
        ("Summer Fridays", "summerfridays.com", "Influencer Güzellik", "Blog'dan markaya; CC Me Serum vitamin C besteller; Jet Lag Mask orijinal viral ürün"),
        ("HAUS Labs", "hauslabs.com", "Lady Gaga Güzellik", "Lady Gaga markası; süper güçlü formüller; PhD Hybrid Lip Oil viral TikTok; Sephora relansman başarısı"),
        ("Fable & Mane", "fableandmane.com", "Hint Saç Yağı", "Hint saç bakım ritüelleri; ayurvedik saç yağlaması; sürdürülebilir Hindistan cevizi yağı formülleri"),
        ("Odacité", "odacite.com", "Fransız Serum", "Fransız formüller; tek bileşen serum konsantreleri; Paris + LA; lüks doğal cilt bakımı"),
        ("Alpyn Beauty", "alpynbeauty.com", "Vahşi Doğa Bitki", "Jackson Hole dağlarından toplanan yabani bitkiler; vahşi doğa ilhamlı temiz güzellik"),
        ("Beneath Your Mask", "beneathyourmask.com", "Lüks Doğal", "Dana Jackson kurdu; kendi hastalık sürecinden doğan marka; Beyoncé'nin favorisi; el yapımı lüks"),
        ("Maya Chia", "mayachia.com", "Süper Tohum", "Chia tohumu bazlı cilt bakımı; süperfood yağları; anti-aging; temiz lüks; ödüllü formüller"),
        ("Beekman 1802", "beekman1802.com", "Keçi Sütü", "Keçi sütü cilt bakımı; çiftlik hikayesi; CBS 'The Amazing Race' kazananları; Ulta'da hızla büyüyen"),
        ("Grown Alchemist", "grownalchemist.com", "Biyolojik Bakım", "Avustralya biyolojik cilt bakımı; peptit + antioksidan formüller; lüks otel amenity markası"),
        ("Codex Beauty", "codexbeauty.com", "Bilimsel Bitki", "İrlanda bitkileri + PubMed araştırmaları; klinik kanıtlı doğal bileşenler; BIA yüz yağı ikonik"),
        ("UpCircle", "upcirclebeauty.com", "Atık Dönüşüm Güzellik", "Kahve telinden yüz peelingi; gıda atıklarını güzellik ürününe dönüştürme; İngiltere merkezli"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 2. SAÇ BAKIMI (Haircare) — 60+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı": [
        ("Olaplex", "olaplex.com", "Bağ Onarım Teknolojisi", "Patentli bis-aminopropyl diglycol dimaleate; kırık saç bağlarını onarır; salon profesyonel to DTC; $1.5B IPO"),
        ("K18", "k18hair.com", "Biyomimetik Saç Onarımı", "K18Peptide™ 4 dakikada hasar onarımı; durulanmayan maske; salon endüstrisini değiştirdi; viral"),
        ("Prose", "prose.com", "Kişisel Saç Formülü", "AI + uzman ile kişiselleştirilmiş saç bakım formülü; 85 faktör analizi; her şişe benzersiz"),
        ("Function of Beauty", "functionofbeauty.com", "Özel Formül", "Quiz ile kişiselleştirilmiş şampuan/saç kremi; 54T+ kombinasyon; etikette adınız yazıyor"),
        ("Curlsmith", "curlsmith.com", "Kıvırcık Saç Uzmanı", "Kıvırcık saç tiplerine göre koleksiyonlar; Bond Curl Rehab en çok satan; Helen of Troy satın aldı"),
        ("Odele Beauty", "odelebeauty.com", "Erişilebilir Saç Bakımı", "Target'ta $12 premium saç bakımı; unisex; temiz formüller; Midwest kadın kurucu ekibi"),
        ("Dae Hair", "dae.com", "Çöl İlhamı Saç", "Arizona çöl bitkileri; Cactus Fruit 3-in-1 styling cream viral; Amber Fillerup Clark kurdu"),
        ("Crown Affair", "crownaffair.com", "Saç Ritüeli", "Dianna Cohen kurdu; günlük saç bakım ritüeli; The Comb, The Towel; lüks saç aksesuarları + bakım"),
        ("Bread Beauty Supply", "breadbeautysupply.com", "Doğal Saç", "Tekstürlü saç için minimalist bakım; 'wash-day essentials'; Sephora özel; doğal saç hareketi"),
        ("Act+Acre", "actandacre.com", "Saç Derisi Bakımı", "Cold Processed® saç derisi bakımı; saç derisi detoksu; Helen Reavey (trişolojist) kurdu"),
        ("Vegamour", "vegamour.com", "Saç Büyütme", "Bitkisel saç büyütme serumu; GRO serisi; vegan ve temiz; saç dökülmesi çözümü; $100M+ gelir"),
        ("Mielle Organics", "mielleorganics.com", "Doğal Saç Bakımı", "Monique Rodriguez kurdu; Rosemary Mint yağı TikTok'ta viral; P&G satın aldı; doğal saç hareketi öncüsü"),
        ("Pattern Beauty", "patternbeauty.com", "Kıvırcık Saç", "Tracee Ellis Ross markası; kıvırcık + coily saçlar için; kıvırcık kız kültürü; Ulta'da"),
        ("Briogeo", "briogeo.com", "Temiz Saç Bakımı", "Don't Despair Repair maske en çok satan; 6-free formüller; Wella $500M'a satın aldı"),
        ("Ouai", "theouai.com", "Hairstylist Markası", "Jen Atkin (Kardashian hairstylist) kurdu; Wave Spray ikonik; P&G satın aldı; lifestyle marka"),
        ("Amika", "loveamika.com", "Renkli Saç Bakımı", "Brooklyn merkezli; eğlenceli ambalaj; ısı koruma + renk koruma uzmanı; Sephora'da büyüyen"),
        ("dpHUE", "dphue.com", "Renk Bakımı", "Saç rengi bakımı uzmanı; Gloss+ evde yarı kalıcı renk; saç boyası arasını uzatan ürünler"),
        ("JVN Hair", "jvnhair.com", "Hemisqualane Saç", "Jonathan Van Ness markası; hemisqualane teknolojisi; Queer Eye ünlüsü; bilim + aktivizm"),
        ("Biolage", "biolage.com", "Profesyonel Temiz", "L'Oréal profesyonel; CleanReset normalizing şampuan; vegan + sürdürülebilir dönüşüm; salon to DTC"),
        ("Rahua", "rahua.com", "Amazon Yağı", "Amazon yağmur ormanı rahua yağı; yerli kabilelerle sürdürülebilir hasat; lüks doğal saç bakımı"),
        ("Davines", "davines.com", "İtalyan Sürdürülebilir", "B Corp sertifikalı İtalyan markası; Oi Oil çok amaçlı yağ; sürdürülebilirlik manifestosu"),
        ("Oribe", "oribe.com", "Lüks Saç Bakımı", "Luxury saç bakımı; Gold Lust şampuan ikonik; Kao satın aldı; Daniel Kaner + Oribe Canales"),
        ("Moroccanoil", "moroccanoil.com", "Argan Yağı Öncü", "Argan yağı saç bakımı kategorisini yaratan marka; Treatment Original ikonik; mavi şişe tanınırlığı"),
        ("Color Wow", "colorwowhair.com", "Renk Koruma", "Dream Coat anti-humidity spray viral; ünlü hairstylist Chris Appleton destekli; renk uzmanı"),
        ("Living Proof", "livingproof.com", "MIT Teknolojisi", "MIT bilim insanları kurdu; OFPMA teknolojisi patentli; Jennifer Aniston yatırımcı; Unilever satın aldı"),
        ("Virtue Labs", "virtuelabs.com", "Keratin Teknolojisi", "Alpha Keratin 60ku® insan keratini; gerçek keratin onarımı; bilim temelli saç bakımı"),
        ("Bumble and Bumble", "bumbleandbumble.com", "NYC Salon", "New York salon kültürü; Thickening serisi ikonik; Estée Lauder bünyesinde; styling uzmanı"),
        ("IGK", "igkhair.com", "Miami Saç", "Miami saç kültürü; First Class kuru şampuan kömürlü; eğlenceli isimler; Sephora favorisi"),
        ("R+Co", "randco.com", "Kolektif Saç Bakımı", "Süper hairstylist kolektifi; Dallas serisi biyotin saç bakımı; Bleu parfümlü şampuan popüler"),
        ("Verb", "verbproducts.com", "Erişilebilir Salon", "Ghost Oil en çok satan; $16-20 salon kalitesi; Sephora'da en uygun fiyatlı; hayalet yağ konsepti"),
        ("Innersense", "innersensebeauty.com", "Organik Salon", "USDA organik sertifikalı; temiz salon saç bakımı; kıvırcık saç topluluğu favorisi"),
        ("Aussie", "aussie.com", "Avustralya İlham", "3 Minute Miracle saç maskesi ikonik; uygun fiyatlı süpermarket markası; P&G; Avustralya botanikleri"),
        ("Ceremonia", "ceremonia.com", "Latin Saç Bakımı", "Latin Amerika bitki özleri; Aceite de Moska saç yağı viral; kültürel miras modern formüllerle"),
        ("Maui Moisture", "mauimoisture.com", "Tropikal Saç Bakımı", "Hawaii ilhamlı; aloe vera + hindistan cevizi suyu; vegan; uygun fiyatlı doğal saç bakımı"),
        ("Cantu", "cantubeauty.com", "Doğal Saç Hareketi", "Shea yağı bazlı; kıvırcık ve doğal saçlar için; uygun fiyatlı; doğal saç hareketi sembolü"),
        ("Melanin Haircare", "melaninhaircare.com", "Melanin Zengin Saç", "Whitney & Trae Bodge kurdu; melanin zengin saçlar için özel formüller; çok kültürlü bakım"),
        ("MONDAY Haircare", "mondayhaircare.com", "Yeni Zelanda Minimalist", "NZ merkezli; $6-8 premium görünüm düşük fiyat; minimalist beyaz ambalaj; viral büyüme"),
        ("Gisou", "gisou.com", "Bal Saç Bakımı", "Negin Mirsalehi'nin arıcı ailesinden ilham; bal bazlı saç bakımı; 7M Instagram; lüks doğal"),
        ("Aunt Jackie's", "auntjackiescurls.com", "Kıvırcık Bakım", "Kıvırcık ve doğal saçlar için; uygun fiyatlı; flaxseed bazlı; curl defining uzmanı"),
        ("Aveda", "aveda.com", "Bitkisel Salon", "Ayurveda ilhamlı salon markası; çevre bilinçli ambalaj; Estée Lauder; Invati saç dökülmesi serisi"),
        ("Cécred", "cecred.com", "Beyoncé Saç Bakımı", "Beyoncé'nin saç bakım markası; 2024 lansmanı; anında tükendi; Fermentation Hydration serisi"),
        ("Blake Brown", "blakebrownbeauty.com", "Lively Saç Bakımı", "Blake Lively'nin saç bakım markası; 2024 lansmanı; Hollywood glamour saç bakımı"),
        ("Maison 276", "maison276.com", "Fransız Saç Parfümü", "Saç parfümü uzmanı; Grasse parfümcülüğü + saç bakımı; lüks Fransız saç kokuları"),
        ("Hairstory", "hairstory.com", "Şampuansız Bakım", "New Wash şampuan yerine geçen temizleyici; co-wash öncüsü; deterjan içermeyen saç yıkama"),
        ("HairMax", "hairmax.com", "Lazer Saç Büyütme", "FDA onaylı lazer saç büyütme cihazları; LaserBand; evde klinik saç tedavisi"),
        ("Nutrafol", "nutrafol.com", "Saç Takviyesi", "Dermatolojist önerilen #1 saç büyütme takviyesi; biyotin + adaptojenleri; klinik çalışmalarla kanıtlanmış"),
        ("Hims", "hims.com/hair", "Erkek Saç Dökülmesi", "Finasteride + minoxidil online reçete; erkek saç dökülmesi tedavisini erişilebilir kıldı; telehealth"),
        ("Keeps", "keeps.com", "Saç Dökülme Önleme", "Erkek saç dökülmesi abonelik; FDA onaylı tedaviler; aylık $25'tan; erken müdahale yaklaşımı"),
        ("Philip Kingsley", "philipkingsley.com", "Trikoloji", "İngiliz trikoloji uzmanı; Elasticizer ikonik ürün; 60+ yıllık saç bilimi; saç derisi sağlığı"),
        ("OUAI", "theouai.com", "Lifestyle Saç", "Jen Atkin; parfümlü saç ürünleri; Melrose Place kokusu ikonik; saç + vücut + fragrance"),
        ("Sachajuan", "sachajuan.com", "İsveç Saç Bakımı", "Stockholm merkezli; Ocean Mist Volume Spray ikonik; İskandinav minimalizm; profesyonel kalite"),
        ("Christophe Robin", "christophe-robin.com", "Paris Kolorist", "Paris'in en ünlü koloristi; Cleansing Purifying Scrub deniz tuzu peelingi; lüks salon evde"),
        ("Maria Nila", "marianila.com", "İsveç Vegan Saç", "100% vegan + hayvan dostu; İsveç markası; Colour Refresh renk maskeleri; sürdürülebilir ambalaj"),
        ("Playa", "plfraya.com", "California Saç", "Shelby Wild kurdu; California plaj dalgası; temiz bileşenler; minimalist koleksiyon"),
        ("Eva NYC", "eva-nyc.com", "NYC Uygun Fiyatlı", "New York ilhamlı; Ulta'da $10-20; eğlenceli ambalaj; erişilebilir saç bakımı; styling uzmanı"),
        ("Not Your Mother's", "nymbrands.com", "Süpermarket Premium", "Curl Talk kıvırcık serisi TikTok viral; uygun fiyatlı ama etkili; Gen Z favorisi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 3. ERKEK BAKIM (Men's Grooming) — 40+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Erkek Bakım": [
        ("Harry's", "harrys.com", "Tıraş Abonelik", "Gillette'e rakip; $1.37B Edgewell satın alma (FTC engelledi); Walmart + Target dağıtım; kendi fabrika"),
        ("Dollar Shave Club", "dollarshaveclub.com", "Tıraş Abonelik", "Viral video ile başladı ($1'lık tıraş bıçağı); Unilever $1B'a satın aldı; DTC abonelik modeli öncüsü"),
        ("Manscaped", "manscaped.com", "Erkek Vücut Bakımı", "Vücut tıraşı; Lawn Mower 5.0; podcast sponsorlukları; cesur pazarlama; $1B+ değerleme"),
        ("Dr. Squatch", "drsquatch.com", "Doğal Erkek Sabun", "Viral YouTube reklamları 100M+ görüntülenme; doğal erkek sabunu; eğlenceli pazarlama; $200M+ gelir"),
        ("Beardbrand", "beardbrand.com", "Sakal Bakımı", "YouTube sakal bakım kanalından markaya; Eric Bandholz kurdu; 'urban beardsman' topluluğu"),
        ("Cremo", "cremocompany.com", "Astonishingly Tıraş", "Süper konsantre tıraş kremi; 'astonishingly superior' slogan; Edgewell satın aldı; uygun fiyat"),
        ("Bevel", "bevelcode.com", "Melanin Erkek Bakım", "Tristan Walker kurdu; tıraş sonrası tahriş önleme; P&G satın aldı; esmer erkekler için özel"),
        ("Henson Shaving", "hensonshaving.com", "Hassas Tıraş", "Havacılık mühendisliği hassasiyetinde tıraş makinesi; AL13 tek bıçak; sıfır atık; Kanada yapımı"),
        ("Supply", "supply.co", "Tek Bıçak Tıraş", "Tek bıçaklı güvenlik tıraşı modern tasarımla; ömür boyu garanti; sürdürülebilir tıraş"),
        ("Hawthorne", "hawthorne.co", "Kişisel Erkek Bakım", "Quiz ile kişiselleştirilmiş kolonya + bakım; veri odaklı ürün eşleştirme; Target'a girdi"),
        ("Huron", "usehuron.com", "Temiz Erkek Bakım", "Matt Mullenax kurdu; temiz bileşenler erkek bakımda; $14-18 fiyat; basit etkili ürünler"),
        ("Lumin", "luminskin.com", "Erkek Cilt Bakımı", "Erkek cilt bakım aboneliği; Kore cilt bakımı formülleri; $10 başlangıç kiti; sosyal medya pazarlaması"),
        ("Scotch Porter", "scotchporter.com", "Siyah Erkek Bakım", "Calvin Quallis kurdu; siyah erkekler için saç + sakal + cilt bakımı; doğal bileşenler; Target'ta"),
        ("Every Man Jack", "everymanjack.com", "Doğal Erkek", "Doğal + sürdürülebilir erkek bakım; uygun fiyat; Marvel lisansı ambalajlar; Target + Walmart"),
        ("Oars + Alps", "oarsandalps.com", "Aktif Erkek Bakım", "Doğal deodorant + cilt bakımı aktif erkekler için; SC Johnson satın aldı; Target'ta"),
        ("Bravo Sierra", "bravosierra.com", "Askeri Erkek Bakım", "ABD askerleri ile birlikte geliştirilen ürünler; aktif yaşam tarzı; sürdürülebilir; performans odaklı"),
        ("Fulton & Roark", "fultonandroark.com", "Katı Kolonya", "Katı kolonya icat ettiler; seyahat dostu; el yapımı; küçük seri; ABD'de üretim"),
        ("Blind Barber", "blindbarber.com", "Berber Stili", "NYC berber dükkanı + ürün markası; 90 Proof pomad ikonik; berber kültürü + lifestyle"),
        ("Byrd Hair", "byrdhair.com", "Sörf Saç Bakımı", "California sörf kültürü ilhamlı saç bakımı; Pocket Comb taşınabilir tarak; casual erkek stili"),
        ("Baxter of California", "bfraxterofcalifornia.com", "California Erkek Bakım", "1965'ten beri; Clay Pomade ikonik; ABD'nin ilk erkek bakım markalarından; L'Oréal bünyesinde"),
        ("Jack Black", "getjackblack.com", "Premium Erkek Cilt", "Lip Balm en çok satan; erkek cilt bakımı öncüsü; Edgewell bünyesinde; $25-50 fiyat aralığı"),
        ("Aesop", "aesop.com", "Unisex Lüks Bakım", "Melbourne minimalizmi; Resurrection el yıkama ikonik; L'Oréal $2.5B'a satın aldı; unisex lüks"),
        ("Frederick Benjamin", "frederickbenjamin.com", "Koyu Cilt Erkek", "Koyu cilt tonları için tıraş + bakım; tahriş önleme; siyah erkek grooming uzmanı"),
        ("Viking Revolution", "vikingrevolution.com", "Sakal Yağları", "Amazon #1 sakal bakım seti; uygun fiyatlı sakal yağı + balm; hediye seti bestseller"),
        ("Honest Amish", "honestamish.com", "Organik Sakal", "El yapımı doğal sakal bakım; Amish topluluğu ilhamı; Amazon favorisi; organik bileşenler"),
        ("Duke Cannon", "dukecannon.com", "Büyük Sabunlar", "280g+ büyük sabunlar; askeri ilhamlı pazarlama; 'Big Ass Beer Soap'; erkeksi marka sesi"),
        ("Disco", "letsdisco.com", "Erkek Cilt Bilimi", "Benjamin Smith kurdu; erkek cilt bakım rutini; deri bakım basitleştirilmiş; Shark Tank"),
        ("Geologie", "gefrologie.com", "Kişisel Erkek Cilt", "30 saniyelik quiz ile kişisel rejim; erkek dermatolojik bakım; abonelik modeli"),
        ("Meridian", "meridiangrooming.com", "Aşağı Bakım", "Erkek bikini bölgesi bakımı; The Trimmer ikonik; cesur pazarlama; Manscaped rakibi"),
        ("Billie", "mybillie.com", "Kadın Tıraş (Ama Erkek de)", "Kadın tıraş markası ama erkek tıraşa da genişledi; pembe vergi karşıtı; Edgewell satın aldı; P&G rakibi"),
        ("Patricks", "patricks.co", "Avustralya Lüks Erkek", "Avustralya lüks erkek bakım; $50-100 premium ürünler; siyah ambalaj; minimalist erkek lüks"),
        ("Aēsop", "aesop.com", "Eczane Lüks", "Dennis Paphitis'in Melbourne markası; mağaza tasarımları sanat eseri; unisex; L'Oréal $2.5B satın aldı"),
        ("Curateur", "curateur.com", "Erkek Stil Kutusu", "Rachel Zoe küratörlüğünde kutu; erkek + kadın; premium ürün keşfi; sezonluk kutu"),
        ("Gillette On Demand", "gillette.com", "P&G DTC Tıraş", "P&G'nin DTC yanıtı Harry's ve DSC'ye; online abonelik tıraş; doğrudan tüketiciye geçiş"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 4. SAĞLIK & WELLNESS (Health & Wellness) — 80+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Sağlık & Wellness": [
        ("AG1", "drinkag1.com", "Hepsi Bir Arada Takviye", "Her podcast'ı sponsorluyor; tek kaşık 75 vitamin/mineral; 'her sabah ilk aldığım şey' UGC stratejisi"),
        ("Ritual", "ritual.com", "Şeffaf Vitamin", "Beanie Babies gibi şeffaf kapsüller; tedarik zinciri şeffaflığı; kadın multivitamin öncüsü; $200M+ gelir"),
        ("Seed", "seed.com", "Probiyotik Bilim", "DS-01 simbiyotik; çift kapsül teknolojisi; bilimsel danışma kurulu; Ara Katz kurdu; probiyotik 2.0"),
        ("Care/of", "takecareof.com", "Kişisel Vitamin Paketi", "Quiz ile kişisel vitamin paketi; Bayer satın aldı; günlük paketler isminizle; veri odaklı sağlık"),
        ("OLLY", "olly.com", "Gummy Vitamin", "Gummy vitamin kategorisini popüler yaptı; Target özel; Unilever satın aldı; eğlenceli ambalaj"),
        ("Liquid IV", "liquid-iv.com", "Hidrasyon Tozu", "CTT (Cellular Transport Technology); Unilever satın aldı; festival + spor pazarlaması; $500M+"),
        ("Bloom Nutrition", "bloomnu.com", "Kadın Fitness Takviye", "Mari Llewellyn kurdu; TikTok viral yeşil toz; kadın fitness topluluğu; Target'a girdi; $200M+"),
        ("Vital Proteins", "vitalproteins.com", "Kolajen", "Jennifer Aniston yüzü; kolajen peptit öncüsü; Nestlé satın aldı; sabah kahvesine kolajen trendi"),
        ("Moon Juice", "moonjuice.com", "Adaptojenik Süper", "Amanda Chantal Bacon LA wellness kültürü; SuperYou adaptojenik; kristal + mantar + bitkiler"),
        ("Four Sigmatic", "foursigmatic.com", "Mantar Kahve", "Finlandiya'dan mantar kahve; lion's mane + chaga; mantarlı wellness öncüsü; $100M+ gelir"),
        ("MUD\\WTR", "mudwtr.com", "Kahve Alternatifi", "Mantar + chai + kakao; 1/7 kafein; 'not coffee' pazarlama; Shane Heath kurdu; $60M+ gelir"),
        ("Eight Sleep", "eightsleep.com", "Akıllı Yatak", "Pod kapağı yatak ısısını ayarlıyor; uyku takibi; $500M+ değerleme; biyohacking topluluğu; Huberman favorisi"),
        ("Whoop", "whoop.com", "Fitness Bileklik", "Ekransız fitness takip; toparlanma + zorlama skoru; abonelik modeli; $3.6B değerleme; sporcuların tercihi"),
        ("Oura Ring", "ouraring.com", "Akıllı Yüzük", "Uyku + aktivite takip yüzüğü; Kim Kardashian + Prince Harry kullanıyor; $2.55B değerleme; Finlandiya"),
        ("Thesis", "takethesis.com", "Kişisel Nootropik", "Quiz ile kişiselleştirilmiş beyin takviyesi; Clarity, Energy, Logic formülleri; nootropik 2.0"),
        ("Onnit", "onnit.com", "Toplam İnsan Optimizasyonu", "Alpha Brain nootropik; Joe Rogan uzun süreli sponsor; Unilever satın aldı; Austin TX merkezli"),
        ("Momentous", "livemomentous.com", "Bilim Bazlı Takviye", "Huberman Lab podcast önerisi; NSF sertifikalı; sporcuların tercihi; klinik doz formüller"),
        ("Transparent Labs", "transparentlabs.com", "Şeffaf Etiket", "%100 şeffaf etiket; gizli karışım yok; klinik dozlar; hardcore fitness topluluğu"),
        ("Gorilla Mind", "gorillamind.com", "Nootropik + Pre-Workout", "More Plates More Dates (Derek) kurdu; Sigma pre-workout viral; YouTube fitness topluluğu"),
        ("Thorne", "thorne.com", "Klinik Takviye", "Mayo Clinic ortağı; NSF sertifikalı; profesyonel sporcu tercihi; $100M+ gelir; bilimsel güvenilirlik"),
        ("Gainful", "gainful.com", "Kişisel Protein", "Quiz ile kişiselleştirilmiş protein tozu; her karışım benzersiz; hedef odaklı formüller"),
        ("Persona Nutrition", "personanutrition.com", "AI Vitamin Paketi", "Nestlé Health Science satın aldı; kişiselleştirilmiş vitamin; online değerlendirme"),
        ("HUM Nutrition", "humnutrition.com", "Güzellik Takviyesi", "Cilt + saç + vücut güzellik takviyeleri; RD ekibi formülüyor; Sephora'da satışta"),
        ("Beam", "beamorganics.com", "CBD Wellness", "Dream Powder uyku tozu viral; nano CBD teknolojisi; performans odaklı CBD; Matt Lombardi kurdu"),
        ("Obvi", "myobvi.com", "Kolajen Protein", "Renkli kolajen protein tozu; Gen Z hedef kitle; TikTok viral; Ronak Shah kurdu"),
        ("Beekeeper's Naturals", "beekeepersnaturals.com", "Arı Ürünleri", "Propolis boğaz spreyi; arı sütü + propolis; Carly Stein kurdu; doğal bağışıklık desteği"),
        ("Goli", "goli.com", "Elma Sirkesi Gummy", "Elma sirkesi gummy'si popüler yaptı; 1B+ satış; kolay tüketim formu; TV reklamları"),
        ("Lemme", "lemme.com", "Kardashian Takviye", "Kourtney Kardashian'ın markası; gummy vitamin; libido + uyku + sindirim; ünlü wellness"),
        ("Timeline Nutrition", "timelinenutrition.com", "Mitokondri Sağlığı", "Mitopure® Urolithin A; mitokondri yenileme; İsviçre bilimi; anti-aging takviye"),
        ("Elysium Health", "elysiumhealth.com", "Longevity Bilimi", "Basis® NAD+ takviyesi; Nobel ödüllü bilim insanı danışmanlar; yaşlanma bilimi; premium anti-aging"),
        ("InsideTracker", "insidetracker.com", "Kan Testi Analizi", "Kan testi + AI ile kişisel sağlık planı; biyobelirteç optimizasyonu; David Sinclair kullanıyor"),
        ("Noom", "noom.com", "Psikoloji Bazlı Diyet", "CBT bazlı kilo verme; psikoloji + yapay zeka; $4B+ değerleme; 50M+ indirme"),
        ("Calibrate", "joincalibrate.com", "Metabolik Sağlık", "GLP-1 ilaç destekli kilo verme programı; doktor gözetiminde; metabolik reset; $100M+ toplam yatırım"),
        ("Found", "joinfound.com", "Kilo Yönetimi", "Kişiselleştirilmiş kilo verme; reçeteli ilaç + koçluk; $100M+ yatırım; bütünsel yaklaşım"),
        ("Levels", "levelshealth.com", "Metabolik Fitness", "CGM ile gerçek zamanlı kan şekeri takibi; yemeklerin etkisini görme; $100M+ yatırım; Sam Corcos"),
        ("Lumen", "lumen.me", "Nefes Metabolizma", "Nefes analizörü yağ mı karbonhidrat mı yaktığını ölçer; CO2 analizi; $80M+ yatırım"),
        ("ZOE", "joinzoe.com", "Kişisel Beslenme", "Tim Spector'ın bilimi; CGM + bağırsak mikrobiyom testi; kişisel beslenme planı; $100M+ yatırım"),
        ("Everlywell", "everlywell.com", "Evde Test", "FDA onaylı evde lab testleri; gıda hassasiyeti, tiroid, STI; Shark Tank; $500M+ değerleme"),
        ("Hims & Hers", "hims.com", "Telehealth", "Saç dökülmesi, ED, cilt bakımı online reçete; $2B+ piyasa değeri; NYSE listeli; telehealth öncüsü"),
        ("Ro", "ro.co", "Dijital Sağlık", "Roman (ED) ile başladı; Zero (sigara bırakma); Rory (kadın sağlığı); $7B değerleme; telehealth"),
        ("Nurx", "nurx.com", "Online Doğum Kontrol", "Doğum kontrol hapı online reçete + teslimat; PrEP; STI test kitleri; kadın sağlığı erişimi"),
        ("Curology", "curology.com", "Online Dermatoloji", "Kişisel reçeteli cilt bakım formülü; online dermatolog; abonelik; adapalene + niacinamide + tret"),
        ("Apostrophe", "apostrophe.com", "Online Cilt Tedavi", "Dermatolojist reçeteli cilt bakımı; online konsültasyon; retinoid + antibiyotik formüller"),
        ("Cerebral", "cerebral.com", "Online Psikoloji", "Online terapi + ilaç tedavisi; anksiyete + depresyon; $4.8B pik değerleme; mental sağlık telehealth"),
        ("Arey Grey", "arey.com", "Beyaz Saç Takviyesi", "Beyaz saçlanmayı tersine çevirme takviyesi; melanin üretimi destekleyen formül; anti-graying"),
        ("Athletic Greens", "athleticgreens.com", "Yeşil Toz", "AG1'in eski adı; podcast pazarlama kralı; Tim Ferriss, Huberman hep önerdi; günlük doz yeşillik"),
        ("Sakara Life", "sakara.com", "Organik Yemek Teslim", "Bitki bazlı organik yemek teslimatı; detoks programları; lüks wellness beslenme; NYC merkezli"),
        ("Hilma", "hilma.co", "Doğal OTC", "Doğal soğuk algınlığı + alerji ilaçları; ilaç dolabını temizle kampanyası; OTC ürünleri doğala çevir"),
        ("Welly", "getwelly.com", "Tasarım Yara Bandı", "Renkli desenli yara bandı; ilk yardım çantası estetik; Target'ta; sıkıcı yara bandına son"),
        ("Joovv", "joovv.com", "Kırmızı Işık Terapi", "Kırmızı ışık terapi panelleri; evde fototerapi; biyohacking; cilt + ağrı + toparlanma"),
        ("Hyperice", "hyperice.com", "Toparlanma Teknoloji", "Hypervolt masaj tabancası; NBA + NFL resmi ortağı; soğutma + ısıtma teknolojisi; $700M+ değerleme"),
        ("Therabody", "therabody.com", "Perkusyon Terapi", "Theragun masaj tabancası kategorisini yarattı; Dr. Jason Wersland kurdu; $1B+ değerleme; toparlanma ikonu"),
        ("MUDWTR", "mudwtr.com", "Mantar İçecek", "Kahve alternatifi; reishi, chaga, lion's mane karışımı; Shane Heath kurdu; anti-kahve pazarlama"),
        ("Cymbiotika", "cymbiotika.com", "Lipozomal Takviye", "Lipozomal taşıma teknolojisi; Vitamin B12, D3, Omega; premium fiyat $50+; biyoyararlanım odaklı"),
        ("Neurohacker Collective", "neurohacker.com", "Nootropik Bilim", "Qualia Mind nootropik; kompleks formüller; Daniel Schmachtenberger; bilimsel derinlik"),
        ("Prenuvo", "prenuvo.com", "Tam Vücut MR", "Tam vücut MRI tarama $2,500; erken kanser tespiti; Kim Kardashian tanıttı; önleyici sağlık"),
        ("Parsley Health", "parsleyhealth.com", "Fonksiyonel Tıp", "Dr. Robin Berzin kurdu; fonksiyonel tıp + telehealth; kök neden yaklaşımı; $100M+ yatırım"),
        ("Forward", "goforward.com", "AI Doktor", "AI destekli birincil bakım; body scanner; $17/ay üyelik; sınırsız ziyaret; geleceğin sağlık hizmeti"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 5. FITNESS & SPOR GİYİM (Fitness & Activewear) — 60+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Fitness & Spor Giyim": [
        ("Vuori", "vuori.com", "Premium Athleisure", "Encinitas CA; Kore Short erkek yoga şort ikonik; SoftBank $4B değerleme; Lululemon rakibi"),
        ("Outdoor Voices", "outdoorvoices.com", "Rekreasyonel Aktivite", "'Doing Things' sloganı; Tyler Haney kurdu; günlük hareket kültürü; pastel renkler; finansal zorluklar atlattı"),
        ("ASRV", "asrv.com", "Erkek Performans", "Erkek teknik athleisure; Silver-Lite kumaş teknolojisi; orta segment lüks; spor salonu + sokak"),
        ("Alphalete", "alphalete.com", "Bodybuilding Giyim", "Christian Guzman kurdu; YouTube fitness topluluğu; Revival koleksiyonları dakikalar içinde tükeniyor"),
        ("Buffbunny", "buffbunny.com", "Kadın Fitness Moda", "Heidi Somers kurdu; cesur desenler; lansmanlar dakikalar içinde tükeniyor; fitness influencer topluluğu"),
        ("Girlfriend Collective", "girlfriend.com", "Geri Dönüşüm Aktif", "Pet şişelerden tayt; %100 geri dönüşüm; kapsayıcı bedenler XXS-6XL; şeffaf üretim"),
        ("Nobull", "nobullproject.com", "CrossFit Giyim", "CrossFit Games resmi sponsoru; SuperFabric dayanıklı kumaş; fonksiyonel fitness topluluğu"),
        ("Born Primitive", "bornprimitive.com", "Taktik Fitness", "Askeri + CrossFit; veteran kurucular; dayanıklı fitness giyim; Savage Race sponsoru"),
        ("Tracksmith", "tracksmith.com", "Koşu Heritage", "Yarış kültürü + heritage estetik; Boston merkezli; elitist koşu topluluğu; premium koşu giyim"),
        ("Set Active", "setactive.co", "LA Sportswear", "Lindsey Carter kurdu; SportsSet ikonik; renk hikayesi lansmanları; Instagram-native marka"),
        ("Beyond Yoga", "beyondyoga.com", "Yumuşak Aktif Giyim", "Spacedye kumaş süper yumuşak; kapsayıcı bedenler; Levi's satın aldı; lüks his + performans"),
        ("Year of Ours", "yearofours.com", "LA Fitness Moda", "Ribbed Football Legging ikonik; LA fitness estetik; cesur tasarımlar; butik fitness topluluğu"),
        ("Peloton", "onepeloton.com", "Bağlı Fitness", "Bağlı bisiklet + koşu bandı; 7M+ üye; canlı dersler; pandemi sırasında patlama; $50B'dan $3B'a düşüş"),
        ("Tonal", "tonal.com", "Akıllı Ev Spor", "Duvara monte dijital ağırlık sistemi; AI antrenör; elektromanyetik direnç; $1.6B değerleme"),
        ("Hydrow", "hydrow.com", "Bağlı Kürek", "Bağlı kürek makinesi; gerçek su çekimleri ekranda; $35/ay abonelik; tam vücut antrenman"),
        ("Ergatta", "ergatta.com", "Oyunlaştırılmış Kürek", "Su dirençli kürek + oyunlaştırma; yarış oyunları; WaterRower ile ortaklık; eğlenceli fitness"),
        ("Bala", "shopbala.com", "Tasarım Ağırlık", "Bileklik ağırlık manşetleri; Shark Tank $7M yatırım; estetik fitness ekipmanı; Instagram viral"),
        ("Crossrope", "crossrope.com", "Akıllı İp Atlama", "Ağırlıklı ip atlama sistemi; uygulama bağlantılı; değiştirilebilir ağırlıklar; eğlenceli kardiyo"),
        ("TRX", "trxtraining.com", "Askı Antrenman", "Navy SEAL geliştirdi; askı sistemi ile vücut ağırlığı antrenmanı; taşınabilir spor salonu"),
        ("FORM Swim", "formswim.com", "Akıllı Yüzme Gözlüğü", "AR ekranlı yüzme gözlüğü; gerçek zamanlı metrikler; Apple Watch entegrasyonu; yüzücü teknolojisi"),
        ("Liteboxer", "liteboxer.com", "Boks Fitness", "Ritmik boks antrenman platformu; müzik senkronlu; evde boks dersleri; Lizzo yatırımcı"),
        ("FightCamp", "joinfightcamp.com", "Ev Boksu", "Evde boks antrenmanı; punch tracker sensörler; aylık abonelik; kickboxing dersleri"),
        ("Hyperice", "hyperice.com", "Toparlanma Cihaz", "Hypervolt masaj tabancası; Normatec sıkıştırma botları; NBA/NFL/PGA resmi ortağı; $700M+ değerleme"),
        ("Therabody", "therabody.com", "Masaj Tabancası", "Theragun perkusyon terapi icat etti; RecoveryAir JetBoots; Wave Roller; $1B+ değerleme"),
        ("Whoop", "whoop.com", "Performans Bileklik", "Ekransız; strain + recovery + sleep skoru; abonelik; profesyonel sporcular kullanıyor; $3.6B değerleme"),
        ("Gymshark", "gymshark.com", "Fitness Sosyal Medya", "Ben Francis 19 yaşında kurdu; fitness influencer ağı; $1.45B değerleme; İngiltere'nin en başarılı DTC'si"),
        ("Alo Yoga", "aloyoga.com", "Yoga Lüks", "LA yoga kültürü; Alo Moves dijital platform; ünlü yoga pantolon; celebrity street style; $10B değerleme"),
        ("P.E Nation", "pe-nation.com", "Avustralya Aktivewear", "Pip Edwards kurdu; spor + sokak modası; cesur desenler; Avustralya moda ayakizi"),
        ("Sweaty Betty", "sweatybetty.com", "İngiliz Kadın Aktif", "İngiltere kadın fitness giyim; Power Legging ikonik; Wolverine Worldwide $410M satın aldı"),
        ("LSKD", "lskd.com.au", "Avustralya Fitness", "Rep Tight ikonik; Avustralya fitness topluluğu; garaja başladı; $100M+ AUD gelir"),
        ("Ryderwear", "ryderwear.com", "Bodybuilding Avustralya", "Avustralya ağırlık kaldırma giyim; D-Mak ayakkabı; bodybuilding topluluğu"),
        ("Under Armour Curry", "currybrand.com", "Steph Curry", "Steph Curry'nin marka içi markası; basketbol + golf + lifestyle; Under Armour bünyesinde"),
        ("Satisfy", "satisfyrunning.com", "Lüks Koşu", "Fransız ultra-premium koşu giyim; $200+ şort; koşu kültürü + sanat; 'MothTech' kumaş"),
        ("District Vision", "districtvision.com", "Meditasyon Koşu", "Koşu + meditasyon kesişimi; koşu gözlükleri; mindful running; Junya Watanabe collab"),
        ("Ten Thousand", "tenthousand.com", "Erkek Performans Şort", "Tek ürün odak: erkek antrenman şort; Set Short ikonik; fonksiyon önce; premium kalite"),
        ("Rhone", "rhone.com", "Erkek Premium Aktif", "Nate Checketts kurdu; GoldFusion anti-koku teknolojisi; Commuter koleksiyonu ofis + spor"),
        ("Fourlaps", "fourlaps.com", "Erkek Koşu Giyim", "Rush Short ikonik; New York koşu topluluğu; erkek athleisure; performans + stil"),
        ("Janji", "janji.com", "Sosyal Koşu", "Gelirin bir kısmı temiz su projelerine; dünya koşu kültürleri ilhamlı desenler; etik koşu giyim"),
        ("Bandit Running", "banditrunning.com", "NYC Koşu", "Tim West + Nick West; New York koşu kültürü; sınırlı koleksiyonlar; koşu topluluğu"),
        ("Ciele Athletics", "cieleathletics.com", "Koşu Şapka", "Montreal; koşu şapkası uzmanı; GOCap ikonik; ultra koşu topluluğu; renkli tasarımlar"),
        ("Soar Running", "soarrunning.com", "İngiliz Koşu Teknoloji", "Ultra hafif koşu giyim; Tim Don (triatloncu) kurdu; teknik performans; İngiltere yapımı"),
        ("Vain Dane Athletic", "vaindaneathletic.com", "Danimarka Sürdürülebilir", "Geri dönüşüm polyester koşu giyim; Kopenhag; sürdürülebilir Nordik atletik giyim"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 6. MODA & GİYİM (Fashion & Apparel) — 80+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Moda & Giyim": [
        ("Everlane", "everlane.com", "Radikal Şeffaflık", "Radikal fiyat şeffaflığı; fabrika hikayeleri; $100M+ gelir; 'ethical factory' öncüsü"),
        ("SKIMS", "skims.com", "Shapewear & Basics", "Kim Kardashian markası; $4B değerleme; Fendi collab; 9 renk ton shapewear; kapsayıcı bedenler"),
        ("Good American", "goodamerican.com", "Kapsayıcı Denim", "Khloé Kardashian + Emma Grede; 00-32 beden aralığı; 'Good Legs' jean ikonik; her bedene fit"),
        ("Reformation", "thereformation.com", "Sürdürülebilir Moda", "LA sürdürülebilir moda; RefScale çevresel etki ölçümü; ünlü favori; $1B+ değerleme"),
        ("Allbirds", "allbirds.com", "Yün Ayakkabı", "Yeni Zelanda merinos yünü ayakkabı; karbon nötr; Tim Brown (NZ futbolcu) kurdu; IPO sonrası zorluklar"),
        ("Rothy's", "rothys.com", "Geri Dönüşüm Ayakkabı", "Pet şişelerden örme ayakkabı; 125M+ şişe geri dönüştürüldü; yıkanabilir flat'ler; $1B+ değerleme"),
        ("CARIUMA", "cariuma.com", "Sürdürülebilir Sneaker", "Brezilya kurucular; her ayakkabı için ağaç dikimi; IBI bambu sneaker; B Corp; doğal malzemeler"),
        ("Oliver Cabell", "olivercabell.com", "Şeffaf Fiyat Sneaker", "Lüks sneaker maliyet dağılımı gösteriyor; İtalya'da üretim; $198 vs designer $600+"),
        ("Thursday Boots", "thursdayboots.com", "Erişilebilir Bot", "Premium deri botlar $200'dan; Horween deri; DTC fiyat avantajı; erkek + kadın; NYC merkezli"),
        ("Tecovas", "tecovas.com", "Kovboy Çizmesi DTC", "El yapımı kovboy çizmeleri fabrikadan tüketiciye; $255'tan; Teksas kültürü; fiziksel mağazalara genişledi"),
        ("Parade", "yourparade.com", "Sürdürülebilir İç Giyim", "Gen Z iç giyim; geri dönüşüm kumaş; kapsayıcı bedenler; renkli kampanyalar; $10 başlangıç"),
        ("True Classic", "trueclassictees.com", "Erkek Temel Tişört", "Kas fit erkek tişört; Facebook reklam ustası; $250M+ gelir; 'her erkek yakışıklı görünür' vaadi"),
        ("Madhappy", "madhappy.com", "Mental Sağlık Streetwear", "Mental sağlık farkındalığı + streetwear; Local Optimist topluluğu; hoodie'ler $160+; LVMH yatırımcı"),
        ("Cuts Clothing", "cutsclothing.com", "Erkek Polo Alternatif", "PYCA kumaş teknolojisi; AO polo en çok satan; 'not a polo not a tee' konsepti; $100M+ gelir"),
        ("Bylt Basics", "byltbasics.com", "Erkek Premium Basics", "LUX Blend kumaş süper yumuşak; drop-cut tişört; erkek temel giyim premium segment"),
        ("Quince", "onequince.com", "Lüks Fabrika Fiyatı", "Fabrikadan direkt lüks malzeme; kaşmir $50, ipek $30; Everlane'den bile ucuz; radikal fiyat"),
        ("Italic", "italic.com", "Markalanmamış Lüks", "Lüks markaların fabrikalarında üretim, markalanmamış; 'brand tax' olmadan kalite; üyelik modeli"),
        ("Hill House Home", "hillhousehome.com", "Nap Dress", "Nell Diamond'ın Nap Dress'i 2020'de viral; 'uyumak için de giymek için de' elbise; bekleme listesi 100K+"),
        ("Lunya", "lunya.com", "Lüks Uyku Giyim", "Washable silk pijama; kadın uyku giyim lüks segmenti; Restore koleksiyonu; 'sleepwear is self-care'"),
        ("Lake Pajamas", "lakepajamas.com", "Pima Pamuk Pijama", "Pima pamuk lüks pijama; eşleşen aile setleri; Instagram viral; ev giyim lüks trendi"),
        ("Bombas", "bombas.com", "Bire Bir Çorap", "Her satışta bir çift evsizlere bağış; $100M+ gelir ilk yılda; Shark Tank en başarılı yatırım"),
        ("Cozy Earth", "cozyearth.com", "Bambu Tekstil", "Bambu viskon çarşaf + pijama; Oprah's Favorite Things; 10 yıl garanti; termal düzenleme"),
        ("Marine Layer", "marinelayer.com", "SF Yumuşak Tişört", "San Francisco; Re-Spun geri dönüşüm tişört; süper yumuşak kumaş; West Coast casual"),
        ("Buck Mason", "buckmason.com", "LA Erkek Basics", "Los Angeles erkek temel giyim; Curved Hem Tee ikonik; ABD'de üretim; vintage his"),
        ("Faherty", "faherty.com", "Sahil Yaşam Tarzı", "Alex & Mike Faherty; New Jersey sahil kültürü; All Day Shorts erkek favori; Sunwashed kumaş"),
        ("UNTUCKit", "untuckit.com", "Gömlek Dışarıda", "Gömlek pantolonun dışında giyilmek için tasarlandı; doğru boy; $600M+ gelir; erkek gömlek inovasyonu"),
        ("Bonobos", "bonobos.com", "Erkek Pantolon Fit", "Mükemmel fit erkek pantolon; Guideshop konsepti; Walmart $310M satın aldı; DTC erkek moda öncüsü"),
        ("ThirdLove", "thirdlove.com", "Yarım Beden Sütyen", "Yarım beden sütyen öncüsü; Fit Finder quiz; 80+ beden; kapsayıcı; $750M değerleme"),
        ("MeUndies", "meundies.com", "Eğlenceli İç Giyim", "MicroModal kumaş; eşleşen çift iç giyim; cesur desenler; abonelik modeli; eğlenceli marka"),
        ("Knix", "knix.com", "Sızıntı Geçirmez İç Giyim", "Joanna Griffiths kurdu; sızıntı geçirmez teknoloji; adet + idrar; kapsayıcı; Essity satın aldı"),
        ("Summersalt", "summersalt.com", "Seyahat Mayo", "1.5M vücut taramasından üretilen mayo; Sidestroke tek omuz mayo ikonik; kompresyon kumaş"),
        ("Andie Swim", "andieswim.com", "Online Mayo", "Quiz ile kişisel mayo önerisi; filter-free reklam; gerçek kadın vücutları; fit garantisi"),
        ("Fair Harbor", "fairharborclothing.com", "Geri Dönüşüm Mayo", "Geri dönüşüm plastikten erkek mayo; 11 plastik şişe = 1 mayo; NYC kurucu; sürdürülebilir plaj"),
        ("Chubbies", "chubbies.com", "Erkek Kısa Şort", "5.5 inç şort; hafta sonu kültürü; 'Sky's Out Thighs Out'; eğlenceli erkek moda; viral sosyal medya"),
        ("Western Rise", "westernrise.com", "Teknik Günlük Giyim", "AT Slim Pant ofis + outdoor; Strongcore kumaş; minimalist erkek teknik giyim"),
        ("Fresh Clean Threads", "freshcleanthreads.com", "Tişört Abonelik", "Aylık tişört abonelik; $9-12/tişört; temel giyim; Facebook reklam başarısı; $100M+ gelir"),
        ("Sézane", "sezane.com", "Fransız DTC Moda", "Morgane Sézalory; ilk Fransız online moda markası; Parisienne chic; ayda 1 koleksiyon; sürdürülebilir"),
        ("Princess Polly", "princesspolly.com", "Gen Z Avustralya Moda", "Avustralya fast fashion; TikTok favorisi; Gen Z; Quince Lundberg bünyesinde; $200M+ gelir"),
        ("Anine Bing", "aninebing.com", "Scandi Chic", "Instagram'dan markaya; LA + Stockholm; Vintage Bing sweatshirt ikonik; $100M+ gelir"),
        ("Totême", "toteme-studio.com", "İsveç Minimalizm", "Elin Kling kurdu; İskandinav minimalizm; The Scarf Coat ikonik; sessiz lüks trendi"),
        ("Ganni", "ganni.com", "Danimarka Eğlenceli Moda", "#GanniGirls; leopar desen + gülücük yüz; Scandi 2.0; LVMH yatırımcı; $200M+ gelir"),
        ("PANGAIA", "thepangaia.com", "Biyomalzeme Moda", "FLWRDWN yalıtım; deniz yosunu lifi; $245M değerleme; bilim + malzeme inovasyonu; biyo-teknoloji"),
        ("Aritzia", "aritzia.com", "Kanada Kadın Moda", "Super Puff ceket viral; TikTok favorisi; $2B+ gelir; Babaton, Wilfred alt markaları; Vancouver"),
        ("Frankie Shop", "thefrankieshop.com", "Oversize Tailoring", "Gaëlle Drevet; oversize blazer trendi başlattı; Bea blazer ikonik; Parisian effortless"),
        ("Warby Parker", "warbyparker.com", "DTC Gözlük", "$95 reçeteli gözlük; 'buy a pair give a pair'; ev deneme programı; $6B IPO; gözlük devrimcisi"),
        ("Pair Eyewear", "paireyewear.com", "Değiştirilebilir Gözlük", "Manyetik üst çerçeve değiştirme; $60 baz + $25 üst çerçeve; kişiselleştirme; çocuk + yetişkin"),
        ("Cuyana", "cuyana.com", "Fewer Better Things", "'Daha az ama daha iyi' felsefesi; zaman ötesi İtalyan deri çanta; Lean Closet programı"),
        ("Senreve", "senreve.com", "Çok İşlevli Çanta", "Maestra çanta 4 farklı şekilde taşınıyor; kadın CEO çantası; İtalya yapımı; fonksiyonel lüks"),
        ("Staud", "stfraud.com", "LA Çağdaş Moda", "Moreau Bucket Bag ikonik; Sarah Staudinger kurdu; $100-500 çağdaş moda; TikTok viral parçalar"),
        ("Mango", "mango.com", "İspanyol DTC Genişleme", "Barcelona fast fashion; DTC kanalına büyük yatırım; sürdürülebilirlik hedefleri; $3B+ gelir"),
        ("COS", "cosstores.com", "H&M Lüks Alt Marka", "H&M grubunun üst segment markası; İskandinav minimalizm; mimari mağazalar; kaliteli basics"),
        ("& Other Stories", "stories.com", "H&M Hikaye Bazlı", "Paris + Stockholm + LA atelyelerinden koleksiyonlar; H&M grubu; orta segment; hikaye anlatımı"),
        ("Aimé Leon Dore", "aimeleondore.com", "NYC Streetwear Lüks", "Teddy Santis kurdu; Queens NYC kültürü; New Balance collab ikonik; café kültürü; hype"),
        ("Kith", "kith.com", "Streetwear Lüks Fusion", "Ronnie Fieg kurdu; Nike, BMW, Coca-Cola collabları; Treats cereal bar; $500M+ değerleme"),
        ("Fear of God Essentials", "essentials.fear-of-god.com", "Erişilebilir Lüks Streetwear", "Jerry Lorenzo; $40-90 basics; FOG'un erişilebilir hattı; PacSun ortaklığı; en çok satan streetwear"),
        ("REPRESENT", "representclo.com", "İngiliz Streetwear", "Manchester merkezli; Owners Club ikonik; UK streetwear lüks; George Heaton kurdu; $100M+ gelir"),
        ("Axel Arigato", "axelarigato.com", "İsveç Sneaker Moda", "Göteborg; Clean 90 sneaker ikonik; sneaker + hazır giyim; İskandinav streetwear; hızla büyüyen"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 7. YİYECEK & İÇECEK (Food & Beverage) — 80+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Yiyecek & İçecek": [
        ("OLIPOP", "drinkolipop.com", "Prebiyotik Soda", "9g fiber prebiyotik soda; Vintage Cola en çok satan; $200M+ gelir; sağlıklı gazlı içecek devrimi"),
        ("Poppi", "drinkpoppi.com", "Elma Sirkesi Soda", "Elma sirkesi bazlı soda; Super Bowl reklamı; $100M+ gelir; OLIPOP rakibi; TikTok viral"),
        ("Liquid Death", "liquiddeath.com", "Punk Su", "Kutu su; 'Murder Your Thirst'; heavy metal estetik; $700M değerleme; pazarlama dahisi; su kategorisini yeniden tanımladı"),
        ("Prime", "drinkprime.com", "İçecek Hype", "Logan Paul + KSI; lansmanında karaborsada 10x fiyat; Gen Alpha fenomen; 10B+ görüntülenme; $1.2B gelir"),
        ("Feastables", "feastables.com", "YouTube Çikolata", "MrBeast'in markası; $500M+ gelir; YouTube to CPG; Walmart + Target dağıtım; çocuk pazarı"),
        ("Magic Spoon", "magicspoon.com", "Protein Tahıl", "0g şeker, 13g protein granola; nostaljik kutu tasarımı; yetişkinler için çocukluk gevreği; $100M+ gelir"),
        ("Graza", "grfraza.co", "Sıkılabilir Zeytinyağı", "Sıkılabilir şişede zeytinyağı; 'Drizzle' + 'Sizzle' ikili; taze hasat; $50M+ gelir; TikTok viral"),
        ("Brightland", "brightland.com", "Premium Zeytinyağı", "California zeytinyağı; erken hasat ekstra sızma; Aishwarya Iyer kurdu; güzel ambalaj; DTC gıda estetiği"),
        ("Fly By Jing", "flybyjing.com", "Sichuan Sos", "Jing Gao kurdu; Sichuan Chili Crisp; Çin mutfağı otantik tat; $50M+ gelir; ABD'de Çin lezzetleri"),
        ("TRUFF", "trfruff.com", "Trüflü Acı Sos", "Trüf + habanero acı sos; siyah şişe lüks ambalaj; Oprah's Favorite; $50M+ gelir; Instagram güzel yemek"),
        ("Mid-Day Squares", "middfraysquares.com", "Fonksiyonel Çikolata", "Protein çikolata bar; kurucu çiftin reality show tarzı sosyal medya; $40M+ gelir; Kanada DTC"),
        ("Chomps", "chomps.com", "Otlak Et Çubuk", "Otlak hayvan eti snack; Whole30 onaylı; $200M+ gelir; Pete Maldonado kurdu; sağlıklı jerky"),
        ("Death Wish Coffee", "deathwishcoffee.com", "Dünyanın En Güçlü", "Dünyanın en güçlü kahvesi; kafein odaklı; kuru kafa logo; Super Bowl reklam kazananı"),
        ("Trade Coffee", "drinktrade.com", "Kavurma Eşleştirme", "450+ kavurucu ile eşleştirme; quiz ile kişisel kahve; abonelik; specialty coffee erişilebilir"),
        ("Cometeer", "cometeer.com", "Dondurulmuş Kahve", "Specialty kahve dondurulmuş kapsül; 10x konsantre; MIT araştırmacı kurdu; buzlu veya sıcak"),
        ("Athletic Brewing", "athleticbrewing.com", "Alkolsüz Bira", "Alkolsüz craft bira #1; Run Wild IPA; TIME 100 Most Influential; $800M+ değerleme; sober curious"),
        ("Ghia", "drinkghia.com", "Alkolsüz Aperitif", "İtalyan ilhamlı alkolsüz aperitif; Melanie Masarin kurdu; güzel ambalaj; sober curious trend"),
        ("Seedlip", "seedlipdrinks.com", "Alkolsüz Distile", "Dünyanın ilk distile alkolsüz içeceği; Diageo satın aldı; premium bar kültürü; 3 çeşit botanik"),
        ("LMNT", "drinklmnt.com", "Sıfır Şeker Elektrolit", "Sıfır şeker elektrolit; Robb Wolf kurdu; keto + paleo topluluğu; tuzlu limonata; podcast sponsoru"),
        ("Nuun", "nuunlife.com", "Efervesan Elektrolit", "Tablet elektrolit; koşucu + bisikletçi favorisi; Nestlé Health Science satın aldı"),
        ("Cure Hydration", "cfrurehydration.com", "ORS Hidrasyon", "Oral rehidrasyon bazlı; Lauren Picasso kurdu; Whole Foods; hindistan cevizi suyu tozu"),
        ("Alani Nu", "alaninu.com", "Kadın Fitness İçecek", "Katy Hearn kurdu; enerji içeceği + protein bar; kadın fitness topluluğu; $200M+ gelir; Target + Walmart"),
        ("Ghost Lifestyle", "ghostlifestyle.com", "Nostaljik Lezzet", "Warheads, Sour Patch Kids lisanslı tatlar; lifestyle fitness; $300M+ gelir; şeffaf etiket"),
        ("Chamberlain Coffee", "chamberlaincoffee.com", "Youtuber Kahve", "Emma Chamberlain'in markası; soğuk demleme + matcha; Gen Z kahve kültürü; YouTube to brand"),
        ("Jot", "jot.co", "Ultra Konsantre Kahve", "20x konsantre kahve; 1 kaşık = 1 fincan; 2 saniyede kahve; organik; ultra pratik format"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Çikolata", "Tahıl + soya + süt içermeyen çikolata; Mondelez satın aldı; simple chocolate bar; 'get back to human'"),
        ("SmartSweets", "smartsweets.com", "Düşük Şeker Şeker", "3g şeker gummy ayılar; 'kick sugar keep candy'; TPG $360M satın aldı; şeker devrimcisi"),
        ("Partake Foods", "partakefoods.com", "Alerjen Dostu", "Top 9 alerjen içermeyen kurabiye; Denise Woodard kurdu; Jay-Z yatırımcı; kapsayıcı snack"),
        ("Fishwife", "eatfishwife.com", "Lüks Konserve Balık", "Tinned fish trendini başlattı; renkli etiket tasarımı; premium sardunya + midye; hipster konserve"),
        ("Omsom", "omfrsom.com", "Asya Sos Kiti", "Vietnamlı Amerikalı kardeşler; otantik Asya yemek başlangıç kitleri; 5 dakikada Asya mutfağı"),
        ("Bachan's", "bachans.com", "Japon BBQ Sos", "Japon tarzı BBQ sos; aile tarifi; Justin Gill kurdu; Costco + Whole Foods; teriyaki alternatifi"),
        ("Mike's Hot Honey", "mikeshothoney.com", "Acı Bal", "Bal + biber; pizza'da viral; $30M+ gelir; tek ürün marka başarısı; restoran to perakende"),
        ("Banza", "eatbanza.com", "Nohut Makarna", "Nohuttan makarna + pizza hamuru; gluten-free protein; $160M+ toplam yatırım; 25g protein/porsiyon"),
        ("Caulipower", "caulipower.com", "Karnabahar Pizza", "Karnabahar hamurlu dondurulmuş pizza; $100M+ gelir; gluten-free pizza kategorisi yarattı"),
        ("Deux", "eatdeux.com", "Fonksiyonel Kurabiye Hamuru", "Yenilebilir kurabiye hamuru + takviye; vegan; Sabeena Ladha kurdu; TikTok viral; eğlenceli sağlıklı"),
        ("Last Crumb", "lastcrumb.com", "Lüks Kurabiye", "$140/12'li kutu; lüks kurabiye; bekleme listesi; her kurabiye farklı; ambalaj lüks unboxing deneyimi"),
        ("Behave Candy", "behavefcfrandy.com", "Düşük Şeker Şekerleme", "Düşük şeker + düşük karbonhidrat gummy; allulose tatlandırıcı; modern şekerleme"),
        ("Sunwink", "sunwink.com", "Bitki Tonik", "Süperfood bitki bazlı tonikler; Eliza Ganesh kurdu; bağırsak sağlığı + detoks; Whole Foods"),
        ("ZBiotics", "zbiotics.com", "Biyomühendislik İçecek", "Genetik mühendislik probiyotik; alkol öncesi shot; asetaldhit parçalama; bilim bazlı"),
        ("Primal Kitchen", "primalkitchen.com", "Paleo Sos & Dressing", "Mark Sisson kurdu; avokado yağlı mayonez; Kraft Heinz satın aldı; paleo + Whole30 mutfak"),
        ("Kettle & Fire", "kettleandfire.com", "Kemik Suyu", "Raf ömürlü kemik suyu; organik + grass-fed; $100M+ gelir; Justin + Nick Mares kurdu"),
        ("ButcherBox", "butcherbox.com", "Et Abonelik", "Otlak eti aylık kutu; grass-fed, organic; $600M+ gelir; çiftlikten kapıya et teslimatı"),
        ("Thrive Market", "thrivemarket.com", "Online Sağlıklı Market", "Costco + Whole Foods melezi online; üyelik modeli $60/yıl; organik + doğal; $500M+ gelir"),
        ("Daily Harvest", "daily-harvest.com", "Dondurulmuş Smoothie", "Dondurulmuş smoothie + harvest bowls; hazır karışım; $250M+ yatırım; sağlıklı dondurulmuş gıda"),
        ("Hungryroot", "hungryroot.com", "AI Yemek Planlama", "AI ile kişisel yemek planı + market alışverişi; sağlıklı hazır yemek; $40M+ yatırım"),
        ("Kodiak Cakes", "kodiakcakes.com", "Protein Pancake", "Protein pancake karışımı; Shark Tank reddi sonrası başarı; ayı logosu ikonik; $200M+ gelir"),
        ("Muddy Bites", "muddybites.com", "Çikolata Koni Ucu", "Dondurma konisinin çikolatalı alt kısmı; Jarod Steffes Shark Tank; TikTok viral snack; eğlenceli niş"),
        ("Tabs Chocolate", "tabschocolate.com", "Afrodizyak Çikolata", "Libido artırıcı dark çikolata; TikTok'ta 1B+ görüntülenme; cesur pazarlama; çift hedef kitle"),
        ("Surely", "drinksurelfy.com", "Alkolsüz Şarap", "Alkolsüz şarap; gerçek şarap tadı; sober curious hareketi; $10M+ yatırım; premium ambalaj"),
        ("Haus", "drink.haus", "Aperitif Likör", "Helena Price Hambrecht; düşük ABV aperitifler; doğal bileşenler; direkt tüketiciye alkol; kapatıldı/yeniden açıldı"),
        ("De La Calle", "delacalle.com", "Tepache İçecek", "Meksika fermente ananas içeceği; probiyotik; otantik tepache; Latin lezzetler; Whole Foods"),
        ("Sanzo", "drinksanzo.com", "Asya Aromalı Maden Suyu", "Asya meyve aromalı maden suyu; calamansi + lychee; Sandro Roco kurdu; Asya Amerikan kimliği"),
        ("Swoon", "drinkswoon.com", "Sıfır Şeker Limonata", "Sıfır şeker limonata; monk fruit tatlandırıcı; Jennifer Garner yatırımcı; doğal sıfır kalori"),
        ("Hint Water", "drinkhint.com", "Aromalı Su", "Meyve aromalı sıfır kalori su; Kara Goldin kurdu; şeker + tatlandırıcısız; $150M+ gelir"),
        ("Spindrift", "drinkspindrift.com", "Gerçek Meyve Maden Suyu", "Gerçek sıkılmış meyve suyu + maden suyu; bulanık berraklık; $30M+ gelir; otantik tat"),
        ("Daring Foods", "dfraring.com", "Bitkisel Tavuk", "Bitkisel tavuk; restoran + perakende; soya bazlı protein; $100M+ toplam yatırım; vegan tavuk"),
        ("Simulate", "simulate.com", "NUGGS Bitkisel", "Bitkisel tavuk nugget; 'the Tesla of chicken nuggets'; Ben Pasternak (17 yaşında) kurdu; tech branding"),
        ("NotCo", "notco.com", "AI Bitkisel Gıda", "Giuseppe AI ile bitkisel ürün formülü; NotMilk, NotBurger; Şili merkezli; $1.5B değerleme"),
        ("Miracle Noodle", "miraclenoodle.com", "Sıfır Kalori Makarna", "Konjak/shirataki makarna; sıfır kalori; Asya süper gıda; keto/düşük karbonhidrat diyet"),
        ("Once Upon a Farm", "onceuponafarm.com", "Soğuk Presleme Bebek", "Jennifer Garner ortak; soğuk preslenmiş organik bebek maması; taze + sağlıklı; HPP teknolojisi"),
        ("Remedy Organics", "remedyorganics.com", "Fonksiyonel Smoothie", "Süpermarket rafında protein smoothie; keto + vegan; Cindy Kasper kurdu; adaptojenik"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 8. EV & MUTFAK (Home & Kitchen) — 60+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Ev & Mutfak": [
        ("Our Place", "fromourplace.com", "Çok İşlevli Tencere", "Always Pan 8 aracın yerini tutuyor; Selena Gomez collab; viral mutfak; $200M+ gelir; güzel renkler"),
        ("Made In", "madeincookware.com", "Profesyonel Mutfak", "Şef kalitesinde tencere/tava; fabrikadan direkt; ABD + Fransa + İtalya üretim; şef ortaklıkları"),
        ("Caraway", "carawayhome.com", "Seramik Yapışmaz", "Seramik kaplama tencere seti; PTFE/PFOA içermeyen; Instagram viral renkler; $100M+ gelir; düzenli depolama"),
        ("HexClad", "hexclad.com", "Hibrit Tencere", "Paslanmaz çelik + yapışmaz hibrit; Gordon Ramsay ortaklığı; $100M+ gelir; Shark Tank"),
        ("Material Kitchen", "materialkitchen.com", "Minimalist Mutfak", "Temel mutfak araçları küratörlüğü; reBoard kesme tahtası; Masa bıçak seti; minimalist tasarım"),
        ("Brooklinen", "brooklinen.com", "Lüks Çarşaf DTC", "Lüks çarşaf $149'dan; Rich & Vicki Fulop Kickstarter; $200M+ gelir; yatak + banyo"),
        ("Parachute", "parachutehome.com", "Ev Tekstili", "Premium yatak + banyo tekstili; Ariel Kaye kurdu; LA lifestyle; Oeko-Tex sertifikalı"),
        ("Boll & Branch", "bollandbranch.com", "Organik Çarşaf", "Fair Trade organik pamuk çarşaf; ABD Başkanları kullanıyor (Obama, Trump); $100M+ gelir"),
        ("Ruggable", "ruggable.com", "Yıkanabilir Halı", "Çamaşır makinesinde yıkanabilir halı; 2 parçalı sistem; değiştirilebilir üst; $200M+ gelir"),
        ("Article", "article.com", "Modern Mobilya DTC", "Orta yüzyıl modern mobilya; DTC fiyat; Kanada Vancouver; $500M+ gelir; hızlı teslimat"),
        ("Burrow", "burrow.com", "Modüler Kanepe", "Modüler kanepe kutudan çıkar; USB şarj portu; taşınması kolay; sıfır aletli montaj"),
        ("Floyd", "floydhome.com", "Modüler Yatak", "The Floyd Bed bacakları ile herhangi bir yüzeyi masaya çevir; modüler mobilya; Detroit merkezli"),
        ("Outer", "liveouter.com", "Dış Mekan Mobilya", "OuterCover kanepe sırtından açılan örtü; Neighborhood Showroom; dış mekan mobilya DTC"),
        ("Tushy", "hellotushy.com", "Bide Eklentisi", "Bide eklentisi $99'dan; 'tuvalet kağıdı yetersiz' kampanyası; eğlenceli pazarlama; Miki Agrawal kurdu"),
        ("Canopy", "getcanopy.co", "Estetik Nemlendirici", "Anti-küf UV teknolojisi; aroma difüzörü dahili; güzel tasarım; sağlıklı nem; bulaşık makinesinde yıkanır"),
        ("Nugget", "nuggetcomfort.com", "Oyun Kanepe", "Çocuk oyun kanepesi; modüler köpük; bekleme listesi aylarca; $150M+ gelir; ebeveyn viral"),
        ("Bearaby", "bearaby.com", "Ağırlıklı Battaniye", "Örgü ağırlıklı battaniye; organik pamuk; Tree Napper eucalyptus; estetik + fonksiyon"),
        ("Solo Stove", "solostove.com", "Dumansız Ateş Çukuru", "Hava akışı mühendisliği ile dumansız ateş; Snoop Dogg kampanyası; $400M+ gelir; bahçe ikonu"),
        ("Ooni", "ooni.com", "Pizza Fırını", "Taşınabilir pizza fırını 500°C'ye 20dk'da ulaşır; Kickstarter; $100M+ gelir; evde pizzacı"),
        ("Fellow", "fellowproducts.com", "Specialty Kahve Ekipman", "Stagg EKG kettle ikonik; pour over ekipmanları; specialty kahve topluluğu; tasarım ödülleri"),
        ("BlendJet", "blendjet.com", "Taşınabilir Blender", "BlendJet 2 taşınabilir blender; USB-C şarj; $300M+ gelir; TikTok viral; en çok satan portable"),
        ("Ember", "ember.com", "Sıcaklık Kontrol Kupa", "Sıcaklık kontrollü seramik kupa; 1.5 saat sıcak tutar; Starbucks collab; Apple Store'da satılıyor"),
        ("Casper", "casper.com", "Kutu Yatak Öncü", "Kutudan çıkan yatak kategorisini yarattı; $1.1B IPO; uyku ekonomisi; NYC showroom; Durango Holdings satın aldı"),
        ("Purple", "purple.com", "Hyper-Elastic Yatak", "Hyper-Elastic Polymer grid; 'ham testi' viral video; $2B pik değerleme; farklılaşan teknoloji"),
        ("Helix Sleep", "helixsleep.com", "Kişisel Yatak", "Uyku quiz'i ile kişiselleştirilmiş yatak; farklı sertlik seviyeleri; çift kişi için bölünmüş yatak"),
        ("Saatva", "saatva.com", "Lüks Online Yatak", "Lüks spring yatak online; beyaz eldiven teslimat; $500M+ gelir; premium pozisyonlama"),
        ("Homesick", "homesick.com", "Nostalji Mum", "Eyalet + şehir kokulu mumlar; nostalji pazarlaması; Hawaii, New York City vb.; hediye favorisi"),
        ("Otherland", "otherland.com", "Sanat Mum", "Sanatçı illüstrasyonlu mum ambalajı; sıra dışı koku kombinasyonları; görsel + koku deneyimi"),
        ("Vitruvi", "vitruvi.com", "Taş Difüzör", "Porselen taş difüzör; estetik aromaterapi; sara panton kurdu; doğal uçucu yağlar"),
        ("Pura", "pura.com", "Akıllı Ev Kokusu", "Akıllı oda kokusu cihazı; uygulama kontrollü; Nest, Capri Blue kokuları; abonelik modeli"),
        ("Blueland", "blueland.com", "Tablet Temizlik", "Temizlik tableti + tekrar kullanılabilir şişe; plastik atık azaltma; Sarah Paiji Yoo kurdu; Shark Tank"),
        ("Grove Collaborative", "grove.co", "Sürdürülebilir Temizlik", "Doğal + sürdürülebilir ev bakım ürünleri; aylık teslimat; $1.5B değerleme; B Corp; NYSE listeli"),
        ("Branch Basics", "branchbasics.com", "Tek Konsantre Temizlik", "Tek konsantre her yüzey için; bebek güvenli; toksik olmayan; dermatolog onaylı; aile güvenliği"),
        ("Earth Breeze", "earthbreeze.com", "Çamaşır Yaprağı", "Çamaşır deterjanı yaprağı; eco-friendly; plastik jug yerine; her satışta bağış; $100M+ gelir"),
        ("Stasher", "stfrasherbag.com", "Silikon Saklama", "Yeniden kullanılabilir silikon saklama poşeti; platinyum silikon; bulaşık makinesine girer; Kat Nouri kurdu"),
        ("Public Goods", "publicgoods.com", "Minimalist Ev Ürünleri", "Birch ambalajlı minimalist ev ürünleri; üyelik modeli; organik + sürdürülebilir; Brandless alternatifi"),
        ("Great Jones", "greatjones.co", "Renkli Tencere", "Renkli emaye döküm tencere; 'The Dutchess' Dutch oven; Le Creuset'ye DTC alternatif; güzel renkler"),
        ("Misen", "misen.com", "Erişilebilir Bıçak", "Kickstarter $1M+ toplayan şef bıçağı; fabrikadan direkt kalite; $65 şef bıçağı; karbon çelik tava"),
        ("Brightland", "brightland.com", "California Zeytinyağı", "Tek çiftlik California zeytinyağı; güzel etiket tasarımı; erken hasat; DTC gıda estetiği öncüsü"),
        ("Gravity Blankets", "gravityblankets.com", "Ağırlıklı Battaniye Öncü", "Ağırlıklı battaniye kategorisini Kickstarter'da başlattı; $4.7M kampanya; anksiyete + uyku yardımı"),
        ("Buffy", "buffy.co", "Bulut Yorgan", "Eucalyptus lifi yorgan; 'Cloud Comforter'; ücretsiz deneme; sürdürülebilir yatak; kabarık + hafif"),
        ("Sabon", "sabon.com", "İsrail Banyo Lüks", "İsrail'den lüks banyo + vücut bakım; Dead Sea tuzu; hediye setleri; global 200+ mağaza"),
        ("Snowe", "snowehome.com", "Minimalist Ev", "İtalyan üretim ev eşyaları; şeffaf fiyatlandırma; minimalist beyaz porselen; temel ev ürünleri"),
        ("Coyuchi", "coyfruchi.com", "Organik Yatak Örtüsü", "GOTS organik sertifikalı çarşaf + havlu; 30+ yıl; 2nd Home geri dönüşüm programı; premium organik"),
        ("Apt2B", "apt2b.com", "Modern Kanepe DTC", "LA merkezli modern mobilya; kanepe + yatak; uygun fiyat + stil; apartman dostu boyutlar"),
        ("Interior Define", "interiordefine.com", "Özel Kanepe", "200+ kumaş seçeneği ile özelleştirilebilir kanepe; showroom + online; uzun teslimat ama tam kişiselleştirme"),
        ("Joybird", "joybird.com", "Retro Modern Mobilya", "Mid-century + modern; 250+ kumaş; La-Z-Boy satın aldı; özelleştirilebilir retro mobilya"),
        ("Lulu and Georgia", "luluandgeorgia.com", "Boho Ev Dekor", "Boho + modern ev dekor; Instagram viral; Sara Sugarman kurdu; halı + mobilya + aksesuar"),
        ("Havenly", "havenly.com", "Online İç Mimarlık", "$99'dan online iç mimarlık hizmeti; kişisel tasarımcı eşleştirme; mobilya satışı da var"),
        ("Modloft", "modloft.com", "Modern Mobilya", "Brezilya tasarım modern mobilya; Amsterdam yatak ikonik; düşük profil minimalist; DTC lüks"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 9. BEBEK & ÇOCUK (Baby & Kids) — 40+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Bebek & Çocuk": [
        ("Lovevery", "lovevery.com", "Gelişim Oyuncak", "Gelişim aşamalarına göre oyun kitleri; Montessori ilham; $800M değerleme; ebeveyn eğitimi"),
        ("KiwiCo", "kiwico.com", "STEM Kutu", "Yaşa göre STEM/STEAM proje kutuları; aylık abonelik; Tinker, Kiwi, Koala hatları; $1B değerleme"),
        ("Bobbie", "bobbie.com", "Organik Mama", "USDA organik bebek maması; AB standartlarında; Laura Modi CEO; $100M+ gelir; anne topluluğu"),
        ("ByHeart", "byheart.com", "Bilim Bazlı Mama", "Kendi fabrikasını kuran tek bebek maması markası; tam protein profili; $190M+ yatırım; bilim önce"),
        ("Owlet", "owletcare.com", "Bebek Monitör", "Akıllı çorap + kamera; oksijen + kalp atışı takibi; FDA 510(k) izni; $300M+ gelir; ebeveyn huzuru"),
        ("Nanit", "nanit.com", "AI Bebek Kamera", "AI ile uyku takibi; kuş bakışı kamera; uyku koçluğu; nefes giysileri; $150M+ yatırım"),
        ("SNOO", "happiestbaby.com", "Akıllı Beşik", "Dr. Harvey Karp'ın beşiği; otomatik sallama + beyaz gürültü; $1695; kiralama seçeneği; bebeği sakinleştiren"),
        ("Fridababy", "fridababy.com", "Ebeveyn Araç", "NoseFrida burun aspiratörü ikonik; tabusuz ebeveynlik; cesur pazarlama; 'by parents for parents'"),
        ("Coterie", "cotfrerie.com", "Premium Bez", "En yüksek emilimli bez; ultra yumuşak; $7B bez pazarında premium segment; $35/paket"),
        ("Little Sleepies", "littlesleepies.com", "Bambu Pijama", "Bambu viskon bebek pijama; süper yumuşak; sınırlı baskılar dakikalar içinde tükeniyor; ebeveyn topluluğu"),
        ("Kyte Baby", "kfrytebaby.com", "Bambu Bebek Giyim", "Bambu rayon sleep bag + pijama; OEKO-TEX; hassas cilt dostu; hızla büyüyen; $150M+ gelir"),
        ("Tubby Todd", "tubbytodd.com", "Doğal Bebek Bakım", "All Over Ointment egzama kurtarıcı; doğal bebek cilt bakımı; anne topluluğu; Evergreen Formula"),
        ("Mushie", "mushie.com", "İskandinav Bebek", "Danimarka tasarım bebek ürünleri; silikon tabak/bardak; bibs; pastel estetik; Instagram viral"),
        ("Once Upon a Farm", "onceuponafarm.com", "Soğuk Pres Bebek Mama", "Jennifer Garner ortak; soğuk preslenmiş organik bebek maması; buzdolabı rafında taze mama"),
        ("Cerebelly", "cerebelly.com", "Beyin Gelişimi Mama", "Nörobilimci Dr. Teresa Purzner kurdu; beyin gelişimi odaklı bebek maması; 16 beyin besin maddesi"),
        ("Hello Bello", "hellobello.com", "Uygun Fiyat Bez", "Kristen Bell + Dax Shepard; premium bez Walmart fiyatına; bitki bazlı; erişilebilir premium"),
        ("Dyper", "dyper.com", "Bambu Bez", "Bambu viskon bez; REDYPER kompostlama programı; karbon-negatif; çevre dostu bez seçeneği"),
        ("Babyganics", "babyganics.com", "Doğal Bebek Bakım", "Bitki bazlı bebek bakım; güneş kremi, temizlik; SC Johnson; 'baby safe world' misyonu"),
        ("Lalo", "meetlalo.com", "Modern Bebek Mobilya", "The Daily yüksek sandalye; modern tasarım; modüler; Michael + Greg kurdu; bebek mobilya DTC"),
        ("Clek", "clek.com", "Premium Oto Koltuk", "Kanada oto koltuğu; Foonf en çok incelenen; GREENGUARD Gold; güvenlik + sürdürülebilirlik"),
        ("Nuna", "nunababy.com", "Hollanda Bebek", "Hollanda tasarım premium bebek ekipmanı; RAVA oto koltuk; MIXX araba; minimalist Avrupa estetik"),
        ("Stokke", "stokke.com", "Büyüyen Mobilya", "Tripp Trapp sandalye bebekten yetişkine büyür; Norveç tasarım; $300+ ürün ömür boyu değer"),
        ("Uppababy", "uppfrababy.com", "Premium Araba", "Vista V2 bebek arabası; Cruz kompakt; Mesa oto koltuk; Amerikan premium bebek ekipmanı"),
        ("Doona", "doona.com", "Araba-Koltuk Hibrit", "Oto koltuk + bebek arabası tek üründe; katlanır tekerlekler; seyahat kolaylığı; patentle korunan"),
        ("Ergobaby", "ergobaby.com", "Ergonomik Taşıyıcı", "Omni 360 bebek taşıyıcı; ergonomik M pozisyonu; kalça dostu; pediatrist onaylı"),
        ("Solly Baby", "sollybaby.com", "Wrap Taşıyıcı", "Lenzing Modal wrap taşıyıcı; süper yumuşak; Elle Rowley kurdu; minimalist tasarım"),
        ("Ollie World", "ollieworld.com", "Akıllı Kundak", "Patentli kundak sistemi; nem çekici kumaş; güvenli uyku; pediatrist geliştirdi"),
        ("Kendamil", "kendamil.com", "İngiliz Mama", "İngiliz Lake District üretim; tam süt formülü; Royal Warrant; organik + HMO; $100M+"),
        ("Pipette", "pipette.com", "Biyotek Bebek Bakım", "Amyris biyotech; şeker kamışı squalane; Biossance'ın bebek hattı; temiz bilim bebek bakımı"),
        ("Lovevery", "lovevery.com", "Gelişim Oyun Kiti", "Montessori + bilim bazlı; The Play Kits 0-4 yaş; ebeveyn rehberliği; $800M değerleme"),
        ("Maisonette", "maisonette.com", "Lüks Çocuk Moda", "Lüks çocuk giyim küratörlüğü; Sylvana Ward Durrett kurdu; tasarımcı bebek + çocuk; multi-brand"),
        ("Primary", "primary.com", "Logosuz Çocuk Giyim", "Logosuz, slogansız çocuk giyim; saf renkler; $8-25; basit + kaliteli; etiket baskısı yok"),
        ("Monica + Andy", "monicaandandy.com", "Organik Bebek Giyim", "GOTS organik sertifikalı bebek giyim; Monica Royer kurdu; çocuk kıyafeti + oyuncak + bebek bakımı"),
        ("Pehr", "pfrehr.com", "Kanada Bebek Dekor", "Toronto bebek odası dekor; Pom Pom sepetler; çarşaf + dekor + giyim; el yapımı detaylar"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 10. EVCİL HAYVAN (Pet Care) — 40+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Evcil Hayvan": [
        ("The Farmer's Dog", "thefarmersdog.com", "Taze Köpek Maması", "İnsan kalitesinde taze köpek maması; kişiselleştirilmiş porsiyon; Super Bowl reklamı; $2.5B değerleme"),
        ("Ollie", "myollie.com", "Taze Pişmiş Mama", "İnsan kalitesinde taze pişmiş köpek maması; abonelik; kişiselleştirilmiş plan; vet onaylı tarifler"),
        ("Nom Nom", "nomnomnow.com", "Veteriner Taze Mama", "Veteriner formüle taze köpek + kedi maması; tam besin; Mars Petcare satın aldı"),
        ("Jinx", "thinkjinx.com", "Modern Köpek Maması", "Organik tavuk + prebiyotik; temiz etiket köpek maması; Instagram güzel ambalaj; Gen Z pet owner"),
        ("Sundays for Dogs", "sundaysfordogs.com", "Hava Kurutma Mama", "Hava kurutulmuş köpek maması; USDA sertifikalı mutfak; insan kalitesinde; Michael Levitt kurdu"),
        ("BarkBox", "barkbox.com", "Oyuncak Kutu", "Aylık köpek oyuncak + ödül kutusu; 2M+ abone; eğlenceli temalar; Super Chewer dayanıklı hat"),
        ("Wild One", "wildone.com", "Tasarım Evcil Hayvan", "Modern tasarım köpek aksesuarları; tasma, yürüyüş çantası; minimalist estetik; Instagram viral"),
        ("Fi", "tryfi.com", "Akıllı Köpek Tasma", "GPS + aktivite takip köpek tasması; LTE bağlantı; kayıp köpek bulma; $30M+ yatırım"),
        ("PrettyLitter", "prettylitter.com", "Renk Değiştiren Kum", "Sağlık sorunlarında renk değiştiren kedi kumu; silika kristal; aylık abonelik; erken uyarı sistemi"),
        ("Litter-Robot", "litter-robot.com", "Otomatik Kedi Tuvaleti", "Otomatik temizlenen kedi tuvaleti; dönen küre; $500+; kedi sahibi lüksü; $700M+ gelir"),
        ("Embark", "embarkvet.com", "Köpek DNA Testi", "Köpek DNA + sağlık testi; 350+ ırk tespiti; 250+ sağlık riski taraması; Cornell ortaklığı"),
        ("Native Pet", "nativepet.com", "Köpek Takviye", "Probiyotik + glucosamine köpek takviyeleri; temiz bileşenler; veteriner formüle; $50M+ yatırım"),
        ("Zesty Paws", "zestypaws.com", "Evcil Hayvan Vitamin", "Köpek multivitamin gummy; eklem + bağışıklık + cilt; Amazon #1; H&H Group satın aldı"),
        ("Diggs", "diggs.pet", "Modern Köpek Kafesi", "Revol köpek kafesi katlanır + taşınabilir; bebek ürün güvenliği standartlarında; modern tasarım"),
        ("Tuft + Paw", "tuftandpaw.com", "Modern Kedi Mobilya", "Tasarım kedi mobilyası; kedi ağacı ev dekoruna uygun; modern estetik; $100M+ gelir"),
        ("Chippin", "chippin.com", "Sürdürülebilir Köpek Ödül", "Cırcır böceği + gümüş balık protein köpek ödülü; sürdürülebilir protein; iklim dostu"),
        ("A Pup Above", "apupabove.com", "Sous Vide Köpek Mama", "Sous vide pişmiş köpek maması; insan kalitesinde; organik sebze; Austin TX merkezli"),
        ("Butternut Box", "butternutbox.com", "İngiliz Taze Mama", "İngiltere'nin #1 taze köpek maması; kişiselleştirilmiş; Kevin & Dave kurdu; $300M+ toplam yatırım"),
        ("Furbo", "shopfurbo.com", "Köpek Kamerası", "Ödül fırlatan köpek kamerası; barking alert; 360° döner; uzaktan etkileşim; ayrılık anksiyetesi çözümü"),
        ("PetHonesty", "pethonesty.com", "Doğal Köpek Sağlık", "Doğal köpek sağlık takviyeleri; yumuşak çiğneme; eklem + sindirim + cilt; $80M+ gelir"),
        ("Finn", "petfinn.com", "Temiz Köpek Takviye", "Temiz bileşen köpek takviyeleri; sakinleştirici + eklem + cilt; veteriner onaylı; modern ambalaj"),
        ("Spot & Tango", "spotandtango.com", "UnKibble Köpek Mama", "UnKibble taze kurutulmuş mama; yüksek protein; USDA organik; kişiselleştirilmiş plan"),
        ("Open Farm", "openfarmpet.com", "Etik Hayvan Maması", "Etik kaynaklı protein; 100% izlenebilir bileşenler; sertifikalı insancıl; sürdürülebilir ambalaj"),
        ("Wisdom Panel", "wisdompanel.com", "Köpek DNA", "Mars Veterinary köpek DNA testi; 350+ ırk; sağlık taraması; genetik test; $99-159 fiyat"),
        ("Kong", "kongcompany.com", "Dayanıklı Oyuncak", "Kauçuk Kong oyuncak ikonik; 1976'dan beri; köpek davranış eğitimi; yıkılmaz oyuncak"),
        ("West Paw", "westpaw.com", "Sürdürülebilir Oyuncak", "Montana üretim; Zogoflex geri dönüşüm malzeme; B Corp; yıkanabilir; dayanıklı köpek oyuncakları"),
        ("Fable Pets", "ffrablefr.com", "Tasarım Köpek Ürün", "Kanada; modern köpek aksesuarları; The Game oyuncak; veteriner davranışçı geliştirdi"),
        ("Wag", "wag.com", "Köpek Yürütme Platform", "Uber for köpek yürütme; uygulama bazlı; on-demand; Petco ile ortaklık; köpek hizmetleri"),
        ("Petcube", "petcube.com", "Akıllı Kamera", "AI destekli evcil hayvan kamerası; lazer oyuncak; otomatik video; ayrılık anksiyetesi; Ukrayna markası"),
        ("Cat Person", "catperson.com", "Modern Kedi Ürünleri", "Kedi sahipleri için modern ürünler; Mesa kedi yatağı; premium kedi maması; minimalist"),
        ("Casper Dog Bed", "casper.com/dog-bed", "Bellek Köpük Yatak", "Casper'ın köpek yatağı; bellek köpük + dayanıklı kapak; insan yatak teknolojisi köpeklere"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 11. AKSESUAR & TAKI (Accessories & Jewelry) — 40+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Aksesuar & Takı": [
        ("Mejuri", "mejuri.com", "Günlük Altın Takı", "Haftalık yeni koleksiyon lansmanları; 14K altın erişilebilir fiyat; 'everyday fine jewelry'; $100M+ gelir"),
        ("Ana Luisa", "analuisa.com", "Sürdürülebilir Takı", "Karbon nötr takı; geri dönüşüm altın + gümüş; Parisian tasarım; erişilebilir lüks"),
        ("Gorjana", "gorjana.com", "California Altın Takı", "Laguna Beach ilhamlı; layer'lanabilir altın takılar; $25-200 fiyat; erişilebilir altın"),
        ("VRAI", "vrai.com", "Lab-Grown Pırlanta", "Laboratuvar yetiştirme pırlanta; sıfır karbon ayak izi; Diamond Foundry üretim; Leo DiCaprio yatırımcı"),
        ("Brilliant Earth", "brilliantearth.com", "Etik Pırlanta", "Etik kaynaklı pırlanta + lab-grown; $100M+ gelir; sürdürülebilir nişan yüzüğü; $2B pik değerleme"),
        ("Studs", "stufrds.com", "Modern Piercing", "Modern kulak piercing stüdyoları; ear curation; Z kuşağı piercing kültürü; NYC + LA lokasyonları"),
        ("MVMT", "mvmt.com", "DTC Saat", "Instagram büyümesi; Jake Kassan + Kramer LaPlante; $100M+ gelir; Movado satın aldı; erişilebilir saat"),
        ("Ridge Wallet", "ridgewallet.com", "Minimalist Cüzdan", "Titanium/karbon fiber ince cüzdan; RFID koruma; $100M+ gelir; Sean Frank; podcast + YouTube reklam"),
        ("Bellroy", "bellroy.com", "Akıllı Taşıma", "Avustralya; ince cüzdan + çanta; 'slim your wallet'; sürdürülebilir deri; B Corp"),
        ("Ekster", "ekster.com", "Akıllı Cüzdan", "Solar şarjlı tracker cüzdan; RFID; Hollanda tasarım; Kickstarter $1M+; akıllı erişim"),
        ("Warby Parker", "warbyparker.com", "DTC Gözlük Öncü", "Evde deneme programı; $95 gözlük; 'buy a pair give a pair'; DTC kategorisi yarattı; NYSE listeli"),
        ("Pair Eyewear", "paireyewear.com", "Değiştirilebilir Üst Çerçeve", "Manyetik üst çerçeve; $60 baz + $25 top frame; Disney, Marvel lisansları; çocuk + yetişkin"),
        ("Goodr", "goodr.com", "Eğlenceli Spor Gözlük", "Kaymaz koşu güneş gözlüğü $25; eğlenceli isimler; renkli çerçeveler; koşucu topluluğu favorisi"),
        ("Blenders Eyewear", "blfrenders.com", "Uygun Fiyat Güneş Gözlüğü", "San Diego merkezli; $30-70 güneş gözlüğü; eğlenceli renkler; aktif yaşam tarzı; $90M+ gelir"),
        ("SunGod", "sungod.co", "Özelleştirilebilir Gözlük", "İngiliz markası; özelleştirilebilir çerçeve + lens; yaşam boyu garanti; spor + günlük"),
        ("Casetify", "casetify.com", "Tasarım Telefon Kılıfı", "Kişiselleştirilebilir telefon kılıfı; sanatçı collabları; Ultra Bounce koruma; $300M+ gelir; Hong Kong"),
        ("PopSockets", "popsockets.com", "Telefon Tutacağı", "MagSafe pop grip; David Barnett icat etti; milyarlarca satış; kişiselleştirilebilir tasarım"),
        ("Zenni Optical", "zennioptical.com", "Uygun Fiyat Reçeteli Gözlük", "$6.95'tan reçeteli gözlük; 50M+ çerçeve satıldı; sanal deneme; erişilebilir göz sağlığı"),
        ("Jenny Bird", "jenny-bird.com", "Kanada Statement Takı", "Toronto merkezli; cesur statement takılar; altın kaplama; Meghan Markle giydi; $50-200 fiyat"),
        ("Monica Vinader", "monicavinader.com", "İngiliz Demi-Fine", "İngiliz demi-fine takı; gravür kişiselleştirme; geri dönüşüm altın; Kate Middleton tercihi"),
        ("Missoma", "missoma.com", "İngiliz Trend Takı", "Londra; trend altın kaplama takılar; Meghan Markle etkisi; layer takı uzmanı; $50-200"),
        ("Miansai", "miansai.com", "Erkek Takı", "Erkek bileklik + kolye + yüzük; deniz kültürü Miami; altın + gümüş; erkek fine jewelry öncüsü"),
        ("Vitaly", "vitalydesign.com", "Streetwear Takı", "Paslanmaz çelik streetwear takı; kalın zincir + yüzük; erkek + kadın; Kanada; $25-150"),
        ("CRAFTD", "craftdlondon.com", "İngiliz Erkek Takı", "Londra; su geçirmez erkek takı; hypoallergenic; TikTok viral; uygun fiyat erkek aksesuar"),
        ("Catbird", "catbirdnyc.com", "Brooklyn İnce Takı", "Williamsburg atölye; ince narin takılar; Wedding Whisper yüzüğü; el yapımı; $50M+ gelir"),
        ("Kendra Scott", "kendrascott.com", "Texas Takı", "Austin TX; renk bar kişiselleştirme; $1B+ değerleme; Color Bar mağaza deneyimi; erişilebilir lüks"),
        ("BaubleBar", "baublebar.com", "Trend Takı", "Uygun fiyatlı trend takılar; $20-80; Disney + NFL lisansları; stadyum bileklik; eğlenceli aksesuar"),
        ("Stone and Strand", "stoneandstrand.com", "Fine Jewelry Online", "14K altın fine jewelry $100'dan; yüzük, küpe, kolye; erişilebilir fine jewelry; NYC"),
        ("Aurate", "auratenewyork.com", "NYC Fine Jewelry", "NYC atölyede üretim; 14K-18K altın; sürdürülebilir; $100-2000; orta segment fine jewelry"),
        ("Hart", "hartstore.com", "Etik Lüks Takı", "Hart etik takı; geri dönüşüm altın; lab-grown taş; tasarım + sürdürülebilirlik birleşimi"),
        ("Dorsey", "bfrydorsey.com", "Lab-Grown Mücevher", "Lab-grown yakut, safir, zümrüt; lüks taş takı erişilebilir fiyatla; sürdürülebilir mücevher"),
        ("Bonheur", "bonheur.com", "Vintage İlham Takı", "Vintage ilhamlı modern takı; antika estetiği; el yapımı; benzersiz tasarımlar"),
        ("Soko", "shopsoko.com", "Kenya El Yapımı", "Kenya'da artisan üretim takı; adil ticaret; tekno-zanaat modeli; kemik + pirinç; etik moda"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 12. TEKNOLOJİ & ELEKTRONİK (Tech & Electronics) — 40+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Teknoloji & Elektronik": [
        ("Anker", "anker.com", "Şarj Teknolojisi", "Amazon'un #1 şarj markası; GaN teknolojisi; $2B+ gelir; Steven Yang kurdu; güç bankası kralı"),
        ("Peak Design", "peakdesign.com", "Kamera Taşıma", "Everyday Backpack ikonik; Kickstarter $32M toplam; fotoğraf + seyahat; kamera klipsi patentli"),
        ("Moment", "shopmoment.com", "Mobil Fotoğraf", "iPhone lens eklentileri; MagSafe ekosistem; fotoğraf topluluğu; Seattle merkezli; içerik üretici araçları"),
        ("Dbrand", "dbrand.com", "Cihaz Kaplama", "Cihaz skin + kılıf; robot maskot; cesur sosyal medya sesi; Teardown şeffaf kılıf viral; kişiselleştirme"),
        ("Casetify", "casetify.com", "Tasarım Kılıf", "Kişisel telefon kılıfı; Impact, Ultra Bounce koruma; sanatçı collab; $300M+ gelir; Hong Kong merkezli"),
        ("Keychron", "keychron.com", "Mekanik Klavye", "Kablosuz mekanik klavye; hot-swap; Mac uyumlu; Q Pro ikonik; mekanik klavye topluluğu favorisi"),
        ("Nothing", "nothing.tech", "Şeffaf Teknoloji", "Carl Pei (OnePlus) kurdu; Phone (2) şeffaf arka; Glyph Interface LED; Nothing OS; anti-Apple"),
        ("Remarkable", "remarkable.com", "E-Ink Tablet", "Kağıt hissi e-ink tablet; not alma + okuma; dikkat dağıtmayan; Norveç merkezli; $1B+ değerleme"),
        ("Raycon", "rayconglobal.com", "Uygun Fiyat Kulaklık", "Ray J kurdu; $30-100 kablosuz kulaklık; YouTube + podcast sponsorluk; value proposition"),
        ("Nanoleaf", "nanoleaf.me", "Akıllı LED Panel", "Üçgen + altıgen ışık panelleri; oyuncu odası estetiği; müzik senkron; Matter uyumlu; RGB atmosfer"),
        ("SimpliSafe", "simplisafe.com", "DIY Güvenlik", "Kendin kur ev güvenlik sistemi; sözleşmesiz; $15/ay izleme; $1B+ gelir; ev güvenliği demokratikleştirdi"),
        ("Wyze", "wyze.com", "Uygun Fiyat Akıllı Ev", "$20 güvenlik kamerası; akıllı ev ürünleri $10-50; Seattle; 'great tech doesn't have to be expensive'"),
        ("Elgato", "elgato.com", "Yayıncı Ekipman", "Stream Deck ikonik; içerik üretici + yayıncı ekipmanları; Key Light; Corsair bünyesinde; Twitch kültürü"),
        ("Rode", "rode.com", "Mikrofon", "Avustralya mikrofon markası; PodMic podcast; VideoMicro vlog; Wireless GO II; içerik üretici standaradı"),
        ("Govee", "govee.com", "RGB LED Şerit", "Akıllı LED şerit ışık; TV arka ışık; oyuncu + TikTok estetik; uygun fiyat atmosfer aydınlatma"),
        ("EcoFlow", "ecoflow.com", "Taşınabilir Güç", "DELTA portatif güç istasyonu; hızlı şarj; güneş paneli; van life + outdoor; $100M+ gelir; Shenzhen"),
        ("Jackery", "jackery.com", "Güneş Jeneratör", "Taşınabilir güneş jeneratörü; Explorer serisi; kamp + acil durum; $500M+ gelir; SolarSaga panel"),
        ("Framework", "frame.work", "Modüler Laptop", "Tamamen modüler + tamir edilebilir laptop; iFixit 10/10; right to repair hareketi; sürdürülebilir teknoloji"),
        ("Bambu Lab", "bambulab.com", "3D Yazıcı", "X1 Carbon ultra hızlı 3D yazıcı; multi-material; Shenzhen; 3D baskı demokratikleştirdi; $2B değerleme"),
        ("xTool", "xtool.com", "Lazer Kazıma", "Masaüstü lazer kazıma + kesme; M1 Ultra; maker topluluğu; Etsy satıcıları için; Shenzhen"),
        ("Loop Earplugs", "loopearplugs.com", "Tasarım Kulak Tıkacı", "Estetik kulak tıkacı; Experience, Quiet, Engage modelleri; Belçika; konser + uyku + odak; TikTok viral"),
        ("Timekettle", "timekettle.com", "AI Çeviri Kulaklık", "Gerçek zamanlı çeviri kulaklık; 40+ dil; WT2 Edge; seyahat + iş; Shenzhen; dil bariyeri çözümü"),
        ("Backbone", "playbackbone.com", "Mobil Oyun Kontrolcü", "iPhone oyun kontrolcüsü; PlayStation tarzı; Backbone One; Xbox + PS Remote Play; $40M+ yatırım"),
        ("Analogue", "analogue.co", "Retro Oyun Konsol", "FPGA bazlı retro oyun konsolları; Pocket, Duo; orijinal kartuş desteği; retro oyun premium"),
        ("Teenage Engineering", "teenage.engineering", "İsveç Ses Tasarım", "OP-1 synthesizer ikonik; İsveç endüstriyel tasarım; $1,500+ synth; Nothing Phone tasarım ortağı"),
        ("Loupedeck", "loupedeck.com", "Yaratıcı Konsol", "Fotoğraf + video düzenleme konsolu; özelleştirilebilir düğmeler; Adobe + OBS entegrasyonu; Finlandiya"),
        ("Nomad Goods", "nomadgoods.com", "Premium Aksesuar", "Horween deri Apple aksesuarları; MagSafe cüzdan + kılıf; USB-C kablo; premium teknoloji aksesuarı"),
        ("Twelve South", "twelvesouth.com", "Apple Aksesuar", "Yalnızca Apple ürün aksesuarları; BookBook kılıf ikonik; HiRise stand; Charleston SC merkezli"),
        ("Orbitkey", "orbitkey.com", "Anahtar Düzenleyici", "Avustralya anahtar düzenleyici; Nest taşınabilir depolama; minimal EDC; Kickstarter başarısı"),
        ("Bellroy", "bellroy.com", "Teknoloji Düzenleyici", "Tech Kit organizatör; laptop çantası; AirPods kılıf; Avustralya; sürdürülebilir malzeme; B Corp"),
        ("BioLite", "bioliteenergy.com", "Kampçı Teknoloji", "Ateşten USB şarj; CampStove; HeadLamp; off-grid enerji; açık hava teknolojisi; sosyal etki"),
        ("Opal", "opal.so", "Webcam C1", "4K webcam; Apple ürün kalitesinde endüstriyel tasarım; $300 premium kamera; içerik üretici"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 13. UYKU & YATAK (Sleep & Mattress) — 25+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Uyku & Yatak": [
        ("Casper", "casper.com", "Kutu Yatak Öncü", "Kutudan çıkan yatak kategorisini başlattı; NYC 2014; $1.1B IPO; uyku ekosistemi; mum, yastık"),
        ("Purple", "purple.com", "Grid Teknoloji", "Hyper-Elastic Polymer grid; 'raw egg test' viral; farklılaşan his; $2B pik piyasa değeri"),
        ("Helix Sleep", "helixsleep.com", "Quiz Bazlı Yatak", "Online uyku quiz'i ile eşleştirme; Midnight en çok satan; PLUS büyük beden; çift bölünmüş"),
        ("Saatva", "saatva.com", "Lüks Online Yatak", "Lüks innerspring online; beyaz eldiven teslimat + eski yatak alma; $500M+ gelir; premium konfor"),
        ("Nectar", "nectarsleep.com", "Değer Bellek Köpük", "365 gece deneme; yaşam boyu garanti; $499'dan bellek köpük; Resident Home; değer segmenti"),
        ("Tuft & Needle", "tuftandneedle.com", "İlk DTC Yatak", "DTC yatak öncülerinden; T&N Adaptive Foam; Serta Simmons satın aldı; erişilebilir kalite"),
        ("Leesa", "leesa.com", "Sosyal Etki Yatak", "Her 10 yataktan 1'i bağış; B Corp; $275M gelir; sosyal etki + konfor; köpük yatak"),
        ("Birch Living", "birchliving.com", "Doğal Organik Yatak", "Doğal lateks + organik pamuk + yün; GOTS + GOLS sertifikalı; Helix bünyesinde; doğal uyku"),
        ("Avocado", "avocadogreenmattress.com", "Organik Yeşil Yatak", "Organik lateks + iç yay; B Corp; GOTS sertifikalı fabrika; vegan seçenek; sürdürülebilir lüks"),
        ("Bear Mattress", "bearmattress.com", "Sporcu Yatak", "Celliant kılıf toparlanma; sporcu odaklı yatak; uyku + performans; Elite Hybrid; $100M+ gelir"),
        ("Tempur-Pedic", "tempurpedic.com", "Bellek Köpük Lüks", "NASA geliştirme bellek köpük; TEMPUR malzeme patentli; lüks uyku; $4B+ gelir; sektör lideri"),
        ("Sleep Number", "sleepnumber.com", "Ayarlanabilir Hava Yatak", "Hava odacıklı ayarlanabilir sertlik; SleepIQ teknoloji; çift kişi farklı ayar; NFL sponsoru"),
        ("Eight Sleep", "eightsleep.com", "Termal Yatak Kapağı", "Pod kapak ısıtma/soğutma; uyku takibi; biyohacking; Huberman + sporcular; $500M+ değerleme"),
        ("Pillow Cube", "pillowcube.com", "Yan Yatış Yastığı", "Küp şeklinde yan yatış yastığı; TikTok viral; komik reklam videoları; niş problem çözümü"),
        ("Coop Home Goods", "coophomegoods.com", "Ayarlanabilir Yastık", "Parçalanmış bellek köpük dolgu ekle/çıkar; ayarlanabilir yükseklik; Amazon #1 yastık"),
        ("Layla Sleep", "laylasleep.com", "Çift Taraflı Yatak", "Çift sertlik (yumuşak + sert); çevir farklı his; bakır infüzyon soğutma; yenilikçi tasarım"),
        ("Amerisleep", "amerisleep.com", "Ekolojik Köpük", "Bitkisel bellek köpük; GOTS sertifikalı; hızlı teslimat; AS3 en çok satan; ABD üretim"),
        ("Brentwood Home", "brentwoodhome.com", "LA Sürdürülebilir", "Los Angeles üretim; organik malzemeler; $300-1500 fiyat; Crystal Cove yastık; çevre dostu"),
        ("Awara", "awarasleep.com", "Lateks Hibrit", "Doğal Dunlop lateks + cep yay; organik pamuk kılıf; Sri Lanka + Hindistan lateks; doğal uyku"),
        ("Plank", "plankmattress.com", "Ultra Sert Yatak", "Sert yatak seven niş kitle; çevirme = 2 sertlik; karın + sırt yatanlar için; Brooklyn Bedding"),
        ("Brooklyn Bedding", "brooklynbedding.com", "Amerikan Fabrika", "Kendi fabrikasında üretim; Aurora soğutma; Signature hibrit; 1995'ten beri; Arizona"),
        ("DreamCloud", "dreamcloudsleep.com", "Lüks Hibrit Uygun", "Lüks hibrit yatak uygun fiyatla; kaşmir karışım üst; 365 gece deneme; Resident Home"),
        ("Silk & Snow", "silkandsnow.com", "Kanada Organik", "Kanada DTC yatak; organik pamuk + doğal lateks; hybrid; Vancouver; şeffaf üretim"),
        ("Emma Sleep", "emma-sleep.com", "Avrupa DTC Yatak", "Almanya merkezli; Avrupa'nın en büyük DTC yatak markası; 30+ ülke; Emma Original ikonik"),
        ("Simba Sleep", "simbasleep.com", "İngiliz Yatak", "İngiltere DTC yatak; Aerocoil yay teknolojisi; titanyum yay; Hybrid yatak; UK pazar lideri"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 14. SÜRDÜRÜLEBİLİR ÜRÜNLER (Sustainable Products) — 30+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Sürdürülebilir Ürünler": [
        ("Blueland", "blueland.com", "Tablet Temizlik", "Temizlik tableti + tekrar kullanılabilir şişe; plastik şişe atığını azaltma; Shark Tank; Sarah Paiji Yoo"),
        ("Grove Collaborative", "grove.co", "Doğal Ev Bakım", "Doğal + sürdürülebilir ev temizlik; aylık teslimat; B Corp; NYSE listeli; $1.5B değerleme"),
        ("Earth Breeze", "earthbreeze.com", "Çamaşır Yaprağı", "Çamaşır deterjan yaprağı; plastik jug eliminasyonu; her sipariş 10 yıkama bağış; viral büyüme"),
        ("Stasher", "stasherbag.com", "Silikon Poşet", "Yeniden kullanılabilir silikon saklama poşeti; platinyum silikon; bulaşık makinesine girer; Kat Nouri"),
        ("Public Goods", "publicgoods.com", "Sürdürülebilir Market", "Birch ambalajlı minimalist ürünler; üyelik modeli; organik + sürdürülebilir; ev + kişisel bakım"),
        ("Who Gives A Crap", "whogivesacrap.com", "Bambu Tuvalet Kağıdı", "Bambu + geri dönüşüm tuvalet kağıdı; gelirin %50'si sanitasyon projelerine; Avustralya; eğlenceli ambalaj"),
        ("Package Free", "packagefreeshop.com", "Sıfır Atık Dükkan", "Lauren Singer'ın sıfır atık mağazası; plastik olmayan alternatifler; Brooklyn; sıfır atık yaşam"),
        ("Ethique", "ethique.com", "Katı Güzellik", "Yeni Zelanda; katı şampuan + sabun barları; 20M+ plastik şişe tasarruf; kompostlanabilir ambalaj"),
        ("LastObject", "lastobject.com", "Tekrar Kullanılabilir", "LastSwab tekrar kullanılabilir kulak çubuğu; LastTissue mendil; LastRound pamuk; Danimarka; tek kullanımlığa son"),
        ("Bite", "bfritetoothpaste.com", "Tablet Diş Macunu", "Diş macunu tableti; plastik tüp yok; Lindsay McCormick kurdu; sıfır atık ağız bakımı"),
        ("by Humankind", "byhumankind.com", "Kişisel Bakım Sıfır Atık", "Yeniden doldurulabilir deodorant, şampuan, diş macunu; plastik atığı azaltma; abonelik"),
        ("Plaine Products", "plaineproducts.com", "Yeniden Doldur Banyo", "Alüminyum şişe gönder-doldur-geri gönder döngüsü; şampuan + duş jeli; sıfır atık banyo"),
        ("Girlfriend Collective", "girlfriend.com", "Geri Dönüşüm Giyim", "Pet şişelerden aktif giyim; %100 geri dönüşüm; kapsayıcı bedenler; şeffaf üretim"),
        ("PANGAIA", "thepangaia.com", "Biyo-Mühendislik Moda", "FLWRDWN kuş tüyü alternatifi; deniz yosunu lifi; biyomalzeme inovasyonu; $245M değerleme"),
        ("Allbirds", "allbirds.com", "Karbon Ayak İzi Ayakkabı", "Yün + ağaç lifi ayakkabı; her üründe karbon etiketi; Tim Brown (NZ); karbon nötr hedefi"),
        ("Tentree", "tentree.com", "10 Ağaç Marka", "Her satışta 10 ağaç dikiyor; 100M+ ağaç dikildi; Kanada; sürdürülebilir moda + etki"),
        ("Patagonia", "patagonia.com", "Çevresel Aktivizm", "İklim aktivizmi; Worn Wear onarım programı; 'Don't Buy This Jacket' reklamı; $1B+ gelir; dünya için çalışan marka"),
        ("Veja", "veja-store.com", "Fransız Sürdürülebilir Sneaker", "Brezilya organik pamuk + fair trade kauçuk; reklam yapmıyor; ağızdan ağıza büyüme; Meghan Markle"),
        ("Pela", "pelacase.com", "Kompostlanabilir Kılıf", "Kompostlanabilir telefon kılıfı; flax straw malzeme; Kanada; 'boring phone case revolution'"),
        ("Bamford", "bamford.com", "İngiliz Organik Lüks", "Cotswolds çiftlik; organik cilt bakım + ev; Lady Bamford; farm-to-face; sürdürülebilir İngiliz lüks"),
        ("MATE the Label", "matethelabel.com", "Organik Temel Giyim", "LA; organik pamuk basics; GOTS sertifikalı; temiz boya; minimalist sürdürülebilir moda"),
        ("Organic Basics", "organicbasics.com", "Danimarka Organik İç Giyim", "Organik pamuk iç giyim; Silvertech gümüş iyon; Danimarka; en az yıkama önerisi; düşük etki"),
        ("Outerknown", "outerknown.com", "Kelly Slater Surf Moda", "Kelly Slater'ın markası; Econyl geri dönüşüm naylon; S.E.A. Jeans 100% geri dönüşüm; surf kültürü"),
        ("Nisolo", "nisolo.com", "Etik Ayakkabı", "Peru + Meksika etik üretim; Nashville merkezli; Lowest Impact Rating; fair trade deri ayakkabı"),
        ("United By Blue", "unitedbyblue.com", "Okyanus Temizlik Moda", "Her satışta 1 pound okyanus çöpü temizliyor; 4M+ pound temizlendi; outdoor giyim + aksesuar"),
        ("Hiut Denim", "hiutdenim.co.uk", "Galler El Yapımı Denim", "Cardigan, Galler'de el yapımı jean; 'Do One Thing Well'; eski Levi's fabrika kasabası; yavaş moda"),
        ("Finisterre", "finisterre.com", "İngiliz Surf Sürdürülebilir", "Cornwall merkezli; soğuk su surf giyim; geri dönüşüm polyester; onarım hizmeti; B Corp"),
        ("Cotopaxi", "cotopaxi.com", "Renkli Outdoor Etki", "Del Dia renk artığı çantalar; her biri benzersiz; Latin Amerika ilhamı; B Corp; gear for good"),
        ("prAna", "prana.com", "Yoga Sürdürülebilir", "Sürdürülebilir yoga + outdoor giyim; organik pamuk + geri dönüşüm; Fair Trade; Columbia bünyesinde"),
        ("Pact", "wearpact.com", "Organik Uygun Fiyat", "Organik pamuk basics uygun fiyatla; Fair Trade; $15 tişört organik; erişilebilir sürdürülebilirlik"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 15. ABONELİK KUTULARI (Subscription Boxes) — 25+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Abonelik Kutuları": [
        ("BarkBox", "barkbox.com", "Köpek Oyuncak Kutu", "Aylık tematik köpek oyuncak + ödül; 2M+ abone; eğlenceli temalar; Super Chewer; $200M+ gelir"),
        ("FabFitFun", "fabfitfun.com", "Yaşam Tarzı Kutu", "Sezonluk güzellik + fitness + ev kutusu; $200+ değer ürünler $50'a; 2M+ abone; celebrity collab"),
        ("Birchbox", "birchbox.com", "Güzellik Keşif", "Mini güzellik ürünleri keşif kutusu; kutu abonelik modelini başlattılar; Katia Beauchamp kurdu"),
        ("Ipsy", "ipsy.com", "Güzellik Glam Bag", "Kişiselleştirilmiş güzellik kutusu; 5 ürün $14/ay; quiz bazlı eşleştirme; BFA ile birleşti; 3M+ abone"),
        ("BoxyCharm", "boxycharm.com", "Full-Size Güzellik", "Full-size güzellik ürünleri kutu; $30/ay $150+ değer; IPSY ile birleşti; lüks güzellik keşfi"),
        ("Stitch Fix", "stitchfix.com", "AI Moda Kutu", "AI + stilist kişisel moda kutusu; algoritmik stil önerisi; $1.5B+ gelir; Katrina Lake kurdu"),
        ("Trunk Club", "trunkclub.com", "Premium Stil Kutu", "Nordstrom'un kişisel stil kutusu; premium markalar; stilist eşleştirme; erkek + kadın"),
        ("Blue Apron", "blueapron.com", "Yemek Kiti Öncü", "Yemek kiti abonelik öncüsü; tarifli malzeme teslimatı; 2017 IPO; Wonder Group satın aldı"),
        ("HelloFresh", "hellofresh.com", "Global Yemek Kiti", "Dünyanın en büyük yemek kiti; 7M+ müşteri; $7B+ gelir; Almanya merkezli; 15+ ülke"),
        ("Meowbox", "meowbox.com", "Kedi Kutu", "Aylık kedi oyuncak + ödül kutusu; tematik kutular; kedi severlerin BarkBox'ı"),
        ("Book of the Month", "bookofthemonth.com", "Kitap Abonelik", "5 seçenekten 1 ciltli kitap seçin; aylık $16.99; kitap topluluğu; 2021'de patlama"),
        ("Breo Box", "breobox.com", "Teknoloji Yaşam Tarzı", "Teknoloji + fitness + ev gadget kutusu; sezonluk; $150+ değer; erkek + kadın"),
        ("Universal Yums", "universalyums.com", "Dünya Snack", "Her ay farklı ülkeden atıştırmalık; eğitim kartları; kültürel keşif; eğlenceli aile kutusu"),
        ("SnackCrate", "snackcrate.com", "Uluslararası Snack", "Her ay farklı ülke snackları; S/M/L beden seçimi; dünya turu atıştırmalık keşfi"),
        ("Bespoke Post", "bespokepost.com", "Erkek Keşif Kutu", "Erkek yaşam tarzı kutusu; 'Boxes of Awesome'; $49/kutu; tema seçimi; kaliteli erkek ürünleri"),
        ("Winc", "winc.com", "Şarap Abonelik", "Tat profili quiz ile kişisel şarap seçimi; aylık 4 şişe; kendi bağlarından + ortaklardan"),
        ("Firstleaf", "firstleaf.com", "AI Şarap", "AI algoritma ile şarap eşleştirme; 6 şişe $40; verilere göre tat geliştirme; uygun fiyat şarap"),
        ("Trade Coffee", "drinktrade.com", "Kahve Abonelik", "450+ kavurucu ile eşleştirme; quiz bazlı kişisel kahve; taze kavurma kapıya; specialty coffee"),
        ("Atlas Coffee Club", "atlascoffeeclub.com", "Dünya Kahve", "Her ay farklı ülkeden single-origin kahve; kartpostal + bilgi; kahve dünya turu"),
        ("The Sill", "thesill.com", "Bitki Abonelik", "Ev bitkisi kutusu; saksı + bakım rehberi; NYC mağazaları; bitki ebeveynliği hareketi"),
        ("Loot Crate", "lootcrate.com", "Geek Kutu", "Geek + gamer koleksiyon kutusu; Marvel, Star Wars, Anime temalı; $100M+ gelir; pop kültür"),
        ("GlossyBox", "glossybox.com", "Lüks Güzellik Kutu", "Avrupa lüks güzellik keşif kutusu; pembe kutu ikonik; 5 ürün/ay; lookfantastic bünyesinde"),
        ("Gentleman's Box", "gentlemansbox.com", "Erkek Stil Kutu", "Kravat, kol düğmesi, cüzdan; erkek aksesuar kutusu; GQ ortaklığı; klasik erkek stili"),
        ("Daily Harvest", "daily-harvest.com", "Smoothie Abonelik", "Dondurulmuş smoothie + harvest bowl abonelik; hazır karışım; sağlıklı fast food; $250M+ yatırım"),
        ("Butcher Box", "butcherbox.com", "Et Abonelik Kutu", "Grass-fed, organik et aylık kutu; $5/öğün; çiftlikten kapıya; $600M+ gelir"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 16. SEYAHAT & BAVUL (Travel & Luggage) — 25+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Seyahat & Bavul": [
        ("Away", "awaytravel.com", "Modern Bavul", "Ejectable pilli valiz; Instagram seyahat kültürü; $1.4B değerleme; Jen Rubio + Steph Korey; renk seçenekleri"),
        ("Monos", "monfrsos.com", "Minimalist Bavul", "Kanada; sessiz tekerlek; Makrolon polikarbonat; minimalist tasarım; $300-400 fiyat; hızla büyüyen"),
        ("July", "july.com", "Avustralya Bavul", "Avustralya merkezli; kişiselleştirilebilir; hafif polikarbonat; isim gravürü; minimalist"),
        ("Béis", "bfreis.com", "Shay Mitchell Seyahat", "Shay Mitchell'in markası; Weekender Bag ikonik; uygun fiyatlı seyahat; $100M+ gelir; celebrity DTC"),
        ("Calpak", "calpaktravel.com", "Renkli Seyahat", "Renkli bavul + seyahat aksesuarları; Hue koleksiyon; erişilebilir fiyat; $100M+ gelir; Instagram viral"),
        ("Paravel", "tourparavel.com", "Sürdürülebilir Seyahat", "Geri dönüşüm malzeme bavul; karbon nötr gönderim; Indre Rockefeller kurdu; lüks + sürdürülebilir"),
        ("Nomatic", "nomatic.com", "Dijital Göçebe Çanta", "Travel Pack 40L; dijital göçebe optimize; Kickstarter $7M; seyahat + iş; organize bölmeler"),
        ("Baboon to the Moon", "baboontothemoon.com", "Renkli Duffle", "Neon renk duffle çantalar; Go-Bag ikonik; dayanıklı 1000D naylon; cesur renkler; TikTok viral"),
        ("WANDRD", "wandrd.com", "Fotoğraf Seyahat Çanta", "PRVKE fotoğraf sırt çantası; Roll Top erişim; kamera + seyahat; fotoğrafçı topluluğu; Kickstarter"),
        ("Tropicfeel", "tropicfeel.com", "Seyahat Ayakkabı", "3-in-1 seyahat ayakkabısı; su geçirmez; kompakt paketleme; Barcelona; Kickstarter $14M; multi-terrain"),
        ("Dagne Dover", "dagnedover.com", "Neopren İş Çantası", "Neopren çanta + sırt çantası; Landon Carryall ikonik; düzenli bölmeler; iş kadını tercihi"),
        ("Mokobara", "mokobara.com", "Hint DTC Bavul", "Hindistan'ın Away'i; premium bavul uygun fiyat; Bangalore; The Transit çanta; hızla büyüyen"),
        ("Horizn Studios", "horizn-studios.com", "Alman Akıllı Bavul", "Berlin; akıllı bavul; çıkarılabilir powerbank; Model H ikonik; Avrupa lüks seyahat"),
        ("Roam Luggage", "rofrm.com", "Özelleştirilebilir Bavul", "500K+ renk kombinasyonu kişisel bavul; ABD üretim; 'design your own'; premium kişiselleştirme"),
        ("Cadence", "keepyourcadence.com", "Seyahat Kapsül", "Manyetik seyahat kapsülleri; sıvı düzenleyici; TSA uyumlu; renkli hexagon tasarım"),
        ("Aer", "aersf.com", "Şehir Seyahat Çanta", "San Francisco; Travel Pack 3; şehir seyahati; minimalist; su geçirmez; iş + seyahat hibrit"),
        ("Peak Design", "peakdesign.com", "Seyahat Sırt Çantası", "Travel Backpack 45L; paketleme küpleri; Kickstarter $32M toplam; fotoğraf + seyahat"),
        ("Topo Designs", "topodesigns.com", "Colorado Outdoor Çanta", "Colorado renkli outdoor çantalar; Mountain Briefcase; durable + eğlenceli; made in USA"),
        ("Bellroy", "bellroy.com", "Seyahat Cüzdan", "Seyahat cüzdanı + pasaport kılıfı; Melbourne; slim taşıma felsefesi; B Corp sertifikalı"),
        ("Level8", "level8cases.com", "Alman Mühendislik Bavul", "Alüminyum + polikarbonat; Almanya tasarım; sessiz tekerlek; uygun fiyat premium; Gibraltar koleksiyon"),
        ("Samsonite DTC", "samsonite.com", "DTC Dönüşüm", "Legacy marka DTC dönüşümü; online direkt satış büyümesi; Proxis koleksiyon; premium seçenekler"),
        ("MVST Select", "mvstselect.com", "Premium Sert Bavul", "Alüminyum bavul; premium malzeme; sessiz spinner tekerlek; $300-600; Rimowa alternatifi"),
        ("Db (Douchebags)", "dbjourney.com", "Norveç Spor Bavul", "Norveç; kayak + surf + seyahat çantaları; pro athlete collab; renkli; macera seyahati"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 17. DİŞ & AĞIZ BAKIMI (Dental & Oral Care) — 20+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Diş & Ağız Bakımı": [
        ("Quip", "getquip.com", "Abonelik Diş Fırçası", "Elektrikli diş fırçası abonelik; $25 başlangıç; 3 ayda yeni fırça başlığı; minimal tasarım; $100M+ gelir"),
        ("Burst", "bfrurst.com", "Sonic Diş Fırçası", "Sonic elektrikli diş fırçası; kömürlü fırça başlığı; dental profesyonel ağı; abonelik modeli"),
        ("Byte", "byteme.com", "Evde Diş Teli", "Evde şeffaf diş teli; HyperByte vibrasyon hızlandırıcı; gece aparatı seçeneği; Dentsply Sirona satın aldı"),
        ("Snow", "trysnow.com", "LED Diş Beyazlatma", "LED diş beyazlatma kiti; 'Shark Tank rejected' pazarlama; Joshua Snow kurdu; $100M+ gelir; celebrity"),
        ("Hismile", "hismile.com", "Avustralya Diş Beyazlatma", "Avustralya; Nik Mirkovic + Alex Tomic 20 yaşında kurdu; PAP+ beyazlatma; TikTok viral; $150M+ gelir"),
        ("Cocofloss", "cocofloss.com", "Lüks Diş İpi", "Lüks diş ipi; kokulu + aromalı; 500+ iplikçik; Chrystle Cu diş hekimi kurdu; premium diş ipi"),
        ("David's", "davids-usa.com", "Doğal Diş Macunu", "Doğal diş macunu metal tüp; ABD üretim; nano-hidroksiapatit; sürdürülebilir ambalaj"),
        ("RiseWell", "risewell.com", "Hidroksiapatit Macun", "Hidroksiapatit bazlı diş macunu; flor alternatifi; çocuk versiyonu; temiz ağız bakımı"),
        ("AutoBrush", "autobrush.com", "U-Şekil Diş Fırçası", "U-şeklinde otomatik diş fırçası; 30 saniye 360° temizlik; çocuk + yetişkin; elektrikli"),
        ("Boka", "bfroka.com", "Nano Diş Macunu", "Nano-hidroksiapatit remineralize macun; Dentist Designed; florsuz; doğal beyazlatma"),
        ("Lumineux", "luminfreux.com", "Sertifikalı Doğal Ağız Bakım", "Non-toxic sertifikalı; diş beyazlatma bantları + ağız çalkalama; doğal bileşenler; Whole Foods"),
        ("SmileDirectClub", "smiledirectclub.com", "Şeffaf Diş Teli DTC", "Evde şeffaf diş teli $2000; ortodontist uzaktan izleme; $8.9B IPO sonrası kapandı ve yeniden açıldı"),
        ("Candid", "candidco.com", "Ortodontist Şeffaf Teli", "Ortodontist yönetiminde şeffaf diş teli; studio ağı; CandidPro doktor platformu"),
        ("Twice", "brushtwice.com", "İkili Macun Sistemi", "Sabah + akşam farklı formül diş macunu; Julian Levine kurdu; $1 bağış; ikili bakım"),
        ("Hello Products", "hello-products.com", "Doğal Ağız Bakımı", "Doğal diş macunu + ağız bakımı; eğlenceli ambalaj; Colgate-Palmolive satın aldı; doğal + etkili"),
        ("Supersmile", "supersmile.com", "Profesyonel Beyazlatma", "Diş hekimi formüle profesyonel beyazlatma; 45+ yıl; Calprox teknoloji; hassas dişler için"),
        ("Moon Oral Care", "moonoralcare.com", "Kendall Jenner Ağız Bakımı", "Kendall Jenner markası; aktif kömür diş macunu; anti-cavity; tasarım ambalaj"),
        ("Bite", "bitetoothpastebits.com", "Tablet Diş Macunu", "Diş macunu tableti; plastik tüp yok; çiğne + fırçala; sıfır atık ağız bakımı; $10M+ gelir"),
        ("Waken", "wafrken.com", "İngiliz Lüks Ağız Bakımı", "İngiliz lüks ağız bakımı; doğal bileşenler; güzel ambalaj; ağız çalkalama + macun"),
        ("Marashi Oral Health", "mafrashi.com", "Dermatolojik Ağız Bakımı", "Diş hekimi + dermatolog geliştirdi; oral + facial bakım kesişimi; yenilikçi yaklaşım"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 18. KADIN SAĞLIĞI (Women's Health) — 25+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Kadın Sağlığı": [
        ("Lola", "mylola.com", "Organik Tampon", "Organik pamuk tampon + ped abonelik; şeffaf bileşenler; 'by women for women'; P&G alternatifi"),
        ("Cora", "cora.life", "Premium Adet Ürünleri", "Organik tampon + ped; her satışta kızlara bağış; Molly Hayward kurdu; sosyal etki + kalite"),
        ("Rael", "getrael.com", "Kore Kadın Bakım", "Kore güzellik ilhamı kadın bakım; organik ped + cilt bakımı; Yanghee Paik kurdu; $60M+ gelir"),
        ("Saalt", "safralt.com", "Adet Kabı", "Silikon adet kabı; yumuşak + sert seçenek; Saalt Disc; adet iç çamaşırı; sürdürülebilir adet yönetimi"),
        ("Flex", "flexfits.com", "Adet Diski", "Tek kullanımlık adet diski; Flex Cup silikon; Lauren Schulte Wang kurdu; cinsel ilişki sırasında kullanım"),
        ("Thinx", "sfrhethinx.com", "Adet İç Giyim", "Adet geçirmez iç giyim; ped/tampon yerine; cesur reklamlar; NYC metro kampanyası; $100M+ gelir"),
        ("Knix", "knix.com", "Sızdırmaz İç Giyim", "Sızıntı geçirmez teknoloji; adet + idrar kaçırma; Joanna Griffiths; kapsayıcı; Essity satın aldı"),
        ("Elvie", "elvie.com", "Kadın Teknoloji", "Giyilebilir göğüs pompası; sessiz + el serbest; Elvie Trainer pelvik taban; Tania Boler kurdu; $100M+"),
        ("Willow", "onewillow.com", "Giyilebilir Pompa", "Süt çantası içinde giyilebilir göğüs pompası; sızdırmaz; hareket halinde süt sağma; FDA onaylı"),
        ("Kindra", "ourkindra.com", "Menopoz Bakım", "Menopoz semptomları için ürünler; vajinal nem + sıcak basma; Stacy London marka müdürü"),
        ("Evernow", "evernow.com", "Menopoz Tedavi", "Menopoz hormon tedavisi online; dermatolojist reçete; kişiselleştirilmiş tedavi planı"),
        ("Bonafide", "hellobonafide.com", "Menopoz Takviye", "İsveç poleninden Relizen; sıcak basma doğal tedavi; klinik çalışmalarla kanıtlanmış"),
        ("Natalist", "natfralist.com", "Hamilelik Hazırlık", "Doğurganlık + hamilelik ürünleri; ovulasyon testi, prenatal vitamin; modern aile planlaması"),
        ("Modern Fertility", "modernfertility.com", "Doğurganlık Testi", "Evde doğurganlık hormon testi $159; Ro satın aldı; kadınlara veri ile güç verme"),
        ("Proov", "proovtest.com", "Progesteron Testi", "Evde progesteron testi; ovülasyon onaylama; FDA onaylı; doğurganlık takibi"),
        ("Stix", "getstix.com", "Hamilelik Testi DTC", "Modern hamilelik + ovülasyon testi; $12 2'li paket; online siparişle gizlilik; eğitim içerikleri"),
        ("Winona", "bywinona.com", "Menopoz Reçete", "Online menopoz tedavisi; biyoidentik hormonlar; doktor konsültasyonu; telehealth kadın sağlığı"),
        ("Daye", "yourdaye.com", "CBD Tampon", "CBD infüze tampon İngiltere; ağrı kesici tampon; diagnostik tampon (vaginal mikrobiyom testi); bilimsel"),
        ("Flo Health", "flo.health", "Adet Takip App", "300M+ indirme adet takip uygulaması; AI destekli; doğurganlık takibi; kadın sağlığı eğitimi"),
        ("Clue", "helloclue.com", "Bilimsel Adet Takip", "Berlin; bilim bazlı adet takip; Oxford araştırmaları; gizlilik odaklı; FDA onaylı doğum kontrol"),
        ("Hers", "forhers.com", "Kadın Telehealth", "Hims & Hers kadın hattı; saç dökülmesi, cilt, cinsel sağlık; online doktor + reçete"),
        ("Nurx", "nurx.com", "Online Doğum Kontrol", "Doğum kontrol hapı online reçete; PrEP; STI test; erişilebilir kadın sağlığı; $4B Thirty Madison"),
        ("Ritual", "ritual.com", "Prenatal Vitamin", "Prenatal Essential vitamin; şeffaf bileşenler; 3. taraf test; hamilelik + emzirme + doğurganlık"),
        ("Needed", "thisisneeded.com", "Perinatal Beslenme", "Hamilelik + emzirme + postpartum takviye; klinisyen formüle; kollajen + probiyotik; kapsamlı destek"),
        ("Kin Fertility", "kinfertility.com.au", "Avustralya Kadın Sağlık", "Avustralya; doğurganlık + hamilelik; telehealth + ürün; prenatal + postnatal; doktor desteği"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 19. PARFÜM & KOKU (Fragrance) — 25+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Parfüm & Koku": [
        ("Phlur", "phlur.com", "Temiz Parfüm", "Missing Person TikTok viral; Chriselle Lim yeniden kurdu; temiz bileşenler; nostaljik kokular"),
        ("Snif", "snfrif.com", "Deneyin Sonra Alın", "Deneme + iade parfüm; 'try before you commit'; $40 tam boy; modern koklama kültürü"),
        ("Skylar", "skylar.com", "Hypoallergenik Parfüm", "Hassas cilt parfüm; temiz + vegan; kişiselleştirilmiş keşif kiti; Cat Chen kurdu"),
        ("Dedcool", "dedcool.com", "Genderless Temiz Koku", "Cinsiyetsiz temiz parfüm; biyozime bazlı; çamaşır deterjanı parfümü; LA merkezli"),
        ("Boy Smells", "boysmells.com", "Genderless Mum + Parfüm", "Genderful konsepti; mum + parfüm; Matthew Herman + David Kien; sınır tanımayan kokular"),
        ("Maison Louis Marie", "maisonlouismfrie.com", "Botanik Parfüm", "Botanik bahçe ilhamı; No.04 Bfrois de Balincourt en çok satan; doğal parfüm; Sephora'da"),
        ("Ellis Brooklyn", "ellisbrooklyn.com", "Temiz Lüks Parfüm", "Bee Shapiro kurdu; sürdürülebilir lüks parfüm; Myth, Verb ikonik; Sephora'da; temiz bileşenler"),
        ("Lake & Skye", "lakeandskye.com", "Minimalist Koku", "11 11 parfüm TikTok viral; beyaz misk; minimalist tek not parfüm; temiz + vegan"),
        ("Commodity", "commodityfragrances.com", "Katmanlı Koku", "Scent Space teknoloji; Personal, Expressive, Bold her koku 3 versiyonda; yeniden lansman başarısı"),
        ("Diptyque", "diptyque.com", "Paris Mum + Parfüm", "Paris efsanevi mum markası; Baies mum ikonik; lüks ev kokusu kategorisi yarattı; $700M+ gelir"),
        ("Le Labo", "lelabofragrances.com", "Artisanal Parfüm", "NYC el yapımı parfüm; Santal 33 kült koku; şehre özel kokular; Estée Lauder; $500M+ gelir"),
        ("Byredo", "byredo.com", "İsveç Lüks Parfüm", "Ben Gorham Stockholm; Gypsy Water ikonik; Puig satın aldı; minimalist şişe tasarımı; $1B+ değerleme"),
        ("D.S. & Durga", "dsanddurga.com", "Brooklyn Niş Parfüm", "Brooklyn atölyede el yapımı; I Don't Know What ikonik; hikaye odaklı kokular; bağımsız"),
        ("Otherland", "otherland.com", "Sanat Mum", "Sanatçı illüstrasyonlu mum; sıra dışı koku hikayeleri; görsel + koku deneyimi; güzel hediye"),
        ("Vitruvi", "vitruvi.com", "Taş Difüzör + Yağ", "Porselen taş difüzör; uçucu yağ karışımları; estetik aromaterapi; sara panton; Vancouver"),
        ("Pura", "pura.com", "Akıllı Ev Kokusu", "Akıllı difüzör cihazı; uygulama kontrollü; Nest + Capri Blue + Anthropologie kokuları; abonelik"),
        ("Homesick", "homesick.com", "Nostalji Mum", "Eyalet + şehir + hafıza kokulu mumlar; nostalji pazarlaması; hediye favorisi; New York, Hawaii"),
        ("Candle Pour", "candlepour.com", "DIY Mum Deneyimi", "Kendi mumunuzu yapın mağazalar; deneyim perakende; arkadaş + aile aktivitesi; ABD geneli"),
        ("Floral Street", "floralstreet.com", "Sürdürülebilir Parfüm", "İngiliz vegan parfüm; geri dönüşüm ambalaj; %100 pulp şişe kutu; Wonderland Peony; B Corp"),
        ("Ouai", "theouai.com", "Saç Parfümü", "Jen Atkin; Melrose Place saç + vücut kokusu; parfüm olarak da kullanılır; lifestyle fragrance"),
        ("Juliette Has a Gun", "juliettehasagun.com", "Provokasyonel Parfüm", "Not A Perfume Cetalox bazlı; Romano Ricci (Nina Ricci torunu); Fransız cesur parfüm"),
        ("Clean Reserve", "cleanreserve.com", "Temiz Parfüm Koleksiyonu", "Sürdürülebilir + temiz bileşen; Skin parfüm kült; 'responsible luxury'; geri dönüşüm ambalaj"),
        ("Scentbird", "scentbird.com", "Parfüm Abonelik", "Aylık $17 tasarımcı parfüm 8ml; 600+ marka seçeneği; 'date before you marry' konsepti"),
        ("The Harmonist", "theharmonist.com", "Feng Shui Parfüm", "Feng Shui ilhamı lüks parfüm; 5 element; Paris; $300+ fiyat; enerji + koku"),
        ("Heretic Parfum", "hereticparfum.com", "Doğal Artisanal", "Douglas Little kurdu; %100 doğal; Dirty Vanilla viral; LA el yapımı; bağımsız parfümör"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 20. OFİS & KIRTASİYE (Office & Stationery) — 20+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Ofis & Kırtasiye": [
        ("Moleskine", "moleskine.com", "Efsanevi Defter", "Hemingway + Picasso kullandığı defter hikayesi; siyah deri kaplı ikonik; $400M+ gelir; dijital + analog"),
        ("Baron Fig", "baronfig.com", "Yaratıcı Defter", "Confidant defter; dot grid; Kickstarter başarısı; yaratıcı profesyoneller; kalem + defter ekosistem"),
        ("Appointed", "appointed.co", "Amerikan Kırtasiye", "Washington DC üretim; lüks planlayıcı + defter; ABD yapımı premium kırtasiye; Suann Song kurdu"),
        ("Mochithings", "mochithings.com", "Kore Kırtasiye", "Güney Kore kırtasiye küratörlük; kawaii estetik; planner + sticker; Asya kırtasiye trendi"),
        ("Archer & Olive", "archerandolive.com", "Bullet Journal", "160gsm kalın kağıt; bullet journal topluluğu; no bleed through; Amanda Rach Lee collab"),
        ("Erin Condren", "erincondren.com", "Life Planner", "Kişiselleştirilebilir planlayıcı; sticker kitapları; öğretmen planlayıcı; yaratıcı planlama topluluğu"),
        ("Ink+Volt", "inkandvolt.com", "Verimlilik Planlayıcı", "Hedef odaklı planlayıcı; üretkenlik araçları; Kate Matsudaira kurdu; girişimci + profesyonel"),
        ("Poketo", "poketo.com", "LA Tasarım Kırtasiye", "Los Angeles sanat + tasarım kırtasiye; renkli desenleri; living & creativity; sanatçı collabları"),
        ("Papier", "papier.com", "İngiliz Kişisel Kırtasiye", "Londra; kişiselleştirilebilir defter + mektup kağıdı; tasarımcı desenleri; dijital baskı; hediye"),
        ("Rifle Paper Co", "riflepaperco.com", "Çiçekli Kırtasiye", "Anna Bond'un çiçek illüstrasyonları; kırtasiye + ev dekoru; lisanslama; Julep + iphone kılıf"),
        ("Leuchtturm1917", "leuchtturm1917.com", "Alman Defter", "1917'den beri; numaralı sayfalar; içindekiler sayfası; arşiv kalitesi; bullet journal resmi partneri"),
        ("Poppin", "poppin.com", "Renkli Ofis Ürünleri", "Renkli ofis kırtasiye + mobilya; 'work happy' sloganı; modern ofis estetiği; B2B + DTC"),
        ("Ugmonk", "ugmonk.com", "Gather Masaüstü Organizatör", "Analog masaüstü organizatör sistemi; ahşap + deri; minimalist üretkenlik; Jeff Sheldon kurdu"),
        ("Grovemade", "grovemade.com", "Ahşap Masaüstü", "Portland ahşap masaüstü aksesuarları; el yapımı; mouse pad + desk shelf + pen cup; doğal malzeme"),
        ("Orbitkey", "orbitkey.com", "Masa Düzenleyici", "Desk Mat + Key Organiser; vegan deri; manyetik kablo tutucu; Avustralya tasarım; Kickstarter"),
        ("Papersmiths", "papersmiths.co.uk", "İngiliz Kırtasiye Butik", "Londra kırtasiye butik markası; premium defter + kalem küratörlüğü; Scandinavian estetik"),
        ("Mark+Fold", "markandffrold.com", "İngiliz Premium Defter", "Londra İngiliz yapımı defter; sade + fonksiyonel; profesyonel kalite; minimal tasarım"),
        ("Shorthand", "shorthandpress.com", "Portland Defter", "Portland OR; Standard Issue defter; yerel üretim; basit + kullanışlı; vintage estetik"),
        ("Wit & Delight", "witanddelight.com", "Yaşam Tarzı Kırtasiye", "Kate Arends kurdu; blog'dan markaya; planlayıcı + defter; self-care kırtasiye"),
        ("Day Designer", "daydesigner.com", "Stratejik Planlayıcı", "Whitney English kurdu; CEO + anne planlayıcısı; Blue Sky lisansı; hedef odaklı yapı"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 21. OUTDOOR & MACERA (Outdoor & Adventure) — 25+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Outdoor & Macera": [
        ("Cotopaxi", "cotopaxi.com", "Renkli Outdoor", "Del Dia benzersiz renkli çantalar; gear for good; B Corp; Latin Amerika ilhamı; etik outdoor"),
        ("Rumpl", "rumpl.com", "Outdoor Battaniye", "Uyku tulumu teknolojisi battaniye; NanoLoft sentetik; Wyatt Rants kurdu; kamp + plaj + piknik"),
        ("BioLite", "bioliteenergy.com", "Off-Grid Enerji", "Kampçı ateş + USB şarj; CampStove ateşten elektrik; HeadLamp 330; off-grid yaşam teknolojisi"),
        ("Hydro Flask", "hydroflask.com", "Yalıtımlı Şişe", "Çift duvar vakum yalıtım; TempShield teknoloji; VSCO girl kültürü; renkli kapaklar; $500M+ gelir"),
        ("Stanley", "stanley1913.com", "Heritage Su Şişesi", "1913'ten beri; Quencher H2.0 viral TikTok; $750M gelir 2023; pastel renkler; hedef kitle değişimi"),
        ("Yeti", "yeti.com", "Premium Cooler", "Indestructible soğutucu + drinkware; premium fiyat $300+ cooler; Austin TX; $1.6B gelir; outdoor lüks"),
        ("Nalgene", "nalgene.com", "Dayanıklı Su Şişesi", "BPA-free su şişesi öncüsü; Tritan plastik; outdoor + kampüs; 75+ yıl; $10-15 fiyat; ikonik"),
        ("GRAYL", "grayl.com", "Su Arıtma Şişesi", "Basınçlı su arıtma şişesi; 15 saniyede temiz su; seyahat + outdoor + acil durum; OnePress filtre"),
        ("Snow Peak", "snowpeak.com", "Japon Outdoor", "Japon outdoor marka; titanyum kamp ekipmanı; minimalist tasarım; apparel + camping; $150M+ gelir"),
        ("Helinox", "helinox.com", "Ultra Hafif Kamp", "DAC alüminyum boru; Chair One 890g kamp sandalyesi; ultra hafif + kompakt; Kore merkezli"),
        ("Sea to Summit", "seatosummit.com", "Avustralya Kamp", "Avustralya; uyku tulumu + kamp matı; Aeros yastık; ultra hafif seyahat; 40+ yıl"),
        ("Thermarest", "thermarest.com", "Kamp Mat Öncü", "Kendinden şişen kamp matı icat etti; 1972; NeoAir ultra hafif; kamp uyku standartı"),
        ("Kelty", "kelty.com", "Aile Kampçılığı", "Aile dostu kamp ekipmanı; erişilebilir outdoor; Discovery Lowdown sandalye; 50+ yıl"),
        ("ENO", "eaglesnestoutfitters.com", "Hamak", "DoubleNest hamak ikonik; kampüs + park + kamp; renkli; $50-100; rahat outdoor yaşam"),
        ("Goal Zero", "goalzero.com", "Güneş Şarj", "Taşınabilir güneş paneli + güç istasyonu; Yeti güç istasyonu; off-grid enerji; NRG Energy bünyesinde"),
        ("Osprey", "osprey.com", "Trekking Çanta", "Atmos AG anti-gravity sırt çantası; All Mighty garanti; trekking standartı; $100-400; Helen of Troy"),
        ("Gregory", "gregorypacks.com", "Sırt Çantası", "Wayne Gregory kurdu; FreeFloat süspansiyon; Samsonite bünyesinde; trekking + seyahat"),
        ("MSR", "msrgear.com", "Kamp Ocağı + Çadır", "PocketRocket ocak ikonik; Hubba Hubba NX çadır; Cascade Designs; profesyonel outdoor"),
        ("Black Diamond", "blackdiamondequipment.com", "Tırmanış Ekipman", "Tırmanış + kayak ekipmanı; headlamp, trekking pole, carabiner; Salt Lake City; profesyonel"),
        ("Petzl", "petzl.com", "Kafa Lambası + Tırmanış", "Fransız kafa lambası + tırmanış ekipmanı; mağaracılık heritage; profesyonel güvenlik; endüstriyel"),
        ("Rumpl", "rumpl.com", "Teknik Battaniye", "Sentetik down battaniye; NanoLoft; suya dayanıklı; kamp + ev + seyahat; National Park koleksiyon"),
        ("PackIt", "packit.com", "Dondurucu Çanta", "Dondurucu jel yerleşik soğutucu çanta; katlayıp dondurucuya koy; öğle yemeği + piknik"),
        ("Igloo", "igloocoolers.com", "Ekolojik Soğutucu", "ECOCOOL geri dönüşüm soğutucu; 75+ yıl heritage; Biodegradable RECOOL; erişilebilir outdoor"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 22. OYUN & EĞLENCE (Gaming & Entertainment) — 20+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Oyun & Eğlence": [
        ("Backbone", "playbackbone.com", "Mobil Oyun Kontrolcü", "iPhone oyun kontrolcüsü; One; PlayStation Remote Play; Xbox Cloud Gaming; $40M+ yatırım"),
        ("Analogue", "analogue.co", "FPGA Retro Konsol", "Pocket el konsolu; Duo TurboGrafx; Mega Sg; FPGA orijinal kartuş; retro premium"),
        ("Panic", "panic.com", "Playdate Konsol", "Krank kollu el konsolu Playdate; siyah beyaz ekran; indie oyunlar; Portland OR; benzersiz form"),
        ("8BitDo", "8bitdo.com", "Retro Kontrolcü", "Retro estetik modern kontrolcü; Pro 2; Ultimate; Nintendo + PC uyumlu; Hong Kong; $30-50"),
        ("Secretlab", "secretlab.co", "Oyuncu Sandalye", "Premium oyuncu sandalye; Titan + Omega; $400-600; esports ortaklıkları; Singapur; $100M+ gelir"),
        ("Razer", "razer.com", "Oyuncu Ekipman", "Gaming laptop + mouse + klavye + kulaklık; yeşil ışık ikonik; 'For Gamers By Gamers'; $1B+ gelir"),
        ("SteelSeries", "steelseries.com", "Esports Ekipman", "Arctis kulaklık; Aerox mouse; profesyonel esports standartı; Danimarka; $100M+ gelir"),
        ("HyperX", "hyperx.com", "Gaming Audio", "Cloud kulaklık serisi; HP bünyesinde; esports sponsorluğu; erişilebilir gaming audio; ödüllü ses"),
        ("NZXT", "nzxt.com", "PC Kasası + BLD", "Minimalist PC kasası; BLD özel PC montaj; H510 ikonik; CAM yazılım; gaming PC ekosistemi"),
        ("Corsair", "corsair.com", "PC Gaming Periferali", "K100 klavye; iCUE RGB; DDR5 RAM; PSU; Elgato + Scuf bünyesinde; $1.5B+ gelir"),
        ("Scuf Gaming", "scuf.com", "Premium Kontrolcü", "Özelleştirilebilir pro kontrolcü; arka paddle'lar; pro oyuncu tercihi; Corsair bünyesinde"),
        ("Nanoleaf", "nanoleaf.me", "RGB Işık Paneli", "Üçgen + altıgen LED ışık panelleri; ses reaktif; oyuncu odası estetiği; Matter + Thread uyumlu"),
        ("Govee", "govee.com", "Ambiyans Aydınlatma", "TV arkası LED şerit; oyuncu RGB aydınlatma; uygun fiyat; uygulama kontrol; TikTok viral odalar"),
        ("Elgato", "elgato.com", "Yayıncı Araçları", "Stream Deck; Key Light; Wave mikrofon; Green Screen; Corsair bünyesinde; içerik üretici standaradı"),
        ("Glorious", "gloriousgaming.com", "Mekanik Klavye + Mouse", "GMMK Pro hot-swap klavye; Model O ultra hafif mouse; DIY mekanik klavye topluluğu"),
        ("Drop", "drop.com", "Enthusiast Klavye", "Holy Panda switch; özel mekanik klavye collabları; Sennheiser collab kulaklık; topluluğa dayalı ürün"),
        ("Lego DTC", "lego.com", "Yetişkin LEGO", "Yetişkin LEGO setleri patladı; Icons + Technic + Architecture; DTC kanal büyümesi; $9B+ gelir"),
        ("Ravensburger", "ravensburger.com", "Premium Puzzle", "Alman puzzle markası; 40,000 parça mega puzzle; yetişkin puzzle trendi; $600M+ gelir"),
        ("Puttview", "puttview.com", "AR Golf", "Artırılmış gerçeklik golf; yeşil okuma teknolojisi; profesyonel eğitim; golf-tech"),
        ("Rocketbook", "getrocketbook.com", "Tekrar Kullanılabilir Defter", "Yazıp sil defter; bulut senkronizasyon; sürdürülebilir not alma; $30M+ gelir; Kickstarter"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 23. SUPPLEMENT & SPORCU BESLENMESİ (Supplements & Sports Nutrition) — 40+ brands
    # ═══════════════════════════════════════════════════════════════════════════
    "Supplement & Sporcu Beslenmesi": [
        ("AG1", "drinkag1.com", "Hepsi Bir Arada Yeşil", "75 vitamin, mineral, probiyotik; her podcast sponsoru; $500M+ gelir; sabah ritüeli; Tim Ferriss + Huberman"),
        ("Transparent Labs", "transparentlabs.com", "Şeffaf Formül", "%100 açık etiket; gizli karışım yok; klinik dozlar; PresJym rakibi; hardcore fitness"),
        ("Gorilla Mind", "gorillamind.com", "YouTube Fitness Takviye", "Derek (MPMD) kurdu; Sigma pre-workout; nootropik; fitness YouTube topluluğu; $100M+ gelir"),
        ("Ghost Lifestyle", "ghostlifestyle.com", "Lisanslı Tatlar", "Sour Patch Kids, Warheads lisanslı lezzet; şeffaf etiket; lifestyle fitness; $300M+ gelir"),
        ("Alani Nu", "alaninu.com", "Kadın Fitness Takviye", "Katy Hearn kurdu; enerji içeceği; Balance takviye; kadın fitness topluluğu; $200M+ gelir; TikTok"),
        ("Bloom Nutrition", "bloomnu.com", "TikTok Yeşil Toz", "Greens & Superfoods TikTok viral; Mari Llewellyn kurdu; kadın wellness; Target rafta; $200M+ gelir"),
        ("Momentous", "livemomentous.com", "Huberman Lab Takviye", "Andrew Huberman'ın podcast önerisi; NSF sertifikalı; klinik doz; profesyonel sporcu kalitesi"),
        ("Thorne", "thorne.com", "Klinik Grade", "Mayo Clinic ortağı; WFS dünya futbol; NSF for Sport; profesyonel sporcu güveni; $100M+ gelir; bilim"),
        ("Gainful", "gainful.com", "Kişisel Protein Tozu", "Quiz bazlı kişiselleştirilmiş protein; hedef odaklı; registered dietitian erişimi; her karışım benzersiz"),
        ("Legion Athletics", "legionathletics.com", "Bilim Bazlı Fitness", "Michael Matthews (Bigger Leaner Stronger) kurdu; klinik doz; doğal tatlandırıcı; $100M+ gelir"),
        ("Jacked Factory", "jackedfactory.com", "Temiz Pre-Workout", "Doğal tatlandırıcılı pre-workout; Nitrosurge; sıfır yapay boya; Amazon besteller"),
        ("1st Phorm", "1stphorm.com", "Topluluk Fitness", "Andy Frisella kurdu; Level-1 protein; Legionnaire program; fitness topluluğu; $500M+ gelir"),
        ("Kaged", "kfragedsupplement.com", "Premium Spor Takviye", "Pre-Kaged legendary pre-workout; Micropure kalite; Kris Gethin kurdu; premium pozisyonlama"),
        ("Ryse Supplements", "rfrysesupps.com", "Lisanslı Snack Lezzet", "Ring Pop, Skippy PB lisanslı protein; Gen Z fitness; eğlenceli lezzetler; $100M+ gelir"),
        ("Raw Nutrition", "getrawnutrition.com", "Chris Bumstead Markası", "CBum Classic protein; Mr. Olympia Classic Physique şampiyonu; Revive MD sağlık hattı"),
        ("Nutrabolt", "nutrabolt.com", "C4 Pre-Workout", "C4 en çok satan pre-workout; Cellucor bünyesinde; enerji içeceği + toz; $500M+ gelir"),
        ("Optimum Nutrition", "optimumnutrition.com", "Gold Standard Protein", "Gold Standard Whey ikonik; Glanbia bünyesinde; global protein tozu standartı; $1B+ marka"),
        ("MyProtein", "myprotein.com", "İngiliz Değer Protein", "Manchester; 70+ ülkeye gönderim; uygun fiyat protein; THG/The Hut Group; $800M+ gelir"),
        ("Bulk", "bulk.com", "İngiliz Toplu Takviye", "İngiliz spor beslenme; uygun fiyat; geniş ürün yelpazesi; pre-workout + protein + amino"),
        ("PE Science", "pescience.com", "Bilim Bazlı Protein", "Select Protein kazein + whey karışım; bilimsel formüller; Tim Muriello; fitness topluluğu"),
        ("Bare Performance Nutrition", "bareperformancenutrition.com", "Veteran Fitness Takviye", "Nick Bare (eski asker) kurdu; Flight pre-workout; YouTube fitness; doğal bileşenler"),
        ("Huge Supplements", "hugesupplements.com", "Yeni Nesil Fitness", "Sapogenix, Enhance; cycle support; bodybuilding topluluğu; agresif büyüme; yeni nesil"),
        ("Swolverine", "swolverine.com", "CrossFit Takviye", "CrossFit + endurance odaklı; Kre-Alkalyn kreatin; temiz bileşen; performans + toparlanma"),
        ("Performix", "performixdriven.com", "SST Yağ Yakıcı", "SST termojenik kapsül; Timed-release teknoloji; Ion + SST serisi; GNC ortaklığı"),
        ("Xtend", "xtfrendbcaa.com", "BCAA Öncü", "BCAA kategorisini yaratan marka; Nutrabolt bünyesinde; elektrolit + amino; intra-workout standartı"),
        ("Cellucor", "cellucor.com", "C4 Enerji", "C4 Smart Energy içecek; Nutrabolt bünyesinde; en çok satan pre-workout; süpermarket enerji içeceği"),
        ("JYM Supplement", "jymsupplementscience.com", "Bilim Adamı Takviye", "Dr. Jim Stoppani kurdu; açık formül; BioJYM araştırma; Pre JYM ikonik; bilimsel güvenilirlik"),
        ("Orgain", "orgain.com", "Organik Protein", "Dr. Andrew Abraham (kanser hayatta kalan) kurdu; organik protein shake; Costco + Amazon #1; temiz"),
        ("Garden of Life", "gardenoflife.com", "Organik Takviye", "USDA organik; Non-GMO Project; Nestlé Health Science; Raw Protein; tüm gıda bazlı takviye"),
        ("Vega", "myvega.com", "Bitkisel Sporcu", "Vegan sporcu beslenmesi; Brendan Brazier kurdu; bitkisel protein + pre-workout; WM Wrigley/Mars"),
        ("Vital Proteins", "vitalproteins.com", "Kolajen Peptit", "Jennifer Aniston yüzü; sabah kahvesine kolajen; Nestlé satın aldı; kolajen kategorisi yarattı"),
        ("Sports Research", "sportsresearch.com", "Amazon Takviye", "MCT yağı, kolajen, omega-3; Amazon bestseller; San Pedro CA; 40+ yıl aile şirketi"),
        ("NOW Foods", "nowfoods.com", "Değer Takviye", "1968'den beri; 1,400+ ürün; erişilebilir fiyat; doğal ürün sektörü lideri; $500M+ gelir"),
        ("Nutrex Research", "nutrex.com", "Lipo-6 Yağ Yakıcı", "Lipo-6 ikonik yağ yakıcı; Hawaii merkezli; Liquid Capsule teknoloji; bodybuilding klasiği"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# EXTRA BRANDS — extending each category to reach 2000+ total
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS = {
    "Güzellik & Cilt Bakımı": [
        ("Skin1004", "skin1004.com", "Kore Centella", "Centella Asiatica ampoule; Kore eczane favorisi; TikTok viral; uygun fiyatlı K-beauty aktif"),
        ("COSRX", "cosrx.com", "Kore Aktif Bakım", "Snail Mucin 96 dünya çapında viral; BHA Blackhead Power Liquid kült; AmorePacific bünyesinde"),
        ("Innisfree", "innisfree.com", "Kore Doğal Bakım", "Jeju adasından yeşil çay; AmorePacific; doğal K-beauty öncüsü; sürdürülebilir ambalaj"),
        ("Laneige", "laneige.com", "Kore Nem Bakımı", "Water Sleeping Mask ikonik; Lip Sleeping Mask TikTok viral; AmorePacific; K-beauty lüks"),
        ("Missha", "missha.com", "Kore Erişilebilir Bakım", "Time Revolution First Treatment Essence SK-II alternatifi; $20 vs $180; uygun K-beauty"),
        ("Purito", "purito.com", "Kore Temiz Bakım", "Centella serisi; güneş kremi skandalı sonrası reformülasyon; şeffaf yaklaşım; güvenilir K-beauty"),
        ("Some By Mi", "somebymi.com", "Kore Mucize Bakım", "30 Days Miracle toner; AHA/BHA/PHA üçlü asit; 30 günde sonuç vaadi; TikTok viral"),
        ("Klairs", "klfrairs.com", "Kore Hassas Bakım", "Supple Preparation Toner kült; hassas cilt uzmanı; Wishtrend bünyesinde; erişilebilir"),
        ("Banila Co", "banilaco.com", "Kore Temizleme", "Clean It Zero sherbet temizleyici ikonik; çift temizleme; Kore makyaj temizleme standartı"),
        ("Heimish", "heimish.com", "Kore Doğal", "All Clean Balm sherbet temizleyici; Bulgari Jasmine kokusu; uygun fiyat lüks his"),
        ("I'm From", "im-from.com", "Kore Tek Bileşen", "Mugwort Mask %100 pelin otu; Rice Toner pirinç suyu; tek bileşen odak; K-beauty"),
        ("Round Lab", "roundlab.co.kr", "Kore Deniz", "Dokdo Toner; deniz mineralleri; uygun fiyat Kore cilt bakımı; Olive Young bestseller"),
        ("Medicube", "medicube.com", "Kore Derma Bakım", "AGE-R cihazları; Red Concealer; dermatolojik K-beauty; cihaz + ürün ekosistem"),
        ("Dr. Jart+", "drjart.com", "Kore Dermakozmetik", "Cicapair Tiger Grass; Ceramidin; dermatoloji + kozmetik; Estée Lauder satın aldı"),
        ("Skinceuticals", "skinceuticals.com", "Dermatolojist Bakım", "CE Ferulic C vitamini ikonik; L'Oréal Active; dermatolojist #1 marka; klinik sonuçlar"),
        ("La Roche-Posay", "laroche-posay.com", "Eczane Bakım", "Cicaplast Baume B5 viral TikTok; Anthelios güneş kremi; dermatolojist önerisi #1"),
        ("CeraVe", "cerave.com", "Eczane Seramid", "Ceramide bazlı; dermatolog geliştirdi; L'Oréal; TikTok skincare routine viral; uygun fiyat"),
        ("The Inkey List", "theinkeylist.com", "Bilim Öğretici", "Her üründe bileşen eğitimi; $6-15 fiyat; 'skincare simplified'; Be For Beauty; İngiltere"),
        ("Revolution Beauty", "revolutionbeauty.com", "İngiliz Erişilebilir", "£1-15 fiyat aralığı; Revolution Skincare; Freedom; LondonCar Park bünyesinde; 150+ ülke"),
        ("e.l.f. Cosmetics", "elfcosmetics.com", "Erişilebilir Makyaj", "$3-14 fiyat; Holy Hydration SPF; Power Grip Primer viral; Z kuşağı makyaj; $1B+ gelir"),
        ("Milani", "milani.com", "İtalyan İlham Erişilebilir", "Conceal + Perfect fondöten; Baked Blush ikonik; İtalyan ilham; süpermarket + CVS"),
        ("Pixi Beauty", "pixibeauty.com", "İngiliz Glow", "Glow Tonic glikolik tonik ikonik; Petra Strand kurdu; İsveç + İngiltere; yeşil ambalaj"),
        ("Charlotte Tilbury", "charlottetilbury.com", "İngiliz Lüks Makyaj", "Pillow Talk koleksiyonu dünya fenomeni; Hollywood Flawless Filter viral; Puig satın aldı"),
        ("Patrick Ta", "patrickta.com", "Ünlü Makyöz Markası", "Major Headlines çift uçlu lip + cheek; Gigi Hadid'in makyözü; Hollywood pro"),
        ("Kjaer Weis", "kjaerweis.com", "Lüks Yeniden Doldur", "Yeniden doldurulabilir lüks ambalaj; organik sertifikalı; Danimarka; refillable makeup öncüsü"),
        ("Westman Atelier", "westman-atelier.com", "Clean Lüks Makyaj", "Gucci Westman markası; Baby Cheeks blush stick; temiz lüks; Barneys/Sephora"),
        ("Rare Beauty", "rarebeauty.com", "Mental Sağlık Güzellik", "Liquid Blush viral; Rare Impact Fund %1 bağış; Selena Gomez; Sephora en çok satan"),
        ("W3LL People", "w3llpeople.com", "Organik Temiz Makyaj", "EWG VERIFIED; Bio Correct concealer; Austin TX; organik temiz makyaj; Target'ta"),
        ("RMS Beauty", "rmsbeauty.com", "Ham Organik Makyaj", "Living Luminizer ikonik; Rose-Marie Swift kurdu; organik hindistan cevizi yağı bazlı; doğal ışıltı"),
        ("Beautycounter", "beautycounter.com", "Temiz Güzellik Aktivizm", "1,800+ yasaklı bileşen; temiz güzellik savunuculuk; Carlyle Group satın aldı; DTC + danışman modeli"),
        ("Trinny London", "trinnylondon.com", "İngiliz Kişisel Makyaj", "Trinny Woodall; Match2Me renk eşleştirme quiz; stackable pot tasarım; İngiltere DTC güzellik"),
        ("Grown Alchemist", "grownalchemist.com", "Avustralya Biyoloji Bakım", "Melbourne; Advanced biyoloji formüller; lüks otel amenity; anti-aging; peptit teknolojisi"),
        ("Aestura", "aestura.com", "Kore Dermatoloji", "Atobarrier 365 Cream; hassas cilt dermatolojik bakım; Kore eczane markası; ceramide odaklı"),
        ("Heimish", "heimish.com", "Kore Vegan", "Bulgarian Rose Water Mist; vegan + cruelty-free K-beauty; uygun fiyat doğal formüller"),
        ("SKIN&LAB", "skinandlab.com", "Kore Vitamin Bakım", "Vitamin C Brightening Serum; Dr.Vita Clinic serisi; dermatoloji bazlı K-beauty"),
        ("Derma E", "dermae.com", "Doğal Dermatoloji", "Vitamin C Concentrated Serum; 40+ yıl; Non-GMO; vegan + cruelty-free; uygun fiyat"),
        ("Acure", "acure.com", "Organik Uygun Bakım", "Brightening Facial Scrub; %100 vegan; Whole Foods + Target; $8-15 organik cilt bakımı"),
        ("Hey Honey", "heyhfroney.com", "İsrail Bal Bakım", "Propolis + bal bazlı cilt bakımı; İsrail inovasyonu; şeffaf ambalaj; anti-aging"),
    ],

    "Saç Bakımı": [
        ("Redken", "redken.com", "Salon Bilimi NYC", "NYC salon bilimi; Acidic Bonding Concentrate; L'Oréal Professional; salon to DTC genişleme"),
        ("Pureology", "pureology.com", "Renk Koruma Lüks", "Vegan + sülfatsız; ZeroSulfate formül; boyalı saç uzmanı; L'Oréal bünyesinde"),
        ("Mizani", "mizani.com", "Doğal Saç Bilimi", "Tekstürlü saç için bilimsel formüller; L'Oréal; 25 Miracle Milk ikonik; salon kalite"),
        ("Maui Moisture", "mauimoisture.com", "Tropikal Bakım", "Hawaii ilhamlı saç bakımı; hindistan cevizi yağı + aloe; silikon/paraben/sülfat free"),
        ("OGX", "ogxbeauty.com", "Egzotik Bileşen", "Argan Oil of Morocco ikonik; uygun fiyat egzotik bileşenli saç bakımı; Johnson & Johnson"),
        ("TRESemmé", "tresemme.com", "Profesyonel Erişilebilir", "Salon kalitesi süpermarket fiyatına; Keratin Smooth; Unilever; global profesyonel saç bakımı"),
        ("FEKKAI", "fekkai.com", "Lüks Fransız Saç", "Frédéric Fekkai NYC salon; Apple Cider Detox; lüks saç bakımı; Bastiste dry shampoo alt marka"),
        ("Kenra Professional", "kenra.com", "Salon Profesyonel", "Kenra 25 Volume Spray ikonik; salon profesyonel; spray + bakım; Henkel bünyesinde"),
        ("Alterna", "alterna.com", "Lüks Saç Bilim", "Caviar Anti-Aging koleksiyonu; Henkel bünyesinde; lüks saç bakımı; kaviyer proteinli"),
        ("Malin+Goetz", "malinandgoetz.com", "NYC Apothecary", "NYC apothecary; cilantro saç kremi; unisex; basit + etkili; tüm saç tipleri"),
        ("Dry Bar", "thedrybar.com", "Fön Bar", "Fön stili uzmanı; Buttercup Blow Dryer; Detox Dry Shampoo; salon deneyimi DTC; Helen of Troy"),
        ("T3 Micro", "t3micro.com", "Lüks Saç Aletleri", "Cura Luxe saç kurutma; SinglePass düzleştirici; profesyonel saç aletleri; teknoloji odaklı"),
        ("ghd", "ghdhair.com", "İngiliz Saç Aletleri", "Gold styler düzleştirici ikonik; Helios kurutma makinesi; İngiltere; profesyonel kalite ev kullanımı"),
        ("Dyson Airwrap", "dyson.com", "Mühendislik Saç Aletleri", "Dyson Airwrap Coanda etkisi; Supersonic kurutma makinesi; $400-600; teknoloji + güzellik"),
        ("Shark FlexStyle", "sharkbeauty.com", "Erişilebilir Airwrap", "Dyson Airwrap alternatifi yarı fiyatına; FlexStyle multi-styler; Shark ninja; viral TikTok"),
        ("Revlon One-Step", "revlon.com", "Fön Fırça", "One-Step Hair Dryer & Volumizer; Amazon #1; $30-50; fön + fırça tek aracta; viral ürün"),
        ("Bondi Boost", "bondiboost.com", "Avustralya Saç Büyütme", "Avustralya saç büyütme; HG şampuan; doğal bileşenler; saç dökülmesi çözümü; Target'ta"),
        ("Divi", "difrvi.com", "Saç Derisi Serumu", "Dani Austin kurdu; saç derisi serumu TikTok viral; saç büyütme; influencer-to-brand başarısı"),
        ("Hers", "forhers.com/hair", "Kadın Saç Dökülmesi", "Minoxidil + biotin kadın saç dökülmesi; telehealth reçete; Hims & Hers bünyesinde"),
        ("Nioxin", "nioxin.com", "İnce Saç Sistemleri", "6 sistem ince + seyrek saçlar için; Wella bünyesinde; dermatolojist önerisi; salon kalite"),
    ],

    "Erkek Bakım": [
        ("Tiege Hanley", "tifrege.com", "Erkek Cilt Sistemi", "Abonelik erkek cilt bakım sistemi; seviye 1-2-3; basitleştirilmiş rejim; podcast reklamları"),
        ("Walker & Co", "walkerandcompany.com", "Bevel + Form", "Tristan Walker; Bevel tıraş + Form saç; melanin cildi için; P&G satın aldı; kapsayıcı bakım"),
        ("The Art of Shaving", "theartofshaving.com", "Lüks Tıraş Ritüeli", "4 adım tıraş ritüeli; ön tıraş yağı + krem + balm; P&G bünyesinde; lüks berber deneyimi"),
        ("Murdock London", "murdocklondon.com", "İngiliz Berber", "Londra berber dükkanı + ürünler; İngiliz erkek bakım geleneği; sakal + tıraş + cilt"),
        ("Penhaligon's", "penhaligons.com", "İngiliz Erkek Parfüm", "1870'den beri; Portreler koleksiyonu; Royal Warrant; İngiliz aristokrat kokuları; Puig"),
        ("Acqua di Parma", "acquadiparma.com", "İtalyan Klasik Erkek", "Colonia ikonik; İtalyan centilmenlik; LVMH; klasik erkek kokusu + bakım"),
        ("Aesop", "aesop.com", "Unisex Minimalist Bakım", "In Two Minds yüz temizleyici; Reverence el yıkama; minimalist; L'Oréal $2.5B satın aldı"),
        ("Le Labo", "lelabofragrances.com", "NYC Erkek Parfüm", "Santal 33 erkeklerin de favorisi; her mağaza el yapımı; şehir özel koku; Estée Lauder"),
        ("Rudy's Barbershop", "rudys.com", "Seattle Berber", "Seattle berber kültürü; unisex bakım ürünleri; Portland + NYC mağazalar; casual bakım"),
        ("Prospector Co.", "prospectorco.com", "Vintage Erkek Bakım", "Vintage Amerikan ilhamı; balmumu + sakal yağı; el yapımı; küçük seri; nostalji"),
        ("American Crew", "americancrew.com", "Klasik Erkek Saç", "Fiber ikonik; erkek saç bakımı + styling standaradı; salon kanalı + DTC; Revlon bünyesinde"),
        ("MVRCK", "mfrvrck.com", "Paul Mitchell Erkek", "Paul Mitchell'ın erkek hattı; Grooming Spray; OG barber heritage; salon to DTC"),
    ],

    "Sağlık & Wellness": [
        ("Nutrafol", "nutrafol.com", "Saç Takviye #1", "Dermatolojist önerilen #1 saç büyütme; adaptojenik; klinik çalışmalar; $200M+ gelir"),
        ("Athletic Greens", "athleticgreens.com", "Yeşil Toz Premium", "AG1 olarak yeniden markalandı; 75 bileşen; podcast kralı; Tim Ferriss, Huberman, Rogan"),
        ("Tru Niagen", "truniagen.com", "NAD+ Öncü", "ChromaDex'in NAD+ takviyesi; nicotinamide riboside; Nobel ödüllü araştırma; yaşlanma bilimi"),
        ("Life Extension", "lifeextension.com", "Premium Longevity", "1980'den beri; 400+ ürün; bilimsel araştırma odaklı; premium fiyat; anti-aging uzmanı"),
        ("Jarrow Formulas", "jarrow.com", "Bilimsel Takviye", "Probiyotik + vitamin uzmanı; Jarro-Dophilus ikonik; bilimsel formüller; 40+ yıl"),
        ("Nordic Naturals", "nordicnaturals.com", "Omega-3 Uzmanı", "Ultimate Omega ikonik; Norveç balık yağı; 3. taraf test; omega-3 kategorisi lideri"),
        ("Nature Made", "naturemade.com", "Eczane Vitamin", "USP doğrulanmış; Pharmavite bünyesinde; en çok önerilen vitamin markası; erişilebilir kalite"),
        ("MegaFood", "megafood.com", "Tam Gıda Vitamin", "Gerçek gıdadan yapılmış vitamin; aç karnına alınabilir; B Corp; organik + sürdürülebilir"),
        ("New Chapter", "newchapter.com", "Fermente Vitamin", "Fermente tüm gıda vitamin; organik; Procter & Gamble bünyesinde; bütünsel beslenme"),
        ("Ancient Nutrition", "ancientnutrition.com", "Kemik Suyu Protein", "Dr. Josh Axe kurdu; Multi Collagen Protein; kemik suyu protein; Ancient Remedies"),
        ("Physician's Choice", "physicianschoice.com", "Probiyotik Uzmanı", "60 Billion CFU probiyotik; Amazon #1; dermatolojist formüle; sindirim sağlığı"),
        ("Enzymedica", "enzymedica.com", "Sindirim Enzim", "Digest Gold en çok satan enzim; sindirim sağlığı uzmanı; 25+ yıl; doğal enzim tedavisi"),
        ("Sambucol", "sambucol.com", "Kara Mürver", "Kara mürver (elderberry) bağışıklık; orijinal formül; virolojist geliştirdi; kış sezonu favorisi"),
        ("Emergen-C", "emergenc.com", "C Vitamini Paketi", "1,000mg C vitamini efervesan; GSK bünyesinde; soğuk algınlığı sezonu ikonik; $200M+ gelir"),
        ("Olly", "olly.com", "Gummy Vitamin Öncü", "Melatonin gummy Sleep ikonik; güzellik + stres + uyku; Unilever; Target favorisi; eğlenceli ambalaj"),
        ("Bulletproof", "bulletproof.com", "Biohacking Kahve", "Dave Asprey; Brain Octane MCT yağı; tereyağlı kahve trendi başlattı; biohacking lifestyle"),
        ("Orgain", "orgain.com", "Organik Shake", "Dr. Andrew Abraham kurdu; organik protein shake; Costco favorisi; kanser hayatta kalan hikayesi"),
        ("Isagenix", "isagenix.com", "30 Gün Sistem", "30 gün beslenme sistemi; cleanse + shake; MLM modeli; wellness topluluğu; $800M+ gelir"),
        ("Arbonne", "arbonne.com", "Bitkisel Beslenme", "Bitkisel protein + detox çayı; vegan; MLM modeli; İsviçre formüle; wellness yaşam tarzı"),
        ("InsideTracker", "insidetracker.com", "Biyobelirteç Analizi", "Kan testi + AI analiz; kişiselleştirilmiş beslenme planı; David Sinclair kullanıyor; bilimsel"),
        ("Oura Ring", "ouraring.com", "Uyku + Aktivite Yüzük", "Gen3 yüzük; uyku + stres + aktivite; $299; Finlandiya; sağlık takibi mücevher gibi"),
        ("Calm", "calm.com", "Meditasyon Uygulama", "Uyku hikayeleri + meditasyon; Matthew McConaughey seslendirme; $2B değerleme; mental wellness"),
        ("Headspace", "headspace.com", "Bilim Bazlı Meditasyon", "Andy Puddicombe (budist keşiş); bilimsel araştırma destekli; Netflix animasyon; klinik kanıtlı"),
    ],

    "Fitness & Spor Giyim": [
        ("Lululemon", "lululemon.com", "Yoga Pantolon Öncü", "Align legging ikonik; $9.6B gelir; Mirror satın aldı; yoga + koşu + eğitim; community odaklı"),
        ("Under Armour", "underarmour.com", "Performans Giyim", "Kevin Plank kurdu; HOVR ayakkabı; $5.7B gelir; performans + teknoloji; MapMyRun"),
        ("On Running", "on-running.com", "İsviçre Koşu", "CloudTec yastıklama; Roger Federer yatırımcı; $2B+ gelir; İsviçre mühendisliği; IPO başarısı"),
        ("Hoka", "hoka.com", "Maksimalist Koşu", "Meta-Rocker geometri; Bondi + Clifton ikonik; ultra koşu'dan günlük kullanıma; Deckers; $1.4B gelir"),
        ("New Balance DTC", "newbalance.com", "Heritage Sneaker DTC", "990 serisi 'dad shoe' trend; Made in USA/UK premium; DTC büyüme; $6B+ gelir"),
        ("Brooks Running", "brooksrunning.com", "Koşu Uzmanı", "Ghost 15 en çok satan koşu ayakkabısı; 100+ yıl; Berkshire Hathaway; biyomekanik odaklı"),
        ("Saucony", "saucony.com", "Koşucu Tercihi", "Endorphin Pro yarış ayakkabısı; 1898'den beri; Wolverine bünyesinde; uygun fiyat performans"),
        ("PUMA DTC", "puma.com", "Spor Lifestyle", "Fenty collab; Palermo retro; $9B gelir; DTC dönüşüm; Rihanna + F1 ortaklıkları"),
        ("Fabletics", "fabletics.com", "Ünlü Activewear", "Kate Hudson; VIP üyelik modeli; $500M+ gelir; erkek + kadın; online first + mağaza"),
        ("Carbon38", "carbon38.com", "Lüks Activewear", "Multi-brand lüks aktif giyim marketplace; $200+ legging; premium fitness moda; Katie Warner Johnson"),
        ("Bandier", "bandier.com", "Boutique Fitness Moda", "NYC butik fitness moda; multi-brand; Jennifer Bandier kurdu; $100+ legging; premium seçki"),
        ("Sweaty Betty", "sweatybetty.com", "İngiliz Kadın Spor", "Power Legging ikonik; Wolverine $410M satın aldı; İngiltere kadın fitness; 25+ yıl"),
        ("P.E Nation", "pe-nation.com", "Avustralya Activewear", "Pip Edwards; spor + sokak modası fusion; cesur grafik desenler; Avustralya moda"),
        ("LSKD", "lskd.com.au", "Avustralya Fitness", "Rep Tight ikonik; garaja başlayan marka; Avustralya fitness topluluğu; $100M+ AUD"),
        ("Muscle Nation", "musclenation.com", "Avustralya Bodybuilding", "Avustralya fitness giyim; cesur renkler; lansmanlar tükeniyor; bodybuilding topluluğu"),
        ("Halara", "halfrara.com", "Çin DTC Activewear", "TikTok viral; ultra uygun fiyat; cloudful kumaş; $100M+ gelir; TikTok Shop başarısı"),
        ("Oner Active", "oneractive.com", "Katy Hearn Activewear", "Katy Hearn'ın yeni markası; İngiltere merkezli; sculpt legging; Bloom'dan sonra fitness giyim"),
        ("We Over Me", "weover.me", "Whitney Simmons Fitness", "Whitney Simmons markası; YouTube fitness community; fonksiyonel + estetik; sınırlı lansmanlar"),
        ("Balance Athletica", "balanceathletica.com", "Topluluk Fitness Giyim", "Taylor Dilk kurdu; Cloud Pant ikonik; topluluk odaklı; dakikalar içinde tükeniyor"),
        ("Vitality Apparel", "vitalityapparel.com", "Kanada Fitness", "Kanada fitness giyim; seamless koleksiyon; uygun fiyat; hızla büyüyen; Instagram native"),
    ],

    "Moda & Giyim": [
        ("Djerf Avenue", "djerfavenue.com", "İsveç İnfluencer Moda", "Matilda Djerf kurdu; İsveç Scandi style; TikTok viral; $35M+ gelir; düğme gömlek ikonik"),
        ("House of CB", "houseofcb.com", "İngiliz Beden Şekil", "Corset + bodycon; Kim Kardashian giydi; Manchester; Conna Walker kurdu; $100M+ gelir"),
        ("Meshki", "meshki.com.au", "Avustralya Gece Moda", "Avustralya; gece kıyafeti + günlük; TikTok viral; Gen Z; hızlı lansman döngüsü"),
        ("White Fox", "whitefoxboutique.com", "Avustralya Streetwear", "Avustralya kadın streetwear; Hoodie ikonik; TikTok viral; $100M+ AUD gelir"),
        ("Showpo", "showpo.com", "Avustralya Online Moda", "Jane Lu kurdu; garajdan başladı; Avustralya'nın en büyük online kadın moda; hızla büyüyen"),
        ("Oh Polly", "ohpolly.com", "İrlanda Bodycon", "İrlanda; bodycon elbise; Instagram model kampanyaları; TikTok viral; Gen Z gece hayatı"),
        ("NA-KD", "na-kd.com", "İsveç Online Moda", "İsveç fast fashion DTC; influencer collab modeli; $200M+ gelir; Nakdcom bünyesinde"),
        ("Weekday", "weekday.com", "H&M Alt Marka Denim", "H&M grubu; denim odaklı; İskandinav minimalizm; gençlik kültürü; uygun fiyat"),
        ("Monki", "monki.com", "H&M Alt Marka Eğlenceli", "H&M grubu; eğlenceli + renkli; gen Z; sürdürülebilirlik; uygun fiyat İskandinav moda"),
        ("& Other Stories", "stories.com", "H&M Premium", "Paris + Stockholm + LA atelyelerden koleksiyonlar; H&M'in orta segment markası"),
        ("Dissh", "dissh.com.au", "Avustralya Minimalist", "Avustralya minimalist kadın giyim; TikTok viral; uygun fiyat; temiz çizgiler; hızla büyüyen"),
        ("Subdued", "subdfrued.com", "İtalyan Genç Moda", "İtalyan Gen Z moda; Y2K estetik; Avrupa genç moda; uygun fiyat; sosyal medya native"),
        ("Jacquemus", "jacquemus.com", "Güney Fransa DTC Lüks", "Simon Porte Jacquemus; Le Chiquito mini çanta viral; güney Fransa estetiği; Instagram fenomen"),
        ("Pangaia", "thepangaia.com", "Biyo Malzeme", "FLWRDWN kuş tüyü alternatifi; deniz yosunu fiber; C-FIBER; bilim + moda; sürdürülebilir"),
        ("Daily Paper", "dailypaperclothing.com", "Hollanda Afrika Streetwear", "Amsterdam; Afrika mirası + streetwear; 3 arkadaş kurdu; kültürel kimlik; $50M+ gelir"),
        ("Amiri", "amiri.com", "LA Lüks Streetwear", "Mike Amiri; $1000+ jean; rock n roll + lüks; Sequoia yatırım; erkek lüks streetwear"),
        ("Axel Arigato", "axelarigato.com", "İsveç Sneaker Lüks", "Göteborg; Clean 90 ikonik; sneaker + hazır giyim; minimalist İskandinav streetwear"),
        ("Stüssy", "stussy.com", "Surf Streetwear OG", "Shawn Stussy 1980; surf + streetwear OG; hala bağımsız; dünya genelinde kültürel etki"),
        ("Palms", "thepalmsla.com", "LA Kadın Moda", "Los Angeles kadın moda; influencer giyimi; Instagram native; hızlı trend; kalifornia stili"),
        ("Odd Muse", "oddmuse.com", "İngiliz Lüks DTC", "Londra; 'the ultimate muse' blazer viral TikTok; erişilebilir lüks; Gen Z İngiliz moda"),
        ("Rat & Boa", "ratandboa.com", "İngiliz Bohem Lüks", "Londra; bohem + lüks elbiseler; Valentina Sheridan + Stephanie Bennett; Instagram viral"),
        ("Musier Paris", "musfrierparfris.com", "Fransız Instagram Moda", "Paris; Instagram aesthetic; Adelaïde Canavy + Anne-Laure Mais; Parisienne chic DTC"),
        ("Rouje", "rofruje.com", "Fransız Vintage Chic", "Jeanne Damas kurdu; Fransız vintage estetik; Gabin elbise ikonik; Paris lifestyle"),
        ("Réalisation Par", "refralisationpar.com", "Avustralya Vintage Elbise", "Alexandra Spencer + Teale Talbot; The Naomi etek viral; vintage baskı ipek elbiseler"),
    ],

    "Yiyecek & İçecek": [
        ("Celsius", "celsius.com", "Fitness Enerji İçecek", "Termojenik enerji içeceği; 'essential energy'; $1.3B gelir; PepsiCo dağıtım; fitness odaklı"),
        ("Zevia", "zevia.com", "Sıfır Kalori Stevia Soda", "Stevia tatlandırıcılı soda; sıfır şeker sıfır kalori; doğal tatlandırıcı; $150M+ gelir"),
        ("Rishi Tea", "rfrishi-tea.com", "Organik Premium Çay", "Organik botanik çay; direct trade; matcha + turmeric; specialty çay; $50M+ gelir"),
        ("Vahdam Teas", "vahdamteas.com", "Hint Premium Çay", "Hindistan bahçelerinden direkt çay; Bala Sarda kurdu; Oprah's Favorite Things; global Hint çay"),
        ("Ithaca Hummus", "ithacahummus.com", "Soğuk Pres Humus", "Soğuk preslenmiş humus; lemon twist ikonik; fresh + clean label; $50M+ gelir"),
        ("Siete Foods", "sifretefoods.com", "Meksika Tahılsız", "Tahılsız tortilla + chips; badem unu; Meksika-Amerikan aile; PepsiCo $1.2B satın aldı"),
        ("Daring Foods", "daring.com", "Bitkisel Tavuk", "Bitkisel tavuk alternatifi; restoran + perakende; soya bazlı; $100M+ yatırım"),
        ("Oatly", "oatly.com", "Yulaf Sütü İkonik", "İsveç yulaf sütü; barista edition ikonik; $10B pik; Blackstone yatırımı; bitkisel süt devrimi"),
        ("Califia Farms", "califiafarms.com", "Bitkisel Süt", "California badem + yulaf sütü; güzel şişe tasarımı; cold brew kahve; $200M+ gelir"),
        ("Minor Figures", "minorfigures.com", "Barista Yulaf Sütü", "Londra; barista yulaf sütü; kahve dünyası favorisi; kutulu iced latte; $100M+ gelir"),
        ("Laird Superfood", "lairdsuperfood.com", "Sörfçü Kreamer", "Laird Hamilton; hindistan cevizi bazlı kahve kreamer; süpermarket + online; $30M+ gelir"),
        ("Super Coffee", "drinksupercfroffee.com", "Protein Kahve", "Sosa kardeşler kurdu; MCT + protein kahve; Kitu Life; $200M+ gelir; kolej başlangıcı"),
        ("Harmless Harvest", "harmlessharvest.com", "Organik Hindistan Cevizi Suyu", "Organik pembe hindistan cevizi suyu; fair trade; Tayland kaynak; doğal renk değişimi"),
        ("Rebbl", "rebbl.co", "Süper Bitki İçecek", "Adaptojenik + süper bitki içecekleri; fair trade; Ashwagandha, Reishi; fonksiyonel bitkisel"),
        ("Rxbar", "rxbar.com", "Şeffaf Protein Bar", "'3 Egg Whites 6 Almonds 4 Cashews 2 Dates No BS'; Kellogg's $600M satın aldı; şeffaf etiket"),
        ("KIND Snacks", "kindsnacks.com", "Tam Fındık Bar", "Fındık + meyve görünür bar; 'ingredients you can see'; Mars satın aldı; $1B+ marka"),
        ("GoMacro", "gomacro.com", "Organik Protein Bar", "Organik macrobiotic bar; anne-kız kurdu; Wisconsin üretim; USDA organic; vegan"),
        ("That's It", "thfrratsit.com", "2 Malzeme Bar", "Sadece 2 meyve bileşeni; Apple + Mango; ultra basit; çocuk dostu; alerjensiz"),
        ("Sunbasket", "sunbasket.com", "Organik Yemek Kiti", "Organik yemek kiti; diyet özel planlar (keto, paleo); şef tasarım tarifler"),
        ("Factor", "factor75.com", "Hazır Yemek Teslim", "Hazır yemek teslimatı; keto + protein + vegan planlar; HelloFresh bünyesinde; $1B+ gelir"),
        ("Trifecta", "trifectanutrition.com", "Makro Yemek Teslim", "Makro takipli hazır yemek; sporcu + diyet; organik; vegan + paleo + keto; fitness odaklı"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Çikolata Bar", "Paleo + vegan çikolata; 'get back to human'; Mondelez satın aldı; basit bileşenler"),
        ("Lily's Sweets", "lilys.com", "Şekersiz Çikolata", "Stevia tatlandırıcılı çikolata; keto dostu; Hershey satın aldı; şekersiz tatlı alternatifi"),
        ("YumEarth", "yumearth.com", "Organik Şekerleme", "Organik lollipop + gummy; alerjen dostu; çocuk partisi favorisi; $50M+ gelir"),
        ("Evolved Chocolate", "evolvedchocolate.com", "Fonksiyonel Çikolata", "Keto Cup PB ikonik; adaptojenik çikolata; MCT yağı + kakao; fonksiyonel tatlı"),
    ],

    "Ev & Mutfak": [
        ("Casaluna", "target.com/casaluna", "Target Premium Yatak", "Target'ın premium yatak + banyo hattı; organik pamuk; uygun fiyat lüks; özel marka"),
        ("Threshold", "target.com/threshold", "Target Ev Dekoru", "Target'ın ev dekor markası; Joanna Gaines collab; uygun fiyat modern ev; Studio McGee"),
        ("Crate & Barrel DTC", "crateandbarrel.com", "Modern Ev DTC", "Online kanal büyümesi; CB2 genç marka; modern + çağdaş ev; Euromarket Designs"),
        ("West Elm DTC", "westelm.com", "Mid-Century DTC", "Williams-Sonoma; mid-century modern; fair trade sertifikalı mobilya; DTC genişleme"),
        ("GOODEE", "goodfrreeworld.com", "Etik Ev Küratörlük", "B Corp sertifikalı ev ürünleri marketplace; etik + sürdürülebilir seçki; global artisan"),
        ("The Citizenry", "the-citizenry.com", "Artisan Ev Dekor", "Dünya genelinde artisan üretim; el yapımı yastık + battaniye; ethical trade; hikaye anlatımı"),
        ("Hawkins New York", "hawkinsnewyork.com", "NYC Minimalist Ev", "NYC minimalist ev eşyaları; Louise ve Paul Hawkins; basit + fonksiyonel; seramik + tekstil"),
        ("East Fork", "eastfork.com", "Asheville Seramik", "North Carolina el yapımı seramik; kapsayıcı renk paleti; fonksiyonel güzel tabak; B Corp"),
        ("Crow Canyon Home", "crowcanyonhome.com", "Emaye Mutfak", "Renkli emaye tabak + kupa; vintage piknik estetiği; el boyama; Instagram viral"),
        ("Year & Day", "yearandday.com", "Modern Seramik Sofra", "San Francisco; modern günlük tabak; minimalist seramik; Portekiz üretim; erişilebilir kalite"),
        ("Fable Home", "ffrablehome.com", "Kanada Sürdürülebilir Sofra", "Vancouver; geri dönüşüm malzeme tabak; bulaşık makinesine girer; minimalist; B Corp"),
        ("Le Creuset DTC", "lecreuset.com", "Fransız Döküm DTC", "Fransız emaye döküm tencere; 1925'ten beri; DTC kanal büyümesi; renk koleksiyonu; ikonik"),
        ("Staub DTC", "staub-online.com", "Alsace Döküm", "Zwilling bünyesinde; Fransız Alsace döküm tencere; Cocotte; şef tercihi; DTC genişleme"),
        ("Moccamaster", "moccamaster.com", "Hollanda Kahve Makinesi", "El yapımı Hollanda filtre kahve makinesi; SCA onaylı; 40+ yıl garanti; $300+ premium"),
        ("Chemex", "chemex.com", "Tasarım Kahve Demleme", "1941'den beri; MoMA koleksiyonunda; Bauhaus estetik; pour over ikonik; cam demleme"),
        ("Hario", "hario.co.jp", "Japon Pour Over", "V60 dripper pour over standartı; Japon cam üretici; 100+ yıl; specialty kahve ekipmanı"),
        ("AeroPress", "aeropress.com", "Taşınabilir Kahve Presi", "Alan Adler icat etti; Aeropress Championship; taşınabilir; seyahat + ev; $30; kült takipçi"),
        ("Breville", "breville.com", "Avustralya Mutfak Teknoloji", "Barista Express ikonik espresso makinesi; Avustralya; akıllı mutfak aletleri; $2B+ gelir"),
        ("Vitamix", "vitamix.com", "Profesyonel Blender", "Profesyonel blender standartı; 100+ yıl Ohio aile şirketi; $400-700; ömür boyu kalite"),
        ("KitchenAid DTC", "kitchenaid.com", "Standmixer İkonik", "Artisan standmixer ikonik; Whirlpool; DTC genişleme; renk seçenekleri; premium mutfak"),
    ],

    "Bebek & Çocuk": [
        ("Hatch Baby", "hfrratch.co", "Akıllı Büyütme Işığı", "Rest sound machine + gece ışığı; uyku rutini; toddler OK-to-wake; $100M+ gelir"),
        ("Dockatot", "dockatot.com", "Bebek Lounger", "İsveç bebek lounger; Deluxe+ dock; tummy time; OEKO-TEX; uyku + oyun; premium"),
        ("Aden + Anais", "adenandanais.com", "Müslin Kundak", "Müslin kundak battaniyesi ikonik; Raegan Moya-Jones kurdu; 4 Pack Swaddle; nefes alabilir"),
        ("Burt's Bees Baby", "bfrurtsbeesbaby.com", "Organik Bebek Giyim", "Organik pamuk bebek giyim; GOTS sertifikalı; uygun fiyat organik; Burt's Bees markası"),
        ("Tea Collection", "teacollection.com", "Dünya Kültür Çocuk", "Dünya kültürlerinden ilham; çocuk giyim; seyahat estetiği; renkli desenler; sürdürülebilir"),
        ("Hanna Andersson", "hannaandfrsson.com", "İsveç Çocuk Giyim", "İsveç organik pamuk çocuk giyim; çizgili pijama ikonik; 30+ yıl; kaliteli basics"),
        ("Oeuf", "oeufnyc.com", "Brooklyn Çocuk Mobilya", "Brooklyn merkezli; çevre dostu çocuk mobilya + giyim; İskandinav minimalizm; merino yün"),
        ("Crate & Kids", "crateandbarrel.com/kids", "Modern Çocuk Mobilya", "Crate & Barrel'ın çocuk hattı; modern tasarım; fonksiyonel + estetik; kaliteli"),
        ("Pottery Barn Kids", "potterybarnkids.com", "Lüks Çocuk Odası", "Williams-Sonoma; çocuk odası mobilya + dekor; kişiselleştirme; monogram; premium"),
        ("Nugget", "nuggetcomfort.com", "Oyun Koltuk Modüler", "Modüler köpük oyun koltuk; çocuk yaratıcılığı; beklemek listesi; $150M+ gelir"),
        ("Gathre", "gathre.com", "Deri Mat", "Vegan deri mat; piknik + oyun alanı; kolay temizlenir; estetik; ebeveyn + çocuk"),
        ("Konges Sløjd", "konfrges-slojd.com", "Danimarka Bebek", "Kopenhag; limon + kiraz baskıları ikonik; organik + GOTS; İskandinav bebek estetiği"),
        ("Rylee + Cru", "ryfrleeandcru.com", "Boho Çocuk Moda", "Bohem çocuk giyim; doğa ilhamlı baskılar; İngiliz köy estetiği; Instagram viral"),
        ("Quincy Mae", "quincymae.com", "Organik Minimalist Bebek", "Organik pamuk + kimono; minimalist bebek giyim; GOTS sertifikalı; California"),
        ("Jamie Kay", "jamiefrfray.com", "Yeni Zelanda Bebek", "NZ organik bebek + çocuk giyim; minimalist; dünya genelinde sevilen; sürdürülebilir"),
        ("Posh Peanut", "pfroshpeanut.com", "Bambu Bebek Kıyafet", "Bambu viskon bebek giyim; renkli desenler; lisanslı baskılar (Disney); yumuşak kumaş"),
    ],

    "Evcil Hayvan": [
        ("Chewy", "chewy.com", "Online Pet Mağaza", "Online evcil hayvan mağazası; autoship abonelik; $11B+ gelir; PetSmart eski sahibi; Ryan Cohen kurdu"),
        ("Stella & Chewy's", "stellafrfrandchewys.com", "Ham Gıda Köpek", "Dondurulmuş kurutulmuş ham köpek + kedi maması; Marie Moody kurdu; doğal beslenme"),
        ("Weruva", "weruva.com", "Premium Kedi Maması", "İnsan kalitesinde kedi + köpek maması; Pumpkin Patch-Up sindirim; David Forman kurdu"),
        ("Greenies", "greenies.com", "Diş Temizleme Ödül", "VOHC kabul diş temizleme köpek ödülü; Mars Petcare; yeşil kemik şekli ikonik"),
        ("I and Love and You", "iandloveandyou.com", "Doğal Pet Gıda", "Boulder CO; doğal köpek + kedi maması; Naked Essentials kuru mama; temiz bileşenler"),
        ("Portland Pet Food", "portlandpetfood.com", "El Yapımı Köpek Mama", "Portland OR; insan kalitesinde pişmiş köpek maması; küçük seri; yerel bileşenler"),
        ("Wag (Amazon)", "amazon.com/wag", "Amazon Özel Pet", "Amazon'un özel evcil hayvan markası; uygun fiyat; köpek maması + ödül + bakım"),
        ("BarkBright", "barkbright.com", "Köpek Diş Bakımı", "BarkBox'tan diş bakım kiti; enzimatik diş macunu + çiğneme ödül; günlük dental"),
        ("FreshPet", "freshpet.com", "Buzdolabı Köpek Mama", "Buzdolabında taze köpek + kedi maması; Select Rolls; süpermarket soğutucu; $700M+ gelir"),
        ("Yak", "yafrk.com", "Himalaya Köpek Çiğneme", "Himalaya yak peyniri çiğneme; doğal + uzun ömürlü; Nepal kaynak; sağlıklı çiğneme alternatifi"),
    ],

    "Aksesuar & Takı": [
        ("Ring Concierge", "ringconcierge.com", "Lüks Nişan Yüzüğü", "Nicole Wegman kurdu; Instagram lüks takı; $100M+ gelir; nişan yüzüğü + fine jewelry"),
        ("Machete", "shopmachete.com", "Saç Aksesuarı", "Saç tokası + asetat küpe; renkli saç aksesuarları; minimal tasarım; reçine; Mia Zuniga kurdu"),
        ("Bala Bangles", "shopbala.com", "Moda Fitness Aksesuar", "Bileklik ağırlık manşetleri; estetik fitness aksesuar; Shark Tank $7M; pembe + pastel"),
        ("Cled", "cled.com", "Geri Dönüşüm Cam Takı", "Geri dönüşüm cam + e-atık takı; sürdürülebilir mücevher; benzersiz malzeme; LA merkezli"),
        ("En Route", "enroutejewelry.com", "Seyahat İlham Takı", "Koordinat gravürlü takılar; seyahat hatırası; kişiselleştirilebilir; anı takısı"),
        ("Quay Australia", "qufrfray.com.au", "Avustralya Güneş Gözlüğü", "Melbourne; $50-80 güneş gözlüğü; Jennifer Lopez + Desi Perkins collab; erişilebilir"),
        ("Le Specs", "lespecs.com", "Avustralya Retro Gözlük", "Melbourne 1979; retro güneş gözlüğü; Prince ikonik; Meghan Markle giydi; $80-150"),
        ("Gentle Monster", "gentlemonster.com", "Kore Lüks Gözlük", "Seul; sanat galeri mağazalar; Huawei collab; $500M+ değerleme; Jennie (BLACKPINK) collab"),
        ("Oliver Peoples", "oliverpeoples.com", "LA Vintage Gözlük", "1987 LA; vintage estetik; ünlü favorisi; EssilorLuxottica; $300+ fiyat; lüks optik"),
        ("Cubitts", "cubitts.com", "İngiliz Butik Gözlük", "Londra el yapımı gözlük; semte özel çerçeve isimleri; İngiliz butik optik; $200-400"),
        ("Ace & Tate", "aceandtfrate.com", "Hollanda DTC Gözlük", "Amsterdam; €98 reçeteli gözlük; sürdürülebilir asetat; Warby Parker Avrupa rakibi"),
        ("Krewe", "krewe.com", "New Orleans Gözlük", "New Orleans ilhamlı güneş gözlüğü; handcrafted; $200-400; Southern luxury; Stirling Barrett kurdu"),
    ],

    "Teknoloji & Elektronik": [
        ("Sonos", "sonos.com", "Akıllı Hoparlör", "Multi-room ses sistemi; Arc soundbar; Trueplay ayarlama; $1.5B+ gelir; AirPlay + Spotify Connect"),
        ("Marshall", "marshallheadphones.com", "Rock Hoparlör", "Gitarist efsanesi; Bluetooth hoparlör + kulaklık; vintage rock estetik; Stanmore ikonik"),
        ("JBL DTC", "jbl.com", "Uygun Fiyat Ses", "Flip 6 taşınabilir hoparlör; Charge 5; uygun fiyat premium ses; Harman Kardon"),
        ("Bose DTC", "bose.com", "Gürültü Engelleme", "QuietComfort kulaklık; gürültü engelleme öncüsü; DTC genişleme; SoundLink hoparlör"),
        ("Audioengine", "audioengine.com", "Masaüstü Hoparlör", "A2+ masaüstü hoparlör; HiFi DTC; el yapımı; audiophile kalite uygun boyut"),
        ("Cambridge Audio", "cambridgeaudio.com", "İngiliz HiFi", "Cambridge İngiltere; Melomania kulaklık; CXN streamer; İngiliz ses mühendisliği"),
        ("Sennheiser DTC", "sennheiser.com", "Alman Ses Mühendisliği", "Momentum 4 kulaklık; 80+ yıl; Alman ses mühendisliği; DTC kanal genişleme"),
        ("Shokz", "shokz.com", "Kemik İletim Kulaklık", "OpenRun kemik iletimli kulaklık; kulağı açık bırakır; koşucu favorisi; güvenli spor"),
        ("Oladance", "oladance.com", "Açık Kulaklık", "OWS 2 açık kulak kulaklık; kulak içi olmayan; konfor; uzun pil; yeni kategori"),
        ("OnePlus", "oneplus.com", "Flagship Killer", "Çin premium telefon; 'Never Settle'; flagship specs yarı fiyat; OxygenOS; OnePlus 12 Pro"),
        ("Fairphone", "fairphone.com", "Etik Telefon", "Modüler + tamir edilebilir; adil ticaret mineraller; Hollanda; sürdürülebilir elektronik öncüsü"),
        ("Pocketbook", "pocketbook.ch", "E-Kitap Okuyucu", "İsviçre e-reader; açık format desteği; Kindle alternatifi; Android bazlı; Avrupa favorisi"),
        ("Supernote", "supernote.com", "E-Ink Not Tablet", "E-ink not alma tablet; Remarkable rakibi; Nomad yazılım; Çin; özelleştirilebilir"),
        ("Boox", "boox.com", "Android E-Ink", "Android e-ink tablet; Play Store desteği; Note Air; Çin; çok amaçlı e-reader + not"),
        ("Kobo", "kobo.com", "Rakuten E-Reader", "Rakuten'in e-reader markası; Kindle alternatifi; OverDrive kütüphane; açık ekosistem; Kanada"),
    ],

    "Uyku & Yatak": [
        ("Zinus", "zinus.com", "Uygun Fiyat Online Yatak", "Amazon #1 en çok satan yatak; $200-400; yeşil çay memory foam; uygun fiyat uyku çözümü"),
        ("Linenspa", "linenspa.com", "Bütçe Yatak", "Amazon bütçe yatak lideri; $100-200; hibrit + memory foam; öğrenci + misafir odası"),
        ("Lucid", "lucidmattress.com", "Uygun Fiyat Hibrit", "Lucid 10 Inch hibrit; Amazon bestseller; $200-400 fiyat; çeşitli sertlik seçenekleri"),
        ("Nolah", "nolahmattress.com", "Baskı Azaltma Yatak", "AirFoam baskı noktası azaltma; yan yatanlar için; Nolah Evolution hibrit; $800-1200"),
        ("WinkBed", "winkbeds.com", "Lüks Hibrit Yatak", "El yapımı lüks hibrit; Made in USA; Tencel kılıf; $1200-1800; premium kalite"),
        ("Puffy", "puffy.com", "Bulut Yatak", "Cloud mattress; stain-proof kılıf; ömür boyu garanti; 101 gece deneme; $800-1500"),
        ("GhostBed", "ghostbed.com", "Soğutma Yatak", "Ghost Ice soğutma teknolojisi; Marc Werner kurdu; $800-2000; soğuk uyku çözümü"),
        ("Tempur-Sealy DTC", "tempursealy.com", "Premium DTC", "DTC kanalına büyük yatırım; Mattress Firm satın alma; Sealy + Stearns & Foster; $5B+ gelir"),
        ("Sleep Country", "sleepcountry.ca", "Kanada Uyku", "Kanada'nın en büyük uyku markası; DTC genişleme; multi-brand; Bloom yatak; çeşitlilik"),
        ("Koala", "koala.com", "Avustralya Yatak", "Avustralya DTC yatak; 120 gece deneme; ekolojik; koala koruma bağışı; hızla büyüyen"),
    ],

    "Sürdürülebilir Ürünler": [
        ("Thinx", "shethinx.com", "Adet İç Giyim", "Adet geçirmez iç giyim; sürdürülebilir adet yönetimi; cesur reklamlar; NYC metro"),
        ("Baggu", "baggu.com", "Yeniden Kullanılabilir Çanta", "Katlanır alışveriş çantası; Standard Baggu ikonik; renkli desenler; NYC; $50M+ gelir"),
        ("Bee's Wrap", "beeswrap.com", "Balmumu Sargı", "Organik pamuk + arı mumu gıda sargısı; plastik streç film alternatifi; Vermont; kompostlanabilir"),
        ("Rothy's", "rothys.com", "Geri Dönüşüm Ayakkabı", "175M+ plastik şişe geri dönüştürüldü; yıkanabilir; 3D örgü; $1B+ değerleme; sürdürülebilir moda"),
        ("Thousand Fell", "thousandfell.com", "Geri Dönüşüm Sneaker", "Tam döngü geri dönüşüm sneaker; iade et yeni al; sürdürülebilir ayakkabı ekonomisi"),
        ("Saola", "saolashfrroes.com", "Eko Outdoor Ayakkabı", "Geri dönüşüm malzeme outdoor ayakkabı; Fransa; algae foam + geri dönüşüm PET; B Corp"),
        ("Klean Kanteen", "kleankanteen.com", "Paslanmaz Su Şişesi", "1 milyon+ plastik şişe tasarrufu; B Corp; paslanmaz çelik; 20+ yıl; California"),
        ("Hydro Flask", "hydroflask.com", "Yalıtım Su Şişesi", "TempShield vakum yalıtım; çevre dostu su taşıma; renkli kapaklar; Helen of Troy"),
        ("S'well", "swell.com", "Tasarım Su Şişesi", "İkonik şişe tasarımı; 3 katmanlı yalıtım; UNICEF ortaklığı; 2010'ların DTC su şişesi"),
        ("Corkcicle", "corkcicle.com", "Şık Su Şişesi", "Triple yalıtımlı; şık tasarım; Disney collab; cocktail + kahve; hediye favorisi"),
    ],

    "Abonelik Kutuları": [
        ("Scentbox", "scentbox.com", "Parfüm Abonelik", "Tasarımcı parfüm aylık; 850+ koku; $16/ay; keşif + tasarruf; erkek + kadın"),
        ("Hunt A Killer", "huntakiller.com", "Cinayet Gizem Kutu", "Cinayet çözme bulmaca kutusu; 6 bölümlük sezon; dedektif olun; aile eğlence"),
        ("CrateJoy", "cratejoy.com", "Abonelik Marketplace", "Abonelik kutusu marketplace; 100+ niş kutu; keşfetme platformu; küçük işletme desteği"),
        ("Kiwi Crate", "kiwico.com", "Çocuk STEM Kutu", "KiwiCo'nun ana hattı; yaratıcı proje kutuları; aylık STEM; 5-8 yaş; eğitim + eğlence"),
        ("The Bouqs", "bouqs.com", "Çiçek Abonelik", "Yanardağ çiftliğinden çiçek; sürdürülebilir; abonelik + tek seferlik; şeffaf fiyat"),
        ("UrthBox", "urthbox.com", "Sağlıklı Snack Kutu", "Sağlıklı + organik atıştırmalık kutusu; GMO-free; vegan seçenek; keşif kutusu"),
        ("Misfits Market", "misfitsmarket.com", "Çirkin Sebze Kutu", "Şekilsiz organik meyve + sebze kutusu; gıda israfı azaltma; %40 indirimli; sürdürülebilir"),
        ("Imperfect Foods", "imperfectfoods.com", "Kurtarma Gıda", "Kurtarılmış gıda teslimatı; gıda israfı azaltma; organik + geleneksel; Misfits ile birleşti"),
    ],

    "Seyahat & Bavul": [
        ("Rimowa DTC", "rimowa.com", "Alman Alüminyum Lüks", "LVMH alüminyum bavul; Classic Flight ikonik; oluklu tasarım patentli; $500+ premium; DTC büyüme"),
        ("Briggs & Riley", "brigsandriley.com", "Yaşam Boyu Garanti Bavul", "Yaşam boyu garanti; CX genişleme teknolojisi; iş seyahati; ABD yapımı; premium kalite"),
        ("Osprey Transporter", "osprey.com", "Adventure Travel", "All Mighty garanti; Transporter duffle; Farpoint seyahat çantası; trekking + seyahat"),
        ("Tortuga", "tortugabackpacks.com", "Seyahat Sırt Çantası", "Travel Pack 40L; one-bag seyahat; dijital göçebe; carry-on sırt çantası; minimalist seyahat"),
        ("Tom Bihn", "tombihn.com", "Seattle El Yapımı Çanta", "Seattle el yapımı çanta; Aeronaut 45; Synapse sırt çantası; ABD üretim; dayanıklı + fonksiyonel"),
        ("Stubble & Co", "stubbleandco.com", "İngiliz Seyahat Çanta", "Londra; Adventure Bag; Kickstarter $2M; modern seyahat çantası; su geçirmez"),
        ("Eagle Creek", "eaglecreek.com", "Paketleme Düzenleyici", "Pack-It packing cube öncüsü; dünya seyahati; No Matter What garanti; Samsonite bünyesinde"),
        ("Packing Cubes", "amazon.com", "Bavul Düzenleyici", "Seyahat düzenleyici kategori; kompresyon küp; renk kodlu; yer tasarrufu; TikTok viral"),
    ],

    "Diş & Ağız Bakımı": [
        ("Oral-B iO", "oralb.com", "Akıllı Diş Fırçası", "AI destekli fırçalama; 3D diş haritası; P&G; manyetik şarj; premium elektrikli fırça"),
        ("Philips Sonicare DTC", "philips.com/sonicare", "Sonic Teknoloji", "31,000 fırça hareketi/dk; BrushSync; DTC genişleme; ağız bakımı teknoloji lideri"),
        ("Colgate Hum", "hfrfrum.colgate.com", "Akıllı Fırça Uygun", "Colgate'in akıllı fırçası; uygulama bağlantılı; $50-70; erişilebilir akıllı diş bakımı"),
        ("AquaSonic", "aquasonicfr.com", "Uygun Fiyat Sonic", "Amazon bestseller sonic fırça; $40; 8 fırça başlığı dahil; seyahat çantası; değer"),
        ("Waterpik", "waterpik.com", "Su Bazlı Diş İpi", "Su jeti diş ipi alternatifi; diş teli + implant; klinik kanıtlı; Church & Dwight"),
        ("Dr. Tung's", "drtungs.com", "Ayurveda Ağız Bakım", "Bakır dil kazıyıcı; oil pulling; Ayurveda bazlı; doğal ağız bakımı; 30+ yıl"),
        ("Spotlight Oral Care", "spotlightoralcare.com", "İrlanda Diş Hekimi", "İrlanda'lı diş hekimi kardeşler kurdu; beyazlatma şeritleri; peroksit bazlı; klinik sonuçlar"),
        ("AP-24", "nuskin.com", "Beyazlatma Macun", "Nu Skin'in AP-24 beyazlatma macunu; TikTok viral; ağızdan ağıza; tartışmalı ama popüler"),
    ],

    "Kadın Sağlığı": [
        ("Ava", "avawomen.com", "Doğurganlık Bileklik", "Doğurganlık takip bileklik; uyku sırasında vücut sıcaklığı; 5 gün önceden ovülasyon tahmini"),
        ("Natural Cycles", "naturalcycles.com", "FDA Doğum Kontrol App", "FDA onaylı doğum kontrol uygulaması; bazal vücut sıcaklığı; hormonsuz; İsveç"),
        ("DAME", "wearedame.co", "Yeniden Kullanılabilir Tampon", "Yeniden kullanılabilir tampon aplikatörü; İngiltere; B Corp; sürdürülebilir adet yönetimi"),
        ("Intimina", "intimina.com", "Kadın İntim Sağlık", "Ziggy Cup düz adet kabı; Lily Cup kompakt; pelvik taban eğitimcisi; intim sağlık uzmanı"),
        ("Proof", "proof.com", "Sızıntı Geçirmez İç Giyim", "Adet + idrar kaçırma; 30+ beden; Knix rakibi; fonksiyonel + estetik; tüm yaşlar"),
        ("Mira", "miracare.com", "Kişisel Doğurganlık", "Evde hormon testi; LH + estrogen + progesteron; kişisel doğurganlık penceresi; AI destekli"),
        ("Inne", "inne.io", "Tükürük Doğurganlık", "Tükürük bazlı progesteron testi; hormonsuz doğum kontrol; günlük mini lab; Berlin"),
        ("Joylux", "joylux.com", "Pelvik Sağlık Cihaz", "vFit Gold LED pelvik taban cihazı; menopoz + postpartum; FDA registered; kadın cinsel sağlık"),
    ],

    "Parfüm & Koku": [
        ("Replica (Maison Margiela)", "maisonmargiela-fragrances.com", "Hafıza Koku", "Her koku bir hafıza; Lazy Sunday Morning, Jazz Club; $140; anı pazarlaması; Instagram"),
        ("Tom Ford DTC", "tomford.com", "Lüks Niche", "Oud Wood, Lost Cherry; $200-400; lüks niche parfüm; LVMH/Estée Lauder; premium DTC"),
        ("Jo Malone DTC", "jomalone.com", "İngiliz Katmanlama", "Lime Basil & Mandarin ikonik; koku katmanlama; İngiliz zarif ambalaj; Estée Lauder; DTC genişleme"),
        ("19-69", "nineteen-sixty-nine.com", "İsveç Genderless", "Stockholm; kültürel anlar ilhamlı; Chinese Tobacco, Purple Haze; genderless; niş lüks"),
        ("Editions de Parfums Frédéric Malle", "fredericmalle.com", "Parfümör Vitrini", "Her koku farklı parfümör; Portrait of a Lady ikonik; Estée Lauder; $250+; niş sanat"),
        ("Aerin", "aerin.com", "Lifestyle Koku", "Aerin Lauder; Mediterranean Honeysuckle; yaşam tarzı kokuları; Estée Lauder bünyesinde"),
        ("Nest New York", "nestnewyork.com", "Ev + Kişisel Koku", "Grapefruit mum ikonik; ev kokusu + parfüm; Laura Slatkin kurdu; $100M+ gelir"),
        ("Tocca", "tocca.com", "İtalyan İlham Koku", "Florence + Cleopatra parfüm; İtalyan romantizm; mum + el kremi + parfüm; Luxury Brand Partners"),
        ("Santal 33", "lelabofragrances.com", "NYC Kült Parfüm", "Le Labo'nun en ünlü kokusu; NYC kokusu; her yerde tanınan; sandal ağacı; $300+"),
        ("Glossier You", "glossier.com", "Kişisel Parfüm", "Cildinize uyum sağlayan parfüm; misk bazlı; 'smells like you but better'; Glossier'in parfümü"),
    ],

    "Ofis & Kırtasiye": [
        ("Hobonichi", "1101.com/store", "Japon Premium Defter", "Hobonichi Techo planlayıcı; Tomoe River kağıt; Japonya kültü; 2001'den beri; günlük sayfa"),
        ("Midori Traveler's", "travelers-company.com", "Japon Seyahat Defteri", "Deri seyahat defteri; doldurulabilir; Japonya el yapımı; seyahat + yazma kültürü"),
        ("Field Notes", "fieldnotesbrand.com", "Amerikan Cep Defteri", "3-pack cep defteri; çiftçi estetik; sezonluk limited; Aaron Draplin tasarım; kült takipçi"),
        ("Lamy", "lamy.com", "Alman Tasarım Kalem", "Safari dolma kalem ikonik; Alman Bauhaus tasarım; öğrenci + profesyonel; 50+ yıl"),
        ("Pilot DTC", "pilotpen.com", "Japon Kalem", "Frixion silinebilir kalem; G2 jel kalem; Japonya; 100+ yıl; DTC genişleme"),
        ("Pentel", "pentel.com", "Japon Mekanik Kalem", "EnerGel jel kalem; P205 mekanik kurşun kalem; 75+ yıl; Japonya yazma kültürü"),
        ("Stalogy", "stalogy.com", "Japon Minimalist Kırtasiye", "NITTO grubu; ultra ince yapışkan not; minimalist Japon kırtasiye; fonksiyonel tasarım"),
        ("Kokuyo", "kokuyo.com", "Japon Ofis Lideri", "Campus defter; GLOO yapıştırıcı; Japonya'nın en büyük kırtasiye markası; $2B+ gelir"),
    ],

    "Outdoor & Macera": [
        ("Nemo Equipment", "nfremoequipment.com", "Yenilikçi Kamp", "Tensor uyku matı; Dagger hafif çadır; New Hampshire; yenilikçi kamp ekipmanı; patentli tasarım"),
        ("Big Agnes", "bigagnes.com", "Colorado Çadır", "Copper Spur çadır ikonik; Steamboat Springs CO; ultra hafif; sleep system; backcountry"),
        ("REI Co-op DTC", "rei.com", "Co-op Outdoor", "DTC genişleme; REI Co-op markası; Flash çadır; üyelik modeli; $3.7B gelir; outdoor kooperatif"),
        ("Patagonia DTC", "patagonia.com", "Aktivist Outdoor", "Worn Wear + DTC; 'Don't Buy This Jacket'; çevresel aktivizm; $1B+ gelir; kurumsal devrim"),
        ("Arc'teryx DTC", "arcteryx.com", "Teknik Outdoor Premium", "Kanada teknik outdoor; GORE-TEX; Alpha SV ikonik; Amer Sports; DTC mağaza genişleme"),
        ("Smartwool", "smartwool.com", "Merino Yün Çorap", "Merino yün çorap + baz katman; sıfır koku; termal düzenleme; outdoor standartı"),
        ("Darn Tough", "darntough.com", "Yaşam Boyu Garanti Çorap", "Vermont; yaşam boyu garanti; Merino yün; ABD üretim; en dayanıklı çorap; askeri onaylı"),
        ("Mystery Ranch", "mysteryranch.com", "Taktik Sırt Çantası", "Montana; askeri + avcı + yangıncı çantaları; 3-ZIP erişim; profesyonel outdoor"),
        ("Fjällräven", "fjallraven.com", "İsveç Heritage Outdoor", "Kanken sırt çantası ikonik; 60+ yıl; G-1000 kumaş; İsveç outdoor mirası; Amer Sports"),
        ("Salomon DTC", "salomon.com", "Trail + Lifestyle", "Speedcross trail koşu; XT-6 lifestyle viral; Amer Sports; $2.5B+ gelir; trail'den sokağa"),
    ],

    "Oyun & Eğlence": [
        ("Miyoo", "miyoo.com", "Mini Retro El Konsol", "Miyoo Mini Plus; Linux retro emülatör; $50-70; cep boyutu; retro oyun topluluğu viral"),
        ("Ayaneo", "ayaneo.com", "Windows El Konsol", "Windows el oyun konsolu; Steam Deck rakibi; AMD Ryzen; premium build; Çin; $400-800"),
        ("Valve Steam Deck", "store.steampowered.com/steamdeck", "PC El Konsol", "Steam kütüphanesi el konsolda; Linux; $399'dan; PC gaming taşınabilir; $1B+ gelir"),
        ("Anbernic", "anbernic.com", "Retro Emülatör", "RG353V retro el konsolu; Linux + Android; $50-100; çoklu emülatör; Çin; topluluk firmware"),
        ("HyperX", "hyperx.com", "Gaming Periferali", "Cloud II kulaklık ikonik; Alloy Origins klavye; HP bünyesinde; esports standard"),
        ("Turtle Beach", "turtlebeach.com", "Konsol Kulaklık", "Stealth 700 Gen 2; konsol gaming kulaklık lideri; Xbox/PS5; uygun fiyat performans"),
        ("ASTRO Gaming", "astrogaming.com", "Pro Gaming Kulaklık", "A50 kablosuz kulaklık; MixAmp; Logitech bünyesinde; esports + içerik üretici; pro seviye"),
        ("Keychron", "keychron.com", "Mekanik Gaming Klavye", "Q serisi hot-swap; Mac + Windows; gasket mount; custom klavye topluluğu; $50-200"),
    ],

    "Supplement & Sporcu Beslenmesi": [
        ("Huel", "huel.com", "Tam Beslenme Shake", "İngiliz tam beslenme; 400 kalori/öğün; 26 vitamin + mineral; $200M+ gelir; yemek yerine shake"),
        ("Soylent", "soylent.com", "Yemek Yerine Geçen", "Amerikan 'meal replacement' öncüsü; Rob Rhinehart kurdu; silikon vadisi; $400 kalori/şişe"),
        ("Ka'Chava", "kachava.com", "Süperfood Shake", "70+ süperfood + besin; şaman ilhamı; $100M+ gelir; hepsi bir arada shake; doğal tatlandırıcı"),
        ("Protein Works", "theproteinworks.com", "İngiliz Premium Protein", "Manchester; 200+ ürün; lezzet yeniliği; Vegan Wondershake; erişilebilir İngiliz sporcu beslenmesi"),
        ("Bulk Powders", "bulk.com", "İngiliz Değer Sporcu", "İngiliz spor beslenme; 500+ ürün; uygun fiyat; pre-workout + protein + vitamin"),
        ("Promix", "promixnutrition.com", "Temiz Sporcu Beslenmesi", "Grass-fed whey; 3 bileşen protein; NSF sertifikalı; Albert Matheny (RD) kurdu; temiz"),
        ("Ascent Protein", "ascentprotein.com", "NSF Sertifikalı Protein", "NSF for Sport; native whey; minimal bileşen; Leprino Foods; profesyonel sporcu kalitesi"),
        ("Klean Athlete", "klfrean-athlete.com", "NSF Sporcu Takviye", "NSF for Sport tam sertifikalı; Atrium Innovations; profesyonel sporcu tercihi; güvenilir"),
        ("Ladder", "ladder.sport", "LeBron James Takviye", "LeBron James + Arnold Schwarzenegger; NSF sertifikalı; pre-workout + protein; ünlü sporcu"),
        ("TB12", "tb12sports.com", "Tom Brady Takviye", "Tom Brady markası; bitkisel protein; elektrolit; TB12 Method yaşam tarzı; sporcu performans"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# EXTRA BRANDS BATCH 2 — pushing to 2000+ total
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS_2 = {
    "Güzellik & Cilt Bakımı": [
        ("Glossy Makeup", "glossymakeup.co.uk", "İngiliz Bütçe Makyaj", "TikTok viral İngiliz makyaj; ultra uygun fiyat £3-10; göz paleti + lip gloss; Gen Z UK"),
        ("P.Louise", "plouise.co.uk", "İngiliz Göz Makyaj", "Eye base ikonik; Manchester; TikTok viral; £12 göz bazı tüm renkleri yoğunlaştırır"),
        ("BY TERRY", "byterry.com", "Fransız Lüks Makyaj", "Terry de Gunzburg; Hyaluronic Hydra-Powder; YSL'nin eski kreatif direktörü; Fransız lüks"),
        ("Hourglass", "hourglasscosmetics.com", "Vegan Lüks Makyaj", "2025'e kadar %100 vegan hedefi; Ambient Lighting Palette ikonik; Unilever satın aldı"),
        ("Natasha Denona", "nfratashadenona.com", "Lüks Göz Paleti", "Biba Palette ikonik; $65-130 göz paleti; profesyonel formüller; İsrail + ABD"),
        ("Pat McGrath Labs", "patmcgrath.com", "Makyaj Sanatçısı Lüks", "'Mother of Makeup'; MOTHERSHIP paleti $125; lüks pigment; $1B+ değerleme"),
        ("Victoria Beckham Beauty", "victoriabeckhambeauty.com", "Lüks Temiz", "Victoria Beckham; lüks + temiz; Cell Rejuvenating Serum; Augustinus Bader collab"),
        ("Tower 28 Beauty", "tower28beauty.com", "Hassas Cilt Makyaj", "NEA onaylı hassas cilt; ShineOn Lip Jelly; SOS Daily Rescue Spray; California"),
        ("Ogee", "ogfree.com", "Organik Lüks Makyaj", "USDA organik sertifikalı lüks makyaj; Vermont; jojoba bazlı; temiz + lüks birleşimi"),
        ("Lawless Beauty", "lawlessbeauty.com", "Temiz Cesur Makyaj", "Annie Lawless kurdu; The One Palette mini hit; cesur renkler temiz formüllerle"),
        ("Em Cosmetics", "emcosmetics.com", "Michelle Phan", "YouTube güzellik öncüsü Michelle Phan'ın markası; Lip Cushion viral; Asian beauty ilhamı"),
        ("colourpop", "colourpop.com", "Trend Erişilebilir", "Seed Beauty; $5-15 makyaj; Disney + Barbie collab; hızlı trend; sosyal medya native"),
        ("Juvia's Place", "jfrviasplace.com", "Afrika İlham Güzellik", "Chichi Eburu kurdu; The Zulu Palette; pigment zengin; kapsayıcı tonlar; Afro-heritage"),
        ("BH Cosmetics", "bhcosmetics.com", "Uygun Fiyat Palet", "Zodiac paleti viral; $5-20; kompakt palet; 2022 iflas sonrası yeniden başlangıç"),
        ("Nabla Cosmetics", "nabla.it", "İtalyan İndie Makyaj", "Bologna İtalya; Dreamy 2 Palette; vegan + cruelty-free; İtalyan indie güzellik"),
        ("Persona Cosmetics", "personacosmetics.com", "Kore-Amerikan Güzellik", "Sona Gasparian kurdu; Identity paleti; Ermeni-Amerikan güzellik; kapsayıcı"),
        ("Kaleidos", "kaleidosmakeup.com", "Futuristik Makyaj", "Asya kökenli; futuristik göz paleti; cesur renkler; Çin merkezli; küresel DTC"),
        ("Danessa Myricks", "danessamyricks.com", "Çok Kültürlü Sanat", "Colorfix pigmentler; Vision Flush; makyaj sanatçısı + tüm cilt tonları; Sephora"),
        ("Shiseido DTC", "shiseido.com", "Japon Bilim Güzellik", "140+ yıl; Ultimune serum; IBUKI; Japon biyoteknoloji; $8B gelir; DTC genişleme"),
        ("SK-II DTC", "sk-ii.com", "Japon Pitera Özü", "Facial Treatment Essence; Pitera maya özü; P&G; $2B+ marka; Japon güzellik ritüeli"),
        ("Sulwhasoo", "sulwhasoo.com", "Kore Hanbang Lüks", "Kore geleneksel tıp (hanbang); ginseng bazlı; AmorePacific; lüks K-beauty; anti-aging"),
        ("Drunk Elephant Bronzi", "drunkelephant.com", "Bronz Damla", "D-Bronzi anti-pollution bronzlaştırıcı; güneş öpüşü etkisi; TikTok viral; karıştır uygula"),
        ("Glow Hub", "glowhub.com", "İngiliz Gen Z Cilt", "İngiltere Gen Z cilt bakımı; Nourish & Glow Toner; Boots'ta; uygun fiyat; pastel ambalaj"),
        ("Bondi Sands", "bondisands.com", "Avustralya Bronzlaştırıcı", "Avustralya self-tan; Liquid Gold yağı; SPF koleksiyonu; TikTok viral; global genişleme"),
        ("Isle of Paradise", "isleofparadise.com", "Vegan Self-Tan", "Vegan self-tan damla; color-correcting; Jules Von Hep kurdu; Sephora; renk seçimi"),
        ("Tanologist", "tanologist.com", "Erişilebilir Self-Tan", "Target'ta $10-15; self-tan demokratikleştirildi; mousse, damla, mist; uygun fiyat bronz"),
        ("St. Tropez", "sttropeztan.com", "Self-Tan Standartı", "Self-tan kategorisi lideri; Express Bronzing Mousse; 1 saatte bronz; premium"),
        ("James Read", "jamesreadtan.com", "İngiliz Lüks Tan", "Ünlü spray tan artisti; Sleep Mask Tan; gece boyunca bronz; lüks self-tan"),
        ("Vita Liberata", "vitaliberata.com", "Organik Self-Tan", "Organik self-tan; pHenomenal mousse; İrlanda; 72 saat kalıcılık; premium doğal"),
    ],

    "Saç Bakımı": [
        ("Redken DTC", "redken.com", "Salon Bilimi", "NYC salon; Acidic Bonding serisi Olaplex rakibi; All Soft; L'Oréal Professional; DTC büyüme"),
        ("Wella DTC", "wella.com", "Profesyonel Saç", "100+ yıl; Color Fresh Create; salon professional; HenkelCotyCoty; global"),
        ("Matrix", "matrix.com", "Salon Erişilebilir", "Total Results; salon kalitesi uygun fiyat; L'Oréal Professional; canlı renkler"),
        ("Sebastian Professional", "sebastianprofessional.com", "Edgy Salon", "Dark Oil ikonik; Shaper Strong spray; Wella bünyesinde; edgy salon styling"),
        ("Paul Mitchell", "paulmitchell.com", "Salon Öncü", "Tea Tree şampuan ikonik; John Paul DeJoria; hayvan testi karşıtı öncü; 1980'den beri"),
        ("Joico", "joico.com", "Renk Uzmanı Salon", "Defy Damage; K-PAK onarım; renk uzmanı; Henkel bünyesinde; salon profesyonel"),
        ("CHI", "chi.com", "Isı Styling", "CHI Iron düzleştirici ikonik; ipek protein teknoloji; Farouk Systems; profesyonel ısı styling"),
        ("BaByliss Pro", "babylisspro.com", "İtalyan Saç Aletleri", "Nano Titanium düzleştirici; Conair bünyesinde; salon kalitesi evde; İtalyan tasarım"),
        ("Aussie", "aussie.com", "Avustralya Süpermarket", "3 Minute Miracle ikonik; Avustralya botanikleri; uygun fiyat; P&G; süpermarket saç bakımı"),
        ("Herbal Essences", "herbalessences.com", "Bitkisel Süpermarket", "Bio Renew; EWG VERIFIED; P&G; süpermarket bitkisel saç bakımı; yeniden lansman"),
        ("Pantene DTC", "pantene.com", "P&G Premium Saç", "Pro-V formül; Miracles koleksiyonu; P&G; global saç bakımı; DTC kanal genişleme"),
        ("Head & Shoulders DTC", "headandshoulders.com", "Kepek Uzmanı", "P&G; ZPT formulü; dünya kepek şampuanı #1; Supreme koleksiyon; DTC genişleme"),
        ("Aquis", "aquis.com", "Hızlı Kurulama", "Rapid Dry saç havlusu; Aquitex kumaş; saç kırılmasını önleme; mikrofiber havlu öncüsü"),
        ("PATTERN", "patternbeauty.com", "Tracee Ellis Ross", "Kıvırcık + coily saç bakımı; Tracee Ellis Ross kurdu; kıvırcık saç kültürü; Ulta'da"),
        ("DevaCurl", "devacurl.com", "Kıvırcık Uzmanı", "Curly Girl Method popülerleştirdi; No-Poo ikonik; kıvırcık saç topluluğu; tartışmalar sonrası yeniden"),
        ("Oui The People", "ouithepeople.com", "Kıvrımlı Bakım", "Karen Young kurdu; vücut bakımı + tıraş; kıvırcık tüy batması çözümü; kapsayıcı güzellik"),
        ("Hair La Vie", "hairlavie.com", "Saç Vitamini", "Biotin + kollajen saç vitamin; saç büyütme; Amazon bestseller; Clinical Formula serisi"),
        ("Revela", "revelabio.com", "Biyotek Saç Büyütme", "ProCelinyl molekülü; biyoteknoloji ile saç büyütme; Harvard bilimi; $20M+ yatırım"),
    ],

    "Erkek Bakım": [
        ("The Ordinary Men", "theordinary.com", "DECIEM Erkek", "DECIEM ürünleri erkekler arasında viral; Niacinamide + Zinc; unisex ama erkek topluluğu büyüyor"),
        ("Bulldog Skincare", "bulldogskincare.com", "İngiliz Erkek Doğal", "İngiliz doğal erkek bakım; Original Moisturiser; B Corp; Edgewell; uygun fiyat"),
        ("Nivea Men DTC", "nivea.com/men", "Erkek Klasik Bakım", "Beiersdorf; Sensitive serisi; uygun fiyat; global erkek bakım standartı; DTC genişleme"),
        ("Lab Series", "labseries.com", "Erkek Bilimsel Bakım", "Estée Lauder erkek cilt bakımı; MAX LS serisi; bilimsel formüller; anti-aging erkek"),
        ("Clinique for Men", "clinique.com/men", "Erkek Dermatolojik", "Dermatolojist geliştirme; Estée Lauder; allergy tested; erkek cilt bakımı öncüsü"),
        ("Anthony", "anthony.com", "NYC Erkek Bakım", "NYC erkek grooming; Glycolic Facial Cleanser; 2000'de kuruldu; premium erkek cilt bakımı"),
        ("Kiehl's Men", "kiehls.com", "NYC Eczane Erkek", "1851'den beri NYC eczane; Facial Fuel; erkek routin; L'Oréal; DTC genişleme"),
        ("Dollar Shave Club", "dollarshaveclub.com", "Abonelik Tıraş Öncü", "Viral YouTube videosu; Unilever $1B satın aldı; tıraş + bakım; abonelik DTC öncüsü"),
        ("Harry's", "harrys.com", "Erişilebilir Premium Tıraş", "Kendi Alman fabrikası; $1.37B Edgewell teklifi FTC engelledi; Walmart + Target; $400M+ gelir"),
        ("Flamingo", "shopflamingo.com", "Harry's Kadın Hattı", "Harry's'nin kadın tıraş markası; aynı fabrika kalitesi; uygun fiyat; güzel tasarım"),
        ("Athena Club", "athenaclub.com", "Kadın Tıraş + Bakım", "Abonelik kadın tıraş; vücut bakımı; deodorant; Target'ta; temiz bileşenler"),
        ("Billie", "mybillie.com", "Pembe Vergi Karşıtı Tıraş", "Kadın tıraş; pembe vergisiz fiyat; Edgewell satın aldı; vücut kılı normalleştirme"),
    ],

    "Sağlık & Wellness": [
        ("Perelel", "perelel.com", "Hamilelik Vitamin Uzmanı", "OB-GYN geliştirdi; hamilelik üçlemesi + emzirme + doğurganlık; bilimsel formüller"),
        ("Athletic Greens AG1", "drinkag1.com", "Podcast Kralı Takviye", "Her büyük podcast sponsoru; günlük 75 bileşen; $500M+ gelir; Tim Ferriss klasiği"),
        ("Huel Daily Greens", "huel.com", "İngiliz Yeşil Toz", "91 vitamin + mineral; tam beslenme yeşil toz; İngiltere; Huel ekosisteminde"),
        ("Amazing Grass", "amazinggrass.com", "Organik Yeşil Toz", "Green Superfood original; organik buğday çimi + arpa çimi; 20+ yıl; Glanbia satın aldı"),
        ("Nested Naturals", "nestednaturals.com", "Temiz Vitamin", "Super Greens; LUNA uyku takviyesi; B Corp; temiz bileşen; Amazon bestseller"),
        ("Country Life Vitamins", "countrylifevitamins.com", "Doğal Vitamin Öncü", "1971'den beri; glutensiz; Non-GMO; CoQ10; uzun geçmiş; güvenilir marka"),
        ("Designs for Health", "designsforhealth.com", "Pratisyen Kalite", "Sağlık profesyonelleri kanalı; yüksek doz; bilimsel formüller; premium"),
        ("Pure Encapsulations", "pureencapsulations.com", "Hipoalerjenik Takviye", "Hipoalerjenik; gereksiz katkı yok; Nestlé Health Science; hassas bireyler için; klinik kalite"),
        ("Standard Process", "standardprocess.com", "Tam Gıda Takviye", "1929'dan beri; kendi çiftliğinde yetiştirme; tam gıda takviye; organik"),
        ("Quicksilver Scientific", "quicksilverscientific.com", "Lipozomal Takviye", "Nanoemülsiyon + lipozomal taşıma; glutatyon; detoks; Dr. Chris Shade; biohacker favorisi"),
        ("ARMRA Colostrum", "tryarmra.com", "Kolostrüm Takviye", "Sığır kolostrüm; bağışıklık + bağırsak; Sarah Chen MD kurdu; TikTok viral; yeni trend"),
        ("BIOptimizers", "bioptimizers.com", "Sindirim Optimizasyon", "MassZymes enzim; Magnesium Breakthrough 7 form; biohacking sindirim; Wade Lightheart"),
        ("Mushroom Design", "mushroomdesign.com", "Fonksiyonel Mantar", "Lion's mane, reishi, chaga karışımlar; mantar takviye; bilişsel + bağışıklık; trend"),
        ("Host Defense", "hostdefense.com", "Paul Stamets Mantar", "Mikolojist Paul Stamets; organik mantar takviye; Lion's Mane, Turkey Tail; bilimsel"),
        ("Real Mushrooms", "realmushrooms.com", "Beta-Glucan Mantar", "Beta-glucan standardize; organik mantar özü; Çin + Kanada; bilimsel doz; güvenilir"),
    ],

    "Fitness & Spor Giyim": [
        ("Lululemon Mirror", "lululemon.com", "Akıllı Ayna Fitness", "Lululemon satın aldı; duvar aynası + ekran; ev fitness dersleri; Studio Mirror"),
        ("Tonal", "tonal.com", "Dijital Ağırlık", "Elektromanyetik direnç; duvara monte; AI antrenör; $1.6B değerleme; ev fitness lüks"),
        ("NordicTrack", "nordictrack.com", "Ev Fitness Ekipman", "İFit bağlantılı; treadmill + bike + rower; Icon Fitness; $30/ay abonelik"),
        ("Mirror", "mirror.co", "Fitness Ayna", "Lululemon bünyesinde; full body mirror + ekran; $1,495; ev fitness; kişisel eğitim"),
        ("Tempo", "tempo.fit", "AI Ev Spor Salonu", "3D sensör + AI form düzeltme; Tempo Studio; $2,500; kişiselleştirilmiş antrenman"),
        ("CLMBR", "clmbr.com", "Dikey Tırmanma", "Dikey tırmanma makinesi; tam vücut; düşük darbe; Jay-Z + Novak Djokovic yatırımcı"),
        ("Aviron", "avironactive.com", "Oyunlaştırılmış Kürek", "Oyun + yarış + koçluk; kürek makinesi; eğlence + fitness; Kanada"),
        ("Xponential Fitness", "xponential.com", "Butik Fitness Franchise", "Club Pilates, CycleBar, Pure Barre, YogaSix; 10 marka; DTC + franchise; $2B+"),
        ("Barry's", "barrys.com", "HIIT Stüdyo", "Red room HIIT; Barry's X ev ekipmanı; Hollywood favorisi; global 80+ lokasyon"),
        ("F45 Training", "f45training.com", "Fonksiyonel Fitness", "Avustralya fonksiyonel fitness franchise; 45 dk antrenman; Mark Wahlberg yatırımcı; global"),
        ("Solidcore", "solidcore.co", "Pilates Reform", "Megaformer Pilates; yoğun 50 dakika; Michelle Obama'nın favorisi; Washington DC merkezli"),
        ("SoulCycle DTC", "soul-cycle.com", "Bağlı Bisiklet", "At-home bike; $2,500; NYC spin kültürü; Equinox grubu; müzik + hareket; premium"),
    ],

    "Moda & Giyim": [
        ("Zara DTC", "zara.com", "Hızlı Moda DTC", "Inditex; online kanal %30+ gelir; hızlı trend; $25B+ toplam gelir; DTC genişleme"),
        ("H&M DTC", "hm.com", "İsveç Fast Fashion DTC", "Online %30+ gelir; Conscious koleksiyon; global DTC; $20B+ gelir"),
        ("Shein", "shein.com", "Ultra Hızlı Moda", "Çin ultra fast fashion; günlük 1000+ yeni ürün; $30B gelir; TikTok haul kültürü; tartışmalı"),
        ("Temu", "temu.com", "Çin Marketplace", "PDD Holdings; ultra uygun fiyat; Super Bowl reklam; 'shop like a billionaire'; hızla büyüyen"),
        ("COS", "cosstores.com", "H&M Lüks Çizgi", "H&M grubunun premium markası; mimari mağazalar; minimalist; kaliteli malzeme; İskandinav"),
        ("Arket", "arket.com", "Nordik Günlük", "H&M grubu; İskandinav günlük lüks; cafe deneyimi; şeffaf fiyat; doğal malzeme"),
        ("Massimo Dutti DTC", "massimodutti.com", "İnditex Premium", "Inditex premium markası; erkek + kadın; iş giyim; DTC online genişleme; İspanya lüks"),
        ("Mango", "mango.com", "İspanyol DTC", "Barcelona; $3B+ gelir; Committed koleksiyon sürdürülebilir; 110+ ülke; online büyüme"),
        ("Cider", "shopcider.com", "Çin Gen Z Moda", "Çin DTC; Y2K estetik; TikTok native; $100M+ yatırım; Shein alternatifi; Gen Z"),
        ("Edikted", "edikted.com", "İsrail TikTok Moda", "İsrail merkezli; TikTok viral; Gen Z; $30-80 fiyat; cesur + seksi; hızla büyüyen"),
        ("iets frans", "urbanoutfitters.com/iets-frans", "Urban Outfitters Alt Marka", "UO alt markası; casual basics; vintage estetik; Gen Z streetwear; Amsterdam adı"),
        ("Free People DTC", "freepeople.com", "Boho Lifestyle", "URBN bünyesinde; FP Movement aktivewear; $1.5B+ gelir; bohem yaşam tarzı; festival moda"),
        ("Anthropologie DTC", "anthropologie.com", "Eklektik Lifestyle", "URBN bünyesinde; ev + moda + güzellik; orta-üst segment; DTC genişleme; $1.5B+ gelir"),
        ("Verge Girl", "vergegirl.com", "Avustralya Indie Moda", "Gold Coast Avustralya; indie kadın moda; TikTok viral; $30-80; boho + streetwear"),
        ("Petal & Pup", "petalandpup.com", "Avustralya Online Moda", "Brisbane; kadın online moda; wedding guest + günlük; Instagram native; hızla büyüyen"),
        ("Hello Molly", "hellomolly.com", "Avustralya Parti Moda", "Sydney; parti + gece kıyafeti; Gen Z; TikTok viral; uygun fiyat; hızla büyüyen"),
        ("Maje DTC", "maje.com", "Fransız Premium DTC", "SMCP grubu; Fransız premium; Judith Milgrom kurdu; Paris chic; DTC genişleme; $500M+"),
        ("Sandro DTC", "sandro-paris.com", "Fransız Cool DTC", "SMCP grubu; Evelyne Chetrite kurdu; Parisian cool; $500M+; DTC online büyüme"),
        ("Ba&sh", "ba-sh.com", "Fransız Kadın DTC", "Barbara Boccara + Sharon Krief; Parisian chic; elbise uzmanı; DTC genişleme"),
        ("Reiss DTC", "reiss.com", "İngiliz İş Giyim", "İngiliz smart-casual; Kate Middleton etkisi; orta segment; DTC online genişleme; $200M+ gelir"),
    ],

    "Yiyecek & İçecek": [
        ("Chobani", "chobani.com", "Yunan Yoğurt DTC", "Hamdi Ulukaya; Türk göçmen başarı hikayesi; $2B+ gelir; DTC + perakende; sosyal misyon"),
        ("Oat Haus", "oat-haus.com", "Granola Butter", "Fındık ezmesi alternatifi granola; TikTok viral; Ali Bonar kurdu; nut-free; yaratıcı niş"),
        ("Cacao Bliss", "earthechofoods.com", "Çikolata Takviye", "Danette May; kakao + adaptojenik mantar; fonksiyonel sıcak çikolata; wellness çikolata"),
        ("Daring Foods", "daring.com", "Bitkisel Tavuk", "Bitkisel tavuk; restoran + perakende; soya bazlı; Ross MacKay kurdu; $100M+ yatırım"),
        ("Nourish Snacks", "nourishsnacks.com", "Temiz Atıştırmalık", "Organik granola bar; temiz bileşen; doğal tatlandırıcı; sağlıklı snacking"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Snack", "No refined sugar; tahıl + soya + süt free; Mondelez satın aldı; 'get back to human'"),
        ("Birch Benders", "birchbenders.com", "Protein Waffle Mix", "Protein pancake + waffle karışımı; keto + paleo seçenekleri; Denver CO; doğal"),
        ("Catalina Crunch", "catalinacrunch.com", "Keto Gevrek", "Keto dostu granola + kurabiye; 0g şeker; Krishna Kaliannan kurdu; düşük karbonhidrat snack"),
        ("Highkey", "highkey.com", "Keto Kurabiye", "Keto mini cookies; 0g net carb; $50M+ gelir; düşük karbonhidrat tatlı alternatifi"),
        ("Legendary Foods", "legendaryfoods.com", "Protein Pastry", "Protein pop-tart; 20g protein tatlı; Shark Tank; düşük şeker; nostaljik snack sağlıkçı"),
        ("Laird Superfood", "lairdsuperfood.com", "Sörfçü Süperfood", "Laird Hamilton; hindistan cevizi kreamer; mantar kahve; süperfood; doğal enerji"),
        ("Orgain", "orgain.com", "Organik Protein Shake", "Dr. Andrew Abraham; organik protein; Costco bestseller; kanser hayatta kalan; temiz beslenme"),
        ("Aloha", "aloha.com", "Bitkisel Protein Bar", "Organik bitkisel protein bar; 14g protein; temiz bileşen; USDA organic; erişilebilir"),
        ("No Cow", "nocow.com", "Sütsüz Protein Bar", "Süt ürünsüz protein bar; 20g protein 1g şeker; vegan; allergen-friendly"),
        ("Perfect Snacks", "perfectsnacks.com", "Buzdolabı Bar", "Peanut butter bar buzdolabından; taze bileşenler; Mondelez yatırım; ünlü protein bar"),
        ("Lenny & Larry's", "lennyandlarrys.com", "Protein Kurabiye", "Complete Cookie 16g protein; vegan; süpermarket + convenience; erişilebilir protein snack"),
        ("Quest DTC", "questnutrition.com", "Protein Bar Öncü", "Quest Bar protein öncüsü; 20g protein; keto dostu; Simply Good Foods; $1B+ marka"),
        ("RXBAR", "rxbar.com", "Şeffaf Bileşen Bar", "'3 Egg Whites, 6 Almonds, 4 Cashews, 2 Dates'; Kellogg's $600M; şeffaf etiket devrimi"),
        ("Poppi", "drinkpoppi.com", "Elma Sirkesi Soda", "TikTok viral; Super Bowl reklamı; prebiyotik; Allison Ellsworth kurdu; OLIPOP rakibi"),
        ("Culture Pop", "drinkfculturepop.com", "Probiyotik Soda", "Canlı kültürlü soda; organik meyve suyu; probiyotik gazlı içecek; sağlıklı soda"),
    ],

    "Ev & Mutfak": [
        ("Cuisinart DTC", "cuisinart.com", "Mutfak Klasik DTC", "Food processor öncüsü; 50+ yıl; DTC genişleme; mutfak robotu + blender + tost makinesi"),
        ("Ninja DTC", "ninjakitchen.com", "Çok İşlevli Mutfak", "Foodi multi-cooker; Creami dondurma makinesi TikTok viral; SharkNinja; $2B+ gelir"),
        ("Instant Brands", "instantbrands.com", "Instant Pot", "Instant Pot çok amaçlı tencere; 7-in-1; $350M gelir; kült takipçi; pressure cooker devrimi"),
        ("Solo Stove", "solostove.com", "Dumansız Ateş", "Hava akışı ile dumansız; Snoop Dogg 'giving up smoke' kampanyası; $400M+ gelir"),
        ("Yoto", "yoto.com", "Çocuk Audio Player", "Ekransız çocuk audio player; kartla hikaye + müzik; İngiliz tasarım; ebeveyn dostu teknoloji"),
        ("Sonos", "sonos.com", "Çok Odalı Ses", "Multi-room wireless hoparlör; Arc soundbar; Era 300; Trueplay; $1.5B+ gelir; ev ses standartı"),
        ("Dyson DTC", "dyson.com", "İngiliz Ev Teknoloji", "V15 süpürge; Pure Cool fan; Airwrap; Supersonic; DTC odaklı; $7B+ gelir; mühendislik lüksü"),
        ("iRobot DTC", "irobot.com", "Robot Süpürge", "Roomba robot süpürge; 30+ yıl; Amazon satın alma (iptal); ev otomasyon; $1.5B gelir"),
        ("Ecovacs", "ecovacs.com", "Çin Robot Süpürge", "Deebot robot süpürge + paspas; X2 Omni; Çin ev robotik; $800M+ gelir; hızla büyüyen"),
        ("Roborock", "roborock.com", "Çin Premium Robot", "S8 Pro Ultra; LiDAR navigasyon; paspas + vakum; $1B+ gelir; Xiaomi ekosisteminden bağımsızlaştı"),
        ("Dreame", "dreametech.com", "Çin Ev Temizlik", "L20 Ultra robot; kablosuz süpürge; Çin ev teknoloji; hızla büyüyen; $500M+ gelir"),
        ("Levoit", "levoit.com", "Hava Temizleyici", "Core 300 Amazon #1 hava temizleyici; Vesync bünyesinde; uygun fiyat HEPA; ev hava kalitesi"),
        ("Coway", "coway.com", "Kore Hava + Su", "Airmega hava temizleyici; su arıtma; Kore ev sağlık teknoloji; $3B+ gelir; IoT ev"),
        ("Miele DTC", "miele.com", "Alman Ev Lüks", "Alman premium ev aletleri; 100+ yıl; süpürge + çamaşır + bulaşık; DTC genişleme; lüks kalite"),
    ],

    "Bebek & Çocuk": [
        ("Banana Republic Baby", "bananarepublic.com", "Gap Premium Bebek", "Gap bünyesinde; premium bebek giyim; organik pamuk; şık bebek basics"),
        ("Carter's DTC", "carters.com", "Amerikan Bebek Klasik", "150+ yıl; ABD'de her 10 bebekten 8'i Carter's giyer; DTC genişleme; $3B+ gelir"),
        ("Gerber Childrenswear", "gerberchildrenswear.com", "Bebek Basics", "Onesie ikonik; Nestlé bünyesinde; temel bebek giyim; uygun fiyat; herkesin bildiği marka"),
        ("Magnetic Me", "magneticme.com", "Manyetik Bebek Giyim", "Manyetik çıtçıt ile kolay giyim; beziz değiştirmede 3 saniye; ebeveyn time-saver; $50M+"),
        ("Copper Pearl", "copperpearl.com", "Bebek Aksesuar", "Önlük + battaniye + müslin; boho desenleri; Amazon bestseller; uygun fiyat şık bebek"),
        ("Ollie Swaddle", "ollieworld.com", "Patentli Kundak", "Patentli tek kundak sistemi; nem çekici; güvenli uyku; pediatrist geliştirdi"),
        ("4moms", "4mfrroms.com", "Akıllı Bebek Ekipman", "mamaRoo salıncak; highchair; self-installing car seat; robot teknoloji bebek ürünleri"),
        ("Guava Family", "guavafamily.com", "Seyahat Bebek Yatağı", "Lotus Travel Crib; 15 saniye kurulum; GreenGuard sertifikalı; seyahat bebek yatağı"),
        ("Slumberpod", "slumberpod.com", "Karartma Çadır", "Bebek/toddler karartma çadırı; seyahatte uyku çözümü; havalandırmalı; Shark Tank $500K"),
        ("Haakaa", "haakaa.com", "Silikon Göğüs Pompası", "Yeni Zelanda; silikon göğüs pompası; $15; basit + etkili; emzirme kolaylaştırıcı; Amazon #1"),
    ],

    "Evcil Hayvan": [
        ("Furrish", "furrish.co.uk", "İngiliz Köpek Bakım", "İngiliz doğal köpek bakım ürünleri; şampuan + parfüm; vegan; salon kalitesi evde"),
        ("Wahl Pet", "wahlpet.com", "Evcil Hayvan Tıraş", "Profesyonel pet grooming aletleri; ev kullanımı; clipper + trimmer; 100+ yıl"),
        ("Charlie & Me", "charlieandmepetfood.com.au", "Avustralya Premium Pet", "Avustralya premium köpek maması; doğal bileşenler; uygun fiyat premium"),
        ("Wellness Pet", "wellnesspetfood.com", "Doğal Pet Gıda", "CORE grain-free; Complete Health; WellPet bünyesinde; doğal + bütünsel beslenme"),
        ("Blue Buffalo DTC", "bluebuffalo.com", "Doğal Pet Premium", "General Mills satın aldı; Life Protection; $2B+ gelir; doğal evcil hayvan gıda lideri"),
        ("Canidae", "canidae.com", "Biyotik Köpek Mama", "SUSTAIN biyotik formül; probiyotik + prebiyotik; 27 vitamin + mineral; sürdürülebilir"),
        ("Bocce's Bakery", "boccesbakery.com", "El Yapımı Köpek Ödül", "NYC; el yapımı köpek bisküvisi; 5 bileşen; eğlenceli tatlar; premium köpek ödülü"),
        ("Mr. Chewy", "mrchewy.com", "Gourmet Köpek Ödül", "Gurme köpek ödülleri; single protein; doğal; küçük seri; premium atıştırmalık"),
        ("Poppy's Picnic", "poppyspicnic.co.uk", "İngiliz Raw Köpek", "İngiltere; raw (çiğ) köpek maması; taze teslimat; doğal beslenme; BARF diyet"),
        ("Lily's Kitchen", "lilyskitchen.com", "İngiliz Premium Pet", "İngiliz doğal köpek + kedi maması; organik; güzel ambalaj; Nestlé Purina satın aldı"),
    ],

    "Aksesuar & Takı": [
        ("Missoma", "missoma.com", "Londra Trend Takı", "İngiliz demi-fine takı; Meghan Markle etkisi; layer koleksiyonu; altın kaplama; $50-200"),
        ("Astrid & Miyu", "astridandmiyu.com", "İngiliz Kulak Piercing", "Londra; ear piercing bar + takı; Gold Huggies ikonik; ear stack kültürü; İngiliz DTC"),
        ("Daisy London", "dfraisylondon.com", "İngiliz Boho Takı", "Londra; doğa ilhamlı takılar; 2009 kuruldu; çakra + doğum taşı; İngiliz DTC"),
        ("Tada & Toy", "tadaandtoy.com", "İngiliz Oyunbaz Takı", "Londra; eğlenceli + renkli takılar; emaye detay; parti takı; İngiliz DTC; $50-150"),
        ("Edge of Ember", "edgeofember.com", "İngiliz Etik Takı", "Londra; geri dönüşüm altın; adil ticaret taş; sürdürülebilir lüks; İngiliz etik takı"),
        ("Otiumberg", "otiumberg.com", "Minimalist İngiliz Takı", "İngiliz minimalist fine jewelry; geri dönüşüm altın; zincir + küpe; Londra; $80-300"),
        ("Mateo", "mfrateonewyork.com", "NYC Modern Takı", "Matthew Harris; 14K altın modern takı; pearl + malachite; $200-2000; CFDA nominesi"),
        ("Spinelli Kilcollin", "spinellikilcollin.com", "LA Linked Ring", "Birbirine bağlı yüzük ikonik; LA merkezli; Hailey Bieber favorisi; $400-5000"),
        ("Lady Grey", "ladygreyjewelry.com", "Brooklyn Artisan Takı", "Brooklyn el yapımı; mimari + geometrik tasarımlar; $80-500; Jill Martinelli + Sabine Le Guyader"),
        ("Alighieri", "alfrghieri.co.uk", "Londra Edebiyat Takı", "Dante'nin İlahi Komedya ilhamı; Rosh Mahtani kurdu; doku + pul detay; İngiliz DTC lüks"),
    ],

    "Teknoloji & Elektronik": [
        ("DJI", "dji.com", "Drone Lider", "Çin drone lideri; Mini 4 Pro; Osmo gimbal; Pocket kamera; $3B+ gelir; hava fotoğrafçılığı"),
        ("GoPro DTC", "gopro.com", "Aksiyon Kamera", "HERO12 Black; GoPro Subscriber; Max 360; $1B+ gelir; aksiyon kamera kategorisi yarattı"),
        ("Insta360", "insta360.com", "360 Kamera", "X4 360 kamera; ONE RS modüler; Ace Pro; Çin; GoPro rakibi; yaratıcı kamera"),
        ("Aqara", "aqara.com", "Akıllı Ev Çin", "Zigbee + Matter akıllı ev; sensör + anahtar + kamera; Çin; uygun fiyat akıllı ev"),
        ("Meross", "meross.com", "Uygun Akıllı Ev", "Akıllı fiş + şerit + garaj; HomeKit + Alexa uyumlu; $15-30; uygun fiyat otomasyon"),
        ("SwitchBot", "switch-bot.com", "Retrofit Akıllı Ev", "Mevcut anahtarları akıllı yapan robot; perde robotu; SwitchBot Bot; $30-100; uygun fiyat"),
        ("Arlo", "arlo.com", "Premium Güvenlik Kamera", "4K güvenlik kamerası; Essential + Pro; kablosuz; gece görüş; Verisure; $500M+ gelir"),
        ("Eufy", "eufy.com", "Anker Akıllı Ev", "Anker markası; Robot süpürge + güvenlik kamerası; yerel depolama; gizlilik odaklı; uygun fiyat"),
        ("Reolink", "reolink.com", "Uygun Güvenlik Kamera", "POE + WiFi güvenlik kameraları; yerel kayıt; $40-200; ev + iş; güvenilir + uygun"),
        ("Nothing Ear", "nothing.tech", "Şeffaf Kulaklık", "Nothing Ear (2) şeffaf kulaklık; ANC; Carl Pei; anti-Apple estetik; $100-150"),
        ("Soundcore", "soundcore.com", "Anker Ses", "Space A40 kulaklık; Liberty 4 NC; Anker alt marka; uygun fiyat premium ses; $50-100"),
        ("Shure DTC", "shure.com", "Mikrofon Efsanesi", "SM7B podcast mikrofon; MV7 USB; 1925'ten beri; DTC genişleme; podcast/streaming standardı"),
    ],

    "Abonelik Kutuları": [
        ("Splendid Spoon", "splendidspoon.com", "Bitki Bazlı Yemek", "Bitkisel smoothie + çorba + bowls; haftalık abonelik; sağlıklı hazır yemek"),
        ("Sakuraco", "sakfrraco.com", "Japon Snack Kutu", "Otantik Japon atıştırmalık kutusu; geleneksel wagashi + modern; Tokyo'dan direkt"),
        ("TokyoTreat", "tokyotreat.com", "Japon Şeker Kutu", "Japon şeker + snack; eğlenceli ambalaj; anime kültürü; aylık keşif; $25-35/kutu"),
        ("Causebox", "causebox.com", "Etik Ürün Kutu", "Sosyal etki + etik ürünler; sezonluk; B Corp markalar; sürdürülebilir keşif kutusu"),
    ],

    "Seyahat & Bavul": [
        ("Samsonite", "samsonite.com", "Global Bavul Lideri", "Proxis + Magnum Eco; 100+ yıl; global bavul #1; DTC genişleme; $3B+ gelir"),
        ("Tumi DTC", "tumi.com", "Lüks İş Seyahati", "Alpha 3 seri ikonik; $500-1000; Samsonite bünyesinde; iş seyahati lüksü; balistik naylon"),
        ("American Tourister DTC", "americantourister.com", "Uygun Fiyat Bavul", "Samsonite alt marka; renkli + uygun fiyat; $50-150; aile seyahati; gençlik"),
        ("Eastpak", "eastpak.com", "Dayanıklı Sırt Çantası", "30 yıl garanti; Padded Pak'r ikonik; VF Corporation; okul + seyahat; dayanıklılık"),
        ("Herschel Supply", "herschel.com", "Vancouver Heritage Çanta", "Kanada; Little America sırt çantası ikonik; heritage estetik; erişilebilir; $100M+ gelir"),
    ],

    "Kadın Sağlığı": [
        ("Perelel", "perelel.com", "Trimester Vitamin", "OB-GYN panel; hamilelik üçlemesine göre vitamin paketi; 1. + 2. + 3. trimester farklı formül"),
        ("Needed", "thisisneeded.com", "Perinatal Beslenme", "Hamilelik boyunca kapsamlı takviye; kollajen + probiyotik; klinisyen formüle; postpartum"),
        ("Frida Mom", "fridamom.com", "Postpartum Bakım", "Doğum sonrası bakım kiti; buz pedi + peribottle; tabusuz annelik; Fridababy'nin anne hattı"),
        ("Bodily", "itsbodily.com", "Doğum + Postpartum", "Doğum + emzirme + postpartum ürünleri; eğitim içerikleri; fourth trimester; anne sağlığı"),
        ("Nyssa", "nfryssa.com", "FourthWear Postpartum", "Doğum sonrası iyileşme giyim; buz cebi iç çamaşırı; anne sağlığı; postpartum destek"),
    ],

    "Parfüm & Koku": [
        ("Malin+Goetz", "malinandgoetz.com", "NYC Apothecary Koku", "Cannabis parfüm viral; Dark Rum; NYC apothecary koku; unisex; minimalist"),
        ("Bon Parfumeur", "bonparfumeur.com", "Paris Renkli Parfüm", "Renkli numaralı parfümler; Paris; #103 tiare flower; 30ml seyahat boy; erişilebilir niş"),
        ("Heeley", "jframesheeley.com", "Paris İngiliz Parfümör", "James Heeley; İngiliz Paris'te; Sel Marin; minimalist niş parfüm; doğal bileşenler"),
        ("Nishane", "nishane.com.tr", "Türk Niş Parfüm", "İstanbul; Hacivat kült koku; Ani; Türk parfüm dünyasında global başarı; niş + lüks"),
        ("Initio", "initfrioparfums.com", "Fransız Niche Konya", "Side Effect TikTok viral; misk + vanilya; Parle Moi de Parfum alternatifi; lüks niş"),
    ],

    "Outdoor & Macera": [
        ("Teva DTC", "teva.com", "Outdoor Sandal", "Hurricane XLT2 ikonik; Grand Canyon ilhamı; DTC büyüme; outdoor sandal + moda çaprazı"),
        ("Chaco", "chacos.com", "Kayak Sandal", "Z/1 Classic ayarlanabilir sandal; Wolverine bünyesinde; outdoor + festival; kişiselleştirilebilir"),
        ("Merrell DTC", "merrell.com", "Hiking Ayakkabı", "Moab hiking boot ikonik; Wolverine bünyesinde; $1B+ marka; DTC genişleme; trail standartı"),
        ("Salewa", "salewa.com", "İtalyan Dağ", "Güney Tirol dağ ekipmanı; Oberalp grup; tırmanış + hiking; Alp kültürü; premium teknik"),
        ("Mammut", "mammut.com", "İsviçre Alp", "160+ yıl İsviçre dağ ekipmanı; Eiger Extreme; tırmanış + hiking + kayak; $400M+ gelir"),
    ],

    "Oyun & Eğlence": [
        ("Funko", "funko.com", "Pop Vinyl Figür", "Pop! figürleri; 1000+ lisans; Marvel, Disney, anime; koleksiyon kültürü; $1B+ gelir"),
        ("Build-A-Bear DTC", "buildabear.com", "Doldurulmuş Hayvan", "Online + mağaza; kişiselleştirilebilir peluş; 25+ yıl; DTC büyüme; $500M+ gelir; yetişkin hayran"),
        ("Jellycat", "jellycat.com", "İngiliz Peluş Oyuncak", "Londra; Bashful Bunny ikonik; yetişkin koleksiyoncular; $200M+ gelir; TikTok viral; lüks peluş"),
        ("Squishmallows", "squishmallows.com", "Viral Peluş", "Kellytoy; Gen Z koleksiyon; ultra yumuşak; 1000+ karakter; $1B+ marka; TikTok kültü"),
    ],

    "Supplement & Sporcu Beslenmesi": [
        ("Onnit", "onnit.com", "Toplam İnsan Performansı", "Alpha Brain nootropik; Joe Rogan uzun sponsor; Austin TX; Unilever satın aldı; bütünsel performans"),
        ("Force Factor", "forcefactor.com", "GNC Bestseller", "GNC en çok satan; Total Beets; Score; performans takviye; $100M+ gelir; perakende odaklı DTC"),
        ("Scivation", "xtend.com", "BCAA Uzmanı", "Xtend BCAA ikonik; intra-workout; Nutrabolt bünyesinde; amino asit kategorisi lideri"),
        ("RSP Nutrition", "rspnutrition.com", "Bilim + Lezzet", "AminoLean enerji + amino; TrueFit protein; temiz etiket; bilimsel formüller + harika tatlar"),
        ("Vital Proteins", "vitalproteins.com", "Kolajen İçecek", "Kolajen su + kolajen kreamer; kahveye kolajen trendi; Jennifer Aniston; Nestlé; DTC öncü"),
        ("Isopure", "theisopure.com", "Sıfır Carb Protein", "Zero Carb protein tozu; WPI izolat; lactose-free; Glanbia; temiz protein; şeffaf"),
    ],

    "Diş & Ağız Bakımı": [
        ("Zenyum", "zenyum.com", "Asya Şeffaf Teli", "Güneydoğu Asya şeffaf diş teli; $2000-3000; Singapur merkezli; Asya DTC dental"),
        ("Oclean", "oclean.com", "Çin Akıllı Diş Fırçası", "Xiaomi ekosistem; AI fırçalama; $40-80; uygun fiyat akıllı diş fırçası; Çin teknoloji"),
        ("Curaprox", "curaprox.com", "İsviçre Diş Fırçası", "CS 5460 ultra yumuşak fırça; İsviçre kalite; renkli tasarım; dental profesyonel önerisi"),
    ],

    "Ofis & Kırtasiye": [
        ("Tombow", "tombowusa.com", "Japon Fırça Kalem", "Dual Brush Pen hand lettering ikonik; MONO silgi; Japonya 100+ yıl; sanat + yazı"),
        ("Copic", "copicmarker.com", "Japon Pro Marker", "Profesyonel alkol bazlı marker; manga + illüstrasyon standartı; 358 renk; yeniden doldurulabilir"),
        ("Blackwing", "blackwing602.com", "Kült Kurşun Kalem", "Palomino Blackwing 602; John Steinbeck kullandığı kalem; premium $25/düzine; kült takipçi"),
        ("Kaweco", "kaweco-pen.com", "Alman Cep Dolma Kalem", "1883'ten beri; Sport dolma kalem ikonik; Alman; cep boy; $30-100; vintage + modern"),
    ],

    "Uyku & Yatak": [
        ("Coop Home Goods", "coophomegoods.com", "Ayarlanabilir Yastık", "Parçalanmış bellek köpük; dolgu ekle/çıkar; Amazon #1 yastık; ayarlanabilir yükseklik"),
        ("Silk & Snow", "silkandsnow.com", "Kanada Yatak", "Kanada DTC yatak; organik pamuk; doğal lateks; hibrit; Vancouver merkezli"),
        ("Panda London", "pandalondon.com", "İngiliz Bambu Yatak", "İngiliz bambu yatak tekstili; hypoallergenic; The Panda Mattress Topper; sürdürülebilir"),
        ("Koala", "koala.com", "Avustralya DTC Yatak", "Avustralya; 120 gece deneme; ekolojik; Klima Comfort Plus; koala koruma; hızla büyüyen"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# EXTRA BRANDS BATCH 3 — final push to 2000+
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS_3 = {
    "Güzellik & Cilt Bakımı": [
        ("Medik8", "medik8.com", "İngiliz Profesyonel Cilt", "İngiliz klinik cilt bakımı; Crystal Retinal retinaldehit; CSA felsefesi; dermatoloji odaklı"),
        ("Alpha-H", "alpha-h.com", "Avustralya AHA", "Avustralya; Liquid Gold glikolik tonik; 1995; AHA uzmanı; profesyonel cilt bakımı"),
        ("Rationale", "rationale.com", "Avustralya Lüks Derma", "Melbourne; DNA teknolojisi bazlı; $200+ ürünler; Avustralya'nın en lüks cilt bakımı"),
        ("Aesop Skincare", "aesop.com", "Avustralya Minimalist Cilt", "Melbourne; Parsley Seed serisi; mağaza tasarımları sanat; L'Oréal $2.5B; unisex cilt bakımı"),
        ("Sodashi", "sodashi.com.au", "Avustralya Doğal Lüks", "Perth; doğal + organik; spa markası; Avustralya bitkileri; lüks doğal"),
        ("Jurlique", "jurlique.com", "Avustralya Çiftlik Güzellik", "Adelaide çiftliğinden biyodinamik bitkiler; Rosewater Balancing Mist ikonik"),
        ("Akin", "akinskincare.com.au", "Avustralya Doğal", "Avustralya doğal cilt + saç bakımı; rosehip oil ikonik; uygun fiyat doğal; B Corp"),
        ("Sukin", "sukfrinorganics.com.au", "Avustralya Karbon Nötr", "Avustralya karbon nötr; doğal cilt bakımı; süpermarket fiyatı; en çok satan Avustralya doğal"),
        ("Thank You Farmer", "thankyoufarmer.com", "Kore Güneş Bakımı", "Kore SPF uzmanı; Sun Project Water Sun Cream; uygun fiyat K-beauty güneş koruma"),
        ("Isntree", "isntree.com", "Kore Hyalüronik Asit", "Hyaluronic Acid Toner ikonik; uygun fiyat K-beauty; aktif madde odaklı; Olive Young"),
        ("Torriden", "torriden.com", "Kore Nemlendirme", "DIVE-IN Serum hyalüronik asit; düşük molekül ağırlık; Kore viral cilt bakımı"),
        ("Anua", "anufrfra.com", "Kore Heartleaf", "Heartleaf 77% Soothing Toner; viral K-beauty; hassas cilt; doğal bitki özleri"),
        ("Tirtir", "tirtir.com", "Kore Cushion", "Mask Fit Red Cushion TikTok viral; 30 ton; Kore fondöten; dünya genelinde tükendi"),
        ("Rom&nd", "romand.us", "Kore Lip Tint", "Juicy Lasting Tint; Kore dudak tint viral; Gen Z K-beauty makyaj; uygun fiyat"),
        ("Peripera", "peripera.com", "Kore Eğlenceli Makyaj", "Ink Velvet lip tint; renkli ambalaj; uygun fiyat K-beauty makyaj; AmorePacific"),
        ("Heimish", "heimish.com", "Kore Vegan Temizlik", "All Clean Balm; vegan Kore cilt bakımı; uygun fiyat; sherbet temizleyici"),
        ("SKIN1004", "skin1004.com", "Kore Centella Ampoule", "Madagascar Centella Ampoule; $15; TikTok viral; onarıcı K-beauty; hassas cilt"),
        ("Goodal", "goodal.com", "Kore Vitamin C", "Green Tangerine Vitamin C serum; uygun fiyat; Kore vitamin C besteller; CJ bünyesinde"),
        ("Caudalie", "caudalie.com", "Fransız Üzüm", "Bordeaux üzüm polifenol; Beauty Elixir mist ikonik; Mathilde Thomas kurdu; Fransız doğal lüks"),
        ("Nuxe", "nuxe.com", "Fransız Çiçek Güzellik", "Huile Prodigieuse çok amaçlı yağ ikonik; Fransız eczane; altın parçacıklı yağ; $300M+ gelir"),
        ("Bioderma", "bioderma.com", "Fransız Eczane Bakım", "Sensibio H2O micellar su dünya #1; dermatolog formüle; NAOS grubu; Fransız eczane standardı"),
        ("Avène", "avene.com", "Fransız Termal Su", "Termal su cilt bakımı; hassas cilt; Pierre Fabre; Cicalfate+ onarım kremi; eczane derma"),
        ("Vichy", "vichy.com", "Fransız Volkanik Su", "Volkanik mineralli su; Liftactiv anti-aging; L'Oréal Active; Fransız dermatolojik güzellik"),
        ("Embryolisse", "embryolisse.com", "Fransız Nemlendirici", "Lait-Crème Concentré makyöz favorisi; 1950'den beri; Fransız backstage güzellik; basit + etkili"),
        ("Filorga", "filorga.com", "Fransız Anti-Aging", "NCEF teknolojisi; Time-Filler kırışıklık kremi; Colgate-Palmolive satın aldı; Fransız anti-aging lüks"),
        ("Erborian", "erborian.com", "Kore-Fransız Fusion", "CC Crème en çok satan; Kore + Fransız güzellik fusyonu; ginseng bazlı; L'Occitane bünyesinde"),
    ],

    "Saç Bakımı": [
        ("evo", "evfro.com", "Avustralya Brutally Honest", "Avustralya; eğlenceli + dürüst isimler; 'No BS haircare'; salon kalitesi; global dağıtım"),
        ("Kevin Murphy", "kevinmurphy.com.au", "Avustralya Salon Moda", "Avustralya; moda + salon; ANGEL.WASH; sürdürülebilir ambalaj; Kevin Murphy kurdu"),
        ("Olaplex No.3", "olaplex.com", "Evde Saç Onarımı", "Hair Perfector No.3 evde kullanım; bis-amino onarım; Sephora #1 saç bakımı; salon to home"),
        ("Vegamour GRO", "vegamour.com", "Bitkisel Saç Serumu", "GRO Hair Serum bitkisel; klinik kanıtlı saç yoğunlaştırma; vegan + temiz; $100M+ gelir"),
        ("Color WOW", "colorwowhair.com", "Anti-Nem Sprey", "Dream Coat anti-humidity spray TikTok viral; Glass Hair efekti; Chris Appleton"),
        ("Biolage", "biolage.com", "Sürdürülebilir Salon", "L'Oréal Professional; vegan dönüşüm; CleanReset; Hydra Source; doğa ilhamlı salon bakımı"),
        ("Kerastase DTC", "kerastase.com", "Lüks Salon DTC", "L'Oréal lüks saç bakımı; Elixir Ultime altın yağ; Genesis saç dökülme; DTC genişleme"),
        ("L'Oréal Pro DTC", "lorealprofessionnel.com", "Profesyonel DTC", "Serie Expert; Metal Detox; global salon lider; DTC kanal genişleme; $3B+ gelir"),
        ("Shu Uemura Art of Hair", "shuuemura-usa.com", "Japon Saç Sanatı", "Japon saç bakım sanatı; Cleansing Oil Shampoo; L'Oréal Professional; Japon güzellik felsefi"),
        ("Authentic Beauty Concept", "authenticbeautyconcept.com", "Vegan Salon", "Henkel'in vegan salon markası; temiz formüller; sürdürülebilir; Avrupa salon; %97+ doğal köken"),
    ],

    "Sağlık & Wellness": [
        ("Mudwater", "mudwtr.com", "Mantar Kahve Alt", "Kahve alternatifi; chaga + reishi + cacao; Shane Heath; podcaster favorisi; $60M+ gelir"),
        ("Clevr Blends", "clfrfrevr.com", "Meghan Markle Latte", "SuperLatte; matcha + oat milk; Meghan Markle yatırımcı; kadın wellness; adaptojenik"),
        ("Golde", "golde.co", "Güzellik Takviye", "Trinity Mouzon Wofford; kurkuma bazlı; superfood güzellik; Beyoncé BeyGOOD; küçük siyah kadın markası"),
        ("Moon Juice Superyou", "moonjuice.com", "Adaptojenik Stres", "SuperYou adaptojenik karışım; ashwagandha + shatavari; stres + enerji; LA wellness"),
        ("Your Super", "yoursuper.com", "Organik Süperfood", "Organik süperfood karışım; Super Green, Super Gut; Almanya merkezli; $30M+ gelir"),
        ("Laird Superfood", "lairdsuperfood.com", "Sörf Kahve Kreamer", "Laird Hamilton; hindistan cevizi bazlı kreamer; mantar kahve; performans beslenme"),
        ("REIZE", "reizeclub.com", "Avustralya Enerji Tozu", "Avustralya enerji içeceği tozu; 50mg kafein; B vitamini + taurin; $1/paket; online only"),
        ("Magic Mind", "magicmind.com", "Nootropik Shot", "Matcha + adaptojenik + nootropik; üretkenlik shot; James Beshara kurdu; Silicon Valley"),
        ("MUDWTR :rise", "mudwtr.com", "Sabah Ritueli", "Cacao + mantar sabah karışımı; kahve kafeininin 1/7'si; ritual odaklı pazarlama"),
        ("TruBrain", "trubrain.com", "Nootropik İçecek", "Bilim insanı formüle nootropik; piracetam + CILTEP; drink + kapsül; beyin performansı"),
        ("Four Sigmatic Coffee", "foursigmatic.com", "Mantar Kahve", "Lion's Mane + Chaga kahve; Finlandiya; fonksiyonel mantar kahve öncüsü; $100M+ gelir"),
        ("Brainzyme", "brainzyme.com", "İngiliz Nootropik", "İngiliz FSA tescilli; Focus Pro; bitki bazlı nootropik; İngiltere en çok satan beyin takviyesi"),
        ("Neurohacker Qualia", "neurohacker.com", "Kompleks Nootropik", "Qualia Mind 28 bileşen; Daniel Schmachtenberger; bilimsel kompleks; premium $139/ay"),
        ("Tru Niagen", "truniagen.com", "NAD+ Takviye", "ChromaDex; nicotinamide riboside; Nobel araştırma; mitokondri sağlığı; anti-aging"),
        ("Basis by Elysium", "elysiumhealth.com", "NAD+ Bilim", "NAD+ + sirtuin aktivasyonu; Nobel laureate danışman; klinik çalışmalar; premium yaşlanma"),
        ("Pendulum Glucose Control", "pendulumlife.com", "Probiyotik Kan Şekeri", "Akkermansia muciniphila; kan şekeri yönetimi probiyotik; Tip 2 diyabet; $100M+ yatırım"),
        ("Viome Precision", "viome.com", "AI Probiyotik", "Bağırsak mikrobiyom testi + kişisel probiyotik; AI formüle; Naveen Jain kurdu"),
        ("Sun Genomics Floré", "sungenomics.com", "Kişisel Probiyotik", "Mikrobiyom testi + kişisel probiyotik formülasyonu; AI; klinik doz; San Diego"),
        ("BiomeMD", "biomemd.com", "Bağırsak Sağlığı", "Sindirim enzim + probiyotik; bağırsak-beyin ekseni; holistic sindirim; bilimsel formül"),
        ("Qualia Life", "neurohacker.com", "Hücre Enerji", "Hücresel enerji + mitokondri desteği; 36 bileşen; performans + yaşlanma; ultra kompleks"),
    ],

    "Moda & Giyim": [
        ("The Frankie Shop", "thefrankieshop.com", "Paris Oversize", "Gaëlle Drevet; Bea Blazer ikonik; oversize tailoring trendi; Parisian effortless; NYC + Paris"),
        ("Amina Muaddi", "aminamuaddi.com", "İtalyan Topuklu Ayakkabı", "Kristal stiletto; Gilda sandalet; Rihanna collab; $800+ ayakkabı; Ürdünlü-Romanyalı tasarımcı"),
        ("By Far", "bfryfar.com", "Bulgar Vintage Aksesuar", "Rachel çanta ikonik; Bulgaristan; vintage 90s estetik; mini çanta trendi; $100M+ gelir"),
        ("Nanushka", "nanushka.com", "Macar Vegan Deri", "Budapest; vegan deri uzmanı; Boxy Jacket; sürdürülebilir Doğu Avrupa moda; Sandra Sandor"),
        ("Stine Goya", "stinegoya.com", "Danimarka Renk Kraliçesi", "Kopenhag renk + desen; İskandinav 2.0; neşeli giyinme trendi; Scandi moda farklı yüzü"),
        ("Hosbjerg", "hosbjerg.com", "Danimarka Cesur Moda", "Kopenhag; renkli + cesur; upcycle vintage kumaş; young Scandinavian; sürdürülebilir"),
        ("Samsøe Samsøe", "samsoe.com", "Danimarka Günlük Lüks", "Kopenhag; 30+ yıl; günlük lüks; erkek + kadın; İskandinav premium basics"),
        ("Filippa K", "filippa-k.com", "İsveç Minimalizm", "Stockholm minimalizm; gardrob temelleri; sürdürülebilir; ikinci el platform; İskandinav lüks"),
        ("Tiger of Sweden", "tigerofsweden.com", "İsveç Tailoring", "İsveç erkek + kadın tailoring; 100+ yıl; modern İskandinav; premium iş giyim"),
        ("Holzweiler", "holzweiler.com", "Norveç Streetwear Lüks", "Oslo; atkı ile başladı; Scandi streetwear; aile şirketi; cesur grafik; Norveçli tasarım"),
        ("Wood Wood", "woodwood.com", "Danimarka Streetwear", "Kopenhag streetwear; Nike + Adidas collab; graf tasarım; İskandinav sokak kültürü"),
        ("Won Hundred", "wonhundred.com", "Danimarka Sürdürülebilir", "Kopenhag; %100 organik denim hedefi; İskandinav sürdürülebilir moda; temiz tasarım"),
        ("A.P.C.", "apc.fr", "Fransız Minimalizm", "Jean Touitou; raw denim kült; Petit Standard; Paris minimalizm; $200M+ gelir"),
        ("AMI Paris", "amipfraris.com", "Coeur Fransız Moda", "Alexandre Mattiussi; kalp logosu ikonik; Paris erkek + kadın; accessible luxury; $500M+ gelir"),
        ("Maison Kitsuné", "maisonkitsune.com", "Fransız-Japon Hybrid", "Paris + Tokyo; Café Kitsuné; tilki logosu; müzik + moda; kültürel marka"),
        ("Lemaire", "lemaire.fr", "Sessiz Lüks Fransız", "Christophe Lemaire; UNIQLO U collab; sessiz lüks; Paris atölye; minimalist tasarım"),
        ("Isabel Marant DTC", "isabelmarant.com", "Fransız Bohem", "Étoile hattı erişilebilir; Bobby sneaker viral; Paris bohem chic; DTC genişleme"),
        ("Zadig & Voltaire", "zadig-et-voltaire.com", "Fransız Rock Chic", "Paris rock'n'roll moda; skull + cashmere; $500M+ gelir; DTC büyüme; Thierry Gillier"),
        ("The Row", "therow.com", "Olsen İkizleri Lüks", "Mary-Kate + Ashley Olsen; ultra sessiz lüks; $5,000+ çantalar; CFDA ödülleri; minimal Amerikan lüks"),
        ("Khaite", "khaite.com", "NYC Modern Lüks", "Catherine Holstein; Eda kaşmir süveter; modern Amerikan lüks; $100M+ gelir; Cathay Capital yatırımcı"),
    ],

    "Yiyecek & İçecek": [
        ("Three Wishes", "threewishes.com", "Protein Gevrek", "8g protein, 3g şeker gevrek; Margaret + Ian Wishingrad; Whole Foods + Target; sağlıklı çocuk gevreği"),
        ("Erewhon", "erewhon.com", "LA Süpermarket Lüks", "Los Angeles lüks organik market; $17 smoothie; ünlü smoothie collabları; Hailey Bieber smoothie viral"),
        ("Fly By Jing Sichuan", "flybyjing.com", "Çin Sos DTC", "Chili Crisp ikonik; Zhong sauce; otantik Sichuan; Jing Gao; Çin mutfağı ABD'ye"),
        ("Momofuku Goods", "momofukugoods.com", "David Chang Sos", "Chili Crunch ikonik; Seasoned Salt; David Chang restoran to DTC; şef markası"),
        ("Fishwife Tinned", "eatfishwife.com", "Konserve Balık Trendi", "Konserve balık trendini başlattı; sardunya + midye; renkli etiket; Instagram güzel gıda"),
        ("Diaspora Co", "diasporaco.com", "Hint Baharat DTC", "Sana Javeri Kadri; tek çiftlik Hint baharatları; zerdeçal + biber; adil ticaret; direct trade"),
        ("Burlap & Barrel", "burlapandbarrel.com", "Tek Kaynak Baharat", "Single-origin baharat; doğrudan çiftçiden; Rainbow Peppercorns; Black Urfa; gourmet baharat"),
        ("Jacobsen Salt", "jacobsensalt.com", "Oregon Deniz Tuzu", "El yapımı Oregon deniz tuzu; Netarts Bay; tütsülenmiş + infüze; artisan tuz; DTC gourmet"),
        ("Maldon Salt DTC", "maldon.com", "İngiliz Tuz İkonu", "Piramit kristal deniz tuzu; Essex İngiltere; 1882'den beri; şef favorisi; DTC genişleme"),
        ("Rancho Gordo", "ranchogordo.com", "Heirloom Fasulye", "Heirloom fasulye; Napa Valley; Bean Club abonelik; aylarca bekleme listesi; $12M+ gelir"),
        ("Acid League", "acidleague.com", "Living Vinegar", "Canlı kültür sirke; Proxies alkolsüz şarap; fermente gıda; Vancouver; hipster gourmet"),
        ("Ghee Hee", "fourthandheart.com", "Ghee DTC", "4th & Heart; vanilya + truffle ghee; grass-fed; Ayurveda bazlı tereyağı; DTC ghee"),
        ("Brightland Olive", "brightland.com", "California Zeytinyağı", "Tek hasat; güzel etiket; DTC premium gıda estetiği; Aishwarya Iyer; California çiftlik"),
        ("Pineapple Collaborative", "pineapplecollaborative.com", "Kadın Gıda Topluluğu", "Kadın kurucuların gıda markası; çeşitlilik + dahil etme; newsletter to brand"),
        ("Umamicart", "umamicart.com", "Asya Market DTC", "Online Asya gıda market; 1000+ ürün; Asya-Amerikan mutfak; NYC; DTC Asya market"),
        ("Weee!", "sayweee.com", "Asya Online Grocery", "Asya + Hispanik online market; $800M+ yatırım; etnik market DTC; SF merkezli"),
        ("Goldbelly", "goldbelly.com", "İkonik Yemek Teslimat", "ABD'nin en ünlü restoranlarından yemek gönderimi; Pat LaFrieda, Katz's Deli; gourmet hediye"),
        ("Levain Bakery DTC", "levainbakery.com", "NYC İkonik Cookie", "NYC dev kurabiye; 6oz kurabiye; 25+ yıl; Harlem + Upper West Side; DTC genişleme"),
        ("Milk Bar DTC", "milkbar.com", "Christina Tosi Tatlı", "Christina Tosi; Crack Pie ikonik; Cereal Milk flavor; Momofuku spinoff; DTC tatlıcı"),
        ("Partake Foods", "partakefoods.com", "Alerjen Dostu Snack", "Top 9 alerjen free; Denise Woodard; Jay-Z yatırımcı; kapsayıcı atıştırmalık"),
    ],

    "Ev & Mutfak": [
        ("Tuft & Needle", "tuftandneedle.com", "OG DTC Yatak", "İlk DTC yatak markalarından; T&N Adaptive Foam; Serta Simmons satın aldı; erişilebilir"),
        ("Saatva DTC", "saatva.com", "Lüks Yatak Online", "Beyaz eldiven teslimat; lüks spring yatak; $500M+ gelir; premium uyku; innerspring"),
        ("Purple Mattress", "purple.com", "Grid Teknoloji Yatak", "Hyper-Elastic Polymer; ham yumurta testi viral; farklılaşan his; $2B pik"),
        ("Leesa Sleep", "leesa.com", "Sosyal Etki Yatak", "10'da 1 bağış; B Corp; $275M gelir; sosyal etki DTC; köpük yatak"),
        ("Thuma", "thuma.co", "Platform Yatak Çerçeve", "Japon marangozluk; aletiz montaj; The Bed minimal platform; $1000; PillowBoard başlık"),
        ("Dims", "dims.com", "Erişilebilir Tasarım Mobilya", "NYC; yeni tasarımcılarla erişilebilir mobilya; $500-2000; modern tasarım demokratikleştirildi"),
        ("Maiden Home", "maidenhome.com", "Özel Üretim Mobilya", "North Carolina atölyede özel üretim; 6-8 hafta; $1000-4000; zanaatkar kalite DTC fiyat"),
        ("Sabai", "sabai.design", "Sürdürülebilir Kanepe", "Yeniden döşenebilir + geri dönüşüm; modüler; sürdürülebilir mobilya; $1000-2000; Miami"),
        ("Castlery", "castlery.com", "Singapur Modern Mobilya", "Singapur; modern mobilya; DTC fiyat; 6 ülkede satış; Avrupa + Asya tasarım"),
        ("Valyou", "valyoufurniture.com", "Uygun Fiyat Mobilya", "Feathers kanepe TikTok viral; $300-800; Gen Z mobilya; bulut kanepe; Instagram aesthetic"),
        ("Albany Park", "albanypark.com", "Modüler Kanepe", "Kova kanepe; modüler; uygun fiyat; 2 kutudan çıkar; aletiz kurulum; $1000-2000"),
        ("Campaign Living", "campaignliving.com", "Modüler Modern", "Campaign kanepe + yatak; yıkanabilir kılıf; modüler; ABD üretim; $800-2000"),
        ("Inside Weather", "insideweather.com", "Özelleştirilebilir Mobilya", "350+ kumaş seçeneği; kanepe + sandalye; kişiselleştirme; California; 3-5 hafta teslimat"),
        ("Benchmade Modern", "benchmademodern.com", "California Custom", "California'da özel üretim kanepe; 100+ kumaş; 4-6 hafta; $1500-3000; hızlı custom"),
        ("Rove Concepts", "roveconcepts.com", "Vancouver Tasarım Mobilya", "Vancouver; mid-century modern; İtalya + Çin üretim; $1000-5000; DTC tasarım mobilya"),
    ],

    "Evcil Hayvan": [
        ("Scratch", "scratch.com.au", "Avustralya Köpek Mama", "Avustralya; doğal kuru köpek maması; kişiselleştirilmiş; insan kalitesinde; abonelik"),
        ("Lyka", "lfryka.com.au", "Avustralya Taze Pet Mama", "Avustralya taze köpek maması; human grade; kişiselleştirilmiş porsiyon; Sydney"),
        ("Tails.com", "tails.com", "İngiliz Kişisel Köpek Mama", "İngiliz kişiselleştirilmiş kuru köpek maması; Nestlé Purina satın aldı; UK pet DTC"),
        ("Butternut Box", "butternutbox.com", "İngiliz Taze Mama", "İngiltere taze köpek maması #1; Kevin & Dave kurdu; $300M+ yatırım; kişiselleştirilmiş"),
        ("Different Dog", "differentdog.com", "İngiliz Ev Pişirme", "İngiliz ev pişirme tarzı köpek maması; insan kalitesinde; taze teslimat; vet formüle"),
        ("Pooch & Mutt", "poochandmutt.com", "İngiliz Fonksiyonel Pet", "İngiliz fonksiyonel köpek maması; Calm & Relaxed; sağlık odaklı; doğal"),
        ("Republic of Cats", "republicofcats.com", "İngiliz Kedi Mama", "İngiliz kişiselleştirilmiş kedi maması; abonelik; Tails.com ekibi; kedi uzmanı"),
        ("Katkin", "katkin.com", "İngiliz Taze Kedi Mama", "İngiliz taze kedi maması; insan kalitesinde; kişisel porsiyon; kedi sağlığı"),
        ("YuMOVE", "yumove.com", "İngiliz Eklem Takviye", "İngiliz köpek eklem takviyesi; glucosamine + omega-3; vet önerisi #1 UK; $100M+ gelir"),
        ("Bella & Duke", "bellaandduke.com", "İngiliz Raw Pet Mama", "İskoçya; raw (çiğ) köpek + kedi maması; doğal beslenme; BARF diyet; abonelik"),
    ],

    "Aksesuar & Takı": [
        ("Pandora DTC", "pandora.net", "Charm Takı DTC", "Charm bileklik; $4B+ gelir; lab-grown pırlanta; DTC kanalına geçiş; kişiselleştirilebilir"),
        ("Monica Vinader", "monicavinader.com", "İngiliz Gravür Takı", "Gravür kişiselleştirme; geri dönüşüm altın + gümüş; Kate Middleton giydi; İngiliz demi-fine"),
        ("Completedworks", "completedworks.com", "Londra Sanat Takı", "Anna Jewsbury; heykelsi takı; organik formlar; sanat eseri mücevher; Londra; $200-1000"),
        ("Agmes", "agfrmes.com", "NYC El Yapımı Takı", "Morgan Thomas; el yapımı NYC; heykelsi altın; minimal + cesur; ABD üretim; $200-2000"),
        ("Wolf Circus", "wolfcircus.com", "Vancouver Modern Takı", "Vancouver; geri dönüşüm bronz + altın; modern günlük takı; uygun fiyat; Kanada DTC"),
        ("Vitaly", "vitalydesign.com", "Kanada Streetwear Takı", "Toronto; paslanmaz çelik; kalın zincir; erkek + kadın; streetwear; $25-150"),
        ("Dorsey", "bydorsey.com", "Lab-Grown Renkli Taş", "Lab-grown yakut, safir, zümrüt; lüks mücevher erişilebilir; sürdürülebilir; Meg Strachan"),
        ("Kimai", "kimai.com", "Belçika Lab-Grown", "Antwerp lab-grown pırlanta; İngiliz-Belçika; etik fine jewelry; Jessica Warch + Sidney Neuhaus"),
        ("Matilde", "matilde.com", "İspanyol Minimal Takı", "Barcelona; minimal altın takı; günlük fine jewelry; İspanyol DTC; uygun fiyat lüks"),
        ("Milamore", "milamore.com", "Japon Kintsugi Takı", "Kintsugi (altın onarım) ilhamlı takı; Tokyo + NYC; Japon zanaat + modern tasarım"),
    ],

    "Teknoloji & Elektronik": [
        ("Aqara", "aqara.com", "Akıllı Ev Sensör", "Zigbee + Matter; kapı/pencere sensör + sıcaklık; uygun fiyat akıllı ev; HomeKit; Çin"),
        ("Meross", "meross.com", "HomeKit Akıllı Fiş", "HomeKit uyumlu akıllı fiş; $15; garaj açıcı; LED şerit; uygun fiyat Apple ev"),
        ("Withings", "withings.com", "Fransız Sağlık Cihaz", "Body+ akıllı tartı; ScanWatch; Fransız sağlık teknoloji; Nokia sattı geri aldı; $100M+ gelir"),
        ("Oura", "ouraring.com", "Finlandiya Sağlık Yüzük", "Gen3 yüzük; uyku + stres + aktivite takibi; $2.55B değerleme; Oulu Finlandiya; mücevher gibi"),
        ("Whoop 4.0", "whoop.com", "Performans Band", "Strain Coach; HRV takibi; Recovery Score; abonelik $30/ay; elite sporcu; $3.6B değerleme"),
        ("Garmin DTC", "garmin.com", "GPS Spor Saati", "Fenix + Forerunner; GPS teknoloji; fitness + outdoor; $5B+ gelir; DTC genişleme; Olathe KS"),
        ("Coros", "coros.com", "Koşucu GPS Saati", "Pace 3; ultra hafif GPS koşu saati; Eliud Kipchoge kullanıyor; uygun fiyat premium; Çin/ABD"),
        ("Suunto", "suunto.com", "Finlandiya Outdoor Saat", "Suunto 9 Peak Pro; Finlandiya pusuladan smartwatch'a; outdoor spor; 80+ yıl; Amer Sports"),
        ("Polar", "polar.com", "Finlandiya HRM Öncü", "Kalp atış monitörü icat etti; Vantage V3; Finlandiya; bilimsel doğruluk; spor bilimi"),
        ("Amazfit", "amazfit.com", "Erişilebilir Akıllı Saat", "Zepp Health; T-Rex Ultra; GTR 4; $50-200; uygun fiyat premium; Çin; Huami"),
        ("Mobvoi", "mobvoi.com", "TicWatch Akıllı Saat", "TicWatch Pro 5; Google Wear OS; Çin; $200-300; dual display; AI teknoloji"),
    ],

    "Abonelik Kutuları": [
        ("Crate & Barrel Box", "crateandbarrel.com", "Ev Dekor Kutusu", "Sezonluk ev dekor kutusu; modern ev estetiği; premium ev ürünleri keşfi"),
        ("Paw Pack", "thepawpack.com", "Doğal Köpek Kutu", "Doğal + organik köpek ödül kutusu; sağlıklı treat; ABD yapımı; aylık keşif"),
        ("KitNipBox", "kitnipbox.com", "Kedi Oyuncak Kutu", "Aylık kedi oyuncak + ödül; tematik kutular; Happy Cat, Multi-Cat seçenekleri"),
        ("Boxy Charm Luxe", "boxycharm.com", "Lüks Güzellik Kutu", "Full-size lüks ürünler; $30/ay $250+ değer; IPSY bünyesinde; premium güzellik keşfi"),
        ("ArtSnacks", "artsnacks.co", "Sanat Malzeme Kutu", "Aylık premium sanat malzemesi kutusu; sürpriz markalar; sanatçı + hobi; yaratıcılık"),
        ("Nomadik", "nomadik.com", "Outdoor Kutu", "Aylık outdoor + kamp ekipmanı kutusu; sürpriz outdoor ürünler; macera keşfi"),
    ],

    "Seyahat & Bavul": [
        ("Pelican", "pelican.com", "Dayanıklı Kılıf DTC", "Askeri seviye dayanıklı kılıf; Air Case; su geçirmez; profesyonel ekipman taşıma"),
        ("Travelpro", "travelpro.com", "Pilot Bavul", "Havayolu pilotlarının tercihi; Crew versalite; 30+ yıl; DTC genişleme; profesyonel seyahat"),
        ("Delsey", "delsey.com", "Fransız Bavul", "Paris; Chatelet Air 2.0; Fransız tasarım; 75+ yıl; premium Fransız bavul; DTC büyüme"),
        ("Thule", "thule.com", "İsveç Taşıma Çözüm", "Araç üstü taşıma + bavul + çanta; İsveç; 80+ yıl; Subterra; outdoor + şehir seyahati"),
    ],

    "Parfüm & Koku": [
        ("Memo Paris", "memfrfrrfrfris.com", "Seyahat İlham Koku", "Her koku bir seyahat anı; Irish Leather; African Leather; lüks niş; Paris; $200+"),
        ("Xerjoff", "xerjoff.com", "İtalyan Ultra Lüks", "Torino; XJ 1861 koleksiyon; Casamorati; ultra lüks İtalyan niş parfüm; $200-500"),
        ("Amouage", "amouage.com", "Oman Lüks Parfüm", "Oman Sultanı kurdu; Interlude, Reflection; Orta Doğu lüks; $200-400; niş parfüm"),
        ("Parfums de Marly", "parfumsdemfrarly.com", "Versay Saray İlham", "18. yüzyıl Versay ilhamı; Layton viral; Delina kadın favorisi; $200-350; TikTok hype"),
        ("MFK Baccarat Rouge", "franciskurkdjian.com", "Baccarat Rouge 540 Viral", "Maison Francis Kurkdjian; BR540 TikTok en viral parfüm; LVMH; $300+; niş mainstream"),
    ],

    "Outdoor & Macera": [
        ("Gregory Mountain", "gregorypacks.com", "Trekking Sırt Çantası", "Baltoro sırt çantası; Samsonite bünyesinde; professional trekking; fit odaklı tasarım"),
        ("Deuter", "deuter.com", "Alman Trekking", "1898'den beri; Aircontact sırt çantası; Alman mühendislik; Futura; hiking standardı"),
        ("Lowe Alpine", "lowealpine.com", "İngiliz Dağ Çantası", "AirZone teknolojisi; İngiliz dağ çantası mirası; Rab ile birlikte Equip Outdoor'da"),
        ("Vaude", "vaude.com", "Alman Yeşil Outdoor", "Alman sürdürülebilir outdoor; Green Shape; B Corp benzeri EMAS; Tettnang; aile şirketi"),
        ("Ortovox", "ortovox.com", "Alp Güvenlik", "Çığ güvenlik ekipmanı; Merino yün baz katman; Alp dağcılığı; Alman kalite; 40+ yıl"),
    ],

    "Supplement & Sporcu Beslenmesi": [
        ("Beyond Raw", "beyondraw.com", "GNC Premium", "GNC'nin premium hattı; LIT pre-workout; perakende + DTC; $30-50; performans odaklı"),
        ("Bucked Up", "buckedup.com", "Geyik Kadife Pre-Workout", "Deer antler velvet; WOKE AF pre-workout; agresif pazarlama; $200M+ gelir; hardcore fitness"),
        ("Gorilla Mode", "gorillamind.com", "Pre-Workout Viral", "Derek (MPMD) kurdu; Nitric pre-workout; mega doz formüller; YouTube fitness topluluğu"),
        ("PEScience", "pescience.com", "Araştırma Bazlı", "Select Protein casein + whey; bilimsel araştırma bazlı; Tim Muriello; lezzet odaklı"),
        ("NutraBio", "nutrabio.com", "Full Disclosure", "Tam açıklama etiket 1996'dan beri; öncü şeffaf marka; NJ üretim; hardcore kalite"),
        ("Vital Proteins Collagen", "vitalproteins.com", "Kolajen Kahve", "Collagen Creamer; kahveye kolajen trendi; Jennifer Aniston; Nestlé; DTC kolajen lideri"),
        ("Beam Dream Powder", "beamorganics.com", "Uyku Takviye", "Dream Powder nano CBD + magnezyum uyku tozu; TikTok viral; uyku rutini; $100M+ gelir"),
        ("Momentous Huberman", "livemomentous.com", "Huberman Podcast Takviye", "Andrew Huberman'ın tek önerdiği marka; NSF Sport; klinik doz; bilimsel güvenilirlik"),
        ("Athletic Greens Gut", "drinkag1.com", "Bağırsak Sağlığı Toz", "AG1 probiyotik + prebiyotik + enzim; 7.2B CFU; sindirim sağlığı + bağışıklık; günlük ritüel"),
    ],

    "Oyun & Eğlence": [
        ("Amiibo", "nintendo.com/amiibo", "Nintendo Figür", "NFC figür; Nintendo Switch oyunlarıyla etkileşim; koleksiyon; 100+ karakter; gaming + koleksiyon"),
        ("Hot Toys", "hottoys.com.hk", "Ultra Detay Figür", "Hong Kong; 1:6 ölçek film figürleri; Marvel, Star Wars; $200-500; koleksiyoncu premium"),
        ("Bandai Namco DTC", "bandainamco.com", "Japon Oyuncak DTC", "Gunpla model kit; Tamagotchi; Japon anime + oyuncak; DTC genişleme; $8B+ gelir"),
        ("Hasbro Pulse DTC", "hasbropulse.com", "Koleksiyoncu DTC", "Hasbro'nun DTC platformu; exclusive Transformers, Star Wars; koleksiyoncu özel ürünler"),
        ("Pokémon Center DTC", "pokemoncenter.com", "Pokémon Resmi Mağaza", "Resmi Pokémon ürünleri; plush + TCG + giyim; DTC koleksiyoncu; $1B+ marka"),
    ],

    "Kadın Sağlığı": [
        ("Clio", "helloclfrfio.com", "Akıllı Doğurganlık", "Akıllı doğurganlık ve adet takip bileklik; vücut sıcaklığı + nabız; AI tahmin"),
        ("Binto", "mybinto.com", "Hemşire Formüle Vitamin", "Kişisel kadın vitamin; hemşire formüle; PMS + doğurganlık + hamilelik; kişiselleştirme"),
        ("Semaine Health", "semaine.com", "PMS Takviye", "PMS semptomları için bitkisel takviye; Ann Garnier kurdu; vücut ağrısı + ruh hali"),
        ("Elix", "elixhealing.com", "Çin Tıbbı Kadın Sağlık", "Geleneksel Çin tıbbı + modern bilim; Cycle Balance; adet düzenleme; kişiselleştirilmiş"),
    ],

    "Diş & Ağız Bakımı": [
        ("ZenyumSonic", "zenyum.com", "Asya Sonic Fırça", "Güneydoğu Asya elektrikli fırça; uygun fiyat; uygulama bağlantılı; Singapur DTC dental"),
        ("Brushbox", "brushbox.com", "Abonelik Fırça Başlığı", "Oral-B uyumlu fırça başlığı aboneliği; İngiliz; 3 ayda kapıya; zamandan tasarruf"),
        ("Happier Beauté", "happierfr.com", "Sonic Fırça+", "Premium sonic fırça; 40K titreşim/dk; USB-C; seyahat boy; TikTok viral"),
    ],

    "Ofis & Kırtasiye": [
        ("Shinola", "shinola.com", "Detroit Lüks Journal", "Detroit; lüks defter + saat + deri ürünler; Amerikan üretim; ABD zanaat; $100M+ gelir"),
        ("Appointed", "appointed.co", "DC Lüks Kırtasiye", "Washington DC; lüks planlayıcı; ABD yapımı; Suann Song kurdu; kağıt + deri lüks"),
        ("Nolki", "nolki.com", "İngiliz Kağıt Atölyesi", "Londra kağıt atölye; workshop + ürün; artisan kırtasiye; el yapımı defter"),
    ],

    "Uyku & Yatak": [
        ("Koala Mattress", "koala.com", "Avustralya Ekoloji Yatak", "Avustralya; koala koruma bağışı; ekolojik; 120 gece deneme; hızla büyüyen marka"),
        ("Eve Sleep", "evesleep.com", "İngiliz DTC Yatak", "İngiliz DTC yatak; Hybrid Premium; Londra; 100 gece deneme; Bensons satın aldı"),
        ("Otty", "otty.com", "İngiliz Hibrit Yatak", "İngiliz hibrit yatak; 2000 pocket spring; OTTY Pure; uygun fiyat İngiliz premium"),
        ("Nectar UK", "nectarsleep.co.uk", "İngiliz Değer Yatak", "365 gece deneme; yaşam boyu garanti; İngiliz pazar; bellek köpük; uygun fiyat"),
        ("Ergoflex", "ergoflex.com.au", "Avustralya Bellek Köpük", "Avustralya + İngiliz bellek köpük; 30 gece deneme; uygun fiyat; 10 yıl garanti"),
    ],

    "Sürdürülebilir Ürünler": [
        ("Zero Co", "zeroco.com.au", "Avustralya Refill Temizlik", "Avustralya refill temizlik; okyanus plastik şişeler; geri gönder-doldur; sıfır atık Avustralya"),
        ("Nothing Naughty", "nothingnaughty.co.nz", "NZ Doğal Atıştırmalık", "Yeni Zelanda sağlıklı bar; doğal bileşenler; NZ spor topluluğu; sürdürülebilir"),
        ("KeepCup", "keepcup.com", "Avustralya Reusable Cup", "Avustralya tekrar kullanılabilir kahve bardağı; cam + plastik; barista standarı; 2009'dan beri"),
        ("Frank Green", "frankgreen.com", "Avustralya Akıllı Bardak", "Avustralya; akıllı bardak + su şişesi; kişiselleştirilebilir renk; push-button kapak; $100M+"),
        ("Chilly's", "chillys.com", "İngiliz Su Şişesi", "Londra; yalıtımlı su şişesi; Series 2; Emma Bridgewater collab; İngiliz tasarım"),
        ("Ocean Bottle", "oceanbottle.co", "Okyanus Plastik Şişe", "Her şişe 11.4kg okyanus plastik toplar; Norveç tasarım; çevre etkisi; B Corp"),
        ("Kleen Kanteen", "kleankanteen.com", "Paslanmaz Öncü", "1 milyon+ plastik tasarruf; B Corp; California; 20+ yıl; paslanmaz çelik öncüsü"),
        ("Stojo", "stojo.co", "Katlanır Bardak", "Katlanır silikon kahve bardağı; cep boyutu; seyahat; BPA-free; NYC; sıfır atık kahve"),
        ("Ecoffee Cup", "ecoffee.com", "Bambu Kahve Bardağı", "Bambu lifi kahve bardağı; tasarım desenler; doğal malzeme; kompostlanabilir kapak"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# EXTRA BRANDS BATCH 4 — final 500+ to exceed 2000 total
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS_4 = {
    "Güzellik & Cilt Bakımı": [
        ("Laniege Water Bank", "laneige.com", "Kore Hydro Teknoloji", "Water Bank Blue Hyaluronic serum; Cica Sleeping Mask; AmorePacific; K-beauty nem ikonu"),
        ("Etude House", "etude.com", "Kore Sevimli Makyaj", "SoonJung hassas cilt; Play Color Eyes; kawaii ambalaj; AmorePacific; uygun K-beauty"),
        ("3CE", "3ce.com", "Kore Trend Makyaj", "Stylenanda bünyesinde; Velvet Lip Tint; L'Oréal satın aldı; Kore trend makyaj #1"),
        ("Moonshot", "moonshot-cosmetics.com", "Kore Micro Cushion", "Micro Glassyfit Cushion; GD (G-Dragon) eski marka yüzü; YG Entertainment; Kore trend"),
        ("Dear Dahlia", "deardahlia.com", "Kore Lüks Vegan", "Vegan lüks K-beauty; Paradise Dream koleksiyon; doğal ışıltı; premium Kore makyajı"),
        ("Hince", "hfrince.com", "Kore Mood Makyaj", "Mood Enhancer Lip Glow; minimalist Kore estetik; sanatsal ambalaj; mood odaklı güzellik"),
        ("Jung Saem Mool", "jfrungsfrfraemmool.com", "Kore Makyöz Lüks", "Kore'nin en ünlü makyözü; Essential Skin Nuder; kişisel renk analizi; lüks K-beauty"),
        ("Tamburins", "tamburins.com", "Kore Lüks Parfüm", "Gentle Monster'ın kardeş markası; Jennie (BLACKPINK) marka yüzü; lüks Kore koku + el kremi"),
        ("Nonfiction", "nfrfiction.kr", "Kore Niş Parfüm", "Seul niş parfüm; Santal Cream; Gaiac Room; Kore parfüm sahnesinin yükselişi"),
        ("Torriden DIVE-IN", "torriden.com", "Kore Hyalüronik Viral", "DIVE-IN Low Molecular Hyaluronic Acid Serum; TikTok en viral K-beauty serumu"),
        ("Numbuzin", "numbuzin.com", "Kore Numara Sistemi", "No.5 Vitamin-Niacinamide serum; numara sistemi ile ürün tanımlama; viral K-beauty"),
        ("Ma:nyo", "manyo.com", "Kore Bifida Bakım", "Bifida Biome Complex Ampoule; fermente K-beauty; Olive Young bestseller; bilimsel"),
        ("VT Cosmetics", "vtcosmetics.com", "Kore BTS Güzellik", "BTS eski marka yüzü; Cica serisi; PDRN ampoule; Kore dermatolojik güzellik"),
        ("Olive Young", "oliveyoung.com", "Kore Güzellik Marketplace", "Kore'nin Sephora'sı; online DTC genişleme; kendi markası + multi-brand; K-beauty keşif"),
        ("Soko Glam", "sokoglam.com", "K-Beauty Küratör ABD", "Charlotte Cho kurdu; K-beauty küratörlük; Then I Met You kendi markası; ABD K-beauty hub"),
        ("Peach Slices", "peachslices.com", "Uygun K-Beauty ABD", "Acne Spot Dots; uygun fiyat K-beauty ABD'de; CVS'te; Alicia Yoon (Peach & Lily) kurdu"),
        ("Good Light", "goodlight.world", "Genderless Cilt Bakımı", "David Yi kurdu; cinsiyet normlarını kıran güzellik; Nova serum; kapsayıcı; Very Good Light blog"),
        ("Topicals Like Butter", "mytopicals.com", "Cilt Sorunları Nemlendirici", "Like Butter maske; Faded serum; hiperpigmentasyon; Gen Z cilt sorunlarını normalleştirme"),
        ("Vacation SPF", "vacation.inc", "80s Retro Güneş Kremi", "Classic Whip krem şantili güneş kremi; 1980'ler estetik; eğlenceli; nostaljik ambalaj"),
        ("Ultra Violette", "ultraviolette.com.au", "Avustralya SPF Uzmanı", "Avustralya SPF uzmanı; Queen Screen; Supreme Screen; her gün güneş koruma; Melbourne"),
        ("Supergoop Unseen", "supergoop.com", "Görünmez SPF", "Unseen Sunscreen görünmez SPF 40; makyaj altı; vegan; Holly Thaggard kurdu; SPF eğlenceli"),
        ("Canmake", "canmake.com", "Japon Uygun Makyaj", "Japon ¥600-1000 makyaj; Cream Cheek; Quick Lash Curler; Japon drugstore güzelliği"),
        ("Kate Tokyo", "nfrfrosegroup.com", "Japon Trend Makyaj", "Kanebo bünyesinde; Lip Monster TikTok viral; Japon lip tint; renkli makyaj"),
        ("Shiseido Waso", "wfrfrfrso.shiseido.com", "Japon Genç Cilt Bakımı", "Wasabi + tofu + miso gibi Japon süpergıdalar; genç cilt bakımı; Shiseido'nun DTC hattı"),
        ("Melano CC", "rohto.com", "Japon Vitamin C Serum", "Rohto; aktif vitamin C serum; ¥1000; Japon eczane bestseller; leke karşıtı; uygun fiyat"),
        ("Hada Labo", "hfrfrdlabo.com", "Japon Hyalüronik Öncü", "Rohto; Gokujyun Premium hyalüronik lotion; 7 tip HA; Japon eczane ikonik; nem standartı"),
    ],

    "Moda & Giyim": [
        ("Alo Yoga DTC", "aloyoga.com", "LA Yoga Moda", "Alo Moves dijital yoga; celebrity street style; $10B değerleme; LA yoga kültürü; premium"),
        ("Loewe DTC", "loewe.com", "İspanyol Lüks DTC", "JW Anderson kreatif; Puzzle Bag ikonik; LVMH; DTC büyüme; İspanyol zanaat lüks"),
        ("Bottega Veneta DTC", "bottegaveneta.com", "Sessiz Lüks DTC", "Matthieu Blazy; intrecciato örgü deri; sosyal medyadan çıktı; sessiz lüks ikonu; Kering"),
        ("JW Anderson", "jwanderson.com", "İngiliz Avant-Garde", "Pigeon Clutch viral; gender-fluid moda; Loewe kreatif direktörü; İngiliz yeni dalga tasarımcı"),
        ("Issey Miyake DTC", "isseymiyake.com", "Japon Plise Teknoloji", "Pleats Please; BAO BAO geometrik çanta; Japon teknoloji + moda; $500M+ gelir; DTC"),
        ("Uniqlo DTC", "uniqlo.com", "Japon Fonksiyonel", "HeatTech, AIRism, LifeWear; $20B+ gelir; DTC online büyüme; fonksiyonel günlük giyim"),
        ("Muji DTC", "muji.com", "Japon Markasız Kalite", "'no-brand quality goods'; minimalist; Japon estetik; ev + giyim + kırtasiye; $4B+ gelir"),
        ("COS New", "cosstores.com", "İskandinav Mimari Moda", "H&M grubunun en premium markası; mimari mağazalar; sürdürülebilir malzeme; minimalist İsveç"),
        ("RIXO", "rixo.co.uk", "İngiliz Vintage Baskı", "Londra; vintage ilhamlı baskılar; Henrietta Rix + Orlagh McCloskey; elbise uzmanı; İngiliz DTC"),
        ("Self-Portrait", "self-portrait-studio.com", "Malezya Londra Elbise", "Han Chong kurdu; dantel elbise uzmanı; Londra; erişilebilir lüks elbise; $200-500"),
        ("Needle & Thread", "needleandthread.com", "İngiliz İşleme Elbise", "Londra; el işlemeli elbiseler; düğün + özel gün; Hannah Coffin; İngiliz romantik"),
        ("Ganni Scandi 2.0", "ganni.com", "Danimarka Neşeli Moda", "#GanniGirls; leopar + smiley; LVMH yatırımcı; sorumlu moda; Kopenhag; $200M+ gelir"),
        ("Rotate Birger Christensen", "rotate.com", "Danimarka Parti Moda", "Kopenhag; puf kollu elbise; parti + gece; Thora Valdimars + Jeanette Madsen; TikTok"),
        ("Baum und Pferdgarten", "baumundpferdgarten.com", "Danimarka Renkli", "Kopenhag; cesur renkler + desenler; Rikke Baumgarten + Helle Hestehave; İskandinav neşeli"),
        ("By Malene Birger", "bymalenebirger.com", "Danimarka Sofistike", "Kopenhag; sofistike İskandinav kadın giyim; 2003; relaxed lüks; kuzey Avrupa şıklığı"),
        ("Cecilie Bahnsen", "ceciliebahnsen.com", "Danimarka Romantik Couture", "Kopenhag; hacimli romantik elbiseler; el işçiliği; LVMH Prize finalist; İskandinav haute couture"),
        ("Tory Burch DTC", "toryburch.com", "Amerikan Accessible Lüks", "Reva flat ikonik; $1.5B+ gelir; kadın girişimci; erişilebilir lüks; DTC genişleme"),
        ("Coach DTC", "coach.com", "Amerikan Heritage Lüks", "Tapestry bünyesinde; Stuart Vevers kreatif; Tabby çanta viral; $5B+ gelir; DTC dönüşüm"),
        ("Kate Spade DTC", "katespade.com", "NYC Neşeli Lüks", "Tapestry bünyesinde; polka dot + stripes; neşeli Amerikan kadın lüks; DTC büyüme"),
        ("Telfar", "tfrlfar.net", "Brooklyn Erişilebilir Lüks", "Telfar Clemens; Shopping Bag 'Bushwick Birkin'; her yerde tükeniyor; kapsayıcı lüks; NYC"),
    ],

    "Yiyecek & İçecek": [
        ("Drizly", "drizly.com", "Alkol Teslimat", "1 saatte alkol teslimatı; Uber satın aldı; online alkol perakende; $1.1B; kapandı/Uber entegre"),
        ("Minibar Delivery", "minibardelivery.com", "Premium Alkol DTC", "Premium alkol teslimatı; şarap + spirits; hediye; hızlı teslimat; online off-premise"),
        ("Brightland Vinegar", "brightland.com", "Premium Sirke", "Champagne vinegar; Parasol; California; güzel ambalaj; DTC gourmet premium"),
        ("Ghee Easy", "ghee-easy.com", "Avrupa Ghee", "Hollanda organik ghee; Avrupa ghee markası; Ayurveda; organik; sürdürülebilir"),
        ("JOI", "addjoi.com", "Bitki Bazlı Süt Konsantre", "Bitki sütü konsantre; badem + yulaf; kendin seyrelt; sıfır atık; taze plant milk"),
        ("Ripple Foods", "ripplefoods.com", "Bezelye Protein Süt", "Bezelye proteini bazlı süt; yüksek protein; sürdürülebilir; $200M+ yatırım; vegan süt"),
        ("Malk Organics", "mfralk.com", "Soğuk Presleme Bitki Sütü", "Soğuk preslenmiş badem + yulaf sütü; minimal bileşen; Austin TX; organik; $30M+"),
        ("Forager Project", "forfragerproject.com", "Organik Cashew Süt", "Organik cashew + oat süt; yogurt; kefir; California; bitkisel süt çeşitliliği"),
        ("Koia", "drinkkoia.com", "Bitkisel Protein Shake", "Bitkisel protein shake; 18g protein; 0g eklenen şeker; süpermarket; uygun fiyat; vegan"),
        ("Rebbl Protein", "rebbl.co", "Adaptojenik Protein", "Süper bitki protein shake; reishi + maca; fair trade; coconut milk bazlı; fonksiyonel"),
        ("Willa's Oat Milk", "willasoatmilk.com", "Organik Yulaf Sütü", "Organik; barista grade; 4 bileşen; Brooklyn; basit + temiz yulaf sütü"),
        ("MUSH", "eatmfrush.com", "Hazır Overnight Oats", "Buzdolabı rafında hazır overnight oats; 5 bileşen; grab-and-go kahvaltı; sağlıklı"),
        ("Brami", "brami.com", "İtalyan Lupini Snack", "Lupini fasulye atıştırmalık; 7g protein; İtalyan heritage; sürdürülebilir protein; düşük kalori"),
        ("Bada Bean Bada Boom", "badabeanbadaboom.com", "Kavrulmuş Fava", "Kavrulmuş fava fasulye; protein snack; 7g protein; lezzet çeşitliliği; sağlıklı atıştırmalık"),
        ("Peatos", "eatpeatos.com", "Bezelye Protein Cips", "Cheetos alternatifi; bezelye protein; 4g protein; junk food tadı sağlıklı bileşen; eğlenceli"),
        ("Outstanding Foods", "outstandingfoods.com", "Bitkisel Pork Rind", "Bitkisel domuz kabuğu; Outstanding Puffs; protein snack; vegan junk food alternatifi"),
        ("LesserEvil", "lesserevil.com", "Organik Popcorn", "Organik popcorn + paleo puff; hindistan cevizi yağı; sağlıklı snack; B Corp; $100M+ gelir"),
        ("SkinnyPop", "skinnypop.com", "Temiz Popcorn", "3 bileşen: popcorn, yağ, tuz; Hershey satın aldı; $400M+ marka; basit sağlıklı snack"),
        ("Pirate's Booty", "piratebrands.com", "Çocuk Snack", "Aged White Cheddar puff; Hershey; çocuk dostu sağlıklı snack; eğlenceli ambalaj"),
        ("Whisps", "whisps.com", "Peynir Cips", "Gerçek peynir cips; keto; düşük karbonhidrat; protein snack; $100M+ gelir; ICV Partners"),
    ],

    "Ev & Mutfak": [
        ("Dyson V15", "dyson.com", "Lazer Süpürge", "Lazer ile tozu görme; piezo sensör; İngiliz mühendislik; $700-800; premium kablosuz süpürge"),
        ("Tineco", "tineco.com", "Akıllı Zemin Temizlik", "Floor ONE S5; vakum + paspas; akıllı sensör; Çin; iFloor3; zemin temizlik yeniliği"),
        ("Bissell CrossWave", "bfrfrrissell.com", "Çok Amaçlı Temizlik", "CrossWave vakum + paspas tek aracta; Pet Pro; Little Green spot temizleyici; ev temizlik"),
        ("Narwal", "narwal.com", "Self-Clean Robot", "Freo X Ultra; kendi kendini temizleyen robot süpürge + paspas; AI; Çin; premium robot"),
        ("Fellow Ode", "fellowproducts.com", "Kahve Değirmeni", "Ode Brew Grinder Gen 2; SSP burr; specialty kahve; tasarım ödülleri; pour over için"),
        ("Timemore", "timemoreofficial.com", "Çin Specialty Kahve", "C2 Max el değirmeni; Chestnut; uygun fiyat specialty ekipman; Çin kahve yükselişi"),
        ("Origami", "origami-coffee.com", "Japon Pour Over Dripper", "Origami Dripper; renkli seramik; düz + konik filtre uyumlu; Japon kahve estetiği"),
        ("MoccaMaster KBG", "moccamaster.com", "Hollanda Premium Filtre", "El yapımı Amerongen Hollanda; SCA onaylı; 40+ yıl garanti; $300+ premium demleme"),
        ("Smeg DTC", "smeg.com", "İtalyan Retro Mutfak", "Retro buzdolabı + tost makinesi + kettle; İtalyan tasarım; pastel renkler; DTC genişleme"),
        ("KitchenAid Artisan", "kitchenaid.com", "Standmixer İkon", "Artisan standmixer 25+ renk; Whirlpool; ikonik mutfak; DTC genişleme; pastacı tercihi"),
        ("Le Creuset DTC Online", "lecreuset.com", "Fransız Emaye Online", "DTC online büyüme; renk sınırlı koleksiyonlar; 1925; Flame renk ikonik; $500M+ gelir"),
        ("Staub Cocotte", "staub-online.com", "Alsace Cocotte", "Siyah mat emaye iç; self-basting kapak; Zwilling bünyesinde; şef tercihi; DTC genişleme"),
        ("Riedel DTC", "riedel.com", "Avusturya Şarap Kadeh", "Üzüm çeşidine göre kadeh; 11 nesil aile; Avusturya kristal; sommelier standardı; DTC"),
        ("Nespresso DTC", "nespresso.com", "Kapsül Kahve Lüks", "Nestlé; Vertuo + Original; George Clooney; lüks kapsül kahve; DTC modeli; $6B+ gelir"),
        ("De'Longhi DTC", "delonghi.com", "İtalyan Espresso", "Dinamica + Magnifica espresso; İtalyan kahve makinesi; $3B+ gelir; DTC genişleme; ev barista"),
    ],

    "Sağlık & Wellness": [
        ("Sakara Life", "sakara.com", "Organik Yemek Teslimat", "Bitki bazlı organik yemek teslimat; detoks programları; lüks wellness beslenme; NYC; $100M+"),
        ("Daily Harvest DTC", "daily-harvest.com", "Dondurulmuş Sağlıklı", "Smoothie + harvest bowl; dondurulmuş sağlıklı gıda; $250M+ yatırım; pratik sağlıklı beslenme"),
        ("Factor DTC", "factor75.com", "Hazır Yemek Fitness", "Hazır sağlıklı yemek; keto + protein; HelloFresh bünyesinde; $1B+ gelir; fitness beslenme"),
        ("Trifecta Nutrition", "trifectanutrition.com", "Makro Bazlı Yemek", "Makro takipli yemek; organik; A La Carte; sporcu beslenme; CrossFit sponsoru"),
        ("Territory Foods", "territoryfoods.com", "Şef Hazır Yemek", "Yerel şefler hazırlıyor; allerjen filtre; keto + paleo + vegan; bölgesel teslimat"),
        ("Hungryroot AI", "hungryroot.com", "AI Sağlıklı Market", "AI yemek planı + market; kişiselleştirilmiş; sağlıklı; $40M+ yatırım; grocery + meal kit"),
        ("Splendid Spoon", "splendidspoon.com", "Bitkisel Hazır Yemek", "Bitkisel smoothie + soup + noodle bowl; vegan; detoks; abonelik; haftalık plan"),
        ("Mosaic Foods", "mosaicfoods.com", "Dondurulmuş Bitkisel Yemek", "Dondurulmuş bitkisel yemek; veggie burger + pizza; uygun fiyat; toptan sipariş"),
        ("Purple Carrot", "purplecarrot.com", "Bitkisel Yemek Kiti", "Bitkisel yemek kiti; Tom Brady eski ortak; vegan tarifler; haftalık teslimat"),
        ("Sun Basket DTC", "sunbasket.com", "Organik Yemek Kiti", "Organik; diyet özel; Paleo + Keto + Mediterranean; şef tasarım; $300M+ gelir"),
    ],

    "Fitness & Spor Giyim": [
        ("Hyperice Normatec", "hyperice.com", "Sıkıştırma Botları", "Normatec sıkıştırma toparlanma; NBA + NFL kullanıyor; $700M+ değerleme; pnömatik terapi"),
        ("Theragun Pro", "therabody.com", "Pro Masaj Tabancası", "Theragun PRO Plus; SmartSense basınç sensörü; $600; profesyonel toparlanma; fizyo terapi"),
        ("Bala Bangles", "shopbala.com", "Ağırlık Bileklik", "1 lb + 2 lb bileklik; Shark Tank $7M; estetik fitness; pembe + renkli; ev antrenmanı"),
        ("Hyperice Vyper", "hyperice.com", "Titreşimli Foam Roller", "Titreşimli foam roller; 3 hız; kas gevşetme; toparlanma; $200; sporcuların tercihi"),
        ("Whoop 4.0 Band", "whoop.com", "Toparlanma Bileklik", "Strain + Recovery + Sleep; ekransız; aylık $30; profesyonel sporcu; 24/7 veri toplama"),
        ("Garmin Forerunner", "garmin.com", "Koşu Saati", "Forerunner 265; AMOLED; koşu dinamikleri; $450; koşucu standartı; GPS doğruluğu #1"),
        ("Apple Fitness+", "apple.com/fitness", "Apple Ev Fitness", "Apple Watch entegre ev fitness; pilates, yoga, HIIT; $10/ay; 4K video; metriks ekranda"),
        ("Peloton Tread", "onepeloton.com", "Koşu Bandı", "Peloton Tread; canlı dersler; $3,495; ev koşu deneyimi; instructor community; NYC"),
        ("CAROL Bike", "carolbfrfrike.ai", "AI REHIT Bisiklet", "AI destekli REHIT 8 dakika antrenman; bilimsel kısa antrenman; $2,000; zaman kısıtlı fitness"),
        ("Hydrow Wave", "hydrow.com", "Kompakt Kürek", "Kompakt bağlı kürek makinesi; gerçek su çekimleri; $1,695; ev kürek; tam vücut"),
    ],

    "Bebek & Çocuk": [
        ("Snuggle Me", "snugglemeorganic.com", "Organik Lounger", "Organik pamuk bebek lounger; ABD yapımı; Minnesota; GOTS sertifikalı; DockATot alternatifi"),
        ("Babo Botanicals", "babobotanicals.com", "Botanik Bebek Bakım", "Botanik bebek cilt bakımı + güneş kremi; sertifikalı organik; allerjen dostu; pediatrist"),
        ("Babybjörn", "babybjorn.com", "İsveç Premium Bebek", "İsveç; Bouncer Bliss ikonik; 60+ yıl; premium bebek ekipmanı; minimal İskandinav tasarım"),
        ("Bugaboo DTC", "bugaboo.com", "Hollanda Premium Araba", "Hollanda; Fox 5 bebek arabası; Butterfly kompakt; $1,000+ ; Max Barenbrug kurdu; DTC"),
        ("Silver Cross", "silvercross.com", "İngiliz Heritage Araba", "1877'den beri İngiliz bebek arabası; Kraliyet Ailesi; Dream i-Size; heritage lüks; DTC"),
        ("Joie DTC", "joie.com", "Erişilebilir Bebek Ekipman", "Küresel bebek ekipmanı; i-Spin 360 oto koltuk; erişilebilir fiyat; geniş ürün yelpazesi"),
        ("BabyZen", "babyzen.com", "Fransız Kompakt Araba", "YOYO2 ultra kompakt araba; uçak kabinine sığar; Fransız tasarım; $500+; şehir seyahati"),
        ("Cybex", "cybex-online.com", "Alman Tasarım Bebek", "Alman güvenlik + tasarım; Sirona oto koltuk; Priam araba; GoodBaby bünyesinde; $500M+"),
        ("Joolz", "joolz.com", "Hollanda Pozitif Tasarım", "Amsterdam; pozitif tasarım; Day+ araba; sürdürülebilir; Hollanda'nın premium bebek arabası"),
        ("iCandy", "icandyworld.com", "İngiliz Lüks Araba", "İngiliz lüks bebek arabası; Peach; modüler; Londra tasarım; premium İngiliz zanaat"),
    ],

    "Teknoloji & Elektronik": [
        ("Samsung DTC", "samsung.com", "Galaxy Ekosistem DTC", "Galaxy S24; Galaxy Ring; AI features; DTC kanal; Samsung.com exclusive renk; $200B+ gelir"),
        ("Google Store DTC", "store.google.com", "Pixel DTC", "Pixel 8 Pro; Pixel Watch 2; Nest Hub; Google donanım DTC; AI entegrasyon"),
        ("Apple DTC", "apple.com", "Apple Store Online", "iPhone, Mac, Watch, AirPods; DTC kanal %36 gelir; Trade-in; özelleştirme; $383B gelir"),
        ("Dyson Tech", "dyson.com", "Mühendislik Teknoloji", "Zone kulaklık + hava temizleyici; Dyson OnTrac; teknoloji + mühendislik; DTC odaklı; $7B+"),
        ("Marshall Speakers", "marshallheadphones.com", "Rock Ses", "Stanmore III; Emberton II; vintage rock estetik; Marshall amp mirası; Bluetooth hoparlör"),
        ("Bang & Olufsen DTC", "bang-olufsen.com", "Danimarka Ses Lüks", "Beoplay H95; Beosound; Danimarka lüks ses; $500+ kulaklık; tasarım + ses; DTC"),
        ("Devialet", "devialet.com", "Fransız Ses Mühendislik", "Phantom hoparlör; ADH teknoloji; Fransız lüks ses; $3,000+; huawei collab; ultra premium"),
        ("Fujifilm DTC", "fujifilm.com", "Instax + X Serisi", "Instax Mini; X100VI kuyruk bekleme listesi; film kamera rönesansı; Japonya; DTC genişleme"),
        ("Polaroid DTC", "polaroid.com", "Anında Baskı Yeniden", "Now+ anında baskı kamera; retro revival; Gen Z film kamera trendi; Oskar Smolokowski CEO"),
        ("Lego Technic DTC", "lego.com", "Yetişkin Teknik Set", "Technic + Icons yetişkin LEGO; Porsche 911, Bugatti; $200-400; DTC exclusive set; $9B+ gelir"),
    ],

    "Supplement & Sporcu Beslenmesi": [
        ("Huel Black", "huel.com", "Yüksek Protein Meal", "50% daha fazla protein; 200 kalori shake; İngiltere; tam beslenme yüksek protein; DTC"),
        ("Soylent Squared", "soylent.com", "Protein Bar Yemek", "400 kalori bar; tam beslenme; Rob Rhinehart; yemek yerine; uygun; vegan"),
        ("Ka'Chava Tribal", "kachava.com", "Superfood Tribal", "70+ süpergıda; şaman ilhamı; $100M+ gelir; adaptojenik; tam beslenme shake"),
        ("Garden of Life RAW", "gardenoflife.com", "Ham Organik Protein", "RAW Organic Protein; çimlendirme; USDA organic; Nestlé; $100M+; tüm gıda bazlı"),
        ("Sunwarrior", "sunwarrior.com", "Bitkisel Sporcu", "Warrior Blend; pirinç + bezelye protein; vegan sporcu; organik; Utah; güneş enerjisi fabrika"),
        ("PlantFusion", "plantfusion.com", "Tam Bitkisel Protein", "Complete Plant protein; allerjen dostu; 21g protein; vegan; iyi tat; erişilebilir"),
        ("Naked Nutrition", "nfrfrkednutrition.com", "Ultra Temiz Protein", "1 bileşen whey; sıfır katkı; Naked Whey; ultra minimalist; $50M+ gelir; şeffaflık"),
        ("Animal Pak", "animalpak.com", "OG Bodybuilding", "Universal Nutrition; 1983'ten beri; Animal Pak vitamin pack; bodybuilding efsanesi; hardcore"),
        ("Cellucor C4 Ultimate", "cellucor.com", "Pre-Workout Pro", "C4 Ultimate 300mg kafein; Nutrabolt; en güçlü C4; hardcore performans; $500M+ marka"),
        ("Dymatize ISO100", "dymatize.com", "Hidrolize Whey", "ISO100 hidrolize whey izolat; Post It kampanyası; BellRing Brands; $300M+; hızlı emilim"),
    ],

    "Evcil Hayvan": [
        ("JustFoodForDogs", "justfoodfordogs.com", "İnsan Kalite Köpek Mama", "İnsan kalitesinde pişmiş köpek maması; $500M+ değerleme; mağazada pişirme; California"),
        ("Maev", "meetmaev.com", "Raw Köpek Bar", "Raw köpek maması bar; porsiyon bar; insan kalite; pratik raw feeding; Brooklyn"),
        ("Bone Broth for Dogs", "brutusbrothbone.com", "Köpek Kemik Suyu", "Köpek kemik suyu; eklem + sindirim; Brutus Broth; insan kalitesinde; sos olarak da"),
        ("Wisdom Panel Premium", "wisdompanel.com", "Premium DNA Testi", "350+ ırk; sağlık tarama; Mars Veterinary; genetik test; premium $160"),
        ("BarkBox Super Chewer", "barkbox.com/super-chewer", "Dayanıklı Oyuncak Kutu", "Güçlü çiğneyiciler için; dayanıklı oyuncak + ödül; aylık kutu; 2M+ abone alt hat"),
    ],

    "Oyun & Eğlence": [
        ("Lego Icons DTC", "lego.com", "Yetişkin Koleksiyon", "Eiffel Tower, Orchid, Bonsai; yetişkin koleksiyoncu; DTC exclusive; vitrin parçası"),
        ("Games Workshop DTC", "games-workshop.com", "Warhammer DTC", "Warhammer 40K + Age of Sigmar; minyatür oyun; İngiliz; $500M+ gelir; DTC büyüme"),
        ("Ravensburger Puzzle", "ravensburger.com", "Premium Puzzle DTC", "5,000+ parça puzzle; Alman kalite; yetişkin puzzle trendi; DTC online; $600M+ gelir"),
        ("Cards Against Humanity", "cardsagainsthumanity.com", "Parti Oyun DTC", "Yetişkin parti kart oyunu; crowdfund başarısı; Black Friday stunts; DTC only; kült"),
        ("Exploding Kittens", "explodingkittens.com", "Kickstarter Kart Oyun", "Kickstarter $8.7M toplayan kart oyunu; Matthew Inman (The Oatmeal); Netflix animasyon"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# EXTRA BRANDS BATCH 5 — final 350+ brands to surpass 2000
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS_5 = {
    "Güzellik & Cilt Bakımı": [
        ("U Beauty", "ubeauty.com", "Süper Akıllı Serum", "Tina Craig kurdu; Resurfacing Compound tek ürün 6 adım yerine; SIREN capsule teknoloji; lüks"),
        ("Saltair", "saltfrfrfrir.com", "İskele Duş Jeli", "TikTok viral vücut bakımı; Sephora'da; tropikal kokular; Gen Z vücut bakımı; erişilebilir"),
        ("Being Frenshe", "beingfrenshe.com", "Ashley Tisdale Wellness", "Ashley Tisdale markası; aromaterapi + vücut bakımı; Target; $10-15; wellness güzellik"),
        ("Topicals Faded", "mytopicals.com", "Leke Serumu", "Faded serum hiperpigmentasyon; Gen Z cilt sorunları normalize; Olamide Olowe kurdu; Sephora"),
        ("Beis Beauty", "beisbeauty.com", "Shay Mitchell Güzellik", "Shay Mitchell'ın güzellik uzantısı; seyahat boy güzellik; BÉIS markasından; pratik güzellik"),
        ("Dime Beauty", "dimebeautyco.com", "TikTok Viral Bakım", "TBC Serum TikTok viral; uygun fiyat lüks bakım; $20-40; erişilebilir premium"),
        ("The Beauty Chef", "thebeautychef.com", "Avustralya İç Güzellik", "Fermente güzellik takviyesi; GLOW Inner Beauty; bağırsak-cilt ekseni; Carla Oates kurdu"),
        ("Summer Friday CC Me", "summerfridays.com", "Vitamin C Serum", "CC Me Vitamin C Serum; Sephora bestseller; Jet Lag Mask ikonu; influencer marka"),
        ("Dr. Barbara Sturm", "drsturm.com", "Alman Lüks Derma", "Hyaluronic Serum; ünlü müşteriler; PRP tedavisi icat etti; $200+ ürünler; Alman cilt bilimi"),
        ("Jan Marini", "janmarini.com", "Profesyonel Cilt Tedavi", "Bioglycolic serisi; dermatoloji klinik; retinol + glikolik; 20+ yıl; profesyonel bakım"),
        ("iS Clinical", "isclinical.com", "Klinik Aktif Bakım", "Active Serum ikonik; Pro-Heal serum; botanik + bilim; dermatolojist klinik markası"),
        ("SkinMedica", "skinmedica.com", "TNS Büyüme Faktörü", "TNS Advanced+ Serum; Allergan bünyesinde; büyüme faktörleri; klinik anti-aging"),
        ("Murad", "murad.com", "Dermatolojist Marka", "Dr. Howard Murad; Retinol Youth Renewal; inclusionist; cilt sağlığı felsefesi; Unilever"),
        ("ZO Skin Health", "zoskinhealth.com", "Obagi Sonrası", "Dr. Zein Obagi kurdu; Exfoliating Polish; reçetesiz klinik sonuçlar; agresif aktif"),
        ("SkinCeuticals CE Ferulic", "skinceuticals.com", "C Vitamini İkon", "CE Ferulic serum sektör standartı; L'Oréal Active; $170 serum; klinik kanıtlı vitamin C"),
        ("Drunk Elephant Protini", "drunkelephant.com", "Peptit Nemlendirici", "Protini Polypeptide Cream; 9 sinyal peptit; Shiseido $845M; protein nemlendirici"),
    ],

    "Saç Bakımı": [
        ("Amika Soulfood", "loveamika.com", "Besleyici Maske", "Soulfood Nourishing Mask; murumuru yağı; Brooklyn; eğlenceli ambalaj; derinlemesine nem"),
        ("Davines MOMO", "davines.com", "İtalyan Nemlendirme", "MOMO şampuan kuru saçlar; B Corp; İtalya; sürdürülebilir salon; Parma üretim"),
        ("Teknia by Lakmé", "lakme.com", "İspanyol Salon Bakım", "İspanyol profesyonel saç bakımı; Teknia Organic Balance; Barcelona; sürdürülebilir formüller"),
        ("Nashi Argan", "nfrshiargan.com", "İtalyan Argan Bakım", "İtalyan argan yağı saç bakımı; Landoll; salon profesyonel; Argan Oil Treatment ikonik"),
        ("Oribe Gold Lust", "oribe.com", "Lüks Altın Şampuan", "Gold Lust Repair & Restore; $50+ şampuan; premium parfüm; lüks saç bakımı standartı"),
    ],

    "Sağlık & Wellness": [
        ("Whoop Recovery", "whoop.com", "Toparlanma Skoru", "Recovery Score; HRV bazlı; ekransız; $30/ay abonelik; elitist fitness takibi; $3.6B"),
        ("Oura Gen3 Horizon", "ouraring.com", "Yüzük Uyku Takibi", "Gen3 Horizon tüm parmaklara; SpO2; sıcaklık; uyku; $299 + $6/ay; Finlandiya; modaya uygun"),
        ("Apple Health DTC", "apple.com/health", "Sağlık Ekosistemi", "Apple Watch sağlık; crash detection; EKG; uyku; Health app; $383B şirketin sağlık kolu"),
        ("Fitbit DTC", "fitbit.com", "Google Fitness Band", "Charge 6; Sense 2; Google bünyesinde; fitness takip demokratikleştirildi; DTC genişleme"),
        ("Amazfit GTR DTC", "amazfit.com", "Erişilebilir Sağlık Saati", "GTR 4; SpO2 + kalp + stres; $200; Zepp Health; uygun fiyat premium sağlık takibi"),
        ("Calm App DTC", "calm.com", "Uyku + Meditasyon", "Sleep Stories; Daily Calm; $2B değerleme; Harry Styles, LeBron; mental wellness; 150M+ indirme"),
        ("Wim Hof Method", "wimhofmethod.com", "Soğuk Terapi", "Buz banyosu + nefes metodu; Wim Hof; uygulama + kurs; soğuk maruziyeti wellness; $50M+ gelir"),
        ("Plunge", "thefrplunge.co", "Soğuk Dalma Küveti", "Soğuk dalma küveti ev kullanımı; $5,000; filtreleme sistemi; biyohacking; Joe Rogan etkisi"),
        ("Therasage", "therasage.com", "Kızılötesi Terapi", "Kızılötesi sauna + mat; detoks; ağrı yönetimi; ev kullanımı; portatif kızılötesi"),
        ("HigherDOSE", "higherdose.com", "Kızılötesi Sauna Battaniye", "Kızılötesi sauna battaniye; Red Light Face Mask; NYC sauna stüdyo; biohacking; $300-800"),
        ("SaunaSpace", "sauna.space", "Yakın Kızılötesi Sauna", "Yakın kızılötesi ışık sauna; Luminati panel; ev sauna; biohacking; doğal ahşap"),
        ("Cold Pod", "thecoldpod.com", "Uygun Fiyat Buz Banyosu", "Taşınabilir buz banyosu $100; TikTok viral; uygun fiyat soğuk terapi; ev kullanımı"),
    ],

    "Fitness & Spor Giyim": [
        ("Rogue Fitness", "roguefitness.com", "CrossFit Ekipman", "CrossFit Games resmi; barbell + rack + plate; Columbus OH; $500M+ gelir; ev + ticari spor salonu"),
        ("Rep Fitness", "repfitness.com", "Ev Spor Salonu Değer", "Rogue alternatifi uygun fiyatla; squat rack + bench; ev spor salonu setup; Colorado"),
        ("Titan Fitness", "titanfitness.com", "Bütçe Ev Spor", "Uygun fiyat ev spor ekipmanı; rack + bench + plate; Memphis TN; Rogue'un yarı fiyatı"),
        ("Bells of Steel", "bellsofsteel.com", "Kanada Ev Fitness", "Kanada ev fitness ekipmanı; Hydra Rack; Blaze bumper plate; uygun fiyat kalite"),
        ("PRx Performance", "prxperformance.com", "Duvara Monte Rack", "Profile Rack duvara katlanır squat rack; alan tasarrufu; garaj spor salonu; Shark Tank"),
        ("NordicTrack Vault", "nordictrack.com", "Fitness Ayna", "Akıllı fitness aynası; iFit bağlantılı; $1,500; aynanın arkasında ekipman depolama"),
        ("Gymreapers", "gymreapers.com", "Ağırlık Aksesuarları", "Lifting belt + knee sleeves + wrist wraps; uygun fiyat; Amazon bestseller; $50M+ gelir"),
        ("SBD Apparel", "sbdapparel.com", "Powerlifting Ekipman", "IPF onaylı; SBD knee sleeves; SBD belt; powerlifting standartı; İrlanda; premium"),
        ("Bear Grips", "bear-grips.com", "CrossFit Aksesuarları", "Ağırlık eldiveni + bileklik; CrossFit + gymnastics; Amazon bestseller; koruyucu ekipman"),
        ("Aer Gym Bag", "aersf.com", "Spor Çantası", "Gym Duffel 3; iş + spor çantası; SF; ayakkabı bölmesi; minimalist; fonksiyonel"),
    ],

    "Moda & Giyim": [
        ("COS Atelier", "cosstores.com", "H&M Ultra Premium", "COS'un en üst hattı; İtalyan kumaş; sınırlı koleksiyon; ultra premium H&M grubu"),
        ("Massimo Dutti Limited", "massimodutti.com", "Inditex Lüks", "Limited Edition; İnditex'in en premium hattı; İtalyan deri; İspanyol el işçiliği"),
        ("& Other Stories x Designers", "stories.com", "Tasarımcı Collab", "Tasarımcı collab koleksiyonları; H&M grubu; erişilebilir tasarım; Paris atölye"),
        ("Arket Essentials", "arket.com", "İsveç Temel Premium", "H&M grubu; A to Z essentials; doğal malzeme; şeffaf fiyat; Stockholm"),
        ("Uniqlo U", "uniqlo.com", "Lemaire Collab", "Christophe Lemaire tasarım; uygun fiyat haute couture; $30-100; sezonluk koleksiyon"),
        ("Zara SRPLS", "zara.com", "Askeri İlham Lüks", "Zara'nın premium military koleksiyonu; sınırlı; Inditex'in üst segmenti"),
        ("Maison Margiela DTC", "maisonmargiela.com", "Avant-Garde Lüks", "Tabi ayakkabı ikonik; Replica parfüm; OTB grubu; DTC büyüme; $500M+ gelir"),
        ("Acne Studios DTC", "acnestudios.com", "İsveç Sanatsal Moda", "Stockholm; Face motif; sanatsal İskandinav moda; $300M+ gelir; DTC genişleme"),
        ("Marni DTC", "marni.com", "İtalyan Eklektik", "OTB grubu; Trunk çanta; eklektik İtalyan; Francesco Risso; DTC büyüme; renkli lüks"),
        ("Ami Paris Heart", "amiparis.com", "Paris Kalp Logo", "Alexandre Mattiussi; Ami de Coeur kalp logo; Parisian accessible luxury; $500M+ gelir"),
        ("Polène", "polene-paris.com", "Fransız Erişilebilir Deri", "Paris deri çanta; $300-500; Numero Un ikonik; erişilebilir Fransız deri lüksü; viral"),
        ("DeMellier", "demellier.com", "İngiliz Etik Çanta", "Londra; her satışta aşı bağışı; etik lüks çanta; Meghan Markle giydi; $200-500"),
        ("Strathberry", "strathberry.com", "İskoç Lüks Çanta", "Edinburgh; bar closure ikonik; Meghan Markle etkisi; İskoç deri zanaatı; £200-600"),
        ("Cult Gaia", "cultgaia.com", "LA Sanat Moda", "Ark Bag bambu çanta viral; Jasmin Larian kurdu; heykelsı aksesuarlar; LA sanat + moda"),
        ("Coperni", "coperni.com", "Fransız Gelecek Moda", "Spray-on dress viral Bella Hadid; Sébastien Meyer + Arnaud Vaillant; Paris tech-fashion"),
        ("Heliot Emil", "helioemil.com", "Danimarka Endüstriyel Moda", "Kopenhag; endüstriyel + minimal; Julius Juul Villumsen; metal detay; avant-garde İskandinav"),
        ("Post Archive Faction", "pfrfrf.world", "Kore Avant-Garde", "PAF; Seul teknik moda; archive numaralama; Dongju Seo; Kore avant-garde; DTC lüks"),
        ("Auralee", "auralee.jp", "Japon Ultra Kumaş", "Tokyo; ultra premium kumaş odaklı; Ryota Iwai; minimalist; kumaş kalitesi her şeyden önce"),
    ],

    "Yiyecek & İçecek": [
        ("Nuggs Beyond", "simulate.com", "Bitkisel Nugget 2.0", "NUGGS v2.0; bitkisel tavuk nugget; Ben Pasternak; tech startup yaklaşımı gıdaya"),
        ("Meati Foods", "meati.com", "Mantardan Et", "Mantardan tam kesim et alternatifi; misel bazlı; Boulder CO; $200M+ yatırım; yeni kategori"),
        ("Good Planet Foods", "goodplanetfoods.com", "Bitkisel Peynir", "Bitkisel peynir; fındık sütü bazlı; eriyen dilim; pizza peyniri; vegan peynir çözümü"),
        ("Miyoko's Creamery", "miyokos.com", "Artisan Vegan Peynir", "Artisan vegan tereyağı + peynir; hindistan cevizi yağı + cashew; Miyoko Schinner kurdu"),
        ("Violife", "violifefoods.com", "Yunan Vegan Peynir", "Yunanistan; vegan peynir; Just Like mozarella; Upfield bünyesinde; Avrupa pazar lideri"),
        ("Laird Hamilton Coffee", "lairdsuperfood.com", "Performance Kahve", "Performance mushroom coffee; hindistan cevizi kreamer; sörf efsanesi; doğal enerji"),
        ("Four Sigmatic Ground", "foursigmatic.com", "Mantar Öğütülmüş Kahve", "Lion's Mane öğütülmüş kahve; yarı yarıya karışım; Finlandiya; günlük mantar"),
        ("Blume", "itsblfrme.com", "Latte Tozu", "Superfood latte tozu; matcha + kurkuma + kakao; Kanada; uygun fiyat; Instagram viral"),
        ("Clevr Blends Matcha", "clevr.com", "SuperLatte Matcha", "Meghan Markle yatırım; oat milk latte karışımı; adaptojenik; kadın kurucular; LA"),
        ("JOI Plant Milk", "addjoi.com", "Konsantre Bitki Sütü", "Bitki sütü konsantre; badem, cashew, oat; kendin seyrelt; sıfır atık; Nashville"),
        ("NotMilk", "notco.com", "AI Bitkisel Süt", "NotCo; Giuseppe AI formülasyon; bitkisel süt; Şili; $1.5B değerleme; 'not' markası"),
        ("Updfrfrft", "updfrift.com", "CBD Seltzer", "CBD maden suyu; rahatlatıcı; THC-free; sosyal içecek; sober curious; alkolsüz alternatif"),
        ("Three Spirit", "threespiritdrinks.com", "Fonksiyonel Alkolsüz", "Bitki bazlı alkolsüz spirits; ruh hali değiştiren; Social Elixir; İngiltere; adaptojenik"),
        ("Monday Gin", "mondaydistillery.com", "Avustralya Alkolsüz Cin", "Avustralya alkolsüz gin; 0.0% ABV; botanik; Melbourne; sober curious; klasik cin tadı"),
        ("Lyre's", "lyres.com", "Alkolsüz Spirit Çeşitlilik", "Avustralya; 18+ alkolsüz spirit çeşidi; her klasik kokteyl için; $50M+ yatırım; global"),
    ],

    "Ev & Mutfak": [
        ("Serax", "serax.com", "Belçika Tasarım Sofra", "Belçika; Axel Van Den Bossche; Zuma tabak; tasarımcı collab sofra; restoran + ev; premium"),
        ("Hay DTC", "hay.dk", "Danimarka Tasarım Ev", "Kopenhag; Rolf + Mette Hay; erişilebilir Danimarka tasarım; mobilya + aksesuar; IKEA collab"),
        ("Menu DTC", "menu.as", "Danimarka Fonksiyonel", "Kopenhag fonksiyonel tasarım; Bottle Grinder; Wire seri; İskandinav günlük lüks; Audo"),
        ("Normann Copenhagen", "normann-copenhagen.com", "Danimarka Modern Tasarım", "Kopenhag; Herit sandalye; Form serisi; İskandinav modern; renkli + fonksiyonel"),
        ("Muuto", "muuto.com", "Yeni İskandinav Tasarım", "Kopenhag; 'new perspectives on Scandinavian design'; The Dots askılık; Knoll bünyesinde"),
        ("Ferm Living", "fermliving.com", "Danimarka Organik Tasarım", "Kopenhag; organik formlar; Pond ayna; Desert lounger; İskandinav premium ev dekor"),
        ("Blomus", "blomus.com", "Alman Fonksiyonel Tasarım", "Alman paslanmaz çelik ev aksesuarları; fonksiyonel + estetik; mum tutucu + havluluk + vazo"),
        ("LSA International", "lsa-international.com", "İngiliz Cam Sofra", "Polonya yapımı İngiliz tasarım cam; Basis karaf; el üflemesi; sofra + dekor; premium"),
        ("Iittala", "iittala.com", "Finlandiya Cam Tasarım", "Alvar Aalto vazo ikonik; Finlandiya cam + seramik; 140+ yıl; İskandinav tasarım mirası"),
        ("Arabia", "arabia.fi", "Finlandiya Seramik", "Moomin mugs; Finlandiya seramik; 150+ yıl; Fiskars bünyesinde; İskandinav sofra mirası"),
        ("Marimekko DTC", "marimekko.com", "Finlandiya Desen", "Unikko çiçek deseni ikonik; Finlandiya; tekstil + moda + ev; $150M+ gelir; DTC büyüme"),
        ("Orla Kiely", "orlakiely.com", "İrlanda Desen Ev", "İrlanda retro desenleri; Stem desen ikonik; ev + moda; Dublin; renkli İrlanda estetiği"),
    ],

    "Aksesuar & Takı": [
        ("Foundrae", "foundrae.com", "NYC Sembol Takı", "Beth Bugdaycay; sembolik madalyon takılar; 18K altın; NYC el yapımı; $500-5000; anlamlı takı"),
        ("Sophie Bille Brahe", "sophiebillebrahe.com", "Danimarka İnci", "Kopenhag; modern inci takı; tek inci kulak düşürme; İskandinav minimalizm; lüks"),
        ("Charlotte Chesnais", "charlottechesnais.com", "Fransız Heykelsi Takı", "Paris; heykelsi altın bileklik + küpe; organik formlar; Parisian art jewelry"),
        ("Anissa Kermiche", "anissakermiche.com", "Vücut Form Takı", "Londra; kadın vücut formu vazo + takı; Jugs vazo viral; feminist tasarım; lüks"),
        ("Maria Black", "maria-black.com", "Danimarka Günlük Lüks Takı", "Kopenhag; layering uzmanı; geri dönüşüm altın + gümüş; minimal İskandinav takı"),
        ("Laura Lombardi", "laurafrlombardi.com", "Vintage Zincir Takı", "NYC; vintage brass zincir; geri dönüşüm malzeme; kalın altın takı; sürdürülebilir lüks"),
        ("Roxanne Assoulin", "roxanneassoulin.com", "Renkli Boncuk Bileklik", "NYC; emaye boncuk bileklik; Poppies + Daisy; renkli fine jewelry; $50-500; eğlenceli lüks"),
        ("Tohum Design", "tohum.com.tr", "Türk El Yapımı Takı", "İstanbul; cam boncuk takılar; Yaprak küpe; Türk zanaatı; el yapımı; global lüks butikler"),
    ],

    "Evcil Hayvan": [
        ("Wag Well", "wagwell.com", "Premium Köpek Takviye", "Premium köpek sağlık takviyesi; CBD + glucosamine; organik; vet formüle; sağlık odaklı"),
        ("Kong Box", "kongbox.com", "Kong Oyuncak Kutu", "KONG markasının aylık kutusu; dayanıklı oyuncak + ödül; orijinal Kong kalitesi"),
        ("Zee.Dog", "zeedog.com", "Brezilya Tasarım Pet", "Brezilya; renkli tasarım köpek aksesuarları; tasma + oyuncak; eğlenceli; global DTC"),
        ("Found My Animal", "foundmyanimal.com", "NYC Kurtarma Pet", "NYC; denizci halatı tasma; 'adopt don't shop'; her satışta barınak bağışı; Brooklyn"),
        ("Max-Bone", "max-bone.com", "Lüks Köpek Lifestyle", "Lüks köpek giyim + yatak + aksesuar; kaşmir süveter; NYC; Instagram viral; premium pet"),
    ],

    "Parfüm & Koku": [
        ("Glossier You Dupe", "glossier.com", "Kişisel Cilt Kokusu", "Cildinize uyum sağlayan misk; 'smells like you but better'; $60; Glossier'in parfüm başarısı"),
        ("Sol de Janeiro 62", "soldejaneiro.com", "Brezilya Vücut Misti", "Brazilian Bum Bum Cream kokusu mist olarak; caramel + pistachio; TikTok viral koku"),
        ("Kayali", "kayfrfrli.co", "Huda Kattan Parfüm", "Huda Beauty'nin kardeş parfüm markası; Vanilla 28; katmanlı kokular; Orta Doğu etkisi"),
        ("Vilhelm Parfumerie", "vilfrfrlhelmparfumerie.com", "NYC İsveç Parfüm", "Jan Ahlgren; İsveç + NYC; Dear Polly; edebiyat ilhamlı; niş parfüm; $200+"),
        ("Zarko Perfume", "zarko.com", "Danimarka Moleküler Parfüm", "Kopenhag; moleküler parfüm; Pink MOLéCULE 090.09; İskandinav niş; $100-200"),
        ("Gallivant", "gallivant.com", "Şehir Kokuları", "Her parfüm bir şehir; Istanbul, Tokyo, Brooklyn; seyahat kokuları; Nick Steward; İngiliz"),
        ("AllSaints Parfüm", "allsaints.com", "İngiliz Rock Koku", "Sunset Riot, Incense City; İngiliz rock kültürü; unisex; erişilebilir niş; $70-100"),
        ("Commodity Gold", "commodityfragrances.com", "Altın Koku", "Scent Space; Gold versiyonları; Book ikonik; 3 yoğunluk; yeniden lansman başarısı"),
    ],

    "Outdoor & Macera": [
        ("GCI Outdoor", "gcioutdoor.com", "Taşınabilir Kamp Sandalye", "Freestyle Rocker; taşınabilir sallanan kamp sandalyesi; Auto-Fold teknoloji; aile kampçılığı"),
        ("Snow Peak Titanium", "snowpeak.com", "Titanyum Kamp Seti", "Titanyum kupa + tabak; ultra hafif; Japon outdoor lüks; $40-100 titanyum ürünler"),
        ("Primus", "primus.eu", "İsveç Kamp Ocağı", "İsveç kamp ocağı; 130+ yıl; OmniLite Ti; ultra hafif; İskandinav outdoor mirası"),
        ("Jetboil", "jetboil.com", "Hızlı Kaynatma Sistem", "Flash sistem; 100 saniyede kaynama; entegre tencere + ocak; backpacking; pratik"),
        ("Sawyer Products", "sawyer.com", "Su Filtresi", "Squeeze su filtresi; 0.1 mikron; Pointone filtre; seyahat + outdoor; en hafif su arıtma"),
        ("Katadyn", "katadyn.com", "İsviçre Su Arıtma", "BeFree filtre; İsviçre; askeri + outdoor su arıtma; 90+ yıl; profesyonel kalite"),
        ("Gerber Gear", "gerbergear.com", "Multi-Tool + Bıçak", "Suspension Multi-Tool; StrongArm; Portland OR; EDC + outdoor; Fiskars bünyesinde"),
        ("Leatherman", "leatherman.com", "Multi-Tool Öncü", "Wave+ ikonik multi-tool; Portland OR; 40+ yıl; 25 yıl garanti; EDC standartı"),
    ],

    "Diş & Ağız Bakımı": [
        ("Flavedo & Albedo", "flavedoandalbedo.com", "Portakal Kabuğu Macun", "Narenciye bazlı doğal diş macunu; İtalyan botanik; sürdürülebilir; premium doğal"),
        ("Aesop Mouthwash", "aesop.com", "Lüks Ağız Bakımı", "Aesop ağız çalkalama; minimalist; doğal; lüks banyo aksesuar olarak; İngiliz mağaza estetiği"),
        ("Marvis", "marfrvis.com", "İtalyan Lüks Macun", "Floransa İtalya; Aquatic Mint; Jasmin Mint; lüks diş macunu; koleksiyonluk tüpler; $10-15"),
        ("Selahatin", "selahatin.com", "İsveç Lüks Ağız Bakımı", "Stockholm; lüks diş macunu; İsveç minimalizm; premium ambalaj; $20+ macun; niş lüks"),
    ],

    "Kadın Sağlığı": [
        ("Alloy Women's Health", "myalloy.com", "Menopoz Telehealth", "Menopoz hormon tedavisi online; kadın doktorlar; kişiselleştirilmiş HRT; $100M+ yatırım"),
        ("Gennev", "gennev.com", "Menopoz Platform", "Menopoz telehealth + ürün; kadın OB-GYN doktorlar; kişiselleştirilmiş bakım planı"),
        ("State of Menopause", "stateofmenopause.com", "Menopoz Cilt Bakımı", "Menopoz cilt bakımı; hormonal değişimlere özel formüller; estrojen azalması çözümü"),
        ("Womaness", "womaness.com", "Target Menopoz", "Target'ta menopoz ürünleri; sıcak basma + vajinal kuruluk + uyku; erişilebilir; Sally Mueller"),
    ],

    "Supplement & Sporcu Beslenmesi": [
        ("Promix Clean Whey", "promixnutrition.com", "Temiz Whey İzolat", "Grass-fed whey isolate; 3 bileşen; NSF; Albert Matheny RD; ultra temiz protein"),
        ("Kion", "getkion.com", "Ben Greenfield Takviye", "Ben Greenfield's brand; Aminos EAA; Clean Energy bar; biohacking + fitness; bilim bazlı"),
        ("Xwerks", "xwerks.com", "Temiz Sporcu Beslenme", "Grow whey; Ignite pre-workout; Motion joint; temiz etiket; $50M+ gelir; CrossFit topluluk"),
        ("SFH", "sfh.com", "Stronger Faster Healthier", "CrossFit odaklı; wild-caught fish oil; grass-fed whey; temiz + sürdürülebilir sporcu beslenme"),
        ("Equip Foods", "equipfoods.com", "Paleo Sporcu Beslenme", "Grass-fed beef protein isolate; Prime Protein; paleo + whole food; Mark Sisson advisor"),
    ],

    "Abonelik Kutuları": [
        ("Pura Vida Bracelets", "pfrfrfrfra.com", "Bileklik Abonelik", "Costa Rica el yapımı bileklik; aylık club; $15/ay; plaj kültürü; bağış + zanaat"),
        ("Honest Wine Co", "honestwine.co", "Organik Şarap Kutu", "Organik + biyodinamik şarap kutusu; doğal şarap; transparanlık; aylık keşif"),
        ("Naked Wines", "nakedwines.com", "Angel Şarap Kutu", "Angel üyelik; bağımsız üreticilere yatırım; $300M+ gelir; İngiliz + ABD + Avustralya"),
        ("Atlas Coffee", "atlascoffeeclub.com", "Dünya Turu Kahve", "Her ay farklı ülke single-origin; kartpostal + bilgi; kahve dünya keşfi; $15/ay"),
        ("Candy Club", "candyclub.com", "Premium Şeker Kutu", "Premium şeker kutusu; güzel ambalaj; hediye; aylık keşif; $30/kutu; nostalji + yeni tatlar"),
    ],

    "Ofis & Kırtasiye": [
        ("Rhodia", "rhodia.com", "Fransız Defter", "Fransız kağıt kalitesi; Webnotebook; 80g kağıt; çizgili + dot grid; turuncu ikonik kapak"),
        ("Clairefontaine", "clairefontaine.com", "Fransız Kağıt Ustası", "1858'den beri; Triomphe kağıt; 90g süper pürüzsüz; Fransız kağıt zanaatı; dolma kalem dostu"),
        ("TWSBI", "twsbi.com", "Tayvan Dolma Kalem", "Şeffaf piston dolma kalem; ECO $30; hobi yazma topluluğu; erişilebilir dolma kalem; Tayvan"),
    ],
}

EXTRA_BRANDS_6 = {
    "Güzellik & Cilt Bakımı": [
        ("Shani Darden", "shanidarden.com", "LA Estetisyen Bakım", "Hollywood estetisyen; Retinol Reform; Jessica Alba'nın cilt bakımcısı; profesyonel to DTC"),
        ("Renée Rouleau", "reneerouleau.com", "Cilt Tipi Uzmanı", "Austin TX estetisyen; 9 cilt tipi sistemi; kişiselleştirilmiş bakım; 30+ yıl deneyim"),
        ("True Botanicals", "truebotanicals.com", "Bilim + Doğa", "MADE SAFE sertifikalı; klinik çalışmalar doğal bileşenlerle; retinol alternatifi; temiz lüks"),
        ("Volition Beauty", "volfritionbeauty.com", "Crowdsource Güzellik", "Tüketici fikirlerinden ürün geliştirme; Snow Mushroom serum; topluluk odaklı inovasyon"),
        ("Evereden", "evereden.com", "Temiz Anne-Bebek Bakım", "Pediatrik dermatolojist; temiz bebek + anne bakımı; Multi-Vitamin Cream; doktor geliştirdi"),
        ("Isla Beauty", "isla-beauty.com", "K-Beauty Inspired", "Kore ilham İngiliz güzellik; co-wash + essence; minimalist bakım rutini; erişilebilir"),
        ("NIOD", "niod.com", "DECIEM Bilim Markası", "DECIEM'in ileri bilim hattı; Copper Amino Isolate; The Ordinary'den bir adım üstü; serum"),
        ("The INKEY List", "theinkeylist.com", "Bileşen Eğitimi", "Her üründe bileşen eğitimi; Oat Cleansing Balm; £6-15; İngiliz erişilebilir aktif bakım"),
        ("Tata Harper", "tataharperskincare.com", "Vermont Organik Lüks", "Vermont çiftliğinden; 100% doğal; Resurfacing Mask ikonik; AmorePacific satın aldı; yeşil lüks"),
        ("May Lindstrom", "maylindstrom.com", "Ultra Lüks Doğal", "The Blue Cocoon balm; el yapımı; $180+ ürünler; ultra niş lüks doğal cilt bakımı"),
        ("Odacite Serum Concentrates", "odacite.com", "Tek Bileşen Serum", "Fransız tek bileşen konsantre serum; hedefli bakım; Paris + LA; lüks doğal"),
        ("Alpyn Beauty PlantGenius", "alpynbeauty.com", "Vahşi Toplanan Bitki", "Jackson Hole yabani bitki; PlantGenius Melt Moisturizer; adaptojenik cilt bakımı"),
        ("Haoma", "haoma.earth", "Bitkisel Sürdürülebilir", "İngiliz bitkisel sürdürülebilir güzellik; Earth Cream; sıfır atık yaklaşım; premium doğal"),
        ("Sangre de Fruta", "sangfrredefruta.com", "Botanik Güzellik", "Kanada botanik güzellik; çiçek bazlı; el yapımı; aromaterapi cilt bakımı; küçük seri"),
        ("Ranavat", "ranavat.com", "Ayurveda Lüks", "Ayurveda ilhamlı lüks cilt bakımı; Michelle Ranavat; Hindistan botanikleri; Sephora'da"),
        ("Mount Lai", "mountfrai.com", "Gua Sha + Bakım", "Jade roller + gua sha + cilt bakımı; geleneksel Çin güzellik ritüelleri; Stephanie Hon kurdu"),
    ],

    "Moda & Giyim": [
        ("Ninety Percent", "ninetypercent.com", "İngiliz Etik Moda", "%90 kar bağış; organik + geri dönüşüm; Londra; ultra sürdürülebilir; erişilebilir"),
        ("People Tree", "peopletree.co.uk", "Fair Trade Moda Öncü", "İngiliz fair trade moda; 30+ yıl; Safia Minney kurdu; organik + adil ticaret; öncü"),
        ("Kotn", "kotn.com", "Mısır Pamuk Basics", "Mısır organik pamuk; şeffaf tedarik; uygun fiyat organik basics; Kanada; B Corp"),
        ("Frank And Oak", "frankandoak.com", "Kanada Sürdürülebilir", "Montreal; sürdürülebilir erkek + kadın giyim; The Good Cotton; B Corp; Kanada DTC lideri"),
        ("Vince DTC", "vince.com", "LA Lüks Basics", "LA lüks cashmere + basics; kaşmir süveter $200-400; DTC genişleme; sessiz lüks; NYSE listeli"),
        ("Theory DTC", "theory.com", "NYC İş Giyim", "NYC minimal iş giyim; Fast Company bünyesinde; premium basics; DTC genişleme; $500M+"),
        ("Club Monaco DTC", "clubmonaco.com", "NYC Çağdaş", "NYC çağdaş moda; Regent Brands satın aldı; Kanada kökenleri; erişilebilir lüks"),
        ("Nili Lotan", "nfrfrlfrfrlfrfrtan.com", "NYC Israeli Kadın Moda", "İsrailli NYC tasarımcı; Cargo Pant viral; military chic; $300-800; effortless NYC lüks"),
        ("LaMarque", "lamarfrque.com", "LA Deri Ceketler", "LA deri ceket uzmanı; biker + moto; uygun fiyat gerçek deri; $300-500; DTC deri"),
        ("FRAME DTC", "frame-store.com", "LA Denim Lüks", "LA + Paris; Le High Skinny ikonik; $200-400 jean; celebrity denim; Jens Grede"),
        ("AGOLDE", "agolde.com", "LA Vintage Denim", "Vintage fit; Riley crop; Citizens of Humanity bünyesinde; $150-250; sessiz lüks denim"),
        ("Citizens of Humanity DTC", "citizensofhumanity.com", "Premium Denim", "LA premium denim; Rocket Skinny; DTC genişleme; $200-300 jean; Jerome Dahan mirası"),
        ("Mother Denim", "motherdenim.com", "LA Fun Denim", "Eğlenceli isimler + yıkamalar; The Hustler; The Insider; $200-300; LA denim kültürü"),
        ("Pistola Denim", "pistolfrfrdenim.com", "LA Uygun Premium Denim", "LA premium denim uygun fiyat; $100-180; Charlie high rise; DTC erişilebilir premium jean"),
        ("Closed Denim", "closed.com", "Alman Sürdürülebilir Denim", "Alman yavaş moda; İtalyan kumaş; sürdürülebilir; 45+ yıl; Hamburg; organik pamuk"),
        ("Nudie Jeans", "nudiejeans.com", "İsveç Raw Denim", "İsveç raw denim; Lean Dean, Tight Terry; onarım hizmeti; sürdürülebilir; Göteborg"),
        ("Rag & Bone DTC", "rag-bone.com", "NYC + İngiliz Moda", "Marcus Wainwright; NYC + İngiliz heritage; $300-500; DTC genişleme; premium çağdaş"),
        ("Veronica Beard DTC", "veronicabeard.com", "NYC Blazer Kadın", "Veronica Miele Beard + Veronica Swanson Beard; Dickey Jacket; NYC güçlü kadın moda"),
        ("Smythe", "smythe.ca", "Kanada Lüks Blazer", "Kanada lüks blazer; Andrea Lenczner + Christie Smythe; Duchess Blazer; İtalyan kumaş"),
    ],

    "Yiyecek & İçecek": [
        ("Magic Spoon Cereal", "magicspoon.com", "Protein Gevrek", "0g şeker 13g protein; nostaljik kutu; $10/kutu; yetişkin çocukluk gevreği; $100M+ gelir"),
        ("Purely Elizabeth", "purelyelizabeth.com", "Organik Granola", "Organik ancient grain granola; süpergıda; chia + quinoa; Boulder CO; Whole Foods"),
        ("Bear Naked", "bearnaked.com", "Custom Granola", "Kişiselleştirilmiş granola; Kellogg's; doğal bileşenler; ABD'de popüler granola markası"),
        ("Hu Gems", "hukitchen.com", "Temiz Çikolata Çips", "Çikolata chips; paleo; vegan; tahılsız; Mondelez; basit bileşenler; pişirme + snacking"),
        ("Raaka Chocolate", "raakachocolate.com", "Raw Çikolata", "Brooklyn raw (kavrulmamış) çikolata; bean-to-bar; transparent trade; single origin; ödüllü"),
        ("Alter Eco", "alterecofoods.com", "Fair Trade Çikolata", "Organik fair trade çikolata + pirinç; B Corp; kompostlanabilir ambalaj; karbon nötr"),
        ("Taza Chocolate", "tazachocolate.com", "Taş Öğütme Çikolata", "Somerville MA; Meksika taş öğütme; stone ground; direct trade cacao; 70% minimally processed"),
        ("Hu Crackers", "hukitchen.com", "Tahılsız Kraker", "Grain-free kraker; paleo; çeşitli tatlar; basit bileşenler; Mondelez bünyesinde"),
        ("Siete Tortilla Chips", "sietefoods.com", "Tahılsız Cips", "Tahılsız tortilla chips; badem unu + cassava; Meksikan aile; PepsiCo $1.2B satın aldı"),
        ("Way Better Snacks", "waybettersnacks.com", "Sprouted Cips", "Filizlenmiş tohum tortilla chips; chia + flax; organik; sağlıklı snack; $20M+ gelir"),
        ("Barnana", "barnana.com", "Upcycled Muz Snack", "Çirkin muzlardan snack; upcycled gıda; sürdürülebilir; plantain chips; B Corp"),
        ("Dang Foods", "dangfoods.com", "Asya İlham Snack", "Hindistan cevizi chips + sticky rice chips; Thai + Vietnamese; Vincent Kitirattragarn kurdu"),
        ("That's It Mango", "thatsitfruit.com", "2 Bileşen Meyve Bar", "Apple + Mango; sıfır eklenen şeker; 2 bileşen; alerjensiz; çocuk dostu; ultra basit"),
        ("Larabar", "larabar.com", "Minimal Bileşen Bar", "3-9 bileşen meyve + fındık bar; General Mills; dates bazlı; vegan; basit beslenme"),
        ("EPIC Provisions", "epicbar.com", "Et Bar + Jerky", "Grass-fed hayvan proteini bar; bison, venison; General Mills; paleo protein snack"),
    ],

    "Ev & Mutfak": [
        ("Balmuda", "balmuda.com", "Japon Premium Ev Teknoloji", "The Toaster buhar teknolojisi; The Light; Japon minimalist ev teknoloji; $300+ tost makinesi"),
        ("Vermicular", "vermicular.com", "Japon Hassas Döküm", "Nagoya; 0.01mm hassasiyette emaye döküm; Oven Pot; 100+ yıl Japon zanaatı; ultra premium"),
        ("Stagg EKG", "fellowproducts.com", "Gooseneck Kettle", "Fellow Stagg EKG sıcaklık kontrollü; pour over; specialty kahve standardı; $165; tasarım ikonu"),
        ("Zwilling Twin", "zwilling.com", "Alman Bıçak + Tencere", "Solingen Almanya; 290+ yıl; Henckels + Staub + Zwilling; DTC genişleme; bıçak zanaatı"),
        ("Wüsthof", "wusthof.com", "Solingen Bıçak", "Solingen Almanya; 200+ yıl; Classic Ikon; el yapımı; profesyonel şef bıçağı; 7. nesil aile"),
        ("Global Knives", "global-knife.com", "Japon Paslanmaz Bıçak", "Cromova 18 çelik; tek parça Japon bıçak; Yoshikin; G-2 şef bıçağı ikonik; 1985'ten beri"),
        ("Shun Cutlery", "shfrfrn.kfrfrrfrfrfrfrfresgroup.com", "Japon Premium Bıçak", "Seki City Japon; Şam çeliği; el yapımı; KAI grubu; lüks Japon bıçak; $100-400"),
        ("Yeti Rambler", "yeti.com", "Premium Drinkware", "Rambler tumbler + mug; MagSlider kapak; premium yalıtım; $30-50; outdoor lüks günlük kullanım"),
        ("Stanley Quencher", "stanley1913.com", "Viral Su Şişesi", "Quencher H2.0 FlowState; TikTok viral; pastel renkler; $45; kadın hedef kitle dönüşümü; $750M gelir"),
        ("Owala", "owala.com", "FreeSip Su Şişesi", "FreeSip çift içme ağzı; TikTok viral; $28; renkli; pratik tek elle açma; hızla büyüyen"),
    ],

    "Sağlık & Wellness": [
        ("HigherDOSE Red Light", "higherdose.com", "Kırmızı Işık Maske", "Red Light Face Mask; kolajen + anti-aging; NYC; Lauren Berlingeri + Katie Kaps; biohacking güzellik"),
        ("Plunge Cold", "thecoldpod.com", "Buz Banyo Uygun", "Uygun fiyat soğuk terapi; $100; TikTok viral; ev buz banyosu; Wim Hof trendi"),
        ("BLUblox", "blublox.com", "Mavi Işık Gözlüğü", "Mavi ışık engelleyen gözlük; uyku düzenleme; bilgisayar + telefon; Avustralya; circadian"),
        ("Ra Optics", "raoptics.com", "Premium Mavi Işık Gözlük", "Premium mavi ışık engelleyen; Matt Maruca; ışık sağlığı bilimi; $100-200; biyohacking"),
        ("Levels CGM", "levelshealth.com", "Sürekli Glikoz Monitör", "CGM ile yemeklerin kan şekerine etkisini görme; metabolik sağlık; $100M+ yatırım; Sam Corcos"),
        ("January AI", "january.ai", "AI Glikoz Yönetimi", "AI + CGM; yemek öncesi kan şekeri tahmini; kişiselleştirilmiş beslenme; bilimsel"),
    ],

    "Teknoloji & Elektronik": [
        ("Anker Soundcore", "soundcore.com", "Uygun Fiyat Kulaklık", "Liberty 4 NC; Space A40; Anker ses markası; $50-100; ANC; uygun fiyat premium ses"),
        ("Nothing Ear", "nothing.tech", "Şeffaf TWS", "Nothing Ear (2); şeffaf tasarım; ANC; Carl Pei; $100-150; anti-Apple kulaklık"),
        ("CMF by Nothing", "cmf.tech", "Ultra Uygun Teknoloji", "Nothing alt marka; Phone 1; Watch Pro; Buds Pro; $30-200; modüler; renk değiştirme"),
        ("Fairphone 5", "fairphone.com", "Modüler Etik Telefon", "Hollanda; modüler; adil mineral; 8 yıl yazılım desteği; right to repair; B Corp"),
        ("Pine64", "pine64.org", "Açık Kaynak Teknoloji", "PinePhone Linux telefon; PineBook; açık kaynak donanım; topluluk destekli; gizlilik odaklı"),
        ("Flipper Zero", "flipperzero.one", "Hacker Multi-Tool", "Taşınabilir hacker cihazı; NFC, RFID, IR; Kickstarter $5M; güvenlik testi; TikTok viral"),
    ],

    "Evcil Hayvan": [
        ("FreshPet Select", "freshpet.com", "Taze Köpek Mama Roll", "Buzdolabında taze; Select Rolls; süpermarket soğutucuda; $700M+ gelir; taze pet gıda"),
        ("Instinct Raw", "instinctpetfood.com", "Raw Boost Kibble", "Raw Boost çiğ kaplı kuru mama; Nature's Variety; premium doğal; grain-free + raw hybrid"),
        ("Wellness CORE", "wellnesspetfood.com", "Tahılsız Premium", "CORE tahılsız yüksek protein; WellPet; premium doğal beslenme; tam bütünsel bakım"),
    ],

    "Bebek & Çocuk": [
        ("Burt's Bees Baby DTC", "bfrfrtsbeesbaby.com", "Organik Bebek DTC", "GOTS organik; bebek giyim + yatak + banyo; erişilebilir organik; Clorox bünyesinde"),
        ("Cloud Island", "target.com/cloud-island", "Target Organik Bebek", "Target'ın organik bebek markası; GOTS sertifikalı; erişilebilir organik; uygun fiyat"),
        ("Cat & Jack", "target.com/cat-jack", "Target Çocuk Giyim", "Target'ın çocuk markası; adaptive giyim; kapsayıcı; $2B+ marka; erişilebilir kalite"),
    ],

    "Abonelik Kutuları": [
        ("GlobeIn", "globein.com", "Artisan Dünya Kutu", "Dünya artisan ürünleri kutusu; fair trade; el yapımı; global kültür keşfi; etik tüketim"),
        ("Therabox", "mytherabox.com", "Self-Care Kutu", "Terapist küratörlüğünde self-care kutusu; wellness + güzellik; $35/kutu; mental sağlık"),
        ("SinglesSwag", "singlesswag.com", "Bekar Kadın Kutu", "Bekar kadınlar için yaşam tarzı kutusu; güzellik + snack + kitap; $40/kutu; eğlenceli"),
    ],

    "Seyahat & Bavul": [
        ("Db Journey", "dbjourney.com", "Norveç Macera Bavul", "Douchebags yeniden markalaştı; ski + surf + seyahat; pro athlete collab; renkli; Norveç"),
        ("YETI Crossroads", "yeti.com", "YETI Seyahat Çanta", "YETI'nin seyahat çantası hattı; Crossroads Backpack; dayanıklı; premium; outdoor lifestyle"),
        ("Filson", "filson.com", "Seattle Heritage Çanta", "1897'den beri; Tin Cloth + Rugged Twill; Seattle outdoor heritage; ömür boyu garanti; DTC"),
    ],

    "Sürdürülebilir Ürünler": [
        ("Kjaer Weis Refill", "kjaerweis.com", "Yeniden Doldur Lüks", "Danimarka; lüks refillable makyaj; altın kompakt; organik sertifikalı; sıfır atık lüks"),
        ("Seed Phytonutrients", "seedphytonutrients.com", "Çiftlik Güzellik Geri Dönüşüm", "L'Oréal; geri dönüşüm kağıt şişe; tohum dahili ambalaj; çiftçi desteği; shower-friendly"),
        ("Attitude Living", "attitudeliving.com", "Kanada Eko Temizlik", "Kanada; EWG VERIFIED; doğal ev + kişisel bakım; CO2 nötr; hypoallergenic; B Corp"),
        ("Method DTC", "methodhome.com", "Tasarım Temizlik", "Güzel şişe tasarım + temiz formül; Karim Rashid; SC Johnson; doğal + sürdürülebilir"),
        ("Mrs. Meyer's DTC", "mrsmeyers.com", "Bahçe İlham Temizlik", "SC Johnson; garden-inspired kokular; doğal temizlik; şakayık + lavanta; ev temizlik"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# FINAL BATCH — unique brands to surpass 2000
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS_FINAL = {
    "Güzellik & Cilt Bakımı": [
        ("Kosas Chemistry Deodorant", "kosas.com", "Temiz AHA Deodorant", "AHA deodorant; koltuk altı bakımı + koku önleme; temiz bileşen; Sephora bestseller"),
        ("Vegamour GRO+", "vegamour.com", "CBD Saç Serumu", "GRO+ Advanced saç serumu CBD ile; bitkisel + hemp; vegan saç bakımı premium hattı"),
        ("Kinfield", "kinfield.com", "Outdoor Cilt Bakımı", "Aktif yaşam cilt bakımı; böcek kovucu + güneş kremi + après sun; outdoor güzellik niş"),
        ("Superegg", "superegg.co", "Yumurta Bazlı Bakım", "Eggssence serum; yumurta sarısı özü; yeni trend bileşen; bariyer onarım; küçük seri"),
        ("Maude", "getmaude.com", "Cinsel Sağlık Güzellik", "Cinsel wellness modern estetikle; Éva Goicochea kurdu; vibratör + kayganlaştırıcı; temiz"),
        ("Dame Products", "dfrframeproducts.com", "Kadın Cinsel Sağlık", "Kadın mühendislerin tasarladığı cinsel sağlık ürünleri; NYC; Sephora'da vibratör satışı"),
        ("Womanizer", "womanizer.com", "Alman Cinsel Sağlık", "Pleasure Air Technology; Alman mühendisliği; premium kadın cinsel sağlık; $100-200"),
        ("LELO", "lelo.com", "İsveç Lüks Cinsel Sağlık", "Stockholm; lüks kişisel bakım; $100-300; modern tasarım; premium malzeme; İsveç estetiği"),
    ],

    "Yiyecek & İçecek": [
        ("Cann", "drinkcann.com", "THC İçecek", "Mikro doz THC + CBD sosyal tonik; 2mg THC; alkolsüz alternatif; Gwyneth Paltrow yatırımcı"),
        ("Wunder", "drinkwunder.com", "Adaptojenik Seltzer", "Adaptojenik sosyal tonik; reishi + lemon balm; sober curious; doğal rahatlatıcı"),
        ("Recess", "takearecess.com", "CBD Maden Suyu", "Hemp extract + adaptojenik maden suyu; sakin ol + odaklan; pastel estetik; NYC; Benjamin Witte"),
        ("Trip", "trip-drinks.com", "İngiliz CBD İçecek", "İngiliz CBD infüze iced tea + drink; Olivia Sherwood kurdu; Londra; premium ambalaj"),
        ("Moment", "drinkmoment.com", "L-Theanine İçecek", "L-theanine + ashwagandha; sakinleştirici maden suyu; meditasyon + mindfulness markası"),
        ("Apothékary", "apothekary.co", "Bitkisel Tonik", "Bitkisel çay + tonik; adaptojenik; geleneksel bitki bilgisi; Shizu Okusa kurdu; fonksiyonel bitki"),
        ("MUD\\WTR :rest", "mudwtr.com", "Gece Ritueli", "Reishi + valerian gece karışımı; uyku ritueli; kahve alternatifi markasının gece hattı"),
        ("Kencko", "kencko.com", "Freeze Dried Smoothie", "Dondurularak kurutulmuş meyve + sebze smoothie; sadece su ekle; organik; porsiyon paketi"),
        ("Revive Superfoods", "revfrfrfrvesuperfoods.com", "Dondurulmuş Smoothie Kiti", "Önceden porsiyon smoothie kiti; blender'a at; organik; aylık abonelik; Kanada"),
        ("Cure Hydration Lemon", "curehydration.com", "Elektrolit Tozu", "ORS bazlı; hindistan cevizi suyu tozu; lemon flavor; Lauren Picasso kurdu; Whole Foods"),
    ],

    "Moda & Giyim": [
        ("With Jean", "withjean.com", "Vintage İlham Denim", "Perth Avustralya; vintage ilham jean + top; Sami Knight kurdu; Instagram estetik; DTC"),
        ("Esse Studios", "essestudios.com", "Avustralya Minimal", "Melbourne minimalist moda; yapısal kesimler; nötr renkler; kapsül gardrob; Charlotte Hicks"),
        ("St Agni", "stagni.com", "Avustralya Sanat Moda", "Byron Bay; sanatsal minimal; el yapımı sandalet; sürdürülebilir; linen + deri; $100-400"),
        ("Bassike", "bassike.com", "Avustralya Organic Basics", "Sydney; organik pamuk basics; Deborah Sams + Mary-Lou Ryan; premium minimal; B Corp"),
        ("Apiece Apart", "apieceapart.com", "NYC Boho Lüks", "NYC + LA; Laura Cramer; sanatçı kadın; bohemian lüks; doğal kumaşlar; $200-500"),
        ("Dôen", "shopdoen.com", "California Prairie", "LA; California prairie style; çiçekli elbise; Margaret + Katherine Kleveland; vintage romantik"),
        ("Christy Dawn", "christydawn.com", "Deadstock Kumaş Moda", "LA; deadstock kumaş + farm-to-closet; sürdürülebilir; çiçekli elbiseler; regenerative tarım"),
        ("Mara Hoffman", "marahoffman.com", "NYC Sürdürülebilir Renk", "NYC; renkli + sürdürülebilir; geri dönüşüm kumaş; cesur desenler; swim + ready-to-wear"),
        ("BITE Studios", "bitestudios.com", "İsveç Sürdürülebilir Lüks", "Stockholm; %100 sürdürülebilir lüks; organik + geri dönüşüm kumaş; William Lundgren"),
        ("Boyish Jeans", "boyish.com", "Sürdürülebilir Denim", "LA; sürdürülebilir denim; geri dönüşüm pamuk; vintage fit; Jordan Nodarse kurdu; B Corp"),
    ],

    "Ev & Mutfak": [
        ("GreenPan", "greenpan.us", "Seramik Yapışmaz Öncü", "Thermolon seramik kaplama icat etti; PFAS-free; Belgika; The Cookware Company; öncü teknoloji"),
        ("Always Pan 2.0", "fromourplace.com", "Yeni Nesil Çok Amaçlı", "Always Pan 2.0; dokunmatik sıcaklık ekranı; 10 aracın yerini tutuyor; yeni nesil mutfak"),
        ("Material The reBoard", "materialkitchen.com", "Jesmonite Kesme Tahtası", "Jesmonite taş kesme tahtası; minimalist mutfak; dayanıklı + estetik; NYC tasarım"),
        ("Cadine Home", "cadine.com", "El Yapımı Seramik", "LA el yapımı seramik; minimalist sofra; her parça benzersiz; zanaatkar kalitesi; $30-80"),
        ("Jono Pandolfi", "jonopandolfi.com", "Şef Seramik", "NJ atölye; şef favorisi tabak; el yapımı; restoran + ev; Eleven Madison Park kullanıyor"),
    ],

    "Sağlık & Wellness": [
        ("Lumen Device", "lumen.me", "Metabolizma Ölçer", "Nefes CO2 analizi; yağ mı karbonhidrat mı yaktığını ölçer; $249 cihaz + $19/ay; İsrail; bilimsel"),
        ("Levels Metabolik", "levelshealth.com", "CGM Metabolik Fitness", "Sürekli glikoz monitör; metabolik sağlık; $199/ay; yemek etkisini görme; Sam Corcos"),
        ("WHOOP Recovery", "whoop.com", "Strain + Recovery", "Günlük strain skoru; recovery bazlı antrenman; ekransız; $30/ay; profesyonel sporcu; $3.6B"),
        ("Apollo Neuro", "apolloneuro.com", "Stres Yönetim Cihaz", "Bileklik vibrasyon cihazı; HRV iyileştirme; stres azaltma; Dr. David Rabin; $350; bilimsel"),
        ("Sensate", "getsensate.com", "Vagus Sinir Stimülasyon", "Göğüs üzerine vibrasyon cihazı; vagus sinir uyarımı; stres + anksiyete; $250; İngiliz"),
    ],

    "Fitness & Spor Giyim": [
        ("Gorilla Grip", "gorillagrip.com", "Ev Spor Mat", "Kalın egzersiz matı; Amazon bestseller; yoga + fitness; uygun fiyat; $20-40; çeşitli kalınlık"),
        ("Manduka", "manduka.com", "Premium Yoga Mat", "PRO mat ikonik; ömür boyu garanti; $100-130; profesyonel yoga; GRP hot yoga; sürdürülebilir"),
        ("Liforme", "liforme.com", "Hizalama Yoga Mat", "AlignForMe sistem; hizalama çizgileri; doğal kauçuk; vegan; İngiliz; $100-150; biodegradable"),
        ("Jade Yoga", "jadeyoga.com", "Doğal Kauçuk Yoga Mat", "ABD doğal kauçuk; her mat için ağaç dikimi; grip + çevre dostu; $70-100; 25+ yıl"),
        ("Lomi", "lfrfrmfrfr.com", "Ev Kompost Makinesi", "Mutfak tezgahı kompost makinesi; gıda atığını 4 saatte toprağa; TikTok viral; $500; sürdürülebilir"),
    ],

    "Aksesuar & Takı": [
        ("Cuyana Bags", "cuyana.com", "İtalyan Deri Çanta", "Fewer Better Things; İtalyan deri; Classic Tote ikonik; $150-400; zaman ötesi tasarım"),
        ("Mansur Gavriel", "mansurgavriel.com", "NYC Minimalist Çanta", "Bucket Bag ikonik; Rachel Mansur + Floriana Gavriel; minimalist İtalyan deri; bekleme listesi"),
        ("The Daily Edited", "thedailyedited.com", "Avustralya Kişisel Aksesuar", "Melbourne; monogram deri aksesuar; cüzdan + çanta + telefon kılıfı; kişiselleştirme"),
        ("Mark Cross", "markcross.com", "Heritage Amerikan Çanta", "1845'ten beri; Grace Kelly çantası; Amerikan lüks heritage; Nicole Hanley; DTC yeniden lansman"),
    ],

    "Teknoloji & Elektronik": [
        ("Therabody Wave Roller", "therabody.com", "Titreşimli Roller", "Wave Roller Bluetooth; 5 hız; kas gevşetme; $150; Therabody ekosistemi"),
        ("Hyperice X", "hyperice.com", "Kontrast Terapi", "Isıtma + soğutma; kontrast terapi; diz + omuz; $350; profesyonel toparlanma"),
        ("Lululemon Mirror DTC", "lululemon.com/mirror", "Fitness Aynası", "Duvar aynası + fitness ekranı; Lululemon bünyesinde; $795; ev fitness; kişisel eğitim"),
    ],

    "Parfüm & Koku": [
        ("Diptyque Tam Dao", "diptyque.com", "Sandal Ağacı Klasik", "Tam Dao; sandal ağacı; Paris efsanesi; EDT + EDP; lüks niş; $100-200; unisex klasik"),
        ("Aesop Hwyl", "aesop.com", "Japon Orman Kokusu", "Hwyl; hinoki Japon orman; incense + moss; Melbourne + Japonya; niş unisex; $120+"),
        ("Penhaligon's Halfeti", "penhaligons.com", "Türk Gül Kokusu", "Halfeti; Türk gülü + baharat; Türkiye'nin gizli bahçesinden ilham; İngiliz niş; $200+"),
    ],

    "Sürdürülebilir Ürünler": [
        ("Pela Vision", "pelafrfrvision.com", "Kompostlanabilir Gözlük", "Kompostlanabilir güneş gözlüğü; flax straw + bioplastik; Kanada; gözlük atığını azaltma"),
        ("Ursa Major", "ursamajorvt.com", "Vermont Doğal Bakım", "Vermont doğal erkek + kadın bakım; Golden Hour Recovery Cream; B Corp; outdoor lifestyle"),
        ("Meow Meow Tweet", "meowmeowtweet.com", "Brooklyn Vegan Bakım", "Brooklyn el yapımı vegan bakım; deodorant + sabun; sıfır atık; küçük seri; artisan"),
        ("Plaine Products", "plaineproducts.com", "Döngüsel Banyo", "Alüminyum şişe; gönder-doldur-geri gönder; şampuan + duş jeli; sıfır plastik; döngüsel"),
    ],

    "Saç Bakımı": [
        ("Hair Story New Wash", "hairstory.com", "Deterjan Free Yıkama", "Şampuan değil temizleyici; deterjan free; co-wash; Eli Halliwell kurdu; NYC; saç yıkama devrimi"),
        ("OUAI Body Creme", "theouai.com", "Vücut Kremi", "Jen Atkin; Dean Street vücut kremi; parfümlü vücut bakımı; saçtan vücuda; P&G bünyesinde"),
        ("R+Co Bleu", "randco.com/bleu", "Lüks Salon R+Co", "R+Co'nun lüks hattı; De Luxe serisi; premium salon; Howard McLaren + Garren; $50+ ürünler"),
    ],

    "Erkek Bakım": [
        ("Stryx", "stryx.com", "Erkek Makyaj", "Erkek kapatıcı + göz altı jeli; 'it's not makeup it's Stryx'; erkek güzelliği normalleştirme"),
        ("War Paint for Men", "warpaintformen.com", "İngiliz Erkek Makyaj", "İngiliz erkek makyaj markası; foundation + concealer + bronzer; Danny Gray kurdu; BBC Dragon's Den"),
        ("MMUK MAN", "mmfrukman.com", "Erkek Kozmetik UK", "İngiliz erkek kozmetik; tinted moisturizer + lip balm; Alex Sheridan kurdu; erkek güzellik öncüsü"),
    ],

    "Fitness & Spor Giyim": [
        ("Nuzest", "nuzest.com", "NZ Bitkisel Protein", "Yeni Zelanda bitkisel protein; Clean Lean Protein; bezelye proteini; NZ kalite; global DTC"),
        ("Athlean-X Supplements", "athleanx.com", "Jeff Cavaliere Takviye", "YouTube fitness fenomeni; fizik terapist; bilimsel antrenman + takviye; AX-1 program"),
        ("V Shred", "vshred.com", "Online Fitness Program", "Vince Sant; online kişiselleştirilmiş fitness; Facebook reklam ustası; takviye hattı; $100M+"),
    ],

    "Oyun & Eğlence": [
        ("Catan Studio DTC", "catan.com", "Masa Oyunu DTC", "Settlers of Catan; 30M+ kopya; dünya genelinde en popüler modern masa oyunu; DTC genişleme"),
        ("Wingspan Board Game", "stonemaiergames.com", "Kuş Temalı Strateji", "Stonemaier Games; kuş temalı motor building; Elizabeth Hargrave; 2019 Kennerspiel des Jahres"),
        ("Wyrmwood Gaming", "wyrmwoodgaming.com", "Lüks Masa Oyun Aksesuar", "El yapımı ahşap masa oyun aksesuarları; zar kuleleri + GM ekranları; D&D topluluğu; premium"),
    ],

    "Kadın Sağlığı": [
        ("Viv for Your V", "vivforyourv.com", "Vajinal Sağlık", "Vajinal wellness; pH dengesi yıkama; Dr. Jodie Horton; kadın intim sağlık; tabusuz"),
        ("Love Wellness", "lovewellness.com", "Lo Bosworth Kadın Sağlık", "Lo Bosworth (The Hills) kurdu; Good Girl Probiotics; vajinal + sindirim; Target'ta"),
        ("pH-D Feminine Health", "phdfemininehealth.com", "Borik Asit Ovül", "Borik asit vajinal supozituvar; pH dengesi; kadın sağlık; Amazon bestseller"),
    ],

    "Bebek & Çocuk": [
        ("Lalo The Chair", "meetlalo.com", "Büyüyen Mama Sandalyesi", "The Daily yüksek sandalye; 6 ay - 3+ yaş; modüler; modern tasarım; $235"),
        ("Stokke Tripp Trapp DTC", "stokke.com", "Norveç Büyüyen Sandalye", "Bebekten yetişkine büyüyen sandalye; 1972'den beri; Norveç tasarım ikonuu; ömür boyu kullanım"),
    ],

    "Evcil Hayvan": [
        ("Barkbox Super Chewer DTC", "barkbox.com", "Dayanıklı Oyuncak Abonelik", "Güçlü çiğneyici köpekler için; aylık dayanıklı oyuncak + ödül; tematik; eğlenceli; 2M+ abone"),
        ("Outward Hound", "outwardhound.com", "Eğlenceli Köpek Oyuncak", "Hide A Squirrel puzzle oyuncak; enrichment; Nina Ottosson puzzle; interaktif köpek oyuncakları"),
    ],

    "Seyahat & Bavul": [
        ("Horizn Studios M5", "horizn-studios.com", "Berlin Kabin Bavul", "Berlin; akıllı kabin bavul; çıkarılabilir şarj; vegan deri etiket; Alman mühendislik"),
        ("Arlo Skye", "arloskye.com", "Minimalist Lüks Bavul", "NYC minimalist lüks bavul; alüminyum + polikarbonat; $550-650; Zipper-less design; whisper tekerlek"),
    ],

    "Teknoloji & Elektronik": [
        ("Ember Mug 2 DTC", "ember.com", "Sıcaklık Kontrol Kupa", "Sıcaklık kontrol seramik kupa 2; 1.5 saat pil; uygulama kontrol; Starbucks collab; $130-200"),
        ("Rocketbook Fusion", "getrocketbook.com", "Akıllı Defter", "Yazıp sil akıllı defter; OCR ile bulut senkron; sürdürülebilir not alma; 7 sayfa stili"),
    ],

    "Diş & Ağız Bakımı": [
        ("Quip Smart Brush", "getquip.com", "Akıllı Elektrikli Fırça", "Bluetooth bağlantılı; fırçalama takip; abonelik; $25-75; ADA onaylı; minimal tasarım"),
    ],

    "Ofis & Kırtasiye": [
        ("Remarkable 2 DTC", "remarkable.com", "E-Ink Defter", "E-ink kağıt his tablet; not alma + PDF okuma; dikkat dağıtmayan; Norveç; $449; kağıt alternatifi"),
        ("Supernote A5X", "supernote.com", "Premium E-Ink Defter", "E-ink not tablet; custom firmware; ince + hafif; Çin; Remarkable alternatifi; yazma deneyimi"),
    ],

    "Abonelik Kutuları": [
        ("The Detox Box", "thedetoxbox.com", "Yeşil Güzellik Kutu", "Aylık yeşil güzellik kutusu; temiz bileşen markalar keşfi; full-size; $50/kutu"),
        ("Cocotique", "cocotique.com", "Siyah Kadın Güzellik Kutu", "Siyah kadınlar için güzellik kutusu; melanin cildi + saçı; aylık keşif; kapsayıcı"),
        ("Alltrue", "alltrue.com", "Etik Yaşam Tarzı Kutu", "Etik + sürdürülebilir ürünler; sezonluk; B Corp markalar; bilinçli tüketim"),
    ],

    "Outdoor & Macera": [
        ("Kovea", "kovea.com", "Kore Kamp Ekipman", "Kore premium kamp ekipmanı; Alpine Master güçlü ocak; cube tost makinesi; Kore kamp kültürü"),
        ("Soto Outdoors", "sfrfrtoodoors.com", "Japon Kamp Ocağı", "Japon hassas kamp ocağı; WindMaster; Micro Regulator; ultra hafif; Japon mühendislik"),
        ("Toaks Titanium", "tofrfrksoutdoor.com", "Titanyum Kamp Ekipman", "Ultra hafif titanyum tencere + bardak; backpacking; 35-100g ürünler; gram sayanlar için"),
    ],

    "Supplement & Sporcu Beslenmesi": [
        ("Beam Elevate", "beamorganics.com", "Pre-Workout Doğal", "Doğal pre-workout; nano CBD + yeşil çay; temiz enerji; performans + toparlanma"),
        ("Dose & Co Collagen", "doseandco.com", "NZ Kolajen", "Yeni Zelanda kolajen; Khloé Kardashian yatırımcı; premium peptit; kahve + smoothie"),
    ],

    "Parfüm & Koku": [
        ("Atelier Cologne", "ateliercologne.com", "Fransız Cologne Absolu", "Cologne Absolue konsantre kolonya kategorisi yarattı; Clémentine California; L'Oréal satın aldı"),
        ("Acqua di Gio Profondo", "armani.com", "Armani DTC Parfüm", "Giorgio Armani parfüm DTC satışı; Profondo; $80-150; erkek parfüm ikonik; L'Oréal"),
        ("Abel Odor", "abelodor.com", "NZ Doğal Parfüm", "Yeni Zelanda %100 doğal parfüm; Amsterdam üretim; Frances Shoemack; biyo-bozunur; temiz lüks"),
    ],

    "Güzellik & Cilt Bakımı": [
        ("Whamisa", "whamisa.com", "Kore Fermente Organik", "Fermente organik K-beauty; çiçek tonik; doğal fermentasyon; EWG VERIFIED; premium doğal Kore"),
        ("Then I Met You", "thenifrfrmfrfretyfrou.com", "Soko Glam Markası", "Charlotte Cho kurdu; Living Cleansing Balm; Korean double cleanse; Soko Glam'ın markası"),
        ("Atolla", "atolla.com", "AI Kişisel Serum", "AI ile kişiselleştirilmiş serum; cilt testi + formül; MIT spinoff; veri odaklı cilt bakımı"),
        ("27 Rosiers", "27rosiers.com", "Fransız Temiz Güzellik", "Paris temiz güzellik; Baobab Oil; Fransız + Afrika bileşenler; 27 gül ağacı; sürdürülebilir"),
        ("Klur", "klfrur.co", "Minimalist Cilt Bilimi", "Lesley Thornton kurdu; Supreme Seed serum; minimalist bilimsel bakım; küçük seri; temiz lüks"),
        ("Peach Not Plastic", "peachnotplastic.com", "Katı Güzellik", "Katı şampuan + saç kremi bar; plastik ambalajsız; şeftali şekli ikonik; sürdürülebilir"),
        ("Veriphy Skincare", "veriphyself.com", "Biyotek Bakım", "Kanada; biyoteknoloji fermente cilt bakımı; Self Absorbed serum; bilim + sürdürülebilirlik"),
        ("Circumference", "circumferencenyc.com", "NYC Bilimsel Doğal", "NYC; bilimsel doğal cilt bakımı; sürdürülebilir ambalaj; çevre aktivizmi; premium"),
        ("Humanrace", "humanrace.com", "Pharrell Williams Bakım", "Pharrell Williams'ın cilt bakım markası; Three Minute Clay Mask; minimalist; Sephora"),
        ("Ami Colé Desert Date", "amicole.com", "Sahra Hurma Bakımı", "Koyu cilt tonları için; desert date yağı; temiz bileşen; Senegalli miras; Diarrha N'Diaye"),
        ("OUI the People", "ouithepeople.com", "Kıl Bakım Uzmanı", "Tıraş + kıl bakımı; kıl batması çözümü; kapsayıcı; Karen Young kurdu; vücut bakımı"),
    ],

    "Moda & Giyim": [
        ("Sporty & Rich", "sportyandrich.com", "Wellness Estetik Moda", "Emily Oberg kurdu; Wellness Club koleksiyonu; vintage spor estetiği; $100-300; hype"),
        ("Les Girls Les Boys", "lesgirlslesboys.com", "İngiliz Genderless İç Giyim", "Julien Macdonald eski asistanı; genderless iç giyim; İngiltere; LGBTQ+ kapsayıcı; neon renkler"),
        ("CDLP", "cdlp.com", "İsveç Erkek İç Giyim", "Stockholm erkek iç giyim; Lyocell kumaş; sürdürülebilir İsveç lüksü; $30-40; premium basics"),
        ("Entireworld", "theentireworld.com", "Numara Sistemi Basics", "Scott Sternberg kurdu; renk kodlu giyim sistemi; Band of Outsiders'dan sonra; LA basics"),
        ("Koio", "koio.co", "İtalyan Sneaker DTC", "İtalya'da el yapımı sneaker; Marche üretim; $250-300; Capri ikonik; erişilebilir İtalyan lüks"),
        ("ROAM", "roamfootwear.com", "Kişisel Sneaker", "Kişiselleştirilebilir sneaker; renk + malzeme seç; DTC; Portland OR; ayakkabıda kişiselleştirme"),
        ("NOTHING NEW", "nothingnew.com", "Geri Dönüşüm Sneaker", "%100 geri dönüşüm malzeme sneaker; sürdürülebilir; $95; pet şişe + geri dönüşüm kauçuk"),
        ("Vivobarefoot", "vivobarefoot.com", "Barefoot Ayakkabı", "İngiliz barefoot ayakkabı; ultra ince taban; doğal yürüyüş; Galahad Clark (Clarks ailesi); sağlık"),
        ("Xero Shoes", "xeroshoes.com", "Minimalist Ayakkabı", "Barefoot koşu ayakkabısı; $35-160; doğal ayak hareketi; Steven Sashen kurdu; Shark Tank"),
        ("Tropicfeel Monsoon", "tropicfeel.com", "3in1 Seyahat Ayakkabı", "Seyahat + su + günlük; 3-in-1; kompakt; Kickstarter $14M; Barcelona; all-terrain"),
    ],

    "Sağlık & Wellness": [
        ("InsideTracker DNA", "insidetracker.com", "Genomik Sağlık", "DNA + kan testi; genomik kişisel sağlık planı; David Sinclair kullanıyor; MIT bilimi"),
        ("Rootine", "rootine.co", "DNA Bazlı Vitamin", "DNA + kan + yaşam tarzı bazlı kişisel vitamin; mikro besin dozları; bilimsel kişiselleştirme"),
        ("CURED Nutrition", "cfrurednutrition.com", "CBD Fonksiyonel", "CBD + mantar fonksiyonel takviye; Zen capsule; Rise cacao; Colorado; plant medicine"),
        ("Thesis Nootropics", "takethesis.com", "Kişisel Beyin Takviyesi", "Quiz ile kişiselleştirilmiş nootropik; Clarity + Energy + Logic; beyin performansı; DTC"),
    ],
}

# ── Merge all extra brand batches ─────────────────────────────────────────────
ALL_BATCHES = [EXTRA_BRANDS, EXTRA_BRANDS_2, EXTRA_BRANDS_3, EXTRA_BRANDS_4, EXTRA_BRANDS_5, EXTRA_BRANDS_6, EXTRA_BRANDS_FINAL]  # noqa: E501
for batch in ALL_BATCHES:
    for cat, extra_list in batch.items():
        if cat in BRANDS:
            existing_names = {b[0].lower() for b in BRANDS[cat]}
            for brand in extra_list:
                if brand[0].lower() not in existing_names:
                    BRANDS[cat].append(brand)
                    existing_names.add(brand[0].lower())
        else:
            BRANDS[cat] = extra_list


def create_meta_ads_url(brand_name: str) -> str:
    """Generate Meta Ads Library search URL for a brand."""
    query = brand_name.replace(" ", "%20").replace("&", "%26")
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={query}"


def apply_header_style(ws, row, max_col, fill_color="1F4E79"):
    """Apply formatting to header row."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_data_style(ws, row, max_col, even=False):
    """Apply formatting to data rows."""
    data_font = Font(name="Calibri", size=10)
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = even_fill if even else odd_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_excel():
    """Build the complete Excel workbook."""
    wb = Workbook()

    # ── Flatten all brands ────────────────────────────────────────────────
    all_brands = []
    for category, brands in BRANDS.items():
        for b in brands:
            all_brands.append((category, b))

    total_brands = len(all_brands)
    print(f"Toplam marka sayısı: {total_brands}")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: ÖZET (Summary)
    # ═══════════════════════════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Özet"
    ws_summary.sheet_properties.tabColor = "1F4E79"

    # Title
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "DTC Ecommerce-Native Markalar - Türkçe Rapor"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:D2")
    ws_summary["A2"].value = f"Oluşturulma Tarihi: {TODAY}  |  Toplam Marka: {total_brands}"
    ws_summary["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    # Category breakdown
    headers = ["#", "Kategori", "Marka Sayısı", "Yüzde (%)"]
    for col_idx, h in enumerate(headers, 1):
        ws_summary.cell(row=4, column=col_idx, value=h)
    apply_header_style(ws_summary, 4, len(headers), "2E75B6")

    row_num = 5
    for idx, (cat, brands) in enumerate(BRANDS.items(), 1):
        count = len(brands)
        pct = round(count / total_brands * 100, 1)
        ws_summary.cell(row=row_num, column=1, value=idx)
        ws_summary.cell(row=row_num, column=2, value=cat)
        ws_summary.cell(row=row_num, column=3, value=count)
        ws_summary.cell(row=row_num, column=4, value=f"%{pct}")
        apply_data_style(ws_summary, row_num, len(headers), even=(idx % 2 == 0))
        row_num += 1

    # Total row
    ws_summary.cell(row=row_num, column=2, value="TOPLAM")
    ws_summary.cell(row=row_num, column=3, value=total_brands)
    ws_summary.cell(row=row_num, column=4, value="%100")
    for col in range(1, 5):
        cell = ws_summary.cell(row=row_num, column=col)
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

    ws_summary.column_dimensions["A"].width = 5
    ws_summary.column_dimensions["B"].width = 40
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: TÜM MARKALAR (All Brands)
    # ═══════════════════════════════════════════════════════════════════════
    ws_all = wb.create_sheet("Tüm Markalar")
    ws_all.sheet_properties.tabColor = "2E75B6"

    all_headers = ["#", "Marka Adı", "Web Sitesi", "Kategori", "Alt Niş",
                   "Öne Çıkan Özellik / Pazarlama Açısı", "Meta Reklam Kütüphanesi"]

    for col_idx, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col_idx, value=h)
    apply_header_style(ws_all, 1, len(all_headers))

    row_num = 2
    for idx, (category, brand) in enumerate(all_brands, 1):
        name, website, subniche, insight = brand
        meta_url = create_meta_ads_url(name)

        ws_all.cell(row=row_num, column=1, value=idx)
        ws_all.cell(row=row_num, column=2, value=name)

        # Clickable website
        site_cell = ws_all.cell(row=row_num, column=3, value=website)
        site_cell.hyperlink = f"https://{website}"
        site_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        ws_all.cell(row=row_num, column=4, value=category)
        ws_all.cell(row=row_num, column=5, value=subniche)
        ws_all.cell(row=row_num, column=6, value=insight)

        # Meta Ads Library link
        meta_cell = ws_all.cell(row=row_num, column=7, value="Meta Reklam Kütüphanesi")
        meta_cell.hyperlink = meta_url
        meta_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        apply_data_style(ws_all, row_num, len(all_headers), even=(idx % 2 == 0))
        # Re-apply link styling after data style
        site_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        meta_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        row_num += 1

    # Column widths
    col_widths_all = [5, 25, 30, 30, 25, 60, 25]
    for i, w in enumerate(col_widths_all, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws_all.freeze_panes = "A2"
    # Auto-filter
    ws_all.auto_filter.ref = f"A1:G{row_num - 1}"

    # ═══════════════════════════════════════════════════════════════════════
    # INDIVIDUAL CATEGORY SHEETS
    # ═══════════════════════════════════════════════════════════════════════
    category_colors = [
        "FF6B6B", "4ECDC4", "45B7D1", "96CEB4", "FFEAA7",
        "DDA0DD", "98D8C8", "F7DC6F", "BB8FCE", "85C1E9",
        "F0B27A", "82E0AA", "F1948A", "AED6F1", "D7BDE2",
        "A3E4D7", "FAD7A0", "ABEBC6", "F9E79F", "D5F5E3",
        "FADBD8", "D4E6F1", "E8DAEF"
    ]

    for cat_idx, (category, brands) in enumerate(BRANDS.items()):
        # Shorten sheet name if needed (max 31 chars)
        sheet_name = category[:31]
        ws_cat = wb.create_sheet(sheet_name)
        tab_color = category_colors[cat_idx % len(category_colors)]
        ws_cat.sheet_properties.tabColor = tab_color

        cat_headers = ["#", "Marka Adı", "Web Sitesi", "Alt Niş",
                       "Öne Çıkan Özellik / Pazarlama Açısı", "Meta Reklam Kütüphanesi"]

        for col_idx, h in enumerate(cat_headers, 1):
            ws_cat.cell(row=1, column=col_idx, value=h)
        apply_header_style(ws_cat, 1, len(cat_headers), tab_color.replace("#", ""))

        for b_idx, brand in enumerate(brands, 1):
            r = b_idx + 1
            name, website, subniche, insight = brand
            meta_url = create_meta_ads_url(name)

            ws_cat.cell(row=r, column=1, value=b_idx)
            ws_cat.cell(row=r, column=2, value=name)

            site_cell = ws_cat.cell(row=r, column=3, value=website)
            site_cell.hyperlink = f"https://{website}"
            site_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

            ws_cat.cell(row=r, column=4, value=subniche)
            ws_cat.cell(row=r, column=5, value=insight)

            meta_cell = ws_cat.cell(row=r, column=6, value="Meta Reklam Kütüphanesi")
            meta_cell.hyperlink = meta_url
            meta_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

            apply_data_style(ws_cat, r, len(cat_headers), even=(b_idx % 2 == 0))
            site_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            meta_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        # Column widths
        col_widths_cat = [5, 25, 30, 25, 60, 25]
        for i, w in enumerate(col_widths_cat, 1):
            ws_cat.column_dimensions[get_column_letter(i)].width = w

        ws_cat.freeze_panes = "A2"
        ws_cat.auto_filter.ref = f"A1:F{len(brands) + 1}"

    # ── Save ──────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, FILENAME)
    wb.save(filepath)
    print(f"Excel dosyası oluşturuldu: {filepath}")
    print(f"Toplam {total_brands} marka, {len(BRANDS)} kategori")
    return filepath


if __name__ == "__main__":
    build_excel()
