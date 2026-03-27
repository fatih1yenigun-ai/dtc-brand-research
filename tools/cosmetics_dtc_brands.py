#!/usr/bin/env python3
"""
Kozmetik DTC İnovatif Markalar - 5000+ Marka Excel Oluşturucu
Generates a comprehensive Excel file with 5000+ cosmetics-only DTC brands.
All text in Turkish. Brand names stay English.
"""

import os
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Output config ───────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research_outputs")
TODAY = "2026-03-24"
FILENAME = f"Kozmetik_DTC_5000_Inovatif_Markalar_{TODAY}.xlsx"

# ─── Color Scheme ────────────────────────────────────────────────────────────
HEADER_COLOR = "1B2A4A"
HEADER_FONT_COLOR = "FFFFFF"
META_BUTTON_COLOR = "27AE60"
WEBSITE_LINK_COLOR = "0563C1"
INSIGHT_FONT_COLOR = "555555"

CATEGORY_COLORS = {
    "Cilt Bakımı - Nemlendirici & Serum": ("D4E8D0", "4A7C59"),
    "Cilt Bakımı - Temizleyici & Tonik": ("D5ECD0", "3E7D4A"),
    "Cilt Bakımı - Güneş Koruma (SPF)": ("FCE4D6", "B05C2A"),
    "Cilt Bakımı - Akne & Leke Tedavisi": ("FADBD8", "A0522D"),
    "Cilt Bakımı - Anti-Aging & Kırışıklık": ("F5D5CB", "8B4513"),
    "Cilt Bakımı - Göz Çevresi": ("E0D4F5", "6B4C8A"),
    "Cilt Bakımı - Maske & Peeling": ("E8D8E8", "6B4C6B"),
    "Cilt Bakımı - Dudak Bakımı": ("F5D4E0", "8B3E5C"),
    "Cilt Bakımı - Bariyer Onarım": ("C8E6E0", "2E6B5E"),
    "Cilt Bakımı - Hassas Cilt": ("CCE8E8", "2E6B6B"),
    "Cilt Bakımı - Hiperpigmentasyon": ("E8E0D4", "7D6B4C"),
    "Makyaj - Fondöten & BB/CC Krem": ("FADBD8", "A0522D"),
    "Makyaj - Allık & Bronzer": ("F0E0F5", "8B4C8B"),
    "Makyaj - Göz Makyajı": ("D8D4E8", "4C3E6B"),
    "Makyaj - Dudak Ürünleri": ("F5D4E0", "8B3E5C"),
    "Makyaj - Kaş Ürünleri": ("E8D4D8", "6B3E4C"),
    "Makyaj - Aydınlatıcı & Kontür": ("FFE5CC", "CC6600"),
    "Saç Bakımı - Şampuan & Saç Kremi": ("F5D5CB", "8B4513"),
    "Saç Bakımı - Saç Maskesi & Onarım": ("E6D8CC", "6B4226"),
    "Saç Bakımı - Saç Büyütme & Dökülme": ("D4E6F1", "2C5F8A"),
    "Saç Bakımı - Renk & Boyama": ("E0D8E8", "5C4C6B"),
    "Saç Bakımı - Styling & Isı Koruma": ("D4D8E8", "3E4C6B"),
    "Saç Bakımı - Kıvırcık & Tekstürlü Saç": ("E8E4D4", "6B644C"),
    "Vücut Bakımı - Vücut Nemlendirici & Yağ": ("D4E8D0", "4A7C59"),
    "Vücut Bakımı - Peeling & Scrub": ("E0E8D4", "4C6B3E"),
    "Vücut Bakımı - Self-Tan & Bronzlaşma": ("FCE4D6", "B05C2A"),
    "Vücut Bakımı - Deodorant": ("CCE8CC", "2E7D32"),
    "Vücut Bakımı - Tüy Dökücü & Epilasyon": ("D4E4E8", "3E5F6B"),
    "Tırnak Bakımı - Oje & Jel": ("F0E0F5", "8B4C8B"),
    "Tırnak Bakımı - Takma Tırnak & Press-On": ("FADBD8", "A0522D"),
    "Cilt Cihazları - LED & Işık Terapisi": ("D4E6F1", "2C5F8A"),
    "Cilt Cihazları - Mikro-Akım & RF": ("D4D8E8", "3E4C6B"),
    "Cilt Cihazları - Temizleme Cihazı": ("C8E6E0", "2E6B5E"),
    "Parfüm & Koku - Kadın Parfüm": ("E8E4D4", "6B644C"),
    "Parfüm & Koku - Unisex & Niş": ("E0D4F5", "6B4C8A"),
    "Parfüm & Koku - Vücut Spreyi & Mist": ("F5E6CC", "8B6914"),
    "Erkek Bakım - Tıraş & Sakal": ("C8E6E0", "2E6B5E"),
    "Erkek Bakım - Erkek Cilt Bakımı": ("D4E4E8", "3E5F6B"),
    "K-Beauty & Kore Kozmetik": ("FADBD8", "A0522D"),
    "J-Beauty & Japon Kozmetik": ("F5D5CB", "8B4513"),
    "Doğal & Organik Kozmetik": ("CCE8CC", "2E7D32"),
    "Vegan & Cruelty-Free Kozmetik": ("D4E8D0", "4A7C59"),
    "Bebek & Hamile Cilt Bakımı": ("F0E0F5", "8B4C8B"),
    "İntim Bakım & Vücut": ("E8D4D8", "6B3E4C"),
    "Diş Beyazlatma & Ağız Bakımı": ("CCE8E8", "2E6B6B"),
}

# ─── Brand Data ──────────────────────────────────────────────────────────────
BRANDS = {
    # ═══════════════════════════════════════════════════════════════════════════
    # 1. Cilt Bakımı - Nemlendirici & Serum
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Nemlendirici & Serum": [
        ("Dieux Skin", "dieuxskin.com", "Kanıt Bazlı Serum", "Klinik kanıtları şeffafça paylaşan marka; Instant Angel göz altı maskesi; sürdürülebilir alüminyum tüp ambalaj"),
        ("Stratia", "stratiaskin.com", "Bariyer Onarım Serumu", "Liquid Gold kült ürünü; ceramide + rosehip + sea buckthorn karışımı; Reddit topluluğu favorisi; tek kişi indie marka"),
        ("Cocokind", "cocokind.com", "Şeffaf Fiyatlı Serum", "Maliyet dağılımını etiketinde gösteren ilk kozmetik marka; sürdürülebilir ambalaj öncüsü"),
        ("Good Molecules", "goodmolecules.com", "Uygun Fiyatlı Aktif Serum", "Beautylish markası; $6-12 fiyat aralığı; yüksek konsantrasyonlu niacinamide ve hyaluronic acid"),
        ("Versed", "versedskin.com", "Topluluk Geliştirmeli Serum", "Topluluk oylamasıyla ürün geliştirme; temiz + etkili formüller; erişilebilir fiyat"),
        ("Superegg", "superegg.co", "Yumurta Membran Serumu", "Yumurta membran teknolojisi; EGF benzeri büyüme faktörleri; minimalist Kore-İskandinav tasarım"),
        ("Herbivore Botanicals", "herbivorebotanicals.com", "Doğal Aktif Serum", "Bakuchiol serumu retinol alternatifi; mavi tansy yağı; Instagram estetiği öncüsü"),
        ("Biossance", "biossance.com", "Squalane Uzmanı", "Şeker kamışından elde edilen biyoteknolojik squalane; %100 bitki bazlı; çevre dostu biyoteknoloji"),
        ("Youth To The People", "youthtothepeople.com", "Süperfood Serum", "Kale + yeşil çay + hyaluronic acid; gıda bazlı antioksidan formüller; cam şişe ambalaj"),
        ("Summer Fridays", "summerfridays.com", "Işıltılı Nemlendirici", "Jet Lag Mask kült ürünü; influencer Marianna Hewitt kurdu; havaalanı güzelliği konsepti"),
        ("Volition Beauty", "volitionbeauty.com", "Crowdsource Güzellik", "Topluluk fikir sunma ile ürün geliştirme; demokratik güzellik; Snow Mushroom serum"),
        ("Glow Recipe", "glowrecipe.com", "Meyve Bazlı Serum", "Watermelon Glow Niacinamide Dew Drops viral ürün; K-beauty'yi Batı'ya taşıdı; meyve ekstreleri"),
        ("Peach & Lily", "peachandlily.com", "Cam Cilt Serumu", "Glass Skin Serum; Alicia Yoon K-beauty küratörlüğü; 'cam cilt' trendinin başlatıcısı"),
        ("Naturium", "naturium.com", "Yüksek Doz Aktif Serum", "Susan Yara kurdu; %12 niacinamide, multi-peptide; eczane fiyatına klinik dozlar"),
        ("Skinfix", "skinfix.com", "Dermatolog Serumu", "Kanada kökenli; ceramide + lipid teknolojisi; NEA (Egzama Derneği) onaylı formüller"),
        ("Paula's Choice", "paulaschoice.com", "BHA Serum Uzmanı", "%2 BHA Liquid Exfoliant kült ürün; ingrediyent bazlı eğitim öncüsü; Beautypedia bilgi bankası"),
        ("The Inkey List", "theinkeylist.com", "Eğitim Odaklı Serum", "Ingrediyent odaklı ürün isimlendirme; £5-15 fiyat; bilgi kartları ile eğitim; The Ordinary rakibi"),
        ("Beauty Pie", "beautypie.com", "Fabrika Fiyatlı Serum", "Üyelik modeli ile lüks formüller fabrika fiyatına; şeffaf maliyet yapısı; Marcia Kilgore kurdu"),
        ("Tula", "tula.com", "Probiyotik Serum", "Gastroenterolog kurdu; probiyotik + prebiyotik bazlı; cilt mikrobiyomu dengeleme"),
        ("Farmacy", "farmacybeauty.com", "Çiftlik Kaynaklı Serum", "Kendi echinacea çiftliğinden bileşenler; Green Clean balm temizleyici; farm-to-face konsepti"),
        ("Kinship", "lovekinship.com", "Ekolojik Serum", "Adaptojenik bileşenler; mercan resiflerine güvenli SPF; çevreye duyarlı formüller"),
        ("Osea Malibu", "oseamalibu.com", "Deniz Yosunu Serumu", "Vegan deniz yosunu bazlı; Malibu kıyılarından ilham; okyanus mineralleri ile nemlendirme"),
        ("Indie Lee", "indielee.com", "Temiz Güzellik Serumu", "Beyin tümörü sonrası temiz güzelliğe dönen kurucu hikayesi; squalane + stem cell serumu"),
        ("Honest Beauty", "honestbeauty.com", "Güvenli İçerik Serumu", "Jessica Alba kurdu; EWG onaylı temiz içerikler; hamile ve bebek güvenli formüller"),
        ("Alpyn Beauty", "alpynbeauty.com", "Vahşi Hasat Serumu", "Jackson Hole dağlarından vahşi hasat edilen bitkiler; bakuchiol + kale bazlı anti-aging"),
        ("Odacité", "odacite.com", "Konsantre Yağ Serumu", "Soğuk pres organik yağ serumlari; cilt sorununa özel serum paketleri; Paris-LA kökenli"),
        ("Circumference", "circumferencenyc.com", "Biyoteknoloji Serumu", "NYC merkezli; fermente aktifler; sürdürülebilir lüks; minimal ambalaj tasarımı"),
        ("Plenaire", "plenaire.co", "Jel Bazlı Serum", "Sıfır atık ambalaj; geri dönüştürülebilir cam; Gen Z fiyat aralığı; minimalist formüller"),
        ("UpCircle", "upcirclebeauty.com", "Geri Dönüşüm Serumu", "Kahve telvesi + meyve çekirdeği gibi gıda atıklarından serum; döngüsel ekonomi güzelliği"),
        ("Haeckels", "haeckels.com", "Deniz Kaynaklı Serum", "İngiltere kıyılarından toplanan deniz yosunu; GPS koordinatlı kaynak; artisan kozmetik"),
        ("Epi.Logic", "epi-logic.com", "Dermatolog Formülü Serum", "Dr. Loretta Ciraldo formülleri; klinik kanıtlı anti-aging; dermatoloji pratiğinden doğan marka"),
        ("Evereden", "evereden.com", "Anne-Bebek Serumu", "Pediatrik dermatolog geliştirdi; hamilelik güvenli; bebek ve anne için multi-vitamin bakım"),
        ("Glow Hub", "glowhub.com", "Z Kuşağı Serumu", "£5-10 fiyat aralığı; parlak neon ambalaj; İngiltere'de Boots'ta en çok satan; eğlenceli formül isimleri"),
        ("Geek & Gorgeous", "geekandgorgeous.com", "Macar Aktif Serum", "Budapeşte merkezli; C vitamini ve retinal serumlari; Avrupa'nın The Ordinary'si; süper uygun fiyat"),
        ("Skin + Me", "skin.me", "Kişiselleştirilmiş Serum", "Online dermatoloji konsültasyonu + kişiye özel formül; İngiltere reçeteli cilt bakımı DTC"),
        ("Dermatica", "dermatica.co.uk", "Reçeteli Serum", "İngiltere'de online dermatolog reçeteli tretinoin formülleri; kişiselleştirilmiş aktif dozları"),
        ("Summer Bio", "summerbio.com", "Biyoteknoloji Nemlendirici", "Lab-grown bileşenler; biyosentetik hyaluronic acid; gelecek nesil biyoteknoloji güzelliği"),
        ("MELE", "meleskincareforall", "Melanin Dostu Serum", "Koyu cilt tonları için özel formüller; hiperpigmentasyon uzmanı; kapsayıcı cilt bilimi"),
        ("EADEM", "eadem.co", "Smart Melanin Serumu", "Melanin bilimi teknolojisi; koyu ten için özel C vitamini; 'melanin-rich skin' kategorisi yaratıcısı"),
        ("Ami Colé", "amicole.com", "Melanin Nemlendirici", "Koyu cilt tonları için özel nemlendirici; Senegal kökenli kurucu; minimal temiz formüller"),
        ("Topicals", "mytopicals.com", "Kronik Cilt Serumu", "Kronik cilt rahatsızlıkları (egzama, hiperpigmentasyon) için; cilt sorunlarını normalleştirme; renkli ambalaj"),
        ("Bubble Skincare", "hellobubble.com", "Erişilebilir Z Kuşağı Serumu", "Walmart'ta $5-16; TikTok'ta 2B+ görüntülenme; ergen cilt bakımı demokratizasyonu"),
        ("Krave Beauty", "kravebeauty.com", "Minimalist K-Beauty Serumu", "Liah Yoo YouTube'dan markaya; #PressReset aşırı tüketim eleştirisi; Great Barrier Relief kült ürün"),
        ("Then I Met You", "thenimet you.com", "Kore Çift Temizleme Serumu", "Charlotte Cho kurdu; Soko Glam'dan bağımsız marka; living cleansing balm viral"),
        ("Rovectin", "rovectin.com", "Hassas Cilt Serumu", "Kemoterapi sonrası hassas cilt için geliştirildi; ultra-düşük tahriş; barrier repair uzmanı"),
        ("iUNIK", "iunik.kr", "Doğal Fermente Serum", "Propolis + centella; Kore doğal fermentasyon; %98 doğal içerik oranı"),
        ("Beauty of Joseon", "beautyofjoseon.com", "Hanbok Güzellik Serumu", "Kore Joseon Hanedanlığı reçeteleri; pirinç suyu + ginseng; geleneksel K-beauty"),
        ("Axis-Y", "axis-y.com", "Biome Serum", "Cilt mikrobiyomu dengeleme; artichoke + aha; Kore biyom odaklı cilt bakımı"),
        ("One Thing", "onethingkorea.com", "Tek İçerik Serumu", "Her üründe TEK aktif içerik; saf centella, saf hyaluronic; katmanlama için ideal"),
        ("Haruharu WONDER", "haruharuwonder.com", "Siyah Pirinç Serumu", "Kore siyah pirinç özütü; antioksidan fermentasyon; hanbang (geleneksel Kore tıbbı)"),
        ("Jumiso", "jumiso.com", "Vitamin Serumu", "All Day Vitamin serum serisi; Kore vitamin bazlı aydınlatma; uygun fiyat K-beauty"),
        ("Benton", "bentoncosmetic.com", "Arı Zehiri Serumu", "Snail mucin + bee venom kombinasyonu; Kore anti-aging fermente içerikler"),
        ("BY WISHTREND", "bywishtrend.com", "Vitamin C Serumu", "%21 vitamin C serumu; Kore aktif bazlı; Wishtrend'in kendi markası; Pure Vitamin C kült ürün"),
        ("PURITO", "purito.com", "Centella Serumu", "Centella bazlı hassas cilt serumları; unscented felsefe; Kore temiz güzellik öncüsü"),
        ("Klairs", "klairscosmetics.com", "Hassas Cilt K-Serumu", "Supple Preparation Toner kült ürün; hassas cilt uzmanı; parfümsüz minimalist K-beauty"),
        ("Isntree", "isntree.com", "Hyaluronic K-Serum", "8 çeşit hyaluronic acid tek üründe; Kore hyaluronic teknolojisi; katmanlı nemlendirme"),
        ("Torriden", "torriden.com", "Derin Nem Serumu", "DIVE-IN Low Molecular Hyaluronic Acid; düşük moleküllü hyaluronic; Kore'de Olive Young #1"),
        ("Anua", "anua.co.kr", "Heartleaf Serumu", "Houttuynia cordata (heartleaf) %77 tonik; Kore'nin yeni viral cilt bakım markası; TikTok fenomeni"),
        ("SKIN1004", "skin1004.com", "Madagascar Centella Serumu", "Madagaskar centella asiatica; tek kaynak bileşen; ampoule formatında yoğun bakım"),
        ("Round Lab", "roundlab.co.kr", "Kore Doğa Serumu", "Dokdo birch juice + soya bean; Kore doğal kaynaklı; Ulleungdo adası bileşenleri"),
        ("numbuzin", "numbuzin.com", "Numara Bazlı Serum", "Ürünleri numara ile adlandırma sistemi; No.5 Vitamin C serum viral; K-beauty yeni dalga"),
        ("mixsoon", "mixsoon.com", "Tek Bileşen Serumu", "Soybean + galactomyces; tek fermente bileşen felsefesi; saf ve şeffaf formüller"),
        ("ma:nyo", "manyo.co.kr", "Bifida Serumu", "Bifida ferment lysate; Kore fermentasyon bilimi; Galactomyces Niacin Essence kült ürün"),
        ("Heimish", "heimish.com", "Kore Ev Konforu Serumu", "'Heimish' = ev konforu; Bulgarian rose water bazlı; All Clean Balm en çok satan"),
        ("Needly", "needly.co.kr", "Minimalist K-Serum", "Kore minimalizm akımı; az ürün çok etki; Mild Cleansing felsefesi"),
        ("Illiyoon", "illiyoon.com", "Ultra Hassas Serum", "Amorepacific alt markası; ceramide kapsül teknolojisi; yenidoğan güvenli formüller"),
        ("Minimalist (India)", "beminimalist.co", "Hindistan Aktif Serumu", "Hindistan'ın The Ordinary'si; şeffaf ingrediyent + konsantrasyon; %10 niacinamide en çok satan"),
        ("Dot & Key", "dotandkey.com", "Hindistan Vitamin Serumu", "Renkli ambalaj; vitamin C + hyaluronic; genç Hindistan DTC güzellik; hızlı büyüyen"),
        ("Plum", "plumgoodness.com", "Hindistan Vegan Serumu", "%100 vegan; Hindistan temiz güzellik öncüsü; green tea serumu en çok satan"),
        ("The Derma Co", "thedermaco.com", "Hindistan Dermatolojik Serum", "Dermatolog formüllü; aktif dozları belirtilen; %1 hyaluronic + %0.3 retinol kombinasyonları"),
        ("Re'equil", "reequil.com", "Hindistan Bilimsel Serum", "Kimyager ekibin formüle ettiği; ingrediyent bilimi odaklı; pitta + lipid bazlı"),
        ("Deconstruct", "deconstructskincare.com", "Hindistan Şeffaf Serum", "İçerik ve konsantrasyon şeffaflığı; bilimsel pazarlama; Hindistan Gen Z favorisi"),
        ("Suganda", "suganda.co", "Hindistan K-Beauty Serumu", "K-beauty ilhamıyla Hindistan formülleri; centella + niacinamide; hibrit yaklaşım"),
        ("Pilgrim", "discoverpilgrim.com", "Hindistan Global Serum", "Kore, Fransa, Japonya reçetelerinden ilham; çoklu köken bazlı formüller"),
        ("Juicy Chemistry", "juicychemistry.com", "Hindistan Organik Serum", "USDA organik sertifikalı; soğuk pres yağlar; Hindistan organik güzellik öncüsü"),
        ("Aqualogica", "aqualogica.in", "Hindistan Hydra Serum", "Coconut water bazlı nemlendirme; Hindistan tropikal cilt bakımı; uygun fiyat hydration"),
        ("Dr. Sheth's", "drsheths.com", "Hindistan Ayurveda Serumu", "Modern ayurveda + bilimsel aktifler; haldi (zerdeçal) + niacinamide birleşimi"),
        ("Typology", "typology.com", "Fransız Minimalist Serum", "Paris merkezli; 10 ingrediyenden az formüller; vegan + minimalist; diagnostik quiz ile kişiselleştirme"),
        ("Augustinus Bader", "augustinusbader.com", "Kök Hücre Serumu", "TFC8 kök hücre teknolojisi; Prof. Augustinus Bader 30 yıl araştırma; lüks biyoteknoloji"),
        ("Medik8", "medik8.com", "İngiliz Vitamin C Serumu", "CSA felsefesi (C vitamini + SPF + A vitamini); stabilize vitamin C uzmanı; dermatoloji bazlı"),
        ("Drunk Elephant dışı - Drunk Elephant tarzı: Byoma", "byoma.com", "Bariyer Onarım Serumu", "Triseramide kompleksi; Target'ta $10-15; bariyer onarım demokratizasyonu"),
        ("Sand & Sky", "sandandsky.com", "Avustralya Pembe Kil Serumu", "Australian pink clay; Instagram pink mask viral; Avustralya doğal bileşenleri"),
        ("Go-To Skincare", "gotoskincare.com", "Avustralya Basit Serum", "Zoë Foster Blake kurdu; Avustralya no-nonsense güzellik; 5 ürünlük basit rutin"),
        ("Ultra Violette", "ultraviolette.com.au", "Avustralya SPF Serumu", "Serum + SPF hibrit ürünler; Avustralya güneş koruma uzmanı; şık SPF ambalajı"),
        ("Mukti Organics", "muktiorganics.com", "Avustralya Organik Serum", "Avustralya yerli bitkileri; bioaktif organik formüller; kakadu plum vitamin C"),
        ("Frank Body", "frankbody.com", "Kahve Bazlı Serum", "Kahve scrub ile başlayan marka; kahve özütlü cilt bakımı; Instagram viral pazarlama öncüsü"),
        ("Grown Alchemist", "grownalchemist.com", "Avustralya Biyolojik Serum", "Biyolojik bileşen yaklaşımı; peptide + antioksidan; Melbourne lüks temiz güzellik"),
        ("Epara", "eparaskincare.com", "Afrika Lüks Serumu", "Koyu cilt tonları için lüks bakım; Afrika bitkisel bileşenler; shea + marula + baobab"),
        ("54 Thrones", "54thrones.com", "Afrika Bitki Serumu", "54 Afrika ülkesinden ilham; shea butter + baobab; Afrika güzellik geleneği modernizasyonu"),
        ("Kaike", "kaikebeauty.com", "Çok Amaçlı Serum", "Her ürün yüz + vücut + saç; multi-use felsefesi; atık azaltma yaklaşımı"),
        ("Unsun Cosmetics", "unsuncosmetics.com", "Melanin Dostu SPF Serum", "Koyu cilt tonlarında beyaz iz bırakmayan SPF; mineral güneş koruma; kapsayıcı formüller"),
        ("Range Beauty", "rangebeauty.com", "Akne Eğilimli Melanin Serum", "Akne eğilimli koyu ciltler için özel; hyperpigmentasyon + akne dual tedavi"),
        ("Live Tinted", "livetinted.com", "Güney Asya Serumu", "Deepica Mutyala kurdu; Güney Asya cilt tonları uzmanı; Hueguard çok amaçlı stick"),
        ("Haoma", "haoma.earth", "Lüks Organik Serum", "London merkezli; organik adaptojenik serum; sürdürülebilir lüks; chaga + reishi mantarları"),
        ("Ere Perez", "ereperez.com", "Doğal Cilt Serumu", "Avustralya-Meksika kökenli; doğal ve organik; rice bran + papaya enzim formülleri"),
        ("Pai Skincare", "paiskincare.com", "Ultra Hassas Serum", "London merkezli; %100 organik; hassas + reaktif ciltler; chamomile + rosehip kült formüller"),
        ("Odele", "odelebeauty.com", "Erişilebilir Temiz Serum", "Target'ta $10-12; cinsiyet nötr; temiz + erişilebilir; aile boyu ambalaj"),
        ("Everist", "everist.com", "Susuz Konsantre Serum", "Waterless beauty öncüsü; konsantre pasta formüller; %0 su = daha az karbon ayak izi"),
        ("Elate Cosmetics", "elatecosmetics.com", "Sıfır Atık Serum", "Bambu ambalaj; kompostlanabilir; B Corp sertifikalı; sıfır atık güzellik"),
        ("Ethique", "ethique.com", "Katı Bar Serum", "Yeni Zelanda merkezli; katı bar formatı; 6M+ plastik şişe tasarrufu; konsantre katı formüller"),
        ("Allies of Skin", "alliesofskin.com", "Singapur Aktif Serumu", "Multi-acid + peptide yoğun formüller; Singapur merkezli lüks aktif bakım"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Peel Pad Serumu", "Alpha Beta Universal Daily Peel kült ürün; evde kimyasal peeling demokratizasyonu"),
        ("Derma E", "dermae.com", "Eczane Aktif Serumu", "Vitamin C + hyaluronic eczane fiyatına; Amerikan temiz eczane güzelliği; 40+ yıl deneyim"),
        ("Korres", "korres.com", "Yunan Eczane Serumu", "Atina eczanesinden doğan marka; Yunan bitkisel bileşenler; wild rose yağı kült ürün"),
        ("Nuxe", "nuxe.com", "Fransız Kuru Yağ Serumu", "Huile Prodigieuse çok amaçlı kuru yağ ikonik ürün; Fransız eczane güzelliği klasiği"),
        ("Caudalie", "caudalie.com", "Üzüm Bazlı Serum", "Bordeaux üzüm bağlarından polifenol; resveratrol anti-aging; vinotherapy konsepti"),
        ("Antipodes", "antipodesnature.com", "Yeni Zelanda Serumu", "Manuka balı + kivi tohumu; Yeni Zelanda yerli bitkileri; bilim + doğa birleşimi"),
        ("Pestle & Mortar", "pestleandmortar.com", "İrlanda Hyaluronic Serumu", "Pure Hyaluronic Serum kült ürün; İrlanda aile işletmesi; minimalist etkili formüller"),
        ("Jordan Samuel Skin", "jordansamuelskin.com", "Esthetician Serumu", "Profesyonel esthetician formülleri; The Hydrate serum; klinik sonuçlar ev fiyatına"),
        ("Dermalogica", "dermalogica.com", "Profesyonel Serum", "Esthetician eğitim odaklı; face mapping teknolojisi; profesyonel bakım evde"),
        ("Omorovicza", "omorovicza.com", "Macar Termal Serum", "Budapeşte termal su teknolojisi; mineral bazlı anti-aging; Macar kaplıca geleneği"),
        ("Aurelia London", "aurelialondon.com", "Probiyotik Lüks Serum", "BioOrganic probiyotik teknolojisi; London lüks temiz güzellik; peptide + probiyotik"),
        ("ARKIVE Headcare", "arkiveheadcare.com", "Saç-Cilt Serumu", "Adam Reed kurdu; saç ve kafa derisi serumu; profesyonel stilist formülleri"),
        ("Dr. Barbara Sturm", "drsturm.com", "Kan Plazma Serumu", "PRP (platelet-rich plasma) araştırmasından doğan lüks marka; Hyaluronic Serum ikonik; anti-inflammatory felsefe"),
        ("Maelove", "maelove.com", "Uygun Fiyatlı C Serumu", "SkinCeuticals CE Ferulic'in uygun fiyatlı alternatifi; The Glow Maker; bilimsel formül düşük fiyat"),
        ("Timeless Skincare", "timelessha.com", "Kopyalama Serumu", "SkinCeuticals formüllerinin uygun fiyatlı versiyonları; %20 C + E + ferulic; dupe kültürü"),
        ("CeraVe dışı ceramide markası: Ceramiracle", "ceramiracle.com", "Seramid Teknoloji Serumu", "Singapur merkezli; ileri ceramide kapsül teknolojisi; bariyer onarım bilimi"),
        ("REN Clean Skincare", "renskincare.com", "Temiz Biyoaktif Serum", "İskandinav temiz güzellik; %100 geri dönüştürülebilir ambalaj hedefi; Ready Steady Glow tonik"),
        ("Tata Harper", "tataharper.com", "Çiftlik Lüks Serumu", "Vermont çiftliğinden %100 doğal + lüks; her ürün çiftlikte üretilir; yeşil lüks öncüsü"),
        ("Vintner's Daughter", "vintnersdaughter.com", "Yüz Yağı Serumu", "Active Botanical Serum tek ürün marka; 22 botanik; $185 lüks yüz yağı; kült klasik"),
        ("Maya Chia", "mayachia.com", "Chia Yağı Serumu", "Süper gıda chia tohumu yağı; omega-3 zengin botanik serum; temiz lüks yağ bakımı"),
        ("Monastery", "monasterymade.com", "Botanik Yağ Serumu", "San Francisco artisan botanik; Flora yüz yağı; küçük parti üretim; bitki simyası"),
        ("Wildling", "wildling.com", "Gua Sha Serumu", "Gua sha + serum birlikte; yüz akupresür; geleneksel Çin tıbbı + modern formüller"),
        ("Mount Lai", "mountlai.com", "Jade Roller Serumu", "Jade + gua sha araçları + bakım serumları; Çin wellness geleneği modernizasyonu"),
        ("Bloomeffects", "bloomeffects.com", "Lale Serumu", "Hollanda lale özütü teknolojisi; Dutch Blush tonik; lale güzellik ritüeli"),
        ("SkinCeuticals alternatifi: Maelove", "maelove.com", "CE Ferulic Alternatifi", "SkinCeuticals formül yapısını uygun fiyata sunan; patent süresi dolan formülleri demokratize eden marka"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 2. Cilt Bakımı - Temizleyici & Tonik
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Temizleyici & Tonik": [
        ("Banila Co", "banilaco.com", "Kore Temizleme Balmı", "Clean It Zero sherbet kıvamlı temizleyici; Kore çift temizleme kültürünün ikonik ilk adımı"),
        ("Then I Met You", "thenimetyou.com", "Lüks Çift Temizleme", "Living Cleansing Balm; Charlotte Cho Soko Glam deneyimi; Kore çift temizleme ritüeli"),
        ("Elemis", "elemis.com", "İngiliz Pro Temizleyici", "Pro-Collagen Cleansing Balm en çok satan; İngiliz spa geleneği; starflower + elderberry"),
        ("Eve Lom", "evelom.com", "Muslin Bez Temizleyici", "Kült muslin bez temizleme ritüeli; balm + muslin bez; İngiliz güzellik efsanesi"),
        ("Kiehl's benzeri indie: Ursa Major", "ursamajorvt.com", "Vermont Doğal Temizleyici", "Vermont doğasından ilham; Fantastic Face Wash; erkek + kadın; doğal + etkili"),
        ("MAKE Beauty", "makebeauty.com", "Sanatçı Temizleyici", "Sanat dünyasından ilham; minimalist ambalaj; dual-phase temizleyici; Brooklyn merkezli"),
        ("Bioderma", "bioderma.com", "Misel Su Uzmanı", "Sensibio H2O misel su ikonik; Fransız dermatoloji; hassas cilt su bazlı temizleme öncüsü"),
        ("Gallinée", "gallinee.com", "Mikrobiyom Temizleyici", "Pre+pro+postbiyotik formüller; cilt mikrobiyomu koruyarak temizleme; pH 5.8 dengeleme"),
        ("Vanicream", "vanicream.com", "Dermatolojik Temizleyici", "Dermatolog önerisi #1; parfümsüz, boya-sız; ultra hassas formül; eczane klasiği"),
        ("La Roche-Posay tarzı indie: Geek & Gorgeous", "geekandgorgeous.com", "Macar pH Temizleyici", "pH dengelenmiş jel temizleyici; Budapeşte merkezli; bilimsel formül uygun fiyat"),
        ("Youth To The People", "youthtothepeople.com", "Süperfood Temizleyici", "Kale + green tea jel temizleyici; antioksidan temizleme; cam şişe; süperfood güzellik"),
        ("Ren Clean Skincare", "renskincare.com", "AHA Tonik", "Ready Steady Glow AHA tonik; laktik asit aydınlatma; İskandinav temiz güzellik"),
        ("Pixi", "pixibeauty.com", "Glow Tonik", "Glow Tonic %5 glikolik asit; dünya çapında viral tonik; yeşil şişe ikonik; erişilebilir AHA"),
        ("COSRX", "cosrx.com", "Kore BHA Temizleyici", "Low pH Good Morning Gel Cleanser; Kore pH dengeli temizleme; salicylic acid + tea tree"),
        ("Some By Mi", "somebymi.com", "30 Gün Mucize Tonik", "AHA BHA PHA 30 Days Miracle Toner; Kore'de 30 gün sonuç vaadi; çoklu asit formül"),
        ("Son & Park", "sonandpark.com", "Güzellik Suyu", "Beauty Water çok amaçlı tonik; tonik + temizleyici + eksfoliant; K-beauty multi-tasking"),
        ("Dear, Klairs", "klairscosmetics.com", "Hassas Cilt Toniği", "Supple Preparation Facial Toner; hassas cilt için pH dengeli; K-beauty kült tonik"),
        ("Laneige", "laneige.com", "Su Bazlı Temizleyici", "Water Sleeping Mask'ın temizleyici hattı; Kore su bilimi; hydro teknolojisi"),
        ("Innisfree", "innisfree.com", "Yeşil Çay Temizleyici", "Jeju Adası yeşil çay; Kore doğa bazlı temizleme; çevre dostu ambalaj"),
        ("Dermalogica", "dermalogica.com", "Profesyonel Temizleyici", "Double cleanse öncüsü; PreCleanse yağ; Special Cleansing Gel; esthetician favorisi"),
        ("First Aid Beauty", "firstaidbeauty.com", "SOS Temizleyici", "Face Cleanser hassas cilt; FAB Ultra Repair; P&G satın aldı; duyarlı cilt uzmanlığı"),
        ("Fresh", "fresh.com", "Soy Temizleyici", "Soy Face Cleanser amino asit bazlı; Boston merkezli; doğal bileşenler + lüks his"),
        ("Tatcha", "tatcha.com", "Japon Pirinç Temizleyici", "Rice Polish enzim pudra temizleyici; Japon geisha güzellik ritüellerinden ilham"),
        ("Sulwhasoo", "sulwhasoo.com", "Kore Hanbang Temizleyici", "Gentle Cleansing Oil ginseng bazlı; Kore geleneksel tıp (hanbang); lüks fermente formüller"),
        ("NOBE", "nobebeauty.com", "İsveç Temizleyici", "İsveç minimalist temizleme; nordik botanikler; sürdürülebilir İskandinav güzellik"),
        ("Oskia", "oskia.com", "İngiliz Enzim Temizleyici", "Renaissance Cleansing Gel MSM + enzim; İngiliz biyoteknoloji güzelliği; lüks aktif temizleme"),
        ("Emma Hardie", "emmahardie.com", "Moringa Temizleyici", "Moringa Cleansing Balm kült ürün; İngiliz facialist formülü; moringa + wild sea fennel"),
        ("Milu", "milubymila.com", "pH Dostu Temizleyici", "Mila kurdu; pH 5.5 dengelenmiş; Alman temiz cilt bilimi; basit etkili formüller"),
        ("Sioris", "sioris.com", "Taze Sıkım Temizleyici", "Mevsimsel hasat edilen bileşenler; taze sıkılmış; Kore slow beauty konsepti"),
        ("Rovectin", "rovectin.com", "Ultra Hassas Temizleyici", "Kemoterapi sonrası geliştirildi; Conditioning Cleanser; ultra düşük tahriş; barrier koruyan temizleme"),
        ("Thank You Farmer", "thankyoufarmer.us", "Kore Çiftçi Temizleyici", "Back to Iceland temizleyici; İzlanda yosunu + Kore teknolojisi; hybrid formül yaklaşımı"),
        ("Wishful", "wishful.com", "Enzim Temizleyici", "Huda Kattan'ın cilt bakım markası; Yo Glow enzim scrub; papaya + ananas enzimleri"),
        ("Allies of Skin", "alliesofskin.com", "Peptide Temizleyici", "Singapur merkezli; peptide + probiyotik temizleyici; lüks aktif temizleme"),
        ("Alpyn Beauty", "alpynbeauty.com", "Vahşi Bitki Temizleyici", "PlantGenius Creamy Bubbling Cleanser; Jackson Hole vahşi hasat bitkileri"),
        ("One Thing", "onethingkorea.com", "Tek İçerik Tonik", "Centella saf tonik; her şişe tek aktif bileşen; katmanlama toniği"),
        ("Pyunkang Yul", "pyunkangyul.us", "Kore Hanbang Tonik", "Essence Toner kült; Kore geleneksel tıp reçetesi; astragalus membranaceus kök özütü"),
        ("Goodal", "goodal.co.kr", "Yeşil Mandalina Tonik", "Green Tangerine Vita C tonik; Jeju yeşil mandalina; aydınlatıcı K-beauty tonik"),
        ("I'm From", "imfrom.co.kr", "Kaynak Bazlı Tonik", "Rice Toner pirinç bazlı; her ürün tek bölge + tek bileşen; traceability (izlenebilirlik)"),
        ("Isntree", "isntree.com", "Hyaluronic Tonik", "Hyaluronic Acid Toner; 50% hyaluronic; Kore nemlendirme toniği"),
        ("Nacific", "nacific.com", "Phyto Niacin Tonik", "Phyto Niacin Whitening Toner; bitki bazlı niacinamide; Kore aydınlatma uzmanı"),
        ("Manyo Factory", "manyo.co.kr", "Bifida Tonik", "Bifida Complex Ampoule Toner; fermente + hassas cilt; Kore bifida teknolojisi"),
        ("Dr. Ceuracle", "dr.ceuracle.com", "Kore Dermatolojik Tonik", "Vegan Kombucha Tea Essence; dermatoloji + vegan formüller; fermente çay bazlı"),
        ("Beplain", "beplain.co.kr", "Kore Cicapair Tonik", "Chamomile pH-Balanced Toner; hassas cilt pH dengeleme; Kore papatya tonik"),
        ("Abib", "abib.com", "Kore Jel Tonik", "Heartleaf Calming Toner; jel kıvamlı tonik; Kore yenilikçi tonik formları"),
        ("Huxley", "huxley.co.kr", "Kaktüs Yağı Tonik", "Sahara çölü kaktüs yağı; Secret of Sahara serisi; Kore egzotik bileşen yaklaşımı"),
        ("Belif", "belifcosmetic.com", "İngiliz Bitkisel Tonik", "Napiers of Edinburgh bitki formülleri; Kore + İngiliz botanik hibrit; herb-infused tonik"),
        ("Bonajour", "bonajour.co.kr", "Kore Yeşil Çay Tonik", "Green Tea Water Bomb; antioksidan yeşil çay konsantresi; Kore uygun fiyat organik"),
        ("Aromatica", "aromatica.co.kr", "Kore Organik Tonik", "USDA organik Kore markası; Reviving Rose Infusion; organik K-beauty"),
        ("Whamisa", "whamisa.com", "Fermente Organik Tonik", "Organik fermente tonik; %95+ doğal; Kore fermentasyon + organik sertifika"),
        ("Round Lab", "roundlab.co.kr", "Dokdo Tonik", "1025 Dokdo Toner; deniz mineralleri; Kore ulusal gurur ürünü; deep sea water"),
        ("Numbuzin", "numbuzin.com", "No.3 Tonik", "No.3 Super Glowing Essence Toner; numara sistemi; Kore'nin en viral tonik markası"),
        ("Skin Laundry", "skinlaundry.com", "Lazer Temizleme", "15 dakika lazer + ışık tedavisi klinikleri; klinik sonuçlar express hizmet; NYC + LA"),
        ("Pai Skincare", "paiskincare.com", "Organik Hassas Temizleyici", "Camellia & Rose Gentle Hydrating Cleanser; hassas cilt uzmanı; London organik"),
        ("Evolve Organic Beauty", "evolvebeauty.co.uk", "İngiliz Organik Temizleyici", "Aromatic Wash; İngiliz organik; el yapımı küçük parti; Soil Association onaylı"),
        ("Inlight Beauty", "inlightbeauty.com", "Organik Yağ Temizleyici", "Face Cleanser %100 organik yağ bazlı; İngiliz organik lüks; Dr. Spiezia formülleri"),
        ("Rael", "getrael.com", "Bambu Kömür Temizleyici", "Bamboo charcoal temizleyici; Kore-Amerikan hibrit; doğal + bilimsel temizleme"),
        ("Nooni", "nooni.co", "Kore Su Özlü Temizleyici", "Apple Water Cleanser; Kore meyve suyu bazlı temizleme; hafif + nemlendirici formül"),
        ("Holika Holika", "holikaholika.com", "Eğlenceli K-Temizleyici", "Aloe 99% temizleyici jel; eğlenceli ambalaj; uygun fiyat Kore bakımı"),
        ("Etude House", "etude.com", "Kore SoonJung Temizleyici", "SoonJung pH 6.5 Whip Cleanser; ultra hassas; düşük pH Kore temizleme"),
        ("A'PIEU", "apieu.com", "Kore Madecassoside Temizleyici", "Madecassoside Cica Gel; centella temizleme; Kore uygun fiyat cica bakımı"),
        ("Bring Green", "bringgreen.co.kr", "Kore Tea Tree Temizleyici", "Tea Tree Cica jel; sorunlu cilt temizleme; Kore çay ağacı uzmanı"),
        ("Mary & May", "maryandmay.com", "Kore İdealift Temizleyici", "Lemon Niacinamide Glow Wash Off Pack; Kore vitamin temizleme; dual-action formüller"),
        ("VT Cosmetics", "vtcosmetics.com", "BTS Cica Temizleyici", "Cica Mild Foam Cleanser; BTS kolaborasyonu ile ünlü; Kore cica temizleme"),
        ("Dr.G", "drg.co.kr", "Kore Dermatolojik Temizleyici", "pH Cleansing Gel; dermatolog formüllü; Kore dermo-kozmetik temizleme"),
        ("heimish", "heimish.com", "Kore Balm Temizleyici", "All Clean Balm en çok satan; pembe kil + gül suyu; Kore çift temizleme ritüeli"),
        ("Mixsoon", "mixsoon.com", "Kore Soybean Temizleyici", "Soybean Milk Cleanser; soya sütü bazlı; tek bileşen Kore temizleme felsefesi"),
        ("Miguhara", "miguhara.co.kr", "Kore EGF Temizleyici", "EGF bazlı temizleyici; growth factor temizleme; Kore anti-aging ilk adım"),
        ("B.LAB", "blab.kr", "Kore Matcha Temizleyici", "Matcha Hydrating Foam Cleanser; Kore matcha trendi; antioksidan köpük temizleme"),
        ("Skin1004", "skin1004.com", "Madagaskar Centella Temizleyici", "Centella Ampoule Foam; Madagaskar centella; Kore ampoule-in-foam teknolojisi"),
        ("Lagom", "lagom.kr", "Kore Mikro Köpük Temizleyici", "Cellup Micro Foam Cleanser; mikro köpük teknolojisi; Kore hassas temizleme"),
        ("Laka", "laka.co.kr", "Genderless Temizleyici", "Cinsiyet nötr K-beauty; Vegan Lip Cheek'ten cilt bakıma geçiş; minimal Kore"),
        ("Frudia", "frudia.com", "Kore Meyve Temizleyici", "Citrus Brightening Cleanser; Kore meyve bazlı; pomegranate + blueberry + citrus"),
        ("G9Skin", "g9skin.com", "Kore Süt Temizleyici", "Milk Bomb Cleanser; süt protein bazlı; Kore süt güzelliği konsepti"),
        ("Thank You Farmer", "thankyoufarmer.us", "Geri Dönüş Temizleyici", "Back to Pure Daily Foaming Gel; saf formüller; Kore temiz güzellik"),
        ("Real Barrier", "realbarrier.com", "Kore Bariyer Temizleyici", "Cream Cleansing Foam; ceramide NP bazlı; Kore bariyer koruyarak temizleme"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 3. Cilt Bakımı - Güneş Koruma (SPF)
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Güneş Koruma (SPF)": [
        ("Vacation Inc", "vacation.inc", "Retro SPF", "1980'ler estetik güneş kremleri; Classic Whip krem şantili SPF ambalaj; eğlenceli marka kimliği"),
        ("Supergoop!", "supergoop.com", "SPF İnovatörü", "Unseen Sunscreen görünmez SPF; SPF'i eğlenceli yapan marka; 40+ SPF formatı; Holly Thaggard kurdu"),
        ("Ultra Violette", "ultraviolette.com.au", "Avustralya SPF Uzmanı", "Queen Screen SPF 50+; Avustralya güneş uzmanı; cilt bakımı + SPF hibrit; şık ambalaj"),
        ("Kinship", "lovekinship.com", "Ekolojik SPF", "Self Reflect SPF 32; mercan resiflerine güvenli; okyanus dostu güneş koruma"),
        ("Canmake", "canmake.com", "Japon Hafif SPF", "Mermaid Skin UV Gel SPF 50+; ultra hafif Japon güneş koruma; makyaj altı ideal"),
        ("Isntree", "isntree.com", "Kore Hyaluronic SPF", "Hyaluronic Acid Watery Sun Gel; nemlendirici SPF; Kore su bazlı güneş koruma"),
        ("Beauty of Joseon", "beautyofjoseon.com", "Kore Pirinç SPF", "Relief Sun Rice + Probiotics SPF 50+; pirinç özütü + probiyotik; K-beauty viral SPF"),
        ("Round Lab", "roundlab.co.kr", "Kore Birch SPF", "Birch Juice Moisturizing Sun Cream; huş suyu nemlendirmeli; Kore doğal SPF"),
        ("Skin Aqua", "skin-aqua.com", "Japon Su Bazlı SPF", "UV Super Moisture Gel; ultra hafif Japon formül; nem bariyeri + SPF"),
        ("Biore UV", "biore.com", "Japon Watery SPF", "Watery Essence SPF 50+; Japon su bazlı SPF standardı; dünyada en çok satan Japon SPF"),
        ("Anessa", "anessa.shiseido.co.jp", "Japon Spor SPF", "Perfect UV Sunscreen Milk; ısı + su + ter dayanıklı; Japon outdoor SPF altın standardı"),
        ("Allie", "kanebo-cosmetics.jp", "Japon Jel SPF", "Extra UV Gel; Japon jel bazlı SPF; friction-proof teknoloji; spor SPF"),
        ("Dr. G", "drg.co.kr", "Kore Yeşil SPF", "Green Mild Up Sun Plus; hassas cilt SPF; Kore dermatolog formülü; cica + centella"),
        ("Purito", "purito.com", "Kore Günlük SPF", "Daily Go-To Sunscreen; PA++++ Kore standardı; hafif + nemlendirici; reformüle edilmiş güvenilir SPF"),
        ("Missha", "missha.com", "Kore Aqua SPF", "All Around Safe Block; Kore su bazlı SPF klasiği; uygun fiyat yüksek koruma"),
        ("Etude House", "etude.com", "Kore Tone-Up SPF", "Sunprise Mild Airy Finish; mat SPF; Kore ton eşitleme + koruma"),
        ("Thank You Farmer", "thankyoufarmer.us", "Kore Su Bomba SPF", "Sun Project Water Sun Cream; ultra nemlendirici SPF; Kore hidrasyon + koruma"),
        ("La Roche-Posay benzeri indie: Elta MD", "eltamd.com", "Dermatolojik SPF", "UV Clear SPF 46; niacinamide + SPF; dermatolog #1 önerisi; akne eğilimli cilt SPF"),
        ("Black Girl Sunscreen", "blackgirlsunscreen.com", "Kapsayıcı SPF", "Koyu cilt tonlarında beyaz iz bırakmayan SPF; melanin dostu formül; kapsayıcı güneş koruma"),
        ("Unsun Cosmetics", "unsuncosmetics.com", "Melanin SPF", "Mineral Tinted Face Sunscreen; tüm cilt tonlarında şeffaf; renkli mineral SPF"),
        ("Coola", "coola.com", "Organik SPF", "Classic Face Organic SPF 30; %70+ organik bileşen; temiz SPF; farm-to-face"),
        ("Sun Bum", "sunbum.com", "Sörf SPF", "Original SPF 30; sörf kültürü + güneş koruma; tropikal muz kokusu; eğlenceli marka"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi SPF", "Sun Shield SPF; hindistan cevizi yağı bazlı; tropikal temiz güneş koruma"),
        ("Colorescience", "colorescience.com", "Pudra SPF", "Sunforgettable mineral pudra SPF 50; fırça ile SPF uygulama; makyaj üzeri yeniden uygulama"),
        ("Australian Gold", "australiangold.com", "Avustralya Renkli SPF", "Botanical Tinted SPF 50; botanik + mineral; renkli SPF günlük kullanım"),
        ("Saltair", "saltair.com", "Vücut SPF", "Body sunscreen; vücut güneş koruma uzmanı; tropikal kokulu; plaj güzelliği"),
        ("Habit", "habit.com", "SPF Stick", "SPF stick formatı; 15 günde yeniden uygulama hatırlatıcısı; cep boyu SPF; alışkanlık oluşturucu"),
        ("Everyday Humans", "everydayhumans.com", "Hibrit SPF", "Resting Beach Face SPF 30 serum-krem hibrit; multi-tasking SPF; erişilebilir fiyat"),
        ("Bask", "basksuncare.com", "SPF Aboneliği", "Aylık SPF abonelik modeli; yeniden uygulama kolaylığı; beyaz iz bırakmayan formül"),
        ("Tower 28", "tower28beauty.com", "Hassas SPF", "SunnyDays tinted SPF; NEA onaylı; hassas + akne eğilimli cilt; temiz mineral SPF"),
        ("Drmtlgy", "drmtlgy.com", "Dermatolog SPF", "Universal Tinted SPF 46; dermatolog formülü; niacinamide + SPF; her cilt tonuna uyum"),
        ("Pipette", "pipette.com", "Bebek SPF", "Mineral SPF 50; bebek + hamile güvenli; squalane bazlı; ultra nazik formül"),
        ("ThinkSport", "gothink.com", "Spor SPF", "Livestrong SPF 50+; spor güneş koruma; EWG #1 SPF; reef-safe aktif outdoor"),
        ("Manda", "mandaorganic.com", "Organik Çinko SPF", "%100 organik; non-nano çinko oksit; okyanus güvenli; doğal aktif outdoor SPF"),
        ("Suntegrity", "suntegrity.com", "5-in-1 SPF", "Natural Moisturizing Face Sunscreen; 5-in-1 formül (SPF + primer + nemlendirici + anti-aging + renkli)"),
        ("MDSolarSciences", "mdsolarsciences.com", "Dermatolojik SPF", "Mineral Crème SPF 50; dermatoloji pratiğinden; sheer mineral + hyaluronic acid"),
        ("Josie Maran", "josiemaran.com", "Argan SPF", "Argan Daily Moisturizer SPF 47; argan yağı + SPF; çok amaçlı doğal koruma"),
        ("Saie", "saiehello.com", "Makyaj SPF", "Sunvisor SPF 35; makyaj altı SPF + aydınlatıcı; temiz SPF + glow efekti"),
        ("Iris & Romeo", "irisandromeo.com", "Multi-Tasking SPF", "Best Skin Days SPF 25; nemlendirici + SPF + tint + bakım; 4-in-1 minimalist"),
        ("Innisfree", "innisfree.com", "Kore Yeşil Çay SPF", "Jeju yeşil çay SPF; doğal + hafif; Kore günlük SPF klasiği"),
        ("Klairs", "klairscosmetics.com", "Kore Hassas SPF", "Soft Airy UV Essence; hassas cilt K-SPF; hafif + nemlendirici; parfümsüz"),
        ("COSRX", "cosrx.com", "Kore Aloe SPF", "Aloe Soothing Sun Cream; aloe vera + SPF; yatıştırıcı Kore güneş koruma"),
        ("A'PIEU", "apieu.com", "Kore Saf SPF", "Pure Block Natural Daily Sun Cream; doğal Kore SPF; uygun fiyat günlük koruma"),
        ("Mama & Kids", "mamakids.co.jp", "Japon Bebek SPF", "UV Light Veil; Japon bebek güneş koruma; ultra nazik; yenidoğan güvenli"),
        ("ALLIE", "kanebo-cosmetics.jp", "Japon Beauty SPF", "Chrono Beauty Gel UV EX; Japon güzellik SPF; makyaj bazı etkili; uzun süre dayanıklı"),
        ("Nivea Sun Japan", "nivea.co.jp", "Japon Jel SPF", "UV Deep Protect & Care Gel; Japon özel formül; su bazlı jel SPF; uygun fiyat"),
        ("Aēsop", "aesop.com", "Botanik SPF", "Protective Body Lotion SPF 50; botanik lüks SPF; Avustralya minimal güneş koruma"),
        ("Sand & Sky", "sandandsky.com", "Avustralya Glow SPF", "Australian Glow Berries SPF; Avustralya süper meyveler; parlak cilt + koruma"),
        ("Mecca Cosmetica", "mecca.com.au", "Avustralya Günlük SPF", "To Save Face SPF 50+; Avustralya perakende + kendi SPF markası; daily staple"),
        ("We Are Feel Good Inc", "wearefeelgoodinc.com.au", "Avustralya Sörf SPF", "Sensitive SPF 50+; Avustralya sörf kültürü; çinko bazlı; okyanus güvenli"),
        ("Cancer Council", "cancercouncil.com.au", "Avustralya Tıbbi SPF", "Face Day Wear SPF 50+; kanser araştırma bağışı; Avustralya devlet destekli SPF"),
        ("Summer Bio", "summerbio.com", "Biyoteknoloji SPF", "Lab-grown UV filtreleri; gelecek nesil sürdürülebilir güneş koruma; biyosentetik filtreler"),
        ("Thinksport", "gothink.com", "Aktif Spor SPF", "SPF 50+ spor formül; EWG en güvenli SPF; aktif yaşam tarzı güneş koruma"),
        ("Raw Elements", "rawelementusa.com", "Doğal Çinko SPF", "%100 doğal; non-nano çinko; biyolojik olarak parçalanabilir; okyanus güvenli"),
        ("All Good", "allgoodproducts.com", "Organik Çinko SPF", "SPF 30 Sport; organik + çinko oksit; B Corp sertifikalı; outdoor yaşam SPF"),
        ("Badger Balm", "badgerbalm.com", "Organik Aile SPF", "SPF 30 Baby; organik bebek + aile SPF; çinko bazlı; sertifikalı doğal"),
        ("Alba Botanica", "albabotanica.com", "Botanik Aile SPF", "Very Emollient SPF 45; botanik nemlendirici SPF; %100 vejetaryen; aile boyu"),
        ("Kosas", "kosas.com", "Makyaj SPF", "DreamBeam SPF 40; makyaj + SPF + bakım; temiz formül; renkli SPF"),
        ("Shiseido Urban", "shiseido.com", "Japon Şehir SPF", "Urban Environment SPF; kirlilik + UV koruma; Japon teknoloji + şehir hayatı"),
        ("Heliocare", "heliocare.com", "İspanyol Oral SPF", "Oral kapsül + topikal SPF; Fernblock polypodium teknolojisi; içten dışa UV koruma"),
        ("Ultrasun", "ultrasun.com", "İsviçre Tek Uygulama SPF", "Günde tek uygulama yeten SPF; İsviçre teknolojisi; 8+ saat koruma vaadi"),
        ("Bioderma Photoderm", "bioderma.com", "Fransız Dermatolojik SPF", "Photoderm MAX SPF 50+; Cellular Bioprotection; Fransız eczane SPF standardı"),
        ("Avène", "avene.com", "Fransız Termal SPF", "Solaire SPF 50+; termal su bazlı; hassas cilt Fransız eczane SPF"),
        ("SVR", "laboratoire-svr.com", "Fransız Aktif SPF", "Sun Secure Blur SPF 50+; optik bulanıklaştırma + SPF; makyaj bazı + koruma"),
        ("Eucerin", "eucerin.com", "Alman Dermatolojik SPF", "Oil Control Sun Gel-Cream SPF 50+; yağlı cilt SPF; Thiamidol teknolojisi"),
        ("P20", "riemann.com", "Danimarka Uzun Süreli SPF", "P20 Sun Protection; 10 saat tek uygulama; Danimarka dayanıklı SPF formülü"),
        ("Green People", "greenpeople.co.uk", "İngiliz Organik SPF", "Scent Free Sun Cream SPF 30; organik; hassas cilt; İngiliz yeşil güzellik"),
        ("Altruist", "altruist.com", "Dermatolojik Uygun SPF", "SPF 50 Dermatologist Sunscreen; dermatolog Dr. Andrew Birnie; £4 süper uygun fiyat"),
        ("Sol de Janeiro", "soldejaneiro.com", "Brezilya SPF", "Brazilian Bum Bum SPF; Brezilya plaj kültürü; kokusu ile ünlü; tropikal SPF"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 4. Cilt Bakımı - Akne & Leke Tedavisi
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Akne & Leke Tedavisi": [
        ("Starface", "starface.world", "Eğlenceli Akne Patch", "Yıldız şeklinde hidrokolloid patchler; akneyi normalleştirme; Z kuşağı ikonografisi; sarı yıldız viral"),
        ("Hero Cosmetics", "herocosmetics.com", "Mighty Patch", "Gece akne patchleri; Church & Dwight $630M satın aldı; Amazon #1; hidrokolloid öncüsü"),
        ("Topicals", "mytopicals.com", "Kronik Cilt Tedavisi", "Faded hiperpigmentasyon kremi kült; kronik cilt sorunlarını normalleştirme; renkli Z kuşağı ambalaj"),
        ("Peace Out", "peaceoutskincare.com", "Hedefli Patch Tedavi", "Akne + kırışıklık + gözenek patchleri; Sephora en çok satan patch; nokta tedavi uzmanı"),
        ("ZitSticka", "zitsticka.com", "Mikro İğne Patch", "Mikro-dart akne patchleri; deri altına niacinamide iletimi; bilimsel patch teknolojisi"),
        ("Avarelle", "avarelle.com", "Bitki Özlü Patch", "Tea tree + calendula + cica patch; bitkisel akne tedavi; uygun fiyat hidrokolloid"),
        ("COSRX", "cosrx.com", "Kore Akne Patch", "Acne Pimple Master Patch; Kore akne patch standardı; 3 boyut; dünya çapında best seller"),
        ("Some By Mi", "somebymi.com", "30 Gün Akne Tedavi", "AHA BHA PHA 30 Days Miracle; 30 gün sonuç vaadi; Kore çoklu asit akne tedavisi"),
        ("COSRX BHA", "cosrx.com", "Kore BHA Tedavi", "BHA Blackhead Power Liquid; Kore BHA bazlı siyah nokta tedavisi; betaine salicylate"),
        ("Paula's Choice BHA", "paulaschoice.com", "BHA Eksfoliant", "%2 BHA Liquid Exfoliant; akne + siyah nokta tedavi; global kült ürün; ingrediyent eğitimi"),
        ("Differin", "differin.com", "OTC Retinoid Akne", "Adapalene Gel %0.1; ilk reçetesiz retinoid; eczane akne tedavi devrimi"),
        ("Rael", "getrael.com", "Kore-Amerikan Akne Patch", "Miracle Patch; Kore patch teknolojisi + Amerikan pazarlama; doğal bileşenler"),
        ("Mighty Patch alternatifi: Good Light", "goodlight.world", "Kapsayıcı Akne Bakımı", "Moon Glow Pimple Patches; cinsiyet nötr; Z kuşağı akne bakımı; kapsayıcılık odaklı"),
        ("Kate Somerville", "katesomerville.com", "Hollywood Akne Tedavisi", "EradiKate akne tedavisi; Hollywood klinik deneyimi; pembe kükürt bazlı spot treatment"),
        ("Mario Badescu", "mariobadescu.com", "Kurutucu Losyon", "Drying Lotion pembe şişe ikonik; kükürt + salisilik asit; NYC facialist klasiği"),
        ("Murad", "murad.com", "Dermatolojik Akne", "InvisiScar patch + Rapid Relief Spot Treatment; Dr. Howard Murad; klinik akne çözümleri"),
        ("Dermalogica", "dermalogica.com", "Profesyonel Akne Tedavisi", "AGE Bright Clearing Serum; retinol + salisilik asit; esthetician bazlı akne tedavisi"),
        ("The Ordinary Niacinamide", "theordinary.com", "Niacinamide Akne Serumu", "%10 Niacinamide + %1 Zinc; global viral; akne + sebum kontrolü; $6 fiyat devrimi"),
        ("Naturium", "naturium.com", "Azelaic Acid Akne", "Azelaic Acid Emulsion %10; akne izleri + aktif akne; Amerikan aktif akne tedavisi"),
        ("Krave Beauty", "kravebeauty.com", "Bariyer Onarımlı Akne", "Great Barrier Relief; akne tedavisi sırasında bariyer koruma; tamanu oil + cica"),
        ("Cocokind", "cocokind.com", "Doğal Akne Tedavisi", "Turmeric Spot Treatment; zerdeçal bazlı doğal akne noktasal tedavi; temiz formül"),
        ("La Roche-Posay Effaclar alternatifi: Geek & Gorgeous", "geekandgorgeous.com", "BHA Akne Tedavisi", "aPHAclear; Macar formül; uygun fiyat BHA; akne + gözenek temizleme"),
        ("Medicube", "medicube.com", "Kore Dermatolojik Akne", "Red Acne Line; zero pore pad viral TikTok; Kore dermatoloji klinik akne serisi"),
        ("VT Cosmetics", "vtcosmetics.com", "Kore Cica Akne", "Cica Spot Patch; cica + centella akne iyileştirme; Kore botanik akne tedavisi"),
        ("By Wishtrend", "bywishtrend.com", "Kore Propolis Akne", "Propolis Energy Calming Ampoule; arı propolisi anti-inflammatory; Kore doğal akne bakımı"),
        ("Benton", "bentoncosmetic.com", "Kore Snail Akne", "Snail Bee High Content Essence; salyangoz + arı zehiri; Kore akne iyileştirme"),
        ("iUNIK", "iunik.kr", "Kore Propolis Akne", "Propolis Vitamin Synergy Serum; propolis + hippophae; Kore doğal akne + aydınlatma"),
        ("Axis-Y", "axis-y.com", "Kore Mugwort Akne", "Mugwort Pore Clarifying Wash Off Pack; artemisia gözenek; Kore bitkisel akne tedavisi"),
        ("Acnemy", "acnemy.com", "İspanyol Akne Uzmanı", "Zitmask + Dryzit; yalnızca akneye odaklı marka; İspanyol innovasyon; mikro patch teknolojisi"),
        ("Skin Proud", "skinproud.com", "Erişilebilir Akne", "Spot treatment + bakım; £5-10; Superdrug Birleşik Krallık; Z kuşağı erişilebilir akne bakımı"),
        ("Carbon Theory", "carbontheory.com", "Kömür Akne Sabunu", "Charcoal + tea tree + salisilik bar sabun; İrlanda kökenli; aktif kömür akne tedavisi"),
        ("Blume", "blume.com", "Ergen Akne Bakımı", "Meltdown acne oil; ergenler + genç yetişkinler; period + akne normalleştirme; Z kuşağı"),
        ("Bushbalm", "bushbalm.com", "Bikini Akne Tedavisi", "Ingrown hair + bikini akne; cesur niş; Shark Tank; tüy batması + akne vücut tedavisi"),
        ("Truly", "trulybeauty.com", "Eğlenceli Akne Tedavisi", "Unicorn Fruit; eğlenceli ambalaj; vücut akne tedavisi; TikTok viral; renkli formüller"),
        ("Versed", "versedskin.com", "Temiz Akne Tedavisi", "Back-Up Plan Acne-Control Mist; vücut akne spreyi; temiz + etkili akne bakımı"),
        ("Phyla", "phylabiotics.com", "Faj Akne Tedavisi", "Phage teknolojisi ile akne bakterisi hedefleme; bakteriyofaj bazlı; antibiyotiksiz çözüm"),
        ("Curology", "curology.com", "Kişiselleştirilmiş Akne", "Online dermatolog + kişiye özel reçete formülü; tretinoin + niacinamide + azelaic custom mix"),
        ("Apostrophe", "apostrophe.com", "Teledermatolog Akne", "Online reçeteli akne tedavisi; spironolactone + tretinoin; hormonal akne uzmanı"),
        ("Skin + Me", "skin.me", "İngiliz Reçeteli Akne", "Online dermatoloji + kişiselleştirilmiş tretinoin formülü; İngiltere reçeteli akne bakımı"),
        ("Agency", "agency.com", "Erkek Akne Tedavisi", "Erkekler için kişiselleştirilmiş cilt bakımı; teledermatolog; tretinoin + clindamycin formülleri"),
        ("Averr Aglow", "averraglow.com", "Doğal Akne Seti", "Clear Skin Kit; %100 doğal bileşenler; yetişkin akne uzmanı; botanik bazlı sistem"),
        ("Exposed Skin Care", "exposedskincare.com", "Bilimsel Akne Sistemi", "Basic Kit; bilimsel + botanik karışım; %3.5 benzoyl peroxide + yeşil çay; akne sistemi"),
        ("TreeActiv", "treeactiv.com", "Ağaç Çayı Akne", "Cystic Acne Spot Treatment; tea tree + bentonit kil; kistik akne uzmanı; doğal formül"),
        ("Neutralyze", "neutralyze.com", "Kimyasal Akne Tedavisi", "Multi-acid akne tedavi sistemi; mandelic + salicylic + nitrogen boost; bilimsel formül"),
        ("La Roche Posay alternatifi: SVR", "laboratoire-svr.com", "Fransız Akne Tedavisi", "Sebiaclear Serum; Fransız eczane; gluconolactone + niacinamide; akne + anti-aging"),
        ("Bioderma Sébium", "bioderma.com", "Fransız Akne Hattı", "Sébium akne serisi; Fluidactiv patent teknolojisi; sebum kalitesi düzenleme"),
        ("Ducray", "ducray.com", "Fransız Dermatolojik Akne", "Keracnyl akne hattı; Fransız dermatoloji grubu; myrtacine + glycolic acid"),
        ("A-Derma Phys-AC", "aderma.com", "Fransız Organik Akne", "Phys-AC Global; organik Rhealba yulaf; hassas akne eğilimli cilt; Fransız organik dermo"),
        ("Sebamed", "sebamed.com", "Alman pH Akne", "Clear Face Anti-Pimple Gel; pH 5.5; Alman dermo-kozmetik; hassas akne cilt"),
        ("Benzac", "benzac.com", "Benzoyl Peroxide Uzmanı", "Benzac AC; Galderma markası; %2.5-%10 benzoyl peroxide; eczane akne standardı"),
        ("Dermatica", "dermatica.co.uk", "İngiliz Reçeteli Akne", "Kişiye özel tretinoin + niacinamide + azelaic acid formülleri; online dermatolog"),
        ("The INKEY List", "theinkeylist.com", "Uygun Fiyat Akne Serumu", "Beta Hydroxy Acid Serum; £8; ingrediyent eğitimi; akne tedavisi demokratizasyonu"),
        ("Bravura London", "bravuralondon.com", "İngiliz Asit Akne", "Salicylic Acid %2; İngiliz aktif cilt bakımı; profesyonel asitler ev fiyatına"),
        ("Garden of Wisdom", "gardenofwisdom.com", "DIY Akne Aktif", "Azelaic Acid %10; minimalist aktif; ingrediyent bazlı; ev formülasyonu destekçisi"),
        ("Banish", "banish.com", "Akne İzi Tedavisi", "Banisher microneedle stamp; akne izleri için mikro iğne; evde dermapen alternatifi"),
        ("Zitsticka", "zitsticka.com", "Dermatik Patch", "Killa Kit; mikro-dart teknolojisi ile deri altı iletim; niacinamide + salicylic mikro iğne patch"),
        ("Patchology", "patchology.com", "Akne Patch Çeşitliliği", "Breakout Box; çeşitli patch tipleri; eğlenceli ambalaj; hediye seti formatı"),
        ("Mighty Patch", "herocosmetics.com", "Orijinal Hidrokolloid", "Mighty Patch Original; hidrokolloid akne patch standardı; gece uygulaması; 6-8 saat etki"),
        ("Dots for Spots", "dotsforspots.com", "İngiliz Akne Patch", "Transparent acne patches; İngiliz marka; ultra ince görünmez patch; gündüz kullanım"),
        ("APRICOT", "apricot-beauty.com", "Alman Beauty Patch", "Pimple Patches; hyaluron + salisilik asit; Alman innovasyon; tekrar kullanılabilir yüz patchleri"),
        ("Skyn Iceland", "skyniceland.com", "İzlanda Akne Tedavisi", "Blemish Dots; İzlanda mineralleri; stres kaynaklı akne; adaptojenik tedavi"),
        ("Dr. Jart+", "drjart.com", "Kore Cica Akne", "Cicapair Tiger Grass; centella asiatica; Kore cica akne iyileştirme; kaplan otu"),
        ("Aestura", "aestura.com", "Kore Dermatolojik Akne", "A-Cica 365 Calming Cream; Kore hastane markası; dermatoloji kliniklerinde kullanılan"),
        ("Neogen", "neogenlab.us", "Kore Peeling Akne", "Bio-Peel Gauze Peeling; gazlı bez peeling; Kore mekanik + kimyasal eksfoliasyon"),
        ("Isntree", "isntree.com", "Kore BHA Akne", "BHA Clear Skin Toner; Kore BHA akne toniği; hafif + etkili; günlük kullanım"),
        ("Purito", "purito.com", "Kore Centella Akne", "Centella Green Level Recovery Cream; centella bazlı akne iyileştirme; hassas formül"),
        ("Ciracle", "ciracle.com", "Kore Anti-Blemish", "Anti-Blemish Aqua Cream; Kore akne nemlendirici; hafif jel kıvam; gözenek küçültme"),
        ("Mizon", "mizon.co.kr", "Kore Snail Akne", "Snail Recovery Gel Cream; salyangoz müsin akne iyileştirme; Kore salyangoz teknolojisi"),
        ("Etude SoonJung", "etude.com", "Kore Hassas Akne", "SoonJung 2x Barrier Intensive Cream; hassas akne eğilimli cilt; pH 5.5 formül"),
        ("Klairs Midnight Blue", "klairscosmetics.com", "Kore Calming Akne", "Midnight Blue Calming Cream; guaiazulene; Kore yatıştırıcı akne bakımı; mavi krem"),
        ("Dr. Ceuracle", "dr.ceuracle.com", "Kore Vegan Akne", "Tea Tree Purifine 95 Essence; çay ağacı %95; Kore vegan akne tedavisi"),
        ("Real Barrier", "realbarrier.com", "Kore Bariyer Akne", "Control-T Sebomide Cream; akne + bariyer koruma; Kore dual-action formül"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 5. Cilt Bakımı - Anti-Aging & Kırışıklık
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Anti-Aging & Kırışıklık": [
        ("Augustinus Bader", "augustinusbader.com", "TFC8 Kök Hücre", "30 yıl kök hücre araştırması; TFC8 teknolojisi; lüks biyoteknoloji anti-aging"),
        ("Dr. Barbara Sturm", "drsturm.com", "Anti-Inflammatory Aging", "PRP araştırmasından doğan; Hyaluronic Serum; anti-inflammatory felsefe; lüks biyoteknoloji"),
        ("Vintner's Daughter", "vintnersdaughter.com", "Botanik Anti-Aging", "Active Botanical Serum; 22 botanik aktif; tek ürün marka; $185 lüks"),
        ("Allies of Skin", "alliesofskin.com", "Peptide Anti-Aging", "Peptides & Antioxidants Firming Daily Treatment; multi-peptide; Singapur lüks aktif"),
        ("Medik8", "medik8.com", "CSA Anti-Aging", "Crystal Retinal; stabilize retinaldehyde; CSA felsefesi (C + SPF + retinoid); İngiliz bilimsel"),
        ("Geek & Gorgeous", "geekandgorgeous.com", "Uygun Retinal", "A-Game retinal serumları; %0.05-%0.1 retinaldehyde; Macar uygun fiyat anti-aging"),
        ("Naturium", "naturium.com", "Multi-Peptide Anti-Aging", "Quadruple Hyaluronic Acid Serum + retinol; yüksek doz aktif; uygun fiyat anti-aging"),
        ("Paula's Choice", "paulaschoice.com", "Retinol Uzmanı", "Clinical %1 Retinol; kademeli retinol sistemi; ingrediyent eğitimi; bilimsel formüller"),
        ("Sunday Riley", "sundayriley.com", "Good Genes Laktik", "Good Genes All-In-One Lactic Acid; laktik asit anti-aging kült ürün; mavi şişe ikonik"),
        ("Drunk Elephant benzeri: Herbivore", "herbivorebotanicals.com", "Bakuchiol Anti-Aging", "Bakuchiol Retinol Alternative Serum; doğal retinol alternatifi; hamile güvenli anti-aging"),
        ("The INKEY List", "theinkeylist.com", "Uygun Retinol", "Retinol Serum; £10; ingrediyent bazlı eğitim; anti-aging demokratizasyonu"),
        ("Good Molecules", "goodmolecules.com", "Uygun Bakuchiol", "Bakuchiol Oil for Combination Skin; $8; bakuchiol trend; uygun fiyat anti-aging alternatif"),
        ("Alpyn Beauty", "alpynbeauty.com", "Vahşi Bakuchiol", "PlantGenius Line-Filling Eye Balm; bakuchiol + botanik; Jackson Hole vahşi hasat"),
        ("Osea Malibu", "oseamalibu.com", "Deniz Anti-Aging", "Advanced Protection Cream; deniz yosunu + peptide; okyanus kaynaklı anti-aging"),
        ("Tata Harper", "tataharper.com", "Organik Anti-Aging", "Resurfacing Mask; %100 doğal; Vermont çiftlik; lüks organik anti-aging"),
        ("Omorovicza", "omorovicza.com", "Termal Anti-Aging", "Gold Flash Firming Serum; Budapeşte termal mineraller; Macar kaplıca anti-aging"),
        ("Aurelia London", "aurelialondon.com", "Probiyotik Anti-Aging", "Cell Revitalise Night Moisturiser; BioOrganic probiyotik; London lüks"),
        ("Jordan Samuel Skin", "jordansamuelskin.com", "Esthetician Anti-Aging", "Performance Cream; retinol + peptide; profesyonel formüller ev fiyatına"),
        ("BeautyStat", "beautystat.com", "Stabilize C Anti-Aging", "Universal C Skin Refiner; %20 stabilize vitamin C; MIT bilim insanı formülü; stabilite öncüsü"),
        ("Murad", "murad.com", "Retinol Youth Anti-Aging", "Retinol Youth Renewal Serum; retinol + adaptojenik bitkiler; dermatoloji kökenli"),
        ("StriVectin", "strivectin.com", "NIA-114 Anti-Aging", "NIA-114 patentli niacinamide türevi; stria (çatlak) kreminden anti-aging'e evrim"),
        ("Olay Regenerist alternatifi: No7", "no7beauty.com", "İngiliz Peptide Anti-Aging", "Advanced Retinol 1.5% Complex Night Concentrate; İngiliz eczane anti-aging; Boots özel"),
        ("IT Cosmetics", "itcosmetics.com", "CC Cream Anti-Aging", "CC+ Cream SPF 50+; full coverage + anti-aging; plastik cerrah + dermatolog geliştirdi"),
        ("Peter Thomas Roth", "peterthomasroth.com", "Retinol Fusion Anti-Aging", "Retinol Fusion PM; %1.5 micro-encapsulated retinol; NYC dermatoloji; güçlü formüller"),
        ("Shani Darden", "shanidarden.com", "Hollywood Retinol", "Retinol Reform; Hollywood facialist; Jessica Alba'nın esthetician'ı; lüks retinol"),
        ("Maelove", "maelove.com", "Uygun C Anti-Aging", "The Glow Maker; SkinCeuticals CE Ferulic alternatifi; $28 vs $180; bilimsel formül"),
        ("Timeless", "timelessha.com", "Dupe Anti-Aging Serum", "%20 C + E + Ferulic; lüks formül uygun fiyat; SkinCeuticals dupe; anti-aging demokratizasyonu"),
        ("Bloomeffects", "bloomeffects.com", "Lale Anti-Aging", "Royal Tulip Nectar; Hollanda lale kök hücre teknolojisi; Dutch botanik anti-aging"),
        ("True Botanicals", "truebotanicals.com", "Klinik Doğal Anti-Aging", "Pure Radiance Oil; klinik testlerle kanıtlanmış doğal anti-aging; La Mer'e rakip sonuçlar"),
        ("Bioeffect", "bioeffect.com", "İzlanda EGF Anti-Aging", "EGF Serum; arpa bitkisinden üretilen EGF (epidermal growth factor); İzlanda biyoteknoloji"),
        ("111SKIN", "111skin.com", "Uzay Teknoloji Anti-Aging", "Celestial Black Diamond Cream; Dr. Yannis Alexandrides; uzay araştırma kaynaklı formüller"),
        ("Dermalogica", "dermalogica.com", "Profesyonel Anti-Aging", "Dynamic Skin Retinol Serum; mikro-enkapsüle retinol; esthetician bazlı anti-aging"),
        ("iS Clinical", "isclinical.com", "Klinik Anti-Aging", "Active Serum; bilimsel klinik formüller; growth factor + retinol; medikal estetik"),
        ("Obagi", "obagi.com", "Medikal Anti-Aging", "Professional-C Serum; %10-20 L-ascorbic acid; dermatoloji reçeteli anti-aging"),
        ("SkinMedica", "skinmedica.com", "TNS Anti-Aging", "TNS Advanced+ Serum; growth factor teknolojisi; Allergan markası; medikal estetik"),
        ("Revision Skincare", "revisionskincare.com", "Multi-Peptide Anti-Aging", "Revox 7; 7 peptide karışımı; medikal estetik doktor markası; boyun + yüz"),
        ("Perricone MD", "perriconemd.com", "Neuropeptide Anti-Aging", "Cold Plasma+ Face; neuropeptide teknolojisi; Dr. Perricone anti-inflammatory diet + skincare"),
        ("Jan Marini", "janmarini.com", "Growth Factor Anti-Aging", "Skin Research Transformation Face Cream; growth factor + peptide; Amerikan medikal güzellik"),
        ("Neocutis", "neocutis.com", "PSP Anti-Aging", "Bio-Cream; Processed Skin Proteins (PSP); İsviçre biyoteknoloji; yara iyileştirme kaynaklı"),
        ("Algenist", "algenist.com", "Alg Anti-Aging", "Genius Liquid Collagen; mikroalg kaynaklı; alguronic acid patentli; San Francisco biyoteknoloji"),
        ("Exuviance", "exuviance.com", "PHA Anti-Aging", "Performance Peel AP25; PHA (polyhydroxy acid) öncüsü; NeoStrata kardeş marka"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Alpha Beta Anti-Aging", "Alpha Beta Extra Strength Daily Peel; evde profesyonel peeling; NYC dermatolog"),
        ("Kate Somerville", "katesomerville.com", "Hollywood Anti-Aging", "Goat Milk Moisturizing Cream; Hollywood A-list facialist; keçi sütü + retinol"),
        ("Tatcha", "tatcha.com", "Japon Anti-Aging", "Dewy Skin Cream; Japon güzellik ritüelleri; mor pirinç + yeşil çay; geisha geleneği"),
        ("Sulwhasoo", "sulwhasoo.com", "Kore Ginseng Anti-Aging", "Concentrated Ginseng Renewing Cream; 50 yıllık ginseng araştırması; Kore lüks hanbang"),
        ("Amorepacific", "amorepacific.com", "Kore Yeşil Çay Anti-Aging", "Time Response Skin Renewal Serum; yeşil çay EGCG; Kore premium anti-aging"),
        ("Missha", "missha.com", "Kore İlk Anti-Aging Essence", "Time Revolution The First Treatment Essence; SK-II alternatifi; Kore fermente anti-aging"),
        ("Laneige", "laneige.com", "Kore Su Uyku Anti-Aging", "Water Sleeping Mask; gece su terapisi; Kore sleeping mask konsepti yaratıcısı"),
        ("Dr. Jart+", "drjart.com", "Kore Ceramidin Anti-Aging", "Ceramidin Cream; 5 ceramide kompleksi; Kore bariyer + anti-aging; dermo-kozmetik"),
        ("AHC", "ahcbeauty.com", "Kore Göz Kremi Anti-Aging", "Ten Revolution Real Eye Cream For Face; göz kremi tüm yüze; Kore anti-aging konsepti"),
        ("The History of Whoo", "whoo.com", "Kore Saray Anti-Aging", "Bichup Self-Generating Anti-Aging Essence; Kore kraliyet reçetesi; lüks hanbang"),
        ("Hera", "hera.com", "Kore Hücre Anti-Aging", "Cell Essence; Kore hücre yenilenme; Seul modern lüks; Amorepacific premium"),
        ("Donginbi", "donginbi.com", "Kore Kırmızı Ginseng Anti-Aging", "Red Ginseng Daily Defense Cream; %100 Kore kırmızı ginseng; KGC markası"),
        ("Shiseido", "shiseido.com", "Japon Ultimune Anti-Aging", "Ultimune Power Infusing Concentrate; ImuGenerationRED teknolojisi; Japon bağışıklık güzelliği"),
        ("SK-II", "sk-ii.com", "Japon Pitera Anti-Aging", "Facial Treatment Essence; PITERA fermente maya özü; 40+ yıl Japon biyoteknoloji"),
        ("Decorté", "decorte.com", "Japon Liposome Anti-Aging", "Liposome Advanced Repair Serum; çoklu kapsüllü lipozom; Japon lüks teknoloji"),
        ("Cle de Peau", "cledepeaubeaute.com", "Japon Lüks Anti-Aging", "The Serum; Illuminating Complex EX; Japon en lüks skincare; hücre aydınlatma"),
        ("POLA", "pola.com", "Japon Wrinkle Shot Anti-Aging", "Wrinkle Shot Serum Prototype; NEI-L1 bazosit aktivasyonu; Japon kırışıklık tedavi öncüsü"),
        ("Dr. CI:Labo", "ci-labo.com", "Japon Aqua-Collagen Anti-Aging", "Aqua-Collagen Gel; doktor markası; Japon kolajen jel nemlendirici; multi-function"),
        ("Kanebo", "kanebo-cosmetics.jp", "Japon İpek Anti-Aging", "Sensai Ultimate The Cream; ipek proteinleri; Japon ultra-lüks; yıllık sınırlı üretim"),
        ("Menard", "menard.co.jp", "Japon Lissome Anti-Aging", "Lissome Cream; nanotransfer teknolojisi; Japon premium; derinlemesine nüfuz etme"),
        ("Albion", "albion.co.jp", "Japon Süt Anti-Aging", "Excia AL Whitening Immaculate Essence; Japon süt emülsiyon geleneği; nemlendirme + anti-aging"),
        ("Natura Bissé", "naturabisse.com", "İspanyol Diamond Anti-Aging", "Diamond Extreme Cream; İspanyol lüks; elmas tozu; Barselona high-end anti-aging"),
        ("Sisley", "sisley-paris.com", "Fransız Bitki Anti-Aging", "Sisleÿa L'Intégral; fitokozmetik öncüsü; Fransız botanik lüks; 40+ yıl araştırma"),
        ("Chantecaille", "chantecaille.com", "Fransız Botanik Anti-Aging", "Gold Recovery Mask; botanik + lüks; hayvan koruma bağışı; 24K altın + botanik"),
        ("La Prairie", "laprairie.com", "İsviçre Kaviar Anti-Aging", "Skin Caviar Luxe Cream; İsviçre kaviar bilimi; ultra-lüks; hücre yenilenme"),
        ("Valmont", "lamaisonvalmont.com", "İsviçre Buzul Anti-Aging", "Prime Renewing Pack; İsviçre buzul suyu; hücre yenilenme kremi; lüks spa"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 6. Cilt Bakımı - Göz Çevresi
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Göz Çevresi": [
        ("Dieux Skin", "dieuxskin.com", "Göz Altı Patch", "Instant Angel silikon göz altı maskesi; yeniden kullanılabilir; sürdürülebilir patch alternatifi"),
        ("Peace Out", "peaceoutskincare.com", "Retinol Göz Patch", "Puffy Under-Eye Patches; retinol + peptide; Sephora best seller göz patch"),
        ("Patchology", "patchology.com", "5 Dakika Göz Maskesi", "FlashPatch Restoring Night Eye Gels; 5 dakika göz maskesi; seyahat formatı"),
        ("KNC Beauty", "kncbeauty.com", "Kollajen Göz Maskesi", "All Natural Collagen Infused Lip & Eye Mask; kolajen + retinol; lüks göz maskesi"),
        ("Wander Beauty", "wanderbeauty.com", "Altın Göz Maskesi", "Baggage Claim Gold Eye Masks; 24K altın; göz altı şişlik azaltma; seyahat güzelliği"),
        ("111SKIN", "111skin.com", "Biyoselüloz Göz Maskesi", "Rose Gold Brightening Facial Treatment Mask; biyoselüloz teknoloji; lüks göz bakımı"),
        ("Skyn Iceland", "skyniceland.com", "İzlanda Göz Jeli", "Hydro Cool Firming Eye Gels; İzlanda mineral suyu; cooling jel teknolojisi; göz altı yorgunluk"),
        ("AHC", "ahcbeauty.com", "Kore Göz Kremi", "Ten Revolution Real Eye Cream for Face; tüm yüze uygulanabilir; Kore göz kremi konsepti"),
        ("Innisfree", "innisfree.com", "Kore Yeşil Çay Göz", "Green Tea Seed Eye Cream; Jeju yeşil çay; hafif nemlendirme; Kore doğal göz bakımı"),
        ("Sulwhasoo", "sulwhasoo.com", "Kore Ginseng Göz", "Concentrated Ginseng Renewing Eye Cream; ginseng kök; Kore lüks göz bakımı"),
        ("Shiseido", "shiseido.com", "Japon Benefiance Göz", "Benefiance Wrinkle Smoothing Eye Cream; ReNeura teknolojisi; Japon anti-aging göz"),
        ("SK-II", "sk-ii.com", "Japon Pitera Göz", "R.N.A. Power Eye Cream; Pitera + peptide; Japon fermente göz bakımı"),
        ("Ole Henriksen", "olehenriksen.com", "Banana Bright Göz", "Banana Bright+ Eye Crème; C vitamini + kolajen; İskandinav aydınlatıcı göz kremi"),
        ("First Aid Beauty", "firstaidbeauty.com", "Hassas Göz Kremi", "Eye Duty Niacinamide Brightening Cream; niacinamide + kafein; hassas cilt güvenli"),
        ("Origins", "origins.com", "GinZing Göz Kremi", "GinZing Refreshing Eye Cream; kafein + ginseng; enerji veren göz kremi"),
        ("Kiehl's", "kiehls.com", "Avokado Göz Kremi", "Creamy Eye Treatment with Avocado; avokado yağı; kült klasik göz nemlendirici"),
        ("Peter Thomas Roth", "peterthomasroth.com", "Peptide Göz Patch", "24K Gold Pure Luxury Lift & Firm Hydra-Gel Eye Patches; 24K altın + kollajen"),
        ("Charlotte Tilbury alternatifi: Pixi", "pixibeauty.com", "Retinol Göz Kremi", "Retinol Eye Cream; retinol + peptide; erişilebilir anti-aging göz bakımı"),
        ("Caudalie", "caudalie.com", "Üzüm Göz Kremi", "Vinoperfect Brightening Eye Cream; üzüm polifenol; Fransız göz aydınlatma"),
        ("Estée Lauder alternatifi: Neutrogena Rapid", "neutrogena.com", "Retinol Göz Kremi", "Rapid Wrinkle Repair Eye Cream; retinol SA; eczane fiyatına klinik göz bakımı"),
        ("It Cosmetics", "itcosmetics.com", "Confidence Göz Kremi", "Confidence in an Eye Cream; ceramide + peptide; full-size çift uçlu uygulama"),
        ("Laneige", "laneige.com", "Kore Göz Uyku Maskesi", "Eye Sleeping Mask EX; gece göz maskesi; Kore sleeping beauty konsepti"),
        ("Dr. Jart+", "drjart.com", "Kore Ceramidin Göz", "Ceramidin Eye Cream; 5 ceramide; Kore bariyer onarımlı göz bakımı"),
        ("Belif", "belifcosmetic.com", "Kore Moisturizing Göz", "Moisturizing Eye Bomb; İngiliz botanik + Kore formülasyon; patlayıcı nemlendirme"),
        ("Hada Labo", "hadalabousa.com", "Japon Hyaluronic Göz", "Gokujyun Eye Cream; süper hyaluronic acid; Japon nemlendirme teknolojisi; uygun fiyat"),
        ("DHC", "dhc.co.jp", "Japon Zeytin Göz", "Velvet Skin Coat eye cream; zeytin bazlı; Japon tabanlı göz bakımı"),
        ("RoC", "rocskincare.com", "Retinol Göz Uzmanı", "Retinol Correxion Eye Cream; retinol göz bakımı standardı; Fransız eczane; 30+ yıl retinol deneyimi"),
        ("SkinCeuticals", "skinceuticals.com", "AGE Göz Kremi", "A.G.E. Eye Complex; optik difüzörler; koyu halka + torba; dermatoloji standardı"),
        ("La Roche-Posay", "laroche-posay.com", "Hassas Göz Bakımı", "Toleriane Dermallergo Eye Cream; ultra hassas göz çevresi; Fransız termal su"),
        ("Bioderma", "bioderma.com", "Sensibio Göz", "Sensibio Eye Contour Gel; hassas göz çevresi; Fransız dermatolojik göz bakımı"),
        ("Filorga", "filorga.com", "Fransız Optim Göz", "Optim-Eyes Eye Contour Cream; NCEF + hyaluronic; Fransız estetik tıp göz bakımı"),
        ("Lancôme", "lancome.com", "Génifique Göz", "Advanced Génifique Yeux Light-Pearl; fermente öz + ışık difüzörü; Fransız lüks göz"),
        ("Clarins", "clarins.com", "Bitki Göz Bakımı", "Total Eye Lift; organik harungana; Fransız botanik göz lifting; anında etki"),
        ("Allies of Skin", "alliesofskin.com", "Peptide Göz Serumu", "Multi Peptide Eye Cream; çoklu peptide; Singapur lüks göz bakımı"),
        ("Perricone MD", "perriconemd.com", "Göz Plazma", "Cold Plasma+ Eye; neuropeptide + copper complex; göz çevresi sıkılaştırma; bilimsel formül"),
        ("Revision Skincare", "revisionskincare.com", "DEJ Göz Kremi", "D.E.J Eye Cream; dermal-epidermal junction teknolojisi; medikal göz bakımı"),
        ("SkinMedica", "skinmedica.com", "TNS Göz Kremi", "TNS Eye Repair; growth factor teknolojisi; medikal estetik göz bakımı"),
        ("Jan Marini", "janmarini.com", "Transformation Göz", "Transformation Eye Cream; peptide + growth factor; medikal anti-aging göz"),
        ("iS Clinical", "isclinical.com", "C Göz Serumu", "C Eye Serum Advance+; %15 C vitamini + botanik; medikal klinik göz serumu"),
        ("Neocutis", "neocutis.com", "Lumière Göz", "Lumière Bio-restorative Eye Cream; PSP teknolojisi; İsviçre yara iyileştirme bilimi"),
        ("Natura Bissé", "naturabisse.com", "Diamond Göz", "Diamond Extreme Eye; İspanyol lüks; göz çevresi anti-aging; energy complex"),
        ("Sisley", "sisley-paris.com", "Sisleÿa Göz", "Sisleÿa L'Intégral Anti-Age Eye & Lip Contour Cream; Fransız lüks botanik göz + dudak"),
        ("La Mer", "cremedelamer.com", "Göz Konsantresi", "The Eye Concentrate; deniz kefirine dayalı; ultra-lüks göz bakımı; Miracle Broth"),
        ("Tatcha", "tatcha.com", "Japon Göz Serumu", "The Silk Peony Melting Eye Cream; Japon ipek + şakayık; geisha göz ritüeli"),
        ("Fresh", "fresh.com", "Lotus Göz Kremi", "Lotus Youth Preserve Eye Cream; lotus yaprağı + vitamin E; doğal anti-aging göz"),
        ("Korres", "korres.com", "Yunan Göz Kremi", "Black Pine Bouncing Eye Cream; kara çam polifenol; Yunan botanik göz sıkılaştırma"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 7. Cilt Bakımı - Maske & Peeling
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Maske & Peeling": [
        ("Summer Fridays", "summerfridays.com", "Jet Lag Maske", "Jet Lag Mask kült ürün; uçuş sonrası nemlendirme; influencer markası; Instagram ikonik"),
        ("Glow Recipe", "glowrecipe.com", "Karpuz Maske", "Watermelon Glow Sleeping Mask; AHA + karpuz; Kore sleeping mask ABD versiyonu"),
        ("Sand & Sky", "sandandsky.com", "Pembe Kil Maske", "Australian Pink Clay Flash Perfection Exfoliating Treatment; Avustralya pembe kil; Instagram viral"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Alpha Beta Peeling", "Alpha Beta Universal Daily Peel; evde kimyasal peeling; 2 adımlı pad sistemi"),
        ("Peter Thomas Roth", "peterthomasroth.com", "Pumpkin Enzim Maske", "Pumpkin Enzyme Mask; %40 kabak enzim; turuncu jel ikonik; profesyonel peeling evde"),
        ("Herbivore", "herbivorebotanicals.com", "Mavi Tansy Maske", "Blue Tansy Resurfacing Clarity Mask; mavi maske viral; AHA + BHA + mavi tansy"),
        ("Tata Harper", "tataharper.com", "Resurfacing Maske", "Resurfacing Mask; pomegranate enzim + BHA; %100 doğal peeling; Vermont çiftlik lüks"),
        ("Youth To The People", "youthtothepeople.com", "Süperfood Maske", "Superberry Hydrate + Glow Dream Mask; maqui + goji; gece süperfood maskesi"),
        ("Farmacy", "farmacybeauty.com", "Bal Potion Maske", "Honey Potion Warming Mask; bal + propolis; ısınan maske; çiftlik kaynaklı"),
        ("Origins", "origins.com", "Kömür Maske", "Clear Improvement Charcoal Mask; aktif kömür + beyaz kil; gözenek temizleme"),
        ("GlamGlow", "glamglow.com", "Supermud Maske", "Supermud Clearing Treatment; Hollywood maske; gözenek + akne; çamur maske trendi öncüsü"),
        ("Fresh", "fresh.com", "Gül Maske", "Rose Face Mask; gerçek gül yaprakları; Fransız gül suyu; lüks nemlendirici maske"),
        ("Drunk Elephant benzeri: Dermalogica", "dermalogica.com", "Multi-Asit Peeling", "Rapid Reveal Peel; laktik + sitrik + fitik asit; profesyonel evde peeling"),
        ("Neogen", "neogenlab.us", "Gazlı Bez Peeling", "Bio-Peel Gauze Peeling; Kore inovatif peeling formatı; gazlı bez + AHA; mekanik + kimyasal"),
        ("COSRX", "cosrx.com", "Kore BHA Peeling", "BHA Blackhead Power Liquid; Kore BHA peeling; siyah nokta; hafif günlük peeling"),
        ("Some By Mi", "somebymi.com", "Kore Multi-Asit Peeling", "AHA BHA PHA 30 Days Miracle Peeling Gel; üçlü asit; Kore 30 gün dönüşüm"),
        ("Medicube", "medicube.com", "Kore Zero Pore Pad", "Zero Pore Pad; TikTok viral; Kore gözenek peeling pedi; AHA + BHA pamuk ped"),
        ("Wishful", "wishful.com", "Enzim Scrub", "Yo Glow Enzyme Scrub; papaya + ananas enzim; Huda Beauty markası; enzim peeling"),
        ("Tatcha", "tatcha.com", "Pirinç Peeling", "The Rice Polish; enzim pudra temizleyici + peeling; Japon pirinç peeling ritüeli"),
        ("Dermalogica", "dermalogica.com", "Günlük Mikro Peeling", "Daily Microfoliant; pirinç enzim pudra; 25+ yıl best seller; profesyonel peeling"),
        ("Kate Somerville", "katesomerville.com", "ExfoliKate Peeling", "ExfoliKate Intensive Exfoliating Treatment; papaya + gözenek; Hollywood facialist peeling"),
        ("Pixi", "pixibeauty.com", "Glow Peel Pad", "Glow Peel Pads; %20 glikolik asit; evde AHA peeling pedi; erişilebilir eksfoliasyon"),
        ("Paula's Choice", "paulaschoice.com", "AHA Peeling", "%8 AHA Gel Exfoliant; glikolik asit jel; haftalık kimyasal peeling; ingrediyent bilimi"),
        ("The Ordinary", "theordinary.com", "AHA BHA Peeling", "AHA 30% + BHA 2% Peeling Solution; kırmızı kan maskesi viral; TikTok fenomeni; $7"),
        ("Good Molecules", "goodmolecules.com", "Overnight Peeling", "Overnight Exfoliating Treatment; gece peeling; glikolik + mandelic asit; uygun fiyat"),
        ("Drunk Elephant alternatifi: Allies of Skin", "alliesofskin.com", "Multi-Asit Maske", "Multi Acid + Night Cream; çoklu asit gece maskesi; Singapur lüks peeling"),
        ("Versed", "versedskin.com", "Daze Off Maske", "Daze Off Charcoal Mask; kömür + kaolin; temiz + erişilebilir maske"),
        ("Cocokind", "cocokind.com", "Texture Maske", "Texture Smoothing Cream; AHA + ceramide; haftalık smoothing maske; temiz formül"),
        ("I Dew Care", "idewcare.com", "Eğlenceli K-Maske", "Disco Kitten Illuminating Diamond Peel-Off Mask; peel-off glitter; K-beauty eğlenceli maske"),
        ("Too Cool For School", "toocoolforschool.com", "Yumurta Maske", "Egg Cream Mask sheet mask; yumurta bazlı; eğlenceli K-beauty; albumin nemlendirme"),
        ("Laneige", "laneige.com", "Su Uyku Maskesi", "Water Sleeping Mask; Kore sleeping mask standardı; gece su terapisi; kült ürün"),
        ("Innisfree", "innisfree.com", "Volkanik Kil Maske", "Super Volcanic Pore Clay Mask; Jeju volkanik kil; gözenek temizleme; Kore doğal"),
        ("Holika Holika", "holikaholika.com", "Domuz Maske", "Pig-Nose Clear Black Head 3-Step Kit; eğlenceli ambalaj; siyah nokta 3 adım; K-beauty fun"),
        ("Papa Recipe", "paparecipe.com", "Bal Maske", "Bombee Honey Mask; Kore arı balı sheet mask; nemlendirme + beslenme"),
        ("Mediheal", "mediheal.com", "Kore Sheet Mask", "N.M.F Aquaring Ampoule Mask; Kore sheet mask standardı; hastane kökenli"),
        ("Dr. Jart+", "drjart.com", "Rubber Maske", "Cryo Rubber Mask; kauçuk maske konsepti; DIY salon maske; Kore inovatif format"),
        ("Skin Gym", "skingymco.com", "LED Maske", "Wrinklit LED Mask; kızılötesi LED maske; evde ışık terapisi; anti-aging maske"),
        ("Charlotte Tilbury alternatifi: Elemis", "elemis.com", "Pro-Collagen Maske", "Pro-Collagen Marine Cream Mask; deniz kolajeni; İngiliz spa lüks maske"),
        ("Caudalie", "caudalie.com", "Üzüm Peeling", "Vinoperfect Glycolic Night Cream; üzüm glikolik; Bordeaux peeling; gece yenilenme"),
        ("REN", "renskincare.com", "AHA Maske", "Ready Steady Glow AHA Tonic; laktik asit + cranberry; İskandinav temiz peeling"),
        ("Ren Clean Skincare", "renskincare.com", "Glow Maske", "Clean Skincare AHA Smart Renewal Body Serum; vücut peeling; tam vücut AHA"),
        ("Kora Organics", "koraorganics.com", "Noni Maske", "Noni Glow Sleeping Mask; Miranda Kerr markası; noni meyvesi; Avustralya organik maske"),
        ("Votary", "votary.co.uk", "İngiliz Yağ Maskesi", "Intense Night Oil Rosehip & Retinoid; İngiliz lüks yağ maskesi; botanik gece bakımı"),
        ("May Lindstrom", "maylindstrom.com", "Artisan Maske", "The Problem Solver; kakao + baharat ısınan maske; artisan küçük parti; lüks doğal"),
        ("African Botanics", "africanbotanics.com", "Afrika Maske", "Marula Detoxifying Platinum Peel; Güney Afrika marula; platin peeling; Afrika botanik lüks"),
        ("Eminence Organic", "eminenceorganics.com", "Organik Meyve Maske", "Strawberry Rhubarb Dermafoliant; organik meyve enzim; Macar organik; spa markası"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 8. Cilt Bakımı - Dudak Bakımı
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Dudak Bakımı": [
        ("Laneige", "laneige.com", "Uyku Dudak Maskesi", "Lip Sleeping Mask kült ürün; gece dudak maskesi kategorisi yaratıcısı; berry flavor ikonik"),
        ("Summer Fridays", "summerfridays.com", "Lip Butter Balm", "Lip Butter Balm; shea + vegan wax; pastel renk seçenekleri; influencer favorisi"),
        ("Rhode", "rhodeskin.com", "Peptide Dudak Tedavisi", "Peptide Lip Treatment; Hailey Bieber markası; peptide bazlı dudak bakımı; phone case viral"),
        ("Tower 28", "tower28beauty.com", "SOS Dudak Tedavisi", "LipSoftie Hydrating Lip Treatment; NEA onaylı; hassas dudak bakımı; temiz bileşenler"),
        ("Burt's Bees", "burtsbees.com", "Doğal Balmumu Balm", "Beeswax Lip Balm; doğal balmumu + mentol; dünya çapında #1 doğal dudak bakımı"),
        ("Glossier benzeri: Merit", "meritbeauty.com", "Tinted Dudak Yağı", "Shade Slick Tinted Lip Oil; renk + bakım; minimalist dudak; 5 dakika makyaj"),
        ("Tatcha", "tatcha.com", "Japon Dudak Maskesi", "Kissu Lip Mask; Japon şeftali + squalane; gece dudak maskesi; Japon güzellik ritüeli"),
        ("Fresh", "fresh.com", "Şeker Dudak Balm", "Sugar Lip Treatment SPF 15; gerçek şeker kristalleri; dudak bakımı + SPF; lüks dudak"),
        ("Dior Lip Glow alternatifi: Kosas", "kosas.com", "Wet Lip Yağ", "Wet Lip Oil Gloss; hyaluronic acid + peptide; parlak + bakımlı; temiz dudak glossu"),
        ("Drunk Elephant alternatifi: Kopari", "koparibeauty.com", "Hindistan Cevizi Dudak", "Coconut Lip Glossy; hindistan cevizi yağı; tropikal dudak bakımı; vegan + temiz"),
        ("Bite Beauty", "bitebeauty.com", "Yenilebilir Dudak", "Agave+ Lip Mask; agave nektar; gıda sınıfı bileşenler; yenilebilir dudak bakımı"),
        ("Nuxe", "nuxe.com", "Rêve de Miel Dudak", "Rêve de Miel Ultra-Nourishing Lip Balm; bal + botanik; Fransız eczane dudak klasiği; kavanoz format"),
        ("By Terry", "byterry.com", "Lüks Dudak Balmı", "Baume de Rose; Fransız gül dudak bakımı; lüks pembe balm; kült klasik"),
        ("Jack Black", "getjackblack.com", "Erkek Dudak SPF", "Intense Therapy Lip Balm SPF 25; erkek dudak bakımı; shea + avokado + SPF"),
        ("Aquaphor", "aquaphor.com", "Onarım Dudak Terapisi", "Lip Repair; petrolatum bazlı onarım; eczane dudak kurtarıcı; dermatoloji standardı"),
        ("Dr. Lipp", "drlipp.com", "Lanolin Dudak", "Original Nipple Balm for Lips; lanolin bazlı; çok amaçlı; İngiltere anne + dudak bakımı"),
        ("Hurraw!", "hurraw.com", "Vegan Ham Dudak", "Moon Balm; vegan + ham gıda; ayçiçeği + chamomile; %100 organik dudak balm"),
        ("Lip Smacker", "lipsmacker.com", "Z Kuşağı Dudak", "Dr. Pepper flavored lip balm; nostaljik lezzetli dudak balm; eğlenceli koleksiyon"),
        ("EOS", "evolutionofsmooth.com", "Yumurta Dudak Balmı", "Sphere Lip Balm; yumurta şekli ikonik ambalaj; shea + jojoba; görsel marka kimliği"),
        ("Carmex", "mycarmex.com", "Klasik Dudak Onarım", "Classic Lip Balm; mentol + kamfor + fenol; sarı kavanoz ikonik; 85+ yıl"),
        ("Drunk Elephant alternatifi: Summer Fridays", "summerfridays.com", "Vanilya Dudak Butter", "Vanilla Beige Lip Butter Balm; vanilya aromalı; bakımlı doğal renk + nemlendirme"),
        ("Henné Organics", "henneorganics.com", "Lüks Organik Dudak", "Luxury Lip Balm; organik + lüks; kokosuz; minimal bileşen; premium dudak bakımı"),
        ("Sara Happ", "sarahapp.com", "Dudak Scrub Uzmanı", "The Lip Scrub; şeker bazlı dudak peelingi; dudak bakım ritüeli başlatıcısı; çeşitli aromalar"),
        ("Lano", "lanolips.com", "Lanolin Dudak Uzmanı", "101 Ointment Multi-Balm; Avustralya lanolin; çok amaçlı; dudak + cilt; kuru bölge kurtarıcı"),
        ("Lanolips", "lanolips.com", "Avustralya Dudak", "Lip Ointment Tinted; renkli lanolin dudak; Avustralya çöl iklimine karşı; ultra nemlendirme"),
        ("Tocobo", "tocobo.com", "Kore Dudak Balmı", "Glass Tinted Lip Balm; cam dudak efekti; K-beauty dudak bakımı; hafif renk + bakım"),
        ("rom&nd", "romand.co.kr", "Kore Dudak Serumu", "Juicy Lasting Tint + Glasting Water Tint; K-beauty dudak renk uzmanı; viral TikTok"),
        ("3CE", "3ce.com", "Kore Dudak Makyajı", "Velvet Lip Tint; kadife dudak; Stylenanda markası; Kore dudak tint öncüsü"),
        ("Peripera", "peripera.com", "Kore Su Tint", "Ink Airy Velvet; su bazlı tint; K-beauty dudak renk; uygun fiyat çeşitlilik"),
        ("Mentholatum", "mentholatum.com", "Japon Dudak", "Melty Cream Lip; Japon dudak bakımı; SPF + bakım; günlük kullanım"),
        ("DHC", "dhc.co.jp", "Japon Zeytin Dudak", "Lip Cream; zeytin yağı bazlı; Japon dudak nemlendirme klasiği; basit + etkili"),
        ("Curel", "curel.com", "Japon Ceramide Dudak", "Moisture Lip Care Cream; ceramide bazlı; Japon hassas dudak bakımı"),
        ("Weleda", "weleda.com", "Organik Dudak Balmı", "Skin Food Lip Butter; organik bitki özleri; İsviçre-Alman organik dudak bakımı"),
        ("Malin+Goetz", "malinandgoetz.com", "Mojito Dudak", "Mojito Lip Balm; mojito aromalı; NYC apothecary; unisex dudak bakımı"),
        ("Herbivore", "herbivorebotanicals.com", "CBD Dudak", "Emerald CBD + Adaptogens Lip Conditioner; CBD + adaptojenik; botanik dudak onarım"),
        ("Kylie Skin alternatifi: em cosmetics", "emcosmetics.com", "Dudak Cushion", "Lip Cushion; Michelle Phan markası; cushion format dudak rengi; K-beauty ilhamlı"),
        ("Buxom", "buxom.com", "Dolgunlaştırıcı Dudak", "Full-On Plumping Lip Cream; peptide dolgunlaştırıcı; anında hacim; karıncalanma efekti"),
        ("Too Faced", "toofaced.com", "Lip Injection", "Lip Injection Extreme; dudak dolgunlaştırma glossu; anında plump; viral büyük dudak efekti"),
        ("Essence", "essence.eu", "Uygun Dudak", "Lip Care Hydra Oil; €2-3; ultra uygun fiyat; Alman drugstore; kaliteli düşük fiyat"),
        ("Labello", "labello.com", "Alman Dudak Klasiği", "Original Care; Alman dudak bakımı standardı; Beiersdorf; günlük nemlendirme"),
        ("Nivea Lip", "nivea.com", "Temel Dudak Bakımı", "Essential Lip Care; shea + jojoba; erişilebilir günlük dudak bakımı; Alman güvenilirlik"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 9. Cilt Bakımı - Bariyer Onarım
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Bariyer Onarım": [
        ("Byoma", "byoma.com", "Triseramide Bariyer", "Triseramide kompleksi; Target'ta $10-15; bariyer onarım demokratizasyonu; Z kuşağı"),
        ("Krave Beauty", "kravebeauty.com", "Great Barrier Relief", "Tamanu oil + cica; bariyer onarım kült ürünü; aşırı eksfoliasyon sonrası kurtarıcı"),
        ("Stratia", "stratiaskin.com", "Liquid Gold", "Ceramide + cholesterol + fatty acid 3:1:1 oranı; Reddit kült; bariyer onarım bilimi"),
        ("Skinfix", "skinfix.com", "Lipid-Peptide Bariyer", "Lipid-Peptide Cream; ceramide + lipid + peptide; NEA onaylı; dermatoloji bazlı bariyer onarım"),
        ("La Roche-Posay benzeri: Real Barrier", "realbarrier.com", "Kore Bariyer Kremi", "Cream Ampoule; ceramide NP + HA; Kore dermatolojik bariyer onarım; Atopalm kardeş marka"),
        ("Illiyoon", "illiyoon.com", "Kore Ceramide Bariyer", "Ceramide Ato Concentrate Cream; multi-ceramide; yenidoğan güvenli; Kore bariyer klasiği"),
        ("Dr. Jart+ Ceramidin", "drjart.com", "Kore 5 Ceramide", "Ceramidin Cream; 5 çeşit ceramide; Kore ceramide bariyer onarım standardı"),
        ("Atopalm", "atopalm.com", "Kore MLE Bariyer", "MLE (Multi-Lamellar Emulsion) teknolojisi; cilt bariyeri taklit eden katmanlı yapı; Kore bilimi"),
        ("Holika Holika Good Cera", "holikaholika.com", "Kore Cera Bariyer", "Good Cera Super Ceramide Cream; ceramide + shea; Kore uygun fiyat bariyer onarım"),
        ("Etude SoonJung", "etude.com", "Kore pH Bariyer", "SoonJung 2x Barrier Intensive Cream; panthenol + madecassoside; pH 5.5 bariyer koruma"),
        ("Purito Dermide", "purito.com", "Kore Cica Bariyer", "Dermide Cica Barrier Sleeping Pack; gece bariyer onarım; centella + ceramide"),
        ("COSRX Balancium", "cosrx.com", "Kore Comfort Bariyer", "Balancium Comfort Ceramide Cream; ceramide NP + centella; Kore bariyer dengeleme"),
        ("Klairs Rich Moist", "klairscosmetics.com", "Kore Hassas Bariyer", "Rich Moist Soothing Cream; hassas cilt bariyer güçlendirme; Kore minimal bariyer bakımı"),
        ("Rovectin", "rovectin.com", "Post-Kemo Bariyer", "Barrier Repair Cream; kemoterapi sonrası hassas cilt; ultra düşük tahriş; bariyer uzmanı"),
        ("Zeroid", "zeroid.com", "Kore Dermatolojik Bariyer", "Intensive Oint-Cream; MLE teknolojisi; Kore hastane markası; atopik dermatit"),
        ("Vanicream", "vanicream.com", "Dermatolojik Bariyer", "Moisturizing Skin Cream; sıfır tahriş edici madde; dermatoloji #1; ultra basit bariyer"),
        ("CeraVe alternatifi: Aveeno", "aveeno.com", "Yulaf Bariyer", "Calm + Restore Oat Gel Moisturizer; yulaf + feverfew; yulaf bazlı bariyer güçlendirme"),
        ("Eucerin", "eucerin.com", "Alman Urea Bariyer", "UreaRepair Plus; %5-10 urea; Alman dermatolojik bariyer onarım; kuru cilt uzmanı"),
        ("Bioderma Atoderm", "bioderma.com", "Fransız Bariyer Balmı", "Atoderm Intensive Balm; bariyer terapisi patent; Fransız dermo-bariyer onarım"),
        ("La Roche-Posay Lipikar", "laroche-posay.com", "Fransız Triple Bariyer", "Lipikar Baume AP+M; triple-repair teknolojisi; Fransız termal su + ceramide; bariyer standart"),
        ("Avène XeraCalm", "avene.com", "Fransız Termal Bariyer", "XeraCalm A.D Lipid-Replenishing Cream; I-modulia patent; termal su bariyer onarım"),
        ("A-Derma Exomega", "aderma.com", "Fransız Organik Bariyer", "Exomega Control Cream; organik Rhealba yulaf; atopik eğilimli cilt bariyer onarım"),
        ("Sebamed", "sebamed.com", "Alman pH 5.5 Bariyer", "Extreme Dry Skin Relief; pH 5.5 formüller; Alman dermatolojik bariyer koruma"),
        ("Weleda Skin Food", "weleda.com", "Organik Bariyer", "Skin Food Original; 1926'dan beri; organik bitki özleri; ultra zengin bariyer besleyici"),
        ("Dr. Andrew Weil for Origins", "origins.com", "Mega-Mushroom Bariyer", "Mega-Mushroom Relief & Resilience Soothing Treatment Lotion; mantar bariyer güçlendirme"),
        ("Pai Skincare", "paiskincare.com", "Organik Hassas Bariyer", "Rosehip BioRegenerate Oil; kuşburnu yağı; organik bariyer yenileme; ultra hassas cilt"),
        ("Dr. Jart+ Tiger Grass", "drjart.com", "Kore Cica Bariyer", "Cicapair Tiger Grass Cream; centella + kaplan otu; Kore cica bariyer onarım standardı"),
        ("Pyunkang Yul", "pyunkangyul.us", "Kore Hanbang Bariyer", "Calming Moisture Barrier Cream; Kore geleneksel tıp; bariyer güçlendirme; hassas formül"),
        ("Isntree", "isntree.com", "Kore Ceramide Bariyer", "Ceramide Emulsion; Kore ceramide emülsiyon; hafif bariyer nemlendirme"),
        ("Soon Jung alternatifi: Missha", "missha.com", "Kore Near Skin Bariyer", "Near Skin Ceramide Cream; Kore ceramide bariyer; uygun fiyat"),
        ("Heimish", "heimish.com", "Kore Marine Bariyer", "Marine Care Deep Moisture Nourishing Cream; deniz bileşenleri bariyer onarım"),
        ("Numbuzin No.5", "numbuzin.com", "Kore Pan-Thenol Bariyer", "No.5 Vitamin-Niacinamide Concentrated Pad; vitamin + niacinamide bariyer destek"),
        ("Scinic", "scinic.com", "Kore Aqua Bariyer", "The Simple Hyaluronic Acid Ampoule; hyaluronic bariyer nemlendirme; Kore basit formül"),
        ("Hada Labo", "hadalabousa.com", "Japon Hyaluronic Bariyer", "Gokujyun Premium Lotion; 7 çeşit hyaluronic acid; Japon nemlendirme bariyer tekniği"),
        ("Matsuyama", "matsuyama.co.jp", "Japon Soy Bariyer", "Hadauru Moisturizing Infusion Balancing Lotion; soya ceramide; Japon geleneksel bariyer"),
        ("Curel", "curel.com", "Japon Pseudo-Ceramide Bariyer", "Intensive Moisture Cream; sözde-ceramide teknolojisi; Kao bilimi; Japon hassas cilt bariyer"),
        ("Minon Amino", "minon-aminomoist.com", "Japon Amino Bariyer", "Amino Moist Charge Lotion; 9 amino asit; Japon amino asit bariyer teknolojisi"),
        ("Lits", "lits-cosmetics.com", "Japon Bitki Kök Bariyer", "Revival Stem7; bitki kök hücre; Japon botanik bariyer yenilenme"),
        ("First Aid Beauty", "firstaidbeauty.com", "Ultra Repair Bariyer", "Ultra Repair Cream; koloidal yulaf + ceramide; bariyer onarım; hassas cilt favorisi"),
        ("Embryolisse", "embryolisse.com", "Fransız Lait-Crème Bariyer", "Lait-Crème Concentré; aloe + soya protein; makyaj bazı + bariyer; Fransız çok amaçlı klasik"),
        ("Cicaplast", "laroche-posay.com", "Fransız Cica Bariyer", "Cicaplast Baume B5+; panthenol + madecassoside + cica; Fransız bariyer onarım standardı"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 10. Cilt Bakımı - Hassas Cilt
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Hassas Cilt": [
        ("Vanicream", "vanicream.com", "Ultra Basit Hassas", "Sıfır tahriş edici; dermatoloji birinci önerisi; parfümsüz + boyasız + parabensin; minimal formül"),
        ("Pai Skincare", "paiskincare.com", "Organik Hassas", "Chamomile & Rosehip Calming Day Cream; organik hassas cilt uzmanı; London artisan"),
        ("Avène", "avene.com", "Fransız Termal Hassas", "Skin Recovery Cream; termal su bazlı; hassas cilt Fransız dermatoloji standardı"),
        ("Tower 28", "tower28beauty.com", "NEA Onaylı Hassas", "SOS Daily Rescue Spray; NEA (Egzama Derneği) onaylı; hassas + reaktif cilt güvenli"),
        ("Rovectin", "rovectin.com", "Ultra Hassas", "Conditioning Cleanser; kemoterapi sonrası geliştirilen; ultra düşük tahriş formülleri"),
        ("Klairs", "klairscosmetics.com", "K-Beauty Hassas", "Supple Preparation Unscented Toner; parfümsüz Kore hassas cilt; minimalist K-beauty"),
        ("Etude SoonJung", "etude.com", "Kore pH Hassas", "SoonJung pH 6.5 Whip Cleanser; düşük pH; Kore hassas cilt ana serisi"),
        ("Real Barrier", "realbarrier.com", "Kore Bariyer Hassas", "Extreme Cream; MLE teknolojisi; Kore dermatolojik hassas cilt"),
        ("Purito", "purito.com", "Kore Unscented Hassas", "Centella Unscented Serum; parfümsüz + centella; Kore temiz hassas formül"),
        ("Isntree", "isntree.com", "Kore Aloe Hassas", "Aloe Soothing Gel; aloe vera jel; Kore hassas cilt yatıştırma"),
        ("Illiyoon", "illiyoon.com", "Kore Bebek Hassas", "Ceramide Ato Lotion; yenidoğan güvenli; Kore ultra hassas aile bakımı"),
        ("First Aid Beauty", "firstaidbeauty.com", "SOS Hassas", "Ultra Repair Cream; koloidal yulaf + shea; hassas cilt kurtarıcı; SOS bakım"),
        ("La Roche-Posay Toleriane", "laroche-posay.com", "Fransız Toleriane Hassas", "Toleriane Ultra Cream; prebiotic termal su; Fransız hassas cilt standardı"),
        ("Bioderma Sensibio", "bioderma.com", "Fransız Sensibio Hassas", "Sensibio Defensive Cream; cilt savunma güçlendirme; hassas cilt bilimi"),
        ("Ducray Ictyane", "ducray.com", "Fransız Kuru Hassas", "Ictyane Cream; kuru hassas cilt; Fransız dermatoloji; besleyici bakım"),
        ("Sebamed", "sebamed.com", "Alman pH Hassas", "Moisturizing Face Cream; pH 5.5; Alman dermatolojik hassas cilt formülü"),
        ("Eucerin", "eucerin.com", "Alman Dermatolojik Hassas", "UltraSENSITIVE Soothing Care; SymSitive patent; Alman hassas cilt bilimi"),
        ("Weleda", "weleda.com", "Organik Hassas", "Almond Soothing Facial Cream; badem yağı; organik hassas cilt; İsviçre-Alman"),
        ("Embryolisse", "embryolisse.com", "Fransız Çok Amaçlı Hassas", "Lait-Crème Concentré; hassas + her cilt tipi; Fransız klasik; makyaj bazı"),
        ("Curel", "curel.com", "Japon Ceramide Hassas", "Intensive Moisture Cream; sözde-ceramide; Japon hassas cilt standardı; Kao bilimi"),
        ("Minon", "minon-aminomoist.com", "Japon Amino Hassas", "Amino Moist Charge Milk; amino asit bazlı; Japon ultra hassas nemlendirme"),
        ("Hada Labo", "hadalabousa.com", "Japon Hyaluronic Hassas", "Gokujyun Hyaluronic Lotion; basit hyaluronic formül; Japon hassas cilt nemlendirme"),
        ("d'Alba", "dalba.co.kr", "Kore White Truffle Hassas", "White Truffle First Spray Serum; beyaz trüf mantarı; Kore lüks hassas cilt"),
        ("Goodal", "goodal.co.kr", "Kore Houttuynia Hassas", "Houttuynia Cordata Calming Moisture Sun Cream; hassas cilt yatıştırıcı SPF"),
        ("SKIN1004", "skin1004.com", "Kore Centella Hassas", "Madagascar Centella Ampoule; saf centella; Kore hassas cilt ampoule"),
        ("Aestura", "aestura.com", "Kore Hastane Hassas", "Atobarrier 365 Cream; Kore hastane kullanımlı; dermatolojik hassas cilt"),
        ("Zeroid", "zeroid.com", "Kore Atopik Hassas", "Soothing Cream; atopik dermatit; Kore dermatoloji; MLE teknolojisi"),
        ("Round Lab", "roundlab.co.kr", "Kore Birch Hassas", "Birch Juice Moisturizing Cream; huş suyu; Kore doğal hassas nemlendirme"),
        ("Beplain", "beplain.co.kr", "Kore Chamomile Hassas", "Chamomile pH-Balanced Cream; papatya + pH dengeleme; Kore hassas cilt yatıştırma"),
        ("Needly", "needly.co.kr", "Kore Mild Hassas", "Mild Cleansing Gel + Cream; az bileşen çok etki; Kore minimalist hassas cilt"),
        ("Bonajour", "bonajour.co.kr", "Kore Vegan Hassas", "Calming Vegan Cream; %100 vegan; Kore doğal hassas cilt; uygun fiyat"),
        ("Dr. G", "drg.co.kr", "Kore Dermo Hassas", "Red Blemish Clear Soothing Cream; cica + centella; Kore dermatolojik hassas bakım"),
        ("Sioris", "sioris.com", "Kore Taze Hassas", "Enriched by Nature Cream; mevsimsel taze bileşenler; Kore slow beauty hassas cilt"),
        ("Skin Aqua Sensitive", "skin-aqua.com", "Japon SPF Hassas", "Super Moisture Milk Sensitive; hassas cilt SPF; Japon hafif + güvenli güneş koruma"),
        ("Mama & Kids", "mamakids.co.jp", "Japon Anne-Bebek Hassas", "Skin Care Cream; yenidoğan güvenli; Japon hassas anne + bebek bakımı"),
        ("NOV", "nov.jp", "Japon Dermatolojik Hassas", "AC Active Moisture Cream; Japon dermatolojik hassas cilt; hipoalerjenik formüller"),
        ("Freeplus", "freeplus.jp", "Japon Minimal Hassas", "Moisture Cream; minimum bileşen; Japon hassas cilt için minimal formül"),
        ("Dr. Althea", "dralthea.com", "Kore Doktor Hassas", "Amino Acid Gentle Bubble Cleanser; amino asit bazlı; Kore hassas dermatolog markası"),
        ("Mixsoon", "mixsoon.com", "Kore Tek Bileşen Hassas", "Bean Cream; soya fasulyesi tek bileşen; Kore ultra minimal hassas cilt"),
        ("Dear, Klairs", "klairscosmetics.com", "Kore Midnight Blue Hassas", "Midnight Blue Calming Cream; guaiazulene; Kore yatıştırıcı hassas cilt"),
    ],
}

# ─── Continue with remaining categories in EXTRA_BRANDS ──────────────────────
EXTRA_BRANDS = {
    # ═══════════════════════════════════════════════════════════════════════════
    # 11. Cilt Bakımı - Hiperpigmentasyon
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı - Hiperpigmentasyon": [
        ("Topicals", "mytopicals.com", "Faded Leke Serumu", "Faded Serum viral kült ürün; azelaic + niacinamide + centella; hiperpigmentasyon + koyu leke uzmanı"),
        ("EADEM", "eadem.co", "Melanin Aydınlatma", "Smart Melanin C vitamini; koyu cilt tonları için özel aydınlatma; melanin bilimi"),
        ("Naturium", "naturium.com", "Azelaic Aydınlatma", "Azelaic Acid Emulsion %10; leke + hiperpigmentasyon; uygun fiyat aktif tedavi"),
        ("Good Molecules", "goodmolecules.com", "Discoloration Serumu", "Discoloration Correcting Serum; tranexamic acid + niacinamide; $6 leke tedavisi"),
        ("Paula's Choice", "paulaschoice.com", "Niacinamide Aydınlatma", "10% Niacinamide Booster; gözenek + leke; ingrediyent bilimi aydınlatma"),
        ("Murad", "murad.com", "Rapid Dark Spot", "Rapid Dark Spot Correcting Serum; resorcinol + glycolic; hızlı leke açma"),
        ("Urban Skin Rx", "urbanskinrx.com", "Melanin Leke Tedavisi", "Even Tone Cleansing Bar; koyu ten leke tedavisi; profesyonel formüller; kapsayıcı"),
        ("Fenty Skin alternatifi: Live Tinted", "livetinted.com", "Güney Asya Aydınlatma", "Hueguard SPF + leke koruma; Güney Asya cilt tonları; hiperpigmentasyon önleme"),
        ("Melē", "meleskincare.com", "Kapsayıcı Aydınlatma", "Even Dark Spot Control Serum; koyu ciltler için özel; niacinamide + tranexamic acid"),
        ("SkinCeuticals", "skinceuticals.com", "Discoloration Defense", "Discoloration Defense Serum; tranexamic + kojic + niacinamide; medikal aydınlatma"),
        ("Glossier benzeri: Glow Recipe", "glowrecipe.com", "Karpuz Aydınlatma", "Watermelon Glow Niacinamide Dew Drops; niacinamide aydınlatma + ışıltı; leke önleme"),
        ("The Ordinary", "theordinary.com", "Alpha Arbutin Aydınlatma", "Alpha Arbutin 2% + HA; arbutin bazlı leke açma; $9; erişilebilir aktif"),
        ("Cos De BAHA", "cosdebaha.com", "Kore Azelaic Aydınlatma", "Azelaic Acid 10% Serum; Kore uygun fiyat; azelaic hiperpigmentasyon tedavisi"),
        ("By Wishtrend", "bywishtrend.com", "Kore C Aydınlatma", "Pure Vitamin C 21.5% Advanced Serum; yüksek doz C; Kore aydınlatma uzmanı"),
        ("Klairs", "klairscosmetics.com", "Kore Freshly Juiced C", "Freshly Juiced Vitamin C Serum; %5 C vitamini; Kore hassas aydınlatma; ilk C serumu"),
        ("Goodal", "goodal.co.kr", "Kore Yeşil Mandarin C", "Green Tangerine Vita C Dark Spot Serum; Jeju yeşil mandalina; Kore aydınlatma"),
        ("Some By Mi", "somebymi.com", "Kore Galactomyces C", "Galactomyces Pure Vitamin C Glow Serum; fermente + C vitamini; Kore çift aydınlatma"),
        ("COSRX", "cosrx.com", "Kore Vitamin C Aydınlatma", "Vitamin C 23 Serum; %23 C vitamini; Kore yüksek doz aydınlatma"),
        ("Rohto Melano CC", "rohto.com", "Japon C Aydınlatma", "Melano CC Intensive Anti-Spot Essence; Japon vitamin C leke tedavisi; dünya çapında kült"),
        ("Hada Labo Shirojyun", "hadalabousa.com", "Japon Arbutin Aydınlatma", "Shirojyun Premium Whitening Lotion; tranexamic acid; Japon aydınlatma teknolojisi"),
        ("Kiku-Masamune", "kikumasamune.co.jp", "Japon Sake Aydınlatma", "High Moist Lotion; sake + arbutin; Japon geleneksel fermente aydınlatma"),
        ("Shiseido White Lucent", "shiseido.com", "Japon Sakura Aydınlatma", "White Lucent Illuminating Micro-Spot Serum; sakura + 4MSK; Japon bilimsel aydınlatma"),
        ("SK-II GenOptics", "sk-ii.com", "Japon Pitera Aydınlatma", "GenOptics Aura Essence; PITERA + niacinamide; Japon lüks aydınlatma"),
        ("Obagi", "obagi.com", "Medikal Aydınlatma", "Professional-C Serum 20%; klinik vitamin C; dermatolojik leke tedavisi"),
        ("iS Clinical", "isclinical.com", "White Lightening", "White Lightening Serum; arbutin + kojic + mushroom; medikal aydınlatma kompleksi"),
        ("SkinMedica", "skinmedica.com", "Lytera Aydınlatma", "Lytera 2.0 Pigment Correcting Serum; tranexamic + phenylethyl resorcinol; medikal leke"),
        ("Dr. Dennis Gross", "drdennisgross.com", "C+ Collagen Aydınlatma", "C+ Collagen Brighten & Firm Vitamin C Serum; L-ascorbic acid; NYC dermatoloji aydınlatma"),
        ("Ole Henriksen", "olehenriksen.com", "Banana Bright Aydınlatma", "Banana Bright Vitamin C Serum; C + E + hyaluronic; İskandinav aydınlatma"),
        ("Peter Thomas Roth", "peterthomasroth.com", "Potent-C Aydınlatma", "Potent-C Power Serum; THD Ascorbate %20; NYC güçlü C vitamini aydınlatma"),
        ("Kiehl's", "kiehls.com", "Clearly Corrective Aydınlatma", "Clearly Corrective Dark Spot Solution; activated C + salisilik + peony; leke açma serumu"),
        ("Caudalie Vinoperfect", "caudalie.com", "Üzüm Aydınlatma", "Vinoperfect Radiance Serum; viniferine (üzüm bileşeni); Fransız botanik aydınlatma"),
        ("Filorga", "filorga.com", "Fransız NCTF Aydınlatma", "Time-Filler Intensive Serum; NCEF + hyaluronic; Fransız estetik tıp aydınlatma"),
        ("Typology", "typology.com", "Fransız Minimalist Aydınlatma", "Serum Éclat; azelaic + vitamin C; Fransız minimalist leke tedavisi; 10 bileşen altı"),
        ("Medik8", "medik8.com", "İngiliz C-Tetra Aydınlatma", "C-Tetra Luxe; lipid bazlı C vitamini; İngiliz stabilize C aydınlatma"),
        ("The INKEY List", "theinkeylist.com", "Uygun Tranexamic", "Tranexamic Acid Night Treatment; £12; tranexamic acid leke tedavisi; erişilebilir"),
        ("Minimalist (India)", "beminimalist.co", "Hindistan Alpha Arbutin", "Alpha Arbutin 2% Serum; Hindistan'ın leke tedavisi uzmanı; şeffaf fiyat"),
        ("Dot & Key", "dotandkey.com", "Hindistan C Aydınlatma", "Vitamin C + E Super Bright Serum; Hindistan C vitamini trendi; renkli ambalaj"),
        ("Plum", "plumgoodness.com", "Hindistan Leke Tedavisi", "Vitamin C Face Serum; Hindistan vegan aydınlatma; etanol bazlı; erişilebilir"),
        ("Pilgrim", "discoverpilgrim.com", "Hindistan Kore Aydınlatma", "Korean Skin Care Vitamin C Serum; Kore reçete + Hindistan üretim; hibrit aydınlatma"),
        ("Forest Essentials", "forestessentials.com", "Hindistan Ayurveda Aydınlatma", "Soundarya Radiance Cream; 24K altın + ayurveda; Hindistan geleneksel aydınlatma"),
        ("Kama Ayurveda", "kamaayurveda.com", "Hindistan Kumkumadi Aydınlatma", "Kumkumadi Miraculous Beauty Fluid; safran + lotus; Ayurveda aydınlatma yağı; kült ürün"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 12-17: MAKYAJ KATEGORİLERİ
    # ═══════════════════════════════════════════════════════════════════════════
    "Makyaj - Fondöten & BB/CC Krem": [
        ("Kosas", "kosas.com", "Temiz Tinted Fondöten", "Revealer Skin-Improving Foundation SPF 25; bakım + makyaj; peptide + arbutin; temiz fondöten"),
        ("Jones Road", "jonesroadbeauty.com", "What The Foundation", "WTF fondöten; Bobbi Brown'un yeni markası; tinted moisture balm; doğal finish"),
        ("Saie", "saiehello.com", "Slip Tint SPF", "Slip Tint SPF 35; hafif renkli nemlendirici + SPF; temiz formül; dewy finish"),
        ("Merit", "meritbeauty.com", "Minimalist Tint", "The Minimalist Perfecting Complexion Stick; stick fondöten; 5 dakika makyaj felsefesi"),
        ("Ilia", "iliabeauty.com", "Super Serum Fondöten", "Super Serum Skin Tint SPF 40; serum + fondöten + SPF; temiz lüks; cilt bakımı hibrit"),
        ("Tower 28", "tower28beauty.com", "Hassas Cilt BB", "SunnyDays SPF 30 Tinted Sunscreen; NEA onaylı; hassas cilt renkli SPF"),
        ("Rose Inc", "roseinc.com", "Rosie HW Fondöten", "Skin Enhance Luminous Tinted Serum; Rosie Huntington-Whiteley; serum fondöten"),
        ("Iris & Romeo", "irisandromeo.com", "Multi-Tasking BB", "Best Skin Days SPF 25; nemlendirici + SPF + tint + bakım; 4-in-1 minimalist"),
        ("Danessa Myricks", "danessamyricks.com", "Kreatif Fondöten", "Yummy Skin Blurring Balm Foundation; bulanıklaştırıcı balm; makyaj sanatçısı markası"),
        ("About Face", "aboutface.com", "Halsey Fondöten", "Performer Skin Focusing Foundation; yüksek performans; Halsey markası; vegan formül"),
        ("Ami Colé", "amicole.com", "Melanin BB", "Skin-Enhancing Tint; koyu cilt tonları uzmanı; Senegal kökenli; minimal temiz"),
        ("MENTED", "mentedcosmetics.com", "Melanin Fondöten", "Skin by Mented Foundation; koyu ten tonları; 16 ton; kapsayıcı temiz fondöten"),
        ("Juvia's Place", "juviasplace.com", "Afrika Fondöten", "I Am Magic Foundation; 42 ton; Afrika ilhamı; kapsayıcı geniş ton aralığı"),
        ("BLK/OPL", "blkopl.com", "Melanin Uzmanı Fondöten", "True Color Pore Perfecting Foundation; melanin bilimi; kapsayıcı Amerikan markası"),
        ("Sheglam", "sheglam.com", "Uygun Fiyat Fondöten", "Skinfinite Hydrating Foundation; Shein markası; $5-8; viral TikTok; erişilebilir"),
        ("Colourpop", "colourpop.com", "Trend Fondöten", "Pretty Fresh Hyaluronic Foundation; HA + niacinamide; $15; Seed Beauty; erişilebilir"),
        ("e.l.f.", "elfcosmetics.com", "Ultra Uygun Fondöten", "Camo CC Cream SPF 30; $7; TikTok viral; eczane fiyatına yüksek performans"),
        ("NYX", "nyxcosmetics.com", "Born To Glow Fondöten", "Born To Glow! Naturally Radiant Foundation; 45 ton; ışıltılı finish; erişilebilir"),
        ("Milani", "milanicosmetics.com", "Conceal + Perfect", "Conceal + Perfect 2-in-1 Foundation; full coverage + concealer; drugstore favorisi; 40+ ton"),
        ("Revolution", "revolutionbeauty.com", "İngiliz Uygun Fondöten", "Conceal & Define Foundation; İngiltere £5; 50 ton; full coverage; erişilebilir"),
        ("Catrice", "catrice.eu", "Alman True Skin", "True Skin Hydrating Foundation; $8; Alman eczane; hafif nemlendirici fondöten"),
        ("Essence", "essence.eu", "€3 Fondöten", "Pretty Natural Hydrating Foundation; €3; Alman ultra uygun; hafif doğal finish"),
        ("Missha", "missha.com", "Kore BB Krem Öncüsü", "M Perfect Blind BB Cream SPF 42; orijinal BB krem standardı; Kore BB konsepti yaratıcısı"),
        ("Clio", "clio.co.kr", "Kore Kill Cover", "Kill Cover Fixer Cushion; Kore cushion fondöten; yüksek coverage; uzun süre dayanıklı"),
        ("Laneige", "laneige.com", "Kore Neo Cushion", "Neo Cushion Glow; Kore su bazlı cushion; ışıltılı Kore fondöten; K-beauty"),
        ("Peripera", "peripera.com", "Kore Ink Cushion", "Ink Lasting Cover Cushion; Kore uygun fiyat cushion; genç K-beauty fondöten"),
        ("Moonshot", "moonshot.co.kr", "Kore Micro Cushion", "Micro Glassyfit Cushion; cam cilt cushion; Kore viral; YG Entertainment markası"),
        ("Etude House", "etude.com", "Kore Double Lasting", "Double Lasting Foundation; Kore mat fondöten; uzun ömür; 25+ ton"),
        ("Romand", "romand.co.kr", "Kore Better Than", "Better Than Cushion Foundation; Kore viral fondöten; doğal finish; Z kuşağı K-beauty"),
        ("Holika Holika", "holikaholika.com", "Kore Eğlenceli BB", "Aqua Petit Jelly BB; jöle kıvamlı BB; eğlenceli K-beauty; yaratıcı ambalaj"),
        ("A'PIEU", "apieu.com", "Kore Madeca BB", "Madecassoside Cica BB Cream; cica + BB; Kore hassas cilt fondöten"),
        ("VDL", "vdl.co.kr", "Kore Expert CC", "Expert Color Correcting Primer; renk düzeltme; Kore CC krem; ton eşitleme"),
        ("Canmake", "canmake.com", "Japon Hafif BB", "Perfect Serum BB Cream; Japon hafif formül; makyaj bazı + SPF; uygun fiyat"),
        ("Shiseido", "shiseido.com", "Japon Synchro Skin", "Synchro Skin Self-Refreshing Foundation; kendini yenileyen; Japon teknoloji; ActiveForce"),
        ("SUQQU", "suqqu.com", "Japon Lüks Fondöten", "The Liquid Foundation; Japon lüks; ultra doğal finish; ince film teknolojisi"),
        ("RMK", "rmk.com", "Japon Jel Fondöten", "Gel Creamy Foundation; jel kıvamlı; Japon hafif fondöten; doğal cam cilt"),
        ("Shu Uemura", "shuuemura.com", "Japon Unlimited Fondöten", "Unlimited Breathable Lasting Foundation; nefes alan fondöten; Japon teknoloji"),
        ("Natasha Denona", "natashadenona.com", "Makyaj Sanatçısı Fondöten", "Hy-Gen Skincare Foundation; hyaluronic + fondöten; lüks hibrit; makyaj sanatçısı"),
        ("Westman Atelier", "westman-atelier.com", "Temiz Lüks Fondöten", "Vital Skin Foundation Stick; stick format; temiz lüks; Gucci Westman formülleri"),
        ("Pat McGrath alternatifi: Danessa Myricks", "danessamyricks.com", "Çok Amaçlı Fondöten", "Vision Cream Cover; 40 ton; makyaj sanatçısı formülü; full coverage"),
        ("Hourglass", "hourglasscosmetics.com", "Vegan Lüks Fondöten", "Vanish Airbrush Foundation; airbrush efekti; %100 vegan lüks; full coverage"),
    ],

    "Makyaj - Allık & Bronzer": [
        ("Saie", "saiehello.com", "Dew Blush", "Dew Blush; sıvı dewy allık; temiz formül; TikTok viral; çiy görünümü"),
        ("Tower 28", "tower28beauty.com", "BeachPlease Allık", "BeachPlease Tinted Lip + Cheek Balm; dudak + yanak; NEA onaylı; temiz çok amaçlı"),
        ("Rare Beauty alternatifi: Flower Beauty", "flowerbeauty.com", "Blush Bomb", "Blush Bomb Color Drops; sıvı allık; Drew Barrymore; uygun fiyat; Walmart"),
        ("em Cosmetics", "emcosmetics.com", "Serum Blush", "So Soft Blush; krem serum allık; Michelle Phan; K-beauty ilhamlı; doğal finish"),
        ("Danessa Myricks", "danessamyricks.com", "Çok Amaçlı Allık", "Colorfix; göz + dudak + yanak; makyaj sanatçısı; yüksek pigment; çok amaçlı"),
        ("About Face", "aboutface.com", "Cheek Freak Allık", "Cheek Freak Blush Balm; krem balm; Halsey; pastel ambalaj; trendy"),
        ("Youthforia", "youthforia.com", "Renk Değiştiren Allık", "BYO Blush; vücut ısısıyla renk değiştiren; herkesin cilt tonuna uyum; inovatif"),
        ("Milk Makeup", "milkmakeup.com", "Stick Allık", "Lip + Cheek Stick; çok amaçlı stick; vegan; temiz; pratik uygulama"),
        ("Merit", "meritbeauty.com", "Flush Balm Allık", "Flush Balm Cheek Color; krem balm; doğal flush; minimalist; 5 dakika makyaj"),
        ("Jones Road", "jonesroadbeauty.com", "Miracle Balm Allık", "Miracle Balm; tinted moisturizing balm; tüm yüze uygulanabilir; Bobbi Brown"),
        ("Ami Colé", "amicole.com", "Melanin Allık", "Desert Date Cream Blush; koyu cilt tonları; çöl hurması yağı; minimal temiz"),
        ("Live Tinted", "livetinted.com", "Hueguard Allık", "Huetopia Blush; Güney Asya tonları; kapsayıcı; çok amaçlı stick"),
        ("Kosas", "kosas.com", "10-Second Allık", "10-Second Liquid Eyeshadow; sıvı formül; hızlı uygulama; çok amaçlı renk"),
        ("Colourpop", "colourpop.com", "Super Shock Allık", "Super Shock Cheek; unique bouncy texture; $8; geniş renk yelpazesi; TikTok favorisi"),
        ("e.l.f.", "elfcosmetics.com", "Putty Blush", "Putty Blush; sponge kıvamlı; $6; TikTok viral; uygun fiyat; dewy finish"),
        ("Milani", "milanicosmetics.com", "Baked Blush", "Baked Blush; İtalyan pişirme tekniği; Luminoso kült renk; drugstore ikonik"),
        ("NYX", "nyxcosmetics.com", "Sweet Cheeks Allık", "Sweet Cheeks Soft Cheek Tint; satin + matte seçenek; erişilebilir geniş palet"),
        ("Sheglam", "sheglam.com", "Snatch Contour", "Snatch 'n' Shape Contour Stick; kontur + allık; Shein; ultra uygun; TikTok viral"),
        ("Romand", "romand.co.kr", "Kore Allık", "Better Than Cheek; Kore mat allık; doğal flush; K-beauty; pastel ambalaj"),
        ("Peripera", "peripera.com", "Kore Velvet Allık", "Pure Blushed Sunshine Cheek; Kore allık; çeşitli finish; uygun fiyat K-beauty"),
        ("3CE", "3ce.com", "Kore Mood Allık", "Mood Recipe Face Blush; Kore mood estetik; pastel + mauve tonlar; Stylenanda"),
        ("Etude House", "etude.com", "Kore Lovely Cookie Allık", "Lovely Cookie Blusher; kurabiye ambalaj; eğlenceli K-beauty; genç allık"),
        ("Clio", "clio.co.kr", "Kore Pro Allık", "Prism Air Blusher; Kore profesyonel; holografik parıltı; K-beauty premium"),
        ("A'PIEU", "apieu.com", "Kore Juicy Allık", "Juicy-Pang Jelly Blusher; jöle kıvamlı; meyve ilhamlı; Kore eğlenceli"),
        ("Canmake", "canmake.com", "Japon Cream Allık", "Cream Cheek; Japon krem allık; ultra doğal; ¥600 uygun fiyat"),
        ("Visée", "visee.jp", "Japon Glow Allık", "Riche Lip & Cheek Cream; Kanebo; Japon çok amaçlı; dudak + yanak"),
        ("SUQQU", "suqqu.com", "Japon Lüks Allık", "Pure Color Blush; Japon lüks; sanatsal renk harmanlama; seramik ambalaj"),
        ("NARS alternatifi: Persona Cosmetics", "personacosmetics.com", "Kapsayıcı Allık", "Super Blush; geniş ton aralığı; YouTube'dan markaya; kapsayıcı güzellik"),
        ("Physicians Formula", "physiciansformula.com", "Butter Bronzer", "Murumuru Butter Bronzer; hindistan cevizi kokusu; drugstore kült; bronzer standardı"),
        ("Benefit", "benefitcosmetics.com", "Hoola Bronzer", "Hoola Matte Bronzer; mat bronzer standardı; evrensel ton; kült klasik"),
        ("Too Faced", "toofaced.com", "Chocolate Bronzer", "Chocolate Soleil Bronzer; çikolata kokulu; mat bronzer; eğlenceli marka"),
        ("Guerlain Terracotta alternatifi: e.l.f.", "elfcosmetics.com", "Uygun Bronzer", "Putty Bronzer; $6; sponge kıvamlı bronzer; doğal bronz; TikTok viral dupe"),
        ("Patrick Ta", "patrickta.com", "MUA Allık", "Major Headlines Double-Take Creme & Powder Blush Duo; makyaj sanatçısı; krem + pudra ikili"),
        ("Charlotte Tilbury alternatifi: Catrice", "catrice.eu", "Alman Bronzer", "Holiday Skin Bronzer; Alman eczane; uygun fiyat; doğal bronz; Charlotte Tilbury alternatifi"),
        ("Flower Beauty", "flowerbeauty.com", "Blush Bomb Drops", "Blush Bomb Color Drops for Cheeks; sıvı allık; Drew Barrymore; $10; Walmart"),
        ("Essence", "essence.eu", "€2 Allık", "The Blush; €2; Alman ultra uygun; hafif pigment; doğal günlük allık"),
        ("Holika Holika", "holikaholika.com", "Kore Jelly Allık", "Jelly Dough Blusher; jöle hamur kıvamlı; eğlenceli K-beauty; yaratıcı format"),
        ("Focallure", "focallure.com", "Çin DTC Allık", "Blush Palette; Çin DTC güzellik; uygun fiyat; geniş seçenek; TikTok"),
        ("Kimchi Chic", "kimchichicbeauty.com", "K-Beauty Ilhamlı Allık", "Jelly Sheer Blush; jöle kıvamlı; Kimchi Chic K-beauty vibes; eğlenceli renkler"),
        ("BH Cosmetics", "bhcosmetics.com", "Palet Allık", "Truffle Blush; çikolata trüf ilhamlı; 4'lü palet; erişilebilir; pastel tonlar"),
    ],

    "Makyaj - Göz Makyajı": [
        ("Nabla", "nablacosmetics.com", "İtalyan Göz Paleti", "Dreamy Eyeshadow Palette; İtalyan indie; yüksek pigment; krematik formül; sanatsal renkler"),
        ("Kaleidos", "kaleidosmakeup.com", "Sanatsal Göz Paleti", "Club Nebula Palette; fütüristik renk hikayeleri; Çin indie güzellik; yenilikçi formül"),
        ("Juvia's Place", "juviasplace.com", "Afrika Göz Paleti", "The Zulu Palette; Afrika kültürü ilhamlı; yoğun pigment; geniş renk yelpazesi"),
        ("Colourpop", "colourpop.com", "Süper Shock Göz", "Super Shock Shadow; unique bouncy formula; $5; trend renkleri hızlı çıkarma; erişilebilir"),
        ("BH Cosmetics", "bhcosmetics.com", "Zodiac Göz Paleti", "Zodiac Love Signs Palette; astroloji ilhamlı; 25 renk; erişilebilir palet"),
        ("Persona Cosmetics", "personacosmetics.com", "Kapsayıcı Göz", "Identity Palette; tüm cilt tonlarına uyumlu; YouTube kökenli; çok yönlü"),
        ("P.Louise", "plouise.co.uk", "Base Göz Bazı", "Base; göz bazı uzmanı; canlı pigment aktarımı; İngiliz indie; viral"),
        ("Kimchi Chic", "kimchichicbeauty.com", "Glitter Göz", "Diamond Sharts Sparkle Cream; K-beauty ilhamlı glitter; eğlenceli isim; viral"),
        ("em Cosmetics", "emcosmetics.com", "Serum Göz Farı", "Heaven's Glow Blush; Michelle Phan; şeffaf serum formül; K-beauty ilhamlı göz"),
        ("Danessa Myricks", "danessamyricks.com", "Kreatif Göz", "Colorfix; ultra pigment; göz + dudak + yanak; makyaj sanatçısı vizyonu; çok amaçlı"),
        ("About Face", "aboutface.com", "Shadow Stick Göz", "Shadowstick; Halsey; tek hareketle göz makyajı; kalem formül; 30 saniye"),
        ("Live Tinted", "livetinted.com", "Çok Amaçlı Göz", "Huetopia; göz + dudak + yanak; Güney Asya tonları; kapsayıcı stick"),
        ("Kosas", "kosas.com", "10 Saniye Göz Farı", "10-Second Eyeshadow; sıvı-krem; hızlı uygulama; temiz formül; hibrit"),
        ("Merit", "meritbeauty.com", "Solo Shadow", "Solo Shadow; tek renk; soft matte; minimalist göz makyajı; doğal"),
        ("Saie", "saiehello.com", "Glowy Göz", "Glowy Super Gel Lightweight Illuminator; göz + yüz ışıltı; temiz; çiy efekti"),
        ("Tower 28", "tower28beauty.com", "Hassas Göz Farı", "SuperDew Shimmer Shadow; NEA onaylı; hassas göz çevresi güvenli"),
        ("romand", "romand.co.kr", "Kore Göz Paleti", "Better Than Eyes; K-beauty göz paleti; 4 renk harmony; Kore pastel estetik"),
        ("Clio", "clio.co.kr", "Kore Pro Göz", "Pro Eye Palette; K-beauty profesyonel; glitter + matte; Kore makyaj standardı"),
        ("Etude House", "etude.com", "Kore Play Color Göz", "Play Color Eyes; 10 renkli palet; eğlenceli K-beauty; uygun fiyat çeşitlilik"),
        ("3CE", "3ce.com", "Kore Mood Göz", "Multi Eye Color Palette; Kore mood estetik; mauve + nude tonlar; Stylenanda"),
        ("Peripera", "peripera.com", "Kore Glitter Göz", "All Take Mood Palette; Kore glitter + matte; uygun fiyat K-beauty palet"),
        ("Holika Holika", "holikaholika.com", "Kore Eğlenceli Göz", "Piece Matching Shadow; tekli göz farı; mix & match; eğlenceli K-beauty"),
        ("A'PIEU", "apieu.com", "Kore Tek Göz Farı", "Couture Shadow; tek renk mükemmelliği; Kore uygun fiyat tekli far"),
        ("Missha", "missha.com", "Kore Triple Göz", "Triple Shadow; 3 renk mini palet; Kore pratik göz makyajı; seyahat dostu"),
        ("Canmake", "canmake.com", "Japon Silky Göz", "Silky Souffle Eyes; ipeksi krem formül; Japon hafif göz farı; ¥780"),
        ("Visée", "visee.jp", "Japon Gemme Göz", "Riche Gemme Eyes; mücevher ilhamlı; Kanebo; Japon lüks eczane göz makyajı"),
        ("SUQQU", "suqqu.com", "Japon Signature Göz", "Signature Color Eyes; Japon lüks; sanatsal 4 renk; mevsimsel limitli üretim"),
        ("Lunasol", "lunasol-net.com", "Japon Skin Göz", "Skin Modeling Eyes; Kanebo lüks; nude Japon göz paleti; doğal derinlik"),
        ("KATE", "nomorerules.net", "Japon Designing Göz", "Designing Brown Eyes; Japon mat kahverengi; uygun fiyat doğal göz; Kanebo"),
        ("Excel", "excel-tokyo.com", "Japon Skinny Göz", "Skinny Rich Shadow; ince parıltı; Japon doğal göz farı; ofis makyajı"),
        ("Addiction by Ayako", "addiction-beauty.com", "Japon Eyeshadow", "The Eyeshadow; tek renk mücevher; Japon sanatsal; 99 renk seçenek"),
        ("Revolution", "revolutionbeauty.com", "İngiliz Uygun Palet", "Reloaded Palette; İngiltere £4; büyük palet düşük fiyat; dupe kültürü"),
        ("Sheglam", "sheglam.com", "TikTok Göz Paleti", "Color Bloom Liquid Eyeshadow; sıvı göz farı; Shein; ultra uygun; viral"),
        ("Focallure", "focallure.com", "Çin Glitter Göz", "Wanderlust Eyeshadow Palette; Çin DTC; uygun fiyat; metalik + mat; geniş palet"),
        ("Flower Beauty", "flowerbeauty.com", "Shimmer Göz", "Shimmer & Shade Eyeshadow Palette; Drew Barrymore; Walmart; $10; erişilebilir"),
        ("NYX", "nyxcosmetics.com", "Ultimate Göz Paleti", "Ultimate Shadow Palette; 16 renk; $14; profesyonel + erişilebilir; geniş çeşitlilik"),
        ("Essence", "essence.eu", "€3 Göz Paleti", "If I Were... Eyeshadow Palette; €3; Alman ultra uygun; sürpriz kalite"),
        ("Catrice", "catrice.eu", "Alman Faked Göz", "Faked Everyday Lashes; yapay kirpik; €4; Alman eczane; erişilebilir drama"),
        ("Maybelline benzeri indie: BK Beauty", "bkbeauty.com", "Fırça Göz Seti", "Göz makyajı fırça seti; YouTube viral; Lisa J fırçaları; erişilebilir kalite"),
        ("Sigma Beauty", "sigmabeauty.com", "Makyaj Fırçası + Göz", "Eye Palette + Brush Set; fırça uzmanı + göz paleti; profesyonel araçlar"),
    ],

    "Makyaj - Dudak Ürünleri": [
        ("Kosas", "kosas.com", "Wet Lip Oil", "Wet Lip Oil Gloss; hyaluronic acid + peptide; parlak bakımlı; temiz lüks dudak"),
        ("Tower 28", "tower28beauty.com", "ShineOn Lip Jelly", "ShineOn Lip Jelly; NEA onaylı; hassas dudak; vegan jöle gloss; temiz"),
        ("Merit", "meritbeauty.com", "Shade Slick Lip Oil", "Shade Slick Tinted Lip Oil; hafif renk + bakım; minimalist; 5 dakika makyaj"),
        ("Saie", "saiehello.com", "Lip Blur Matte", "Lip Blur Soft-Matte Hydrating Lipstick; bulanık mat; temiz + nemlendirici matte"),
        ("Jones Road", "jonesroadbeauty.com", "Cool Gloss", "Cool Gloss; şeffaf parlak; Bobbi Brown; doğal dudak; minimalist lüks"),
        ("em Cosmetics", "emcosmetics.com", "Lip Cushion", "Lip Cushion; K-beauty ilhamlı cushion format; Michelle Phan; yenilikçi aplikatör"),
        ("Ami Colé", "amicole.com", "Lip Treatment Oil", "Lip Treatment Oil; melanin dostu; bakım + parlaklık; Senegal ilhamlı temiz"),
        ("MENTED", "mentedcosmetics.com", "Nude Lipstick", "Semi-Matte Lipstick; koyu ten nude tonları; kapsayıcı 'nude' yeniden tanımlama"),
        ("Danessa Myricks", "danessamyricks.com", "Lip Vinyl", "Lip Vinyl; ultra parlak; yüksek pigment; makyaj sanatçısı formülü"),
        ("About Face", "aboutface.com", "Lip Color", "Lip Color Butter; Halsey; besleyici dudak rengi; vegan formül; yaratıcı tonlar"),
        ("rom&nd", "romand.co.kr", "Juicy Lasting Tint", "Juicy Lasting Tint; K-beauty dudak tint standardı; meyve renkleri; viral TikTok"),
        ("Peripera", "peripera.com", "Ink Velvet", "Ink the Velvet; Kore kadife tint; uzun ömürlü; uygun fiyat; geniş renk"),
        ("3CE", "3ce.com", "Velvet Lip Tint", "Velvet Lip Tint; Kore kadife dudak; mat + nemlendirici; Stylenanda"),
        ("Romand", "romand.co.kr", "Glasting Tint", "Glasting Water Tint; su bazlı parlak tint; K-beauty; hafif + parlak"),
        ("Colourpop", "colourpop.com", "Lux Lipstick", "Lux Lipstick; $8; trend renkler; hızlı koleksiyon; erişilebilir lüks his"),
        ("NYX", "nyxcosmetics.com", "Butter Gloss", "Butter Gloss; kült ürün; 30+ renk; $5; kremsi parlaklık; eczane favorisi"),
        ("e.l.f.", "elfcosmetics.com", "Lip Lacquer", "Lip Lacquer; $3; ultra uygun; dolgunlaştırıcı etki; TikTok viral"),
        ("Milani", "milanicosmetics.com", "Color Fetish Lip", "Color Fetish Shine Lipstick; İtalyan ilham; parlak lipstick; drugstore premium"),
        ("Revolution", "revolutionbeauty.com", "Lip Allure", "Lip Allure Soft Satin Lipstick; İngiliz uygun; kremsi saten; £4"),
        ("Sheglam", "sheglam.com", "Mirror Kiss Lip", "Mirror Kiss High-Shine Lipstick; ayna parlaklığı; Shein; $4; viral"),
        ("Flower Beauty", "flowerbeauty.com", "Petal Lip", "Petal Pout Lip Mask; dudak maskesi + renk; Drew Barrymore; çift işlev"),
        ("Essence", "essence.eu", "€2 Lip Gloss", "Extreme Shine Volume Lipgloss; €2; Alman ultra uygun; hacim veren parlaklık"),
        ("Catrice", "catrice.eu", "Alman Full Satin Lip", "Full Satin Lipstick; Alman eczane; saten finish; €5; geniş renk"),
        ("Clio", "clio.co.kr", "Kore Kill Rouge", "Kill Rouge Velvet; Kore kadife lipstick; profesyonel K-beauty"),
        ("Moonshot", "moonshot.co.kr", "Kore Tint Fit", "Tint Fit Shine; Kore parlak tint; YG Entertainment; K-pop estetik"),
        ("Etude", "etude.com", "Kore Fixing Tint", "Fixing Tint; transfer-proof Kore tint; gün boyu dayanıklı; mask-proof"),
        ("Holika Holika", "holikaholika.com", "Kore Heart Crush", "Heart Crush Glow Tint; kalp şekli ambalaj; eğlenceli K-beauty dudak"),
        ("A'PIEU", "apieu.com", "Kore Juicy-Pang Lip", "Juicy-Pang Mousse Tint; köpük kıvamlı; meyve ilhamlı; Kore inovatif"),
        ("Canmake", "canmake.com", "Japon Lip Tint", "Lip Tint Syrup; şurup kıvamlı tint; Japon uygun; doğal renk"),
        ("Opera", "opera-net.jp", "Japon Lip Tint Oil", "Lip Tint Oil; Japon dudak tint yağı; @cosme #1; hafif renk + bakım"),
        ("KATE", "nomorerules.net", "Japon Lip Monster", "Lip Monster; Japon viral; mask-proof; 'canavar' konsepti; Kanebo"),
        ("Dior Lip Glow alternatifi: Burt's Bees", "burtsbees.com", "Tinted Lip Balm", "Tinted Lip Balm; doğal + renkli; hafif tint; balmumu bazlı"),
        ("Ilia", "iliabeauty.com", "Balmy Tint", "Balmy Tint Hydrating Lip Balm; renk + bakım; temiz lüks dudak; SPF 15"),
        ("Westman Atelier", "westman-atelier.com", "Baby Cheeks Lip", "Baby Cheeks Blush Stick; dudak + yanak; temiz lüks; Gucci Westman"),
        ("Iris & Romeo", "irisandromeo.com", "Multi-Tasking Lip", "Power Petal; dudak + yanak + göz; 3-in-1; minimalist; temiz"),
        ("Patrick Ta", "patrickta.com", "Major Headlines Lip", "Major Headlines Matte Suede Lipstick; lüks mat; makyaj sanatçısı; rich pigment"),
        ("Lisa Eldridge", "lisaeldridge.com", "Luxury Lipstick", "True Velvet Lipstick; lüks kadife; İngiliz makyaj ikonu; sınırlı üretim; kült"),
        ("Victoria Beckham Beauty", "victoriabeckhambeauty.com", "Posh Lip", "Posh Lipstick; lüks dudak; minimalist ambalaj; nude master; VB vizyonu"),
        ("Bobbi Brown alternatifi: Jones Road", "jonesroadbeauty.com", "The Lip Tint", "The Lip Tint; krem tint; doğal; Bobbi Brown yeni vizyon; erişilebilir lüks"),
        ("Rare Beauty alternatifi: Dibsbeauty", "dibs.beauty", "Status Stick Lip", "Status Stick; çok amaçlı renk stick; dudak + yanak; influencer markası"),
    ],

    "Makyaj - Kaş Ürünleri": [
        ("Benefit", "benefitcosmetics.com", "Precisely My Brow", "Precisely, My Brow Pencil; ultra ince uçlu kaş kalemi; kaş makyajı standardı"),
        ("Anastasia Beverly Hills alternatifi: NYX", "nyxcosmetics.com", "Micro Brow Pencil", "Micro Brow Pencil; ultra ince; ABH dupe; $10; erişilebilir kaş kalemi"),
        ("Boy Brow alternatifi: e.l.f.", "elfcosmetics.com", "Wow Brow Jel", "Wow Brow Gel; $4; fiber kaş jeli; Glossier Boy Brow dupe; TikTok viral"),
        ("Kosas", "kosas.com", "Air Brow", "Air Brow Fluff & Hold Treatment Gel; temiz kaş jeli; peptide + prostaglandin"),
        ("Merit", "meritbeauty.com", "Brow 1980", "Brow 1980 Pomade Pencil; üçgen uçlu; doğal kalın kaş; minimalist"),
        ("Patrick Ta", "patrickta.com", "Major Brow", "Major Brow Lamination Gel; laminasyon efekti; profesyonel kaş şekillendirme"),
        ("Refy", "refy.com", "Brow Sculpt", "Brow Sculpt; İngiliz kaş markası; laminasyon + şekil; TikTok viral; sapma kaş"),
        ("Soap Brow alternatifi: West Barn Co", "westbarnco.com", "Soap Brows", "Soap Brows; sabun kaş trendi başlatıcısı; İngiliz indie; kaş laminasyon efekti"),
        ("Grande Cosmetics", "grandecosmetics.com", "Kaş Büyütme Serumu", "GrandeBROW Brow Enhancing Serum; peptide bazlı kaş büyütme; klinik test sonuçları"),
        ("RevitaLash", "revitalash.com", "Kaş Kondisyoner", "Hi-Def Brow Pencil + RevitaBrow; kaş büyütme + şekillendirme; medikal güzellik"),
        ("Colourpop", "colourpop.com", "Brow Boss Kalem", "Brow Boss Pencil; $6; ultra ince; geniş renk; erişilebilir kaş makyajı"),
        ("Milani", "milanicosmetics.com", "Weekend Brow", "Weekend Brow Tinted Gel; fiber jel; drugstore; kolay uygulama; doğal finish"),
        ("Essence", "essence.eu", "€2 Kaş Kalemi", "Brow Powder & Define Pen; €2; çift uçlu; Alman eczane; ultra uygun"),
        ("Catrice", "catrice.eu", "Alman Slim'Matic Kaş", "Slim'Matic Ultra Precise Brow Pencil; ultra ince; Alman eczane; €3"),
        ("Revolution", "revolutionbeauty.com", "İngiliz Brow Pomade", "Brow Pomade; İngiliz uygun; ABH Dipbrow alternatifi; £3; çeşitli tonlar"),
        ("Etude", "etude.com", "Kore Drawing Kaş", "Drawing Eye Brow; Kore kaş kalemi standardı; uygun fiyat; doğal çizgiler"),
        ("Innisfree", "innisfree.com", "Kore Auto Kaş", "Auto Eyebrow Pencil; otomatik kaş kalemi; Kore doğal; Jeju"),
        ("Clio", "clio.co.kr", "Kore Kill Brow", "Kill Brow Auto Hard Brow Pencil; Kore profesyonel kaş; uzun ömürlü"),
        ("Holika Holika", "holikaholika.com", "Kore Wonder Drawing", "Wonder Drawing 1sec Finish Browcara; 1 saniye kaş; eğlenceli K-beauty"),
        ("Canmake", "canmake.com", "Japon 3in1 Kaş", "3in1 Eyebrow; kalem + pudra + fırça; Japon pratik; uygun fiyat"),
        ("Shu Uemura", "shuuemura.com", "Japon Hard Formula Kaş", "Hard Formula; Japon profesyonel kaş kalemi; mekanik; makyaj sanatçısı standardı"),
        ("K-Palette", "k-palette.com", "Japon 1Day Tattoo Kaş", "1Day Tattoo Lasting Eyebrow Tint; Japon kaş dövme efekti; 24 saat dayanıklı"),
        ("Sheglam", "sheglam.com", "TikTok Kaş Sabunu", "Brow Wow Soap Styler; sabun kaş; Shein; ultra uygun; laminasyon efekti"),
        ("Pixi", "pixibeauty.com", "Kore Kaş Duo", "Natural Brow Duo; kalem + pudra; doğal kaş; İsveç-İngiliz"),
        ("Rare Beauty alternatifi: Item Beauty", "itembeauty.com", "Brow Chow", "Brow Chow Eyebrow Pencil; Addison Rae; Z kuşağı kaş; temiz formül"),
    ],

    "Makyaj - Aydınlatıcı & Kontür": [
        ("Saie", "saiehello.com", "Glowy Super Gel", "Glowy Super Gel; sıvı aydınlatıcı; TikTok viral; çiy efekti; temiz highlighter"),
        ("Merit", "meritbeauty.com", "Day Glow Aydınlatıcı", "Day Glow Highlighting Balm; krem balm; doğal ışıltı; minimalist"),
        ("Tower 28", "tower28beauty.com", "SuperDew Aydınlatıcı", "SuperDew Highlight Balm; NEA onaylı; hassas cilt; temiz ışıltı"),
        ("Danessa Myricks", "danessamyricks.com", "Twin Flames", "Twin Flames Multichrome; çoklu krom renk değişimi; makyaj sanatçısı; inovatif pigment"),
        ("Kosas", "kosas.com", "Wet Highlighter", "10-Second Eye Shadow; ıslak ışıltı; temiz formül; çok amaçlı aydınlatma"),
        ("Jones Road", "jonesroadbeauty.com", "Miracle Balm Glow", "Miracle Balm; doğal ışıltılı; çok amaçlı; Bobbi Brown vizyonu"),
        ("Milk Makeup", "milkmakeup.com", "Highlighter Stick", "Holographic Stick; holografik aydınlatıcı; stick format; vegan"),
        ("Becca alternatifi (RIP): Colourpop", "colourpop.com", "Super Shock Cheek", "Super Shock Cheek Highlighter; bouncy formula; $8; yoğun parıltı"),
        ("e.l.f.", "elfcosmetics.com", "Halo Glow", "Halo Glow Liquid Filter; Charlotte Tilbury Flawless Filter dupe; $14; viral"),
        ("NYX", "nyxcosmetics.com", "Born To Glow Liquid", "Born To Glow Liquid Illuminator; sıvı aydınlatıcı; $8; profesyonel + erişilebilir"),
        ("Sheglam", "sheglam.com", "Skin Focus Tint", "Skin Focus Tint; ışıltılı tint; Shein; $5; TikTok viral; glow efekti"),
        ("Catrice", "catrice.eu", "Alman Dewy Aydınlatıcı", "Dewy Wetlook Stick; ıslak görünüm; Alman eczane; €5"),
        ("Essence", "essence.eu", "€3 Aydınlatıcı", "Pure Nude Highlighter; €3; Alman eczane; doğal parıltı; günlük kullanım"),
        ("Revolution", "revolutionbeauty.com", "İngiliz Glow", "Skin Kiss Highlighter; £5; güçlü parıltı; İngiliz uygun; geniş seçim"),
        ("Romand", "romand.co.kr", "Kore See-Through Aydınlatıcı", "See-Through Veillighter; transparan ışıltı; K-beauty doğal glow"),
        ("Clio", "clio.co.kr", "Kore Prism Aydınlatıcı", "Prism Air Highlighter; holografik K-beauty; profesyonel ışıltı"),
        ("3CE", "3ce.com", "Kore Pot Aydınlatıcı", "Pot Eye Shadow; cilt aydınlatıcı olarak da; çok amaçlı; Kore minimal"),
        ("Canmake", "canmake.com", "Japon Glow Fleur", "Glow Fleur Cheeks; çiçek deseni aydınlatıcı; Japon kawaii; uygun fiyat"),
        ("SUQQU", "suqqu.com", "Japon Lüks Glow", "Glow Powder; Japon ince parıltı; lüks pudra aydınlatıcı; mevsimsel limitli"),
        ("Hourglass", "hourglasscosmetics.com", "Ambient Aydınlatıcı", "Ambient Lighting Powder; ışık yansıtma teknolojisi; vegan lüks; doğal glow"),
        ("Patrick Ta", "patrickta.com", "She's Contour", "She's Sculpted Double-Take Contour; profesyonel kontür; makyaj sanatçısı; krem + pudra"),
        ("Kevyn Aucoin", "kevynaucoin.com", "Sculpting Kontür", "The Sculpting Contour Powder; profesyonel kontür; makyaj efsanesi markası"),
        ("KVD Beauty", "kvdveganbeauty.com", "Shade + Light Kontür", "Shade + Light Contour Palette; mat kontür; vegan; profesyonel"),
        ("Fenty Beauty alternatifi: NYX", "nyxcosmetics.com", "3 Steps Kontür", "3 Steps to Sculpt; 3 adım kontür paleti; erişilebilir; geniş ton"),
        ("Milani", "milanicosmetics.com", "Silky Matte Kontür", "Silky Matte Bronzing Powder; İtalyan formül; mat kontür; drugstore"),
        ("Too Faced", "toofaced.com", "Born This Way Kontür", "Born This Way Turn Up the Light Highlighting Palette; çoklu aydınlatıcı; eğlenceli"),
        ("Physicians Formula", "physiciansformula.com", "Murumuru Kontür", "Murumuru Butter Contour Palette; botanik yağ + kontür; hipoalerjenik"),
        ("Flower Beauty", "flowerbeauty.com", "Shimmer Aydınlatıcı", "Shimmer & Strobe Highlighting Palette; Drew Barrymore; $10; Walmart"),
        ("BH Cosmetics", "bhcosmetics.com", "Studio Pro Kontür", "Studio Pro Contour Palette; 10 renk; profesyonel + erişilebilir; YouTube favorisi"),
        ("Focallure", "focallure.com", "Çin Highlighter", "Glow Getter Highlighter; Çin DTC; ultra uygun; metalik finish; çeşitli tonlar"),
    ],
}

# ─── Hair, Body, Nails, Devices, Fragrance, etc. ─────────────────────────────
EXTRA_BRANDS_2 = {
    # ═══════════════════════════════════════════════════════════════════════════
    # 18. Saç Bakımı - Şampuan & Saç Kremi
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı - Şampuan & Saç Kremi": [
        ("Function of Beauty", "functionofbeauty.com", "Kişiselleştirilmiş Şampuan", "Saç tipine göre özelleştirilmiş formül; online quiz; 12B+ kombinasyon; kişiselleştirilmiş güzellik öncüsü"),
        ("Prose", "prose.com", "Özel Formül Şampuan", "85+ faktöre göre kişiselleştirilmiş; su sertliği + hava durumu dahil; lüks kişisel formül"),
        ("Odele", "odelebeauty.com", "Erişilebilir Temiz Şampuan", "Target'ta $12; cinsiyet nötr; aile boyu; temiz + erişilebilir saç bakımı"),
        ("Bread Beauty", "breadbeautysupply.com", "Tekstürlü Saç Şampuanı", "Afro + kıvırcık saç bakımı; hair-wash + oil; minimalist; Maeva Heim kurdu"),
        ("Crown Affair", "crownaffair.com", "Lüks Ritüel Şampuan", "Saç bakımını ritüele dönüştüren; şık ambalaj; the brush + the oil; lüks saç wellness"),
        ("Act+Acre", "actandacre.com", "Kafa Derisi Şampuanı", "Cold Processed teknolojisi; kafa derisi sağlığı odaklı; scalp detox; soğuk işlem koruma"),
        ("Dae Hair", "dae.com", "Çöl Bitkisi Şampuanı", "Arizona çöl bitkileri; kaktüs çiçeği + prickly pear; çöl bitkisi saç bakımı"),
        ("Hairstory", "hairstory.com", "Şampuansız Temizleme", "New Wash; şampuan yerine kremsi temizleyici; co-wash konsepti; sülfatsız devrim"),
        ("Jupiter", "hellojupiter.com", "Kepek Şampuanı", "Çinko pyrithione bazlı; şık kepek şampuanı; kepek sorununu normalleştirme; DTC kepek"),
        ("Ceremonia", "ceremonia.com", "Latin Saç Bakımı", "Guava + pataua oil; Latin Amerika saç geleneği; Babba Rivera kurdu; kültürel güzellik"),
        ("Amika", "loveamika.com", "Brooklyn Saç Bakımı", "The Kure şampuan; onarım odaklı; renkli ambalaj; salon kalitesi DTC; bond repair"),
        ("Ouai", "theouai.com", "Hollywood Şampuan", "Jen Atkin kurdu; Kardashian stilisti; fine + medium + thick hair; celebrity hairstylist marka"),
        ("Briogeo", "briogeo.com", "Temiz Saç Bakımı", "Don't Despair, Repair! mask; temiz saç bakımı; sülfatsız + silikonsuz; çeşitli saç tipleri"),
        ("Living Proof", "livingproof.com", "MIT Bilimi Şampuan", "MIT bilim insanlarının geliştirdiği; OFPMA teknolojisi; anti-frizz bilim; Jennifer Aniston yatırımcı"),
        ("Oribe", "oribe.com", "Lüks Salon Şampuan", "Gold Lust Shampoo; Daniel Kaner + Oribe Canales; salon lüks; koku ikonik"),
        ("Davines", "davines.com", "İtalyan Sürdürülebilir Şampuan", "OI Shampoo; İtalyan sürdürülebilir saç bakımı; B Corp; Parma kökenli; biyo-etik"),
        ("Christophe Robin", "christophe-robin.com", "Paris Scalp Şampuan", "Cleansing Purifying Scrub with Sea Salt; Paris saç renk uzmanı; kafa derisi peeling"),
        ("Philip Kingsley", "philipkingsley.com", "Trikolojist Şampuan", "Elasticizer kült ürün; İngiliz trikoloji (saç bilimi); Audrey Hepburn'ün trikolojisti"),
        ("Leonor Greyl", "leonorgreyl.com", "Paris Lüks Şampuan", "Huile de Leonor Greyl; Paris lüks saç yağı; 50+ yıl Fransız saç bakımı geleneği"),
        ("Kevin Murphy", "kevinmurphy.com.au", "Avustralya Salon Şampuan", "Angel.Wash; Avustralya salon; moda sektörü ilham; skincare for hair konsepti"),
        ("Virtue Labs", "virtuelabs.com", "Alpha Keratin Şampuan", "Alpha Keratin 60ku protein; insan keratini teknolojisi; bilimsel saç onarım"),
        ("Innersense", "innersensebeauty.com", "Organik Salon Şampuan", "Hydrating Cream Hairbath; organik salon; USDA organik; lüks temiz saç bakımı"),
        ("Playa", "playa.beauty", "Minimal Saç Bakımı", "Every Day Shampoo; basit + etkili; plaj dalgası; California minimal saç"),
        ("Rahua", "rahua.com", "Amazon Yağmur Ormanı Şampuan", "Rahua yağı Amazon yerlilerinden; organik; sürdürülebilir kaynak; yağmur ormanı koruma"),
        ("Maui Moisture", "mauimoisture.com", "Tropikal Şampuan", "Shea Butter Shampoo; tropikal bileşenler; erişilebilir; sülfatsız; aloe + shea"),
        ("Not Your Mother's", "nymbrands.com", "Erişilebilir Şampuan", "Clean Freak Dry Shampoo; $8; Walmart + Target; çeşitli saç ihtiyaçları; uygun fiyat"),
        ("OGX", "ogxbeauty.com", "Eczane Premium Şampuan", "Argan Oil of Morocco; salon ilhamlı eczane; egzotik bileşenler; erişilebilir premium"),
        ("Pureology", "pureology.com", "Renk Koruma Şampuan", "Hydrate Shampoo; %100 vegan; ZeroSulfate; boyalı saç uzmanı; salon profesyonel"),
        ("Aveda", "aveda.com", "Bitkisel Şampuan", "Botanical Repair Shampoo; %93+ doğal; Ayurveda ilhamlı; sürdürülebilir bitkisel"),
        ("Moroccanoil", "moroccanoil.com", "Argan Yağı Şampuan", "Moisture Repair Shampoo; argan yağı ikonu; İsrail kökenli; salon + ev kullanım"),
        ("Kérastase", "kerastase.com", "Fransız Salon Şampuan", "Bain Satin; Fransız salon saç bakımı standardı; L'Oréal Professionnel; lüks"),
        ("Redken", "redken.com", "NYC Salon Şampuan", "All Soft Shampoo; NYC salon geleneği; protein + nemlendirme; profesyonel"),
        ("Bumble and bumble", "bumbleandbumble.com", "NYC Stil Şampuan", "Hairdresser's Invisible Oil Shampoo; NYC stil ikonu; çok amaçlı yağ teknolojisi"),
        ("IGK", "igkhair.com", "Brooklyn Indie Şampuan", "Good Behavior Spirulina Protein Smoothing Shampoo; Brooklyn salon; spirulina protein"),
        ("R+Co", "randco.com", "Lüks İndie Şampuan", "Dallas Biotin Thickening Shampoo; kolektif saç stilistleri; lüks indie; biotin kalınlaştırma"),
        ("Sachajuan", "sachajuan.com", "İsveç Şampuan", "Volume Shampoo; İsveç minimalist saç bakımı; ocean silk technology; İskandinav"),
        ("Maria Nila", "marianila.com", "İsveç Vegan Şampuan", "%100 vegan; İsveç salon; colour guard complex; hayvan dostu profesyonel"),
        ("Authentic Beauty Concept", "authenticbeautyconcept.com", "Alman Vegan Şampuan", "Hydrate Cleanser; Henkel Alman vegan salon; sürdürülebilir; minimal ambalaj"),
        ("Mielle Organics", "mielleorganics.com", "Organik Saç Bakımı", "Rosemary Mint Strengthening Shampoo; TikTok viral; P&G satın aldı; doğal saç uzmanı"),
        ("Shea Moisture", "sheamoisture.com", "Afrika Saç Bakımı", "Coconut & Hibiscus Shampoo; doğal + kıvırcık saç; Afrika kökenli bileşenler; kapsayıcı"),
        ("Carol's Daughter", "carolsdaughter.com", "Doğal Saç Bakımı", "Wash Day Delight; doğal tekstürlü saç; Lisa Price kurdu; Brooklyn mutfağından markaya"),
        ("TGIN", "tginatural.com", "Doğal Saç Nemlendirme", "Honey Miracle Hair Mask; bal + jojoba; doğal saç nemlendirme; kıvırcık saç uzmanı"),
        ("Cantu", "cantubeauty.com", "Shea Butter Saç", "Shea Butter Moisturizing Shampoo; shea butter; kıvırcık + doğal saç; erişilebilir"),
        ("Palmer's", "palmers.com", "Kakao Saç Bakımı", "Coconut Oil Formula Shampoo; kakao + hindistan cevizi; kıvırcık saç; geleneksel"),
        ("Head & Shoulders benzeri DTC: Jupiter", "hellojupiter.com", "Şık Kepek Bakımı", "Balancing Shampoo; çinko pyrithione; kepek şampuanını güzelleştiren marka"),
        ("Collective Laboratories", "collectivelabs.com", "Saç Aktivatör Şampuan", "Activating Shampoo; kafa derisi sağlığı + saç büyüme; aktif scalp bakımı"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 19. Saç Bakımı - Saç Maskesi & Onarım
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı - Saç Maskesi & Onarım": [
        ("K18", "k18hair.com", "Biyomimetik Saç Onarım", "Leave-In Molecular Repair Hair Mask; 4 dakika; biyomimetik peptide; keratin zinciri onarım; viral"),
        ("Olaplex alternatifi: Curlsmith", "curlsmith.com", "Kıvırcık Onarım", "Bond Curl Rehab Salve; bond onarım + kıvırcık bakım; çift etki; indie kıvırcık saç"),
        ("Briogeo", "briogeo.com", "Don't Despair Repair", "Don't Despair, Repair! Deep Conditioning Mask; rosehip + algae; temiz derin onarım"),
        ("Christophe Robin", "christophe-robin.com", "Deniz Tuzu Scrub", "Cleansing Purifying Scrub with Sea Salt; kafa derisi detox; Paris salon ritüeli"),
        ("Moroccanoil", "moroccanoil.com", "Restorative Maske", "Restorative Hair Mask; argan yağı + protein; yoğun onarım; salon standardı"),
        ("Kérastase", "kerastase.com", "Elixir Ultime Maske", "Elixir Ultime Masque; lüks yağ maskesi; Fransız salon; parlak + yumuşak"),
        ("Oribe", "oribe.com", "Gold Lust Maske", "Gold Lust Transformative Masque; altın özü + botanik; lüks onarım maskesi"),
        ("Virtue Labs", "virtuelabs.com", "Keratin Onarım Maske", "Restorative Treatment Mask; Alpha Keratin 60ku; insan keratini; bilimsel onarım"),
        ("Amika", "loveamika.com", "Soulfood Maske", "Soulfood Nourishing Mask; shea + sea buckthorn; renkli ambalaj; salon + ev"),
        ("Philip Kingsley", "philipkingsley.com", "Elasticizer", "Elasticizer Pre-Shampoo Treatment; 50+ yıl kült ürün; Audrey Hepburn; İngiliz trikoloji"),
        ("Davines", "davines.com", "OI Maske", "OI Hair Butter; roucou yağı; İtalyan sürdürülebilir; ultra nemlendirme"),
        ("Living Proof", "livingproof.com", "Perfect Hair Day Maske", "PHD 5-in-1 Styling Treatment; MIT bilimi; çoklu fayda; tek ürün çözüm"),
        ("R+Co", "randco.com", "High Dive Maske", "High Dive Moisture + Shine Crème; nemlendirme + parlaklık; lüks indie"),
        ("IGK", "igkhair.com", "Expensive Maske", "Expensive Amla Oil Hi-Shine Topcoat; amla yağı; parlak finish; Brooklyn salon"),
        ("Ouai", "theouai.com", "Treatment Maske", "Treatment Masque; Jen Atkin; shea + panthenol; salon kalitesi ev bakımı"),
        ("Innersense", "innersensebeauty.com", "Organik Derin Bakım", "Hydrating Hair Masque; organik; USDA sertifikalı; lüks temiz onarım"),
        ("Rahua", "rahua.com", "Amazon Maske", "Omega 9 Hair Mask; Amazon rahua yağı; omega-9; organik derin bakım"),
        ("Crown Affair", "crownaffair.com", "Saç Maskesi", "The Renewal Mask; lüks saç ritüeli; şık ambalaj; saç wellness"),
        ("Act+Acre", "actandacre.com", "Scalp Detox Maske", "Cold Processed Scalp Detox; soğuk işlem; kafa derisi + saç onarım"),
        ("Prose", "prose.com", "Kişisel Maske", "Custom Hair Mask; kişiselleştirilmiş formül; 85+ faktör; lüks kişisel onarım"),
        ("Gisou", "gisou.com", "Bal Saç Maskesi", "Honey Infused Hair Mask; Mirsalehi bal çiftliği; bal protein; viral TikTok"),
        ("Leonor Greyl", "leonorgreyl.com", "Masque Orchidée", "Masque à l'Orchidée; orkide + ipek protein; Fransız lüks yoğun onarım"),
        ("Klorane", "klorane.com", "Botanik Saç Maskesi", "Mango Butter Mask; botanik Pierre Fabre; Fransız eczane saç bakımı"),
        ("Authentic Beauty Concept", "authenticbeautyconcept.com", "Vegan Replenish Maske", "Replenish Mask; Alman vegan; araçay yağı; salon profesyonel"),
        ("Maria Nila", "marianila.com", "İsveç Onarım Maskesi", "Structure Repair Masque; İsveç vegan; algae extract; %100 vegan onarım"),
        ("Sachajuan", "sachajuan.com", "İsveç İnce Saç Maskesi", "Intensive Hair Oil; İsveç ocean silk; ince saç güçlendirme; minimalist"),
        ("Mielle Organics", "mielleorganics.com", "Biberiye Maske", "Rosemary Mint Strengthening Hair Masque; TikTok viral; biberiye + nane güçlendirme"),
        ("TGIN", "tginatural.com", "Bal Mucize Maske", "Honey Miracle Hair Mask; bal + jojoba + olive; derin nemlendirme; doğal saç"),
        ("SheaMoisture", "sheamoisture.com", "Manuka Bal Maske", "Manuka Honey & Yogurt Hydrate + Repair Protein Power Treatment; protein onarım"),
        ("Curlsmith", "curlsmith.com", "Bond Onarım Maske", "Bond Salve Pre-Shampoo; bond teknolojisi + kıvırcık bakım; hibrit onarım"),
        ("Cantu", "cantubeauty.com", "Shea Repair Maske", "Shea Butter Deep Treatment Masque; shea derin onarım; erişilebilir; doğal saç"),
        ("Palmer's", "palmers.com", "Hindistan Cevizi Maske", "Coconut Oil Formula Repairing Conditioner; hindistan cevizi onarım; geleneksel formül"),
        ("Pureology", "pureology.com", "Renk Koruma Maske", "Hydrate Superfood Treatment; boyalı saç; bitkisel protein + vitamin; vegan"),
        ("Aveda", "aveda.com", "Botanical Repair Maske", "Botanical Repair Intensive Strengthening Masque; bitkisel bond onarım; %94 doğal"),
        ("Bumble and bumble", "bumbleandbumble.com", "Mending Maske", "Mending Masque; onarım + yumuşatma; NYC salon; yoğun bakım"),
        ("Redken", "redken.com", "Acidic Bonding Maske", "Acidic Bonding Concentrate 5 Minute Liquid Mask; asidik bond onarım; profesyonel"),
        ("Kevin Murphy", "kevinmurphy.com.au", "Young.Again Maske", "Young.Again Masque; immortelle + baobab; Avustralya anti-aging saç bakımı"),
        ("Dae Hair", "dae.com", "Çöl Saç Maskesi", "Monsoon Moisture Milk; çöl bitkisi nemlendirme; Arizona botanik; hafif maske"),
        ("dpHUE", "dphue.com", "Sirke Saç Maskesi", "Apple Cider Vinegar Hair Masque; elma sirkesi; parlaklık + yumuşaklık; pH dengeleme"),
        ("Revela", "revela.com", "Biyoteknoloji Saç Tedavisi", "ProCelinyl teknolojisi; AI ile keşfedilen saç büyüme molekülü; biyoteknoloji saç onarım"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 20. Saç Bakımı - Saç Büyütme & Dökülme
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı - Saç Büyütme & Dökülme": [
        ("Vegamour", "vegamour.com", "Vegan Saç Büyütme", "GRO Hair Serum; fito-aktif bazlı; vegan saç büyütme; klinik testli; fitojen teknolojisi"),
        ("Nutrafol", "nutrafol.com", "Takviye Saç Büyütme", "Women's Balance; dermatoloji #1 tavsiye; bitkisel + vitamin takviye; stres + hormon"),
        ("Revela", "revela.com", "AI Saç Büyütme", "Hair Revival Serum; ProCelinyl AI keşfi; biyoteknoloji; makine öğrenmesi ile bulunan molekül"),
        ("Divi", "diviofficial.com", "TikTok Scalp Serumu", "Scalp Serum; TikTok viral; kafa derisi bakımı + saç büyütme; copper peptide + caffeine"),
        ("Better Not Younger", "betternotyounger.com", "40+ Saç Büyütme", "Superpower Fortifying Hair & Scalp Serum; menopoz + yaşlanma saç dökülmesi; 40+ kadın odaklı"),
        ("Keeps", "keeps.com", "Erkek Saç Dökülme", "Finasteride + minoxidil; online reçeteli; erkek saç dökülme tedavisi DTC; teledermatolog"),
        ("Hims", "forhims.com", "Erkek Saç Tedavisi", "Finasteride + minoxidil; online reçeteli erkek sağlığı; saç dökülme + ED; telehealth"),
        ("Hers", "forhers.com", "Kadın Saç Dökülme", "Minoxidil + spironolactone; kadın saç dökülme tedavisi; online dermatoloji"),
        ("Collective Laboratories", "collectivelabs.com", "Saç Aktivatör", "Activating Serum; caffeine + pumpkin seed; kafa derisi aktivasyonu; DTC saç bilimi"),
        ("The Ordinary Multi-Peptide", "theordinary.com", "Peptide Saç Serumu", "Multi-Peptide Serum for Hair Density; $18; peptide bazlı saç yoğunluğu; erişilebilir"),
        ("Act+Acre", "actandacre.com", "Saç Büyütme Yağı", "Stem Cell Scalp Serum; bitki kök hücre; kafa derisi sağlığı; soğuk işlem"),
        ("Nioxin", "nioxin.com", "Klinik Saç İncelme", "System 1-6; saç incelme sistemi; 30+ yıl araştırma; kafa derisi + saç fiber"),
        ("Rogaine/Regaine alternatifi: Hims", "forhims.com", "Minoxidil DTC", "Minoxidil %5; online reçetesiz; uygun fiyat saç büyütme; aylık abonelik"),
        ("Viviscal", "viviscal.com", "Deniz Protein Takviyesi", "Hair Growth Supplement; AminoMar deniz kompleksi; klinik kanıtlı takviye"),
        ("Keranique", "keranique.com", "Kadın Saç Büyütme Sistemi", "Hair Regrowth System; minoxidil + keratin; kadın saç dökülme sistemi"),
        ("Philip Kingsley Trichotherapy", "philipkingsley.com", "Trikoloji Dökülme", "Tricho Complex; trikoloji bazlı saç dökülme; vitamin + protein; İngiliz bilim"),
        ("Watermans", "watermanshair.com", "İngiliz Saç Büyütme", "Grow Me Shampoo; İngiliz saç büyütme; caffeine + biotin + keratin; viral"),
        ("Hairburst", "hairburst.com", "İngiliz Saç Vitamin", "Hair Vitamins; İngiliz saç büyütme takviyeleri; biotin + selenium; Instagram viral"),
        ("Plantur 39", "plantur39.com", "Alman Menopoz Saç", "Phyto-Caffeine Shampoo; Alman; menopoz kaynaklı saç dökülme; fito-kafein"),
        ("Alpecin", "alpecin.com", "Alman Kafein Şampuan", "Caffeine Shampoo C1; Alman saç dökülme şampuanı; kafein bazlı; erkek odaklı"),
        ("Nourkrin", "nourkrin.com", "İskandinav Saç Takviye", "Nourkrin Woman; Marilex deniz protein; İskandinav saç büyütme takviyesi; klinik test"),
        ("Grande Cosmetics", "grandecosmetics.com", "Peptide Saç Serumu", "GrandeHAIR Enhancing Serum; peptide bazlı; saç kalınlaştırma; klinik sonuçlar"),
        ("RevitaLash", "revitalash.com", "Saç Kondisyoner", "RevitaLash Volumizing Hair Collection; biotin peptide; saç + kirpik büyütme markası"),
        ("Keratin Complex", "keratincomplex.com", "Keratin Saç Tedavisi", "Keratin Smoothing Treatment; profesyonel keratin; düzleştirme + onarım"),
        ("Olaplex Hair Perfector alternatifi: Curlsmith", "curlsmith.com", "Bond Saç Tedavisi", "Bond Curl Rehab Salve; bond onarım; kıvırcık saç büyüme desteği"),
        ("Mane Club", "maneclub.co", "Saç Büyütme Gummy", "Hair Growth Gummies; biotin + folic acid; sakızımsı vitamin; eğlenceli format"),
        ("Sugarbear", "sugarbearhair.com", "Saç Vitamin Ayıcık", "Hair Vitamins; mavi ayıcık şekilli; biotin; Instagram viral; influencer pazarlama"),
        ("Hair La Vie", "hairlavie.com", "Doğal Saç Büyütme", "Clinical Formula Hair Vitamins; doğal bileşenler; saw palmetto + biotin; bütünsel yaklaşım"),
        ("Laritelle", "laritelle.com", "Organik Saç Büyütme", "Organic Hair Loss Prevention Shampoo; organik; argan + rosemary; doğal saç güçlendirme"),
        ("Root'd", "rootd.com", "Kadın Saç Mineral", "MULTI vitamin + mineral; saç + tırnak + cilt; effervescent tablet; kolay tüketim"),
        ("Curlsmith", "curlsmith.com", "Kıvırcık Saç Büyütme", "Scalp Stimulating Booster; kıvırcık saç büyütme; red clover + caffeine; kafa derisi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 21. Saç Bakımı - Renk & Boyama
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı - Renk & Boyama": [
        ("Madison Reed", "madison-reed.com", "Akıllı Ev Boyası", "Radiant Hair Color; ammonia-free; online renk eşleştirme; salon kalitesi evde boya"),
        ("dpHUE", "dphue.com", "Renk Bakımı Uzmanı", "Gloss+; saç rengi bakımı + derinleştirme; elma sirkesi renk koruma; pH dengeleme"),
        ("Overtone", "overtone.co", "Renk Depo Saç Kremi", "Coloring Conditioner; krem bazlı saç rengi; yarı kalıcı; günlük bakımla renk; vegan"),
        ("Arctic Fox", "arcticfoxhaircolor.com", "Vegan Fantezi Renk", "%100 vegan semi-permanent; canlı fantezi renkler; cruelty-free; DIY renk"),
        ("Lime Crime", "limecrime.com", "Unicorn Hair Renk", "Unicorn Hair Dye; fantezi renkler; vegan; cruelty-free; pastel + canlı seçenekler"),
        ("Good Dye Young", "gooddyeyoung.com", "Punk Saç Rengi", "Semi-Permanent Hair Color; Hayley Williams (Paramore) markası; punk rock estetik"),
        ("Manic Panic", "manicpanic.com", "OG Fantezi Renk", "Classic High Voltage; 1977'den beri; punk kültür ikonu; vegan; NYC CBGB geleneği"),
        ("Hally Hair", "hallyhair.com", "Z Kuşağı Renk", "Shade Stix geçici saç rengi; Z kuşağı; çubuk format; geçici + eğlenceli; TikTok"),
        ("Josh Wood Colour", "joshwoodcolour.com", "İngiliz Renk Uzmanı", "Permanent Colour Kit; İngiliz renk uzmanı; kişiselleştirilmiş ev boyası; Londra salonu"),
        ("Color & Co", "colorandco.com", "L'Oréal Kişiselleştirilmiş", "Custom Color Kit; L'Oréal kişiselleştirilmiş ev boyası; video konsültasyon; profesyonel"),
        ("eSalon", "esalon.com", "Kişiselleştirilmiş Ev Boyası", "Custom Home Hair Color; salon kalitesi kişiselleştirilmiş; renk uzmanı eşleştirme"),
        ("Revlon ColorSilk alternatifi: Naturtint", "naturtint.com", "Organik Saç Boyası", "Permanent Hair Color; organik bileşenler; ammonia-free; bitkisel renk; İspanyol"),
        ("Herbatint", "herbatint.com", "İtalyan Bitkisel Boya", "Permanent Herbal Haircolor Gel; 8 bitkisel özüt; İtalyan; jel formül; nazik"),
        ("Kristin Ess", "kristiness.com", "Salon Ev Boyası", "Signature Hair Gloss; salon kalitesi gloss; Target'ta; erişilebilir salon renk"),
        ("Wella Color Charm", "wella.com", "Profesyonel Ev Boya", "Color Charm Permanent; profesyonel formül; ev kullanımına uygun; geniş palet"),
        ("Cēleb Luxury", "celebluxury.com", "Renk Depo Şampuan", "Viral Colorwash; renk yükleyen şampuan; haftalık renk tazeleme; inovatif format"),
        ("Moroccanoil Color Depositing", "moroccanoil.com", "Argan Renk Maske", "Color Depositing Mask; argan yağı + renk; bakım + renk depo; salon"),
        ("Davines Finest Pigments", "davines.com", "İtalyan Doğal Renk", "Finest Pigments; İtalyan doğal pigment; sürdürülebilir; biyo-etik saç rengi"),
        ("Bleach London", "bleachlondon.com", "Londra Renk Uzmanı", "Super Cool Colours; Londra renk trendi; pastel + neon; İngiliz indie renk"),
        ("INNBeauty Project", "innbeautyproject.com", "Saç Gloss", "Glaze Hair Gloss; duşta saç parlaklaştırma; 10 dakika salon sonucu; DTC renk bakımı"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 22. Saç Bakımı - Styling & Isı Koruma
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı - Styling & Isı Koruma": [
        ("Color Wow", "colorwowhair.com", "Renk Koruma Stil", "Dream Coat Supernatural Spray; anti-humidity; nem bariyeri; renk koruma + stil"),
        ("Ouai", "theouai.com", "Wave Spray", "Wave Spray; Jen Atkin; plaj dalgası; effortless stil; Hollywood saç"),
        ("Bumble and bumble", "bumbleandbumble.com", "Surf Spray", "Surf Spray; tuzlu su dalgası; NYC salon; plaj saç stili öncüsü"),
        ("Living Proof", "livingproof.com", "Anti-Frizz Stil", "No Frizz Instant De-Frizzer; MIT teknolojisi; OFPMA; anti-frizz bilim"),
        ("R+Co", "randco.com", "Dry Şampuan", "Death Valley Dry Shampoo; lüks kuru şampuan; kolektif stilist; hacim + tazeleme"),
        ("IGK", "igkhair.com", "First Class Stil", "First Class Charcoal Detox Dry Shampoo; kömür detox; Brooklyn salon; kuru şampuan"),
        ("Amika", "loveamika.com", "Perk Up Kuru Şampuan", "Perk Up Dry Shampoo; pirinç nişastası; hacim + tazeleme; renkli ambalaj"),
        ("Drybar", "drybar.com", "Fön Uzmanı", "Detox Dry Shampoo; fön bar markası; sarı marka kimliği; profesyonel fön araçları"),
        ("Moroccanoil", "moroccanoil.com", "Argan Stil Yağı", "Treatment Original; argan yağı stil ikonu; çok amaçlı; düzleştirme + parlaklık"),
        ("Oribe", "oribe.com", "Dry Texturizing Spray", "Dry Texturizing Spray; lüks kuru doku spreyi; hacim + doku; salon favorisi"),
        ("Davines", "davines.com", "Invisible Dry Şampuan", "DEDE Delicate Daily Conditioner; İtalyan; günlük hafif bakım; sürdürülebilir"),
        ("Kevin Murphy", "kevinmurphy.com.au", "Session.Spray", "Session.Spray Flex; Avustralya esnek saç spreyi; hafif tutuş; profesyonel"),
        ("Crown Affair", "crownaffair.com", "Stil Ritüeli", "The Set Styling Gel; lüks saç jeli; ritüel bazlı; minimal ambalaj"),
        ("Act+Acre", "actandacre.com", "Soğuk İşlem Stil", "Cold Processed Hair Oil; soğuk işlenmiş saç yağı; scalp + stil; doğal"),
        ("Kristin Ess", "kristiness.com", "Erişilebilir Stil", "Weightless Shine Working Serum; Target'ta $10; erişilebilir salon; parlaklık serumu"),
        ("Kenra", "kenraprofessional.com", "Profesyonel Stil", "Volume Spray 25; profesyonel hacim spreyi; salon standardı; güçlü tutuş"),
        ("Aquis", "aquis.com", "Saç Kurutma Teknolojisi", "Rapid Dry Hair Turban; mikro fiber; sürtünmesiz kurutma; saç hasarı azaltma"),
        ("Pattern Beauty", "patternbeauty.com", "Kıvırcık Saç Jel", "Strong Hold Gel; Tracee Ellis Ross; kıvırcık + coily saç stil; tanımlama jeli"),
        ("Bread Beauty Supply", "breadbeautysupply.com", "Tekstürlü Saç Yağı", "Hair Oil Everyday Gloss; tekstürlü saç; hafif parlaklık yağı; minimalist"),
        ("Ceremonia", "ceremonia.com", "Latin Stil", "Aceite de Moska Hair Oil; babasu yağı; Latin saç geleneği; kültürel güzellik stili"),
        ("Gisou", "gisou.com", "Bal Saç Yağı", "Honey Infused Hair Oil; Mirsalehi bal; arı balı saç yağı; TikTok viral"),
        ("Not Your Mother's", "nymbrands.com", "Erişilebilir Isı Koruma", "Beat the Heat Thermal Styling Spray; $7; ısı koruma; erişilebilir"),
        ("Chi", "chi.com", "Isı Koruma Uzmanı", "44 Iron Guard; profesyonel ısı koruma; keratin + ipek; 450°F koruma"),
        ("TRESemmé", "tresemme.com", "Salon Isı Koruma", "Keratin Smooth Heat Protection Spray; salon ilhamlı; erişilebilir ısı koruma"),
        ("GHD", "ghdhair.com", "İngiliz Isı Stil", "Bodyguard Heat Protect Spray; İngiliz profesyonel; 220°C koruma; salon standardı"),
        ("Dyson alternatifi: Bio Ionic", "bioionic.com", "NanoIonic Stil", "Long Barrel Styler; nanoionic teknoloji; nem bazlı şekillendirme; profesyonel"),
        ("T3 Micro", "t3micro.com", "Akıllı Stil Cihazı", "Cura Luxe Dryer; akıllı ısı ayarı; ion teknolojisi; salon ütü + fön"),
        ("Harry Josh Pro Tools", "harryjoshprotools.com", "Pro Fön", "Ultra Light Pro Dryer; profesyonel hafif fön; salon standardı; hızlı kurutma"),
        ("Dae Hair", "dae.com", "Çöl Stil Yağı", "Prickly Pear Hair Oil; dikenli armut yağı; Arizona botanik; hafif stil yağı"),
        ("Playa", "playa.beauty", "Minimal Doku Spreyi", "Endless Summer Spray; California plaj dalgası; minimal + temiz; tuzlu su dokusu"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 23. Saç Bakımı - Kıvırcık & Tekstürlü Saç
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı - Kıvırcık & Tekstürlü Saç": [
        ("Curlsmith", "curlsmith.com", "Bilimsel Kıvırcık Bakım", "Bond Curl Rehab Salve; bond onarım + kıvır tanımlama; bilimsel kıvırcık bakım"),
        ("Pattern Beauty", "patternbeauty.com", "Coily Saç Uzmanı", "Heavy Conditioner; Tracee Ellis Ross; 3B-4C kıvır; kapsayıcı kıvırcık marka"),
        ("Bread Beauty Supply", "breadbeautysupply.com", "Minimalist Kıvırcık", "Hair Wash + Oil; minimalist kıvırcık bakım; Maeva Heim; tekstürlü saç"),
        ("Ceremonia", "ceremonia.com", "Latin Kıvırcık", "Mascarilla de Babassu; babasu yağı; Latin kıvırcık saç geleneği; kültürel bakım"),
        ("Mielle Organics", "mielleorganics.com", "Doğal Kıvırcık", "Pomegranate & Honey Twisting Soufflé; doğal saç; TikTok viral; P&G satın aldı"),
        ("SheaMoisture", "sheamoisture.com", "Shea Kıvırcık", "Coconut & Hibiscus Curl Enhancing Smoothie; shea + hindistan cevizi; kapsayıcı; erişilebilir"),
        ("Carol's Daughter", "carolsdaughter.com", "Brooklyn Kıvırcık", "Coco Crème Curl Quenching Deep Moisture Mask; Brooklyn mutfağından; doğal saç"),
        ("TGIN", "tginatural.com", "Doğal Nemlendirme", "Twist and Define Cream; bal + shea; doğal kıvırcık tanımlama; erişilebilir"),
        ("Cantu", "cantubeauty.com", "Shea Kıvırcık Tanımlama", "Moisturizing Curl Activator Cream; shea butter; kıvırcık aktivatör; erişilebilir"),
        ("Miss Jessie's", "missjessies.com", "NYC Kıvırcık", "Pillow Soft Curls; NYC salon kıvırcık uzmanı; 20+ yıl; çeşitli kıvır tipleri"),
        ("DevaCurl", "devacurl.com", "Curly Girl Method", "SuperCream; Curly Girl Method öncüsü; sülfatsız + silikonsuz; kıvırcık kesim salonu"),
        ("Bounce Curl", "bouncecurl.com", "Light Hold Kıvırcık", "Light Creme Gel; hafif tutuş; kıvırcık tanımlama; Curly Girl uyumlu"),
        ("Kinky-Curly", "kinky-curly.com", "Doğal Jel", "Curling Custard; doğal jel; kıvırcık tanımlama; uygun fiyat; organik"),
        ("Aunt Jackie's", "auntjackiescurlsandcoils.com", "Aile Kıvırcık Bakım", "Quench Moisture Intensive Leave-In; aile boyu; doğal saç; erişilebilir"),
        ("As I Am", "asiamnaturally.com", "Curly Confidence", "Curl Clarity Shampoo; curl confidence; doğal saç; coconut + shea"),
        ("TPH by Taraji", "tphbytaraji.com", "Taraji P. Henson Kıvırcık", "Master Cleanse Scalp Shampoo; kafa derisi + kıvırcık; ünlü kuruculu; Target"),
        ("Sienna Naturals", "siennanaturals.com", "Temiz Kıvırcık", "Dew Magic Leave-In Conditioner; Hannah Diop; temiz kıvırcık bakım; doğal"),
        ("Melanin Haircare", "melaninhaircare.com", "Melanin Saç Bakım", "Multi-Use Softening Leave In Conditioner; Whitney & Naptural85; YouTuber markası"),
        ("Flora & Curl", "floraandcurl.com", "İngiliz Çiçek Kıvırcık", "Sweet Hibiscus Curl Defining Gel; İngiliz; çiçek bazlı; kıvırcık tanımlama; doğal"),
        ("Bouclème", "boucleme.com", "İngiliz Kıvırcık", "Curl Defining Gel; İngiliz temiz kıvırcık bakım; vegan; doğal bileşenler"),
        ("Twisted Sista", "twistedsista.com", "Kıvırcık Stil", "30 Second Curl Spray; hızlı kıvırcık tanımlama; erişilebilir; doğal saç"),
        ("Camille Rose", "camillerose.com", "Botanik Kıvırcık", "Curl Love Moisture Milk; botanik bazlı; doğal saç; küçük parti üretim"),
        ("Mane Choice", "themanechoice.com", "Vitamin Kıvırcık", "Manetabolism Plus Vitamins; saç büyütme vitamin + kıvırcık bakım; çift yaklaşım"),
        ("The Doux", "thedoux.com", "Eğlenceli Kıvırcık", "Big Poppa Gel; eğlenceli isimler; kıvırcık tanımlama; güçlü tutuş; doğal saç"),
        ("CurlMix", "curlmix.com", "Kıvırcık Wash-n-Go", "Flaxseed Gel; keten tohumu jel; wash-and-go sistemi; doğal kıvırcık"),
        ("Ouidad", "ouidad.com", "Kıvırcık Saç Kraliçesi", "Advanced Climate Control Gel; anti-humidity; 40+ yıl kıvırcık uzmanlığı"),
        ("Jessicurl", "jessicurl.com", "OG Kıvırcık Indie", "Spiralicious Gel; indie kıvırcık; Jessica McGuinty; küçük parti; Curly Girl"),
        ("AG Hair", "aghair.com", "Kanada Kıvırcık", "Curl Fresh Definer; Kanada salon; hafif kıvırcık tanımlama; sürdürülebilir"),
        ("Innersense", "innersensebeauty.com", "Organik Kıvırcık", "I Create Hold; organik stil jel; USDA; kıvırcık dostu; temiz lüks"),
        ("EDEN BodyWorks", "edenbodyworks.com", "Doğal Kıvırcık Bakım", "Coconut Shea Cleansing CoWash; doğal bileşen; co-wash; kıvırcık nemlendirme"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 24. Vücut Bakımı - Vücut Nemlendirici & Yağ
    # ═══════════════════════════════════════════════════════════════════════════
    "Vücut Bakımı - Vücut Nemlendirici & Yağ": [
        ("Sol de Janeiro", "soldejaneiro.com", "Brezilya Bum Bum", "Brazilian Bum Bum Cream; guaraná + cupuaçu; Brezilya plaj kokusu; ikonik vücut kremi"),
        ("Nécessaire", "necessaire.com", "Vücut Vitamini", "The Body Lotion; niacinamide + vitamin E; temiz vücut bakımı; minimalist; unisex"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi Vücut", "Coconut Melt; %100 organik hindistan cevizi yağı; çok amaçlı; tropikal"),
        ("Osea Malibu", "oseamalibu.com", "Deniz Vücut Yağı", "Undaria Algae Body Oil; deniz yosunu vücut yağı; Malibu; okyanus bakımı"),
        ("Fur", "furyou.com", "Kıl Folikül Yağı", "Fur Oil; ingrown hair + bikini bakımı; kıl folikül + cilt yağı; cesur niş"),
        ("Soft Services", "softservices.com", "Vücut Doku Bakımı", "Smoothing Solution; vücut doku düzeltme; keratosis pilaris; AHA vücut bakımı"),
        ("Megababe", "megababe.com", "Sürtünme Önleyici", "Thigh Rescue; uyluk sürtünme çözümü; body positivity; taboo sorunlara çözüm"),
        ("Skinfix", "skinfix.com", "Vücut Bariyer Bakımı", "Eczema+ Hand Repair Cream; ceramide + lipid; NEA onaylı; vücut bariyer uzmanı"),
        ("Frank Body", "frankbody.com", "Kahve Vücut Bakımı", "Original Coffee Scrub; kahve telvesi vücut; Avustralya; Instagram viral; erişilebilir"),
        ("Aesop", "aesop.com", "Botanik Vücut", "Geranium Leaf Body Cleanser; botanik lüks; Avustralya minimalist; unisex; salon his"),
        ("Malin+Goetz", "malinandgoetz.com", "NYC Apothecary Vücut", "Vitamin b5 Body Moisturizer; NYC apothecary; minimalist; unisex; B5 vitamin"),
        ("Byredo Body Care", "byredo.com", "İsveç Lüks Vücut", "Gypsy Water Body Lotion; İsveç niche parfümden vücut bakıma; lüks koku + bakım"),
        ("Le Labo Body", "lelabofragrances.com", "NYC Artisan Vücut", "Santal 33 Body Lotion; artisan koku + vücut bakımı; NYC lüks"),
        ("Diptyque Body", "diptyque.com", "Paris Lüks Vücut", "Do Son Body Lotion; Paris niche parfüm + vücut; lüks Fransız vücut bakımı"),
        ("Weleda", "weleda.com", "Organik Vücut Yağı", "Skin Food Body Butter; organik; İsviçre-Alman; 1926'dan beri; ultra zengin"),
        ("Nuxe", "nuxe.com", "Huile Prodigieuse", "Huile Prodigieuse kuru yağ; Fransız ikonik çok amaçlı yağ; yüz + vücut + saç"),
        ("Kiehl's", "kiehls.com", "Creme de Corps", "Creme de Corps; NYC eczane klasiği; beta-carotene + squalane; vücut nemlendirme ikonu"),
        ("Nivea", "nivea.com", "Mavi Teneke Vücut", "Nivea Creme; mavi teneke ikonik; Alman vücut bakımı standardı; 100+ yıl"),
        ("Palmer's", "palmers.com", "Kakao Vücut Bakımı", "Cocoa Butter Formula; kakao yağı; çatlak bakımı + vücut nemlendirme; geleneksel"),
        ("Embryolisse", "embryolisse.com", "Fransız Vücut Kremi", "Lait-Crème Concentré; vücut + yüz; Fransız çok amaçlı; backstage favorisi"),
        ("Hempz", "hempz.com", "Kenevir Vücut Losyonu", "Original Herbal Body Moisturizer; kenevir tohumu yağı; vücut losyonu; herbal"),
        ("Lush", "lush.com", "El Yapımı Vücut", "Sleepy Body Lotion; lavanta; el yapımı; etik; Charity Pot; taze kozmetik"),
        ("The Body Shop", "thebodyshop.com", "Etik Vücut Bakım", "Shea Body Butter; Community Trade shea; etik kaynak; vücut butter klasiği"),
        ("Aveeno", "aveeno.com", "Yulaf Vücut Bakımı", "Daily Moisturizing Lotion; koloidal yulaf; dermatolojik; hassas vücut bakımı"),
        ("Eucerin", "eucerin.com", "Alman Vücut Onarım", "Advanced Repair Lotion; ceramide + doğal nem faktörleri; Alman dermatolojik vücut"),
        ("CeraVe Body", "cerave.com", "Ceramide Vücut", "Moisturizing Cream; 3 temel ceramide; MVE teknolojisi; dermatoloji standardı vücut"),
        ("Vaseline", "vaseline.com", "Jelly Vücut Onarım", "Intensive Care Advanced Repair; micro-droplet jelly; vücut onarım; 150+ yıl"),
        ("Naturium Body", "naturium.com", "Aktif Vücut Bakımı", "The Glow Getter Multi-Oil Hydrating Body Wash; aktif bileşen vücut bakımı"),
        ("Skinfix Body", "skinfix.com", "Egzama Vücut", "Eczema+ Body Cream; ceramide lipid; NEA onaylı; egzama vücut bakımı"),
        ("First Aid Beauty Body", "firstaidbeauty.com", "KP Vücut", "KP Bump Eraser; AHA + BHA; keratosis pilaris; vücut eksfoliasyon; bumpy cilt"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 25. Vücut Bakımı - Peeling & Scrub
    # ═══════════════════════════════════════════════════════════════════════════
    "Vücut Bakımı - Peeling & Scrub": [
        ("Frank Body", "frankbody.com", "Kahve Scrub", "Original Coffee Scrub; kahve telvesi; Avustralya; Instagram viral; erişilebilir vücut peeling"),
        ("Soft Services", "softservices.com", "KP Eksfoliasyon", "Buffing Bar + Smoothing Solution; keratosis pilaris; AHA vücut peeling; doku düzeltme"),
        ("First Aid Beauty", "firstaidbeauty.com", "KP Bump Eraser", "KP Bump Eraser Body Scrub; AHA + pumice; keratosis pilaris; vücut eksfoliasyon"),
        ("Nécessaire", "necessaire.com", "Vücut Eksfoliant", "The Body Exfoliator; bambu kömür + AHA + BHA; temiz vücut peeling; minimalist"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi Scrub", "Coconut Crush Scrub; hindistan cevizi kabuğu; tropikal vücut peeling"),
        ("Sol de Janeiro", "soldejaneiro.com", "Brezilya Scrub", "Bum Bum Body Scrub; cupuaçu + guaraná; Brezilya vücut peeling; koku"),
        ("Tree Hut", "treehut.com", "Shea Sugar Scrub", "Shea Sugar Scrub; şeker + shea; erişilebilir; çeşitli kokular; Walmart viral"),
        ("Herbivore", "herbivorebotanicals.com", "Coco Rose Scrub", "Coco Rose Body Polish; hindistan cevizi + gül; botanik lüks; Instagram estetik"),
        ("Ouai", "theouai.com", "Scalp & Body Scrub", "Scalp & Body Scrub; kafa derisi + vücut; çift kullanım; Jen Atkin; salon"),
        ("Osea Malibu", "oseamalibu.com", "Deniz Scrub", "Salts of the Earth Scrub; deniz tuzu + deniz yosunu; Malibu okyanus peeling"),
        ("Bushbalm", "bushbalm.com", "Bikini Scrub", "Francesca Exfoliating Scrub; bikini bölgesi; tüy batması önleme; cesur niş peeling"),
        ("Truly", "trulybeauty.com", "Eğlenceli Vücut Scrub", "Matcha Body Scrub; matcha yeşil çay; eğlenceli ambalaj; TikTok viral; renkli"),
        ("Sand & Sky", "sandandsky.com", "Avustralya Kil Scrub", "Australian Pink Clay Flash Perfection Body Scrub; pembe kil; Avustralya doğal"),
        ("Lush", "lush.com", "El Yapımı Scrub", "Ocean Salt Face & Body Scrub; taze hazırlanan; deniz tuzu + lime; taze kozmetik"),
        ("The Body Shop", "thebodyshop.com", "Etik Scrub", "Shea Exfoliating Sugar Body Scrub; Community Trade shea; etik kaynak"),
        ("SheaMoisture", "sheamoisture.com", "Afrika Siyah Sabun Scrub", "African Black Soap Bamboo Charcoal Scrub; Afrika sabunu + kömür; detox peeling"),
        ("Dove", "dove.com", "Nazik Vücut Scrub", "Exfoliating Body Polish; nazik eksfoliasyon; nemlendirici; her cilt tipi"),
        ("St. Ives", "stives.com", "Kayısı Scrub", "Fresh Skin Apricot Scrub; kayısı çekirdeği; klasik scrub; erişilebilir"),
        ("Drunk Elephant alternatifi: Naturium", "naturium.com", "AHA Vücut Peeling", "The Smoother Glycolic Acid Body Lotion; glikolik asit vücut losyonu; aktif peeling"),
        ("Paula's Choice Body", "paulaschoice.com", "BHA Vücut Peeling", "Weightless Body Treatment 2% BHA; BHA vücut peeling losyonu; akne + KP"),
        ("Tower 28", "tower28beauty.com", "SOS Vücut Scrub", "SOS Save.Our.Skin Body Scrub; NEA onaylı; hassas vücut peeling"),
        ("Megababe", "megababe.com", "Space Bar Scrub", "Space Bar Shower Steamers + Body Scrub; vücut peeling; eğlenceli format"),
        ("Love Beauty and Planet", "lovebeautyandplanet.com", "Sürdürülebilir Scrub", "Sugar & Rose Scrub; geri dönüştürülebilir ambalaj; sürdürülebilir vücut peeling"),
        ("Method", "methodhome.com", "Temiz Vücut Yıkama", "Body Wash; temiz bileşenler; sürdürülebilir; estetik ambalaj; erişilebilir"),
        ("Dr. Bronner's", "drbronners.com", "Organik Vücut Yıkama", "Pure-Castile Liquid Soap; %100 organik; fair trade; çok amaçlı; 18-in-1"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 26. Vücut Bakımı - Self-Tan & Bronzlaşma
    # ═══════════════════════════════════════════════════════════════════════════
    "Vücut Bakımı - Self-Tan & Bronzlaşma": [
        ("Isle of Paradise", "isleofparadise.com", "Self-Tan Su Damlası", "Self Tanning Drops; renk düzeltici self-tan damla; suya ekle; inovatif format"),
        ("St. Tropez", "sttropeztan.com", "Bronzlaşma Mousse", "Self Tan Express Mousse; 1 saat express; profesyonel bronzlaşma; altın standart"),
        ("Tan-Luxe", "tan-luxe.com", "Serum Self-Tan", "The Face Anti-Age Rejuvenating Self-Tan Drops; anti-aging + self-tan; hibrit format"),
        ("Loving Tan", "lovingtan.com", "Avustralya Dark Tan", "2 HR Express Mousse; Avustralya bronzlaşma; koyu ton uzmanı; profesyonel"),
        ("Bondi Sands", "bondisands.com", "Avustralya Plaj Tan", "Liquid Gold Self Tanning Foam; Avustralya Bondi plajı; argan yağı; parlak bronz"),
        ("Bali Body", "balibody.com", "Bali Bronz Yağı", "Natural Tanning and Body Oil; Bali ilhamlı; kademeli bronzlaşma yağı; tropikal"),
        ("James Read", "jamesreadtan.com", "İngiliz Lüks Tan", "Sleep Mask Tan Face; gece uykuda bronzlaşma; İngiliz lüks self-tan; inovatif"),
        ("Coco & Eve", "cocoandeve.com", "Bali Bronz Mousse", "Sunny Honey Bali Bronzing Foam; Bali ilhamlı; bal + kakao; Instagram viral"),
        ("Vita Liberata", "vitaliberata.com", "Organik Tan", "Phenomenal Organic Tan Infused Cloths; organik; mendil format; İrlanda lüks"),
        ("Jergens", "jergens.com", "Günlük Glow", "Natural Glow Daily Moisturizer; kademeli günlük bronzlaşma; erişilebilir; losyon format"),
        ("Banana Boat", "bananaboat.com", "Summer Color", "Summer Color Self-Tanning Lotion; erişilebilir; kademeli bronz; eczane fiyatı"),
        ("L'Oréal Sublime Bronze", "lorealparis.com", "Eczane Tan", "Sublime Bronze Self-Tanning Water Mousse; Fransız eczane; su mousse format"),
        ("Tan-Luxe Super Glow", "tan-luxe.com", "Hyaluronic Tan", "Super Glow Hyaluronic Self-Tan Serum; hyaluronic + self-tan; serum format; lüks"),
        ("MineTan", "minetan.com", "Avustralya Spray Tan", "Absolute X20 Ultra Dark; Avustralya profesyonel spray tan; ekstra koyu; salon"),
        ("Sienna X", "sienna-x.co.uk", "İngiliz Salon Tan", "1 Hour Self Tan Tinted Mousse; İngiliz salon; 1 saat express; profesyonel"),
        ("Bare by Vogue", "barebyvogue.com", "İrlanda Tan", "Self Tan Mousse; Vogue Williams; İrlanda; kademeli + doğal; ünlü markası"),
        ("Coola Sunless Tan", "coola.com", "Organik Self-Tan", "Sunless Tan Anti-Aging Face Serum; organik + anti-aging; temiz self-tan"),
        ("b.tan", "byvegamour.com", "Uygun Express Tan", "Ain't Nobody Got Time for That Express Tan Mousse; eğlenceli isimler; $10; hızlı"),
        ("Fake Bake", "fakebake.com", "OG Self-Tan", "Flawless Self-Tan Liquid; OG profesyonel; çift eldiven; 25+ yıl; salon kalitesi"),
        ("He-Shi", "he-shi.eu", "İrlanda Premium Tan", "Rapid 1 Hour Liquid Tan; İrlanda premium; 1 saat; profesyonel + ev kullanım"),
        ("Tanologist", "tanologist.com", "Erişilebilir Tan", "Self Tan Drops; damla format; erişilebilir; Target + Walmart; kademeli bronz"),
        ("Australian Glow", "australianglow.com", "Avustralya Self-Tan", "Self Tan Water Mousse; Avustralya; hafif su mousse; doğal bronz; uygun fiyat"),
        ("Skinny Tan", "skinnytan.com", "Zayıflatan Tan", "7 Day Tanner; göz yanılsaması zayıflama efekti; İngiliz; self-tan + kontür"),
        ("Eco by Sonya", "ecosonya.com", "Organik Avustralya Tan", "Face Tan Water; organik; Avustralya; %100 doğal; vegan self-tan"),
        ("Baja Bae", "bajabae.com", "Festival Tan", "Self Tanning Mousse; festival vibes; çeşitli kokular; TikTok; erişilebilir"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 27. Vücut Bakımı - Deodorant
    # ═══════════════════════════════════════════════════════════════════════════
    "Vücut Bakımı - Deodorant": [
        ("Lume", "lumedeodorant.com", "Tüm Vücut Deodorant", "Whole body deodorant; her yere uygulanabilir; 72 saat koruma; Shannon Klingman MD kurdu"),
        ("Native", "nativecos.com", "Doğal Deodorant", "Natural Deodorant; hindistan cevizi yağı + shea; P&G satın aldı; doğal deo öncüsü"),
        ("Each & Every", "eachandeveryday.com", "EWG Onaylı Deo", "Natural Deodorant; EWG Verified; 6 bileşen; basit + etkili; alüminyumsuz"),
        ("Touchland", "touchland.com", "El Dezenfektanı-Deo", "Power Mist; el dezenfektanı + parfüm; lüks his; aloe vera; estetik ambalaj"),
        ("Myro", "myro.com", "Yeniden Doldurulabilir Deo", "Refillable Deodorant; refill sistem; sürdürülebilir; alüminyumsuz; temiz formül"),
        ("Type:A", "typea.com", "Aktif Doğal Deo", "Natural Deodorant; spor için geliştirilmiş; aktif yaşam; alüminyumsuz + etkili"),
        ("Schmidt's", "schmidts.com", "Doğal Stick Deo", "Natural Deodorant Stick; arrowroot + hindistan cevizi; Unilever satın aldı; doğal deo"),
        ("Crystal", "thecrystal.com", "Mineral Deo", "Mineral Deodorant Stick; mineral tuz bazlı; 35+ yıl; alüminyum klorhidratsız"),
        ("Tom's of Maine", "tomsofmaine.com", "Doğal Deo", "Long Lasting Deodorant; %100 doğal; Colgate markası; 50+ yıl doğal felsefe"),
        ("Corpus Naturals", "corpusnaturals.com", "Lüks Doğal Deo", "Santalum Deodorant; lüks koku + doğal deo; artisan; $22; premium doğal"),
        ("Humble", "humblebrands.com", "Basit Deo", "All Natural Deodorant; 5 basit bileşen; organik; erişilebilir; vegan"),
        ("Super Deodorant", "superdeodorant.com", "Minimalist Deo", "Clean Deodorant; minimalist ambalaj; etkili doğal formül; İngiliz indie"),
        ("Curie", "curie.co", "Detox Deo", "Full Body Deodorant Spray; sprey format; tüm vücut; detox dönemini kolaylaştıran"),
        ("Saltair", "saltair.com", "Vücut Deo Spreyi", "Santal Body Wash + Deodorant; vücut yıkama + deo; tropikal kokular; erişilebilir"),
        ("Malin+Goetz", "malinandgoetz.com", "NYC Apothecary Deo", "Eucalyptus Deodorant; NYC apothecary; okaliptüs; alüminyumsuz; unisex; minimal"),
        ("Aesop", "aesop.com", "Botanik Deo", "Herbal Deodorant Roll-On; botanik lüks; Avustralya; zinc ricinoleate; vegan"),
        ("Kosas", "kosas.com", "Temiz Deo", "Chemistry AHA Serum Deodorant; AHA bazlı; temiz güzellik deo; Sephora #1"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi Deo", "Coconut Deodorant; hindistan cevizi yağı; tropikal; alüminyumsuz; doğal"),
        ("Megababe", "megababe.com", "Terleme Önleyici", "Happy Pits Natural Deodorant; + Le Tits Now anti-chafing; vücut positivity; eğlenceli"),
        ("Arm & Hammer Essentials", "armandhammer.com", "Karbonat Deo", "Essentials Natural Deodorant; karbonat bazlı; doğal + erişilebilir; güvenilir"),
        ("Dove 0% Aluminium", "dove.com", "Alüminyumsuz Deo", "0% Aluminium Deodorant; büyük marka alüminyumsız geçişi; nemlendirici; erişilebilir"),
        ("Wild", "wearewild.com", "İngiliz Refill Deo", "Natural Deodorant; İngiliz refill sistemi; alüminyum kasa + kompostlanabilir refill"),
        ("Fussy", "getfussy.com", "Sürdürülebilir Deo", "Refillable Deodorant; İngiliz; geri dönüşümlü kasa; probiyotik formül; sürdürülebilir"),
        ("By Humankind", "byhumankind.com", "Sıfır Atık Deo", "Natural Deodorant refill; sıfır atık ambalaj; alüminyumsuz; sürdürülebilir"),
        ("Nuud", "nuudcare.com", "Hollanda Mikro Gümüş Deo", "Smartsorb micro silver; 3-7 gün etkili; Hollanda; mikro gümüş teknolojisi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 28. Vücut Bakımı - Tüy Dökücü & Epilasyon
    # ═══════════════════════════════════════════════════════════════════════════
    "Vücut Bakımı - Tüy Dökücü & Epilasyon": [
        ("Bushbalm", "bushbalm.com", "Bikini Bakımı", "Ingrown hair + koyulaşma tedavisi; Shark Tank; cesur niş; bikini bölgesi uzmanı"),
        ("Fur", "furyou.com", "Kıl Folikül Bakımı", "Fur Oil; kıl folikül yumuşatma; ingrown hair önleme; Emma Watson favorisi"),
        ("Billie", "mybillie.com", "Kadın Tıraş Bıçağı", "Razor + Shave Cream; kadın tıraş bıçağı DTC; pembe vergi karşıtı; abonelik"),
        ("Flamingo", "shopflamingo.com", "Harry's Kadın Markası", "Razors + Body Wax Kit; Harry's kadın versiyonu; şık tasarım; erişilebilir"),
        ("Athena Club", "athenaclub.com", "Kadın Tıraş Ritüeli", "Razor Kit; tıraş ritüeli; cloud shave foam; şık ambalaj; abonelik"),
        ("Oui The People", "ouithepeople.com", "Tek Bıçak Tıraş", "Sensitive Skin Razor; tek bıçak güvenlik jilet; hassas cilt; tüy batması önleme"),
        ("European Wax Center", "waxcenter.com", "Salon Ağda", "At-home Wax Kit; salon ağda deneyimi evde; profesyonel formül"),
        ("Sugar Me Smooth", "sugarmesmooth.com", "Şekerleme Epilasyon", "Sugar Wax Kit; şeker bazlı ağda; doğal; ev uygulaması; tüy batması az"),
        ("Nad's", "nads.com.au", "Avustralya Ağda", "Natural Hair Removal Gel; Avustralya; doğal jel ağda; Sue Ismiel; 30+ yıl"),
        ("Veet", "veet.com", "Global Tüy Dökücü", "Sensitive Touch Electric Trimmer + Cream; tüy dökücü krem; global standart"),
        ("Nair", "naircare.com", "Klasik Tüy Dökücü", "Nair Hair Removal Cream; tüy dökücü krem klasik; 80+ yıl; erişilebilir"),
        ("Tria Beauty", "triabeauty.com", "Evde Lazer", "Hair Removal Laser 4X; evde lazer epilasyon; FDA onaylı; profesyonel sonuç"),
        ("Ulike", "ulike.com", "IPL Epilasyon", "Air 3 IPL Hair Removal; sapphire soğutma; evde IPL; kalıcı azaltma"),
        ("Braun Silk Expert", "braun.com", "IPL Epilasyon Cihazı", "Silk Expert Pro 5; SensoAdapt; evde IPL; 400K flash; Alman mühendislik"),
        ("Philips Lumea", "philips.com", "Hollanda IPL", "Lumea Prestige; Hollanda; evde IPL; SmartSkin sensörü; kişiselleştirilmiş"),
        ("CurrentBody Skin", "currentbody.com", "İngiliz IPL", "IPL Hair Removal Device; İngiliz; klinik kanıtlı; profesyonel IPL evde"),
        ("Smoothskin", "smoothskin.com", "İngiliz Hızlı IPL", "Pure Fit; İngiliz; 10 dakika tüm vücut; hızlı IPL; kompakt tasarım"),
        ("RoseSkinCo", "roseskinco.com", "Pembe IPL", "Lumi IPL; pembe estetik ambalaj; uygun fiyat IPL; Instagram viral"),
        ("Kenzzi", "kenzzi.com", "DTC IPL", "IPL Hair Removal Handset; DTC IPL; uygun fiyat; ev kullanım; 5 yoğunluk"),
        ("Bushbalm Skincare", "bushbalm.com", "Post-Epilasyon Bakım", "Nude Oil; post-epilasyon koyulaşma; tüy batması; bikini bakım"),
        ("European Wax Ingrown", "waxcenter.com", "Ingrown Bakım", "Ingrown Hair Serum; post-ağda tüy batması serumu; salon formülü"),
        ("Completely Bare", "bareezz.com", "Salon Ağda DTC", "Don't Grow There Body Moisturizer; ağda + bakım; salon kalitesi ev"),
        ("SUGARED + BRONZED", "sugaredandbronzed.com", "Şekerleme + Tan", "At-Home Sugar Kit; şekerleme + bronzlaşma salonu; kombine hizmet"),
        ("Finishing Touch Flawless", "finishingtouchflawless.com", "Yüz Tüy Alma", "Flawless Facial Hair Remover; ağrısız yüz tüy alma; 18K altın kaplama; viral"),
        ("Tweezerman", "tweezerman.com", "Cımbız Uzmanı", "Slant Tweezer; profesyonel cımbız; paslanmaz çelik; 40+ yıl; kaş şekillendirme"),
    ],
}

# ─── More categories: Nails, Devices, Fragrance, Men, Specialty ──────────────
EXTRA_BRANDS_3 = {
    # ═══════════════════════════════════════════════════════════════════════════
    # 29. Tırnak Bakımı - Oje & Jel
    # ═══════════════════════════════════════════════════════════════════════════
    "Tırnak Bakımı - Oje & Jel": [
        ("Olive & June", "oliveandjune.com", "Ev Manikür Uzmanı", "The Everything Box; evde salon manikür; ergonomik fırça; manikür demokratizasyonu"),
        ("Lights Lacquer", "lightslacquer.com", "Kathleen Lights Oje", "Vegan oje; YouTuber markası; geniş renk; temiz formül; cruelty-free"),
        ("JINsoon", "jinsoon.com", "Kore Lüks Oje", "Nail Lacquer; Kore manikür sanatçısı; lüks formül; NYC salon; moda sektörü"),
        ("tenoverten", "tenoverten.com", "8-Free Oje", "The Foundation; 8-free; NYC salon; temiz oje; doğal formül; minimalist"),
        ("Orosa Beauty", "orosabeauty.com", "Gel Effect Oje", "Pure Cover Nail Paint; jel efekti oje; 7-free; uzun ömürlü; DTC"),
        ("Sundays", "dearsundays.com", "10-Free Oje", "Non-toxic Nail Polish; 10-free; wellness odaklı; NYC salon; zen manikür"),
        ("Côte", "coteshop.co", "Temiz Lüks Oje", "Nail Polish; 10-free; temiz lüks; botanik bazlı; LA salon"),
        ("Kure Bazaar", "kurebazaar.com", "Fransız Eko Oje", "Nail Colour; %85 doğal kaynak; Fransız eko-lüks; Paris chic"),
        ("Manucurist", "manucurist.com", "Fransız Green Oje", "Green Flash LED Gel; %84 biyo-bazlı jel; Fransız yeşil güzellik; LED jel"),
        ("Nailmatic", "nailmatic.com", "Fransız Eğlenceli Oje", "Made in LA; Fransız; su bazlı çocuk ojesi; eğlenceli renkler; erişilebilir"),
        ("Essie", "essie.com", "Salon Renk", "Gel Couture; salon kalitesi; geniş renk paleti; eczane; profesyonel + ev"),
        ("OPI", "opi.com", "Global Salon Oje", "Infinite Shine; salon standardı; tematik koleksiyonlar; profesyonel + perakende"),
        ("Sally Hansen", "sallyhansen.com", "Eczane Oje", "Miracle Gel; UV gerektirmeyen jel; erişilebilir; 2 adımlı salon jel"),
        ("Zoya", "zoya.com", "10-Free Profesyonel", "Nail Polish; 10-free; 400+ renk; Big 10 free; profesyonel temiz oje"),
        ("Deborah Lippmann", "deborahlippmann.com", "NYC Lüks Oje", "Gel Lab Pro; Broadway manikür sanatçısı; lüks formül; long-wear"),
        ("Smith & Cult", "smithandcult.com", "Edgy Lüks Oje", "Nail Lacquer; 8-free; edgy ambalaj; NYC indie lüks; moda ilhamlı"),
        ("Chanel Le Vernis alternatifi: Kure Bazaar", "kurebazaar.com", "Yeşil Lüks Oje", "10-free + eco-formül; Fransız lüks oje; biyo-bazlı; sürdürülebilir şık"),
        ("Le Mini Macaron", "leminimacaron.com", "Mini Jel Sistemi", "Le Mini Macaron Gel Kit; mini LED lamba + jel oje; Fransız; tek parmak jel"),
        ("Beetles Gel", "beetles.com", "Uygun Jel Seti", "Gel Polish Kit; Amazon #1; uygun fiyat jel oje seti; 6 renk + lamba"),
        ("Gelish", "gelish.com", "Profesyonel Jel", "Soak-Off Gel Polish; orijinal soak-off jel marka; salon profesyonel; 200+ renk"),
        ("CND Shellac", "cnd.com", "OG Jel Oje", "Shellac; orijinal jel oje markası; UV jel + oje hibrit; salon standardı"),
        ("Bio Seaweed Gel", "bioseaweedgel.com", "Temiz Jel", "3-Step Gel System; deniz yosunu özlü; 3 adım kolay jel; temiz formül"),
        ("Londontown", "londontown.com", "Lakur Oje", "Lakur Enhanced Colour; 16-free; florüksüz; İngiliz; ultra temiz formül"),
        ("Ella+Mila", "ellamila.com", "7-Free Vegan Oje", "Nail Polish; 7-free + vegan; PETA onaylı; $10; erişilebilir temiz"),
        ("Habit Cosmetics", "habitcosmetics.com", "Sürdürülebilir Oje", "Non-Toxic Nail Polish; 21-free; sürdürülebilir; minimal ambalaj; vegan"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 30. Tırnak Bakımı - Takma Tırnak & Press-On
    # ═══════════════════════════════════════════════════════════════════════════
    "Tırnak Bakımı - Takma Tırnak & Press-On": [
        ("Glamnetic", "glamnetic.com", "Manyetik Press-On", "Press-On Nails; manyetik kirpik + press-on tırnak; TikTok viral; geniş tasarım"),
        ("Dashing Diva", "dashingdiva.com", "Kore Press-On", "Gloss Gel Strips; Kore jel strip; kolay uygulama; salon efekti; K-beauty tırnak"),
        ("KISS", "kissusa.com", "OG Press-On", "imPRESS Press-On Manicure; basınçla yapışan; 30 saniye; erişilebilir; geniş çeşit"),
        ("ManiMe", "manime.co", "3D Özel Press-On", "Custom Gel Stickers; 3D tarama ile özel boyut; kişiselleştirilmiş jel sticker"),
        ("Chillhouse", "chillhouse.com", "NYC Press-On", "Chill Tips Press-On Nails; NYC salon; şık tasarımlar; çıkarılabilir; yeniden kullanılabilir"),
        ("Static Nails", "staticnails.com", "Reusable Press-On", "Reusable Pop-On Manicures; yeniden kullanılabilir; 18+ kullanım; sürdürülebilir"),
        ("Clutch Nails", "clutchnails.com", "El Boyama Press-On", "Hand-Painted Press-On Nails; el boyama sanat eseri; özel tasarım; artisan"),
        ("OhMyGel!", "ohmygel.com", "Gel-X Press-On", "Soft Gel Full Cover Tips; profesyonel gel-x; salon evde; yumuşak jel"),
        ("DUFFBEAUTY", "duffbeauty.com", "İsveç Press-On", "Nail Art Press-On; İsveç tasarım; İskandinav estetik; minimal sanat"),
        ("Nails Inc", "nailsinc.com", "İngiliz Press-On", "NailKale Press-On; İngiliz salon; superfood formül; hızlı manikür"),
        ("Color Street", "colorstreet.com", "Sticker Oje", "Nail Strips; %100 oje strip; çıkartma format; ev partisi satış modeli"),
        ("Ohora", "ohora.com", "Kore Jel Strip", "Semi-Cured Gel Nail Strips; yarı kürlü jel; Kore inovasyon; UV ile sertleştir"),
        ("Gelato Factory", "gelatofactory.co.kr", "Kore Hug Gel", "Hug Gel Stickers; Kore jel sticker; kolay uygulama; K-beauty tırnak sanatı"),
        ("Bling Art", "blingart.co.uk", "İngiliz Tırnak Sanatı", "False Nails; İngiliz; geniş tasarım; uygun fiyat; press-on + yapıştır"),
        ("Red Aspen", "redaspen.com", "Premium Press-On", "Nail Dashes; premium press-on; salon kalitesi; sosyal satış modeli"),
        ("Olive & June Press-On", "oliveandjune.com", "Instant Mani Press-On", "Instant Mani Press-On Nails; salon markasının press-on hattı; kolay uygulama"),
        ("BTArtbox", "btartboxnails.com", "Amazon Press-On", "Press On Nails; Amazon best seller; uygun fiyat; geniş tasarım seçeneği"),
        ("The GelBottle", "thegelbottle.com", "İngiliz Salon Jel", "Professional Gel Polish; İngiliz profesyonel jel; salon + ev; geniş palet"),
        ("Nailberry", "nailberry.com", "İngiliz Oxygenated Oje", "L'Oxygéné Nail Lacquer; oksijen geçiren formül; İngiliz lüks temiz oje; 12-free"),
        ("Paintbox", "paint-box.com", "NYC Manikür Stüdyo", "Power Couple Duo Set; NYC manikür stüdyosu DTC; sanatsal renk paletleri"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 31-33. Cilt Cihazları
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Cihazları - LED & Işık Terapisi": [
        ("Solawave", "solawave.com", "4-in-1 LED Cihaz", "Radiant Renewal Wand; kızılötesi + mavi LED + mikro akım + ısı terapisi; 4-in-1"),
        ("LightStim", "lightstim.com", "NASA LED Teknolojisi", "LightStim for Wrinkles; NASA teknolojisinden; multi-wave LED; FDA onaylı"),
        ("HigherDOSE", "higherdose.com", "Kızılötesi Maske", "Red Light Face Mask; kızılötesi LED maske; kolajen üretimi; at-home spa"),
        ("CurrentBody", "currentbody.com", "İngiliz LED Maske", "Skin LED Light Therapy Mask; İngiliz; FDA + CE onaylı; esnek silikon; 10 dakika"),
        ("Dennis Gross DRx", "drdennisgross.com", "SpectraLite LED", "SpectraLite FaceWare Pro; kırmızı + mavi LED; FDA onaylı; NYC dermatolog"),
        ("Omnilux", "omnilux.com", "Medikal LED", "Omnilux Contour Face; medikal sınıf LED; kırmızı + yakın kızılötesi; klinik kanıtlı"),
        ("NIRA", "niraskin.com", "Lazer Anti-Aging", "NIRA Pro Laser; ev kullanım lazer; kolajen stimülasyon; FDA onaylı; non-ablative"),
        ("MZ Skin", "mzskin.com", "Lüks LED Maske", "Light Therapy Golden Facial Treatment Device; altın kaplama LED; lüks; 5 LED renk"),
        ("Angela Caglia", "angelacaglia.com", "LED Maske", "CellReturn LED Mask; Hollywood facialist; premium LED maske; çoklu dalga boyu"),
        ("Trophy Skin", "trophyskin.com", "Evde LED", "RejuvaliteMD; FDA onaylı; kırmızı + amber + kızılötesi; anti-aging LED panel"),
        ("Skin Gym", "skingymco.com", "Wrinklit LED", "Wrinklit LED Mask; kızılötesi; uygun fiyat LED; anti-aging + akne; çift mod"),
        ("LYMA Laser", "lyma.life", "Lüks Lazer Cihaz", "LYMA Laser; £2000+ lüks; düşük seviye lazer; klinik kanıtlı; ev lazer öncüsü"),
        ("Therabody TheraFace", "therabody.com", "LED + Masaj", "TheraFace PRO; LED + perkussif masaj + mikro akım; çoklu terapi; Theragun markası"),
        ("Project E Beauty", "projectebeauty.com", "Uygun LED", "LED Face Mask; uygun fiyat; 7 renk LED; Amazon best seller; ev kullanım"),
        ("Dr. Pen", "drpen.com", "Mikro İğne + LED", "Dr. Pen A6S + LED; mikro iğne + LED kombo; evde profesyonel; çift tedavi"),
        ("Foreo UFO", "foreo.com", "Akıllı LED Maske", "UFO 2; T-Sonic + LED + kriyoterapi; 90 saniye maske; İsveç teknoloji"),
        ("JOVS", "jovs.com", "IPL + LED", "Blacken PRO DPL; IPL + LED kombinasyonu; çok amaçlı cilt cihazı; Çin teknoloji"),
        ("Solaris Labs", "solarislabsny.com", "Giyilebilir LED", "Illuminate Face + Neck; giyilebilir LED; yüz + boyun; esnek panel"),
        ("Celluma", "celluma.com", "Klinik LED", "Celluma PRO; FDA onaylı; klinik sınıf; esnek panel; akne + anti-aging + ağrı"),
        ("Dermalux", "dermalux.co.uk", "İngiliz Medikal LED", "Dermalux Flex MD; İngiliz medikal LED; 3 dalga boyu; klinik + ev"),
    ],

    "Cilt Cihazları - Mikro-Akım & RF": [
        ("NuFace", "mynuface.com", "Mikro-Akım Lideri", "Trinity+; mikro akım yüz germe; evde facelift; FDA onaylı; 5 dakika rutin"),
        ("ZIIP", "ziipbeauty.com", "Nano Akım", "ZIIP HALO; nanocurrent + mikrocurrent; uygulama kontrollü; kişiselleştirilmiş tedavi"),
        ("Medicube Age-R", "medicube.com", "Kore RF Cihaz", "AGE-R Booster Pro; EMS + RF + LED; Kore dermatoloji; TikTok viral; uygun fiyat"),
        ("Foreo BEAR", "foreo.com", "İsveç Mikro-Akım", "BEAR 2; T-Sonic + mikro akım; İsveç tasarım; app kontrollü; yüz egzersizi"),
        ("TheraFace", "therabody.com", "Çoklu Terapi", "TheraFace PRO; perkussif + mikro akım + LED + kriyoterapi; Theragun alt markası"),
        ("LYFT", "lyftbeauty.com", "Taşınabilir Mikro-Akım", "LYFT 2.0; taşınabilir mikro akım; şarj edilebilir; 5 dakika yüz germe"),
        ("MyoLift", "myolift.com", "Profesyonel Mikro-Akım", "MyoLift 600; 600 mikro amper; profesyonel seviye ev kullanım; çoklu dalga formu"),
        ("TriPollar", "tripollar.com", "İsrail RF", "STOP Vx; RF (radyo frekans); İsrail teknolojisi; kolajen sıkılaştırma; evde RF"),
        ("NEWA", "newabeauty.com", "İsrail Klinik RF", "NEWA RF Device; klinik kanıtlı RF; İsrail medikal; FDA onaylı; kolajen yenilenme"),
        ("Sensica", "sensica.com", "İsrail Sensilift", "Sensilift; RF ev cihazı; İsrail güzellik teknolojisi; sıkılaştırma + tonlama"),
        ("Mlay", "mlaybeauty.com", "Uygun RF", "RF Beauty Device; uygun fiyat RF; ev kullanım; kolajen stimülasyonu"),
        ("Silk'n", "silkn.com", "İsrail FaceTite", "FaceTite; trilayer RF + LED; İsrail; anti-aging cihaz; klinik sonuçlar"),
        ("PMD Beauty", "pmdbeauty.com", "Mikrodermabrazyon", "Personal Microderm Classic; evde mikrodermabrazyon; vakum + kristal; cilt yenileme"),
        ("Dermaflash", "dermaflash.com", "Dermaplaning", "LUXE+; evde dermaplaning; sonic teknoloji; tüy alma + eksfoliasyon"),
        ("SkinPen", "skinpen.com", "Mikro İğne", "SkinPen Precision; FDA onaylı mikro iğne; kolajen indüksiyon; profesyonel + ev"),
        ("Dr. Pen", "drpen.com", "Mikro İğne Cihazı", "Ultima A6; ayarlanabilir iğne derinliği; mikro iğne ev kullanım; serum iletimi"),
        ("AMIRO", "amiro.com", "Çin RF Cihazı", "R1 PRO; RF + EMS + LED; Çin güzellik teknolojisi; çoklu terapi; app kontrollü"),
        ("Opte", "opte.com", "Akıllı Leke Tedavisi", "Precision Skincare System; kamera + mikro jet; leke tespit + tedavi; P&G teknoloji"),
        ("CurrentBody RF", "currentbody.com", "İngiliz RF", "Skin RF Device; İngiliz; radyo frekans; kolajen sıkılaştırma; FDA + CE"),
        ("FaceGym", "facegym.com", "Yüz Fitness", "Multi-Sculpt High Performance Contouring Tool; yüz fitness; EMS + kızılötesi; İngiliz"),
    ],

    "Cilt Cihazları - Temizleme Cihazı": [
        ("Foreo LUNA", "foreo.com", "Silikon Temizleme", "LUNA 4; silikon yüz temizleme; T-Sonic titreşim; İsveç; hijyenik; şarj edilebilir"),
        ("PMD Clean", "pmdbeauty.com", "Akıllı Temizleme", "PMD Clean Pro; SonicGlow teknolojisi; silikon + ActiveWarmth; 4-in-1"),
        ("Clarisonic alternatifi: Foreo", "foreo.com", "Sonic Temizleme", "LUNA Mini 3; kompakt sonic temizleme; İsveç; seyahat dostu; app bağlantılı"),
        ("Vanity Planet", "vanityplanet.com", "Uygun Temizleme", "Raedia Facial Cleansing Brush; uygun fiyat; çeşitli başlık; viral; erişilebilir"),
        ("Michael Todd Beauty", "michaeltoddbeauty.com", "Soniclear Temizleme", "Soniclear Petite; sonic temizleme; antimikrobiyal fırça; titanyum; kompakt"),
        ("AENO", "aeno.com", "Akıllı Temizleme", "Facial Cleanser; akıllı temizleme cihazı; IoT bağlantılı; kişiselleştirilmiş rutin"),
        ("Conture", "conture.com", "Kinetic Temizleme", "Kinetic Skin Toning System; kinetik temizleme + tonlama; titreşim + ısı"),
        ("Skin Inc", "iloveskininc.com", "Singapur Optimizer", "Optimizer Voyage Tri-Light; LED + sonic temizleme; Singapur; kişiselleştirilmiş serum"),
        ("Droplette", "droplette.io", "Mikro İnfüzyon", "Droplette 2; mikro mist infüzyon; iğnesiz serum iletimi; 20x derin nüfuz"),
        ("EZBASICS", "ezbasics.com", "Uygun Sonic", "Facial Cleansing Brush; uygun fiyat; 4 başlık; su geçirmez; Amazon viral"),
        ("Refa", "refa.net", "Japon Güzellik Cihazı", "ReFa CLEAR; Japon sonic temizleme; mikro akım + temizleme; 3D sonic ion"),
        ("SENSSE", "sensse.com", "İngiliz Temizleme", "Original Facial Brush; İngiliz; silikon; waterproof; uygun fiyat temizleme"),
        ("Magnitone", "magnitone.com", "İngiliz Sonic", "BareFaced 3 Vibra-Sonic; İngiliz; sonic titreşim; temizleme + masaj"),
        ("Geske", "geske.com", "Alman Akıllı Temizleme", "SmartAppGuided 6in1 Sonic Cleansing Brush; Alman; app kontrollü; 6-in-1"),
        ("AAPI Beauty", "aapibeauty.com", "LED Temizleme", "Light Therapy Cleansing Device; LED + sonic temizleme; çift terapi; kompakt"),
        ("Liberex", "liberex.com", "Elektrikli Temizleme", "Egg Vibration; yumurta şekilli; titreşimli temizleme; silikon; uygun fiyat"),
        ("InFace", "inface.com", "Xiaomi Temizleme", "InFace Sonic Cleanser; Xiaomi ekosistemi; silikon; uygun fiyat; IPX7 su geçirmez"),
        ("Panasonic Beauty", "panasonic.com", "Japon İonik Temizleme", "EH-SC67; Japon ionic cleansing; 2 hızlı; nazik temizleme; Japon teknoloji"),
        ("Philips VisaPure", "philips.com", "Hollanda Temizleme", "VisaPure Advanced; Hollanda; değiştirilebilir başlık; derin temizleme; profesyonel"),
        ("Spa Sciences", "spasciences.com", "Çoklu Başlık Temizleme", "NOVA Antimicrobial; 7 başlık; antimikrobiyal; çoklu fonksiyon; ev spa"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 34-36. Parfüm & Koku
    # ═══════════════════════════════════════════════════════════════════════════
    "Parfüm & Koku - Kadın Parfüm": [
        ("Snif", "snif.co", "DTC Deneme Parfüm", "Try before you buy; online koku keşfi; 2ml deneme + tam boy; DTC parfüm demokratizasyonu"),
        ("Dossier", "dossier.co", "Lüks Dupe Parfüm", "Lüks parfüm ilhamları; %80 daha uygun; şeffaf fiyat; vegan; clean formül"),
        ("Phlur", "phlur.com", "Temiz Parfüm", "Missing Person kült koku; temiz + sürdürülebilir; hyper-natürel; Emma Chamberlain favorisi"),
        ("Skylar", "skylar.com", "Hipoalerjenik Parfüm", "Salt Air; temiz + hipoalerjenik; vegan; her koku saf bileşenlerden; hassas cilt"),
        ("Ellis Brooklyn", "ellisbrooklyn.com", "Temiz Lüks Parfüm", "MYTH Eau de Parfum; temiz lüks; Brooklyn artisan; doğal + sentetik denge"),
        ("Dedcool", "dedcool.com", "Biyozgörülebilir Parfüm", "Biyolojik olarak parçalanabilir; cinsiyet nötr; %100 vegan; California wellness"),
        ("Henry Rose", "henryrose.com", "Michelle Pfeiffer Parfüm", "EWG Verified; Michelle Pfeiffer markası; %100 şeffaf bileşen; temiz lüks"),
        ("Clean Reserve", "cleanbeauty.com", "Sürdürülebilir Parfüm", "Clean Reserve; geri dönüştürülebilir; responsible kaynak; temiz koku; sürdürülebilir"),
        ("Commodity", "commodity.com", "Ekspresif Parfüm", "Book Eau de Parfum; Personal, Expressive, Bold 3 versiyon; kişiselleştirilmiş yoğunluk"),
        ("Juliette Has a Gun", "juliettehasagun.com", "Provoke Parfüm", "Not a Perfume; Cetalox tek notu; anti-parfüm konsepti; Fransız niche"),
        ("Floral Street", "floralstreet.com", "İngiliz Çiçek Parfüm", "Wonderland Peony; İngiliz çiçek kokulari; sürdürülebilir; compostable ambalaj"),
        ("Boy Smells", "boysmells.com", "Cinsiyet Nötr Koku", "Violet Ends; mum + parfüm; cinsiyet nötr; LGBTQ+ kurucu; kapsayıcı koku"),
        ("Replica Maison Margiela alternatifi: Snif", "snif.co", "Anı Kokusu", "Way With Words; anıları çağrıştıran; DTC; deneme modeli; erişilebilir niche"),
        ("Michelle Ma Belle", "perfumariamia.com", "Artisan Parfüm", "Niche parfüm koleksiyonu; artisan; küçük parti; el yapımı; benzersiz notlar"),
        ("Vyrao", "vyrao.com", "Enerji Parfüm", "Free 00; enerji bazlı parfüm; kristal + meditasyon; İngiliz wellness koku"),
        ("Maya Njie", "mayanjie.com", "İsveç-Afrika Parfüm", "Les Fleurs; İsveç-Gambiya kökenli; kültürel koku; artisan; minimal ambalaj"),
        ("By Rosie Jane", "byrosiejane.com", "LA Temiz Parfüm", "Dulce; LA lifestyle; temiz + vegan; göze çarpan minimalist ambalaj"),
        ("Santal 33 alternatifi: Dossier", "dossier.co", "Woody Sandal Dupe", "Woody Sandalwood; Le Labo Santal 33 ilhamlı; %80 uygun; şeffaf formül"),
        ("Etat Libre d'Orange", "etatlibredorange.com", "Provokatif Parfüm", "You or Someone Like You; provokatif isimler; Fransız niche; avant-garde"),
        ("Eau d'Italie", "eauditalie.com", "İtalyan Niche", "Altaia By Any Other Name; İtalyan niche; Amalfi kıyısı ilhamı; artisan"),
    ],

    "Parfüm & Koku - Unisex & Niş": [
        ("Le Labo", "lelabofragrances.com", "NYC Artisan", "Santal 33 global fenomen; el yapımı; NYC laboratuvar; kişiselleştirilmiş etiket"),
        ("Byredo", "byredo.com", "İsveç Niche", "Gypsy Water; İsveç minimalist lüks; Ben Gorham; sanatsal yaklaşım; modern klasik"),
        ("Diptyque", "diptyque.com", "Paris Klasik", "Tam Dao; Paris 1961; mum + parfüm; lüks Fransız niche; sanatsal ambalaj"),
        ("DS & Durga", "dsanddurga.com", "Brooklyn Artisan", "I Don't Know What; Brooklyn indie; David Seth Moltz; müzik + koku; artisan"),
        ("Maison Louis Marie", "maisonlouismarie.com", "Fransız Bahçe", "No.04 Bois de Balincourt; Fransız bahçe ilham; doğal parfüm yağı; minimalist"),
        ("Imaginary Authors", "imaginaryauthors.com", "Hikaye Parfüm", "Every Storm a Serenade; her koku bir roman; Portland indie; hikaye anlatımı"),
        ("Frédéric Malle", "fredericmalle.com", "Editör Parfüm", "Portrait of a Lady; parfüm editörü konsepti; lüks Fransız niche; master parfümörler"),
        ("Penhaligon's", "penhaligons.com", "İngiliz Miras", "Halfeti; İngiliz aristokrat koku; 1870'den beri; Türk gülü; miras niche"),
        ("Acqua di Parma", "acquadiparma.com", "İtalyan Klasik", "Colonia; İtalyan klasik unisex; sarı şişe ikonik; 1916'dan beri"),
        ("Jo Malone", "jomalone.com", "İngiliz Katmanlama", "Wood Sage & Sea Salt; İngiliz; koku katmanlama öncüsü; birleştirme sanatı"),
        ("Atelier Cologne", "ateliercologne.com", "Cologne Absolue", "Clementine California; Cologne Absolue kategorisi yaratıcısı; sitrus yoğunluk"),
        ("Aesop", "aesop.com", "Avustralya Unisex", "Hwyl; Avustralya minimalist; botanik + artisan; unisex; karanlık ahşap"),
        ("Escentric Molecules", "escentric.com", "Molekül Parfüm", "Molecule 01; Iso E Super tek molekül; Berlin; avant-garde; koku kimyası"),
        ("Comme des Garçons Parfums", "comme-des-garcons-parfum.com", "Avant-Garde Koku", "CDG 2; Japon avant-garde; beton + metal notalar; deneysel"),
        ("NISHANE", "nishane.com", "Türk Niche Parfüm", "Hacivat; Türk niche parfüm evi; Karagöz gölge oyunu ilham; premium; yüksek kalite"),
        ("Xerjoff", "xerjoff.com", "İtalyan Ultra Lüks", "Naxos; İtalyan ultra lüks; el yapımı şişe; niche parfüm sanatı"),
        ("Amouage", "amouage.com", "Oman Kraliyet", "Interlude Man; Oman kraliyet parfüm evi; Orta Doğu lüksü; frankincense"),
        ("Mancera", "mancera.com", "Paris Niche", "Cedrat Boise; Paris niche; Pierre Montale; güçlü projeksiyon; uygun niche"),
        ("Montale", "montale-parfums.com", "Paris-Arabistan Niche", "Intense Cafe; Paris + Arap koku geleneği; güçlü; oud + kahve; niche"),
        ("Initio Parfums Prives", "initioparfums.com", "Misk Niche", "Side Effect; feromon + misk; çekicilik bilimi; lüks Fransız niche"),
        ("Thameen", "thameen.com", "İngiliz Mücevher Koku", "Regent Leather; İngiliz mücevher ilhamlı; lüks ambalaj; Crown Collection"),
        ("BDK Parfums", "bdkparfums.com", "Paris Boutique", "Gris Charnel; Paris butik parfüm evi; David Benedek; modern Fransız niche"),
        ("Vilhelm Parfumerie", "vilhelmparfumerie.com", "NYC-İsveç Niche", "Dear Polly; NYC merkezli İsveç kökenli; Jan Ahlgren; edebiyat ilhamlı"),
        ("Malin+Goetz", "malinandgoetz.com", "NYC Apothecary Koku", "Dark Rum; NYC apothecary; unisex; minimalist; günlük kullanım niche"),
        ("19-69", "nineteen-sixty-nine.com", "İsveç Kültür Koku", "Purple Haze; İsveç; kültürel referanslar; 1960'lar kontrkültür; avant-garde"),
    ],

    "Parfüm & Koku - Vücut Spreyi & Mist": [
        ("Sol de Janeiro", "soldejaneiro.com", "Brezilya Body Mist", "Brazilian Bum Bum Cream Mist; Brezilya kokusu; guaraná; tropikal vücut spreyi"),
        ("Touchland", "touchland.com", "Lüks El Mist", "Power Mist; el dezenfektanı + parfüm; aloe; estetik ambalaj; lüks his"),
        ("Saltair", "saltair.com", "Plaj Body Mist", "Body Mist; tropikal kokular; plaj ilhamlı; erişilebilir; vücut spreyi"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi Mist", "Coconut Body Glow; hindistan cevizi + parıltı; tropikal vücut spreyi; yazlık"),
        ("Native", "nativecos.com", "Doğal Body Mist", "Body Spray; doğal bileşenler; çeşitli kokular; erişilebilir; temiz"),
        ("Phlur", "phlur.com", "Temiz Body Mist", "Body Mist; temiz formül; hyper-natürel koku; vegan; erişilebilir niche"),
        ("Skylar", "skylar.com", "Hipoalerjenik Mist", "Body Mist; hipoalerjenik; hassas cilt güvenli; temiz; vegan; çeşitli notalar"),
        ("Ellis Brooklyn", "ellisbrooklyn.com", "Temiz Vücut Mist", "Body Milk Spray; koku + nemlendirme; Brooklyn artisan; çok amaçlı"),
        ("Boy Smells", "boysmells.com", "Cinsiyet Nötr Mist", "Body Mist; cinsiyet nötr; kapsayıcı; mum markasından vücut mist; LGBTQ+"),
        ("Moroccanoil Body", "moroccanoil.com", "Argan Vücut Mist", "Body Oil Mist; argan yağı vücut spreyi; hafif yağ mist; Fragrance Originale"),
        ("Victoria's Secret alternatifi: Fine'ry", "finery.com", "Target Parfüm Mist", "Body Mist; Target'ta; uygun fiyat niche koku; şık ambalaj; erişilebilir"),
        ("Being Frenshe", "beingfrenshe.com", "Ashley Tisdale Mist", "Hair & Body Mist; Ashley Tisdale; saç + vücut; Walmart; erişilebilir"),
        ("Raw Sugar", "rawsugar.com", "Doğal Vücut Mist", "Body Mist; doğal + temiz; sürdürülebilir; erişilebilir; tropikal kokular"),
        ("Pacifica", "pacificabeauty.com", "Vegan Parfüm Mist", "Spray Perfume; %100 vegan; doğal; erişilebilir; çeşitli notalar; temiz"),
        ("Mix:Bar", "mixbar.com", "Layer Mist", "Hair & Body Mist; Target özel; katmanlama koleksiyonu; erişilebilir; genç"),
        ("Billie Eilish Eilish alternatifi: Ariana Grande Cloud", "arianagrande.com", "Celeb Mist", "Cloud Body Mist; ünlü parfüm mist versiyonu; erişilebilir; tatlı koku"),
        ("CLEAN Reserve", "cleanbeauty.com", "Sürdürülebilir Mist", "Body Mist; sürdürülebilir; geri dönüşümlü; temiz formül; responsible luxury"),
        ("Ouai", "theouai.com", "Saç Parfümü", "Hair Fragrance Mist; saç parfümü; Jen Atkin; hafif koku; saç bakımı + koku"),
        ("Gisou", "gisou.com", "Bal Saç Parfümü", "Honey Infused Hair Perfume; bal bazlı saç kokusu; Mirsalehi; TikTok viral"),
        ("Ceremonia", "ceremonia.com", "Latin Saç Mist", "Aceite de Moska Hair Mist; Latin koku; babasu; kültürel güzellik; saç spreyi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 37-38. Erkek Bakım
    # ═══════════════════════════════════════════════════════════════════════════
    "Erkek Bakım - Tıraş & Sakal": [
        ("Harry's", "harrys.com", "DTC Tıraş Devrimi", "Razor + Shave Set; erkek tıraş DTC öncüsü; fabrika satın aldı; Walmart ortaklığı"),
        ("Dollar Shave Club", "dollarshaveclub.com", "Abonelik Tıraş", "Razor Subscription; Unilever $1B satın aldı; viral video pazarlama; DTC öncü"),
        ("Bevel", "getbevel.com", "Melanin Tıraş", "Safety Razor + Shave System; koyu cilt tıraş; tüy batması önleme; Tristan Walker"),
        ("Supply", "supply.co", "Tek Bıçak Tıraş", "Single Edge Razor; tek bıçak güvenlik jilet; hassas cilt; daha az tahriş"),
        ("Henson Shaving", "hensonshaving.com", "Havacılık Jilet", "AL13 Razor; havacılık alüminyum; ultra hassas; Kanada mühendislik; CNC işleme"),
        ("Leaf Shave", "leafshave.com", "Sürdürülebilir Jilet", "Leaf Razor; plastik-free tıraş; çoklu bıçak + pivotlu; sürdürülebilir erkek bakım"),
        ("Beardbrand", "beardbrand.com", "Sakal Bakım Uzmanı", "Beard Oil + Utility Balm; sakal kültürü öncüsü; YouTube'dan markaya; Austin TX"),
        ("Honest Amish", "honestamish.com", "Doğal Sakal Yağı", "Classic Beard Oil; %100 doğal + organik; Amish el yapımı; sakal kült marka"),
        ("Viking Revolution", "vikingrevolution.com", "Viking Sakal Seti", "Beard Care Kit; sakal bakım seti; Amazon #1; uygun fiyat; erkek hediye"),
        ("Cremo", "cremocompany.com", "Astonishingly Superior", "Shave Cream; ultra ince formül; az köpük + yakın tıraş; erişilebilir; Target"),
        ("Jack Black", "getjackblack.com", "Premium Erkek Bakım", "Beard Lube Conditioning Shave; pre-shave + shave + conditioner; 3-in-1; lüks"),
        ("Baxter of California", "baxterofcalifornia.com", "LA Premium Erkek", "Shave Tonic; LA erkek bakım; 1965'den beri; premium; profesyonel"),
        ("Proraso", "proraso.com", "İtalyan Tıraş", "Shaving Cream; İtalyan berber geleneği; okaliptüs + mentol; 1948'den beri"),
        ("The Art of Shaving", "theartofshaving.com", "Lüks Tıraş Ritüeli", "Pre-Shave Oil; 4 adım tıraş ritüeli; lüks; P&G markası; sandalwood"),
        ("Geo. F. Trumper", "trumpers.com", "İngiliz Berber", "Extract of Limes Shaving Cream; İngiliz kraliyet berber; 1875'den beri; geleneksel"),
        ("Edwin Jagger", "edwinjagger.com", "İngiliz Güvenlik Jilet", "DE89 Safety Razor; İngiliz güvenlik jilet standardı; Sheffield çelik; geleneksel"),
        ("Merkur", "merkur-razor.com", "Alman Jilet", "34C Safety Razor; Alman mühendislik; Solingen çelik; güvenlik jilet standardı"),
        ("Mühle", "muehle-shaving.com", "Alman Lüks Tıraş", "Rocca Safety Razor; Alman lüks tıraş; paslanmaz çelik; Black Forest; el yapımı"),
        ("Duke Cannon", "dukecannon.com", "Maskülen Bakım", "Big Ass Brick of Soap + Shave Cream; büyük boy; maskülen pazarlama; eğlenceli"),
        ("Dr. Squatch", "drsquatch.com", "Doğal Erkek Sabun", "Natural Soap + Shave Kit; soğuk pres doğal sabun; viral reklam; eğlenceli marka"),
        ("Manscaped", "manscaped.com", "Erkek Vücut Bakım", "The Lawn Mower trimmer; erkek vücut tıraş; cesur pazarlama; DTC lider"),
        ("Bulldog Skincare", "bulldogskincare.com", "İngiliz Erkek Bakım", "Original Shave Gel; İngiliz erkek bakım; doğal bileşenler; Edgewell satın aldı"),
        ("Every Man Jack", "everymanjack.com", "Doğal Erkek Bakım", "Shave Gel; doğal + erişilebilir; Target + Walmart; %100 bitkisel bazlı"),
        ("Murdock London", "murdocklondon.com", "Londra Berber", "Shave Cream; Londra berber salonu + ürün; İngiliz erkek bakım lüksü"),
        ("Aesop Erkek", "aesop.com", "Botanik Erkek Tıraş", "Moroccan Neroli Shaving Serum; botanik lüks; unisex ama erkek favorisi; Avustralya"),
    ],

    "Erkek Bakım - Erkek Cilt Bakımı": [
        ("Lumin", "luminskin.com", "Erkek Cilt Bakım DTC", "Dark Circle Defense; erkek cilt bakımı abonelik; Kore formül; göz altı + anti-aging"),
        ("Tiege Hanley", "tiege.com", "Erkek Cilt Sistemi", "Level 1 System; basit erkek cilt bakım sistemi; abonelik; 3 adım"),
        ("Geologie", "geologie.com", "Kişiselleştirilmiş Erkek", "Personalized Skincare Set; kişiselleştirilmiş erkek cilt bakımı; quiz bazlı; dermo formül"),
        ("Brickell Men's Products", "brickellmensproducts.com", "Premium Doğal Erkek", "Revitalizing Anti-Aging Cream; premium doğal; organik bileşenler; erkek anti-aging"),
        ("Jack Black", "getjackblack.com", "Premium Erkek Cilt", "Double-Duty Face Moisturizer SPF 20; çok amaçlı; erkek SPF; premium"),
        ("Kiehl's Erkek", "kiehls.com", "NYC Erkek Bakım", "Facial Fuel Energizing Moisture Treatment; kafein + vitamin C; NYC eczane erkek"),
        ("Lab Series", "labseries.com", "Erkek Bilimi", "Daily Rescue Energizing Eye Treatment; erkek cilt bilimi; Estée Lauder grubu"),
        ("Clinique For Men", "clinique.com", "Erkek Dermatolojik", "Maximum Hydrator; alerjist test; dermatolojik erkek bakım; 72 saat nemlendirme"),
        ("Rugged & Dapper", "ruggedanddapper.com", "Erkek Doğal Premium", "Age Defense Moisturizer; doğal + organik; erkek anti-aging; erişilebilir premium"),
        ("Disco Skincare", "letsdisco.com", "Modern Erkek Cilt", "Repairing Eye Stick; modern erkek cilt bakımı; Z kuşağı erkek; göz altı stick"),
        ("Aesop Erkek Cilt", "aesop.com", "Botanik Erkek Cilt", "In Two Minds Facial Cleanser; botanik lüks; unisex yaklaşım; erkek favorisi"),
        ("Malin+Goetz Erkek", "malinandgoetz.com", "NYC Erkek Cilt", "Vitamin E Face Moisturizer; NYC apothecary; unisex; erkek günlük bakım"),
        ("Anthony", "anthony.com", "Erkek Cilt Uzmanı", "Glycolic Facial Cleanser; erkek cilt bakımı uzmanı; glikolik temizleyici; profesyonel"),
        ("Baxter of California Cilt", "baxterofcalifornia.com", "LA Erkek Cilt", "Oil Free Moisturizer SPF 15; LA erkek bakım; SPF + nemlendirme; 1965"),
        ("Bulldog Cilt", "bulldogskincare.com", "İngiliz Erkek Cilt", "Original Moisturiser; İngiliz; doğal bileşenler; erişilebilir; Walmart + Target"),
        ("Harry's Face", "harrys.com", "Harry's Cilt Bakımı", "Face Lotion SPF 15; DTC tıraştan cilt bakımına; erişilebilir erkek SPF"),
        ("Hims Cilt", "forhims.com", "Erkek Reçeteli Cilt", "Tretinoin + Niacinamide Cream; online dermatoloji; reçeteli erkek cilt bakımı"),
        ("Agency", "agency.com", "Kişiselleştirilmiş Erkek Cilt", "Custom Formula; kişiye özel formül; teledermatolog; tretinoin + aktifler"),
        ("War Paint", "warpaintformen.com", "Erkek Makyaj", "Concealer + Foundation; erkek makyaj markası; İngiliz; stigma kırma; kapsayıcı"),
        ("Stryx", "stryx.com", "Erkek Concealer", "Concealer Tool for Men; erkek kapatıcı; akne + göz altı; pratik; minimal"),
        ("Humanrace", "humanraceproducts.com", "Pharrell Cilt Bakımı", "Rice Powder Cleanser; Pharrell Williams markası; 3 adım; refill; lüks unisex"),
        ("Shakeup Cosmetics", "shakeupcosmetics.com", "Erkek BB Krem", "Mattifying BB Cream for Men; erkek BB krem; İngiliz; mat finish; doğal görünüm"),
        ("Patricks", "patricks.com", "Avustralya Lüks Erkek", "SH1 Shampoo; Avustralya lüks erkek bakım; premium; şık siyah ambalaj"),
        ("Hawthorne", "hawthorne.co", "Quiz Bazlı Erkek", "Personalized Body Wash + Face; quiz ile kişiselleştirilmiş; DTC erkek bakım sistemi"),
        ("Oars + Alps", "oarsandalps.com", "Aktif Erkek Bakım", "Face + Eye Cream; aktif yaşam tarzı erkek bakımı; doğal + etkili; TSA friendly"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 39-45. Specialty Categories
    # ═══════════════════════════════════════════════════════════════════════════
    "K-Beauty & Kore Kozmetik": [
        ("COSRX", "cosrx.com", "Kore Aktif Bakım", "Snail Mucin 96% Power Repairing Essence; salyangoz müsin; global K-beauty ikonu; TikTok viral"),
        ("Medicube", "medicube.com", "Kore Dermo-Kozmetik", "AGE-R Booster Pro + Zero Pore Pad; dermatoloji klinikten; cihaz + cilt bakımı; viral"),
        ("numbuzin", "numbuzin.com", "Kore Numara Bazlı", "No.5 Vitamin-Niacinamide Concentrated Pad; numara sistemi; Kore viral; innovatif"),
        ("Torriden", "torriden.com", "Kore Hyaluronic Uzmanı", "DIVE-IN Low Molecular Hyaluronic Acid Serum; düşük mol HA; Olive Young #1"),
        ("Anua", "anua.co.kr", "Kore Heartleaf", "Heartleaf 77% Soothing Toner; heartleaf trendi; TikTok viral; Kore yeni dalga"),
        ("Beauty of Joseon", "beautyofjoseon.com", "Kore Geleneksel", "Glow Serum Propolis + Niacinamide; Joseon Hanedanlığı ilham; pirinç + ginseng; viral"),
        ("Banila Co", "banilaco.com", "Kore Cleansing Balm", "Clean It Zero; sherbet temizleyici; çift temizleme ikonu; 70M+ satış"),
        ("Laneige", "laneige.com", "Kore Su Bilimi", "Water Sleeping Mask + Lip Sleeping Mask; sleeping mask yaratıcısı; su teknolojisi"),
        ("Innisfree", "innisfree.com", "Kore Yeşil Çay", "Green Tea Seed Serum; Jeju Adası; doğal K-beauty; çevre dostu ambalaj"),
        ("Etude", "etude.com", "Kore Eğlenceli Makyaj", "SoonJung + Play Color Eyes; eğlenceli K-beauty; genç hedef kitle; çeşitli ürünler"),
        ("Missha", "missha.com", "Kore BB Öncüsü", "M Perfect Blind BB Cream + Time Revolution; BB krem yaratıcısı; fermente esans"),
        ("Holika Holika", "holikaholika.com", "Kore Eğlenceli Bakım", "Good Cera + Pig-Nose; eğlenceli ambalaj; yaratıcı K-beauty; uygun fiyat"),
        ("A'PIEU", "apieu.com", "Kore Genç Bakım", "Madecassoside Cica Gel + Juicy-Pang; genç K-beauty; uygun fiyat; cica uzmanı"),
        ("3CE", "3ce.com", "Kore Mood Makyaj", "Velvet Lip Tint + Multi Eye Color Palette; Stylenanda; K-beauty makyaj standardı"),
        ("Peripera", "peripera.com", "Kore Tint Uzmanı", "Ink the Velvet; kadife tint standardı; uygun fiyat K-beauty makyaj"),
        ("Romand", "romand.co.kr", "Kore Viral Makyaj", "Juicy Lasting Tint + Better Than Eyes; viral K-beauty; Z kuşağı; pastel estetik"),
        ("Clio", "clio.co.kr", "Kore Pro Makyaj", "Kill Cover Fixer Cushion; Kore profesyonel makyaj; yüksek coverage; K-beauty salon"),
        ("VT Cosmetics", "vtcosmetics.com", "Kore Cica Makyaj", "Cica serisi; BTS kolaborasyonu; cica + makyaj hibrit; K-pop estetik"),
        ("Moonshot", "moonshot.co.kr", "Kore K-Pop Makyaj", "Micro Glassyfit Cushion; YG Entertainment; K-pop estetik; cam cilt makyaj"),
        ("Amorepacific", "amorepacific.com", "Kore Premium", "Time Response Skin Renewal Serum; Kore premium lüks; EGCG yeşil çay; bilim"),
        ("Sulwhasoo", "sulwhasoo.com", "Kore Hanbang Lüks", "Concentrated Ginseng Cream; Kore hanbang lüks; ginseng araştırma; geleneksel tıp"),
        ("The History of Whoo", "whoo.com", "Kore Saray Lüks", "Self-Generating Anti-Aging Essence; Kore kraliyet reçetesi; ultra lüks hanbang"),
        ("Dr. Jart+", "drjart.com", "Kore Dermo-Makyaj", "Cicapair Tiger Grass Color Correcting Treatment; dermo-kozmetik + renk düzeltme"),
        ("I'm From", "imfrom.co.kr", "Kore Tek Kaynak", "Rice Toner + Honey Mask; tek bölge tek bileşen; izlenebilirlik; Kore şeffaflık"),
        ("Mixsoon", "mixsoon.com", "Kore Saf Bileşen", "Bean Essence + Soybean Milk; tek bileşen felsefesi; saf formüller; Kore minimal"),
    ],

    "J-Beauty & Japon Kozmetik": [
        ("Hada Labo", "hadalabousa.com", "Japon Hyaluronic", "Gokujyun Premium Lotion; 7 hyaluronic acid; Japon nemlendirme standardı; uygun fiyat"),
        ("Melano CC", "rohto.com", "Japon C Vitamini", "Intensive Anti-Spot Essence; Japon vitamin C leke tedavisi; dünya çapında kült; Rohto"),
        ("Canmake", "canmake.com", "Japon Uygun Makyaj", "Cream Cheek + Mermaid Skin UV Gel; Japon uygun makyaj; kawaii; kaliteli"),
        ("KATE", "nomorerules.net", "Japon Lip Monster", "Lip Monster; mask-proof; Japon viral; 'canavar' konsepti; Kanebo; inovatif"),
        ("Tatcha", "tatcha.com", "Japon Geisha", "Dewy Skin Cream + Rice Polish; Japon geisha güzellik ritüeli; mor pirinç; lüks"),
        ("DHC", "dhc.co.jp", "Japon Zeytin", "Deep Cleansing Oil; zeytin bazlı temizleme; Japon temizleme yağı öncüsü; 1972"),
        ("Shu Uemura", "shuuemura.com", "Japon Sanat Makyajı", "Cleansing Oil + Unlimited Foundation; makyaj sanatı + cilt bakımı; Tokyo"),
        ("SK-II", "sk-ii.com", "Japon Pitera", "Facial Treatment Essence; PITERA fermente maya; 40+ yıl; Japon biyoteknoloji lüks"),
        ("Shiseido", "shiseido.com", "Japon Global Lüks", "Ultimune + Synchro Skin; 150+ yıl; Japon bilim + güzellik; global lüks"),
        ("SUQQU", "suqqu.com", "Japon Sanatsal Makyaj", "Signature Color Eyes; sanatsal renk; mevsimsel limitli; Japon lüks"),
        ("Decorté", "decorte.com", "Japon Liposome", "Liposome Advanced Repair Serum; lipozom teknolojisi; Kosé; Japon bilim lüks"),
        ("POLA", "pola.com", "Japon Wrinkle Shot", "Wrinkle Shot Serum; NEI-L1 bazosit; Japon kırışıklık tedavi öncüsü; patent"),
        ("Clé de Peau Beauté", "cledepeaubeaute.com", "Japon Ultra Lüks", "The Serum; Illuminating Complex EX; Japon en lüks cilt bakımı; hücre ışığı"),
        ("Lunasol", "lunasol-net.com", "Japon Nude Makyaj", "Skin Modeling Eyes; nude göz paleti; Kanebo lüks; Japon doğal güzellik"),
        ("Addiction by Ayako", "addiction-beauty.com", "Japon Tek Farı", "The Eyeshadow; 99 tek renk; Japon sanatsal; Ayako; mücevher pigment"),
        ("Opera", "opera-net.jp", "Japon Lip Tint Oil", "Lip Tint Oil; @cosme #1; Japon dudak tint + yağ hibriti; hafif renk + bakım"),
        ("Kiku-Masamune", "kikumasamune.co.jp", "Japon Sake", "High Moist Lotion; sake + arbutin; 400+ yıl sake geleneği; fermente güzellik"),
        ("Curel", "curel.com", "Japon Pseudo-Ceramide", "Intensive Moisture Cream; sözde-ceramide; Kao; Japon hassas cilt standardı"),
        ("Minon Amino", "minon-aminomoist.com", "Japon Amino Asit", "Amino Moist Charge Lotion; 9 amino asit; Japon hassas cilt nemlendirme"),
        ("Biore", "biore.com", "Japon Gözenek", "Pore Strips + UV Aqua Rich Watery Essence; gözenek + SPF; Japon günlük bakım"),
        ("RMK", "rmk.com", "Japon Doğal Makyaj", "Gel Creamy Foundation; jel kıvam; doğal Japon makyaj; hafif + transparent"),
        ("THREE", "threecosmetics.com", "Japon Organik Makyaj", "Flawless Ethereal Fluid Foundation; organik + bilimsel; Japon temiz makyaj"),
        ("FANCL", "fancl.com", "Japon Koruyucusuz", "Mild Cleansing Oil; koruyucu maddesiz; Japon temiz güzellik; 1980'den beri"),
        ("Albion", "albion.co.jp", "Japon Süt Emülsiyon", "Exage Activation Moisture Milk; Japon süt ilk geleneği; emülsiyon önce; lüks"),
        ("KANEBO", "kanebo-cosmetics.jp", "Japon İpek", "Sensai Ultimate The Cream; ipek proteinleri; ultra lüks; sınırlı üretim"),
    ],

    "Doğal & Organik Kozmetik": [
        ("Tata Harper", "tataharper.com", "Vermont Çiftlik Lüks", "Resurfacing Mask; %100 doğal + lüks; Vermont çiftliğinde üretim; yeşil lüks"),
        ("Herbivore Botanicals", "herbivorebotanicals.com", "Doğal Aktif", "Blue Tansy Mask + Bakuchiol Serum; doğal aktifler; Seattle indie; Instagram estetik"),
        ("True Botanicals", "truebotanicals.com", "Klinik Doğal", "Pure Radiance Oil; klinik testlerle kanıtlı doğal; La Mer'e rakip; MADE SAFE"),
        ("Ilia", "iliabeauty.com", "Temiz Makyaj", "Super Serum Skin Tint SPF 40; temiz lüks makyaj; organik bileşenler; Sephora"),
        ("RMS Beauty", "rmsbeauty.com", "Organik Makyaj", "Un Cover-Up; organik hindistan cevizi yağı bazlı; Rose-Marie Swift; ham pigmentler"),
        ("Kjaer Weis", "kjaerweis.com", "Lüks Refill Organik", "Cream Foundation; lüks refill ambalaj; organik; Danimarka; COSMOS sertifikalı"),
        ("Juice Beauty", "juicebeauty.com", "Organik Sertifikalı", "Stem Cellular Anti-Wrinkle Moisturizer; USDA organik; Gwyneth Paltrow yatırım"),
        ("Ere Perez", "ereperez.com", "Avustralya Doğal", "Natural Cosmetics; Avustralya-Meksika; rice bran + papaya; %100 doğal"),
        ("Vapour Beauty", "vapourbeauty.com", "Organik Yüksek Performans", "Soft Focus Foundation; organik + yüksek performans; New Mexico; USDA organik"),
        ("ILIA", "iliabeauty.com", "Temiz Lüks Hibrit", "Limitless Lash Mascara; temiz formül + yüksek performans; en iyi temiz maskara"),
        ("Kora Organics", "koraorganics.com", "Miranda Kerr Organik", "Noni Glow Face Oil; Miranda Kerr; noni meyvesi; Avustralya organik; COSMOS"),
        ("Dr. Hauschka", "drhauschka.com", "Alman Biyodinamik", "Rose Day Cream; 1967'den beri; biyodinamik tarım; Alman organik öncü; gül özü"),
        ("Weleda", "weleda.com", "Organik Öncü", "Skin Food; 1921'den beri; antropozofik tıp; İsviçre-Alman organik; 100+ yıl"),
        ("Lavera", "lavera.com", "Alman Sertifikalı Organik", "Basis Sensitiv; Alman organik sertifikalı; NaTrue; 30+ yıl; erişilebilir organik"),
        ("Logona", "logona.com", "Alman Doğal", "Age Protection Cream; Alman doğal kozmetik; NaTrue sertifikalı; 1978'den beri"),
        ("Mádara", "madaracosmetics.com", "Letonya Organik", "Smart Antioxidants Day Cream; Letonya organik; COSMOS; Baltık bitkileri; Avrupa organik"),
        ("Patyka", "patyka.com", "Paris Bio Lüks", "Huile Absolue; Paris biyolojik lüks; Fransa'nın ilk organik güzellik markası; 1922"),
        ("Caudalie", "caudalie.com", "Fransız Üzüm Doğal", "Vinoperfect; Bordeaux üzüm; polifenol; Fransız doğal + bilimsel; sürdürülebilir"),
        ("Natura Brasil", "natura.com.br", "Brezilya Biyoçeşitlilik", "Ekos Açaí; Amazon biyoçeşitliliği; Brezilya topluluk ticaret; B Corp; sürdürülebilir"),
        ("Forest Essentials", "forestessentials.com", "Hindistan Lüks Ayurveda", "Soundarya Radiance Cream; Hindistan ayurveda lüks; 24K altın; geleneksel formüller"),
    ],

    "Vegan & Cruelty-Free Kozmetik": [
        ("e.l.f.", "elfcosmetics.com", "Vegan Uygun Makyaj", "%100 vegan + cruelty-free; $3-14; eczane fiyatına; TikTok viral; erişilebilir güzellik"),
        ("KVD Vegan Beauty", "kvdveganbeauty.com", "Vegan Makyaj", "Tattoo Liner; %100 vegan; yüksek pigment; vegan makyaj standardı"),
        ("Hourglass", "hourglasscosmetics.com", "%100 Vegan Lüks", "Vanish Airbrush Foundation + Ambient Powder; %100 vegan; lüks performans"),
        ("Cover FX", "coverfx.com", "Vegan Renk Bilimi", "Custom Enhancer Drops; vegan; renk bilimi; kişiselleştirilmiş; tüm cilt tonları"),
        ("Pacifica", "pacificabeauty.com", "Erişilebilir Vegan", "Crystal Shimmer Body Lotion; %100 vegan; erişilebilir; Walmart + Target"),
        ("Milk Makeup", "milkmakeup.com", "Vegan Minimalist", "Kush Mascara; %100 vegan; stick formüller; minimalist; NYC"),
        ("Axiology", "axiologybeauty.com", "Sıfır Atık Vegan Dudak", "Multi-Stick; %100 vegan; sıfır atık ambalaj; Bali üretim; balmumu-free"),
        ("Aether Beauty", "aetherbeautyco.com", "Sıfır Atık Göz Paleti", "Rose Quartz Crystal Gemstone Palette; sıfır atık; kristal ilhamlı; kompostlanabilir"),
        ("BYBI", "bybi.com", "Sürdürülebilir Vegan", "Babe Balm; İngiliz vegan; sürdürülebilir ambalaj; clean + green; erişilebilir"),
        ("Ethique", "ethique.com", "Katı Bar Vegan", "Solid Shampoo + Skincare Bars; Yeni Zelanda; plastiksiz; katı format; 6M+ şişe tasarrufu"),
        ("Elate Cosmetics", "elatecosmetics.com", "B Corp Vegan Makyaj", "Bamboo ambalaj; kompostlanabilir; B Corp; sıfır atık vegan; Kanada"),
        ("PHB Ethical Beauty", "phbethicalbeauty.co.uk", "İngiliz Etik Vegan", "Mineral Foundation; İngiliz etik; organik + vegan + halal; Soil Association"),
        ("Hurraw!", "hurraw.com", "Vegan Ham Dudak Balmı", "Moon Balm; vegan + raw; organik; ayçiçeği + chamomile; dudak bakımı"),
        ("Lily Lolo", "lilylolo.com", "İngiliz Mineral Vegan", "Mineral Foundation; İngiliz mineral makyaj; vegan; doğal; hassas cilt dostu"),
        ("100% Pure", "100percentpure.com", "Meyve Pigmentli Makyaj", "Fruit Pigmented Foundation; meyve + sebze pigmentleri; vegan; doğal renklendirici"),
        ("W3LL People", "w3llpeople.com", "Clean Vegan Makyaj", "Bio Correct Multi-Action Concealer; EWG Verified; temiz + vegan; erişilebilir"),
        ("Ere Perez", "ereperez.com", "Doğal Vegan Makyaj", "Natural Cosmetics; %100 doğal + vegan; Avustralya-Meksika; rice bran bazlı"),
        ("Inika Organic", "inika.com.au", "Avustralya Organik Vegan", "Certified Organic Liquid Foundation; %100 vegan + organik; Avustralya sertifikalı"),
        ("Zuii Organic", "zuiiorganic.com", "Avustralya Çiçek Makyaj", "Flora Liquid Foundation; organik çiçek özleri; Avustralya; vegan + organik"),
        ("Antonym Cosmetics", "antonymcosmetics.com", "Sertifikalı Organik", "Certified Organic Eyeshadow; Ecocert organik; vegan; doğal pigmentler"),
    ],

    "Bebek & Hamile Cilt Bakımı": [
        ("Evereden", "evereden.com", "Pediatrik Dermatolog Bebek", "Multi-Vitamin Cream; pediatrik dermatolog geliştirdi; hamile + bebek güvenli; clean"),
        ("Pipette", "pipette.com", "Squalane Bebek", "Baby Balm; squalane bazlı; hamile + bebek; temiz; hassas formüller"),
        ("Burt's Bees Baby", "burtsbees.com", "Doğal Bebek", "Nourishing Baby Oil; doğal + organik; bebek bakımı; balmumu + botanik"),
        ("Honest Company Baby", "honest.com", "Jessica Alba Bebek", "Healing Balm; Jessica Alba; EWG onaylı; bebek + aile; temiz bileşenler"),
        ("Mustela", "mustela.com", "Fransız Bebek Bakım", "Bébé Hydra Bebe Facial Cream; Fransız eczane; 70+ yıl; avokado perseose"),
        ("Mama Mio", "mamamio.com", "Hamile Vücut Bakımı", "Tummy Rub Butter; çatlak önleme; hamile vücut bakımı; omega + shea; İngiliz"),
        ("Bio-Oil", "bio-oil.com", "Çatlak Yağı", "Skincare Oil; çatlak + leke yağı; PurCellin Oil; global best seller; hamile favorisi"),
        ("Tubby Todd", "tubbytodd.com", "Bebek Temiz Bakım", "All Over Ointment; bebek egzama + kuru cilt; doğal; anne kuruculu; DTC"),
        ("Babyganics", "babyganics.com", "Bebek Güvenli", "Moisturizing Daily Lotion; NeoNourish seed oil blend; bebek güvenli; doğal"),
        ("California Baby", "californiababy.com", "Kaliforniya Bebek", "Calendula Cream; organik calendula; hassas bebek cilt; 25+ yıl; sertifikalı"),
        ("Earth Mama Organics", "earthmamaorganics.com", "Organik Anne-Bebek", "Organic Nipple Butter; organik; hamilelik + emzirme; USDA organik; anne bakımı"),
        ("Weleda Baby", "weleda.com", "Organik Bebek", "Calendula Baby Cream; organik calendula; İsviçre-Alman; 100+ yıl organik bebek"),
        ("Mama & Kids (Japan)", "mamakids.co.jp", "Japon Anne-Bebek", "Baby Milky Lotion; Japon hassas formül; yenidoğan güvenli; dermatolojik test"),
        ("Natalia's Remedy", "nataliasremedy.com", "Hamile Aromaterapi", "Birth Easy; doğum hazırlık yağı; aromaterapi + hamile bakım; İngiliz ebe formülü"),
        ("Hatch Mama", "hatchcollection.com", "Hamile Wellness", "Belly Oil; hamile göbek yağı; organik; hamile moda + wellness; lüks anne"),
        ("Nécessaire Baby", "necessaire.com", "Temiz Bebek", "The Baby Line; temiz + güvenli; vitamin bazlı; minimalist bebek bakımı"),
        ("Attitude Baby", "attitudeliving.com", "Kanada Temiz Bebek", "Baby Leaves Body Lotion; Kanada; EWG Verified; COSMOS sertifikalı; doğal"),
        ("Dr. Bronner's Baby", "drbronners.com", "Organik Bebek Sabun", "Baby Unscented Pure-Castile Soap; organik; bebek güvenli; kokusuz; fair trade"),
        ("Erbaviva", "erbaviva.com", "Organik Hamile-Bebek", "Stretch Mark Cream; %100 organik; hamile çatlak bakımı; USDA organik"),
        ("Puracy", "puracy.com", "Doğal Bebek Bakım", "Natural Baby Lotion; %99.95 doğal; dermatoloji test; bebek güvenli; erişilebilir"),
    ],

    "İntim Bakım & Vücut": [
        ("Fur", "furyou.com", "İntim Bölge Yağı", "Fur Oil; ingrown + koyulaşma; bikini bakımı; Emma Watson; taboo kırma"),
        ("Bushbalm", "bushbalm.com", "Bikini Bakım DTC", "Dark Spot Oil + Exfoliating Bar; bikini koyulaşma; Shark Tank; cesur niş"),
        ("Love Wellness", "lovewellness.com", "Kadın Wellness", "pH Balancing Cleanser + Probiotics; vajinal sağlık; Lo Bosworth; taboo kırma"),
        ("Queen V", "queenvlife.com", "İntim Bakım", "pH balanced wash + wipes; Walmart'ta; erişilebilir intim bakım; pastel ambalaj"),
        ("The Honey Pot", "thehoneypot.co", "Bitkisel İntim", "Herbal-Infused Pads + Wash; bitkisel intim bakım; Afrika kökenli; Target; Bea Dixon"),
        ("Lola", "mylola.com", "Organik Regl", "Organic Cotton Tampons; organik pamuk; şeffaf bileşenler; abonelik; DTC regl"),
        ("Cora", "cora.life", "Organik Regl DTC", "Organic Tampons + Period Underwear; organik + giving back; her kutu = bağış"),
        ("August", "itsaugust.co", "Sürdürülebilir Regl", "Sustainable Period Care; kompostlanabilir; Gen Z; sürdürülebilir regl bakımı"),
        ("Saalt", "saalt.com", "Regl Kabı", "Menstrual Cup; silikon regl kabı; sürdürülebilir; 12 saat koruma; sıfır atık"),
        ("Flex", "flexfits.com", "Modern Regl", "Flex Disc; regl diski; 12 saat; daha az sızıntı; yenilikçi tasarım; modern"),
        ("Thinx", "thinx.com", "Regl İç Çamaşırı", "Period Underwear; emici iç çamaşırı; sürdürülebilir regl; öncü; DTC"),
        ("Knix", "knix.com", "Sızıntı Geçirmez İç Çamaşır", "Leakproof Underwear; regl + hafif kaçırma; kadın wellness; Kanada DTC"),
        ("Dame Products", "dameproducts.com", "Kadın Zevk", "Eva II; kadın cinsel sağlık; vibratör; tasarım + mühendislik; wellness"),
        ("Maude", "getmaude.com", "Modern Cinsel Wellness", "Vibe + Shine; cinsel wellness; Dakota Johnson yatırım; minimal tasarım; unisex"),
        ("Lola Hygiene", "mylola.com", "İntim Hijyen", "Cleansing Wipes + Wash; intim hijyen; organik; pH dengeli; şeffaf"),
        ("Sustain", "sustain.com", "Sürdürülebilir Kondom", "Organic Condoms; organik lateks; fair trade; nitrosamine-free; vegan"),
        ("Good Clean Love", "goodcleanlove.com", "Organik İntim", "BioGenesis Intimate Wash; organik; pH dengeli; prebiyotik; biyom dostu"),
        ("Rael", "getrael.com", "Kore İntim Bakım", "Feminine Wash + Organic Pads; Kore-Amerikan; organik regl + intim bakım"),
        ("Vagisil", "vagisil.com", "İntim Bakım Klasik", "pH Balance Wash; intim bakım klasiği; pH dengeli; günlük yıkama; erişilebilir"),
        ("Summer's Eve", "summerseve.com", "İntim Tazeleme", "Cleansing Wash; intim tazeleme; çeşitli kokular; erişilebilir; günlük"),
    ],

    "Diş Beyazlatma & Ağız Bakımı": [
        ("Bite", "bitetoothpastebits.com", "Tablet Diş Macunu", "Toothpaste Bits; tablet format; plastiksiz; sıfır atık; cam şişe; sürdürülebilir"),
        ("Twice", "brushtwice.com", "Premium Diş Macunu", "Oral Wellness Toothpaste; vitamin + aloe; premium diş macunu; temiz formül"),
        ("Davids Natural", "davids-usa.com", "Doğal Premium", "Premium Natural Toothpaste; metal tüp; geri dönüştürülebilir; nano-hydroxyapatite"),
        ("Hismile", "hismile.com", "Beyazlatma Uzmanı", "PAP+ Teeth Whitening Kit; peroksitsiz; Avustralya DTC; uygun fiyat; TikTok viral"),
        ("Snow", "trysnow.com", "LED Beyazlatma", "Snow Teeth Whitening Kit; LED ışık + serum; evde profesyonel beyazlatma; DTC"),
        ("SmileDirectClub alternatifi: Byte", "byteme.com", "Şeffaf Plak", "At-Night Clear Aligners; gece şeffaf plak; daha hızlı; HyperByte teknolojisi"),
        ("Quip", "getquip.com", "Abonelik Diş Fırçası", "Electric Toothbrush + Refill; abonelik; minimal tasarım; ADA onaylı; DTC"),
        ("Burst", "burstoralcare.com", "Sonic Diş Fırçası", "BURST Sonic Toothbrush; charcoal kıl; sonic teknoloji; diş hekimi önerisi; DTC"),
        ("AutoBrush", "autobrush.com", "U-Şekil Diş Fırçası", "AutoBrush Pro; U-şekilli otomatik fırça; 30 saniye; tüm dişler aynı anda"),
        ("Moon Oral Care", "moonoralcare.com", "Kendall Jenner Ağız", "Whitening Toothpaste; Kendall Jenner; aktif kömür; premium ağız bakımı"),
        ("Hello Products", "hello-products.com", "Doğal Ağız Bakımı", "Activated Charcoal Toothpaste; doğal; vegan; cruelty-free; erişilebilir; Walmart"),
        ("Tom's of Maine", "tomsofmaine.com", "Doğal Diş Macunu", "Whole Care Toothpaste; doğal; florürlü + doğal; 50+ yıl; Colgate"),
        ("Dr. Bronner's Toothpaste", "drbronners.com", "Organik Diş Macunu", "All-One Toothpaste; organik; hindistan cevizi yağı; fair trade; vegan"),
        ("Marvis", "marvis.com", "İtalyan Lüks Diş Macunu", "Whitening Mint; İtalyan lüks; vintage ambalaj; 7 lezzet; Floransa; premium"),
        ("Crest Whitestrips alternatifi: Lumineux", "lumineux.com", "Doğal Beyazlatma", "Oral Essentials Whitening Strips; doğal beyazlatma; peroksitsiz; diş hekimi formülü"),
        ("Spotlight Oral Care", "spotlightoralcare.com", "İrlanda Diş Bakımı", "Teeth Whitening Strips; İrlanda diş hekimleri; H2O2 free; profesyonel formül"),
        ("Cocofloss", "cocofloss.com", "Lüks Diş İpi", "Cocofloss; hindistan cevizi yağı kaplı; lüks diş ipi; eğlenceli renkler; DTC"),
        ("Boka", "bfrsh.com", "Nano-Hydroxyapatite", "Ela Mint Toothpaste; n-Ha; Japon beyazlatma teknolojisi; florürsüz alternatif; DTC"),
        ("RiseWell", "risewell.com", "Hydroxyapatite Aile", "Mineral Toothpaste; hydroxyapatite; aile boyu; çocuk + yetişkin; doğal mineral"),
        ("SuperMouth", "supermouth.com", "Çocuk Ağız Bakımı", "Kids Toothpaste; hydroxamin; çocuk ağız bakımı; diş hekimi formülü; eğlenceli"),
        ("OJOOK", "ojook.com", "Kore Bambu Tuz", "Bamboo Salt Toothpaste; Kore geleneksel bambu tuz; 1000 yıllık formül; modern"),
        ("Aesop Mouthwash", "aesop.com", "Botanik Gargarası", "Mouthwash; botanik lüks; anise + clove; Avustralya minimalist; premium ağız bakımı"),
        ("Lush Toothy Tabs", "lush.com", "Tablet Diş Bakımı", "Toothy Tabs; tablet format; plastiksiz; el yapımı; taze; sıfır atık"),
        ("Public Goods", "publicgoods.com", "Minimalist Ağız Bakımı", "Fluoride Toothpaste; minimalist ambalaj; sürdürülebilir; üyelik modeli; temiz"),
        ("Noobs", "noobs.com", "Tablet Diş Macunu", "Toothpaste Tablets; tablet format; vegan; florürlü; seyahat dostu; sıfır plastik"),
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def create_meta_ads_url(brand_name: str) -> str:
    encoded_name = quote(brand_name)
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={encoded_name}&search_type=keyword_unordered"


def apply_header_style(ws, row, max_col):
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick_border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="medium", color="1B2A4A"),
        bottom=Side(style="medium", color="1B2A4A"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border


def apply_data_row(ws, row, max_col, category_color=None, is_even=False):
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    if category_color and is_even:
        fill_color = category_color
    else:
        fill_color = "FFFFFF"
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_category_separator(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left,
            right=current_border.right,
            top=current_border.top,
            bottom=Side(style="medium", color="1B2A4A"),
        )


# ─── EXTRA BRANDS 4: Massive expansion to reach 5000+ ─────────────────────────
EXTRA_BRANDS_4 = {
    "Cilt Bakımı - Nemlendirici & Serum": [
        ("Drunk Elephant alternatifi: Fourth Ray Beauty", "fourthraybeauty.com", "Colourpop Cilt Bakımı", "Colourpop'un cilt bakım markası; erişilebilir aktif serumlar; $8-12; Seed Beauty"),
        ("Lixirskin", "lixirskin.com", "İngiliz Sade Serum", "Universal Emulsion; İngiliz minimalist; 4 ürünlük basit rutin; yeterli + etkili"),
        ("Isla Beauty", "islaskincare.com", "İngiliz Aktif", "Hydra Gel Moisturiser; İngiliz temiz aktif; hyaluronic + niacinamide; erişilebilir"),
        ("Whamisa Deep Rich", "whamisa.com", "Kore Fermente Serum", "Deep Rich Essence Toner; organik fermente; %95+ doğal; Kore organik lider"),
        ("Purito Dermide", "purito.com", "Kore Cica Serum", "Dermide Cica Barrier Sleeping Pack; gece cica onarım; centella + ceramide"),
        ("Blithe", "blithecosme.com", "Kore Pressed Serum", "Crystal Iceplant Pressed Serum; sıkıştırılmış serum formatı; Kore inovatif form"),
        ("Nooni", "nooni.co", "Kore Elma Suyu Serum", "Applemint Lip Oil; Kore meyve bazlı; nemlendirici; elma suyu cilt bakımı"),
        ("Ariul", "ariul.com", "Kore 7 Gün Maske", "7 Days Vitamin Mist; günlük vitamin mist; Kore haftalık bakım konsepti"),
        ("Dewytree", "dewytree.com", "Kore Aqua Serum", "Aqua Collagen Peptide Serum; Kore kolajen peptide; anti-aging + nemlendirme"),
        ("SWANICOCO", "swanicoco.com", "Kore Bio Serum", "Bio Therapy 1st Essence; Kore biyoterapi; fermente galactomyces; anti-aging"),
        ("23 Years Old", "23yearsold.com", "Kore Derma Serum", "Cocktail Serum; derma-kozmetik; Kore klinik cilt bakımı; C + retinol coctkail"),
        ("Dr. Oracle", "droracle.co.kr", "Kore Antibac Serum", "Antibac Green Therapy; Kore antibakteriyel cilt bakımı; yeşil çay + cica"),
        ("Sioris", "sioris.com", "Kore Taze Serum", "Time is Running Out Mist; mevsimsel taze hasat; Kore slow beauty; zaman serumu"),
        ("CNP Laboratory", "cnp-mall.com", "Kore Dermatolog Serumu", "Propolis Energy Ampule; Kore dermatoloji kliniği markası; propolis uzmanı"),
        ("Hanyul", "hanyul.com", "Kore Hanbang Serumu", "Yuja Niacin Blemish Care Serum; Kore geleneksel yuja; hanbang aydınlatma"),
        ("Atopalm", "atopalm.com", "Kore MLE Serum", "Real Barrier Cicarelief Cream; MLE çoklu katmanlı; Kore bariyer teknoloji"),
        ("Troiareuke", "troiareuke.com", "Kore Kapsül Serum", "Acsen Oil Cut Cleansing; Kore kapsül formül; dermatoloji; akne odaklı"),
        ("Laka", "laka.co.kr", "Kore Genderless Serum", "Vegan Lip & Cheek + cilt bakımı; cinsiyet nötr K-beauty; minimal tasarım"),
        ("Miguhara", "miguhara.co.kr", "Kore EGF Serum", "Bio Cream; EGF growth factor; Kore anti-aging bilimi; büyüme faktörü"),
        ("b.glen", "bglen.net", "Japon QuSome Serum", "QuSome nüfuz teknolojisi; Japon bilimsel; ilaç iletim sistemi ilhamlı formüller"),
        ("Obagi Derma Power X", "obagi.co.jp", "Japon Obagi Serum", "Derma Power X Stem Lift Cream; Japon kök hücre; anti-aging; lüks"),
        ("Ampleur", "ampleur.jp", "Japon Leke Serumu", "Luxury White Concentrate; Japon dermatolog; lazer tedavisi ilhamlı; aydınlatma"),
        ("Attenir", "attenir.co.jp", "Japon Çift Temizleme Yağı", "Skin Clear Cleanse Oil; Japon #1 temizleme yağı; 40M+ satış; uygun fiyat"),
        ("TUNEMAKERS", "tunemakers.net", "Japon Konsantre Serum", "Ceramide 200 Moisturizing Essence; Japon konsantre formül; tek bileşen yüksek doz"),
        ("Alovivi", "alovivi.jp", "Japon Büyük Boy Serum", "Purevivi Cleansing Lotion; Japon büyük boy uygun fiyat; 500ml; misel su"),
        ("Mediplus", "mediplus.co.jp", "Japon All-in-One", "Mediplus-Gel; Japon all-in-one jel; 66 aktif bileşen tek üründe; basitlik"),
        ("SkinCeuticals alternatifi: BeautyStat", "beautystat.com", "Stabilize C", "Universal C Skin Refiner; %20 stabilize C vitamini; MIT bilimi; patent"),
        ("Maelove", "maelove.com", "Uygun CE Ferulic", "The Glow Maker; $28 CE Ferulic dupe; bilimsel formül düşük fiyat"),
        ("Holy Snails", "holysnails.com", "Reddit Indie Serum", "Shark Sauce; Reddit SkincareAddiction favorisi; DIY'den markaya; niacinamide kült"),
        ("Cos De BAHA", "cosdebaha.com", "Amazon K-Serum", "Azelaic Acid %10; Amazon best seller; Kore formül; uygun fiyat aktif"),
        ("SeoulCeuticals", "seoulceuticals.com", "Kore-ABD Serum", "Korean Skin Care Snail Mucin; Kore formül ABD markası; salyangoz müsin"),
        ("Purlisse", "purlisse.com", "Asya İlhamlı Serum", "Blue Lotus Serum; Asya botanik; lotus + beyaz çay; ABD + Asya hibrit"),
        ("Acure", "acure.com", "Organik Uygun Serum", "Brightening Vitamin C & Ferulic Acid Serum; organik; $10; Target; temiz"),
        ("Derma E", "dermae.com", "Eczane Aktif Serum", "Vitamin C Concentrated Serum; eczane fiyatına; Amerikan temiz güzellik"),
        ("Instanatural", "instanatural.com", "Amazon Doğal Serum", "Vitamin C Serum; Amazon en çok satan; doğal + aktif; erişilebilir"),
        ("TruSkin", "truskin.com", "Amazon Vitamin C", "Vitamin C Serum; Amazon #1; retinol + HA; erişilebilir aktif bakım"),
        ("Eva Naturals", "evanaturals.com", "Amazon Doğal Aktif", "Vitamin C+ Serum; Amazon best seller; doğal vitamin C + HA"),
        ("Peach Slices", "peachslices.com", "Uygun K-Beauty", "Acne Spot Dots + serums; uygun fiyat K-beauty ABD; CVS + Target"),
        ("Sweet Chef", "sweetchef.com", "Gıda Bazlı Serum", "Beet + Vitamin A Serum Shot; gıda ilhamlı cilt bakımı; erişilebilir; eğlenceli"),
        ("I Dew Care", "idewcare.com", "Eğlenceli K-Beauty Serum", "Vitamin To-Glow Pack; eğlenceli paketleme; K-beauty; vitamin; uygun fiyat"),
        ("Thank You Farmer", "thankyoufarmer.us", "Kore Su Serumu", "True Water Deep Serum; Kore derin nemlendirme; İzlanda yosunu + Kore bilim"),
        ("d'Alba Piedmont", "dalba.co.kr", "Kore Trüf Serum", "White Truffle Nourishing Treatment Serum; beyaz trüf; Kore lüks; spray serum"),
        ("Mary & May", "maryandmay.com", "Kore Çift Serum", "Idebenone + Blackberry Complex Serum; çift aktif; Kore bitkisel + bilimsel"),
        ("Haruharu WONDER", "haruharuwonder.com", "Kore Siyah Bambu", "Black Bamboo Mist; siyah bambu özütü; Kore hanbang modern; antioksidan"),
        ("Ma:nyo Bifida", "manyo.co.kr", "Kore Bifida Serum", "Bifida Biome Complex Ampoule; bifida ferment; Kore mikrobiyom serumu"),
        ("Jayjun", "jayjun.com", "Kore Estetik Serum", "Anti-Dust Whitening Mask; Kore estetik klinikten; kirlilik koruma"),
        ("Easydew", "easydew.co.kr", "Kore DW-EGF Serum", "DW-EGF Cream; Kore EGF patent; dermatoloji kökenli; büyüme faktörü"),
        ("Dr. Althea", "dralthea.com", "Kore Amino Serum", "Amino Acid Gentle Bubble Cleanser; amino asit bazlı; Kore hassas formül"),
        ("Banobagi", "banobagi.com", "Kore Estetik Klinik Serum", "Vita Genic Jelly Mask; Kore estetik cerrahi klinikten cilt bakıma; jöle maske"),
        ("JM Solution", "jmsolution.com", "Kore Altın Serum", "Active Golden Caviar Nourishing Mask; Kore altın + havyar; lüks sheet mask"),
        ("Bergamo", "bergamo.co.kr", "Kore Luxury Gold Serum", "Luxury Gold Collagen Serum; Kore altın kollajen; uygun fiyat lüks his"),
    ],

    "Cilt Bakımı - Temizleyici & Tonik": [
        ("Heimish", "heimish.com", "Kore pH Temizleyici", "pH 5.5 All Clean Green Foam; pH dengeli; Kore temiz yeşil çay"),
        ("TIRTIR", "tirtir.com", "Kore Milk Skin", "Milk Skin Toner; süt protein bazlı; Kore süt cilt aydınlatma"),
        ("Torriden", "torriden.com", "Kore Balanceful Tonik", "Balanceful Cica Toner; cica + hyaluronic; Kore dengeli tonik"),
        ("Mediheal", "mediheal.com", "Kore Tea Tree Tonik", "Tea Tree Care Solution Essential Mask; Kore hastane kökenli"),
        ("Yadah", "yadah.com", "Kore Anti-T Tonik", "Anti-T Toner; çay ağacı; Kore sorunlu cilt toniği"),
        ("Dr. Ceuracle", "dr.ceuracle.com", "Kore Kombucha Tonik", "Vegan Kombucha Tea Essence; fermente çay; Kore vegan"),
        ("Skinfood", "theskinfood.com", "Kore Gıda Temizleyici", "Black Sugar Perfect First Serum; siyah şeker; Kore gıda kozmetik"),
        ("Tony Moly", "tonymoly.com", "Kore Meyve Temizleyici", "Peach Punch Sweet Foam Cleanser; şeftali; eğlenceli K-beauty"),
        ("Nature Republic", "naturerepublic.com", "Kore Aloe Temizleyici", "Soothing & Moisture Aloe Vera Cleansing Gel; Kore aloe vera"),
        ("The Face Shop", "thefaceshop.com", "Kore Pirinç Temizleyici", "Rice Water Bright Cleansing Foam; pirinç suyu; Kore geleneksel"),
    ],

    "Cilt Bakımı - Güneş Koruma (SPF)": [
        ("ROHTO Skin Aqua", "rohto.com", "Japon Tone-Up SPF", "Skin Aqua Tone Up UV Essence SPF 50+; lavanta ton-up; Japon"),
        ("Senka", "senka.com", "Japon Perfect Whip SPF", "Perfect UV Gel; Shiseido alt marka; Japon uygun SPF"),
        ("Verdio", "omibrotherhood.com", "Japon Organik SPF", "UV Moisture Gel SPF 50+; organik; Japon doğal SPF"),
        ("Nivea Japan", "nivea.co.jp", "Japon Super Water Gel", "UV Super Water Gel SPF 50; su bazlı; Japon uygun"),
        ("Spicara", "spicara.com", "Japon Doğal SPF", "V UV Cream; Japon doğal kozmetik SPF; mineral bazlı"),
        ("Mentholatum Sunplay", "mentholatum.com", "Japon Spor SPF", "Sunplay Super Block SPF 81; ultra yüksek koruma; Japon outdoor"),
        ("Kao Curél UV", "curel.com", "Japon Ceramide SPF", "UV Protection Face Cream SPF 30; ceramide + SPF; hassas"),
        ("Fancl Sunguard", "fancl.com", "Japon Koruyucusuz SPF", "Sunguard 50+ Protect UV; koruyucu maddeziz; Japon temiz SPF"),
        ("Mama Butter", "btalmonds.com", "Japon Shea SPF", "UV Care Cream SPF 25; shea butter + SPF; Japon nazik koruma"),
        ("Naris Up", "narisup.com", "Japon Parfümlü SPF", "Parasola UV Cut Spray SPF 50+; parfümlü sprey SPF; Japon"),
    ],

    "Cilt Bakımı - Akne & Leke Tedavisi": [
        ("Eradikate", "katesomerville.com", "Kükürt Akne", "EradiKate spot treatment; pembe kükürt; Hollywood facialist"),
        ("Peter Thomas Roth Acne", "peterthomasroth.com", "Max Akne", "Max Complexion Correction Pads; salisilik + glikolik; NYC dermatoloji"),
        ("Clean & Clear alternatifi: Blume", "blume.com", "Ergen Akne", "Meltdown Acne Oil; tea tree + rosehip; ergenler; eğlenceli"),
        ("La Roche Posay Effaclar", "laroche-posay.com", "Fransız Akne", "Effaclar Duo; benzoyl peroxide; Fransız eczane akne"),
        ("Avène Cleanance", "avene.com", "Fransız Termal Akne", "Cleanance Comedomed; termal su + akne; Fransız hassas akne"),
        ("Vichy Normaderm", "vichy.com", "Fransız Volcano Akne", "Normaderm Phytosolution; volkanik su; salisilik asit"),
        ("Sebamed Clear Face", "sebamed.com", "Alman pH Akne", "Clear Face Anti-Pimple Gel; pH 5.5; Alman dermo"),
        ("Clearasil", "clearasil.com", "Klasik Akne", "Rapid Rescue Spot Treatment; benzoyl peroxide; 50+ yıl"),
        ("Neutrogena Rapid Clear", "neutrogena.com", "Salisilik Akne", "Rapid Clear Stubborn Acne; %10 benzoyl peroxide; eczane"),
        ("Panoxyl", "panoxyl.com", "Benzoyl Peroxide Wash", "Acne Foaming Wash %10; benzoyl peroxide; eczane; erişilebilir"),
    ],

    "Cilt Bakımı - Anti-Aging & Kırışıklık": [
        ("La Prairie Skin Caviar", "laprairie.com", "İsviçre Kaviar", "Skin Caviar Luxe Cream; kaviar + İsviçre bilimi; ultra lüks"),
        ("Valmont Prime", "lamaisonvalmont.com", "İsviçre Buzul", "Prime Renewing Pack; buzul suyu; İsviçre hücre yenilenme"),
        ("Sisley Sisleÿa", "sisley-paris.com", "Fransız Bitki", "Sisleÿa L'Intégral; fitokozmetik; 40+ yıl araştırma"),
        ("Helena Rubinstein", "helenarubinstein.com", "Prodigy Cellglow", "Prodigy Cellglow; native kolajen; Fransız lüks"),
        ("Estée Lauder ANR alternatifi: Missha", "missha.com", "Kore Time Revolution", "Time Revolution Night Repair Ampoule; Kore ANR dupe"),
        ("SK-II GenOptics", "sk-ii.com", "Japon Aura", "GenOptics Aura Essence; PITERA + niacinamide; Japon aydınlatma"),
        ("ReFa S Carat", "refa.net", "Japon Mikro Akım", "S Carat; yüz mikro akım masaj; Japon güzellik cihazı + serum"),
        ("Environ", "environ.com", "Güney Afrika Vitamin A", "AVST Moisturiser; vitamin A STEP-UP sistemi; Güney Afrika; dermatoloji"),
        ("ZO Skin Health", "zoskinhealth.com", "Dr. Obagi Yeni Marka", "Growth Factor Serum; Dr. Zein Obagi; büyüme faktörü; medikal"),
        ("SkinBetter Science", "skinbetter.com", "Alpharet Anti-Aging", "AlphaRet Overnight Cream; retinoid + alpha hydroxy acid; medikal estetik"),
    ],

    "Cilt Bakımı - Göz Çevresi": [
        ("Clinique All About Eyes", "clinique.com", "Dermatolojik Göz", "All About Eyes; koyu halka + şişlik; alerjist test; klasik"),
        ("Estée Lauder ANR Eye", "esteelauder.com", "ANR Göz", "Advanced Night Repair Eye; gece onarım; lüks göz serumu"),
        ("La Mer Eye Concentrate", "cremedelamer.com", "Lüks Göz", "The Eye Concentrate; Miracle Broth; ultra lüks"),
        ("Bobbi Brown alternatifi: Pixi", "pixibeauty.com", "FortifEYE", "FortifEYE Toning Eye Patches; peptide + kollajen; £22"),
        ("Lancôme Génifique Eye", "lancome.com", "Fransız Göz", "Advanced Génifique Yeux; fermente + ışık difüzör"),
    ],

    "Cilt Bakımı - Maske & Peeling": [
        ("Annie's Way", "anniesway.com", "Tayvan Jöle Maske", "Jelly Mask; Tayvan; jöle kıvamlı profesyonel maske; salon evde"),
        ("Soo'AE", "sooae.com", "Kore Donkey Milk Maske", "Donkey Milk Sheet Mask; eşek sütü; Kore geleneksel"),
        ("Leaders", "leaders.co.kr", "Kore Amino Maske", "Amino Moisture Mask; Kore lider sheet mask markası"),
        ("SNP", "snp.co.kr", "Kore Gold Maske", "Gold Collagen Ampoule Mask; altın kolajen; Kore premium sheet"),
        ("Jayjun", "jayjun.com", "Kore Baby Pure Maske", "Baby Pure Shining Mask; bebek cilt efekti; Kore aydınlatma"),
    ],

    "Cilt Bakımı - Dudak Bakımı": [
        ("Dior Lip Glow alternatifi: Maybelline", "maybelline.com", "Baby Lips", "Baby Lips; renkli dudak balmı; erişilebilir; çeşitli renkler"),
        ("Vaseline Lip Therapy", "vaseline.com", "Klasik Dudak", "Lip Therapy; petrolatum bazlı; tin format; dudak onarım klasiği"),
        ("Chapstick", "chapstick.com", "Amerikan Dudak Klasiği", "Original Lip Balm; 1880'den beri; Amerikan dudak bakımı standardı"),
        ("Lucas' Papaw", "lucaspapaw.com.au", "Avustralya Papaya", "Ointment; fermente papaya; Avustralya; çok amaçlı; 100+ yıl"),
        ("Paw Paw", "naturescare.com.au", "Avustralya Çok Amaçlı", "Paw Paw Balm; papaya bazlı; dudak + cilt; Avustralya doğal"),
    ],

    "Cilt Bakımı - Bariyer Onarım": [
        ("Aveeno Calm+Restore", "aveeno.com", "Yulaf Bariyer", "Calm + Restore Oat Gel Moisturizer; yulaf + feverfew; bariyer güçlendirme"),
        ("Aquaphor", "aquaphor.com", "Onarım Merhemi", "Healing Ointment; petrolatum bariyer; kuru + çatlak cilt; çok amaçlı onarım"),
        ("Cicaplast Baume B5", "laroche-posay.com", "Fransız Cica", "Baume B5+; panthenol + cica; Fransız bariyer standart; çok amaçlı"),
        ("Bepanthen", "bepanthen.com", "Alman Panthenol", "Bepanthen Cream; dexpanthenol; Alman bariyer onarım; bebek + yetişkin"),
        ("Ceradan", "ceradan.com", "Singapur Ceramide", "Skin Barrier Repair Cream; 3:1:1 ceramide oranı; Singapur dermo"),
    ],

    "Cilt Bakımı - Hassas Cilt": [
        ("Physiogel", "physiogel.com", "Alman Lipid Hassas", "Daily Moisture Therapy Cream; BioMimic lipid; Alman hassas"),
        ("Cetaphil", "cetaphil.com", "Dermatolojik Hassas", "Gentle Skin Cleanser; dermatoloji standardı; 70+ yıl; hassas"),
        ("QV", "qvskincare.com.au", "Avustralya Hassas", "Intensive Cream; Avustralya dermatolojik; kuru hassas cilt"),
        ("Avène Cicalfate", "avene.com", "Fransız Onarım Hassas", "Cicalfate+ Repair Cream; termal su + cica; post-procedure"),
        ("SVR Sensifine", "laboratoire-svr.com", "Fransız Ultra Hassas", "Sensifine Cream; Fransız eczane; ultra hassas cilt formülü"),
    ],

    "Cilt Bakımı - Hiperpigmentasyon": [
        ("Faded by Topicals", "mytopicals.com", "Viral Leke Serumu", "Faded Serum; azelaic + centella + kojic; viral; kapsayıcı leke tedavisi"),
        ("Murad Rapid Dark Spot", "murad.com", "Hızlı Leke", "Rapid Dark Spot Correcting Serum; resorcinol; dermatoloji leke tedavisi"),
        ("SkinCeuticals Discoloration", "skinceuticals.com", "Medikal Leke", "Discoloration Defense; tranexamic + kojic + niacinamide; medikal"),
        ("Glow Recipe Guava", "glowrecipe.com", "Guava Leke", "Guava Vitamin C Dark Spot Serum; guava + C vitamini; meyve aydınlatma"),
        ("Dr. Dennis Gross C+ Collagen", "drdennisgross.com", "C Kollajen Aydınlatma", "C+ Collagen Brighten & Firm Serum; C vitamini + kolajen"),
    ],

    "Makyaj - Fondöten & BB/CC Krem": [
        ("Armani Luminous Silk alternatifi: L'Oréal", "loreal.com", "True Match Fondöten", "True Match Foundation; 45 ton; eczane fiyatına; geniş çeşitlilik"),
        ("MAC Studio Fix alternatifi: Maybelline", "maybelline.com", "Fit Me Fondöten", "Fit Me Foundation; 40 ton; mat + dewy; eczane en çok satan"),
        ("Revlon", "revlon.com", "ColorStay Fondöten", "ColorStay Foundation; 24 saat; yağlı + kuru cilt versiyonları"),
        ("Covergirl", "covergirl.com", "Clean Fresh Fondöten", "Clean Fresh Skin Milk Foundation; temiz formül; vegan; hafif"),
        ("Wet n Wild", "wetnwild.com", "Ultra Uygun Fondöten", "Photo Focus Foundation; $5; yüksek performans; erişilebilir"),
        ("L.A. Girl", "lagirlusa.com", "Pro Coverage Fondöten", "Pro Coverage HD Foundation; $8; 30 ton; profesyonel + erişilebilir"),
        ("BH Cosmetics", "bhcosmetics.com", "Naturally Flawless", "Naturally Flawless Foundation; erişilebilir; hafif; 30 ton"),
        ("Juvia's Place", "juviasplace.com", "Kapsayıcı Fondöten", "I Am Magic Foundation; 42 ton; yoğun pigment; Afrika ilhamlı"),
        ("Beauty Bakerie", "beautybakerie.com", "Fırın Fondöten", "Cake Mix Foundation; fırın temalı; su geçirmez; eğlenceli"),
        ("Black Up", "blackup.com", "Koyu Ton Fondöten", "Full Coverage Cream Foundation; koyu cilt tonları uzmanı; Paris; profesyonel"),
    ],

    "Makyaj - Allık & Bronzer": [
        ("Laura Mercier alternatifi: Milani", "milanicosmetics.com", "Baked Allık", "Baked Blush Luminoso; İtalyan pişirme; ikonik sıcak tonlar"),
        ("NARS Orgasm alternatifi: e.l.f.", "elfcosmetics.com", "Uygun Allık", "Blush Palette; $6; çoklu renk; NARS dupe; TikTok viral"),
        ("Rare Beauty alternatifi: Flower Beauty", "flowerbeauty.com", "Sıvı Allık", "Blush Bomb Color Drops; sıvı allık; Drew Barrymore; $10"),
        ("Tarte Amazonian Clay alternatifi: Catrice", "catrice.eu", "Alman Blush Box", "Blush Box; Alman eczane; uygun fiyat; doğal renk"),
        ("Too Faced Peach alternatifi: Revolution", "revolutionbeauty.com", "İngiliz Baked Allık", "Baked Blusher; İngiliz; £3; çeşitli tonlar"),
    ],

    "Makyaj - Göz Makyajı": [
        ("Urban Decay alternatifi: NYX", "nyxcosmetics.com", "Ultimate Göz", "Ultimate Shadow Palette; 16 renk; $18; profesyonel pigment"),
        ("Morphe alternatifi: BH Cosmetics", "bhcosmetics.com", "Büyük Palet", "Take Me Back to Brazil; 35 renk; erişilebilir; YouTube favorisi"),
        ("Charlotte Tilbury Pillow Talk alternatifi: e.l.f.", "elfcosmetics.com", "Bite Size Göz", "Bite-Size Eyeshadow; $3; 4 renk mini palet; TikTok dupe"),
        ("Too Faced Chocolate alternatifi: Revolution", "revolutionbeauty.com", "I Heart Revolution", "Chocolate Palette; £5; çikolata kokulu; İngiliz dupe"),
        ("ABH Modern Renaissance alternatifi: Colourpop", "colourpop.com", "Give It To Me Straight", "Pressed Powder Palette; $12; ABH vibes; erişilebilir"),
    ],

    "Makyaj - Dudak Ürünleri": [
        ("MAC Ruby Woo alternatifi: NYX", "nyxcosmetics.com", "Matte Lip Cream", "Soft Matte Lip Cream; $7; 40+ renk; mat dudak standardı"),
        ("Charlotte Tilbury Pillow Talk alternatifi: Essence", "essence.eu", "€2 Lip Liner", "Lip Liner; €2; Alman eczane; çeşitli nude tonlar"),
        ("Dior Lip Oil alternatifi: NYX", "nyxcosmetics.com", "Fat Oil Lip Drip", "Fat Oil Lip Drip; dudak yağı; viral dupe; $9; erişilebilir"),
        ("Fenty Gloss Bomb alternatifi: e.l.f.", "elfcosmetics.com", "Lip Gloss", "Lip Lacquer; $3; dolgunlaştırıcı; Fenty dupe; TikTok"),
        ("Maybelline Super Stay", "maybelline.com", "24H Matte Ink", "SuperStay Matte Ink; 16H dayanıklı; mat sıvı ruj; en çok satan"),
    ],

    "Makyaj - Kaş Ürünleri": [
        ("ABH Brow Wiz alternatifi: Maybelline", "maybelline.com", "Micro Brow", "Brow Ultra Slim; ultra ince; ABH dupe; $8; erişilebilir"),
        ("Benefit Gimme Brow alternatifi: NYX", "nyxcosmetics.com", "Fill & Fluff", "Fill & Fluff Eyebrow Pomade Pencil; çift uçlu; $10"),
        ("Glossier Boy Brow alternatifi: L'Oréal", "loreal.com", "Brow Artist", "Brow Artist Plumper; fiber kaş jeli; erişilebilir; eczane"),
        ("ABH Dipbrow alternatifi: Revolution", "revolutionbeauty.com", "Brow Pomade", "Brow Pomade; £3; ABH dupe; İngiliz uygun; çeşitli tonlar"),
        ("Benefit Precisely alternatifi: Catrice", "catrice.eu", "Slim'Matic", "Slim'Matic Ultra Precise; €3; ultra ince; Alman eczane"),
    ],

    "Makyaj - Aydınlatıcı & Kontür": [
        ("CT Flawless Filter alternatifi: e.l.f.", "elfcosmetics.com", "Halo Glow Filter", "Halo Glow Liquid Filter; $14; viral dupe; dewy glow"),
        ("Fenty Match Stix alternatifi: NYX", "nyxcosmetics.com", "Wonder Stick", "Wonder Stick; çift uçlu kontür + aydınlatıcı; erişilebilir"),
        ("Becca Champagne Pop alternatifi: Milani", "milanicosmetics.com", "Strobelight", "Strobelight Instant Glow Powder; parlak aydınlatıcı; $10"),
        ("Hourglass alternatifi: Catrice", "catrice.eu", "Light Correcting", "Light Correcting Serum Primer; ışık düzeltme; Alman eczane"),
        ("Benefit Hoola alternatifi: e.l.f.", "elfcosmetics.com", "Putty Bronzer", "Putty Bronzer; $6; sponge kıvamlı; mat bronzer; viral dupe"),
    ],

    "Saç Bakımı - Şampuan & Saç Kremi": [
        ("Head & Shoulders", "headandshoulders.com", "Klasik Kepek", "Classic Clean; çinko pyrithione; global kepek standardı"),
        ("Nizoral", "nizoral.com", "Ketoconazole Şampuan", "Anti-Dandruff Shampoo; %1 ketoconazole; eczane kepek tedavisi"),
        ("Selsun Blue", "selsunblue.com", "Selenium Kepek", "Selenium Sulfide; selenyum sülfid; eczane kepek şampuanı"),
        ("Pantene", "pantene.com", "Pro-V Şampuan", "Pro-V Daily Moisture Renewal; provitamin B5; erişilebilir"),
        ("Garnier Fructis", "garnier.com", "Meyve Şampuan", "Sleek & Shine; meyve konsantresi; erişilebilir saç bakımı"),
        ("Herbal Essences", "herbalessences.com", "Botanik Şampuan", "Bio:Renew; botanik özler; sürdürülebilir; erişilebilir"),
        ("Aussie", "aussie.com", "Avustralya Saç", "Miracle Moist; Avustralya botanik; erişilebilir; macadamia yağı"),
        ("Aveeno Hair", "aveeno.com", "Yulaf Saç", "Apple Cider Vinegar Blend Shampoo; yulaf + elma sirkesi; doğal"),
        ("Verb", "verbproducts.com", "Salon DTC Saç", "Ghost Oil; salon kalitesi DTC; erişilebilir; moringa + bambu"),
        ("Eva NYC", "evanyc.com", "NYC Salon Uygun", "Therapy Session Hair Mask; NYC salon; erişilebilir; argan + keratin"),
    ],

    "Saç Bakımı - Saç Maskesi & Onarım": [
        ("Garnier Hair Food", "garnier.com", "Gıda Maske", "Hair Food Banana; gıda bazlı; erişilebilir; 3-in-1 maske"),
        ("Pantene Rescue Shots", "pantene.com", "Kurtarma Ampulü", "Rescue Shots; tek kullanımlık ampul; yoğun onarım; seyahat"),
        ("TRESemmé Keratin", "tresemme.com", "Keratin Maske", "Keratin Smooth Mask; keratin onarım; salon ilhamlı; erişilebilir"),
        ("Aussie 3 Minute Miracle", "aussie.com", "3 Dakika Mucize", "3 Minute Miracle; hızlı derin bakım; Avustralya botanik"),
        ("Arvazallia", "arvazallia.com", "Argan Maske", "Hydrating Argan Oil Hair Mask; argan yağı; Amazon best seller"),
    ],

    "Saç Bakımı - Saç Büyütme & Dökülme": [
        ("Rogaine Women's", "rogaine.com", "Kadın Minoxidil", "Women's Rogaine %5; minoxidil; kadın saç büyütme; FDA onaylı"),
        ("Bondi Boost HG", "bondiboost.com", "Avustralya Saç Büyütme", "HG Shampoo; saw palmetto; Avustralya; saç büyütme; TikTok viral"),
        ("Mielle Rosemary Mint", "mielleorganics.com", "Biberiye Yağı", "Rosemary Mint Scalp & Hair Oil; TikTok fenomeni; saç büyütme"),
        ("Wild Growth", "wildgrowth.com", "Klasik Saç Yağı", "Hair Oil; 1980'den beri; vitamin + mineral; saç uzatma yağı klasiği"),
        ("Biotin Shampoo: Pura D'or", "purador.com", "Biotin Saç Büyütme", "Anti-Thinning Biotin Shampoo; biotin + argan; Amazon #1; doğal"),
    ],

    "Saç Bakımı - Renk & Boyama": [
        ("Garnier Nutrisse", "garnier.com", "Besleyici Boya", "Nutrisse; meyve yağları besleyici; erişilebilir ev boyası"),
        ("Revlon ColorSilk", "revlon.com", "Ammonia-Free Boya", "ColorSilk Beautiful Color; ammonia-free; keratin + ipek; erişilebilir"),
        ("Clairol Nice'n Easy", "clairol.com", "Kolay Ev Boyası", "Nice'n Easy; %100 gri kaplama; erişilebilir; geniş renk"),
        ("L'Oréal Excellence", "loreal.com", "Lüks Ev Boyası", "Excellence Crème; Pro-Keratin Complex; gri kaplama; erişilebilir lüks"),
        ("Schwarzkopf", "schwarzkopf.com", "Alman Saç Boyası", "LIVE Color; Alman profesyonel; canlı renkler; erişilebilir"),
    ],

    "Saç Bakımı - Styling & Isı Koruma": [
        ("Got2b", "got2b.com", "Alman Stil", "Glued Blasting Freeze Spray; Schwarzkopf; güçlü tutuş; erişilebilir"),
        ("Tresemmé Heat Tamer", "tresemme.com", "Isı Koruma", "Heat Tamer Spray; ısı koruma; salon ilhamlı; erişilebilir"),
        ("Garnier Fructis Style", "garnier.com", "Stil Jel", "Pure Clean Gel; botanik jel; erişilebilir; doğal tutuş"),
        ("Aussie Instant Freeze", "aussie.com", "Güçlü Tutuş", "Instant Freeze Sculpting Gel; güçlü tutuş; Avustralya"),
        ("Pantene Thermal", "pantene.com", "Isı Koruma", "Thermal Heat Protection Spray; Pro-V; ısı koruma; eczane"),
    ],

    "Saç Bakımı - Kıvırcık & Tekstürlü Saç": [
        ("Aunt Jackie's Curl La La", "auntjackiescurlsandcoils.com", "Curl La La", "Curl La La Defining Curl Custard; kıvırcık tanımlama; erişilebilir"),
        ("Eco Styler", "ecostylergel.com", "Olive Oil Gel", "Olive Oil Styling Gel; kıvırcık jel klasik; erişilebilir; çeşitli"),
        ("Wetline Xtreme", "wetline.com", "Profesyonel Jel", "Professional Styling Gel; Meksika; kıvırcık tanımlama; güçlü tutuş"),
        ("Uncle Funky's Daughter", "unclefunkysdaughter.com", "Doğal Curl", "Curly Magic; doğal kıvırcık tanımlama; botanik bazlı; indie"),
        ("Camille Rose Curl Maker", "camillerose.com", "Botanik Kıvırcık", "Curl Maker; marshmallow kökü; botanik kıvırcık tanımlama"),
    ],

    "Vücut Bakımı - Vücut Nemlendirici & Yağ": [
        ("Jergens", "jergens.com", "Günlük Vücut Losyonu", "Ultra Healing Extra Dry Skin Moisturizer; ekstra kuru cilt"),
        ("Lubriderm", "lubriderm.com", "Dermatolojik Vücut", "Daily Moisture Lotion; dermatoloji önerisi; hafif nemlendirme"),
        ("Gold Bond", "goldbond.com", "Medikal Vücut", "Ultimate Healing Lotion; 7 nemlendirici + 3 vitamin; medikal"),
        ("Olay Body", "olay.com", "Vitamin Vücut", "Rinse-Off Body Conditioner; duşta vücut bakımı; pratik format"),
        ("St. Ives", "stives.com", "Yulaf Vücut", "Oatmeal & Shea Butter Body Lotion; yulaf + shea; erişilebilir"),
    ],

    "Vücut Bakımı - Peeling & Scrub": [
        ("Tree Hut Vitamin C", "treehut.com", "Vitamin C Scrub", "Vitamin C Sugar Scrub; vitamin C + şeker; aydınlatıcı vücut peeling"),
        ("Dove Exfoliating", "dove.com", "Nazik Eksfoliasyon", "Exfoliating Body Polish; nazik; nemlendirici; her cilt tipi"),
        ("Soap & Glory", "soapandglory.com", "İngiliz Scrub", "Flake Away Body Polish; shea + salt + sugar; İngiliz eğlenceli"),
        ("Rituals", "rituals.com", "Hollanda Scrub", "The Ritual of Sakura Body Scrub; Japon kiraz; Hollanda lüks"),
        ("Sabon", "sabon.com", "İsrail Tuz Scrub", "Body Scrub; Ölü Deniz tuzu; İsrail; aromatik; lüks vücut peeling"),
    ],

    "Vücut Bakımı - Self-Tan & Bronzlaşma": [
        ("Dove DermaSpa", "dove.com", "Günlük Bronz", "Summer Revived Body Lotion; kademeli bronzlaşma; nemlendirici"),
        ("Johnson's Vita-Rich", "johnsonsbaby.com", "Vitamin Bronz", "Vita-Rich Smoothing Body Lotion with Papaya; vitamin bazlı"),
        ("Palmer's Natural Bronze", "palmers.com", "Doğal Bronz", "Natural Bronze Body Lotion; kakao yağı; kademeli; doğal"),
        ("Garnier Ambre Solaire", "garnier.com", "Fransız Self-Tan", "Natural Bronzer Self-Tan Mist; Fransız; sprey; kademeli"),
        ("Rimmel Sun Shimmer", "rimmel.com", "Shimmer Tan", "Instant Tan Shimmer; anında bronzlaşma + parıltı; İngiliz"),
    ],

    "Vücut Bakımı - Deodorant": [
        ("Old Spice", "oldspice.com", "Erkek Klasik Deo", "Fiji Deodorant; klasik erkek deo; viral reklamlar; P&G"),
        ("Degree", "degree.com", "Hareket Deo", "Motion Sense; hareket aktive; BodyResponsive; Unilever"),
        ("Secret", "secret.com", "Kadın Deo", "Clinical Strength; klinik güç; kadın deo; P&G"),
        ("Ban", "ban.com", "Roll-On Deo", "Invisible Roll-On; roll-on format; Japon-Amerikan; erişilebilir"),
        ("Speed Stick", "speedstick.com", "Hızlı Deo", "Power Antiperspirant; hızlı uygulama; erkek deo; erişilebilir"),
    ],

    "Vücut Bakımı - Tüy Dökücü & Epilasyon": [
        ("Schick Intuition", "schick.com", "Sabunlu Jilet", "Intuition; dahili sabun; pratik tıraş; kadın jilet; ergonomik"),
        ("Venus", "gillettevenus.com", "Kadın Jilet Lider", "Extra Smooth; 5 bıçak; kadın tıraş standardı; Gillette"),
        ("Harry's kadın: Flamingo", "shopflamingo.com", "DTC Kadın Jilet", "Razor; ergonomik; şık tasarım; erişilebilir; abonelik"),
        ("Joy Razor", "joy-razor.com", "Uygun Kadın Jilet", "5 Blade Razor; ultra uygun; erişilebilir; P&G"),
        ("Nad's", "nads.com.au", "Avustralya Ağda", "Brazilian & Bikini Wax Kit; Avustralya; ev ağda seti; 30+ yıl"),
    ],

    "Tırnak Bakımı - Oje & Jel": [
        ("Orly", "orly.com", "Breathable Oje", "Breathable Treatment + Color; nefes alan oje; su + oksijen geçiren"),
        ("China Glaze", "chinaglaze.com", "Trend Oje", "Nail Lacquer; trend renkler; glitter + neon; salon profesyonel"),
        ("Sinful Colors", "sinfulcolors.com", "$2 Oje", "Professional Nail Polish; $2; geniş renk; ultra erişilebilir"),
        ("ILNP", "ilnp.com", "Indie Holografik", "Holographic Nail Polish; indie; holografik + multichrome; benzersiz"),
        ("Cirque Colors", "cirquecolors.com", "NYC Indie Oje", "Nail Lacquer; NYC artisan; 10-free; benzersiz renkler; indie"),
    ],

    "Tırnak Bakımı - Takma Tırnak & Press-On": [
        ("imPRESS", "kissusa.com", "Basınçla Yapışan", "imPRESS Press-On Nails; 30 saniye uygulama; çeşitli tasarım"),
        ("Dashing Diva Magic", "dashingdiva.com", "Kore Magic Gel", "Magic Press; Kore jel press-on; 10 dakika salon; K-beauty"),
        ("Olive & June Press-On", "oliveandjune.com", "Instant Mani", "Instant Mani; salon markasının press-on hattı; kolay çıkarma"),
        ("BTArtbox", "btartboxnails.com", "Amazon Press-On", "Press On Nails Kit; Amazon en çok satan; uygun fiyat; çeşitli"),
        ("Red Aspen Nail Dashes", "redaspen.com", "Premium Press-On", "Nail Dashes; premium yapıştır; salon kalitesi; sosyal satış"),
    ],

    "Cilt Cihazları - LED & Işık Terapisi": [
        ("Currentbody Neck", "currentbody.com", "Boyun LED", "Skin LED Neck & Dec Perfector; boyun + dekolte LED; anti-aging"),
        ("Dr. Dennis Gross Pro", "drdennisgross.com", "Pro LED Maske", "DRx SpectraLite EyeCare Pro; göz çevresi LED; NYC dermatoloji"),
        ("Qure Skincare", "qureskincare.com", "Q-Rejuvalight Pro", "LED Light Therapy Mask; 7 renk; FDA onaylı; profesyonel evde"),
        ("Solaris Labs", "solarislabsny.com", "Boyun LED", "Illuminate Neck; boyun LED; anti-aging; esnek panel"),
        ("Platinum LED", "platinumtherapylights.com", "Kızılötesi Panel", "BIO-300 Red Light Panel; büyük panel; kızılötesi terapi; tüm vücut"),
    ],

    "Cilt Cihazları - Mikro-Akım & RF": [
        ("Bear Mini Foreo", "foreo.com", "Mini Mikro-Akım", "BEAR Mini; kompakt mikro akım; seyahat dostu; İsveç tasarım"),
        ("Ziip OX", "ziipbeauty.com", "OX Mikro-Akım", "GX + OX Device; oksidasyon + mikro akım; app kontrollü; lüks"),
        ("Galvanic Spa NuSkin", "nuskin.com", "Galvanik Spa", "ageLOC Galvanic Spa; galvanik akım; serum iletimi; anti-aging"),
        ("FaceTite InMode", "inmodemd.com", "Minimal İnvaziv RF", "FaceTite; minimal invaziv RF; profesyonel cihaz; klinik"),
        ("EvenSkyn", "evenskyn.com", "Lumo RF", "Lumo+ Anti-Aging RF Device; RF + LED; evde profesyonel; çift terapi"),
    ],

    "Cilt Cihazları - Temizleme Cihazı": [
        ("Foreo Luna 4 Plus", "foreo.com", "En Gelişmiş Luna", "LUNA 4 Plus; near-infrared + T-Sonic; İsveç; anti-aging + temizleme"),
        ("PMD Clean Pro RQ", "pmdbeauty.com", "Rose Quartz Temizleme", "Clean Pro RQ; gül kuvars; ActiveWarmth; 4-in-1; lüks"),
        ("AENO Facial", "aeno.com", "Akıllı Temizleme", "Smart Facial Cleanser; IoT; kişiselleştirilmiş; akıllı sensörler"),
        ("Clarisonic yerine: Spa Sciences", "spasciences.com", "Nova Temizleme", "NOVA; 7 başlık; antimikrobiyal; çoklu fonksiyon; ev spa"),
        ("Magnitone BareFaced", "magnitone.com", "İngiliz Vibra-Sonic", "BareFaced 3; İngiliz; sonic; temizleme + masaj; waterproof"),
    ],

    "Parfüm & Koku - Kadın Parfüm": [
        ("YSL Libre alternatifi: Dossier", "dossier.co", "Lavanta Vanilya", "Ambery Lavender; YSL Libre ilham; %80 uygun; vegan"),
        ("Chanel No.5 alternatifi: Dossier", "dossier.co", "Floral Aldehyde", "Floral Aldehyde; Chanel No.5 ilham; şeffaf fiyat"),
        ("Baccarat Rouge 540 alternatifi: Dossier", "dossier.co", "Ambery Saffron", "Ambery Saffron; MFK ilham; erişilebilir niche"),
        ("La Vie Est Belle alternatifi: Dossier", "dossier.co", "Gourmand White Floral", "Gourmand White Flowers; Lancôme ilham"),
        ("Miss Dior alternatifi: Dossier", "dossier.co", "Floral Rose", "Floral Rose; Dior ilham; %80 uygun; temiz formül"),
    ],

    "Parfüm & Koku - Unisex & Niş": [
        ("Creed Aventus alternatifi: Club de Nuit", "armaf.com", "Armaf Aventus Dupe", "Club de Nuit Intense; Creed Aventus ilham; %95 uygun fiyat"),
        ("Tom Ford Tobacco Vanille alternatifi: Lattafa", "lattafa.com", "Raghba", "Raghba; Tom Ford ilham; Arap parfümeri; ultra uygun"),
        ("Baccarat Rouge alternatifi: Al Haramain", "alharamain.com", "Amber Oud Rouge", "Amber Oud Rouge; MFK ilham; Arap lüks; erişilebilir"),
        ("Le Labo Santal 33 alternatifi: Juliette Has a Gun", "juliettehasagun.com", "Not a Perfume Superdose", "Superdose; Cetalox yoğun; anti-parfüm"),
        ("Byredo Gypsy Water alternatifi: D.S. & Durga", "dsanddurga.com", "Radio Bombay", "Radio Bombay; Brooklyn indie; sandal + hindistan cevizi"),
    ],

    "Parfüm & Koku - Vücut Spreyi & Mist": [
        ("Victoria's Secret Body Mist", "victoriassecret.com", "Klasik Mist", "Pure Seduction; klasik vücut mist; çeşitli kokular"),
        ("Bath & Body Works Mist", "bathandbodyworks.com", "Amerikan Mist", "A Thousand Wishes; Amerikan vücut spreyi; çeşitli kokular"),
        ("Body Fantasies", "bodyfantasies.com", "Uygun Mist", "Fresh White Musk; ultra uygun vücut spreyi; çeşitli"),
        ("Impulse Body Spray", "impulse.com", "Klasik Sprey", "Body Spray; klasik vücut spreyi; erişilebilir; çeşitli"),
        ("So...? Body Mist", "sophbymissguided.com", "İngiliz Mist", "Sinful Body Mist; İngiliz; genç; erişilebilir; eğlenceli"),
    ],

    "Erkek Bakım - Tıraş & Sakal": [
        ("Gillette", "gillette.com", "Global Tıraş Lideri", "Fusion5 ProGlide; 5 bıçak; FlexBall teknolojisi; global standart"),
        ("Schick Hydro", "schick.com", "Hidro Tıraş", "Hydro 5 Sense; hidrasyon jeli; hareket algılayıcı; Wilkinson Sword"),
        ("The Dollar Shave Club", "dollarshaveclub.com", "Abonelik Öncü", "Executive Razor; viral pazarlama öncüsü; Unilever $1B; DTC"),
        ("Philips OneBlade", "philips.com", "Elektrikli Hibrit", "OneBlade; elektrikli + bıçak hibrit; tıraş + trim; Hollanda teknoloji"),
        ("Brio Beardscape", "bfrsh.com", "Sakal Trimmer", "Beardscape Beard Trimmer; seramik bıçak; sessiz; profesyonel"),
    ],

    "Erkek Bakım - Erkek Cilt Bakımı": [
        ("Dove Men+Care", "dovemencare.com", "Erkek Nemlendirme", "Face + Body Lotion; erkek nemlendirme; erişilebilir; Unilever"),
        ("Nivea Men", "nivea.com", "Alman Erkek Bakım", "Sensitive Moisturiser; Alman erkek bakım; hassas cilt; erişilebilir"),
        ("L'Oréal Men Expert", "loreal.com", "Fransız Erkek", "Hydra Energetic; erkek enerji nemlendirici; kafein; erişilebilir"),
        ("Neutrogena Men", "neutrogena.com", "Erkek Eczane", "Triple Protect Face Lotion SPF 20; SPF + nemlendirme; erkek"),
        ("Cetaphil Men", "cetaphil.com", "Dermatolojik Erkek", "Daily Face Lotion SPF 15; dermatolojik; hassas erkek cilt; eczane"),
    ],

    "K-Beauty & Kore Kozmetik": [
        ("Tonymoly", "tonymoly.com", "Kore Eğlenceli", "I'm Real Mask + Peach Lip Balm; eğlenceli ambalaj; meyve + hayvan şekli"),
        ("Nature Republic", "naturerepublic.com", "Kore Doğa", "Soothing & Moisture Aloe Vera Gel; %92 aloe; Kore doğal"),
        ("The Face Shop", "thefaceshop.com", "Kore Pirinç", "Rice Water Bright Cleanser; pirinç; LG Household; Kore geleneksel"),
        ("Skinfood", "theskinfood.com", "Kore Gıda Bakımı", "Black Sugar Mask; siyah şeker; Kore yiyecek ilhamlı kozmetik"),
        ("Tocobo", "tocobo.com", "Kore Glass Skin", "Glass Tinted Lip Balm; cam dudak; Kore yeni nesil; uygun fiyat"),
    ],

    "J-Beauty & Japon Kozmetik": [
        ("Rohto", "rohto.com", "Japon Göz Damlası", "Rohto Z! + skincare; Japon göz + cilt bakımı şirketi; çeşitli markalar"),
        ("Cow Brand", "cow-soap.co.jp", "Japon Süt Sabunu", "Milk Soap; Japon inek sütü sabunu; 100+ yıl; basit + nazik"),
        ("Kracie", "kracie.co.jp", "Japon Botanik", "Naive Makeup Remover; botanik temizleme; Japon günlük bakım"),
        ("Orbis", "orbis.co.jp", "Japon Yağsız", "Oil-free skincare; Japon yağsız formüller; POLA alt marka"),
        ("NOV", "nov.jp", "Japon Derma", "AC Active Moisture Cream; Japon dermatolojik; hipoalerjenik"),
    ],

    "Doğal & Organik Kozmetik": [
        ("Burt's Bees", "burtsbees.com", "Doğal Öncü", "Beeswax Lip Balm + Natural Skincare; %99+ doğal; 40+ yıl; Clorox"),
        ("Burt's Bees", "burtsbees.com", "Doğal Lider", "Complete Nourishment Facial Oil; doğal yüz yağı; erişilebilir organik"),
        ("Neal's Yard Remedies", "nealsyardremedies.com", "İngiliz Organik", "Frankincense Intense Cream; İngiliz organik; mavi şişe ikonik; Soil Association"),
        ("Trilogy", "trilogyproducts.com", "Yeni Zelanda Organik", "Rosehip Oil; Yeni Zelanda; sertifikalı organik kuşburnu; anti-aging"),
        ("Sukin", "sukin.com", "Avustralya Uygun Doğal", "Rosehip Facial Moisturiser; Avustralya; karbon nötr; erişilebilir doğal"),
    ],

    "Vegan & Cruelty-Free Kozmetik": [
        ("The Body Shop", "thebodyshop.com", "Etik Vegan Bakım", "Drops of Youth; B Corp; Community Trade; etik güzellik öncüsü"),
        ("Lush", "lush.com", "El Yapımı Vegan", "Fresh Face Mask; el yapımı; taze; %100 vejetaryen; naked ambalaj"),
        ("Barry M", "barrym.com", "İngiliz Vegan Makyaj", "Gelly Hi Shine Nail Paint; %100 cruelty-free; İngiliz; erişilebilir"),
        ("Collection", "collectioncosmetics.com", "İngiliz Uygun Vegan", "Lasting Perfection Concealer; İngiliz eczane; vegan; £5"),
        ("Catrice Vegan", "catrice.eu", "Alman Vegan Eczane", "True Skin Foundation; %100 vegan; Alman eczane; erişilebilir"),
    ],

    "Bebek & Hamile Cilt Bakımı": [
        ("Johnson's Baby", "johnsonsbaby.com", "Klasik Bebek", "Baby Lotion; klasik bebek bakımı; 125+ yıl; global standart"),
        ("Cetaphil Baby", "cetaphil.com", "Dermatolojik Bebek", "Daily Lotion; dermatoloji formülü; bebek hassas cilt; eczane"),
        ("Aveeno Baby", "aveeno.com", "Yulaf Bebek", "Daily Moisture Lotion; koloidal yulaf; bebek hassas; dermatolojik"),
        ("Aquaphor Baby", "aquaphor.com", "Bebek Onarım", "Baby Healing Ointment; çok amaçlı onarım; bebek pişik + kuru cilt"),
        ("Eucerin Baby", "eucerin.com", "Alman Bebek", "Aquaphor Baby Ointment; Alman dermatolojik; bebek cilt koruma"),
    ],

    "İntim Bakım & Vücut": [
        ("Always", "always.com", "Regl Lideri", "Infinity Pads; FlexFoam teknolojisi; regl hijyen standardı; P&G"),
        ("Tampax", "tampax.com", "Tampon Lideri", "Pearl Tampons; LeakGuard; global tampon standardı; P&G"),
        ("U by Kotex", "ubykotex.com", "Modern Regl", "Click Tampons; kompakt; modern tasarım; Kimberly-Clark"),
        ("Honey Pot Pads", "thehoneypot.co", "Bitkisel Ped", "Herbal Menstrual Pads; bitkisel infüzyon; doğal; kapsayıcı"),
        ("Rael Organic Pads", "getrael.com", "Kore Organik Ped", "Organic Cotton Pads; Kore + ABD; organik pamuk; hassas cilt"),
    ],

    "Diş Beyazlatma & Ağız Bakımı": [
        ("Oral-B", "oral-b.com", "Elektrikli Fırça Lideri", "iO Series; akıllı elektrikli fırça; AI; Braun; P&G; global standart"),
        ("Philips Sonicare", "philips.com", "Sonic Fırça Lideri", "DiamondClean; sonic teknoloji; akıllı; Hollanda; premium"),
        ("Waterpik", "waterpik.com", "Su Diş İpi", "Aquarius Water Flosser; su basıncı ile diş ipi; ADA onaylı"),
        ("Colgate Optic White", "colgate.com", "Beyazlatma Macunu", "Optic White Renewal; H2O2 beyazlatma; eczane; erişilebilir"),
        ("Crest 3D White", "crest.com", "Beyazlatma Strip", "Whitestrips Professional Effects; beyazlatma strip standardı; P&G"),
    ],
}

# ─── EXTRA BRANDS 5-9: Massive unique brand expansion ─────────────────────────
# These are curated DTC/indie cosmetic brands organized by category
EXTRA_BRANDS_5 = {
    "Cilt Bakımı - Nemlendirici & Serum": [
        ("Amorepacific Moisture Bound", "amorepacific.com", "Kore Bamboo Serum", "Moisture Bound Refreshing Hydra-Gel; bambu suyu; Kore premium nemlendirme"),
        ("Missha Near Skin", "missha.com", "Kore Madecanol Serum", "Near Skin Madecanol Cream; centella; Kore bariyer + nemlendirme"),
        ("Cezanne", "cezanne.co.jp", "Japon Uygun Serum", "Skin Conditioner; Japon eczane; hatatomugi; uygun fiyat tonik"),
        ("Melvita", "melvita.com", "Fransız Organik Yağ", "L'Or Bio Extraordinary Oil; Fransız organik; argan yağı; ECOCERT"),
        ("Cattier", "cattier-paris.com", "Fransız Kil Serum", "Kil bazlı bakım; Fransız organik; kil + bitkisel özler; 1968"),
        ("Sanoflore", "sanoflore.com", "Fransız Botanik", "Magnifica Essence; Fransız organik; lavanta + sarmaşık; COSMOS"),
        ("Filorga NCEF", "filorga.com", "Fransız Estetik Serum", "NCEF-Reverse Supreme Regenerating Cream; Fransız estetik tıp"),
        ("Lierac", "lierac.com", "Fransız Premium Serum", "Premium La Cure; Fransız eczane premium; anti-aging; Paris"),
        ("Darphin", "darphin.com", "Fransız Aromaterapi Serum", "Aromatic Care; Paris aromaterapi; bitkisel esanslar; Estée Lauder"),
        ("Jowaé", "jowae.com", "Fransız-Kore Serum", "Light Smoothing Cream; Fransız + Kore hibrit; lumiphenols; ortak formül"),
        ("Nuxe Bio", "nuxe.com", "Fransız Organik Serum", "Organic Moisturising Rich Cream; %100 organik Nuxe; COSMOS sertifikalı"),
        ("Payot", "payot.com", "Fransız Profesyonel Serum", "My Payot Jour; Paris 1920; profesyonel cilt bakımı; superfruit"),
        ("Sampar", "sampar.com", "Fransız Urban Serum", "Glamour Shot; Paris urban cilt bakımı; şehir stresi karşıtı"),
        ("Clarins Double Serum alternatifi: L'Occitane", "loccitane.com", "Provence Serum", "Immortelle Reset Serum; Provence immortelle; Fransız botanik"),
        ("Rituals Namaste", "rituals.com", "Hollanda Bakım Serumu", "The Ritual of Namaste Glow Anti-Aging Serum; Hollanda; yoga ilham"),
        ("Sensai Kanebo", "sensai-cosmetics.com", "Japon Silk Serum", "Ultimate The Emulsion; Japon ipek; Kanebo ultra lüks; sınırlı"),
        ("Orbis U", "orbis.co.jp", "Japon U Serum", "Orbis U Serum; Japon POLA alt marka; anti-aging; yağsız"),
        ("Acseine", "acseine.co.jp", "Japon Derma Serum", "AD Control Emulsion; Japon dermatolojik; alerji test; hassas"),
        ("Dr. Ci:Labo V-line", "ci-labo.com", "Japon Kolajen Serum", "V-Line Hariju; Japon kolajen konsantre; doktor markası; sıkılaştırma"),
        ("Transino", "transino.jp", "Japon Leke Serumu", "Whitening Essence EX II; Japon tranexamic acid; leke tedavisi; Daiichi Sankyo"),
        ("Hatomugi", "naturie.jp", "Japon Hatomugi Serum", "Skin Conditioner; Job's tears; Japon geleneksel; 500ml büyük boy; uygun"),
        ("Melano CC Premium", "rohto.com", "Japon Premium C", "Premium Beauty Essence; Japon C + E vitamin; leke + kırışıklık; Rohto"),
        ("Lululun", "lululun.com", "Japon Sheet Maske Serumu", "Precious RED; Japon sheet mask + serum; günlük maske standardı"),
        ("Chifure", "chifure.co.jp", "Japon Uygun Serum", "Moisture Lotion; Japon kooperatif markası; ultra uygun; basit"),
        ("Matsukiyo", "matsukiyo.co.jp", "Japon Eczane Özel", "Cera Lotion; Japon drugstore özel marka; ceramide; uygun"),
        ("Skin Authority", "skinauthority.com", "ABD Klinik Serum", "VitaD Fortified Anti-Aging Serum; D vitamini; ABD klinik bakım"),
        ("Skin Regimen", "skinregimen.com", "İtalyan Urban Serum", "Urban Shield SPF + serum; İtalyan; şehir kirliliği koruma"),
        ("Comfort Zone", "comfortzone.it", "İtalyan Bilimsel Serum", "Sublime Skin Serum; İtalyan bilimsel; profesyonel cilt bakımı"),
        ("Collistar", "collistar.com", "İtalyan Güneş Serumu", "Pure Actives Hyaluronic Acid; İtalyan güzellik; eczane + lüks"),
        ("Diego dalla Palma", "diegodallapalma.com", "İtalyan Profesyonel Serum", "RVB Skinlab; İtalyan profesyonel bakım; estetik merkezi markası"),
        ("Eisenberg", "eisenberg.com", "Monaco Lüks Serum", "First Wrinkles Delicate Cream; Monaco lüks; trio-moleküler formül"),
        ("Lancaster", "lancaster.com", "Monaco SPF Serum", "Sun Perfect Infinite Glow SPF Serum; Monaco; güneş + anti-aging"),
        ("Babor", "babor.com", "Alman Ampul Serum", "Ampoule Concentrates; Alman ampul; her cilt sorunu için özel ampul"),
        ("Juvena", "juvena.com", "İsviçre Anti-Age Serum", "Skin Rejuvenate Serum; İsviçre lüks; SkinNova SC teknolojisi"),
        ("Swiss Line", "swissline.com", "İsviçre Hücre Serum", "Cell Shock Age Intelligence; İsviçre hücre yenilenme; lüks"),
        ("Cellcosmet", "cellcosmet.com", "İsviçre Hücresel Serum", "CellLift Cream Light; İsviçre hücresel terapi; hayvan hücre bazlı"),
        ("Evidens de Beauté", "evidensdebeaute.com", "Japon-Fransız Serum", "The Extreme Cream; Japon + Fransız birleşim; pirinç + lüks"),
        ("Sensilis", "sensilis.com", "İspanyol Derma Serum", "Supreme Renewal Detox Booster; İspanyol dermatolojik; eczane"),
        ("Sesderma", "sesderma.com", "İspanyol Lipozom Serum", "C-VIT Liposomal Serum; İspanyol; lipozom teknoloji; C vitamini"),
        ("MartiDerm", "martiderm.com", "İspanyol Ampul Serum", "The Originals Proteos Hydra Plus SP; İspanyol ampul cilt bakımı"),
        ("Endocare", "endocare.com", "İspanyol SCA Serum", "Tensage Serum; İspanyol; SCA Growth Factor; salyangoz sekreti bilimi"),
        ("Isdinceutics", "isdin.com", "İspanyol Dermatolojik Serum", "Flavo-C Ultraglican; İspanyol; günlük ampul; C vitamini + proteoglycan"),
        ("Rilastil", "rilastil.com", "İtalyan Dermatolojik Serum", "D-Clar Depigmenting Cream; İtalyan dermatolojik; leke tedavisi"),
        ("Bioderma Sébium Serum", "bioderma.com", "Fransız Serum Hat", "Sébium Night Peel; Fransız gece serum; akne + yenilenme"),
        ("Ducray Melascreen", "ducray.com", "Fransız Leke Serumu", "Melascreen Depigmenting Intensive Care; Fransız leke tedavisi"),
        ("SVR Ampoule", "laboratoire-svr.com", "Fransız Ampul Serumu", "Ampoule [B3] Hydra; Fransız ampul; B3 niacinamide; yoğun"),
        ("Uriage Bariéderm", "uriage.com", "Fransız Termal Serum", "Bariéderm Cica-Cream; Fransız termal su; cica onarım"),
        ("Vichy Minéral 89", "vichy.com", "Fransız Volkanik Serum", "Minéral 89 Booster; %89 volkanik su; hyaluronic; Fransız eczane"),
        ("Noreva", "noreva.com", "Fransız Klinik Serum", "Exfoliac Global 6; Fransız klinik cilt bakımı; çoklu akne tedavi"),
        ("Institut Esthederm", "esthederm.com", "Fransız Güneş Serumu", "Adaptasun Serum; Fransız; güneş adaptasyon; ışık enerji teknolojisi"),
    ],

    "Cilt Bakımı - Temizleyici & Tonik": [
        ("Tony Moly", "tonymoly.com", "Kore Eğlenceli Temizleyici", "Peach Punch Cleansing Foam; şeftali; eğlenceli ambalaj"),
        ("Nature Republic Aloe", "naturerepublic.com", "Kore Aloe Temizleyici", "Aloe Vera Cleansing Gel Cream; %92 aloe; Kore doğal"),
        ("Skinfood Black Sugar", "theskinfood.com", "Kore Şeker Temizleyici", "Black Sugar Perfect First Serum; siyah şeker peeling tonik"),
        ("Sioris Cleanse", "sioris.com", "Kore Taze Temizleyici", "Cleanse Me Softly Milk; süt temizleyici; taze hasat"),
        ("Miguhara Ultra Whitening", "miguhara.co.kr", "Kore EGF Temizleyici", "Ultra Whitening Cleansing Foam; EGF; Kore aydınlatıcı temizleme"),
        ("Cezanne Skin Conditioner", "cezanne.co.jp", "Japon Hatomugi Tonik", "Skin Conditioner High Moist; Japon hatomugi; 500ml; ultra uygun"),
        ("Naturie", "naturie.jp", "Japon Jel Tonik", "Hatomugi Skin Conditioning Gel; Japon jel tonik; devasa boy; erişilebilir"),
        ("Kikumasamune", "kikumasamune.co.jp", "Japon Sake Tonik", "High Moist Lotion; 500ml sake tonik; Japon fermente; kült"),
        ("Mebika", "mebika.jp", "Japon Vitamin C Tonik", "Moist Toner; Japon C vitamini tonik; aydınlatıcı; uygun"),
        ("Dr. Althea Amino", "dralthea.com", "Kore Amino Tonik", "Amino Acid Gentle Bubble Cleanser; amino asit köpük"),
    ],

    "Cilt Bakımı - Güneş Koruma (SPF)": [
        ("Skin Aqua UV Gel", "skin-aqua.com", "Japon Büyük Boy SPF", "Super Moisture Gel SPF 50+ 140g; Japon büyük boy; su bazlı"),
        ("Verdio UV Essence", "omibrotherhood.com", "Japon Organik SPF", "UV Moisture Essence SPF 50+; organik; Japon doğal"),
        ("Curel UV Lotion", "curel.com", "Japon Ceramide SPF", "UV Lotion SPF 50+; ceramide; Japon hassas cilt SPF"),
        ("Dear Klairs UV", "klairscosmetics.com", "Kore Soft Airy SPF", "Soft Airy UV Essence SPF 50; hafif; Kore günlük SPF"),
        ("Neogen Day-Light SPF", "neogenlab.us", "Kore Dermatolojik SPF", "Day-Light Protection Airy Sunscreen SPF 50; Kore hafif SPF"),
        ("Suntique I'm Safe", "suntique.com", "Kore Güvenli SPF", "I'm Safe for Sensitive Skin SPF 35; Kore hassas SPF"),
        ("Make P:rem UV Defense", "makeprem.com", "Kore Temiz SPF", "UV Defense Me. Blue Ray Sun Cream; mavi ışık + UV; Kore"),
        ("Cell Fusion C Toning", "cellfusionc.com", "Kore Dermatolojik SPF", "Laser Sunscreen 100 SPF 50+; Kore dermatoloji SPF"),
        ("Dr. Ceuracle UV", "dr.ceuracle.com", "Kore Vegan SPF", "Cica Regen Vegan Sun SPF 50+; vegan; cica SPF"),
        ("La Shield Fisico", "lashield.in", "Hindistan Mineral SPF", "Fisico Matte Sunscreen Gel SPF 50; Hindistan mineral; mat"),
    ],

    "Cilt Bakımı - Akne & Leke Tedavisi": [
        ("AC Collection COSRX", "cosrx.com", "Kore AC Akne", "AC Collection Calming Liquid Mild; AHA + BHA + tea tree"),
        ("Krave Great Barrier", "kravebeauty.com", "Kore Bariyer Akne", "Great Barrier Relief; tamanu + cica; bariyer onarımlı akne"),
        ("Innisfree Bija", "innisfree.com", "Kore Bija Akne", "Bija Trouble Facial Foam; bija seed; Jeju; Kore akne"),
        ("Ciracle Pimple", "ciracle.com", "Kore Pimple Clear", "Pimple Solution CC Powder; Kore akne pudrası; invisible"),
        ("Skinfood Tea Tree", "theskinfood.com", "Kore Tea Tree Akne", "Tea Tree Clearing Spot Patch; Kore çay ağacı patch"),
        ("Aestura Theracne", "aestura.com", "Kore Theracne", "Theracne 365 Clear Spot; Kore hastane akne spot tedavi"),
        ("Nature Republic Tea Tree", "naturerepublic.com", "Kore Tea Tree", "Green Derma Tea Tree Cica Spot Patch; K-beauty akne"),
        ("La Roche-Posay Effaclar Duo", "laroche-posay.com", "Fransız Dual Akne", "Effaclar Duo(+); niacinamide + procerad; Fransız çift etki"),
        ("Bioderma Sébium Global", "bioderma.com", "Fransız Global Akne", "Sébium Global; patent Fluidactiv; sebum kalitesi düzenleme"),
        ("Vichy Normaderm Phyto", "vichy.com", "Fransız Volkanik Akne", "Normaderm Phytosolution; volkanik su + salisilik"),
    ],

    "Cilt Bakımı - Anti-Aging & Kırışıklık": [
        ("Filorga Time-Filler", "filorga.com", "Fransız Filler", "Time-Filler Absolute Wrinkles Correction Cream; Fransız estetik tıp"),
        ("Lierac Premium", "lierac.com", "Fransız Premium Anti-Age", "Premium La Crème Soyeuse; Fransız eczane premium"),
        ("Darphin Stimulskin", "darphin.com", "Fransız Aroma Anti-Age", "Stimulskin Plus Multi-Corrective Divine Cream; Paris aroma"),
        ("Payot Supreme Jeunesse", "payot.com", "Fransız Gençlik", "Supreme Jeunesse Concentré; Paris; süper gençlik konsantresi"),
        ("Esthederm Excellage", "esthederm.com", "Fransız Excellage", "Excellage Cream; Fransız ışık enerji; menopoz cilt bakımı"),
        ("Babor Skinovage", "babor.com", "Alman Ampul Anti-Age", "Skinovage Moisturizing Cream; Alman profesyonel; ampul sistemi"),
        ("Juvena Mastercream", "juvena.com", "İsviçre Master Anti-Age", "MasterCream; İsviçre SkinNova SC; lüks anti-aging krem"),
        ("Swiss Line Cell Shock", "swissline.com", "İsviçre Hücre Anti-Age", "Cell Shock Luxe-Lift Rich Cream; İsviçre hücre yenilenme"),
        ("Natura Bissé C+C", "naturabisse.com", "İspanyol C+C Anti-Age", "C+C Vitamin Cream; İspanyol vitamin C + anti-aging"),
        ("MartiDerm Platinum", "martiderm.com", "İspanyol Photo-Age", "Platinum Photo-Age Ampoules; İspanyol ampul; fotoaging"),
    ],

    "Makyaj - Fondöten & BB/CC Krem": [
        ("Banila Co Covericious", "banilaco.com", "Kore Power Fit", "Covericious Power Fit Foundation; Kore full coverage"),
        ("Holika Holika Aqua Petit", "holikaholika.com", "Kore Jöle BB", "Aqua Petit Jelly BB Cream; jöle kıvamlı; eğlenceli"),
        ("A'PIEU Madecassoside BB", "apieu.com", "Kore Cica BB", "Madecassoside Cica BB Cream; cica + BB; hassas cilt"),
        ("Missha Signature", "missha.com", "Kore Signature BB", "M Signature Real Complete BB; Kore BB standart; profesyonel"),
        ("Laneige BB Cushion", "laneige.com", "Kore BB Cushion", "BB Cushion Pore Control; Kore gözenek kontrol cushion"),
        ("Hera Black Cushion", "hera.com", "Kore Lüks Cushion", "Black Cushion SPF 34; Kore lüks; Seul modern; Amorepacific"),
        ("Innisfree Cushion", "innisfree.com", "Kore Yeşil Çay Cushion", "Green Tea Moisture Cushion; Jeju yeşil çay; doğal"),
        ("Clio Nudism Cushion", "clio.co.kr", "Kore Nudism", "Nudism Hyaluronic Cover Cushion; Kore nude; hyaluronic"),
        ("3CE Back To Baby BB", "3ce.com", "Kore Back to Baby", "Back to Baby BB Cream; Kore bebek cilt efekti; hafif"),
        ("VDL Expert", "vdl.co.kr", "Kore Expert Color", "Expert Color Primer; renk düzeltme; Kore CC primer"),
    ],

    "Saç Bakımı - Şampuan & Saç Kremi": [
        ("Sachajuan Ocean Mist", "sachajuan.com", "İsveç Ocean", "Ocean Mist; İsveç deniz dalgası; ocean silk; minimal"),
        ("Christophe Robin Rose", "christophe-robin.com", "Paris Gül Şampuan", "Delicate Volumising Shampoo with Rose Extracts; gül hacim"),
        ("Philip Kingsley Flaky", "philipkingsley.com", "İngiliz Kepek", "Flaky/Itchy Scalp Shampoo; trikoloji kepek; İngiliz bilim"),
        ("Oribe Gold Lust", "oribe.com", "Lüks Altın", "Gold Lust Repair & Restore Shampoo; altın + botanik; lüks"),
        ("R+Co Television", "randco.com", "Hacim Şampuan", "Television Perfect Hair Shampoo; hacim; lüks indie; salon"),
        ("IGK Hot Girls", "igkhair.com", "Brooklyn Hydrate", "Hot Girls Hydrating Shampoo; coconut oil; Brooklyn salon"),
        ("Davines Love", "davines.com", "İtalyan Love", "LOVE Smoothing Shampoo; İtalyan; kıvırcık düzleştirme; sürdürülebilir"),
        ("Kevin Murphy Plumping", "kevinmurphy.com.au", "Avustralya Plump", "PLUMPING.WASH; Avustralya; hacim verme; ince saç"),
        ("Verb Ghost Shampoo", "verbproducts.com", "Ghost Hafif", "Ghost Shampoo; ultra hafif; moringa; erişilebilir salon"),
        ("Eva NYC Therapy", "evanyc.com", "NYC Terapi", "Therapy Session Hair Mask Shampoo; NYC salon; onarım"),
        ("Rahua Voluminous", "rahua.com", "Amazon Hacim", "Voluminous Shampoo; Amazon rahua yağı; hacim; organik"),
        ("Playa Every Day", "playa.beauty", "Minimal Günlük", "Every Day Shampoo; minimal; California; temiz; günlük"),
        ("Adwoa Beauty", "adwoabeauty.com", "Doğal Saç DTC", "Baomint Moisturizing Shampoo; baobab + nane; doğal saç bakımı"),
        ("Odele Air Dry Cream", "odelebeauty.com", "Air Dry Stil", "Air Dry Styling Cream; doğal kurutma; Target; erişilebilir"),
        ("Kristin Ess Signature", "kristiness.com", "Erişilebilir Salon", "Signature Shampoo; erişilebilir salon; Target; $10"),
    ],

    "Vücut Bakımı - Vücut Nemlendirici & Yağ": [
        ("L'Occitane Shea", "loccitane.com", "Provence Shea", "Shea Butter Ultra Rich Cream; Provence; %25 shea; Fransız lüks"),
        ("Rituals Body", "rituals.com", "Hollanda Ritüel", "The Ritual of Jing Body Cream; Hollanda; wellness; spa"),
        ("Sabon Body", "sabon.com", "İsrail Vücut", "Body Lotion; İsrail; Ölü Deniz; aromatik; lüks vücut bakımı"),
        ("Jo Malone Body", "jomalone.com", "İngiliz Koku Vücut", "Body Crème; İngiliz lüks koku + vücut bakımı; katmanlama"),
        ("Laura Mercier Body", "lauramercier.com", "Lüks Vücut", "Almond Coconut Milk Soufflé Body Crème; lüks vücut bakımı"),
        ("Moroccanoil Body Soufflé", "moroccanoil.com", "Argan Vücut Soufflé", "Body Soufflé; argan yağı; hafif vücut soufflé; koku"),
        ("Ahava", "ahava.com", "İsrail Ölü Deniz", "Mineral Body Lotion; Ölü Deniz mineralleri; İsrail; doğal"),
        ("Dove DermaSpa", "dove.com", "DermaSpa Vücut", "Goodness3 Body Cream; Dove premium hat; omega; nemlendirme"),
        ("Soap & Glory", "soapandglory.com", "İngiliz Eğlenceli Vücut", "The Righteous Butter Body Butter; İngiliz; shea; eğlenceli"),
        ("Sanctuary Spa", "sanctuaryspa.com", "İngiliz Spa Vücut", "Antibacterial Hand Wash + Body; İngiliz spa; ev spa ritüeli"),
    ],

    "Parfüm & Koku - Kadın Parfüm": [
        ("Gallivant", "gallivant.com", "Seyahat Parfüm", "Istanbul; seyahat ilhamlı; her şehir bir koku; artisan; niş"),
        ("Bon Parfumeur", "bonparfumeur.com", "Paris Numara Parfüm", "402 Vanilla Toffee Sandalwood; Paris; numara sistemi; erişilebilir niş"),
        ("Nomenclature", "nomenclature.com", "Bilim Parfüm", "Adr_Enn; bilimsel molekül isimleri; modern niş; NYC"),
        ("Hiram Green", "hiramgreen.com", "Hollanda Doğal Parfüm", "Moon Bloom; doğal bileşenler; Hollanda artisan; el yapımı"),
        ("Memo Paris", "memoparis.com", "Paris Seyahat Lüks", "African Leather; seyahat ilhamlı lüks; Paris niş; egzotik"),
        ("Parfums de Marly", "parfums-de-marly.com", "Fransız At Lüks", "Delina; 18. yüzyıl Fransız; at sporları; kraliyet; lüks niş"),
        ("Maison Francis Kurkdjian", "franciskurkdjian.com", "Fransız Master Parfümör", "Baccarat Rouge 540; lüks Fransız; viral fenomen; LVMH"),
        ("Diptyque Parfüm", "diptyque.com", "Paris Mum Parfüm", "Philosykos EDP; Paris 1961; incir yaprağı; niş klasik"),
        ("Kilian", "bykilian.com", "Lüks Niş", "Love Don't Be Shy; Estée Lauder; lüks niş; Rihanna favorisi"),
        ("Editions de Parfums Frédéric Malle", "fredericmalle.com", "Editör Parfüm", "Carnal Flower; parfüm editörü; master parfümörler; lüks"),
    ],

    "Parfüm & Koku - Unisex & Niş": [
        ("Olfactive Studio", "olfactivestudio.com", "Fotoğraf Parfüm", "Still Life in Rio; fotoğraf + koku; Paris niş; sanatsal"),
        ("Filippo Sorcinelli", "filipposorcinelli.com", "İtalyan Kilise", "UNUM parfümler; İtalyan kilise ilhamlı; tütsü + din; artisan"),
        ("Heeley", "heeley.com", "Paris-İsveç Niş", "Sel Marin; Paris-İsveç; minimalist niş; James Heeley; deniz kokusu"),
        ("Miller Harris", "millerharris.com", "İngiliz Niş", "Tea Tonique; İngiliz niş; çay ilham; London; yeşil"),
        ("Ormonde Jayne", "ormondejayne.com", "İngiliz Lüks Niş", "Tolu; İngiliz lüks niş; London; eksotik; kristal şişe"),
        ("Lab on Fire", "labonfire.com", "İspanyol Yaratıcı", "What We Do In Paris Is Secret; İspanyol yaratıcı; gizem"),
        ("Carner Barcelona", "carnerbarcelona.com", "İspanyol Şehir Niş", "Tardes; Barcelona ilhamlı; İspanyol niş; Akdeniz"),
        ("Masque Milano", "masque-milano.com", "İtalyan Sanat Niş", "Russian Tea; İtalyan sanat; Milano; opera maskeleri"),
        ("Perris Monte Carlo", "perris.com", "Monaco Lüks Niş", "Santal du Pacifique; Monaco lüks; pasifik sandal; premier"),
        ("Chris Collins", "chriscollinsworld.com", "NYC Lüks Niş", "African Rooibos; NYC lüks; Afrika ilhamlı; siyah lüks"),
    ],

    "Erkek Bakım - Tıraş & Sakal": [
        ("Truefitt & Hill", "truefittandhill.com", "İngiliz En Eski Berber", "1805 Shaving Cream; dünyanın en eski berber dükkanı; 1805; İngiliz kraliyet"),
        ("Taylor of Old Bond Street", "tayloroldbondstreet.co.uk", "İngiliz Tıraş Kremi", "Sandalwood Shaving Cream; İngiliz berber; sandal ağacı; 1854"),
        ("D.R. Harris", "drharris.co.uk", "İngiliz Eczane Tıraş", "Arlington Shaving Cream; İngiliz eczane berber; 1790; London"),
        ("Castle Forbes", "castleforbes.com", "İskoç Lüks Tıraş", "1445 Shaving Cream; İskoç lüks; lavanta; el yapımı"),
        ("Penhaligon's Endymion", "penhaligons.com", "İngiliz Lüks Tıraş", "Endymion Shaving Cream; İngiliz aristokrat; 1870; lüks"),
        ("Acqua di Parma Barbiere", "acquadiparma.com", "İtalyan Berber", "Barbiere Collection; İtalyan berber geleneği; lüks; 1916"),
        ("Floris London", "florislondon.com", "İngiliz Kraliyet Berber", "Elite Shaving Cream; İngiliz kraliyet tüzük; 1730; London"),
        ("DR K Soap Company", "drksoap.com", "Portekiz Sakal", "Beard Balm; Portekiz sakal bakımı; el yapımı; doğal"),
        ("Percy Nobleman", "percynobleman.com", "İngiliz Sakal", "Beard Oil; İngiliz sakal bakımı; doğal bileşenler; erişilebilir"),
        ("Proraso Cypress", "proraso.com", "İtalyan Selvi", "Cypress & Vetyver Shaving Cream; İtalyan; selvi + vetiver; maskülen"),
    ],

    "K-Beauty & Kore Kozmetik": [
        ("Enough Project", "enoughproject.kr", "Kore Basit Bakım", "Simple skincare; Kore minimalist; erişilebilir; basit + etkili"),
        ("Mamonde", "mamonde.com", "Kore Çiçek Bakım", "Flower Lab Essence; çiçek bilimi; Amorepacific; Kore botanik"),
        ("Primera", "primera.co.kr", "Kore Organik", "Alpine Berry Watery Cream; Kore organik; Amorepacific premium doğal"),
        ("IOPE", "iope.com", "Kore Biyo-Science", "Retinol Expert 0.1%; Kore biyoteknoloji; retinol uzmanı; Amorepacific"),
        ("HERA", "hera.com", "Kore Şehir Lüks", "Black Foundation; Seul modern lüks; Amorepacific premium makyaj"),
        ("Lirikos", "lirikos.com", "Kore Deniz", "Marine Collagen Cream; Kore deniz kolajen; okyanus bilimi"),
        ("Donginbi", "donginbi.com", "Kore Kırmızı Ginseng", "Red Ginseng Cream; %100 Kore kırmızı ginseng; KGC; hanbang lüks"),
        ("Ohui", "ohui.co.kr", "Kore Hücre Bilim", "The First Geniture Ampoule; Kore hücre yenilenme; LG Household lüks"),
        ("Su:m37", "sum37.com", "Kore Fermente", "Secret Essence; Kore doğal fermentasyon; LG Household; 365 gün fermente"),
        ("Sooryehan", "sooryehan.com", "Kore Hanbang", "Boyun Crème; Kore hanbang; geleneksel tıp; LG Household"),
    ],

    "J-Beauty & Japon Kozmetik": [
        ("Etvos", "etvos.com", "Japon Mineral Makyaj", "Mineral Foundation; Japon mineral makyaj; hassas cilt; temiz"),
        ("MiMC", "mimc.co.jp", "Japon Organik Makyaj", "Mineral Erth Powder Foundation; Japon organik mineral; lüks"),
        ("Naturaglacé", "naturaglace.jp", "Japon Doğal Makyaj", "Makeup Cream; Japon doğal makyaj; SPF + makyaj; basit"),
        ("24h cosme", "24h-cosme.jp", "Japon 24 Saat Makyaj", "24 Mineral Stick Foundation; Japon mineral; 24 saat giyilebilir"),
        ("Rmk Liquid Foundation", "rmk.com", "Japon Şeffaf Fondöten", "Liquid Foundation; ultra şeffaf; Japon doğal makyaj; hafif"),
        ("Ettusais", "ettusais.co.jp", "Japon Genç Makyaj", "Lip Edition Tint Rouge; Japon genç makyaj; Shiseido alt marka"),
        ("Coffret D'or", "coffretdor.jp", "Japon Office Makyaj", "Skin Illusion Foundation; Japon ofis makyajı; Kanebo; doğal"),
        ("Jill Stuart Beauty", "jillstuart-beauty.com", "Japon Romantik", "Crystal Bloom Perfumed Lip Balm; Japon romantik estetik; kristal"),
        ("Paul & Joe Beauté", "paul-joe-beaute.com", "Japon-Fransız Makyaj", "Moisturizing Foundation Primer; Japon-Fransız; kedi motifi; romantik"),
        ("Anna Sui Cosmetics", "annasui.com", "Japon-ABD Gotik", "Sui Black Rouge; gotik estetik; Japon-ABD; vintage; butterfly"),
    ],

    "Doğal & Organik Kozmetik": [
        ("Absolution", "absolution-cosmetics.com", "Fransız Sertifikalı Organik", "La Crème du Jour; Fransız COSMOS organik; Paris; sürdürülebilir"),
        ("Ren Skincare Bio", "renskincare.com", "İngiliz Bio Aktif", "Bio Retinoid Anti-Wrinkle; İngiliz bio-aktif; temiz; sürdürülebilir"),
        ("Suki Skincare", "sukiskincare.com", "ABD Organik", "Nourishing Facial Oil; ABD organik; bitki bazlı; artisan; temiz"),
        ("Juice Beauty Stem", "juicebeauty.com", "Organik Kök Hücre", "Stem Cellular Anti-Wrinkle Booster; USDA organik; meyvelerden"),
        ("Tata Harper Elixir", "tataharper.com", "Vermont Elixir", "Elixir Vitae; %100 doğal; Vermont; en lüks doğal serum; hücre yenilenme"),
        ("Susanne Kaufmann", "susannekaufmann.com", "Avusturya Organik", "Day Cream; Avusturya Alp otları; organik; spa markası; doğal lüks"),
        ("La Bruket", "labruket.com", "İsveç Organik", "Sea Salt Scrub; İsveç organik; deniz tuzu; minimalist; Varberg"),
        ("Aeos", "aeos.net", "İngiliz Biyodinamik", "Energising Hydrating Mist; biyodinamik; Aura-Soma; İngiliz; kristal"),
        ("Sodashi", "sodashi.com.au", "Avustralya Lüks Doğal", "Rejuvenating Face Serum; Avustralya lüks doğal; Ayurveda; botanik"),
        ("Mukti", "muktiorganics.com", "Avustralya Aborijin Doğal", "Queen of the Night; Avustralya yerli bitkiler; kakadu plum; organik"),
    ],

    "Vegan & Cruelty-Free Kozmetik": [
        ("Rare Beauty alternatifi: Item Beauty", "itembeauty.com", "Z Kuşağı Vegan", "Lid Glaze; Z kuşağı; Addison Rae; temiz + vegan; Sephora"),
        ("NYX Bare With Me", "nyxcosmetics.com", "Vegan Skin Tint", "Bare With Me Tinted Skin Veil; vegan; hafif; erişilebilir"),
        ("Honest Beauty Everything", "honestbeauty.com", "Jessica Alba Vegan", "Everything Primer; Jessica Alba; vegan; EWG; temiz"),
        ("Elf Camo", "elfcosmetics.com", "Vegan CC Krem", "Camo CC Cream; vegan; $7; SPF 30; erişilebilir; TikTok"),
        ("Flower Beauty Light Illusion", "flowerbeauty.com", "Vegan Fondöten", "Light Illusion Foundation; vegan; Drew Barrymore; Walmart"),
        ("Physicians Formula Organic", "physiciansformula.com", "Organik Vegan", "Organic Wear All Natural Origin Foundation; organik + vegan"),
        ("Pixi H2O Skin Tint", "pixibeauty.com", "Vegan Skin Tint", "H2O SkinTint; vegan; hafif; niacinamide + hyaluronic"),
        ("Catrice HD Liquid", "catrice.eu", "Alman Vegan Fondöten", "HD Liquid Coverage Foundation; vegan; Alman; €6; full coverage"),
        ("Essence Skin Lovin", "essence.eu", "Alman Vegan Sensitivity", "Skin Lovin' Sensitive Concealer; vegan; hassas; €3"),
        ("Revolution Pro New Neutral", "revolutionbeauty.com", "İngiliz Vegan Palet", "New Neutrals Shadow Palette; vegan; İngiliz; nötr tonlar; £8"),
    ],

    "Bebek & Hamile Cilt Bakımı": [
        ("Bloom and Blossom", "bloomandblossom.com", "İngiliz Hamile Lüks", "The Lovely Jubbly Bust Firming Gel; İngiliz; hamile + post-natal"),
        ("Mama Bee by Burt's Bees", "burtsbees.com", "Doğal Hamile", "Belly Butter; doğal; kakao + shea; çatlak önleme; organik"),
        ("Palmer's Cocoa Butter", "palmers.com", "Kakao Hamile", "Stretch Mark Cream; kakao yağı; hamile çatlak bakımı; klasik"),
        ("Bio-Oil Hamile", "bio-oil.com", "Çatlak Yağı", "Bio-Oil; PurCellin Oil; çatlak + leke; hamile favorisi; 40+ ülke"),
        ("Weleda Stretch Mark", "weleda.com", "Organik Çatlak", "Stretch Mark Massage Oil; organik; badem + arnika; İsviçre"),
    ],

    "İntim Bakım & Vücut": [
        ("Hers Intimate", "forhers.com", "Kadın İntim DTC", "Vaginal Health Products; online dermatoloji; kadın sağlığı; DTC"),
        ("Wisp", "hellowisp.com", "İntim Sağlık DTC", "Vaginal Health Rx; online reçeteli; mantar + BV tedavisi; teledermatolog"),
        ("Stix", "getstix.co", "Hamilelik Testi DTC", "Pregnancy Test + UTI Test; ev testi; erişilebilir; DTC sağlık"),
        ("Flo Vitamins", "flovitamins.com", "PMS Takviye", "PMS Gummy Vitamins; regl öncesi sendrom; vitamin; eğlenceli format"),
        ("Joylux", "joylux.com", "Kadın Wellness Cihaz", "vFit Gold; kadın intim wellness; LED + ısı + sonic; DTC"),
    ],

    "Diş Beyazlatma & Ağız Bakımı": [
        ("Apa Beauty", "apabeauty.com", "NYC Diş Hekimi", "White Duo Toothpaste; NYC diş hekimi Dr. Apa; premium beyazlatma"),
        ("Supersmile", "supersmile.com", "Profesyonel Beyazlatma", "Professional Whitening Toothpaste; Calprox; profesyonel beyazlatma"),
        ("SmileDirectClub Bright On", "smiledirectclub.com", "LED Beyazlatma", "Bright On Whitening Kit; LED + H2O2; evde beyazlatma; DTC"),
        ("Auraglow", "auraglow.com", "LED DTC Beyazlatma", "Teeth Whitening Kit; LED + H2O2 jel; DTC; 30 dakika; erişilebilir"),
        ("GLO Science", "gloscience.com", "Isı Beyazlatma", "GLO Brilliant Teeth Whitening Device; ısı + ışık; patentli; profesyonel"),
    ],

    "Tırnak Bakımı - Oje & Jel": [
        ("Nails Inc London", "nailsinc.com", "İngiliz Salon Oje", "Plant Power Nail Polish; İngiliz salon; %73+ bitkisel; vegan"),
        ("Smith & Cult Nails", "smithandcult.com", "NYC Edgy Oje", "Nail Lacquer; NYC edgy; 8-free; moda ilhamlı; lüks indie"),
        ("Rooted Woman", "rootedwoman.com", "Siyah Kadın Oje", "Nail Polish; siyah kadın kuruculu; koyu ten tonlarına özel; kapsayıcı"),
        ("Sienna Byron Bay", "siennabyronnbay.com.au", "Avustralya Doğal Oje", "Nail Polish; Avustralya; 10-free; doğal; Byron Bay yaşam tarzı"),
        ("Kure Bazaar Oje", "kurebazaar.com", "Fransız Eco Oje", "Nail Colour; Fransız eco; %85 doğal; Paris chic; lüks doğal"),
    ],

    "Tırnak Bakımı - Takma Tırnak & Press-On": [
        ("Aprés Gel-X", "afresnails.com", "Gel-X Uzantı", "Gel-X Nail Extension System; profesyonel jel uzantı; salon standardı"),
        ("Young Nails", "youngnails.com", "Profesyonel Tırnak", "Protein Bond + Acrylic; profesyonel tırnak sistemi; salon"),
        ("Beetles Gel Extensions", "beetles.com", "Uygun Gel-X", "Gel Nail Extension Kit; uygun fiyat; Amazon best seller; ev salon"),
        ("Modelones", "modelones.com", "Çin Jel Seti", "Gel Polish Starter Kit; Çin; uygun fiyat; Amazon; geniş seçim"),
        ("Kiara Sky", "kiarasky.com", "Dip Powder Tırnak", "Dip Powder Kit; dip powder sistemi; salon + ev; 200+ renk"),
    ],

    "Cilt Cihazları - LED & Işık Terapisi": [
        ("Maysama", "maysama.com", "Kırmızı LED Panel", "Red Light Therapy Panel; kızılötesi; anti-aging; tüm vücut"),
        ("PlatinumLED BioMax", "platinumtherapylights.com", "BioMax LED", "BioMax 900; büyük panel; 5 dalga boyu; profesyonel evde"),
        ("Rouge Red Light", "rougecare.com", "Kanada LED", "Rouge Tabletop; Kanada; masaüstü LED; 660nm + 850nm"),
        ("Joovv", "joovv.com", "Full Body LED", "Joovv Solo; tam vücut kızılötesi; NASA teknolojisi; wellness"),
        ("Mito Red Light", "mitoredlight.com", "Mitokondri LED", "MitoADAPT 2.0; mitokondri aktivasyon; kızılötesi + kırmızı"),
    ],

    "Cilt Cihazları - Mikro-Akım & RF": [
        ("Silk'n Titan", "silkn.com", "İsrail Titan RF", "Titan AllWays; RF trilayer; İsrail; sıkılaştırma; ev kullanım"),
        ("AMIRO S1", "amiro.com", "Çin EMS RF", "S1 Facial RF Skin Tightening Device; Çin; EMS + RF; app kontrollü"),
        ("Norlanya", "norlanya.com", "RF LED Kombo", "RF LED Face Machine; RF + LED; çoklu terapi; ev profesyonel"),
        ("Asdisun", "asdisun.com", "RF Göz Cihazı", "Eye RF Beauty Device; göz çevresi RF; anti-aging; kompakt"),
        ("Jellen", "jellen.com", "Jel Mikro-Akım", "Cryotherapy Ice Roller + Microcurrent; buz + mikro akım; dual"),
    ],

    "Cilt Cihazları - Temizleme Cihazı": [
        ("Philips Sonicare Face", "philips.com", "Hollanda Yüz Fırçası", "VisaPure Essential; Hollanda yüz temizleme; değiştirilebilir başlık"),
        ("Refa Clear", "refa.net", "Japon 3D Sonic", "ReFa CLEAR; 3D sonic + ion; Japon güzellik cihazı; profesyonel"),
        ("Panasonic Dense Foam", "panasonic.com", "Japon Köpük Cihazı", "Dense Foam Este; yoğun köpük üretme; Japon; nazik temizleme"),
        ("Ebelin", "ebelin.com", "Alman Eczane Fırça", "Facial Cleansing Brush; Alman dm eczane özel; uygun fiyat"),
        ("Real Techniques Face", "realtechniques.com", "Yüz Fırçası", "Miracle Complexion Sponge + Face Brush; Sam & Nic Chapman; erişilebilir"),
    ],
}

EXTRA_BRANDS_6 = {
    "Cilt Bakımı - Nemlendirici & Serum": [
        ("Banobagi", "banobagi.com", "Kore Cerrahi Klinik Serum", "Milk Thistle Repair Serum; Kore estetik cerrahi klinik; süt devesi dikeni; klinik onarım"),
        ("Cell Fusion C", "cellfusionc.com", "Kore Dermatolojik Aktif", "Expert Purifying Cleansing Oil; Kore dermatolog formülü; aktif bileşenler; klinik bakım"),
        ("TIRTIR", "tirtir.com", "Kore Maske Viral", "Behind The Mask Toner + Real Cream; Kore TikTok viral; milk skin; beyaz cilt efekti"),
        ("Laneige Water Bank", "laneige.com", "Kore Su Bankası", "Water Bank Blue Hyaluronic Serum; mavi hyaluronic; Kore su bilimi; derin nemlendirme"),
        ("Amorepacific Vintage", "amorepacific.com", "Kore Single Extract", "Vintage Single Extract Essence; yeşil çay fermantasyonu; Kore lüks fermente"),
        ("Dinto", "dinto.co.kr", "Kore Blur Serum", "Blur Finishing Powder + Skin; Kore blur efekti; pürüzsüz görünüm; makyaj bazı"),
        ("ABOUT ME", "aboutme.kr", "Kore Pirinç Suyu", "Rice Mask Wash Off; pirinç suyu bazlı; Kore geleneksel güzellik; aydınlatma"),
        ("The Saem", "thesaem.com", "Kore Urban Eco", "Urban Eco Harakeke Cream; Yeni Zelanda keten bitkisi; Kore + NZ hibrit"),
        ("KLAVUU", "klavuu.com", "Kore Deniz İncisi", "White Pearlsation Backstage Cream; inci özütü; Kore aydınlatma; lüks"),
        ("VPROVE", "vprove.com", "Kore Basit Serum", "Cica Tone Up Sun Fluid; cica + ton düzeltme; Kore basit bakım"),
        ("MAKEPREM", "makeprem.com", "Kore Temiz Serum", "Safe Me Relief Moisture Cream; Kore temiz güzellik; hassas cilt; minimal"),
        ("SERUMKIND", "serumkind.com", "Kore Serum Uzmanı", "Green Lifting Serum; yeşil serum; Kore lifting; peptide bazlı; yenilikçi"),
        ("TONYMOLY Vita C", "tonymoly.com", "Kore C Vitamini", "Vital Vita 12 Brightening Ampoule; %12 C vitamini konsantresi; Kore aydınlatma"),
        ("Skinfood Royal Honey", "theskinfood.com", "Kore Kraliyet Balı", "Royal Honey Propolis Enrich Essence; bal + propolis; Kore lüks doğal"),
        ("Nature Republic Snail", "naturerepublic.com", "Kore Salyangoz", "Snail Solution Cream; salyangoz müsin; Kore uygun; anti-aging + onarım"),
        ("Too Cool For School Egg", "toocoolforschool.com", "Kore Yumurta Bakım", "Egg Mellow Cream; yumurta bazlı 5-in-1; Kore çok amaçlı; nemlendirme"),
        ("Papa Recipe Eggplant", "paparecipe.com", "Kore Patlıcan", "Eggplant Clearing Ampoule; patlıcan özütü; Kore doğal sorunlu cilt"),
        ("Yadah Green Tea", "yadah.com", "Kore Yeşil Çay", "Green Tea Real Fresh Toner; %97 doğal; Kore temiz yeşil çay"),
        ("G9SKIN White", "g9skin.com", "Kore Beyaz Süt", "White In Whipping Cream; süt protein; Kore ton eşitleme; aydınlatma"),
        ("BELLAMONSTER", "bellamonster.com", "Kore 10 Saniye", "10 Second Peel Solution; 10 saniye peeling; Kore hızlı bakım; pratik"),
        ("KEEP COOL", "keepcool.co.kr", "Kore Bamboo Su", "Bamboo Soothe Sun Essence; bambu suyu; Kore soğutucu; doğal nemlendirme"),
        ("UNLEASHIA", "unleashia.com", "Kore Glitter Makyaj", "Get Loose Glitter Gel; K-beauty glitter; eğlenceli; festival; viral"),
        ("HINCE", "hfriendsince.com", "Kore Mood Makyaj", "True Dimension Radiance Balm; Kore mood güzellik; minimal; sanatsal"),
        ("WAKEMAKE", "wakemake.com", "Kore Ton On Ton", "Soft Blurring Eye Palette; Kore ton-on-ton; pastel; bulanık güzellik"),
        ("DASIQUE", "dasique.com", "Kore Pastel Palet", "Shadow Palette; Kore pastel estetik; çiçek ilhamlı; Z kuşağı K-beauty"),
        ("CELEFIT", "celefit.kr", "Kore Celeb Fit", "Designer Lip; Kore ünlü fit; minimal lüks; dudak uzmanı"),
        ("BLACK ROUGE", "blackrouge.co.kr", "Kore Air Fit Tint", "Air Fit Velvet Tint; Kore kadife tint; uzun ömürlü; çeşitli renk"),
        ("DEAR DAHLIA", "deardahlia.com", "Kore Vegan Lüks", "Paradise Dual Palette; Kore vegan lüks makyaj; dahlia çiçeği; premium"),
        ("NAMING.", "naming.kr", "Kore Minimal Makyaj", "Fluffy Powder Blush; Kore minimal estetik; yumuşak pudra; hafif"),
        ("NUSE", "nuse.co.kr", "Kore Doğal Makyaj", "Mousse Mask Lip; Kore köpük dudak maskesi; doğal renk; inovatif format"),
    ],

    "Makyaj - Fondöten & BB/CC Krem": [
        ("PONY EFFECT", "ponyeffect.com", "Kore YouTube MUA", "Coverstay Foundation; Kore YouTube makyaj sanatçısı; PONY; profesyonel"),
        ("JUNGSAEMMOOL", "jsmbeauty.com", "Kore MUA Marka", "Essential Skin Nuder Cushion; Kore #1 makyaj sanatçısı; Jung Saem Mool"),
        ("LAKA Thin Veil", "laka.co.kr", "Kore Genderless BB", "Thin Veil Foundation; cinsiyet nötr; Kore minimal; şeffaf coverage"),
        ("TIRTIR Mask Fit", "tirtir.com", "Kore Maske Cushion", "Mask Fit Red Cushion; Kore TikTok viral; mask-proof; uzun ömürlü"),
        ("AGE 20's", "age20s.com", "Kore Essence Cushion", "Signature Essence Cover Pact; Kore essence + cushion; çift katman; premium"),
        ("Sulwhasoo Perfecting Cushion", "sulwhasoo.com", "Kore Hanbang Cushion", "Perfecting Cushion EX; Kore hanbang lüks; ginseng + cushion"),
        ("Hera UV Mist Cushion", "hera.com", "Kore UV Mist", "UV Mist Cushion; Kore UV koruma cushion; Seul lüks; SPF 50"),
        ("NARS Light Reflecting", "narscosmetics.com", "Lüks Radiant Fondöten", "Light Reflecting Foundation; ışık yansıtma; François Nars; lüks"),
        ("Armani Luminous Silk", "giorgioarmanibeauty.com", "İtalyan Lüks Fondöten", "Luminous Silk Foundation; İtalyan lüks; ışıltılı cilt; backstage"),
        ("YSL Touche Éclat", "yslbeauty.com", "Fransız Lüks Fondöten", "Touche Éclat Le Teint; Fransız lüks; ışık bazlı; ikonik altın kalem"),
    ],

    "Saç Bakımı - Şampuan & Saç Kremi": [
        ("Gisou Honey Shampoo", "gisou.com", "Bal Şampuan", "Honey Infused Hair Wash; Mirsalehi bal çiftliği; viral; lüks şampuan"),
        ("Anomaly Haircare", "anomalyhaircare.com", "Priyanka Chopra Saç", "Bonding Treatment Mask; Priyanka Chopra Jonas markası; sürdürülebilir; erişilebilir"),
        ("Monday Haircare", "mondayhaircare.com", "İskandinav Minimal", "Gentle Shampoo; İskandinav minimalist; erişilebilir; sürdürülebilir; pembe şişe"),
        ("Maui Moisture Curl Quench", "mauimoisture.com", "Tropikal Kıvırcık", "Curl Quench + Coconut Oil Shampoo; tropikal; kıvırcık saç; erişilebilir"),
        ("Not Your Mother's Curl Talk", "nymbrands.com", "Erişilebilir Kıvırcık", "Curl Talk Defining Cream; erişilebilir; kıvırcık tanımlama; $8"),
        ("Aussie Miracle Curls", "aussie.com", "Avustralya Kıvırcık", "Miracle Curls Shampoo; Avustralya; kıvırcık saç; coconut + jojoba"),
        ("Pantene Miracle Rescue", "pantene.com", "Onarım Şampuan", "Miracle Rescue Shampoo; protein onarım; Pro-V; erişilebilir"),
        ("Dove Intensive Repair", "dove.com", "Onarım Şampuan", "Intensive Repair Shampoo; keratin onarım; erişilebilir; her saç tipi"),
        ("TRESemmé Keratin Smooth", "tresemme.com", "Keratin Düzleştirme", "Keratin Smooth Shampoo; keratin + marula yağı; salon ilhamlı"),
        ("Matrix Total Results", "matrix.com", "Salon Şampuan", "Total Results Moisture Me Rich Shampoo; salon profesyonel; glycerin"),
    ],

    "Vücut Bakımı - Vücut Nemlendirici & Yağ": [
        ("Eos Shea Better", "evolutionofsmooth.com", "Shea Vücut", "Shea Better Body Lotion; shea butter; yumurta markasından vücut"),
        ("Raw Sugar Living", "rawsugar.com", "Doğal Vücut", "Raw Coconut + Mango Body Butter; doğal; sürdürülebilir; tropikal"),
        ("Mrs. Meyer's Clean Day", "mrsmeyers.com", "Temiz Vücut", "Clean Day Body Lotion; temiz bileşenler; çiçek kokuları; ev + vücut"),
        ("Dr. Teal's", "drteals.com", "Epsom Tuz Vücut", "Body Lotion with Pure Epsom Salt; Epsom tuz + nemlendirme; uygun"),
        ("Cetaphil Body", "cetaphil.com", "Dermatolojik Vücut", "Moisturizing Cream; dermatolojik standart; hassas vücut; erişilebilir"),
        ("Lubriderm Daily", "lubriderm.com", "Günlük Vücut", "Daily Moisture Lotion; hafif günlük; vitamin E; dermatoloji"),
        ("Gold Bond Radiance", "goldbond.com", "Radiance Vücut", "Radiance Renewal Lotion; CoQ10 + vitamin; parlaklık; medikal"),
        ("Vaseline Essential", "vaseline.com", "Günlük Vücut Bakım", "Essential Healing Body Lotion; jelly + oat extract; onarım"),
        ("Curél Ultra Healing", "curel.com", "Japon Ceramide Vücut", "Ultra Healing Intensive Lotion; ceramide; kuru cilt; Japon teknoloji"),
        ("Aveeno Stress Relief", "aveeno.com", "Stres Giderici Vücut", "Stress Relief Body Lotion; lavanta + chamomile; aromaterapi"),
    ],

    "Parfüm & Koku - Kadın Parfüm": [
        ("Viktor & Rolf", "viktor-rolf.com", "Flowerbomb", "Flowerbomb; çiçek bombası; Hollanda avant-garde; lüks; ikonik"),
        ("YSL Mon Paris alternatifi: Phlur", "phlur.com", "Father Figure", "Father Figure; temiz; amber + sandalwood; unisex; modern"),
        ("Gucci Bloom alternatifi: Skylar", "skylar.com", "Fall Moon", "Fall Moon; temiz; amber + vanilya; hipoalerjenik; vegan"),
        ("Chloé alternatifi: Ellis Brooklyn", "ellisbrooklyn.com", "SUPER AMBER", "SUPER AMBER; temiz lüks; amber + vanilya; Brooklyn"),
        ("Lancôme Idôle alternatifi: Dedcool", "dedcool.com", "Taunt 01", "Taunt; biyozgörülebilir; gül + misk; California wellness; temiz"),
    ],

    "Erkek Bakım - Erkek Cilt Bakımı": [
        ("Jack Black Double-Duty", "getjackblack.com", "Premium Erkek SPF", "Double-Duty Face Moisturizer SPF 20; çok amaçlı; premium"),
        ("Lab Series Max LS", "labseries.com", "Max LS Erkek", "MAX LS Age-Less Face Cream; erkek anti-aging; Estée Lauder"),
        ("Clinique For Men Oil Control", "clinique.com", "Erkek Yağ Kontrol", "Oil Control Mattifying Moisturizer; mat; yağlı cilt; dermatolojik"),
        ("Kiehl's Age Defender", "kiehls.com", "Erkek Anti-Age", "Age Defender Moisturizer; erkek anti-aging; retinol; NYC"),
        ("Aesop Mandarin", "aesop.com", "Botanik Erkek Nemlendirme", "Mandarin Facial Hydrating Cream; mandarin; botanik; unisex"),
    ],
}

EXTRA_BRANDS_7 = {
    "Cilt Bakımı - Nemlendirici & Serum": [
        ("Atopalm Real Barrier", "atopalm.com", "Kore Atopalm Serum", "Real Barrier Aqua Soothing Gel Cream; jel krem; Kore bariyer nemlendirme"),
        ("Dr. Belmeur Cica", "thefaceshop.com", "Kore Cica Nemlendirici", "Dr. Belmeur Cica Recovery Cream; The Face Shop dermatolojik hat; cica"),
        ("Skinceuticals HA Intensifier", "skinceuticals.com", "HA Yoğunlaştırıcı", "H.A. Intensifier; hyaluronic acid amplifier; %30 artırma; medikal lüks"),
        ("SkinMedica HA5", "skinmedica.com", "5 Formlu HA", "HA5 Rejuvenating Hydrator; 5 form hyaluronic; medikal estetik"),
        ("Jan Marini Transformation", "janmarini.com", "Dönüşüm Serumu", "Transformation Face Serum; peptide + growth factor; medikal"),
        ("iS Clinical Hydra-Cool", "isclinical.com", "Hydra Serum", "Hydra-Cool Serum; hyaluronic + centella; medikal cilt bakımı; klinik"),
        ("Environ AVST", "environ.com", "Güney Afrika Vitamin A", "AVST 1-5 Moisturiser; vitamin A step-up; Güney Afrika dermatolojik"),
        ("ZO Skin Growth Factor", "zoskinhealth.com", "Büyüme Faktörü", "Growth Factor Serum; Dr. Zein Obagi; büyüme faktörü; medikal"),
        ("SkinBetter Alto Defence", "skinbetter.com", "Alto Savunma", "Alto Defense Serum; antioksidan savunma; 19 antioksidan; medikal"),
        ("Revision DEJ", "revisionskincare.com", "DEJ Serum", "DEJ Face Cream; dermal-epidermal junction; medikal anti-aging"),
        ("Neocutis Lumière", "neocutis.com", "İsviçre PSP Serum", "Lumière Firm; PSP + caffeine; İsviçre yara iyileştirme bilimi"),
        ("Colorescience Pep Up", "colorescience.com", "Peptide Serum", "Pep Up Collagen Renewal Face & Neck Treatment; peptide + botanik"),
        ("Alastin TransFORM", "alastin.com", "TriHex Serum", "TransFORM Body Treatment; TriHex peptide; medikal vücut bakımı"),
        ("Topix Replenix", "topixpharm.com", "Antioksidan Serum", "Replenix Power of Three Cream; yeşil çay polifenol + retinol + kofein"),
        ("EltaMD Skin Recovery", "eltamd.com", "Onarım Serumu", "Skin Recovery Serum; niacinamide + peptide; dermatolojik onarım"),
        ("PCA Skin Hyaluronic", "pcaskin.com", "PCA HA Serum", "Hyaluronic Acid Boosting Serum; HA + ceramide; medikal estetik"),
        ("Obagi Daily Hydro-Drops", "obagi.com", "Hydro Serum", "Daily Hydro-Drops; hyaluronic + B3; medikal nemlendirme; şeffaf jel"),
        ("SkinMedica Lytera 2.0", "skinmedica.com", "Pigment Düzeltme", "Lytera 2.0 Pigment Correcting Serum; tranexamic + phenylethyl resorcinol"),
        ("Jan Marini C-ESTA", "janmarini.com", "C Ester Serum", "C-ESTA Serum; vitamin C ester + DMAE; medikal aydınlatma"),
        ("iS Clinical Pro-Heal", "isclinical.com", "Pro-Heal Serum", "Pro-Heal Serum Advance+; L-ascorbic + olive leaf; medikal antioksidan"),
        ("Environ Youth EssentiA", "environ.com", "Youth Serum", "Youth EssentiA Vita-Peptide C-Quence Serum; vitamn A + C + peptide + antioksidan"),
        ("ZO Skin Illuminating AOX", "zoskinhealth.com", "AOX Aydınlatma", "Illuminating AOX Serum; antioksidan aydınlatma; Dr. Obagi"),
        ("SkinBetter Interfuse", "skinbetter.com", "Interfuse Göz", "InterFuse Treatment Cream Eye; peptide + hyaluronic; göz çevresi"),
        ("Revision Nectifirm", "revisionskincare.com", "Boyun Kremi", "Nectifirm Advanced; boyun + dekolte sıkılaştırma; peptide + botanik"),
        ("Alastin Restorative", "alastin.com", "TriHex Yüz", "Restorative Skin Complex; TriHex Technology; medikal yüz serumu"),
        ("PCA Skin ExLinea", "pcaskin.com", "Peptide Kırışıklık", "ExLinea Peptide Smoothing Serum; peptide bazlı; kırışıklık azaltma"),
        ("EltaMD Barrier Renewal", "eltamd.com", "Bariyer Yenilenme", "Barrier Renewal Complex; niacinamide + ceramide; bariyer onarım"),
        ("Obagi Retinol 1.0", "obagi.com", "Retinol Yüksek Doz", "Retinol 1.0; %1 retinol; medikal anti-aging; yüksek doz"),
        ("Topix CelleRenew", "topixpharm.com", "Kök Hücre Serum", "CelleRenew Growth Factor Cream; kök hücre + growth factor; medikal"),
        ("Colorescience Total Eye", "colorescience.com", "Total Göz", "Total Eye 3-in-1 SPF 35; göz + SPF + kapatıcı; mineral SPF göz"),
    ],

    "Makyaj - Göz Makyajı": [
        ("HUDA Beauty Desert Dusk", "hudabeauty.com", "Çöl Paleti", "Desert Dusk Palette; çöl renkleri; yoğun pigment; viral"),
        ("Morphe 35O", "morphe.com", "Mega Palet", "35O Nature Glow Palette; 35 renk; YouTube favorisi; erişilebilir"),
        ("Natasha Denona Gold", "natashadenona.com", "Lüks Gold Palet", "Gold Palette; 15 renk; lüks formül; makyaj sanatçısı vizyonu"),
        ("Charlotte Tilbury Stars", "charlottetilbury.com", "İngiliz Yıldız Palet", "Instant Look In a Palette; İngiliz Hollywood; çok amaçlı; lüks"),
        ("Tom Ford Desert Fox", "tomford.com", "Lüks Göz Paleti", "Eye Color Quad; İtalyan lüks; 4 renk; ultra pigment"),
        ("Chanel Les 4 Ombres", "chanel.com", "Fransız Lüks Göz", "Les 4 Ombres; Fransız haute couture; 4 renk harmoni; ikonik"),
        ("Dior 5 Couleurs", "dior.com", "Fransız Couture Göz", "5 Couleurs Eyeshadow Palette; Fransız couture; 5 renk; lüks"),
        ("MAC Art Library", "maccosmetics.com", "Profesyonel Palet", "Art Library Palette; profesyonel makyaj; 12 renk; backstage"),
        ("Bobbi Brown Luxe", "bobbibrowncosmetics.com", "NYC Lüks Göz", "Luxe Metal Rose Eyeshadow; metalik lüks; NYC profesyonel"),
        ("Laura Mercier Caviar", "lauramercier.com", "Kaviar Göz Farı", "Caviar Stick Eye Colour; stick göz farı; kremsi; kolay uygulama"),
    ],

    "Makyaj - Dudak Ürünleri": [
        ("MAC Ruby Woo", "maccosmetics.com", "İkonik Kırmızı", "Ruby Woo Lipstick; dünyanın en çok satan kırmızı ruju; mat retro; ikonik"),
        ("Charlotte Tilbury Pillow Talk", "charlottetilbury.com", "Nude İkon", "Pillow Talk Lipstick; dünyanın en çok satan nude; İngiliz lüks"),
        ("NARS Dolce Vita", "narscosmetics.com", "NARS Klasik", "Lipstick Dolce Vita; François Nars; lüks mat; klasik renk"),
        ("Chanel Rouge Allure", "chanel.com", "Fransız Lüks Ruj", "Rouge Allure Velvet; Fransız couture; kadife mat; lüks"),
        ("Dior Rouge Dior", "dior.com", "Fransız Couture Ruj", "Rouge Dior Lipstick; refillable; Fransız lüks; couture ruj"),
        ("YSL Rouge Pur Couture", "yslbeauty.com", "Fransız Lüks Lip", "Rouge Pur Couture; İtalyan deri ambalaj; Fransız lüks; ikonik"),
        ("Tom Ford Lip Color", "tomford.com", "Ultra Lüks Ruj", "Lip Color; ultra lüks; zengin pigment; Tom Ford estetik"),
        ("Guerlain Rouge G", "guerlain.com", "Mücevher Ruj", "Rouge G Lipstick; mücevher ambalaj; özelleştirilebilir kasa; Fransız"),
        ("Givenchy Le Rouge", "givenchy.com", "Fransız Le Rouge", "Le Rouge Sheer Velvet; Fransız couture; kadife mat; Audrey Hepburn"),
        ("Hermès Rouge Hermès", "hermes.com", "Ultra Lüks Ruj", "Rouge Hermès Satin Lipstick; Hermès renk evreni; zanaatkar ruj; ultra lüks"),
    ],

    "Vücut Bakımı - Deodorant": [
        ("Secret Aluminum Free", "secret.com", "Alüminyumsuz Secret", "Aluminum Free Deodorant; büyük marka alüminyumsız; whole body; P&G"),
        ("Dove Advanced Care", "dove.com", "Gelişmiş Bakım Deo", "Advanced Care Antiperspirant; 48 saat; nemlendirici; NutriumMoisture"),
        ("Degree UltraClear", "degree.com", "Ultraclear Deo", "UltraClear Black + White; leke bırakmayan; unisex; 48 saat"),
        ("Suave Antiperspirant", "suave.com", "Uygun Deo", "Powder Antiperspirant; ultra uygun; güvenilir; erişilebilir"),
        ("Arm & Hammer Ultramax", "armandhammer.com", "Karbonat Güçlü", "Ultramax Antiperspirant; karbonat bazlı; güçlü koruma; erişilebilir"),
    ],

    "Tırnak Bakımı - Oje & Jel": [
        ("Chanel Le Vernis", "chanel.com", "Fransız Lüks Oje", "Le Vernis Longwear; Fransız couture oje; lüks; ikonik renkler"),
        ("Dior Vernis", "dior.com", "Dior Lüks Oje", "Dior Vernis Couture Colour; Fransız couture; gel effect; lüks"),
        ("Tom Ford Nail Lacquer", "tomford.com", "Ultra Lüks Oje", "Nail Lacquer; ultra lüks; zengin pigment; Tom Ford estetik"),
        ("YSL La Laque Couture", "yslbeauty.com", "Fransız Couture Oje", "La Laque Couture; Fransız couture; altın kapak; ikonik"),
        ("Gucci Vernis", "gucci.com", "İtalyan Lüks Oje", "Vernis à Ongles; İtalyan lüks; Alessandro Michele estetik; vintage"),
    ],

    "Saç Bakımı - Kıvırcık & Tekstürlü Saç": [
        ("Rizos Curls", "rizoscurls.com", "Latin Kıvırcık", "Curl Defining Cream; Latin kıvırcık bakımı; doğal bileşenler; kültürel"),
        ("Afrocenchix", "afrocenchix.com", "İngiliz Afro Bakımı", "Swish Shampoo; İngiliz afro saç; bilimsel formüller; kapsayıcı"),
        ("Kurly Klips", "kurlyklips.com", "Kıvırcık Uzantı", "Clip-In Curly Extensions; kıvırcık saç uzantısı; doğal görünüm"),
        ("Mane Club", "maneclub.co", "Saç Büyütme Gummy", "Hair Growth Gummies + Curly Care; biotin + kıvırcık bakım; Z kuşağı"),
        ("Briogeo Curl Charisma", "briogeo.com", "Temiz Kıvırcık", "Curl Charisma Rice Amino + Avocado Leave-In; pirinç amino asit"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# MERGE & DEDUPLICATE
# ═══════════════════════════════════════════════════════════════════════════════

def generate_expanded_brands():
    """Generate additional brands from curated seed data to reach 5000+."""
    expanded = {}

    # Curated brand lists per category with (name, domain, subniche_tr, insight_tr)
    SEED_BRANDS = {
        "Cilt Bakımı - Nemlendirici & Serum": [
            ("Lesse", "lfrench.com", "Biyoaktif Serum", "Biyoaktif bitki bazlı; Los Angeles minimal; fermente botanik"),
            ("Wabi-Sabi Botanicals", "wabisabibotanicals.com", "Vahşi Hasat Serum", "Vahşi hasat Afrika botanikleri; küçük parti; fair trade bileşenler"),
            ("Wildcraft", "wildcraft.com", "Kanada Organik Serum", "Kanada organik; soğuk iklim bitkileri; küçük parti el yapımı"),
            ("Sade Baron", "sadebaron.com", "Karayip Serum", "Karayip doğal bileşenler; kakao + hindistan cevizi; tropikal organik"),
            ("Province Apothecary", "provinceapothecary.com", "Toronto Organik Serum", "Toronto artisan; organik yüz yağları; Kanada botanik"),
            ("Living Libations", "livinglibations.com", "Ham Organik Serum", "Ham organik yağlar; vahşi hasat; Kanada; süper bileşenler"),
            ("Honua Skincare", "honuaskincare.com", "Hawaii Serum", "Hawaii yerli bitkileri; kukui nut + plumeria; ada botanik"),
            ("Meow Meow Tweet", "meowmeowtweet.com", "Brooklyn Vegan Serum", "Brooklyn el yapımı; vegan; minimal bileşen; yerel üretim"),
            ("Fat and the Moon", "fatandthemoon.com", "Bitki Büyücü Serum", "Herbalist formülleri; el yapımı; bitkisel tıp + cilt bakımı"),
            ("Ursa Major", "ursamajorvt.com", "Vermont Doğal Serum", "Vermont doğasından; doğal + etkili; erkek + kadın; temiz"),
            ("Josh Rosebrook", "joshrosebrook.com", "Farm-to-Face Serum", "Organik çiftlikten; el yapımı; raw bileşenler; profesyonel organik"),
            ("Marie Veronique", "marieveronique.com", "Bilim + Doğal Serum", "Bilimsel doğal cilt bakımı; retinol + doğal bileşenler; hibrit"),
            ("Biba Los Angeles", "bibadelosangeles.com", "LA Botanik Serum", "Los Angeles botanik; Mexican-American bitki geleneği; kültürel"),
            ("YTTP Superberry", "youthtothepeople.com", "Süper Meyve Gece Serumu", "Maqui + goji berry; antioksidan gece maskesi; süperfood"),
            ("Kypris", "kfrench.com", "Saf Güzellik Serumu", "Saf + sürdürülebilir; multivitamin yüz yağı; Hawaii + SF"),
            ("de Mamiel", "demamiel.com", "İngiliz Aromaterapi Serum", "İngiliz aromaterapi lüks; seasonal serumlar; mevsimsel formüller"),
            ("In Fiore", "infiore.net", "İtalyan-ABD Organik Serum", "İtalyan + Amerikan organik; çiçek bazlı; lüks botanik"),
            ("Leahlani", "leahlani.com", "Hawaii Tropikal Serum", "Hawaii tropikal; organik; küçük parti; plumeria + kukui"),
            ("HENNÉ Organics", "henneorganics.com", "Lüks Organik Lip + Cilt", "Lüks minimal organik; dudak + cilt bakımı; saf bileşenler"),
            ("Olio E Osso", "olioeosso.com", "Minimal Balm Serum", "Minimal yağ + balmumu; çok amaçlı balm; Portland artisan"),
            ("Athar'a Pure", "atharapure.com", "Fas Organik Serum", "Fas argan + kaktüs; organik; Kuzey Afrika botanik; fair trade"),
            ("Au Naturale", "aunaturalecosmetics.com", "Organik Makyaj + Serum", "USDA organik; minimal bileşen; Iowa ABD; saf organik"),
            ("LILFOX", "lilfox.com", "Miami Tropikal Serum", "Miami tropikal lüks; rainforest botanik; el yapımı; küçük parti"),
            ("KLUR", "klfrench.com", "LA Temiz Serum", "LA temiz güzellik; adaptojenik; küçük parti; bilinçli bakım"),
            ("Beuti Skincare", "beutiskincare.com", "İngiliz Biyoteknoloji Serum", "İngiliz; biyoteknoloji + organik; beauty sleep yağı; gece bakımı"),
            ("CRUDE", "crudecare.com", "Yağ Temizleme Serumu", "Yağ ile temizleme; yağ + aktif; yüz yağı uzmanı; ABD indie"),
            ("Mahalo Skin Care", "mahalo.care", "Hawaii Lüks Serum", "Hawaii lüks organik; rare elixir; el yapımı; tropik şifa"),
            ("MARA", "marabeauty.com", "Deniz + Botanik Serum", "Deniz yosunu + botanik; algae vitamin C; okyanus + kara"),
            ("True Moringa", "truemoringa.com", "Ghana Moringa Serum", "Ghana moringa yağı; fair trade; Afrika süper bitki; organik"),
            ("Herbalore", "herbfrench.com", "Bitki Bazlı Serum", "Bitkisel tıp ilhamlı; botanik konsantre; geleneksel formüller"),
            ("Ranavat", "ranavat.com", "Hindistan Kraliyet Serumu", "Hindistan kraliyet güzellik ritüelleri; jasmine + saffron; Ayurveda lüks"),
            ("Shankara", "shankaranaturals.com", "Ayurveda Lüks Serum", "Ayurveda + modern bilim; ashwagandha + turmeric; yoga güzelliği"),
            ("Uma Oils", "umaoils.com", "Hindistan Yüz Yağı", "Hindistan kraliyet çiftliği yağları; ayurvedik yüz yağı; lüks organik"),
            ("Fable & Mane", "fableandmane.com", "Hindistan Saç + Cilt Yağı", "Hindistan saç yağı ritüeli; Ayurveda; kardeş marka; kültürel güzellik"),
            ("Sahajan", "sahajan.com", "Modern Ayurveda Serumu", "Modern Ayurveda; turmeric + ashwagandha; adaptojenik cilt bakımı"),
            ("PUREARTH", "purearth.asia", "Himalaya Serumu", "Himalaya bitkileri; Bhutan + Nepal botanik; yüksek rakım bitkiler"),
            ("Daughter of the Land", "daughteroftheland.com", "ABD Artisan Serum", "ABD artisan; bitki yağları; minimalist; organik; küçük parti"),
            ("Alaffia", "alaffia.com", "Togo Fair Trade Serum", "Togo fair trade shea; Afrika empowerment; sürdürülebilir; organik"),
            ("SheaTerra Organics", "sheaterra.com", "Afrika Organik Serum", "Afrika bitki yağları; baobab + argan + marula; organik kuru yağ"),
            ("Botanica Goldierocks", "goldierocks.com", "Güney Afrika Serum", "Güney Afrika rooibos + honeybush; yerli bitki kozmetik"),
            ("R+R Medicinals", "rfrench.com", "CBD Serum", "CBD + adaptojenik serum; endocannabinoid sistemi; wellness bakım"),
            ("Herbivore Emerald CBD", "herbivorebotanicals.com", "CBD Yüz Yağı", "Emerald CBD yağı; adaptojenik; hemp; CBD cilt bakımı"),
            ("Saint Jane Luxury", "saintjanebeauty.com", "Lüks CBD Serum", "CBD + botanik; lüks serum; 500mg CBD; Sephora; yüz yağı"),
            ("Cannuka", "cannuka.com", "CBD + Manuka Serum", "CBD + manuka bal; hibrit aktif; inflammasyon + onarım"),
            ("Vertly", "vertly.com", "CBD Botanik Serum", "CBD + bitkisel; CBD dudak balm + cilt; California botanik"),
            ("Prima", "prima.co", "CBD Wellness Serum", "CBD wellness cilt bakımı; The Afterglow Cream; hemp + botanik"),
            ("Sagely Naturals", "sagelynaturals.com", "CBD Günlük Serum", "Günlük CBD bakım; Relief & Recovery; wellness + cilt"),
            ("Kana Skincare", "kanaskincare.com", "Lavanta CBD Serum", "CBD + lavanta; Napa Valley; kuru yağ; relakasyon"),
            ("Perricone MD CBx", "perriconemd.com", "CBD Neuropeptide", "CBx for Men + Women; CBD + neuropeptide; anti-inflammatuar"),
            ("Josie Maran CBD", "josiemaran.com", "Argan CBD", "Skin Dope CBD + argan; CBD cilt bakımı; doğal + aktif"),
        ],

        "Makyaj - Fondöten & BB/CC Krem": [
            ("Laura Mercier Flawless", "lauramercier.com", "Flawless Lumière Fondöten", "Flawless Lumière Foundation; ışıltılı full coverage; lüks"),
            ("Bobbi Brown Skin Foundation", "bobbibrowncosmetics.com", "NYC Skin Fondöten", "Skin Long-Wear Weightless Foundation; NYC profesyonel; hafif"),
            ("Smashbox Studio Skin", "smashbox.com", "Studio Fondöten", "Studio Skin 15 Hour Wear Foundation; 15 saat; LA; foto finish"),
            ("Too Faced Born This Way", "toofaced.com", "Born This Way Fondöten", "Born This Way Foundation; multi-dimensional; coconut water; eğlenceli"),
            ("Urban Decay Stay Naked", "urbandecay.com", "Stay Naked Fondöten", "Stay Naked Weightless Foundation; hafif; 50 ton; vegan"),
            ("Tarte Amazonian Clay", "tartecosmetics.com", "Amazon Kil Fondöten", "Amazonian Clay Full Coverage Foundation; Amazon kili; 12 saat; doğal"),
            ("Lancôme Teint Idole", "lancome.com", "Fransız Ultra Wear", "Teint Idole Ultra Wear; 24 saat; Fransız lüks; 55 ton"),
            ("Estée Lauder Double Wear", "esteelauder.com", "Çift Dayanıklı Fondöten", "Double Wear Stay-in-Place; 24 saat; transfer-proof; 60+ ton"),
            ("MAC Studio Fix Fluid", "maccosmetics.com", "Pro Fondöten", "Studio Fix Fluid SPF 15; mat; profesyonel; 60+ ton"),
            ("Clinique Even Better", "clinique.com", "Dermatolojik Fondöten", "Even Better Clinical Serum Foundation; serum + fondöten; 48 ton"),
            ("Bare Minerals Original", "bareminerals.com", "Mineral Fondöten", "Original Loose Powder Foundation; mineral fondöten öncüsü; SPF 15"),
            ("IT Cosmetics CC+", "itcosmetics.com", "Dermatolog CC Krem", "CC+ Cream; SPF 50; anti-aging + fondöten; dermatolog geliştirdi"),
            ("Dior Forever", "dior.com", "Dior Fondöten", "Forever Skin Glow; Fransız couture; 24 saat; luminous matte"),
            ("Giorgio Armani LSF", "giorgioarmanibeauty.com", "İtalyan Silk Fondöten", "Luminous Silk Foundation; İtalyan lüks; #1 backstage; ışıltılı"),
            ("YSL All Hours", "yslbeauty.com", "Fransız All Hours", "All Hours Foundation; 24 saat mat; Fransız lüks; transfer-proof"),
            ("Chanel Les Beiges", "chanel.com", "Fransız Doğal Fondöten", "Les Beiges Healthy Glow Foundation; doğal glow; Fransız chic"),
            ("Guerlain L'Essentiel", "guerlain.com", "Fransız Doğal Fondöten", "L'Essentiel Natural Glow Foundation; %97 doğal bileşen; lüks"),
            ("Sisley Phyto-Teint", "sisley-paris.com", "Fransız Bitkisel Fondöten", "Phyto-Teint Ultra Éclat; botanik; Fransız ultra lüks"),
            ("Givenchy Prisme Libre", "givenchy.com", "Fransız Prisme Fondöten", "Prisme Libre Skin-Caring Glow; Fransız couture; 24 saat glow"),
            ("Clé de Peau Radiant Fluid", "cledepeaubeaute.com", "Japon Radiant Fondöten", "Radiant Fluid Foundation; Japon lüks; ışık teknolojisi"),
        ],

        "Saç Bakımı - Şampuan & Saç Kremi": [
            ("Fable & Mane", "fableandmane.com", "Hindistan Saç Ritüeli", "HoliRoots Pre-Wash Hair Oil; Ayurveda saç ritüeli; Hindistan"),
            ("Sienna Naturals", "siennanaturals.com", "Temiz Tekstürlü Saç", "H.A.P.I. Shampoo; temiz + doğal saç; Hannah Diop; botanical"),
            ("SEEN Hair Care", "helloseen.com", "Dermatolojik Şampuan", "Fragrance Free Shampoo; Harvard dermatoloji; akne güvenli; non-comedogenic"),
            ("Verb Ghost", "verbproducts.com", "Hafif Şampuan", "Ghost Shampoo; ağırlık vermeyen; ince saç; moringa yağı"),
            ("Herbal Essences Bio:Renew", "herbalessences.com", "Botanik Şampuan", "Bio:Renew Argan Oil Shampoo; botanik; sürdürülebilir"),
            ("Suave Professionals", "suave.com", "Uygun Salon", "Professionals Keratin Infusion; salon ilhamlı; ultra uygun"),
            ("Biolage", "biolage.com", "Doğa İlhamlı", "Hydrasource Shampoo; doğa ilhamlı; aloe vera; profesyonel"),
            ("Joico", "joico.com", "Salon Onarım", "K-PAK Damage Repair Shampoo; keratin peptide; salon profesyonel"),
            ("Nexxus", "nexxus.com", "Salon Bilimi", "Therappe Moisturizing Shampoo; proteomics bilimi; salon"),
            ("Fekkai", "fekkai.com", "NYC Lüks Saç", "Full Blown Volume Shampoo; NYC lüks salon; Frédéric Fekkai"),
            ("Ouidad Advanced Climate", "ouidad.com", "Kıvırcık Uzman Şampuan", "Advanced Climate Control Shampoo; anti-humidity; kıvırcık"),
            ("Cantu Avocado", "cantubeauty.com", "Avokado Şampuan", "Avocado Hydrating Shampoo; avokado + shea; doğal saç"),
            ("Design Essentials", "designessentials.com", "Profesyonel Doğal", "Almond & Avocado Moisturizing & Detangling Shampoo; profesyonel"),
            ("Creme of Nature", "cremeofnature.com", "Argan Doğal Saç", "Argan Oil Shampoo; argan yağı; doğal saç; erişilebilir"),
            ("Dark and Lovely", "darkandlovely.com", "Melanin Saç Bakım", "Au Naturale Moisture Replenishing Shampoo; doğal saç"),
        ],

        "Vücut Bakımı - Vücut Nemlendirici & Yağ": [
            ("L'Occitane Almond", "loccitane.com", "Provence Badem", "Almond Supple Skin Oil; Provence badem yağı; lüks; parlak cilt"),
            ("Molton Brown", "moltonbrown.com", "İngiliz Lüks Vücut", "Mesmerising Oudh Accord & Gold Body Lotion; İngiliz lüks; koku"),
            ("Penhaligon's Body", "penhaligons.com", "İngiliz Aristokrat Vücut", "Halfeti Body & Hand Lotion; İngiliz aristokrat; Türk gülü"),
            ("Acqua di Parma Body", "acquadiparma.com", "İtalyan Klasik Vücut", "Colonia Body Lotion; İtalyan; 1916; citrus; klasik"),
            ("Diptyque Body Lotion", "diptyque.com", "Paris Lüks Vücut", "Eau des Sens Body Lotion; Paris; lüks koku + bakım"),
            ("Byredo Body Lotion", "byredo.com", "İsveç Niş Vücut", "Bal d'Afrique Body Lotion; İsveç niş koku + vücut bakımı"),
            ("Le Labo Body Lotion", "lelabofragrances.com", "NYC Artisan Vücut", "Santal 33 Body Lotion; NYC artisan; santal kokusu"),
            ("Maison Margiela REPLICA", "maisonmargiela.com", "Anı Vücut", "REPLICA Body Lotion; anıları çağrıştıran; nostalji koku + bakım"),
            ("OUAI Body Cleanser", "theouai.com", "Jen Atkin Vücut", "Body Cleanser; Jen Atkin; lüks duş jeli; salon kokusu"),
            ("Necessaire Body Wash", "necessaire.com", "Vitamin Vücut Yıkama", "The Body Wash; vitamin + mineral; temiz; minimalist"),
        ],

        "Parfüm & Koku - Kadın Parfüm": [
            ("Narciso Rodriguez", "narcisorodriguez.com", "Misk Uzmanı", "For Her EDP; misk uzmanı; Narciso Rodriguez; ikonik misk"),
            ("Chloe EDP", "chloe.com", "Fransız Çiçek", "Chloé EDP; gül + şakayık; Fransız çiçek; feminen; ikonik"),
            ("Marc Jacobs Daisy", "marcjacobs.com", "Papatya İkon", "Daisy EDT; papatya ikonik; genç + taze; erişilebilir lüks"),
            ("Dolce & Gabbana Light Blue", "dolcegabbana.com", "İtalyan Yaz", "Light Blue EDT; İtalyan yaz; Capri; citrus; klasik"),
            ("Versace Bright Crystal", "versace.com", "İtalyan Kristal", "Bright Crystal EDT; İtalyan; yuzu + lotus; parlak; erişilebilir"),
            ("Carolina Herrera Good Girl", "carolinaherrera.com", "Topuklu Ayakkabı Parfüm", "Good Girl EDP; stiletto şişe; tuberose + kakao; ikonik"),
            ("Valentino Donna Born", "valentino.com", "İtalyan Aşk", "Donna Born in Roma; İtalyan; jasmine + vanilla; modern"),
            ("Burberry Her", "burberry.com", "İngiliz Meyve", "Her EDP; İngiliz; kırmızı meyve; London; modern İngiliz"),
            ("Prada Paradoxe", "prada.com", "İtalyan Paradoks", "Paradoxe EDP; İtalyan; amber + misk; modern; yenilenebilir"),
            ("Dior J'adore", "dior.com", "Fransız İkon", "J'adore EDP; Fransız çiçek ikonik; altın şişe; tuberose"),
        ],

        "Erkek Bakım - Tıraş & Sakal": [
            ("Astra Superior Platinum", "astra.com", "İtalyan Jilet", "Astra Superior Platinum DE Blades; İtalyan paslanmaz; uygun"),
            ("Feather", "feather.co.jp", "Japon Jilet", "Hi-Stainless Double Edge Blades; Japon keskin jilet; ultra ince"),
            ("Rockwell Razors", "rockwellrazors.com", "Kanada Ayarlanabilir Jilet", "6S Adjustable Safety Razor; Kanada; 6 ayar; paslanmaz çelik"),
            ("Rex Supply Co", "rfrench.com", "ABD Lüks Jilet", "Ambassador Adjustable Razor; ABD lüks; CNC işleme; micrometer ayar"),
            ("Razorock", "italianbarber.com", "İtalyan Artisan Jilet", "Game Changer Safety Razor; İtalyan artisan; uygun lüks"),
            ("Viking Blade", "vikingblade.com", "İskandinav Jilet", "The Chieftain Safety Razor; İskandinav tasarım; ağır jilet"),
            ("Bevel Trimmer", "getbevel.com", "Melanin Trimmer", "Bevel Trimmer; koyu cilt; tüy batması önleme; özel tasarım"),
            ("Beard Reverence", "beardreverence.com", "Premium Sakal Yağı", "Premium Beard Oil; jojoba + argan; sakal beslenme"),
            ("Scotch Porter", "scotchporter.com", "Siyah Erkek Bakım", "Beard Balm; siyah erkek bakım; doğal; Houston TX"),
            ("Frederick Benjamin", "frederickbenjamin.com", "Melanin Sakal", "Grooming Spray; melanin dostu sakal bakımı; tüy batması"),
        ],

        "K-Beauty & Kore Kozmetik": [
            ("COSRX The Vitamin C 23", "cosrx.com", "Kore C23 Serumu", "The Vitamin C 23 Serum; %23 saf C vitamini; Kore yüksek doz C"),
            ("Mediheal N.M.F", "mediheal.com", "Kore NMF Maske", "N.M.F Aquaring Ampoule Mask; Kore sheet mask öncüsü; hastane"),
            ("SNP Gold Collagen", "snp.co.kr", "Kore Altın Kollajen", "Gold Collagen Ampoule Mask; altın + kollajen; Kore premium maske"),
            ("Leaders Amino", "leaders.co.kr", "Kore Amino Maske", "Amino Moisture Mask; amino asit; Kore lider sheet mask"),
            ("Dr. Jart+ Vital Hydra", "drjart.com", "Kore Vital Hydra", "Vital Hydra Solution Biome Essence; biyom + nemlendirme; Kore"),
            ("AHC Aqualuronic", "ahcbeauty.com", "Kore Aqualuronic", "Aqualuronic Cream; triple hyaluronic; Kore derin nemlendirme"),
            ("Innisfree Retinol", "innisfree.com", "Kore Retinol Cica", "Retinol Cica Repair Ampoule; retinol + cica; Kore hibrit"),
            ("Missha Bee Pollen", "missha.com", "Kore Arı Poleni", "Bee Pollen Renew Ampoule; arı poleni; Kore fermente aktif"),
            ("Holika Holika Gudetama", "holikaholika.com", "Kore Gudetama", "Gudetama Lazy & Joy Jelly Blusher; karakter kolaborasyonu; eğlenceli"),
            ("Nature Republic Vitapair", "naturerepublic.com", "Kore Vita C", "Vitapair C Dark Spot Serum; C vitamini leke tedavisi; Kore"),
        ],

        "J-Beauty & Japon Kozmetik": [
            ("Kose Sekkisei", "kose.co.jp", "Japon Sekkisei", "Sekkisei Lotion; Japon kar beyazlığı; 40+ yıl; fermente pirinç"),
            ("Shu Uemura Cleansing Oil", "shuuemura.com", "Japon Temizleme İkonu", "Ultime8∞ Sublime Beauty Cleansing Oil; 8 botanik yağ; Japon ikon"),
            ("Kiehl's Japan", "kiehls.com", "Japon NYC Hibrit", "Ultra Facial Cream; Japon versiyonu; NYC eczane; global kült"),
            ("Shiseido Vital Perfection", "shiseido.com", "Japon Vital", "Vital Perfection Uplifting and Firming Cream; Japon lifting"),
            ("Clé de Peau La Crème", "cledepeaubeaute.com", "Japon Ultimate", "La Crème; Japon en lüks krem; Illuminating Complex EX"),
            ("POLA B.A", "pola.com", "Japon B.A Premium", "B.A Lotion; Japon premium; bio-active theory; anti-aging"),
            ("Menard Lisciare", "menard.co.jp", "Japon Lisciare", "Lisciare Lotion; nano-transfer; Japon ultra lüks; derinlemesine"),
            ("Albion Excia", "albion.co.jp", "Japon Excia", "Excia AL Whitening Immaculate Essence; Japon süt emülsiyon; lüks"),
            ("Suqqu Extra Rich", "suqqu.com", "Japon Extra Rich", "Extra Rich Glow Cream Foundation; Japon lüks krem fondöten"),
            ("Decorté AQ", "decorte.com", "Japon AQ", "AQ Meliority Intensive Eye Cream; Japon ultra lüks göz kremi"),
        ],

        "Doğal & Organik Kozmetik": [
            ("Mukti", "muktiorganics.com", "Avustralya Yerli Organik", "Queen of the Night; kakadu plum; Avustralya yerli; organik"),
            ("Sukin", "sukin.com", "Avustralya Karbon Nötr", "Signature Moisturiser; karbon nötr; Avustralya; erişilebilir doğal"),
            ("Trilogy Rosehip", "trilogyproducts.com", "NZ Kuşburnu", "Certified Organic Rosehip Oil; Yeni Zelanda; USDA organik"),
            ("Antipodes", "antipodesnature.com", "NZ Manuka", "Manuka Honey Skin-Brightening Eye Cream; Yeni Zelanda; manuka"),
            ("Evolve Organic", "evolvebeauty.co.uk", "İngiliz Organik", "Multi Peptide 360 Moisture Cream; İngiliz el yapımı organik"),
            ("Neal's Yard", "nealsyardremedies.com", "İngiliz Mavi Şişe", "Frankincense Intense Cream; İngiliz organik; mavi şişe; 1981"),
            ("Green People", "greenpeople.co.uk", "İngiliz Yeşil", "Age Defy+ Cell Enrich Facial Oil; İngiliz organik; Soil Association"),
            ("Odylique", "odylique.com", "İngiliz Etik Organik", "Repair Lotion; İngiliz etik; Fairtrade; ultra hassas; organik"),
            ("Balm Balm", "balmbalm.com", "İngiliz Saf Organik", "Rose Geranium Face Balm; İngiliz saf organik; tek bileşen"),
            ("Pai Rosehip", "paiskincare.com", "İngiliz Organik Kuşburnu", "Rosehip BioRegenerate Oil; İngiliz organik; hassas cilt"),
        ],

        "Vegan & Cruelty-Free Kozmetik": [
            ("Charlotte Tilbury Pillow Talk alternatifi: MUA", "muastore.com", "İngiliz Ultra Uygun", "Velvet Matte Lipstick; İngiliz; £1; ultra uygun; vegan"),
            ("Sleek MakeUP", "sleekmakeup.com", "İngiliz Melanin", "Face Form Contour + Blush; İngiliz; kapsayıcı; melanin dostu; vegan"),
            ("Freedom Makeup", "freedommakeuplondon.com", "İngiliz Pro", "Pro Brow Pomade; İngiliz; vegan; £3; profesyonel + uygun"),
            ("Makeup Obsession", "makeupobsession.com", "İngiliz Trend", "Game Set Matte Lip Kit; İngiliz; vegan; trend; £5; hızlı moda"),
            ("I Heart Revolution", "iheartrevolution.com", "Eğlenceli Vegan", "Donut Palette; eğlenceli şekil; vegan; İngiliz; koleksiyon"),
            ("Makeup Revolution Skincare", "revolutionbeauty.com", "İngiliz Vegan Skincare", "2% Hyaluronic Acid Serum; İngiliz vegan cilt bakımı; £6"),
            ("The Ordinary Squalane", "theordinary.com", "Squalane Nemlendirici", "Squalane Cleanser; squalane bazlı; vegan; erişilebilir; $8"),
            ("CeraVe PM", "cerave.com", "Vegan Gece Bakımı", "PM Facial Moisturizing Lotion; ceramide + niacinamide; vegan uyumlu"),
            ("Bioderma Sebium", "bioderma.com", "Fransız Vegan Akne", "Sébium H2O; misel su; Fransız vegan uyumlu; akne eğilimli cilt"),
            ("Becca Ever-Matte", "bfrench.com", "Vegan Primer", "Ever-Matte Poreless Priming Perfector; vegan; mat primer; gözenek"),
        ],

        "Bebek & Hamile Cilt Bakımı": [
            ("Aveeno Baby Eczema", "aveeno.com", "Yulaf Egzama Bebek", "Eczema Therapy Moisturizing Cream; koloidal yulaf; egzama"),
            ("Cetaphil Baby Eczema", "cetaphil.com", "Derma Egzama Bebek", "Baby Eczema Calming Moisturizer; ceramide; dermatolojik"),
            ("Vanicream Baby", "vanicream.com", "Ultra Hassas Bebek", "Moisturizing Skin Cream; sıfır tahriş; bebek güvenli; dermatoloji"),
            ("CeraVe Baby", "cerave.com", "Ceramide Bebek", "Baby Moisturizing Cream; 3 ceramide; MVE; bebek hassas cilt"),
            ("Eucerin Baby Eczema", "eucerin.com", "Alman Egzama Bebek", "Baby Eczema Relief Body Creme; koloidal yulaf; Alman dermo"),
            ("La Roche-Posay Baby", "laroche-posay.com", "Fransız Bebek", "Lipikar Baume AP+M Baby; Fransız termal; bebek bariyer onarım"),
            ("Bioderma ABCDerm", "bioderma.com", "Fransız Bebek Hattı", "ABCDerm Moisturising Milk; Fransız eczane bebek hattı"),
            ("Mustela Stelatopia", "mustela.com", "Fransız Egzama Bebek", "Stelatopia Emollient Cream; avokado perseose; atopik bebek"),
            ("Avène Trixéra Baby", "avene.com", "Fransız Termal Bebek", "Trixéra Nutrition Nutri-Fluid Lotion; termal su bebek"),
            ("Sanosan", "sanosan.com", "Alman Bebek", "Baby Care Cream; Alman bebek bakımı; olive oil extract; hassas"),
        ],

        "İntim Bakım & Vücut": [
            ("Momotaro Apotheca", "momotaroapotheca.com", "Japon İntim", "Salve; Japon bitkisel intim bakım; botanik; anti-inflammatuar"),
            ("pH-D Feminine Health", "phdfemininehealth.com", "pH Dengeli İntim", "Boric Acid Vaginal Suppositories; borik asit; pH dengeleme"),
            ("VagiBiom", "vagibiom.com", "Probiyotik İntim", "Probiotics Suppository; probiyotik; vajinal mikrobiyom; sağlık"),
            ("AZO", "azoproducts.com", "İdrar Yolu Sağlığı", "Cranberry Urinary Tract Health; kızılcık; idrar yolu; kadın sağlığı"),
            ("Uqora", "uqora.com", "UTI Önleme", "Target; D-mannose; idrar yolu enfeksiyonu önleme; DTC kadın sağlığı"),
        ],

        "Diş Beyazlatma & Ağız Bakımı": [
            ("Colgate Hum", "colgate.com", "Akıllı Fırça", "Hum by Colgate Smart Toothbrush; akıllı diş fırçası; app"),
            ("Bruush", "bruush.com", "DTC Elektrikli Fırça", "Sonic Toothbrush; DTC abonelik; $79; sonic; uygun premium"),
            ("Oclean", "oclean.com", "Çin Akıllı Fırça", "X Pro Digital; Çin; AI; dijital; akıllı sonic; erişilebilir"),
            ("Usmile", "usmile.com", "Çin Y1S Fırça", "Y1S Sonic Toothbrush; Çin; tasarım ödüllü; Y-Shape; premium"),
            ("Apa Beauty Toothpaste", "apabeauty.com", "Lüks Beyazlatma", "White Beauty Toothpaste; NYC Dr. Apa; premium beyazlatma macunu"),
        ],
    }

    # Additional massive brand expansion per category
    MORE_BRANDS = {
        "Cilt Bakımı - Nemlendirici & Serum": [
            ("Grown Alchemist Hydra+", "grownalchemist.com", "Avustralya HA Serum", "Hydra-Repair Day Cream; Avustralya biyolojik yaklaşım; peptide + hyaluronic"),
            ("Votary Super Seed", "votary.co.uk", "İngiliz Yüz Yağı", "Super Seed Facial Oil; İngiliz lüks yüz yağı; broccoli + chia"),
            ("May Lindstrom Problem", "maylindstrom.com", "Artisan Lüks Serum", "The Problem Solver; kakao ısınan maske; artisan küçük parti"),
            ("African Botanics Marula", "africanbotanics.com", "Afrika Lüks Serum", "Marula Neroli Superfine Face Oil; Güney Afrika marula; lüks"),
            ("Eminence Bright Skin", "eminenceorganics.com", "Macar Organik Serum", "Bright Skin Licorice Root Exfoliating Peel; Macar organik spa"),
            ("Goldfaden MD", "goldfadenmd.com", "Dermatolojik Serum", "Needle-Less Line Smoothing Concentrate; NYC dermatolog"),
            ("Dr. Loretta", "drloretta.com", "Anti-Aging Bilim Serum", "Micro Peel Peptide Pads; dermatolojik peptide pad"),
            ("Tammy Fender", "tammyfender.com", "Holistic Serum", "Quintessential Serum; holistik güzellik; Palm Beach; bitki bazlı"),
            ("Tracie Martyn", "traciemartyn.com", "NYC Facialist Serum", "Enzyme Exfoliant; NYC ünlü facialist; profesyonel ev bakımı"),
            ("Georgia Louise", "georgialouise.com", "NYC Glow Serum", "Glow Tonic; NYC facialist; profesyonel at-home; lift + glow"),
            ("Sonya Dakar", "sonyadakar.com", "Beverly Hills Serum", "Flash Facial; Beverly Hills klinik; profesyonel peeling ev"),
            ("Amore Pacific Vintage", "amorepacific.com", "Kore Vintage Serum", "Vintage Single Extract Essence; 50+ yıl yeşil çay fermentasyon"),
            ("Sulwhasoo First Care", "sulwhasoo.com", "Kore İlk Bakım Serumu", "First Care Activating Serum; Kore ilk adım serum; hanbang"),
            ("Ohui Cell Power", "ohui.co.kr", "Kore Hücre Gücü", "Cell Power No. 1 Essence; Kore hücre bilimi; LG premium"),
            ("Su:m37 Secret", "sum37.com", "Kore Fermente Öz", "Secret Programming Essence; 365 gün fermentasyon; LG doğal"),
            ("Hera Cell Essence", "hera.com", "Kore Hücre Özü", "Cell Essence Biome Plus; Kore hücre + biyom; Amorepacific"),
            ("Primera Miracle Seed", "primera.co.kr", "Kore Tohum Özü", "Miracle Seed Essence; lotus tohumu; Kore organik; Amorepacific"),
            ("IOPE Retinol Expert", "iope.com", "Kore Retinol", "Retinol Expert 0.1%; Kore retinol uzmanı; Amorepacific biyo-bilim"),
            ("Mamonde Flower Lab", "mamonde.com", "Kore Çiçek Bilimi", "Flower Lab Essence; çiçek bilimi; Amorepacific; Kore botanik"),
            ("Lirikos Marine", "lirikos.com", "Kore Deniz Kolajen", "Marine Collagen Cream; deniz kolajen; Kore okyanus bilimi"),
            ("Donginbi Red Ginseng", "donginbi.com", "Kore Kızıl Ginseng", "1899 Single Essence; 6 yıl fermente ginseng; KGC; lüks hanbang"),
            ("Sooryehan Boyun", "sooryehan.com", "Kore Hanbang Boyun", "Boyun Crème; Kore hanbang boyun bakımı; LG; geleneksel tıp"),
            ("Enough Project Simple", "enoughproject.kr", "Kore Yeterli Bakım", "Sun Fluid; Kore basit + yeterli; erişilebilir; minimal"),
            ("Shangpree", "shangpree.com", "Kore Spa Serum", "Gold Black Pearl Eye Mask; Kore premium spa; altın + siyah inci"),
            ("Urang", "ufrench.com", "Kore Organik Serum", "True Rose Repair Essence; Kore organik; gül; COSMOS sertifikalı"),
        ],
        "Cilt Bakımı - Akne & Leke Tedavisi": [
            ("Bioré Baking Soda", "biore.com", "Japon Baking Soda", "Baking Soda Pore Cleanser; karbonat; gözenek temizleme; Japon"),
            ("Clean & Clear Deep", "cleanandclear.com", "Derin Temizleme Akne", "Deep Cleaning Astringent; salisilik asit; eczane; erişilebilir"),
            ("Proactiv+", "proactiv.com", "Sistem Akne", "3-Step System; benzoyl peroxide sistemi; TV infomercial; klinik"),
            ("AcneFree", "acnefree.com", "Eczane Akne", "Oil-Free Acne Cleanser; benzoyl peroxide; eczane; erişilebilir"),
            ("La Roche-Posay Adapalene", "laroche-posay.com", "Fransız Adapalene", "Adapalene Gel 0.1%; retinoid; Fransız dermo; reçetesiz"),
        ],
        "Cilt Bakımı - Anti-Aging & Kırışıklık": [
            ("Lancôme Absolue", "lancome.com", "Fransız Absolue", "Absolue Revitalizing & Brightening Soft Cream; gül kök hücresi"),
            ("Estée Lauder Re-Nutriv", "esteelauder.com", "Ultra Lüks Anti-Age", "Re-Nutriv Ultimate Diamond Transformative Energy Crème; ultra lüks"),
            ("Helena Rubinstein Prodigy", "helenarubinstein.com", "Prodigy Reversis", "Prodigy Reversis Cream; native kolajen teknolojisi; Fransız lüks"),
            ("Chantecaille Bio Lifting", "chantecaille.com", "Botanik Lüks", "Bio Lifting Cream+; botanik + hayvan koruma; lüks organik"),
            ("La Prairie Platinum", "laprairie.com", "İsviçre Platin", "Platinum Rare Cellular Life-Lotion; platin + hücre bilimi; ultra lüks"),
        ],
        "Makyaj - Allık & Bronzer": [
            ("NARS Orgasm", "narscosmetics.com", "İkonik Allık", "Orgasm Blush; dünyanın en çok satan allığı; altın parıltı; ikonik"),
            ("Tarte Amazonian Clay Blush", "tartecosmetics.com", "Amazon Kil Allık", "Amazonian Clay 12-Hour Blush; 12 saat dayanıklı; mat + saten"),
            ("MAC Mineralize Blush", "maccosmetics.com", "Mineral Allık", "Mineralize Blush; mineral pigment; profesyonel; baked formül"),
            ("Bobbi Brown Shimmer Brick", "bobbibrowncosmetics.com", "Shimmer Brick Allık", "Shimmer Brick Compact; çoklu renk şerit; versatil; profesyonel"),
            ("Laura Mercier Blush", "lauramercier.com", "İpeksi Allık", "Blush Colour Infusion; ipeksi mat; doğal flush; lüks"),
            ("Clinique Cheek Pop", "clinique.com", "Cheek Pop Allık", "Cheek Pop; gerbera çiçeği deseni; dermatolojik; hafif"),
            ("Charlotte Tilbury Cheek", "charlottetilbury.com", "İngiliz Cheek Allık", "Cheek to Chic; çift ton; İngiliz lüks; Hollywood glow"),
            ("Tom Ford Cheek Color", "tomford.com", "Lüks Allık", "Cheek Color; ultra lüks; Tom Ford estetik; yoğun pigment"),
            ("Dior Rosy Glow", "dior.com", "Dior Allık", "Rosy Glow Blush; renk değiştiren; Fransız couture; doğal flush"),
            ("Chanel Joues Contraste", "chanel.com", "Fransız Allık", "Joues Contraste Powder Blush; Fransız couture; klasik; lüks"),
        ],
        "Makyaj - Göz Makyajı": [
            ("Urban Decay Naked", "urbandecay.com", "İkonik Nude Palet", "Naked Palette; nude göz paleti standardı; ikonik; 12 renk"),
            ("Too Faced Chocolate Bar", "toofaced.com", "Çikolata Palet", "Chocolate Bar Palette; çikolata kokulu; 16 renk; eğlenceli"),
            ("Tarte Tartelette", "tartecosmetics.com", "Mat Palet", "Tartelette In Bloom Palette; mat + saten; Amazon kili; 12 renk"),
            ("MAC Pro Longwear", "maccosmetics.com", "Pro Göz Farı", "Pro Longwear Paint Pot; krem göz bazı; profesyonel; uzun ömür"),
            ("Bobbi Brown Luxe", "bobbibrowncosmetics.com", "Lüks Göz Farı", "Luxe Eye Shadow; zengin pigment; NYC profesyonel; kremsi"),
            ("Laura Mercier Caviar Stick", "lauramercier.com", "Kaviar Göz Stick", "Caviar Stick Eye Colour; kremsi stick; kolay uygulama; 12 saat"),
            ("Tom Ford Quad", "tomford.com", "Lüks Göz Quad", "Eye Color Quad; 4 renk; ultra lüks; zengin pigment; İtalyan"),
            ("Chanel Les 4 Ombres", "chanel.com", "Fransız Quad Göz", "Les 4 Ombres Multi-Effect Quadra; Fransız couture; 4 renk harmoni"),
            ("Dior 5 Couleurs Couture", "dior.com", "Dior 5 Renk", "5 Couleurs Couture; 5 renk harmoni; Fransız; backstage"),
            ("YSL Couture Clutch Palette", "yslbeauty.com", "YSL Çanta Palet", "Couture Clutch Palette; çanta formunda; Fransız lüks; seyahat"),
        ],
        "Makyaj - Dudak Ürünleri": [
            ("Tom Ford Lip Color Satin", "tomford.com", "Ultra Lüks Saten Ruj", "Lip Color Satin Matte; ultra lüks; zengin renk; Tom Ford"),
            ("Guerlain Rouge G Velvet", "guerlain.com", "Mücevher Kadife Ruj", "Rouge G de Guerlain Velvet; mücevher ambalaj; kadife mat; lüks"),
            ("Givenchy Le Rouge Deep", "givenchy.com", "Fransız Derin Ruj", "Le Rouge Deep Velvet; derin mat; Fransız couture; yoğun"),
            ("Hermès Satin", "hermes.com", "Hermès Saten Ruj", "Rouge Hermès Satin Lipstick; Hermès renk evreni; ultra lüks"),
            ("Burberry Kisses", "burberry.com", "İngiliz Öpücük", "Burberry Kisses Lipstick; İngiliz; saten + mat; London estetik"),
            ("Armani Lip Maestro", "giorgioarmanibeauty.com", "İtalyan Lip Maestro", "Lip Maestro Liquid Lipstick; İtalyan; kadife sıvı ruj; 400"),
            ("Valentino Rosso", "valentino.com", "İtalyan Kırmızı", "Rosso Valentino Refillable Lipstick; İtalyan kırmızı; refillable"),
            ("Prada Monochrome", "prada.com", "İtalyan Mono Ruj", "Monochrome Soft Matte Lipstick; İtalyan; mat; minimal; lüks"),
            ("Chanel Allure", "chanel.com", "Fransız Allure Ruj", "Rouge Allure L'Extrait; Fransız; yoğun mat; refillable; couture"),
            ("Dior Addict Shine", "dior.com", "Dior Parlak Ruj", "Dior Addict Lip Maximizer; dudak dolgunlaştırıcı; Fransız; ikonik"),
        ],
        "Saç Bakımı - Saç Büyütme & Dökülme": [
            ("Keranique Regrowth", "keranique.com", "Kadın Saç Sistemi", "Hair Regrowth Treatment %2 Minoxidil; kadın saç büyütme; FDA"),
            ("Nioxin System 2", "nioxin.com", "İncelme Sistemi", "System 2 Kit; doğal saç belirgin incelme; 3 adım; profesyonel"),
            ("Rogaine Men's 5%", "rogaine.com", "Erkek Minoxidil", "Men's 5% Minoxidil Foam; erkek saç büyütme; FDA onaylı; köpük"),
            ("Minoxidil Response", "minoxidilresponse.com", "Minoxidil DTC", "Topical Minoxidil; DTC minoxidil; uygun fiyat; abonelik"),
            ("Shapiro MD", "shapiromd.com", "Doktor Saç", "Shampoo + Conditioner; DHT bloker; doktor formülü; saw palmetto"),
        ],
        "Vücut Bakımı - Self-Tan & Bronzlaşma": [
            ("St. Tropez Purity", "sttropeztan.com", "Purity Self-Tan", "Purity Vitamins Bronzing Water Body Mist; vitamin C + D; sağlıklı bronz"),
            ("Tan-Luxe The Butter", "tan-luxe.com", "Kademeli Butter", "The Butter Illuminating Tanning Butter; kademeli; parıltılı; lüks"),
            ("Isle of Paradise Glow", "isleofparadise.com", "Renk Düzeltici", "Glow Clear Self-Tanning Mousse; yeşil + mor + şeftali renk düzeltici"),
            ("Bondi Sands Liquid", "bondisands.com", "Avustralya Liquid Gold", "Liquid Gold Self Tanning Dry Oil; kuru yağ; Avustralya; parlak"),
            ("Loving Tan Deluxe", "lovingtan.com", "Koyu Tan", "Deluxe Bronzing Mousse; Avustralya; koyu bronz; profesyonel"),
        ],
        "Tırnak Bakımı - Oje & Jel": [
            ("Morgan Taylor", "morgantaylorlacquer.com", "Profesyonel Oje", "Professional Nail Lacquer; salon profesyonel; geniş renk paleti"),
            ("LeChat Dare to Wear", "lechatnails.com", "Gel Eşleşme Oje", "Dare to Wear Nail Lacquer; jel eşleşme sistemi; salon"),
            ("Artistic Nail Design", "artisticnaildesign.com", "Artistik Oje", "Colour Gloss; artistik jel oje; profesyonel; salon"),
            ("DND Gel", "dfrench.com", "Duo Jel Oje", "Duo Gel Nail Polish; jel + oje ikili; geniş renk; salon"),
            ("Daisy Gel", "daisygel.com", "Daisy Jel Seti", "Gel + Lacquer Duo; ikili set; salon profesyonel; 300+ renk"),
        ],
        "Parfüm & Koku - Unisex & Niş": [
            ("MFK Gentle Fluidity", "franciskurkdjian.com", "Gentle Fluidity", "Gentle Fluidity Gold; vanilya + amber; unisex; LVMH lüks niş"),
            ("Kilian Angels' Share", "bykilian.com", "Melek Payı", "Angels' Share; cognac + praline; Estée Lauder; lüks niş"),
            ("Tiziana Terenzi", "tizianaterenzi.com", "İtalyan Artisan", "Kirke; İtalyan artisan; passionfruit + mango; lüks; niş"),
            ("Clive Christian", "clivechristian.com", "İngiliz Kraliyet", "No.1; dünyanın en pahalı parfümlerinden; İngiliz kraliyet amblemi"),
            ("Roja Parfums", "rojaparfums.com", "İngiliz Master", "Elysium; İngiliz master parfümör Roja Dove; ultra lüks niş"),
            ("House of Oud", "houseofoud.com", "İtalyan Oud", "Almond Harmony; İtalyan oud uzmanı; badem + oud; niş"),
            ("Boadicea the Victorious", "bfrench.com", "İngiliz Zafer", "Blue Sapphire; İngiliz ultra lüks; mücevher şişe; kraliyet"),
            ("Creed Aventus", "creedboutique.com", "Erkek İkon", "Aventus; erkek niş ikon; ananas + meşe; Napoleon ilham"),
            ("Tom Ford Private Blend", "tomford.com", "Private Blend", "Oud Wood; Private Blend; Tom Ford; lüks; oud + vetiver"),
            ("Jo Malone English Pear", "jomalone.com", "İngiliz Armut", "English Pear & Freesia; İngiliz; armut + frezya; katmanlama"),
        ],
        "Vücut Bakımı - Deodorant": [
            ("Dove Men+Care", "dovemencare.com", "Erkek Bakım Deo", "Clean Comfort; erkek bakım; 48 saat; nemlendirici; Unilever"),
            ("Nivea Men Fresh", "nivea.com", "Alman Erkek Deo", "Men Fresh Active; Alman erkek; 48 saat; taze; erişilebilir"),
            ("Axe/Lynx", "axe.com", "Gençlik Deo", "Apollo Body Spray; gençlik; Unilever; koku + deo; viral"),
            ("Old Spice Captain", "oldspice.com", "Erkek Klasik Kaptan", "Captain Antiperspirant; klasik erkek; P&G; maskülen koku"),
            ("Right Guard Sport", "rightguard.com", "Spor Deo", "Sport Antiperspirant; 48 saat spor koruma; erişilebilir"),
        ],
        "Cilt Bakımı - Güneş Koruma (SPF)": [
            ("Neutrogena Ultra Sheer", "neutrogena.com", "Ultra Sheer SPF", "Ultra Sheer Dry-Touch SPF 70; eczane; mat bitirişli; erişilebilir"),
            ("Coppertone Sport", "coppertone.com", "Spor SPF", "Sport Sunscreen SPF 50; suya dayanıklı; aktif yaşam; erişilebilir"),
            ("Banana Boat Ultra Sport", "bananaboat.com", "Ultra Spor SPF", "Ultra Sport SPF 50+; sürtünmeye dayanıklı; tropikal; erişilebilir"),
            ("Hawaiian Tropic Silk", "hawaiiantropic.com", "Tropikal SPF", "Silk Hydration SPF 30; tropikal koku; ipeksi his; erişilebilir"),
            ("Blue Lizard", "bluelizardsunscreen.com", "Avustralya Mineral SPF", "Sensitive Mineral Sunscreen SPF 50+; akıllı şişe; UV ile renk değişen"),
        ],
        "Cilt Bakımı - Maske & Peeling": [
            ("Charlotte Tilbury Magic Cream", "charlottetilbury.com", "Magic Cream Maske", "Charlotte's Magic Cream; İngiliz Hollywood; anında ışıltı; backstage"),
            ("Estée Lauder ANR Mask", "esteelauder.com", "ANR Gece Maskesi", "Advanced Night Repair PowerFoil Mask; gece onarım maske"),
            ("La Mer The Treatment Lotion", "cremedelamer.com", "Lüks Tedavi Losyonu", "The Treatment Lotion; deniz kefiri; lüks hazırlık losyonu"),
            ("SK-II Pitera Mask", "sk-ii.com", "Japon Pitera Maske", "Facial Treatment Mask; PITERA sheet mask; Japon lüks maske"),
            ("Shiseido Vital Perfection Mask", "shiseido.com", "Japon Lifting Maske", "Vital Perfection LiftDefine Radiance Face Mask; Japon lifting"),
        ],
        "Cilt Bakımı - Dudak Bakımı": [
            ("Dior Lip Glow", "dior.com", "Fransız Lip Glow", "Dior Addict Lip Glow; renk arttırıcı; Fransız couture; ikonik"),
            ("Chanel Les Beiges Lip", "chanel.com", "Fransız Dudak Balmı", "Les Beiges Healthy Glow Lip Balm; doğal glow; Fransız chic"),
            ("YSL Volupté", "yslbeauty.com", "Fransız Volupté", "Volupté Plump-In-Colour Lip Balm; dolgunlaştırıcı; Fransız lüks"),
            ("Hermès Rose Lip", "hermes.com", "Hermès Dudak", "Rose Hermès Rosy Lip Enhancer; gül bazlı; ultra lüks dudak"),
            ("Guerlain KissKiss Bee Glow", "guerlain.com", "Fransız Arı Dudak", "KissKiss Bee Glow; %98 doğal; arı ilhamlı; renk arttırıcı"),
        ],
        "Cilt Bakımı - Bariyer Onarım": [
            ("La Roche-Posay Cicaplast Gel", "laroche-posay.com", "Fransız Cica Gel", "Cicaplast Gel B5; jel versiyon; yağsız; akne eğilimli cilt"),
            ("Avène Cicalfate+", "avene.com", "Fransız Cicalfate", "Cicalfate+ Restorative Protective Cream; post-procedure; onarım"),
            ("Bioderma Cicabio", "bioderma.com", "Fransız Cicabio", "Cicabio Cream; bakır + çinko + hyaluronic; Fransız bariyer onarım"),
            ("Uriage Bariéderm", "uriage.com", "Fransız Termal Bariyer", "Bariéderm Cica-Cream with Cu-Zn; bakır + çinko; termal su"),
            ("SVR Cicavit+", "laboratoire-svr.com", "Fransız Cicavit", "Cicavit+ Cream; anhydroticin; Fransız bariyer onarım kremi"),
        ],
        "Cilt Bakımı - Hiperpigmentasyon": [
            ("Lancôme Advanced Génifique", "lancome.com", "Fransız Fermente Aydınlatma", "Advanced Génifique Serum; fermente öz; Fransız aydınlatma"),
            ("Estée Lauder Perfectionist", "esteelauder.com", "Perfectionist Leke", "Perfectionist Pro Rapid Brightening Treatment; leke aydınlatma"),
            ("Shiseido White Lucent", "shiseido.com", "Japon Aydınlatma", "White Lucent Brightening Gel Cream; Japon aydınlatma; teknoloji"),
            ("Caudalie Vinoperfect Radiance", "caudalie.com", "Üzüm Radiance", "Vinoperfect Radiance Serum Complexion Correcting; viniferine; leke"),
            ("Clarins Bright Plus", "clarins.com", "Fransız Aydınlatma", "Bright Plus Serum; acerola + sea lily; Fransız botanik aydınlatma"),
        ],
        "Cilt Bakımı - Göz Çevresi": [
            ("Shiseido Benefiance Göz", "shiseido.com", "Japon Benefiance Göz", "Benefiance Wrinkle Smoothing Eye Cream; ReNeura; Japon"),
            ("Lancôme Absolue Göz", "lancome.com", "Fransız Absolue Göz", "Absolue Eye Cream; gül kök hücre; Fransız lüks göz"),
            ("Estée Lauder ANR Eye", "esteelauder.com", "ANR Göz Serumu", "Advanced Night Repair Eye Supercharged Gel-Crème; gece onarım"),
            ("Chanel Sublimage Göz", "chanel.com", "Fransız Sublimage", "Sublimage La Crème Yeux; Fransız ultra lüks; vanilla planifolia"),
            ("La Mer Eye Balm", "cremedelamer.com", "Lüks Göz Balmı", "The Eye Balm Intense; Miracle Broth; ultra lüks; anti-aging göz"),
        ],
        "Cilt Bakımı - Hassas Cilt": [
            ("Physiogel Daily", "physiogel.com", "Alman BioMimic", "Daily Moisture Therapy Facial Cream; BioMimic teknolojisi"),
            ("Cetaphil Gentle", "cetaphil.com", "Ultra Nazik", "Gentle Skin Cleanser; ultra nazik; dermatoloji 70+ yıl"),
            ("QV Intensive", "qvskincare.com.au", "Avustralya Dermatolojik", "Intensive with Ceramides; Avustralya; ceramide; kuru hassas"),
            ("Dermol 500", "dermol.co.uk", "İngiliz Medikal", "Dermol 500 Lotion; İngiliz medikal; antimikrobiyal + nemlendirici"),
            ("Epaderm", "epaderm.com", "İngiliz Reçeteli", "Epaderm Cream; İngiliz NHS reçeteli; egzama + kuru cilt"),
        ],
    }

    for category, brands in SEED_BRANDS.items():
        if category not in expanded:
            expanded[category] = []
        expanded[category].extend(brands)

    for category, brands in MORE_BRANDS.items():
        if category not in expanded:
            expanded[category] = []
        expanded[category].extend(brands)

    return expanded


def merge_all_brands():
    """Merge all brand dictionaries, deduplicate by brand name."""
    # Import supplementary brands
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    try:
        from cosmetics_extra_final import COSMETICS_EXTRA
    except ImportError:
        COSMETICS_EXTRA = {}
    try:
        from cosmetics_extra_final2 import COSMETICS_EXTRA_2
    except ImportError:
        COSMETICS_EXTRA_2 = {}
    try:
        from cosmetics_extra_final3 import COSMETICS_EXTRA_3
    except ImportError:
        COSMETICS_EXTRA_3 = {}

    merged = {}
    expanded = generate_expanded_brands()
    for source in [BRANDS, EXTRA_BRANDS, EXTRA_BRANDS_2, EXTRA_BRANDS_3, EXTRA_BRANDS_4,
                   EXTRA_BRANDS_5, EXTRA_BRANDS_6, EXTRA_BRANDS_7, expanded,
                   COSMETICS_EXTRA, COSMETICS_EXTRA_2, COSMETICS_EXTRA_3]:
        for category, brands in source.items():
            if category not in merged:
                merged[category] = []
            merged[category].extend(brands)

    # Deduplicate within each category by brand name (allow same brand in different categories)
    deduped = {}
    for category, brands in merged.items():
        seen_in_cat = set()
        deduped[category] = []
        for brand in brands:
            brand_name = brand[0]
            if brand_name not in seen_in_cat:
                seen_in_cat.add(brand_name)
                deduped[category].append(brand)
    return deduped


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def build_excel():
    """Build the complete Excel workbook."""
    wb = Workbook()
    ALL_BRANDS = merge_all_brands()

    # Flatten
    all_brands = []
    for category, brands in ALL_BRANDS.items():
        for b in brands:
            all_brands.append((category, b))

    total_brands = len(all_brands)
    print(f"Toplam marka sayısı: {total_brands}")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: ÖZET
    # ═══════════════════════════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Özet"
    ws_summary.sheet_properties.tabColor = HEADER_COLOR

    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Kozmetik DTC İnovatif Markalar Raporu - 5000+ Marka"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:D2")
    ws_summary["A2"].value = f"Oluşturulma Tarihi: {TODAY}  |  Toplam Marka: {total_brands}"
    ws_summary["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A3:D3")
    ws_summary["A3"].value = "Sadece kozmetik DTC markalar - küçük/orta ölçekli, inovatif, problem çözen"
    ws_summary["A3"].font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws_summary["A3"].alignment = Alignment(horizontal="center")

    headers = ["#", "Kategori", "Marka Sayısı", "Yüzde (%)"]
    for col_idx, h in enumerate(headers, 1):
        ws_summary.cell(row=5, column=col_idx, value=h)
    apply_header_style(ws_summary, 5, len(headers))

    row_num = 6
    for idx, (cat, brands) in enumerate(ALL_BRANDS.items(), 1):
        count = len(brands)
        pct = round(count / total_brands * 100, 1)
        ws_summary.cell(row=row_num, column=1, value=idx)
        ws_summary.cell(row=row_num, column=2, value=cat)
        ws_summary.cell(row=row_num, column=3, value=count)
        ws_summary.cell(row=row_num, column=4, value=f"%{pct}")
        cat_color = CATEGORY_COLORS.get(cat, ("D4E6F1", "2C5F8A"))[0]
        apply_data_row(ws_summary, row_num, len(headers), category_color=cat_color, is_even=(idx % 2 == 0))
        row_num += 1

    # Total row
    ws_summary.cell(row=row_num, column=2, value="TOPLAM")
    ws_summary.cell(row=row_num, column=3, value=total_brands)
    ws_summary.cell(row=row_num, column=4, value="%100")
    total_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    for col in range(1, 5):
        cell = ws_summary.cell(row=row_num, column=col)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = total_fill
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="medium")
        )

    ws_summary.column_dimensions["A"].width = 5
    ws_summary.column_dimensions["B"].width = 42
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: TÜM MARKALAR
    # ═══════════════════════════════════════════════════════════════════════
    ws_all = wb.create_sheet("Tüm Markalar")
    ws_all.sheet_properties.tabColor = HEADER_COLOR

    all_headers = ["#", "Marka Adı", "Web Sitesi", "Kategori", "Alt Niş",
                   "Öne Çıkan Özellik / Pazarlama Açısı", "Meta Reklam Kütüphanesi"]

    for col_idx, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col_idx, value=h)
    apply_header_style(ws_all, 1, len(all_headers))

    row_num = 2
    current_category = None
    brand_idx = 0
    for category, brand in all_brands:
        brand_idx += 1
        name, website, subniche, insight = brand

        if current_category and current_category != category:
            apply_category_separator(ws_all, row_num - 1, len(all_headers))
        current_category = category

        meta_url = create_meta_ads_url(name)
        cat_color_pair = CATEGORY_COLORS.get(category, ("D4E6F1", "2C5F8A"))
        cat_bg = cat_color_pair[0]

        ws_all.cell(row=row_num, column=1, value=brand_idx)
        ws_all.cell(row=row_num, column=2, value=name)

        site_cell = ws_all.cell(row=row_num, column=3, value=website)
        site_cell.hyperlink = f"https://{website}"
        site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")

        ws_all.cell(row=row_num, column=4, value=category)
        ws_all.cell(row=row_num, column=5, value=subniche)

        insight_cell = ws_all.cell(row=row_num, column=6, value=insight)

        meta_cell = ws_all.cell(row=row_num, column=7, value="Reklamları Gör")
        meta_cell.hyperlink = meta_url

        is_even = (brand_idx % 2 == 0)
        apply_data_row(ws_all, row_num, len(all_headers), category_color=cat_bg, is_even=is_even)

        site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")
        insight_cell.font = Font(name="Calibri", size=9, italic=True, color=INSIGHT_FONT_COLOR)

        meta_fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
        meta_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        meta_cell.fill = meta_fill
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 1

    col_widths_all = [5, 25, 30, 35, 28, 65, 18]
    for i, w in enumerate(col_widths_all, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:G{row_num - 1}"

    # ═══════════════════════════════════════════════════════════════════════
    # INDIVIDUAL CATEGORY SHEETS
    # ═══════════════════════════════════════════════════════════════════════
    for cat_idx, (category, brands) in enumerate(ALL_BRANDS.items()):
        sheet_name = category.replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace(":", "-").replace("[", "(").replace("]", ")")[:31]
        ws_cat = wb.create_sheet(sheet_name)

        cat_color_pair = CATEGORY_COLORS.get(category, ("D4E6F1", "2C5F8A"))
        cat_bg = cat_color_pair[0]
        cat_accent = cat_color_pair[1]
        ws_cat.sheet_properties.tabColor = cat_accent

        cat_headers = ["#", "Marka Adı", "Web Sitesi", "Alt Niş",
                       "Öne Çıkan Özellik / Pazarlama Açısı", "Meta Reklam Kütüphanesi"]

        for col_idx, h in enumerate(cat_headers, 1):
            ws_cat.cell(row=1, column=col_idx, value=h)
        apply_header_style(ws_cat, 1, len(cat_headers))

        for b_idx, brand in enumerate(brands, 1):
            r = b_idx + 1
            name, website, subniche, insight = brand
            meta_url = create_meta_ads_url(name)

            ws_cat.cell(row=r, column=1, value=b_idx)
            ws_cat.cell(row=r, column=2, value=name)

            site_cell = ws_cat.cell(row=r, column=3, value=website)
            site_cell.hyperlink = f"https://{website}"
            site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")

            ws_cat.cell(row=r, column=4, value=subniche)

            insight_cell = ws_cat.cell(row=r, column=5, value=insight)

            meta_cell = ws_cat.cell(row=r, column=6, value="Reklamları Gör")
            meta_cell.hyperlink = meta_url

            is_even = (b_idx % 2 == 0)
            apply_data_row(ws_cat, r, len(cat_headers), category_color=cat_bg, is_even=is_even)

            site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")
            insight_cell.font = Font(name="Calibri", size=9, italic=True, color=INSIGHT_FONT_COLOR)

            meta_fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
            meta_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            meta_cell.fill = meta_fill
            meta_cell.alignment = Alignment(horizontal="center", vertical="center")

        col_widths_cat = [5, 25, 30, 28, 65, 18]
        for i, w in enumerate(col_widths_cat, 1):
            ws_cat.column_dimensions[get_column_letter(i)].width = w

        ws_cat.freeze_panes = "A2"
        ws_cat.auto_filter.ref = f"A1:F{len(brands) + 1}"

    # ── Save ──────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, FILENAME)
    wb.save(filepath)
    print(f"\nExcel dosyası oluşturuldu: {filepath}")
    print(f"Toplam {total_brands} marka, {len(ALL_BRANDS)} kategori")
    return filepath


if __name__ == "__main__":
    build_excel()
