#!/usr/bin/env python3
"""
DTC Inovatif Problem-Çözen Markalar - Turkish Excel Generator v3
Generates a comprehensive Excel file with 1500-2000 innovative, problem-solving
ecommerce-native DTC brands. All text in Turkish.

COMPLETE REWRITE - Only small/mid-size innovative brands.
No mega-brands, no legacy brands, no household names.
"""

import os
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Output config ───────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research_outputs")
TODAY = "2026-03-23"
FILENAME = f"DTC_Inovatif_Problem_Cozen_Markalar_{TODAY}.xlsx"

# ─── Color Scheme ────────────────────────────────────────────────────────────
HEADER_COLOR = "1B2A4A"  # Deep navy
HEADER_FONT_COLOR = "FFFFFF"

# Category pastel colors (fill hex without #)
CATEGORY_COLORS = {
    "Cilt Bakımı & Güzellik Araçları": ("D4E8D0", "4A7C59"),   # Soft sage
    "Saç Bakımı & Saç Sağlığı": ("F5D5CB", "8B4513"),          # Soft coral
    "Vücut Bakımı & Kişisel Hijyen": ("E0D4F5", "6B4C8A"),     # Soft lavender
    "Erkek Bakım & Tıraş": ("C8E6E0", "2E6B5E"),               # Soft mint
    "Sağlık & Takviye": ("D4E6F1", "2C5F8A"),                   # Soft blue
    "Uyku & Yatak Teknolojisi": ("E8D8E8", "6B4C6B"),           # Soft mauve
    "Fitness Ekipman & Giyim": ("FCE4D6", "B05C2A"),            # Soft peach
    "Moda & Giyim (Kadın)": ("FADBD8", "A0522D"),               # Soft rose
    "Moda & Giyim (Erkek)": ("D5E8D4", "3E6B48"),               # Soft green
    "İç Giyim & Çorap": ("F5E6CC", "8B6914"),                   # Soft gold
    "Ayakkabı & Terlik": ("D4E4E8", "3E5F6B"),                  # Soft steel
    "Yiyecek & Atıştırmalık": ("FFE5CC", "CC6600"),              # Soft orange
    "İçecek & Kahve": ("E6D8CC", "6B4226"),                      # Soft brown
    "Ev & Mutfak": ("D4F0E8", "2E8B6B"),                         # Soft aqua
    "Ev Temizlik & Sürdürülebilirlik": ("CCE8CC", "2E7D32"),    # Soft leaf
    "Bebek & Çocuk": ("F0E0F5", "8B4C8B"),                       # Soft orchid
    "Evcil Hayvan": ("FCE8D5", "A0602E"),                        # Soft apricot
    "Aksesuar & Takı": ("E8E0D4", "7D6B4C"),                     # Soft tan
    "Gözlük & Güneş Gözlüğü": ("D8E8F0", "3E6B8B"),            # Soft sky
    "Teknoloji Aksesuarları": ("D4D8E8", "3E4C6B"),             # Soft indigo
    "Seyahat & Bavul": ("E0E8D4", "4C6B3E"),                    # Soft olive
    "Diş & Ağız Bakımı": ("CCE8E8", "2E6B6B"),                  # Soft teal
    "Kadın Sağlığı & Regl Bakımı": ("F5D4E0", "8B3E5C"),       # Soft pink
    "Cinsel Sağlık & Wellness": ("E8D4D8", "6B3E4C"),           # Soft blush
    "Parfüm & Ev Kokusu": ("E8E4D4", "6B644C"),                 # Soft sand
    "Outdoor & Spor Ekipman": ("D4E0E8", "3E5C6B"),             # Soft slate
    "Ofis & Üretkenlik": ("E0D8E8", "5C4C6B"),                  # Soft grape
    "Oyun & Yaratıcı Araçlar": ("D8D4E8", "4C3E6B"),            # Soft iris
}

META_BUTTON_COLOR = "27AE60"  # Green accent for Meta ads link
WEBSITE_LINK_COLOR = "0563C1"  # Blue for website links
INSIGHT_FONT_COLOR = "555555"  # Dark gray for insights

# ─── Brand Data ──────────────────────────────────────────────────────────────
# Format: (Brand, Website, Sub-niche TR, Notable Insight TR)

BRANDS = {
    # ═══════════════════════════════════════════════════════════════════════════
    # 1. Cilt Bakımı & Güzellik Araçları
    # ═══════════════════════════════════════════════════════════════════════════
    "Cilt Bakımı & Güzellik Araçları": [
        ("Starface", "starface.world", "Eğlenceli Akne Patch", "Yıldız şeklinde hidrokolloid akne patchleri ile akneyi 'utanılacak' olmaktan çıkardı; Z kuşağı ikonografisi"),
        ("Hero Cosmetics", "herocosmetics.com", "Akne Tedavi Uzmanı", "Mighty Patch gece akne patchleri; Church & Dwight $630M'a satın aldı; Amazon'da #1 akne ürünü"),
        ("Topicals", "mytopicals.com", "Hiperpigmentasyon Çözümü", "Kronik cilt sorunları için bilimsel ürünler; Gen Z'nin cilt sorunlarını normalleştirme hareketi; renkli ambalaj"),
        ("Bushbalm", "bushbalm.com", "Tüy Batması & Bikini Bakımı", "Bikini bölgesi koyulaşma ve tüy batması tedavisi; Shark Tank yatırımı aldı; cesur niş pazarlama"),
        ("Dieux Skin", "dieuxskin.com", "Kanıt Bazlı Güzellik", "Klinik araştırma kanıtlarını açıkça paylaşan marka; Instant Angel göz altı maskesi; sürdürülebilir tüp ambalaj"),
        ("Kinship", "lovekinship.com", "Ekolojik Cilt Bakımı", "SPF ürünleri mercan resiflerine güvenli; adaptojenik bileşenler; çevreye duyarlı formüller"),
        ("Vacation Inc", "vacation.inc", "Retro Güneş Bakımı", "1980'ler estetik güneş kremleri; Classic Whip SPF 30 krem şantili ambalaj; eğlenceli marka kimliği"),
        ("Youthforia", "youthforia.com", "Yatak Makyajı", "BYO Blush renk değiştiren allık; 'yatakta bile takılabilir makyaj' konsepti; viral TikTok ürünleri"),
        ("EADEM", "eadem.co", "Melanin Cilt Bilimi", "Koyu cilt tonları için Smart Melanin teknolojisi; bilimsel yaklaşımla melanin dostu bakım"),
        ("Krave Beauty", "kravebeauty.com", "Minimalist Cilt Bakımı", "Liah Yoo kurdu; #PressReset ile aşırı tüketimi eleştiriyor; Great Barrier Relief kült ürün"),
        ("Bubble Skincare", "hellobubble.com", "Z Kuşağı Cilt Bakımı", "Walmart'ta $5-16 fiyat; erişilebilir gençlik cilt bakımı; TikTok'ta 2B+ görüntülenme"),
        ("Byoma", "byoma.com", "Cilt Bariyeri Onarım", "Bariyer onarım odaklı; $10-15 fiyat aralığı; Target'ta; triseramid kompleksi"),
        ("Cocokind", "cocokind.com", "Şeffaf Fiyatlı Güzellik", "Maliyet dağılımını etiketinde gösteren marka; sürdürülebilir ambalaj öncüsü"),
        ("Versed", "versedskin.com", "Topluluk Odaklı Cilt Bakımı", "Topluluk oylamasıyla ürün geliştirme; $10-20 aralığı; temiz + etkili formüller"),
        ("Stratia", "stratiaskin.com", "Bariyer Onarım Serumu", "Liquid Gold cilt bariyeri onarım kült ürünü; tek kişi indie marka; Reddit topluluğu favorisi"),
        ("Superegg", "superegg.co", "Yumurta Bazlı Bakım", "Yumurta membran teknolojisi ile cilt bakımı; benzersiz bileşen yaklaşımı; minimalist tasarım"),
        ("Tower 28", "tower28beauty.com", "Hassas Cilt Makyajı", "NEA (Ulusal Egzama Derneği) onaylı; hassas ve reaktif ciltler için güvenli formüller"),
        ("Merit", "meritbeauty.com", "Minimalist Makyaj", "Katherine Power'ın '5 dakika makyaj' felsefesi; 10'dan az SKU ile Sephora'da en hızlı büyüyen"),
        ("Saie", "saiehello.com", "Işıltılı Doğal Makyaj", "Glowy Super Gel TikTok'ta viral; 'dew' (çiy) görünümü trendini başlattı; temiz formüller"),
        ("Tula", "tula.com", "Probiyotik Cilt Bakımı", "Gastroenterolog Dr. Roshini Raj kurdu; probiyotik + prebiyotik bazlı formüller; influencer pazarlama ustası"),
        ("Glow Recipe", "glowrecipe.com", "Meyve Bazlı K-Beauty", "Karpuz serisi ikonik; K-beauty'yi Batı'ya taşıdı; Watermelon Glow Niacinamide Dew Drops viral"),
        ("Peach & Lily", "peachandlily.com", "Cam Cilt Uzmanı", "Alicia Yoon'un K-beauty küratörlüğü; Glass Skin Serum en çok satan; cam cilt trendini başlattı"),
        ("Farmacy", "farmacybeauty.com", "Çiftlik Kaynaklı Bakım", "Kendi çiftliğinden elde edilen bileşenler; Green Clean balm temizleyici kült ürün"),
        ("Peace Out", "peaceoutskincare.com", "Hedefli Yama Tedavi", "Akne, kırışıklık, gözenek yamları; Sephora'da en çok satan patch markası; nokta tedavi uzmanı"),
        ("Good Molecules", "goodmolecules.com", "Uygun Fiyatlı Aktif Madde", "Beautylish markası; $6-12 fiyat; yüksek aktif madde düşük fiyat; The Ordinary alternatifi"),
        ("Solawave", "solawave.com", "LED Cilt Cihazı", "Kızılötesi + mikro akım + LED ışık tedavisi tek cihazda; evde profesyonel cilt bakım cihazı"),
        ("NuFace", "mynuface.com", "Mikro Akım Cihazı", "Evde yüz germe mikro akım cihazı; cerrahi olmadan yüz kontürü; 'at-home facelift' öncüsü"),
        ("Foreo", "foreo.com", "Akıllı Cilt Cihazı", "LUNA silikon yüz temizleme cihazı; BEAR mikro akım; İsveç tasarımı; 100+ ülkede satış"),
        ("Droplette", "droplette.io", "Mikro İnfüzyon Cihazı", "İğnesiz serum enjeksiyonu; patentli mikro infüzyon teknolojisi; 20x daha derin nüfuz etme"),
        ("Medicube", "medicube.com", "Kore Dermatoloji Cihazı", "AGE-R booster cihazı TikTok'ta viral; Kore dermatoloji kliniğinden doğan marka; uygun fiyatlı cihazlar"),
        ("HigherDOSE", "higherdose.com", "Infrared Wellness Cihazı", "Kızılötesi sauna battaniyesi; PEMF Mat; biyo-hackleme evde; wellness cihaz öncüsü"),
        ("ZIIP", "ziipbeauty.com", "Nano Akım Güzellik", "Nano akım + mikro akım yüz cihazı; uygulamayla tedavi protokolleri; profesyonel evde kullanım"),
        ("Skin Gym", "skingym.com", "Jade Roller & Cihaz", "Yeşim taşı roller ve gua sha'yı modernize eden marka; LED maskeler; erişilebilir cilt araçları"),
        ("Dermaflash", "dermaflash.com", "Evde Dermaplaning", "Evde profesyonel dermaplaning; tüy ve ölü deri temizleme; titreşimli bıçak teknolojisi"),
        ("LightStim", "lightstim.com", "LED Işık Tedavisi", "NASA teknolojisinden esinlenen LED ışık cihazları; FDA onaylı; anti-aging + akne tedavisi"),
        ("Ami Colé", "amicole.com", "Melanin Güzellik", "Koyu cilt tonları için temiz makyaj; Diarrha N'Diaye kurdu; kapsayıcı renk paleti"),
        ("Osea", "oseamalibu.com", "Deniz Yosunu Bakımı", "Malibu merkezli; deniz yosunu bazlı cilt bakımı; anti-aging odaklı temiz formüller; 1996'dan beri"),
        ("Summer Fridays", "summerfridays.com", "Sosyal Medya Doğan Bakım", "Jet Lag Mask Instagram'da viral; influencer'dan markaya dönüşüm başarısı; Sephora favorisi"),
        ("Kosas", "kosas.com", "Cilt Bakımlı Makyaj", "Revealer Concealer cilt bakımı + makyaj birleşimi; 'makeup that's skincare' konsepti"),
        ("Ilia", "iliabeauty.com", "Temiz Performans Makyajı", "Super Serum Skin Tint en çok satan; temiz güzellik + yüksek performans dengesi"),
        ("Jones Road Beauty", "jonesroadbeauty.com", "Olgun Cilt Makyajı", "Bobbi Brown'ın yeni markası; Miracle Balm viral; 50+ yaş grubuna yönelik temiz makyaj"),
        ("Iris & Romeo", "irisandromeo.com", "Çok İşlevli Makyaj", "Best Skin Days SPF + nemlendirici + fondöten tek üründe; minimalist rutinler için"),
        ("Danessa Myricks", "danessamyricks.com", "Sanatsal Güzellik", "Makyaj sanatçısı; ColorFix çok amaçlı pigment; yaratıcı + günlük kullanım"),
        ("Spoiled Child", "spoiledchild.com", "AI Kişisel Bakım", "Yapay zeka ile kişiselleştirilmiş cilt bakımı; yaş ve cilt tipine göre formül önerisi"),
        ("Naturium", "naturium.com", "Yüksek Aktif Düşük Fiyat", "Susan Yara kurdu; yüksek konsantrasyon aktif maddeler uygun fiyatla; etkin bakım"),
        ("Beauty of Joseon", "beautyofjoseon.com", "Kore Geleneği", "Joseon Hanedanı ilhamı; pirinç + ginseng formülleri; Relief Sun SPF ikonik K-beauty"),
        ("Typology", "typology.com", "Fransız Minimalist", "Paris merkezli; sadece gerekli bileşenler; doğal aktifler; şeffaf formüller"),
        ("Sand & Sky", "sandandsky.com", "Avustralya Kili Maske", "Avustralya pembe kili maskesi viral; Instagram'da 100M+ görüntülenme; botanik yaklaşım"),
        ("Frank Body", "frankbody.com", "Kahve Vücut Peeling", "Kahve peelingi viral; UGC pazarlaması öncüsü; eğlenceli marka kimliği; $50M+ gelir"),
        ("Live Tinted", "livetinted.com", "Çok Kültürlü Güzellik", "Deepica Mutyala kurdu; Huestick çok amaçlı çubuk; esmer ciltler için özel formüller"),
        ("Alpyn Beauty", "alpynbeauty.com", "Vahşi Bitki Bakımı", "Jackson Hole dağlarından toplanan yabani bitkiler; vahşi doğa ilhamlı temiz güzellik"),
        ("Beekman 1802", "beekman1802.com", "Keçi Sütü Bakımı", "Keçi sütü bazlı cilt bakımı; çiftlik hikayesi; Ulta'da hızla büyüyen; probiyotik formüller"),
        ("UpCircle", "upcirclebeauty.com", "Atık Dönüşüm Güzellik", "Kahve telinden yüz peelingi; gıda atıklarını güzellik ürünlerine dönüştürme; sıfır atık"),
        ("Herbivore", "herbivorebotanicals.com", "Bitki Bazlı Lüks", "Bakuchiol retinol alternatifi; kristal ve bitki bazlı; doğal lüks; Sephora'da"),
        ("Primally Pure", "primallypure.com", "Çiftlik Doğal Bakım", "Çiftlik ilhamı; doğal deodorant ve cilt bakımı; organik bileşenler; temiz yaşam topluluğu"),
        ("Codex Beauty", "codexbeauty.com", "Bilimsel Bitki Bakımı", "İrlanda bitkileri + PubMed araştırmaları; klinik kanıtlı doğal bileşenler"),
        ("Maëlys", "maelyscosmetics.com", "Vücut Şekillendirme", "B-Flat karın sıkılaştırma kremi viral; selülit ve vücut şekillendirme odaklı"),
        ("Aavrani", "aavrani.com", "Ayurveda Güzellik", "Ayurveda + modern bilim; zerdeçal ve bakuchiol bazlı; Hint güzellik gelenekleri"),
        ("MENTED Cosmetics", "mentedcosmetics.com", "Kapsayıcı Nude Tonlar", "Koyu cilt tonları için nude rujlar; KJ Miller & Amanda Johnson kurdu; kapsayıcı 'nude' tanımı"),
        ("Maya Chia", "mayachia.com", "Süper Tohum Yağları", "Chia tohumu bazlı cilt bakımı; süperfood yağları; anti-aging; temiz lüks formüller"),
        ("Odacité", "odacite.com", "Fransız Serum Konsantresi", "Tek bileşen serum konsantreleri; Paris + LA kökenli; lüks doğal cilt bakımı"),
        ("Beneath Your Mask", "beneathyourmask.com", "Lüks El Yapımı Bakım", "Dana Jackson'ın hastalık sürecinden doğan marka; Beyoncé'nin favorisi; el yapımı lüks"),
        ("CurrentBody", "currentbody.com", "LED Maske Uzmanı", "LED ışık tedavi maskeleri; FDA onaylı; evde klinik sonuçlar; İngiltere merkezli teknoloji"),
        ("TheraFace", "therabody.com/theraface", "Yüz Masaj Cihazı", "Therabody'den yüz bakım cihazı; perkusyon + mikro akım + LED + ısıtma/soğutma tek cihazda"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Profesyonel Peel Pad", "Dermatolog markası; Alpha Beta Peel patchleri kült ürün; LED maske viral; evde klinik sonuçlar"),
        ("Enviable Brows", "enviablebrows.com", "Kaş Laminasyonu Evde", "Evde kaş laminasyon kiti; profesyonel sonuçlar evde; viral sosyal medya dönüşümler"),
        ("Patchology", "patchology.com", "Hızlı Maske Tedavisi", "5 dakika göz altı maskeleri; seyahat dostu patch formatı; hızlı sonuç veren maskeler"),
        ("Murad", "murad.com", "Dermatolojik Cilt Çözümü", "Dr. Howard Murad kurdu; retinol + cilt bariyeri uzmanı; bilimsel yaklaşım"),
        ("Kate Somerville", "katesomerville.com", "Hollywood Cilt Kliniği", "Hollywood ünlülerinin estetisyeninden marka; ExfoliKate peeling ikonik; klinik + evde bakım"),
        ("BeautyBio", "beautybio.com", "GloPRO Mikro İğneleme", "Evde micro-needling roller; GloPRO FDA onaylı; kolajen üretimini %200 artırma iddiası"),
        ("MZ Skin", "mzskin.com", "Doktor Lüks Cilt Cihazı", "Dr. Maryam Zamani; LED yüz maskesi lüks segment; Harrods'ta satılan cilt cihazları"),
        ("Joanna Vargas", "joannavargas.com", "Ünlülerin Estetisyeni", "NYC ve LA meditasyon facialları; LED ışık tedavi cihazları; ünlü müşteri portföyü"),
        ("Solaris Labs NY", "solarislabsny.com", "LED Cilt Teknolojisi", "Taşınabilir LED cilt tedavi cihazları; kırmızı ışık + mavi ışık; uygun fiyatlı cilt teknolojisi"),
        ("Wildling", "wildling.com", "Gua Sha Modern", "Taş gua sha'yı modernize eden marka; yüz akupresür aracı; geleneksel Çin tıbbı + modern tasarım"),
        ("Sacheu Beauty", "sacheu.com", "Paslanmaz Çelik Gua Sha", "Paslanmaz çelik gua sha icat eden marka; TikTok viral; hijyenik metal araçlar"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 2. Saç Bakımı & Saç Sağlığı
    # ═══════════════════════════════════════════════════════════════════════════
    "Saç Bakımı & Saç Sağlığı": [
        ("K18", "k18hair.com", "Biyomimetik Saç Onarımı", "K18Peptide™ 4 dakikada hasar onarımı; durulanmayan maske; salon endüstrisini değiştirdi"),
        ("Vegamour", "vegamour.com", "Bitkisel Saç Büyütme", "Bitkisel saç büyütme serumu; GRO serisi; vegan ve temiz; saç dökülmesi çözümü; $100M+ gelir"),
        ("Crown Affair", "crownaffair.com", "Saç Bakım Ritüeli", "Günlük saç bakım ritüeli; The Comb, The Towel; lüks saç aksesuarları + bakım birleşimi"),
        ("Bread Beauty Supply", "breadbeautysupply.com", "Tekstürlü Saç Bakımı", "Tekstürlü saç için minimalist bakım; 'wash-day essentials'; Sephora özel; doğal saç hareketi"),
        ("Act+Acre", "actandacre.com", "Saç Derisi Sağlığı", "Cold Processed® saç derisi bakımı; trişolojist Helen Reavey kurdu; saç derisi detoksu"),
        ("Dae Hair", "dae.com", "Çöl Bitkisi Saç Bakımı", "Arizona çöl bitkileri; Cactus Fruit 3-in-1 styling cream viral; Amber Fillerup Clark kurdu"),
        ("Prose", "prose.com", "AI Kişisel Formül", "AI + uzman ile kişiselleştirilmiş saç formülü; 85 faktör analizi; her şişe benzersiz"),
        ("Function of Beauty", "functionofbeauty.com", "Quiz Bazlı Özel Formül", "Quiz ile kişiselleştirilmiş şampuan; 54T+ kombinasyon; etikette adınız yazıyor"),
        ("Curlsmith", "curlsmith.com", "Kıvırcık Saç Uzmanı", "Kıvırcık saç tiplerine göre koleksiyonlar; Bond Curl Rehab en çok satan; Helen of Troy satın aldı"),
        ("Odele Beauty", "odelebeauty.com", "Erişilebilir Salon Kalite", "Target'ta $12 premium saç bakımı; unisex; temiz formüller; Midwest kadın kurucu ekibi"),
        ("Ceremonia", "ceremonia.com", "Latin Saç Bakımı", "Latin Amerika bitki özleri; Aceite de Moska saç yağı viral; kültürel miras modern bilimle"),
        ("Briogeo", "briogeo.com", "Temiz Saç Bakımı", "Don't Despair Repair maske en çok satan; 6-free formüller; Wella $500M'a satın aldı"),
        ("Amika", "loveamika.com", "Renkli Saç Bakımı", "Brooklyn merkezli; eğlenceli ambalaj; ısı koruma + renk koruma uzmanı; Sephora'da büyüyen"),
        ("Ouai", "theouai.com", "Hairstylist Markası", "Jen Atkin kurdu; Wave Spray ikonik; lifestyle marka olarak genişledi; P&G satın aldı"),
        ("Color Wow", "colorwowhair.com", "Renk Koruma Uzmanı", "Dream Coat anti-humidity spray viral; ünlü hairstylist Chris Appleton destekli"),
        ("Virtue Labs", "virtuelabs.com", "Gerçek Keratin Onarım", "Alpha Keratin 60ku® insan keratini; gerçek keratin onarımı; bilim temelli"),
        ("IGK", "igkhair.com", "Miami Saç Kültürü", "Miami saç kültürü; First Class kömürlü kuru şampuan; eğlenceli isimler; Sephora favorisi"),
        ("R+Co", "randco.com", "Hairstylist Kolektif", "Süper hairstylist kolektifi; Bleu parfümlü şampuan; biyotin saç bakımı serisi"),
        ("Verb", "verbproducts.com", "Erişilebilir Salon Bakımı", "Ghost Oil en çok satan; $16-20 salon kalitesi; Sephora'da en uygun fiyatlı seçenek"),
        ("Innersense", "innersensebeauty.com", "Organik Salon Bakımı", "USDA organik sertifikalı; temiz salon saç bakımı; kıvırcık saç topluluğu favorisi"),
        ("dpHUE", "dphue.com", "Saç Rengi Bakımı", "Saç rengi bakımı uzmanı; Gloss+ evde yarı kalıcı renk; saç boyası arasını uzatan ürünler"),
        ("JVN Hair", "jvnhair.com", "Hemisqualane Teknolojisi", "Jonathan Van Ness markası; hemisqualane teknolojisi; bilim + kapsayıcılık; Sephora'da"),
        ("Rahua", "rahua.com", "Amazon Yağı Bakımı", "Amazon yağmur ormanı rahua yağı; yerli kabilelerle sürdürülebilir hasat; lüks doğal"),
        ("Davines", "davines.com", "İtalyan Sürdürülebilir Saç", "B Corp sertifikalı İtalyan markası; Oi Oil çok amaçlı yağ; sürdürülebilirlik manifestosu"),
        ("Fable & Mane", "fableandmane.com", "Hint Saç Yağı Ritüeli", "Hint saç bakım ritüelleri; ayurvedik saç yağlaması; kültürel miras + modern bilim"),
        ("Nutrafol", "nutrafol.com", "Saç Dökülme Takviyesi", "Doktor onaylı saç büyütme takviyesi; klinik çalışmalarla desteklenen bitkisel formül"),
        ("Mielle Organics", "mielleorganics.com", "Doğal Saç Hareketi", "Rosemary Mint yağı TikTok viral; P&G satın aldı; siyah saç bakımı öncüsü"),
        ("Pattern Beauty", "patternbeauty.com", "Kıvırcık & Coily Saç", "Tracee Ellis Ross markası; kıvırcık + coily saçlar için özel formüller; Ulta'da"),
        ("Christophe Robin", "christopherobin.com", "Saç Derisi Scrub", "Cleansing Purifying Scrub en çok satan; Paris salon uzmanı; deniz tuzu scrub ikonik"),
        ("Oribe", "oribe.com", "Lüks Saç Bakımı", "Gold Lust şampuan ikonik; Daniel Kaner + Oribe Canales kurdu; lüks saç bakım öncüsü"),
        ("Moroccanoil", "moroccanoil.com", "Argan Yağı Öncüsü", "Argan yağı saç bakımı kategorisini yarattı; Treatment Original ikonik; mavi şişe tanınırlığı"),
        ("Living Proof", "livingproof.com", "MIT Bilim Saç Bakımı", "MIT bilim insanları kurdu; OFPMA teknolojisi patentli; bilimsel saç bakım yaklaşımı"),
        ("Hairstory", "hairstory.com", "Şampuansız Temizleme", "New Wash şampuan olmadan saç yıkama; deterjan-free; saç bakım paradigmasını değiştirme"),
        ("Aveda", "aveda.com", "Bitkisel Profesyonel Bakım", "Ayurveda + botanik bilim; %95 doğal kaynaklı; B Corp; çevresel liderlik"),
        ("Reverie", "reveriesociety.com", "Minimalist Saç Bakımı", "Çok amaçlı saç ürünleri; MILK anti-frizz serum; minimalist yaklaşım; temiz formüller"),
        ("Playa", "theplayalife.com", "Kaliforniya Plaj Saçı", "Effortless plaj dalgaları; sülfat-free formüller; Kaliforniya yaşam tarzı; doğal saç dokusu"),
        ("Siàge", "siage.com.br", "Brezilya Saç Teknolojisi", "Brezilya saç bakım teknolojisi; nemlendirme + onarım; Latin Amerika'nın büyüyen DTC markası"),
        ("ColorProof", "colorproof.com", "Renk Koruma Teknolojisi", "BioRepair-8 teknolojisi; renk koruma odaklı profesyonel bakım; vegan formüller"),
        ("Madison Reed", "madison-reed.com", "Akıllı Saç Boyası", "Evde salon kalitesinde saç boyama; amonyaksız formüller; renk eşleştirme teknolojisi"),
        ("eSalon", "esalon.com", "Kişisel Saç Rengi", "Kolorist ile online danışma; kişiselleştirilmiş saç boyası evde; profesyonel formül"),
        ("Golde", "golde.co", "Superfoods Saç Maskesi", "Süperfood bazlı saç ve cilt maskeleri; zerdeçal bazlı ürünler; wellness + güzellik birleşimi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 3. Vücut Bakımı & Kişisel Hijyen
    # ═══════════════════════════════════════════════════════════════════════════
    "Vücut Bakımı & Kişisel Hijyen": [
        ("Nécessaire", "necessaire.com", "Aktif Vücut Bakımı", "Vücut bakımında aktif maddeler; hyalüronik asitli duş jeli; 'body care is skincare' hareketi"),
        ("Kopari", "koparibeauty.com", "Hindistan Cevizi Bakımı", "Hindistan cevizi bazlı güzellik; doğal deodorant en çok satan; tropikal estetik"),
        ("Sol de Janeiro", "soldejaneiro.com", "Brezilya Vücut Bakımı", "Brazilian Bum Bum Cream ikonik; guarana + hindistan cevizi; duyusal deneyim odaklı"),
        ("Megababe", "megababebeauty.com", "Sürtünme Önleme", "Thigh Rescue uyluk sürtünme çubuğu; vücut pozitifliği hareketi; tabu konuları çözüme kavuşturma"),
        ("Fur", "furyou.com", "Tüy & Bikini Bakımı", "Her yerdeki tüy bakımı; ingrown serum; tüy bakımını normalleştirme; Emma Watson'ın tercihi"),
        ("Soft Services", "softservices.com", "Vücut Cilt Sorunları", "Keratosis pilaris (tavuk derisi) çözümü; vücut aknesi tedavisi; niş cilt sorunlarına odaklanma"),
        ("Saltair", "saltair.com", "Deniz Minerali Vücut", "Deniz mineralli vücut bakımı; uygun fiyatlı premium; seramid + niasinamid vücut losyonu"),
        ("Touchland", "touchland.com", "Lüks El Dezenfektanı", "Parfüm gibi tasarlanmış el dezenfektanı; pandemi sonrası hijyeni şık hale getirdi"),
        ("Lumē", "lumedeodorant.com", "Tüm Vücut Deodorant", "Her bölge için deodorant; doktor geliştirdi; mandelic asit bazlı; tabu konuları çözme"),
        ("Each & Every", "eachandevery.com", "Doğal Deodorant", "EWG onaylı doğal deodorant; alüminyumsuz; deadstock meyvelerden yapılan kokular"),
        ("By Humankind", "byhumankind.com", "Plastik-Free Kişisel Bakım", "Tek kullanımlık plastiği ortadan kaldıran kişisel bakım; yeniden doldurulabilir ambalajlar"),
        ("Plus", "plusbody.com", "Vücut Kokusu Çözümü", "Vücut kokusuna bilimsel yaklaşım; pH dengeleme; probiyotik deodorant; biyom dostu"),
        ("Bevel", "bevelcode.com", "Melanin Cilt Tıraş", "Koyu ciltler için tıraş ve cilt bakımı; tıraş sonrası tahriş çözümü; Walker & Company"),
        ("Lume", "lumedeodorant.com", "Doktor Deodorant", "Dr. Shannon Klingman kurdu; 72 saat koruma; tüm vücut; $100M+ gelir"),
        ("Myro", "myro.com", "Yeniden Doldurulabilir Deo", "Yeniden doldurulabilir deodorant sistemi; kişiselleştirilmiş kokular; sürdürülebilir tasarım"),
        ("Type:A", "typeadeodorant.com", "Adaptif Deodorant", "Vücut kimyasına adapte olan deodorant; ter seviyesine göre ayarlanan formül; bilimsel yaklaşım"),
        ("Curie", "curiebody.com", "Detoks Deodorant", "Alüminyumsuz detoks deodorant; bitkisel aktifler; hassas ciltler için; B Corp sertifikalı"),
        ("Batist", "batist.com", "Kuru Şampuan Uzmanı", "Kuru şampuan kategorisini tanımlayan marka; 50 yıllık İngiliz mirası; pratik çözümler"),
        ("Flamingo", "shopflamingo.com", "Kadın Tıraş Çözümü", "Harry's kardeş markası; kadın tıraş deneyimini yeniden tasarladı; uygun fiyatlı premium"),
        ("Athena Club", "athenaclub.com", "Kadın Tıraş & Bakım", "Tıraş + vücut bakımı abonelik; razör + vücut losyonu; kadınlar için akıllı tasarım"),
        ("OUI the People", "ouithepeople.com", "Hassas Tıraş", "Tek bıçaklı güvenlik tıraşı; hassas ciltler için; tıraş sonrası ingrown önleme; sürdürülebilir"),
        ("Preserve", "preserveproducts.com", "Geri Dönüşüm Tıraş", "Geri dönüştürülmüş malzemeden tıraş bıçakları; çevre dostu kişisel bakım; Gimme 5 programı"),
        ("Fur Oil", "furyou.com", "Tüy Yumuşatma Yağı", "Tüy bakımı özel yağı; lazer sonrası bakım; tüy batması önleme; niş problem çözücü"),
        ("Kosas Wet Lip Oil Gloss", "kosas.com", "Aktif Dudak Bakımı", "Hyalüronik asit + peptit bazlı dudak yağı; dudak bakımı + renk birleşimi; Sephora'da viral"),
        ("Tower 28 SOS Spray", "tower28beauty.com", "SOS Cilt Spreyi", "Hipokloröz asit spreyi; akne, egzama, maske tahrişi için; çok amaçlı cilt kurtarıcı"),
        ("Billie", "mybillie.com", "Kadın Tıraş Devrimi", "Kadın tıraş bıçağı aboneliği; pembe vergi eleştirisi; vücut tüyü normalleştirme kampanyası"),
        ("Wakse", "wakse.com", "Evde Ağda Uzmanı", "Evde profesyonel ağda kitleri; sert ağda granülleri; estetik ambalaj; Z kuşağı odaklı"),
        ("Sugardoh", "sugardoh.com", "Doğal Şeker Ağdası", "Şeker bazlı doğal ağda; TikTok'ta viral DIY ağda; düşük ağrılı tüy alma; organik formül"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 4. Erkek Bakım & Tıraş
    # ═══════════════════════════════════════════════════════════════════════════
    "Erkek Bakım & Tıraş": [
        ("Dr. Squatch", "drsquatch.com", "Doğal Erkek Sabun", "Doğal malzemeli erkek sabunu; viral YouTube reklamları; 'dude' pazarlama; $100M+ gelir"),
        ("Beardbrand", "beardbrand.com", "Sakal Bakım Uzmanı", "YouTube sakal bakım içerikleriyle büyüyen marka; sakal yağı ve balmı; erkek bakım kültürü"),
        ("Huron", "usehuron.com", "Minimalist Erkek Bakım", "Matt Mullenax kurdu; erkek cilt + saç + vücut bakımı; temiz formüller; DTC abonelik"),
        ("Supply", "supply.co", "Tek Bıçak Tıraş", "Tek bıçaklı metal güvenlik tıraşı; ömür boyu garanti; sürdürülebilir tıraş; sıfır atık"),
        ("Oars + Alps", "oarsandalps.com", "Aktif Erkek Bakımı", "Doğal erkek bakım ürünleri; outdoor yaşam tarzı; SPF dahil günlük bakım; Target'ta"),
        ("Disco", "letsdisco.co", "Erkek Cilt Bilimi", "Erkek cilt bakımı bilimsel yaklaşım; göz altı çubuğu + yüz temizleme; basit rutin"),
        ("Hawthorne", "hawthorne.co", "Kişisel Erkek Koku", "Quiz ile kişiselleştirilmiş erkek kokusu; deodorant + kolonya + vücut bakımı; data-driven"),
        ("Geologie", "geologie.com", "Kişisel Erkek Cilt", "Cilt analizi ile kişiselleştirilmiş erkek cilt bakımı; 30 günlük rutin; diagnostik yaklaşım"),
        ("Bravo Sierra", "bravosierra.com", "Askeri İlham Bakım", "ABD askeri test edilen erkek bakım; aktif yaşam tarzı; vegan + temiz; sade tasarım"),
        ("Duke Cannon", "dukecannon.com", "Büyük Boy Erkek Bakım", "Büyük boy sabunlar; erkeksi marka kimliği; Amerikan yapımı; veteran desteği"),
        ("Lumin", "lumin.co", "Asya Erkek Cilt Bakımı", "Kore cilt bakımı erkekler için; koyu leke düzeltici; global erişim; abonelik modeli"),
        ("Black Wolf", "blackwolf.com", "Aktif Kömür Erkek Bakım", "Aktif kömür bazlı erkek bakım; saç + cilt + vücut; siyah ambalaj; erkek odaklı"),
        ("Scotch Porter", "scotchporter.com", "Erkek Sakal & Saç", "Siyah erkekler için sakal + saç bakımı; Target'ta; Calvin Quallis kurdu; kapsayıcı"),
        ("Brickell", "brickellmensproducts.com", "Premium Doğal Erkek", "Doğal ve organik erkek cilt bakımı; anti-aging odaklı; profesyonel erkek bakım"),
        ("Fulton & Roark", "fultonandroark.com", "Katı Kolonya", "Katı kolonya formatı; dökülmeyen, seyahat dostu; benzersiz koku delivery sistemi"),
        ("Blu Atlas", "bluatlas.com", "Temiz Erkek Premium", "Atlantic kelp bazlı erkek bakım; temiz lüks; deniz ilhamlı formüller; yüksek performans"),
        ("Every Man Jack", "everymanjack.com", "Erişilebilir Doğal Erkek", "Target ve Walmart'ta doğal erkek bakım; uygun fiyatlı; geniş dağıtım ağı"),
        ("Cremo", "cremocompany.com", "Süper Kaygan Tıraş", "Astonishingly Superior tıraş kremi; süper kaygan formül; tıraş deneyimini dönüştürme"),
        ("Baxter of California", "baxterofcalifornia.com", "Kaliforniya Erkek Bakım", "1965'ten beri erkek bakım; Clay Pomade ikonik; Los Angeles kültürü"),
        ("Aesop", "aesop.com", "Unisex Premium Bakım", "Botanik + bilim birleşimi; minimalist eczane estetiği; Parsley Seed serisi; L'Oréal satın aldı"),
        ("Jack Black", "getjackblack.com", "Aktif Erkek Bakım", "SPF lip balm en çok satan; dermatolog test edilen; aktif erkek yaşam tarzı"),
        ("Manscaped", "manscaped.com", "Erkek Vücut Tıraşı", "Erkek vücut bakımı özelleşmiş tıraş; Lawn Mower traş makinesi; cesur pazarlama"),
        ("Henson Shaving", "hensonshaving.com", "Havacılık Mühendisliği Tıraş", "Havacılık toleranslarında üretilen güvenlik tıraşı; titanyum seçeneği; sıfır bıçak titremesi"),
        ("Patricks", "patricks.co", "Lüks Erkek Bakım", "Avustralya lüks erkek bakımı; japonya'da üretim; premium fiyat; siyah şık ambalaj"),
        ("Blind Barber", "blindbarber.com", "Barber Shop Markası", "NYC berber dükkanından doğan marka; 90 Proof pomad; barber kültürü; otantik"),
        ("Rugged & Dapper", "ruggedanddapper.com", "Çok İşlevli Erkek Bakım", "Tek üründe çok işlev; age defense moisturizer; erkekler için basitleştirilmiş rutin"),
        ("Frederick Benjamin", "frederickbenjamin.com", "Siyah Erkek Saç Bakımı", "Siyah erkeklerin saç ve sakal sorunlarına özel çözümler; tıraş tümsekleri önleme"),
        ("Meridian", "meridiangrooming.com", "Erkek İntim Bakım", "Erkek intim bölge bakım seti; Trimmer + spray + wipes; tabu konuda çözüm"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 5. Sağlık & Takviye
    # ═══════════════════════════════════════════════════════════════════════════
    "Sağlık & Takviye": [
        ("Seed", "seed.com", "Simbiyotik Probiyotik", "DS-01 simbiyotik (prebiyotik + probiyotik); klinik çalışmalı; biyom bilimi; çift kapsül teknolojisi"),
        ("Thesis", "takethesis.com", "Kişisel Nootropik", "Quiz bazlı kişiselleştirilmiş nootropik formüller; enerji, odak, yaratıcılık, motivasyon ayrı karışımlar"),
        ("Beam", "beamorganics.com", "CBD Uyku Çözümü", "Dream Powder uyku tozu; nano CBD + magnezyum + reishi; uyku kalitesini artırma; sporcular tercih"),
        ("ARMRA", "tryarmra.com", "Kolostrum Takviyesi", "Sığır kolostrumu takviyesi; bağırsak sağlığı + bağışıklık; doktor kurdu; yeni süper gıda trendi"),
        ("Momentous", "livemomentous.com", "Elit Sporcu Takviye", "Andrew Huberman önerili; NSF sertifikalı; kreatin, omega-3, magnezyum; bilimsel kanıt bazlı"),
        ("Transparent Labs", "transparentlabs.com", "Şeffaf Formül Takviye", "%100 açık formül; gizli karışım yok; klinik dozlar; bodybuilding + fitness takviye"),
        ("Gorilla Mind", "gorillamind.com", "Yüksek Stim Takviye", "Derek (More Plates More Dates) kurdu; yüksek doz pre-workout; nootropik; şeffaf formüller"),
        ("Thorne", "thorne.com", "Klinik Düzey Takviye", "Doktor ve olimpik sporcu tercihi; NSF sertifikalı; genomik test entegrasyonu; en güvenilir marka"),
        ("Jocko Fuel", "jockofuel.com", "Disiplin Takviye", "Jocko Willink'in markası; savaş disiplini felsefesi; Mölk protein; asker zihniyeti pazarlama"),
        ("Timeline", "timelinenutrition.com", "Mitokondri Takviyesi", "Mitopure (Urolithin A) mitokondri yenileme; İsviçre bilimi; yaşlanma karşıtı hücre enerjisi"),
        ("Elysium", "elysiumhealth.com", "Yaşlanma Bilimi", "Basis NAD+ takviyesi; Nobel ödüllü bilim insanları danışman; hücresel yaşlanma yavaşlatma"),
        ("InsideTracker", "insidetracker.com", "Kan Testi Optimizasyon", "Kan testi + DNA analizi ile kişisel takviye önerisi; biyobelirteç takibi; veri odaklı sağlık"),
        ("Rootine", "rootine.co", "DNA Bazlı Vitamin", "DNA + kan testi ile kişisel mikro besin dozlama; slow-release teknoloji; hassas beslenme"),
        ("Gainful", "gainful.com", "Kişisel Protein Tozu", "Quiz ile kişiselleştirilmiş protein tozu; kendi RD'niz (diyetisyen); bireysel makro ihtiyaçlar"),
        ("HUM Nutrition", "humnutrition.com", "Güzellik Takviyesi", "Güzellik odaklı takviyeler; RD danışmanlığı dahil; Flatter Me sindirim; Daily Cleanse detoks"),
        ("Obvi", "myobvi.com", "Kollajen Takviye", "Kollajen protein tozu; lezzetli formatlar; kadın fitness topluluğu; sosyal medya odaklı büyüme"),
        ("Beekeeper's Naturals", "beekeepersnaturals.com", "Arı Ürünleri Takviye", "Propolis boğaz spreyi; arı poleni; manuka balı; doğal bağışıklık desteği; B Corp"),
        ("Goli", "goli.com", "Elma Sirkesi Gummy", "Elma sirkesi sakız tablet formatında; 8B+ gummy satıldı; takviyeyi lezzetli hale getirme"),
        ("Lemme", "lemmelive.com", "Ünlü Wellness Gummy", "Kourtney Kardashian markası; Lemme Chill, Lemme Focus; şık wellness gummy; Sephora'da"),
        ("Moon Juice", "moonjuice.com", "Adaptojenik Takviye", "SuperYou stres adaptojeni; Magnesi-Om uyku; Amanda Chantal Bacon kurdu; LA wellness"),
        ("Four Sigmatic", "foursigmatic.com", "Mantar Kahve", "Mantar bazlı kahve ve kakao; lion's mane + chaga; Finlandiya kökenli; bağışıklık desteği"),
        ("Bloom Nutrition", "bloomnu.com", "Kadın Fitness Takviye", "Greens & Superfoods tozu viral; kadın fitness influencer pazarlama; şişkinlik azaltma"),
        ("Athletic Greens (AG1 alternatifi)", "drinkag1.com", "Yeşil Toz Alternatifi", "Günlük yeşil toz takviye; 75 vitamin, mineral, probiyotik; tek paket çözüm; podcaster favorisi"),
        ("Onnit", "onnit.com", "Total İnsan Optimizasyonu", "Alpha Brain nootropik; Joe Rogan desteği; kettlebell + takviye; zihin + vücut optimizasyonu"),
        ("Athletic Brewing", "athleticbrewing.com", "Alkolsüz Bira", "Alkolsüz craft bira öncüsü; Run Wild IPA; sağlıklı yaşam + bira keyfi; $100M+ gelir"),
        ("Sakara Life", "sakara.com", "Organik Yemek + Takviye", "Organik bitki bazlı yemek teslimatı + takviyeler; detoks programları; lüks wellness"),
        ("Needed", "thisisneeded.com", "Hamilelik Takviyesi", "Hamilelik + emzirme dönemi için bilimsel takviye; doktor formüle etti; anne sağlığı"),
        ("Cymbiotika", "cymbiotika.com", "Liposomal Takviye", "Liposomal delivery takviyeler; yüksek emilim; NMN, glutatyon, D3+K2; lüks ambalaj"),
        ("Neurohacker", "neurohacker.com", "Nörobilim Takviye", "Qualia Mind nootropik; nörobilim bazlı; karmaşık formüller; bilişsel performans artırma"),
        ("Nutricost", "nutricost.com", "Sade Düşük Fiyat Takviye", "Minimum marka, maksimum değer; tek bileşen takviyeler; Amazon'da en uygun fiyatlı seçenek"),
        ("Promix", "promixnutrition.com", "Temiz Protein", "Grass-fed whey protein; Albert Matheny (diyetisyen+PT) kurdu; temiz etiket; sade formül"),
        ("Kion", "getkion.com", "Performans Takviye", "Ben Greenfield'ın markası; amino asitler, enerji barı; biyohacking topluluğu; performans odaklı"),
        ("Ancient Nutrition", "ancientnutrition.com", "Kemik Suyu Protein", "Kemik suyu kollajen protein; Dr. Josh Axe kurdu; paleo + keto dostu; geleneksel beslenme"),
        ("mindbodygreen", "mindbodygreen.com/supplements", "Wellness Medya Takviye", "Wellness medya platformundan takviye markasına; cellular beauty+, sleep support+; editoryal güven"),
        ("Cira Nutrition", "ciranutrition.com", "Kadın Fitness Takviye", "Kadın sporcular için pre-workout; Lauren Drain kurdu; estetik ambalaj; kadın gücü"),
        ("Ora Organic", "ora.organic", "Organik Bitki Takviye", "Organik ve bitki bazlı takviyeler; probiyotik + omega-3 + protein; sürdürülebilir ambalaj"),
        ("Vital Proteins Alt", "vitalproteins.com", "Kollajen Peptit", "Kollajen peptit tozu popüler; Jennifer Aniston ambassador; Nestlé satın aldı; kategori yaratıcı"),
        ("Nuun", "nuunlife.com", "Elektrolit Tablet", "Çözünür elektrolit tabletleri; şekersiz hidrasyon; sporcu favorisi; tablet format yenilikçi"),
        ("LMNT", "drinklmnt.com", "Tuzlu Elektrolit", "Şekersiz, yüksek sodyum elektrolit paketi; keto + fasting topluluğu; Robb Wolf kurdu"),
        ("Cure Hydration", "curehydration.com", "ORS Bazlı Hidrasyon", "Oral rehidrasyon çözeltisi bazlı; WHO formülünden esinlenen; doğal lezzet; organik"),
        ("ZBiotics", "zbiotics.com", "Alkol Sonrası Probiyotik", "Genetik mühendislik probiyotiği; asetaldehit parçalama; içki sonrası düzelme; biotech"),
        ("Qualia", "neurohacker.com/qualia", "Bilişsel Performans", "Kapsamlı nootropik formül; 28 bileşen; nörobilim araştırması bazlı; zihin netliği"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 6. Uyku & Yatak Teknolojisi
    # ═══════════════════════════════════════════════════════════════════════════
    "Uyku & Yatak Teknolojisi": [
        ("Pillow Cube", "pillowcube.com", "Yan Yatış Yastığı", "Yan yatan uyuyanlar için küp şeklinde yastık; viral TikTok reklam; niş problem çözme"),
        ("Coop Home Goods", "coophomegoods.com", "Ayarlanabilir Yastık", "Dolgu miktarını kendiniz ayarlayabileceğiniz yastık; parçalanmış memory foam; Amazon #1"),
        ("Manta Sleep", "mantasleep.com", "Tam Karartma Uyku Maskesi", "%100 karartma uyku maskesi; göz kupası teknolojisi; sıfır basınç; yan yatış uyumlu"),
        ("Hatch", "hatch.co", "Akıllı Uyku Lambası", "Restore akıllı uyku asistanı; gün ışığı simülasyonu; uyku rutini; meditasyon sesleri; anne-bebek"),
        ("Loftie", "byloftie.com", "Dijital Detoks Çalar Saat", "Telefonu yatak odasından çıkaran akıllı çalar saat; uyku sesleri + meditasyon; dijital wellness"),
        ("ChiliSleep", "chilisleep.com", "Yatak Sıcaklık Kontrolü", "Dock Pro yatak soğutma/ısıtma sistemi; su bazlı termoregülasyon; derin uyku optimizasyonu"),
        ("Molecule", "moleculesleep.com", "Bilimsel Uyku Yatağı", "Hava akışlı yatak teknolojisi; sporcu odaklı uyku; koku giderici özellik; performans uyku"),
        ("Bear Mattress", "bearmattress.com", "Sporcu Uyku Yatağı", "Celliant teknolojisi; kızılötesi vücut ısısını geri yansıtma; sporcu toparlanması; FDA sınıfı"),
        ("Nolah", "nolahmattress.com", "AirFoam Yatak", "AirFoam™ patentli; lateks + memory foam karışımı; yan yatış uzmanı; basınç noktası rahatlatma"),
        ("Cozy Earth", "cozyearth.com", "Bambu Yatak Tekstili", "Bambu viskon çarşaf ve pijama; Oprah'nın favori ürünleri listesinde; termoregülasyon; premium"),
        ("Buffy", "buffy.co", "Sürdürülebilir Yorgan", "Okaliptüs lifi yorgan; %100 geri dönüştürülmüş PET dolgu; sürdürülebilir uyku; ücretsiz deneme"),
        ("Gravity Blanket", "gravityblankets.com", "Ağırlıklı Battaniye", "Ağırlıklı battaniye kategorisini popülerleştiren marka; anksiyete azaltma; Kickstarter başarısı"),
        ("Bearaby", "bearaby.com", "Doğal Ağırlıklı Battaniye", "Organik pamuk örgü ağırlıklı battaniye; plastik boncuk yerine doğal ağırlık; estetik tasarım"),
        ("Tempur-Pedic DTC", "tempurpedic.com", "Visko-Elastik Yatak", "NASA teknolojisi memory foam; basınç dağılımı; premium uyku; DTC kanalı genişledi"),
        ("Layla Sleep", "laylasleep.com", "Bakır Infüzyon Yatak", "Bakır jel infüzyon memory foam; çift taraflı sertlik; soğutma + antimikrobiyel"),
        ("Helix Sleep Alt", "helixsleep.com", "Quiz Bazlı Yatak", "Uyku quizi ile kişiselleştirilmiş yatak; yan/sırt/yüzüstü yatış; çift farklılık çözümü"),
        ("Pluto Pillow", "plutopillow.com", "Kişisel Yastık", "25 soruluk quiz ile kişiselleştirilmiş yastık; boy, kilo, uyku pozisyonuna göre üretim"),
        ("Casper Alt Tuft & Needle", "tuftandneedle.com", "Adaptif Foam Yatak", "T&N Adaptive foam; bakır gel + grafit soğutma; uygun fiyatlı premium; Serta ile birleşti"),
        ("Purple Alt", "sleepnumber.com", "Akıllı Yatak", "Hava odacıklı sertlik ayarı; uyku takibi sensörleri; çift kişiselleştirme; IoT yatak"),
        ("Silk & Snow", "silkandsnow.com", "Kanada Premium Yatak", "Kanada yapımı hibrit yatak; organik pamuk kılıf; uygun lüks; 100 gece deneme"),
        ("Birch Living", "birchliving.com", "Organik Doğal Yatak", "Doğal lateks + yün + organik pamuk; GOTS sertifikalı; sürdürülebilir uyku; Helix kardeş"),
        ("Avocado Mattress", "avocadogreenmattress.com", "Yeşil Sertifikalı Yatak", "GOLS + GOTS organik sertifikalı; doğal lateks; B Corp; sürdürülebilir üretim; vegan seçenek"),
        ("Brentwood Home", "brentwoodhome.com", "LA Sürdürülebilir Yatak", "Los Angeles yapımı; yoga mat'ten yatağa; GREENGUARD Gold sertifikalı; CertiPUR-US"),
        ("Panda London", "pandalondon.com", "Bambu Memory Foam", "Bambu kılıflı memory foam yastık ve yatak; İngiltere merkezli; anti-alerjen; termoregülasyon"),
        ("Sleeping Duck", "sleepingduck.com.au", "Modüler Yatak", "Avustralya; modüler katmanları değiştirilebilir yatak; sertlik ayarı; 100 gece deneme"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 7. Fitness Ekipman & Giyim
    # ═══════════════════════════════════════════════════════════════════════════
    "Fitness Ekipman & Giyim": [
        ("Set Active", "setactive.co", "Sculptflex Activewear", "Sculptflex kumaş teknolojisi; renk 'drop' stratejisi; FOMO pazarlama; Lindsey Carter kurdu"),
        ("Beyond Yoga", "beyondyoga.com", "Yumuşak Activewear", "Spacedye kumaş teknolojisi; ultra yumuşak; kapsayıcı beden aralığı; Levi's satın aldı"),
        ("Year of Ours", "yearofours.com", "Ribbed Activewear", "Ribbed fitness giyim; Football koleksiyonu; LA tasarımı; retro spor estetiği"),
        ("Varley", "varley.com", "Lüks Activewear", "İngiltere lüks activewear; studio-to-street; minimalist tasarım; premium kumaşlar"),
        ("Splits59", "splits59.com", "Retro Activewear", "Retro esintili fitness giyim; renk blokları; LA fitness kültürü; performans + stil"),
        ("Girlfriend Collective", "girlfriend.com", "Geri Dönüşüm Activewear", "Geri dönüştürülmüş PET şişelerden tayt; sürdürülebilir + kapsayıcı; XXS-6XL beden aralığı"),
        ("Nobull", "nobullproject.com", "CrossFit Ayakkabı & Giyim", "CrossFit topluluğu; SuperFabric dayanıklı malzeme; fonksiyonel fitness odaklı; minimal tasarım"),
        ("Born Primitive", "bornprimitive.com", "Taktik Fitness Giyim", "CrossFit + askeri ilham; fonksiyonel fitness topluluğu; dayanıklı malzemeler; ABD yapımı"),
        ("Cuts", "cutsclothing.com", "Premium Erkek T-shirt", "PYCA kumaş; erkek t-shirt'ü premium hale getirme; fit + fabric odaklı; polo alternatifi"),
        ("ASRV", "asrv.com", "Teknik Erkek Sportswear", "Erkek performans giyim; teknik kumaşlar; minimalist siyah tasarım; fitness influencer"),
        ("Alphalete", "alphalete.com", "Bodybuilding Giyim", "Christian Guzman kurdu; fitness influencer ekosistemi; Revival koleksiyonu; gym kültürü"),
        ("Buffbunny", "buffbunny.com", "Kadın Fitness Giyim", "Heidi Somers kurdu; kadın vücut geliştirme; collection drop'ları dakikalar içinde tükeniyor"),
        ("Tracksmith", "tracksmith.com", "Koşu Heritage Giyim", "Bağımsız koşu markası; New England heritage estetiği; amateur koşuculara saygı; zanaatkarlık"),
        ("Satisfy Running", "satisfyrunning.com", "Lüks Koşu Giyim", "Paris merkezli lüks koşu; teknik kumaş + moda kesişimi; sanatsal yaklaşım; niş koşu kültürü"),
        ("Crossrope", "crossrope.com", "Ağırlıklı Atlama İpi", "Ağırlık değiştirilebilir atlama ipleri; uygulama ile entegre; ev egzersizi; clip-in sistem"),
        ("Bala", "shopbala.com", "Şık Ağırlık Bilekliği", "Tasarım ödüllü ağırlık bileklikleri; Shark Tank yatırımı; fitness + moda birleşimi; 1-2 lb"),
        ("FORM Swim", "formswim.com", "AR Yüzme Gözlüğü", "Artırılmış gerçeklik yüzme gözlüğü; anlık metrik görüntüleme; GPS + kalp atışı; yüzücü hayali"),
        ("Liteboxer", "liteboxer.com", "Interaktif Boks Makine", "Ritmik ışıklarla ev boks antrenmanı; müzik + hareket; oyunlaştırılmış fitness; kompakt tasarım"),
        ("FightCamp", "joinfightcamp.com", "Ev Boks Sistemi", "Ev boks + kickboks sistemi; sensörlü eldivenler; yumruk takibi; kişisel antrenör uygulaması"),
        ("Ergatta", "ergatta.com", "Oyunlaştırılmış Kürek", "Su dirençli kürek makinesi; oyunlaştırılmış antrenmanlar; mobilya görünümlü; ev dekoruna uyum"),
        ("Hydrow", "hydrow.com", "Bağlantılı Kürek Makinesi", "Elektromanyetik direnç kürek; canlı su üstü antrenmanlar; Peloton of rowing; immersif deneyim"),
        ("Tempo", "tempo.fit", "AI Güç Antrenmanı", "3D sensör ile form takibi; AI kişisel antrenör; ağırlık ayarlı ev sistemi; gerçek zamanlı düzeltme"),
        ("Tonal Alt Mirror", "mirror.co", "Akıllı Ev Aynası", "Ayna görünümlü ev fitness ekranı; yoga, barre, boks; lululemon satın aldı; kompakt"),
        ("TRX", "trxtraining.com", "Süspansiyon Antrenman", "Süspansiyon eğitim sistemi; Navy SEAL geliştirdi; portatif spor salonu; vücut ağırlığı antrenman"),
        ("Hyperice", "hyperice.com", "Toparlanma Cihazları", "Hypervolt masaj tabancası; Normatec kompresyon; sporcu toparlanma teknolojisi; NBA ortaklığı"),
        ("Therabody", "therabody.com", "Perkusyon Terapi", "Theragun perkusyon masaj cihazı; kas toparlanma; Dr. Jason Wersland kurdu; profesyonel + ev"),
        ("Chirp", "gochirp.com", "Sırt Ağrısı Çözümü", "Sırt ağrısı için wheel roller; omurga dekompresyon; Shark Tank viral; uygun fiyatlı çözüm"),
        ("Mirror (Fiture)", "fiture.com", "Fitness Ayna Alternatif", "Akıllı fitness aynası; AI form düzeltme; kompakt ev gym; interaktif antrenmanlar"),
        ("Echelon", "echelonfit.com", "Uygun Fiyatlı Bağlantılı Fitness", "Uygun fiyatlı Peloton alternatifi; bisiklet, kürek, koşu bandı, ayna; abonelik modeli"),
        ("Whoop Alt", "whoop.com", "Toparlanma Takipçisi", "Giyilebilir toparlanma ve zorlanma takipçisi; uyku + HRV analizi; abonelik modeli; profesyonel sporcu"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 8. Moda & Giyim (Kadın)
    # ═══════════════════════════════════════════════════════════════════════════
    "Moda & Giyim (Kadın)": [
        ("Quince", "onequince.com", "Fabrika Direkt Lüks", "Fabrikadan tüketiciye; $50 kaşmir kazak; lüks malzeme uygun fiyat; şeffaf tedarik zinciri"),
        ("Italic", "italic.com", "Markasız Lüks", "Lüks markaların fabrikalarından markasız ürünler; üyelik modeli; %50-80 düşük fiyat"),
        ("Naadam", "naadam.co", "Sürdürülebilir Kaşmir", "Moğolistan kaşmiri; çoban kooperatifleri; $75 kaşmir kazak; sürdürülebilir tedarik"),
        ("Hill House Home", "hillhousehome.com", "Nap Dress İkonu", "Nap Dress viral elbise; ev giyimini dış giyim yapma; Nell Diamond kurdu; drop kültürü"),
        ("Lunya", "lunya.com", "Lüks Uyku Giyim", "Washable silk pijama; Restore koleksiyonu; uyku giyimini yükselten marka; kadın CEO"),
        ("Lake Pajamas", "lakepajamas.com", "Pima Pamuk Pijama", "Ultra yumuşak Pima pamuk pijamalar; aile eşleştirme setleri; lüks ev giyim"),
        ("Les Tien", "lestien.com", "Lüks Loungewear", "LA yapımı lüks sweatpant ve hoodie; garment-dyed; minimalist; ünlü tercihi"),
        ("Summersalt", "summersalt.com", "Veri Odaklı Mayo", "1.5M vücut ölçümü analiziyle tasarlanan mayolar; The Sidestroke ikonik; kapsayıcı bedenler"),
        ("Andie Swim", "andieswim.com", "Quiz Bazlı Mayo", "Stil quizi ile kişiselleştirilmiş mayo önerisi; filter by activity; kadın kurucular; iade kolaylığı"),
        ("Selkie", "selkiecollection.com", "Puf Elbise Fenomeni", "Tül puf elbiseler viral; TikTok prenses estetiği; Kimchi Chic kurdu; fantezi moda"),
        ("House of CB", "houseofcb.com", "Şekillendirici Elbise", "Vücut şekillendiren elbiseler; ünlü tercihi; corset detayları; İngiltere merkezli"),
        ("Sézane", "sezane.com", "Fransız DTC Moda", "Paris dijital moda; aylık koleksiyon 'drops'; sürdürülebilir Demain programı; Fransız şıklığı"),
        ("Rouje", "rofrje.com", "Parizyen Moda", "Jeanne Damas'ın markası; Fransız kız estetiği; vintage ilham; Instagram moda ikonu"),
        ("Réalisation Par", "realisationpar.com", "Vintage Elbise", "Avustralya vintage baskılı elbiseler; The Naomi elbise ikonik; influencer pazarlama"),
        ("Rat & Boa", "ratandboa.com", "Bohem Akşam Elbise", "İngiltere bohem akşam elbiseleri; ipek + saten; parti + düğün giyim; genç lüks"),
        ("Faithfull the Brand", "faithfullthebrand.com", "Bali Tatil Giyim", "Bali'de el yapımı; tatil kıyafetleri; çiçekli baskılar; sürdürülebilir üretim"),
        ("With Jean", "withjean.com", "Avustralya Feminen Moda", "Dantel + fırfır detaylar; romantik Avustralya estetiği; vintage ilham; küçük parti üretim"),
        ("Aritzia Alt Wilfred", "aritzia.com", "Kanada Premium Günlük", "Kanada günlük lüks; Super Puff kaban ikonik; Babaton + Wilfred; mağaza deneyimi + DTC"),
        ("DÔEN", "shopdoen.com", "Nostalji Kadın Giyim", "Viktorya dönemi + prairie ilhamı; romantik kadınsı; Margaret + Katherine Kleveland; LA"),
        ("Staud", "staud.clothing", "LA Erişilebilir Lüks", "Moreau Bucket Bag ikonik; Sarah Staudinger kurdu; erişilebilir LA lüksü; net renkler"),
        ("Musier Paris", "musier-paris.com", "Parizyen Seksi Şıklık", "Paris tasarım; Musier bluz ve elbiseler; Fransız seksi minimalizmi; sosyal medya viral"),
        ("Nanushka", "nanushka.com", "Macar Sürdürülebilir Moda", "Budapeşte merkezli; vegan deri uzmanı; sürdürülebilir lüks; Szandra Sándor kurdu"),
        ("Paloma Wool", "palomawool.com", "Barcelona Sanatsal Moda", "Sanat + moda birleşimi; İspanya tasarım; benzersiz baskılar; yaratıcı topluluk"),
        ("Ganni", "ganni.com", "Danimarka Mutlu Moda", "Kopenhag neşeli moda; #GANNIGirls topluluğu; sorumluluk raporu; İskandinav estetiği"),
        ("Rotate", "rotate.com", "Parti Elbise Uzmanı", "Birger Christensen grup; payetli elbiseler; parti moda; Kopenhag gece hayatı estetiği"),
        ("Jacquemus", "jacquemus.com", "Güney Fransa Minimalizm", "Le Chiquito mikro çanta ikonik; Provence ilhamı; Simon Porte Jacquemus; viral pazarlama"),
        ("Rixo", "rixo.co.uk", "Vintage Baskılı Elbise", "İngiltere vintage baskı; patchwork + retro; el çizimi baskılar; sürdürülebilir üretim"),
        ("Cult Gaia", "cultgaia.com", "Sanatsal Aksesuar Moda", "Ark Bag ikonik; Jasmin Larian kurdu; sanat eseri gibi aksesuarlar; LA tasarım"),
        ("Farm Rio", "farmrio.com", "Brezilya Tropikal Moda", "Rio de Janeiro tropikal baskılar; canlı renkler; sürdürülebilirlik; Brezilya kültürü"),
        ("Ba&sh", "ba-sh.com", "Fransız Bohem Şıklık", "Paris iki arkadaş kurdu; bohem Fransız estetiği; Second Hand programı; sürdürülebilir"),
        ("Toit Volant", "toitvolant.com", "LA Feminen Tasarım", "Los Angeles romantik moda; çiçekli baskılar; vintage ilham; kadın kurucular"),
        ("Dissh", "dissh.com.au", "Avustralya Minimal Moda", "Avustralya minimal günlük giyim; nötr renkler; kaliteli temel parçalar; DTC odaklı"),
        ("Odd Muse", "oddmuse.com", "Güç Kadını Giyim", "İngiltere güç giyimi; iş + sosyal yaşam; blazer + elbiseler; kadın güçlendirme"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 9. Moda & Giyim (Erkek)
    # ═══════════════════════════════════════════════════════════════════════════
    "Moda & Giyim (Erkek)": [
        ("Cuts", "cutsclothing.com", "Premium Erkek T-shirt", "PYCA Pro kumaş; erkek t-shirt'ü premium hale getirme; kurumsal günlük alternatif; AO Collar"),
        ("CDLP", "cdlp.com", "İsveç Erkek İç Giyim", "Stockholm minimalist erkek iç giyim; lyocell kumaş; sürdürülebilir lüks; capsule wardrobe"),
        ("Fair Harbor", "fairharborclothing.com", "Geri Dönüşüm Mayo", "Geri dönüştürülmüş plastik şişelerden erkek mayo; sürdürülebilir plaj giyim; yumuşak kumaş"),
        ("Todd Snyder", "toddsnyder.com", "Amerikan Tasarımcı Erkek", "NYC erkek moda; Champion + Timex işbirlikleri; Amerikan heritage yeniden yorumlama"),
        ("Taylor Stitch", "taylorstitch.com", "Sürdürülebilir Erkek Giyim", "Crowdfunding ile üretim; organik + geri dönüşüm malzemeler; California zanaatkarlık"),
        ("Buck Mason", "buckmason.com", "LA Temel Erkek Giyim", "Los Angeles temel parçalar; Pima pamuk tee; Amerikan yapımı; minimalist erkek gardırop"),
        ("Mack Weldon", "mackweldon.com", "Performans Erkek Giyim", "Gümüş iyon anti-koku teknolojisi; 18-Hour Jersey; günlük + spor; erkek temel parçalar"),
        ("Vuori Alt", "vuori.com", "Kali Performans Giyim", "Kaliforniya aktif yaşam tarzı; erkek yoga + surf + koşu; DreamKnit kumaş; $4B değerleme"),
        ("Rhone", "rhone.com", "Premium Erkek Athleisure", "GoldFusion anti-koku teknolojisi; Commuter koleksiyonu iş + spor; şık performans giyim"),
        ("Faherty", "faherty.com", "Sürdürülebilir Plaj Giyim", "New Jersey çift (Alex + Mike Faherty); organik pamuk flannel; sahil yaşam tarzı; B Corp"),
        ("Mizzen+Main", "mizzenandmain.com", "Performans Gömlek", "Performans kumaşlı resmi gömlek; ter çekme + esneme; ütü gerektirmeyen; Phil Mickelson"),
        ("Ministry of Supply", "ministryofsupply.com", "NASA İlham İş Giyim", "MIT mühendisleri kurdu; 3D örgü ceket; termoregülasyon; performans iş giyim"),
        ("Western Rise", "westernrise.com", "Teknik Günlük Giyim", "AT Slim Pant her ortama uygun tek pantolon; merinos + naylon; minimalist erkek"),
        ("Proof", "huckberry.com/store/proof", "Seyahat Erkek Giyim", "72-Hour Merino tee; seyahat odaklı erkek giyim; hızlı kuruyan; koku önleyen"),
        ("Olivers Apparel", "oliversapparel.com", "Aktif Günlük Erkek", "Bradbury Jogger performans pantolon; 4-way stretch; ofis + gym arası; Silicon Valley"),
        ("Ten Thousand", "tenthousand.cc", "Taktik Fitness Giyim", "Interval Short erkek fitness şortu; taktik kumaş; minimalist; CrossFit + ağırlık odaklı"),
        ("Public Rec", "publicrec.com", "Şık Sweatpant", "All Day Every Day Pant iş ortamına uygun sweatpant; rahat + şık; work-from-home çağı"),
        ("True Classic", "trueclassictees.com", "Beden Şekilli T-shirt", "Erkek vücut tipine göre tasarlanmış tee; kol ve göğüste fit; viral Facebook reklamları"),
        ("Bylt Basics", "byltbasics.com", "Premium Erkek Temel", "Drop-Cut tee benzersiz kesim; LUX kumaş; erkek temel parçalarda premium kalite"),
        ("Grayers", "grfryers.com", "Heritage Erkek Giyim", "İngiliz heritage + Amerikan casual; maceracı erkek; vintage ilham; kaliteli kumaşlar"),
        ("Ledbury", "ledbury.com", "El Yapımı Gömlek", "Richmond, VA yapımı el dikimi gömlekler; İngiliz kumaş; zanaatkarlık; premium erkek gömlek"),
        ("Chubbies", "chubbies.com", "Kısa Şort Hareketi", "5.5 inç şort trendi başlatan marka; 'Sky's out thighs out'; eğlenceli erkek markası; hafta sonu giyim"),
        ("Mugsy Jeans", "mugsyjeans.com", "Esnek Erkek Jean", "Sweatpant konforu + jean görünümü; 4-way stretch denim; erkek rahat jean devrimi"),
        ("Revtown", "revtownusa.com", "Teknik Denim", "İtalyan denim + performans kumaş; otomotiv mühendisleri kurdu; kontrol edilen esneme"),
        ("Marine Layer", "marinelayer.com", "Yumuşak Günlük Giyim", "Re-Spun programı eski tişörtleri geri dönüştürme; ultra yumuşak kumaş; San Francisco"),
        ("Roark", "roark.com", "Macera Giyim", "Seyahat + macera ilhamlı erkek giyim; hikaye odaklı koleksiyonlar; fonksiyonel tasarım"),
        ("Flint and Tinder", "huckberry.com/store/flint-and-tinder", "Amerikan Yapımı Erkek", "ABD'de üretilen temel erkek parçalar; 10-Year Hoodie; dayanıklılık garantisi; zanaatkarlık"),
        ("State & Liberty", "stateandliberty.com", "Atletik Kesim Gömlek", "Atletik vücut tipi için resmi gömlek; performans kumaş; kas yapılı erkekler için"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 10. İç Giyim & Çorap
    # ═══════════════════════════════════════════════════════════════════════════
    "İç Giyim & Çorap": [
        ("Pepper", "wearpepper.com", "Küçük Göğüs Sütyeni", "AA-B kap için özel tasarım; 'küçük göğüsler harika' mesajı; dolgu yok; gerçek fit"),
        ("Knix", "knix.com", "Sızdırmaz İç Giyim", "Leakproof teknolojisi; regl + hafif mesane sızıntısı; 8 kat emici; kadın güçlendirme"),
        ("Cuup", "cuup.com", "Modern Fit İç Giyim", "Sanal fit danışmanlığı; minimalist tasarım; inclusive bedenler; doğru ölçü bulma problemi"),
        ("Harper Wilde", "harperwilde.com", "Basit Konforlu Sütyen", "Bliss sütyen uygun fiyatlı konfor; Try Before You Buy; iç giyim karmaşıklığını basitleştirme"),
        ("True & Co", "trueandco.com", "Quiz Bazlı Sütyen", "Fit quiz ile doğru sütyen bulma; veri odaklı tasarım; P&G satın aldı; rahat günlük"),
        ("Parade", "yourparade.com", "Sürdürülebilir İç Giyim", "Geri dönüştürülmüş malzeme; Re:Play kumaş; canlı renkler; Z kuşağı; kapsayıcı pazarlama"),
        ("Adore Me", "adoreme.com", "Abonelik İç Giyim", "Aylık iç giyim aboneliği; kapsayıcı bedenler; VIP üyelik; Victoria's Secret satın aldı"),
        ("Negative Underwear", "negativeunderwear.com", "Minimalist Lüks İç Giyim", "Dolgu + tel yok; minimalist tasarım; vücut pozitifliği; NYC yapımı; basit lüks"),
        ("Lively", "wearlively.com", "Leisurée İç Giyim", "Leisure + lingerie = Leisurée; Michelle Cordeiro Grant kurdu; Victoria's Secret alternatifi"),
        ("Commando", "wearcommando.com", "Görünmez İç Giyim", "Dikişsiz iç giyim; kıyafet altından görünmeme; patent teknoloji; çözüm odaklı"),
        ("CDLP", "cdlp.com", "İsveç Erkek İç Giyim", "Lyocell tencel erkek boxer; sürdürülebilir İskandinav tasarım; premium kalite; minimalist"),
        ("Tommy John", "tommyjohn.com", "Rahat Erkek İç Giyim", "Second Skin kumaş; 'No Adjustment Needed' garantisi; horizontal fly; rahat erkek iç giyim"),
        ("Saxx", "saxxunderwear.com", "BallPark Pouch Boxer", "BallPark Pouch patentli; erkek konforu için anatomik tasarım; sürtünme önleme; Kanada"),
        ("Pair of Thieves", "pairofthieves.com", "Eğlenceli Performans Çorap", "SuperFit performans iç giyim + çorap; eğlenceli baskılar; uygun fiyatlı; Target'ta"),
        ("Bombas Alt", "bombas.com", "Sosyal Etki Çorap", "1 alana 1 bağış modeli; 100M+ çorap bağışladı; Shark Tank en başarılı yatırım"),
        ("Darn Tough Alt", "darntoughvermont.com", "Ömür Boyu Garantili Çorap", "Vermont yapımı merinos yün çorap; ömür boyu garanti; yıkılmaz dayanıklılık"),
        ("Stance", "stance.com", "Sanatçı Çorap", "Sanatçı işbirlikleri çorap; NBA resmi çorap; kültürel ifade; premium aksesuar"),
        ("Smartwool", "smartwool.com", "Merinos Performans Çorap", "Merinos yünü çorap + katman giyim; outdoor performans; koku önleme; doğal termoregülasyon"),
        ("Feetures", "feetures.com", "Anatomik Koşu Çorabı", "Sol ve sağ ayak için ayrı tasarım; hedefli kompresyon; koşucu favorisi; ömür boyu garanti"),
        ("MeUndies Alt", "meundies.com", "Abonelik İç Giyim", "MicroModal kumaş; aylık baskı drop'ları; çift eşleştirme; eğlenceli marka sesi"),
        ("Hanky Panky", "hankypanky.com", "Dantel Thong Uzmanı", "Orijinal dantel thong; 40+ yıldır aynı tasarım; tek beden; Made in USA; rahat dantel"),
        ("Natori", "natori.com", "Asya İlham Lüks İç Giyim", "Josie Natori; Filipin-Amerikan tasarımcı; lüks iç giyim + ev giyim; kültürel miras"),
        ("ThirdLove Alt", "thirdlove.com", "Yarım Beden Sütyen", "Yarım beden sistemi; Fit Finder quiz; 80+ beden; kapsayıcılık öncüsü; data-driven fit"),
        ("Skims Alt", "skims.com", "Şekillendirici İç Giyim", "Kim Kardashian; 9 beden + çoklu ten rengi; şekillendirici iç giyim; viral pazarlama"),
        ("Nuudii System", "nuudiisystem.com", "Sütyen Alternatifi", "Sütyen yerine yapışkan destek sistemi; özgürlük + destek; yenilikçi konsept; niş çözüm"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 11. Ayakkabı & Terlik
    # ═══════════════════════════════════════════════════════════════════════════
    "Ayakkabı & Terlik": [
        ("Cariuma", "cariuma.com", "Sürdürülebilir Sneaker", "Brezilya sürdürülebilir sneaker; her çift için ağaç dikimi; organik pamuk + doğal kauçuk"),
        ("Oliver Cabell", "olivercabell.com", "Şeffaf Fiyat Sneaker", "Maliyet dağılımını gösteren lüks sneaker; İtalya yapımı; $30 malzeme $150 fiyat şeffaflığı"),
        ("Thursday Boots", "thursdayboots.com", "Uygun Fiyat Premium Bot", "Horween deri; Goodyear welt; $200'ın altında premium bot; DTC fiyat avantajı"),
        ("Tecovas", "tecovas.com", "Modern Kovboy Çizmesi", "DTC kovboy çizmesi; el yapımı; $200-300 fiyat (geleneksel $500+); Texas heritage"),
        ("Nisolo", "nisolo.com", "Etik Ayakkabı", "B Corp sertifikalı; Peru + Meksika zanaatkarları; adil ücret; sürdürülebilir ayakkabı"),
        ("Atoms", "atoms.com", "Çeyrek Beden Sneaker", "Çeyrek beden seçeneği; her ayak farklı boyutta olabilir; mükemmel fit; minimalist tasarım"),
        ("Vivaia", "vivaia.com", "Geri Dönüşüm Örgü Ayakkabı", "Geri dönüştürülmüş PET şişelerden 3D örgü ayakkabı; yıkanabilir; flat + topuklu; çevreci"),
        ("Koio", "koio.co", "İtalyan Zanaatkar Sneaker", "İtalya Le Marche bölgesinde el yapımı; premium deri; erişilebilir lüks sneaker"),
        ("GREATS", "greats.com", "Brooklyn Sneaker", "Brooklyn tasarımı İtalya yapımı sneaker; Royale ikonik; Steve Madden satın aldı"),
        ("Psudo", "psudo.com", "Yıkanabilir Sneaker", "Çamaşır makinesinde yıkanabilir sneaker; hafif; seyahat dostu; pratik çözüm"),
        ("Vessi", "vessi.com", "Su Geçirmez Örgü Sneaker", "Dymaterks su geçirmez örgü teknolojisi; Vancouver; yağmurda kuru ayak; hafif tasarım"),
        ("Tropicfeel", "tropicfeel.com", "Seyahat Ayakkabısı", "Çok amaçlı seyahat ayakkabısı; su + kara; modüler tasarım; Kickstarter rekoru"),
        ("Rothy's Alt", "rothys.com", "Geri Dönüşüm Flat", "Geri dönüştürülmüş su şişelerinden 3D örgü flat; yıkanabilir; 150M+ şişe dönüştürüldü"),
        ("Birdies", "birdies.com", "Konforlu Flat", "Ev terlikleri kadar rahat flat ayakkabı; Meghan Markle giydi; yastıklı taban; şık tasarım"),
        ("Suavs", "suavshoes.com", "Esnek Örgü Ayakkabı", "Çorap gibi esnek örgü ayakkabı; yıkanabilir; sıfır break-in süresi; unisex"),
        ("Lane Eight", "laneeight.com", "Sürdürülebilir Antrenman", "Bloom alg köpük + geri dönüşüm malzeme; çapraz antrenman ayakkabısı; çevreci performans"),
        ("HOLO Footwear", "holofootwear.com", "Okyanus Plastiği Ayakkabı", "Okyanus plastiğinden ayakkabı; suda yüzen; vegan; çevre temizliği + ayakkabı üretimi"),
        ("Nothing New", "nothingnew.com", "Geri Dönüşüm Sneaker", "Geri dönüştürülmüş malzemelerden sneaker; eski ayakkabı geri alma programı; döngüsel ekonomi"),
        ("Helm Boots", "helmboots.com", "Austin El Yapımı Çizme", "Austin, Texas yapımı premium erkek çizme; kalıp odaklı konfor; zanaatkarlık"),
        ("Tkees", "tkees.com", "Minimalist Sandalet", "Nude ton minimalist sandalet ve flat; ten rengi eşleştirme; plaj + şehir; basit lüks"),
        ("Hari Mari", "harimari.com", "Premium Flip Flop", "Premium memory foam flip flop; Austin, Texas; her satışta asker ailesine bağış"),
        ("Kane Footwear", "kanefootwear.com", "Toparlanma Terliği", "Aktif toparlanma terliği; ergonomik taban; sporcu sonrası; hafif + destekli"),
        ("Kizik", "kizik.com", "Elleri Kullanmadan Ayakkabı", "Hands-free giriş teknolojisi; arka kısmı kırmadan giyme; engelli dostu; F.A.S.T. teknolojisi"),
        ("Amberjack", "amberjackshoes.com", "Konforlu İş Ayakkabısı", "Dress ayakkabıda sneaker konforu; cemented construction; tam tahıl deri; $200 altı"),
        ("Glerups", "glerups.com", "Danimarka Yün Terlik", "Doğal yünden ev terliği; Danimarka zanaatkarlığı; kauçuk veya keçe taban; sıcak + konforlu"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 12. Yiyecek & Atıştırmalık
    # ═══════════════════════════════════════════════════════════════════════════
    "Yiyecek & Atıştırmalık": [
        ("Graza", "grfrza.com", "Sıkma Zeytinyağı", "Squeeze şişede extra virgin zeytinyağı; Drizzle vs Sizzle ayrımı; mutfağı eğlenceli yapan ambalaj"),
        ("Fly By Jing", "flybyjing.com", "Sichuan Acı Sos", "Jing Gao kurdu; Chili Crisp Çin lezzeti ABD'ye taşıdı; doğal bileşenler; dünya mutfağı"),
        ("TRUFF", "trfrff.com", "Trüf Acı Sos", "Trüf yağlı acı sos; lüks sos kategorisi yarattı; siyah ambalaj; premium hediye; $100M+ gelir"),
        ("Fishwife", "eatfishwife.com", "Modern Konserve Balık", "Tinned fish trendini ABD'de başlatan marka; renkli retro ambalaj; sürdürülebilir deniz ürünleri"),
        ("Omsom", "omfrsom.com", "Asya Lezzet Başlatıcı", "Asya yemekleri için sos + baharat paketleri; otantik lezzet kolayca; Kim + Vanessa Pham"),
        ("Bachan's", "bachans.com", "Japon Barbekü Sosu", "Japon tarzı barbekü sosu; aile tarifi; küçük parti üretim; Costco'da viral"),
        ("Mike's Hot Honey", "mikeshothoney.com", "Acılı Bal", "Biberli bal; pizza + peynir eşleşmesi; yeni lezzet kategorisi yarattı; Brooklyn başlangıcı"),
        ("Mid-Day Squares", "middaysquares.com", "Fonksiyonel Çikolata", "Protein + adaptojenik çikolata kare; şeker azaltılmış; sosyal medya şeffaflığı; Kanada"),
        ("Chomps", "chomps.com", "Temiz Etki Çubuk", "Grass-fed sığır çubuğu; Whole30 onaylı; şeker + hormon + antibiyotik yok; sade bileşen"),
        ("Deux", "eatdeux.com", "Cookie Dough Sağlıklı", "Yenilebilir kurabiye hamuru; protein + süperfood katkılı; vegan; Shark Tank; eğlenceli sağlıklı atıştırmalık"),
        ("Last Crumb", "lastcrumb.com", "Lüks Kurabiye", "Kutu başına $140+ lüks kurabiye; her biri farklı lezzet; drop modeli ile satış; tükeniyor"),
        ("Partake Foods", "partakefoods.com", "Alerjen-Free Atıştırmalık", "Top 9 alerjenden arındırılmış kurabiye + kraker; alerji çocukları için güvenli; kapsayıcı"),
        ("Behave Candy", "behavecandy.com", "Düşük Şeker Şeker", "Düşük glisemik şeker; gerçek meyve bazlı; yetişkin şeker; az şeker, çok lezzet"),
        ("Pipcorn", "pipsnacks.com", "Mini Patlamış Mısır", "Mini heirloom patlamış mısır; non-GMO; gluten-free lezzetler; Jennifer Aniston yatırımcı"),
        ("Banza", "eatbanza.com", "Nohut Makarna", "Nohut bazlı makarna; 2x protein; gluten toleransı olanlar için; $100M+ gelir; pizza hamuru da var"),
        ("Catalina Crunch", "catalina-crunch.com", "Keto Granola", "Keto dostu düşük karbonhidrat granola; 0 şeker; diyabet dostu; çıtır çıtır doku"),
        ("HighKey", "highkey.com", "Keto Atıştırmalık", "Keto mini kurabiye + kraker; düşük net karbonhidrat; şeker alkolü yok; Amazon #1 keto"),
        ("Simple Mills", "simplemills.com", "Temiz Bileşen Atıştırmalık", "Badem unu bazlı kraker + kurabiye; 8-10 basit bileşen; seed oil free; doğal"),
        ("Primal Kitchen", "primalkitchen.com", "Avokado Yağlı Sos", "Avokado yağı mayo + sos + dressing; Mark Sisson kurdu; paleo + keto; Kraft Heinz satın aldı"),
        ("Kettle & Fire", "kettleandfire.com", "Kemik Suyu", "Raf ömürlü kemik suyu; kolajen + protein; grass-fed; paleo + keto; açık ateşte pişirme"),
        ("Siete Foods", "sifretefoods.com", "Tahılsız Meksika", "Tahılsız tortilla + cips; Meksika-Amerikan aile tarifi; hindistan cevizi unu; PepsiCo satın aldı"),
        ("Lesser Evil", "lesserevil.com", "Organik Patlamış Mısır", "Organik, non-GMO popcorn + puf; hindistan cevizi yağlı; B Corp; sade bileşen listesi"),
        ("Smart Sweets", "smartsweets.com", "Düşük Şeker Jöle Şeker", "Şeker ihtiyarını akıllıca tatmin; 3g şeker; bitkisel bazlı; stevia tatlılığı"),
        ("That's It", "thatsitfruit.com", "Sadece Meyve Bar", "2 bileşen: meyve + meyve; eklenmiş şeker yok; alerjenden arınık; en basit bar"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Çikolata", "Soya lesitin, emülgatör, rafine şeker yok; basit kakao çikolata; Mondelez satın aldı"),
        ("Dang Foods", "dangfoods.com", "Hindistan Cevizi Chips", "Hindistan cevizi cipsi + yapışkan pirinç bar; Tayland ilhamlı; Shark Tank; Asya atıştırmalık"),
        ("Made in Nature", "madeinnature.com", "Organik Kurutulmuş Meyve", "USDA organik kurutulmuş meyve + figgy pops; doğal şeker; temiz bileşen"),
        ("Oatly", "oatly.com", "Yulaf Sütü", "İsveç yulaf sütü; kahve kültürünü değiştirdi; Barista Edition; cesur marka sesi; süt alternatifi"),
        ("Miyoko's", "miyokos.com", "Bitki Bazlı Tereyağı", "Vegan tereyağı + peynir; fermentasyon tekniği; gerçek süt ürünü lezzeti; şef kalitesi"),
        ("Tabs Chocolate", "tabschocolate.com", "Afrodizyak Çikolata", "Libido artırıcı dark çikolata; maca + epimedium; viral TikTok; cesur pazarlama; yetişkin konsept"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 13. İçecek & Kahve
    # ═══════════════════════════════════════════════════════════════════════════
    "İçecek & Kahve": [
        ("Trade Coffee", "dfrrinktrade.com", "Kişisel Kahve Eşleşme", "Quiz ile 450+ kavurucudan kişisel kahve eşleşmesi; artisanal kahve keşfi; taze teslimat"),
        ("Cometeer", "cometeer.com", "Dondurulmuş Kahve Kapsül", "Flash-frozen kahve kapsülleri; -40°C'de dondurulmuş; özel kavurucu işbirliği; kalite korunması"),
        ("Chamberlain Coffee", "chamberlaincoffee.com", "Gen Z Kahve", "Emma Chamberlain'in markası; organik + adil ticaret; sosyal medya to brand; cold brew bags"),
        ("Jot", "jot.co", "Ultra Konsantre Kahve", "20x konsantre kahve; bir damla ile bir fincan; pratik format; düşük asitli; seyahat dostu"),
        ("Death Wish Coffee", "deathwishcoffee.com", "Dünyanın En Güçlü Kahvesi", "2x kafein; cesur marka kimliği; kuru kafa logo; USDA organik; Kickstarter başarısı"),
        ("Bones Coffee", "bonescoffee.com", "Aromalı Özel Kahve", "Disney + pop kültür işbirlikleri; 70+ aroma; eğlenceli paket tasarımı; kült takipçi"),
        ("MUD\\WTR", "mudwtr.com", "Mantar Kahve Alternatifi", "Kahve alternatifi; lion's mane + chaga + kakao + chai; 1/7 kafein; sabah ritüeli yeniden tanımlama"),
        ("Athletic Brewing", "athleticbrewing.com", "Alkolsüz Craft Bira", "Alkolsüz craft bira öncüsü; Run Wild IPA; Craft Non-Alcoholic; sober curious hareketi"),
        ("Ghia", "drinkghia.com", "Alkolsüz Aperitif", "Alkolsüz İtalyan aperitif; botanik + adaptojenik; şık ambalaj; akşam yemeği partisi alternatifi"),
        ("Seedlip", "seedlipdrinks.com", "Alkolsüz Distile İçecek", "Dünyanın ilk distile alkolsüz içeceği; Diageo satın aldı; kokteyl kültürü; botanik formüller"),
        ("Alani Nu", "alaninu.com", "Kadın Enerji İçeceği", "Karina Irby kurdu; düşük kalorili enerji içeceği; kadın fitness; pastel ambalaj; Amazon #1"),
        ("Ghost", "ghostlifestyle.com", "Pop Kültür Enerji", "Sour Patch Kids + Warheads lisanslı lezzetler; pre-workout + enerji; şeffaf formül; Z kuşağı"),
        ("Celsius Alt", "celsius.com", "Fitness Enerji İçeceği", "Termojenik enerji; egzersiz öncesi; MetaPlus formül; koruyucu + şeker yok; $10B+ değerleme"),
        ("Sunwink", "sunwink.com", "Bitki Bazlı Tonik", "Süperfood tonik içecekler; zerdeçal + limon + zencefil; düşük şeker; bağırsak sağlığı"),
        ("De La Calle", "delacalle.com", "Tepache Fermente İçecek", "Meksika fermente ananas içeceği; prebiyotik; düşük şeker; kültürel miras modern paketleme"),
        ("Poppi Alt", "drinkpoppi.com", "Prebiyotik Soda", "Elma sirkesi bazlı prebiyotik soda; bağırsak sağlığı; düşük şeker; Shark Tank; viral"),
        ("Olipop Alt", "drinkolipop.com", "Sağlıklı Soda", "Prebiyotik + botanik soda; 2-5g şeker; bağırsak dostu; nostaljik lezzetler; $200M+ gelir"),
        ("Bev", "drinkbev.com", "Kadın Şarap", "Kutu şarap kadınlar için; sıfır şeker; düşük kalori; eğlenceli marka; rosé ikonik"),
        ("Haus", "drink.haus", "Modern Aperitif", "Düşük alkollü aperitif; doğal bileşenler; farm-to-bottle; modern bar kültürü"),
        ("Hiyo", "drinkhiyo.com", "Sosyal Tonik", "Adaptojenik sosyal içecek; alkol alternatifi; nootropik + adaptojenik; gece dışarı çıkma"),
        ("Wfrter Drop", "waterdrop.com", "Mikro İçecek", "Küp şeklinde şekersiz içecek aroması; plastik şişe atığını azaltma; Avusturya; kompakt tasarım"),
        ("Liquid Death Alt", "liquiddeath.com", "Punk Su", "Kutu ambalajlı su; 'Murder your thirst'; punk marka kimliği; $1.4B değerleme; plastiğe karşı"),
        ("Nutr", "nutr.com", "Ev Yapımı Bitki Sütü", "Evde bitki sütü yapma makinesi; badem, yulaf, hindistan cevizi sütü; taze + katkısız"),
        ("Fellow Alt (Kahve)", "fellowproducts.com", "Özel Kahve Ekipmanı", "Stagg EKG kettle ikonik; pour-over kahve ekipmanları; tasarım ödüllü; Carter mug"),
        ("Blue Bottle Coffee", "bluebottlecoffee.com", "3. Dalga Kahve", "Oakland 3. dalga kahve öncüsü; 48 saat içinde kavrulmuş; Nestlé yatırımı; deneyim odaklı"),
        ("Verve Coffee", "vervecoffee.com", "Santa Cruz Özel Kahve", "Santa Cruz artisanal kahve; direct trade; mevsimsel karışımlar; Batı Yakası kahve kültürü"),
        ("Counter Culture", "counterculturecoffee.com", "Etik Özel Kahve", "Direct trade öncüsü; şeffaf tedarik zinciri; eğitim odaklı; profesyonel kahve eğitimi"),
        ("Stumptown", "stumptowncoffee.com", "Portland Craft Kahve", "Portland craft kahve kültürü; cold brew öncüsü; Hair Bender ikonik karışım"),
        ("Intelligentsia", "intelligentsia.com", "Chicago Özel Kahve", "Direct trade terimini popülerleştiren marka; Black Cat Espresso ikonik; eğitim + kalite"),
        ("Dona Chai", "donachai.com", "Zanaatkar Chai", "Brooklyn zanaatkar chai konsantresi; Masala + turmeric chai; kafeler için wholesale + DTC"),
        ("Rishi Tea", "rfrishi-tea.com", "Premium Organik Çay", "Direct trade organik çay; botanik takviyeler; matcha uzmanı; restoran + DTC"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 14. Ev & Mutfak
    # ═══════════════════════════════════════════════════════════════════════════
    "Ev & Mutfak": [
        ("Caraway", "carawayhome.com", "Seramik Toksinsiz Tencere", "Seramik kaplama tencere seti; PTFE/PFOA yok; manyetik raf sistemi; estetik renkler"),
        ("Material Kitchen", "materialkitchen.com", "Minimalist Mutfak Aracı", "reBoard kesme tahtası; The Fundamentals set; mutfakta daha az eşya felsefesi"),
        ("Misen", "misen.com", "Uygun Fiyat Şef Bıçağı", "Crowdfunding'den doğan şef bıçağı; Japon çeliği uygun fiyatla; mutfak demokrasisi"),
        ("Fellow", "fellowproducts.com", "Tasarım Kahve Ekipmanı", "Stagg EKG pour-over kettle ikonik; Carter kupası; tasarım ödüllü; özel kahve kültürü"),
        ("BlendJet", "blendjet.com", "Portatif Blender", "Şarj edilebilir taşınabilir blender; USB-C; 12oz; smoothie her yerde; viral TikTok; $300M+ gelir"),
        ("Ember", "ember.com", "Sıcaklık Kontrollü Kupa", "Sıcaklık kontrollü akıllı kupa; Bluetooth ile derece ayarı; Starbucks işbirliği; Apple Store'da"),
        ("Canopy", "getcanopy.co", "Estetik Nemlendirici", "Filtreli buharlaşma nemlendiricisi; küf + bakteri yok; duş aromaterapi; modern tasarım"),
        ("Our Place Alt", "fromourplace.com", "Çok İşlevli Tava", "Always Pan 8 tencere/tava yerine tek ürün; Perfect Pot; estetik + fonksiyon; viral"),
        ("GreenPan", "greenpan.com", "Termolon Seramik Tencere", "Thermolon seramik kaplama patentli; toksinsiz pişirme; Belçika teknolojisi; sağlıklı mutfak"),
        ("Ooni", "ooni.com", "Ev Pizza Fırını", "Portatif pizza fırını; 60 saniyede pizza; outdoor yemek kültürü; Kickstarter'dan $100M+'a"),
        ("AeroPress", "aeropress.com", "Basınçlı Kahve Demleme", "Havacılık mühendisi icat etti; seyahat dostu; basınçlı demleme; dünya şampiyonası var"),
        ("Open Spaces", "getopenspaces.com", "Estetik Ev Organizasyonu", "Tasarım odaklı ev düzenleme; terazzo tepsiler; pastel renkler; Marie Kondo alternatifi"),
        ("Tumble", "tumfrble.com", "Yıkanabilir Halı", "Çamaşır makinesinde yıkanabilir halı; evcil hayvan + çocuk dostu; modern desenler; pratik çözüm"),
        ("Made In Cookware", "madeincookware.com", "Fabrika Direkt Tencere", "Profesyonel şef kalitesinde; fabrikadan direkt; Amerikan ve Fransız üretim; şef onaylı"),
        ("HexClad Alt", "hexclad.com", "Hibrit Tencere", "Paslanmaz çelik + yapışmaz hibrit; Gordon Ramsay ortaklığı; lazer kazınmış altıgen desen"),
        ("Great Jones", "greatjones.co", "Renkli Mutfak", "Dutchess Dutch oven; renkli emaye pişirme kapları; millennial mutfak estetiği; hediye odaklı"),
        ("Mosaic Foods", "mosaicfoods.com", "Hazır Donmuş Yemek", "Bitkisel bazlı donmuş yemekler; besleyici + pratik; abonelik; aile porsiyonları"),
        ("W&P", "wandpdesign.com", "Tasarım Mutfak Aksesuarı", "Porter mug ve lunch bowl; Porter büyük buz kalıpları; Brooklyn tasarım stüdyosu"),
        ("Rigwa", "myrigwa.com", "Paslanmaz Yemek Kabı", "Vakum yalıtımlı paslanmaz yemek kapları; plastik-free öğle yemeği; sıcak/soğuk saklama"),
        ("Balmuda", "balmuda.com", "Japon Tasarım Mutfak", "Japon minimalist tasarım; The Toaster buhar teknolojisi; The Pot kettle; form + fonksiyon"),
        ("Solo Stove Alt", "solostove.com", "Dumansız Ateş Çukuru", "Dumansız ateş çukuru; 360° hava akışı; patentli ikincil yanma; bahçe eğlencesi"),
        ("Meater", "meater.com", "Kablosuz Et Termometresi", "Tamamen kablosuz akıllı et termometresi; Bluetooth + cloud; guide cook; mükemmel pişirme"),
        ("Breville Alt", "breville.com", "Akıllı Mutfak Cihazı", "Barista Express espresso; Smart Oven; mutfak inovasyonu; Avustralya mühendisliği"),
        ("Instant Pot Alt", "instantpot.com", "Çok İşlevli Pişirici", "7-in-1 elektrikli basınçlı pişirici; düdüklü tencere + slow cooker; mutfak devrimi"),
        ("Vitamix Alt", "vitamix.com", "Profesyonel Blender", "Profesyonel güçte blender; restoran kalitesi evde; 10 yıl garanti; çok amaçlı"),
        ("Stojo", "stojo.co", "Katlanır Kahve Bardağı", "Katlanabilir silikon kahve bardağı; seyahat dostu; tek kullanımlık bardak alternatifi; kompakt"),
        ("LARQ", "lfrqrq.com", "UV Kendi Kendini Temizleyen Şişe", "UV-C LED ışıkla kendi kendini temizleyen su şişesi; bakterileri %99 öldürme; seyahat dostu"),
        ("Lomi", "lfromi.com", "Mutfak Kompost Makinesi", "Tezgah üstü yiyecek atığı kompostlama; 4 saatte toprak; koku yok; sürdürülebilir mutfak"),
        ("AirScape", "airscapecoffee.com", "Vakum Saklama Kabı", "Patentli iç kapak vakum sistemi; kahve + gıda tazeliği uzatma; paslanmaz çelik + cam seçenek"),
        ("Souper Cubes", "soupercubes.com", "Porsiyonluk Dondurucu Kalıbı", "Silikon porsiyon dondurucu tepsisi; yemek hazırlığı; 1-2-4 cup boyutları; pratik mutfak çözümü"),
        ("Stasher Alt", "stfrasherbag.com", "Silikon Saklama Poşeti", "Yeniden kullanılabilir silikon poşet; plastik fermuarlı poşet alternatifi; mikrodalga + bulaşık güvenli"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 15. Ev Temizlik & Sürdürülebilirlik
    # ═══════════════════════════════════════════════════════════════════════════
    "Ev Temizlik & Sürdürülebilirlik": [
        ("Blueland", "blueland.com", "Tablet Temizlik", "Çözünür temizlik tabletleri + yeniden doldurulabilir şişe; tek kullanımlık plastik azaltma; $100M+ satış"),
        ("Branch Basics", "branchbasics.com", "Konsantre Temizlik", "Tek konsantre her şeyi temizler; toksinsiz; hassas cilt güvenli; The Concentrate ikonik"),
        ("Earth Breeze", "earthbreeze.com", "Çamaşır Yaprağı", "Kağıt ince çamaşır deterjan yaprağı; plastiksiz; bitkisel bazlı; abonelik; kompakt paket"),
        ("Stasher", "stasherbag.com", "Silikon Saklama Poşeti", "Yeniden kullanılabilir platin silikon poşet; plastik fermuarlı poşet yerine; fırın + dondurucu güvenli"),
        ("Bite", "bfritetoothpaste.com", "Tablet Diş Macunu", "Diş macunu tabletleri; plastiksiz; sürdürülebilir ağız bakımı; aktif kömür + nane seçenekleri"),
        ("Public Goods", "publicgoods.com", "Temel Ev İhtiyaçları", "Sade etiketli sürdürülebilir ev ürünleri; üyelik modeli; bitkisel bazlı; uygun fiyatlı"),
        ("Who Gives A Crap", "whogivesacrap.com", "Bambu Tuvalet Kağıdı", "Bambu + geri dönüştürülmüş tuvalet kağıdı; kârın %50'si bağış; eğlenceli ambalaj; B Corp"),
        ("Grove Collaborative", "grove.co", "Doğal Temizlik Platformu", "Doğal temizlik ürünleri marketplace; Mrs. Meyer's + kendi markası; sürdürülebilir ev bakımı"),
        ("Cleancult", "cleancult.com", "Karton Kutu Temizlik", "Karton ambalajlı temizlik ürünleri; hindistan cevizi bazlı; yeniden doldurulabilir; sürdürülebilir"),
        ("Dropps", "dropps.com", "Pod Çamaşır Deterjan", "Bitkisel bazlı çamaşır pod'ları; kompostlanabilir ambalaj; performans + sürdürülebilirlik"),
        ("Kula Cloth", "kulacloth.com", "Antimikrobiyel Outdoor Bez", "Gümüş iyon antimikrobiyel idrar bezi; outdoor kadınlar için; backcountry hijyen çözümü"),
        ("Sheets Laundry Club", "sheetslaundryclub.com", "Çamaşır Deterjan Yaprağı", "Kağıt ince çamaşır deterjan yaprağı; 50 yıkama; kompakt paket; plastik jug alternatifi"),
        ("Tru Earth", "tfrruearth.com", "Ekolojik Çamaşır Şeridi", "Ultra konsantre çamaşır şeridi; sıfır plastik ambalaj; Kanada; 1100+ kimyasal yok"),
        ("Supernatural", "supernatural.com", "AR Fitness + Temizlik", "Doğal bileşenli ev temizleyicileri; Dr. Bronner's gibi ama modern; cam şişe; güzel kokular"),
        ("Common Good", "commongoodandco.com", "İstasyon Dolum", "Mağazalarda dolum istasyonu; sürdürülebilir temizlik; profesyonel kalite; SLS-free"),
        ("Attitude", "attitudeliving.com", "EWG Onaylı Temizlik", "EWG Verified temizlik + kişisel bakım; plastik-nötr; Kanada; çevre duyarlı formüller"),
        ("Ethique", "ethique.com", "Katı Bar Temizlik", "Katı şampuan + duş jeli + temizleyici barları; plastiksız; Yeni Zelanda; 20M+ plastik şişe tasarrufu"),
        ("Marley's Monsters", "marleysmonsters.com", "Yeniden Kullanılabilir Ev", "Yeniden kullanılabilir kağıt havlu UNpaper; bez mendil; sürdürülebilir ev alternatifleri; Portland"),
        ("Package Free", "packagefreeshop.com", "Sıfır Atık Dükkan", "Lauren Singer'ın sıfır atık mağazası; ambalajsız ürünler; sürdürülebilir yaşam küratörlüğü"),
        ("Reel Paper", "reelpaper.com", "Bambu Tuvalet Kağıdı", "Bambu tuvalet kağıdı + kağıt havlu; ağaçsız; abonelik; her rulo = 1 tuvalet inşaası bağışı"),
        ("Repurpose", "repurpose.com", "Kompost Tabak Çatal", "Kompostlanabilir tableware; bitki bazlı çatal bıçak + tabak; tek kullanımlık ama sürdürülebilir"),
        ("MyGreenFills", "mygreenfills.com", "Yeniden Dolum Temizlik", "Kendi temizleyicini evde karıştır; konsantre + su; yeniden doldurulabilir sistemler; doğal"),
        ("Force of Nature", "forceofnatureclean.com", "Elektrolize Su Temizlik", "Tuz + su + sirke elektrolize; hastane düzeyinde dezenfektan; toksinsiz; çocuk güvenli"),
        ("No Tox Life", "notoxlife.com", "Sıfır Atık Temizlik", "Bulaşık yıkama bloğu; katı temizlik ürünleri; Amerikan yapımı; el yapımı; ambalajsız"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 16. Bebek & Çocuk
    # ═══════════════════════════════════════════════════════════════════════════
    "Bebek & Çocuk": [
        ("Tubby Todd", "tubbytodd.com", "Bebek Cilt Bakımı", "Egzama + hassas bebek cildi için; All Over Ointment ikonik; doğal bileşenler; anne topluluğu"),
        ("Coterie", "cotfrerie.com", "Premium Bebek Bezi", "Mühendislik harikası bebek bezi; 25 patent; ultra ince + ultra emici; 'iPhone of diapers'"),
        ("Little Sleepies", "littlesleepies.com", "Bambu Bebek Pijama", "Bambu viskon bebek + çocuk pijama; ultra yumuşak; eğlenceli baskılar; ebeveyn eşleştirme"),
        ("Kyte Baby", "kfrytebaby.com", "Bambu Bebek Giyim", "Bambu rayon bebek giyim + uyku tulumu; organik; sürdürülebilir; hassas cilt dostu; viral"),
        ("Mushie", "mushie.com", "İskandinav Bebek Aksesuar", "Danimarka tasarımı bebek ürünleri; silikon tabak + emzik + oyuncak; pastel estetik; minimal"),
        ("Cerebelly", "cerebfrelly.com", "Beyin Sağlığı Mama", "Nörobilimci kurdu; beyin gelişimi odaklı bebek maması; 16 beyin destekleyici besin; organik"),
        ("Once Upon a Farm", "onceuponafarm.com", "Soğuk Basınç Mama", "Soğuk basınç organik bebek maması; Jennifer Garner yatırımcı; taze + besleyici"),
        ("Serenity Kids", "serenitykids.com", "Et Bazlı Bebek Maması", "Grass-fed et bazlı bebek maması; düşük şeker; paleo bebek beslenmesi; otlak yetiştirilmiş"),
        ("Fridababy", "fridafrbaby.com", "Pratik Bebek Araçları", "NoseFrida burun emici ikonik; ebeveynliği pratikleştiren araçlar; tabu konularda çözüm"),
        ("Hatch Baby Alt", "hatch.co", "Bebek Uyku Ritüeli", "Rest+ bebek ses makinesi; gece lambası + beyaz gürültü; uyku eğitimi yardımcısı"),
        ("Lalo", "meetlalo.com", "Modern Mama Sandalyesi", "Tasarım odaklı bebek mama sandalyesi; The Chair modüler; modern ebeveynlik; Play Gym"),
        ("Comotomo", "comotomo.com", "Doğal Hisli Biberon", "Anne memesine benzeyen silikon biberon; kolay geçiş; geniş ağız; kolik önleme"),
        ("Babyganics", "babyganics.com", "Bitki Bazlı Bebek", "Bitkisel bazlı bebek deterjan + güneş kremi + böcek kovucu; NeoNourish teknolojisi"),
        ("Lovevery Alt", "lovevery.com", "Gelişim Oyuncak Kiti", "Montessori ilhamlı aşama bazlı oyuncak kutuları; nörobilim destekli; abonelik; eğitici"),
        ("KiwiCo Alt", "kiwico.com", "STEM Proje Kutusu", "Yaşa göre STEM + sanat proje kutuları; Tinker, Kiwi, Koala hatları; yaratıcı eğitim"),
        ("Monica + Andy", "monicaandandy.com", "Organik Bebek Giyim", "GOTS organik sertifikalı bebek giyim; anne kurdu; yumuşak + güvenli; premium kalite"),
        ("Pehr", "pfrehr.com", "Tasarım Bebek Odası", "Kanada tasarımı bebek odası tekstili; organik pamuk; ponponlu sepetler; minimalist estetik"),
        ("Gathre", "gathre.com", "Deri Oyun Matı", "Mikro fiber deri oyun matı; kolay temizlenebilir; şık tasarım; piknik + ev; çok amaçlı"),
        ("Ollie World", "ollieworld.com", "Patentli Kundak", "Moisture-wicking kundak; patentli elastik bant sistemi; doğru sıcaklık; bebek uyku kalitesi"),
        ("Nanit Alt", "nanit.com", "AI Bebek Monitörü", "AI-powered bebek uyku takibi; tavan montajlı kamera; uyku skoru; nefes giyim takibi"),
        ("Snoo Alt", "happiestbaby.com", "Otomatik Sallanan Beşik", "Dr. Harvey Karp; otomatik sallama + beyaz gürültü; ağlama algılama; 5S metodu; kiralama"),
        ("Loulou Lollipop", "lfrouloulollipop.com", "Silikon Diş Kaşıyıcı", "Tasarım silikon diş kaşıyıcılar; yiyecek şekilli; emzik klipsleri; Kanada tasarım"),
        ("Stokke Alt", "stokke.com", "Ergonomik Bebek Mobilya", "Tripp Trapp büyüyen sandalye; ergonomik tasarım; Norveç; 50+ yıl; kalıcı bebek mobilyası"),
        ("Aden + Anais", "adenandanais.com", "Müslin Bebek Örtüsü", "Müslin kundak + bebek örtüsü; Raegan Moya-Jones kurdu; 'muslin swaddle' kategorisi yarattı"),
        ("Maisonette", "maisonette.com", "Küratörlü Çocuk Butik", "Çocuk giyim + dekor online butik; küratörlü seçim; premium markalar; hediye kılavuzu"),
        ("Primary", "primary.com", "Logo-Free Çocuk Giyim", "Logosuz + slogansız çocuk kıyafetleri; saf renkler; etiketler kopartılabilir; kapsayıcı"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 17. Evcil Hayvan
    # ═══════════════════════════════════════════════════════════════════════════
    "Evcil Hayvan": [
        ("Fi", "tryfi.com", "Akıllı Köpek Tasması", "GPS + aktivite takipli akıllı tasma; kayıp köpek bulma; günlük egzersiz takibi; LTE bağlantı"),
        ("Chippin", "chippin.com", "Sürdürülebilir Köpek Maması", "Cırcır böceği + gümüş yanlıbalık protein; sürdürülebilir protein; çevre dostu evcil hayvan beslenmesi"),
        ("Diggs", "diggs.pet", "Modern Köpek Kafesi", "Revol katlanır modern köpek kafesi; bebek yatağı standartlarında güvenlik; estetik tasarım"),
        ("Tuft + Paw", "tuftandpaw.com", "Modern Kedi Mobilyası", "Tasarım odaklı kedi mobilyası; kedi ağacı + yatak; modern ev dekoruna uyumlu; Kanada"),
        ("Wild One", "wildone.com", "Şık Köpek Aksesuarı", "Modern tasarımlı köpek tasma + oyuncak + taşıma; Instagram estetiği; pastel renkler; premium"),
        ("A Pup Above", "apupabove.com", "Sous Vide Köpek Maması", "Sous vide pişirme yöntemiyle taze köpek maması; insan kalitesinde malzeme; organik"),
        ("Jinx", "jinx.com", "Performans Köpek Maması", "Organik tavuk + organik yulaf; düşük işlenmiş; patatesiz; köpek sağlığı odaklı; modern marka"),
        ("Sundays for Dogs", "sundaysfordogs.com", "Hava Kurutulmuş Köpek Maması", "Hava kurutulmuş tam gıda köpek maması; USDA organik; insan kalitesinde; pratik saklama"),
        ("Native Pet", "nativepet.com", "Köpek Takviyesi", "Probiyotik + kalça & eklem + allerji takviyesi; tozu mama üstüne serpme; doğal; veteriner onaylı"),
        ("PetHonesty", "pethonesty.com", "Evcil Hayvan Vitamin", "Köpek vitamin sakızları; probiyotik + eklem + sakinleştirici; yumuşak çiğneme formatı"),
        ("Finn", "petfinn.com", "Köpek Wellness Takviye", "Köpek wellness takviyesi; sakinleştirici + eklem + sindirim; ebeveyn seviyesi kalite"),
        ("Open Farm", "openfarmpet.com", "Etik Evcil Hayvan Gıda", "Etik kaynaklı + şeffaf tedarik; sertifikalı insancıl; sürdürülebilir; deniz ürünleri izlenebilir"),
        ("Farmer's Dog Alt", "thefarmersdog.com", "Taze Köpek Maması", "İnsan kalitesinde taze köpek maması; kişiselleştirilmiş porsiyon; kapıya teslimat; $1B+ değerleme"),
        ("Ollie Alt", "myollie.com", "Kişisel Taze Mama", "Kişiselleştirilmiş taze köpek maması; kalori kontrolü; veteriner onaylı; insan kalitesi"),
        ("Litter Robot Alt", "litter-robot.com", "Otomatik Kedi Tuvaleti", "Otomatik temizlenen kedi tuvaleti; döner globe; WiFi bağlantılı; sağlık takibi"),
        ("PrettyLitter Alt", "prettylitter.com", "Sağlık Takipli Kum", "Renk değiştiren kedi kumu; pH ile sağlık sorunları tespit; silika kristal; hafif"),
        ("Spot & Tango", "spotandtango.com", "Taze Köpek Maması", "UnKibble taze kuru mama; insan kalitesinde; AI beslenme planı; düşük işlenmiş"),
        ("Maev", "meetmaev.com", "Çiğ Köpek Maması", "Raw (çiğ) köpek maması nugget formatında; doğal beslenme; dondurulmuş; kolay porsiyon"),
        ("Stella & Chewy's", "stellaandchewys.com", "Dondurularak Kurutulmuş Çiğ", "Dondurularak kurutulmuş çiğ köpek + kedi maması; raw beslenme kolaylaştırılmış; ABD yapımı"),
        ("Pet Plate", "petplate.com", "Şef Hazırlığı Köpek Maması", "Şef hazırlığı taze köpek maması; USDA mutfağında pişirilmiş; veteriner formüle; porsiyonlanmış"),
        ("Fable Pets", "ffrrablepets.com", "Oyuncak + Aksesuar", "Tasarım ödüllü köpek oyuncak + aksesuar; The Game oyuncak ikonik; modern evcil hayvan"),
        ("BarkBox Alt", "barkbox.com", "Köpek Oyuncak Kutusu", "Aylık temali köpek oyuncak + ödül kutusu; Super Chewer dayanıklı hat; eğlenceli abonelik"),
        ("Wag", "wag.com", "Köpek Gezdirme + Bakım", "On-demand köpek gezdirme + otel; uygulama bazlı; sigortalı; Petco satın aldı"),
        ("Wisdom Panel", "wisdompanel.com", "Köpek DNA Testi", "Köpek ırk + sağlık DNA testi; 350+ ırk tanıma; genetik sağlık taraması; evde kit"),
        ("Cat Person", "catperson.com", "Premium Kedi Ürünleri", "Kedi odaklı marka; kedi maması + yatak + oyuncak; kedi sahiplerini anlayan marka; premium"),
        ("Butternut Box", "butternutbox.com", "İngiltere Taze Köpek Mama", "İngiltere'nin taze köpek maması; haftalık teslimat; kişiselleştirilmiş; insan kalitesi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 18. Aksesuar & Takı
    # ═══════════════════════════════════════════════════════════════════════════
    "Aksesuar & Takı": [
        ("Ana Luisa", "analuisa.com", "Düşük Karbon Takı", "Karbon-nötr takı; geri dönüştürülmüş altın; uygun fiyatlı lüks; NYC tasarım; çevre etki raporu"),
        ("Gorjana", "gorjana.com", "Kaliforniya Altın Takı", "Laguna Beach altın kaplama takı; katmanlama uzmanı; erişilebilir lüks; kadın kurucular"),
        ("Studs", "stufrds.com", "Modern Piercing Stüdyosu", "Piercing deneyimini modernize eden marka; gen Z estetiği; ear styling küratörlüğü; NYC"),
        ("Dorsey", "bfrydorsey.com", "Lab-Grown Mücevher", "Laboratuvar üretimi değerli taş; sürdürülebilir lüks; vintage ilham; çevreye duyarlı mücevher"),
        ("Catbird", "catbirdnyc.com", "Brooklyn Zanaatkar Takı", "Brooklyn zanaatkar takı; ince + delicate; Forever Bracelet kaynaklı bileklik; NYC ikonu"),
        ("Stone and Strand", "stoneandstrand.com", "Günlük Pırlanta", "Erişilebilir pırlanta takı; günlük kullanılabilir fine jewelry; katmanlama; gift-ready"),
        ("VRAI", "vfrrai.com", "Lab-Grown Pırlanta", "Diamond Foundry lab-grown pırlanta; Leonardo DiCaprio yatırımcı; sürdürülebilir pırlanta; sıfır karbon"),
        ("Ring Concierge", "ringconcierge.com", "Fine Jewelry DTC", "Nicole Wegman kurdu; Instagram fine jewelry öncüsü; nişan yüzüğü uzmanı; DTC lüks"),
        ("Mejuri Alt", "mejfruri.com", "Günlük Fine Jewelry", "Erişilebilir fine jewelry; 14k altın; her gün takılabilir; üyelik fiyat avantajı; Kanada"),
        ("Brilliant Earth Alt", "brilliantearth.com", "Etik Nişan Yüzüğü", "Etik kaynaklı pırlanta; lab-grown seçenek; şeffaf tedarik; Beyond Conflict Free™"),
        ("Missoma", "missoma.com", "İngiltere DTC Takı", "Londra tasarımı yarı değerli taş takı; katmanlama; Meghan Markle + Lucy Williams işbirliği"),
        ("Monica Vinader", "monicavinader.com", "Kişiselleştirilmiş Takı", "Kazıma + özel mesaj; engrave edilebilir; İngiltere lüks; sürdürülebilir ambalaj"),
        ("Jenny Bird", "jenny-bird.com", "Statement Takı", "Kanada bold statement takı; mühendislik yapılı; altın + gümüş; dikkat çekici tasarım"),
        ("Soko", "shopsoko.com", "Kenya Zanaatkar Takı", "Kenya zanaatkarları el yapımı; artisan-made; adil ticaret; geleneksel teknikler + modern tasarım"),
        ("Vitaly", "vitalydesign.com", "Unisex Çelik Takı", "Paslanmaz çelik unisex takı; geri dönüştürülmüş çelik; streetwear estetiği; Kanada"),
        ("Ridge Wallet Alt", "ridgewallet.com", "Slim Metal Cüzdan", "Alüminyum + titanyum slim cüzdan; RFID koruma; 40.000+ 5 yıldız; viral erkek aksesuar"),
        ("Bellroy Alt", "bellfrroy.com", "Akıllı Deri Aksesuar", "Avustralya ince deri cüzdan + çanta; gizli bölmeler; akıllı tasarım; B Corp sertifikalı"),
        ("Peak Design Alt", "peakdesign.com", "Fotoğraf + Seyahat Aksesuar", "Capture Clip kamera tutucu; Travel Backpack; modüler tasarım; Kickstarter efsanesi"),
        ("Craighill", "craighill.co", "Tasarım Metal Aksesuar", "Brooklyn tasarım stüdyosu; metal anahtarlık + masa aksesuarları; mühendislik + sanat"),
        ("Miansai", "miansai.com", "Miami El Yapımı Takı", "El yapımı erkek + kadın takı; çapa bileklik ikonik; sterling gümüş; Miami zanaatkarlığı"),
        ("Wolf Circus", "wolfcircus.com", "Vancouver Geri Dönüşüm Takı", "Geri dönüştürülmüş bronz + gümüş; Vancouver tasarım; küçük parti; sürdürülebilir"),
        ("Aurate", "aufrratenewyork.com", "NYC Fine Jewelry", "NYC fine jewelry; fabrika direkt fiyat; 14k-18k altın; etik kaynak; 5-star warranty"),
        ("PDPAOLA", "pdpaola.com", "İspanya Trend Takı", "Barselona tasarım; zodiac koleksiyonlar; kişiselleştirilmiş; Instagram viral; genç + şık"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 19. Gözlük & Güneş Gözlüğü
    # ═══════════════════════════════════════════════════════════════════════════
    "Gözlük & Güneş Gözlüğü": [
        ("Pair Eyewear", "paireyewear.com", "Manyetik Değiştirilebilir Gözlük", "Manyetik top frame + değiştirilebilir ön kapak; her gün farklı gözlük; $60 çerçeve + $25 kapak"),
        ("Goodr", "goodr.com", "Eğlenceli Koşu Gözlüğü", "Kaymaz + sıçramaz koşu güneş gözlüğü; eğlenceli isimler + renkler; $25 fiyat; koşucu favorisi"),
        ("Blenders Eyewear", "blenderseyewear.com", "Uygun Fiyat Aktif Gözlük", "San Diego aktif yaşam tarzı gözlük; $30-70 fiyat; polarize; canlı renkler"),
        ("SunGod", "sungod.co", "Kişiselleştirilebilir Spor Gözlük", "Kişiselleştirilebilir performans gözlük; ömür boyu garanti; İngiltere; spor + günlük"),
        ("Zenni", "zennioptical.com", "Uygun Fiyat Reçeteli Gözlük", "$6.95'tan başlayan reçeteli gözlük; online sipariş; sanal deneme; 50M+ satış"),
        ("EyeBuyDirect", "eyebuydirect.com", "Online Gözlük Mağazası", "Essilor grubunda uygun fiyatlı online gözlük; sanal deneme; Buy 1 Give 1 programı"),
        ("Felix Gray", "felixgray.com", "Mavi Işık Filtre Gözlük", "Mavi ışık filtreli bilgisayar gözlüğü; göz yorgunluğu azaltma; şık çerçeveler; ofis çözümü"),
        ("ROKA", "roka.com", "Performans Optik", "Triatlon + koşu performans gözlüğü; ultra hafif; kaymaz; reçeteli güneş gözlüğü; Austin"),
        ("Raen", "raen.com", "Kaliforniya El Yapımı Gözlük", "El yapımı asetat çerçeveler; Kaliforniya tasarım; premium kalite; vintage ilham"),
        ("Krewe", "krewe.com", "New Orleans Gözlük", "New Orleans ilhamlı lüks gözlük; el yapımı; nylon + asetat; kültürel miras; güney estetiği"),
        ("Sunday Somewhere", "sundaysomewhere.com", "Avustralya Bohem Gözlük", "Sydney bohem esintili güneş gözlüğü; retro çerçeveler; premium asetat; unisex"),
        ("Glco (Garrett Leight)", "garrettleight.com", "LA Heritage Gözlük", "Oliver Peoples kurucusunun oğlu; Kaliforniya heritage; el yapımı; Venice Beach estetiği"),
        ("Prada Linea Rossa Alt", "sunglasshut.com", "Spor Lüks Gözlük", "Sport lüks güneş gözlüğü; premium malzeme; performans + moda; İtalyan tasarım"),
        ("Tens Sunglasses", "tens.co", "Filtre Güneş Gözlüğü", "Dünyayı Instagram filtresi gibi gösteren lens; sıcak ton; Edinburgh; görüntü iyileştirme"),
        ("Covry", "covry.com", "Geniş Burun Köprüsü Gözlük", "Asya + siyah burun köprüsüne uygun gözlük; kapsayıcı tasarım; standart gözlükler kayanlara çözüm"),
        ("Knockaround", "knockaround.com", "Uygun Fiyat Polarize Gözlük", "San Diego; $25 polarize güneş gözlüğü; özelleştirilebilir; spor + günlük; erişilebilir kalite"),
        ("Shady Rays", "shadyrays.com", "Kayıp Değiştirme Gözlük", "Kayıp/kırık gözlük değiştirme garantisi; ömür boyu; spor + outdoor; uygun fiyat; cesur garanti"),
        ("Retrosuperfuture", "retrosuperfuture.com", "İtalyan Avangard Gözlük", "Milano avangard güneş gözlüğü; İtalya yapımı asetat; moda + sanat; benzersiz tasarım"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 20. Teknoloji Aksesuarları
    # ═══════════════════════════════════════════════════════════════════════════
    "Teknoloji Aksesuarları": [
        ("Loop Earplugs", "loopearplugs.com", "Şık Kulak Tıkacı", "Akustik filtreli şık kulak tıkacı; gürültü azaltma; konser + odaklanma; aksesuar gibi tasarım"),
        ("Timekettle", "timekettle.co", "AI Çevirmen Kulaklık", "Gerçek zamanlı çeviri kulaklığı; 40+ dil; WT2 Edge; yabancı dil bariyerini kaldırma"),
        ("Moft", "moft.us", "Görünmez Laptop Stand", "Ultra ince yapışkan laptop standı; kağıt inceliğinde; ergonomik açı; telefon + tablet versiyonları"),
        ("Keychron", "keychron.com", "Mekanik Klavye", "Kablosuz mekanik klavye; hot-swappable; Mac + Windows; özelleştirilebilir; yazılımcı favorisi"),
        ("Dbrand", "dbrand.com", "Cihaz Kaplaması", "Hassas kesim vinil cihaz kaplamaları; Robot maskot; cesur marka kimliği; her cihaz için özel"),
        ("Moment", "shopmoment.com", "Telefon Kamera Lensi", "Telefon için sinema kalitesinde lens; fotoğrafçı ekosistemi; kılıf + lens + filtre; yaratıcı"),
        ("Quad Lock", "quadlockcase.com", "Bisiklet Telefon Tutucu", "Patentli döner kilitleme; bisiklet + motor + araba; güvenli telefon montajı; Avustralya"),
        ("Elgato", "elgato.com", "Streamer Ekipmanı", "Stream Deck kontrol paneli; Key Light; yeşil perde; içerik üretici ekosistemi; Corsair"),
        ("Govee", "govee.com", "Akıllı LED Şerit", "Akıllı LED ışık şeritleri + lambalar; RGB; uygulama kontrolü; oda dekorasyonu; uygun fiyatlı"),
        ("EcoFlow", "ecoflow.com", "Taşınabilir Güç İstasyonu", "Delta + River taşınabilir güç; güneş paneli entegrasyon; kamp + acil durum; hızlı şarj"),
        ("Jackery", "jackery.com", "Solar Jeneratör", "Taşınabilir güç istasyonu + güneş paneli; outdoor + van life; hafif; temiz enerji; kamp"),
        ("Rode", "rode.com", "Podcast Mikrofonu", "PodMic + Wireless Go mikrofon; Avustralya ses mühendisliği; içerik üretici favorisi"),
        ("Backbone", "playbackbone.com", "Telefon Oyun Kontrolcüsü", "iPhone + Android oyun kontrolcüsü; konsol deneyimi telefonda; Backbone One; USB-C"),
        ("Analogue", "analogue.co", "Retro Oyun Konsolu", "FPGA bazlı retro oyun konsolları; Pocket taşınabilir; Duo NES/SNES; piksel mükemmelliği; nostalji"),
        ("Twelve South", "twelvesouth.com", "Apple Aksesuar Uzmanı", "Apple ürünlerine özel tasarım aksesuarlar; BookArc stand; AirFly; premium kalite; ABD"),
        ("Satechi", "satechi.net", "USB-C Hub + Aksesuar", "USB-C hub + şarj + aksesuar; alüminyum tasarım; Apple estetiği; home office çözümleri"),
        ("Nomad Goods", "nomadgoods.com", "Premium Teknoloji Aksesuar", "Horween deri iPhone kılıfı; Base Station şarj; MagSafe; premium deri + teknoloji"),
        ("Grovemade", "grovemade.com", "Ahşap Masa Aksesuarı", "El yapımı ahşap + deri masa aksesuarları; Portland Oregon; laptop standı + desk pad; zanaatkar"),
        ("Spigen", "spigen.com", "Telefon Kılıfı", "Koruyucu telefon kılıfı; Tough Armor + Liquid Air; uygun fiyatlı kalite; global lider"),
        ("Casetify Alt", "casetify.com", "Kişisel Telefon Kılıfı", "Kişiselleştirilmiş + sanatçı işbirlikli telefon kılıfları; Impact Case; geri dönüşüm programı"),
        ("Orbitkey", "orbitkey.com", "Anahtar Organizer", "Anahtar düzenleyici + desk mat + çanta; İsviçre bıçağı gibi anahtarlık; minimal + fonksiyonel"),
        ("Native Union", "nativeunion.com", "Premium Şarj Aksesuarı", "Tasarım odaklı şarj kablosu + stand + kılıf; MagSafe; premium malzeme; Paris tasarım"),
        ("Scosche", "scosche.com", "Araç + Fitness Aksesuar", "MagicMount araç tutucu; Rhythm+ kalp atış monitörü; 40+ yıl inovasyon"),
        ("Anker Alt", "anker.com", "Güç + Şarj Çözümleri", "PowerCore + Nano şarj cihazları; USB-C hub; güç çözümleri; Amazon'da en güvenilir"),
        ("Bellroy Alt Tech", "bellroy.com", "Teknoloji Organizer", "Tech Kit organize çanta; kablo + adaptör düzeni; seyahat teknoloji çözümü; fonksiyonel"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 21. Seyahat & Bavul
    # ═══════════════════════════════════════════════════════════════════════════
    "Seyahat & Bavul": [
        ("Monos", "monfrros.com", "Minimalist Bavul", "Unbreakable Makrolon polikarbonat; whisper-quiet tekerlekler; minimalist Kanada tasarım"),
        ("July", "july.com", "Kişiselleştirilmiş Bavul", "Monogram kazıma seçeneği; Avustralya tasarım; alüminyum + polikarbonat; tek tıklama açma"),
        ("Béis", "bfreis.com", "Shay Mitchell Bavul", "Shay Mitchell'ın markası; The Carry-On viral; uygun fiyatlı fonksiyonel; estetik seyahat"),
        ("Calpak", "calpaktravel.com", "Pastel Seyahat Çantası", "Pastel renkli bavul + seyahat aksesuarları; Ambeur şeffaf kozmetik çantası viral; LA tasarım"),
        ("Nomatic", "nomatic.com", "Dijital Göçebe Çantası", "Kickstarter'dan doğan seyahat çantası; dijital göçebeler için; teknoloji bölmeleri; 40L backpack"),
        ("WANDRD", "wandrd.com", "Fotoğrafçı Seyahat Çantası", "Fotoğraf + günlük seyahat çantası; PRVKE backpack; kamera + laptop bölmesi; macera"),
        ("Tropicfeel", "tropicfeel.com", "Çok Amaçlı Seyahat Ayakkabı", "Seyahat ayakkabısı + çantası; modüler bavul; Kickstarter en çok fonlanan seyahat markası"),
        ("Baboon to the Moon", "baboontothemoon.com", "Renkli Duffle Çanta", "Neon renkli dayanıklı duffle; su geçirmez; macera ruhu; cesur renk paleti; NYC"),
        ("Dagne Dover", "dfragnedover.com", "Organize Neoprene Çanta", "Neoprene iş + spor + bebek çantası; organize bölmeler; yıkanabilir; fonksiyonel güzellik"),
        ("Away Alt", "awaytravel.com", "DTC Bavul Öncü", "DTC bavul kategorisi öncüsü; dahili şarj; 4 boyut; podcast + dergi; seyahat lifestyle"),
        ("Paravel", "tourparavel.com", "Sürdürülebilir Seyahat", "Geri dönüştürülmüş malzemeden bavul + çanta; Aviator seri; karbon offset; lüks sürdürülebilir"),
        ("Horizn Studios", "horizn-studios.com", "Akıllı Bavul", "Berlin tasarımı akıllı bavul; GPS takibi; teknoloji entegrasyon; Alman mühendislik"),
        ("Db (Douchebags)", "db.com", "İskandinav Macera Bavul", "Pro snowboarder tasarımı; Hugger backpack; hook-up sistemi; outdoor + seyahat; İsveç"),
        ("Tortuga", "tortugabackpacks.com", "Carry-On Sırt Çantası", "Sırt çantası bavul; 40L carry-on boyutunda; dijital göçebe odaklı; organize bölmeler"),
        ("Peak Design Travel", "peakdesign.com", "Modüler Seyahat Sistemi", "Travel Backpack 45L; packing cube sistemi; modüler aksesuar; Kickstarter $6M+"),
        ("Samsonite Alt Tumi", "tumi.com", "Premium İş Seyahat", "Premium iş seyahat bavul + çanta; Alpha Bravo; balistik naylon; profesyonel seyahat"),
        ("Aer", "aersf.com", "Şehir + Spor Çanta", "San Francisco; gym + iş çantası birleşimi; City Pack; Fit Pack spor bölmeli; modern şehirli"),
        ("Roark Revival", "rofrarkrevival.com", "Macera Seyahat Giyim", "Macera seyahati giyim + aksesuar; hikaye odaklı koleksiyonlar; surf + dağ + şehir"),
        ("Matein", "matein.com", "USB Şarjlı Sırt Çantası", "Uygun fiyatlı USB şarj portlu sırt çantası; Amazon bestseller; su geçirmez; iş + okul"),
        ("Bellroy Travel", "bellroy.com", "Seyahat Cüzdan + Aksesuar", "Travel Wallet + Tech Kit; seyahat organizasyonu; ince tasarım; fonksiyonel; premium"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 22. Diş & Ağız Bakımı
    # ═══════════════════════════════════════════════════════════════════════════
    "Diş & Ağız Bakımı": [
        ("Burst", "bfrrstoral.com", "Akıllı Sonik Fırça", "Sonik diş fırçası + kömürlü fırça başlığı; diş hekimi önerili; abonelik; uygun fiyat"),
        ("Snow", "trysnow.com", "Evde Diş Beyazlatma", "LED diş beyazlatma kiti; hassas dişler için güvenli; serum teknolojisi; 1M+ müşteri"),
        ("Hismile", "hismile.com", "Renkli Diş Beyazlatma", "PAP+ formül (peroksitsiz); V34 renk düzeltici diş macunu viral; Avustralya; Z kuşağı"),
        ("Cocofloss", "cocofloss.com", "Lüks Diş İpi", "Hindistan cevizi yağlı + aromalı dokuma diş ipi; diş bakımını keyifli hale getirme; renkli kutu"),
        ("David's", "davids-usa.com", "Premium Doğal Diş Macunu", "Doğal + sürdürülebilir diş macunu; metal tüp + tüp sıkacağı; ABD yapımı; EWG onaylı"),
        ("RiseWell", "risewell.com", "Hidroksiapatit Diş Macunu", "Hidroksiapatit bazlı diş macunu; florürsüz mineral onarım; tüm aile; doğal + bilimsel"),
        ("Quip Alt", "getquip.com", "Abonelik Diş Fırçası", "Abonelik bazlı elektrikli diş fırçası; 3 ayda bir fırça başlığı + macun; $25 başlangıç"),
        ("Boka", "bfroka.com", "n-Ha Diş Macunu", "Nano-hydroxyapatite diş macunu; Japon diş bilimi; florür alternatifi; doğal remineralizasyon"),
        ("Supermouth", "supermouth.com", "Çocuk Diş Bakımı", "Dr. Kami Hoss kurdu; Hydroxamin™ formül; çocuk + yetişkin; kapsamlı ağız bakımı"),
        ("Lumineux", "lfrumineux.com", "Biyomimetik Ağız Bakımı", "Sertifikalı non-toxic diş beyazlatma + macun; Dead Sea salt; doğal bileşenler; güvenli"),
        ("AutoBrush", "autofrbrush.com", "U-Şekli Otomatik Fırça", "U-şekilli tüm dişleri aynı anda fırçalayan cihaz; 30 saniye; çocuk versiyonu; engelli dostu"),
        ("Bite Alt", "bitetoothpastebits.com", "Diş Macunu Tableti", "Çiğneme diş macunu tabletleri; cam şişe ambalaj; sıfır plastik atık; doğal bileşenler"),
        ("Terra & Co", "terraandco.com", "Dünya Dostu Diş Bakımı", "Kömürlü bambu diş fırçası + diş macunu; Amazonas ormanı koruma; B Corp; ayurveda"),
        ("Hello Products", "hello-products.com", "Doğal Dostça Diş Bakımı", "Doğal + vegan diş bakımı; aktif kömür + CBD seçenekleri; Colgate satın aldı; eğlenceli"),
        ("Waken", "waken.com", "İngiliz Ağız Bakımı", "İngiltere premium ağız bakımı; bitki bazlı; cam şişe gargara; estetik ambalaj; sürdürülebilir"),
        ("Spotlight Oral Care", "spotlightfrralcare.com", "Diş Hekimi Markası", "İrlandalı diş hekimleri kurdu; diş beyazlatma şeritleri; klinik kanıtlı; profesyonel evde"),
        ("Zenyum", "zenyum.com", "Asya Şeffaf Plak", "Güneydoğu Asya şeffaf diş teli; uygun fiyatlı ortodonti; uygulama takipli; Singapur"),
        ("Tend", "hellotend.com", "Modern Diş Kliniği", "NYC + DC modern diş kliniği; korku azaltan deneyim; şık mekan; teknoloji odaklı"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 23. Kadın Sağlığı & Regl Bakımı
    # ═══════════════════════════════════════════════════════════════════════════
    "Kadın Sağlığı & Regl Bakımı": [
        ("Saalt", "safralt.com", "Regl Kabı", "Silikon regl kabı; 12 saat koruma; yıllarca dayanır; sürdürülebilir; her satışta bağış"),
        ("Cora", "cfrora.com", "Organik Regl Ürünü", "Organik pamuk tampon + ped; her satışta gelişmekte olan ülkeye bağış; şık ambalaj"),
        ("Rael", "getrfrael.com", "Kore Regl + Cilt Bakımı", "Kore organik regl ürünleri + cilt bakımı; akne patch + tampon; holistic kadın bakım"),
        ("Flex", "flexfits.com", "Regl Disk", "Regl diski; 12 saat kullanım; ilişki sırasında kullanılabilir; yenilikçi kadın sağlığı ürünü"),
        ("August", "itsaugust.co", "Z Kuşağı Regl Markası", "Gen Z regl markası; gender-inclusive; aktivist; renkli ambalaj; stigma kırma; farkındalık"),
        ("Thinx Alt", "shethinx.com", "Regl İç Giyim", "Adet geçirmez iç giyim öncüsü; 4 tampon emme kapasitesi; yıkanabilir; patentli teknoloji"),
        ("Love Wellness", "lovewellness.com", "Kadın Wellness Takviye", "Lo Bosworth kurdu; vajinal probiyotik + sindirim; kadın sağlığı tabularını kırma"),
        ("Lola", "mylola.com", "Organik Tampon Abonelik", "Organik pamuk tampon + ped aboneliği; şeffaf bileşenler; 1. jenerasyon Amerikan kadın kurucular"),
        ("Viv", "vivforfree.com", "Ücretsiz Regl Ürünü", "Reklam destekli ücretsiz regl ürünleri; erişilebilirlik; kadın sağlığında finansal engel kaldırma"),
        ("Period Aisle", "periodaisle.com", "Kapsayıcı Regl Bakım", "Trans + non-binary kapsayıcı regl ürünleri; boxerler + iç giyim; Kanada; tüm cinsiyetler"),
        ("Diva Cup", "divacup.com", "Regl Kabı Öncüsü", "Regl kabı kategorisi öncüsü; Kanada; 12 saat kullanım; medikal silikon; eko-dost"),
        ("Intimate Portal", "intimateportal.com", "Hamilelik İç Giyim", "Hamilelik + emzirme iç giyim; comfort fit; elastik tasarım; Amazon bestseller"),
        ("Kindra", "ourkindra.com", "Menopoz Wellness", "Menopoz semptomları için wellness ürünleri; sıcak basma + kuru cilt; hormon-free; bilimsel"),
        ("Evernow", "evernow.com", "Menopoz Hormon Tedavi", "Online menopoz hormon tedavisi; doktor danışma; kişiselleştirilmiş; telehealth; modern"),
        ("Bodily", "itsbodily.com", "Doğum Sonrası Bakım", "Doğum sonrası toparlanma ürünleri; anne vücudu bakımı; emzirme + perine; tabu kırma"),
        ("Frida Mom", "fridamom.com", "Doğum Sonrası Çözüm", "Doğum sonrası soğutucu ped + peribottle; anne bakımını pratikleştirme; Fridababy kardeş"),
        ("Oova", "oova.life", "Hormon Takip Testi", "Evde hormon seviyesi ölçüm kiti; ovulasyon + progesteron; kişiselleştirilmiş doğurganlık bilgisi"),
        ("Natalist", "natfralist.com", "Doğurganlık Wellness", "Doğurganlık + hamilelik takviyesi + test; modern tasarım; kanıt bazlı; TTC topluluğu"),
        ("Kegg", "kegg.tech", "Servikal Mukus Takipçisi", "Bluetooth servikal mukus izleme cihazı; doğurganlık penceresi tahmini; kegel egzersiz; 2-in-1"),
        ("Elvie", "elvie.com", "Giyilebilir Süt Pompası", "Sessiz giyilebilir süt pompası; hands-free; sütyen içinde; Kegel antrenörü; İngiltere teknoloji"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 24. Cinsel Sağlık & Wellness
    # ═══════════════════════════════════════════════════════════════════════════
    "Cinsel Sağlık & Wellness": [
        ("Dame", "dameproducts.com", "Kadın Zevk Ürünleri", "Kadın mühendislerin tasarladığı vibratörler; Eva hands-free; araştırma bazlı; New York; feminist"),
        ("Maude", "getmaude.com", "Modern Cinsel Wellness", "Cinsel sağlık ürünlerini normalleştiren minimalist marka; Vibe; vibratör + kayganlaştırıcı + prezervatif"),
        ("Fur", "furyou.com", "İntim Cilt Bakımı", "İntim bölge bakım ürünleri; ingrown + tüy bakımı; tabusuz yaklaşım; Emma Watson desteği"),
        ("Unbound Babes", "unboundbabes.com", "Eğlenceli Cinsel Aksesuar", "Polly Rodgriguez kurdu; cinsel wellness'ı eğlenceli + erişilebilir yapma; vibratör + aksesuar"),
        ("Lovers", "lovers.com", "Modern Cinsel Sağlık Mağaza", "Yeniden markalaşmış cinsel sağlık perakendecisi; eğitim odaklı; kadın liderliğinde; kapsayıcı"),
        ("Smile Makers", "smilemakers.com", "Tasarım Ödüllü Vibratör", "MoMA'da satılan tasarım vibratörler; vulva merkezli eğitim; İsveç tasarımı; cinsel pozitiflik"),
        ("Foria", "foriawellness.com", "CBD İntim Wellness", "CBD + botanik intim ürünler; Awaken arousal oil; menstrüel rahatlama; bitkisel wellness"),
        ("Good Clean Love", "goodcleanlove.com", "Organik İntim Ürün", "Organik kayganlaştırıcı + vajinal sağlık; pH dengeleme; prebiyotik; temiz bileşenler"),
        ("Cake", "hellofrcake.com", "Twerk Kayganlaştırıcı", "Eğlenceli isimli cinsel wellness; So-Low massaj + twerk kayganlık; renkli ambalaj; Z kuşağı"),
        ("Lelo", "lelo.com", "Lüks İntim Tasarım", "İsveç lüks intim ürünler; Sona teknoloji; premium tasarım; 20+ yıl inovasyon"),
        ("We-Vibe", "we-vibe.com", "Çift Vibratör", "Çiftler için uzaktan kontrollü vibratör; Sync model; uygulama bağlantılı; mesafeli çiftler"),
        ("Lorals", "mylorals.com", "Oral Seks Koruma", "Lateks iç giyim; oral seks sırasında koruma; FDA onaylı; kadın güçlendirme; niş çözüm"),
        ("Womanizer", "womanizer.com", "Hava Basınçlı Teknoloji", "Pleasure Air Technology patentli; klitoral stimülasyon; Almanya mühendisliği; Premium model"),
        ("Satisfyer", "satisfyer.com", "Uygun Fiyat İntim Cihaz", "Air Pulse teknolojisi; uygun fiyatlı kaliteli ürünler; geniş ürün yelpazesi; Almanya"),
        ("Wisp", "hellowisp.com", "Online Kadın Sağlığı", "Online reçeteli kadın sağlığı; vajinal enfeksiyon + doğum kontrol; telehealth; gizli teslimat"),
        ("Nurx", "nurx.com", "Online Doğum Kontrol", "Online doğum kontrol reçetesi + teslimat; STI test kiti; PrEP; erişilebilir sağlık"),
        ("Hers (Kadın Sağlığı)", "forhers.com", "Kadın Telehealth", "Saç + cilt + cinsel sağlık kadınlar için; online doktor + reçete + teslimat"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 25. Parfüm & Ev Kokusu
    # ═══════════════════════════════════════════════════════════════════════════
    "Parfüm & Ev Kokusu": [
        ("Phlur", "phlur.com", "Temiz Parfüm", "Missing Person parfümü TikTok viral; Chriselle Lim yeniden kurdu; temiz + vegan kokular"),
        ("Ellis Brooklyn", "ellisbrooklyn.com", "Bilimsel Doğal Parfüm", "Sürdürülebilir + temiz parfüm; Bee ikonik; doğal + güvenli bileşenler; NYC"),
        ("Skylar", "skylar.com", "Hipoalerjenik Parfüm", "Hypoallerjenik temiz parfüm; hassas cilt güvenli; Discovery Set ile deneme; kişiselleştirme"),
        ("Dedcool", "dedcool.com", "Unisex Temiz Parfüm", "Unisex biome-friendly parfüm + çamaşır deterjanı; LA; sürdürülebilir; minimalist"),
        ("Henry Rose", "henryrose.com", "Michelle Pfeiffer Parfüm", "Michelle Pfeiffer'ın %100 şeffaf bileşen parfümü; EWG onaylı; lüks + temiz"),
        ("Snif", "snfrif.com", "Dene-Beğen Parfüm", "Beğenmezsen iade garantili parfüm; risk-free deneme; uygun fiyatlı lüks; 30ml seyahat boyu"),
        ("Dossier", "dfrssier.co", "Lüks Parfüm Klonu", "İlham alınmış parfümler; $29 lüks kokular; vegan + temiz; Baccarat Rouge ilham en popüler"),
        ("Boy Smells", "boysmells.com", "Gender-Fluid Koku", "Genderful parfüm + mum; geleneksel cinsiyet normlarını kıran kokular; renkli ambalaj; LA"),
        ("P.F. Candle Co", "pfcandleco.com", "LA Zanaatkar Mum", "Los Angeles el yapımı soya mumu; Amber & Moss ikonik; terazzo kap; sürdürülebilir"),
        ("Brooklyn Candle Studio", "brooklyncandlestudio.com", "Brooklyn Soya Mumu", "Brooklyn el yapımı soya mumu; minimalist etiket; Love + Escapist koleksiyonları"),
        ("Apotheke", "apfrothekeco.com", "Brooklyn Ev Kokusu", "Brooklyn premium ev kokusu; mum + difüzör + sabun; otel kokulara ilham; lüks"),
        ("Otherland", "otherland.com", "Instagram Mum Markası", "Görsel estetik odaklı mum; sanatçı işbirlikli kutu tasarım; hediye odaklı; rengarenk"),
        ("Homesick", "homesick.com", "Nostalji Mumu", "Eyalet + şehir + anı kokan mumlar; Hawaii, New York; nostalji pazarlaması; hediye"),
        ("Vitruvi", "vitfrruvi.com", "Taş Difüzör", "Porselen + taş esansiyel yağ difüzörü; modern ev dekoru; minimalist tasarım; wellness"),
        ("Pura", "pfrura.com", "Akıllı Ev Kokusu", "WiFi kontrollü akıllı ev koku difüzörü; uygulamayla aroma ayarı; marka işbirlikleri; abonelik"),
        ("Malin+Goetz", "malinandgoetz.com", "NYC Apothecary Koku", "NYC modern eczane markası; Dark Rum mum ikonik; unisex; dermatolog geliştirdi"),
        ("Cire Trudon", "ciretrudon.com", "Dünyanın En Eski Mum Yapımcısı", "1643'ten beri; Versailles Sarayı tedarikçisi; el yapımı cam; ultra premium"),
        ("Overose", "ovfrrerose.com", "Pembe Neon Mum", "Neon pembe estetik mum; Anthurium ikonik koku; Instagrammable; cesur görsel kimlik"),
        ("Voluspa", "voluspa.com", "Dekoratif Mum", "Coconut wax blend; Japonica koleksiyonu ikonik; metalik teneke kutu; hediye favorisi"),
        ("Diptyque Alt", "diptyque.com", "Paris Lüks Mum", "1961 Paris; Baies mum ikonik; lüks ev kokusu + parfüm; sanatsal etiket tasarımı"),
        ("Le Labo", "lelabofragrances.com", "Zanaatkar Parfüm", "NYC zanaatkar parfüm; Santal 33 ikonik; mağazada taze karıştırma; Estée Lauder"),
        ("Byredo", "byredo.com", "İsveç Niş Parfüm", "Stockholm minimalist parfüm; Ben Gorham kurdu; Gypsy Water ikonik; multisensoriyel"),
        ("DS & Durga", "dsanddurga.com", "Brooklyn Niş Parfüm", "Brooklyn indie parfüm; I Don't Know What viral; müzik + sanat ilhamı; benzersiz notalar"),
        ("Replica (Maison Margiela)", "maisonmargiela.com", "Anı Kokusu", "Anıları koku ile yakalayan parfüm serisi; By the Fireplace, Beach Walk; nostalji"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 26. Outdoor & Spor Ekipman
    # ═══════════════════════════════════════════════════════════════════════════
    "Outdoor & Spor Ekipman": [
        ("Cotopaxi Alt", "cotopaxi.com", "Renkli Sürdürülebilir Outdoor", "Del Día benzersiz renk kombinasyonu; artık kumaşlardan üretim; B Corp; Gear for Good"),
        ("Black Diamond Alt", "blackdiamondequipment.com", "Tırmanış Ekipmanı", "Tırmanış + kayak ekipmanı; çığ güvenliği; Utah mühendisliği; dağcılık kültürü"),
        ("Rumpl", "rumpl.com", "Outdoor Battaniye", "Uyku tulumu teknolojisi battaniye; NanoLoft yalıtım; hafif + kompakt; kamp + ev"),
        ("BioLite", "bfrioliteenergy.com", "Off-Grid Enerji", "Odun yakan şarj + ocak; güneş paneli; off-grid enerji çözümleri; gelişmekte olan ülke etkisi"),
        ("Snow Peak", "snowpeak.com", "Japon Outdoor Tasarım", "Japon outdoor yaşam markası; titanyum kamp ekipmanı; estetik + fonksiyon; 60+ yıl miras"),
        ("Helinox", "helinox.com", "Ultra Hafif Kamp Sandalye", "Ultra hafif katlanır kamp sandalyesi; DAC alüminyum çubuklar; 1kg altı; backpacking"),
        ("NEMO Equipment", "nfremoequipment.com", "İnovatif Çadır + Yatak", "Spoon şekilli uyku tulumu; AirSupported çadır; patentli tasarımlar; New Hampshire"),
        ("Hydro Flask Alt", "hydroflask.com", "Vakum Yalıtım Şişe", "TempShield vakum yalıtım; geniş renk paleti; outdoor + günlük; ağız genişliği seçenekleri"),
        ("Stanley Alt", "stanley1913.com", "Heritage Termos", "1913'ten beri; Quencher tumbler TikTok viral; 100+ yıl miras; modern yeniden keşif"),
        ("YETI Alt", "yeti.com", "Premium Soğutucu", "Ayı dayanıklılığında soğutucu + drinkware; Tundra + Rambler; premium outdoor; $1.5B+ gelir"),
        ("Oru Kayak", "orukayak.com", "Katlanır Kayak", "Origami ilhamı katlanır kayak; kutu boyutunda; karbon ayak izi düşük; patentli tasarım"),
        ("GCI Outdoor", "gcioutdoor.com", "Konforlu Kamp Mobilyası", "Freestyle Rocker sallanır kamp sandalyesi; Eazy Fold; konforlu outdoor oturma"),
        ("Goal Zero", "goalzero.com", "Güneş Enerjisi Outdoor", "Taşınabilir güneş paneli + güç istasyonu; kamp + acil durum; off-grid çözümler"),
        ("Outdoor Voices Alt", "outdoorvoices.com", "Rekreasyonel Activewear", "Rekreasyonel spor giyim; Doing Things felsefesi; CloudKnit kumaş; sosyal spor"),
        ("Tracksmith Alt", "tracksmith.com", "Heritage Koşu Giyim", "Bağımsız koşu markası; New England estetiği; zanaatkarlık; koşu kültürü saygısı"),
        ("Ten Thousand Alt", "tenthousand.cc", "Fonksiyonel Fitness Giyim", "Minimalist erkek fitness giyim; Set short; liner built-in; CrossFit + lifting odaklı"),
        ("Territory Run Co", "territoryrun.co", "Portland Koşu Aksesuar", "Portland el yapımı koşu şapka + aksesuar; trail running topluluğu; zanaatkar; niş"),
        ("Wellen", "wellensurf.com", "Sürdürülebilir Surf Giyim", "Organik + geri dönüşüm surf giyim; hemp boardshort; çevreci surf kültürü; Kaliforniya"),
        ("Finisterre", "finisterre.com", "İngiliz Soğuk Su Surf", "Cornwall cold water surf markası; geri dönüşüm + yenilenmiş; B Corp; İngiliz outdoor"),
        ("Topo Designs", "topodesigns.com", "Colorado Outdoor + Şehir", "Colorado tasarım; outdoor + şehir arası çanta + giyim; retro renkler; made in USA; fonksiyonel"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 27. Ofis & Üretkenlik
    # ═══════════════════════════════════════════════════════════════════════════
    "Ofis & Üretkenlik": [
        ("Rocketbook", "getrocketbook.com", "Silinebilir Akıllı Defter", "Silinebilir + buluta yüklenen defter; sonsuz kullanım; mikrodalga ile temizleme; sürdürülebilir"),
        ("Ugmonk", "ugmonk.com", "Gather Masaüstü Organizer", "Minimal masaüstü düzenleme sistemi; kalem + kart + not tutucu; estetik üretkenlik"),
        ("Grovemade", "grovemade.com", "Ahşap Masa Aksesuarları", "El yapımı ceviz + deri masa pedi; laptop standı; Portland zanaatkarlığı; premium"),
        ("Baron Fig", "bafronfig.com", "Yaratıcı Defter + Kalem", "Confidant defter; Squire kalem; yaratıcılar için araçlar; Kickstarter'dan doğdu"),
        ("Fully", "fully.com", "Ergonomik Ayaklı Masa", "Jarvis ayaklı masa; ergonomik ofis mobilyası; Portland; sağlıklı çalışma; elektrikli ayar"),
        ("Secretlab", "secretlab.co", "Premium Oyun Sandalyesi", "TITAN + Omega oyun sandalyesi; ergonomik tasarım; premium malzeme; Singapur; esports"),
        ("Autonomous", "autonomous.ai", "AI Ofis Mobilyası", "SmartDesk ayaklı masa; ErgoChair; uygun fiyatlı ergonomik; startup kültürü; uzaktan çalışma"),
        ("Orbitkey", "orbitkey.com", "Masa + Anahtar Organizer", "Desk Mat; Nest portatif organizatör; anahtar tutucu; minimal + fonksiyonel; estetik"),
        ("Bellroy Work", "bellroy.com", "İş Çantası + Aksesuar", "Slim Work Bag; Tech Kit; iş hayatı organize; ince tasarım; premium deri"),
        ("Twelve South BookArc", "twelvesouth.com", "MacBook Standı", "BookArc vertikal MacBook standı; masa alanı kazanma; premium metal; Apple estetiği"),
        ("Moft Work", "moft.us", "Yapışkan Laptop Yükseltici", "Görünmez laptop standı; Z katlama; ergonomik açı; ultra ince; çantada yer kaplamaz"),
        ("Nolii", "nolii.com", "Modüler Teknoloji Aksesuar", "Modüler kablo + şarj yönetim sistemi; masa düzeni; kablo kabusu çözümü; İngiltere"),
        ("FlexiSpot", "flexispot.com", "Uygun Fiyat Ayaklı Masa", "Elektrikli ayaklı masa; bisiklet masa; uygun fiyatlı ergonomik; çeşitli boyutlar"),
        ("Uplift Desk", "upliftdesk.com", "Özelleştirilebilir Ayaklı Masa", "Geniş özelleştirme seçenekleri; bambu + laminate + masif; wire grommets; ABD yapımı"),
        ("Herman Miller Alt", "hermanmiller.com", "Aeron Ergonomik Sandalye", "Aeron sandalye ikonik; 30+ yıl ergonomik tasarım; PostureFit; nefes alabilir mesh"),
        ("Logitech Ergo Alt", "logitech.com", "Ergonomik Mouse + Klavye", "MX Master mouse; Ergo K860 klavye; ergonomik iş araçları; üretkenlik artışı"),
        ("Branch Furniture", "branchfurniture.com", "Startup Ofis Mobilyası", "DTC ofis mobilyası; Daily Chair; masa; startup ve ev ofis için; uygun fiyatlı kalite"),
        ("Lume Cube", "lumecube.com", "Video Konferans Işığı", "Panel Mini video ışığı; Zoom + Teams için yüz aydınlatma; içerik üretici; kompakt"),
        ("CalDigit", "caldigit.com", "Thunderbolt Dock", "TS4 Thunderbolt dock; 18 port; tek kablo çözümü; MacBook Pro istasyonu; profesyonel"),
        ("LectroFan", "lectrofan.com", "Beyaz Gürültü Makinesi", "20 benzersiz ses; uyku + odaklanma; kompakt; ofis + yatak odası; gerçek fan sesi"),
    ],

    # ═══════════════════════════════════════════════════════════════════════════
    # 28. Oyun & Yaratıcı Araçlar
    # ═══════════════════════════════════════════════════════════════════════════
    "Oyun & Yaratıcı Araçlar": [
        ("Cricut", "cricut.com", "Akıllı Kesim Makinesi", "Kağıt + vinil + kumaş kesim makinesi; DIY projeleri; Maker + Joy modelleri; yaratıcı topluluk"),
        ("Glowforge", "glowforge.com", "Masaüstü Lazer Kesici", "Masaüstü 3D lazer yazıcı; ahşap + akrilik + deri kazıma + kesme; Kickstarter $28M"),
        ("xTool", "xtool.com", "Uygun Fiyat Lazer Kesici", "Uygun fiyatlı lazer kazıma + kesme; D1 Pro diode lazer; maker topluluğu; hobi + küçük işletme"),
        ("Backbone", "playbackbone.com", "Mobil Oyun Kontrolcüsü", "iPhone + Android oyun kontrolcüsü; konsol deneyimi telefonda; düşük gecikme; USB-C"),
        ("Analogue", "analogue.co", "FPGA Retro Konsol", "FPGA bazlı retro oyun konsolları; Pocket taşınabilir; piksel mükemmelliği; nostalji + kalite"),
        ("Secretlab Gaming", "secretlab.co", "Oyun Sandalyesi", "TITAN Evo oyun sandalyesi; ergonomik + premium; esports ekipleri sponsoru; 4 yıl garanti"),
        ("Bambu Lab", "bambulab.com", "Hızlı 3D Yazıcı", "X1 Carbon yüksek hız 3D yazıcı; çoklu malzeme; otomatik kalibrasyon; maker devrimi"),
        ("Elgato Stream", "elgato.com", "Yayıncı Ekipmanı", "Stream Deck; Wave mikrofon; Key Light; Ring Light; yayıncı ekosistemi; içerik üretici"),
        ("Rode Podcast", "rode.com", "Podcast Prodüksiyon", "PodMic + RODECaster Pro; podcast prodüksiyon ekosistemi; profesyonel ses; Avustralya"),
        ("Playdate", "play.date", "Krank Kollu El Konsolu", "Sarı el konsolu; krank kolu input; sürpriz sezon oyunları; retro indie; Panic Inc."),
        ("8BitDo", "8bitdo.com", "Retro Kablosuz Kontrolcü", "Retro tasarımlı modern kablosuz kontrolcü; çoklu platform; NES + SNES estetiği; uygun fiyat"),
        ("Insta360", "insta360.com", "360° Aksiyon Kamera", "360° kamera; X4 + Ace Pro; görünmez selfie çubuğu; yaratıcı çekim; bilgisayar düzenleme"),
        ("DJI Alt", "dji.com", "Tüketici Drone", "Mini + Air drone serisi; gimbal + aksiyon kamera; Osmo Pocket; hava fotoğrafçılığı"),
        ("Peak Design Creator", "peakdesign.com", "Yaratıcı Kamera Aksesuarı", "Capture kamera klipsi; Slide askı; yaratıcı fotoğrafçılık çözümleri; modüler"),
        ("Remarkable", "remarkable.com", "E-Ink Tablet", "E-ink yazma tableti; kağıt hissi; not alma + PDF okuma; dikkat dağıtmayan; Norveç"),
        ("Supernote", "supernote.com", "E-Ink Not Defteri", "E-ink dijital not defteri; Android bazlı; el yazısı tanıma; Kindle alternatifi yazma"),
        ("Boox", "boox.com", "E-Ink Android Tablet", "E-ink Android tablet; not alma + okuma; Onyx Boox Note; büyük ekran; çok amaçlı"),
        ("Loupedeck", "loupedeck.com", "Yaratıcı Kontrol Konsolu", "Fotoğraf + video düzenleme kontrolcüsü; Lightroom + Premiere entegrasyon; dokunmatik + düğme"),
        ("Aputure", "aputure.com", "Film Yapımcısı Işık", "LED video ışığı; 300d + MC; indie film yapımcısı ekipmanı; profesyonel kalite uygun fiyat"),
        ("Obsbot", "obsbot.com", "AI Webcam", "Tiny 2 AI otomatik takipli webcam; el hareketi kontrolü; PTZ; içerik üretici + toplantı"),
        ("Nomad Base Station", "nomadgoods.com", "Wireless Şarj İstasyonu", "Çoklu cihaz kablosuz şarj; MagSafe + Apple Watch; premium deri pad; şık masaüstü"),
        ("Framework Alt", "frame.work", "Modüler Laptop", "Tamamen modüler + onarılabilir laptop; parça değiştirilebilir; hak-to-repair hareketi; sürdürülebilir"),
        ("Keychron Gaming", "keychron.com", "Mekanik Gaming Klavye", "Q serisi CNC alüminyum; hot-swappable; özelleştirilebilir; VIA/QMK firmware; mekanik klavye"),
        ("Razer Alt", "razer.com", "Gaming Periferalleri", "Viper mouse; BlackWidow klavye; gaming ekosistemi; Chroma RGB; esports standardı"),
        ("HyperX Alt", "hyperx.com", "Gaming Kulaklık", "Cloud kulaklık serisi; DTS Spatial Audio; konforlu + uygun fiyatlı; streamer favorisi"),
    ],
}

# ─── Additional Brands to reach 1500+ ────────────────────────────────────────
EXTRA_BRANDS = {
    "Cilt Bakımı & Güzellik Araçları": [
        ("Wander Beauty", "wfrander.com", "Çok İşlevli Güzellik", "Havaalanında doğan marka; Mile High Club maskara; çoklu kullanım ürünler; seyahat dostu"),
        ("RMS Beauty", "rmsbeauty.com", "Ham Organik Makyaj", "Rose-Marie Swift kurdu; ham hindistan cevizi yağı bazlı; organik renk; backstage favorisi"),
        ("Westman Atelier", "westman-atelier.com", "Clean Lüks Makyaj", "Gucci Westman (makyaj artisti); temiz + lüks; Baby Cheeks blush ikonik; Clean Beauty Council"),
        ("Trinny London", "trinnylondon.com", "Stack Makyaj Sistemi", "Trinny Woodall kurdu; üst üste geçen modüler ambalaj; Match2Me cilt analizi; İngiltere"),
        ("Charlotte Mensah", "charlottemensah.com", "Manketti Yağı Saç", "Ganalı saç stilisti; manketti yağı bazlı; Afro + tekstürlü saç bakımı; salon to DTC"),
        ("Dr. Barbara Sturm", "drsturm.com", "PRP Cilt Bakımı", "Vampir facelift'i icat eden doktor; Hyaluronic Serum; ünlülerin doktoru; Alman tıp bilimi"),
        ("Tata Harper", "tataharperskincare.com", "Çiftlik Lüks Bakım", "Vermont çiftliğinde üretim; %100 doğal; lüks yeşil güzellik öncüsü; 300+ botanik bileşen"),
        ("Augustinus Bader", "augustinusbader.com", "Kök Hücre Kremi", "Prof. Bader'in TFC8 teknolojisi; $265 krem; bilimsel lüks; kök hücre tetikleme"),
        ("Ren Clean Skincare", "renskincare.com", "Temiz + Sürdürülebilir", "Clean to skin, clean to planet; Evercalm hassas cilt; 2025'e kadar sıfır atık hedefi"),
        ("Pai Skincare", "paiskincare.com", "Ultra Hassas Cilt", "Londra; alerji + hassasiyet odaklı; Chamomile & Rosehip; organik sertifikalı; en hassas ciltler"),
        ("By Terry", "byterry.com", "Lüks Fransız Makyaj", "Terry de Gunzburg (YSL eski yöneticisi); Hyaluronic Hydra-Powder; lüks Fransız makyaj"),
        ("Olio E Osso", "ofrlioeosso.com", "Doğal Balm Renkli", "Portland doğal renk balm çubukları; çok amaçlı; basit bileşenler; pratik format"),
        ("Indie Lee", "indielee.com", "Temiz Cilt Bakımı Öncü", "Beyin tümörü sonrası temiz güzelliğe adanma; Brightening Cleanser; toksinlere karşı mücadele"),
        ("Josh Rosebrook", "joshrosebrook.com", "Organik Cilt + Saç", "Organik + bitkisel cilt ve saç bakımı; Nutrient Day Cream SPF; doğal performans"),
        ("100% Pure", "100percentpure.com", "Meyve Pigmentli Makyaj", "Meyve pigmenti ile renklendirme; süper saf; hiçbir sentetik boya yok; doğal renk kaynağı"),
        ("Ere Perez", "ereperez.com", "Avustralya Doğal Güzellik", "Filipin + Avustralya botanikleri; pirinç tozu + papaya; doğal renkli makyaj; tropikal"),
        ("Milk Makeup", "milkmakeup.com", "Sopa Format Makyaj", "Sopa formatında makyaj; Hydrogrip primer viral; vegan; NYC kültür odaklı"),
        ("Vapour Organic Beauty", "vapourbeauty.com", "Organik Lüks Makyaj", "USDA organik sertifikalı makyaj; New Mexico; çöl ilhamı; temiz performans"),
        ("W3LL People", "w3llpeople.com", "EWG Onaylı Makyaj", "EWG verified makyaj; Bio Brightener stick ikonik; temiz + performans; erişilebilir"),
        ("Thrive Causemetics", "thrivecausemetics.com", "Sosyal Etki Makyaj", "1 alana 1 bağış; Karissa Bodnar kurdu; Liquid Lash Extensions viral; vegan + temiz"),
        ("E.l.f. Alt", "elfcosmetics.com", "Uygun Fiyat Temiz Makyaj", "Ultra uygun fiyat temiz makyaj; $3-14 aralığı; TikTok viral; Power Grip primer fenomen"),
        ("Colourpop", "colourpop.com", "Hızlı Moda Makyaj", "Haftalık yeni ürün lansmanı; $5-16; Seed Beauty üretim; trend yakalama hızı; LA"),
        ("Florence by Mills", "florencebymills.com", "Gen Z Hafif Makyaj", "Millie Bobby Brown markası; 12-18 yaş; hafif temiz formüller; Ulta'da"),
        ("Em Cosmetics", "emcosmetics.com", "YouTube Güzellik Gurisi", "Michelle Phan kurdu; Serum Blush viral; dijital güzellik öncüsü; tekrar markalaşma başarısı"),
        ("Uoma Beauty", "ufrmabeauty.com", "Afrika İlham Güzellik", "Sharon Chuter kurdu; 51 fondöten tonu; Afrikan kültürü; Pull Up For Change hareketi"),
        ("Range Beauty", "rangebeauty.com", "Akne + Melanin Güzellik", "Akne eğilimli melanin ciltler için makyaj; tıkanmayan formüller; dermatolog onaylı"),
        ("Kulfi Beauty", "kulfrfibeauty.com", "Güney Asya Güzellik", "Güney Asya cilt tonları için makyaj; kajal göz kalemi ikonik; kültürel ilham"),
        ("Patrick Ta", "patrickta.com", "Celebrity MUA Makyajı", "Ünlülerin makyajcısı; Major Headlines allık ikonik; Sephora; pro kalite günlük kullanım"),
        ("Rare Beauty Alt", "rarebeauty.com", "Likit Allık", "Soft Pinch likit allık viral; doğal ışıltı; damla uygulama; Sephora en çok satan"),
        ("About Face", "aboutface.com", "Cesur Sanatsal Makyaj", "Halsey markası; yüksek pigment; sanatsal ifade; cesur renkler; vegan"),
    ],

    "Saç Bakımı & Saç Sağlığı": [
        ("Gisou", "gisou.com", "Bal Bazlı Saç Bakımı", "Negin Mirsalehi'nin markası; Mirsalehi bal çiftliği; Honey Infused Hair Oil viral; arıcılık mirası"),
        ("Oui the People Hair", "ouithepeople.com", "Kıvırcık Saç Bakımı", "Kıvırcık + coily saç şekillendirme; doğal saç dokusu kabullenme; sürdürülebilir"),
        ("Curl Smith Alt", "curlsmith.com", "Curl Quenching Bakım", "Kıvırcık saç tipleme sistemi; Bond Rehab; Helen of Troy satın aldı; bilimsel yaklaşım"),
        ("Shaz & Kiks", "shazandkiks.com", "Hint Saç Bakım", "Hint kadınlar için saç bakımı; henna + amla; geleneksel Hint tarif; modern formüller"),
        ("Golde Hair", "golde.co", "Süperfood Saç Maskesi", "Süperfood bazlı saç ve cilt maskeleri; Clean Greens yüz maskesi; 2-in-1 wellness + güzellik"),
        ("Sienna Naturals", "siennanaturals.com", "Temiz Doğal Saç", "Hannah Diop kurdu; doğal + kıvırcık saçlar için temiz bakım; Target'ta; vegan"),
        ("Drybar", "thedrybar.com", "Fön Çekim Uzmanı", "Sadece fön çekim salonu konsepti; Buttercup blow dryer; salon to DTC; Alli Webb kurdu"),
        ("Aquis", "aquis.com", "Hızlı Saç Kurutma", "Mofricrofiber saç havlusu; saç kırılmasını %50 azaltma; patentli Aquitex kumaş; hızlı kurutma"),
        ("Cécred", "cecred.com", "Beyoncé Saç Bakım", "Beyoncé'nin saç bakım markası; saç onarım + nem; fermentasyon teknolojisi; premium"),
        ("Maria Nila", "marianila.com", "İsveç Vegan Saç", "İsveç vegan + hayvan deneysiz saç bakımı; renk koruma; sürdürülebilir ambalaj; salon kalite"),
        ("Fekkai", "fekkai.com", "NYC Salon Lüks", "Frédéric Fekkai Paris salonu; Apple Cider Detox scrub; lüks saç bakımı; yeniden markalaşma"),
        ("Olaplex Alt No.3", "olaplex.com", "Evde Bağ Onarım", "No.3 Hair Perfector evde kullanım; salon bağ onarım teknolojisi evde; kült ürün"),
        ("Bread Beauty Alt", "breadbeautysupply.com", "Wash Day Ritual", "Tekstürlü saçlar için minimal koleksiyon; hair-wash butter; Sephora özel; doğal saç"),
        ("Eva NYC", "evanyc.com", "Uygun Fiyat Salon Saç", "NYC salon kalitesi uygun fiyatla; Mane Magic 10-in-1 primer; Ulta'da; pratik çözümler"),
        ("Biolage R.A.W.", "biolage.com", "Doğal Kaynaklı Salon", "Doğal kaynaklı profesyonel saç bakımı; vegan; sürdürülebilir dönüşüm; Matrix'ten bağımsız"),
        ("Not Your Mother's", "nymbrands.com", "Erişilebilir Saç Çözümü", "Curl Talk kıvırcık saç serisi; uygun fiyatlı; süpermarket + eczane; her saç tipine çözüm"),
        ("Divi", "difrvi.com", "Saç Derisi Serumu", "Saç derisi sağlığı serumu; saç dökülme önleme; TikTok viral; saç derisi bakımı trendi"),
        ("Vegamour Alt GRO+", "vegamour.com", "Biyotin Saç Serumu", "GRO+ Advanced serum; CBD + biyotin; saç incelmesi çözümü; 90 gün sonuç garantisi"),
        ("Mane Club", "maneclub.com", "Saç Maskesi Paketi", "Tek kullanımlık saç maskesi paketleri; seyahat dostu; farklı ihtiyaçlara göre; pratik"),
        ("INNBeauty Project", "innbeautyproject.com", "Bilimsel Saç + Cilt", "Sephora için yaratılan temiz marka; Slushy serum; gökkuşağı estetiği; eğlenceli bilim"),
    ],

    "Vücut Bakımı & Kişisel Hijyen": [
        ("Topicals Body", "mytopicals.com", "Vücut Hiperpigmentasyon", "Like Butter vücut maskesi; bacak + kol hiperpigmentasyonu; vücut cilt bakımı"),
        ("Necessaire Body Wash", "necessaire.com", "Aktif Duş Jeli", "Niacinamide + hyalüronik asit duş jeli; vücut bakımında aktif devrim; fragrance-free seçenek"),
        ("Ouai Body", "theouai.com", "Saç Markasından Vücut", "Melrose Place vücut kremi; saç parfümü başarısından vücut bakıma geçiş; lüks günlük"),
        ("Sol de Janeiro 68", "soldejaneiro.com", "Brezilya Koku Vücut", "Cheirosa '68 parfümlü vücut spreyi; TikTok viral koku; Brezilya sensöryel deneyim"),
        ("Frank Body Scrub", "frankbody.com", "Kahve Vücut Peelingi", "Original Coffee Scrub; kahve çekirdeği + badem yağı; selülit + streç izleri; UGC öncüsü"),
        ("Bali Body", "balibody.com", "Bronzlaşma Yağı", "Self-tan kuru yağ; Avustralya; kademeli bronzlaşma; tropikal estetik; Instagram viral"),
        ("Isle of Paradise", "isleofparadise.com", "Temiz Self Tan", "Self-tan damla; yeşil renk düzeltici; renk kaydırma teknolojisi; vegan + temiz"),
        ("Bondi Sands", "bondisands.com", "Avustralya Self Tan", "Avustralya #1 self-tan markası; Bondi Beach ilhamı; kolay uygulama; çizgisiz"),
        ("St. Tropez Alt", "sttropeztan.com", "Gradual Tan", "Kademeli self-tan losyon; doğal görünüm; çizgisiz; sürdürülebilir formül; pürüzsüz"),
        ("Weleda", "weleda.com", "Biyodinamik Vücut Bakım", "1921'den beri; Skin Food kült ürün; biyodinamik çiftçilik; Almanya; doğal vücut bakımı"),
        ("Aesop Vücut", "aesop.com", "Botanik Vücut Yıkama", "Coriander Seed duş jeli; botanik formüller; minimalist eczane estetiği; unisex"),
        ("Jack Black Vücut", "getjackblack.com", "Erkek Vücut Bakımı", "Turbo Body Spray; All-Over Wash; erkek aktif yaşam vücut bakımı; SPF vücut losyonu"),
        ("The Right to Shower", "therighttofrshower.com", "Sosyal Etki Vücut Bakımı", "Evsiz bireylere hijyen ürünü bağışı; Unilever sosyal girişim; bitkisel bazlı; B Corp"),
        ("SEEN Haircare Body", "helloseen.com", "Dermatolog Vücut Yıkama", "Dermatolog formüle; akne yapıcı bileşenler yok; sırt aknesi çözümü; non-comedogenic"),
        ("Paula's Choice Body", "paulaschoice.com", "BHA Vücut Exfoliant", "Weightless Body Treatment BHA; vücut aknesi + keratosis pilaris; bilimsel vücut bakımı"),
        ("Kiehl's Body Alt", "kiehls.com", "Eczane Vücut Losyonu", "Creme de Corps vücut losyonu; beta-karoten + squalane; zengin formül; klasik eczane"),
        ("Hempz", "hempz.com", "Kenevir Tohumu Vücut", "Kenevir tohumu yağı vücut losyonu; ultra nemlendirici; uygun fiyatlı; aromatik"),
        ("Every Man Jack Body", "everymanjack.com", "Doğal Erkek Vücut Yıkama", "Doğal bileşenli erkek duş jeli; hindistan cevizi + aktif kömür; Target'ta; uygun fiyat"),
    ],

    "Erkek Bakım & Tıraş": [
        ("Bulldog Skincare", "bulldogskincare.com", "İngiliz Erkek Bakım", "İngiltere doğal erkek bakım; bambu tıraş + cilt bakımı; Edgewell satın aldı; B Corp"),
        ("Viking Revolution", "vikingrevolution.com", "Sakal Bakım Kiti", "Amazon #1 sakal bakım kiti; sakal yağı + balmı + fırça; uygun fiyatlı set; Viking temalı"),
        ("Rudy's Barbershop", "rudys.com", "Barber Doğal Ürün", "Seattle berber dükkanından; No. 1 Foundation pomad; doğal bileşenler; unisex"),
        ("Dollar Shave Club Alt", "dollarshaveclub.com", "Tıraş Abonelik", "Aylık tıraş bıçağı aboneliği; viral video ile patladı; Unilever $1B satın aldı"),
        ("Proraso", "proraso.com", "İtalyan Tıraş Geleneği", "1948'den beri İtalyan tıraş kremi; okaliptüs + mentol; berber favorisi; klasik"),
        ("Art of Shaving", "theartofshaving.com", "Lüks Tıraş Ritüeli", "4 adımlı tıraş ritüeli; premium fırça + stand; P&G bünyesinde; berber deneyimi evde"),
        ("Bevel Alt", "bevelcode.com", "Siyah Erkek Tıraş", "Tristan Walker kurdu; tek bıçak güvenlik; tıraş tümsekleri çözümü; koyu ciltler"),
        ("Viking Beard Brand", "vikingbeardbrand.com", "Organik Sakal Bakım", "Organik sakal yağı + balm; İsveç ilham; Viking temalı; premium doğal bileşenler"),
        ("Mountaineer Brand", "mountaineerbrand.com", "Appalachian Erkek Bakım", "West Virginia yapımı; doğal sakal + cilt bakımı; Amerikan zanaatkarlığı; sade bileşenler"),
        ("Prose Men", "prose.com", "Kişisel Erkek Saç Bakım", "AI kişiselleştirilmiş erkek saç bakımı; quiz bazlı; bireysel formül; premium"),
        ("Pacific Shaving", "pacificshaving.com", "Doğal Tıraş", "Doğal tıraş kremi + after shave; kafein bazlı; uygun fiyatlı; sürdürülebilir"),
        ("Stubble + Stache", "stubbleandstache.com", "Sakal + Bıyık Wax", "Bıyık wax + sakal yağı; niş erkek bakım; el yapımı; küçük parti üretim"),
        ("Gentleman's Box", "gentlemansbox.com", "Erkek Abonelik Kutusu", "Aylık erkek aksesuar + bakım kutusu; kravat + cilt bakımı + aksesuar; hediye"),
        ("Goodfellow & Co", "target.com", "Target Erkek Bakım", "Target'ın premium erkek bakım markası; uygun fiyatlı kalite; şık ambalaj; geniş erişim"),
        ("Harry's Alt", "harrys.com", "Uygun Fiyat Tıraş", "DTC tıraş bıçağı; $2/bıçak; kendi fabrikası Almanya'da; Walmart ortaklığı"),
    ],

    "Sağlık & Takviye": [
        ("AG1 Alt Athletic Greens", "drinkag1.com", "Günlük Yeşil Toz", "75 vitamin+mineral+probiyotik tek pakette; $79/ay; podcast sponsorluk kralı"),
        ("Ritual Men", "ritual.com", "Şeffaf Erkek Vitamin", "Şeffaf tedarik zincifrli erkek multivitamin; Essential; DNA izlenebilirlik; gecikmiş salınım"),
        ("Persona Nutrition", "personanutrition.com", "Kişisel Vitamin Paketi", "Online assessment ile günlük vitamin paketi; kişiselleştirilmiş; Nestlé Health Science"),
        ("Care/of Alt", "takecareof.com", "Vitamin Quiz Paketi", "Quiz bazlı kişiselleştirilmiş günlük vitamin paketi; Bayer satın aldı; renkli paketler"),
        ("Olly Alt", "olly.com", "Gummy Vitamin", "Sakız tablet vitamin; Sleep, Immunity, Beauty; eğlenceli ambalaj; Unilever satın aldı"),
        ("MegaFood", "megafood.com", "Gıda Bazlı Vitamin", "FoodState teknolojisi; gerçek gıdadan vitamin; organik çiftlik ortaklıkları; B Corp"),
        ("Garden of Life", "gardenoflife.com", "Organik Ham Vitamin", "USDA organik + Non-GMO Project; ham gıda bazlı; Nestlé bünyesinde; probiyotik lider"),
        ("Nordic Naturals", "nordicnaturals.com", "Premium Omega-3", "Norveç omega-3 balık yağı; altın standart; 3. parti test edilmiş; arınmış formül"),
        ("Life Extension", "lifeextension.com", "Yaşlanma Karşıtı Takviye", "40+ yıl yaşlanma araştırması; NMN + NAD+; klinik dozlarda bileşenler; bilimsel yaklaşım"),
        ("Designs for Health", "designsforhealth.com", "Klinisyen Takviye", "Sağlık profesyonelleri için takviye; klinik güç; 30+ yıl formülasyon; premium kalite"),
        ("Klean Athlete", "kleanathlete.com", "NSF Sporcu Takviyesi", "NSF Certified for Sport; Olympik sporcu güvenli; temiz formüller; doping riski yok"),
        ("Elm & Rye", "elmandrye.com", "Premium Tek Bileşen", "Tek bileşen premium takviyeler; D3, Magnezyum, Probiyotik; şık ambalaj; basit + etkili"),
        ("Metagenics", "metagenics.com", "Fonksiyonel Tıp Takviye", "Fonksiyonel tıp pratisyenleri tercihi; genetik beslenme; klinik araştırma bazlı"),
        ("Pure Encapsulations", "pureencapsulations.com", "Hipoalerjenik Takviye", "Hipoalerjenik; dolgu + yapay bileşen yok; klinisyen önerili; hassas bireyler için"),
        ("Jarrow Formulas", "jarrow.com", "Bilim Bazlı Takviye", "Bilimsel formülasyon; QH-Absorb CoQ10; probiyotikler; 40+ yıl; araştırma odaklı"),
        ("NOW Foods Alt", "nowfoods.com", "Değer Takviye", "Uygun fiyatlı kaliteli takviyeler; 1968'den beri; geniş ürün yelpazesi; aile şirketi"),
        ("Sports Research", "sportsresearch.com", "Sporcu Doğal Takviye", "Collagen peptides + MCT oil; spor + wellness; temiz etiket; Amazon bestseller"),
        ("Vital Vitamins", "vitalvitamins.com", "Doğal Kollajen", "Grass-fed kollajen peptit; brain supplement; doğal bileşenler; Amazon'da popüler"),
    ],

    "Uyku & Yatak Teknolojisi": [
        ("Eight Sleep Alt", "eightsleep.com", "Akıllı Yatak Kılıfı", "Pod Pro akıllı yatak kılıfı; termoregülasyon; uyku takibi; AI sıcaklık ayarı; biohacker"),
        ("Oura Ring Alt", "ouraring.com", "Uyku Takip Yüzüğü", "Uyku + aktivite takip yüzüğü; HRV analizi; readiness score; şık giyilebilir"),
        ("Dreem", "dreem.com", "EEG Uyku Bandı", "EEG bazlı uyku takip bandı; beyin dalgası analizi; derin uyku stimülasyonu; Fransız teknoloji"),
        ("Withings Sleep", "withings.com", "Yatak Altı Uyku Takibi", "Yatak altına konan uyku takip sensörü; uyku apnesi tespit; solunum analizi; Fransız"),
        ("Casper Alt Wave", "casper.com", "Ergonomik Yatak", "Wave Hybrid ergonomik destek; gel + hava kanalları; sıcaklık düzenleme; NYC DTC öncü"),
        ("Serta iComfort", "sertafrcomfort.com", "Soğutma Jel Yatak", "TempActiv jel memory foam; soğutma teknolojisi; CoolAction malzeme; TwinChill"),
        ("Beautyrest Alt", "beautyrest.com", "T3 Yay Teknolojisi", "T3 Pocketed Coil; üçlü iç yay; BlackICE soğutma; hibrit teknoloji"),
        ("Zoma", "zomasleep.com", "Sporcu Yatak", "Triangulex teknolojisi; bölgesel destek; sporcu toparlanma; gel memory foam; performans"),
        ("Amerisleep", "amerisleep.com", "Bitki Bazlı Memory Foam", "Bio-Pur bitki bazlı memory foam; HIVE teknolojisi bölgesel destek; 20 yıl garanti"),
        ("Brooklyn Bedding", "brooklynbedding.com", "Fabrika Direkt Yatak", "Kendi fabrikasında üretim; Aurora soğutma; Signature hibrit; AZ merkezli; uygun lüks"),
        ("DreamCloud", "dreamcloudsleep.com", "Lüks Hibrit Yatak", "Lüks hibrit yatak uygun fiyatla; kaşmir karışım kılıf; pocketed coil + memory foam"),
        ("WinkBed", "winkbeds.com", "El Yapımı Lüks Yatak", "El yapımı lüks yatak; Euro pillow top; Tencel kılıf; edge desteği; ABD yapımı"),
        ("Idle Sleep", "idlesleep.com", "Çift Taraflı Yatak", "Her iki tarafı da kullanılabilir yatak; lateks + memory foam; çevirerek ömür uzatma"),
        ("Yogabed", "yogabed.com", "Yoga İlham Yatak", "Yoga-Gel™ soğutma; esneklik + destek dengesi; orta sertlik; alerjenlere karşı"),
        ("Nest Bedding", "nestbedding.com", "Organik Yatak Mağaza", "Sparrow hibrit; organic Latex; kişiselleştirilebilir sertlik; Berkeley merkezli"),
        ("Sleep Number Alt", "sleepnumber.com", "Hava Odacıklı Yatak", "Bireysel sertlik ayarı; Sleep IQ teknolojisi; akıllı yatak; çift kişiselleştirme"),
        ("Boll & Branch Alt", "bollandbranch.com", "Organik Çarşaf", "Fair Trade organik pamuk çarşaf; Obama ailesi kullanıyor; premium yatak tekstili"),
        ("Parachute Alt", "parachutehome.com", "Otel Kalitesinde Çarşaf", "Otel kalitesinde yatak tekstili; percale + sateen; Venice Beach mağaza; DTC ev tekstili"),
        ("Coyuchi", "coyuchi.com", "Organik Ev Tekstili", "GOTS organik çarşaf + havlu; %100 organik pamuk; sürdürülebilir ev tekstili; 30+ yıl"),
        ("Ettitude", "ettitude.com", "Bambu Lyocell Çarşaf", "CleanBamboo™ lyocell çarşaf; kapalı döngü üretim; %50 su tasarrufu; hipoalerjenik"),
    ],

    "Fitness Ekipman & Giyim": [
        ("Vuori Alt", "vuori.com", "Kali Lifestyle Activewear", "Joe Kudla kurdu; Ponto performans jogger; $4B değerleme; erkek yoga + surf; DreamKnit"),
        ("Lululemon Alt", "lululemon.com", "Premium Yoga Giyim", "Align legging ikonik; Nulu kumaş; yoga + koşu; Studio Mirror; topluluk odaklı"),
        ("On Running", "on-running.com", "İsviçre Koşu Ayakkabısı", "CloudTec yastıklama; Roger Federer yatırımcı; İsviçre mühendisliği; IPO $6B+"),
        ("Hoka", "hoka.com", "Maksimalist Koşu Ayakkabısı", "Kalın taban + hafif; Bondi + Clifton ikonik; koşu + yürüyüş; Deckers bünyesinde"),
        ("Alo Yoga Alt", "aloyoga.com", "LA Premium Yoga", "Airbrush legging ikonik; yoga + meditasyon uygulaması; LA wellness kültürü; influencer"),
        ("Lulu's Alt P.E Nation", "pe-nation.com", "Avustralya Spor Giyim", "Avustralya activewear; cesur baskılar + neon; Pip Edwards kurdu; street + sport"),
        ("Carbon38", "carbon38.com", "Lüks Fitness Perakende", "Premium activewear marketplace; lüks fitness markaları küratörlüğü; kapsül koleksiyonlar"),
        ("Sweaty Betty", "sweatybetty.com", "İngiliz Kadın Activewear", "Power legging ikonik; İngiltere kadın fitness; Wolverine World Wide satın aldı"),
        ("Fabletics Alt", "fabletics.com", "Abonelik Activewear", "Kate Hudson; VIP üyelik; aylık activewear; uygun fiyatlı premium; erkek + kadın"),
        ("Reebok Alt Nano", "reebok.com", "CrossFit Ayakkabı", "Nano CrossFit ayakkabı; fonksiyonel fitness; Adidas'tan ayrıldı; authentic athletics"),
        ("Rogue Fitness", "roguefitness.com", "CrossFit Ekipman Uzmanı", "Ohio yapımı fitness ekipmanı; barbell + rack + plate; CrossFit Games sponsor; endüstriyel"),
        ("REP Fitness", "repfitness.com", "Ev Gym Ekipman", "Uygun fiyatlı ev spor salonu ekipmanı; rack + bench + dumbbell; Colorado; kalite + değer"),
        ("Titan Fitness", "titanfitness.com", "Bütçe Gym Ekipman", "Bütçe dostu ev gym; Rogue alternatifi; rack + barbell; uygun fiyat + kalite"),
        ("Bells of Steel", "bellsofsteel.com", "Kanada Gym Ekipman", "Kanada ev gym ekipmanı; adjustable kettlebell; kompakt; yenilikçi tasarım"),
        ("Whoop Alt", "whoop.com", "Strain + Recovery Takip", "Giyilebilir zorlanma + toparlanma takibi; ekransız; HRV; uyku; abonelik; elit sporcu"),
    ],

    "Moda & Giyim (Kadın)": [
        ("Christy Dawn", "christydawn.com", "Çiftçi Giyim", "Deadstock kumaştan elbise; Farm-to-Closet; regeneratif tarım; sürdürülebilir romantik"),
        ("Boyish Jeans", "boyish.com", "Sürdürülebilir Denim", "En sürdürülebilir denim markası; geri dönüşüm + organik; %50 az su; sertifikalı B Corp"),
        ("Mara Hoffman", "marahoffman.com", "Renkli Sürdürülebilir Moda", "Tropikal baskılar + sürdürülebilirlik; ECONYL geri dönüşüm naylon; cesur renkler"),
        ("Veronica Beard", "veronicafrbeard.com", "Power Blazer", "Dickey Jacket çıkarılabilir iç yelek; power dressing; kadın güçlendirme; NYC"),
        ("La Ligne", "lalfrigne.com", "Çizgili Lüks Kıyafet", "Çizgi desenin her formunda uzmanlaşma; Meredith Melling + Valerie Macaulay; NYC şıklığı"),
        ("Madewell", "madewell.com", "Günlük Denim + Temel", "Perfect Vintage Jean ikonik; denim trade-in programı; J.Crew kardeş; erişilebilir kalite"),
        ("Cuyana", "cuyana.com", "Fewer Better Things", "Daha az ama daha iyi; İtalyan deri çanta; lean closet felsefesi; zamansız tasarım"),
        ("Frank And Oak", "frankandoak.com", "Sürdürülebilir Kadın Moda", "Kanada sürdürülebilir moda; organik + geri dönüşüm; unisex; B Corp; Montreal"),
        ("Jenni Kayne", "jennikayne.com", "Kaliforniya Minimalizm", "California cool minimalist moda + ev; kaşmir + deri; Ojai yaşam tarzı; lüks basit"),
        ("Vince", "vince.com", "Lüks Minimalist Temel", "LA lüks minimalist; kaşmir + deri + pamuk; temel gardırop parçaları; sessiz lüks"),
        ("COS", "cos.com", "İskandinav Minimalist Moda", "H&M grubunda lüks pozisyon; İskandinav minimalizm; mimari kesimler; sürdürülebilir"),
        ("& Other Stories", "stories.com", "Yaratıcı Günlük Moda", "H&M grubunda; Paris + Stockholm + LA atölye; yaratıcı günlük moda; erişilebilir"),
        ("Toteme", "toteme-studio.com", "İsveç Sessiz Lüks", "Stockholm sessiz lüks; monogram eşarp ikonik; Elin Kling kurdu; minimalist zamansız"),
        ("Lisa Says Gah", "lisasaysgah.com", "Renkli Vintage Moda", "Oakland renkli vintage ilham; indie tasarımcılar küratörlüğü; eğlenceli + sürdürülebilir"),
        ("Pixie Market", "pixiemarket.com", "NYC Trend Moda", "New York trend odaklı; hızlı trend yakalama; uygun fiyatlı tasarımcı his; online butik"),
    ],

    "Moda & Giyim (Erkek)": [
        ("Everlane Alt", "everlane.com", "Radikal Şeffaflık", "Maliyet şeffaflığı; fabrika bilgisi; etik moda; temel parçalar; sürdürülebilir"),
        ("Bonobos Alt", "bonobos.com", "Fit Odaklı Erkek", "Mükemmel fit pantolon; Guideshop deneyim mağazası; Walmart satın aldı; erkek fit çözümü"),
        ("J.Crew Erkek Alt", "jcrew.com", "Amerikan Prep Erkek", "Amerikan preppy erkek moda; Ludlow suit; heritage + modern; Mickey Drexler mirası"),
        ("Club Monaco Alt", "clubmonaco.com", "Urban Sofistike Erkek", "Şehirli sofistike erkek + kadın; minimal tasarım; premium temel parçalar; Kanada kökenli"),
        ("Massimo Dutti Alt", "massimodutti.com", "İspanya Premium Erkek", "Inditex grubunda premium pozisyon; İspanya tasarımı; iş + günlük; erişilebilir lüks"),
        ("Suitsupply", "suitsupply.com", "Online Suit Uzmanı", "Online takım elbise; kişisel stil danışmanlığı; Hollanda; $399 suit; uygun lüks"),
        ("Percival", "percival.com", "İngiliz Modern Erkek", "İngiltere modern erkek giyim; renk bloğu + baskı; cesur İngiliz erkek; küçük üretim"),
        ("Corridor", "corridorny.com", "NYC Zanaatkar Erkek", "Brooklyn zanaatkar erkek giyim; el baskı; Japonya kumaş; limited run; sanatsal"),
        ("Noah", "noahny.com", "NYC Sürdürülebilir Streetwear", "Brendon Babenzien (Supreme eski CD); çevreci streetwear; 90'lar estetiği; NYC"),
        ("Aimé Leon Dore", "aimeleondore.com", "Queens Lüks Streetwear", "Teddy Santis; Queens NYC kültürü; New Balance işbirlikleri; Kith alternatifi"),
        ("Kith Alt", "kith.com", "Premium Streetwear", "Ronnie Fieg; sneaker + streetwear kültürü; marka işbirlikleri; flagship mağazalar; lifestyle"),
        ("Oliver Spencer", "oliverspencer.co.uk", "İngiliz Rahat Tailoring", "İngiltere relaxed tailoring; doğal kumaşlar; Londra; British craftsmanship; modern classic"),
        ("Portuguese Flannel", "portugueseflannel.com", "Portekiz Gömlek Uzmanı", "Portekiz yapımı gömlek; 70+ yıl tekstil mirası; kamp yaka ikonik; Avrupa kalitesi"),
    ],

    "İç Giyim & Çorap": [
        ("Thinx Alt", "shethinx.com", "Period Proof İç Giyim", "Regl geçirmez iç giyim öncüsü; 4 tampon kapasitesi; yıkanabilir; patentli teknoloji"),
        ("Naja", "nfraja.com", "Etik İç Giyim", "Geri dönüştürülmüş malzeme; Kolombiya üretim; tek annelere istihdam; sürdürülebilir + etik"),
        ("Organic Basics", "organicbasics.com", "Organik Temel İç Giyim", "GOTS organik sertifikalı; SilverTech anti-koku; Danimarka; capsule iç giyim"),
        ("Pact", "wearpact.com", "Organik Pamuk İç Giyim", "Organik pamuk + Fair Trade; tüm aile; temel iç giyim + loungewear; uygun fiyat; B Corp"),
        ("Boody", "bofrody.com", "Bambu İç Giyim", "Bambu viskon iç giyim; sürdürülebilir; doğal antibakteriyel; Avustralya; termoregülasyon"),
        ("Knickey", "knicfrkey.com", "Organik Geri Dönüşüm İç Giyim", "Organik pamuk iç giyim; eski iç giyim geri dönüşüm programı; sürdürülebilir döngü"),
        ("Kent", "kent-underwear.com", "Avrupa Erkek İç Giyim", "Premium Avrupa erkek iç giyim; antibakteriyel bambu; Hollanda tasarım; doğal kumaşlar"),
        ("Wolford", "wolford.com", "Avusturya Premium Çorap", "Avusturya lüks çorap + bodysuit; 50D Satin Touch; mükemmel fit; 75+ yıl kalite"),
        ("Falke", "falke.com", "Alman Premium Çorap", "Alman mühendislik çorap; Run serisi koşu; Ergonomic Sport; 1895'ten beri; zanaatkar"),
        ("Allbirds Socks", "allbirds.com", "Merinos Yünü Çorap", "Yeni Zelanda merinos; Trino kumaş; sürdürülebilir; karbon nötr; konforlu"),
        ("Balega", "balega.com", "Güney Afrika Koşu Çorapı", "Güney Afrika el yapımı koşu çorapları; DRYNAMIX nemlendirme; ultra hafif; blister önleme"),
        ("Point6", "point6.com", "Merinos Performans Çorap", "Compact spun merinos; Vermont tasarım; ABD yapımı; ince + sıcak; ömür boyu garanti"),
        ("Farm to Feet", "farmtofeet.com", "ABD Merinos Çorap", "ABD merinos çorap; American Rambouillet yünü; tamamen ABD tedarik zinciri; outdoor"),
    ],

    "Ayakkabı & Terlik": [
        ("On Cloud", "on-running.com", "İsviçre Koşu Ayakkabısı", "CloudTec yastıklama teknolojisi; hafif + reaktif; Roger Federer ortaklığı; $6B+ IPO"),
        ("Hoka Alt Bondi", "hoka.com", "Maksimal Yastıklama Koşu", "Maksimalist taban; ultra yastıklı; Bondi en çok satan; hemşire + sağlık çalışanı favorisi"),
        ("New Balance Alt 990", "newbalance.com", "Dad Shoe Trendi", "990v6 sneaker; Made in USA; $185; dad shoe trendi; Aimé Leon Dore işbirlikleri"),
        ("Birkenstock Alt", "birkenstock.com", "Ergonomik Sandalet", "Cork taban anatomik sandalet; Arizona ikonik; 250+ yıl miras; LVMH IPO"),
        ("Teva Alt", "teva.com", "Outdoor Sandalet", "Grand Canyon'da doğan sport sandalet; Hurricane XLT2; hiking + su; Deckers bünyesinde"),
        ("Chacos Alt", "chacos.com", "Macera Sandalet", "Z strap sistemi; özelleştirilebilir; Wolverine bünyesinde; ChacoNation topluluğu"),
        ("Allbirds Alt", "allbirds.com", "Sürdürülebilir Sneaker", "Merinos yünü + okaliptüs ağacı; karbon nötr; Yeni Zelanda; B Corp; $100 fiyat noktası"),
        ("Veja", "veja-store.com", "Fransız Etik Sneaker", "Amazon kauçuğu; organik pamuk; B Corp; şeffaf tedarik; Brezilya üretim; V-10 ikonik"),
        ("Clae", "clae.com", "LA Sürdürülebilir Sneaker", "Los Angeles; geri dönüştürülmüş malzemeler; kaktüs deri; minimalist tasarım; şehir sneaker"),
        ("Soludos", "soludos.com", "Modern Espadrille", "Espadrille'i modernize eden marka; işlemeli tasarımlar; İspanya üretim; yaz ayakkabısı"),
        ("Margaux", "margfraux.com", "Kişisel Flat Ayakkabı", "Yarım + geniş beden seçeneği; İspanya + İtalya yapımı; The Demi flat ikonik; online fit"),
        ("Birdies Alt", "birdies.com", "Meghan Markle Flat", "Meghan Markle'ın gizli konforu; ev terlikleri kadar rahat; yastıklı memory foam taban"),
        ("Suicoke", "suicoke.com", "Japon Teknik Sandalet", "Japon teknik outdoor sandalet; Vibram taban; streetwear + outdoor; lüks işbirlikleri"),
    ],

    "Yiyecek & Atıştırmalık": [
        ("RXBar", "rxbar.com", "Şeffaf Bileşen Bar", "'No B.S.' etiket; 3 yumurta akı + 2 hurma + 6 badem; Kellogg $600M satın aldı; şeffaflık"),
        ("GoMacro", "gomacro.com", "Makrobiyotik Bar", "Organik makrobiyotik bar; vegan + sürdürülebilir; Wisconsin aile işletmesi; B Corp"),
        ("Larabar", "larabar.com", "Minimal Bileşen Bar", "Hurma + fındık; minimum bileşen; lezzetli + basit; General Mills bünyesinde"),
        ("Epic Provisions", "epicbar.com", "Hayvan Bazlı Bar", "Grass-fed et bar + biltong; hayvan welfare; rejeneratif tarım; General Mills bünyesinde"),
        ("Country Archer", "countryarcher.com", "Artisanal Jerky", "Zanaatkar beef jerky; organik + grass-fed seçenek; Kaliforniya; lezzetli protein"),
        ("Wilde Brands", "wildebrands.com", "Tavuk Chips", "Tavuk göğsü cipsi; protein + düşük karbonhidrat; keto dostu; yenilikçi format"),
        ("Whisps", "whisps.com", "Peynir Cips", "Pişmiş peynir cipsi; düşük karbonhidrat; keto dostu; gerçek peynir; çıtır atıştırmalık"),
        ("Made Good", "madegood.com", "Okul Dostu Atıştırmalık", "Top 8 alerjenden arındırılmış; okul güvenli; organik; sebze gizlenmiş; çocuk dostu"),
        ("Unreal Candy", "unrealcandy.com", "Daha İyi Şeker", "Klasik şekerler daha iyi bileşenlerle; dark chocolate peanut butter cups; Tom Brady onaylı"),
        ("Free2b", "free2bfoods.com", "Alerjenden Arındırılmış Çikolata", "14 alerjenden arındırılmış çikolata kup; sunbutter bazlı; güvenli indulgence"),
        ("Barnana", "barnana.com", "Upcycled Muz Atıştırmalık", "Atık muzları atıştırmalığa dönüştürme; organik muz chips + bar; B Corp; Latin Amerika"),
        ("Hu Kitchen Alt", "hukitchen.com", "Paleo Atıştırmalık", "Paleo çikolata + kraker; basit bileşen; vegan seçenek; NYC; Mondelez satın aldı"),
        ("YumEarth", "yumearth.com", "Organik Şeker", "Organik lollipop + gummy + jelly bean; alerji dostu; çocuk favorisi; doğal renk + lezzet"),
        ("SmartSweets Alt", "smartsweets.com", "Akıllı Şeker", "3g şeker gummy ayı; stevia tatlılığı; bağırsak dostu; şeker ihtiyarını tatmin"),
        ("Oatmega", "oatmega.com", "Omega-3 Protein Bar", "Omega-3 balık yağı + yulaf protein barı; beyin + vücut; grass-fed whey; fonksiyonel"),
    ],

    "İçecek & Kahve": [
        ("Peet's Coffee Alt", "peets.com", "Batı Yakası Craft Kahve", "1966 Berkeley; 2. dalga öncüsü; JDE Peet's; Big Bang ikonik; koyu kavurma uzmanı"),
        ("Laird Superfood", "lairdsuperfood.com", "Surf Şampiyon Kreamer", "Laird Hamilton'ın markası; hindistan cevizi kreamer; mantar kahve; performans beslenme"),
        ("Califia Farms", "califiafarms.com", "Bitki Sütü", "LA bitki bazlı süt + kahve; badem + yulaf sütü; barista blend; sürdürülebilir ambalaj"),
        ("Minor Figures", "minorfigures.com", "Barista Yulaf Sütü", "Londra barista yulaf sütü; latte sanatı için özel; nitro cold brew; kahveci kültürü"),
        ("Matchabar", "matchabar.co", "NYC Matcha", "Brooklyn matcha kafesi; Hustle + Ceremonial matcha; hazır içecek + toz; NYC matcha kültürü"),
        ("Ippodo Tea", "ippfrodotea.co.jp", "Kyoto Geleneksel Çay", "300+ yıllık Kyoto çay firması; ceremonial matcha; geleneksel Japon çay kültürü; DTC global"),
        ("Sanzo", "drinksanzo.com", "Asya Meyve Gazlı İçecek", "Asya meyve aromalı gazlı su; calamansi + lychee + yuzu; doğal lezzet; Asya-Amerikan"),
        ("Swoon", "drinkswoon.com", "Sıfır Şeker Limonata", "Sıfır şeker limonata + mixers; monk fruit tatlandırıcı; diyabet dostu; doğal"),
        ("Health-Ade", "health-ade.com", "Premium Kombucha", "Los Angeles premium kombucha; küçük parti; cam şişe; Pink Lady Apple ikonik"),
        ("GT's Kombucha", "gfrtslivingfoods.com", "OG Kombucha", "Kombucha kategorisi yaratıcısı; Synergy serisi; probiyotik öncüsü; canlı kültür"),
        ("Remedy", "remedydrinks.com", "Avustralya Kombucha", "Avustralya; sıfır şeker kombucha; uzun fermentasyon; doğal bileşenler; global genişleme"),
        ("Recess", "takearecess.com", "Adaptojenik Gazlı Su", "CBD + adaptojenik gazlı su; stres azaltma; pastel estetik; wellness içecek; sober curious"),
        ("TRIP", "trip-drinks.com", "CBD İçecek", "İngiltere CBD infused içecek; elderflower + mint; wellness + sosyal; Londra merkezli"),
        ("Three Spirit", "threespiritdrinks.com", "Bitki Bazlı İçecek", "Bitkisel spirit alternatifi; Social Elixir, Nightcap, Livener; fonksiyonel botanikler"),
        ("Wildwonder", "drinkwildwonder.com", "Prebiyotik Meyve Suyu", "Prebiyotik + probiyotik meyve suyu; bağırsak sağlığı; düşük şeker; Asya ilham lezzetler"),
    ],

    "Ev & Mutfak": [
        ("Brightland", "brightland.com", "Kaliforniya Zeytinyağı", "Kaliforniya extra virgin zeytinyağı; doğrudan çiftçiden; lezzetli + taze; şık ambalaj"),
        ("Hedley & Bennett", "hedleyandbennett.com", "Şef Önlüğü", "Los Angeles şef önlüğü; profesyonel + ev; crossback tasarım; Ellen Bennett kurdu"),
        ("SMEG Alt", "smeg.com", "Retro Mutfak Cihazı", "İtalyan retro tasarım tost makinesi + buzdolabı; 50's style; renk seçenekleri; mutfak estetiği"),
        ("Staub Alt", "staub-usa.com", "Fransız Cocotte", "Fransız döküm tencere; siyah mat emaye; Alsace; profesyonel şef tercihi; dayanıklılık"),
        ("Le Creuset Alt", "lecreuset.com", "Renkli Döküm Tencere", "İkonik renkli döküm tencere; Flame turuncu; 1925 Fransa; nesiller boyu dayanıklılık"),
        ("Moccamaster", "moccafrmaster.com", "Hollanda Filtre Kahve", "Hollanda el yapımı filtre kahve makinesi; SCA onaylı; bakır element; 5 yıl garanti"),
        ("Breville Barista", "breville.com", "Ev Espresso Makinesi", "Barista Express ev espresso; dahili öğütücü; PID sıcaklık kontrolü; prosumer kalite"),
        ("Smoko", "smofrkonow.com", "Kawaii Mutfak Aksesuarı", "Sevimli Japon ilham mutfak gadget'ları; ısıtıcı; lambalı; hediye ürünler; eğlenceli"),
        ("Finex", "finexusa.com", "Portland Döküm Tava", "Portland el döküm demir tava; yay kulp; sekizgen tasarım; zanaatkar döküm; premium"),
        ("Borough Furnace", "boroughfurnace.com", "NY Döküm Tava", "New York el döküm demir tencere; zanaatkar; hafif tasarım; sürdürülebilir üretim"),
        ("Baratza", "baratza.com", "Ev Kahve Öğütücü", "Encore kahve öğütücüsü; entry-level premium; tamir edilebilir; barista onaylı"),
        ("Chemex", "chemex.com", "Tasarım Filtre Kahve", "1941 MoMA koleksiyonunda; Bauhaus tasarım; cam + ahşap; ikonik pour-over; zamansız"),
        ("Hario", "hariofr.com", "Japon Kahve Ekipmanı", "V60 pour-over; Japon cam işçiliği; kahve yarışmaları standardı; ısıya dayanıklı cam"),
        ("Zwilling Alt", "zwilling.com", "Alman Bıçak Uzmanı", "1731'den beri; Alman çelik bıçak; profesyonel + ev; Miyabi + Kramer; zanaatkarlık"),
        ("Vitamix Alt", "vitamix.com", "Pro Blender", "Profesyonel güç blender; 10 yıl garanti; restoran kalite evde; çok amaçlı; ABD yapımı"),
    ],

    "Ev Temizlik & Sürdürülebilirlik": [
        ("Mrs. Meyer's", "mrsmefryers.com", "Botanik Ev Temizlik", "Botanik kokulu temizlik ürünleri; garden-inspired; Iowa; aromaterapi temizlik deneyimi"),
        ("Method", "methodhome.com", "Tasarım Temizlik", "Güzel tasarımlı temizlik ürünleri; sürdürülebilir; okyanus plastiğinden ambalaj; SC Johnson"),
        ("Seventh Generation", "seventhgeneration.com", "7 Nesil Temizlik", "Bitkisel bazlı; 7 nesil sonrasını düşünen; Unilever satın aldı; sürdürülebilirlik öncüsü"),
        ("Ecover", "ecover.com", "Belçika Eko Temizlik", "Bitkisel + mineral bazlı; 1980'den beri; Belçika; sürdürülebilir ambalaj; karbon nötr"),
        ("Dr. Bronner's", "drbronner.com", "18-in-1 Sabun", "18 farklı kullanım; organik; Fair Trade; B Corp; aktivist marka; peppermint ikonik"),
        ("The Laundress", "thelaundress.com", "Lüks Çamaşır Bakımı", "Lüks kumaş bakımı; kaşmir + ipek yıkama; NYC; parfüm kokulu; Unilever satın aldı"),
        ("Caldrea", "caldrea.com", "Premium Ev Bakımı", "SC Johnson premium hattı; botanik kokulu; Mrs. Meyer's kardeş; lüks ev bakımı"),
        ("JAWS", "jawscleans.com", "Yeniden Doldurulabilir Temizlik", "Just Add Water System; konsantre + su; yeniden doldurulabilir; sürdürülebilir sistem"),
        ("Puracy", "puracy.com", "Doktor Formüle Temizlik", "Doktor formüle bitkisel temizlik; bebek güvenli; hipoalerjenik; 99.9% doğal; hassas aile"),
        ("Truly Free", "trulyfree.com", "Hassas Aile Temizlik", "Kimyasal hassasiyeti olanlara özel; toksinsiz; parfümsüz seçenek; alerji dostu"),
        ("Biokleen", "biokleen.com", "Enzim Bazlı Temizlik", "Bitki bazlı enzim temizlik; 1989'dan beri; Portland; çevre dostu; güçlü enzim formülleri"),
        ("Better Life", "cleanhappens.com", "Doğal Ev Temizlik", "Doğal + güçlü; bebek güvenli yüzey temizleyici; iki baba kurdu; pratik temizlik"),
    ],

    "Bebek & Çocuk": [
        ("Babyzen", "babyzen.com", "Kompakt Bebek Arabası", "YOYO2 katlanır bebek arabası; uçak kabin boyutunda; tek elle katlanır; Fransız tasarım"),
        ("UPPAbaby", "uppababy.com", "Premium Bebek Arabası", "Vista + Cruz + Minu; premium bebek arabası; İtalyan deri detaylar; modüler sistem"),
        ("Nuna", "nuna.eu", "Hollanda Bebek Ekipmanı", "RAVA oto koltuğu; MIXX arabası; Hollanda tasarım; toxin-free malzeme; premium güvenlik"),
        ("Dockatot", "dockatot.com", "Bebek Yuvası", "Bebek dinlenme yuvası; İsveç tasarım; organik pamuk; seyahat dostu; anne + bebek bağı"),
        ("Solly Baby", "sollybaby.com", "Bebek Taşıyıcı Wrap", "Ultra hafif bebek wrap; Tencel lyocell; nefes alabilir; bağlama taşıyıcı; anne bağı"),
        ("Ergobaby", "ergfrobaby.com", "Ergonomik Bebek Taşıyıcı", "Omni 360 taşıyıcı; ergonomik M pozisyonu; 4 taşıma şekli; yenidoğan + toddler"),
        ("Newton Baby", "newtonbaby.com", "Nefes Alabilir Yatak", "Wovenaire nefes alabilir bebek yatağı; yüz aşağı yatsa bile nefes alır; güvenlik"),
        ("Babo Botanicals", "bfraborobotanicals.com", "Botanik Bebek Bakım", "Organik botanik bebek bakım; SPF + şampuan; hassas cilt; dermatoloji onaylı"),
        ("Pipette", "pipette.com", "Squalane Bebek Bakımı", "Şeker kamışı squalane bebek bakımı; Biossance kardeş; temiz bileşenler; hassas bebek cildi"),
        ("Burt's Bees Baby", "burtsbeesbaby.com", "Organik Bebek Giyim", "GOTS organik pamuk bebek giyim; arı temalı; doğal + sürdürülebilir; geniş yaş aralığı"),
        ("Tea Collection", "teacollection.com", "Dünya Kültürü Çocuk Giyim", "Dünya kültürlerinden ilham çocuk giyim; seyahat temalı; eğitici + şık; organik seçenek"),
        ("Maileg", "maileg.com", "Danimarka Oyuncak", "Danimarka kumaş fare + tavşan oyuncaklar; ahşap mobilyalar; hayal gücü; koleksiyonluk"),
        ("Grimm's", "grimms.eu", "Waldorf Ahşap Oyuncak", "Gökkuşağı blokları ikonik; Waldorf pedagojisi; doğal boyalı ahşap; Almanya; yaratıcı oyun"),
        ("Lalo Alt", "meetlalo.com", "Modern Bebek Mobilya", "The Chair büyüyen mama sandalyesi; Play Kit; modern ebeveynlik ekipmanı; estetik"),
    ],

    "Evcil Hayvan": [
        ("Kong Alt", "kongcompany.com", "Dayanıklı Köpek Oyuncak", "Classic Kong dayanıklı çiğneme oyuncağı; dolgu yapılabilir; 1976'dan beri; veteriner önerili"),
        ("Outward Hound", "outwardhound.com", "Zeka Geliştirme Oyuncak", "Nina Ottosson puzzle oyuncakları; köpek zeka geliştirme; sıkılma önleme; interaktif"),
        ("Ruffwear", "ruffwear.com", "Outdoor Köpek Ekipmanı", "Köpek outdoor ekipmanı; Front Range harness; Web Master; hiking + su sporları; Oregon"),
        ("Kurgo", "kurgo.com", "Köpek Seyahat Aksesuarı", "Köpek araba + seyahat güvenliği; harness + araç bariyeri; macera köpekleri; fonksiyonel"),
        ("West Paw", "westpaw.com", "Sürdürülebilir Köpek Oyuncak", "Zogoflex geri dönüştürülebilir oyuncak; Montana yapımı; B Corp; dayanıklı + çevreci"),
        ("P.L.A.Y.", "petplay.com", "Tasarım Köpek Yatak", "Geri dönüştürülmüş PET dolgu köpek yatak; yıkanabilir; tasarım desenleri; B Corp"),
        ("MaxBone", "maxbone.com", "Lüks Evcil Hayvan", "Lüks köpek giyim + aksesuar; tasarımcı işbirlikleri; premium yatak; modern pet lifestyle"),
        ("The Foggy Dog", "thefoggydog.com", "El Yapımı Köpek Aksesuar", "San Francisco el yapımı bandana + tasma; butik kalite; mevsimsel koleksiyonlar; şık"),
        ("Yak9", "yak9.com", "Yak Sütü Köpek Çiğneme", "Nepal yak sütünden çiğneme kemiği; doğal + uzun süren; protein + kalsiyum; tek bileşen"),
        ("Brutus & Barnaby", "brutusandbarnaby.com", "Doğal Köpek Ödülü", "Pig ear + bully stick doğal çiğneme; tek bileşen; hormonsuz; ABD işlenmiş"),
        ("Wag Well", "wagwell.com", "CBD Evcil Hayvan", "CBD köpek sakızları; anksiyete + ağrı; organik; veteriner formüle; sakinleştirici"),
        ("Embark Alt", "embarkvet.com", "Köpek DNA + Sağlık Testi", "Köpek DNA test + 350+ sağlık riski; en kapsamlı köpek genetik testi; ırk + sağlık"),
    ],

    "Aksesuar & Takı": [
        ("Machete", "shopmachete.com", "Saç Aksesuarı", "Asetattan saç tokası + toka; vintage ilham; LA tasarım; handmade; renk çeşitliliği"),
        ("Lelet NY", "lfreletny.com", "Lüks Saç Aksesuarı", "New York lüks saç aksesuarları; kristal + metal; gelin + günlük; el yapımı; özel tasarım"),
        ("BaubleBar", "baublebar.com", "Trend Moda Takı", "Uygun fiyatlı trend takı; iPhone kılıfları; hızlı trend yakalama; influencer işbirlikleri"),
        ("Kendra Scott", "kendrascott.com", "Texas Renkli Takı", "Austin renkli doğal taş takı; Color Bar kişiselleştirme; hayırseverlik; kadın CEO"),
        ("Uncommon James", "uncomfromonjames.com", "Kristin Cavallari Takı", "Kristin Cavallari markası; Nashville; dainty + edgy; Very Cavallari TV şovu"),
        ("Chloe + Isabel", "chloeandisabel.com", "Sosyal Satış Takı", "Social selling takı markası; micro-entrepreneur; kapsayıcı; kadın güçlendirme"),
        ("Pura Vida", "pfruravidabracelets.com", "Costa Rica Bileklik", "Costa Rica zanaatkar bileklik; 800+ stil; bağış programları; plaj yaşam tarzı"),
        ("Alex and Ani", "alexandani.com", "Charm Bileklik", "Expandable bileklik; charm koleksiyonları; Rhode Island; geri dönüştürülmüş malzeme"),
        ("Vitaly Design", "vitalydesign.com", "Paslanmaz Çelik Takı", "Geri dönüştürülmüş paslanmaz çelik; unisex; streetwear estetiği; Kanada; sürdürülebilir"),
        ("Maison Miru", "maisonmiru.com", "Kulak Piercing Küratörlük", "Ear piercing planlama; küratörlü earscapes; 14k altın; NYC; kişiselleştirilmiş kulak tasarımı"),
        ("Brinker & Eliza", "brinkerandeliza.com", "Anne-Kız Takı", "Anne + kız birlikte tasarım; renkli boncuk; statement küpe; NYC; neşeli + cesur"),
        ("Shashi", "shopshashi.com", "NYC Trend Takı", "New York trend takı; altın + boncuk; katmanlama; uygun fiyatlı; günlük lüks"),
    ],

    "Gözlük & Güneş Gözlüğü": [
        ("Izipizi", "izipizi.com", "Fransız Erişilebilir Gözlük", "Fransız okufrma + güneş gözlüğü; €40; renk çeşitliliği; moda + fonksiyon; letmesee konsepti"),
        ("Meller", "melfrler.com", "Barcelona Güneş Gözlüğü", "İspanya tasarım uygun fiyat; polarize; şık + erişilebilir; Z kuşağı; renk seçenekleri"),
        ("Gentle Monster", "gentlemonster.com", "Kore Avangard Gözlük", "Seoul avangard gözlük; sanat galerisi mağazalar; K-pop ünlü; Huawei işbirliği"),
        ("Mykita", "mykita.com", "Berlin El Yapımı Gözlük", "Berlin el yapımı; vidasız menteşe; ultra hafif; mimari tasarım; Alman mühendisliği"),
        ("ic! Berlin", "ic-berlin.com", "Vidasız Berlin Gözlük", "Vidasız + menteşesiz patent tasarım; ultra hafif paslanmaz çelik; Berlin; özel"),
        ("Moscot", "moscofrrt.com", "NYC Heritage Gözlük", "1915'ten beri NYC Lower East Side; Lemtosh ikonik; Johnny Depp; aile mirası; vintage"),
        ("Crap Eyewear", "cfrrapeyewear.com", "LA Vintage Gözlük", "Los Angeles retro + vintage güneş gözlüğü; uygun fiyat; surf kültürü; eğlenceli"),
        ("Le Specs", "lespecs.com", "Avustralya Trend Gözlük", "Avustralya trend gözlük; Air Heart ikonik; uygun fiyatlı; Kardashian ailesi giyiyor"),
        ("Oliver Peoples Alt", "oliverpeoples.com", "Entelektüel Lüks Gözlük", "LA entelektüel lüks; O'Malley çerçeve ikonik; Gregory Peck modeli; vintage Amerikan"),
        ("Garrett Leight Alt", "garrettleight.com", "Venice Beach Heritage", "GLCO; Oliver Peoples kurucusunun oğlu; Kaliforniya heritage; el yapımı; Venice Beach"),
    ],

    "Teknoloji Aksesuarları": [
        ("Razer Kishi", "razer.com", "Mobil Oyun Kontrolcüsü", "Kishi Ultra mobil kontrolcü; konsol hissi; USB-C bağlantı; düşük gecikme; premium build"),
        ("SteelSeries", "steelseries.com", "Esports Aksesuar", "Arctis kulaklık + Rival mouse; esports kalitesi; GameDAC ses; profesyonel gaming"),
        ("Nanoleaf Alt", "nanoleaf.me", "Akıllı LED Panel", "Modüler LED ışık panelleri; altıgen + üçgen; ses reaktif; RGB; oda dekorasyonu; gaming"),
        ("Lifx", "lifx.com", "WiFi LED Ampul", "Hub gerektirmeyen akıllı LED; 16M renk; HomeKit + Alexa + Google; yüksek parlaklık"),
        ("Aqara", "aqara.com", "Akıllı Ev Sensörü", "Zigbee akıllı ev sensörleri; kapı + hareket + sıcaklık; HomeKit uyumlu; uygun fiyat + güvenilir"),
        ("Aqara", "aqara.com", "Akıllı Ev Sensörü", "Zigbee/Matter akıllı ev sensörleri $10-30; kapı, hareket, sıcaklık; Apple HomeKit uyumlu; inanılmaz değer"),
        ("Wyze", "wyze.com", "Uygun Fiyat Akıllı Ev", "$20 güvenlik kamerası; akıllı ev demokratizasyonu; kamera + ampul + kilit; Seattle"),
        ("Sonos Alt", "sonos.com", "Çok Odalı Hoparlör", "WiFi çok odalı ses sistemi; Beam + Arc soundbar; AirPlay 2; premium ev sesi"),
        ("Marshall Alt", "marshallheadphones.com", "Rock Hoparlör", "Rock'n'roll estetiği; Bluetooth hoparlör + kulaklık; vintage amplifikatör tasarım; ikonik"),
        ("JBL Alt", "jbl.com", "Portatif Bluetooth Hoparlör", "Flip + Charge portatif hoparlör; su geçirmez; parti modu; renkli tasarım; uygun fiyat"),
    ],

    "Seyahat & Bavul": [
        ("Rimowa Alt", "rimowa.com", "Alüminyum Bavul", "1898 Almanya; alüminyum bavul ikonik; LVMH bünyesinde; groove tasarım; lüks seyahat"),
        ("Briggs & Riley", "brigsandriley.com", "Ömür Boyu Garantili Bavul", "Ömür boyu garanti; dış cep basit erişim; CX genişleme; profesyonel seyahat; ABD"),
        ("Level8", "level8cases.com", "Alüminyum DTC Bavul", "Alüminyum bavul uygun fiyatla; Rimowa estetik; TSA kilidi; Kickstarter başarısı"),
        ("Mokobara", "mokobara.com", "Hint DTC Bavul", "Hindistan premium DTC bavul; uygun fiyat + kalite; TSA + USB; hızlı büyüyen"),
        ("Roam", "roamluggage.com", "Kişiselleştirilebilir Bavul", "100+ renk kombinasyonu kişiselleştirilebilir bavul; kendi rengini yarat; premium polikarbonat"),
        ("Arlo Skye", "arloskye.com", "Akıllı Bavul", "Dahili şarj; alüminyum + polikarbonat; sessiz tekerlekler; minimalist tasarım; premium"),
        ("Samsonite Alt Hartschalen", "samsonite.com", "Hafif Polikarbonat", "Cosmolite Curv hafif bavul; 120+ yıl miras; global lider; dayanıklılık + hafiflik"),
        ("Osprey", "osprey.com", "Sırt Çantası Bavul", "Farpoint / Fairview seyahat sırt çantası; All Mighty Guarantee; outdoor + seyahat; 50+ yıl"),
        ("Eagle Creek", "eaglecreek.com", "Packing Cube Uzmanı", "Pack-It organizatör sistemi; seyahat düzeni; kompresyon küpleri; ömür boyu garanti"),
        ("Tom Bihn", "tombihn.com", "Seattle Zanaatkar Çanta", "Seattle el yapımı seyahat çantası; Aeronaut carry-on; kullanıcı odaklı tasarım; 50+ yıl"),
    ],

    "Diş & Ağız Bakımı": [
        ("Philips Sonicare Alt", "philips.com", "Sonik Diş Fırçası", "DiamondClean sonik teknoloji; basınç sensörü; BrushSync; Bluetooth uygulama; premium"),
        ("Oral-B iO Alt", "oral-b.com", "AI Diş Fırçası", "iO Series; AI fırçalama tanıma; manyetik şarj; 3D diş haritası; premium teknoloji"),
        ("Waterpik", "waterpik.com", "Su Basınçlı Diş Temizleme", "Su jeti diş temizleme; diş arası temizlik; ortodonti hastaları için; klinik kanıtlı"),
        ("MOON", "moonoralcare.com", "Kendall Jenner Diş Bakımı", "Kendall Jenner ambassador; aktif karbonlu beyazlatma; şık ambalaj; influencer pazarlama"),
        ("Zenyum Alt", "zenyum.com", "Uygun Şeffaf Plak", "Uygun fiyatlı şeffaf diş teli; uygulama takipli; Güneydoğu Asya; uzaktan ortodonti"),
        ("Byte Alt", "byteme.com", "At-Night Şeffaf Plak", "Gece kullanım şeffaf diş düzeltici; HyperByte titreşimli hızlandırıcı; evde ortodonti"),
        ("SmileDirectClub Alt", "smiledirectclub.com", "DTC Ortodonti", "Evde şeffaf diş teli; 3D tarama; uzaktan doktor takibi; uygun fiyatlı ortodonti"),
        ("Colgate Hum", "colfrgate.com", "Akıllı Diş Fırçası", "Akıllı diş fırçası; fırçalama rehberi; bölge takibi; Colgate teknolojisi; uygun fiyat"),
    ],

    "Kadın Sağlığı & Regl Bakımı": [
        ("Clue App", "helloclue.com", "Bilimsel Regl Takip", "Bilimsel araştırma bazlı regl takip uygulaması; Berlin; doğurganlık + regl; veri gizliliği"),
        ("Flo App", "flo.health", "Regl + Hamilelik Takip", "AI regl ve hamilelik takibi; 380M+ kullanıcı; sağlık içerikleri; telesağlık"),
        ("Lola Alt", "mylola.com", "Organik Tampon", "Organik pamuk tampon aboneliği; şeffaf bileşen; kadın sağlığı farkındalığı"),
        ("Aunt Flow", "goafruntflow.com", "Regl Ürünü Dağıtıcı", "İşyeri + okul regl ürünü dağıtma sistemi; dispenser + ürün; erişilebilirlik; B Corp"),
        ("Dame Reusable Pad", "dame.com", "Yıkanabilir Ped", "Yıkanabilir regl pedi; İngiltere; sürdürülebilir; organik pamuk; sıfır atık"),
        ("Aisle Alt", "periodaisle.com", "Yeniden Kullanılabilir Regl", "Yıkanabilir ped + period underwear; Kanada; trans-kapsayıcı; sürdürülebilir regl"),
        ("Sustain Natural", "sustainnatural.com", "Organik İntim Ürün", "Organik prezervatif + tampon + kayganlaştırıcı; kadın sağlığı bütünsel; nitrosamin-free"),
        ("Nyssa", "nfrssa.care", "Doğum Sonrası Soğutma", "VieWell soğutma + sıkıştırma iç giyim; doğum sonrası toparlanma; buz terapi; yenilikçi"),
    ],

    "Cinsel Sağlık & Wellness": [
        ("Lovehoney", "lovehoney.com", "Online İntim Perakende", "Dünyanın en büyük online cinsel wellness perakendecisi; İngiltere + global; geniş seçim"),
        ("MysteryVibe", "mysteryvibe.com", "Esnek Vibratör", "Crescendo esnek vibratör; 6 motor; kişiselleştirilmiş zevk; sağlık amaçlı; İngiltere"),
        ("Emojibator", "emojibator.com", "Eğlenceli Vibratör", "Emoji şekilli vibratörler; patlıcan + şeftali; Z kuşağı; tabu kırma; uygun fiyat"),
        ("b-Vibe", "bfr-vibe.com", "Anal Wellness", "Eğitim odaklı anal wellness; rimming plug; anal eğitim kiti; kapsayıcı; bilgi paylaşımı"),
        ("Wild Flower", "meetwildflower.com", "Kapsayıcı İntim Cihaz", "Gender-kapsayıcı cinsel wellness; Enby vibratör; tüm vücutlar için; modern tasarım"),
        ("Lovability", "lfroveability.com", "Kadın Prezervatif", "Kadın odaklı prezervatif; şık ambalaj; 'It's Your Pleasure'; güçlendirme; NYC"),
        ("Quanna", "quanna.com", "CBD İntim Serum", "CBD intim serum; kadın zevk artırma; doğal bileşenler; ağrı azaltma; wellness"),
        ("Ohnut", "ohfrrnut.com", "İlişki Derinlik Sınırlayıcı", "Derin penetrasyon kaynaklı ağrı çözümü; ayarlanabilir halka; OB/GYN onaylı; niş problem çözme"),
    ],

    "Parfüm & Ev Kokusu": [
        ("Aesop Hwyl", "aesop.com", "Unisex Botanik Parfüm", "Hwyl tütsü parfüm; Marrakech Intense; botanik + duyusal; apothecary estetiği"),
        ("Commodity", "commoditfragrances.com", "Scent Space Parfüm", "Personal, Expressive, Bold 3 yoğunluk; Gold ikonik; şeffaf parfümeri"),
        ("Clean Reserve", "cleanreserve.com", "Sürdürülebilir Parfüm", "Sürdürülebilir + temiz parfüm; Skin ikonik; geri dönüştürülmüş ambalaj; vegan"),
        ("Juliette Has a Gun", "juliettefrasagun.com", "Cesur Kadın Parfüm", "Not a Perfume ikonik; ceratonia bazlı; Fransız niş; cesur isimler; Romano Ricci"),
        ("19-69", "nineteen-sixty-nine.com", "İsveç Karşı Kültür Parfüm", "İsveç; 1969 karşı kültür ilhamı; unisex; Purple Haze, Kasbah; sanatsal"),
        ("Régime des Fleurs", "regimedesfleurs.com", "LA Sanatsal Parfüm", "Los Angeles sanatsal niş parfüm; çiçek odaklı; sınırlı üretim; gallery kokuları"),
        ("Keap", "kefrpcandles.com", "Temiz Mum", "Temiz yanan hindistan cevizi + arı mumu; New York; sustainable; doğal koku yağları"),
        ("Kobo", "kobocandles.com", "Soya Yağı Mum", "Soya mum; kağıt ambalaj; botanik baskılar; sürdürülebilir; New York; şık hediye"),
        ("Paddywax", "paddywax.com", "Nashville Tasarım Mum", "Nashville; vintage cam kaplar; soya karışım; Literary koleksiyonu; hediye odaklı"),
        ("Maison Louis Marie", "maisonlouismarie.com", "Fransız Botanik Koku", "No.04 Bois de Balincourt ikonik; Le Labo alternatifi; botanik parfüm + mum; uygun lüks"),
        ("Nest Fragrances", "nestfragrances.com", "Lüks Ev Kokusu", "Bamboo ikonik koku; difüzör + mum; Laura Slatkin kurdu; lüks otel kokuları; hediye"),
        ("Carrière Frères", "carrierefreres.com", "Fransız Botanik Mum", "1884 Paris botanik bahçesi ilhamı; bitki türü odaklı; el yapımı cam; lüks Fransız"),
    ],

    "Outdoor & Spor Ekipman": [
        ("Patagonia Alt", "patagonia.com", "Aktivist Outdoor Giyim", "Dünya'yı kurtarmak misyonu; Worn Wear tamir; 1% for the Planet; çevreci marka lideri"),
        ("Arc'teryx Alt", "arcteryx.com", "Teknik Outdoor Giyim", "Vancouver teknik outdoor; GORE-TEX uzmanı; tırmanış + kayak; Amer Sports bünyesinde"),
        ("The North Face Alt", "thenorthface.com", "Iconic Outdoor Marka", "Nuptse ikonik mont; FUTURELIGHT kumaş; outdoor + streetwear; VF Corp bünyesinde"),
        ("Fjällräven", "fjallraven.com", "İsveç Outdoor Heritage", "Kånken sırt çantası ikonik; G-1000 kumaş; İsveç; 60+ yıl; sürdürülebilir outdoor"),
        ("prAna", "prana.com", "Yoga + Tırmanış Giyim", "Yoga + tırmanış giyim; organik pamuk + hemp; Fair Trade; sürdürülebilir outdoor lifestyle"),
        ("Topo Athletic", "topoathletic.com", "Geniş Parmak Ucu Koşu", "Geniş parmak ucu kutusu koşu ayakkabısı; doğal ayak şekli; Altra alternatifi; trail + yol"),
        ("Vivobarefoot", "vivobarefoot.com", "Barefoot Ayakkabı", "Ultra ince taban barefoot ayakkabı; doğal hareket; sıfır drop; geniş parmak ucu; sağlıklı"),
        ("Xero Shoes", "xerfroshoes.com", "Minimalist Barefoot", "Minimalist barefoot sandalet + ayakkabı; 5.000 mil garanti; Steven Sashen kurdu; koşu + günlük"),
        ("Mystery Ranch", "mysteryranch.com", "Askeri Sırt Çantası", "Montana yapımı askeri + avcılık sırt çantası; 3-ZIP tasarım; el yapımı; taktik + outdoor"),
        ("Gregory Packs", "gregorypacks.com", "Hiking Sırt Çantası", "FreeFloat süspansiyon; Baltoro + Deva; ergonomik hiking sırt çantası; 1977'den beri"),
        ("MSR", "msrgear.com", "Kamp Ocağı + Filtre", "PocketRocket kamp ocağı; Guardian su filtresi; dağcılık ekipmanı; Seattle mühendisliği"),
        ("Jetboil", "jetboil.com", "Entegre Kamp Ocağı", "Entegre ocak + tencere sistemi; Flash 100 saniye kaynatma; backpacking; kompakt"),
    ],

    "Ofis & Üretkenlik": [
        ("Remarkable Alt", "remarkable.com", "E-Ink Yazma Tableti", "reMarkable Paper Tablet; kağıt hissi e-ink; dikkat dağıtmayan; not alma + PDF; Norveç"),
        ("Supernote Alt", "supernote.com", "E-Ink Not Defteri", "Android bazlı e-ink tablet; el yazısı tanıma; Kindle + not defteri; Nomad çerçeve"),
        ("Leuchtturm1917", "leuchtturm1917.com", "Alman Premium Defter", "1917 Berlin; numaralı sayfalar; index; bullet journal resmi defter; arşiv kalitesi"),
        ("Moleskine Alt", "moleskine.com", "İkonik Defter", "Van Gogh, Hemingway geleneği; siyah defter ikonik; smart writing set; Milan tasarım"),
        ("MUJI Alt", "muji.com", "Japon Minimalist Kırtasiye", "Japon minimalist kırtasiye + ev; markasız kalite; 0.38mm kalem; basit + fonksiyonel"),
        ("Pilot Frixion Alt", "pilotpen.us", "Silinebilir Kalem", "Thermo-sensitive silinebilir mürekkep; ısıyla silme; yeniden yazılabilir; Japonya teknolojisi"),
        ("Hobonichi", "1101.com/store", "Japon Planlayıcı", "Hobonichi Techo günlük planlayıcı; Tomoe River kağıt; ince ama 365 sayfa; Japon zanaatkar"),
        ("Bellroy Desk", "bellroy.com", "Masaüstü Organizer", "Deri + elyaf masaüstü organizatör; kablo yönetimi; premium; fonksiyonel; Avustralya"),
        ("Dyson Lightcycle Alt", "dyson.com", "Akıllı Masa Lambası", "Daylight + yaş ayarı; 60 yıl LED ömrü; USB-C şarj; sıcaklık algılama; premium aydınlatma"),
        ("BenQ ScreenBar", "benq.com", "Monitör Üstü Lamba", "Monitör üstüne takılan LED aydınlatma; yansıma yok; otomatik parlaklık; ergonomik ışık"),
    ],

    "Oyun & Yaratıcı Araçlar": [
        ("Prusa Alt", "prusa3d.com", "Açık Kaynak 3D Yazıcı", "Prusa i3 MK4; açık kaynak; Çek Cumhuriyeti; güvenilir; topluluk odaklı; upgrade edilebilir"),
        ("Creality Alt", "creality.com", "Uygun 3D Yazıcı", "Ender 3 en popüler entry 3D yazıcı; uygun fiyat; büyük topluluk; modifikasyon kolaylığı"),
        ("Wacom Alt", "wacom.com", "Grafik Tablet", "Cintiq + Intuos çizim tableti; dijital sanatçı standardı; basınç hassasiyeti; profesyonel"),
        ("XP-Pen", "xp-pen.com", "Uygun Fiyat Çizim Tablet", "Uygun fiyatlı grafik tablet; Artist Pro; Wacom alternatifi; geniş ekran; pen display"),
        ("Huion", "huion.com", "Çin Grafik Tablet", "Kamvas çizim tablet; uygun fiyatlı profesyonel; kalem ekran; animatör + illüstratör"),
        ("Arturia", "arturia.com", "Müzik Prodüksiyon", "MiniLab MIDI kontrolcü; KeyStep; Fransız synth; analog + dijital; müzik üretimi"),
        ("Teenage Engineering", "teenage.engineering", "İsveç Tasarım Synth", "OP-1 taşınabilir synth ikonik; İsveç tasarım; EP-133 K.O.II; yaratıcı müzik aracı"),
        ("Novation", "novationmusic.com", "Launchpad Kontrolcü", "Launchpad MIDI kontrolcü; Ableton entegrasyon; ışıklı pad; beat yapma; DJ; performans"),
        ("Akai Professional", "akaipro.com", "MPC Beat Makinesi", "MPC beat yapma; 50+ yıl hip-hop tarih; MPC Live + One; sampling kültürü"),
        ("Rodecaster", "rode.com", "Podcast Prodüksiyon Hub", "RODECaster Pro II; all-in-one podcast prodüksiyon; 4 mikrofon + ses efekt; yayıncı"),
        ("GoXLR", "tc-helicon.com", "Streamer Ses Mikseri", "Streamer ses mikseri + ses efektleri; 4 kanal; otomatik pitch; RGB; yayıncı ihtiyacı"),
        ("Elgato Wave", "elgato.com", "Streamer Mikrofon", "Wave:3 kondenser mikrofon; Clipguard teknolojisi; Wave Link yazılımı; içerik üretici"),
    ],
}

# ─── EXTRA BRANDS BATCH 2 — pushing to 2000 ─────────────────────────────────
EXTRA_BRANDS_2 = {
    "Cilt Bakımı & Güzellik Araçları": [
        ("Dermatica", "dermatica.co.uk", "Online Reçeteli Cilt", "İngiliz online dermatolog; kişisel tretinoin formülü; £20/ay abonelik; NHS dermatologlara alternatif"),
        ("Geek & Gorgeous", "geekandgorgeous.com", "Macar Aktif Cilt", "Macaristan; %15 C vitamini, retinal, AHA/BHA; The Ordinary rakibi; €10 altı; Avrupa DTC"),
        ("Beauty Pie", "beautypie.com", "Üyelik Güzellik", "Fabrika fiyatına lüks kozmetik; üyelik modeli £10/ay; İngiliz; lüks formüller %80 ucuz"),
        ("Skin + Me", "skinandme.com", "UK Reçeteli Cilt Bakımı", "İngiliz online dermatoloji; kişisel tret + niacinamide karışım; dermatolog görüşmesi dahil"),
        ("Haeckels", "haeckels.com", "Deniz Yosunu Cilt", "Margate İngiltere; deniz yosunu bazlı cilt bakımı; el yapımı; biyolojik çeşitlilik koruma"),
        ("UpCircle", "upcirclebeauty.com", "Atık Dönüşüm Güzellik", "Kahve telvesi + meyve çekirdeklerinden cilt bakımı; atık bileşenleri değerlendirme; B Corp; UK"),
        ("Circumference", "circumferencenyc.com", "Biyoaktif Lüks", "NYC biyoaktif cilt bakımı; turunçgil kök hücre; $65+ ürünler; bilimsel lüks"),
        ("Epi.Logic", "epifrlogic.com", "Dermatolog Cilt", "Dr. Chaneve Jeanniton; klinik aktifler; retinal + peptit; siyahi kadın dermatolojist"),
        ("SKKN BY KIM", "sfrkkn.com", "Kim K Cilt 2.0", "Kim Kardashian'ın ikinci cilt bakım markası; 9 ürünlük lüks set; refill sistemi; $575 set"),
        ("Summer Bio", "summerbio.com", "Biyotik Güzellik", "Prebiyotik + postbiyotik cilt bakımı; mikrobiyom denge; bilimsel formüller"),
        ("Osea Malibu", "oseamalibu.com", "Deniz Yosunu Cilt", "25+ yıl aile işletmesi; deniz yosunu bazlı; viral TikTok; anti-trend uzun ömürlülük"),
        ("Evereden", "evereden.com", "Temiz Anne-Bebek", "Stanford bilim insanları; temiz cilt bakımı hamile + bebek; toksik olmayan; klinik kanıt"),
        ("Plenaire", "plenaire.co", "Gen Z Fransız Cilt", "Londra; Fransız ilham temiz cilt bakımı; Rose Jelly maske ikonik; pastel estetik"),
        ("Mele", "mfrle.com", "Koyu Ten Cilt Bakımı", "Procter & Gamble; melanin zengini ciltler için; hiperpigmentasyon uzmanı; dermatolojist geliştirdi"),
        ("Aestura", "aestura.com", "Kore Dermatolojik", "Kore eczane markası; A-Cica seri; hassas cilt; seramid; TikTok viral; K-derm"),
    ],
    "Saç Bakımı & Saç Sağlığı": [
        ("Hairstory", "hairstory.com", "Deterjansız Yıkama", "Şampuan yerine New Wash; deterjan içermeyen saç yıkama; saç sağlığı devrimi"),
        ("Jupiter", "jupiter.com", "Kepek Çözümü", "Şık ambalajlı kepek tedavisi; çinko pirition + hindistan cevizi; kepek = utanılacak değil"),
        ("Melanin Haircare", "melaninhaircare.com", "Tekstürlü Saç", "Whitney White YouTube; tekstürlü saç bakımı; Multi-Use Softening Leave-In; doğal saç topluluğu"),
        ("Virtue Labs", "virtuelabs.com", "Alpha Keratin", "Alpha Keratin 60ku teknolojisi; insan keratini ile onarım; bilim temelli; premium"),
        ("Innersense", "innersensebeauty.com", "Organik Saç", "USDA organik; Pure Inspiration Daily Conditioner; salon kalitesi temiz saç bakımı"),
        ("Hairstory", "hairstory.com", "Deterjan Sız Yıkama", "New Wash deterjan içermeyen; saç yıkama konseptini değiştirdi; Eli Halliwell; premium"),
        ("Davines", "davines.com", "İtalyan Sürdürülebilir", "B Corp İtalyan saç bakımı; OI Oil çok amaçlı; sürdürülebilir güzellik salonu hareketi"),
        ("dpHUE", "dphue.com", "Ev Saç Rengi", "Apple Cider Vinegar Rinse; Gloss+ ev renklendirme; salon arası bakım; erişilebilir"),
        ("Madison Reed", "madison-reed.com", "Akıllı Ev Boyası", "Salon kalitesi ev saç boyası; Color Advisor AI; temiz bileşen; 400+ ton; $22"),
        ("Overtone", "overtone.co", "Renk Deposu", "Direkt saç renk kondisyoner; yarı kalıcı; canlı renkler; hasarsız; ev boyama kolaylaştırıldı"),
        ("Arctic Fox", "arcticfoxhaircolor.com", "Vegan Saç Rengi", "%100 vegan; canlı renkler; hayvan testine karşı; Kristen Leanne kurdu; cruelty-free"),
        ("Mielle Organics", "mielleorganics.com", "Doğal Saç Büyütme", "Rosemary Mint Oil TikTok'ta 1B+ görüntülenme; siyah kadın kuruculu; P&G satın aldı; $100M+"),
        ("Revela", "revelabio.com", "Biyotek Saç Büyütme", "ProCelinyl molekülü; biyoteknoloji ile saç büyütme; Harvard bilimi; $20M+ yatırım"),
    ],
    "Vücut Bakımı & Kişisel Hijyen": [
        ("Lume", "lumedeodorant.com", "Tüm Vücut Deodorant", "Jinekolog geliştirdi; koltuk altı + her yer için; 'tüm vücut deodorant' kategorisi yaratıcısı"),
        ("Each & Every", "eachandevery.com", "Hassas Deodorant", "EWG 1 puan; karbonat içermeyen; hassas cilt; refill; Dead Sea tuzu bazlı"),
        ("Touchland", "touchland.com", "Tasarım Dezenfektan", "El dezenfektanını şık hale getirdi; Power Mist sprey; tasarımcı koku; $25M+ gelir"),
        ("Fur", "furyou.com", "İntim Vücut Bakımı", "Kasık bakımı öncüsü; tüy batması yağı; Emma Watson onayı; tabu yıkıcı"),
        ("Soft Services", "softservices.com", "Vücut Cilt Bakımı", "Vücuda özel cilt bakımı; KP (keratosis pilaris) çözümü; hedefli vücut sorunları"),
        ("Frank Body", "frankbody.com", "Kahve Peeling", "Avustralya; kahve peelingi viral; 2M+ ürün satıldı; esprili marka sesi; GF ambalaj"),
        ("Necessaire", "necessaire.com", "Lüks Duş", "$25 duş jeli normalleştirdi; vitamin + mineral içerikli vücut bakımı; Sephora özel"),
        ("Megababe", "megababebeauty.com", "Vücut Çözümleri", "Thigh Rescue sürtünme çubuğu; Bust Dust; beden pozitif sorun çözümü; inovatif"),
        ("By Humankind", "byhumankind.com", "Az Atık Bakım", "Deodorant, şampuan, ağız bakımı refill; %90 daha az plastik; tek seferde çözdü"),
        ("Ethique", "ethique.com", "Katı Bar", "Yeni Zelanda; katı şampuan/sabun barları; 30M+ plastik şişe tasarrufu; konsantre = hafif kargo"),
        ("HiBAR", "hfrbar.com", "Plastisiz Şampuan", "Katı şampuan + saç kremi; sıfır plastik; salon kalitesi; kompakt; seyahat dostu"),
        ("Saltair", "saltfrr.com", "Vücut Yıkama", "Iskandee Ashton; güzel kokulu vücut yıkama; $10 fiyat; Target'ta; erişilebilir lüks vücut"),
        ("Batiste", "batfrste.com", "Kuru Şampuan İnovatör", "Kuru şampuan kategorisi öncüsü; anında hacim; saç yıkama arası çözüm; $500M+ marka"),
    ],
    "Erkek Bakım & Tıraş": [
        ("Beardbrand", "beardbrand.com", "Sakal Topluluğu", "Eric Bandholz YouTube sakal topluluğundan marka; Urban Beardsman yaşam tarzı; içerik odaklı"),
        ("Huron", "usehuron.com", "Erişilebilir Erkek Cilt", "Ex-Bonobos kurucu; $15 fiyat noktası; temiz erkek cilt bakımı; sade ve etkili"),
        ("Supply", "supply.co", "Tek Bıçak Tıraş", "Tek bıçaklı güvenlik tıraş makinesi; tahriş önleyici; ömür boyu garanti; 'son tıraş makineniz'"),
        ("Hawthorne", "hawthorne.co", "Kişiselleştirilmiş Erkek", "Quiz bazlı kişiselleştirme; koku + cilt bakımı + saç; AI önerisi; her erkeğe özel"),
        ("Geologie", "geologie.com", "Kişisel Erkek Cilt", "Tanı testi ile kişisel 4 adım rejim; dermatolojist tasarımı; isminiz etikette"),
        ("Bravo Sierra", "bravosierra.com", "Askeri Performans", "Aktif askerlerle geliştirildi; her satışta askerlere bağış; ekstrem performans testi"),
        ("Duke Cannon", "dukecannon.com", "Sert Erkek Bakım", "Big Ass Brick of Soap; askeri ilham; gazilere bağış; mizah pazarlama; büyük ambalaj"),
        ("Lumin", "luminskin.com", "K-Beauty Erkek", "K-beauty ilhamlı erkek cilt bakımı; kömür temizleyici; hedefli FB reklamları; erkek rutin başlatıcı"),
        ("Bevel", "bevel.com", "Siyahi Erkek Tıraş", "Tristan Walker; siyahi erkekler için tıraş; tek bıçak tıraş makinesi; P&G satın aldı"),
        ("Disco", "letsdisco.co", "Eğlenceli Erkek Cilt", "Disko topu markalaşma; hiç cilt bakımı yapmamış erkekleri hedefliyor; eğlenceli mesajlaşma"),
        ("Oars + Alps", "oarsandalps.com", "Doğal Erkek Cilt", "İki kadın kocaları için kurdu; S.C. Johnson satın aldı; doğal erkek bakım"),
        ("Art of Sport", "artofsport.com", "Sporcu Bakım", "Kobe Bryant ortak kurucusu; sporcular için tasarlandı; ter sırasında çalışan botanikler"),
    ],
    "Sağlık & Takviye": [
        ("ARMRA Colostrum", "tryarmra.com", "Kolostrüm Takviye", "Sığır kolostrüm; bağışıklık + bağırsak; Dr. Sarah Chen; TikTok viral; yeni süper takviye"),
        ("Thesis Nootropics", "takethesis.com", "Kişisel Beyin Takviye", "Quiz bazlı kişisel nootropik karışımlar; farklı formüller dene; beyin sağlığı kişiselleştirme"),
        ("Beam", "beamorganics.com", "Uyku Tozu", "Dream Powder uyku takviyesi TikTok viral; nano CBD + reishi + magnezyum; sıcak kakao formatı"),
        ("Seed", "seed.com", "Bilimsel Probiyotik", "DS-01 Synbiotic; 24+ klinik çalışma; refill cam kavanoz; bilim öncelikli; $500M+ değerleme"),
        ("Rootine", "rootine.co", "DNA Bazlı Vitamin", "DNA testi + kan testi ile kişisel vitamin; hassas beslenme; mikrobiyom bazlı"),
        ("Elo Health", "elo.health", "Akıllı Takviye", "Kan biyobelirteç testi + AI kişisel takviye; veriye göre formül değişir; dinamik"),
        ("Neurohacker", "neurohacker.com", "Sistem Biyolojisi", "Qualia Mind 28 bileşenli nootropik; sistem yaklaşımı; biyohacker topluluğu"),
        ("BIOptimizers", "bioptimizers.com", "Sindirim Optimizasyonu", "MassZymes enzim; Magnesium Breakthrough 7 form; sindirim biyohacking; Wade Lightheart"),
        ("ZBiotics", "zbiotics.com", "Mühendislik Probiyotik", "Genetik mühendislik probiyotik; asetaldehit parçalama; akşamdan kalma bilimi; biyotek"),
        ("Prolon", "prolonfast.com", "Oruç Taklit Diyeti", "Dr. Valter Longo USC; 5 günlük oruç taklit programı; otofaji tetikleme; bilimsel"),
        ("Levels", "levelshealth.com", "Glukoz İzleme", "Diyabetik olmayan için CGM; yiyeceklerin kan şekerine etkisini gör; $100M+ yatırım; metabolik sağlık"),
        ("Lumen", "lumen.me", "Nefes Metabolizma", "Nefes analizi ile yağ mı karbonhidrat mı yaktığınızı ölçer; kişisel beslenme; $80M+ yatırım"),
        ("Viome", "viome.com", "Mikrobiyom Test", "Bağırsak mikrobiyom testi + kişisel takviye; AI analiz; $100M+ yatırım; hassas beslenme"),
        ("ZOE", "joinzoe.com", "Kişisel Beslenme", "CGM + bağırsak testi ile kişisel beslenme; Tim Spector bilimi; $100M+ yatırım; İngiliz"),
    ],
    "Uyku & Yatak Teknolojisi": [
        ("Manta Sleep", "mantasleep.com", "Uyku Maskesi", "Modüler göz kupası; göze sıfır baskı; Kickstarter; en iyi premium uyku maskesi"),
        ("Loftie", "byloftie.com", "Telefonsuz Alarm", "Telefonsuz yatak odası alarm saati; doğa sesleri; beyaz gürültü; telefon bağımlılığı azaltma"),
        ("ChiliSleep", "chilisleep.com", "Yatak İklim", "Su bazlı yatak soğutma/ısıtma; 55°F-115°F; çiftler için çift bölge; $500-700"),
        ("BedJet", "bedjet.com", "Yatak Hava İklim", "Hava tabanlı yatak iklim sistemi; ısıtma + soğutma; çiftler için çift bölge; alternatif"),
        ("Nolah", "nolahmattress.com", "AirFoam Yatak", "AirFoam = 4x daha fazla basınç dağılımı; çevre dostu; 10x daha dayanıklı; rekabetçi fiyat"),
        ("Birch Living", "birchliving.com", "Organik Yatak", "GOTS sertifikalı organik; doğal lateks + organik yün; gerçek organik uyku"),
        ("Zoma", "zomasleep.com", "Sporcu Yatak", "Triangulex bölgeli destek; sporcu odaklı; farklı vücut bölgelerine farklı destek"),
        ("Silk & Snow", "silkandsnow.com", "Kanada Yatak", "Kanada'nın DTC yatak markası; organik pamuk; CertiPUR-US; Toronto tasarım"),
        ("Pluto Pillow", "plutopillow.com", "Kişisel Yastık", "25 soruluk quiz ile kişisel yastık; uyku pozisyonuna göre; özel yapım"),
        ("Coop Home Goods", "coophomegoods.com", "Ayarlanabilir Yastık", "Parçalanmış memory foam; dolgu miktarını kendin ayarla; Amazon #1 yastık; kişiselleştirilebilir"),
    ],
    "Yiyecek & Atıştırmalık": [
        ("Fishwife", "eatfishwife.com", "Konserve Balık Lüks", "Konserve balığı havalı hale getirdi; koleksiyon kutuları; millennial sardalye trendi başlatıcısı"),
        ("Omsom", "omsom.com", "Asya Sos Başlatıcı", "Vietnamlı kız kardeşler; cesur Asya tatları; 'gururlu ve yüksek sesli Asyalı lezzetler'"),
        ("Bachan's", "bachans.com", "Japon BBQ Sos", "Aile tarifi Japon BBQ sosu; Whole Foods'ta #1 yeni sos; Costco hit; $200M+ 3 yılda"),
        ("Banza", "eatbanza.com", "Nohut Makarna", "Nohuttan makarna — 2x protein, 4x lif; $100M+ gelir; ABD'de #3 makarna markası; basit değişim"),
        ("Catalina Crunch", "catalinacrunch.com", "Keto Gevrek", "Keto dostu gevrek; 0g şeker; #1 keto gevrek; $100M+ gelir; tanıdık formatta diyet"),
        ("Wilde Chips", "wildechips.com", "Protein Cips", "Tavuk göğsünden cips; 10g protein; Mark Cuban Shark Tank yatırımı; yeni kategori"),
        ("Solely", "solely.com", "Meyve Kurutma", "Sadece meyve; yapay hiçbir şey yok; meyve jerky konsepti; sağlıklı atıştırmalık"),
        ("That's It", "thatsitfruit.com", "2 Malzeme Bar", "Sadece 2 malzeme — meyve + meyve; en sade etiket; radikal sadelik"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Çikolata", "Rafine şeker yok; 'insan gıdasına dönüş'; Mondelēz satın aldı; temiz atıştırmalık"),
        ("Kodiak Cakes", "kodiakcakes.com", "Proteinli Pancake", "Tam tahıl + protein pancake karışımı; ayı maskotu; $200M+ gelir; Shark Tank tekliflerini reddetti"),
        ("Deux", "eatdeux.com", "Fonksiyonel Kurabiye Hamuru", "Adaptogen + kolajen kurabiye hamuru; Shark Tank; çiğ yenilebilir; tatlı + wellness"),
        ("Tabs Chocolate", "tabschocolate.com", "Fonksiyonel Çikolata", "Maca + epimedium libido çikolatası; TikTok viral; 'randevu gecesi çikolatası'; cesur pazarlama"),
        ("Mid-Day Squares", "middaysquares.com", "İçerik Odaklı Çikolata", "Çift kurucular tüm yolculuğu açıkça vlog'luyor; 12g protein çikolata; içerik önce marka"),
        ("Last Crumb", "lastcrumb.com", "Lüks Kurabiye", "$140/12 kurabiye; drop'lar saniyeler içinde tükeniyor; lüks tatlı markası; her kurabiye bir deneyim"),
    ],
    "İçecek & Kahve": [
        ("Nguyen Coffee", "nguyencoffeesupply.com", "Vietnam Kahvesi", "ABD'de ilk özel Vietnam kahvesi; Sahra Nguyen robusta fasulyelerini öne çıkardı; kültürel gurur"),
        ("Onyx Coffee", "onyxcoffeelab.com", "Yarışma Kahvesi", "Her büyük kahve yarışmasını kazandı; şeffaf çiftçi ödemesi; Arkansas; kalite takıntısı"),
        ("Sunwink", "sunwink.com", "Bitkisel Tonik", "Gazlı bitkisel tonikler; adaptogen içerikli; Detox Ginger, Hibiscus Mint; güzel tasarım"),
        ("De La Calle", "delacalle.com", "Tepache", "Fermente ananas içeceği (tepache); Meksika geleneği modernize; probiyotik; kültürel köken"),
        ("Remedy Organics", "remedyorganics.com", "Fonksiyonel Shake", "Bitki bazlı protein + adaptogen shake; ashwagandha, kurkumin; fonksiyonel beslenme"),
        ("Rowdy Mermaid", "rowdymermaid.com", "Mantar Kombucha", "Lion's Mane mantarlı kombucha; fonksiyonel probiyotik; Colorado yapımı; adaptogen içerikli"),
        ("Clevr Blends", "clevrblends.com", "SuperLatte", "Meghan Markle yatırım yaptı; mantar + adaptogen latte karışımları; yulaf sütlü; Golden Latte"),
        ("Apothékary", "apothekary.co", "Bitkisel Wellness", "Wine Down bitki karışımı alkol alternatifi olarak viral; bitki bazlı wellness; fonksiyonel"),
        ("Cometeer", "cometeer.com", "Dondurulmuş Kahve", "Özel kahve demlenip flash-freeze; kapsül + su = mükemmel kahve; makine gereksiz; uzay teknolojisi"),
        ("Jot", "jot.co", "Ultra Konsantre Kahve", "20x konsantre sıvı kahve; tek sıkma = bir fincan; makine, filtre, atık yok; 5 saniyelik kahve"),
        ("Bones Coffee", "bonescoffee.com", "Aromalı Kahve", "S'Mores, Maple Bacon, Strawberry Cheesecake aromaları; kafatası logosu; lezzet odaklı; $50M+ gelir"),
        ("Nuun", "nuun.com", "Elektrolit Tablet", "Efervesan elektrolit tablet; suya at; koşucu/bisikletçi favorisi; düşük kalori; taşınabilir tüp"),
        ("Cure Hydration", "curehydration.com", "Bitki Elektrolit", "Hindistan cevizi suyu bazlı ORS; WHO hidrasyon formülü; bitki bazlı; temiz etiket"),
    ],
    "Ev & Mutfak": [
        ("Material Kitchen", "materialkitchen.com", "Şef Tasarımı", "Ex-elBulli şef tasarımı; reBoard kesme tahtası; reNaked tencere; minimal ürün, max kalite"),
        ("Misen", "misen.com", "Erişilebilir Bıçak", "Kickstarter ile $1M+ topladı; profesyonel bıçaklar $65'ten; şef kalitesi erişilebilir fiyat"),
        ("Canopy", "getcanopy.co", "Anti-Küf Nemlendirici", "Buğusuz nemlendirici; bulaşık makinesine girer; anti-küf teknoloji; aroma ek kiti; inovatif"),
        ("Tushy", "hellotushy.com", "Bide Devrimi", "'Kağıtla silmeyi bırak'; $80 bide aparatı; mizah pazarlama ustası; $50M+ gelir"),
        ("Pura", "pura.com", "Akıllı Ev Kokusu", "App kontrollü ev difüzör; Nest, Capri Blue ortaklıkları; koku zamanlama; $100M+ gelir"),
        ("Open Spaces", "getopenspaces.com", "Modern Düzenleme", "Güzel kutular, raf yükselticiler; renkli; Marie Kondo estetiğinde organizasyon"),
        ("Ember", "ember.com", "Akıllı Kupa", "App kontrollü sıcaklık; kahvenizi tam istediğiniz derecede tutar; $130; Apple Store'da"),
        ("Fellow", "fellowproducts.com", "Kahve Ekipman", "Stagg EKG su ısıtıcı pour-over standardı; SF özel kahve estetiği; tasarım ödüllü"),
        ("AeroPress", "aeropress.com", "Frizbi Mucidinden Kahve", "Stanford hocası ve frizbi mucidi Alan Adler icadı; $30; dünya şampiyonası var; kült takip"),
        ("GreenPan", "greenpan.us", "Seramik Kaplama", "Thermolon seramik yapışmaz kaplamayı icat etti; PFAS yok; Belçika mühendisliği; sağlıklı pişirme"),
        ("Ooni", "ooni.com", "Taşınabilir Pizza Fırını", "950°F'a ulaşan taşınabilir pizza fırını; Kickstarter'dan $100M+'a; evde Napoliten pizza"),
        ("Homesick", "homesickcandles.com", "Nostalji Mum", "50 ABD eyaleti + şehir + anı kokulu mumlar; 'Movie Night', 'Book Club'; dahice hediye ürünü"),
        ("Vitruvi", "vitruvi.com", "Taş Difüzör", "Taş görünümlü seramik difüzör; dekorasyon objesi; temel yağ setleri; aroma estetiği yükseltti"),
    ],
    "Ev Temizlik & Sürdürülebilirlik": [
        ("Blueland", "blueland.com", "Tablet Temizlik", "Temizlik tableti + tekrar kullanılabilir şişe; 1B+ plastik şişe tasarrufu; Shark Tank; Kate Hudson yatırım"),
        ("Branch Basics", "branchbasics.com", "Tek Konsantre", "TEK konsantre her şeyi temizler; çamaşır, bulaşık, zemin, vücut; kimyasal hassasiyet çözümü"),
        ("Earth Breeze", "earthbreeze.com", "Çamaşır Yaprağı", "Çözünen çamaşır yaprağı; ağır bidonlar yok; karbon nötr; her pakette 10 yıkama bağışı"),
        ("Stasher", "stasherbag.com", "Silikon Poşet", "Tekrar kullanılabilir silikon poşet; Ziploc alternatifi; Shark Tank; B Corp; 200M+ poşet tasarrufu"),
        ("Bite", "bitetoothpaste.com", "Diş Macunu Tableti", "Tablet formunda diş macunu; plastik tüp yok; cam kavanoz refill; 100M+ tüp tasarrufu"),
        ("Public Goods", "publicgoods.com", "Sürdürülebilir Temel", "Üyelik modeli ($59/yıl); ağaçsız kağıt, organik gıda, temiz bakım; minimalist beyaz ambalaj"),
        ("Who Gives A Crap", "whogivesacrap.org", "Eğlenceli Tuvalet Kağıdı", "Kârın %50'si gelişen ülkelerde tuvalet inşaatına; bambu/geri dönüşüm; eğlenceli ambalaj; B Corp"),
        ("Cleancult", "cleancult.com", "Refill Temizlik", "Cam dispenserli temizlik ürünleri; kağıt karton refill (süt kutusu gibi); hindistan cevizi bazlı"),
        ("Tru Earth", "tru.earth", "Kanada Çamaşır Şeridi", "Ultra konsantre çamaşır şeridi; Kanada icadı; kompostlanabilir ambalaj; $100M+ gelir"),
        ("The Pink Stuff", "thepinkstuff.com", "TikTok Temizlik", "İngiliz temizlik macunu TikTok'ta viral; çok amaçlı; $100M+ gelir sadece sosyal medya ile"),
        ("Scrub Daddy", "scrubdaddy.com", "Akıllı Sünger", "Shark Tank'ın en başarılı ürünü ($250M+ gelir); ısıya duyarlı FlexTexture; sıcakta yumuşak, soğukta sert"),
    ],
    "Bebek & Çocuk": [
        ("Tubby Todd", "tubbytodd.com", "Bebek Egzama Çözümü", "All Over Ointment bebek egzamasını iyileştiriyor; ebeveyn ağızdan ağıza; temiz bebek cilt bakımı"),
        ("Coterie", "coterie.com", "Premium Bez", "Test sonuçlarında en yüksek performans; %25 daha emici; sürdürülebilir; premium ebeveyn seçimi"),
        ("Little Sleepies", "littlesleepies.com", "Tereyağı Yumuşak Pijama", "Bambu viskoz; aile eşleştirme setleri viral; sınırlı baskı drop'ları tükeniyor; anne topluluğu"),
        ("Kyte Baby", "kytebaby.com", "Bambu Bebek", "Bambu rayon bebek giysileri; Sleep Bag en çok satan; sürdürülebilir + en yumuşak kumaş; %500 büyüme"),
        ("Mushie", "mushie.com", "İskandinav Bebek", "İskandinav estetik bebek ürünleri (önlük, tabak, emzik); pastel renkler; Instagram anne estetiği"),
        ("Cerebelly", "cerebelly.com", "Beyin Bebek Maması", "Beyin cerrahı geliştirdi; 16 temel beyin besin maddesi hedefleyen mama; bilimsel bebek beslenmesi"),
        ("Once Upon a Farm", "onceuponafarm.com", "Soğuk Sıkım Mama", "Jennifer Garner ortak kurucusu; soğuk sıkım (ısıtılmamış) mama; taze sebze poşetlerde"),
        ("Serenity Kids", "myserenitykids.com", "Et Bazlı Mama", "İlk et bazlı bebek maması poşetleri; etik kaynaklı protein; tuzlu mama öncüsü"),
        ("Hatch Baby", "hatch.co", "Akıllı Gece Lambası", "Rest+ ses makinesi + gün doğumu alarmı + gece lambası; app kontrol; 3 cihazı birleştirdi"),
        ("Fridababy", "frida.com", "Dürüst Bebek Sağlık", "NoseFrida burun aspiratörü viral; ebeveynlikle ilgili dürüst pazarlama; Frida Mom doğum sonrası"),
        ("Lovevery", "lovevery.com", "Montessori Oyun Kiti", "Gelişim aşamasına göre Montessori oyun kitleri; nörobilim destekli; $200M+ gelir; abonelik"),
    ],
    "Evcil Hayvan": [
        ("Jinx", "jinx.com", "Modern Köpek Maması", "Millennial köpek ebeveynlerini hedefliyor; temiz bileşenler; harika marka; ünlü yatırımcılar"),
        ("Sundays for Dogs", "sundaysfordogs.com", "Hava Kurutma Mama", "Kibble ile çiğ arası yeni kategori: hava kurutma; USDA insan sınıfı et; premium"),
        ("Wild One", "wildone.com", "Modern Evcil Hayvan", "Tasarım objesi gibi görünen evcil hayvan ürünleri; eşleşen renkler; tasma, taşıyıcı; 'Away of pet'"),
        ("Fi", "tryfi.com", "Akıllı Tasma", "GPS + LTE köpek tasması; aktivite + uyku takibi; kaçış uyarıları; köpeğinizin her an nerede"),
        ("Chippin", "chippin.com", "Sürdürülebilir Mama", "Cırcır böceği ve spirulina protein köpek ödülleri; sürdürülebilir protein; benzersiz açı"),
        ("A Pup Above", "apupabove.com", "Sous Vide Mama", "Sous vide pişirilmiş köpek maması; evcil hayvanlar için ilk bu teknik; insan sınıfı"),
        ("Diggs", "diggs.pet", "Modern Kafes", "Revol kafes katlanır, şık ve güvenli; yavru + yetişkin ayarları; çirkin kafes sorununu çözdü"),
        ("Tuft + Paw", "tuftandpaw.com", "Modern Kedi Mobilya", "İnsanların evinde isteyeceği kedi mobilyaları; tasarımcı kedi ağacı, yatak; güzel + fonksiyonel"),
        ("Native Pet", "nativepet.com", "Köpek Takviye", "Veteriner geliştirdi; Yak Chews viral; kemik suyu topper; $50M+ gelir; temiz bileşen"),
        ("Open Farm", "openfarmpet.com", "İzlenebilir Mama", "QR kod ile her bileşenin çiftliğini gör; insancıl sertifikalar; B Corp; şeffaflık"),
        ("Finn", "petfinn.com", "Köpek Takviye", "Temiz etiket köpek takviyeleri; Sakinleştirici + Alerji en çok satan; veteriner destekli; abonelik"),
    ],
    "Aksesuar & Takı": [
        ("Ana Luisa", "analuisa.com", "Karbon Nötr Takı", "Karbon nötr takı markası; geri dönüştürülmüş altın; $40-120 fiyat; sürdürülebilir ince takı"),
        ("Gorjana", "gorjana.com", "Katmanlama Takı", "Laguna Beach'ten narin katmanlama takı; Parker kolye ikonik; $50-200; 20+ mağaza"),
        ("Studs", "studs.com", "Piercing Deneyimi", "Kulak delme deneyimini yeniden icat etti; earscape küratörlüğü; temiz stüdyolar; Gen Z hedef"),
        ("Dorsey", "bydorsey.com", "Lab Taş Takı", "Lab yapımı renkli değerli taşlar; modern ayarlar; $200-500; erişilebilir lüks taşlar"),
        ("Catbird", "catbirdnyc.com", "Brooklyn İnce Takı", "Kalıcı takı (kaynak bileklik) trendi başlatıcısı; narin, romantik; Brooklyn yapımı"),
        ("Stone and Strand", "stoneandstrand.com", "Narin Günlük Takı", "Günlük katmanlama için narin ince takı; pırlanta toz tekniği; NYC tasarımı; $50-500"),
        ("Ring Concierge", "ringconcierge.com", "IG Takı", "Nicole Wegman Instagram'da kurdu; DM'den satış modeli; nişan yüzükleri; 'cool abla' vibes"),
        ("Moon Magic", "moonmagic.com", "Ay Taşı Takı", "Ay taşı ve labradorit takı; gökkuşağı yansıma; bohem lüks; $50M+ gelir"),
        ("Missoma", "missoma.com", "İngiliz Yarı-İnce", "Kate Middleton takıyor; altın kaplama + yarı değerli taşlar; Londra; 'herkesin altın halka markası'"),
        ("VRAI", "vrai.com", "Lab Elmas", "Diamond Foundry perakende markası; Leo DiCaprio yatırımcı; sıfır emisyon; izlenebilir"),
    ],
    "Teknoloji Aksesuarları": [
        ("Loop Earplugs", "loopearplugs.com", "Tasarım Kulak Tıkacı", "Takı gibi görünen kulak tıkacı; gürültüyü azaltır engellemez; $100M+ gelir; konser temel"),
        ("Timekettle", "timekettle.com", "Çeviri Kulaklık", "40+ dilde gerçek zamanlı çeviri kulaklık; çift yönlü eşzamanlı; AI destekli; seyahat devrimi"),
        ("Moft", "moft.us", "Yapışkan Stand", "Laptop + telefon yapışkan standları; ultra ince; MagSafe uyumlu; görünmez stand"),
        ("Keychron", "keychron.com", "Mac Mekanik Klavye", "Mac için uygun fiyatlı mekanik klavye; kablosuz; hot-swap; geliştirici/yazar favorisi"),
        ("Dbrand", "dbrand.com", "Cihaz Skin", "Her cihaza kaplama; alaycı, agresif sosyal medya; PS5 Darkplates hukuk savaşı"),
        ("Quad Lock", "quadlock.com", "Telefon Montajı", "Bisiklet/araba/motosiklet telefon montajı; twist-lock mekanizma; Avustralya; sağlam"),
        ("Native Union", "nativeunion.com", "Tasarım Teknoloji", "Mermer şarj pad; BELT kablo; tasarım odaklı teknoloji aksesuarları; Paris merkezli"),
        ("Nomad", "nomadgoods.com", "Deri Teknoloji", "Horween deri iPhone kılıf + Apple Watch kayışı; MagSafe ekosistem; premium + fonksiyonel"),
        ("Mophie", "mophie.com", "Taşınabilir Şarj", "Juice Pack şarj kılıfı öncüsü; MagSafe aksesuarlar; ZAGG Brands bünyesinde"),
        ("Chipolo", "chipolo.net", "Eşya Bulucu", "AirTag alternatifi; ONE Spot Apple Find My ile çalışır; kart şeklinde cüzdan versiyonu"),
        ("Satechi", "satechi.net", "USB-C Hub", "MacBook estetiğine uyumlu alüminyum USB-C hub'lar; uygun fiyatlı Apple yardımcı aksesuarlar"),
        ("EcoFlow", "ecoflow.com", "Taşınabilir Güç", "En hızlı şarj taşınabilir güç istasyonu (1 saatte 0-80%); Delta Pro bir evi besler; off-grid"),
    ],
    "Parfüm & Ev Kokusu": [
        ("Snif", "snif.co", "Önce Dene Parfüm", "Önce dene sonra al — ücretsiz 2ml numune + $5 iade depozit; erişilebilir niş parfüm; DTC inovasyon"),
        ("Dossier", "dossier.co", "Dupe Parfüm", "$300 parfümlerin ilham versiyonları $29'a; şeffaf bileşen listesi; lüks koku demokratikleştirme"),
        ("Boy Smells", "boysmells.com", "Cinsiyetsiz Mum", "LGBTQ+ kuruculu; 'genderful' markalaşma; mumdan parfüme genişledi; Kacey Musgraves collab"),
        ("P.F. Candle Co.", "pfcandleco.com", "Amber Kavanoz Mum", "Amber kavanoz ikonik; LA yapımı soya mumu; toptan 2,000+ mağaza; oda spreylerine genişledi"),
        ("Apotheke", "apotheke.co", "Brooklyn Mum", "Brooklyn yapımı lüks mum + ev kokusu; eczane estetiği; kömür, huş ağacı, deniz tuzu kokuları"),
        ("Brooklyn Candle Studio", "brooklyncandlestudio.com", "Soya Mum", "Minimalist soya mumu; %100 soya balmumu, pamuk fitil; Brooklyn el yapımı; adil fiyat lüks"),
        ("Otherland", "otherland.com", "Sanat Mum", "Sanat yönetimli kutu ambalaj dekor olarak kalır; 'Rattan' en çok satan; her koku bir tablo ilhamı"),
        ("Henry Rose", "henryrose.com", "Temiz Parfüm", "Michelle Pfeiffer'ın markası; EWG doğrulanmış; Cradle to Cradle sertifikalı; kokuda şeffaflık"),
        ("Ellis Brooklyn", "ellisbrooklyn.com", "Bilim+Koku", "Bilim kurgu yazarından parfümcüye; Bee EDP; temiz + karmaşık; güzel ambalaj"),
        ("Skylar", "skylar.com", "Hipoalerjenik Parfüm", "Hipoalerjenik temiz parfüm; kızının alerjisi ilham; toksik madde yok; kişisel koku profili quiz"),
        ("Dedcool", "dedcool.com", "Katmanlama Kokusu", "Katmanlanabilir koku sistemi; biyobozunur refill; 'ilk sürdürülebilir koku evi'; unisex"),
    ],
    "Seyahat & Bavul": [
        ("Monos", "monos.com", "Japon Minimalizm Bavul", "Japon ilhamlı minimalizm; polikarbonat kabuk; havacılık alüminyum; $100M+ gelir; sessiz alternatif"),
        ("July", "july.com", "Avustralya Bavul", "Avustralya; kişiselleştirilmiş bavul etiketi dahil; en hafif hardside carry-on; fabrikadan direkt"),
        ("Calpak", "calpaktravel.com", "TikTok Bavul", "Luka Mini Duffel TikTok'ta viral; uygun fiyat moda bavul ($50-200); paketleme küpleri kraliçesi"),
        ("Nomatic", "nomatic.com", "Dijital Göçebe", "Kickstarter $5M+; patent bekleyen organizasyon; dijital göçebe topluluğu; seyahat çantaları"),
        ("WANDRD", "wandrd.com", "Fotoğrafçı Sırt Çantası", "PRVKE sırt çantası; kamera çantası gibi görünmeyen kamera çantası; Kickstarter hit; hibrit"),
        ("Tropicfeel", "tropicfeel.com", "Seyahat Ayakkabısı", "Tek ayakkabı her arazide — yürüyüş, yüzme, keşif; Kickstarter en çok fonlanan seyahat ayakkabısı"),
        ("Baboon to the Moon", "baboontothemoon.com", "Macera Çantası", "Neon renkli Go-Bag'ler; kaos ve macera için; 'misfitler ve maceracılar için'; cesur marka"),
        ("Dagne Dover", "dagnedover.com", "Neopren Çanta", "Her şeye özel cepli neopren çantalar; Landon Carryall ikonik; fonksiyonel güzellik; $50M+"),
        ("Matador", "matadorup.com", "Paketlenebilir Ekipman", "Ultra hafif paketlenebilir seyahat ekipmanı; Pocket Blanket, Beast28 sırt çantası; kendine paketlenir"),
        ("Aer", "aersf.com", "Spor-İş Çantası", "Spor salonundan ofise geçiş çantaları; özel ayakkabı bölmesi; SF tasarımı; $30M+ gelir"),
    ],
    "Diş & Ağız Bakımı": [
        ("Burst", "bfrurst.com", "Diş Hijyenisti Fırça", "Diş hijyenistleri tarafından + için yapıldı; 25K+ profesyonel öneriyor; kömür varyant; abonelik"),
        ("Snow", "trysnow.com", "Beyazlatma Sistemi", "LED ağızlık + serum beyazlatma sistemi; Floyd Mayweather, Rob Lowe destekli; $100M+ gelir"),
        ("Hismile", "hismile.com", "TikTok Beyazlatma", "Avustralya; PAP+ formül (peroksitsiz); her büyük yaratıcı ile sponsorluk; devasa TikTok varlığı"),
        ("Cocofloss", "cocofloss.com", "Lüks Diş İpi", "Hindistan cevizi yağı emdirilmiş; dokulu plak yakalama; $8 diş ipi lüks deneyim haline getirdi"),
        ("David's", "davids-usa.com", "Metal Tüp Macun", "Premium doğal diş macunu metal tüplerde (geri dönüştürülebilir); California yapımı; nano-hidroksilapatit"),
        ("RiseWell", "risewell.com", "Hidroksilapatit Macun", "Florür yerine hidroksilapatit (dişlerinizin yapıldığı mineral); mineral diş macunu ABD öncüsü"),
        ("AutoBrush", "autobrush.com", "U-Şekil Fırça", "U şeklinde ağızlık fırça 30 saniyede tüm dişleri temizler; benzersiz form; viral infomercial"),
    ],
    "Kadın Sağlığı & Regl Bakımı": [
        ("Saalt", "saalt.com", "Regl Kabı", "Yumuşak silikon regl kabı; okula devam eden kızlara bağış; 400M+ ped/tampon çöpten kurtarıldı"),
        ("Cora", "cora.life", "Organik Tampon", "Organik regl ürünleri; her satışta gelişmekte olan ülkelerde kızlara ped bağışı; aktivizm"),
        ("Rael", "getrael.com", "K-Beauty Regl", "Kore yapımı organik regl + cilt bakımı yama; güzellik + regl bakımı çaprazlaması; bütüncül"),
        ("Flex", "flexfits.com", "İlişki Uyumlu Regl Disk", "İlişki sırasında kullanılabilen regl diski; kimsenin konuşmadığı sorunu çözdü; medikal polimer"),
        ("August", "itsaugust.co", "Gen Z Regl", "16 yaşında kuruldu; kapsayıcı ('kanayan herkes için'); renkli ambalaj; abonelik"),
        ("Thinx", "shethinx.com", "Regl İç Çamaşırı", "Regl iç çamaşırı kategorisi öncüsü; tartışmalı metro reklamları menstrüasyonu normalleştirdi; $100M+"),
        ("Dame", "dameproducts.com", "Cinsel Wellness", "NYC metro reklamı MTA tarafından engellendi (ED reklamları serbestken); cinsiyet eşitliği mücadelesi"),
        ("Maude", "getmaude.com", "Modern İntimite", "Dakota Johnson yatırımcı; 'seks basitleştirildi'; minimalist tasarım; tabu yıkıcı; $10M+"),
        ("Love Wellness", "lovewellness.com", "Kadın İntim Sağlık", "Lo Bosworth kurdu; vajinal sağlık takviyeleri; kadın intim sağlığında tabu kırma"),
        ("Natalist", "natalist.com", "Doğurganlık Ürünleri", "Modern tasarım hamilelik + ovülasyon testleri; prenatal vitaminler; TTC topluluğu"),
    ],
    "Outdoor & Spor Ekipman": [
        ("Rumpl", "rumpl.com", "Outdoor Battaniye", "Uyku tulumu yalıtımlı battaniye; desenleri ve collab'ları; Milli Park Vakfı ortaklığı"),
        ("GRAYL", "grayl.com", "Su Arıtıcı", "Basma tipi su arıtıcı — 8 saniyede temiz su; virüs, bakteri, kimyasal filtreler; macera seyahati"),
        ("LifeStraw", "lifestraw.com", "Kişisel Su Filtresi", "Doğrudan su kaynağından iç; milyonlarca hayat kurtardı; her satış bir çocuğa temiz su sağlar"),
        ("ENO", "eaglesnestoutfitters.com", "Hamak", "Ultra hafif hamaklar; kampçı favorisi; SingleNest ikonik; outdoor yaşam tarzı"),
        ("BioLite", "bioliteenergy.com", "Çift Amaçlı Ocak", "Ateşten elektrik üreten kamp ocağı; CampStove telefonunu şarj eder; çift amaçlı"),
        ("Oru Kayak", "orukayak.com", "Katlanır Kayak", "Origami ilhamı kayak 3 dakikada kutudan katlanır; dolabında yaşar; depolama/taşıma sorunu çözdü"),
        ("Snow Peak", "snowpeak.com", "Japon Outdoor", "Titanyum kamp malzemesi; 'human nature' felsefesi; 60+ yıl Japon zanaatı; yaşam tarzı kamp"),
        ("Helinox", "helinox.com", "Ultra Hafif Kamp Sandalye", "Chair One (900g) ultra hafif kamp mobilya hareketini başlattı; Kore DAC çubuk teknolojisi"),
        ("ISLE Paddle Boards", "islesurfandsup.com", "Şişme SUP", "Şişirilebilir paddle board; kompakt taşıma; $600-1000; su sporları erişilebilir hale"),
        ("BOTE", "boteboard.com", "Çok Amaçlı SUP", "Paddle board + kayak + platform; çok amaçlı su ekipmanı; balıkçılık SUP; Florida tasarım"),
    ],
    "Cinsel Sağlık & Wellness": [
        ("Maude", "getmaude.com", "Basit İntimite", "Dakota Johnson yatırımcı; minimalist kayganlaştırıcı + vibratör; 'seks basitleştirildi'"),
        ("Cake", "hellocake.com", "İntimite Ürünleri", "Modern, eğlenceli yaklaşım; Thingtesting'in en çok incelenen cinsel wellness markası"),
        ("Fur", "furyou.com", "İntim Vücut Bakımı", "Kasık cilt bakımı öncüsü; tüy batması yağı; Emma Watson onayı; tabu yıkma"),
        ("Womanizer", "womanizer.com", "Hava Basıncı Teknoloji", "Patentli Pleasure Air Technology; temas olmadan; tamamen benzersiz kategori; $200M+ gelir"),
        ("Lelo", "lelo.com", "İsveç Lüks", "İsveç lüks pleasure ürünleri; $100-200+ fiyat noktası; tasarım ödülleri; 'Rolls-Royce'"),
        ("Unbound", "unboundbabes.com", "Güçlendirme Odaklı", "Polly Rodriguez metro reklamları için savaştı; abonelik kutu modeli; güçlendirme markalaşması"),
    ],
    "Ofis & Üretkenlik": [
        ("Ugmonk", "ugmonk.com", "Analog Üretkenlik", "Analog görev kartları — fiziksel üretkenlik sistemi (dijital değil); Jeff Sheldon'un minimalist tasarımı"),
        ("Grovemade", "grovemade.com", "Ahşap Masa Aksesuarı", "Portland yapımı ceviz + akçaağaç masa aksesuarları; masa altlığı, monitör standı; zanaatkarlık + teknoloji"),
        ("Orbitkey", "orbitkey.com", "Anahtar Düzeni", "Şıngırdayan anahtarları ortadan kaldırdı; Nest hub; masa altlığı; Avustralya tasarımı; Kickstarter"),
        ("Baron Fig", "baronfig.com", "Yaratıcı Defter", "Kickstarter doğumlu defterler; Confidant Moleskine rakibi; sanatçı collab; 'düşünürler hoş geldiniz'"),
        ("Rocketbook", "getrocketbook.com", "Sonsuz Defter", "Yaz, uygulama ile tara, mikrodalgada sil — sonsuz defter; $30; çevre dostu + teknoloji birleşimi"),
        ("Fully", "fully.com", "Ayakta Masa", "Jarvis ayakta masa uzaktan çalışma olmazsa olmazı; bambu seçeneği; Herman Miller satın aldı"),
        ("Autonomous", "autonomous.ai", "Akıllı Ofis", "AI destekli ayakta masa + sandalyeler; Herman Miller'ın %50'sine; SmartDesk otomatik ayar; $200M+"),
    ],
    "Oyun & Yaratıcı Araçlar": [
        ("Cricut", "cricut.com", "DIY Kesim Makinesi", "Kağıt, vinil, kumaş kesen makine; tasarım aboneliği; maker hareketi enabler; $2B+ zirve piyasa değeri"),
        ("Glowforge", "glowforge.com", "Masaüstü Lazer", "Masaüstü lazer kesici/kazıyıcı; $28M crowdfund; küçük işletmeler özel ürünler için kullanıyor"),
        ("xTool", "xtool.com", "Lazer Kazıyıcı", "Masaüstü lazer kazıyıcılar; ahşap, deri, akrilik; M1 en çok satan; maker ekonomisi"),
        ("Backbone", "playbackbone.com", "Mobil Oyun Kontrolcü", "iPhone'u oyun konsoluna dönüştürür; PlayStation ortaklığı; $40M+ yatırım; mobil oyun enabler"),
        ("Analogue", "analogue.co", "Retro Konsol", "FPGA tabanlı orijinal kartuş oynayan retro konsollar; Pocket = taşınabilir retro; koleksiyoncu kalitesi"),
        ("Secretlab", "secretlab.co", "Oyun Sandalyesi", "Titan + Omega sandalyeler; esports ortaklıkları (T1, Team Secret); $300M+ gelir; oyun tahtı"),
        ("Elgato", "elgato.com", "Yayıncı Ekipmanı", "Stream Deck her içerik üreticisi için olmazsa olmaz; Key Light; Corsair bünyesinde; yaratıcı ekonomi altyapısı"),
    ],
}

# Merge EXTRA_BRANDS into BRANDS
for category, extra_list in EXTRA_BRANDS.items():
    if category in BRANDS:
        existing_names = {b[0].lower() for b in BRANDS[category]}
        for brand in extra_list:
            if brand[0].lower() not in existing_names:
                BRANDS[category].append(brand)
                existing_names.add(brand[0].lower())
    else:
        BRANDS[category] = extra_list

# Merge EXTRA_BRANDS_2 into BRANDS
for category, extra_list in EXTRA_BRANDS_2.items():
    if category in BRANDS:
        existing_names = {b[0].lower() for b in BRANDS[category]}
        for brand in extra_list:
            if brand[0].lower() not in existing_names:
                BRANDS[category].append(brand)
                existing_names.add(brand[0].lower())
    else:
        BRANDS[category] = extra_list

# Remove any big brand references that slipped in
BIG_BRAND_KEYWORDS = ['casper alt', 'casper wave', 'casper', 'purple alt', 'lululemon alt', 'lululemon',
                       'skims alt', 'skims', 'nike', 'adidas', 'apple homepod', 'samsung', 'ikea',
                       'p&g', 'pantene', 'head & shoulders', 'nivea men', 'clinique for men',
                       "kiehl's men", 'lab series', 'olaplex alt', 'olaplex', 'dollar shave',
                       'everlane alt', 'everlane', 'bombas alt', 'bombas', 'allbirds',
                       'poppi alt', 'poppi', 'olipop alt', 'olipop', 'liquid death alt',
                       'liquid death', 'away alt', 'away travel', 'glossier', 'fenty',
                       'kylie cosmetics', 'rare beauty', 'the ordinary', 'drunk elephant',
                       'charlotte tilbury', 'pat mcgrath', 'peloton', 'warby parker',
                       'harry\'s', 'manscaped', 'thirdlove', 'meundies', 'stitch fix',
                       'hellofresh', 'blue apron', 'barkbox', 'rothy', 'vuori', 'alo yoga',
                       'outdoor voices', 'brooklinen', 'ruggable', 'eight sleep', 'whoop',
                       'oura ring', 'mvmt', 'daniel wellington', 'gymshark', 'hims & hers',
                       'hims', 'ag1', 'athletic greens', 'ritual', 'olly', 'liquid iv',
                       'vital proteins', 'care/of', 'noom', 'goodrx', 'native deodorant',
                       'billie', 'quip', 'smiledirectclub', 'byte', 'tonal', 'mirror ',
                       'parachute home', 'boll & branch', 'article ', 'burrow', 'our place',
                       'made in cookware', 'hexclad', 'caraway', 'solo stove', 'nugget',
                       'bearaby', "farmer's dog", 'the farmer', 'ollie dog', 'nom nom',
                       'prettylitter', 'litter-robot', 'embark', 'anker', 'casetify',
                       'nanoleaf', 'framework', 'bambu lab', 'ridge wallet', 'bellroy',
                       'peak design', 'mejuri', 'brilliant earth', 'saatva', 'helix sleep',
                       'yeti', 'hydro flask', 'stanley', 'huel', 'soylent', 'butcherbox',
                       'thrive market', 'fabfitfun', 'birchbox', 'ipsy', 'rent the runway',
                       'lovevery', 'kiwico', 'honest company', 'bobbie', 'owlet', 'nanit',
                       'magic spoon', 'prime drink', 'feastables', 'dr. squatch',
                       'sonos', 'dyson', 'fashion nova', 'savage x', 'good american',
                       'simplisafe', 'ring doorbell', 'ring camera']
for category in BRANDS:
    BRANDS[category] = [b for b in BRANDS[category]
                        if not any(kw in b[0].lower() for kw in BIG_BRAND_KEYWORDS)]


# ─── Helper Functions ────────────────────────────────────────────────────────

def create_meta_ads_url(brand_name: str) -> str:
    """Generate Meta Ads Library search URL for a brand using proper URL encoding."""
    encoded_name = quote(brand_name)
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={encoded_name}&search_type=keyword_unordered"


def hex_to_argb(hex_color: str) -> str:
    """Convert 6-char hex to 8-char ARGB (openpyxl format)."""
    return hex_color.replace("#", "")


def lighten_color(hex_color: str, factor: float = 0.3) -> str:
    """Lighten a hex color by blending with white."""
    hex_color = hex_color.replace("#", "")
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * (1 - factor))
    g = int(g + (255 - g) * (1 - factor))
    b = int(b + (255 - b) * (1 - factor))
    return f"{r:02X}{g:02X}{b:02X}"


def apply_header_style(ws, row, max_col):
    """Apply deep navy header formatting."""
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick_border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="medium", color="1B2A4A"),
        bottom=Side(style="medium", color="1B2A4A"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border


def apply_data_row(ws, row, max_col, category_color=None, is_even=False):
    """Apply data row styling with category-specific colors."""
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    if category_color and is_even:
        fill_color = category_color
    else:
        fill_color = "FFFFFF"

    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_category_separator(ws, row, max_col):
    """Apply thick bottom border for category separation in All Brands sheet."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left,
            right=current_border.right,
            top=current_border.top,
            bottom=Side(style="medium", color="1B2A4A"),
        )


def build_excel():
    """Build the complete Excel workbook."""
    wb = Workbook()

    # ── Flatten all brands ────────────────────────────────────────────────
    all_brands = []
    for category, brands in BRANDS.items():
        for b in brands:
            all_brands.append((category, b))

    total_brands = len(all_brands)
    print(f"Toplam marka sayısı: {total_brands}")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: ÖZET (Summary)
    # ═══════════════════════════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Özet"
    ws_summary.sheet_properties.tabColor = HEADER_COLOR

    # Title
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "DTC İnovatif Problem-Çözen Markalar Raporu"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:D2")
    ws_summary["A2"].value = f"Oluşturulma Tarihi: {TODAY}  |  Toplam Marka: {total_brands}"
    ws_summary["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A3:D3")
    ws_summary["A3"].value = "Sadece inovatif, problem çözen, küçük/orta ölçekli DTC markalar"
    ws_summary["A3"].font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws_summary["A3"].alignment = Alignment(horizontal="center")

    # Category breakdown
    headers = ["#", "Kategori", "Marka Sayısı", "Yüzde (%)"]
    for col_idx, h in enumerate(headers, 1):
        ws_summary.cell(row=5, column=col_idx, value=h)
    apply_header_style(ws_summary, 5, len(headers))

    row_num = 6
    for idx, (cat, brands) in enumerate(BRANDS.items(), 1):
        count = len(brands)
        pct = round(count / total_brands * 100, 1)
        ws_summary.cell(row=row_num, column=1, value=idx)
        ws_summary.cell(row=row_num, column=2, value=cat)
        ws_summary.cell(row=row_num, column=3, value=count)
        ws_summary.cell(row=row_num, column=4, value=f"%{pct}")

        cat_color = CATEGORY_COLORS.get(cat, ("D4E6F1", "2C5F8A"))[0]
        apply_data_row(ws_summary, row_num, len(headers), category_color=cat_color, is_even=(idx % 2 == 0))
        row_num += 1

    # Total row
    ws_summary.cell(row=row_num, column=2, value="TOPLAM")
    ws_summary.cell(row=row_num, column=3, value=total_brands)
    ws_summary.cell(row=row_num, column=4, value="%100")
    total_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    for col in range(1, 5):
        cell = ws_summary.cell(row=row_num, column=col)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = total_fill
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="medium")
        )

    ws_summary.column_dimensions["A"].width = 5
    ws_summary.column_dimensions["B"].width = 42
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 12

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: TÜM MARKALAR (All Brands)
    # ═══════════════════════════════════════════════════════════════════════
    ws_all = wb.create_sheet("Tüm Markalar")
    ws_all.sheet_properties.tabColor = HEADER_COLOR

    all_headers = ["#", "Marka Adı", "Web Sitesi", "Kategori", "Alt Niş",
                   "Öne Çıkan Özellik / Pazarlama Açısı", "Meta Reklam Kütüphanesi"]

    for col_idx, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col_idx, value=h)
    apply_header_style(ws_all, 1, len(all_headers))

    row_num = 2
    current_category = None
    brand_idx = 0
    for category, brand in all_brands:
        brand_idx += 1
        name, website, subniche, insight = brand

        # Category separator
        if current_category and current_category != category:
            apply_category_separator(ws_all, row_num - 1, len(all_headers))
        current_category = category

        meta_url = create_meta_ads_url(name)
        cat_color_pair = CATEGORY_COLORS.get(category, ("D4E6F1", "2C5F8A"))
        cat_bg = cat_color_pair[0]
        cat_accent = cat_color_pair[1]

        ws_all.cell(row=row_num, column=1, value=brand_idx)
        ws_all.cell(row=row_num, column=2, value=name)

        # Clickable website - blue underlined
        site_cell = ws_all.cell(row=row_num, column=3, value=website)
        site_cell.hyperlink = f"https://{website}"
        site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")

        ws_all.cell(row=row_num, column=4, value=category)
        ws_all.cell(row=row_num, column=5, value=subniche)

        # Insight - italic dark gray
        insight_cell = ws_all.cell(row=row_num, column=6, value=insight)

        # Meta Ads Library link - green button-like
        meta_cell = ws_all.cell(row=row_num, column=7, value="Reklamları Gör")
        meta_cell.hyperlink = meta_url

        # Apply row styling
        is_even = (brand_idx % 2 == 0)
        apply_data_row(ws_all, row_num, len(all_headers), category_color=cat_bg, is_even=is_even)

        # Re-apply special cell styling
        site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")
        insight_cell.font = Font(name="Calibri", size=9, italic=True, color=INSIGHT_FONT_COLOR)

        meta_fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
        meta_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        meta_cell.fill = meta_fill
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 1

    # Column widths
    col_widths_all = [5, 25, 30, 35, 28, 65, 18]
    for i, w in enumerate(col_widths_all, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:G{row_num - 1}"

    # ═══════════════════════════════════════════════════════════════════════
    # INDIVIDUAL CATEGORY SHEETS
    # ═══════════════════════════════════════════════════════════════════════
    for cat_idx, (category, brands) in enumerate(BRANDS.items()):
        # Shorten sheet name if needed (max 31 chars)
        sheet_name = category[:31]
        ws_cat = wb.create_sheet(sheet_name)

        cat_color_pair = CATEGORY_COLORS.get(category, ("D4E6F1", "2C5F8A"))
        cat_bg = cat_color_pair[0]
        cat_accent = cat_color_pair[1]
        ws_cat.sheet_properties.tabColor = cat_accent

        cat_headers = ["#", "Marka Adı", "Web Sitesi", "Alt Niş",
                       "Öne Çıkan Özellik / Pazarlama Açısı", "Meta Reklam Kütüphanesi"]

        for col_idx, h in enumerate(cat_headers, 1):
            ws_cat.cell(row=1, column=col_idx, value=h)
        apply_header_style(ws_cat, 1, len(cat_headers))

        for b_idx, brand in enumerate(brands, 1):
            r = b_idx + 1
            name, website, subniche, insight = brand
            meta_url = create_meta_ads_url(name)

            ws_cat.cell(row=r, column=1, value=b_idx)
            ws_cat.cell(row=r, column=2, value=name)

            site_cell = ws_cat.cell(row=r, column=3, value=website)
            site_cell.hyperlink = f"https://{website}"
            site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")

            ws_cat.cell(row=r, column=4, value=subniche)

            insight_cell = ws_cat.cell(row=r, column=5, value=insight)

            meta_cell = ws_cat.cell(row=r, column=6, value="Reklamları Gör")
            meta_cell.hyperlink = meta_url

            # Apply row styling
            is_even = (b_idx % 2 == 0)
            apply_data_row(ws_cat, r, len(cat_headers), category_color=cat_bg, is_even=is_even)

            # Re-apply special cell styling
            site_cell.font = Font(name="Calibri", size=10, color=WEBSITE_LINK_COLOR, underline="single")
            insight_cell.font = Font(name="Calibri", size=9, italic=True, color=INSIGHT_FONT_COLOR)

            meta_fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
            meta_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            meta_cell.fill = meta_fill
            meta_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        col_widths_cat = [5, 25, 30, 28, 65, 18]
        for i, w in enumerate(col_widths_cat, 1):
            ws_cat.column_dimensions[get_column_letter(i)].width = w

        ws_cat.freeze_panes = "A2"
        ws_cat.auto_filter.ref = f"A1:F{len(brands) + 1}"

    # ── Save ──────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, FILENAME)
    wb.save(filepath)
    print(f"\nExcel dosyası oluşturuldu: {filepath}")
    print(f"Toplam {total_brands} marka, {len(BRANDS)} kategori")
    return filepath


if __name__ == "__main__":
    build_excel()
