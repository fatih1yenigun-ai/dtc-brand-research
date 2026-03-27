#!/usr/bin/env python3
"""
Marketing Angle Analyzer — Reusable tool for analyzing brand marketing angles.
Reads brand data, clusters by marketing angle, and outputs analysis to Excel.
Can be used on any brand dataset with (brand, website, sub_niche, insight) tuples.
"""

import os
import re
from datetime import datetime
from collections import Counter, defaultdict
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# ─── Marketing Angle Taxonomy ────────────────────────────────────────────────
# Each angle: (display_name_tr, keywords, demand_signal_description_tr)

MARKETING_ANGLES = {
    # ═══ MATERIAL & TECHNOLOGY ═══
    "bambu_tekstil": {
        "name_tr": "Bambu Tekstil",
        "name_en": "Bamboo Textile",
        "keywords": ["bambu", "bamboo", "bambou", "CleanBamboo"],
        "desc_tr": "Bambu fiber/viskon bazlı ürünler — soğutma, antibakteriyel, eko",
        "group": "Malzeme & Teknoloji",
    },
    "organik_sertifikali": {
        "name_tr": "Organik & Sertifikalı",
        "name_en": "Organic & Certified",
        "keywords": ["organik", "organic", "GOTS", "GOLS", "OEKO-TEX", "sertifika"],
        "desc_tr": "GOTS/OEKO-TEX sertifikalı organik ürünler",
        "group": "Malzeme & Teknoloji",
    },
    "keten_linen": {
        "name_tr": "Keten / Linen",
        "name_en": "Linen / Flax",
        "keywords": ["keten", "linen", "flax", "French flax", "Belgian flax"],
        "desc_tr": "Keten bazlı ürünler — doğal, nefes alabilir, vintage estetik",
        "group": "Malzeme & Teknoloji",
    },
    "ipek_silk": {
        "name_tr": "İpek / Silk",
        "name_en": "Silk / Mulberry",
        "keywords": ["ipek", "silk", "mulberry", "momme", "saten"],
        "desc_tr": "İpek/saten ürünler — saç & cilt bakımı, lüks his",
        "group": "Malzeme & Teknoloji",
    },
    "memory_foam": {
        "name_tr": "Memory Foam / Köpük",
        "name_en": "Memory Foam",
        "keywords": ["memory foam", "foam", "köpük", "viskoelastik"],
        "desc_tr": "Hafıza köpüğü bazlı uyku ürünleri",
        "group": "Malzeme & Teknoloji",
    },
    "merinos_yun": {
        "name_tr": "Merinos & Yün",
        "name_en": "Merino & Wool",
        "keywords": ["merinos", "merino", "yün", "wool", "yünü", "alpaka", "alpaca", "cashmere", "kaşmir"],
        "desc_tr": "Yün/merinos/alpaka — sıcaklık, doğal fiber, heritage",
        "group": "Malzeme & Teknoloji",
    },
    "tencel_okaliptus": {
        "name_tr": "TENCEL / Okaliptüs",
        "name_en": "TENCEL / Eucalyptus",
        "keywords": ["tencel", "lyocell", "okaliptüs", "eucalyptus"],
        "desc_tr": "TENCEL/okaliptüs lifi — eko + soğutma + yumuşaklık",
        "group": "Malzeme & Teknoloji",
    },

    # ═══ FUNCTIONAL BENEFIT ═══
    "sogutma_termal": {
        "name_tr": "Soğutma & Termal Düzenleme",
        "name_en": "Cooling & Thermal Regulation",
        "keywords": ["soğut", "cool", "serin", "sıcaklık düzenl", "termal", "temperature", "gece terlemesi", "night sweat"],
        "desc_tr": "Sıcaklık düzenleyen, soğutan ürünler — gece terlemesi çözümü",
        "group": "Fonksiyonel Fayda",
    },
    "agirlikli_anksiyete": {
        "name_tr": "Ağırlıklı / Anksiyete Azaltma",
        "name_en": "Weighted / Anxiety Relief",
        "keywords": ["ağırlıklı", "weighted", "ağırlık", "cam boncuk", "anksiyete", "anxiety", "terapi"],
        "desc_tr": "Ağırlıklı battaniye/yorgan — stres ve anksiyete azaltma",
        "group": "Fonksiyonel Fayda",
    },
    "hipoalerjen_temiz": {
        "name_tr": "Hipoalerjenik & Temiz",
        "name_en": "Hypoallergenic & Clean",
        "keywords": ["hipoalerjenik", "hypoallerjenik", "allerj", "allergen", "antibakteriyel", "antimikrobiyal"],
        "desc_tr": "Allerjen-free, antibakteriyel, hijyenik ürünler",
        "group": "Fonksiyonel Fayda",
    },
    "ortopedik_saglik": {
        "name_tr": "Ortopedik & Sağlık",
        "name_en": "Orthopedic & Health",
        "keywords": ["ortopedik", "ergonomik", "servikal", "boyun", "omurga", "postür", "ağrı", "kiropraktör", "fizyoterapi"],
        "desc_tr": "Boyun/sırt ağrısı, ergonomik destek, sağlık odaklı",
        "group": "Fonksiyonel Fayda",
    },
    "uyku_optimizasyonu": {
        "name_tr": "Uyku Optimizasyonu",
        "name_en": "Sleep Optimization",
        "keywords": ["uyku", "sleep", "horlama", "anti-horlama", "circadian", "melatonin", "deep sleep"],
        "desc_tr": "Uyku kalitesini artıran ürünler — biohacking, sleep tech",
        "group": "Fonksiyonel Fayda",
    },
    "alan_tasarrufu": {
        "name_tr": "Alan Tasarrufu & Organizasyon",
        "name_en": "Space Saving & Organization",
        "keywords": ["alan tasarrufu", "space sav", "kompakt", "compact", "katlanır", "katlanabil", "stackable", "modüler", "modular"],
        "desc_tr": "Küçük alan çözümleri, katlanır/modüler ürünler",
        "group": "Fonksiyonel Fayda",
    },
    "cok_fonksiyonlu": {
        "name_tr": "Çok Fonksiyonlu / Multi-Use",
        "name_en": "Multi-Functional",
        "keywords": ["çok kullanım", "çok fonksiyon", "multi-", "dual-use", "3-in-1", "4-in-1", "8-in-1", "çift taraf"],
        "desc_tr": "Birden fazla fonksiyonu olan ürünler",
        "group": "Fonksiyonel Fayda",
    },
    "yikanabilir_kolay": {
        "name_tr": "Yıkanabilir & Kolay Bakım",
        "name_en": "Washable & Easy Care",
        "keywords": ["yıkanabilir", "washable", "makinede", "wrinkle-free", "kolay bakım", "easy care"],
        "desc_tr": "Makinede yıkanabilir, kolay bakım ürünleri",
        "group": "Fonksiyonel Fayda",
    },

    # ═══ SUSTAINABILITY & VALUES ═══
    "surdurulebilir_eko": {
        "name_tr": "Sürdürülebilir & Eko",
        "name_en": "Sustainable & Eco",
        "keywords": ["sürdürülebilir", "geri dönüşüm", "recycl", "eko", "eco-", "çevreci", "zero waste", "sıfır atık", "karbon", "carbon"],
        "desc_tr": "Geri dönüşüm malzeme, karbon-nötr, zero waste",
        "group": "Sürdürülebilirlik & Değerler",
    },
    "fair_trade_etik": {
        "name_tr": "Fair Trade & Etik Üretim",
        "name_en": "Fair Trade & Ethical",
        "keywords": ["fair trade", "adil ticaret", "etik", "ethical", "zanaatkar", "artisan", "el yapımı", "el dokuması", "handmade"],
        "desc_tr": "Adil ticaret, zanaatkar desteği, etik üretim zinciri",
        "group": "Sürdürülebilirlik & Değerler",
    },
    "sosyal_misyon": {
        "name_tr": "Sosyal Misyon / Bağış",
        "name_en": "Social Mission / Give-Back",
        "keywords": ["bağış", "donate", "one-for-one", "sosyal etki", "social impact", "misyon", "kadın güçlendirme"],
        "desc_tr": "Her satışta bağış, sosyal etki, topluluk desteği",
        "group": "Sürdürülebilirlik & Değerler",
    },

    # ═══ MARKETING & GROWTH MODEL ═══
    "tiktok_viral": {
        "name_tr": "TikTok / Viral Sosyal Medya",
        "name_en": "TikTok / Viral Social",
        "keywords": ["tiktok", "viral", "TikTok"],
        "desc_tr": "TikTok veya sosyal medyada viral olan ürünler",
        "group": "Pazarlama & Büyüme",
    },
    "instagram_pinterest": {
        "name_tr": "Instagram / Pinterest Estetik",
        "name_en": "Instagram / Pinterest Aesthetic",
        "keywords": ["instagram", "pinterest", "estetik", "aesthetic", "flat-lay", "influencer"],
        "desc_tr": "Görsel estetik odaklı, Instagram/Pinterest pazarlama",
        "group": "Pazarlama & Büyüme",
    },
    "celebrity_kollab": {
        "name_tr": "Celebrity / Influencer Kollab",
        "name_en": "Celebrity / Influencer Collab",
        "keywords": ["celebrity", "ünlü", "kollab", "collab", "Kardashian", "Oprah", "Shark Tank", "kickstarter"],
        "desc_tr": "Ünlü onayı, Shark Tank, Kickstarter, influencer iş birliği",
        "group": "Pazarlama & Büyüme",
    },
    "kisisellestirme": {
        "name_tr": "Kişiselleştirme / Custom",
        "name_en": "Personalization / Custom",
        "keywords": ["kişiselleş", "custom", "monogram", "personali", "özel ölçü", "quiz"],
        "desc_tr": "Kişiye özel ürün, monogram, quiz-based eşleştirme",
        "group": "Pazarlama & Büyüme",
    },
    "subscription_refill": {
        "name_tr": "Abonelik / Refill Model",
        "name_en": "Subscription / Refill",
        "keywords": ["abonelik", "subscription", "refill", "yeniden doldur"],
        "desc_tr": "Tekrarlayan gelir modeli — abonelik veya refill",
        "group": "Pazarlama & Büyüme",
    },

    # ═══ TARGET AUDIENCE ═══
    "bebek_anne_odakli": {
        "name_tr": "Bebek & Anne Odaklı",
        "name_en": "Baby & Mom Focused",
        "keywords": ["bebek", "baby", "hamile", "emzirme", "anne", "mom", "nursery", "newborn", "yenidoğan"],
        "desc_tr": "Bebek ürünleri, hamile/anne odaklı, nursery",
        "group": "Hedef Kitle",
    },
    "minimalist_quiet_luxury": {
        "name_tr": "Minimalist / Quiet Luxury",
        "name_en": "Minimalist / Quiet Luxury",
        "keywords": ["minimalist", "minimal", "sade", "quiet luxury", "sessiz lüks", "basic"],
        "desc_tr": "Minimalist estetik, sessiz lüks, az ama kaliteli",
        "group": "Hedef Kitle",
    },
    "otel_spa_deneyimi": {
        "name_tr": "Otel & Spa Deneyimi Evde",
        "name_en": "Hotel & Spa Experience at Home",
        "keywords": ["otel", "hotel", "5 yıldız", "resort", "spa"],
        "desc_tr": "Otel/spa kalitesini eve taşıma — aspirasyonel yaşam",
        "group": "Hedef Kitle",
    },
    "boho_eklektik": {
        "name_tr": "Boho / Eklektik",
        "name_en": "Boho / Eclectic",
        "keywords": ["boho", "bohemian", "eklektik", "eclectic", "makrome", "macrame", "tassel", "fringe"],
        "desc_tr": "Boho chic, eklektik, makrome, festival estetiği",
        "group": "Hedef Kitle",
    },

    # ═══ ORIGIN & HERITAGE ═══
    "iskandinav_tasarim": {
        "name_tr": "İskandinav Tasarım",
        "name_en": "Scandinavian Design",
        "keywords": ["iskandinav", "scandinavian", "danimarka", "danish", "isveç", "swedish", "norveç", "finlandiya", "kopenhag", "stockholm"],
        "desc_tr": "İskandinav minimal tasarım, hygge, fonksiyonellik",
        "group": "Köken & Miras",
    },
    "japon_minimalizm": {
        "name_tr": "Japon Minimalizm",
        "name_en": "Japanese Minimalism",
        "keywords": ["japon", "japan", "tokyo", "muji", "wabi-sabi"],
        "desc_tr": "Japon minimalizm, wabi-sabi, fonksiyon > form",
        "group": "Köken & Miras",
    },
    "amerikan_heritage": {
        "name_tr": "Amerikan Heritage",
        "name_en": "American Heritage",
        "keywords": ["amerikan miras", "american heritage", "ABD üretimi", "made in USA", "maine", "new england"],
        "desc_tr": "Amerikan üretim mirası, New England, made in USA",
        "group": "Köken & Miras",
    },
    "avrupa_zanaat": {
        "name_tr": "Avrupa Zanaatı",
        "name_en": "European Craftsmanship",
        "keywords": ["İtalya", "italyan", "italian", "fransız", "french", "paris", "alman", "german", "portekiz", "portuguese", "İngiltere", "ingiliz", "british"],
        "desc_tr": "Avrupa zanaatı — İtalyan, Fransız, Alman, İngiliz kalite",
        "group": "Köken & Miras",
    },

    # ═══ SMART / IOT ═══
    "akilli_ev_iot": {
        "name_tr": "Akıllı Ev / IoT",
        "name_en": "Smart Home / IoT",
        "keywords": ["akıllı", "smart", "IoT", "AI", "app kontrol", "sensör", "bluetooth", "wifi", "alexa", "google home"],
        "desc_tr": "Akıllı ev teknolojisi, app kontrol, sensör, otomasyon",
        "group": "Teknoloji",
    },
}


def analyze_brands(brands_dict):
    """
    Analyze all brands and return marketing angle data.
    brands_dict: {category: [(brand, website, sub_niche, insight), ...]}
    """
    all_brands = []
    for cat, brands in brands_dict.items():
        for b in brands:
            all_brands.append({
                "brand": b[0], "website": b[1], "sub_niche": b[2],
                "insight": b[3], "category": cat
            })

    results = {}
    for angle_id, angle in MARKETING_ANGLES.items():
        matched_brands = []
        for b in all_brands:
            text = (b["insight"] + " " + b["sub_niche"] + " " + b["category"]).lower()
            for kw in angle["keywords"]:
                if kw.lower() in text:
                    matched_brands.append(b)
                    break
        results[angle_id] = {
            "info": angle,
            "count": len(matched_brands),
            "pct": len(matched_brands) / len(all_brands) * 100 if all_brands else 0,
            "brands": matched_brands,
            "categories": Counter([b["category"] for b in matched_brands]),
        }

    return results, all_brands


def add_analysis_to_workbook(wb, results, total_brands):
    """Add analysis sheets to an existing workbook."""

    # ── Styles ───────────────────────────────────────────────────────────────
    DARK = "0D1B2A"
    ACCENT = "1B4965"
    GREEN = "2ECC71"
    ORANGE = "E67E22"
    RED = "E74C3C"
    LIGHT_BG = "F8F9FA"

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    # Sort results by count descending
    sorted_results = sorted(results.items(), key=lambda x: x[1]["count"], reverse=True)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Pazarlama Açısı Analizi (Marketing Angle Analysis)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Pazarlama Açısı Analizi", 0)
    ws.sheet_properties.tabColor = "E74C3C"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="PAZARLAMA AÇISI & TALEP ANALİZİ")
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
    title_cell.fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    sub = ws.cell(row=2, column=1, value=f"Toplam {total_brands} marka analiz edildi | {datetime.now().strftime('%Y-%m-%d')}")
    sub.font = Font(name="Calibri", color="95A5A6", size=11)
    sub.fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["#", "Pazarlama Açısı", "Grup", "Marka Sayısı", "Oran (%)", "Talep Gücü", "Açıklama", "Örnek Markalar"]
    widths = [5, 32, 25, 14, 12, 14, 55, 55]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w

    # Data rows
    row = 5
    for i, (angle_id, data) in enumerate(sorted_results, 1):
        info = data["info"]
        count = data["count"]
        pct = data["pct"]

        # Demand signal color
        if pct >= 8:
            demand = "🔴 ÇOK YÜKSEK"
            demand_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        elif pct >= 5:
            demand = "🟠 YÜKSEK"
            demand_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
        elif pct >= 3:
            demand = "🟡 ORTA"
            demand_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        elif pct >= 1.5:
            demand = "🟢 GELİŞEN"
            demand_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        else:
            demand = "⚪ NİŞ"
            demand_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")

        alt_fill = PatternFill(
            start_color="FFFFFF" if row % 2 == 0 else "F8F9FA",
            end_color="FFFFFF" if row % 2 == 0 else "F8F9FA",
            fill_type="solid"
        )

        # Sample brands (top 5)
        sample = ", ".join([b["brand"] for b in data["brands"][:5]])

        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=info["name_tr"]).font = Font(bold=True, size=10)
        ws.cell(row=row, column=3, value=info["group"]).font = Font(size=10, color="555555")
        ws.cell(row=row, column=4, value=count).alignment = center
        ws.cell(row=row, column=4).font = Font(bold=True, size=11)
        ws.cell(row=row, column=5, value=round(pct, 1)).alignment = center
        ws.cell(row=row, column=6, value=demand).alignment = center
        ws.cell(row=row, column=6).fill = demand_fill
        ws.cell(row=row, column=7, value=info["desc_tr"]).font = Font(size=9, color="555555")
        ws.cell(row=row, column=7).alignment = wrap
        ws.cell(row=row, column=8, value=sample).font = Font(size=9, color="777777")
        ws.cell(row=row, column=8).alignment = wrap

        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
            if c not in [6]:
                ws.cell(row=row, column=c).fill = alt_fill

        row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{row - 1}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Grup Bazlı Özet (Group Summary)
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Grup Bazlı Özet", 1)
    ws2.sheet_properties.tabColor = "3498DB"

    # Group data
    groups = defaultdict(list)
    for angle_id, data in sorted_results:
        g = data["info"]["group"]
        groups[g].append((data["info"]["name_tr"], data["count"], data["pct"]))

    ws2.merge_cells("A1:E1")
    t = ws2.cell(row=1, column=1, value="GRUP BAZLI PAZARLAMA AÇISI ÖZETİ")
    t.font = Font(bold=True, color="FFFFFF", size=14)
    t.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    t.alignment = Alignment(horizontal="center")

    row = 3
    for group_name, angles in groups.items():
        total_group = sum(a[1] for a in angles)
        # Group header
        ws2.merge_cells(f"A{row}:E{row}")
        gh = ws2.cell(row=row, column=1, value=f"  {group_name} ({total_group} toplam marka)")
        gh.font = Font(bold=True, color="FFFFFF", size=12)
        gh.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
        gh.alignment = Alignment(vertical="center")
        row += 1

        # Sub-headers
        for col, h in enumerate(["Pazarlama Açısı", "Marka Sayısı", "Oran (%)", "Talep Gücü", "Bar"], 1):
            c = ws2.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, size=10, color="555555")
            c.border = thin_border
        row += 1

        for name, cnt, pct in sorted(angles, key=lambda x: x[1], reverse=True):
            ws2.cell(row=row, column=1, value=name).font = Font(size=10)
            ws2.cell(row=row, column=2, value=cnt).alignment = center
            ws2.cell(row=row, column=3, value=round(pct, 1)).alignment = center
            if pct >= 8:
                ws2.cell(row=row, column=4, value="ÇOK YÜKSEK").font = Font(color=RED, bold=True, size=9)
            elif pct >= 5:
                ws2.cell(row=row, column=4, value="YÜKSEK").font = Font(color=ORANGE, bold=True, size=9)
            elif pct >= 3:
                ws2.cell(row=row, column=4, value="ORTA").font = Font(color="F39C12", size=9)
            else:
                ws2.cell(row=row, column=4, value="GELİŞEN/NİŞ").font = Font(color=GREEN, size=9)
            ws2.cell(row=row, column=4).alignment = center
            bar = "█" * max(1, int(pct / 1.5))
            ws2.cell(row=row, column=5, value=bar).font = Font(color="3498DB", size=9)
            for c in range(1, 6):
                ws2.cell(row=row, column=c).border = thin_border
            row += 1
        row += 1

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 30

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Açı Detayları (per angle, brand lists)
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Açı Detay — Marka Listesi", 2)
    ws3.sheet_properties.tabColor = "27AE60"

    ws3.merge_cells("A1:G1")
    t3 = ws3.cell(row=1, column=1, value="HER PAZARLAMA AÇISI İÇİN MARKA LİSTESİ")
    t3.font = Font(bold=True, color="FFFFFF", size=14)
    t3.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    t3.alignment = Alignment(horizontal="center")

    detail_headers = ["#", "Marka", "Web Sitesi", "Kategori", "Pazarlama Açısı", "Öne Çıkan Özellik", "Meta Reklamları"]
    detail_widths = [5, 25, 30, 30, 25, 55, 20]
    for col, (h, w) in enumerate(zip(detail_headers, detail_widths), 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border
        ws3.column_dimensions[get_column_letter(col)].width = w

    row = 4
    idx = 1
    for angle_id, data in sorted_results:
        if data["count"] == 0:
            continue
        for b in data["brands"]:
            ws3.cell(row=row, column=1, value=idx).alignment = center
            ws3.cell(row=row, column=2, value=b["brand"]).font = Font(bold=True, size=10)
            url = f"https://{b['website']}" if not b['website'].startswith("http") else b['website']
            wc = ws3.cell(row=row, column=3, value=b["website"])
            wc.hyperlink = url
            wc.font = Font(color="3498DB", underline="single", size=9)
            ws3.cell(row=row, column=4, value=b["category"]).font = Font(size=9)
            ws3.cell(row=row, column=5, value=data["info"]["name_tr"]).font = Font(size=9, color="555555")
            ws3.cell(row=row, column=6, value=b["insight"]).font = Font(size=8, color="666666")
            ws3.cell(row=row, column=6).alignment = wrap

            meta_url = f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={quote(b['brand'])}&search_type=keyword_unordered"
            mc = ws3.cell(row=row, column=7, value="Reklamları Gör")
            mc.hyperlink = meta_url
            mc.font = Font(color="FFFFFF", bold=True, size=9)
            mc.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
            mc.alignment = center

            for c in range(1, 8):
                ws3.cell(row=row, column=c).border = thin_border
            row += 1
            idx += 1

    ws3.freeze_panes = "A4"
    ws3.auto_filter.ref = f"A3:G{row - 1}"

    return wb


if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from yuvacim_home_brands import BRANDS

    OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research_outputs")
    TODAY = datetime.now().strftime("%Y-%m-%d")

    print("Analiz başlıyor...")
    results, all_brands = analyze_brands(BRANDS)

    print(f"Toplam {len(all_brands)} marka analiz edildi")
    print(f"Toplam {len(MARKETING_ANGLES)} pazarlama açısı tarandı")

    # Create workbook with analysis
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    wb = add_analysis_to_workbook(wb, results, len(all_brands))

    filepath = os.path.join(OUTPUT_DIR, f"Yuvacim_Pazarlama_Acisi_Analizi_{TODAY}.xlsx")
    wb.save(filepath)
    print(f"\nAnaliz dosyası: {filepath}")
    print(f"Toplam {sum(d['count'] for d in results.values())} marka-açı eşleşmesi bulundu")

    # Also print top 10 for quick view
    sorted_r = sorted(results.items(), key=lambda x: x[1]["count"], reverse=True)
    print("\n" + "=" * 60)
    print("EN YÜKSEK TALEP PAZARLAMA AÇILARI")
    print("=" * 60)
    for i, (aid, data) in enumerate(sorted_r[:15], 1):
        print(f"{i:2d}. {data['info']['name_tr']:35s} {data['count']:4d} marka ({data['pct']:5.1f}%)")
