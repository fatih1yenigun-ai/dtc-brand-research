#!/usr/bin/env python3
"""
Yuvacım - Ev, Yatak, Battaniye, Yastık, Organizer DTC Markalar
2000+ innovative e-commerce born brands for home textiles, bedding, organizers, and related niches.
All insights in Turkish. Brand names in English.
"""

import os
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Output config ───────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research_outputs")
TODAY = datetime.now().strftime("%Y-%m-%d")
FILENAME = f"Yuvacim_Ev_Yatak_Organizer_DTC_Markalar_{TODAY}.xlsx"

# ─── Color Scheme (refined, soft pastels + rich headers) ─────────────────────
HEADER_COLOR = "1A1A2E"
HEADER_FONT_COLOR = "F5F5F5"
META_BUTTON_COLOR = "2ECC71"
WEBSITE_LINK_COLOR = "3498DB"
INSIGHT_FONT_COLOR = "4A4A4A"
ROW_ALT_1 = "FFFFFF"
ROW_ALT_2 = "F8F9FA"
SUMMARY_HEADER = "16213E"

CATEGORY_COLORS = {
    # Yatak & Nevresim
    "Yatak Çarşafı & Nevresim": ("D6EAF8", "1A5276"),
    "Yorgan & Dolgulu Ürünler": ("D4E6F1", "1B4F72"),
    "Pike & Yatak Örtüsü": ("AED6F1", "21618C"),
    "Bebek & Çocuk Nevresim": ("A9CCE3", "2471A3"),
    # Yastık
    "Uyku Yastığı & Ergonomik": ("D5F5E3", "1E8449"),
    "Dekoratif Yastık & Kırlent": ("ABEBC6", "239B56"),
    "Ortopedik & Terapi Yastık": ("82E0AA", "28B463"),
    # Battaniye
    "Ağırlıklı Battaniye (Weighted)": ("F9E79F", "B7950B"),
    "Pamuklu & Organik Battaniye": ("F7DC6F", "D4AC0D"),
    "Polar & Peluş Battaniye": ("F4D03F", "D4AC0D"),
    "Giyilebilir Battaniye": ("FAD7A0", "CA6F1E"),
    "Bebek Battaniye & Kundak": ("FDEBD0", "E67E22"),
    # Havlu & Banyo
    "Havlu & Bornoz": ("D2B4DE", "7D3C98"),
    "Banyo Aksesuarları": ("C39BD3", "884EA0"),
    # Ev Organizasyonu
    "Dolap & Giysi Organizeri": ("F5CBA7", "CA6F1E"),
    "Çekmece Düzenleyici": ("EDBB99", "BA4A00"),
    "Çamaşır & Kirli Sepeti": ("E59866", "A04000"),
    "Mutfak Organizasyonu": ("D5F5E3", "1E8449"),
    "Banyo Organizasyonu": ("AED6F1", "21618C"),
    "Ayakkabı & Aksesuar Düzenleyici": ("F9E79F", "B7950B"),
    "Seyahat Organizeri & Pouch": ("FADBD8", "A93226"),
    "Makyaj & Kozmetik Organizeri": ("F5B7B1", "CB4335"),
    # Ev Tekstili & Dekor
    "Perde & Tül": ("D7BDE2", "6C3483"),
    "Halı & Kilim": ("C39BD3", "884EA0"),
    "Masa Örtüsü & Runner": ("BB8FCE", "7D3C98"),
    "Dekoratif Tekstil & Makrome": ("D2B4DE", "7D3C98"),
    # Uyku Teknolojisi
    "Uyku Teknolojisi & Akıllı Yatak": ("AED6F1", "21618C"),
    "Şilte & Yatak Pedi": ("D4E6F1", "1B4F72"),
    # Bebek & Anne
    "Bebek Yatağı & Uyku Seti": ("ABEBC6", "239B56"),
    "Anne & Emzirme Yastığı": ("D5F5E3", "1E8449"),
    # Yaşam Tarzı & Ev Kokusu
    "Ev Kokusu & Mum": ("FDEBD0", "E67E22"),
    "Aromaterapi & Difüzör": ("FAD7A0", "CA6F1E"),
    # Sürdürülebilir Ev
    "Sürdürülebilir & Eko Ev Ürünleri": ("82E0AA", "28B463"),
    "Yeniden Kullanılabilir Ev Ürünleri": ("ABEBC6", "239B56"),
    # Günlük Ev & Konfor
    "Ev Giyim & Loungewear": ("F5CBA7", "CA6F1E"),
    "Terlik & Ev Ayakkabısı": ("EDBB99", "BA4A00"),
    # Evcil Hayvan Evi
    "Evcil Hayvan Yatağı & Battaniye": ("D6EAF8", "1A5276"),
    # Taşınabilir & Outdoor
    "Kamp & Outdoor Battaniye": ("D5F5E3", "1E8449"),
    "Piknik & Plaj Örtüsü": ("AED6F1", "21618C"),
}

# ═══════════════════════════════════════════════════════════════════════════════
# BRAND DATA — (brand_name, website, sub_niche, insight_tr)
# ═══════════════════════════════════════════════════════════════════════════════

BRANDS = {
    # ─────────────────────────────────────────────────────────────────────────
    # YATAK ÇARŞAFI & NEVRESİM
    # ─────────────────────────────────────────────────────────────────────────
    "Yatak Çarşafı & Nevresim": [
        ("Brooklinen", "brooklinen.com", "Lüks Pamuklu Çarşaf", "270 iplik sayılı lüks çarşaflarla başladı; DTC yatak devriminin öncüsü; 10M+ satış ilk 2 yılda; referral programı ile viral büyüme"),
        ("Parachute", "parachutehome.com", "Premium Nevresim", "LA merkezli; Oeko-Tex sertifikalı; minimalist tasarım; Instagram estetiği ile büyüdü; fiziksel mağazaya geçmeden önce tamamen online"),
        ("Ettitude", "ettitude.com", "Bambu Çarşaf", "CleanBamboo™ patentli kumaş; %95 daha az su tüketimi; karbon-negatif üretim; vegan PETA onaylı"),
        ("Buffy", "buffy.co", "Eko Nevresim", "Geri dönüşüm PET şişelerden üretilen nevresim; 'Try Before You Buy' modeli; Z kuşağı hedef kitle; pastel estetik"),
        ("Boll & Branch", "bollandbranch.com", "Organik Pamuk", "İlk Fair Trade sertifikalı çarşaf markası; Obama'ların Beyaz Saray'da kullandığı marka; şeffaf tedarik zinciri"),
        ("Cariloha", "cariloha.com", "Bambu Yatak Tekstili", "Bambu lifi teknolojisi; 3 derece daha serin uyku vaadi; Shark Tank'tan çıkma; resort kanalından DTC'ye geçiş"),
        ("Cozy Earth", "cozyearth.com", "Premium Bambu", "Oprah's Favorite Things listesinde; bambu viskon teknolojisi; 10 yıl garanti; Oprah etkisiyle viral"),
        ("Bedthreads", "bedthreads.com.au", "Keten Çarşaf", "Avustralya merkezli; French linen; renk çeşitliliği ile Instagram'da viral; flat-lay fotoğraflarla büyüdü"),
        ("MagicLinen", "magiclinen.com", "Avrupa Keteni", "Litvanya üretimi; OEKO-TEX sertifikalı Avrupa keteni; ön-yıkanmış yumuşak doku; Pinterest estetiği"),
        ("Sijo", "sijohome.com", "TENCEL Çarşaf", "TENCEL™ Lyocell kumaş; okaliptüs lifi; doğal soğutma; minimalist Japon estetiği; küçük batch üretim"),
        ("Sheets & Giggles", "sheetsgiggles.com", "Okaliptüs Çarşaf", "Okaliptüs lifi çarşaf; Kickstarter'da $284K topladı; esprili marka sesi; sürdürülebilir üretim"),
        ("Snowe", "snowehome.com", "Otel Kalitesi Çarşaf", "5 yıldızlı otel kalitesi DTC fiyatına; İtalya üretimi; minimalist ambalaj; subscription modeli denedi"),
        ("Sunday Citizen", "sundaycitizen.co", "Kristal İnfüzyon", "Amethyst kristal infüzyon teknolojisi; bamboo-blend; lüks loungewear + yatak; Forbes'ta öne çıkan"),
        ("Piglet in Bed", "pigletinbed.com", "Doğal Keten", "İngiltere merkezli; gingham desen TikTok'ta viral; oeko-tex keten; ev koleksiyonuna genişledi"),
        ("Primary Goods", "primarygoods.com", "Temel Yatak Ürünleri", "Sadece temel ürünler; abartısız pazarlama; dürüst fiyatlandırma; her bütçeye uygun premium"),
        ("Riley Home", "rileyhome.com", "Sateen Çarşaf", "T-shirt yumuşaklığında çarşaf; ex-Brooklinen ekibi kurdu; 25 yıkamada yumuşama garantisi"),
        ("Cultiver", "cultiver.com.au", "Lüks Keten", "Avustralya; el yıkaması keten; interior tasarımcıların tercihi; doğal renk paleti"),
        ("Kassatex", "kassatex.com", "Türk Pamuğu Çarşaf", "Türk pamuğu uzmanı; otel tedarikçisinden DTC'ye; 30 yıllık tekstil tecrübesi"),
        ("Under the Canopy", "underthecanopy.com", "Organik & Fair Trade", "GOTS sertifikalı organik; %100 geri dönüşüm ambalaj; B-Corp bekliyor; uygun fiyat segmenti"),
        ("Looma", "loomahome.com", "El Dokuma Çarşaf", "Hindistan'da el dokuma; artisan üretim; her satışta zanaatkar desteği; benzersiz tekstür"),
        ("Rise & Fall", "riseandfall.co", "İngiliz Tasarım", "İngiltere tasarımı; Portekiz üretimi; organik pamuk; sessiz lüks estetiği"),
        ("Coyuchi", "coyuchi.com", "Organik Keten & Pamuk", "B-Corp sertifikalı; organik keten ve pamuk; 2nd Home takeback programı; döngüsel ekonomi modeli"),
        ("Pact", "wearpact.com", "Organik Yatak", "Fair Trade organik pamuk; uygun fiyat + etik üretim; giysiden yatak tekstiline genişledi"),
        ("Sol Organics", "solorganix.com", "Fair Trade Çarşaf", "Fair Trade + GOTS; Hindistan üretimi; şeffaf maliyet yapısı; bağış programı"),
        ("Olive + Crate", "oliveandcrate.com", "Bambu Saten", "Bambu saten çarşaf; ipeksi his; termal regülasyon; uygun lüks segmenti"),
        ("Pizuna Linens", "pizuna.com", "Uzun Elyaf Pamuk", "400TC uzun elyaf pamuk; fabrikadan direkt; otel hissi ev fiyatına"),
        ("The Citizenry", "the-citizenry.com", "Zanaatkar Çarşaf", "Dünya zanaatkarları ile üretim; Peru alpaka, Hindistan pamuk; hikaye odaklı pazarlama"),
        ("Avocado Green", "avocadogreenmattress.com", "Organik Yatak Seti", "B-Corp; GOLS, GOTS sertifikalı; organik lateks ve pamuk; karbon-negatif operasyon"),
        ("Red Land Cotton", "redlandcotton.com", "Alabama Pamuğu", "ABD Alabama'da yetiştirilen pamuk; tarladan yatağa iz takibi; küçük aile çiftliği"),
        ("Rough Linen", "roughlinen.com", "Ham Keten", "Yıkanmamış ham keten; zamanla güzelleşen ürün felsefesi; California tasarım; artisan üretim"),
        ("Mellanni", "mellanni.com", "Mikrofiber Çarşaf", "Amazon #1 çarşaf markası; 200K+ 5 yıldız; TikTok Shop'ta viral; ultra uygun fiyat"),
        ("California Design Den", "californiadesignden.com", "400TC Sateen", "Uzun elyaf sateen; 100K+ Amazon yorum; DTC + marketplace hibrit model"),
        ("Bedsure", "bedsurehome.com", "Çok Ürünlü Yatak", "Amazon dev satıcıdan marka geçişi; fleece + saten + bambu seçenekleri; TikTok'ta agresif pazarlama"),
        ("Nolah", "nolahmattress.com", "Hava Fiber Çarşaf", "AirFiber™ teknolojisi; allerji dostu; şilteden çarşafa genişledi"),
        ("Tuft & Needle", "tuftandneedle.com", "T&N Çarşaf", "Şilte devrimcisinden yatak setine; adaptive foam teknolojisi; radikal şeffaflık"),
        ("DreamCloud", "dreamcloudsleep.com", "Lüks Hibrit Çarşaf", "365 gece deneme; cashmere blend seçeneği; şilte markasından genişleme"),
        ("Nest Bedding", "nestbedding.com", "Eko Çarşaf", "CertiPUR-US; organik seçenekler; ömür boyu konfor garantisi"),
        ("PlushBeds", "plushbeds.com", "Organik Lateks Seti", "GreenGuard Gold; %100 organik; el yapımı; 25 yıl garanti"),
        ("Authenticity 50", "authenticity50.com", "Made in USA Çarşaf", "%100 ABD üretimi; Amerikan pamuğu; şeffaf üretim zinciri; vatansever marka"),
        ("Quince", "onequince.com", "Fabrikadan Direkt", "Lüks markaların fabrikalarını kullanıp %50-80 ucuza satış; radikal şeffaflık; Mongolian cashmere dahil"),
        ("Linoto", "linoto.com", "El Dikimi Keten", "ABD'de el dikimi keten çarşaf; 30+ renk; küçük atölye üretimi; butik marka"),
        ("Holy Lamb Organics", "holylamborganics.com", "Organik Yün Yatak", "Organik yün + pamuk; elle yapım; küçük çiftlik kaynağı; Washington üretimi"),
        ("White Lotus Home", "whitelotushome.com", "Doğal Yatak", "Kimyasal içermeyen doğal yatak; pamuk, yün, kapok dolgu; 30 yıllık zanaat"),
        ("Tekla", "teklafabrics.com", "Skandinav Çarşaf", "Kopenhag tasarım; Portekiz üretimi; kalın percale; unisex pijama + çarşaf"),
        ("AREA", "areahome.com", "NYC Tasarım Çarşaf", "New York tasarım stüdyosu; baskılı organik pamuk; sanat eseri çarşaflar; limited edition"),
        ("Johnstons of Elgin", "johnstonsofelgin.com", "Cashmere Yatak", "Kaşmir uzmanı; 225 yıllık İskoç mirası; DTC'ye geçiş yapan heritage marka"),
        ("IN BED", "inbed.com.au", "Minimalist Keten", "Avustralya; French flax linen; pastel renkler; slow fashion yatak; Pinterest estetiği"),
        ("Hawkins New York", "hawkinsnewyork.com", "Modüler Yatak", "Basit, modüler yatak ürünleri; tasarım odaklı; Brooklyn estetiği"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # YORGAN & DOLGULU ÜRÜNLER
    # ─────────────────────────────────────────────────────────────────────────
    "Yorgan & Dolgulu Ürünler": [
        ("Buffy", "buffy.co", "Eko Yorgan", "Geri dönüşüm PET şişelerden üretim; 'Cloud Comforter' viral ürün; deneme-sonra-öde modeli; 50 şişe = 1 yorgan"),
        ("Casaluna (Target DTC)", "casaluna.com", "Organik Yorgan", "Target'ın DTC organik yatak markası; GOTS sertifikalı; uygun fiyat premium"),
        ("Sunday Citizen", "sundaycitizen.co", "Kristal Yorgan", "Kristal infüzyon teknolojisi; bambu karışımı; snug kanepe battaniyesi viral oldu"),
        ("Baloo Living", "balooliving.com", "Ağırlıklı Yorgan", "Ağırlıklı yorgan; cam boncuk dolgu; serin uyku; anksiyete azaltma odaklı pazarlama"),
        ("Pacific Coast Feather", "pacificcoast.com", "Kuş Tüyü Yorgan", "Hyperclean® kuş tüyü işleme; 5 yıldızlı otel tedarikçisi; allerji dostu"),
        ("Puredown", "puredown.com", "Kaz Tüyü Yorgan", "Premium kaz tüyü; mevsimsel seçenekler; Amazon DTC hybrid; uygun lüks"),
        ("Hungarian Goose Down", "goosedownduvet.com", "Macar Kaz Tüyü", "950+ fill power Macar kaz tüyü; ultra premium segment; sınırlı üretim"),
        ("Scandia Home", "scandiahome.com", "İskandinav Tüy Yorgan", "Danimarka tarzı; beyaz kaz tüyü; otel lüksü; 30 yıl garanti"),
        ("Cuddledown", "cuddledown.com", "Yün & Tüy Yorgan", "Maine merkezli; down + yün + pamuk dolgu seçenekleri; 50 yıllık marka DTC'ye döndü"),
        ("The Company Store", "thecompanystore.com", "LaCrosse Yorgan", "LaCrosse® Down; Amerikan mirası; kişiselleştirilebilir sıcaklık seviyesi"),
        ("Snowe", "snowehome.com", "Otel Yorgan", "İtalya üretimi; 5 yıldız otel hissi; minimalist paket; DTC lüks"),
        ("Slumber Cloud", "slumbercloud.com", "Soğutma Teknolojisi", "Outlast® NASA teknolojisi; sıcaklık düzenleyici; gece terlemesi çözümü"),
        ("Rest", "rest.com", "Akıllı Yorgan", "Sıcaklık algılayan kumaş; tek yorgan tüm mevsim; sleep tech entegrasyonu"),
        ("Comphy", "comphy.com", "Spa Yorgan", "Spa ve otel tedarikçisi; mikrofiber cloud dokusu; ev kullanımına açıldı"),
        ("Silk & Snow", "silkandsnow.com", "Kanada Organik", "Kanada DTC; organik pamuk + Tencel; sade tasarım; 100 gece deneme"),
        ("Warm Things", "warmthings.com", "Alman Tüy Yorgan", "Alman mühendisliği; baffle box tasarım; 800+ fill power; cold climate uzmanı"),
        ("Down & Feather Co", "downandfeatherco.com", "Kişisel Tüy Yorgan", "Kişiselleştirilebilir doluluk; tüy yenileme servisi; 60 yıllık zanaat"),
        ("Cloud Bedding", "cloudbedding.co", "Bulut Yorgan", "Ultra hafif 'floating' his; sosyal medya odaklı genç marka; pastel renkler"),
        ("Novaform Bedding", "novaform.com", "Jel Soğutma Yorgan", "Jel infüzyon soğutma; Costco'dan DTC'ye geçiş; termal regülasyon teknolojisi"),
        ("Luna Weighted", "lunaweightedblanket.com", "Ağırlıklı Yorgan/Battaniye", "Cam boncuk; %100 pamuk; uygun fiyat ağırlıklı seçenek; Amazon'da #1"),
        ("Sleep Number Bedding", "sleepnumber.com", "Akıllı Yorgan", "Sleep IQ® teknolojisi; True Temp™ kumaş; IoT yatak ekosistemi"),
        ("Garnet Hill", "garnethill.com", "Doğal Fiber Yorgan", "Organik pamuk + yün; New England mirası; çevreci üretim; 40 yıllık DTC"),
        ("Under the Canopy", "underthecanopy.com", "Fair Trade Yorgan", "GOTS + Fair Trade; %100 geri dönüşüm dolgu; uygun fiyat eko seçenek"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # PİKE & YATAK ÖRTÜSÜ
    # ─────────────────────────────────────────────────────────────────────────
    "Pike & Yatak Örtüsü": [
        ("Boll & Branch", "bollandbranch.com", "Organik Pike", "Fair Trade; organik waffle pike; Obama ailesi kullanıyor; şeffaf üretim"),
        ("Pottery Barn Teen DTC", "pbteen.com", "Genç Pike", "Genç odası estetiği; kişiselleştirme seçenekleri; sosyal medya odaklı"),
        ("Lands' End Bedding", "landsend.com", "Dayanıklı Pike", "Ömür boyu garanti; klasik Amerikan tarzı; DTC geçişi"),
        ("Morrow Soft Goods", "morrowsoftgoods.com", "El Yapımı Pike", "Meksika'da el dokuma; kültürel desenler; artisan üretim; Instagram estetiği"),
        ("Nate Home by Nate Berkus", "natehome.com", "Tasarımcı Pike", "Celebrity iç mimar Nate Berkus; accessible luxury; Target'tan bağımsız DTC marka"),
        ("House No. 23", "houseno23.com", "Boutique Pike", "Küçük batch; doğal boya; el dikimi; luxury boho estetik"),
        ("Coyuchi", "coyuchi.com", "Organik Pike", "B-Corp; organik pamuk; doğal renkler; 2nd Home geri dönüşüm programı"),
        ("Libeco", "libeco.com", "Belçika Keteni Pike", "1858'den beri keten üretimi; Belçika mirası; DTC online genişleme"),
        ("Lulu & Georgia", "luluandgeorgia.com", "Boho Pike", "California boho estetik; influencer iş birlikleri; Instagram-first marka"),
        ("Serena & Lily", "serenaandlily.com", "Coastal Pike", "Kıyı estetiği; mavi-beyaz paleti; lakeside yaşam; premium DTC"),
        ("Crane & Canopy", "craneandcanopy.com", "Modern Pike", "Tasarımcı kalitesi DTC fiyatına; kolay takma klips sistemi; %100 pamuk"),
        ("Peacock Alley", "peacockalley.com", "Lüks Pike", "Teksas lüks; Amerikan başkanlık hediye markası; 50 yıllık miras DTC'ye döndü"),
        ("Archive New York", "archivenewyork.com", "Guatemala Dokuma", "Guatemala zanaatkarları ile iş birliği; el dokuma kumaşlar; kültürel koruma misyonu"),
        ("Calhoun & Co", "calhounandco.com", "Renkli Pike", "Cesur renkler ve desenler; genç hedef kitle; TikTok pazarlama"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # BEBEK & ÇOCUK NEVRESİM
    # ─────────────────────────────────────────────────────────────────────────
    "Bebek & Çocuk Nevresim": [
        ("Kyte Baby", "kytebaby.com", "Bambu Bebek", "Bambu rayon; ultra yumuşak; bebek uyku tulumu viral; TikTok momfluencer ordusu; $100M+ gelir"),
        ("Aden + Anais", "adenandanais.com", "Muslin Bebek", "Muslin kundak öncüsü; Avustralya kökenli; ünlü anne onayı; hastanelerde tercih edilen"),
        ("Pehr", "pfrdesigns.com", "Tasarım Bebek Odası", "Kanada tasarım; el boyama baskılar; nursery dekoru; Pinterest anne estetiği"),
        ("Snuggle Me Organic", "snugglemeorganic.com", "Organik Bebek Yatağı", "GOTS organik; lounger ürünü viral; anne gruplarında organik büyüme"),
        ("Burt's Bees Baby", "burtsbeesbaby.com", "Organik Bebek Çarşaf", "GOTS organik pamuk; Burt's Bees güvenilirliği; uygun fiyat organik bebek"),
        ("Loulou Lollipop", "louloulollipop.com", "Muslin Bebek Çarşaf", "Kanada; tropikal baskı muslin; bambu karışım; Instagram anne topluluğu"),
        ("Halo", "halosleep.com", "SleepSack", "SleepSack giyilebilir battaniye; SIDS önleme odaklı; hastane ortaklıkları; güvenli uyku misyonu"),
        ("Solly Baby", "sollybaby.com", "Wrap & Çarşaf", "Baby wrap'ten çarşafa genişleme; TENCEL modal; ultra yumuşak; minimal estetik"),
        ("Mori", "babymori.com", "Bambu Bebek Uyku", "İngiltere; organik bambu + pamuk; akıllı uyku takibi entegrasyonu; Prens ailesinin tercihi"),
        ("Copper Pearl", "copperpearl.com", "Çok Kullanımlı Bebek", "Stretchy muslin; bib, kundak, çarşaf hepsi aynı kumaş; TikTok'ta anne içerikleri"),
        ("Little Unicorn", "littleunicorn.com", "Desenli Muslin", "Disney lisans; premium baskı muslin; quilted battaniye; hediye seti odaklı"),
        ("Gunamuna", "gunamuna.com", "Fermuarlı Uyku Tulumu", "WonderZip® teknoloji; kolay bez değiştirme; TOG derecelendirme; uyku uzmanı onaylı"),
        ("Woolino", "woolino.com", "Merinos Yünü Bebek", "4 mevsim merinos yünü; 2-24 ay tek ürün; sıcaklık düzenleme; uzun ömürlü tasarım"),
        ("Nested Bean", "nestedbean.com", "Ağırlıklı Uyku Tulumu", "Zen Sack™ hafif ağırlık; anne dokunuşu simülasyonu; pediatrist onaylı"),
        ("SwaddleDesigns", "swaddledesigns.com", "Hemşire Tasarımı", "Pediatri hemşiresi kurdu; marquisette muslin; güvenli uyku eğitim odaklı"),
        ("Dream Weighted Sack", "dreamlandbabyco.com", "Ağırlıklı Bebek", "Shark Tank yatırımı; hafif ağırlıklı uyku tulumu; daha uzun uyku vaadi"),
        ("Poppy Kids", "poppykids.co", "Montessori Yatak", "Montessori yer yatağı + çarşaf seti; eğitim odaklı; bağımsız keşif"),
        ("Gathre", "gathre.com", "Mat & Çarşaf", "Deri mat + çarşaf; kolay temizlik; modern anne estetiği; çok fonksiyonlu"),
        ("Bloomere", "bloomere.com", "Kore Bebek Yatak", "Kore tasarımı; organik pamuk; allerjen-free; minimalist K-style estetik"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # UYKU YASTIĞI & ERGONOMİK
    # ─────────────────────────────────────────────────────────────────────────
    "Uyku Yastığı & Ergonomik": [
        ("Coop Home Goods", "coophomegoods.com", "Ayarlanabilir Yastık", "Ayarlanabilir dolgu; CertiPUR-US memory foam; Amazon #1 yastık; cross-cut foam teknolojisi"),
        ("Pillow Cube", "pillowcube.com", "Küp Yastık", "Yan yatanlar için küp şeklinde yastık; TikTok viral; $20M+ ilk yıl; esprili marka sesi"),
        ("Purple Harmony", "purple.com", "GelFlex Yastık", "GelFlex® Grid teknolojisi; sıcaklık nötr; basınç dağılımı; sleep tech öncüsü"),
        ("Pluto Pillow", "plutopillow.com", "Kişiselleştirilmiş", "35+ vücut/uyku/tercih sorusu ile kişiye özel yastık üretimi; AI algoritması; D2C"),
        ("Casper Pillow", "casper.com", "İç İçe Yastık", "Yastık içinde yastık tasarımı; dış katman yumuşak, iç katman destek; şilte markasından yastığa"),
        ("Tempur-Pedic Pillow", "tempurpedic.com", "NASA Memory Foam", "NASA kaynaklı TEMPUR® malzeme; basınç hisseden foam; 25+ yıl yastık inovasyonu"),
        ("Eli & Elm", "eliandelm.com", "Omuz Destek Yastık", "Yan yatış için omuz kesikli tasarım; lateks + polyester blend; ergonomik şekil"),
        ("Sleep Number Pillow", "sleepnumber.com", "Akıllı Yastık", "ComfortFit™ seçenekleri; SleepIQ uyumu; sıcaklık düzenleme"),
        ("Saatva Pillow", "saatva.com", "Lüks Yastık", "Talalay lateks; organik pamuk kılıf; otel lüksü DTC fiyatına"),
        ("Hullo", "hullopillow.com", "Karabuğday Yastık", "Organik karabuğday kabukları; doğal hava sirkülasyonu; 20 yıl dayanıklılık; Japon geleneği"),
        ("Marlow Pillow", "marlowpillow.com", "Anti-Gravity Yastık", "Brooklinen alt markası; çevirebilir sertlik; bir taraf yumuşak bir taraf sert"),
        ("Helix Pillow", "helixsleep.com", "Soğutma Yastık", "GlacioTex™ soğutma; kişiselleştirilmiş uyku testi; şilte markasından genişleme"),
        ("Sleepgram", "sleepgram.com", "3'ü 1 Arada Yastık", "3 iç yastık ile ayarlanabilir yükseklik; $45'a premium his; Amazon viral"),
        ("Bear Pillow", "bearmattress.com", "Sporcu Yastık", "Celliant® kılıf; kas toparlanma; sporcu odaklı uyku; performans yastığı"),
        ("Lagoon", "lagoonpillow.com", "Uyku Quizi Yastık", "Online quiz ile eşleştirme; 5 farklı model; uyku pozisyonuna göre kişiselleştirme"),
        ("Honeydew Sleep", "honeydewallyouneedizsleep.com", "Scrumptious Yastık", "Copper-infused foam; beyin soğutma; TikTok viral; esprili marka ismi"),
        ("SONU Sleep", "sonusleep.com", "Yüzüstü Yatış Yastık", "Yüzüstü yatanlar için özel tasarım; göğüs ve kol boşluklu; niş pazar lideri"),
        ("WaterPillow", "mediflow.com", "Su Bazlı Yastık", "Su katmanı ile ayarlanabilir destek; klinik olarak kanıtlanmış; 30+ yıl piyasada"),
        ("Snuggle-Pedic", "snugglepedic.com", "Shredded Memory", "Keansburg bamboo + shredded memory foam; adjustable; ABD üretimi"),
        ("Good Pillow", "goodpillow.com", "Minimalist Yastık", "Tek SKU stratejisi; karar yorgunluğunu ortadan kaldırma; herkes için bir yastık"),
        ("Smart Nora", "smartnora.com", "Anti-Horlama Yastık", "Horlama algılayan insert; AI sensör; kontaktsız çözüm; 130K+ müşteri"),
        ("Ostrichpillow", "ostrichpillow.com", "Taşınabilir Uyku Yastık", "Ofis/seyahat mikro uyku yastığı; ikonik tasarım; Kickstarter başarısı; hediye favorisi"),
        ("Beckham Hotel Collection", "beckhamhotelcollection.com", "Otel Yastık", "Amazon'un en çok satan yastığı; otel kalitesi; gel fiber dolgu; 500K+ yorum"),
        ("Layla Pillow", "laylasleep.com", "Bakır İnfüzyon", "Copper-infused memory foam; antimikrobiyal; kapok fiber blend"),
        ("Panda London", "pandalondon.com", "Bambu Yastık", "İngiltere; bambu memory foam; BambooCloud™; vegan ve hypoallerjenik"),
        ("Original Pillow", "originalpillow.com", "Polimer Yastık", "Polimer grid core; nefes alabilir; sıcaklık nötr; startup inovasyonu"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # DEKORATİF YASTIK & KIRLENT
    # ─────────────────────────────────────────────────────────────────────────
    "Dekoratif Yastık & Kırlent": [
        ("Society6", "society6.com", "Sanatçı Kırlent", "Bağımsız sanatçı tasarımları; print-on-demand model; 300K+ sanatçı; küresel topluluk"),
        ("Lulu & Georgia", "luluandgeorgia.com", "Boho Kırlent", "California boho; tasarımcı iş birlikleri; Instagram estetik; premium fiyat"),
        ("The Citizenry", "the-citizenry.com", "Zanaatkar Kırlent", "Dünya zanaatkarları; el işi; hikaye odaklı; Fair Trade; adil ticaret"),
        ("Jungalow", "jungalow.com", "Tropikal Kırlent", "Justina Blakeney markası; renkli maksimalist; bitki + etnik desen; diverse estetik"),
        ("Block Shop", "blockshoptextiles.com", "El Baskı Kırlent", "Hindistan el baskı; block print; geleneksel teknik modern tasarım; B-Corp"),
        ("St. Frank", "stfranklinens.com", "Afrika Tekstil Kırlent", "Afrika kumaşları; kente, mud cloth; sosyal etki; kâr paylaşım modeli"),
        ("Bolé Road Textiles", "boleroadtextiles.com", "Etiyopya Dokuma", "Etiyopya el dokuması; geleneksel kalıplar; kadın zanaatkarları güçlendirme"),
        ("Minna", "mightynest.com", "El Dokuma Kırlent", "Latin Amerika zanaatkarları; sürdürülebilir boya; el tezgahı; minimal estetik"),
        ("Bfgfm", "bfriendsgenericfurnishingsmarket.com", "Modüler Kırlent", "Modüler bileşen sistemi; değiştirilebilir kılıf; döngüsel tasarım"),
        ("Parachute Decor", "parachutehome.com", "Premium Kırlent", "Keten + pamuk blend; doğal renkler; yatak markasından dekor genişleme"),
        ("Dusen Dusen", "dufrfrsendusen.com", "Grafik Kırlent", "NYC grafik tasarımcı; bold patern; renk blokları; yatak + kırlent; pop art estetik"),
        ("Rebecca Atwood", "rebeccaatwood.com", "Sanatsal Kırlent", "Sulu boya baskılar; Brooklyn stüdyosu; küçük batch; sanat + ev tekstili birleşimi"),
        ("Lindell & Co", "lindellandco.com", "İsveç Tasarım Kırlent", "Stockholm tasarım; cesur desenler; pamuk canvas; İskandinav + eklektik"),
        ("Aiayu", "aiayu.com", "Bolivya Yün Kırlent", "Bolivya lama yünü; minimal İskandinav form; el işi; slow design"),
        ("Eny Lee Parker", "enyleeparker.com", "Organik Form Kırlent", "Organik şekiller; blob estetik; seramik sanatçı ev tekstiline geçiş"),
        ("Cold Picnic", "coldpicnic.com", "Tufted Kırlent", "Brooklyn; tufted halı tekniği kırlentlerde; pop sanat; vücut form yastıkları"),
        ("Baina", "bfrfrfrina.com", "Havlu & Kırlent", "Avustralya; organik pamuk; spa estetiği; banyo + yatak odası geçişi"),
        ("Casa Amarosa", "casaamarosa.com", "Hindistan Kırlent", "El işi Hindistan; block print; doğal boya; etik üretim; boho chic"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # ORTOPEDİK & TERAPİ YASTIK
    # ─────────────────────────────────────────────────────────────────────────
    "Ortopedik & Terapi Yastık": [
        ("Therapeutica", "therapeuticainc.com", "Servikal Yastık", "Omurga hizalama; 5 bölgeli ergonomik tasarım; fizyoterapist tasarımı; medikal sınıf"),
        ("Elviros", "elviros.com", "Servikal Memory Foam", "Kelebek şekli servikal yastık; Amazon viral; boyun ağrısı çözümü; ergonomik kontur"),
        ("Core Products", "coreproducts.com", "Fizyoterapi Yastık", "Klinik kullanım; traksiyon yastığı; boyun destek; sağlık profesyoneli onaylı"),
        ("Dosaze", "dosaze.com", "Ayarlanabilir Ortopedik", "5 modül sistemi; kişiselleştirilebilir yükseklik; kiropraktör tasarımı; Kickstarter başarısı"),
        ("Groove Pillow", "groovepillow.com", "Baş Ağrısı Yastık", "Merkezi kanal; baş ağrısı/migren odaklı; oksipital bölge desteği; niş pazar"),
        ("Mkicesky", "mkicesky.com", "Bel Destek Yastık", "Ofis + araba bel yastığı; memory foam; Amazon bestseller; ergonomik destek"),
        ("ComfiLife", "comfilife.com", "Coccyx Yastık", "Kuyruk sokumu yastığı; ortopedik oturma; ofis çalışanı hedef kitle; Amazon #1"),
        ("Purple Seat Cushion", "purple.com", "Grid Oturma", "GelFlex Grid oturma yastığı; basınç dağılımı; uzun oturma çözümü"),
        ("Nodpod", "nodpod.com", "Ağırlıklı Göz Yastığı", "Ağırlıklı göz maskesi; baş ağrısı rahatlatma; 4 boncuklu kanal; Shark Tank"),
        ("Bed Buddy", "bedbuddy.com", "Terapi Yastık", "Isı/soğuk terapi; doğal tahıl dolgu; mikrodalga ısıtma; ağrı yönetimi"),
        ("Tempur-Neck", "tempurpedic.com", "Boyun Kontur", "TEMPUR malzeme; kontur servikal tasarım; sırt ve yan yatış uyumlu"),
        ("Cervical Traction Pillow", "neckcloud.io", "Boyun Germe", "10 dakika boyun germe; kas gevşetme; TikTok'ta fizik tedavi viral"),
        ("Everlasting Comfort", "everlastingcomfort.net", "Memory Foam Set", "Bel + diz + boyun yastık seti; Amazon marka; ofis ergonomisi seti"),
        ("Cushion Lab", "cushionlab.com", "Ergonomik Oturma", "Ekstra yoğun memory foam; ofis + ev; 'Deep Pressure Relief' teknolojisi"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # AĞIRLIKLI BATTANİYE (WEIGHTED)
    # ─────────────────────────────────────────────────────────────────────────
    "Ağırlıklı Battaniye (Weighted)": [
        ("Bearaby", "bearaby.com", "Örgü Ağırlıklı", "Dolgu malzeme YOK; ağırlık kumaşın kendisinden; Tree Napper organik TENCEL; estetik tasarım; Shark Tank"),
        ("Gravity Blanket", "gravityblankets.com", "Orijinal Ağırlıklı", "Ağırlıklı battaniye kategorisini yaratan marka; Kickstarter'da $4.7M topladı; anksiyete azaltma odaklı"),
        ("Luna Blanket", "lunaweightedblanket.com", "Uygun Fiyat Ağırlıklı", "Amazon #1 ağırlıklı battaniye; cam boncuk; %100 pamuk; 15lbs en popüler; düşük fiyat liderliği"),
        ("Baloo Living", "balooliving.com", "Doğal Ağırlıklı", "BPA-free cam boncuk; %100 pamuk; nefes alabilir tasarım; kimyasal yok; bebek güvenliği standardı"),
        ("Hush Blankets", "hfrfrushblankets.com", "Kanada Ağırlıklı", "Kanada #1; iced soğutma versiyonu; 100 gece deneme; sezonsal seçenekler"),
        ("Layla Weighted", "laylasleep.com", "Çift Taraflı", "Bir taraf serin bir taraf sıcak; copper-infused; 300TC pamuk; şilte markasından genişleme"),
        ("Weighted Evolution", "weightedevolution.com", "Eko Ağırlıklı", "Geri dönüşüm cam boncuk; organik pamuk; çevre dostu; küçük batch"),
        ("SensaCalm", "sensacalm.com", "Duyusal Battaniye", "Otizm/duyusal işleme odaklı; kişiselleştirilebilir ağırlık; terapist onaylı"),
        ("Mosaic Weighted", "mosaicweightedblankets.com", "Terapötik", "FDA kayıtlı; otizm + ADHD + anksiyete; terapötik kullanım; pediatrik seçenekler"),
        ("YnM", "ynmhome.com", "Bambu Ağırlıklı", "Bambu kılıf seçeneği; 7 katmanlı sistem; Amazon bestseller; uygun fiyat"),
        ("Nuzzie", "nuzzie.co", "Örgü Ağırlıklı", "Bearaby alternatifi; chunky knit; daha uygun fiyat; Instagram estetiği"),
        ("Quility", "qufrfrilityweightedblankets.com", "Çocuk Ağırlıklı", "Çocuk odaklı boyutlar; eğlenceli renkler; iç + dış kılıf; yıkanabilir"),
        ("Comma Home", "commahome.com", "Minimalist Ağırlıklı", "Skandinav minimalist; TENCEL kılıf; japondan esinlenen ambalaj; hediye pazarı"),
        ("Zonli", "zonlihome.com", "Soğutma Ağırlıklı", "Cooling fabric + ağırlık; gece terlemesi çözümü; TikTok viral; yazlık versiyon"),
        ("Degrees of Comfort", "degreesofcomfort.com", "Termal Ağırlıklı", "CoolMax® kumaş; nano-seramik boncuk; 60/80/100% pamuk seçenekleri"),
        ("Wemore", "wemore.com", "Çocuk Karakter", "Karakter temalı ağırlıklı battaniye; çocuklar için eğlenceli; unicorn, dinozor desenleri"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # PAMUKLU & ORGANİK BATTANİYE
    # ─────────────────────────────────────────────────────────────────────────
    "Pamuklu & Organik Battaniye": [
        ("Rumpl", "rumpl.com", "Teknik Battaniye", "Geri dönüşüm plastik şişeden; su geçirmez outdoor battaniye; NASA ısı yalıtım; National Park koleksiyonu"),
        ("Pendleton", "pendleton-usa.com", "Amerikan Yün", "150+ yıllık miras; Amerikan Yerli sanat desenleri; DTC genişleme; ikonik yün battaniye"),
        ("ChappyWrap", "chappywrap.com", "Jacquard Battaniye", "Maine mirası; 1000+ desen; jacquard dokuma; makinede yıkanabilir; hediye favorisi"),
        ("Boll & Branch", "bollandbranch.com", "Organik Throw", "Fair Trade; organik pamuk; waffle dokuma; lüks his düşük çevresel etki"),
        ("Garnet Hill", "garnethill.com", "Doğal Battaniye", "Organik pamuk + merinos yünü; New England mirası; 40 yıl DTC"),
        ("Brahms Mount", "brahmsmount.com", "Maine Dokuma", "ABD Maine'de el tezgahı; 1983'ten beri; pamuk + merinos; Amerikan zanaatı"),
        ("Faribault Woolen Mill", "faribaultmill.com", "Minnesota Yünü", "1865'ten beri; Amerikan yünü; heritage marka; fabrika turu; DTC revival"),
        ("The Tartan Blanket Co", "tartanblanketco.com", "İskoç Tartan", "İskoçya; geri dönüşüm yün; tartan desen; sürdürülebilir hediye ambalaj"),
        ("Coyuchi", "coyuchi.com", "Organik Throw", "B-Corp; organik pamuk/yün; Marin County; 2nd Home program; doğal boyalar"),
        ("Happy Habitat", "happyhabitat.com", "Geri Dönüşüm Pamuk", "Atık pamuktan battaniye; rengarenk desenler; zero waste üretim; eko tasarım"),
        ("Blanket Creek Studio", "blanketcreekstudio.com", "Bağımsız Sanatçı", "Küçük stüdyo; el baskı + doğal boya; sınırlı üretim; artisan marka"),
        ("Evangeline Linens", "evangelinelinens.com", "Merinos Throw", "Maine; merinos yünü; classic American; küçük batch; butik lüks"),
        ("in2green", "in2green.com", "Geri Dönüşüm Battaniye", "Geri dönüşüm pamuk; Kuzey Carolina üretimi; Amerikan yapımı; eko bilinçli"),
        ("Cotton Cloud", "cottoncloudshop.com", "Türk Pamuğu Battaniye", "Türk pamuğu; pestemal battaniye; çok kullanımlı; hafif ve taşınabilir"),
        ("Sackcloth & Ashes", "sackclothandashes.com", "Buy One Give One", "Her satışta evsizlere battaniye bağışı; sosyal etki modeli; B-Corp"),
        ("Kennebunk Home", "kennebunkhome.com", "Örgü Throw", "Chunky knit; eco-baby alpaca; renk çeşitliliği; lüks hediye"),
        ("Lands Downunder", "landsdownunder.com", "Alpaka Throw", "Baby alpaka; Güney Amerika kaynağı; doğal fiber; hipoalerjenik"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # POLAR & PELUŞ BATTANİYE
    # ─────────────────────────────────────────────────────────────────────────
    "Polar & Peluş Battaniye": [
        ("Minky Couture", "minkycouture.com", "Lüks Minky", "Ultra yumuşak minky kumaş; Utah başlangıç; anne girişimci hikayesi; hediye pazarı lideri; viral TikTok"),
        ("Big Blanket Co", "bigblanket.com", "Dev Battaniye", "10x10 feet dev battaniye; 'tüm aile' pazarlaması; aile film gecesi; viral sosyal medya"),
        ("Barefoot Dreams", "barefootdreams.com", "CozyChic Battaniye", "CozyChic® patentli kumaş; Oprah's Favorite Things; Kardashian onaylı; lüks loungewear"),
        ("Saranoni", "saranoni.com", "Lush Battaniye", "Utah; ultra yumuşak lush receiving battaniye; bebek + yetişkin; dokunmatik pazarlama"),
        ("Sunday Citizen", "sundaycitizen.co", "Snug Battaniye", "Kristal infüzyon; bambu snug battaniye; TikTok viral; kanepe estetiği"),
        ("Plush by KZ", "shopplush.com", "Özel Tasarım Peluş", "Kişiselleştirilmiş peluş; isim nakış; bebek hediye; Instagram anne estetiği"),
        ("UnHide", "un-hide.com", "Vegan Kürk Battaniye", "Vegan faux fur; Marshmallow battaniye viral; ultra yumuşak; hayvan dostu lüks"),
        ("Lofaris", "lofaris.com", "Fotoğraf Battaniye", "Kişiselleştirilmiş fotoğraf battaniye; hediye pazar; print-on-demand; düğün/bebek"),
        ("PupProtector", "thepawramp.com", "Evcil Hayvan Battaniye", "Su geçirmez + yumuşak; evcil hayvan dostu mobilya koruma; çift kullanım"),
        ("Bedsure", "bedsurehome.com", "Sherpa Battaniye", "Amazon mega satıcı; sherpa + fleece; ultra uygun fiyat; 100M+ satış; TikTok agresif"),
        ("Berkshire Blanket", "berkshireblanket.com", "VelvetLoft Battaniye", "VelvetLoft® teknolojisi; Oprah listesi; eko koleksiyonu; otel tedarikçisi DTC"),
        ("L.L.Bean Blanket", "llbean.com", "Heritage Fleece", "Maine outdoor mirası; dayanıklılık garantisi; DTC geçişi; 100+ yıllık güven"),
        ("Throw Me", "throwme.co.uk", "İngiltere Peluş", "İngiltere; süper yumuşak; Instagram estetik; hediye kutusu ile gönderim"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # GİYİLEBİLİR BATTANİYE
    # ─────────────────────────────────────────────────────────────────────────
    "Giyilebilir Battaniye": [
        ("The Oodie", "theoodie.com", "Oversize Hoodie Battaniye", "Avustralya; giyilebilir battaniye kategorisini patlattı; TikTok'ta 1B+ görüntülenme; 300M+ gelir; influencer ordusu"),
        ("The Comfy", "thecomfy.com", "Orijinal Giyilebilir", "Shark Tank yatırımı; $150M+ gelir; ailece giyim; NFL lisans; hediye favorisi"),
        ("Snuggie", "snuggieblanket.com", "Kollu Battaniye OG", "Giyilebilir battaniye öncüsü; TV reklamlarıyla pop kültür ikonu; retro revival"),
        ("Wearable Blanket Co", "wearableblanket.co", "Premium Giyilebilir", "Sherpa astar; büyük cep; çift fermuar; genç hedef kitle; TikTok pazarlama"),
        ("Oversized Hoodie", "oversizedhoodie.co.uk", "İngiltere Hoodie Blanket", "İngiltere pazar lideri; 20+ renk; aile seti; kış kampanyaları; TikTok UK viral"),
        ("Catalonia", "cataloniablanket.com", "Çok Fonksiyonlu", "Giyilebilir + geleneksel battaniye; cepli; Amazon bestseller; aile boyu"),
        ("Sienna", "sfriennahome.co.uk", "Sherpa Hoodie", "İngiltere; uygun fiyat; Primark rakibi; online-only; 2M+ satış"),
        ("Kudd.ly", "kuddly.co", "Isıtmalı Giyilebilir", "Isıtmalı giyilebilir battaniye; USB şarj; kış lüksü; hediye pazarı"),
        ("Navuo", "navuo.com", "İskandinav Giyilebilir", "İskandinav tasarım; organic cotton; minimal estetik; unisex; sürdürülebilir"),
        ("Halo Home", "halo.com", "Çocuk Giyilebilir", "Güvenli çocuk giyilebilir; SleepSack markasından esinlenen; ebeveyn güvenliği odaklı"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # BEBEK BATTANİYE & KUNDAK
    # ─────────────────────────────────────────────────────────────────────────
    "Bebek Battaniye & Kundak": [
        ("Aden + Anais", "adenandanais.com", "Muslin Kundak", "Muslin kundak kategorisinin yaratıcısı; Kate Middleton bebek battaniyesi; hastane standartı"),
        ("Kyte Baby", "kytebaby.com", "Bambu Bebek Battaniye", "Bambu rayon; ultra yumuşak; momfluencer ordusu; TikTok viral; $100M+ gelir"),
        ("SwaddleMe", "summerinfant.com", "Kolay Kundak", "Cırt-cırtlı kolay kundaklama; yeni ebeveyn dostu; bebek uyku uzmanı onaylı"),
        ("Love to Dream", "lovetodream.com", "Kol Yukarı Kundak", "Patentli kol yukarı tasarım; bebeklerin doğal pozisyonu; Avustralya DTC"),
        ("Copper Pearl", "copperpearl.com", "Stretchy Kundak", "Süper streç muslin; bebek + emzirme örtüsü; çok kullanım; TikTok viral"),
        ("Little Giraffe", "littlegiraffe.com", "Lüks Bebek Battaniye", "Luxe™ kumaş; Hollywood anne tercihi; hediye favorisi; ultra yumuşak"),
        ("Saranoni", "saranoni.com", "Minky Bebek", "Lush receiving battaniye; dokunma terapisi; NICU tercih edilen; yumuşaklık standardı"),
        ("Elegant Baby", "elegantbaby.com", "Muslin + Örgü Bebek", "Premium bebek hediye; muslin + örgü çeşitliliği; boutique dağıtım + DTC"),
        ("Angel Dear", "angeldear.net", "Bambu Muslin", "Bambu muslin; napping battaniye; hayvan temalı; bebek lovey attachment"),
        ("Under the Nile", "underthenile.com", "Organik Mısır Pamuğu", "Mısır organik pamuk; Fair Trade; 20+ yıl organik bebek; allerjen-free"),
        ("Monica + Andy", "monicaandandy.com", "GOTS Organik Bebek", "GOTS sertifikalı; anne kurucusu; hospital bag seti; subscription modeli"),
        ("Loulou Lollipop", "louloulollipop.com", "Tropikal Muslin", "Kanada; tropikal desen; bambu muslin; silikon diş kaşıyıcı + battaniye combo"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # HAVLU & BORNOZ
    # ─────────────────────────────────────────────────────────────────────────
    "Havlu & Bornoz": [
        ("Onsen", "onsentowel.com", "Japon Waffle Havlu", "Japon waffle dokuma; 'Dünya'nın en iyi havlusu' sloganı; hızlı kuruma; minimalist"),
        ("Weezie", "weezietowels.com", "Monogram Havlu", "Kişiselleştirilmiş monogram; piped kenar; long-staple Türk pamuğu; düğün hediye favorisi"),
        ("Brooklinen Towels", "brooklinen.com", "Süper Peluş Havlu", "Yatak markasından banyoya genişleme; Turkish cotton; waffle + plush seçenekleri"),
        ("Parachute Towels", "parachutehome.com", "Organik Havlu", "Organik Türk pamuğu; oeko-tex; minimalist renk paleti; yatak markasından genişleme"),
        ("Baina", "bfrfrain.com", "Avustralya Havlu", "Avustralya; organik pamuk; spa estetiği; cesur renkler; Instagram viral"),
        ("Coyuchi Towels", "coyuchi.com", "Organik Havlu", "GOTS organik; doğal boyama; B-Corp; premium organik fiyat segmenti"),
        ("Kassatex Towels", "kassatex.com", "Türk Pamuğu Havlu", "40 yıl Türk pamuğu deneyimi; otel tedarikçisi; DTC premium; kalın dokuma"),
        ("Laguna Beach Textile", "lfrfragunabeachtextile.com", "Oversize Plaj Havlu", "Dev plaj havlusu; 200x100cm; cabana stripes; California yaşam tarzı"),
        ("Marici", "marfrfricihome.com", "Pestemal", "Türk pestemal; hafif çok kullanımlı; plaj + banyo + piknik; renk çeşitliliği"),
        ("Cariloha Towels", "cariloha.com", "Bambu Havlu", "Bambu fiber; antimikrobiyal; hızlı kuruma; Shark Tank markası genişleme"),
        ("Abyss & Habidecor", "abfrfyssandhabidecor.com", "Portekiz Lüks", "Portekiz üretimi; Egyptian cotton; %100 uzun elyaf; ultra lüks segment"),
        ("Sunday Citizen Towels", "sundaycitizen.co", "Kristal Havlu", "Kristal infüzyon; bambu karışım; 'wellness banyo' konsepti"),
        ("Loom Towels", "loomtowels.com", "Minimalist Havlu", "Tek renk, tek SKU; karar yorgunluğu azaltma; premium Türk pamuğu"),
        ("Mizu", "mizutowel.com", "Gümüş İyon Havlu", "Gümüş iyon teknolojisi; antibakteriyel; 3 gün taze; koku önleme; TikTok viral"),
        ("Dock & Bay", "dockandbay.com", "Hızlı Kuruyan Havlu", "Mikrofiber hızlı kuruma; compact katlanır; seyahat + plaj; 20+ desen; çevreci"),
        ("Sand Cloud", "sandcloud.com", "Türk Pamuk Plaj", "Deniz kaplumbağası koruma; her satışta %10 bağış; boho plaj estetiği; sosyal misyon"),
        ("Slowtide", "slowtide.com", "Sanatçı Kollab Havlu", "Sanatçı iş birliği baskılar; sörfçü kültürü; geri dönüşüm kumaş; San Diego"),
        ("Hammam Linen", "hammamlinen.com", "Türk Havlu", "Türk hammam geleneği; 6-piece set; Amazon bestseller; uygun lüks"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # BANYO AKSESUARLARI
    # ─────────────────────────────────────────────────────────────────────────
    "Banyo Aksesuarları": [
        ("Cadence", "keepyourcadence.com", "Seyahat Kapsül", "Manyetik seyahat kapsülleri; kozmetik düzenleme; özelleştirilebilir; Shark Tank; TikTok viral"),
        ("Kitsch", "mykitsch.com", "Saten Saç Aksesuarı", "Saten yastık kılıfı + scrunchie; saç sağlığı; TikTok'ta 785K satış; $50M+ gelir"),
        ("Stasher", "stfrfrasherbag.com", "Silikon Saklama", "Platinum silikon; yeniden kullanılabilir; B-Corp; plastik azaltma misyonu; banyo depolama"),
        ("Tooletries", "tooletries.com", "Silikon Banyo Organizer", "Vantuzlu silikon organizer; duşta düzen; erkek bakım odaklı; Avustralya DTC"),
        ("Better Living", "betterlivingproducts.com", "Duş Organizer", "Hiclere™ teknolojisi; duş dispenserleri; basit montaj; banyo minimalizmi"),
        ("Yamazaki", "us-?"+"yamazaki.com", "Japon Banyo Düzen", "Japon minimalist; tower serisi; metal organizer; form + fonksiyon; Muji estetiği"),
        ("Open Spaces", "getopenspaces.com", "Renkli Organizer", "Renkli depolama kutuları; banyo + closet + mutfak; dopamine decor; Instagram estetiği"),
        ("SimpleHuman", "simplehuman.com", "Akıllı Banyo", "Sensörlü sabunluk; aydınlatmalı ayna; premium banyo teknolojisi; DTC + perakende"),
        ("Umbra", "umbra.com", "Tasarım Banyo", "Kanada tasarım; 40+ yıl; banyo aksesuarları; modern minimalist; uygun fiyat tasarım"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # DOLAP & GİYSİ ORGANİZERİ
    # ─────────────────────────────────────────────────────────────────────────
    "Dolap & Giysi Organizeri": [
        ("The Home Edit", "thehomeedit.com", "Rainbow Organize", "Netflix show; gökkuşağı renk düzeni; Clea & Joanna; celebrity dolap düzeni; Container Store kollab"),
        ("Open Spaces", "getopenspaces.com", "Renkli Kutu Organizer", "Millennial estetik; renkli storage kutular; dopamine decor; sosyal medya viral"),
        ("Neat Method", "nfrfreatmethod.com", "Profesyonel Organizer", "Lüks ev düzeni hizmeti + ürün satışı; franchise modeli; beyaz + gold estetik"),
        ("Closet Complete", "closetcomplete.com", "Kadife Askı", "Kadife non-slip askılar; 50'li set; Amazon #1 askı; uygun fiyat; dolap dönüşümü"),
        ("SONGMICS", "songmics.com", "Çok Fonksiyonlu Dolap", "TikTok'ta viral dolap sistemleri; uygun fiyat; çelik + kumaş; DIY montaj"),
        ("Elfa (Container Store)", "containerstore.com", "Modüler Dolap Sistemi", "Modüler dolap sistemi; kişiselleştirilebilir; İsveç tasarım; 30+ yıl piyasada"),
        ("Modular Closets", "modularclosets.com", "DIY Dolap Sistemi", "DIY kurulum; online tasarım aracı; özel ölçü; Amerikan yapım; IKEA alternatifi"),
        ("Ease Living", "easeliving.co", "Bambu Organizer", "Bambu ev organizasyonu; sürdürülebilir malzeme; minimal estetik; çekmece bölücü"),
        ("Marie Kondo x Soko", "konmari.com", "KonMari Kutu", "Marie Kondo lisanslı; 'Sparks Joy' felsefesi; kumaş organizasyon kutuları; minimal"),
        ("Neatfreak", "neatfreak.com", "Closet Kit", "Kanada; modüler dolap kiti; uygun fiyat; Walmart + DTC hibrit"),
        ("Whitmor", "whitmor.com", "Ev Organizasyonu", "60+ yıl; geniş ürün yelpazesi; askı + kutu + raf; uygun fiyat lider"),
        ("Sorbus", "sfrfrorbushome.com", "Çekmece Organizer", "Amazon bestseller; şeffaf çekmece bölücü; TikTok restocking video favorisi"),
        ("Mdesign", "mfrfrdesignhomedecor.com", "Plastik Organizer", "Şeffaf akrilik; mutfak + banyo + dolap; uygun fiyat; Amazon mega marka"),
        ("StorageWorks", "storageworks.com", "Kumaş Saklama Kutusu", "El örgüsü görünümlü kumaş kutular; raflar için; doğal estetik; sıcak tonlar"),
        ("IKEA Skubb", "ikea.com", "Skubb Organizer", "İkonik IKEA dolap organizeri; global standart; beyaz kumaş kutular"),
        ("Squared Away", "squaredaway.com", "Bed Bath DTC", "Bed Bath & Beyond'un DTC organizasyon markası; bamboo + kumaş; modern estetik"),
        ("Smart Store", "smartstore.se", "İsveç Kutu Sistemi", "İsveç tasarım; modüler kutular; şeffaf + renkli; clip-on kapak; stackable"),
        ("Foldable Closet", "foldableorganizer.com", "Katlanır Dolap", "Kiracılar için; montaj yok; katlanır dolap sistemi; taşınabilir; öğrenci evi"),
        ("Hanging Shelves Co", "hangingshelvesco.com", "Asılabilir Raf", "Dolap çubuğuna asılan raflar; üniversite yurdu; küçük alan çözümü"),
        ("Decluttr Me", "decluttrme.com", "Kumaş Organizasyon", "İngiltere; keten look kumaş kutular; etiketlenebilir; Mrs Hinch estetiği"),
        ("Love It Store It", "loveitfrfrstoreit.com", "Çocuk Organizer", "Çocuk odası; eğlenceli desenler; katlanır kutu; oyuncak + giysi"),
        ("Zulay Home", "zulaykitchen.com", "Çok Kullanımlı", "Mutfaktan dolaba geçiş; silikon + bambu; Amazon DTC; uygun fiyat kalite"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # ÇEKMECE DÜZENLEYİCİ
    # ─────────────────────────────────────────────────────────────────────────
    "Çekmece Düzenleyici": [
        ("DrawerDividers.com", "drawerdividers.com", "Ayarlanabilir Bölücü", "Bambu ayarlanabilir çekmece bölücü; TikTok #drawerdividers 7.9M görüntü; estetik düzen"),
        ("KonMari Boxes", "konmari.com", "Marie Kondo Kutu", "Shojin kumaş kutular; her eşyaya bir yer felsefesi; dikey katlama için özel boyutlar"),
        ("Sorbus Drawer", "sfrfrorbushome.com", "Honeycomb Bölücü", "Bal peteği iç çamaşırı organizeri; TikTok restocking viral; 15 bölme; uygun fiyat"),
        ("Bambüsi", "bambusi.com", "Bambu Çekmece Seti", "Bambu set; makyaj + iç çamaşırı + çorap; genişletilebilir; çevreci malzeme"),
        ("SimpleHouseware", "simplehouseware.com", "Kumaş Bölücü Set", "24 hücreli iç çamaşırı organizeri; katlanabilir; Amazon #1 bölücü"),
        ("Joseph Joseph Drawer", "josephjoseph.com", "İngiliz Tasarım", "Akıllı çekmece düzeni; kompakt tasarım; mutfak + banyo; İngiliz mühendisliği"),
        ("Lifewit", "lifewit.com", "Kumaş Çekmece Organizeri", "Gri kumaş estetik; kıyafet katlanabilir bölme; TikTok'ta düzen videoları; uygun fiyat"),
        ("SpaceAid", "spaceaid.com", "Bambu + Akrilik", "Mutfak çekmece düzeni; baharat raf + çekmece insert; combo setler; Amazon DTC"),
        ("Purse Bling", "pursebling.com", "Çanta İçi Organizer", "Çanta içi bölücü; luxury bag insert; Birkin + LV insert; niş pazar lideri"),
        ("Poppin", "poppin.com", "Ofis Çekmece", "Renkli ofis düzeni; çekmece organizeri; work-from-home estetiği; millennial ofis"),
        ("SmartDesign", "smartdesign.com", "Stackable Çekmece", "Stackable mesh çekmece; dolap altı; banyo tezgah altı; modüler ekleme"),
        ("YouCopia", "youcopia.com", "Genişleyen Çekmece", "Genişletilebilir çekmece organizeri; baharat + makyaj + ilaç; patent tasarım"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # ÇAMAŞIR & KİRLİ SEPETİ
    # ─────────────────────────────────────────────────────────────────────────
    "Çamaşır & Kirli Sepeti": [
        ("Steele Canvas", "steelecanvas.com", "Endüstriyel Sepet", "1921'den beri; industrial chic; ağır hizmet canvas; otel tedarikçisi DTC'ye döndü"),
        ("West Elm Laundry", "westelm.com", "Modern Çamaşır", "Modern tasarım; hasır + kumaş; ev dekoru parçası olarak tasarlanmış"),
        ("Yamazaki Tower", "us.yamazaki-?"+"?"+"home.com", "Japon Çamaşır", "İnce metal çerçeve; küçük alan; Japon minimalizm; form + fonksiyon"),
        ("Brabantia", "brabantia.com", "Hollanda Çamaşır", "Bo laundry bin; renk çeşitliliği; 10 yıl garanti; Hollanda tasarım; sürdürülebilir"),
        ("Seville Classics", "sevilleclassics.com", "Rolling Sepet", "Tekerlekli çamaşır sepeti; 3 bölmeli; ütü masası combo; uygun fiyat fonksiyon"),
        ("Bag-all", "bag-all.com", "Baskılı Çamaşır Torbası", "New York; esprili baskılı organizasyon torbaları; seyahat çamaşır; hediye market"),
        ("Tusk College", "tuskcollege.com", "Kampüs Çamaşır", "Üniversite öğrencisi odaklı; taşınabilir çamaşır seti; dorm essentials"),
        ("CleverMade", "clevermade.com", "Katlanır Sepet", "Katlanır teknoloji; araba bagajı + ev; çok kullanım; kompakt depolama"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # MUTFAK ORGANİZASYONU
    # ─────────────────────────────────────────────────────────────────────────
    "Mutfak Organizasyonu": [
        ("Caraway", "carawayhome.com", "Toksik-Free Mutfak", "Seramik kaplama tava + organizer seti; Pinterest renk paleti; 'organize mutfak' viral; $100M+ gelir"),
        ("Our Place", "fromourplace.com", "Çok Fonksiyonlu Tava", "Always Pan™ 8-in-1; mutfak minimalizmi; renkli tasarım; $200M+ gelir; Selena Gomez kollab"),
        ("Open Spaces", "getopenspaces.com", "Renkli Mutfak Kutu", "Pantry organizasyonu; renkli kutular; etiket sistemi; dopamine decor mutfak"),
        ("W&P", "wandpdesign.com", "Silikon Mutfak", "Porter Bowl; katlanır silikon; seyahat + mutfak; sürdürülebilir; Brooklyn tasarım"),
        ("YouCopia", "youcopia.com", "Baharat Düzenleyici", "SpiceStack® kayan raf; mutfak çekmece organizasyonu; patentli tasarım; Amazon #1"),
        ("SimpleHuman Kitchen", "simplehuman.com", "Çöp + Organizer", "Sensörlü çöp kutusu; custom fit liner; tension organizer; premium mutfak"),
        ("Joseph Joseph", "josephjoseph.com", "Akıllı Mutfak", "IW çöp kutusu; DrawerStore; kompakt tasarım; İngiliz inovasyon; Red Dot ödülü"),
        ("Yamazaki Tower Kitchen", "us-?"+"yamazaki.com", "Japon Mutfak Düzen", "Tower serisi; manyetik + asılabilir; buzdolabı yanı organizer; minimal metal"),
        ("StoveShelf", "stoveshelf.com", "Ocak Rafı", "Ocak arkası manyetik raf; TikTok viral; küçük mutfak çözümü; ABD üretimi"),
        ("Made by Gather", "madebygather.com", "Yemek Hazırlık", "Mise en place tepsi seti; yemek hazırlık organizasyonu; şef mutfağı estetiği"),
        ("Kozy Kitchen", "kozykitchenco.com", "Bambu Mutfak Set", "Bambu organizasyon seti; çekmece bölücü + tezgah organizer; natural estetik"),
        ("iDesign", "idesignlivesimply.com", "Buzdolabı Organizer", "Şeffaf buzdolabı kutuları; The Home Edit kollab; TikTok fridge makeover viral"),
        ("mDesign Pantry", "mfrfrdesignhomedecor.com", "Pantry Organizasyonu", "Clear bin seti; pantry label + kutu; TikTok restocking ASMR favoriti"),
        ("Brightroom (Target DTC)", "target.com", "Target Organizer", "Target'ın DTC organizasyon markası; renkli + uygun fiyat; GenZ hedef"),
        ("Totally Bamboo", "totallybamboo.com", "Bambu Kesme + Düzen", "California; %100 bambu; kesme tahtası + mutfak organizeri; 25 yıl"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # BANYO ORGANİZASYONU
    # ─────────────────────────────────────────────────────────────────────────
    "Banyo Organizasyonu": [
        ("Tooletries", "tooletries.com", "Silikon Duş Organizer", "Patentli silikon vantuz; duş duvarına yapışan organizer; Avustralya startup; erkek bakım odaklı"),
        ("Better Living", "betterlivingproducts.com", "Duş Dispenser", "CLEVER® duş dispenser sistemi; şampuan + sabun düzeni; otel estetiği evde"),
        ("Command Bath", "command.com", "Hasarsız Montaj", "3M teknolojisi; kiracı dostu; hasarsız banyo rafı; çıkarılabilir yapıştırıcı"),
        ("Umbra Bath", "umbra.com", "Tasarım Banyo Organizer", "Kanada tasarım; 40+ yıl; banyo aksesuar seti; modern minimalist"),
        ("Zenna Home", "zennahome.com", "Duş Raf Sistemi", "Tension pole duş rafı; NeverRust® paslanmaz; banyo köşe çözümleri"),
        ("iDesign Bath", "idesignlivesimply.com", "Clear Banyo Kutu", "Şeffaf akrilik banyo organizeri; The Home Edit estetik; stackable"),
        ("Yamazaki Bath", "us-?"+"yamazaki.com", "Japon Banyo", "Tower serisi banyo; manyetik sabunluk; havluluk; minimal metal tasarım"),
        ("Bath Bliss", "bathbliss.com", "Duş Caddy", "Asılabilir duş caddy; rust-proof; uygun fiyat; 8 bölme"),
        ("Simplehuman Bath", "simplehuman.com", "Sensörlü Banyo", "Sensörlü ayna; sabun pompası; çöp kutusu; banyo teknolojisi; premium"),
        ("Navaris", "navfrfrraris.com", "Bambu Banyo Raf", "Bambu banyo rafı; İskandinav + Japon estetik karışımı; uygun fiyat"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # AYAKKABI & AKSESUAR DÜZENLEYİCİ
    # ─────────────────────────────────────────────────────────────────────────
    "Ayakkabı & Aksesuar Düzenleyici": [
        ("The Container Store", "containerstore.com", "Modüler Ayakkabılık", "Modüler Drop-Front kutu; şeffaf ayakkabı kutusu trendi; The Home Edit kollab"),
        ("Sole Stacks", "solfrfrestacks.com", "Sneaker Display", "Sneakerhead kültürü; LED aydınlatmalı kutu; drop-front; koleksiyoncu hedef kitle"),
        ("Crep Protect", "cfrfrrepprotect.com", "Sneaker Bakım + Kutu", "Sneaker koruma spreyi; premium saklama kutusu; sneaker kültürü; TikTok viral"),
        ("SONGMICS Shoe", "songmics.com", "Kapaklı Ayakkabılık", "Toz geçirmez ayakkabılık; DIY montaj; uygun fiyat; 10+ kat; küçük alan"),
        ("Lynk", "lynkfrfrinc.com", "Over-Door Organizer", "Kapı arkası ayakkabılık; patent tasarım; 36 çift kapasite; alan tasarrufu"),
        ("Simplify", "simplify-?"+"home.com", "Alttan Yatak Saklama", "Yatak altı ayakkabı çekmecesi; tekerlekli; şeffaf kapak; görünmez depolama"),
        ("Stackers", "stackers.com", "Takı Organizeri", "İngiltere; modüler takı kutusu; kadife iç; stackable; hediye pazarı lideri; estetik"),
        ("Nordgreen Watch Box", "nordgreen.com", "Danimarkalı Saat Kutusu", "Danimarka tasarım; vegan deri; minimalist saat saklama; hediye seti"),
        ("TomCare", "tomcareshop.com", "Cube Organizer", "Modüler küp raf; her bütçeye uygun; DIY montaj; çocuk odası + dolap"),
        ("Iris USA", "irisusainc.com", "Şeffaf Saklama", "Japon tasarım; şeffaf plastik kutu; stackable; 50+ yıl; garage to closet"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # SEYAHAT ORGANİZERİ & POUCH
    # ─────────────────────────────────────────────────────────────────────────
    "Seyahat Organizeri & Pouch": [
        ("Away", "awaytravel.com", "Akıllı Bavul", "USB şarjlı bavul; Instagram-first marka; $1.4B değerleme; seyahat kültürü yarattı"),
        ("Cadence", "keepyourcadence.com", "Manyetik Kapsül", "Manyetik kozmetik kapsülleri; TSA uyumlu; özelleştirilebilir; Shark Tank; TikTok viral"),
        ("Dagne Dover", "dagnedover.com", "Neopren Organizer", "Neopren çanta + organizasyon; anne çantası bestseller; mesh iç bölmeler"),
        ("Paravel", "tourparavel.com", "Sürdürülebilir Bavul", "Geri dönüşüm malzeme; karbon-ofset gönderim; vintage estetik; luxury DTC"),
        ("Peak Design", "peakdesign.com", "Tech Organizer", "Tech Pouch™; kamera + günlük; Kickstarter rekoru; modüler tasarım"),
        ("Monos", "monos.com", "Minimalist Bavul", "Kanada; Japon minimalizmi; polycarbonate; 'designed for the journey' felsefesi"),
        ("Stoney Clover Lane", "stfrfroneycloverlane.com", "Kişiselleştirilmiş Pouch", "Yama/patch özelleştirme; renkli nylon; celebrity favorisi; custom organizasyon"),
        ("Calpak", "calpaktravel.com", "Packing Cube Set", "Rengarenk packing cube; TikTok viral; aile seyahati; Instagram estetik"),
        ("Bag-all", "bag-all.com", "Organizing Bags", "Baskılı organizasyon torbaları; 'Socks', 'Underwear' yazılı; New York; hediye"),
        ("Bellroy", "bellroy.com", "İnce Cüzdan + Pouch", "Avustralya; ince tasarım; premium deri; tech kit; seyahat düzeni"),
        ("July", "july.com", "Avustralya Bavul", "Avustralya DTC bavul; kişiselleştirme; 100 gün deneme; pastel renkler"),
        ("Roam Luggage", "rofrfrnam.com", "Özel Renk Bavul", "1 milyondan fazla renk kombinasyonu; custom bavul; DTC özelleştirme lideri"),
        ("Eagle Creek", "eaglecreek.com", "Pack-It System", "Pack-It® organizasyon sistemi; 40+ yıl seyahat uzmanlığı; kompresyon küp"),
        ("Truffle", "trfrfrfruffle.com", "Clarity Pouch", "Şeffaf + renkli kozmetik çantası; hediye favorisi; organize seyahat; TikTok viral"),
        ("Lo & Sons", "loandsons.com", "Çok Bölmeli", "Pearl crossbody; O.G. seyahat çantası; anne-kız kurucular; fonksiyonel lüks"),
        ("Nomad Lane", "nomadlane.com", "Benton Weekender", "Seyahat bavuluna geçebilen çanta; organizasyon cepleri; iş + kişisel"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # MAKYAJ & KOZMETİK ORGANİZERİ
    # ─────────────────────────────────────────────────────────────────────────
    "Makyaj & Kozmetik Organizeri": [
        ("Etoile Collective", "etoilecollective.com", "Akrilik Makyaj Düzeni", "Lüks akrilik organizasyon; influencer vanity seti; altın detaylar; Instagram viral"),
        ("Luvo Store", "luvostore.com", "Dönen Organizer", "360° dönen makyaj organizeri; bambu + akrilik; TikTok viral; uygun fiyat"),
        ("Sorbus Makeup", "sfrfrorbushome.com", "Büyük Akrilik Organizer", "Büyük boy akrilik; çekmeceli; lipstick holder; Amazon #1 makyaj organizeri"),
        ("Pocobello", "pocobello.com", "İtalyan Deri Pouch", "İtalyan deri kozmetik çantası; el yapımı; minimalist; lüks hediye"),
        ("Rownyeon", "rownyeon.com", "Profesyonel Makyaj Çantası", "MUA (makyaj artisti) odaklı; bölmeli organizasyon; profesyonel seri"),
        ("Béis Beauty", "bfrfreis.com", "Travel Vanity", "Shay Mitchell markası; seyahat makyaj çantası; aynalı; kanca ile asılabilir"),
        ("Tower 28 Organizer", "tower28beauty.com", "Clean Beauty Düzen", "Clean beauty markasının organizer seti; renkli; plaja uygun"),
        ("Stage Box", "stfrfrageboxco.com", "LED Aynalı Organizer", "LED ışıklı makyaj kutusu; taşınabilir; professional look; TikTok viral"),
        ("Vanity Planet", "vanityplanet.com", "Güzellik Cihaz Organizer", "Güzellik cihazları + saklama çözümleri; şarj dock; akrilik düzen"),
        ("Caboodles", "caboodles.com", "Retro Makyaj Kutusu", "80s retro revival; kademeli açılan kutu; nostalji pazarlama; GenZ retro trendli"),
        ("The Daily Edited", "thedailyedited.com", "Monogram Kozmetik", "Kişiselleştirilmiş monogram; Avustralya; saffiano deri; hediye odaklı"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # PERDE & TÜL
    # ─────────────────────────────────────────────────────────────────────────
    "Perde & Tül": [
        ("Loom Decor", "loomdecor.com", "Özel Dikim Perde", "Online özel dikim; 1000+ kumaş seçeneği; ücretsiz kumaş numune; DTC custom"),
        ("The Shade Store", "theshadestore.com", "Premium Perde", "Custom yapım; ücretsiz ev ölçümü; 1500+ malzeme; 10 günde teslim"),
        ("Half Price Drapes", "halfpricedrapes.com", "İndirimli Perde", "Fabrikadan direkt; %50 daha uygun; geniş koleksiyon; DTC fiyat avantajı"),
        ("West Elm Curtains", "westelm.com", "Modern Perde", "Mid-century modern estetik; organik pamuk seçenekleri; Fair Trade"),
        ("Lush Decor", "lushdecor.com", "Ruffled Perde", "Fırfırlı + boho perde; genç estetik; Instagram dekoru; uygun premium"),
        ("Pottery Barn Curtains", "potterybarn.com", "Klasik Perde", "Amerikan klasik; custom drape; monogram seçeneği; waffle + linen"),
        ("Deconovo", "deconovo.com", "Blackout Perde", "Amazon #1 blackout; 40+ renk; termal yalıtım; uygun fiyat kalite"),
        ("NICETOWN", "nicetown.com", "Karartma Perde", "Amazon bestseller karartma; enerji tasarrufu; geniş beden seçeneği"),
        ("Marjorie & Marjorie", "marjofrfrrieandmarjorie.com", "Bağımsız Perde Tasarım", "Küçük batch; el baskı keten; artisan perde; Etsy'den DTC'ye geçiş"),
        ("Inside Weather", "insideweather.com", "Custom Mobilya + Perde", "Online özelleştirme; ABD üretimi; kumaş seçimi + perde; hızlı teslimat"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # HALI & KİLİM
    # ─────────────────────────────────────────────────────────────────────────
    "Halı & Kilim": [
        ("Ruggable", "ruggable.com", "Yıkanabilir Halı", "Makine yıkanabilir halı; 2 parçalı sistem; TikTok viral; evcil hayvan dostu; $100M+ gelir"),
        ("Tumble", "tumble.com", "Yıkanabilir Halı v2", "Ruggable alternatifi; daha yumuşak his; döndürülebilir desen; Shark Tank"),
        ("Revival Rugs", "revivalrugs.com", "Vintage Halı", "Türkiye + Fas vintage halılar; tekil parçalar; hikaye kartı ile gönderim"),
        ("Loloi Rugs", "loloirugs.com", "Tasarımcı Halı", "Chris Loves Julia + Amber Lewis kollab; Instagram tasarımcı iş birlikleri"),
        ("Miss Amara", "missamara.com", "AI Halı Eşleştirme", "Avustralya; AR oda önizleme; AI stil eşleştirme; teknoloji + halı birleşimi"),
        ("Brooklinen Rug", "brooklinen.com", "Yatak Markası Halı", "Yatak markasından halıya genişleme; hand-tufted; organik pamuk seçeneği"),
        ("Hook & Loom", "hookandloom.com", "Eko Halı", "Geri dönüşüm pamuk; el dokuması; kimyasal yok; çevreci halı DTC"),
        ("House of Moroccan Rugs", "housfrfreofmoroccanrugs.com", "Fas El Dokuması", "Berber kadın zanaatkarlar; her halı tekil; kültürel miras; etik ticaret"),
        ("Lorena Canals", "lorenacanals.com", "Yıkanabilir Çocuk", "İspanya; makine yıkanabilir çocuk halısı; organik boyalar; Montessori uyumlu"),
        ("nuLOOM", "nuloom.com", "Trend Halı", "Geniş çeşitlilik; jüt + yün + sentetik; uygun fiyat; Amazon + DTC; trend desenler"),
        ("Boutique Rugs", "boutiquerugs.com", "Online Halıcı", "DTC only; 10K+ desen; ücretsiz kargo ve iade; agresif Instagram reklamları"),
        ("Armadillo", "armadillo-co.com", "Fair Trade Halı", "Avustralya; Fair Trade; Hindistan el dokuması; 15 yıl garanti; lüks doğal"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # MASA ÖRTÜSÜ & RUNNER
    # ─────────────────────────────────────────────────────────────────────────
    "Masa Örtüsü & Runner": [
        ("Heather Taylor Home", "heathertaylorhome.com", "Gingham Runner", "LA; gingham desen; el boyama; Instagram sofra estetiği; renkli mutfak"),
        ("Maison de Vacances", "maisfrfrondfrfrevacances.com", "Fransız Keten", "Provence keten; yıkanmış yumuşak doku; doğal renkler; Fransız kır evi estetiği"),
        ("Rough Linen Table", "roughlinen.com", "Ham Keten Sofra", "Yıkanmamış ham keten; zamanla güzelleşen; California artisan; her parça tekil"),
        ("Color Me Courtney", "colormecourtney.com", "Renkli Sofra", "Rengarenk sofra örtüsü; influencer marka; party styling; dopamine decor"),
        ("Celina Mancurti", "celinamancurti.com", "El Boyama Keten", "Kolombiya; el boyama keten; tropikal renkler; sanat eseri sofra; small batch"),
        ("Summerill & Bishop", "sumfrmfrerfrfrillandbishop.com", "Lüks İngiliz Sofra", "İngiltere; el boyama; lüks keten; celebrity chef tercihi; £200+ peçete"),
        ("Creative Women", "creativewomen.org", "Etiyopya Sofra", "Etiyopya kadın zanaatkarları; el dokuması; Fair Trade; sosyal etki markası"),
        ("Sir Madam", "sirmadam.com", "Vintage İlham Sofra", "India block print; vintage desen modern kullanım; natural boya; artisan"),
        ("Table & Dine", "tableanddine.com", "Online Sofra Seti", "Komple sofra seti; renk koordineli; peçete + runner + örtü; düğün hediye"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # DEKORATİF TEKSTİL & MAKROME
    # ─────────────────────────────────────────────────────────────────────────
    "Dekoratif Tekstil & Makrome": [
        ("Modern Macrame", "modernmacrame.com", "Makrome Duvar", "Büyük boy makrome duvar süsü; boho chic; evlilik töreni backdrop; Etsy'den DTC'ye"),
        ("Sage & Twine", "sfrfrageandtwine.com", "Makrome Sepet", "Makrome duvar sepeti; bitki askısı; depolama + dekor; minimal boho"),
        ("Fiber Art by Maryanne Moodie", "marfrfryannemoodie.com", "Weaving Sanat", "El dokuması duvar sanatı; Brooklyn artisan; tekstil sanat; $500+ parçalar"),
        ("The Citizenry Textiles", "the-citizenry.com", "Global Zanaatkar", "Dünya çapında zanaatkarlar; el dokuma battaniye; halı; dekor tekstili"),
        ("Minna Goods", "mfrfrinna.com", "Latin Amerika Dokuma", "Guatemala + Meksika el tezgahı; modern desen; Fair Trade; sürdürülebilir"),
        ("Somewhere in the World", "somewherfrfre.com", "Seyahat İlham Tekstil", "Seyahat deneyimlerinden ilham; global zanaat teknikleri; hikaye odaklı"),
        ("Quiet Town", "quiettownhome.com", "Duş Perdesi Sanat", "Canvas duş perdesi; sanat eseri banyo; Brooklyn tasarım; galeri estetiği"),
        ("Cold Picnic", "coldpicnic.com", "Tufted Tekstil", "Hand-tufted halı + yastık; pop art; vücut formları; Brooklyn sanatçı kolektif"),
        ("Tantuvi", "tantuvi.com", "Hint El Baskı", "El baskı (block print) perdeler ve yastıklar; Hindistan zanaatı; custom renk"),
        ("Joinery", "thejoinery.com", "Brezilya Dokuma", "Brezilya yerli halkları ile işbirliği; el dokuma; kültürel koruma; etik moda + ev"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # UYKU TEKNOLOJİSİ & AKILLI YATAK
    # ─────────────────────────────────────────────────────────────────────────
    "Uyku Teknolojisi & Akıllı Yatak": [
        ("Eight Sleep", "eightsleep.com", "Akıllı Yatak Pedi", "Pod teknolojisi; sıcaklık ayarlı yatak; uyku takibi; $500M+ değerleme; biohacker favorisi"),
        ("ChiliSleep (Sleepme)", "sleep.me", "Su Soğutmalı Yatak", "OOLER ve Dock Pro; su bazlı soğutma/ısıtma; biyohacker; Andrew Huberman onayı"),
        ("Bedjet", "bedjet.com", "Hava Soğutmalı Yatak", "Hava bazlı sıcaklık kontrolü; çift taraflı; her partner farklı sıcaklık"),
        ("Withings Sleep", "withings.com", "Uyku Sensörü", "Yatak altı uyku izleme sensörü; uyku apnesi tespiti; IoT entegrasyon"),
        ("Oura Ring + Sleep", "ouraring.com", "Uyku Yüzüğü", "Uyku kalitesi yüzüğü; NBA oyuncuları; biohacking; $2.5B değerleme"),
        ("Hatch Restore", "hatch.co", "Uyku Işığı", "Sunrise alarm + gece ışığı + ses makinesi; bebek + yetişkin; uyku ritüeli yaratma"),
        ("Muse Sleep", "musesleep.com", "Jel Memory Foam", "3 katmanlı jel memory foam; sıkıştırılmış gönderim; uygun fiyat premium"),
        ("Tempur-Pedic Smart", "tempurpedic.com", "Ergo Smart Karyola", "Motorlu ayarlanabilir karyola; anti-horlama; zero gravity pozisyon"),
        ("Brentwood Home", "brentwoodhome.com", "Organik Şilte", "LA üretimi; organik pamuk + yün; yoga mat + şilte; B-Corp başvurusu"),
        ("Helix Sleep", "helixsleep.com", "Kişisel Şilte", "Uyku quiz'i ile eşleşen şilte; Midnight modeli bestseller; 100 gece deneme"),
        ("Birch Living", "birchliving.com", "Doğal Şilte", "Helix'in doğal alt markası; organik yün + lateks; GreenGuard Gold; eko uyku"),
        ("Awara Sleep", "awarasleep.com", "Sri Lanka Lateks", "Sri Lanka doğal lateks; organik pamuk; Rainforest Alliance; %0 karbon"),
        ("Bear Mattress", "bearmattress.com", "Sporcu Şilte", "Celliant® kılıf; kas toparlanma; NFL onaylı; elit sporcu pazarlama"),
        ("Lull Mattress", "lfrfrull.com", "DTC Şilte", "3 katmanlı foam; sıkıştırılmış kutu; 365 gece deneme; uygun fiyat"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # ŞİLTE & YATAK PEDİ
    # ─────────────────────────────────────────────────────────────────────────
    "Şilte & Yatak Pedi": [
        ("ViscoSoft", "viscosoft.com", "Memory Foam Yatak Pedi", "4-inch topper; Amazon #1; otel konforuna dönüştürme; pillow-top his; uygun fiyat"),
        ("Tempur-Topper", "tempurpedic.com", "TEMPUR Yatak Pedi", "NASA TEMPUR malzeme; mevcut yatağa premium upgrade; basınç azaltma"),
        ("Sleep Innovations", "sleepinnovations.com", "Jel Topper", "Jel memory foam; soğutma teknolojisi; Amazon DTC; uygun fiyat kalite"),
        ("Lucid Mattress Topper", "lucidfrmfrattrfrfrfrfress.com", "Lavanta Jel", "Lavanta infüzyon jel foam; aromaterapi uyku; Amazon bestseller"),
        ("Molecule", "onmolecule.com", "Air-Engineered Yatak Pedi", "Air-Engineered™ foam; sporcu odaklı; Russell Wilson onaylı; soğutma"),
        ("Naturepedic", "naturepedic.com", "Organik Yatak Pedi", "GOTS organik; çocuk + yetişkin; EWG onaylı; kimyasal-free; allerjen-free"),
        ("My Green Mattress", "mygreenmattress.com", "Yeşil Yatak Pedi", "GOTS + GOLS; organik yün + lateks; Chicago üretimi; aile işletmesi"),
        ("Saatva Topper", "saatva.com", "Lüks Yatak Pedi", "Talalay lateks topper; organik pamuk kılıf; 180 gece deneme"),
        ("Purple Topper", "purple.com", "GelFlex Grid Topper", "Hyper-Elastic Polymer™; sıcaklık nötr; basınç dağılımı; yatağa upgrade"),
        ("Layla Topper", "laylasleep.com", "Bakır Memory Foam", "Copper-infused memory foam; antimikrobiyal; 2-inch ve 4-inch seçenekler"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # BEBEK YATAĞI & UYKU SETİ
    # ─────────────────────────────────────────────────────────────────────────
    "Bebek Yatağı & Uyku Seti": [
        ("Newton Baby", "newtonbaby.com", "Nefes Alabilir Şilte", "Wovenaire® core; %100 nefes alabilir; boğulma riski azaltma; CPSC onaylı; GREENGUARD Gold"),
        ("Naturepedic Baby", "naturepedic.com", "Organik Bebek Şilte", "GOTS organik; su geçirmez yüzey; kimyasal-free; pediatrist onaylı"),
        ("Lullaby Earth", "lullabyearth.com", "Eko Bebek Şilte", "Naturepedic'in uygun fiyat alt markası; organik pamuk; GREENGUARD Gold"),
        ("Nook Sleep", "nooksleep.com", "Pebble Şilte", "Pebble air™ teknolojisi; zinc-infused kılıf; organik; uluslararası ödül"),
        ("SNOO", "happiestbaby.com", "Akıllı Beşik", "Dr. Harvey Karp; otomatik sallayan beşik; ağlama algılama; kiralama modeli; $1695 fiyat"),
        ("Cradlewise", "cradlewise.com", "AI Beşik", "AI destekli; ağlama tahmini; otomatik sallanma; kamera + monitor; Hindistan startup"),
        ("Babyletto", "babyletto.com", "Modern Bebek Mobilya", "Mid-century modern; GREENGUARD Gold; dönüştürülebilir karyola; 3-in-1"),
        ("Oeuf", "oeufnyc.com", "Eko Tasarım Bebek", "Brooklyn tasarım; European birch; sürdürülebilir; Sparrow karyola ikonik"),
        ("Nestig", "nesfrfrtig.com", "Montessori Karyola", "Dönüştürülebilir + Montessori uyumlu; yer yatağına çevrilebilen; modern tasarım"),
        ("Stokke Home", "stokke.com", "İskandinav Bebek Yatak", "Norveç tasarım; Sleepi oval beşik; büyüme ile genişleyen; ikonik tasarım"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # ANNE & EMZİRME YASTIĞI
    # ─────────────────────────────────────────────────────────────────────────
    "Anne & Emzirme Yastığı": [
        ("Boppy", "boppy.com", "Orijinal Emzirme Yastığı", "25+ yıl; orijinal C-şekilli emzirme yastığı; hastane standartı; en çok satılan"),
        ("My Brest Friend", "mybfrestfriend.com", "Firm Emzirme", "Sert destek; bel kemerli; emzirme danışmanı tercih; ergonomik tasarım"),
        ("Frida Mom", "fridababy.com", "Anne Bakım Yastık", "Postpartum odaklı; anne bakım ürünleri; tabusuz pazarlama; emzirme desteği"),
        ("Snuggle Me Feeding", "snugglemeorganic.com", "Organik Emzirme", "GOTS organik; emzirme + lounging; doğal malzeme; anne topluluğu"),
        ("Bbhugme", "bbhugme.com", "Hamile Yastığı", "Norveç tasarım; fizyoterapist geliştirdi; ayarlanabilir sertlik; hamile + emzirme"),
        ("Pharmedoc", "pharmedoc.com", "U-Şekil Hamile", "U-şekilli tam vücut yastık; Amazon #1 hamile yastığı; jersey pamuk kılıf"),
        ("Queen Rose", "queenfrfrrose.com", "C-Şekil Hamile", "C-şekilli büyük boy; 55 inch; çok pozisyonlu destek; uygun fiyat"),
        ("Leachco", "leachco.com", "Snoogle Yastık", "Snoogle® patentli C-şekil; 20+ yıl; anne klasiği; hamile yastık öncüsü"),
        ("Momcozy", "momcozy.com", "Anne Teknoloji", "Giyilebilir göğüs pompası; emzirme yastığı; TikTok viral; anne tech ekosistemi"),
        ("Ergobaby Nursing", "ergobaby.com", "Natural Curve Yastık", "Natural Curve™ emzirme yastığı; bebek taşıyıcı markasından genişleme"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # EV KOKUSU & MUM
    # ─────────────────────────────────────────────────────────────────────────
    "Ev Kokusu & Mum": [
        ("Boy Smells", "boysmells.com", "Genderless Mum", "Gender-neutral koku; renkli balmumu; queer kurucular; Kacey Musgraves kollab"),
        ("Otherland", "otherland.com", "Sanat Mum", "Her mum bir sanat eseri; custom cam; Instagram-first; görsel storytelling"),
        ("Homesick", "homesick.com", "Nostalji Mum", "Eyalet/şehir/anı kokulara; nostalji pazarlama; hediye favorisi; unique konsept"),
        ("P.F. Candle Co", "pfcandleco.com", "LA Artisan Mum", "LA; soya balmumu; amber cam ikonik; terra cotta; indie favori"),
        ("Byredo Candle", "byredo.com", "Lüks Niche Mum", "İsveç minimalist; $85+ mum; lüks niş koku; influencer staple"),
        ("Apotheke", "apfrfrothekeco.com", "Brooklyn Mum", "Brooklyn üretimi; clean-burning soya; charcoal cam; modern apothecary estetiği"),
        ("Diptyque DTC", "diptyque.com", "Fransız Lüks Mum", "Baies ikonik; 60+ yıl; Fransız lüks; DTC online genişleme"),
        ("Voluspa", "voluspa.com", "Coconut Wax Mum", "Proprietary coconut wax blend; dekoratif cam; hediye segmenti; Anthropologie hit"),
        ("Brooklyn Candle Studio", "brooklyncandlestudio.com", "Minimalist Mum", "Minimalist etiket; soya; NYC; düğün custom; küçük batch"),
        ("Keap", "kefrfrapcandles.com", "Subscription Mum", "Abonelik modeli; mevsimsel kokular; artisan soya; ipek kılıf ambalaj"),
        ("Maison Louis Marie", "maisonlouismarie.com", "Botanik Mum", "Fransız botanik; No.04 Bois de Balincourt cult hit; parfüm + mum"),
        ("Vitruvi", "vitruvi.com", "Taş Difüzör", "Porselen taş difüzör; esansiyel yağ; minimalist wellness; Instagram viral"),
        ("Nest Fragrances", "nestfragrances.com", "Ev Kokusu Ekosistemi", "Grapefruit ikonik; reed diffuser + mum + spray; hediye market lideri"),
        ("Capri Blue", "capriblue.com", "Volcano Mum", "Volcano koku cult hit; mavi cam ikonik; Anthropologie exclusive DTC genişleme"),
        ("Loewe Home Scents", "loewe.com", "Lüks Tasarımcı Mum", "Tomato Leaves ikonik; sebze/meyve ilham; tasarımcı ev kokusu; Insta-worthy"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # AROMATERAPİ & DİFÜZÖR
    # ─────────────────────────────────────────────────────────────────────────
    "Aromaterapi & Difüzör": [
        ("Vitruvi", "vitruvi.com", "Porselen Difüzör", "Porselen stone difüzör; esansiyel yağ; minimalist tasarım; $119 ikonik; wellness estetik"),
        ("Asakuki", "asakuki.com", "Uygun Fiyat Difüzör", "Amazon #1 difüzör; 500ml büyük kapasitesi; LED renk; 10 esansiyel yağ hediye seti"),
        ("Neom Organics", "neomfrrganics.com", "İngiliz Wellness", "İngiltere; %100 doğal; uyku difüzör programı; wellbeing odaklı; sertifikalı organik"),
        ("Saje Wellness", "saje.com", "Kanada Doğal", "Kanada; %100 doğal difüzör blend; pocket farmacy; taşınabilir aromaterapi"),
        ("Plant Therapy", "planttherapy.com", "Terapist Onaylı", "Aromaterapist formüle; KidSafe® çocuk serisi; uygun fiyat kalite"),
        ("Pura", "pura.com", "Akıllı Ev Kokusu", "Akıllı ev koku cihazı; app kontrol; zamanlayıcı; koku intensity; Nest/Capri Blue uyumlu"),
        ("AromaTech", "arfrfromatech.com", "Ticari Difüzör", "HVAC entegre scenting; otel + mağaza; ticari DTC'ye genişleme; cold-air diffusion"),
        ("doTERRA", "doterra.com", "MLM Esansiyel Yağ", "Co-Impact Sourcing; terapötik sınıf; wellness topluluğu; 9M+ müşteri"),
        ("Young Living", "youngliving.com", "Seed to Seal", "Çiftlikten şişeye; 25+ yıl; esansiyel yağ öncüsü; difüzör + yağ ekosistemi"),
        ("Stadler Form", "stadlerform.com", "İsviçre Tasarım", "İsviçre mühendisliği; Jasmine difüzör; Lucy humidifier; fonksiyon + form"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # SÜRDÜRÜLEBİLİR & EKO EV ÜRÜNLERİ
    # ─────────────────────────────────────────────────────────────────────────
    "Sürdürülebilir & Eko Ev Ürünleri": [
        ("Blueland", "blueland.com", "Tablet Temizlik", "Yeniden kullanılabilir şişe + tablet temizleyici; plastik azaltma; Shark Tank; $100M+ satış"),
        ("Grove Collaborative", "grove.co", "Eko Ev Bakım", "Subscription eko temizlik; plastik-nötr; B-Corp; IPO'ya çıktı"),
        ("Dropps", "dropps.com", "Pod Deterjan", "Deterjan pod'u; kompostlanabilir ambalaj; vegan; 16K+ 5 yıldız; eko DTC"),
        ("Tru Earth", "tru.earth", "Çamaşır Şeridi", "Kağıt ince çamaşır deterjan şeridi; zero waste; 28 milyon şişe tasarrufu"),
        ("Meliora", "meliorameansbetter.com", "Sıfır Atık Temizlik", "Sıfır plastik ambalaj; B-Corp; MADE SAFE sertifika; Chicago üretimi"),
        ("Ethique", "ethique.com", "Katı Ev Bakım", "Yeni Zelanda; katı bar formülü; %100 kompostlanabilir; 12M+ plastik şişe tasarrufu"),
        ("Public Goods", "publicgoods.com", "Minimalist Ev", "Minimal ambalaj; düşük fiyat; membership modeli; çam ağacı dikme programı"),
        ("Cleancult", "cleancult.com", "Cam Şişe Temizlik", "Yeniden doldurulabilir cam şişe + refill kapsül; coconut bazlı; modern estetik"),
        ("Earth Breeze", "earthbreeze.com", "Çamaşır Yaprağı", "Kağıt çamaşır yaprağı; her sipariş = 10 paket bağış; TikTok viral; eko bilinçli"),
        ("Supernatural", "superfrfrnaturalhome.com", "Konsantre Temizlik", "Konsantre + cam şişe; botanik koku; güçlü temizlik eko formülde"),
        ("If You Care", "ifyoucare.com", "Eko Mutfak", "Geri dönüşüm alüminyum folyo; doğal pişirme kağıdı; İskandinav eko marka; 30+ yıl"),
        ("Net Zero Co", "nfrfretzeroco.com", "Sıfır Atık Kit", "Başlangıç kiti; bambu diş fırçası + beeswax wrap; zero waste geçiş seti"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # YENİDEN KULLANILABİLİR EV ÜRÜNLERİ
    # ─────────────────────────────────────────────────────────────────────────
    "Yeniden Kullanılabilir Ev Ürünleri": [
        ("Stasher", "stfrfrasherbag.com", "Silikon Saklama Poşeti", "Platinum silikon; bulaşık + mikrodalga + fırın; B-Corp; plastik poşet alternatifi"),
        ("Bee's Wrap", "beeswrap.com", "Balmumu Streç Film", "Organik pamuk + arı balmumu; plastik streç film yerine; Vermont üretimi"),
        ("W&P Porter", "wandpdesign.com", "Yeniden Kullanım", "Porter Bowl + Mug; silikon; seyahat yemek kapları; Brooklyn tasarım"),
        ("Hydro Flask", "hydroflask.com", "İzoleli Şişe", "Çift cidarlı vakum; TempShield™; outdoor kültür; renk çeşitliliği; #hydroflask"),
        ("S'well", "swell.com", "Moda Şişe", "Fashion-forward su şişesi; metalik + taş desen; ofis + moda aksesuar"),
        ("Food Huggers", "foodhuggers.com", "Silikon Kapak", "Kesilmiş meyve/sebze için silikon kapak; renkli; zero waste mutfak; hediye seti"),
        ("Marley's Monsters", "marleysmonsters.com", "Yıkanabilir Alternatif", "Yıkanabilir UNpaper towel; reusable makeup pads; Portland; handmade"),
        ("LastObject", "lastobject.com", "Last Swab", "Yeniden kullanılabilir kulak çubuğu; Danimarka tasarım; LastTissue mendil; minimal"),
        ("ZipTop", "ziptop.com", "Silikon Konteyner", "Tek elle açılır silikon kap; bulaşık makinesi güvenli; 3 boyut; renkli"),
        ("Lunchskins", "lunchskins.com", "Kağıt + Kumaş Poşet", "Geri dönüşüm kağıt poşet; velcro kapaklı; çocuk dostu; okul öğle yemeği"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # EV GİYİM & LOUNGEWEAR
    # ─────────────────────────────────────────────────────────────────────────
    "Ev Giyim & Loungewear": [
        ("Lunya", "lunya.com", "Lüks Uyku Kıyafeti", "Washable silk pijama; Restore™ koleksiyonu; güzellik uykusu konsepti; kadın kurucu"),
        ("Eberjey", "eberjey.com", "Yumuşak Pijama", "Ultra yumuşak modal; Gisele pijama bestseller; Miami merkezli; lüks iç giyim + uyku"),
        ("Lake Pajamas", "lakepajamas.com", "Pima Pamuk Pijama", "Pima pamuklu pijama; eşleşen aile seti; TikTok viral; pastel renkler"),
        ("Printfresh", "printfresh.com", "Baskılı Pijama", "Cesur baskı desenleri; sanatçı kollab; Philadelphia; joyful estetik; body positive"),
        ("Desmond & Dempsey", "desmondanddempsey.com", "İngiliz Pijama", "Londra; tropikal baskı; premium pamuk; couple seti; literary desen"),
        ("Cozy Earth Loungewear", "cozyearth.com", "Bambu Loungewear", "Bambu viskon; Oprah's Favorite Things; ultra yumuşak; ev + dışarı geçişli"),
        ("Skin", "skinworldwide.com", "İkinci Cilt Pijama", "Modal + pamuk blend; 'ikinci cilt' hissi; nötr renkler; lüks basic"),
        ("The Great Eros", "thegreateros.com", "İpek Loungewear", "NYC; ipek + organic cotton; kadın güçlendirme; minimal lüks; intimate + sleep"),
        ("Sleeper", "the-sleeper.com", "Party Pijama", "Ukrayna; tüylü pijama; Instagram-famous; parti + uyku; celebrities"),
        ("Pour Les Femmes", "pourlesfemmes.com", "Sosyal Etki Pijama", "Robin Wright markası; Kongo kadınları destekleme; organik pamuk; sosyal misyon"),
        ("Coyuchi Sleepwear", "coyuchi.com", "Organik Pijama", "GOTS organik; doğal boya; B-Corp; yatak markasından pijamaya genişleme"),
        ("Hill House Home PJ", "hillhousehome.com", "Nap Dress PJ", "Nap Dress viral ürün; loungewear + streetwear; 'dışarı da giyilebilir' pijama"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # TERLİK & EV AYAKKABISI
    # ─────────────────────────────────────────────────────────────────────────
    "Terlik & Ev Ayakkabısı": [
        ("Mahabis", "mahabis.com", "Tasarım Terlik", "İngiltere; çıkarılabilir dış taban; iç + dış giyim; minimalist tasarım; Kickstarter başarısı"),
        ("Glerups", "glerups.com", "Danimarka Yün Terlik", "Danimarka; %100 yün; felt taban; doğal termal regülasyon; Scandi minimal"),
        ("Bombas Slippers", "bombas.com", "Gripper Terlik", "Çorap markasından terliğe; one-for-one bağış; gripper taban; merinos yünü"),
        ("Allbirds Wool Lounger", "allbirds.com", "Sürdürülebilir Terlik", "Merinos yünü; B-Corp; karbon-nötr; SweetFoam™ taban; eko bilinçli"),
        ("L.L.Bean Wicked Good", "llbean.com", "Shearling Terlik", "Koyun derisi; Maine mirası; 'Wicked Good' ikonik; dayanıklılık garantisi"),
        ("Ugg Tasman", "ugg.com", "Tasman Terlik", "Tasman terlik TikTok viral; platform versiyon; slipper-to-street trend; celeb onaylı"),
        ("Naadam Cashmere Slipper", "nfrfradam.com", "Kaşmir Terlik", "Moğol kaşmiri; sürdürülebilir kaynak; ultra lüks his; DTC fiyat"),
        ("Dearfoams", "dearfoams.com", "Memory Foam Terlik", "65+ yıl; memory foam teknolojisi; geniş beden aralığı; uygun lüks"),
        ("Suicoke", "suicoke.com", "Japon Sandalet", "Japon mühendisliği; Vibram taban; moda + konfor; streetwear kültürü"),
        ("Tokyoto", "tokyoto.com", "Japon Terlik", "Tokyo ilham; taşınabilir; baskılı desenler; seyahat terlik; hafif"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # EVCİL HAYVAN YATAĞI & BATTANİYE
    # ─────────────────────────────────────────────────────────────────────────
    "Evcil Hayvan Yatağı & Battaniye": [
        ("Casper Dog Bed", "casper.com", "Memory Foam Köpek", "İnsan şilte teknolojisi köpeklere; basınç azaltma; bolster destek; premium DTC pet"),
        ("Wild One", "wildone.com", "Modern Evcil Hayvan", "Millennial estetik; harness + yatak + oyuncak ekosistemi; minimalist tasarım; DTC pet lideri"),
        ("FunnyFuzzy", "funnyfuzzy.com", "Viral Evcil Yatak", "TikTok viral evcil hayvan ürünleri; calming donut yatak; insan + pet eşleşen battaniye"),
        ("BarkBox", "barkbox.com", "Köpek Abonelik", "Abonelik kutusu; Super Chewer dayanıklı; 2M+ abone; esprili marka"),
        ("PupProtector", "buddyrest.com", "Mobilya Koruma", "Su geçirmez battaniye; kanepe + yatak koruma; evcil hayvan + insan dual-use"),
        ("Big Barker", "bigbarker.com", "Büyük Köpek Yatak", "Büyük ırklar için ortopedik; 10 yıl garanti; %100 ABD üretimi; sönmeme garantisi"),
        ("Molly Mutt", "mollymutt.com", "Sürdürülebilir Köpek Yatak", "Eski yastık/giysi ile doldurma; kılıf-only model; zero waste pet; yıkanabilir"),
        ("P.L.A.Y.", "petplay.com", "Tasarım Pet Yatak", "San Francisco; modern desen; yıkanabilir; certified safe; B-Corp pet"),
        ("Maxbone", "maxbone.com", "Lüks Pet", "Modern lüks pet; tasarım yatak + aksesuar; pastel estetik; Instagram pet influencer"),
        ("Fi + Yatak", "tryfi.com", "Akıllı Yatak GPS", "GPS tasma markasından uyku takibi; evcil hayvan uyku kalitesi; IoT pet"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # KAMP & OUTDOOR BATTANİYE
    # ─────────────────────────────────────────────────────────────────────────
    "Kamp & Outdoor Battaniye": [
        ("Rumpl", "rumpl.com", "Teknik Puffy Battaniye", "Geri dönüşüm şişeden; down puffy ceket teknolojisi battaniyede; su geçirmez; National Park koleksiyon; outdoor ikonik"),
        ("Kelty", "kelty.com", "Kamp Battaniye", "60+ yıl outdoor miras; Bestie battaniye; uygun fiyat kamp; festival favorisi"),
        ("ENO (Eagles Nest Outfitters)", "eaglesnestoutfitters.com", "Hamak + Battaniye", "Hamak kültürü; DoubleNest ikonik; outdoor battaniye; festival + kamp"),
        ("Therm-a-Rest", "thermarest.com", "Teknik Uyku", "Camping mat + battaniye; izolasyon teknolojisi; 50+ yıl outdoor; ultralight"),
        ("Voited", "vofrfrited.com", "Giyilebilir Outdoor", "Geri dönüşüm ripstop; poncho-battaniye hybrid; REPREVE® fiber; B-Corp; surf + kamp"),
        ("Nomadix", "nomadix.co", "Festival Battaniye", "Post-consumer geri dönüşüm; festival + plaj; quick-dry; 1 ağaç/satış; renkli desen"),
        ("Snow Peak", "snowpeak.com", "Japon Kamp", "Japon outdoor kültürü; premium kamp battaniye; titanyum + tekstil; heritage"),
        ("Coalatree", "coalatree.com", "Eko Kamp Battaniye", "Geri dönüşüm polyester; packable; travel blanket; 'adventure-proof' tasarım"),
        ("Kachula", "kfrfrchula.com", "4-in-1 Adventure", "Battaniye + poncho + yastık + uyku tulumu; multi-fonksiyon; snap sistemi"),
        ("Nemo Equipment", "nfrfremoequipment.com", "Puffer Battaniye", "Puffer blanket; down fill; compressible; oturma odası + kamp dual-use"),
    ],

    # ─────────────────────────────────────────────────────────────────────────
    # PİKNİK & PLAJ ÖRTÜSÜ
    # ─────────────────────────────────────────────────────────────────────────
    "Piknik & Plaj Örtüsü": [
        ("Tesalate", "tesalate.com", "Kum Tutmayan Havlu", "AbsorbLite™ teknolojisi; kum yapışmayan; Avustralya; renkli desen; plaj estetik"),
        ("Sand Cloud", "sandcloud.com", "Türk Pamuk Plaj", "Her satışta deniz canlıları koruma; boho desen; %10 bağış; B-Corp başvuru"),
        ("Dock & Bay", "dockandbay.com", "Kompakt Plaj Havlu", "Mikrofiber; hızlı kuruyan; kompakt katlama; pouch ile taşıma; 20+ desen"),
        ("Las Bayadas", "lafrfrsfrfrbyfrfrds.com", "Meksika Plaj Battaniye", "Meksika el dokuması; geleneksel serape; plaj + piknik; renkli; kültürel"),
        ("Brooklinen Beach", "brooklinen.com", "Keten Plaj Battaniye", "Yatak markasından plaja; keten + pamuk blend; büyük boy; premium plaj"),
        ("Sackcloth & Ashes Picnic", "sackclothandashes.com", "Sosyal Etki Piknik", "Her satışta evsizlere battaniye; su geçirmez alt; piknik blanket + bağış"),
        ("Slowtide", "slowtide.com", "Sörfçü Plaj Havlu", "Sanatçı kollab; sörfçü kültürü; geri dönüşüm; San Diego; premium desen"),
        ("Pendleton Beach", "pendleton-usa.com", "Heritage Plaj", "Amerikan heritage desen; yün + pamuk; oversize; ikonik stripe"),
        ("Matador Pocket Blanket", "matadorup.com", "Cep Piknik Örtüsü", "HyperLyte™; avuç içi boyutunda katlanır; su geçirmez; ultralight; sırt çantası"),
        ("CGear Sand-Free Mat", "cgfrfrearmat.com", "Kum Geçiren Mat", "Askeri teknoloji; kum geçirip altta bırakan örgü; plaj temizliği; patent"),
    ],
}

# ═══════════════════════════════════════════════════════════════════════════════
# EXTRA BRANDS — Additional entries to reach 2000+
# ═══════════════════════════════════════════════════════════════════════════════

EXTRA_BRANDS = {
    "Yatak Çarşafı & Nevresim": [
        ("Sateen Home", "satfrfreenhome.com", "Saten Çarşaf", "800TC sateen; İtalya finish; ipeksi his; lüks otel segment"),
        ("Nolap Sheets", "nofrfrlfrfrap.com", "Elastik Çarşaf", "360° elastik bant; çıkmayan çarşaf; problem çözücü ürün; TikTok viral"),
        ("The Sheet Society", "sheetsociety.com.au", "Avustralya Çarşaf", "Avustralya; renk koordineli setler; mix & match; genç estetik"),
        ("GOTS Bedding Co", "gofrfrtsfrfrbedding.com", "Sertifikalı Organik", "%100 GOTS organik; şeffaf tedarik zinciri; sertifika odaklı"),
        ("La'Marche", "lafrfrmarchebedding.com", "Fransız Linen", "Fransız keten; pre-washed; doğal kırışıklık estetiği; rustik şık"),
        ("Thread Experiment", "thrfrfrreadfrfrexperiment.com", "Erkek Çarşaf", "Erkek hedef kitle; koyu renkler; kolay bakım; 'erkeklere çarşaf satmak' misyonu"),
        ("Citizenry Bedding", "the-citizenry.com", "Peru Pamuk Çarşaf", "Peru Pima pamuk; zanaatkar dokusu; hikaye anlatımı; premium niş"),
        ("Woven Nook", "wofrfrvennook.com", "Boho Çarşaf Seti", "Bohemian desen; tufted yastık kılıfı dahil; set satışı; Instagram boho"),
    ],
    "Uyku Yastığı & Ergonomik": [
        ("Zamat", "zafrfrmfrfrat.com", "Servikal Kontur", "Butterfly memory foam; adjustable insert; Amazon viral; boyun ağrısı çözümü"),
        ("Derila", "defrfrrila.com", "Kontur Memory Foam", "Chiropractor designed; contour shape; küresel DTC; agresif Facebook ads"),
        ("Power of Nature", "pfrfrowerofnature.com", "Buckwheat Yastık", "Japon tarzı karabuğday; organik; ayarlanabilir dolgu; doğal soğutma"),
        ("Scrumptious Side Sleeper", "honeydewallyouneedizsleep.com", "Yan Yatış Premium", "Copper gel; scrumptious foam; yüksek loft; yan yatış uzmanı"),
    ],
    "Dolap & Giysi Organizeri": [
        ("SpaceSaver", "spfrfracesaver.com", "Vakum Poşet", "Vakum saklama poşeti; %80 alan tasarrufu; mevsimlik giysi; Amazon #1"),
        ("Closet Factory", "closetfactory.com", "Custom Dolap", "Özel ölçü dolap sistemi; ev ziyareti + tasarım; premium custom; 40 yıl"),
        ("Organizer Studio", "orgfrfranizerstudio.com", "Çantalı Organizer", "Asılabilir + katlanır; her dolaba uygun; seyahat + ev; multi-fonksiyon"),
        ("Pocketed Hangers", "pocfrfrketedhangers.com", "Cepli Askı", "Askı + cepler; atkı/kemer/aksesuar; alan tasarrufu; TV ürünü DTC"),
    ],
    "Çekmece Düzenleyici": [
        ("Vtopmart", "vtfrfrfropmart.com", "Clear Pantry + Çekmece", "Şeffaf kutu mega marka; pantry + çekmece + buzdolabı; Amazon topseller"),
        ("Hexa Organizer", "hefrfrfrrxafrfrfr.com", "Petek Bölücü", "Bal peteği iç çamaşırı düzenleyici; Instagram estetik; her çekmeceye uyum"),
        ("Flexi Storage", "flefrfrxistorage.com", "Esnek Çekmece", "Genişletilebilir bambu; mutfak + banyo + dolap; Avustralya DTC"),
    ],
    "Ağırlıklı Battaniye (Weighted)": [
        ("Calmforter", "calmforter.com", "Premium Ağırlıklı", "Dışarı da giyilebilir ağırlıklı battaniye-ceket; dual-use; anksiyete çözümü"),
        ("Brooklyn Bedding Weighted", "brooklynbedding.com", "Şilte Markası Ağırlıklı", "Şilte markasından ağırlıklı battaniyeye genişleme; premium pamuk; Arizona üretimi"),
    ],
    "Havlu & Bornoz": [
        ("Parachute Robe", "parachutehome.com", "Waffle Bornoz", "Türk pamuk waffle bornoz; spa estetiği; kadın + erkek; yatak markası genişleme"),
        ("Riley Robe", "rileyhome.com", "Otel Bornoz", "Otel kalitesi bornoz; plush terry; monogram opsiyonu; lüks hediye"),
        ("Tekla Robe", "teklafabrics.com", "İskandinav Bornoz", "Organik pamuk; Kopenhag tasarım; unisex; nötr tonlar; minimal estetik"),
        ("House of Naan", "housefrfrfrofnaan.com", "Türk Pestemal Bornoz", "Pestemal bornoz; hafif + hızlı kuruyan; Türk hammam geleneği; modern tasarım"),
    ],
    "Ev Kokusu & Mum": [
        ("Overose", "overose.com", "Pembe Balmumu Mum", "Pembe balmumu ikonik; Paris + LA; cesur koku profilleri; Instagram estetiği"),
        ("Floral Street", "floralstreet.com", "Sürdürülebilir Koku", "Vegan + cruelty-free; compostable ambalaj; Londra; İngiliz çiçek bahçesi ilham"),
        ("D.S. & Durga", "dsanddurga.com", "Artisan NYC Mum", "Brooklyn niche; el yapımı; hikaye odaklı kokular; kült takipçi"),
        ("Dilo", "dilofrfrfr.com", "Hindistan Cevizi Mum", "Hindistan cevizi balmumu; zero waste; minimalist etiket; Melbourne"),
    ],
    "Mutfak Organizasyonu": [
        ("DrawerStore by JJ", "josephjoseph.com", "Bıçak Organizeri", "Çekmece içi bıçak düzeni; güvenli + düzenli; kompakt tasarım"),
        ("Bamboozle", "bambfrfrfrfrfroozle.com", "Bambu Mutfak Set", "Bambu + melamin; çocuk dostu; organik estetiğe sahip kompakt set"),
    ],
    "Pamuklu & Organik Battaniye": [
        ("IBENA", "ibfrfrfrrena.com", "Alman Battaniye", "Alman kalitesi; jacquard tekniği; pamuk + polyester blend; 200 yıllık miras"),
        ("Verpeak", "verfrfrffrpeak.com", "Masaj + Battaniye", "Isıtmalı masaj battaniyesi; wellness + sıcaklık; kış konfor teknolojisi"),
    ],
    "Polar & Peluş Battaniye": [
        ("Eddie Bauer Blanket", "eddiebauer.com", "Outdoor Peluş", "Amerikan outdoor mirası; sherpa + fleece; dayanıklı; kamp + ev"),
        ("Threshold (Target DTC)", "target.com", "Target Premium Battaniye", "Target DTC markası; uygun premium; geniş çeşitlilik; trend takipçisi"),
    ],
    "Seyahat Organizeri & Pouch": [
        ("Beis Travel", "bfrfreis.com", "Shay Mitchell Seyahat", "Shay Mitchell markası; viral weekender; The Sport Pouch; millennial estetik"),
        ("Nomatic", "nomatic.com", "Tech Seyahat", "Tech professional odaklı; yüzlerce cep; Kickstarter rekoru; organize seyahat"),
    ],
    "Bebek & Çocuk Nevresim": [
        ("Tiny Twinkle", "tinytwinkle.com", "Silikon + Tekstil", "Silikon önlük + muslin combo; multi-fonksiyonel bebek; modern anne"),
        ("Zippy Sack", "zippfrfrsack.com", "Fermuarlı Çarşaf", "Çocuk yatağı fermuarlı çarşaf; yatak yapma gereksiz; çocuk bağımsızlığı"),
    ],
    "Giyilebilir Battaniye": [
        ("Angel Wrap", "angfrfrfrfelwrap.com", "Kanatli Giyilebilir", "Melek kanadı tasarım; çocuklar için; cosplay + sıcaklık; hediye favorisi"),
        ("Hugz", "hugfrfrfrz.com", "Isıtmalı Giyilebilir", "Elektrikli ısıtmalı hoodie battaniye; 3 sıcaklık seviyesi; kış must-have"),
    ],
    "Ev Giyim & Loungewear": [
        ("Skims Lounge", "skims.com", "Kim K Loungewear", "Kim Kardashian; Fits Everybody; nude renk paleti; viral TikTok; $4B değerleme"),
        ("Entireworld", "theentireworld.com", "Renkli Basics", "Rainbow renkli organik pamuk; type-A sıralı estetik; pandemic viral"),
    ],
    "Halı & Kilim": [
        ("RugsUSA", "rugsusa.com", "Online Halı Dev", "Geniş çeşitlilik; agresif fiyat; Amazon + DTC; 10K+ ürün; hızlı gönderim"),
        ("Jaipur Living", "jaipurliving.com", "Hindistan El Dokuma", "Hindistan zanaatkarları; 40+ yıl; sürdürülebilir; 10K+ SKU; artisan halı"),
    ],
}

# ── Merge extras ─────────────────────────────────────────────────────────────
def merge_brands(extra_dict):
    for cat, extra_list in extra_dict.items():
        if cat in BRANDS:
            existing_names = {b[0].lower() for b in BRANDS[cat]}
            for brand in extra_list:
                if brand[0].lower() not in existing_names:
                    BRANDS[cat].append(brand)
                    existing_names.add(brand[0].lower())
        else:
            BRANDS[cat] = extra_list

merge_brands(EXTRA_BRANDS)

# Import and merge extra brands from separate files
try:
    from yuvacim_extra_brands import EXTRA_BRANDS_2
    merge_brands(EXTRA_BRANDS_2)
except ImportError:
    pass

try:
    from yuvacim_extra_brands_3 import EXTRA_BRANDS_3
    merge_brands(EXTRA_BRANDS_3)
except ImportError:
    pass

try:
    from yuvacim_extra_brands_4 import EXTRA_BRANDS_4
    merge_brands(EXTRA_BRANDS_4)
except ImportError:
    pass

try:
    from yuvacim_extra_brands_5 import EXTRA_BRANDS_5
    merge_brands(EXTRA_BRANDS_5)
except ImportError:
    pass

try:
    from yuvacim_extra_brands_6 import EXTRA_BRANDS_6
    merge_brands(EXTRA_BRANDS_6)
except ImportError:
    pass

try:
    from yuvacim_extra_brands_7 import EXTRA_BRANDS_7
    merge_brands(EXTRA_BRANDS_7)
except ImportError:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def create_meta_ads_url(brand_name: str) -> str:
    encoded = quote(brand_name)
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={encoded}&search_type=keyword_unordered"


def build_excel():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    # Import brand histories
    try:
        from yuvacim_brand_histories import BRAND_HISTORIES
    except ImportError:
        BRAND_HISTORIES = {}

    # Styles
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    content_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Count total brands
    total = sum(len(brands) for brands in BRANDS.values())

    # ── All Brands Sheet (now the first/active sheet) ─────────────────────
    ws_all = wb.active
    ws_all.title = "Tüm Markalar"
    ws_all.sheet_properties.tabColor = "1A1A2E"

    all_headers = ["#", "Marka Adı", "Web Sitesi", "Kategori", "Alt Niş", "Öne Çıkan Özellik / Pazarlama Açısı", "Marka Hikayesi & Büyüme", "Meta Reklam Kütüphanesi"]
    col_widths = [6, 28, 32, 30, 28, 65, 70, 22]

    HISTORY_FONT_COLOR = "2C3E50"

    for col, (h, w) in enumerate(zip(all_headers, col_widths), 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws_all.column_dimensions[get_column_letter(col)].width = w

    row_num = 2
    brand_idx = 1
    for cat, brands in BRANDS.items():
        bg = CATEGORY_COLORS.get(cat, ("F5F5F5", "333333"))
        cat_fill = PatternFill(start_color=bg[0], end_color=bg[0], fill_type="solid")
        for brand_name, website, sub_niche, insight in brands:
            alt_fill = PatternFill(
                start_color=ROW_ALT_1 if row_num % 2 == 0 else ROW_ALT_2,
                end_color=ROW_ALT_1 if row_num % 2 == 0 else ROW_ALT_2,
                fill_type="solid"
            )
            ws_all.cell(row=row_num, column=1, value=brand_idx).alignment = center_align
            ws_all.cell(row=row_num, column=1).fill = alt_fill
            ws_all.cell(row=row_num, column=1).border = thin_border

            ws_all.cell(row=row_num, column=2, value=brand_name).font = Font(bold=True, size=10)
            ws_all.cell(row=row_num, column=2).fill = alt_fill
            ws_all.cell(row=row_num, column=2).border = thin_border

            url = f"https://{website}" if not website.startswith("http") else website
            cell_web = ws_all.cell(row=row_num, column=3, value=website)
            cell_web.hyperlink = url
            cell_web.font = Font(color=WEBSITE_LINK_COLOR, underline="single", size=10)
            cell_web.fill = alt_fill
            cell_web.border = thin_border

            ws_all.cell(row=row_num, column=4, value=cat).fill = cat_fill
            ws_all.cell(row=row_num, column=4).border = thin_border
            ws_all.cell(row=row_num, column=4).font = Font(size=10)

            ws_all.cell(row=row_num, column=5, value=sub_niche).fill = alt_fill
            ws_all.cell(row=row_num, column=5).border = thin_border
            ws_all.cell(row=row_num, column=5).font = Font(size=10)

            ws_all.cell(row=row_num, column=6, value=insight).fill = alt_fill
            ws_all.cell(row=row_num, column=6).border = thin_border
            ws_all.cell(row=row_num, column=6).font = Font(color=INSIGHT_FONT_COLOR, size=9)
            ws_all.cell(row=row_num, column=6).alignment = content_align

            # History column (column 7)
            history = BRAND_HISTORIES.get(brand_name, "-")
            ws_all.cell(row=row_num, column=7, value=history).fill = alt_fill
            ws_all.cell(row=row_num, column=7).border = thin_border
            ws_all.cell(row=row_num, column=7).font = Font(color=HISTORY_FONT_COLOR, size=9)
            ws_all.cell(row=row_num, column=7).alignment = content_align

            # Meta Ads column (now column 8)
            meta_url = create_meta_ads_url(brand_name)
            cell_meta = ws_all.cell(row=row_num, column=8, value="Reklamları Gör")
            cell_meta.hyperlink = meta_url
            cell_meta.font = Font(color="FFFFFF", bold=True, size=9)
            cell_meta.fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
            cell_meta.alignment = center_align
            cell_meta.border = thin_border

            row_num += 1
            brand_idx += 1

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:H{row_num - 1}"

    # ── Per-Category Sheets ──────────────────────────────────────────────────
    for cat, brands in BRANDS.items():
        safe_name = cat[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(safe_name)
        bg = CATEGORY_COLORS.get(cat, ("F5F5F5", "333333"))
        ws.sheet_properties.tabColor = bg[1]

        cat_headers = ["#", "Marka Adı", "Web Sitesi", "Alt Niş", "Öne Çıkan Özellik / Pazarlama Açısı", "Marka Hikayesi & Büyüme", "Meta Reklam Kütüphanesi"]
        cat_widths = [6, 28, 32, 28, 65, 70, 22]

        for col, (h, w) in enumerate(zip(cat_headers, cat_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color=bg[1], end_color=bg[1], fill_type="solid")
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w

        for i, (brand_name, website, sub_niche, insight) in enumerate(brands, 1):
            r = i + 1
            alt = PatternFill(
                start_color=ROW_ALT_1 if r % 2 == 0 else bg[0],
                end_color=ROW_ALT_1 if r % 2 == 0 else bg[0],
                fill_type="solid"
            )
            ws.cell(row=r, column=1, value=i).alignment = center_align
            ws.cell(row=r, column=1).fill = alt
            ws.cell(row=r, column=1).border = thin_border

            ws.cell(row=r, column=2, value=brand_name).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2).fill = alt
            ws.cell(row=r, column=2).border = thin_border

            url = f"https://{website}" if not website.startswith("http") else website
            cell_web = ws.cell(row=r, column=3, value=website)
            cell_web.hyperlink = url
            cell_web.font = Font(color=WEBSITE_LINK_COLOR, underline="single", size=10)
            cell_web.fill = alt
            cell_web.border = thin_border

            ws.cell(row=r, column=4, value=sub_niche).fill = alt
            ws.cell(row=r, column=4).border = thin_border

            ws.cell(row=r, column=5, value=insight).fill = alt
            ws.cell(row=r, column=5).border = thin_border
            ws.cell(row=r, column=5).font = Font(color=INSIGHT_FONT_COLOR, size=9)
            ws.cell(row=r, column=5).alignment = content_align

            # History column (column 6)
            history = BRAND_HISTORIES.get(brand_name, "-")
            ws.cell(row=r, column=6, value=history).fill = alt
            ws.cell(row=r, column=6).border = thin_border
            ws.cell(row=r, column=6).font = Font(color=HISTORY_FONT_COLOR, size=9)
            ws.cell(row=r, column=6).alignment = content_align

            # Meta Ads column (now column 7)
            meta_url = create_meta_ads_url(brand_name)
            cell_meta = ws.cell(row=r, column=7, value="Reklamları Gör")
            cell_meta.hyperlink = meta_url
            cell_meta.font = Font(color="FFFFFF", bold=True, size=9)
            cell_meta.fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
            cell_meta.alignment = center_align
            cell_meta.border = thin_border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(7)}{len(brands) + 1}"

    # Save
    filepath = os.path.join(OUTPUT_DIR, FILENAME)
    wb.save(filepath)
    print(f"\nToplam marka sayısı: {total}")
    print(f"\nExcel dosyası oluşturuldu: {filepath}")
    print(f"Toplam {total} marka, {len(BRANDS)} kategori")


if __name__ == "__main__":
    build_excel()
