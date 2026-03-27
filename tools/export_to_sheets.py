"""
Tool: Export Normalized Products to Google Sheets
==================================================
Reads normalized product data and writes it to a multi-tab Google Sheet
or local Excel file as fallback.

Input:  .tmp/normalized_products.json
Output: Google Sheet (5 tabs) or .tmp/product_research.xlsx

Usage:
    python3 tools/export_to_sheets.py
    python3 tools/export_to_sheets.py --test
    python3 tools/export_to_sheets.py --excel-only
"""

import argparse
import json
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(PROJECT_ROOT / ".env")

TMP_DIR = PROJECT_ROOT / ".tmp"
NORMALIZED_FILE = TMP_DIR / "normalized_products.json"
EXCEL_OUTPUT = TMP_DIR / "product_research.xlsx"

# Column definitions for each tab
ALL_PRODUCTS_COLUMNS = [
    "Source", "Product Name", "Product Name (CN)", "Price (USD)",
    "Price (Local)", "Currency", "Total Sold", "7d Sales", "30d Sales",
    "Rating", "Reviews", "Seller", "Location", "MOQ",
    "Ad Duration (days)", "Cross-Platform", "Matched Platforms",
    "Keyword", "Product URL", "Image URL",
]

PIPIADS_COLUMNS = [
    "Product Name", "Advertiser", "Price (USD)", "Total Sold",
    "Ad Impressions", "Ad Likes", "Ad Comments", "Ad Shares",
    "First Seen", "Last Seen", "Duration (days)",
    "Creative URL", "Product URL", "Keyword",
]

DOUYIN_COLUMNS = [
    "Product Name", "Product Name (CN)", "Seller",
    "Price (CNY)", "Price (USD)", "Total Sold",
    "Video Views", "Video Likes", "Video Comments", "Video Shares",
    "Product URL", "Keyword",
]

TAOBAO_COLUMNS = [
    "Product Name (CN)", "Price Range (CNY)", "Price (USD)",
    "MOQ", "Orders", "Total Sold", "Rating", "Reviews",
    "Supplier", "Location", "Years", "Verified",
    "Product URL", "Keyword",
]

SUMMARY_COLUMNS = [
    "Keyword", "PiPiAds Hits", "Douyin Hits", "1688 Hits",
    "Cross-Platform Matches", "Avg Price (USD)", "Top Seller",
]


def load_normalized():
    if not NORMALIZED_FILE.exists():
        print(f"ERROR: {NORMALIZED_FILE} not found.")
        print("Run tools/normalize_data.py first.")
        sys.exit(1)
    with open(NORMALIZED_FILE, encoding="utf-8") as f:
        return json.load(f)


def product_to_all_row(p):
    matched = ", ".join(p.get("matched_platforms", []))
    price_local = p.get("price", 0)
    return [
        p.get("source_platform", ""),
        p.get("product_name", ""),
        p.get("product_name_cn", ""),
        p.get("price_usd", 0),
        price_local,
        p.get("price_currency", ""),
        p.get("total_sold", 0),
        p.get("sales_7d", 0),
        p.get("sales_30d", 0),
        p.get("rating", 0),
        p.get("review_count", 0),
        p.get("seller_name") or p.get("advertiser_name", ""),
        p.get("seller_location", ""),
        p.get("moq", 0),
        p.get("ad_duration_days", 0),
        "Yes" if p.get("cross_platform_match") else "",
        matched,
        p.get("search_keyword", ""),
        p.get("product_url", ""),
        p.get("image_url", ""),
    ]


def product_to_pipiads_row(p):
    return [
        p.get("product_name", ""),
        p.get("advertiser_name", ""),
        p.get("price_usd", 0),
        p.get("total_sold", 0),
        p.get("ad_impressions", 0),
        p.get("ad_likes", 0),
        p.get("ad_comments", 0),
        p.get("ad_shares", 0),
        p.get("ad_first_seen", ""),
        p.get("ad_last_seen", ""),
        p.get("ad_duration_days", 0),
        p.get("ad_creative_url", ""),
        p.get("product_url", ""),
        p.get("search_keyword", ""),
    ]


def product_to_douyin_row(p):
    return [
        p.get("product_name", ""),
        p.get("product_name_cn", ""),
        p.get("seller_name", ""),
        p.get("price", 0),
        p.get("price_usd", 0),
        p.get("total_sold", 0),
        p.get("video_views", 0),
        p.get("video_likes", 0),
        p.get("video_comments", 0),
        p.get("video_shares", 0),
        p.get("product_url", ""),
        p.get("search_keyword", ""),
    ]


def product_to_taobao_row(p):
    price_range = ""
    if p.get("price_range_min") or p.get("price_range_max"):
        price_range = f"¥{p.get('price_range_min', 0)} - ¥{p.get('price_range_max', 0)}"
    return [
        p.get("product_name_cn") or p.get("product_name", ""),
        price_range,
        p.get("price_usd", 0),
        p.get("moq", 0),
        p.get("order_count", 0),
        p.get("total_sold", 0),
        p.get("rating", 0),
        p.get("review_count", 0),
        p.get("seller_name", ""),
        p.get("seller_location", ""),
        p.get("supplier_years", 0),
        "Yes" if p.get("supplier_verified") else "",
        p.get("product_url", ""),
        p.get("search_keyword", ""),
    ]


def build_summary(products):
    """Build per-keyword summary rows."""
    keywords = {}
    for p in products:
        kw = p.get("search_keyword", "Unknown")
        if kw not in keywords:
            keywords[kw] = {
                "pipiads": 0, "douyin": 0, "1688": 0,
                "cross_platform": 0, "price_sum": 0, "price_count": 0,
                "sellers": {},
            }
        s = keywords[kw]
        platform = p.get("source_platform", "")
        if platform in s:
            s[platform] += 1
        if p.get("cross_platform_match"):
            s["cross_platform"] += 1
        if p.get("price_usd", 0) > 0:
            s["price_sum"] += p["price_usd"]
            s["price_count"] += 1
        seller = p.get("seller_name") or p.get("advertiser_name", "")
        if seller:
            s["sellers"][seller] = s["sellers"].get(seller, 0) + 1

    rows = []
    for kw, s in sorted(keywords.items()):
        avg_price = round(s["price_sum"] / s["price_count"], 2) if s["price_count"] > 0 else 0
        top_seller = max(s["sellers"].items(), key=lambda x: x[1])[0] if s["sellers"] else ""
        rows.append([
            kw, s["pipiads"], s["douyin"], s["1688"],
            s["cross_platform"], avg_price, top_seller,
        ])
    return rows


def write_tab(spreadsheet, tab_name, columns, rows):
    """Write a tab to Google Sheets (create or clear)."""
    try:
        ws = spreadsheet.worksheet(tab_name)
        ws.clear()
    except Exception:
        ws = spreadsheet.add_worksheet(title=tab_name, rows=max(len(rows) + 10, 50), cols=len(columns))

    ws.update(range_name="A1", values=[columns])
    if rows:
        ws.update(range_name="A2", values=rows)

    # Format header
    col_letter = chr(64 + len(columns)) if len(columns) <= 26 else "Z"
    ws.format(f"A1:{col_letter}1", {"textFormat": {"bold": True}})
    ws.freeze(rows=1)
    return ws


def export_to_google_sheets(data, test_only=False):
    import gspread
    from google.oauth2.service_account import Credentials

    creds_file = PROJECT_ROOT / os.getenv("GOOGLE_SHEETS_CREDENTIALS_FILE", "credentials.json")
    sheet_id = os.getenv("GOOGLE_SHEET_ID")

    if not creds_file.exists():
        print(f"Credentials file not found: {creds_file}")
        return False
    if not sheet_id:
        print("GOOGLE_SHEET_ID not set in .env")
        return False

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(creds_file), scopes=scopes)
    gc = gspread.authorize(creds)

    try:
        spreadsheet = gc.open_by_key(sheet_id)
    except Exception as e:
        print(f"ERROR: Cannot open sheet '{sheet_id}': {e}")
        return False

    if test_only:
        print(f"Google Sheets OK: '{spreadsheet.title}'")
        return True

    products = data.get("products", [])
    pipiads_products = [p for p in products if p.get("source_platform") == "pipiads"]
    douyin_products = [p for p in products if p.get("source_platform") == "douyin"]
    taobao_products = [p for p in products if p.get("source_platform") == "1688"]

    # Tab 1: All Products
    print("  Writing 'All Products' tab...")
    all_rows = [product_to_all_row(p) for p in products]
    write_tab(spreadsheet, "All Products", ALL_PRODUCTS_COLUMNS, all_rows)

    # Tab 2: PiPiAds Results
    if pipiads_products:
        print("  Writing 'PiPiAds Results' tab...")
        pipiads_rows = [product_to_pipiads_row(p) for p in pipiads_products]
        write_tab(spreadsheet, "PiPiAds Results", PIPIADS_COLUMNS, pipiads_rows)

    # Tab 3: Douyin Results
    if douyin_products:
        print("  Writing 'Douyin Results' tab...")
        douyin_rows = [product_to_douyin_row(p) for p in douyin_products]
        write_tab(spreadsheet, "Douyin Results", DOUYIN_COLUMNS, douyin_rows)

    # Tab 4: 1688 Suppliers
    if taobao_products:
        print("  Writing '1688 Suppliers' tab...")
        taobao_rows = [product_to_taobao_row(p) for p in taobao_products]
        write_tab(spreadsheet, "1688 Suppliers", TAOBAO_COLUMNS, taobao_rows)

    # Tab 5: Summary
    print("  Writing 'Summary' tab...")
    summary_rows = build_summary(products)
    write_tab(spreadsheet, "Summary", SUMMARY_COLUMNS, summary_rows)

    print(f"\nExported {len(products)} products to Google Sheets: '{spreadsheet.title}'")
    print(f"  PiPiAds: {len(pipiads_products)} | Douyin: {len(douyin_products)} | 1688: {len(taobao_products)}")
    return True


def export_to_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    products = data.get("products", [])
    pipiads_products = [p for p in products if p.get("source_platform") == "pipiads"]
    douyin_products = [p for p in products if p.get("source_platform") == "douyin"]
    taobao_products = [p for p in products if p.get("source_platform") == "1688"]

    header_font = Font(bold=True)
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def write_sheet(ws, columns, rows, highlight_col=None):
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
        for row in rows:
            ws.append(row)
        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, min(len(str(cell.value or "")), 50))
            ws.column_dimensions[col_letter].width = max_len + 2

    # All Products
    ws_all = wb.active
    ws_all.title = "All Products"
    all_rows = [product_to_all_row(p) for p in products]
    write_sheet(ws_all, ALL_PRODUCTS_COLUMNS, all_rows)

    # Highlight cross-platform matches
    for row_idx, p in enumerate(products, start=2):
        if p.get("cross_platform_match"):
            for col_idx in range(1, len(ALL_PRODUCTS_COLUMNS) + 1):
                ws_all.cell(row=row_idx, column=col_idx).fill = match_fill

    # PiPiAds
    if pipiads_products:
        ws_pp = wb.create_sheet("PiPiAds Results")
        write_sheet(ws_pp, PIPIADS_COLUMNS, [product_to_pipiads_row(p) for p in pipiads_products])

    # Douyin
    if douyin_products:
        ws_dy = wb.create_sheet("Douyin Results")
        write_sheet(ws_dy, DOUYIN_COLUMNS, [product_to_douyin_row(p) for p in douyin_products])

    # 1688
    if taobao_products:
        ws_tb = wb.create_sheet("1688 Suppliers")
        write_sheet(ws_tb, TAOBAO_COLUMNS, [product_to_taobao_row(p) for p in taobao_products])

    # Summary
    ws_sum = wb.create_sheet("Summary")
    summary_rows = build_summary(products)
    write_sheet(ws_sum, SUMMARY_COLUMNS, summary_rows)

    wb.save(str(EXCEL_OUTPUT))
    print(f"Exported {len(products)} products to {EXCEL_OUTPUT}")
    print(f"  PiPiAds: {len(pipiads_products)} | Douyin: {len(douyin_products)} | 1688: {len(taobao_products)}")


def main():
    parser = argparse.ArgumentParser(description="Export normalized products to Google Sheets")
    parser.add_argument("--test", action="store_true", help="Test Google Sheets connection only")
    parser.add_argument("--excel-only", action="store_true", help="Force Excel output")
    args = parser.parse_args()

    if args.test:
        if not export_to_google_sheets(None, test_only=True):
            print("Google Sheets not configured. Excel mode available.")
        return

    data = load_normalized()
    products = data.get("products", [])
    print(f"Export to Sheets")
    print(f"  Products: {len(products)}")
    print()

    if args.excel_only:
        export_to_excel(data)
    elif not export_to_google_sheets(data):
        print("Google Sheets not available, falling back to Excel...")
        export_to_excel(data)

    print("Done.")


if __name__ == "__main__":
    main()
