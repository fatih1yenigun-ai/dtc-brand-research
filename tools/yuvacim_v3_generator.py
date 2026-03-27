#!/usr/bin/env python3
"""
Yuvacım v3 — Post-2018 Ecommerce-Native Brands Only
Generator script that imports brand data from batch files and builds Excel.
"""
import os
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research_outputs")
TODAY = datetime.now().strftime("%Y-%m-%d")
FILENAME = f"Yuvacim_Post2018_Ecommerce_Markalar_{TODAY}.xlsx"

HEADER_COLOR = "0D1B2A"
HEADER_FONT_COLOR = "F5F5F5"
META_BUTTON_COLOR = "2ECC71"
WEBSITE_LINK_COLOR = "3498DB"
INSIGHT_FONT_COLOR = "4A4A4A"
HISTORY_FONT_COLOR = "6C3483"

CATEGORY_COLORS = {
    "Yatak Çarşafı & Nevresim": ("D6EAF8", "1A5276"),
    "Yorgan & Dolgulu Ürünler": ("D4E6F1", "1B4F72"),
    "Yastık & Uyku Aksesuarı": ("D5F5E3", "1E8449"),
    "Ağırlıklı Battaniye": ("F9E79F", "B7950B"),
    "Battaniye & Throw": ("F7DC6F", "D4AC0D"),
    "Giyilebilir Battaniye": ("FAD7A0", "CA6F1E"),
    "Havlu & Bornoz & Pestemal": ("D2B4DE", "7D3C98"),
    "Dolap & Giysi Organizeri": ("F5CBA7", "CA6F1E"),
    "Çekmece Düzenleyici": ("EDBB99", "BA4A00"),
    "Mutfak Organizasyonu": ("ABEBC6", "239B56"),
    "Banyo Organizasyonu": ("AED6F1", "21618C"),
    "Seyahat Organizeri": ("FADBD8", "A93226"),
    "Makyaj & Kozmetik Organizeri": ("F5B7B1", "CB4335"),
    "Ev Kokusu & Mum": ("FDEBD0", "E67E22"),
    "Halı & Kilim": ("C39BD3", "884EA0"),
    "Perde & Ev Tekstili": ("D7BDE2", "6C3483"),
    "Dekoratif Yastık & Kırlent": ("BB8FCE", "7D3C98"),
    "Bebek Uyku & Tekstil": ("A9CCE3", "2471A3"),
    "Anne & Hamile Ürünleri": ("82E0AA", "28B463"),
    "Ev Giyim & Loungewear": ("F5CBA7", "CA6F1E"),
    "Terlik & Ev Ayakkabısı": ("EDBB99", "BA4A00"),
    "Uyku Teknolojisi": ("AED6F1", "21618C"),
    "Sürdürülebilir Ev Ürünleri": ("D5F5E3", "1E8449"),
    "Evcil Hayvan Yatak & Tekstil": ("D6EAF8", "1A5276"),
    "Outdoor & Piknik Tekstil": ("ABEBC6", "239B56"),
}

# Master brand dict — populated by imports
BRANDS = {}

def merge_brands(extra):
    # Handle both dict format and flat list format
    if isinstance(extra, list):
        # Flat list: each item is (name, website, category, insight, history)
        grouped = {}
        for item in extra:
            cat = item[2]
            brand_tuple = (item[0], item[1], item[2], item[3], item[4] if len(item) > 4 else "-")
            grouped.setdefault(cat, []).append(brand_tuple)
        extra = grouped

    for cat, extra_list in extra.items():
        if cat in BRANDS:
            existing = {b[0].lower() for b in BRANDS[cat]}
            for brand in extra_list:
                if brand[0].lower() not in existing:
                    BRANDS[cat].append(brand)
                    existing.add(brand[0].lower())
        else:
            BRANDS[cat] = list(extra_list)

def create_meta_ads_url(brand_name):
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={quote(brand_name)}&search_type=keyword_unordered"

def build_excel():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    # Import TQS calculator and brand metrics
    from tqs_calculator import calc_tqs, tqs_to_conversion, estimate_revenue, get_niche_from_category
    from brand_metrics import BRAND_METRICS, CATEGORY_AOV_DEFAULTS

    hdr_font = Font(name="Calibri", bold=True, color=HEADER_FONT_COLOR, size=11)
    hdr_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"), bottom=Side(style="thin", color="E0E0E0"),
    )
    content_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    money_font = Font(bold=True, color="27AE60", size=10)
    traffic_font = Font(bold=True, color="2980B9", size=10)
    tqs_high = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    tqs_mid = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    tqs_low = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    no_data_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")

    def get_brand_row_data(brand_name, cat):
        """Get metrics for a brand, calculate TQS and niche-adjusted revenue."""
        m = BRAND_METRICS.get(brand_name, {})
        niche = get_niche_from_category(cat)
        if m:
            founded = m.get("founded", "")
            traffic = m.get("traffic", "")
            bounce = m.get("bounce_pct", 50)
            pages = m.get("pages_visit", 3)
            sess = m.get("session_sec", 120)
            aov = m.get("aov", CATEGORY_AOV_DEFAULTS.get(cat, 80))
            tqs = calc_tqs(bounce, pages, sess)
            conv = tqs_to_conversion(tqs, niche)
            est_rev = round(estimate_revenue(traffic, aov, conv)) if traffic else ""
            return founded, traffic, aov, tqs, conv, est_rev
        else:
            aov = CATEGORY_AOV_DEFAULTS.get(cat, 80)
            return "", "", aov, "", "", ""

    # ── Tüm Markalar ────────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "Tüm Markalar"
    ws_all.sheet_properties.tabColor = HEADER_COLOR

    all_headers = [
        "#", "Marka Adı", "Web Sitesi", "Kategori", "Alt Niş",
        "Kuruluş Yılı", "Aylık Trafik", "Tahmini AOV ($)",
        "TQS", "Tahmini Dönüşüm %", "Tahmini Aylık Gelir ($)",
        "Öne Çıkan Özellik / Pazarlama Açısı", "Marka Hikayesi & Büyüme",
        "Meta Reklam Kütüphanesi"
    ]
    col_widths = [5, 28, 28, 26, 24, 12, 14, 14, 8, 14, 18, 55, 55, 18]

    for col, (h, w) in enumerate(zip(all_headers, col_widths), 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = thin_border
        ws_all.column_dimensions[get_column_letter(col)].width = w

    row_num = 2
    brand_idx = 1
    brands_with_data = 0
    for cat, brands in BRANDS.items():
        bg = CATEGORY_COLORS.get(cat, ("F5F5F5", "333333"))
        cat_fill = PatternFill(start_color=bg[0], end_color=bg[0], fill_type="solid")
        for b in brands:
            brand_name, website, sub_niche, insight = b[0], b[1], b[2], b[3]
            history = b[4] if len(b) > 4 else "-"
            founded, traffic, aov, tqs, conv, est_rev = get_brand_row_data(brand_name, cat)
            if traffic: brands_with_data += 1

            alt_fill = PatternFill(
                start_color="FFFFFF" if row_num % 2 == 0 else "F8F9FA",
                end_color="FFFFFF" if row_num % 2 == 0 else "F8F9FA", fill_type="solid"
            )

            # Col 1: #
            ws_all.cell(row=row_num, column=1, value=brand_idx).alignment = center_align
            ws_all.cell(row=row_num, column=1).fill = alt_fill; ws_all.cell(row=row_num, column=1).border = thin_border
            # Col 2: Brand Name
            ws_all.cell(row=row_num, column=2, value=brand_name).font = Font(bold=True, size=10)
            ws_all.cell(row=row_num, column=2).fill = alt_fill; ws_all.cell(row=row_num, column=2).border = thin_border
            # Col 3: Website
            url = f"https://{website}" if not website.startswith("http") else website
            cw = ws_all.cell(row=row_num, column=3, value=website)
            cw.hyperlink = url; cw.font = Font(color=WEBSITE_LINK_COLOR, underline="single", size=10)
            cw.fill = alt_fill; cw.border = thin_border
            # Col 4: Category
            ws_all.cell(row=row_num, column=4, value=cat).fill = cat_fill
            ws_all.cell(row=row_num, column=4).border = thin_border; ws_all.cell(row=row_num, column=4).font = Font(size=10)
            # Col 5: Sub-niche
            ws_all.cell(row=row_num, column=5, value=sub_niche).fill = alt_fill
            ws_all.cell(row=row_num, column=5).border = thin_border; ws_all.cell(row=row_num, column=5).font = Font(size=10)
            # Col 6: Founded Year
            c6 = ws_all.cell(row=row_num, column=6, value=founded if founded else "-")
            c6.alignment = center_align; c6.fill = alt_fill; c6.border = thin_border
            c6.font = Font(bold=True, size=10) if founded else Font(color="CCCCCC", size=9)
            # Col 7: Monthly Traffic
            c7 = ws_all.cell(row=row_num, column=7, value=traffic if traffic else "-")
            c7.alignment = center_align; c7.border = thin_border
            if traffic:
                c7.number_format = '#,##0'; c7.font = traffic_font; c7.fill = alt_fill
            else:
                c7.fill = no_data_fill; c7.font = Font(color="CCCCCC", size=9)
            # Col 8: Estimated AOV
            c8 = ws_all.cell(row=row_num, column=8, value=aov)
            c8.alignment = center_align; c8.number_format = '$#,##0'; c8.fill = alt_fill; c8.border = thin_border
            # Col 9: TQS
            c9 = ws_all.cell(row=row_num, column=9, value=tqs if tqs else "-")
            c9.alignment = center_align; c9.border = thin_border
            if tqs:
                if tqs >= 7: c9.fill = tqs_high; c9.font = Font(bold=True, color="1E8449", size=10)
                elif tqs >= 4: c9.fill = tqs_mid; c9.font = Font(bold=True, color="B7950B", size=10)
                else: c9.fill = tqs_low; c9.font = Font(bold=True, color="C0392B", size=10)
            else:
                c9.fill = no_data_fill; c9.font = Font(color="CCCCCC", size=9)
            # Col 10: Conversion %
            c10 = ws_all.cell(row=row_num, column=10, value=f"{conv}%" if conv else "-")
            c10.alignment = center_align; c10.fill = alt_fill; c10.border = thin_border
            # Col 11: Est Monthly Revenue
            c11 = ws_all.cell(row=row_num, column=11, value=est_rev if est_rev else "-")
            c11.alignment = center_align; c11.border = thin_border
            if est_rev:
                c11.number_format = '$#,##0'; c11.font = money_font; c11.fill = alt_fill
            else:
                c11.fill = no_data_fill; c11.font = Font(color="CCCCCC", size=9)
            # Col 12: Insight
            ws_all.cell(row=row_num, column=12, value=insight).fill = alt_fill
            ws_all.cell(row=row_num, column=12).border = thin_border
            ws_all.cell(row=row_num, column=12).font = Font(color=INSIGHT_FONT_COLOR, size=9)
            ws_all.cell(row=row_num, column=12).alignment = content_align
            # Col 13: History
            ws_all.cell(row=row_num, column=13, value=history).fill = alt_fill
            ws_all.cell(row=row_num, column=13).border = thin_border
            ws_all.cell(row=row_num, column=13).font = Font(color=HISTORY_FONT_COLOR, size=9)
            ws_all.cell(row=row_num, column=13).alignment = content_align
            # Col 14: Meta Ads
            meta_url = create_meta_ads_url(brand_name)
            cm = ws_all.cell(row=row_num, column=14, value="Reklamları Gör")
            cm.hyperlink = meta_url
            cm.font = Font(color="FFFFFF", bold=True, size=9)
            cm.fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
            cm.alignment = center_align; cm.border = thin_border

            row_num += 1; brand_idx += 1

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:N{row_num - 1}"
    total = brand_idx - 1

    # ── Per-Category Sheets ──────────────────────────────────────────────────
    for cat, brands in BRANDS.items():
        safe = cat[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(safe)
        bg = CATEGORY_COLORS.get(cat, ("F5F5F5", "333333"))
        ws.sheet_properties.tabColor = bg[1]
        cat_headers = [
            "#", "Marka Adı", "Web Sitesi", "Alt Niş",
            "Kuruluş Yılı", "Aylık Trafik", "AOV ($)", "TQS", "Dönüşüm %", "Tahmini Gelir ($)",
            "Öne Çıkan Özellik", "Marka Hikayesi",
            "Meta Reklam Kütüphanesi"
        ]
        cat_widths = [5, 26, 28, 24, 10, 13, 12, 8, 10, 16, 50, 50, 18]
        for col, (h, w) in enumerate(zip(cat_headers, cat_widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = PatternFill(start_color=bg[1], end_color=bg[1], fill_type="solid")
            c.alignment = hdr_align; c.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w

        for i, b in enumerate(brands, 1):
            r = i + 1
            brand_name, website, sub_niche, insight = b[0], b[1], b[2], b[3]
            history = b[4] if len(b) > 4 else "-"
            founded, traffic, aov, tqs, conv, est_rev = get_brand_row_data(brand_name, cat)

            alt = PatternFill(
                start_color="FFFFFF" if r % 2 == 0 else bg[0],
                end_color="FFFFFF" if r % 2 == 0 else bg[0], fill_type="solid"
            )
            ws.cell(row=r, column=1, value=i).alignment = center_align; ws.cell(row=r, column=1).fill = alt; ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2, value=brand_name).font = Font(bold=True, size=10); ws.cell(row=r, column=2).fill = alt; ws.cell(row=r, column=2).border = thin_border
            url = f"https://{website}" if not website.startswith("http") else website
            cw = ws.cell(row=r, column=3, value=website); cw.hyperlink = url
            cw.font = Font(color=WEBSITE_LINK_COLOR, underline="single", size=10); cw.fill = alt; cw.border = thin_border
            ws.cell(row=r, column=4, value=sub_niche).fill = alt; ws.cell(row=r, column=4).border = thin_border
            # Founded
            c5 = ws.cell(row=r, column=5, value=founded if founded else "-")
            c5.alignment = center_align; c5.fill = alt; c5.border = thin_border
            # Traffic
            c6 = ws.cell(row=r, column=6, value=traffic if traffic else "-")
            c6.alignment = center_align; c6.border = thin_border
            if traffic: c6.number_format = '#,##0'; c6.font = traffic_font; c6.fill = alt
            else: c6.fill = no_data_fill
            # AOV
            c7 = ws.cell(row=r, column=7, value=aov); c7.alignment = center_align; c7.number_format = '$#,##0'; c7.fill = alt; c7.border = thin_border
            # TQS
            c8 = ws.cell(row=r, column=8, value=tqs if tqs else "-"); c8.alignment = center_align; c8.border = thin_border
            if tqs:
                if tqs >= 7: c8.fill = tqs_high
                elif tqs >= 4: c8.fill = tqs_mid
                else: c8.fill = tqs_low
            else: c8.fill = no_data_fill
            # Conv
            c9 = ws.cell(row=r, column=9, value=f"{conv}%" if conv else "-"); c9.alignment = center_align; c9.fill = alt; c9.border = thin_border
            # Revenue
            c10 = ws.cell(row=r, column=10, value=est_rev if est_rev else "-"); c10.alignment = center_align; c10.border = thin_border
            if est_rev: c10.number_format = '$#,##0'; c10.font = money_font; c10.fill = alt
            else: c10.fill = no_data_fill
            # Insight
            ws.cell(row=r, column=11, value=insight).fill = alt; ws.cell(row=r, column=11).border = thin_border
            ws.cell(row=r, column=11).font = Font(color=INSIGHT_FONT_COLOR, size=9); ws.cell(row=r, column=11).alignment = content_align
            # History
            ws.cell(row=r, column=12, value=history).fill = alt; ws.cell(row=r, column=12).border = thin_border
            ws.cell(row=r, column=12).font = Font(color=HISTORY_FONT_COLOR, size=9); ws.cell(row=r, column=12).alignment = content_align
            # Meta
            meta_url = create_meta_ads_url(brand_name)
            cm = ws.cell(row=r, column=13, value="Reklamları Gör"); cm.hyperlink = meta_url
            cm.font = Font(color="FFFFFF", bold=True, size=9)
            cm.fill = PatternFill(start_color=META_BUTTON_COLOR, end_color=META_BUTTON_COLOR, fill_type="solid")
            cm.alignment = center_align; cm.border = thin_border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:M{len(brands) + 1}"

    filepath = os.path.join(OUTPUT_DIR, FILENAME)
    wb.save(filepath)
    print(f"\nToplam marka sayısı: {total}")
    print(f"Trafik verisi olan: {brands_with_data}")
    print(f"Excel: {filepath}")
    print(f"Toplam {total} marka, {len(BRANDS)} kategori")
    return filepath

if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    # Import all batch files
    for i in range(1, 20):
        try:
            mod = __import__(f"yuvacim_v3_batch{i}")
            merge_brands(mod.BATCH_DATA)
            print(f"Batch {i} yüklendi")
        except ImportError:
            break
    build_excel()
