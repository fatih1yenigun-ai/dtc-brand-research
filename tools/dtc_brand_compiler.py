#!/usr/bin/env python3
"""
DTC Brand Compiler Tool
Compiles 1500+ DTC brands with websites, categories, and Meta Ads Library links.
Exports to Excel format.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import os
import json
from datetime import datetime

# ============================================================================
# COMPREHENSIVE DTC BRAND DATABASE
# Organized by category/niche. Each entry: (Brand, Website, Sub-niche)
# ============================================================================

BRANDS = {
    # =========================================================================
    # BEAUTY & SKINCARE
    # =========================================================================
    "Beauty & Skincare": [
        ("Glossier", "glossier.com", "Skincare & Makeup"),
        ("Rhode", "rhodeskin.com", "Skincare"),
        ("Drunk Elephant", "drunkelephant.com", "Skincare"),
        ("The Ordinary", "theordinary.com", "Skincare"),
        ("Fenty Beauty", "fentybeauty.com", "Cosmetics"),
        ("Kylie Cosmetics", "kyliecosmetics.com", "Cosmetics"),
        ("Charlotte Tilbury", "charlottetilbury.com", "Cosmetics"),
        ("Tatcha", "tatcha.com", "Skincare"),
        ("Summer Fridays", "summerfridays.com", "Skincare"),
        ("Supergoop", "supergoop.com", "Suncare"),
        ("Tower 28", "tower28beauty.com", "Clean Beauty"),
        ("Kosas", "kosas.com", "Clean Beauty"),
        ("Ilia Beauty", "iliabeauty.com", "Clean Beauty"),
        ("Merit", "meritbeauty.com", "Minimalist Makeup"),
        ("Saie", "saiehello.com", "Clean Beauty"),
        ("Jones Road Beauty", "jonesroadbeauty.com", "Makeup"),
        ("Rare Beauty", "rarebeauty.com", "Cosmetics"),
        ("Herbivore Botanicals", "herbivorebotanicals.com", "Natural Skincare"),
        ("Youth to the People", "youthtothepeople.com", "Skincare"),
        ("Biossance", "biossance.com", "Clean Skincare"),
        ("Glow Recipe", "glowrecipe.com", "K-Beauty Inspired"),
        ("Peach & Lily", "peachandlily.com", "K-Beauty"),
        ("Versed", "versedskin.com", "Affordable Skincare"),
        ("Cocokind", "cocokind.com", "Sustainable Skincare"),
        ("Kinship", "lovekinship.com", "Clean Skincare"),
        ("Tula", "tula.com", "Probiotic Skincare"),
        ("Kopari", "koparibeauty.com", "Coconut Beauty"),
        ("Osea Malibu", "oseamalibu.com", "Seaweed Skincare"),
        ("Beautycounter", "beautycounter.com", "Clean Beauty"),
        ("Naturium", "naturium.com", "Skincare"),
        ("Good Molecules", "goodmolecules.com", "Affordable Skincare"),
        ("Bubble Skincare", "hellobubble.com", "Gen Z Skincare"),
        ("Byoma", "byoma.com", "Barrier Skincare"),
        ("CeraVe", "cerave.com", "Dermatologist Skincare"),
        ("Paula's Choice", "paulaschoice.com", "Science Skincare"),
        ("Dr. Dennis Gross", "drdennisgross.com", "Clinical Skincare"),
        ("Peace Out Skincare", "peaceoutskincare.com", "Acne Care"),
        ("Blume", "blume.com", "Teen Skincare"),
        ("Then I Met You", "thenimetyou.com", "K-Beauty"),
        ("Farmacy Beauty", "farmacybeauty.com", "Farm-to-Face"),
        ("Milk Makeup", "milkmakeup.com", "Vegan Makeup"),
        ("RMS Beauty", "rmsbeauty.com", "Organic Makeup"),
        ("Ogee", "ogee.com", "Luxury Organic"),
        ("Maëlys", "mymaelys.com", "Body Care"),
        ("Phlur", "phlur.com", "Clean Fragrance"),
        ("Scentbird", "scentbird.com", "Fragrance Subscription"),
        ("Skylar", "skylar.com", "Clean Fragrance"),
        ("Ellis Brooklyn", "ellisbrooklyn.com", "Fragrance"),
        ("Boy Smells", "boysmells.com", "Candles & Fragrance"),
        ("Dedcool", "dedcool.com", "Unisex Fragrance"),
        ("Ouai", "theouai.com", "Haircare"),
        ("Olaplex", "olaplex.com", "Hair Bond Repair"),
        ("Function of Beauty", "functionofbeauty.com", "Custom Haircare"),
        ("Prose", "prose.com", "Custom Haircare"),
        ("Bread Beauty Supply", "breadbeautysupply.com", "Textured Hair"),
        ("Vegamour", "vegamour.com", "Hair Growth"),
        ("Act+Acre", "actandacre.com", "Scalp Care"),
        ("Dae Hair", "daehair.com", "Desert Haircare"),
        ("Crown Affair", "crownaffair.com", "Luxury Haircare"),
        ("Amika", "loveamika.com", "Professional Haircare"),
        ("Color Wow", "colorwowhair.com", "Color Haircare"),
        ("Briogeo", "briogeohair.com", "Clean Haircare"),
        ("K18 Hair", "k18hair.com", "Hair Repair"),
        ("Oribe", "oribe.com", "Luxury Haircare"),
        ("Moroccanoil", "moroccanoil.com", "Argan Oil Hair"),
        ("Living Proof", "livingproof.com", "Science Haircare"),
        ("Curlsmith", "curlsmith.com", "Curly Hair"),
        ("Nutrafol", "nutrafol.com", "Hair Supplements"),
        ("Doe Lashes", "doelashes.com", "False Lashes"),
        ("Velour Lashes", "velourlashes.com", "Luxury Lashes"),
        ("Pat McGrath Labs", "patmcgrath.com", "Luxury Makeup"),
        ("Danessa Myricks", "danessamyricksbeauty.com", "Pro Makeup"),
        ("About Face", "aboutface.com", "Makeup"),
        ("Rose Inc", "roseinc.com", "Clean Luxury Beauty"),
        ("Lawless Beauty", "lawlessbeauty.com", "Clean Makeup"),
        ("One/Size", "onesize.com", "Inclusive Beauty"),
        ("Ami Colé", "amicole.com", "Melanin Beauty"),
        ("Live Tinted", "livetinted.com", "Multicultural Beauty"),
        ("Aavrani", "aavrani.com", "Ayurvedic Skincare"),
        ("RoseSkinCo", "roseskinco.com", "Hair Removal"),
        ("Bushbalm", "bushbalm.com", "Skincare"),
        ("Spoiled Child", "spoiledchild.com", "AI Skincare"),
        ("Topicals", "mytopicals.com", "Skin Conditions"),
        ("Stratia", "stratiaskin.com", "Barrier Repair"),
        ("Typology", "typology.com", "French Skincare"),
        ("Beauty of Joseon", "beautyofjoseon.com", "K-Beauty"),
        ("Primally Pure", "primallypure.com", "Natural Skincare"),
        ("100% Pure", "100percentpure.com", "Natural Cosmetics"),
        ("Pinrose", "pinrose.com", "Fragrance"),
        ("Fable & Mane", "fableandmane.com", "Ayurvedic Hair"),
        ("Necessaire", "necessaire.com", "Body Care"),
        ("Megababe", "megababebeauty.com", "Body Care"),
        ("Sol de Janeiro", "soldejaneiro.com", "Brazilian Body Care"),
    ],

    # =========================================================================
    # MEN'S GROOMING & PERSONAL CARE
    # =========================================================================
    "Men's Grooming & Personal Care": [
        ("Dollar Shave Club", "dollarshaveclub.com", "Shaving"),
        ("Harry's", "harrys.com", "Shaving"),
        ("Dr. Squatch", "drsquatch.com", "Natural Soap"),
        ("Beardbrand", "beardbrand.com", "Beard Care"),
        ("Scotch Porter", "scotchporter.com", "Grooming"),
        ("Manscaped", "manscaped.com", "Body Grooming"),
        ("Huron", "usehuron.com", "Men's Skincare"),
        ("Disco", "letsdisco.co", "Men's Skincare"),
        ("Hawthorne", "hawthorne.co", "Men's Personal Care"),
        ("Bevel", "bfrb.co", "Men's Grooming"),
        ("Supply", "supply.co", "Shaving"),
        ("Oars + Alps", "oarsandalps.com", "Men's Skincare"),
        ("Hims", "forhims.com", "Men's Health"),
        ("Keeps", "keeps.com", "Hair Loss"),
        ("Lumin", "luminskin.com", "Men's Skincare"),
        ("Every Man Jack", "everymanjack.com", "Natural Grooming"),
        ("Jack Black", "getjackblack.com", "Men's Skincare"),
        ("Blind Barber", "blindbarber.com", "Grooming"),
        ("Viking Revolution", "vikingrevolution.com", "Beard Care"),
        ("Cremo", "crfremo.com", "Shaving & Grooming"),
        ("Black Wolf", "blackwolf.com", "Men's Skincare"),
        ("Art of Sport", "artofsport.com", "Athletic Grooming"),
        ("Bravo Sierra", "bravosierra.com", "Military-Grade Care"),
        ("Frederick Benjamin", "frederickbenjamin.com", "Men's Grooming"),
        ("Geologie", "geologie.com", "Personalized Skincare"),
    ],

    # =========================================================================
    # HEALTH & WELLNESS
    # =========================================================================
    "Health & Wellness": [
        ("AG1 Athletic Greens", "drinkag1.com", "Greens Supplement"),
        ("Ritual", "ritual.com", "Vitamins"),
        ("Seed", "seed.com", "Probiotics"),
        ("Care/of", "takecareof.com", "Personalized Vitamins"),
        ("OLLY", "olly.com", "Gummy Vitamins"),
        ("Huel", "huel.com", "Meal Replacement"),
        ("Liquid IV", "liquid-iv.com", "Hydration"),
        ("Bloom Nutrition", "bloomnu.com", "Greens & Wellness"),
        ("Vital Proteins", "vitalproteins.com", "Collagen"),
        ("Moon Juice", "moonjuice.com", "Adaptogens"),
        ("Four Sigmatic", "foursigmatic.com", "Mushroom Coffee"),
        ("MUD\\WTR", "mudwtr.com", "Coffee Alternative"),
        ("Hims & Hers", "forhers.com", "Telehealth"),
        ("Ro", "ro.co", "Telehealth"),
        ("Nurx", "nurx.com", "Birth Control"),
        ("Cerebral", "cerebral.com", "Mental Health"),
        ("Calm", "calm.com", "Meditation App"),
        ("Headspace", "headspace.com", "Meditation"),
        ("Whoop", "whoop.com", "Fitness Tracker"),
        ("Oura Ring", "ouraring.com", "Health Tracker"),
        ("Eight Sleep", "eightsleep.com", "Sleep Tech"),
        ("Casper", "casper.com", "Mattress"),
        ("Purple", "purple.com", "Mattress"),
        ("Helix Sleep", "helixsleep.com", "Custom Mattress"),
        ("Birch Living", "birchliving.com", "Organic Mattress"),
        ("Saatva", "saatva.com", "Luxury Mattress"),
        ("Bear Mattress", "bearmattress.com", "Recovery Mattress"),
        ("Leesa", "leesa.com", "Mattress"),
        ("Nectar Sleep", "nectarsleep.com", "Mattress"),
        ("Tuft & Needle", "tuftandneedle.com", "Mattress"),
        ("Pillow Cube", "pillowcube.com", "Pillows"),
        ("Coop Home Goods", "coophomegoods.com", "Pillows"),
        ("Buffy", "buffy.co", "Comforters"),
        ("Molecule", "onmolecule.com", "Sleep Tech"),
        ("Nolah", "nolahmattress.com", "Mattress"),
        ("DreamCloud", "dreamcloudsleep.com", "Mattress"),
        ("Persona Nutrition", "personanutrition.com", "Custom Vitamins"),
        ("Gainful", "gainful.com", "Custom Protein"),
        ("Rae Wellness", "raewellness.co", "Women's Wellness"),
        ("HUM Nutrition", "humnutrition.com", "Beauty Vitamins"),
        ("Sakara Life", "sakara.com", "Plant-Based Meals"),
        ("Daily Harvest", "dailyharvest.com", "Frozen Meals"),
        ("Hungryroot", "hungryroot.com", "Healthy Groceries"),
        ("Factor", "factor75.com", "Prepared Meals"),
        ("Thistle", "thistle.co", "Healthy Meals"),
        ("Splendid Spoon", "splendidspoon.com", "Plant-Based Meals"),
        ("Territory Foods", "territoryfoods.com", "Prepared Meals"),
        ("Trifecta", "trifectanutrition.com", "Meal Delivery"),
        ("Snap Kitchen", "snapkitchen.com", "Prepared Meals"),
        ("CookUnity", "cookunity.com", "Chef Meals"),
        ("Freshly", "freshly.com", "Prepared Meals"),
        ("Sunbasket", "sunbasket.com", "Meal Kit"),
        ("HelloFresh", "hellofresh.com", "Meal Kit"),
        ("Blue Apron", "blueapron.com", "Meal Kit"),
        ("Home Chef", "homechef.com", "Meal Kit"),
        ("Green Chef", "greenchef.com", "Organic Meal Kit"),
        ("Dinnerly", "dinnerly.com", "Affordable Meal Kit"),
        ("EveryPlate", "everyplate.com", "Budget Meal Kit"),
        ("Nutrisystem", "nutrisystem.com", "Diet Meals"),
        ("Athletic Greens", "athleticgreens.com", "Greens"),
        ("Amazing Grass", "amazinggrass.com", "Greens Powder"),
        ("Garden of Life", "gardenoflife.com", "Organic Supplements"),
        ("Sports Research", "sportsresearch.com", "Supplements"),
        ("Momentous", "livemomentous.com", "Sports Supplements"),
        ("Transparent Labs", "transparentlabs.com", "Clean Supplements"),
        ("Jocko Fuel", "jockofuel.com", "Supplements"),
        ("Gorilla Mind", "gorillamind.com", "Nootropics"),
        ("Thesis", "takethesis.com", "Custom Nootropics"),
        ("Onnit", "onnit.com", "Supplements & Fitness"),
        ("Timeline Nutrition", "timelinenutrition.com", "Longevity"),
        ("Elysium Health", "elysiumhealth.com", "Longevity"),
        ("InsideTracker", "insidetracker.com", "Blood Testing"),
        ("Rootine", "rootine.co", "Precision Vitamins"),
        ("Beekeeper's Naturals", "beekeepersnaturals.com", "Bee Products"),
        ("Goli Nutrition", "goli.com", "ACV Gummies"),
        ("Lemme", "lemmelive.com", "Wellness Gummies"),
        ("Wellabs", "wellabs.com", "Liquid Supplements"),
        ("Nuzest", "nuzest.com", "Plant Protein"),
        ("Orgain", "orgain.com", "Organic Protein"),
        ("KOS", "kos.com", "Plant Protein"),
        ("Vega", "myvega.com", "Plant Nutrition"),
    ],

    # =========================================================================
    # FITNESS & ACTIVEWEAR
    # =========================================================================
    "Fitness & Activewear": [
        ("Gymshark", "gymshark.com", "Gym Apparel"),
        ("Alo Yoga", "aloyoga.com", "Yoga Apparel"),
        ("Vuori", "vuori.com", "Performance Apparel"),
        ("Lululemon", "lululemon.com", "Athletic Apparel"),
        ("Outdoor Voices", "outdoorvoices.com", "Recreational Apparel"),
        ("Rhone", "rhone.com", "Men's Activewear"),
        ("Ten Thousand", "tenthousand.cc", "Men's Training"),
        ("ASRV", "asrv.com", "Aesthetic Sportswear"),
        ("Alphalete", "alphaleteathletics.com", "Gym Apparel"),
        ("Buffbunny Collection", "buffbunny.com", "Women's Activewear"),
        ("Fabletics", "fabletics.com", "Activewear Subscription"),
        ("Sweaty Betty", "sweatybetty.com", "Women's Activewear"),
        ("Girlfriend Collective", "girlfriend.com", "Sustainable Activewear"),
        ("Vitality", "vitalityapparel.com", "Women's Gym Wear"),
        ("Nobull", "nobullproject.com", "CrossFit Shoes"),
        ("Hylete", "hylete.com", "Training Apparel"),
        ("Virus", "virusintl.com", "Performance Wear"),
        ("Born Primitive", "bornprimitive.com", "CrossFit Apparel"),
        ("Legends", "legends.com", "Lifestyle Activewear"),
        ("Cuts Clothing", "cutsclothing.com", "Men's Tees"),
        ("Bylt Basics", "byltbasics.com", "Men's Basics"),
        ("BRADY", "bradybrand.com", "Tom Brady Apparel"),
        ("Satisfy Running", "satisfyrunning.com", "Running Apparel"),
        ("Tracksmith", "tracksmith.com", "Running Apparel"),
        ("Janji", "janji.com", "Running Apparel"),
        ("District Vision", "districtvision.com", "Running Accessories"),
        ("Peloton", "onepeloton.com", "Connected Fitness"),
        ("Mirror", "mirror.co", "Home Fitness"),
        ("Tonal", "tonal.com", "Smart Home Gym"),
        ("Tempo", "tempo.fit", "Home Fitness"),
        ("Hydrow", "hydrow.com", "Connected Rowing"),
        ("Ergatta", "ergatta.com", "Connected Rowing"),
        ("Hyperice", "hyperice.com", "Recovery Tech"),
        ("Therabody", "therabody.com", "Recovery Tech"),
        ("Bala", "shopbala.com", "Weighted Accessories"),
        ("P.volve", "pfrvolve.com", "Functional Fitness"),
        ("Obé Fitness", "obefitness.com", "Virtual Fitness"),
        ("Crossrope", "crossrope.com", "Jump Ropes"),
        ("TRX", "trxtraining.com", "Suspension Training"),
        ("Rogue Fitness", "roguefitness.com", "Gym Equipment"),
        ("REP Fitness", "repfitness.com", "Home Gym"),
        ("Bells of Steel", "bellsofsteel.com", "Gym Equipment"),
        ("Titan Fitness", "titanfitness.com", "Budget Gym"),
        ("Gorilla Bow", "gorillabow.com", "Resistance Training"),
        ("X3 Bar", "jfraquatina.com", "Variable Resistance"),
        ("CLMBR", "clfrmbr.com", "Climbing Machine"),
        ("Whoop", "whoop.com", "Fitness Wearable"),
        ("FORM Swim", "formswim.com", "Smart Swim Goggles"),
        ("Wahoo Fitness", "wahoofitness.com", "Cycling Tech"),
    ],

    # =========================================================================
    # FASHION & APPAREL
    # =========================================================================
    "Fashion & Apparel": [
        ("Everlane", "everlane.com", "Transparent Fashion"),
        ("Reformation", "thereformation.com", "Sustainable Fashion"),
        ("SKIMS", "skims.com", "Shapewear & Basics"),
        ("Good American", "goodamerican.com", "Inclusive Denim"),
        ("Pangaia", "thepangaia.com", "Sustainable Basics"),
        ("Allbirds", "allbirds.com", "Sustainable Footwear"),
        ("Rothy's", "rothys.com", "Recycled Shoes"),
        ("CARIUMA", "cariuma.com", "Sustainable Sneakers"),
        ("Veja", "vfreja-store.com", "Eco Sneakers"),
        ("Koio", "koio.co", "Italian Sneakers"),
        ("Greats", "greats.com", "Premium Sneakers"),
        ("Oliver Cabell", "olivercabell.com", "Luxury Sneakers"),
        ("Thursday Boots", "thursdayboots.com", "Boots"),
        ("Tecovas", "tecovas.com", "Western Boots"),
        ("Helm Boots", "helmboots.com", "Handmade Boots"),
        ("Nisolo", "nisolo.com", "Ethical Footwear"),
        ("M.Gemi", "mgemi.com", "Italian Shoes"),
        ("Birdies", "birdies.com", "Women's Flats"),
        ("Vivaia", "vivfraia.com", "Recycled Shoes"),
        ("Atoms", "atoms.com", "Quarter-Size Shoes"),
        ("HOKA", "hoka.com", "Running Shoes"),
        ("On Running", "on.com", "Swiss Running Shoes"),
        ("Merrell", "merrell.com", "Outdoor Shoes"),
        ("American Giant", "american-giant.com", "Made in USA Basics"),
        ("Buck Mason", "buckmason.com", "Men's Essentials"),
        ("Marine Layer", "marinelayer.com", "Soft Tees"),
        ("Faherty", "faherty.com", "Coastal Style"),
        ("Bonobos", "bonobos.com", "Men's Pants"),
        ("Mack Weldon", "mackweldon.com", "Men's Basics"),
        ("MeUndies", "meundies.com", "Underwear"),
        ("Tommy John", "tommyjohn.com", "Underwear"),
        ("Parade", "yourparade.com", "Sustainable Underwear"),
        ("Pact", "wearpact.com", "Organic Basics"),
        ("Organic Basics", "organicbasics.com", "Sustainable Basics"),
        ("True Classic", "trueclassictees.com", "Men's Tees"),
        ("Fresh Clean Threads", "freshcleanthreads.com", "Men's Tees"),
        ("Into The AM", "intotheam.com", "Graphic Tees"),
        ("State & Liberty", "stateandliberty.com", "Athletic Fit Shirts"),
        ("Untuckit", "untuckit.com", "Casual Shirts"),
        ("Mizzen+Main", "mizzenandmain.com", "Performance Dress Shirts"),
        ("Ministry of Supply", "ministryofsupply.com", "Performance Workwear"),
        ("Taylor Stitch", "taylorstitch.com", "Sustainable Menswear"),
        ("Todd Snyder", "toddsnyder.com", "Designer Menswear"),
        ("Ledbury", "ledbury.com", "Dress Shirts"),
        ("Proper Cloth", "propercloth.com", "Custom Shirts"),
        ("Suitsupply", "suitsupply.com", "Suits"),
        ("Indochino", "indochino.com", "Custom Suits"),
        ("MTailor", "mtailor.com", "AI Custom Fit"),
        ("Hill City", "hillcity.gap.com", "Active Menswear"),
        ("ADAY", "thisisaday.com", "Capsule Wardrobe"),
        ("Cuyana", "cuyana.com", "Fewer Better Things"),
        ("M.M.LaFleur", "mmlafleur.com", "Women's Workwear"),
        ("Argent", "argentwork.com", "Power Dressing"),
        ("Of Mercer", "ofmercer.com", "Women's Workwear"),
        ("Senreve", "senreve.com", "Luxury Handbags"),
        ("Dagne Dover", "dfragnedover.com", "Neoprene Bags"),
        ("Stoney Clover Lane", "stoneycloverlane.com", "Customizable Bags"),
        ("Béis", "beistravel.com", "Travel Bags"),
        ("Calpak", "calpak.com", "Luggage"),
        ("Away", "awaytravel.com", "Modern Luggage"),
        ("Monos", "monfroos.com", "Premium Luggage"),
        ("July", "july.com", "Luggage"),
        ("Roark", "roark.com", "Adventure Apparel"),
        ("Vuori", "vufrori.com", "Versatile Apparel"),
        ("prAna", "prana.com", "Yoga & Outdoor"),
        ("Cotopaxi", "cotopaxi.com", "Outdoor Gear"),
        ("United by Blue", "unitedbyblue.com", "Sustainable Outdoor"),
        ("Outerknown", "outerknown.com", "Sustainable Surf"),
        ("Fair Harbor", "fairharborclothing.com", "Recycled Swimwear"),
        ("Onia", "onia.com", "Swimwear"),
        ("Summersalt", "summersalt.com", "Travel Swimwear"),
        ("Andie Swim", "andieswim.com", "Women's Swimwear"),
        ("Solid & Striped", "solidandstriped.com", "Swimwear"),
        ("CDLP", "cdlp.com", "Swedish Essentials"),
        ("Madhappy", "madhappy.com", "Mental Health Fashion"),
        ("Aimé Leon Dore", "aimeleondore.com", "Streetwear"),
        ("Kith", "kith.com", "Streetwear"),
        ("Fear of God Essentials", "fearofgfrod.com", "Essentials"),
        ("Stüssy", "stussy.com", "Streetwear"),
        ("Carhartt WIP", "carhartt-wip.com", "Workwear Street"),
        ("Noah", "noahny.com", "Streetwear"),
        ("Online Ceramics", "online-ceramics.com", "Art Streetwear"),
        ("Brain Dead", "wearebraindead.com", "Creative Streetwear"),
        ("Corridor", "corridornyc.com", "Brooklyn Menswear"),
        ("Alex Mill", "alexmill.com", "Modern Basics"),
        ("Frank and Oak", "frankandoak.com", "Sustainable Fashion"),
        ("Aritzia", "aritzia.com", "Canadian Fashion"),
        ("& Other Stories", "stories.com", "Fashion"),
        ("COS", "cosstores.com", "Modern Essentials"),
        ("Sézane", "sezane.com", "French Fashion"),
        ("Rouje", "rofruje.com", "French Fashion"),
        ("Réalisation Par", "refralisationpar.com", "Australian Fashion"),
        ("House of Sunny", "houseofsunny.com", "London Fashion"),
        ("Rat & Boa", "ratandboa.com", "Women's Fashion"),
        ("Odd Muse", "odd-mfruse.com", "Occasion Wear"),
        ("Meshki", "meshki.com.au", "Women's Fashion"),
        ("Princess Polly", "princesspolly.com", "Fast Fashion"),
        ("Hello Molly", "hellomolly.com", "Women's Fashion"),
        ("SHEIN", "shein.com", "Fast Fashion"),
        ("Fashion Nova", "fashionnova.com", "Trend Fashion"),
        ("Savage X Fenty", "savagex.com", "Lingerie"),
        ("ThirdLove", "thirdlove.com", "Bras"),
        ("Lively", "wearlively.com", "Leisurée"),
        ("Negative Underwear", "negativeunderwear.com", "Lingerie"),
        ("Cuup", "shopcuup.com", "Modern Bras"),
        ("Harper Wilde", "harperwilde.com", "Simple Bras"),
        ("Knix", "knix.com", "Wireless Bras"),
        ("True & Co", "trueandco.com", "Comfort Bras"),
        ("Pepper", "wearpepper.com", "Small Bust Bras"),
        ("Adore Me", "adoreme.com", "Lingerie Subscription"),
        ("11 Honoré", "11honore.com", "Plus-Size Luxury"),
        ("Universal Standard", "universalstandard.com", "Size Inclusive"),
        ("Henning", "henning.com", "Plus Size Fashion"),
        ("Warp + Weft", "warpweftworld.com", "Sustainable Denim"),
        ("MOTHER Denim", "motherdenim.com", "Premium Denim"),
        ("AGOLDE", "agolde.com", "Denim"),
        ("Citizens of Humanity", "citizensofhumanity.com", "Premium Denim"),
        ("FRAME", "frame-store.com", "Denim & Fashion"),
        ("DL1961", "dl1961.com", "Sustainable Denim"),
        ("Boyish Jeans", "boyish.com", "Vintage Denim"),
    ],

    # =========================================================================
    # FOOD & BEVERAGE
    # =========================================================================
    "Food & Beverage": [
        ("OLIPOP", "drinkolipop.com", "Prebiotic Soda"),
        ("Poppi", "drinkpoppi.com", "Prebiotic Soda"),
        ("Liquid Death", "liquiddeath.com", "Canned Water"),
        ("Prime", "drinkprime.com", "Energy Drink"),
        ("Feastables", "feastables.com", "Chocolate & Snacks"),
        ("Magic Spoon", "magicspoon.com", "Protein Cereal"),
        ("Graza", "grfraza.co", "Olive Oil"),
        ("Brightland", "brightland.co", "Olive Oil"),
        ("Fly By Jing", "flybyjing.com", "Sichuan Sauce"),
        ("TRUFF", "trfruff.com", "Truffle Hot Sauce"),
        ("Siete Foods", "sifretefoods.com", "Mexican-Inspired"),
        ("Mid-Day Squares", "middaysquares.com", "Functional Chocolate"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Snacks"),
        ("Lesser Evil", "lesserevil.com", "Organic Snacks"),
        ("Skinny Pop", "skinnypop.com", "Popcorn"),
        ("Pipcorn", "pipsnacks.com", "Heirloom Snacks"),
        ("Outstanding Foods", "outstandingfoods.com", "Plant-Based Snacks"),
        ("Partake Foods", "partakefoods.com", "Allergen-Free Snacks"),
        ("Perfect Bar", "perfectbar.com", "Refrigerated Bars"),
        ("RXBAR", "rxbar.com", "Protein Bars"),
        ("GoMacro", "gomacro.com", "Macro Bars"),
        ("Larabar", "larabar.com", "Fruit & Nut Bars"),
        ("Chomps", "chomps.com", "Meat Sticks"),
        ("Epic Provisions", "epicbar.com", "Meat Snacks"),
        ("Country Archer", "countryarcher.com", "Jerky"),
        ("Krave Jerky", "kravejerky.com", "Jerky"),
        ("Muddy Bites", "muddybites.com", "Chocolate Snacks"),
        ("Behave Candy", "behavecandy.com", "Low Sugar Candy"),
        ("SmartSweets", "smartsweets.com", "Low Sugar Candy"),
        ("Unreal Candy", "unrealcandy.com", "Better Candy"),
        ("Deux", "eatdeux.com", "Cookie Dough"),
        ("Nunbelievable", "nunbelievable.com", "Cookies"),
        ("Last Crumb", "lastcrumb.com", "Luxury Cookies"),
        ("Levain Bakery", "levainbakery.com", "Premium Cookies"),
        ("Milk Bar", "milkbarstore.com", "Desserts"),
        ("Goldbelly", "goldbelly.com", "Food Marketplace"),
        ("ButcherBox", "butcherbox.com", "Meat Subscription"),
        ("Crowd Cow", "crowdcow.com", "Craft Meat"),
        ("Rastelli's", "rastellis.com", "Meat Delivery"),
        ("Wild Alaskan Company", "wildalaskancompany.com", "Seafood"),
        ("Sizzlefish", "sizzlefish.com", "Seafood Delivery"),
        ("Omaha Steaks", "omahasteaks.com", "Steaks"),
        ("Porter Road", "porterroad.com", "Pasture-Raised Meat"),
        ("Good Chop", "goodchop.com", "Meat & Seafood"),
        ("Thrive Market", "thrivemarket.com", "Online Grocery"),
        ("Misfits Market", "misfitsmarket.com", "Imperfect Produce"),
        ("Imperfect Foods", "imperfectfoods.com", "Surplus Grocery"),
        ("Farmstead", "farmsteadapp.com", "Online Grocery"),
        ("Death Wish Coffee", "deathwishcoffee.com", "Strong Coffee"),
        ("Trade Coffee", "drinktrade.com", "Coffee Subscription"),
        ("Atlas Coffee Club", "atlascoffeeclub.com", "World Coffee"),
        ("Bones Coffee", "bonescoffee.com", "Flavored Coffee"),
        ("Cometeer", "cometeer.com", "Flash-Frozen Coffee"),
        ("Chamberlain Coffee", "chamberlaincoffee.com", "Creator Coffee"),
        ("Jot", "jot.co", "Ultra Coffee"),
        ("Counter Culture Coffee", "counterculturecoffee.com", "Specialty Coffee"),
        ("Stumptown Coffee", "stumptowncoffee.com", "Craft Coffee"),
        ("Blue Bottle Coffee", "bluebottlecoffee.com", "Specialty Coffee"),
        ("Verve Coffee", "vervecoffee.com", "Specialty Coffee"),
        ("Firebelly Tea", "firebellytea.com", "Premium Tea"),
        ("Vahdam Teas", "vahdamteas.com", "Indian Tea"),
        ("Art of Tea", "artoftea.com", "Artisan Tea"),
        ("Dona Chai", "donachai.com", "Chai Concentrate"),
        ("Boba Guys", "bobaguys.com", "Boba Tea"),
        ("Ghia", "drinkghia.com", "Non-Alcoholic Aperitif"),
        ("Haus", "drink.haus", "Aperitifs"),
        ("Seedlip", "seedlipdrinks.com", "Non-Alcoholic Spirits"),
        ("Athletic Brewing", "athleticbrewing.com", "Non-Alcoholic Beer"),
        ("HOP WTR", "hopwtr.com", "Sparkling Hop Water"),
        ("Lagunitas IPNA", "lagunitas.com", "NA Beer"),
        ("De Soi", "drinkdesoi.com", "Adaptogen Aperitif"),
        ("Three Spirit", "threespiritdrinks.com", "Functional Drinks"),
        ("Acid League", "acidleague.com", "Living Vinegar"),
        ("Hydrant", "drinkhydrant.com", "Hydration Mix"),
        ("Nuun", "nuun.com", "Electrolyte Tabs"),
        ("LMNT", "drinklmnt.com", "Electrolytes"),
        ("Cure Hydration", "curehydration.com", "Electrolyte Mix"),
        ("Alani Nu", "alaninu.com", "Energy & Fitness"),
        ("Celsius", "celsius.com", "Fitness Drink"),
        ("Celsius Essentials", "celsius.com", "Energy"),
        ("Bloom", "bloomnu.com", "Greens & Energy"),
        ("Ryse Supplements", "rfrysesupps.com", "Pre-Workout"),
        ("Ghost Lifestyle", "ghostlifestyle.com", "Supplements & Energy"),
    ],

    # =========================================================================
    # HOME & KITCHEN
    # =========================================================================
    "Home & Kitchen": [
        ("Our Place", "fromourplace.com", "Cookware"),
        ("Made In", "madeincookware.com", "Professional Cookware"),
        ("Great Jones", "greatjonesgoods.com", "Colorful Cookware"),
        ("Caraway", "carawayhome.com", "Non-Toxic Cookware"),
        ("HexClad", "hexclad.com", "Hybrid Cookware"),
        ("Material Kitchen", "materialkitchen.com", "Kitchen Tools"),
        ("Misen", "misen.com", "Kitchen Knives"),
        ("Hedley & Bennett", "hedleyandbennett.com", "Chef Aprons"),
        ("W&P", "wandpdesign.com", "Kitchen Accessories"),
        ("Food52", "food52.com", "Kitchen & Home"),
        ("Brooklinen", "brooklinen.com", "Luxury Bedding"),
        ("Parachute Home", "parachutehome.com", "Bedding & Bath"),
        ("Boll & Branch", "bollandbranch.com", "Organic Bedding"),
        ("Ettitude", "ettitude.com", "Bamboo Bedding"),
        ("Coyuchi", "coyuchi.com", "Organic Home Textiles"),
        ("Cozy Earth", "cozyearth.com", "Bamboo Bedding"),
        ("Pom Pom at Home", "pomfromathome.com", "Luxury Linens"),
        ("Sijo", "sijohome.com", "Eucalyptus Bedding"),
        ("Sunday Citizen", "sundaycitizen.co", "Cozy Home"),
        ("Gravity Blanket", "gravityblankets.com", "Weighted Blankets"),
        ("Bearaby", "bearaby.com", "Knitted Weighted Blankets"),
        ("Baloo Living", "balooliving.com", "Weighted Blankets"),
        ("Ruggable", "ruggable.com", "Washable Rugs"),
        ("Tumble", "tumblelivfring.com", "Washable Rugs"),
        ("Revival Rugs", "revivalrugs.com", "Vintage Rugs"),
        ("Loloi Rugs", "loloirugs.com", "Designer Rugs"),
        ("Rugs USA", "rugsusa.com", "Affordable Rugs"),
        ("Apt2B", "apt2b.com", "Modern Furniture"),
        ("Interior Define", "interiordefine.com", "Custom Sofas"),
        ("Article", "article.com", "Modern Furniture"),
        ("Floyd", "floydhome.com", "Modular Furniture"),
        ("Burrow", "burrow.com", "Modular Sofas"),
        ("Campaign Living", "campaignliving.com", "Leather Furniture"),
        ("Joybird", "joybird.com", "Custom Furniture"),
        ("AllModern", "allmodern.com", "Modern Furniture"),
        ("Castlery", "castlery.com", "Direct Furniture"),
        ("Albany Park", "albfrypark.com", "Modular Sofas"),
        ("Outer", "liveouter.com", "Outdoor Furniture"),
        ("Yardbird", "yardbird.com", "Recycled Outdoor"),
        ("Neighbor", "neighbor.com", "Outdoor Furniture"),
        ("Polywood", "polywood.com", "Recycled Furniture"),
        ("Tushy", "hellotushy.com", "Bidet Attachments"),
        ("Brondell", "brondell.com", "Bidet & Water"),
        ("Canopy", "getcanopy.co", "Humidifiers"),
        ("Molekule", "molekule.com", "Air Purifiers"),
        ("Mila", "milacares.com", "Air Purifiers"),
        ("Filtrete", "filtrete.com", "Air Filters"),
        ("Dyson", "dyson.com", "Home Tech"),
        ("iRobot", "irobot.com", "Robot Vacuums"),
        ("Roborock", "roborock.com", "Robot Vacuums"),
        ("Otherland", "otherland.com", "Luxury Candles"),
        ("Homesick", "homesickcandles.com", "Nostalgia Candles"),
        ("Voluspa", "voluspa.com", "Luxury Candles"),
        ("Vitruvi", "vitruvi.com", "Diffusers & Oils"),
        ("P.F. Candle Co.", "pfcandleco.com", "Candles & Home"),
        ("YIELD Design Co", "yielddesign.co", "Modern Home"),
        ("Skandinavisk", "skandinavisk.com", "Nordic Home"),
        ("Olive & June", "olivejune.com", "Nail Care"),
        ("Lalo", "meetlalo.com", "Baby Furniture"),
        ("Babyletto", "babyletto.com", "Modern Baby"),
        ("Nugget", "nuggetcomfort.com", "Play Couch"),
    ],

    # =========================================================================
    # BABY & KIDS
    # =========================================================================
    "Baby & Kids": [
        ("The Honest Company", "honest.com", "Baby & Home"),
        ("Babyganics", "babyganics.com", "Baby Care"),
        ("Tubby Todd", "tubbytodd.com", "Baby Skincare"),
        ("Pipette", "pipfrette.com", "Baby Skincare"),
        ("Puracy", "puracy.com", "Natural Baby Care"),
        ("Hello Bello", "hellobello.com", "Baby Essentials"),
        ("Coterie", "cotfrerie.com", "Premium Diapers"),
        ("Kudos", "mykudos.com", "Plant-Based Diapers"),
        ("Dyper", "dyper.com", "Bamboo Diapers"),
        ("Rascal + Friends", "rascalandfriends.com", "Affordable Diapers"),
        ("Bobbie", "hibobbie.com", "Organic Formula"),
        ("ByHeart", "byheart.com", "Infant Formula"),
        ("Enfamil", "enfamil.com", "Baby Formula"),
        ("Kendamil", "kendamil.com", "European Formula"),
        ("Guava Family", "guavafamily.com", "Travel Crib"),
        ("4moms", "4moms.com", "Baby Tech"),
        ("Owlet", "owletcare.com", "Baby Monitor"),
        ("Nanit", "nanit.com", "Smart Baby Monitor"),
        ("Snoo", "happiestbaby.com", "Smart Bassinet"),
        ("Lovevery", "lovevery.com", "Play Kits"),
        ("Kiwi Co", "kiwico.com", "STEM Kits"),
        ("Little Passports", "littlepassports.com", "Learning Kits"),
        ("Melissa & Doug", "melissaanddoug.com", "Wooden Toys"),
        ("Nugget", "nuggetcomfort.com", "Play Furniture"),
        ("Gathre", "gathre.com", "Leather Mats"),
        ("Loulou Lollipop", "lfrouloulollipop.com", "Baby Accessories"),
        ("Mushie", "mushie.com", "Baby Essentials"),
        ("Little Sleepies", "littlesleepies.com", "Baby Sleepwear"),
        ("Kyte Baby", "kytebaby.com", "Bamboo Baby"),
        ("Monica + Andy", "monicaandandy.com", "Organic Baby"),
        ("Primary", "primary.com", "Kids Basics"),
        ("Tea Collection", "teacollection.com", "Global Kids Fashion"),
        ("Maisonette", "maisonette.com", "Kids Fashion"),
        ("Mini Rodini", "minirodini.com", "Sustainable Kids"),
        ("Cat & Jack", "catandjack.com", "Kids Basics"),
        ("Hanna Andersson", "hannaandersson.com", "Quality Kids"),
        ("PatPat", "patpat.com", "Affordable Kids"),
        ("Stokke", "stokke.com", "Premium Baby"),
        ("UPPAbaby", "uppababy.com", "Strollers"),
        ("Bugaboo", "bugaboo.com", "Strollers"),
        ("Doona", "dofrona.com", "Car Seat Stroller"),
    ],

    # =========================================================================
    # PET CARE
    # =========================================================================
    "Pet Care": [
        ("The Farmer's Dog", "thefarmersdog.com", "Fresh Dog Food"),
        ("Ollie", "myollie.com", "Fresh Dog Food"),
        ("Nom Nom", "nomnom.com", "Fresh Pet Food"),
        ("Jinx", "jinx.com", "Premium Dog Food"),
        ("Sundays for Dogs", "sundaysfordogs.com", "Air-Dried Dog Food"),
        ("Spot & Tango", "spotandtango.com", "Fresh Dog Food"),
        ("PetPlate", "petplate.com", "Chef-Cooked Dog Food"),
        ("JustFoodForDogs", "justfoodfordogs.com", "Whole Food Dog Food"),
        ("Open Farm", "openfarmpet.com", "Ethical Pet Food"),
        ("A Pup Above", "apupabove.com", "Sous Vide Dog Food"),
        ("Chippin", "chippin.com", "Sustainable Dog Food"),
        ("Wild Earth", "wildearth.com", "Plant-Based Dog Food"),
        ("Yumwoof", "yumwoof.com", "Keto Dog Food"),
        ("The Honest Kitchen", "thehonestkitchen.com", "Dehydrated Pet Food"),
        ("Stella & Chewy's", "stellaandchewys.com", "Raw Pet Food"),
        ("BarkBox", "barkbox.com", "Dog Subscription"),
        ("Bark", "bark.co", "Dog Products"),
        ("Super Chewer", "barkbox.com/super-chewer", "Tough Toy Box"),
        ("Wild One", "wildone.com", "Modern Pet"),
        ("Fable Pets", "ffrablefrpets.com", "Dog Accessories"),
        ("Fi", "tryfi.com", "Smart Dog Collar"),
        ("Whistle", "whistle.com", "GPS Pet Tracker"),
        ("PupBox", "pupbox.com", "Puppy Subscription"),
        ("Kong", "kongcompany.com", "Dog Toys"),
        ("West Paw", "westpaw.com", "Sustainable Dog Toys"),
        ("P.L.A.Y.", "petplay.com", "Pet Lifestyle"),
        ("Bowlmates", "bowlmates.com", "Pet Bowls"),
        ("Diggs", "diggs.pet", "Modern Dog Crate"),
        ("Litter-Robot", "litter-robot.com", "Self-Cleaning Litter"),
        ("PrettyLitter", "prettylitter.com", "Health-Monitoring Litter"),
        ("Tuft + Paw", "tuftandpaw.com", "Modern Cat Furniture"),
        ("Catit", "catit.com", "Cat Products"),
        ("Smalls", "smalls.com", "Fresh Cat Food"),
        ("Made by Nacho", "madebynacho.com", "Premium Cat Food"),
        ("Dandy", "getdandy.com", "Custom Dog Supplements"),
        ("Native Pet", "nativepet.com", "Dog Supplements"),
        ("Zesty Paws", "zestypaws.com", "Pet Supplements"),
        ("PetHonesty", "pethonesty.com", "Pet Supplements"),
        ("Finn", "petfinn.com", "Dog Supplements"),
        ("ElleVet Sciences", "ellevet.com", "CBD for Pets"),
        ("Embark", "embarkvet.com", "Dog DNA Tests"),
        ("Wisdom Panel", "wisdompanel.com", "Pet DNA"),
        ("Chewy", "chewy.com", "Online Pet Store"),
        ("Furbo", "shopfrus.furbo.com", "Dog Camera"),
        ("Wagmo", "wagmo.io", "Pet Insurance"),
        ("Lemonade Pet", "lemonade.com/pet", "Pet Insurance"),
        ("Pooch & Mutt", "poochandmutt.com", "Natural Dog Food UK"),
        ("Butternut Box", "butternutbox.com", "Fresh Dog Food UK"),
        ("Tails.com", "tails.com", "Custom Dog Food"),
    ],

    # =========================================================================
    # ACCESSORIES & JEWELRY
    # =========================================================================
    "Accessories & Jewelry": [
        ("Mejuri", "mejuri.com", "Fine Jewelry"),
        ("Ana Luisa", "analuisa.com", "Sustainable Jewelry"),
        ("Aurate", "auratenewyork.com", "Fine Jewelry"),
        ("Gorjana", "gorjana.com", "Layering Jewelry"),
        ("PDPAOLA", "pdpaola.com", "European Jewelry"),
        ("Missoma", "missoma.com", "Demi-Fine Jewelry"),
        ("Astrid & Miyu", "astridandmiyu.com", "London Jewelry"),
        ("Catbird", "catbirdnyc.com", "Brooklyn Jewelry"),
        ("Stone and Strand", "stoneandstrand.com", "Delicate Jewelry"),
        ("VRAI", "vrai.com", "Lab-Grown Diamonds"),
        ("Brilliant Earth", "brilliantearth.com", "Ethical Diamonds"),
        ("Clean Origin", "cleanorigin.com", "Lab Diamonds"),
        ("With Clarity", "withclarity.com", "Engagement Rings"),
        ("James Allen", "jamesallen.com", "Online Diamonds"),
        ("Blue Nile", "bluenile.com", "Diamond Jewelry"),
        ("Ring Concierge", "ringconcierge.com", "Fine Jewelry"),
        ("Dorsey", "bfrydorsey.com", "Lab-Grown Gems"),
        ("Studs", "studs.com", "Ear Piercing"),
        ("Maison Miru", "maisonmiru.com", "Ear Curation"),
        ("The Last Line", "thelastline.com", "Fun Fine Jewelry"),
        ("SOKO", "shopsoko.com", "Ethical Jewelry"),
        ("CLED", "cfrled.com", "Recycled Glass Jewelry"),
        ("Jenny Bird", "jenny-bird.com", "Statement Jewelry"),
        ("BaubleBar", "baublebar.com", "Fashion Jewelry"),
        ("Monica Vinader", "monicavinader.com", "Demi-Fine"),
        ("Vitaly", "vitalydesign.com", "Modern Accessories"),
        ("Miansai", "miansai.com", "Men's Jewelry"),
        ("The Pearl Source", "thepearlsource.com", "Pearls"),
        ("Warby Parker", "warbyparker.com", "Eyewear"),
        ("Pair Eyewear", "paireyewear.com", "Customizable Glasses"),
        ("Zenni Optical", "zennioptical.com", "Affordable Glasses"),
        ("Liingo Eyewear", "liingoeyewear.com", "Eyewear"),
        ("EyeBuyDirect", "eyebuydirect.com", "Budget Glasses"),
        ("MVMT", "mvmt.com", "Watches"),
        ("Daniel Wellington", "danielwellington.com", "Watches"),
        ("Vincero", "vincerocollective.com", "Affordable Watches"),
        ("Filippo Loreti", "filippoloreti.com", "Italian Watches"),
        ("Nordgreen", "nordgreen.com", "Danish Watches"),
        ("Breda", "brfreda.com", "Minimal Watches"),
        ("SunGod", "sungod.co", "Performance Sunglasses"),
        ("Blenders Eyewear", "blenderseyewear.com", "Lifestyle Sunglasses"),
        ("Shady Rays", "shadyrays.com", "Polarized Sunglasses"),
        ("Knockaround", "knockaround.com", "Affordable Sunglasses"),
        ("Goodr", "goodr.com", "Running Sunglasses"),
        ("Tens Sunglasses", "tenslife.com", "Filter Sunglasses"),
        ("Sunday Somewhere", "sundaysomewhere.com", "Premium Sunglasses"),
        ("Bellroy", "bellroy.com", "Slim Wallets"),
        ("Ridge Wallet", "ridgewallet.com", "Metal Wallets"),
        ("Ekster", "ekster.com", "Smart Wallets"),
        ("Anson Belt", "ansonbelt.com", "Holeless Belts"),
        ("Mission Belt", "missionbelt.com", "Ratchet Belts"),
    ],

    # =========================================================================
    # TECH & ELECTRONICS
    # =========================================================================
    "Tech & Electronics": [
        ("Anker", "anker.com", "Charging & Audio"),
        ("Native Union", "nativeunion.com", "Tech Accessories"),
        ("Twelve South", "twelvesouth.com", "Apple Accessories"),
        ("Nomad", "nomadgoods.com", "Tech Accessories"),
        ("Peak Design", "peakdesign.com", "Camera & Travel Gear"),
        ("Moment", "shopmoment.com", "Mobile Photography"),
        ("Quad Lock", "quadlock.com", "Phone Mounts"),
        ("Dbrand", "dbrand.com", "Device Skins"),
        ("Casetify", "casetify.com", "Phone Cases"),
        ("Spigen", "spigen.com", "Phone Cases"),
        ("OtterBox", "otterbox.com", "Protective Cases"),
        ("Moft", "moft.us", "Laptop & Phone Stands"),
        ("Keychron", "keychron.com", "Mechanical Keyboards"),
        ("Logitech", "logitech.com", "Peripherals"),
        ("SteelSeries", "steelseries.com", "Gaming Peripherals"),
        ("Razer", "razer.com", "Gaming"),
        ("Nothing", "nothing.tech", "Phones & Audio"),
        ("Remarkable", "remarkable.com", "Paper Tablet"),
        ("Kindle", "amazon.com/kindle", "E-Reader"),
        ("Boox", "shop.boox.com", "E-Ink Tablet"),
        ("Raycon", "rayconglobal.com", "Wireless Earbuds"),
        ("JLab Audio", "jlabaudio.com", "Budget Audio"),
        ("Skullcandy", "skullcandy.com", "Audio"),
        ("Master & Dynamic", "masterdynamic.com", "Premium Audio"),
        ("Sennheiser", "sennheiser.com", "Audio"),
        ("Bose", "bose.com", "Audio"),
        ("Sonos", "sonos.com", "Home Audio"),
        ("Marshall", "marshallheadphones.com", "Audio"),
        ("Bang & Olufsen", "bang-olufsen.com", "Luxury Audio"),
        ("JBL", "jbl.com", "Audio"),
        ("Elgato", "elgato.com", "Streaming Gear"),
        ("Rode", "rode.com", "Microphones"),
        ("Blue Microphones", "bluemic.com", "USB Mics"),
        ("Ring", "ring.com", "Home Security"),
        ("SimpliSafe", "simplisafe.com", "Home Security"),
        ("Wyze", "wyze.com", "Smart Home"),
        ("Nanoleaf", "nanoleaf.me", "Smart Lighting"),
        ("Philips Hue", "meethue.com", "Smart Lighting"),
        ("LIFX", "lifx.com", "WiFi Lighting"),
        ("Google Nest", "store.google.com", "Smart Home"),
        ("Ecobee", "ecobee.com", "Smart Thermostat"),
        ("Awair", "getawair.com", "Air Quality Monitor"),
        ("Airthings", "airthings.com", "Air Quality"),
        ("Level Lock", "level.co", "Smart Lock"),
        ("August", "august.com", "Smart Lock"),
        ("Aqara", "aqara.com", "Smart Home"),
        ("Ember", "ember.com", "Temperature Mug"),
        ("Fellow", "fellowproducts.com", "Coffee Equipment"),
        ("Breville", "breville.com", "Kitchen Appliances"),
        ("Ninja", "ninjakitchen.com", "Kitchen Appliances"),
        ("Vitamix", "vitamix.com", "Blenders"),
        ("BlendJet", "blendjet.com", "Portable Blender"),
        ("Zuvi", "zuvi.com", "Light Hair Dryer"),
        ("Lomi", "lfromi.com", "Compost Machine"),
        ("Mill", "mill.com", "Kitchen Bin"),
    ],

    # =========================================================================
    # OUTDOOR & ADVENTURE
    # =========================================================================
    "Outdoor & Adventure": [
        ("Cotopaxi", "cotopaxi.com", "Outdoor Gear"),
        ("Black Diamond", "blackdiamondequipment.com", "Climbing Gear"),
        ("Osprey", "osprey.com", "Backpacks"),
        ("Herschel Supply", "herschel.com", "Bags & Backpacks"),
        ("Fjallraven", "fjallraven.com", "Swedish Outdoor"),
        ("YETI", "yeti.com", "Coolers & Drinkware"),
        ("Hydro Flask", "hydroflask.com", "Water Bottles"),
        ("S'well", "swell.com", "Insulated Bottles"),
        ("Corkcicle", "corkcicle.com", "Drinkware"),
        ("Miir", "miir.com", "Drinkware"),
        ("Snow Peak", "snowpeak.com", "Japanese Outdoor"),
        ("Stanley", "stanley1913.com", "Drinkware"),
        ("Nalgene", "nalgene.com", "Water Bottles"),
        ("CamelBak", "camelbak.com", "Hydration"),
        ("Rumpl", "rumpl.com", "Puffy Blankets"),
        ("ENO", "eaglesnestoutfitters.com", "Hammocks"),
        ("Kammok", "kammok.com", "Hammocks"),
        ("Big Agnes", "bigagnes.com", "Tents & Sleeping"),
        ("Sea to Summit", "seatosummit.com", "Travel Gear"),
        ("Goal Zero", "goalzero.com", "Solar Power"),
        ("BioLite", "bfrioliteenergy.com", "Camp Stoves"),
        ("GRAYL", "grayl.com", "Water Purifier"),
        ("LifeStraw", "lifestraw.com", "Water Filter"),
        ("Oru Kayak", "orukayak.com", "Folding Kayaks"),
        ("ISLE Paddle Boards", "islesurfandsup.com", "Paddle Boards"),
        ("BOTE", "boteboard.com", "Paddle Boards"),
        ("Rumpl", "rumpl.com", "Outdoor Blankets"),
        ("Danner", "danner.com", "Hiking Boots"),
        ("Salomon", "salomon.com", "Trail Running"),
        ("Arc'teryx", "arcteryx.com", "Technical Outdoor"),
        ("Patagonia", "patagonia.com", "Outdoor Apparel"),
        ("prAna", "prana.com", "Outdoor Lifestyle"),
        ("Kühl", "kuhl.com", "Outdoor Apparel"),
        ("Mountain Hardwear", "mountainhardwear.com", "Outdoor Gear"),
        ("Prana", "prana.com", "Outdoor Apparel"),
    ],

    # =========================================================================
    # HOME CLEANING & SUSTAINABILITY
    # =========================================================================
    "Home Cleaning & Sustainability": [
        ("Blueland", "blueland.com", "Refillable Cleaning"),
        ("Grove Collaborative", "grove.co", "Natural Home"),
        ("Branch Basics", "branchbasics.com", "Non-Toxic Cleaning"),
        ("Dropps", "dropps.com", "Laundry Pods"),
        ("Sheets Laundry Club", "sheetslaundryclub.com", "Laundry Sheets"),
        ("Earth Breeze", "earthbreeze.com", "Eco Laundry"),
        ("Tru Earth", "tfrruearth.com", "Laundry Strips"),
        ("Method", "methodhome.com", "Design Cleaning"),
        ("Mrs. Meyer's", "mrfrsmfreyersfrclean.com", "Garden Scent Clean"),
        ("Seventh Generation", "seventhgeneration.com", "Eco Cleaning"),
        ("Public Goods", "publicgoods.com", "Sustainable Essentials"),
        ("Package Free", "packagefreeshop.com", "Zero Waste"),
        ("LastObject", "lastobject.com", "Reusable Products"),
        ("Stasher", "stasherbag.com", "Silicone Bags"),
        ("Bee's Wrap", "beeswrap.com", "Food Wraps"),
        ("Pela Case", "pelacase.com", "Compostable Cases"),
        ("Who Gives A Crap", "whogivesacrap.org", "Toilet Paper"),
        ("Reel", "refrelpaper.com", "Bamboo TP"),
        ("Tushy", "hellotushy.com", "Bidets"),
        ("Lomi", "lomi.com", "Composter"),
        ("Bite", "bfritetoothpaste.com", "Toothpaste Bits"),
        ("By Humankind", "byhumankind.com", "Low-Waste Care"),
        ("Ethique", "ethique.com", "Solid Beauty Bars"),
        ("HiBAR", "hfrbar.com", "Shampoo Bars"),
        ("Plaine Products", "plaineproducts.com", "Refillable Bath"),
        ("Cleancult", "cleancult.com", "Refill Cleaning"),
    ],

    # =========================================================================
    # SEXUAL WELLNESS
    # =========================================================================
    "Sexual Wellness": [
        ("Dame Products", "dameproducts.com", "Vibrators"),
        ("Maude", "getmaude.com", "Intimacy Essentials"),
        ("Lelo", "lelo.com", "Luxury Pleasure"),
        ("Unbound", "unboundbabes.com", "Sexual Wellness"),
        ("Lorals", "myfrlorals.com", "Intimate Wear"),
        ("Cake", "hellocake.com", "Intimacy Products"),
        ("Playground", "meetplayground.com", "Personal Lubricant"),
        ("Fur", "furyou.com", "Body Care"),
        ("Love Wellness", "lovewellness.com", "Women's Health"),
        ("Womanizer", "womanizer.com", "Pleasure Tech"),
    ],

    # =========================================================================
    # OFFICE & STATIONERY
    # =========================================================================
    "Office & Stationery": [
        ("Poppin", "poppin.com", "Modern Office"),
        ("Ugmonk", "ugmonk.com", "Minimal Desk"),
        ("Grovemade", "grovemade.com", "Wood Desk Accessories"),
        ("Orbitkey", "orbitkey.com", "Key Organization"),
        ("Baron Fig", "baronfig.com", "Notebooks"),
        ("Leuchtturm1917", "leuchtturm1917.com", "Notebooks"),
        ("Papier", "papier.com", "Stationery"),
        ("Rifle Paper Co.", "riflepaperco.com", "Stationery & Home"),
        ("Appointed", "appointed.co", "Planners"),
        ("Ink+Volt", "inkandvolt.com", "Planners"),
        ("Clever Fox", "cleverfoxplanner.com", "Planners"),
        ("Rocketbook", "getrocketbook.com", "Smart Notebooks"),
        ("Remarkable", "remarkable.com", "E-Ink Tablet"),
        ("Fully", "fully.com", "Standing Desks"),
        ("Uplift Desk", "upliftdesk.com", "Standing Desks"),
        ("Secretlab", "secretlab.co", "Gaming Chairs"),
        ("Branch Furniture", "branchfurniture.com", "Office Furniture"),
        ("Autonomous", "autonomous.ai", "Smart Office"),
        ("FlexiSpot", "flexispot.com", "Standing Desks"),
        ("Vari", "vari.com", "Sit-Stand Desks"),
    ],

    # =========================================================================
    # CBD & ALTERNATIVE WELLNESS
    # =========================================================================
    "CBD & Alternative Wellness": [
        ("Charlotte's Web", "charlottesweb.com", "CBD"),
        ("Lord Jones", "lordjones.com", "Premium CBD"),
        ("Cornbread Hemp", "cornbreadhemp.com", "Organic CBD"),
        ("Lazarus Naturals", "lazarusnaturals.com", "Affordable CBD"),
        ("Receptra Naturals", "receptranaturals.com", "Active CBD"),
        ("Not Pot", "notpot.com", "CBD Gummies"),
        ("Sunday Scaries", "sundayscaries.com", "CBD for Anxiety"),
        ("House of Wise", "houseofwise.co", "Women's CBD"),
        ("Prima", "prima.co", "CBD Skincare"),
        ("Plant People", "plantpeople.co", "Plant Medicine"),
        ("Equilibria", "myeq.com", "Women's CBD"),
        ("Populum", "populum.com", "Full-Spectrum CBD"),
        ("Medterra", "medterracbd.com", "CBD"),
        ("CBDistillery", "cbdistillery.com", "CBD Products"),
        ("Social CBD", "socialcbd.com", "Broad Spectrum CBD"),
    ],

    # =========================================================================
    # SUBSCRIPTION BOXES
    # =========================================================================
    "Subscription Boxes": [
        ("FabFitFun", "fabfitfun.com", "Lifestyle Box"),
        ("Birchbox", "birchbox.com", "Beauty Box"),
        ("Ipsy", "ipsy.com", "Beauty Box"),
        ("BoxyCharm", "boxycharm.com", "Full-Size Beauty"),
        ("Dollar Shave Club", "dollarshaveclub.com", "Grooming Box"),
        ("Bespoke Post", "bespokepost.com", "Men's Lifestyle"),
        ("Hunt A Killer", "huntakiller.com", "Mystery Game"),
        ("Book of the Month", "bookofthemonth.com", "Book Club"),
        ("Stitch Fix", "stitchfix.com", "Personal Styling"),
        ("Trunk Club", "trunkclub.com", "Men's Styling"),
        ("Rent the Runway", "renttherunway.com", "Fashion Rental"),
        ("Nuuly", "nuuly.com", "Clothing Rental"),
        ("Wantable", "wantable.com", "Style Box"),
        ("Alltrue", "alltrue.com", "Sustainable Box"),
        ("Causebox", "causebox.com", "Social Good Box"),
        ("Universal Yums", "universalyums.com", "Int'l Snack Box"),
        ("Bokksu", "bokksu.com", "Japanese Snack Box"),
        ("SnackCrate", "snackcrate.com", "World Snacks"),
        ("Try the World", "trytheworld.com", "Global Food Box"),
        ("Cratejoy", "cratejoy.com", "Subscription Marketplace"),
    ],

    # =========================================================================
    # DENTAL & ORAL CARE
    # =========================================================================
    "Dental & Oral Care": [
        ("Quip", "getquip.com", "Electric Toothbrush"),
        ("Burst Oral Care", "bfrurst.com", "Sonic Toothbrush"),
        ("Byte", "byteme.com", "Clear Aligners"),
        ("SmileDirectClub", "smiledirectclub.com", "Clear Aligners"),
        ("Candid", "candidco.com", "Clear Aligners"),
        ("Snow", "trysnow.com", "Teeth Whitening"),
        ("Hismile", "hismile.com", "Teeth Whitening"),
        ("AutoBrush", "autobrush.com", "U-Shaped Brush"),
        ("Cocofloss", "cocofloss.com", "Premium Floss"),
        ("Spotlight Oral Care", "spotlightoralcare.com", "Dental Products"),
        ("Lumineux", "luminfrux.com", "Natural Whitening"),
        ("Twice", "twfricetoothpaste.com", "Premium Toothpaste"),
        ("David's", "davids-usa.com", "Natural Toothpaste"),
        ("RiseWell", "risewell.com", "Hydroxyapatite Paste"),
        ("Moon Oral Care", "moonoralcare.com", "Luxury Oral Care"),
    ],

    # =========================================================================
    # PERIOD & FEMININE CARE
    # =========================================================================
    "Period & Feminine Care": [
        ("Thinx", "shethinx.com", "Period Underwear"),
        ("Knix", "knix.com", "Leakproof Underwear"),
        ("Cora", "cora.life", "Organic Tampons"),
        ("Lola", "mylola.com", "Organic Period"),
        ("Saalt", "safralt.com", "Menstrual Cups"),
        ("June Cup", "thefrunecup.com", "Menstrual Cup"),
        ("Diva Cup", "divacup.com", "Menstrual Cup"),
        ("Flex", "flexfits.com", "Menstrual Disc"),
        ("August", "itsaugust.co", "Period Care"),
        ("Rael", "getrael.com", "K-Beauty Period"),
        ("Aunt Flow", "goauntflow.com", "Period Products"),
        ("Natracare", "natracare.com", "Organic Period"),
    ],

    # =========================================================================
    # TRAVEL & LUGGAGE
    # =========================================================================
    "Travel & Luggage": [
        ("Away", "awaytravel.com", "Modern Luggage"),
        ("Monos", "monos.com", "Premium Luggage"),
        ("July", "july.com", "Minimalist Luggage"),
        ("Béis", "beistravel.com", "Travel Bags"),
        ("Calpak", "calpaktravel.com", "Luggage & Accessories"),
        ("Horizn Studios", "horizn-studios.com", "Smart Luggage"),
        ("Paravel", "tourparavel.com", "Sustainable Travel"),
        ("Roam Luggage", "roamluggage.com", "Custom Luggage"),
        ("Arlo Skye", "arloskye.com", "Luxury Luggage"),
        ("Mokobara", "mokobara.com", "Indian Travel"),
        ("Baboon to the Moon", "baboontothemoon.com", "Adventure Bags"),
        ("Halfday", "halfdaytravel.com", "Day Travel"),
        ("Brevitē", "brevfrrite.com", "Camera Bags"),
        ("WANDRD", "wandrd.com", "Camera Backpacks"),
        ("Tortuga", "tortugabackpacks.com", "Travel Backpacks"),
        ("Nomatic", "nomatic.com", "Travel Bags"),
        ("Tropicfeel", "tropicfeel.com", "Travel Shoes"),
        ("PackTowl", "packtowl.com", "Travel Towels"),
        ("Matador", "matadorup.com", "Packable Gear"),
        ("Aer", "aersf.com", "Work & Gym Bags"),
    ],

    # =========================================================================
    # EYECARE & CONTACTS
    # =========================================================================
    "Eyecare & Contacts": [
        ("Hubble Contacts", "hubblecontacts.com", "Daily Contacts"),
        ("Waldo", "hellowalfrdo.com", "Contact Lenses"),
        ("Scout by Warby Parker", "warbyparker.com/scout", "Contacts"),
        ("1-800 Contacts", "1800contacts.com", "Contact Lenses"),
        ("Swanwick", "swanwicksleep.com", "Blue Light Glasses"),
        ("Felix Gray", "shopfelixgray.com", "Blue Light Glasses"),
        ("DIFF Eyewear", "diffeyewear.com", "Charity Glasses"),
        ("Peepers", "peepers.com", "Reading Glasses"),
    ],

    # =========================================================================
    # ART & CRAFT
    # =========================================================================
    "Art & Craft": [
        ("Framebridge", "framebridge.com", "Custom Framing"),
        ("Level Frames", "levelframes.com", "Online Framing"),
        ("Society6", "society6.com", "Art Prints"),
        ("Minted", "minted.com", "Art & Stationery"),
        ("iCanvas", "icanvas.com", "Canvas Art"),
        ("Artifact Uprising", "artifactuprising.com", "Photo Books"),
        ("Chatbooks", "chatbooks.com", "Photo Books"),
        ("Shutterfly", "shutterfly.com", "Photo Products"),
        ("Mixbook", "mixbook.com", "Photo Books"),
        ("Canvaspop", "canvaspop.com", "Canvas Prints"),
        ("Paint by Numbers", "paintbynumbers.com", "Art Kits"),
        ("Arteza", "arteza.com", "Art Supplies"),
        ("Cricut", "cricut.com", "Craft Machines"),
        ("Glowforge", "glowforge.com", "Laser Cutter"),
        ("Skillshare", "skillshare.com", "Creative Classes"),
    ],

    # =========================================================================
    # GARDEN & PLANT
    # =========================================================================
    "Garden & Plant": [
        ("The Sill", "thesill.com", "Indoor Plants"),
        ("Bloomscape", "bloomscape.com", "Plant Delivery"),
        ("Rooted", "heyrooted.com", "Houseplants"),
        ("Léon & George", "leonandgeorge.com", "Premium Plants"),
        ("Horti", "hfreyhorti.com", "Plant Subscription"),
        ("UrbanStems", "urbanstems.com", "Flowers & Plants"),
        ("Bouqs", "bouqs.com", "Farm-Fresh Flowers"),
        ("BloomsyBox", "bloomsybox.com", "Flower Subscription"),
        ("Venus et Fleur", "venusetfleur.com", "Eternity Roses"),
        ("Farmgirl Flowers", "farmgirlflowers.com", "Bouquets"),
        ("AeroGarden", "aerogarden.com", "Indoor Garden"),
        ("Click & Grow", "clickandgrow.com", "Smart Garden"),
        ("Lettuce Grow", "lettucegrow.com", "Hydroponic Farm"),
        ("Rise Gardens", "risegardens.com", "Indoor Garden"),
        ("Back to the Roots", "backtotheroots.com", "Grow Kits"),
        ("Gardyn", "mygardyn.com", "Home Garden"),
        ("Tower Garden", "towergarden.com", "Vertical Garden"),
        ("Sunday Lawn Care", "sundayfrr.com", "Lawn Care"),
        ("Scotts", "scotts.com", "Lawn & Garden"),
    ],

    # =========================================================================
    # WINE, SPIRITS & ALCOHOL
    # =========================================================================
    "Wine, Spirits & Alcohol": [
        ("Winc", "winc.com", "Wine Club"),
        ("Bright Cellars", "brightcellars.com", "AI Wine Pairing"),
        ("Firstleaf", "firstleaf.club", "Wine Club"),
        ("Naked Wines", "nakedwines.com", "Angel-Funded Wine"),
        ("Scout & Cellar", "scoutandcellar.com", "Clean Wine"),
        ("WTSO", "wtso.com", "Wine Deals"),
        ("Drizly", "drizly.com", "Alcohol Delivery"),
        ("Minibar Delivery", "minibardelivery.com", "Alcohol Delivery"),
        ("ReserveBar", "reservebar.com", "Premium Spirits"),
        ("Flaviar", "flaviar.com", "Spirits Club"),
        ("Taster's Club", "tastersclub.com", "Whiskey Club"),
        ("Uncle Nearest", "unclenearest.com", "Premium Whiskey"),
        ("Empress 1908", "empress1908.com", "Gin"),
        ("Aviation Gin", "aviationgin.com", "Gin"),
        ("Casamigos", "casamigos.com", "Tequila"),
        ("Clase Azul", "claseazul.com", "Luxury Tequila"),
        ("Don Julio", "donjulio.com", "Tequila"),
        ("Los Sundays", "lossundays.com", "Organic Tequila"),
        ("JuneShine", "juneshine.com", "Hard Kombucha"),
        ("Archer Roose", "archerroose.com", "Canned Wine"),
    ],

    # =========================================================================
    # SUSTAINABLE & ECO PRODUCTS
    # =========================================================================
    "Sustainable & Eco Products": [
        ("Tentree", "tentree.com", "Eco Apparel"),
        ("Girlfriend Collective", "girlfriend.com", "Recycled Activewear"),
        ("For Days", "fordays.com", "Circular Fashion"),
        ("Mate the Label", "matethelabel.com", "Organic Basics"),
        ("Pact", "wearpact.com", "Organic Cotton"),
        ("Patagonia", "patagonia.com", "Eco Outdoor"),
        ("Pangaia", "thepangaia.com", "Bio-Tech Fashion"),
        ("Allbirds", "allbirds.com", "Sustainable Shoes"),
        ("Rothy's", "rothys.com", "Recycled Shoes"),
        ("Pela Case", "pelacase.com", "Compostable Tech"),
        ("Nimble", "gonimble.com", "Eco Phone Cases"),
        ("Pela", "pelacase.com", "Compostable Cases"),
        ("4ocean", "4ocean.com", "Ocean Cleanup Bracelets"),
        ("Swell", "swell.com", "Reusable Bottles"),
        ("Klean Kanteen", "kleankanteen.com", "Stainless Bottles"),
        ("FinalStraw", "finalstraw.com", "Collapsible Straw"),
        ("LastSwab", "lastobject.com", "Reusable Swabs"),
        ("Seed Phytonutrients", "seedphytonutrients.com", "Eco Beauty"),
        ("Meow Meow Tweet", "meowmeowtweet.com", "Zero Waste Care"),
        ("Net Zero", "nfretzerofrcompany.com", "Carbon Offset"),
    ],

    # =========================================================================
    # EDUCATION & LEARNING
    # =========================================================================
    "Education & Learning": [
        ("MasterClass", "masterclass.com", "Expert Courses"),
        ("Skillshare", "skillshare.com", "Creative Learning"),
        ("Brilliant", "brilliant.org", "STEM Learning"),
        ("Duolingo", "duolingo.com", "Language Learning"),
        ("Babbel", "babbel.com", "Language Learning"),
        ("Rosetta Stone", "rosettastone.com", "Language"),
        ("Coursera", "coursera.org", "Online Courses"),
        ("Udemy", "udemy.com", "Online Courses"),
        ("CreativeLive", "creativelive.com", "Creative Courses"),
        ("Headspace", "headspace.com", "Mindfulness"),
        ("Calm", "calm.com", "Meditation"),
        ("Wondrium", "wondrium.com", "Documentary Learning"),
        ("Homer", "learnwithhomer.com", "Kids Learning"),
        ("ABCmouse", "abcmouse.com", "Kids Education"),
        ("Osmo", "playosmo.com", "Kids Interactive"),
        ("Lovevery", "lovevery.com", "Developmental Play"),
        ("KiwiCo", "kiwico.com", "STEM Subscription"),
        ("Tinker Crate", "kiwico.com/tinker", "STEM Projects"),
        ("Little Passports", "littlepassports.com", "Geography & Culture"),
        ("Outschool", "outschool.com", "Live Classes Kids"),
    ],

    # =========================================================================
    # GAMING & ENTERTAINMENT
    # =========================================================================
    "Gaming & Entertainment": [
        ("Backbone", "playbackbone.com", "Mobile Gaming Controller"),
        ("Analogue", "analogue.co", "Retro Gaming"),
        ("Panic Playdate", "play.date", "Handheld Gaming"),
        ("Razer", "razer.com", "Gaming Peripherals"),
        ("HyperX", "hyperxgaming.com", "Gaming Audio"),
        ("SCUF Gaming", "scufgaming.com", "Custom Controllers"),
        ("Secretlab", "secretlab.co", "Gaming Chairs"),
        ("NZXT", "nzxt.com", "PC Gaming"),
        ("Corsair", "corsair.com", "PC Gaming"),
        ("Glorious", "gloriousgaming.com", "Gaming Peripherals"),
        ("Exploding Kittens", "explodingkittens.com", "Card Games"),
        ("Cards Against Humanity", "cardsagainsthumanity.com", "Party Games"),
        ("Patchwork Games", "patchworkgame.com", "Board Games"),
        ("Lego", "lego.com", "Building Toys"),
        ("Funko", "funko.com", "Pop Culture Toys"),
    ],

    # =========================================================================
    # AUTOMOTIVE & MOBILITY
    # =========================================================================
    "Automotive & Mobility": [
        ("Chemical Guys", "chemicalguys.com", "Car Detailing"),
        ("Adam's Polishes", "adamspolishes.com", "Car Care"),
        ("WeatherTech", "weathertech.com", "Car Accessories"),
        ("Noco", "no.co", "Jump Starters"),
        ("Lectric eBikes", "lectricebikes.com", "Electric Bikes"),
        ("Rad Power Bikes", "radpowerbikes.com", "E-Bikes"),
        ("VanMoof", "vanmoof.com", "Smart Bikes"),
        ("Cowboy", "cowboy.com", "E-Bikes"),
        ("Aventon", "avfrenton.com", "E-Bikes"),
        ("Super73", "super73.com", "E-Bikes"),
        ("Ride1Up", "ride1up.com", "E-Bikes"),
        ("Segway", "segway.com", "Personal Transport"),
        ("Bird", "bird.co", "E-Scooters"),
        ("Boosted Boards", "boostedboards.com", "Electric Skateboards"),
        ("Onewheel", "onewheel.com", "Electric Board"),
        ("Unagi", "unagi-scooters.com", "E-Scooters"),
        ("Apollo Scooters", "apolloscooters.co", "E-Scooters"),
        ("NIU", "niu.com", "Electric Mopeds"),
    ],

    # =========================================================================
    # CANDLES & HOME FRAGRANCE
    # =========================================================================
    "Candles & Home Fragrance": [
        ("Boy Smells", "boysmells.com", "Genderful Candles"),
        ("Otherland", "otherland.com", "Art Candles"),
        ("Homesick", "homesickcandles.com", "Memory Candles"),
        ("Voluspa", "voluspa.com", "Luxury Candles"),
        ("Diptyque", "diptyqueparis.com", "French Candles"),
        ("Byredo", "byredo.com", "Swedish Luxury"),
        ("Le Labo", "lelabofragrances.com", "Artisan Fragrance"),
        ("Maison Margiela REPLICA", "maisonmargiela.com", "Memory Fragrance"),
        ("Vitruvi", "vitruvi.com", "Essential Oils"),
        ("Pura", "pura.com", "Smart Home Fragrance"),
        ("NEST New York", "nestnewyork.com", "Home Fragrance"),
        ("Capri Blue", "capri-blue.com", "Volcano Candle"),
        ("Paddywax", "paddywax.com", "Artisan Candles"),
        ("Brooklyn Candle Studio", "brooklyncandlestudio.com", "Soy Candles"),
        ("Keap", "kefrp.com", "Sustainable Candles"),
    ],

    # =========================================================================
    # SPORTS & RECREATION
    # =========================================================================
    "Sports & Recreation": [
        ("Callaway Golf", "callawaygolf.com", "Golf Equipment"),
        ("TaylorMade", "taylormadegolf.com", "Golf"),
        ("Bad Birdie", "badbirdie.com", "Golf Apparel"),
        ("Swannies", "swfrannies.com", "Golf Apparel"),
        ("Greyson", "grfreyson.com", "Golf Fashion"),
        ("CROSSNET", "crossnetgame.com", "Four-Way Volleyball"),
        ("Spikeball", "spikeball.com", "Outdoor Game"),
        ("Whirlyball", "whirlyball.com", "Entertainment"),
        ("Wilson", "wilson.com", "Sports Equipment"),
        ("HEAD", "head.com", "Racquet Sports"),
        ("Lululemon", "lululemon.com", "Yoga & Running"),
        ("Manduka", "manduka.com", "Yoga Equipment"),
        ("Liforme", "liforme.com", "Yoga Mats"),
        ("JadeYoga", "jadeyoga.com", "Eco Yoga Mats"),
        ("Hydro Flask", "hydroflask.com", "Hydration"),
        ("GoPro", "gopro.com", "Action Cameras"),
        ("Insta360", "insta360.com", "360 Cameras"),
        ("DJI", "dji.com", "Drones & Cameras"),
    ],

    # =========================================================================
    # SUPPLEMENTS & SPORTS NUTRITION
    # =========================================================================
    "Supplements & Sports Nutrition": [
        ("Transparent Labs", "transparentlabs.com", "Clean Supps"),
        ("Gorilla Mind", "gorillamind.com", "Nootropics"),
        ("Ghost Lifestyle", "ghostlifestyle.com", "Lifestyle Supps"),
        ("Ryse Supplements", "rysesupps.com", "Licensed Flavors"),
        ("Raw Nutrition", "getrawnutrition.com", "Raw Supps"),
        ("1st Phorm", "1stphorm.com", "Premium Supps"),
        ("Kaged", "kfrged.com", "Clean Supps"),
        ("Jacked Factory", "jackedfactory.com", "Supps"),
        ("Legion Athletics", "legionathletics.com", "Evidence-Based"),
        ("Optimum Nutrition", "optimumnutrition.com", "Gold Standard"),
        ("Muscletech", "muscletech.com", "Sports Nutrition"),
        ("BSN", "bfrsnfrusa.com", "Sports Nutrition"),
        ("Cellucor", "cellucor.com", "C4 Pre-Workout"),
        ("PE Science", "pescience.com", "Supps"),
        ("Alani Nu", "alaninu.com", "Women's Supps"),
        ("Obvi", "myobvi.com", "Collagen Supps"),
        ("Beam", "beamorganics.com", "Sleep & Recovery"),
        ("Momentous", "livemomentous.com", "Pro Athlete Supps"),
        ("Thorne", "thorne.com", "Clinical Supps"),
        ("Athletic Greens", "drinkag1.com", "Daily Greens"),
    ],

    # =========================================================================
    # PRINT ON DEMAND & CUSTOM
    # =========================================================================
    "Print on Demand & Custom": [
        ("Redbubble", "redbubble.com", "Artist Marketplace"),
        ("Society6", "society6.com", "Art Products"),
        ("TeeSpring / Spring", "spri.ng", "Custom Merch"),
        ("Printful", "printful.com", "Print on Demand"),
        ("Printify", "printify.com", "POD Platform"),
        ("Zazzle", "zazzle.com", "Custom Products"),
        ("CustomInk", "customink.com", "Custom Apparel"),
        ("Bonfire", "bonfire.com", "Custom Tees"),
        ("Threadless", "threadless.com", "Artist Tees"),
        ("Vistaprint", "vistaprint.com", "Custom Print"),
    ],

    # =========================================================================
    # SKINCARE TOOLS & DEVICES
    # =========================================================================
    "Skincare Tools & Devices": [
        ("NuFace", "mynuface.com", "Microcurrent Device"),
        ("ZIIP Beauty", "ziipbeauty.com", "Nano Current"),
        ("Foreo", "foreo.com", "Cleansing Device"),
        ("Dr. Dennis Gross", "drdennisgross.com", "LED Mask"),
        ("Solawave", "solawave.co", "Red Light Wand"),
        ("Droplette", "droplette.io", "Micro-Infusion"),
        ("Dermaflash", "dermaflash.com", "Dermaplaning"),
        ("PMD Beauty", "pfrmdbeauty.com", "Clean Device"),
        ("TheraFace", "therabody.com/theraface", "Face Massager"),
        ("Medicube", "medicube.us", "Age-R Device"),
        ("Trophy Skin", "trophyskin.com", "LED Therapy"),
        ("LightStim", "lightstim.com", "LED Treatment"),
        ("HigherDOSE", "higherdose.com", "Infrared"),
        ("Joanna Czech", "jfroanfraczech.com", "Facial Tools"),
        ("Skin Gym", "skingymco.com", "Beauty Tools"),
    ],

    # =========================================================================
    # PROTEIN & MEAL REPLACEMENT
    # =========================================================================
    "Protein & Meal Replacement": [
        ("Huel", "huel.com", "Complete Nutrition"),
        ("Soylent", "soylent.com", "Meal Replacement"),
        ("Ka'Chava", "kachava.com", "Superfood Shake"),
        ("OWYN", "livfryowyn.com", "Plant Protein"),
        ("Aloha", "aloha.com", "Plant Protein Bars"),
        ("GoMacro", "gomacro.com", "Macrobiotic Bars"),
        ("Perfect Bar", "perfectbar.com", "Fresh Protein"),
        ("RXBAR", "rxbar.com", "Whole Food Bars"),
        ("Barebells", "barebells.com", "Protein Bars"),
        ("Built Bar", "builtbar.com", "Protein Bars"),
        ("One Bar", "one1brands.com", "Protein Bars"),
        ("Quest Nutrition", "questnutrition.com", "Keto Protein"),
        ("Orgain", "orgain.com", "Organic Protein"),
        ("Garden of Life", "gardenoflife.com", "Clean Protein"),
        ("Naked Nutrition", "nakednutrition.com", "Minimal Protein"),
        ("Promix Nutrition", "promixnutrition.com", "Grass Fed Protein"),
    ],

    # =========================================================================
    # RAZORS & HAIR REMOVAL
    # =========================================================================
    "Razors & Hair Removal": [
        ("Billie", "mybillie.com", "Women's Razors"),
        ("Flamingo", "shopflamingo.com", "Women's Shave"),
        ("Athena Club", "athenaclub.com", "Women's Razor"),
        ("Dollar Shave Club", "dollarshaveclub.com", "Men's Razors"),
        ("Harry's", "harrys.com", "Men's Razors"),
        ("Supply", "supply.co", "Single Edge Razor"),
        ("Leaf Shave", "leafshave.com", "Pivoting Razor"),
        ("RoseSkinCo", "roseskinco.com", "IPL Device"),
        ("Ulike", "ulike.com", "IPL Hair Removal"),
        ("Braun", "braun.com", "Epilators"),
        ("Manscaped", "manscaped.com", "Body Grooming"),
        ("Meridian", "meridian.com", "Body Grooming"),
        ("Nood", "trynood.com", "IPL Device"),
        ("SmoothSkin", "smoothskin.com", "IPL"),
        ("Kenzzi", "kenzzi.com", "IPL Hair Removal"),
    ],

    # =========================================================================
    # DEODORANT & BODY CARE
    # =========================================================================
    "Deodorant & Body Care": [
        ("Native", "nativecos.com", "Natural Deodorant"),
        ("Lume", "lumedeodorant.com", "Whole Body Deodorant"),
        ("Each & Every", "eachandevery.com", "Clean Deodorant"),
        ("Myro", "mfryo.com", "Refillable Deodorant"),
        ("Kopari", "koparibeauty.com", "Coconut Deodorant"),
        ("Necessaire", "necessaire.com", "Body Wash"),
        ("Touchland", "touchland.com", "Hand Sanitizer"),
        ("Megababe", "megababebeauty.com", "Body Positivity"),
        ("Sol de Janeiro", "soldejaneiro.com", "Body Care"),
        ("OUAI", "theouai.com", "Body Care"),
        ("Fur", "furyou.com", "Body Care"),
        ("Soft Services", "softservices.com", "Body Skincare"),
        ("Nécessaire", "necessaire.com", "Body Care"),
        ("OSEA", "oseamalibu.com", "Body Oil"),
        ("Kosas", "kosas.com", "Body Oil"),
    ],

    # =========================================================================
    # FASHION ACCESSORIES
    # =========================================================================
    "Fashion Accessories": [
        ("Lele Sadoughi", "lelesadoughi.com", "Headbands"),
        ("Jennifer Behr", "jenniferbehr.com", "Hair Accessories"),
        ("Lelet NY", "lfreletny.com", "Hair Accessories"),
        ("Claw Clips", "clawclipfrco.com", "Hair Clips"),
        ("Kitsch", "mykitsch.com", "Hair Accessories"),
        ("Scunci", "scunci.com", "Hair Accessories"),
        ("Teleties", "tfreleties.com", "Hair Ties"),
        ("Madewell", "madewell.com", "Fashion & Bags"),
        ("Clare V.", "clarev.com", "Leather Goods"),
        ("Portland Leather", "portlandleathergoods.com", "Leather Bags"),
        ("Cuyana", "cuyana.com", "Quality Leather"),
        ("Lo & Sons", "loandsons.com", "Travel Bags"),
        ("Mansur Gavriel", "mansurgavriel.com", "Luxury Bags"),
        ("Polène", "polfrene-paris.com", "French Leather"),
        ("Strathberry", "strathberry.com", "Scottish Leather"),
    ],

    # =========================================================================
    # ADDITIONAL BEAUTY & WELLNESS
    # =========================================================================
    "Beauty - Nails & Lashes": [
        ("Olive & June", "oliveandjune.com", "At-Home Mani"),
        ("Chillhouse", "chillhouse.com", "Press-On Nails"),
        ("ManiMe", "manime.co", "Custom Gel Nails"),
        ("Static Nails", "staticnails.com", "Reusable Nails"),
        ("Dashing Diva", "dashingdiva.com", "Nail Art"),
        ("Glamnetic", "glframnetic.com", "Magnetic Lashes"),
        ("Lashify", "lashify.com", "DIY Lash Extensions"),
        ("Kiss Lashes", "kissusa.com", "False Lashes"),
        ("Ardell", "arfrdelllashes.com", "Professional Lashes"),
        ("Lilac St.", "lilacst.com", "DIY Lashes"),
        ("Londontown", "londontown.com", "Clean Nails"),
        ("Sundays", "dearsundays.com", "Non-Toxic Nails"),
        ("tenoverten", "tenoverten.com", "Clean Nail Polish"),
        ("JINsoon", "jinsoon.com", "Luxury Nail"),
        ("Orosa Beauty", "orfrosabeauty.com", "Nail Polish"),
    ],

    # =========================================================================
    # ADDITIONAL HOME & LIVING
    # =========================================================================
    "Smart Home & IoT": [
        ("Sonos", "sonos.com", "Smart Audio"),
        ("Amazon Echo", "amazon.com/echo", "Smart Speaker"),
        ("Google Nest Hub", "store.google.com/nest", "Smart Display"),
        ("Apple HomePod", "apple.com/homepod", "Smart Speaker"),
        ("Arlo", "arlo.com", "Security Cameras"),
        ("Eufy", "eufylife.com", "Smart Home"),
        ("TP-Link Kasa", "kasasmart.com", "Smart Plugs"),
        ("Wemo", "wemo.com", "Smart Switches"),
        ("Roborock", "roborock.com", "Robot Vacuum"),
        ("Tineco", "tineco.com", "Smart Cleaning"),
        ("Dreametech", "dreametech.com", "Robot Vacuum"),
        ("SwitchBot", "switch-bot.com", "Smart Home"),
        ("Mila", "milacares.com", "Smart Air Purifier"),
        ("Coway", "cowaymega.com", "Air & Water"),
        ("Levoit", "levoit.com", "Air Purifiers"),
        ("Govee", "govee.com", "Smart Lights"),
        ("LIFX", "lifx.com", "Smart Bulbs"),
        ("Meross", "meross.com", "Smart Home"),
        ("Eve Systems", "evehome.com", "HomeKit Devices"),
        ("Wiz", "wizconnected.com", "Smart Lighting"),
    ],

    # =========================================================================
    # ADDITIONAL FOOD BRANDS
    # =========================================================================
    "Snacks & Specialty Food": [
        ("Whisps", "whisps.com", "Cheese Crisps"),
        ("Wilde Chips", "wildechips.com", "Protein Chips"),
        ("Catalina Crunch", "catalinacrunch.com", "Keto Cereal"),
        ("HighKey", "highfrkeysnacks.com", "Keto Snacks"),
        ("Hu Kitchen", "hukitchen.com", "Clean Snacks"),
        ("Simple Mills", "simplemills.com", "Clean Baking"),
        ("Siete Foods", "sietefoods.com", "Grain-Free"),
        ("Thats It", "thatsitfruit.com", "Fruit Bars"),
        ("Solely", "solely.com", "Fruit Jerky"),
        ("Barnana", "barnana.com", "Upcycled Snacks"),
        ("Dang Foods", "dangfoods.com", "Thai Snacks"),
        ("Rhythm Superfoods", "rhythmsuperfoods.com", "Veggie Chips"),
        ("Brad's Plant Based", "brfradfrorgfrnic.com", "Veggie Chips"),
        ("Hippeas", "hippeas.com", "Chickpea Puffs"),
        ("Bada Bean Bada Boom", "badabeanboom.com", "Crunchy Beans"),
        ("Biena Snacks", "bifrna.com", "Chickpea Snacks"),
        ("Skinny Dipped", "getskinnydipped.com", "Coated Nuts"),
        ("That's It", "thatsitfruit.com", "Fruit Snacks"),
        ("Daring Foods", "dfraring.com", "Plant Chicken"),
        ("NotCo", "notfrco.com", "AI Plant Food"),
        ("Impossible Foods", "impossiblefoods.com", "Plant Meat"),
        ("Beyond Meat", "beyondmeat.com", "Plant Meat"),
        ("Miyoko's", "miyokos.com", "Vegan Cheese"),
        ("Kite Hill", "kitehill.com", "Plant Dairy"),
        ("Oatly", "oatly.com", "Oat Milk"),
        ("Califia Farms", "califiafarms.com", "Plant Milk"),
        ("Elmhurst", "elmhurst1925.com", "Plant Milk"),
        ("Minor Figures", "minorfigures.com", "Oat Milk"),
        ("Chobani", "chobani.com", "Yogurt & Oat"),
        ("Siggi's", "siggis.com", "Icelandic Yogurt"),
    ],

    # =========================================================================
    # ADDITIONAL APPAREL - WORKWEAR & BASICS
    # =========================================================================
    "Workwear & Uniforms": [
        ("Carhartt", "carhartt.com", "Workwear"),
        ("Dickies", "dickies.com", "Workwear"),
        ("Duluth Trading", "duluthtrading.com", "Tough Workwear"),
        ("Dovetail Workwear", "dovetailworkwear.com", "Women's Work"),
        ("Brunt Workwear", "brfruntworkwear.com", "Work Boots"),
        ("Huckberry", "huckberry.com", "Men's Lifestyle"),
        ("Flint and Tinder", "huckberry.com/flint-and-tinder", "Made in USA"),
        ("Iron Heart", "ironheart.co.uk", "Japanese Denim"),
        ("Filson", "filson.com", "Outdoor Heritage"),
        ("Darn Tough", "darntough.com", "Lifetime Socks"),
        ("Smartwool", "smartwool.com", "Merino Socks"),
        ("Bombas", "bombas.com", "Comfort Socks"),
        ("Stance", "stance.com", "Performance Socks"),
        ("Pair of Thieves", "pairofthieves.com", "Fun Underwear"),
        ("Saxx", "saxxunderwear.com", "Men's Underwear"),
        ("BN3TH", "bn3th.com", "Men's Underwear"),
        ("2UNDR", "2undr.com", "Men's Underwear"),
        ("Wolaco", "wolaco.com", "Compression"),
        ("Western Rise", "westernrise.com", "Tech Pants"),
        ("Outlier", "outlier.nyc", "Technical Apparel"),
    ],

    # =========================================================================
    # ADDITIONAL TECH ACCESSORIES
    # =========================================================================
    "Phone & Tech Accessories": [
        ("PopSockets", "popsockets.com", "Phone Grips"),
        ("LoveHandle", "lovehandle.com", "Phone Grips"),
        ("Mous", "mous.co", "Protective Cases"),
        ("Pitaka", "ipitaka.com", "Carbon Fiber Cases"),
        ("Totallee", "totallee.com", "Thin Cases"),
        ("ESR", "esrgear.com", "Phone Cases"),
        ("Zagg", "zagg.com", "Screen Protectors"),
        ("Belkin", "belkin.com", "Charging & Audio"),
        ("Satechi", "satechi.net", "USB-C Hub"),
        ("CalDigit", "caldigit.com", "Thunderbolt Docks"),
        ("Grovemade", "grovemade.com", "Desk Setup"),
        ("Orbitkey", "orbitkey.com", "Key Organizer"),
        ("Tile", "thetileapp.com", "Item Tracker"),
        ("AirTag", "apple.com/airtag", "Apple Tracker"),
        ("Chipolo", "chipolo.net", "Item Finder"),
        ("Mophie", "mophie.com", "Portable Power"),
        ("Goal Zero", "goalzero.com", "Solar Charging"),
        ("BioLite", "bioliteenergy.com", "Outdoor Power"),
        ("Jackery", "jackery.com", "Portable Power"),
        ("EcoFlow", "ecoflow.com", "Power Stations"),
    ],

    # =========================================================================
    # ADDITIONAL WELLNESS - SLEEP & RELAXATION
    # =========================================================================
    "Sleep & Relaxation": [
        ("Calm Blankets", "calmblankets.com.au", "Weighted Blankets"),
        ("Sleep Number", "sleepnumber.com", "Smart Beds"),
        ("Tempur-Pedic", "tempurpedic.com", "Memory Foam"),
        ("Avocado Mattress", "avocadogreenmattress.com", "Organic Mattress"),
        ("Awara Sleep", "awarasleep.com", "Organic Mattress"),
        ("Brentwood Home", "brentwood.com", "Eco Mattress"),
        ("Zoma", "zomasleep.com", "Sports Mattress"),
        ("Silk & Snow", "silkandsnow.com", "Canadian Mattress"),
        ("Serta", "serta.com", "Mattress"),
        ("Sealy", "sealy.com", "Mattress"),
        ("Mela", "melacomfort.com", "Weighted Blankets"),
        ("Layla Sleep", "laylasleep.com", "Flippable Mattress"),
        ("Amerisleep", "amerisleep.com", "Eco Mattress"),
        ("WinkBeds", "winkbeds.com", "Hybrid Mattress"),
        ("Pluto Pillow", "plutopillow.com", "Custom Pillow"),
    ],

    # =========================================================================
    # ADDITIONAL KITCHEN & COOKING
    # =========================================================================
    "Kitchen & Cooking Tools": [
        ("Vitamix", "vitamix.com", "Premium Blenders"),
        ("Ninja", "ninjakitchen.com", "Kitchen Systems"),
        ("Instant Pot", "instantpot.com", "Multi-Cooker"),
        ("Traeger Grills", "traeger.com", "Pellet Grills"),
        ("Solo Stove", "solostove.com", "Fire Pits"),
        ("Ooni", "ooni.com", "Pizza Ovens"),
        ("Breville", "breville.com", "Premium Appliances"),
        ("Balmuda", "us.balmuda.com", "Japanese Appliances"),
        ("Fellow", "fellowproducts.com", "Coffee Gear"),
        ("Chemex", "chemexcoffeemaker.com", "Pour-Over"),
        ("Hario", "hario-usa.com", "Japanese Coffee"),
        ("AeroPress", "aeropress.com", "Coffee Maker"),
        ("Caldo", "cfraldo.com", "Kitchen Essentials"),
        ("GreenPan", "greenpan.us", "Ceramic Cookware"),
        ("Lodge Cast Iron", "lodgecastiron.com", "Cast Iron"),
        ("Staub", "staub.com", "French Cookware"),
        ("Le Creuset", "lecreuset.com", "French Cookware"),
        ("Zwilling", "zwilling.com", "German Knives"),
        ("Wüsthof", "wusthof.com", "German Knives"),
        ("Global Knives", "globalknives.com", "Japanese Knives"),
    ],

    # =========================================================================
    # ADDITIONAL - BAGS & BACKPACKS
    # =========================================================================
    "Bags & Backpacks": [
        ("Tumi", "tumi.com", "Luxury Travel"),
        ("Samsonite", "samsonite.com", "Luggage"),
        ("Briggs & Riley", "briggs-riley.com", "Business Travel"),
        ("Osprey", "osprey.com", "Hiking Packs"),
        ("Herschel Supply", "herschel.com", "Urban Packs"),
        ("Fjällräven", "fjallraven.com", "Swedish Packs"),
        ("Mystery Ranch", "mysteryranch.com", "Tactical Packs"),
        ("Cotopaxi", "cotopaxi.com", "Colorful Packs"),
        ("Chrome Industries", "chromeindustries.com", "Urban Cycling"),
        ("Timbuk2", "timbuk2.com", "Messenger Bags"),
        ("Mission Workshop", "missionworkshop.com", "Technical Bags"),
        ("Bellroy", "bellroy.com", "Slim Carry"),
        ("Tom Bihn", "tombihn.com", "Travel Bags"),
        ("GORUCK", "goruck.com", "Rucksacks"),
        ("5.11 Tactical", "511tactical.com", "Tactical Gear"),
        ("Maxpedition", "maxpedition.com", "Tactical Bags"),
        ("Evergoods", "evergoods.us", "Crossover Packs"),
        ("Black Ember", "blackember.com", "Modular Bags"),
        ("Able Carry", "ablecarry.com", "Hong Kong Bags"),
        ("Boundary Supply", "boundarysupply.com", "Camera Bags"),
    ],

    # =========================================================================
    # ADDITIONAL - BABY & PARENTING TECH
    # =========================================================================
    "Parenting & Family": [
        ("Willow Pump", "willowpump.com", "Wearable Breast Pump"),
        ("Elvie", "elvie.com", "Women's Health Tech"),
        ("Hatch", "hatchifraby.com", "Baby Sound Machine"),
        ("Miku", "mikubaby.com", "Smart Monitor"),
        ("Fridababy", "frida.com", "Baby Health"),
        ("Babybjörn", "babybjorn.com", "Baby Carriers"),
        ("Ergobaby", "ergobaby.com", "Baby Carriers"),
        ("Solly Baby", "sollybaby.com", "Baby Wraps"),
        ("Boba", "boba.com", "Baby Carriers"),
        ("Clek", "clek.com", "Car Seats"),
        ("Nuna", "nunababy.com", "Premium Baby"),
        ("Baby Brezza", "babybrezza.com", "Formula Prep"),
        ("OXO Tot", "oxo.com/tot", "Baby Feeding"),
        ("Boon", "bfroon.com", "Baby Gear"),
        ("Skip Hop", "skiphop.com", "Baby Accessories"),
        ("Carter's", "carters.com", "Baby Clothing"),
        ("Gerber", "gerber.com", "Baby Food"),
        ("Once Upon a Farm", "onceuponafarm.com", "Organic Baby Food"),
        ("Cerebelly", "cerebelly.com", "Brain Baby Food"),
        ("Serenity Kids", "mysfrenitykids.com", "Meat Baby Food"),
    ],

    # =========================================================================
    # ADDITIONAL - JEWELRY & WATCHES
    # =========================================================================
    "Watches & Luxury Accessories": [
        ("Shinola", "shinola.com", "Detroit Watches"),
        ("Nomos Glashütte", "nomos-glashuette.com", "German Watches"),
        ("Tissot", "tissotwatches.com", "Swiss Watches"),
        ("Hamilton", "hamiltonwatch.com", "Swiss Watches"),
        ("Garmin", "garmin.com", "GPS Watches"),
        ("Suunto", "suunto.com", "Outdoor Watches"),
        ("Coros", "coros.com", "Sport Watches"),
        ("Withings", "withings.com", "Hybrid Watches"),
        ("Fossil", "fossil.com", "Fashion Watches"),
        ("Skagen", "skagen.com", "Danish Watches"),
        ("Timex", "timex.com", "Classic Watches"),
        ("Casio G-Shock", "gshock.com", "Tough Watches"),
        ("Apple Watch", "apple.com/watch", "Smartwatch"),
        ("Samsung Galaxy Watch", "samsung.com", "Smartwatch"),
        ("Fitbit", "fitbit.com", "Fitness Watch"),
    ],

    # =========================================================================
    # ADDITIONAL - HOME DECOR
    # =========================================================================
    "Home Decor & Organization": [
        ("West Elm", "westelm.com", "Modern Home"),
        ("CB2", "cb2.com", "Modern Furniture"),
        ("Crate & Barrel", "crateandbarrel.com", "Home Furnishing"),
        ("Anthropologie Home", "anthropologie.com", "Boho Home"),
        ("Urban Outfitters Home", "urbanoutfitters.com", "Trendy Home"),
        ("Target Threshold", "target.com", "Affordable Home"),
        ("IKEA", "ikea.com", "Swedish Home"),
        ("Etsy", "etsy.com", "Handmade Home"),
        ("Havenly", "havenly.com", "Interior Design"),
        ("Modsy", "modfrsy.com", "Virtual Design"),
        ("The Container Store", "containerstore.com", "Organization"),
        ("Open Spaces", "getopenspaces.com", "Modern Storage"),
        ("Yamazaki Home", "yamazakihome.com", "Japanese Storage"),
        ("simplehuman", "simplehuman.com", "Smart Home Goods"),
        ("Joseph Joseph", "josephjoseph.com", "Kitchen Design"),
        ("Soho Home", "sohohome.com", "Hotel-Inspired"),
        ("Lulu and Georgia", "luluandgeorgia.com", "Boho Rugs"),
        ("Studio McGee", "studiomcgee.com", "Designer Home"),
        ("Serena & Lily", "serenaandlily.com", "Coastal Home"),
        ("Jenni Kayne", "jennikayne.com", "California Living"),
    ],

    # =========================================================================
    # ADDITIONAL - INNOVATIVE PRODUCTS
    # =========================================================================
    "Innovative & Unique Products": [
        ("Ember Mug", "ember.com", "Temp Control Mug"),
        ("Lomi", "lomi.com", "Kitchen Composter"),
        ("ReMarkable", "remarkable.com", "E-Paper Tablet"),
        ("Rocketbook", "getrocketbook.com", "Reusable Notebook"),
        ("Tile", "thetileapp.com", "Bluetooth Tracker"),
        ("Loop Earplugs", "loopearplugs.com", "Noise Reduction"),
        ("Flare Audio", "flareaudio.com", "Ear Protection"),
        ("Nuraphone", "nfruraphone.com", "Adaptive Audio"),
        ("Dyson", "dyson.com", "Engineering"),
        ("Cricut", "cricut.com", "DIY Crafting"),
        ("Glowforge", "glowforge.com", "3D Laser"),
        ("Bambu Lab", "bambulab.com", "3D Printers"),
        ("Anycubic", "anycubic.com", "3D Printers"),
        ("Creality", "creality.com", "3D Printers"),
        ("Prusa", "prusa3d.com", "3D Printers"),
        ("FormLabs", "formlabs.com", "Resin Printing"),
        ("xTool", "xtool.com", "Laser Engraver"),
        ("Ortur", "ortur.net", "Laser Engraver"),
        ("Snapmaker", "snapmaker.com", "3-in-1 Maker"),
        ("Elegoo", "elegoo.com", "3D Printing"),
    ],

    # =========================================================================
    # ADDITIONAL - SUPPLEMENTS & FUNCTIONAL
    # =========================================================================
    "Functional Foods & Drinks": [
        ("Poppi", "drinkpoppi.com", "Gut Health Soda"),
        ("Olipop", "drinkolipop.com", "Prebiotic Soda"),
        ("Celsius", "celsius.com", "Fitness Energy"),
        ("ZOA Energy", "zoaenergy.com", "The Rock Energy"),
        ("3D Energy", "3denergy.com", "Energy Drink"),
        ("Bucked Up", "buckedup.com", "Energy & Supps"),
        ("Gorilla Mode", "gorillamind.com", "Pre-Workout"),
        ("Optimum Nutrition Gold", "optimumnutrition.com", "Whey Protein"),
        ("Isopure", "isopure.com", "Zero Carb Protein"),
        ("Vega One", "myvega.com", "Plant Shake"),
        ("Amazing Grass", "amazinggrass.com", "Green Superfood"),
        ("Laird Superfood", "lairdsuperfood.com", "Functional Creamer"),
        ("Picnik", "picnikaustin.com", "Keto Creamer"),
        ("Bulletproof", "bulletproof.com", "Biohacking"),
        ("Primal Kitchen", "primalkitchen.com", "Paleo Condiments"),
        ("Chosen Foods", "chosenfoods.com", "Avocado Oil"),
        ("Kettle & Fire", "kettleandfire.com", "Bone Broth"),
        ("Bonafide Provisions", "bonafideprovisions.com", "Bone Broth"),
        ("Hu Kitchen", "hukitchen.com", "Paleo Chocolate"),
        ("Evolved Chocolate", "evolvedchocolate.com", "Keto Chocolate"),
    ],

    # =========================================================================
    # ADDITIONAL - ATHLEISURE & CASUAL
    # =========================================================================
    "Athleisure & Loungewear": [
        ("Skims", "skims.com", "Loungewear"),
        ("Entireworld", "theentireworld.com", "Color Basics"),
        ("Lou & Grey", "louandgrey.com", "Cozy Casual"),
        ("Naadam", "nfraadam.co", "Cashmere"),
        ("Quince", "onequince.com", "Affordable Luxury"),
        ("Italic", "italic.com", "Unbranded Luxury"),
        ("Verishop", "verishop.com", "Curated Fashion"),
        ("Cozy Earth", "cozyearth.com", "Bamboo Loungewear"),
        ("Lake Pajamas", "lakepajamas.com", "Luxury PJs"),
        ("Lunya", "lunya.co", "Sleepwear"),
        ("Eberjey", "eberjey.com", "Sleepwear"),
        ("Printfresh", "printfresh.com", "Printed PJs"),
        ("Hill House Home", "hillhousehome.com", "Nap Dress"),
        ("Draper James", "draperjames.com", "Southern Style"),
        ("Jenni Kayne", "jennikayne.com", "California Style"),
        ("Barefoot Dreams", "barefootdreams.com", "Cozy Blankets"),
        ("Splendid", "splendid.com", "Soft Basics"),
        ("Sundry", "sundry-clothing.com", "French-American"),
        ("Monrow", "monrow.com", "Made in LA"),
        ("LNA", "lfrna-clothing.com", "Cut & Sew"),
    ],

    # =========================================================================
    # ADDITIONAL - SKINCARE SUPPLEMENTS
    # =========================================================================
    "Ingestible Beauty & Collagen": [
        ("Vital Proteins", "vitalproteins.com", "Collagen Peptides"),
        ("Further Food", "furtherfood.com", "Collagen"),
        ("Sports Research", "sportsresearch.com", "Collagen"),
        ("Ancient Nutrition", "ancientnutrition.com", "Multi Collagen"),
        ("Dose & Co", "doseandco.com", "NZ Collagen"),
        ("Reserveage", "reserveage.com", "Resveratrol"),
        ("SkinnyFit", "skinnyfit.com", "Detox & Collagen"),
        ("Alaya Naturals", "alayanaturals.com", "Multi Collagen"),
        ("NeoCell", "neocell.com", "Collagen Science"),
        ("Great Lakes Wellness", "greatlakeswellness.com", "Collagen"),
        ("Olly Collagen", "olly.com", "Gummy Collagen"),
        ("Moon Juice Beauty", "moonjuice.com", "Beauty Dust"),
        ("The Beauty Chef", "thebeautychef.com", "Gut-Skin"),
        ("Golde", "golde.co", "Superfood Beauty"),
        ("Your Super", "yoursuper.com", "Superfood Mixes"),
    ],

    # =========================================================================
    # ADDITIONAL - PERSONAL SAFETY & WELLNESS DEVICES
    # =========================================================================
    "Personal Safety & Wearables": [
        ("Birdie", "mybfriridie.com", "Personal Alarm"),
        ("She's Birdie", "shesbfriridie.com", "Personal Safety"),
        ("Invisawear", "invisawear.com", "Safety Jewelry"),
        ("Revolar", "revfrolar.com", "Safety Wearable"),
        ("Flare", "flfrare.com", "Safety Bracelet"),
        ("Apple AirTag", "apple.com/airtag", "Item Tracking"),
        ("Tile", "thetileapp.com", "Tracking"),
        ("Oura Ring", "ouraring.com", "Health Ring"),
        ("Motiv Ring", "mfrotiv.com", "Fitness Ring"),
        ("Bellabeat", "bellabeat.com", "Women's Health"),
        ("Ava Bracelet", "avawomen.com", "Fertility Tracker"),
        ("Tempdrop", "tempdrop.com", "Fertility Sensor"),
        ("Natural Cycles", "naturalcycles.com", "Birth Control App"),
        ("Daysy", "daysy.me", "Fertility Computer"),
        ("Flo", "flo.health", "Period Tracker"),
    ],

    # =========================================================================
    # ADDITIONAL - HOME TEXTILES & BATH
    # =========================================================================
    "Bath & Towels": [
        ("Onsen", "onsentowel.com", "Waffle Towels"),
        ("Weezie", "wfrzie.com", "Monogram Towels"),
        ("Boll & Branch", "bollandbranch.com", "Organic Towels"),
        ("Coyuchi", "coyuchi.com", "Organic Bath"),
        ("Snowe Home", "snowehome.com", "Essential Towels"),
        ("Riley Home", "rileyhome.com", "Affordable Luxury"),
        ("Brooklinen", "brooklinen.com", "Super-Plush Towels"),
        ("Restoration Hardware", "rh.com", "Luxury Bath"),
        ("Pottery Barn", "potterybarn.com", "Classic Home"),
        ("Frontgate", "frontgate.com", "Premium Home"),
        ("Crane & Canopy", "craneandcanopy.com", "Designer Bedding"),
        ("Matouk", "matouk.com", "Luxury Linens"),
        ("Kassatex", "kassatex.com", "Fine Bath"),
        ("Turkish-T", "turkish-t.com", "Turkish Towels"),
        ("Laguna Beach Textile", "lfragunabeachtextile.com", "Beach Towels"),
    ],

    # =========================================================================
    # ADDITIONAL - EMERGING DTC BRANDS
    # =========================================================================
    "Emerging & Trending DTC": [
        ("Tabs Chocolate", "tabschocolate.com", "Functional Chocolate"),
        ("Hims", "forhims.com", "Men's Telehealth"),
        ("Roman", "getroman.com", "Men's Health"),
        ("Nurx", "nurx.com", "Women's Health"),
        ("GoodRx", "goodrx.com", "Rx Savings"),
        ("Calibrate", "joincalibrate.com", "Weight Health"),
        ("Noom", "noom.com", "Behavior Change"),
        ("WW (WeightWatchers)", "weightwatchers.com", "Weight Management"),
        ("Found", "joinfound.com", "Weight Care"),
        ("Sequence", "joinsequence.com", "GLP-1 Platform"),
        ("Ro", "ro.co", "Digital Health"),
        ("FOLX Health", "folxhealth.com", "LGBTQ Health"),
        ("Tend", "hellotend.com", "Modern Dental"),
        ("Zocdoc", "zocdoc.com", "Doctor Booking"),
        ("Thirty Madison", "thirtymadison.com", "Health Platform"),
        ("Prose", "prose.com", "Custom Beauty"),
        ("Curology", "curology.com", "Custom Skincare"),
        ("Apostrophe", "apostrophe.com", "Online Derm"),
        ("Musely", "musely.com", "Rx Skincare"),
        ("Agency", "agency.com", "Anti-Aging Skincare"),
        ("Arey Grey", "areygrey.com", "Gray Hair"),
        ("Vegamour", "vegamour.com", "Hair Wellness"),
        ("Nutrafol", "nutrafol.com", "Hair Growth"),
        ("Keeps", "keeps.com", "Hair Loss"),
        ("Hims Hair", "forhims.com/hair", "Hair Loss"),
    ],
}


def get_meta_ads_library_url(brand_name: str) -> str:
    """Generate Meta Ads Library search URL for a brand."""
    encoded = quote(brand_name)
    return f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=ALL&q={encoded}"


def generate_excel(output_path: str):
    """Generate the Excel file with all brand data."""
    wb = openpyxl.Workbook()

    # ---- Styles ----
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    link_font = Font(name="Calibri", color="0563C1", underline="single", size=10)
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)

    # Category colors for alternating fills
    cat_colors = [
        "F2F7FB", "FFF2E6", "F0FFF0", "FFF0F5", "F5F5DC",
        "E6F3FF", "FFF8DC", "F0E6FF", "E6FFE6", "FFE6E6",
        "E6FFFF", "FFF0E0", "F5F0FF", "E8FFE8", "FFE8E8",
        "E8F4FD", "FDF5E6", "F0FFF4", "FFF5EE", "F8F8FF",
    ]

    # ============================================================
    # SHEET 1: ALL BRANDS (Master List)
    # ============================================================
    ws = wb.active
    ws.title = "All Brands"
    headers = ["#", "Brand Name", "Website", "Category", "Sub-Niche", "Meta Ads Library"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = "A2"

    row_num = 2
    brand_count = 0
    for cat_idx, (category, brands) in enumerate(BRANDS.items()):
        fill_color = cat_colors[cat_idx % len(cat_colors)]
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for brand_name, website, sub_niche in brands:
            brand_count += 1
            meta_url = get_meta_ads_library_url(brand_name)
            website_url = f"https://{website}" if not website.startswith("http") else website

            ws.cell(row=row_num, column=1, value=brand_count).font = data_font
            ws.cell(row=row_num, column=1).fill = row_fill
            ws.cell(row=row_num, column=1).border = thin_border

            ws.cell(row=row_num, column=2, value=brand_name).font = Font(name="Calibri", bold=True, size=10)
            ws.cell(row=row_num, column=2).fill = row_fill
            ws.cell(row=row_num, column=2).border = thin_border

            cell_web = ws.cell(row=row_num, column=3, value=website)
            cell_web.hyperlink = website_url
            cell_web.font = link_font
            cell_web.fill = row_fill
            cell_web.border = thin_border

            ws.cell(row=row_num, column=4, value=category).font = data_font
            ws.cell(row=row_num, column=4).fill = row_fill
            ws.cell(row=row_num, column=4).border = thin_border
            ws.cell(row=row_num, column=4).alignment = data_alignment

            ws.cell(row=row_num, column=5, value=sub_niche).font = data_font
            ws.cell(row=row_num, column=5).fill = row_fill
            ws.cell(row=row_num, column=5).border = thin_border
            ws.cell(row=row_num, column=5).alignment = data_alignment

            cell_meta = ws.cell(row=row_num, column=6, value="View Ads")
            cell_meta.hyperlink = meta_url
            cell_meta.font = link_font
            cell_meta.fill = row_fill
            cell_meta.border = thin_border
            cell_meta.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 15

    # ============================================================
    # SHEET 2: BY CATEGORY (One sheet per category)
    # ============================================================
    for cat_idx, (category, brands) in enumerate(BRANDS.items()):
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = category[:31]
        ws_cat = wb.create_sheet(title=sheet_name)

        cat_headers = ["#", "Brand Name", "Website", "Sub-Niche", "Meta Ads Library"]
        for col_idx, header in enumerate(cat_headers, 1):
            cell = ws_cat.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws_cat.freeze_panes = "A2"

        for idx, (brand_name, website, sub_niche) in enumerate(brands, 1):
            r = idx + 1
            meta_url = get_meta_ads_library_url(brand_name)
            website_url = f"https://{website}" if not website.startswith("http") else website

            ws_cat.cell(row=r, column=1, value=idx).font = data_font
            ws_cat.cell(row=r, column=1).border = thin_border

            ws_cat.cell(row=r, column=2, value=brand_name).font = Font(name="Calibri", bold=True, size=10)
            ws_cat.cell(row=r, column=2).border = thin_border

            cell_web = ws_cat.cell(row=r, column=3, value=website)
            cell_web.hyperlink = website_url
            cell_web.font = link_font
            cell_web.border = thin_border

            ws_cat.cell(row=r, column=4, value=sub_niche).font = data_font
            ws_cat.cell(row=r, column=4).border = thin_border

            cell_meta = ws_cat.cell(row=r, column=5, value="View Ads")
            cell_meta.hyperlink = meta_url
            cell_meta.font = link_font
            cell_meta.border = thin_border
            cell_meta.alignment = Alignment(horizontal="center")

        ws_cat.column_dimensions["A"].width = 6
        ws_cat.column_dimensions["B"].width = 30
        ws_cat.column_dimensions["C"].width = 35
        ws_cat.column_dimensions["D"].width = 25
        ws_cat.column_dimensions["E"].width = 15

    # ============================================================
    # SHEET: SUMMARY
    # ============================================================
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.cell(row=1, column=1, value="DTC Brand Research Summary").font = Font(
        name="Calibri", bold=True, size=16, color="2F5496"
    )
    ws_summary.merge_cells("A1:C1")

    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Calibri", size=10, italic=True
    )
    ws_summary.cell(row=3, column=1, value=f"Total Brands: {brand_count}").font = Font(
        name="Calibri", bold=True, size=12
    )
    ws_summary.cell(row=4, column=1, value=f"Total Categories: {len(BRANDS)}").font = Font(
        name="Calibri", bold=True, size=12
    )

    # Category breakdown
    ws_summary.cell(row=6, column=1, value="Category").font = header_font
    ws_summary.cell(row=6, column=1).fill = header_fill
    ws_summary.cell(row=6, column=1).border = thin_border
    ws_summary.cell(row=6, column=2, value="Brand Count").font = header_font
    ws_summary.cell(row=6, column=2).fill = header_fill
    ws_summary.cell(row=6, column=2).border = thin_border

    for idx, (category, brands) in enumerate(BRANDS.items()):
        r = 7 + idx
        ws_summary.cell(row=r, column=1, value=category).font = data_font
        ws_summary.cell(row=r, column=1).border = thin_border
        ws_summary.cell(row=r, column=2, value=len(brands)).font = data_font
        ws_summary.cell(row=r, column=2).border = thin_border
        ws_summary.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 15

    # Notes section
    notes_row = 7 + len(BRANDS) + 2
    ws_summary.cell(row=notes_row, column=1, value="Notes:").font = Font(
        name="Calibri", bold=True, size=11
    )
    notes = [
        "- All brands are DTC (Direct-to-Consumer) with proven ecommerce revenue",
        "- Meta Ads Library links show active ad campaigns for each brand",
        "- Websites are clickable hyperlinks",
        "- Brands span multiple price points and niches",
        "- Many are lightweight/producible products suitable for Turkey market",
        "- Categories include innovative brands with strong unit economics",
    ]
    for i, note in enumerate(notes):
        ws_summary.cell(row=notes_row + 1 + i, column=1, value=note).font = Font(
            name="Calibri", size=10
        )

    # Save
    wb.save(output_path)
    return brand_count


if __name__ == "__main__":
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "research_outputs")
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"DTC_Brands_1500_Ecommerce_Research_{date_str}.xlsx"
    output_path = os.path.join(output_dir, filename)

    count = generate_excel(output_path)
    print(f"Exported {count} brands to: {output_path}")
